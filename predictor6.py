# @title
"""
매수/매도 앙상블 — 메타 그리드 자동 튜닝 + MDD 제한 + 손절매 + ANCHOR 보정
====================================================================
★ 기능
   1. META_GRID — 점수화/풀 구성 변수를 자동 그리드 탐색
   2. MIN_TRADES_DAILY — 일별 거래수 하한
   3. MAX_DRAWDOWN_LIMIT_PCT — 일별 MDD 한도
   4. 점수화 결과 캐싱
   5. 신호 충돌 해결 — 강도(count/K) 큰 쪽 우선
   6. Balanced Accuracy 평가 + plain accuracy 참고
   7. ★ Tolerance Band 정렬 — 정확도 top-band 이내 중 수익률 최대 (SELECTION_TOLERANCE)
   8. 손절매 (STOP_LOSS_PCT)
   9. ★ ANCHOR 자동 계산 (AUTO_ANCHOR) — 백테스트 데이터로 최적 매수/매도 정답일 자동 산출
"""

import warnings; warnings.filterwarnings('ignore')
import os, math, time, itertools
from datetime import datetime, timedelta
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = '/content/ensemble_analysis'   # ★ 로컬(Colab 세션) 저장 — Drive 저장 안 함

def _resolve_output_dir(target):
    """OUTPUT_DIR 해석 + 폴더 생성. (Google Drive 저장 안 함 — 로컬에만 저장 후 다운로드)"""
    if target is None:
        return '/content' if os.path.isdir('/content') else os.getcwd()
    # ★ Drive 경로가 지정돼 있어도 Drive에 저장하지 않고 로컬로 대체
    if target.startswith('/content/drive/'):
        local_alt = '/content/ensemble_analysis' if os.path.isdir('/content') else os.path.join(os.getcwd(), 'ensemble_analysis')
        print(f"  ℹ Drive 저장 비활성화 — 로컬 폴더에 저장 후 다운로드: {local_alt}")
        target = local_alt
    # 폴더 없으면 생성
    try:
        os.makedirs(target, exist_ok=True)
    except Exception as e:
        print(f"  ⚠ {target} 생성 실패: {e} → /content 로 fallback")
        return '/content' if os.path.isdir('/content') else os.getcwd()
    return target

SCRIPT_DIR = _resolve_output_dir(OUTPUT_DIR)
OUTPUT_DIR = SCRIPT_DIR   # ★ 이후 모든 참조(재현 모드 등)가 로컬 폴더를 가리키도록 동기화
print(f"  📂 출력 폴더(로컬): {SCRIPT_DIR}  — 실행 종료 후 자동 다운로드됩니다")

try:
    from numba import njit
    HAS_NUMBA = True
    print("  ℹ numba 감지 — JIT 가속")
except ImportError:
    HAS_NUMBA = False
    def njit(*args, **kw):
        if args and callable(args[0]): return args[0]
        return lambda f: f
    print("  ⚠ numba 미설치.  pip install numba 권장")


# ════════════════════════════════════════════════════════════════
#                            설정
# ════════════════════════════════════════════════════════════════
EVAL_START          = '2025-01-01'

OOS_ENABLED         = False          # ★ 끔(요청): OOS 미사용, 전체수익 최고 K만
OOS_START           = None           # OOS 미사용

HORIZON_DAYS        = 1
DRAWDOWN_LIMIT_BUY  = 0.02
RUNUP_LIMIT_SELL    = 0.02

# ★ 요청: 신호 다음날 '1~10% 이상' 상승/하락 예측 성공률로 지표 선출.
#   아래 리스트의 각 한도(상승=매수, 하락=매도)로 성공률을 따로 계산해 '최적 한도'를 탐색.
#   (성공 판정: HORIZON_DAYS 이내 종가가 +한도 이상 오르면 매수성공 / -한도 이상 내리면 매도성공)
STAGE_SUCCESS_LIMIT = [0.01, 0.02, 0.03]   # ★ 1~5% (요청: 1~10%에서 축소)
SEARCH_SUCCESS_LIMIT = True        # True면 위 리스트 전부 탐색해 최적 한도 선정

N_THRESHOLDS        = 1000
MAX_INDICATORS      = 3000

# ★ 성공률 우선 풀 선출 (요청) — 점수가 아니라 '성공률'로 먼저 지표를 선발한 뒤 그리드.
#   목적: pct(분위)가 달라 따로 나오던 고성공 지표를 누락 없이 한 풀에 모으고,
#         선발 기준을 점수(Wilson)→성공률 우선으로 바꿈. 가짜 100% 방지 위해 표본 가드 둠.
POOL_SELECT_BY_SUCCESS = True      # True면 풀을 성공률 우선으로 선출(아래 기준), False면 기존 점수순.
POOL_SUCCESS_MIN_RATE  = 0.60      # 성공률 컷오프 (요청: 0.60)
POOL_SUCCESS_MIN_SIG   = 10        # ★ 최소 신호수(요청: 신호 10개 '초과') — 소표본 가짜 100% 방지
POOL_SUCCESS_WIDE_PCT  = (0, 100)  # ★ 풀 평가용 분위 (요청: 0,100 전체 탐색)
POOL_SUCCESS_K_FLOOR   = 2         # ★ 성공률 우선 시 K 하한 — 정예(희소) 풀은 소수 동의로도 신호나야 거래 발생.
# ★ ⓑ 순신호 점수가중 (요청) — net을 '단순 개수' 대신 '지표 점수(성공률) 가중합'으로.
NET_SIGNAL_WEIGHTED    = True
NET_SIGNAL_WEIGHT_COL  = 'success_rate'   # 가중치로 쓸 컬럼: 'success_rate'(성공률) 또는 'score'(Wilson점수)
# ★ 요청: 윌슨 최적값 정해진 뒤, '성공률 비례 점수 가중치'도 몇 개 정해서 윌슨처럼 반복 탐색.
#   각 스킴은 성공률 p를 가중치로 변환하는 지수 g: weight = p**g (g=1이면 성공률 그대로, 클수록 고성공 더 강조).
NET_WEIGHT_SCHEMES   = [0.5, 1.0, 1.5, 2.0, 2.5, 3.0]  # ★ g=0 제거(성공률 무시로 낙폭↑ 문제)
SEARCH_WEIGHT_SCHEME = True        # True면 위 스킴 전부 탐색해 최적 가중 선정
# ★ 요청: 한 지표가 임계치별로 여러 성공률을 가지면, '그날 켜진 것 중 가장 높은 성공률'을 그 지표 점수로 적용.
#   (예: 임계 T1→90%, T2→80%. 90% 신호 뜨면 90% 적용, 90% 꺼지고 80% 켜지면 80% 적용)
NET_MULTI_THRESHOLD_WEIGHT = True

# ★ 미래 예측 정확도 강화 (요청) — 지표 신호 생성·평가 방식 보강.
#   목적: 과거 적합이 아니라 '미래에도 유지되는 신호'를 우대해 매수/매도 시점 예측력↑.
# (1) z-스코어 신호: 절대 임계 대신 롤링 z=(x-평균)/표준편차에도 임계를 걸어 신호 생성.
#     시장 레짐이 바뀌어도 '평균 대비 몇 시그마'는 의미가 유지됨 → 미래 일반화에 강함.
USE_ZSCORE_SIGNAL   = False    # True면 각 지표를 절대임계 + z스코어임계 두 방식으로 평가
ZSCORE_WINDOW       = 60      # z스코어 롤링 창(거래일). 약 3개월.
ZSCORE_THRESHOLDS   = [-2.5, -2.0, -1.5, -1.0, -0.5, 0.5, 1.0, 1.5, 2.0, 2.5]  # z 임계 후보 (요청 확장)
# (2) OOS 안정성 가중: 지표 점수를 '전체기간 Wilson'에만 의존하지 말고,
#     기간을 앞/뒤로 나눠 둘 다 좋은 지표(시간적으로 안정)에 가산점 → 과최적화 억제.
USE_OOS_STABILITY   = False   # ★ 요청: OOS 관련 기능 전부 OFF
OOS_STABILITY_WEIGHT = 0.3    # 안정성 가중 강도 (0=기존과 동일, 1=안정성 절반 반영)

# ★ 큰 상승/하락 적중 가산점 (요청) — 지표 신호 중 '큰 움직임'을 맞춘 비율이 높을수록
#   점수에 가산. 매수지표: 신호 뒤 큰 상승을 맞춘 비율, 매도지표: 큰 하락을 맞춘 비율.
#   목적: 자잘한 적중만 많은 지표보다 '굵직한 변동을 잡는' 지표를 우대 → 실전 수익 직결.
USE_BIG_MOVE_BONUS   = True   # True면 큰움직임 적중비율을 점수에 가산
BIG_MOVE_THRESHOLD   = 0.03   # 신호 뒤 horizon 내 유리방향 최대변동이 이 값(3%) 이상이면 '큰 움직임'
BIG_MOVE_BONUS_WEIGHT = 0.5   # 가산 강도: score *= (1 + W * 큰움직임적중비율)

# ════════════════════════════════════════════════════════════════
# ★ 지표 선출·점수 개선 (요청) — 매매 규칙(KL 순신호)은 유지, 지표 선별만 개선.
#   각 아이디어는 개별 flag로 켜고 끔. A/B 검증(SELECTION_AB_VERIFY)이 자동으로 켜져서
#   각 flag의 ON/OFF 성적(수익률·MDD·보유중하락)을 엑셀 시트에 남긴다.
# ════════════════════════════════════════════════════════════════
# [SEL-1] 풀 정렬 기준 선택 — 현재는 raw success_rate. 대안: expected(SR×평균움직임),
#   wilson(하한 = 표본크기 반영), skill(SR − 기저확률 = 순 알파).
POOL_RANK_BY = 'success'        # 'success' | 'expected' | 'wilson' | 'skill'
# [SEL-2] 신호 후 '불리방향' 감점 — 성공률은 좋아도 신호 후 큰 손실을 겪는 지표 강등.
#   avg_adverse = 신호 후 horizon 내 평균 최악 불리방향(매수=최저하락, 매도=최고상승).
USE_ADVERSE_PENALTY   = True
ADVERSE_PENALTY_WEIGHT = 0.5     # score *= max(0.5, 1 − W × |avg_adverse|/limit)
# [SEL-3] 지표당 최대 임계 개수 — 지표당 상위 K개(성공률순)만 유지 → 풀 노이즈↓, 다양성↑.
MAX_THRESHOLDS_PER_INDICATOR = 3   # 0 = 무제한 (기존과 동일)
# [SEL-4] 홀드아웃 감쇠 감점 — 훈련 vs 홀드아웃 스킬 감쇠가 크면 강등. enrichment의
#   skill/skill_holdout 활용.  감쇠 = max(0, skill − skill_holdout).
USE_HOLDOUT_DECAY_PENALTY = True
HOLDOUT_DECAY_WEIGHT      = 1.0  # score *= max(0.5, 1 − W × 감쇠)
# [SEL-5] 신호수 하한 적응 — 절대치 10 대신 데이터 길이의 X% 이상 요구 (더 안정된 소표본 방어).
USE_ADAPTIVE_MIN_SIG = False
ADAPTIVE_MIN_SIG_PCT = 0.03       # 전체 거래일의 3% 이상 (예: 500일→15개)

# ★ A/B 검증 (요청) — 각 개선 아이디어의 ON/OFF 성적을 KL 백테스트로 비교해 엑셀 시트로 저장.
#   Baseline = 모든 신규 flag OFF (원본 로직) → 각 flag 하나씩 켜서 개별 기여 측정
#   →  전부 켠 조합(현재 defaults)까지 총 N+2회 KL 백테스트.  시간 소폭 증가.
SELECTION_AB_VERIFY = True         # False면 A/B 검증 건너뛰고 실행 로그만 남김
SELECTION_AB_QUICK  = False        # True면 각 flag 개별 ON만 하고 조합은 생략 (시간 절약)

# ════════════════════════════════════════════════════════════════
# ★ 미래 예측 로직 개선 (요청) — 리드타임 탐색 · 스킬 필터 · 홀드아웃 가드 · 검증 시트
#   목적: '반등/하락 며칠 전 신호가 가장 정확한지'를 지표별로 찾아 체결 시점에 정렬하고,
#         성공 정의를 '기저확률(그냥 아무 날이나 골라도 맞는 확률)을 초과하는가'로 보강해
#         과거 적합이 아닌 미래 예측력을 우대. 각 적용사항은 엑셀 '검증_예측로직' 시트에서 검증.
# ════════════════════════════════════════════════════════════════
# (1) 리드타임(선행일) 탐색 — 지표별로 '신호 후 h일 이내 한도 도달' 성공률을 h 후보 전부에서 측정.
#     단, h가 길수록 성공률이 기계적으로 올라가므로(도달 기회↑) 반드시 기저확률을 빼서 비교:
#     스킬(h) = 성공률(h) − 기저확률(h).  best_lead = 스킬 최대 h → '몇 일 전 신호가 가장 정확한지'.
LEAD_TIME_SEARCH    = True
LEAD_HORIZONS       = [1, 2, 3, 4, 5, 7, 10]   # 선행일(h) 후보 (거래일)
# (2) 신호 지연 정렬 — best_lead가 긴 지표(예: 3일 전에 미리 켜지는 지표)는 신호를 d일 늦춰
#     체결(다음날) 직전에 켜지도록 정렬. 훈련구간 스킬이 최소 개선폭 이상 & 홀드아웃 스킬이
#     유지될 때만 채택(과최적화 방지). 지연은 sig[t-d]→t 이므로 미래참조 없음.
LEAD_SHIFT_ENABLED  = True
LEAD_SHIFT_MAX      = 5        # 지연 최대 일수
LEAD_SHIFT_MIN_GAIN = 0.05     # 지연 채택 최소 스킬 개선 (훈련구간, 5%p)
LEAD_SHIFT_HO_TOL   = 0.02     # 지연 채택 시 홀드아웃 스킬 허용 하락폭 (2%p 이내)
# (2b) ★ 선출 단계 리드 반영 — 지연 탐색을 '풀 선출(점수화) 안'에서 수행.
#     핵심: '3일 전에만 정확한' 지표는 다음날(1일) 성공률이 0%라 기존 선출에서 아예 탈락함.
#     선출 시 지연 d(0~MAX)를 함께 탐색해 최적 d로 성공률을 계산하면 이런 선행 지표가
#     처음부터 풀에 들어온다. (d 선택은 훈련구간 스킬 + 홀드아웃 가드 — 과최적화 방지)
LEAD_SELECT_IN_SCORING = True
# (3) 삼중배리어 판정 (리드 탐색용) — h일 이내 '유리 한도'가 '불리 한도'보다 먼저 도달해야 성공.
#     긴 h에서 '먼저 -3% 빠졌다가 나중에 +1% 회복'을 성공으로 세는 왜곡 제거 (실전 손절 현실 반영).
#     ※ 체결 지평(HORIZON_DAYS=1) 성공 판정은 기존 정의 그대로 유지 — 풀 선출 기준 불변.
LEAD_TRIPLE_BARRIER = True
# (4) 스킬 필터 — 성공률이 높아도 기저확률(시장이 원래 그만큼 오르내림)을 못 넘으면 예측력 0.
#     체결 지평 기준 스킬 = 성공률 − 기저확률 이 하한 미만이면 풀에서 제외. (정렬은 기존대로 성공률순 유지)
POOL_REQUIRE_SKILL  = True
POOL_MIN_SKILL      = 0.0      # 스킬 하한 (0 = 기저확률 초과만 요구)
# (5) 홀드아웃 가드 — 기간 뒤쪽 일부를 홀드아웃으로 떼어 훈련/홀드아웃 성공률을 따로 계산(보고),
#     홀드아웃 스킬이 크게 음수인(미래에 안 통한) 지표만 제외. 홀드아웃 신호 3개 미만이면 판정 보류(통과).
#     ※ 최종 K/L·수익 선정은 기존대로 전체구간 — OOS 선정 기능(끔)과 무관한 '지표 수준' 가드.
POOL_HOLDOUT_FRACTION  = 0.30  # 뒤쪽 30%를 홀드아웃으로
POOL_HOLDOUT_GUARD     = True
POOL_HOLDOUT_MIN_SKILL = -0.10 # 홀드아웃 스킬 허용 하한 (-10%p 미만이면 제외)
# (6) 검증 시트 — 위 (1)~(5)가 의도대로 동작하는지 실행 시점에 자체 점검해 엑셀에 기록.
#     성공률 독립 재계산 일치, KL 수익 재계산 일치, 기저확률 표, 지표별 리드/스킬/홀드아웃,
#     순열(원형시프트) 검정(우연 대비 우위 백분위) 포함.
VERIFY_SHEET_ENABLED = True
VERIFY_PERM_N        = 60      # 순열 검정 횟수 (지표당)
VERIFY_PERM_MAX_ROWS = 40      # 순열 검정 대상 풀 행 수 상한 (한쪽당, 시간 보호)

K_BUY_RANGE         = [i for i in range(10, 100)]
K_SELL_RANGE        = [i for i in range(10, 100)]
VOTE_RATIO_BUY      = [0.2, 0.25, 0.3, 0.35, 0.4, 0.45, 0.5,
                       0.55, 0.6, 0.65, 0.7, 0.75, 0.8, 0.85]
VOTE_RATIO_SELL     = [0.2, 0.25, 0.3, 0.35, 0.4, 0.45, 0.5,
                       0.55, 0.6, 0.65, 0.7, 0.75, 0.8, 0.85]

COST_PER_TRADE      = 0.004

MIN_TRADES_DAILY    = 10
MAX_DRAWDOWN_LIMIT_PCT = 0.03   # ★ 분수(0.03=3%). 최대낙폭 3% 이내

STOP_LOSS_PCT       = 0.05
# ★ IC 기반 지표 선정 (A+B+C) — 성공률 대체.  기본 OFF (실험적).
#   ON 시: |IC|·|OOS IC|·|OOS IR|·FDR 조건으로 지표 선정 + OOS IC를 순신호 가중치로.
#   ⚠ 375일 규모 데이터는 통계적 검정력이 낮아 노이즈가 함께 통과할 수 있음. 참고용.
USE_IC_SELECTION    = False   # True로 켜면 IC 방식, False면 기존 성공률 방식
IC_MIN              = 0.05
OOS_IC_MIN          = 0.03
OOS_IR_MIN          = 0.3
FDR_ALPHA           = 0.10
FDR_ENABLED         = True    # False면 IC/IR 조건만
PURGED_K_FOLDS      = 5
PURGED_EMBARGO      = 5

# ★ 선정 우선순위
#   'winrate_return'     : 일별거래 승률 최고 -10%p 밴드 → 그중 누적수익 최고 (요청, 기본)
#   'sell_mdd_return'    : 매도성공률 최고 -2%p 밴드 → MDD 최저 -1%p 밴드 → 누적수익 최고
#   'avgband_mdd_return' : 평균성공률 top band 안에서 → MDD 가장 낮은(0.1%p 단위 동률이면 수익)
#   'independent'     : 매수성공률 최고 설정 + 매도성공률 최고 설정을 따로 찾아 합침.
#                       성공률은 매수/매도가 독립이므로 각각 최댓값을 동시 달성. 수익·MDD는 따라옴.
#                       (MDD 한도가 켜져 있으면 '한도 통과분 중' 각각 최고를 찾음)
#   'stability'       : 매도성공·매수성공·누적수익·MDD방어를 종합한 안정성 점수 최대
#   'sell_buy_return' : 매도성공률 → 매수성공률 → 누적수익 순 (밴드 줄세우기)
#   'balacc_return'   : 기존 (평균 BalAcc → 수익률)
SELECTION_PRIORITY = 'winrate_return'

# ★ 안정성 종합 점수 가중치 (SELECTION_PRIORITY='stability'일 때 사용)
#   (매도성공률, 매수성공률, 누적수익, MDD방어) — 합이 1이 되도록 자동 정규화됨.
#   가중 기하평균이라 한 요소라도 후보군 내 최저면 점수가 크게 깎임 → 골고루 좋은 조합 선호.
#   하락 회피 강화하려면 매도성공률(첫째)·MDD방어(넷째) 비중을 올리세요.
STABILITY_WEIGHTS = (0.30, 0.25, 0.20, 0.25)

# ★ 가중 투표 (변경2) — 지표 성공률(Wilson)에 비례해 표 가중
#   USE_WEIGHTED_VOTE=False면 기존 일반 투표(모두 1표)
#   가중 강도는 점수 분포 편차에 따라 자동 조절(데이터가 결정), 상한 WEIGHT_MAX_RATIO
USE_WEIGHTED_VOTE = True
WEIGHT_MAX_RATIO  = 1.6   # (구) 선형 가중 상한 — 아래 희석방지 장치가 켜지면 보조 역할

# ★ 점수 희석 방지 장치 (요청) — 자잘한 다수 지표가 합쳐져 강한 소수 지표를 덮지 않도록,
#   가중치를 '절대 점수 차이'에 지수적으로 반응시킨다(소프트맥스형). 강한 지표일수록 표가
#   급증해 합산에서 살아남는다. 단 한 지표가 전부를 지배하지 않게 상·하한으로 '적당히' 제한.
#     - VOTE_WEIGHT_TEMP: 온도(작을수록 강한지표에 표 더 몰림). 0.15≈적당.
#     - VOTE_STRONG_CAP : 최고지표가 가질 수 있는 표 상한(평균=1 기준 배수). 과도지배 방지.
#   끄려면 USE_ANTIDILUTION=False (그러면 기존 선형 WEIGHT_MAX_RATIO 가중).
USE_ANTIDILUTION  = True
VOTE_WEIGHT_TEMP  = 0.15
VOTE_STRONG_CAP   = 4.0

SELECTION_TOLERANCE = 0.04

# ★ 'sell_mdd_return' 모드 밴드 폭
SELL_SUCCESS_TOLERANCE = 0.02   # 매도성공률 최고에서 이 차이(2%p)까지 후보
MDD_TOLERANCE          = 0.01   # MDD 최저에서 이 차이(1%p)까지 후보 (그중 수익 최고 선택)

# ★ 'winrate_return' 모드 밴드 폭 (요청) — 일별거래 승률 최고에서 이 차이(10%p)까지 후보
WINRATE_TOLERANCE      = 0.04

# ★ 승률 후보 실거래 검증 (요청) — 그리드는 빠른 근사라 실제 일별거래와 MDD·승률·수익이
#   다를 수 있음. 그래서 승률 상위 후보만 골라 '실제 일별 백테스트'를 돌려 진짜 수치를
#   구하고, 그중 실제 누적수익이 가장 높은 조합을 최종 선정한다.
VERIFY_BY_DAILY_BACKTEST = False  # ★ 끔(요청): net>K 시스템 사용 → 그리드 실거래검증(수십분) 불필요. 폴백 선정 사용.
VERIFY_TOP_N             = 0      # ★ 실거래검증 OFF라 미사용(죽은 값)

# ★ 실거래 검증 후, '실제 최대 거래손실'이 이 값 이하(더 안전)인 후보만 선정 대상으로 (요청).
#   예: -0.03 이면 실제 단일거래 최대손실이 -3%보다 깊지 않은 조합만 후보.
#   None 이면 이 필터를 끈다. (단, 필터로 후보가 0개면 자동으로 필터를 완화해 최선을 고름)
VERIFY_MAX_DRAWDOWN_LIMIT = -0.03

# ★ 최종 선정 2차 밴드 (요청) — 실제 승률 밴드 후보 중, 실제 '평균 성공률' 최고에서
#   이 값(3%p) 이내를 다시 후보로 두고, 그중 실제 수익률 최고를 선정.
VERIFY_AVG_SUCCESS_TOLERANCE = 0.4

# ★ 수익 동률 판정 (요청) — 실제 수익률 차이가 이 값(3%p) 이내면 '동률'로 보고,
#   그중 실제 평균 성공률이 더 높은 조합을 선정.
VERIFY_RETURN_TIE = 0.03

# ★ 최종 선정 2차 밴드 (요청) — 실제 승률 밴드 후보 중, 실제 '평균 매칭률'(매수·매도 앵커 매칭)
#   최고에서 이 값(2%p) 이내를 다시 후보로 두고, 그중 실제 수익률 최고를 선정.
VERIFY_MATCH_TOLERANCE = 0.4

# ★ Buy&Hold 미달 조합 제외 (전략 누적수익이 B&H 이하면 후보에서 버림)
EXCLUDE_BELOW_BH = False

ANCHOR_MATCH_PRIORITY = False   # ★ 선정은 평균성공+MDD+수익 기준으로. 매칭 우선이 그걸 덮지 않도록 OFF
ANCHOR_MATCH_TOLERANCE = 0.10

ANCHOR_MODE = False   # ★ 끔(요청): 앵커 미사용 → plain BalAcc로 선정

AUTO_ANCHOR = False   # ★ 끔(요청): 앵커 정답일 자동계산 안 함
AUTO_ANCHOR_WINDOW     = 1
AUTO_ANCHOR_LOOKFORWARD = 1
AUTO_ANCHOR_MIN_RISE   = 0.01
AUTO_ANCHOR_MIN_DROP   = 0.01
AUTO_ANCHOR_PRICE_TOLERANCE = 0.01
AUTO_ANCHOR_MAX_DATES  = None

# ★ 앵커 매칭 타이밍 윈도우 (요청) — 신호는 정답일보다 먼저 떠야 정답일에 포지션이 맞음.
#   매칭 판정 시 '정답일 당일 ~ +N일' 사이에 실제 포지션이 맞으면 매칭으로 인정.
#   신호→익일체결 지연 + 앵커 정답일이 며칠에 걸친 바닥/천장 구간인 점을 반영.
#   0이면 당일만, 1이면 당일+익일, 2면 +2일까지. (충돌·실제매매는 그대로 반영)
ANCHOR_MATCH_WINDOW = 2

# ★ 매도 매칭의 '고점 전 보유' 조회 범위 (요청) — 고점 직전 W일만 보면 너무 빡빡해
#   매도매칭이 50%대로 낮게 나옴. 고점까지 오는 과정 최근 N일 중 보유한 적 있으면
#   '고점까지 들고 왔다'로 인정 (그 뒤 고점 근처 청산이면 매칭).
ANCHOR_SELL_HOLD_LOOKBACK = 5

# ★ 충돌 처리 정책 (요청) — '매도 우선':
#   True: 매수·매도 신호 충돌 시, 현금이면 매수 보류(잘못된 매수 상쇄),
#         보유면 청산(올바른 매도는 안 상쇄). 위험회피적, 손실 방지 우선.
#   False: 기존 강도 비교(b_strength vs s_strength).
CONFLICT_SELL_PRIORITY = False

# ★ 최종 선정 우선순위 (요청) — 매도 정확도 우선 → 매수 정확도 → 수익:
#   1) 실제 매도 성공률(정확도) 최고에서 -SELL_ACC_TOL 범위
#   2) 그중 실제 매수 성공률 최고에서 -BUY_ACC_TOL 범위
#   3) 그중 실제 수익 최고
VERIFY_SELL_ACC_TOLERANCE = 0.01   # 매도 정확도 1차 밴드 (3%p)
VERIFY_BUY_ACC_TOLERANCE  = 0.01   # 매수 정확도 2차 밴드 (3%p)

# ★ CatBoost 액션 보정 (요청) — 선정된 조합의 신호로 백테스트한 뒤, 앵커 정답과 안 맞는
#   매수/매도 액션을 머신러닝으로 보정해본다.
#   - 피처: 매수/매도 신호 강도, 최근 수익/변동성 등 (그 시점까지 정보만)
#   - 타겟: 앵커 정답(올라야 할 날=보유, 내려야 할 날=현금)
#   - walk-forward(시계열 분할)로 학습→이후구간 적용. 미래참조를 피하려 과거로만 학습.
#   ⚠ 앵커는 사후적 정답이라, 보정 결과는 '참고용'. 실전 일반화는 별도 검증 필요.
USE_CATBOOST_CORRECTION = False  # ★ 요청: CatBoost 보정 제거 — 앵커 기반 가중치 보정만 사용
CATBOOST_MIN_TRAIN_DAYS = 120    # 최소 학습일수 (이만큼 쌓인 뒤부터 보정 적용)
CATBOOST_PROB_THRESHOLD = 0.5    # 보유 확률이 이 값 이상이면 '보유'로 보정
CATBOOST_TARGET_HORIZON = 5      # ★ 타겟: 미래 N일 수익 (앵커 아님 — 실전 검증 가능한 라벨)
CATBOOST_OOS_FRACTION   = 0.2    # ★ 마지막 N%는 학습에서 빼고 '진짜 미래'로 성능 측정

# ★ OOS(out-of-sample) 검증 기간 (요청) — 최근 이 개월은 학습/탐색에서 빼고,
#   후보 그리드를 이 기간에서만 백테스트해 'OOS 수익/정확도'를 측정.
#   OOS 수익이 가장 높은 조합을 최종 선정 → 미래 일반화에 가까운 선택.
OOS_MONTHS = 1                   # 최근 N개월을 OOS로 (이전달~현재)
OOS_SELECT_BY_OOS_RETURN = False # ★ 요청: OOS 관련 기능 전부 OFF

# ★ 앵커 미매칭 보정 (요청) — 최적 그리드 선정 후, CatBoost 보정 '전'에 실행.
#   충돌 패배·간발의 차로 앵커와 안 맞은 매수/매도를 '지표 가중치 조정'으로 보정.
#   단, 조정 후 수익이 떨어지면 롤백(기존 그리드 유지) → 미매칭만 보정, 수익 손해 없음.
USE_ANCHOR_MATCH_CORRECTION = True
# ★ 검증 후보 보정 통합 (요청) — 실거래 검증 후보 중 수익 상위 N개에 지표 가중치 보정을
#   적용하고, 보정 후 수익/손실/승률을 재계산. 보정 후 수익이 가장 높은 조합을 선정.
#   (모든 후보에 보정하면 너무 느려서 수익 상위 N개로 제한)
CORRECT_VERIFY_TOP_N = 0         # 수익 상위 몇 개 후보에 보정을 적용할지
CORRECT_PROFIT_FLOOR = 0.01      # 거래 진단: 수익 이 값(1%) 이하 + 손실 거래를 '틀린 거래'로
SELECT_BY_CORRECTED_RETURN = True  # True면 보정 후 수익 최고로 선정
# ★ 후보 선정 필터를 '보정 후' 수치로 (요청) — 최대거래손실 한도·승률 밴드를 보정 적용된
#   조합은 보정후 수치로 판정. 보정 안 된 조합은 실제 수치 그대로. SELECT_BY_CORRECTED_RETURN과 짝.
SELECT_FILTER_BY_CORRECTED = True
# ★ 보정후 승률 하한 (요청) — 보정후 실제 거래 승률이 이 값 이상인 조합 중 보정후 수익 최고를
#   1차 선정. 이후 그 승률에서 WINRATE_TOLERANCE 이내 후보와 수익 비교해 더 높으면 재선정.
SELECT_WIN_FLOOR = 0.90

ANCHOR_BUY_DATES = [
]
ANCHOR_SELL_DATES = [
]

SELECT_BY           = 'total_return'
TOP_N_GRID_OUT      = 10000
# ★ 그리드 행별 net>K를 상위 몇 행만 계산할지 (메모리/시간 보호). net>K 시트는 합친 풀로 별도 계산되므로
#   이 값은 '그리드 시트 🎯K 표시'용일 뿐. 낮출수록 빠르고 가벼움. 0이면 그리드 net>K 완전 생략.
KNET_GRID_ROW_CAP   = 100

META_GRID = {
    # ★ staged 방식의 '시작값'. 단계 탐색은 STAGE_PCT_RANGE / STAGE_WILSON_Z /
    #   STAGE_CORR_LIMIT 후보들을 순서대로 돌리며 좁힌다 (모든 조합 X).
    'wilson_z':    [1.95],
    'pct_range':   [(0, 100)],
    'min_signals': [10],
    'corr_limit':  [0.2],
    'top_n_pool':  [100],
}

STAGED_META_TUNE = False  # ★ 끔(요청): 단계적 튜닝은 옛 그리드-투표용 → 그리드 8회 재실행 낭비. k순신호는 corr만 자체 탐색.
# ★ 그리드-투표 K탐색 최소화 (요청) — k순신호만 사용 시 True. K_buy×K_sell 99×99(≈700초) → 1×1로 축소.
#   그리드-투표 시트(현재 포지션 등)는 대략값이 되지만, net>K 시트는 합친 풀로 정확. 실행 대폭 단축.
SKIP_GRID_VOTE = True
                          #   단계에서 돌린 결과들을 한 엑셀에 모두 모아 최종 판단.
STAGE_PCT_RANGE   = [(0, 100)]
STAGE_WILSON_Z    = [1.95]
STAGE_WILSON_REFINE_STEP = 0
STAGE_CORR_LIMIT  = [0.2]

# ============================================================
# ★ 테스트 모드 — 빠르게 동작만 확인할 때 True. (정식 분석은 False)
#   지표수·임계수·한도수·윌슨후보를 줄여 수 분 내로 1회 돌게 함.
# ============================================================
TEST_MODE = False
if TEST_MODE:
    MAX_INDICATORS      = 150          # 1000 → 150 (후보 지표 대폭 축소)
    N_THRESHOLDS        = 120          # 800 → 120 (임계 후보 축소)
    STAGE_SUCCESS_LIMIT = [0.02, 0.03] # 1~5% → 2개만 (한도 탐색 빠르게)
    STAGE_WILSON_Z      = [1.95]       # 윌슨 1개
    STAGE_PCT_RANGE     = [(0, 100)]
    STAGE_CORR_LIMIT    = [0.2]
    CORRECT_VERIFY_TOP_N = 0           # 보정 검증 끔(이미 0)
    print("⚡ TEST_MODE ON — 축소 설정(지표150/임계120/한도2개/윌슨1개)로 빠른 확인용 실행")
# ============================================================

PREFILTER_ENABLED          = False
PREFILTER_MIN_CORR         = 0.005
PREFILTER_MIN_VARIANCE_REL = 1e-6
PREFILTER_MAX_NAN_RATIO    = 0.5

AUTO_DOWNLOAD_EXCEL = True

# ★ 재현 방식 (요청) — 모드 2/3가 예전에 정확히 맞던 방식과 동일하게.
#   재현 때 '원본 엑셀의 마지막 날짜까지'로 데이터를 잘라 compute_features에 넘긴다.
#   → 전체구간 순위(rank)·정규화 지표가 원본과 '같은 범위'로 계산돼 신호가 정확히 일치.
#   (재현이 하루 더 받아 순위가 다시 매겨지던 문제 해결. 연장은 안 함 = 원본 날짜까지 정확 재현)
REPLAY_MATCH_ORIGINAL_RANGE = True

# ★ 재현 후 새 거래일 연장 (요청) — 기본 끔(OFF). 정확 재현이 우선.
#   켜면(True) 과거(원본 마지막날)까지는 정확 재현 + 그 이후 새 거래일을 이어 계산해 덧붙인다.
#   ⚠ 단, 전체구간 순위 지표가 있으면 '정확 재현'과 '연장'을 100% 동시 보장하기 어렵다.
#   새 데이터까지 보려면 '새 분석'을 한 번 더 돌리는 것이 가장 정확하다.
REPLAY_EXTEND_TO_TODAY = False

# 데이터 스냅샷(.pkl 저장) 방식 — 기본 끔. (원본 범위 자르기로 충분)
#   True로 켜면 분석 때 feat/close를 ..._data.pkl로 저장하고 재현 때 그걸 그대로 사용.
SAVE_DATA_SNAPSHOT  = False
REPLAY_USE_SNAPSHOT = False

# ★ summary(전체 티커 요약) 파일 생성 여부 — 요청으로 기본 OFF.
#   끄면 종목별 분석 엑셀만 만들고, ensemble_summary_all_tickers.xlsx는 안 만든다.
WRITE_SUMMARY_FILE = False

# ════════════════════════════════════════════════════════════════
#   ★ 실행 모드 — input() 대신 여기서 직접 지정 (요청)
#     RUN_MODE 로 동작을 고른다. Colab에서 이 값만 바꿔 실행하면 됨.
# ════════════════════════════════════════════════════════════════
#   RUN_MODE = 1 : 새 분석 실행
#            = 2 : 기존 결과의 '그리드 번호'로 일별 Excel 재현
#            = 3 : 티커의 '가장 최근 분석 엑셀'에서 ★최적 조합 자동 재현 (지표·변수 그대로)
#            = 4 : 드라이브 폴더의 '티커별 가장 최근' 엑셀들을 모두 ★최적 재현(현재까지) + 요약
RUN_MODE = 3

# ── 모드 1 (새 분석) 설정 ──
#   분석할 티커 목록. None 이면 코드의 기본 TICKERS 사용.
RUN_TICKERS = None                 # 예: ['STX', 'GOOG', 'ORCL']  또는  None
#   기존 요약 엑셀의 변수값을 사용할지 (True/False)
RUN_USE_EXCEL_OVERRIDES = False

# ── 모드 2 (그리드 번호 재현) 설정 ──
RUN_REPLAY_FILE = 'ensemble_search_VRT_2026-06-01.xlsx'   # OUTPUT_DIR 안의 파일명
RUN_REPLAY_GRID_NUMBER = 14                                # 재현할 그리드 번호

# ── 모드 3 (최근 엑셀 ★최적 자동 재현) 설정 ──
RUN_REPLAY_TICKER = 'STX'          # 재현할 티커 (가장 최근 일반 분석 엑셀을 자동으로 찾음)

# ── 모드 4 (드라이브 폴더 일괄 ★재현 + 요약) 설정 ──
#   드라이브의 ensemble_analysis 폴더에서 '티커별 가장 최근' 분석 엑셀을 모두 찾아,
#   각 ★최적 조합을 '현재까지' 일별 백테스트로 재현하고 티커별 요약 엑셀까지 만든다.
#   결과물은 '날짜 폴더'(드라이브폴더/오늘날짜)에 저장. (먼저 drive.mount 필요)
RUN_MODE4_DRIVE_DIR   = '/content/drive/MyDrive/ensemble_analysis'  # 드라이브 폴더
RUN_MODE4_DATE_FOLDER = None       # 출력 날짜 폴더명. None=오늘 날짜 자동


def wilson_lower(k, n, z):
    if n == 0: return 0.0
    p = k / n
    denom  = 1.0 + z * z / n
    center = p + z * z / (2 * n)
    half   = z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n))
    return max(0.0, (center - half) / denom)


def compute_vote_weights(scores, max_ratio=1.6):
    """성공률(Wilson 점수)에 비례한 투표 가중치. 데이터가 강도를 결정.
       - 점수 분포 편차 크면 차등↑, 작으면 균등에 가깝게
       - 최고/최저 비율 max_ratio 이하 (온건)
       - 평균 1로 정규화 → 기존 vote 임계값 스케일 그대로 호환"""
    s = np.clip(np.asarray(scores, dtype=np.float64), 0.0, 1.0)
    n = len(s)
    if n == 0:
        return np.ones(0, dtype=np.float64)
    spread = float(s.max() - s.min())
    if spread < 1e-9:
        return np.ones(n, dtype=np.float64)

    # ★ 희석 방지 장치 (요청) — 절대 점수차에 지수 반응(소프트맥스형). 강한 지표 표가 급증해
    #   자잘한 다수에 묻히지 않음. 상·하한으로 단일 지표의 과도지배만 막는다(적당히).
    if globals().get('USE_ANTIDILUTION', False):
        T   = max(float(globals().get('VOTE_WEIGHT_TEMP', 0.15)), 1e-6)
        cap = float(globals().get('VOTE_STRONG_CAP', 4.0))
        w = np.exp((s - s.max()) / T)      # 0<w<=1, 최고점수=1 (수치 안정)
        w = w / w.mean()                   # 평균 1 정규화 (vote 비율 스케일 호환)
        w = np.clip(w, 1.0 / cap, cap)     # 과도지배/과도소멸 방지
        w = w / w.mean()
        return w.astype(np.float64)

    # (기존) 선형 가중 — 최고/최저 비율 max_ratio 이하
    z = (s - s.min()) / spread
    beta = min(1.0, float(np.std(s)) / 0.10)
    w = 1.0 + z * (max_ratio - 1.0) * beta
    w = w / w.mean()
    return w.astype(np.float64)


def compute_stability_scores(df, weights=None):
    """후보 DataFrame에 '안정성 종합 점수'(stability_score) 컬럼을 붙여 반환.
       구성: 매도성공률, 매수성공률, 누적수익(후보군 min-max), MDD방어(후보군 min-max).
       가중 기하평균 — 한 요소라도 후보군 내 최저 수준이면 점수가 크게 깎여
       '종합적으로 안정적인'(골고루 좋은) 조합이 상위로 온다.
       거래수는 별도 MIN_TRADES_DAILY 필터로 이미 보장되므로 점수에는 안 넣음."""
    import numpy as np
    if weights is None:
        weights = globals().get('STABILITY_WEIGHTS', (0.30, 0.25, 0.20, 0.25))
    w = np.asarray(weights, dtype=np.float64)
    w = w / w.sum()
    out = df.copy()
    if len(out) == 0:
        out['stability_score'] = []
        return out
    ret_col = 'total_return' if 'total_return' in out.columns else 'combined_return'
    rets = out[ret_col].astype(float).values
    mdds = out['max_drawdown'].astype(float).values   # 음수, 0에 가까울수록 좋음
    ret_lo, ret_hi = float(np.min(rets)), float(np.max(rets))
    mdd_worst, mdd_best = float(np.min(mdds)), float(np.max(mdds))

    def _nrm(x, lo, hi):
        if hi - lo < 1e-12: return 0.5
        return min(1.0, max(0.0, (x - lo) / (hi - lo)))

    eps = 1e-6
    scores = np.zeros(len(out), dtype=np.float64)
    sell = out['sell_success_rate'].astype(float).values
    buy  = out['buy_success_rate'].astype(float).values
    for i in range(len(out)):
        s = min(1.0, max(0.0, sell[i]))
        b = min(1.0, max(0.0, buy[i]))
        r = _nrm(rets[i], ret_lo, ret_hi)
        m = _nrm(mdds[i], mdd_worst, mdd_best)   # MDD 방어력 (0근처=1, 최악=0)
        comps = (max(s, eps), max(b, eps), max(r, eps), max(m, eps))
        scores[i] = float(np.prod([c ** wi for c, wi in zip(comps, w)]))
    out['stability_score'] = scores
    return out


def _bh_sum_return(close_arr):
    a = np.asarray(close_arr, dtype=np.float64)
    if len(a) < 2:
        return 0.0
    rets = a[1:] / a[:-1] - 1.0
    rets = rets[np.isfinite(rets)]
    return float(np.sum(rets))


def _bh_up_sum_return(close_arr):
    a = np.asarray(close_arr, dtype=np.float64)
    if len(a) < 2:
        return 0.0
    rets = a[1:] / a[:-1] - 1.0
    rets = rets[np.isfinite(rets)]
    return float(np.sum(rets[rets > 0.0]))


def _optimal_swing_points(prices, cost, price_tolerance=0.01):
    """진짜 최대 수익 스윙의 저점(매수)/고점(매도)을 찾는다 (요청).
       anchor_strategy_return과 동일한 로직: 단기 저점에서 사서 오르는 한 보유,
       단기 고점에서 청산. 단 거래비용(2*cost)을 넘는 상승 구간만 채택.
       그 저점/고점 직전 ±price_tolerance 이내 가격이면 같은 앵커로 포함.
       반환: (buy_idx_set, sell_idx_set) — 우선순위 1 (최우선 매칭 대상)
    """
    n = len(prices)
    buy_idx = set(); sell_idx = set()
    i = 0
    while i < n - 1:
        if prices[i+1] > prices[i]:
            j = i                      # j = 국소 저점
            k = j
            while k < n - 1 and prices[k+1] >= prices[k]:
                k += 1                 # k = 국소 고점
            gain = prices[k] / prices[j] - 1.0
            net = (1.0 - cost) * (1.0 + gain) * (1.0 - cost) - 1.0
            if net > 0:                # 비용 넘는 상승 구간만 채택
                base_b = prices[j]
                base_s = prices[k]
                # 저점 j + 직전 ±tol 이내 (j 주변, 같은 가격대)
                for m in range(max(0, j - 5), j + 1):
                    if base_b > 0 and abs(prices[m] / base_b - 1.0) <= price_tolerance:
                        buy_idx.add(m)
                # 고점 k + 직전 ±tol 이내
                for m in range(max(0, k - 5), k + 1):
                    if base_s > 0 and abs(prices[m] / base_s - 1.0) <= price_tolerance:
                        sell_idx.add(m)
            i = k + 1
        else:
            i += 1
    return buy_idx, sell_idx


def auto_compute_anchor_dates(dates, close, *,
                                window=5,
                                lookforward=5,
                                min_rise_after_buy=0.03,
                                min_drop_after_sell=0.03,
                                price_tolerance=0.005,
                                max_dates=None,
                                return_priority=False):
    """앵커 정답일 생성 (요청 개정):
       (우선순위1) 진짜 최대 수익 스윙의 저점(매수)/고점(매도) + 직전 ±tol 이내.
       (우선순위2) 기존 방식: 좌우 window 최저/최고 & lookforward 내 min_rise/drop 이상.
       두 종류를 합치되, 우선순위1을 최우선 매칭 대상으로 표시한다.
       return_priority=True면 (buy_dates, sell_dates, prio_dict) 반환.
       prio_dict: {'buy_p1':set, 'sell_p1':set, 'buy_p2':set, 'sell_p2':set} (날짜 문자열)
    """
    prices = close.values.astype(np.float64)
    n = len(prices)
    cost = float(globals().get('COST_PER_TRADE', 0.004))

    # ── 우선순위1: 최대 수익 스윙 저점/고점 (+ 직전 ±tol) ──
    sw_buy_idx, sw_sell_idx = _optimal_swing_points(prices, cost, price_tolerance)

    # ── 우선순위2: 기존 1% 이상 앵커 ──
    base_buys = []
    base_sells = []
    for j in range(window, n - lookforward):
        win_lo = max(0, j - window)
        win_hi = min(n - 1, j + window)
        win = prices[win_lo:win_hi + 1]
        cur = prices[j]
        future = prices[j:j + lookforward + 1]
        if cur == win.min() and len(future) > 1:
            max_rise = future.max() / cur - 1.0
            if max_rise >= min_rise_after_buy:
                base_buys.append((j, max_rise))
        if cur == win.max() and len(future) > 1:
            max_drop = future.min() / cur - 1.0
            if max_drop <= -min_drop_after_sell:
                base_sells.append((j, -max_drop))
    base_buys.sort(key=lambda p: -p[1])
    base_sells.sort(key=lambda p: -p[1])
    if max_dates is not None:
        base_buys  = base_buys[:max_dates]
        base_sells = base_sells[:max_dates]

    p2_buy_idx = set(); p2_sell_idx = set()
    n_buy_base = 0; n_sell_base = 0
    for j, _ in base_buys:
        base_p = prices[j]
        if base_p <= 0: continue
        n_buy_base += 1
        for k in range(max(0, j - window), min(n, j + window + 1)):
            if prices[k] > 0 and abs(prices[k] / base_p - 1.0) <= price_tolerance:
                p2_buy_idx.add(k)
    for j, _ in base_sells:
        base_p = prices[j]
        if base_p <= 0: continue
        n_sell_base += 1
        for k in range(max(0, j - window), min(n, j + window + 1)):
            if prices[k] > 0 and abs(prices[k] / base_p - 1.0) <= price_tolerance:
                p2_sell_idx.add(k)

    # 우선순위1이 우선 — p2에서 p1과 겹치는 건 p1로 귀속
    p2_buy_only  = p2_buy_idx  - sw_buy_idx
    p2_sell_only = p2_sell_idx - sw_sell_idx

    all_buy_idx  = sw_buy_idx  | p2_buy_only
    all_sell_idx = sw_sell_idx | p2_sell_only

    def _to_dates(idx_set):
        return sorted([dates[k].strftime('%Y-%m-%d') for k in idx_set])

    buy_dates  = _to_dates(all_buy_idx)
    sell_dates = _to_dates(all_sell_idx)

    print(f"     앵커: [우선순위1 최대수익스윙] 매수 {len(sw_buy_idx)}일 / 매도 {len(sw_sell_idx)}일,  "
          f"[우선순위2 {min_rise_after_buy*100:.0f}%이상] 매수 +{len(p2_buy_only)}일 / 매도 +{len(p2_sell_only)}일  "
          f"(직전 ±{price_tolerance*100:.1f}% 포함)")

    if return_priority:
        prio = {
            'buy_p1':  set(_to_dates(sw_buy_idx)),
            'sell_p1': set(_to_dates(sw_sell_idx)),
            'buy_p2':  set(_to_dates(p2_buy_only)),
            'sell_p2': set(_to_dates(p2_sell_only)),
        }
        return buy_dates, sell_dates, prio
    return buy_dates, sell_dates


@njit
def _eval_buy_signals(close_arr, signal_arr, horizon, dd_limit, anchor_buy_arr):
    # ★ 기준(요청): 매수신호 적중 = 신호 후 'horizon일 이내'에 종가가 신호일 종가 대비
    #   +dd_limit 이상 상승(기간 내 최고가 기준). horizon=1이면 '다음날', 5면 '5일 이내 어느 날이든 도달'.
    #   ★ 앵커 오버라이드 없음 — 실제 도달 여부만으로 판정.
    #   anchor_buy_arr 인자는 호출 호환 위해 남겨두나 미사용.
    n = close_arr.shape[0]
    h = horizon if horizon >= 1 else 1
    ns = 0; ok = 0; sum_ret = 0.0
    for i in range(n - 1):
        if signal_arr[i] != 1: continue
        base_p = close_arr[i]
        if base_p <= 0.0: continue
        end = i + h
        if end > n - 1: end = n - 1
        if end <= i: continue
        max_ret = -1.0e18
        for j in range(i + 1, end + 1):
            r = close_arr[j] / base_p - 1.0
            if r > max_ret: max_ret = r
        ns += 1; sum_ret += max_ret
        if max_ret >= dd_limit:
            ok += 1
    return ns, ok, sum_ret


@njit(cache=True)
def _eval_sell_signals(close_arr, signal_arr, horizon, ru_limit, anchor_sell_arr):
    # ★ 기준(요청): 매도신호 적중 = 신호 후 'horizon일 이내'에 종가가 신호일 종가 대비
    #   -ru_limit 이상 하락(기간 내 최저가 기준). horizon=1이면 '다음날', 5면 '5일 이내'.
    #   anchor_sell_arr 인자는 호출 호환 위해 남겨두나 미사용.
    n = close_arr.shape[0]
    h = horizon if horizon >= 1 else 1
    ns = 0; ok = 0; sum_ret = 0.0
    for i in range(n - 1):
        if signal_arr[i] != 1: continue
        base_p = close_arr[i]
        if base_p <= 0.0: continue
        end = i + h
        if end > n - 1: end = n - 1
        if end <= i: continue
        min_ret = 1.0e18
        for j in range(i + 1, end + 1):
            r = close_arr[j] / base_p - 1.0
            if r < min_ret: min_ret = r
        ns += 1; sum_ret += min_ret
        if min_ret <= -ru_limit:
            ok += 1
    return ns, ok, sum_ret


@njit(cache=True)
def _eval_big_move_hits(close_arr, signal_arr, horizon, big_thr, is_buy):
    """신호 중 '큰 움직임'을 맞춘 비율 계산용 (요청).
       매수신호(is_buy=1): 익일 진입가 대비 이후 horizon 내 최대 상승률이 big_thr 이상이면 적중.
       매도신호(is_buy=0): 익일 기준가 대비 이후 horizon 내 최대 하락률이 -big_thr 이하면 적중
         (= 매도해서 큰 하락을 피했다).
       반환: (n_signals, n_big_hits)."""
    n = close_arr.shape[0]
    ns = 0; hit = 0
    for i in range(n - 1):
        if signal_arr[i] != 1: continue
        if i + 1 >= n: break
        base_p = close_arr[i + 1]
        if base_p <= 0.0: continue
        end = i + 1 + horizon
        if end >= n: end = n - 1
        if end <= i + 1: continue
        if is_buy == 1:
            mx = base_p
            for j in range(i + 2, end + 1):
                if close_arr[j] > mx: mx = close_arr[j]
            ns += 1
            if mx / base_p - 1.0 >= big_thr: hit += 1
        else:
            mn = base_p
            for j in range(i + 2, end + 1):
                if close_arr[j] < mn: mn = close_arr[j]
            ns += 1
            if mn / base_p - 1.0 <= -big_thr: hit += 1
    return ns, hit


@njit(cache=True)
def _eval_avg_adverse(close_arr, signal_arr, horizon, is_buy):
    """★ [SEL-2] 신호 후 '불리방향' 평균 (지표 선출 감점용).
       매수신호(is_buy=1): 신호일 종가 대비 horizon 내 최저 하락률의 평균 (음수, 유리하면 0).
       매도신호(is_buy=0): 신호일 종가 대비 horizon 내 최고 상승률의 평균 (양수, 유리하면 0).
       반환: (n_signals, avg_adverse). n_signals=0면 avg_adverse=0.
       — 성공률만 좋아도 신호 후 큰 손실을 겪는 지표는 이 값으로 강등된다."""
    n = close_arr.shape[0]
    h = horizon if horizon >= 1 else 1
    ns = 0; sum_adv = 0.0
    for i in range(n - 1):
        if signal_arr[i] != 1: continue
        base_p = close_arr[i]
        if base_p <= 0.0: continue
        end = i + h
        if end > n - 1: end = n - 1
        if end <= i: continue
        if is_buy == 1:
            worst = 0.0                # 최저 하락 (음수 or 0)
            for j in range(i + 1, end + 1):
                r = close_arr[j] / base_p - 1.0
                if r < worst: worst = r
            sum_adv += worst
        else:
            worst = 0.0                # 최고 상승 (양수 or 0) - 매도인데 오르면 불리
            for j in range(i + 1, end + 1):
                r = close_arr[j] / base_p - 1.0
                if r > worst: worst = r
            sum_adv += worst
        ns += 1
    if ns == 0: return 0, 0.0
    return ns, sum_adv / ns


@njit(cache=True)
def _compute_safe_arrays(close_arr, horizon, dd_limit, ru_limit):
    """
    ★ 정답일 — 신호 후 'horizon일 이내' 종가 변동 기준.
      safe_buy[i]=1  : horizon일 이내 최고가가 신호일 종가 대비 +dd_limit 이상 상승 → '올랐어야'(매수 정답)
      safe_sell[i]=1 : horizon일 이내 최저가가 -ru_limit 이상 하락 → '내렸어야'(매도 정답)
      둘 다 아니면 evaluable=0 → 성공/실패 평가에서 제외.
      (매수·매도 둘 다 가능하면 둘 다 1 — 변동성 큰 구간)
      horizon=1이면 다음날만, 5면 5일 이내.
    """
    n = close_arr.shape[0]
    h = horizon if horizon >= 1 else 1
    safe_buy  = np.zeros(n, dtype=np.uint8)
    safe_sell = np.zeros(n, dtype=np.uint8)
    evaluable = np.zeros(n, dtype=np.uint8)
    for i in range(n - 1):
        base = close_arr[i]
        if base <= 0.0:
            continue
        end = i + h
        if end > n - 1: end = n - 1
        if end <= i: continue
        max_ret = -1.0e18
        min_ret = 1.0e18
        for j in range(i + 1, end + 1):
            r = close_arr[j] / base - 1.0
            if r > max_ret: max_ret = r
            if r < min_ret: min_ret = r
        if max_ret >= dd_limit:
            safe_buy[i] = 1; evaluable[i] = 1
        if min_ret <= -ru_limit:
            safe_sell[i] = 1; evaluable[i] = 1
        # 둘 다 미달이면 evaluable 0 (평가 제외)
    return safe_buy, safe_sell, evaluable


def _compute_anchor_arrays(dates, anchor_buy_dates, anchor_sell_dates):
    n = len(dates)
    anchor_buy  = np.zeros(n, dtype=np.uint8)
    anchor_sell = np.zeros(n, dtype=np.uint8)

    def _to_norm_set(date_list):
        s = set()
        for d in date_list:
            try:
                ts = pd.Timestamp(d).normalize()
                s.add(ts)
            except Exception:
                pass
        return s

    buy_set  = _to_norm_set(anchor_buy_dates)
    sell_set = _to_norm_set(anchor_sell_dates)

    norm_dates = pd.DatetimeIndex(dates).normalize()
    for i in range(n - 1):
        nd = norm_dates[i + 1]
        if nd in buy_set:  anchor_buy[i]  = 1
        if nd in sell_set: anchor_sell[i] = 1

    n_buy_total  = len(buy_set)
    n_sell_total = len(sell_set)
    n_buy_matched  = int(anchor_buy.sum())
    n_sell_matched = int(anchor_sell.sum())
    if n_buy_total > 0 and n_buy_matched < n_buy_total:
        print(f"  ⚠ ANCHOR: 매수 정답일 {n_buy_total}개 중 {n_buy_matched}개만 데이터 범위 내")
    if n_sell_total > 0 and n_sell_matched < n_sell_total:
        print(f"  ⚠ ANCHOR: 매도 정답일 {n_sell_total}개 중 {n_sell_matched}개만 데이터 범위 내")
    return anchor_buy, anchor_sell


def _apply_anchor_correction(safe_buy, safe_sell, evaluable,
                              anchor_buy, anchor_sell):
    safe_buy_c  = np.maximum(safe_buy,  anchor_buy)
    safe_sell_c = np.maximum(safe_sell, anchor_sell)
    eval_c = np.maximum(evaluable, np.maximum(anchor_buy, anchor_sell))
    if len(eval_c) > 0:
        eval_c[-1] = 0
    return safe_buy_c, safe_sell_c, eval_c


@njit(cache=True)
def _simulate_ensemble(close_arr, buy_sig_mat, sell_sig_mat,
                        vote_buy, vote_sell, cost,
                        safe_buy, safe_sell, evaluable,
                        stop_loss_pct,
                        anchor_buy_arr, anchor_sell_arr,
                        buy_w, sell_w):
    n   = close_arr.shape[0]
    K_b = buy_sig_mat.shape[1]
    K_s = sell_sig_mat.shape[1]
    pos = 0; ent = 0.0; entidx = -1
    sum_daily = 0.0
    nt = 0; nw = 0; sr = 0.0; sqr = 0.0
    me = 1.0; md = 0.0
    n_eval = 0
    n_buy_correct = 0
    n_sell_correct = 0
    n_buy_tp = 0; n_buy_fp = 0; n_buy_fn = 0
    n_sell_tp = 0; n_sell_fp = 0; n_sell_fn = 0
    n_stop_triggered = 0
    n_anchor_buy_total = 0
    n_anchor_buy_matched = 0
    n_anchor_sell_total = 0
    n_anchor_sell_matched = 0
    use_anchor_buy  = anchor_buy_arr.shape[0]  == n
    use_anchor_sell = anchor_sell_arr.shape[0] == n
    for i in range(n):
        if pos == 1 and entidx >= 0 and i > entidx and i >= 1:
            if close_arr[i - 1] > 0.0:
                sum_daily += close_arr[i] / close_arr[i - 1] - 1.0

        stopped_today = False
        if pos == 1 and stop_loss_pct > 0.0 and ent > 0.0:
            cur_p = close_arr[i]
            if cur_p > 0.0 and (cur_p / ent - 1.0) <= -stop_loss_pct:
                ret = cur_p / ent - 1.0
                sum_daily -= cost
                nt += 1
                if ret > 0.0: nw += 1
                sr += ret; sqr += ret * ret
                pos = 0
                ent = 0.0
                entidx = -1
                n_stop_triggered += 1
                stopped_today = True

        cur_eq = 1.0 + sum_daily
        if cur_eq > me: me = cur_eq
        dd = cur_eq / me - 1.0
        if dd < md: md = dd

        if i >= n - 1:
            if use_anchor_buy and anchor_buy_arr[i] == 1:
                n_anchor_buy_total += 1
                if pos == 1: n_anchor_buy_matched += 1
            if use_anchor_sell and anchor_sell_arr[i] == 1:
                n_anchor_sell_total += 1
                if pos == 0: n_anchor_sell_matched += 1
            continue

        b_count = 0.0
        for k in range(K_b):
            if buy_sig_mat[i, k] == 1: b_count += buy_w[k]
        s_count = 0.0
        for k in range(K_s):
            if sell_sig_mat[i, k] == 1: s_count += sell_w[k]

        b_on = b_count >= vote_buy
        s_on = s_count >= vote_sell

        if evaluable[i] == 1:
            n_eval += 1
            sb = safe_buy[i] == 1
            ss = safe_sell[i] == 1
            if (b_on and sb) or ((not b_on) and (not sb)):
                n_buy_correct += 1
            if (s_on and ss) or ((not s_on) and (not ss)):
                n_sell_correct += 1
            if b_on and sb: n_buy_tp += 1
            elif b_on and (not sb): n_buy_fp += 1
            elif (not b_on) and sb: n_buy_fn += 1
            if s_on and ss: n_sell_tp += 1
            elif s_on and (not ss): n_sell_fp += 1
            elif (not s_on) and ss: n_sell_fn += 1

        if stopped_today: continue

        if K_b > 0:
            b_strength = b_count / K_b
        else:
            b_strength = 0.0
        if K_s > 0:
            s_strength = s_count / K_s
        else:
            s_strength = 0.0

        if pos == 0:
            if b_on:
                if s_on and s_strength > b_strength:
                    pass
                else:
                    p = close_arr[i + 1]
                    if p > 0.0:
                        ent = p; pos = 1; entidx = i + 1
                        sum_daily -= cost
        else:
            if s_on:
                if b_on and b_strength > s_strength:
                    pass
                else:
                    p = close_arr[i + 1]
                    if p > 0.0 and ent > 0.0:
                        ret = p / ent - 1.0
                        if close_arr[i] > 0.0:
                            sum_daily += close_arr[i + 1] / close_arr[i] - 1.0
                        sum_daily -= cost
                        nt += 1
                        if ret > 0.0: nw += 1
                        sr += ret; sqr += ret * ret
                    pos = 0
                    ent = 0.0
                    entidx = -1

        if use_anchor_buy and anchor_buy_arr[i] == 1:
            n_anchor_buy_total += 1
            if pos == 1: n_anchor_buy_matched += 1
        if use_anchor_sell and anchor_sell_arr[i] == 1:
            n_anchor_sell_total += 1
            if pos == 0: n_anchor_sell_matched += 1
    if pos == 1 and ent > 0.0:
        ret = close_arr[n - 1] / ent - 1.0
        sum_daily -= cost
        nt += 1
        if ret > 0.0: nw += 1
        sr += ret; sqr += ret * ret
        cur_eq = 1.0 + sum_daily
        if cur_eq > me: me = cur_eq
        dd = cur_eq / me - 1.0
        if dd < md: md = dd
    return (sum_daily, nt, nw, sr, sqr, md,
            n_eval, n_buy_correct, n_sell_correct, n_stop_triggered,
            n_buy_tp, n_buy_fp, n_buy_fn,
            n_sell_tp, n_sell_fp, n_sell_fn,
            n_anchor_buy_total, n_anchor_buy_matched,
            n_anchor_sell_total, n_anchor_sell_matched)


def prefilter_indicators(feat, close, *,
                          horizon=None,
                          min_corr=None,
                          min_variance_rel=None,
                          max_nan_ratio=None,
                          verbose=True):
    if horizon          is None: horizon          = HORIZON_DAYS
    if min_corr         is None: min_corr         = PREFILTER_MIN_CORR
    if min_variance_rel is None: min_variance_rel = PREFILTER_MIN_VARIANCE_REL
    if max_nan_ratio    is None: max_nan_ratio    = PREFILTER_MAX_NAN_RATIO

    future_ret = close.shift(-horizon) / close - 1
    future_ret = future_ret.dropna()

    n_orig = len(feat.columns)
    keep = []
    drop_nan = drop_var = drop_corr = 0
    n_days = len(feat)

    for col in feat.columns:
        x = feat[col]
        nan_ratio = x.isna().sum() / max(n_days, 1)
        if nan_ratio > max_nan_ratio:
            drop_nan += 1
            continue

        x_valid = x.dropna()
        if len(x_valid) < 50:
            drop_nan += 1
            continue

        std_v  = float(x_valid.std())
        mean_v = float(x_valid.mean())
        if std_v < 1e-12:
            drop_var += 1
            continue
        rel_var = std_v / max(abs(mean_v), 1e-10)
        if abs(mean_v) > 1e-10 and rel_var < min_variance_rel:
            drop_var += 1
            continue

        common = x_valid.index.intersection(future_ret.index)
        if len(common) < 50:
            keep.append(col)
            continue
        try:
            corr = float(x.loc[common].corr(future_ret.loc[common]))
        except Exception:
            corr = np.nan
        if pd.isna(corr) or abs(corr) < min_corr:
            drop_corr += 1
            continue

        keep.append(col)

    if verbose:
        n_kept = len(keep)
        pct_kept = n_kept / max(n_orig, 1) * 100
        print(f"  🔍 지표 사전 필터: {n_orig} → {n_kept}개 ({pct_kept:.1f}% 유지)")
        print(f"     제거 — NaN과다: {drop_nan}, 저분산: {drop_var}, 저상관(|r|<{min_corr}): {drop_corr}")
        if n_kept < 30:
            print(f"     ⚠ 유지 지표 너무 적음. PREFILTER_MIN_CORR를 낮추거나 PREFILTER_ENABLED=False 권장")

    return feat[keep]


def _select_indicators(feat, max_n=None):
    valid = []
    for col in feat.columns:
        x = feat[col].values.astype(np.float64)
        m = ~np.isnan(x)
        if m.sum() < 100: continue
        v = x[m]
        sd = float(np.std(v))
        if sd < 1e-12: continue
        cv = sd / max(abs(float(np.mean(v))), 1e-10)
        valid.append((col, cv))
    if max_n and len(valid) > max_n:
        valid.sort(key=lambda p: -p[1])
        valid = valid[:max_n]
    return [c for c, _ in valid]


def _rolling_zscore(x, window):
    """롤링 z-스코어 = (x - 이동평균) / 이동표준편차. 미래참조 없음(과거 window만)."""
    s = pd.Series(x)
    mu = s.rolling(window, min_periods=max(10, window//3)).mean()
    sd = s.rolling(window, min_periods=max(10, window//3)).std(ddof=0)
    z = ((s - mu) / sd.replace(0.0, np.nan)).values
    return z


# ══════════════════════════════════════════════════════════════════════
#  ★ 미래 예측 로직 개선 — 리드타임/스킬/홀드아웃 핵심 함수 (요청)
# ══════════════════════════════════════════════════════════════════════

@njit(cache=True)
def _fwd_hit_flags(close_arr, horizon, limit, is_buy, use_barrier):
    """각 날 i의 '정답 깃발' 사전계산 — hit[i]=1이면 그날 신호가 켜졌을 때 성공.
       - use_barrier=0: 기존 정의 그대로 — horizon일 이내 유리방향 최대변동 ≥ limit.
         (매수: 최고 종가 상승률 ≥ +limit / 매도: 최저 종가 하락률 ≤ -limit)
       - use_barrier=1: 삼중배리어 — 유리 한도가 '불리 한도보다 먼저' 도달해야 성공.
       ev[i]=1 은 평가가능(뒤에 볼 날이 있음). 마지막 날은 평가불가.
       ※ use_barrier=0 이면 _eval_buy_signals/_eval_sell_signals와 판정이 정확히 일치
         (검증 시트의 '성공률 독립 재계산' 교차검증에 사용)."""
    n = close_arr.shape[0]
    hit = np.zeros(n, dtype=np.uint8)
    ev  = np.zeros(n, dtype=np.uint8)
    h = horizon if horizon >= 1 else 1
    for i in range(n - 1):
        base = close_arr[i]
        if base <= 0.0:
            continue
        end = i + h
        if end > n - 1: end = n - 1
        if end <= i: continue
        ev[i] = 1
        if use_barrier == 1:
            for j in range(i + 1, end + 1):
                r = close_arr[j] / base - 1.0
                if is_buy == 1:
                    if r <= -limit: break          # 불리 배리어 먼저 → 실패
                    if r >= limit:
                        hit[i] = 1; break
                else:
                    if r >= limit: break
                    if r <= -limit:
                        hit[i] = 1; break
        else:
            best = -1.0e18
            for j in range(i + 1, end + 1):
                r = close_arr[j] / base - 1.0
                rr = r if is_buy == 1 else -r
                if rr > best: best = rr
            if best >= limit:
                hit[i] = 1
    return hit, ev


_HITF_CACHE = {}

def _hit_flags_cached(close_arr, horizon, limit, is_buy, use_barrier):
    """(_fwd_hit_flags 결과 캐시) — 내용 기반 키. 같은 종가·조건이면 재사용."""
    n = len(close_arr)
    key = (n,
           float(close_arr[0]) if n else None,
           float(close_arr[-1]) if n else None,
           int(horizon), round(float(limit), 8), int(is_buy), int(use_barrier))
    c = _HITF_CACHE.get(key)
    if c is not None:
        return c
    hit, ev = _fwd_hit_flags(close_arr, int(horizon), float(limit), int(is_buy), int(use_barrier))
    if len(_HITF_CACHE) > 600:
        _HITF_CACHE.clear()
    _HITF_CACHE[key] = (hit, ev)
    return hit, ev


def _shift_signal_forward(sig, d):
    """신호를 d일 늦춤(sig[t-d]가 t에 반영). 앞은 0 채움 — 미래참조 없음."""
    d = int(d)
    if d <= 0:
        return sig
    out = np.zeros_like(sig)
    out[d:] = sig[:-d]
    return out


def _success_on(hit, ev, sig, lo=0, hi=None):
    """구간 [lo,hi)에서의 (신호수, 성공수, 성공률). sig=None이면 무조건부(=기저확률)."""
    hi = len(hit) if hi is None else hi
    m = ev[lo:hi].astype(bool)
    if sig is not None:
        m = m & (np.asarray(sig[lo:hi]) == 1)
    nn = int(m.sum())
    ok = int(np.asarray(hit[lo:hi])[m].sum()) if nn else 0
    return nn, ok, (ok / nn if nn else float('nan'))


def _lead_profile(close_arr, sig, limit, is_buy, horizons, use_barrier, min_n=5):
    """리드타임 프로파일 — 각 h에 대해 (h, 신호수, 성공률, 기저확률, 스킬).
       스킬 = 성공률(h) − 기저확률(h): h가 길수록 성공률이 저절로 오르는 착시를 제거."""
    prof = []
    for h in horizons:
        hit, ev = _hit_flags_cached(close_arr, int(h), float(limit),
                                    1 if is_buy else 0, 1 if use_barrier else 0)
        nn, ok, sr = _success_on(hit, ev, sig)
        _, _, base = _success_on(hit, ev, None)
        sk = (sr - base) if (nn >= min_n and np.isfinite(sr) and np.isfinite(base)) else float('nan')
        prof.append({'h': int(h), 'n': nn, 'sr': sr, 'base': base, 'skill': sk})
    return prof


def _best_shift_for_side(sig_arr, hit0, ev0, *, split, min_sig, d_max, min_gain, ho_tol):
    """★ 신호 지연 d(0~d_max) 최적화 — 벡터화 (신호 인덱스만 이동, O(신호수×d)).
       규칙: 훈련구간 스킬(성공률−기저확률)이 d=0 대비 min_gain 이상 개선 &
             홀드아웃 스킬이 유지(허용 하락 ho_tol 이내)될 때만 채택. 반환: 최적 d."""
    idx = np.flatnonzero(np.asarray(sig_arr) == 1)
    if len(idx) == 0 or d_max <= 0:
        return 0
    n = len(sig_arr)
    ev_b = np.asarray(ev0).astype(bool); hit_a = np.asarray(hit0)
    # 세그먼트 기저확률 (무조건부)
    def _base(lo, hi):
        m = ev_b[lo:hi]
        return (float(hit_a[lo:hi][m].mean()) if m.sum() else np.nan)
    b_tr = _base(0, split); b_ho = _base(split, n)

    def _stats(d):
        j = idx + d
        j = j[(j < n)]
        j = j[ev_b[j]]
        jt = j[j < split]; jh = j[j >= split]
        n_tr = len(jt); n_ho = len(jh)
        sr_tr = float(hit_a[jt].mean()) if n_tr else np.nan
        sr_ho = float(hit_a[jh].mean()) if n_ho else np.nan
        return n_tr, sr_tr, n_ho, sr_ho

    n_tr0, sr_tr0, n_ho0, sr_ho0 = _stats(0)
    sk_tr0 = (sr_tr0 - b_tr) if (n_tr0 > 0 and np.isfinite(sr_tr0) and np.isfinite(b_tr)) else -1e18
    sk_ho0 = ((sr_ho0 - b_ho) if (n_ho0 >= 3 and np.isfinite(sr_ho0) and np.isfinite(b_ho)) else None)
    best_d = 0; best_sk = sk_tr0
    for d in range(1, int(d_max) + 1):
        n_tr, sr_tr, n_ho, sr_ho = _stats(d)
        if n_tr < min_sig or not np.isfinite(sr_tr) or not np.isfinite(b_tr):
            continue
        sk_tr = sr_tr - b_tr
        if sk_tr - sk_tr0 < min_gain:
            continue
        if sk_ho0 is not None and n_ho >= 3 and np.isfinite(sr_ho) and np.isfinite(b_ho):
            if (sr_ho - b_ho) < sk_ho0 - ho_tol:
                continue      # 홀드아웃에서 무너지는 지연은 미채택
        if sk_tr > best_sk + 1e-12:
            best_sk = sk_tr; best_d = d
    return best_d


def enrich_pool_with_lead_and_skill(feat, close, pool_df, is_buy, *, verbose=True):
    """★ 풀 후보 행에 미래예측 지표를 부여 + 필터 (요청):
       - best_lead / lead_profile : 지표별 최적 선행일 (스킬 최대 h)
       - lead_shift               : 채택된 신호 지연 d (훈련 스킬 개선 & 홀드아웃 유지 시만)
       - base_rate / skill / lift : 체결 지평(HORIZON_DAYS) 기준 기저확률·스킬·배율
       - sr_train / sr_holdout / skill_holdout : 앞(1-f)/뒤(f) 분할 성공률 (미래 유지력)
       필터: POOL_REQUIRE_SKILL(스킬<하한 제외), POOL_HOLDOUT_GUARD(홀드아웃 스킬<하한 제외).
       풀이 다 비면 필터 미적용 원본으로 폴백(거래 0 방지). 정렬은 성공률순 유지."""
    g = globals()
    if pool_df is None or len(pool_df) == 0:
        return pool_df
    do_lead   = bool(g.get('LEAD_TIME_SEARCH', False))
    do_shift  = bool(g.get('LEAD_SHIFT_ENABLED', False))
    req_skill = bool(g.get('POOL_REQUIRE_SKILL', False))
    guard     = bool(g.get('POOL_HOLDOUT_GUARD', False))
    if not (do_lead or do_shift or req_skill or guard):
        return pool_df
    close_arr = np.asarray(pd.Series(close).values, dtype=np.float64)
    n = len(close_arr)
    hz = max(1, int(g.get('HORIZON_DAYS', 1)))
    horizons = sorted(set(int(h) for h in (g.get('LEAD_HORIZONS') or [1, 2, 3, 5])) | {hz})
    use_bar = bool(g.get('LEAD_TRIPLE_BARRIER', False))
    f_hold = min(max(float(g.get('POOL_HOLDOUT_FRACTION', 0.30)), 0.0), 0.9)
    split = min(max(int(round(n * (1.0 - f_hold))), 1), n)
    min_skill    = float(g.get('POOL_MIN_SKILL', 0.0))
    ho_min_skill = float(g.get('POOL_HOLDOUT_MIN_SKILL', -0.10))
    shift_max  = max(0, int(g.get('LEAD_SHIFT_MAX', 5)))
    shift_gain = float(g.get('LEAD_SHIFT_MIN_GAIN', 0.05))
    ho_tol     = float(g.get('LEAD_SHIFT_HO_TOL', 0.02))
    min_sig    = int(g.get('POOL_SUCCESS_MIN_SIG', 10))
    default_limit = float(DRAWDOWN_LIMIT_BUY if is_buy else RUNUP_LIMIT_SELL)

    all_rows = []; kept = []
    n_drop_skill = 0; n_drop_hold = 0; n_shift = 0
    for _, row in pool_df.iterrows():
        rd = row.to_dict()
        try:
            _sl = rd.get('sel_limit', default_limit)
            limit = float(_sl) if _sl is not None and np.isfinite(float(_sl)) else default_limit
        except Exception:
            limit = default_limit
        rd['sel_limit'] = limit
        try:
            sig = _to_signal_array_raw(feat, rd)     # 지연 미적용 원신호
        except Exception:
            all_rows.append(rd); kept.append(rd); continue

        # ① 리드타임 프로파일 (삼중배리어 판정) — '몇 일 전 신호가 가장 정확한가'
        best_h = hz; best_sk = -1e18
        if do_lead:
            prof = _lead_profile(close_arr, sig, limit, is_buy, horizons, use_bar,
                                 min_n=max(3, min_sig // 2))
            for p in prof:
                if np.isfinite(p['skill']) and p['skill'] > best_sk + 1e-12:
                    best_sk = p['skill']; best_h = p['h']
            rd['best_lead'] = int(best_h)
            rd['lead_profile'] = ' | '.join(
                (f"{p['h']}일:{p['sr']*100:.0f}%"
                 f"({p['skill']*100:+.0f}p)" if np.isfinite(p['skill']) else f"{p['h']}일:—")
                for p in prof)

        # ② 체결 지평(hz) 기준 훈련/홀드아웃/전체 통계 (기존 성공 정의 그대로, 배리어 없음)
        #    선출 단계(LEAD_SELECT_IN_SCORING)에서 이미 지연 d0가 채택된 행은 d0 반영 신호가 기준.
        try:
            d0 = int(rd.get('lead_shift', 0) or 0)
        except Exception:
            d0 = 0
        d0 = max(0, min(d0, shift_max))
        sig_eff = _shift_signal_forward(sig, d0) if d0 else sig
        hit0, ev0 = _hit_flags_cached(close_arr, hz, limit, 1 if is_buy else 0, 0)
        def _seg(sg):
            n_tr, _, sr_tr = _success_on(hit0, ev0, sg, 0, split)
            _,   _, b_tr   = _success_on(hit0, ev0, None, 0, split)
            n_ho, _, sr_ho = _success_on(hit0, ev0, sg, split, n)
            _,   _, b_ho   = _success_on(hit0, ev0, None, split, n)
            n_all, ok_all, sr_all = _success_on(hit0, ev0, sg)
            _, _, b_all = _success_on(hit0, ev0, None)
            return dict(n_tr=n_tr, sr_tr=sr_tr, b_tr=b_tr,
                        n_ho=n_ho, sr_ho=sr_ho, b_ho=b_ho,
                        n=n_all, ok=ok_all, sr=sr_all, base=b_all)
        s0 = _seg(sig_eff)

        # ③ 신호 지연 정렬 — 선출 단계에서 이미 탐색했으면(d0 채택 여부와 무관) 재탐색 생략.
        #    선출 단계 탐색이 꺼진 경우에만 여기서 추가 지연을 탐색 (기준: 훈련 스킬 개선 & 홀드아웃 유지)
        d_best = d0; s_best = s0
        scoring_searched = bool(g.get('LEAD_SELECT_IN_SCORING', True)) and do_shift
        if do_shift and shift_max > d0 and not scoring_searched:
            d_cap = min(shift_max - d0, max(int(rd.get('best_lead', hz)) - 1 - d0, 0))
            sk_tr0 = (s0['sr_tr'] - s0['b_tr']) if np.isfinite(s0['sr_tr']) else -1e18
            sk_ho0 = (s0['sr_ho'] - s0['b_ho']) if (s0['n_ho'] >= 3 and np.isfinite(s0['sr_ho'])) else None
            cur_best_tr = sk_tr0
            for d in range(1, d_cap + 1):
                sd = _seg(_shift_signal_forward(sig, d0 + d))
                if sd['n_tr'] < min_sig or not np.isfinite(sd['sr_tr']):
                    continue
                sk_tr_d = sd['sr_tr'] - sd['b_tr']
                if sk_tr_d - sk_tr0 < shift_gain:
                    continue
                if sk_ho0 is not None and sd['n_ho'] >= 3 and np.isfinite(sd['sr_ho']):
                    if (sd['sr_ho'] - sd['b_ho']) < sk_ho0 - ho_tol:
                        continue            # 홀드아웃에서 무너지는 지연은 미채택
                if sk_tr_d > cur_best_tr + 1e-12:
                    cur_best_tr = sk_tr_d; d_best = d0 + d; s_best = sd
        rd['lead_shift'] = int(d_best)
        if d_best > 0:
            n_shift += 1
            if d_best > d0:                       # 여기서 추가 채택된 경우만 이전 성공률 기록
                rd['sr_preshift'] = rd.get('success_rate')
            rd['n_signals']   = int(s_best['n'])
            rd['n_success']   = int(s_best['ok'])
            if np.isfinite(s_best['sr']):
                rd['success_rate'] = float(s_best['sr'])

        st = s_best
        rd['base_rate']     = float(st['base']) if np.isfinite(st['base']) else np.nan
        rd['skill']         = (float(st['sr'] - st['base'])
                               if (np.isfinite(st['sr']) and np.isfinite(st['base'])) else np.nan)
        rd['lift']          = (float(st['sr'] / st['base'])
                               if (np.isfinite(st['sr']) and st['base'] and st['base'] > 0) else np.nan)
        rd['sr_train']      = float(st['sr_tr']) if np.isfinite(st['sr_tr']) else np.nan
        rd['sr_holdout']    = float(st['sr_ho']) if np.isfinite(st['sr_ho']) else np.nan
        rd['n_holdout']     = int(st['n_ho'])
        rd['skill_train']   = (float(st['sr_tr'] - st['b_tr'])
                               if (np.isfinite(st['sr_tr']) and np.isfinite(st['b_tr'])) else np.nan)
        rd['skill_holdout'] = (float(st['sr_ho'] - st['b_ho'])
                               if (st['n_ho'] >= 3 and np.isfinite(st['sr_ho']) and np.isfinite(st['b_ho']))
                               else np.nan)
        all_rows.append(rd)

        # ④ 필터 — 스킬 하한 / 홀드아웃 가드
        if req_skill and np.isfinite(rd['skill']) and rd['skill'] < min_skill:
            n_drop_skill += 1; continue
        if guard and np.isfinite(rd['skill_holdout']) and rd['skill_holdout'] < ho_min_skill:
            n_drop_hold += 1; continue
        kept.append(rd)

    out = pd.DataFrame(kept)
    fell_back = False
    k_floor = max(1, int(g.get('POOL_SUCCESS_K_FLOOR', 2)))
    if len(out) == 0 or ('indicator' in out.columns and out['indicator'].nunique() < k_floor):
        out = pd.DataFrame(all_rows)      # 폴백 — 거래 0 방지 (필터 미적용, 지표는 부여됨)
        fell_back = True
    if len(out) and 'success_rate' in out.columns:
        out = out.sort_values('success_rate', ascending=False, kind='mergesort').reset_index(drop=True)
    side = 'buy' if is_buy else 'sell'
    stats = g.setdefault('_ENRICH_STATS', {})
    stats[side] = dict(n_in=len(pool_df), n_out=len(out),
                       drop_skill=n_drop_skill, drop_holdout=n_drop_hold,
                       n_shift=n_shift, fell_back=fell_back, split=split, n_days=n)
    if verbose:
        _lb = '매수' if is_buy else '매도'
        print(f"    ↳ {_lb} 풀 예측력 보강: {len(pool_df)}행 → {len(out)}행 "
              f"(스킬미달 -{n_drop_skill} / 홀드아웃 -{n_drop_hold} / 지연채택 {n_shift}행"
              f"{' / ⚠폴백-필터해제' if fell_back else ''})")
    return out


def _stability_adjusted_score(close_arr, sig_arr, horizon, limit, anchor_arr,
                              wilson_z, is_buy, min_signals):
    """지표 신호의 점수 = Wilson 하한 × (1 + w·시간안정성).
       시간안정성: 신호일을 앞/뒤 절반으로 나눠 각각 성공률을 구하고,
         둘 다 좋고 서로 비슷할수록(미래에도 유지될 가능성↑) 가산점.
       반환: (n, ok, sum, score). 신호 부족 시 score=음수로 사실상 제외.
    """
    evalf = _eval_buy_signals if is_buy else _eval_sell_signals
    n_all, ok_all, sum_all = evalf(close_arr, sig_arr, horizon, limit, anchor_arr)
    if n_all < min_signals:
        return n_all, ok_all, sum_all, -1.0
    base = wilson_lower(ok_all, n_all, wilson_z)

    # ★ 큰 움직임 적중 가산 (요청) — 신호 중 '큰 상승/하락'을 맞춘 비율이 높을수록 점수↑.
    #   자잘한 적중만 많은 지표보다 굵직한 변동을 잡는 지표를 우대 → 실전 수익 직결.
    if globals().get('USE_BIG_MOVE_BONUS', False):
        big_thr = float(globals().get('BIG_MOVE_THRESHOLD', 0.03))
        bw      = float(globals().get('BIG_MOVE_BONUS_WEIGHT', 0.5))
        bn, bhit = _eval_big_move_hits(close_arr, sig_arr, horizon, big_thr,
                                       1 if is_buy else 0)
        if bn > 0:
            big_ratio = bhit / bn          # 0~1
            base = base * (1.0 + bw * big_ratio)

    if not globals().get('USE_OOS_STABILITY', False):
        return n_all, ok_all, sum_all, base
    # 앞/뒤 절반으로 신호 분할 (시간 순)
    nd = len(sig_arr); half = nd // 2
    sig_a = sig_arr.copy(); sig_a[half:] = 0     # 앞 절반만
    sig_b = sig_arr.copy(); sig_b[:half] = 0     # 뒤 절반만
    na, oka, _ = evalf(close_arr, sig_a, horizon, limit, anchor_arr)
    nb, okb, _ = evalf(close_arr, sig_b, horizon, limit, anchor_arr)
    if na < 3 or nb < 3:
        return n_all, ok_all, sum_all, base   # 한쪽 표본 부족 → 안정성 평가 불가, 기본점수
    ra = oka / na; rb = okb / nb
    # 안정성 = 두 기간 모두 0.5 넘고 + 서로 가까움. [0,1] 정규화.
    both_good = min(ra, rb)                       # 둘 중 나쁜 쪽 (낮으면 불안정)
    consistency = 1.0 - min(1.0, abs(ra - rb) / 0.5)  # 차이 작을수록 1
    stability = max(0.0, (both_good - 0.5) * 2.0) * consistency  # 0~1
    w = float(globals().get('OOS_STABILITY_WEIGHT', 0.3))
    score = base * (1.0 + w * stability)
    return n_all, ok_all, sum_all, score


def evaluate_buy_sell_scores(feat, close, *, indicators,
                              n_thresholds, pct_low, pct_high,
                              horizon, dd_limit, ru_limit,
                              min_signals, wilson_z,
                              anchor_buy_arr=None, anchor_sell_arr=None):
    close_arr = close.values.astype(np.float64)
    n_days    = len(close_arr)
    pcts      = np.linspace(pct_low, pct_high, n_thresholds)
    use_z     = globals().get('USE_ZSCORE_SIGNAL', False)
    z_window  = int(globals().get('ZSCORE_WINDOW', 60))
    z_thrs    = globals().get('ZSCORE_THRESHOLDS', [-2.0, 2.0])

    if anchor_buy_arr is None:
        anchor_buy_arr = np.zeros(0, dtype=np.uint8)
    if anchor_sell_arr is None:
        anchor_sell_arr = np.zeros(0, dtype=np.uint8)

    if HAS_NUMBA:
        zero_sig = np.zeros(n_days, dtype=np.uint8)
        _eval_buy_signals(close_arr, zero_sig, horizon, dd_limit, anchor_buy_arr)
        _eval_sell_signals(close_arr, zero_sig, horizon, ru_limit, anchor_sell_arr)

    # ★ (2b) 선출 단계 리드 탐색 준비 — '며칠 전 신호'가 정확한 지표는 다음날 성공률이
    #    낮아 기존 선출에서 탈락하므로, 선출 시 지연 d(0~MAX)를 함께 탐색한다.
    #    d 판정 기준은 보고 성공률과 동일(비배리어 hit 플래그) → 채택 d의 성공률이
    #    _stability_adjusted_score 결과와 정확히 일치. 훈련구간 스킬 + 홀드아웃 가드.
    lead_sel = bool(globals().get('LEAD_SELECT_IN_SCORING', True)) and \
               bool(globals().get('LEAD_SHIFT_ENABLED', True))
    if lead_sel:
        _d_max   = int(globals().get('LEAD_SHIFT_MAX', 5))
        _min_g   = float(globals().get('LEAD_SHIFT_MIN_GAIN', 0.05))
        _ho_tol  = float(globals().get('LEAD_SHIFT_HO_TOL', 0.02))
        _ho_frac = float(globals().get('POOL_HOLDOUT_FRACTION', 0.30))
        _split   = max(30, int(round(n_days * (1.0 - _ho_frac))))
        _bhit, _bev = _hit_flags_cached(close_arr, horizon, dd_limit, 1, False)
        _shit, _sev = _hit_flags_cached(close_arr, horizon, ru_limit, 0, False)

    buy_rows = []; sell_rows = []
    for col in indicators:
        x = feat[col].values.astype(np.float64)
        valid = ~np.isnan(x)
        if valid.sum() < 100: continue

        # ── 신호 후보 목록: (방향라벨, 임계라벨, 임계값, 신호배열) ──
        sig_specs = []
        # (a) 기존 절대 임계 (백분위 기반)
        for p in pcts:
            thr = float(np.nanpercentile(x, p))
            sig_specs.append(('>=', float(p), thr, ((x >= thr) & valid).astype(np.uint8)))
            sig_specs.append(('<=', float(p), thr, ((x <= thr) & valid).astype(np.uint8)))
        # (b) z-스코어 임계 (롤링 정규화) — 미래 일반화에 강함 (요청)
        if use_z:
            z = _rolling_zscore(x, z_window)
            zvalid = ~np.isnan(z)
            for zt in z_thrs:
                if zt >= 0:
                    s = ((z >= zt) & zvalid).astype(np.uint8); d = 'z>='
                else:
                    s = ((z <= zt) & zvalid).astype(np.uint8); d = 'z<='
                sig_specs.append((d, float(zt), float(zt), s))

        for direction, plabel, thr, sig_arr in sig_specs:
            if int(sig_arr.sum()) < min_signals: continue
            if lead_sel:
                db = _best_shift_for_side(sig_arr, _bhit, _bev, split=_split,
                                          min_sig=min_signals, d_max=_d_max,
                                          min_gain=_min_g, ho_tol=_ho_tol)
                ds = _best_shift_for_side(sig_arr, _shit, _sev, split=_split,
                                          min_sig=min_signals, d_max=_d_max,
                                          min_gain=_min_g, ho_tol=_ho_tol)
            else:
                db = ds = 0
            sig_b = _shift_signal_forward(sig_arr, db) if db else sig_arr
            bn, bok, bsum, bscore = _stability_adjusted_score(
                close_arr, sig_b, horizon, dd_limit, anchor_buy_arr,
                wilson_z, True, min_signals)
            if bn >= min_signals and bscore >= 0:
                buy_rows.append((col, direction, plabel, thr,
                                  bn, bok, bok / bn,
                                  float(bsum / bn) if bn else 0.0, bscore, int(db)))
            sig_s = _shift_signal_forward(sig_arr, ds) if ds else sig_arr
            sn, sok, ssum, sscore = _stability_adjusted_score(
                close_arr, sig_s, horizon, ru_limit, anchor_sell_arr,
                wilson_z, False, min_signals)
            if sn >= min_signals and sscore >= 0:
                sell_rows.append((col, direction, plabel, thr,
                                   sn, sok, sok / sn,
                                   float(ssum / sn) if sn else 0.0, sscore, int(ds)))

    cols = ['indicator', 'direction', 'pct_label', 'threshold',
            'n_signals', 'n_success', 'success_rate', 'avg_extreme', 'score',
            'lead_shift']
    buy_df  = pd.DataFrame(buy_rows,  columns=cols)
    sell_df = pd.DataFrame(sell_rows, columns=cols)
    if globals().get('POOL_SELECT_BY_SUCCESS', False):
        # ★ 성공률 우선(요청) + 절충(요청) — 메타조합(wilson_z)마다 풀이 달라지도록:
        #   ① 지표당 '최고 성공률' 1행만 남김(중복 임계 제거).
        #   ② 성공률 컷(POOL_SUCCESS_MIN_RATE) 통과 지표를 '우선' 묶음으로 (컷 유지).
        #   ③ 각 묶음 안 정렬을 'Wilson 하한(wilson_z 반영)'으로 → 표본 적은 고성공 지표는
        #      wilson_z가 높을수록 뒤로 밀려 top_n 멤버가 조합마다 달라짐.
        #   하드 컷오프는 두지 않음(풀이 비어 거래 0 되지 않게 best-available).
        def _succ_rank(df):
            if df is None or len(df) == 0:
                return (df.sort_values('score', ascending=False).reset_index(drop=True)
                        if df is not None else df)
            d = df.sort_values(['success_rate', 'score'], ascending=[False, False])
            d = d.drop_duplicates('indicator', keep='first').reset_index(drop=True)
            try:
                _z = float(wilson_z)
                p = d['success_rate'].values.astype(float)
                nn = np.maximum(d['n_signals'].values.astype(float), 1.0)
                denom  = 1.0 + _z * _z / nn
                center = p + _z * _z / (2.0 * nn)
                margin = _z * np.sqrt(np.clip(p * (1.0 - p) / nn + _z * _z / (4.0 * nn * nn), 0.0, None))
                _cut = float(globals().get('POOL_SUCCESS_MIN_RATE', 0.6))
                d = d.assign(_wlb=(center - margin) / denom,
                             _above=(d['success_rate'].values >= _cut).astype(int))
                d = d.sort_values(['_above', '_wlb', 'success_rate'],
                                  ascending=[False, False, False]).reset_index(drop=True)
                d = d.drop(columns=['_wlb', '_above'])
            except Exception:
                pass
            return d
        buy_df  = _succ_rank(buy_df)
        sell_df = _succ_rank(sell_df)
    else:
        buy_df  = buy_df.sort_values('score', ascending=False).reset_index(drop=True)
        sell_df = sell_df.sort_values('score', ascending=False).reset_index(drop=True)
    return buy_df, sell_df


# ══════════════════════════════════════════════════════════════════════
#  ★ 통계적 지표 선정 도구 (A+B+C) — 성공률 대신 IC/p-value/OOS IC 사용
# ══════════════════════════════════════════════════════════════════════

def _spearman_ic(x, y):
    """Spearman rank correlation (순위 상관). NaN 안전. |x| < 5면 NaN 반환."""
    x = np.asarray(x, float); y = np.asarray(y, float)
    m = np.isfinite(x) & np.isfinite(y)
    if m.sum() < 5:
        return np.nan
    xr = pd.Series(x[m]).rank().values
    yr = pd.Series(y[m]).rank().values
    if xr.std() < 1e-9 or yr.std() < 1e-9:
        return 0.0
    return float(np.corrcoef(xr, yr)[0, 1])


def _ic_pvalue(ic, n):
    """Spearman IC의 양측 p-value (t-분포 근사). n=표본수."""
    if not np.isfinite(ic) or n < 5:
        return 1.0
    from scipy import stats
    t = ic * np.sqrt(max(n - 2, 1) / max(1 - ic * ic, 1e-9))
    return float(2 * (1 - stats.t.cdf(abs(t), df=max(n - 2, 1))))


def _bh_fdr(pvals, alpha=0.05):
    """Benjamini-Hochberg FDR 보정. p-value 배열 → 통과 여부 배열."""
    p = np.asarray(pvals, float); n = len(p)
    order = np.argsort(p); ranked = p[order]
    thresh = alpha * (np.arange(1, n + 1) / n)
    passed_sorted = ranked <= thresh
    if not passed_sorted.any():
        return np.zeros(n, bool)
    kmax = np.max(np.where(passed_sorted)[0])
    out = np.zeros(n, bool); out[order[:kmax + 1]] = True
    return out


def _purged_kfold_ic(sig_vals, fwd_ret, k=5, embargo=5):
    """Purged K-fold OOS IC (Marcos López de Prado 방식).
       훈련/검증 사이 embargo 기간을 두어 정보 누출 방지.
       반환: (mean_oos_ic, std_oos_ic, ir)"""
    n = len(sig_vals)
    if n < k * 20:
        return np.nan, np.nan, np.nan
    fold_size = n // k
    ics = []
    for fi in range(k):
        v_start = fi * fold_size
        v_end = min((fi + 1) * fold_size, n)
        # 검증 구간 + 앞뒤 embargo 제외한 훈련
        mask = np.ones(n, bool)
        mask[max(0, v_start - embargo):min(n, v_end + embargo)] = False
        # 검증 IC (훈련구간 정보는 어차피 이 계산에 안 쓰임 — 순수 통계만 측정)
        v = _spearman_ic(sig_vals[v_start:v_end], fwd_ret[v_start:v_end])
        if np.isfinite(v):
            ics.append(v)
    if len(ics) < 2:
        return np.nan, np.nan, np.nan
    m = float(np.mean(ics)); s = float(np.std(ics))
    ir = m / (s + 1e-9)
    return m, s, ir


def compute_indicator_stats(feat, close, *, indicators=None, horizon=1,
                            fdr_alpha=0.05, k_folds=5, embargo=5):
    """★ 각 지표의 IC / p-value / FDR통과 / OOS IC / IR 을 계산.
       - IC: Spearman(지표값 오늘, 수익률 다음날)
       - p-value: Spearman 유의성 (t-분포 근사)
       - FDR: Benjamini-Hochberg 다중검정 보정
       - OOS IC/IR: Purged K-fold (embargo=5일)
       반환: DataFrame(indicator, ic, pvalue, fdr_pass, oos_ic, oos_ir, n_obs)"""
    r = close.pct_change(horizon).shift(-horizon).values   # 다음 h일 수익률
    if indicators is None:
        indicators = list(feat.columns)
    rows = []
    for ind in indicators:
        x = feat[ind].values
        m = np.isfinite(x) & np.isfinite(r)
        n_obs = int(m.sum())
        if n_obs < 30:
            rows.append({'indicator': ind, 'ic': np.nan, 'pvalue': 1.0,
                         'oos_ic': np.nan, 'oos_ir': np.nan, 'n_obs': n_obs})
            continue
        ic = _spearman_ic(x[m], r[m])
        pv = _ic_pvalue(ic, n_obs)
        oos_ic, _, oos_ir = _purged_kfold_ic(x[m], r[m], k=k_folds, embargo=embargo)
        rows.append({'indicator': ind, 'ic': ic, 'pvalue': pv,
                     'oos_ic': oos_ic, 'oos_ir': oos_ir, 'n_obs': n_obs})
    df = pd.DataFrame(rows)
    # FDR 보정
    valid = df['pvalue'].notna() & (df['pvalue'] <= 1)
    passed = np.zeros(len(df), bool)
    if valid.sum() > 0:
        passed[valid.values] = _bh_fdr(df.loc[valid, 'pvalue'].values, alpha=fdr_alpha)
    df['fdr_pass'] = passed
    return df


def select_pool_by_ic(feat, close, *, indicators=None, horizon=1,
                      ic_min=0.03, oos_ic_min=0.01, oos_ir_min=0.3,
                      fdr_alpha=0.05, k_folds=5, embargo=5,
                      max_pool=100, direction='auto'):
    """★ IC 기반 지표 선정 (A+B+C 통합).
       조건: (|IC| ≥ ic_min) AND (FDR 통과) AND (|OOS IC| ≥ oos_ic_min) AND (|OOS IR| ≥ oos_ir_min)
       direction='buy'→IC 양수만(오르는 신호), 'sell'→IC 음수만(내리는 신호), 'auto'→절댓값.
       반환: 선정 지표 DataFrame (indicator, ic, pvalue, oos_ic, oos_ir, direction)."""
    stats = compute_indicator_stats(feat, close, indicators=indicators, horizon=horizon,
                                    fdr_alpha=fdr_alpha, k_folds=k_folds, embargo=embargo)
    df = stats.dropna(subset=['ic', 'oos_ic']).copy()
    if direction == 'buy':
        df = df[(df['ic'] >= ic_min) & (df['oos_ic'] >= oos_ic_min) & (df['oos_ir'] >= oos_ir_min)]
    elif direction == 'sell':
        df = df[(df['ic'] <= -ic_min) & (df['oos_ic'] <= -oos_ic_min) & (df['oos_ir'] <= -oos_ir_min)]
    else:
        df = df[(df['ic'].abs() >= ic_min) & (df['oos_ic'].abs() >= oos_ic_min) & (df['oos_ir'].abs() >= oos_ir_min)]
    df = df[df['fdr_pass']]
    if direction == 'buy':
        df = df.sort_values('oos_ic', ascending=False)
    elif direction == 'sell':
        df = df.sort_values('oos_ic', ascending=True)
    else:
        df = df.reindex(df['oos_ic'].abs().sort_values(ascending=False).index)
    return df.head(max_pool).reset_index(drop=True)


def select_pool_by_success(feat, close, *, indicators, n_thresholds,
                           horizon, dd_limit, ru_limit, wilson_z=1.0):
    """★ 성공률 우선 풀 선출 (요청).
       - 넓은 분위(POOL_SUCCESS_WIDE_PCT) + z-스코어 후보 전체로 모든 지표를 평가 →
         pct(분위)가 달라 따로 나오던 고성공 지표를 누락 없이 한 번에 평가(z스코어는 pct무관).
       - 표본 가드: 신호수 >= POOL_SUCCESS_MIN_SIG (가짜 100% 방지).
       - 성공률 >= POOL_SUCCESS_MIN_RATE 만 후보로.
       - 성공률 내림차순 정렬(동률은 점수=Wilson 하한으로 타이브레이크).
       - 지표당 1행(최고 성공률)만 남긴 dedup 버전도 함께 반환(시트/표시용).
       반환: (buy_full, sell_full, buy_dedup, sell_dedup)
         *_full   : 성공률순 정렬된 전체 후보행 (diversify_candidates에 그대로 투입 → 풀)
         *_dedup  : 지표당 최고성공 1행 (성공률 우선 선출 시트용)
    """
    lo, hi = POOL_SUCCESS_WIDE_PCT
    bdf, sdf = evaluate_buy_sell_scores(
        feat, close, indicators=indicators, n_thresholds=n_thresholds,
        pct_low=lo, pct_high=hi, horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
        min_signals=POOL_SUCCESS_MIN_SIG, wilson_z=wilson_z,
        anchor_buy_arr=None, anchor_sell_arr=None,   # 오버라이드 제거됨 → 앵커 무관(순수 다음날 ±1%)
    )

    def _filt(df):
        if df is None or len(df) == 0:
            return df.copy() if df is not None else df
        d = df[(df['n_signals'] >= POOL_SUCCESS_MIN_SIG) &
               (df['success_rate'] >= POOL_SUCCESS_MIN_RATE)].copy()
        # 성공률 우선, 동률이면 점수(Wilson 하한)로 — 가짜 100% 강등
        d = d.sort_values(['success_rate', 'score'], ascending=[False, False]).reset_index(drop=True)
        return d

    buy_full  = _filt(bdf)
    sell_full = _filt(sdf)
    # 지표당 1행(성공률 최고) — 이미 성공률순 정렬이라 first가 최고
    buy_dedup  = (buy_full.drop_duplicates('indicator', keep='first').reset_index(drop=True)
                  if buy_full is not None and len(buy_full) else buy_full)
    sell_dedup = (sell_full.drop_duplicates('indicator', keep='first').reset_index(drop=True)
                  if sell_full is not None and len(sell_full) else sell_full)
    return buy_full, sell_full, buy_dedup, sell_dedup


def _diversify_keep_thresholds(feat, score_df, *, top_n, corr_limit):
    """상관 다변화 — 단, '같은 지표의 여러 임계'는 유지(다중임계), 서로 다른 상관 지표만 제거.
       score_df는 success_rate 내림차순 가정. 유니크 지표 top_n개까지."""
    if score_df is None or len(score_df) == 0:
        return score_df
    kept = []; seen_sig = {}; seen_order = []; rejected = set()
    for _, row in score_df.iterrows():
        ind = row['indicator']
        if ind in seen_sig:              # 같은 지표 다른 임계 → 유지 (다중임계)
            kept.append(row.to_dict()); continue
        if ind in rejected:              # 이미 상관중복으로 버린 지표 → 스킵
            continue
        if len(seen_order) >= top_n:     # 유니크 지표 수 상한
            rejected.add(ind); continue
        try:
            sig = _to_signal_array(feat, row).astype(np.float64)
        except Exception:
            continue
        redundant = False
        for pind in seen_order:
            ps = seen_sig[pind]
            mask = ~np.isnan(sig) & ~np.isnan(ps)
            if mask.sum() < 10: continue
            a = sig[mask]; b = ps[mask]
            if a.std() < 1e-9 or b.std() < 1e-9: continue
            if abs(float(np.corrcoef(a, b)[0, 1])) >= corr_limit:
                redundant = True; break
        if redundant:
            rejected.add(ind); continue
        kept.append(row.to_dict()); seen_sig[ind] = sig; seen_order.append(ind)
    return pd.DataFrame(kept)


def _compute_avg_adverse_for_pool(feat, close, pool_df, is_buy, horizon=1):
    """[SEL-2] 풀의 각 (지표,임계) 행에 avg_adverse 열 추가.
       매수: 신호 후 horizon 내 최저하락률 평균 (음수), 매도: 최고상승률 평균 (양수)."""
    if pool_df is None or len(pool_df) == 0:
        return pool_df
    close_arr = np.asarray(close.values, dtype=np.float64)
    out = pool_df.copy()
    advs = []
    for _, row in out.iterrows():
        try:
            sig = _to_signal_array(feat, row).astype(np.uint8)
            n_, adv = _eval_avg_adverse(close_arr, sig, int(horizon), 1 if is_buy else 0)
            advs.append(float(adv) if n_ > 0 else 0.0)
        except Exception:
            advs.append(0.0)
    out['avg_adverse'] = advs
    return out


def _rank_pool_by_selection(pool_df, is_buy, *, dd_limit=0.01, verbose=False):
    """★ [SEL-1/2/4] 풀을 개선된 정렬 기준으로 재정렬.
       POOL_RANK_BY: 'success' | 'expected' | 'wilson' | 'skill'
       USE_ADVERSE_PENALTY: True면 신호 후 불리방향 크기로 감점
       USE_HOLDOUT_DECAY_PENALTY: True면 훈련→홀드아웃 스킬 감쇠로 감점
       원본 success_rate 컬럼은 보존, 'rank_score' 컬럼을 추가해 그 값으로 정렬.
       enrichment가 부여한 skill/skill_holdout/base_rate/avg_adverse가 없으면 성공률로 폴백."""
    g = globals()
    if pool_df is None or len(pool_df) == 0:
        return pool_df
    df = pool_df.copy()
    rank_by = str(g.get('POOL_RANK_BY', 'success')).lower()

    # 기본 점수 계산
    sr = df['success_rate'].astype(float)
    n_sig = df['n_signals'].astype(float).replace(0, 1)
    if rank_by == 'expected':
        # 기대수익 = 성공률 × 평균 유리방향 (avg_extreme는 항상 존재)
        base = sr * df.get('avg_extreme', sr * 0.02).astype(float).abs()
    elif rank_by == 'wilson':
        # Wilson 하한(z=1.96 근사) — 표본 크기 반영
        z = 1.96
        p = sr; n = n_sig
        base = (p + z*z/(2*n) - z*np.sqrt((p*(1-p)+z*z/(4*n))/n)) / (1 + z*z/n)
    elif rank_by == 'skill':
        # 스킬 = 성공률 - 기저확률 (enrichment 있어야, 없으면 성공률로 폴백)
        if 'base_rate' in df.columns:
            base = sr - df['base_rate'].astype(float).fillna(0)
        else:
            base = sr
    else:  # 'success'
        base = sr

    # [SEL-2] 불리방향 감점
    if g.get('USE_ADVERSE_PENALTY', False) and 'avg_adverse' in df.columns:
        w = float(g.get('ADVERSE_PENALTY_WEIGHT', 0.5))
        # |avg_adverse| / dd_limit — 1.0이면 손실이 이익 목표와 같은 크기
        adv_ratio = df['avg_adverse'].astype(float).abs() / max(float(dd_limit), 1e-6)
        penalty = np.clip(1.0 - w * adv_ratio, 0.5, 1.0)
        base = base * penalty

    # [SEL-4] 홀드아웃 감쇠 감점
    if g.get('USE_HOLDOUT_DECAY_PENALTY', False) and 'skill' in df.columns and 'skill_holdout' in df.columns:
        w = float(g.get('HOLDOUT_DECAY_WEIGHT', 1.0))
        decay = (df['skill'].astype(float).fillna(0) - df['skill_holdout'].astype(float).fillna(0)).clip(lower=0)
        penalty = np.clip(1.0 - w * decay, 0.5, 1.0)
        base = base * penalty

    df['rank_score'] = base.astype(float)
    df = df.sort_values(['rank_score', 'success_rate', 'n_signals'],
                        ascending=[False, False, False]).reset_index(drop=True)
    if verbose:
        print(f"    · 풀 재정렬: 기준={rank_by}"
              f"{' +불리감점' if g.get('USE_ADVERSE_PENALTY') else ''}"
              f"{' +홀드감쇠감점' if g.get('USE_HOLDOUT_DECAY_PENALTY') else ''} "
              f"({len(df)}행)")
    return df


def _limit_thresholds_per_indicator(pool_df, max_per=None, verbose=False):
    """★ [SEL-3] 지표당 최대 임계 수 제한. 성공률 기준 상위 K개만 유지.
       max_per=0 또는 None이면 무제한 (기존 동작)."""
    g = globals()
    k = int(max_per if max_per is not None else g.get('MAX_THRESHOLDS_PER_INDICATOR', 0) or 0)
    if pool_df is None or len(pool_df) == 0 or k <= 0:
        return pool_df
    n_before = len(pool_df)
    df = pool_df.sort_values('success_rate', ascending=False)
    df = df.groupby('indicator', as_index=False, sort=False).head(k).reset_index(drop=True)
    if verbose:
        print(f"    · 지표당 임계 상위 {k}개 유지: {n_before}행 → {len(df)}행")
    return df



def select_pool_combined(feat, close, *, indicators, n_thresholds, horizon, wilson_z=1.0, corr_limit=None):
    """★ 요청: 1~5% 각 한도로 성공풀 선출 → 하나로 합친 '다중임계' 풀.
       각 한도의 full 풀(지표당 여러 임계)을 concat → (indicator,threshold) 중복은 최고 success_rate 1행
       → 상관 다변화(같은 지표 여러 임계 유지, 다른 상관 지표만 제거)로 수정전처럼 정예화.
       반환 (buy_combined, sell_combined)."""
    limits = list(globals().get('STAGE_SUCCESS_LIMIT', [DRAWDOWN_LIMIT_BUY]))
    bparts, sparts = [], []
    for L in limits:
        try:
            bf, sf, _, _ = select_pool_by_success(
                feat, close, indicators=indicators, n_thresholds=n_thresholds,
                horizon=horizon, dd_limit=L, ru_limit=L, wilson_z=wilson_z)
        except Exception:
            continue
        if bf is not None and len(bf):
            bf = bf.copy(); bf['sel_limit'] = L; bparts.append(bf)
        if sf is not None and len(sf):
            sf = sf.copy(); sf['sel_limit'] = L; sparts.append(sf)
    def _comb(parts):
        if not parts:
            return None
        allp = pd.concat(parts, ignore_index=True)
        allp = allp.sort_values('success_rate', ascending=False)   # 수정전: raw 성공률 정렬
        allp = allp.drop_duplicates(['indicator', 'threshold'], keep='first').reset_index(drop=True)
        return allp
    buy_all = _comb(bparts); sell_all = _comb(sparts)
    # ★ 상관 다변화 (수정전과 동일 기준) — 고성공 지표는 다 남기되 중복 상관만 제거
    _tn = int(globals().get('TOP_N_POOL', globals().get('MAX_POOL', 100)) or 100)
    _cl = float(corr_limit if corr_limit is not None else (globals().get('STAGE_CORR_LIMIT') or [0.2])[0])
    buy_c = _diversify_keep_thresholds(feat, buy_all, top_n=_tn, corr_limit=_cl)
    sell_c = _diversify_keep_thresholds(feat, sell_all, top_n=_tn, corr_limit=_cl)
    # ★ 미래 예측력 보강 (요청) — 리드타임 탐색·신호 지연 정렬·스킬 필터·홀드아웃 가드.
    #   net>K/KL이 실제로 쓰는 풀에만 적용 (그리드-투표 경로는 기존 그대로).
    try:
        buy_c  = enrich_pool_with_lead_and_skill(feat, close, buy_c,  True)
        sell_c = enrich_pool_with_lead_and_skill(feat, close, sell_c, False)
    except Exception as _ee:
        print(f"    ⚠ 풀 예측력 보강 실패(원본 풀 유지): {_ee}")

    # ★ [SEL-2] 매수/매도 avg_adverse 부여를 여기서(A/B 스냅샷 이전) 해서 baseline·모든 변형에서 재사용
    try:
        _hz = int(globals().get('HORIZON_DAYS', 1))
        buy_c  = _compute_avg_adverse_for_pool(feat, close, buy_c,  True,  _hz)
        sell_c = _compute_avg_adverse_for_pool(feat, close, sell_c, False, _hz)
    except Exception:
        pass

    # ★ A/B 스냅샷: 재정렬/제한 前 상태 저장 (A/B 검증이 이 시점의 풀을 기준으로 각 변형을 비교)
    try:
        globals()['_KNET_PRERANK_POOL'] = (buy_c.copy() if buy_c is not None else None,
                                            sell_c.copy() if sell_c is not None else None)
    except Exception:
        pass

    # ★ 지표 선출 개선 (요청) — enrichment 뒤에 신규 정렬/필터 적용.
    _sel_verbose = bool(globals().get('SELECTION_VERBOSE', True))
    _limit = float((globals().get('STAGE_SUCCESS_LIMIT') or [0.01])[0])
    try:
        # [SEL-3] 지표당 임계 개수 제한
        buy_c  = _limit_thresholds_per_indicator(buy_c,  verbose=_sel_verbose)
        sell_c = _limit_thresholds_per_indicator(sell_c, verbose=_sel_verbose)
        # [SEL-1/2/4] 최종 정렬
        buy_c  = _rank_pool_by_selection(buy_c,  True,  dd_limit=_limit, verbose=_sel_verbose)
        sell_c = _rank_pool_by_selection(sell_c, False, dd_limit=_limit, verbose=_sel_verbose)
    except Exception as _rke:
        print(f"    ⚠ 지표 선출 개선 실패(원본 풀 유지): {_rke}")
    return buy_c, sell_c


def _build_pool_by_success(feat, close, *, indicators, n_thresholds, horizon, ticker):
    """기존 방식: 1~5% 통합 다중임계 풀 + (wilson×corr) 순차 탐색으로 k순신호 전체수익 최고 조합 선택."""
    _limits = list(globals().get('STAGE_SUCCESS_LIMIT', [DRAWDOWN_LIMIT_BUY]))
    _wzs = list(globals().get('STAGE_WILSON_Z') or [1.95])
    _cls = list(globals().get('STAGE_CORR_LIMIT') or [0.2])
    print(f"\n  ── 1~5% 통합 풀 선출 + (wilson 순차→corr) 'k순신호 전체수익 최고' 선택 "
          f"— 한도 {[f'{x*100:.0f}%' for x in _limits]} 통합 ──")
    _score_cache = {}
    def _score(_wz, _cl):
        _ck = (round(float(_wz), 4), round(float(_cl), 4))
        if _ck in _score_cache: return _score_cache[_ck]
        _cb, _cs = select_pool_combined(feat, close, indicators=indicators,
                                        n_thresholds=n_thresholds, horizon=horizon,
                                        wilson_z=_wz, corr_limit=_cl)
        if _cb is None or _cs is None or len(_cb) == 0 or len(_cs) == 0:
            _score_cache[_ck] = (None, None, -1e18); return _score_cache[_ck]
        _nsd = _net_signal_k_search(feat, close, _cb, _cs, ticker=ticker,
                                    oos_start=None, search_counts=True)
        _score_cache[_ck] = (_cb, _cs, (_nsd['full_cum'] if _nsd else -1e18))
        return _score_cache[_ck]
    _cl0 = _cls[0]; _best_wz = _wzs[0]; _wz_sc = -1e18
    for _wz in _wzs:
        _cb, _cs, _sc = _score(_wz, _cl0)
        _np = _cb['indicator'].nunique() if _cb is not None else 0
        _ns = _cs['indicator'].nunique() if _cs is not None else 0
        print(f"    [wilson={_wz}, corr={_cl0}] 전체수익 {_sc*100:+8.2f}% | 매수{_np}/매도{_ns}"
              f"{'  ★최고' if _sc > _wz_sc else ''}")
        if _sc > _wz_sc: _wz_sc = _sc; _best_wz = _wz
    print(f"    → 최고 wilson = {_best_wz} (전체수익 {_wz_sc*100:+.2f}%)")
    _best = None; _best_sc = -1e18; _bcl = _cls[0]
    _shown = {(round(float(_best_wz), 4), round(float(_cl0), 4))}
    for _cl in _cls:
        _cb, _cs, _sc = _score(_best_wz, _cl)
        _np = _cb['indicator'].nunique() if _cb is not None else 0
        _ns = _cs['indicator'].nunique() if _cs is not None else 0
        if (round(float(_best_wz), 4), round(float(_cl), 4)) not in _shown:
            print(f"    [wilson={_best_wz}, corr={_cl}] 전체수익 {_sc*100:+8.2f}% | 매수{_np}/매도{_ns}"
                  f"{'  ★최고' if _sc > _best_sc else ''}")
            _shown.add((round(float(_best_wz), 4), round(float(_cl), 4)))
        if _sc > _best_sc: _best_sc = _sc; _best = (_cb, _cs); _bcl = _cl
    if _best is not None:
        globals()['_KNET_MULTI_POOL'] = (ticker, _best[0], _best[1])
        print(f"  ★ k순신호 최적: wilson={_best_wz}, corr={_bcl} → 전체수익 {_best_sc*100:+.2f}% "
              f"(매수 {_best[0]['indicator'].nunique()}지표 / 매도 {_best[1]['indicator'].nunique()}지표, 다중임계)")
    else:
        globals()['_KNET_MULTI_POOL'] = (ticker, None, None)
        print(f"  ⚠ 합친 풀 생성 실패")
    return globals()['_KNET_MULTI_POOL']


def _build_and_pick_knet_pool(feat, close, *, indicators, n_thresholds, horizon, ticker):
    """★ 지표 풀 선정. USE_IC_SELECTION 스위치로 두 방식 분기.
       - False (기본): 기존 성공률+wilson×corr 순차 탐색 (안전, 검증됨)
       - True  (실험): IC/OOS IC/OOS IR/FDR 기반 통계적 선정 (A+B+C)"""
    globals().pop('_KNET_REPLAY_FIXED', None)
    globals().pop('_KNET_KL_FIXED', None)

    if not bool(globals().get('USE_IC_SELECTION', False)):
        # ── 기존 방식 (성공률 + wilson×corr 순차) ──
        return _build_pool_by_success(feat, close, indicators=indicators,
                                      n_thresholds=n_thresholds, horizon=horizon, ticker=ticker)

    # ── IC 기반 방식 (A+B+C) ──
    _ic_min = float(globals().get('IC_MIN', 0.05))
    _oos_ic_min = float(globals().get('OOS_IC_MIN', 0.015))
    _oos_ir_min = float(globals().get('OOS_IR_MIN', 0.3))
    _fdr_a = float(globals().get('FDR_ALPHA', 0.05))
    _kf = int(globals().get('PURGED_K_FOLDS', 5))
    _emb = int(globals().get('PURGED_EMBARGO', 5))
    _maxp = int(globals().get('TOP_N_POOL', 100))

    print(f"\n  ── IC 기반 지표 선정 (A+B+C) — |IC|≥{_ic_min}, |OOS IC|≥{_oos_ic_min}, "
          f"|OOS IR|≥{_oos_ir_min}, FDR α={_fdr_a}, k={_kf}, embargo={_emb} ──")
    _stats = compute_indicator_stats(feat, close, indicators=indicators, horizon=horizon,
                                     fdr_alpha=_fdr_a, k_folds=_kf, embargo=_emb)

    def _pick(direction):
        df = _stats.dropna(subset=['ic', 'oos_ic']).copy()
        if direction == 'buy':
            df = df[(df['ic'] >= _ic_min) & (df['oos_ic'] >= _oos_ic_min) & (df['oos_ir'] >= _oos_ir_min)]
        else:
            df = df[(df['ic'] <= -_ic_min) & (df['oos_ic'] <= -_oos_ic_min) & (df['oos_ir'] <= -_oos_ir_min)]
        if bool(globals().get('FDR_ENABLED', True)):
            df = df[df['fdr_pass']]
        if direction == 'buy':
            df = df.sort_values('oos_ic', ascending=False).head(_maxp).reset_index(drop=True)
        else:
            df = df.sort_values('oos_ic', ascending=True).head(_maxp).reset_index(drop=True)
        # 기존 시스템 호환 형식 (indicator, direction, threshold, success_rate=|OOS IC|)
        if len(df) == 0:
            return None
        pool_rows = []
        for _, r in df.iterrows():
            med = float(np.nanmedian(feat[r['indicator']].values))
            pool_rows.append({
                'indicator': r['indicator'],
                'direction': ('>=' if direction == 'buy' else '<='),
                'threshold': med,
                'success_rate': float(abs(r['oos_ic'])),  # B: OOS IC를 가중치로
                'ic': float(r['ic']), 'pvalue': float(r['pvalue']),
                'oos_ic': float(r['oos_ic']), 'oos_ir': float(r['oos_ir']),
                'n_signals': int(r['n_obs']),
            })
        return pd.DataFrame(pool_rows)

    _cb = _pick('buy'); _cs = _pick('sell')
    _nb = 0 if _cb is None else len(_cb)
    _ns = 0 if _cs is None else len(_cs)
    print(f"  ★ IC 선정 결과: 매수 {_nb}지표 / 매도 {_ns}지표 (전체 {len(_stats)}개 중)")
    if _cb is not None and len(_cb) > 0:
        print(f"    [매수 상위 5] " + " | ".join(
            f"{r['indicator'][:18]}(IC={r['ic']:+.3f},OOS={r['oos_ic']:+.3f})"
            for _, r in _cb.head(5).iterrows()))
    if _cs is not None and len(_cs) > 0:
        print(f"    [매도 상위 5] " + " | ".join(
            f"{r['indicator'][:18]}(IC={r['ic']:+.3f},OOS={r['oos_ic']:+.3f})"
            for _, r in _cs.head(5).iterrows()))

    if (_cb is None or len(_cb) == 0) or (_cs is None or len(_cs) == 0):
        print(f"  ⚠ IC 기준 통과 지표 부족 → 성공률 기반 폴백")
        _cb2, _cs2 = select_pool_combined(feat, close, indicators=indicators,
                                          n_thresholds=n_thresholds, horizon=horizon,
                                          wilson_z=1.95, corr_limit=0.2)
        if _cb is None or len(_cb) == 0: _cb = _cb2
        if _cs is None or len(_cs) == 0: _cs = _cs2

    globals()['_KNET_MULTI_POOL'] = (ticker, _cb, _cs)
    return globals()['_KNET_MULTI_POOL']


_ZCACHE = {}

def _free_global_caches(*, keep_pool_map=False):
    """전역 캐시·맵을 비워 메모리 누적을 막는다 (Colab 여러 번 실행 시 끊김 방지).
       - _ZCACHE: z-스코어 캐시 (지표×기간 배열 — 가장 큼)
       - _LAST_FULL_CAND_MAP: 메타조합별 전체 후보 풀 (수천 행 × 메타키 — 큼)
       - _LAST_POOL_MAP: 메타조합별 선정 풀 (엑셀 재현용 — keep_pool_map이면 유지)
       각 티커/분석 시작 시 호출. gc까지 강제해 즉시 회수."""
    import gc
    g = globals()
    if '_ZCACHE' in g:
        g['_ZCACHE'].clear()
    if '_HITF_CACHE' in g:
        g['_HITF_CACHE'].clear()
    if '_ENRICH_STATS' in g:
        g['_ENRICH_STATS'].clear()
    if '_LAST_FULL_CAND_MAP' in g:
        g['_LAST_FULL_CAND_MAP'].clear()
    if not keep_pool_map and '_LAST_POOL_MAP' in g:
        g['_LAST_POOL_MAP'].clear()
    gc.collect()


def _get_zscore_cached(feat, col, window):
    """feat[col]의 z-스코어를 캐싱해 반환.
       ★ 키를 '내용 기반'으로 (요청) — 예전엔 id(feat)였는데, 파이썬 객체 id는 GC 후
       재사용돼 '다른 feat'인데 같은 키가 되어 stale z-스코어를 반환할 수 있었다(검증≠최종 원인).
       이제 (길이, 시작/끝 인덱스, 해당 컬럼 첫/끝값, 지표명, window)로 키를 만들어,
       데이터가 같으면 재사용·다르면 반드시 새로 계산한다."""
    try:
        idx = feat.index
        col_vals = feat[col].values
        n = len(col_vals)
        # 내용 식별자 — 동일 데이터면 동일, 한 행만 달라도 달라짐
        ckey = (n,
                (idx[0] if n else None), (idx[-1] if n else None),
                (float(col_vals[0]) if n and col_vals[0] == col_vals[0] else None),
                (float(col_vals[-1]) if n and col_vals[-1] == col_vals[-1] else None))
    except Exception:
        ckey = (id(feat),)   # 안전 폴백
    key = (ckey, col, window)
    c = _ZCACHE.get(key)
    if c is not None and len(c) == len(feat):
        return c
    z = _rolling_zscore(feat[col].values.astype(np.float64), window)
    # 캐시 폭주 방지: 항목이 너무 많으면 비움
    if len(_ZCACHE) > 6000:
        _ZCACHE.clear()
    _ZCACHE[key] = z
    return z


def _to_signal_array_raw(feat, row):
    """지연(lead_shift) 미적용 원신호 — 리드탐색/지연탐색 내부용."""
    d = row['direction']
    # z-스코어 신호 (롤링 정규화) — 절대임계와 달리 분포 적응적, 미래 일반화에 강함
    if d == 'z>=' or d == 'z<=':
        z = _get_zscore_cached(feat, row['indicator'], int(globals().get('ZSCORE_WINDOW', 60)))
        zvalid = ~np.isnan(z)
        if d == 'z>=':
            return ((z >= row['threshold']) & zvalid).astype(np.uint8)
        return ((z <= row['threshold']) & zvalid).astype(np.uint8)
    # 기존 절대 임계
    x = feat[row['indicator']].values.astype(np.float64)
    valid = ~np.isnan(x)
    if d == '>=':
        return ((x >= row['threshold']) & valid).astype(np.uint8)
    return ((x <= row['threshold']) & valid).astype(np.uint8)


def _to_signal_array(feat, row):
    """행의 신호 배열. ★ lead_shift(신호 지연 정렬)가 있으면 d일 늦춰 반환 —
       리드타임 탐색으로 '너무 일찍 켜지는' 지표를 체결 직전에 정렬 (sig[t-d]→t, 미래참조 없음).
       net>K·KL·지표매트릭스·재현 등 모든 소비처가 이 함수를 쓰므로 지연이 일관 적용된다."""
    sig = _to_signal_array_raw(feat, row)
    try:
        d = row.get('lead_shift', 0)
        d = 0 if d is None or (isinstance(d, float) and np.isnan(d)) else int(d)
    except Exception:
        d = 0
    if d > 0:
        sig = _shift_signal_forward(sig, d)
    return sig


def diversify_candidates(feat, score_df, *, top_n, corr_limit):
    if len(score_df) == 0: return score_df.copy()
    pool = []
    pool_signals = []
    pool_indicators = set()
    for _, row in score_df.iterrows():
        if len(pool) >= top_n: break
        if row['indicator'] in pool_indicators: continue
        sig = _to_signal_array(feat, row).astype(np.float64)
        redundant = False
        for prev_sig in pool_signals:
            mask = ~np.isnan(sig) & ~np.isnan(prev_sig)
            if mask.sum() < 10: continue
            a = sig[mask]; b = prev_sig[mask]
            if a.std() < 1e-9 or b.std() < 1e-9: continue
            c = float(np.corrcoef(a, b)[0, 1])
            if abs(c) >= corr_limit:
                redundant = True; break
        if redundant: continue
        pool.append(row.to_dict())
        pool_signals.append(sig)
        pool_indicators.add(row['indicator'])
    return pd.DataFrame(pool)


def _select_with_tolerance(df, tolerance,
                            primary='avg_success_rate',
                            secondary='combined_return'):
    """
    SELECTION_PRIORITY='sell_buy_return'이면:
      1차 매도성공률(top-tol 밴드) → 2차 매수성공률(top-tol 밴드) → 3차 누적수익 최대
    아니면 기존: primary(top-tol 밴드) 중 secondary 최대
    """
    if len(df) == 0: return None
    if secondary not in df.columns:
        secondary = 'total_return'

    prio = globals().get('SELECTION_PRIORITY', 'balacc_return')

    # ★ 일별거래 승률 밴드(-10%p) → 그중 누적수익 최고 (요청)
    #   1) 거래 승률(수익 낸 거래 비율) 최고에서 10%p 안의 후보
    #   2) 그중 누적수익이 가장 높은 것
    if prio == 'winrate_return' and 'win_rate' in df.columns:
        win_tol = globals().get('WINRATE_TOLERANCE', 0.10)
        best_win = df['win_rate'].max()
        band = df[df['win_rate'] >= best_win - win_tol].copy()
        if len(band) == 0:
            band = df[df['win_rate'] == best_win].copy()
        ret_col = 'total_return' if 'total_return' in band.columns else 'combined_return'
        band = band.sort_values([ret_col, 'win_rate'], ascending=False)
        return band.index[0]

    # ★ 매도성공 밴드(-2%p) → MDD최저 밴드(-1%p) → 누적수익 최고 (요청, 기본)
    #   1) 매도성공률 최고에서 2%p 안의 후보
    #   2) 그중 MDD 가장 낮은(0에 가까운)에서 1%p 안의 후보 (비슷한 낙폭은 묶음)
    #   3) 그중 누적수익이 가장 높은 것
    if prio == 'sell_mdd_return' and 'sell_success_rate' in df.columns \
       and 'max_drawdown' in df.columns:
        sell_tol = globals().get('SELL_SUCCESS_TOLERANCE', 0.02)
        mdd_tol  = globals().get('MDD_TOLERANCE', 0.01)
        best_sell = df['sell_success_rate'].max()
        b1 = df[df['sell_success_rate'] >= best_sell - sell_tol].copy()
        if len(b1) == 0:
            b1 = df[df['sell_success_rate'] == best_sell].copy()
        best_mdd = b1['max_drawdown'].max()           # 가장 0에 가까운(덜 빠진)
        b2 = b1[b1['max_drawdown'] >= best_mdd - mdd_tol].copy()
        if len(b2) == 0:
            b2 = b1
        ret_col = 'total_return' if 'total_return' in b2.columns else 'combined_return'
        b2 = b2.sort_values([ret_col, 'sell_success_rate'], ascending=False)
        return b2.index[0]

    # ★ 평균성공 밴드 → MDD최저 → 수익최대 모드 (요청, 기본).
    #   1) 평균성공률 최고에서 tolerance(밴드) 안의 후보를 모음
    #   2) 그중 MDD가 가장 낮은(0에 가까운) 것 — 0.1%p 단위로 같으면 '비슷'으로 보고
    #   3) 그 동률 그룹에서 누적수익이 가장 높은 것
    if prio == 'avgband_mdd_return' and 'avg_success_rate' in df.columns \
       and 'max_drawdown' in df.columns:
        best_avg = df['avg_success_rate'].max()
        band = df[df['avg_success_rate'] >= best_avg - tolerance].copy()
        if len(band) == 0:
            band = df[df['avg_success_rate'] == best_avg].copy()
        band['_mdd_r'] = band['max_drawdown'].round(3)   # 0.1%p 단위
        best_mdd_r = band['_mdd_r'].max()                # 가장 0에 가까운(덜 빠진)
        mdd_grp = band[band['_mdd_r'] == best_mdd_r]
        ret_col = 'total_return' if 'total_return' in mdd_grp.columns else 'combined_return'
        mdd_grp = mdd_grp.sort_values([ret_col, 'avg_success_rate'], ascending=False)
        return mdd_grp.index[0]

    # ★ 독립 최적화 모드 — 매수성공률 최고 설정 + 매도성공률 최고 설정을 따로 찾아 합침.
    #   (매수성공률은 K_buy/vote_buy만으로, 매도성공률은 K_sell/vote_sell만으로 결정되므로
    #    둘은 독립. 각각 최고를 골라 한 조합으로 합치면 매수·매도 성공률 모두 최댓값 달성.)
    #   합친 (K_buy*, vote_buy*, K_sell*, vote_sell*) 행은 grid가 모든 조합을 돌렸으므로
    #   이미 df 안에 존재 → 그 행을 찾아 반환. 수익·MDD는 그 조합의 실제값을 그대로 씀.
    if prio == 'independent' and 'buy_success_rate' in df.columns \
       and 'sell_success_rate' in df.columns \
       and all(c in df.columns for c in ('K_buy','vote_buy','K_sell','vote_sell')):
        # 1) 매수성공률 최고 → 동률이면 매수 신호가 더 자주 뜨는(vote_buy/K_buy 낮은) 쪽
        best_buy_sr = df['buy_success_rate'].max()
        buy_cand = df[df['buy_success_rate'] >= best_buy_sr - 1e-9].copy()
        buy_cand['_vote_ratio_b'] = buy_cand['vote_buy'] / buy_cand['K_buy'].clip(lower=1)
        buy_cand = buy_cand.sort_values(['_vote_ratio_b'], ascending=True)
        Kb_star = int(buy_cand.iloc[0]['K_buy']); vb_star = int(buy_cand.iloc[0]['vote_buy'])
        # 2) 매도성공률 최고 → 동률이면 매도 신호가 더 자주 뜨는 쪽
        best_sell_sr = df['sell_success_rate'].max()
        sell_cand = df[df['sell_success_rate'] >= best_sell_sr - 1e-9].copy()
        sell_cand['_vote_ratio_s'] = sell_cand['vote_sell'] / sell_cand['K_sell'].clip(lower=1)
        sell_cand = sell_cand.sort_values(['_vote_ratio_s'], ascending=True)
        Ks_star = int(sell_cand.iloc[0]['K_sell']); vs_star = int(sell_cand.iloc[0]['vote_sell'])
        # 3) 합친 조합 행 찾기
        merged = df[(df['K_buy']==Kb_star) & (df['vote_buy']==vb_star) &
                    (df['K_sell']==Ks_star) & (df['vote_sell']==vs_star)]
        if len(merged) > 0:
            return merged.index[0]
        # 합친 조합이 df에 없으면(필터로 빠졌거나 그리드 누락): 매수최고 행으로 폴백
        return buy_cand.index[0]

    # ★ 안정성 종합 점수 모드 — 매도·매수성공·수익·MDD방어 가중 기하평균 최대
    if prio == 'stability' and 'sell_success_rate' in df.columns \
       and 'buy_success_rate' in df.columns and 'max_drawdown' in df.columns:
        scored = compute_stability_scores(df)
        scored = scored.sort_values(
            ['stability_score', 'sell_success_rate', secondary], ascending=False)
        return scored.index[0]

    if prio == 'sell_buy_return' \
       and 'sell_success_rate' in df.columns and 'buy_success_rate' in df.columns:
        best_sell = df['sell_success_rate'].max()
        b1 = df[df['sell_success_rate'] >= best_sell - tolerance]
        if len(b1) == 0: b1 = df[df['sell_success_rate'] == best_sell]
        best_buy = b1['buy_success_rate'].max()
        b2 = b1[b1['buy_success_rate'] >= best_buy - tolerance]
        if len(b2) == 0: b2 = b1[b1['buy_success_rate'] == best_buy]
        b2_sorted = b2.sort_values(
            [secondary, 'sell_success_rate', 'buy_success_rate'], ascending=False)
        return b2_sorted.index[0]

    best_primary = df[primary].max()
    band_min = best_primary - tolerance
    band = df[df[primary] >= band_min]
    if len(band) == 0:
        band = df[df[primary] == best_primary]
    band_sorted = band.sort_values([secondary, primary], ascending=False)
    return band_sorted.index[0]


def _select_with_anchor_priority(df, *,
                                  anchor_tolerance, balacc_tolerance,
                                  match_col='anchor_avg_match_rate',
                                  balacc_col='avg_success_rate',
                                  return_col='combined_return'):
    if len(df) == 0: return None
    if return_col not in df.columns:
        return_col = 'total_return'
    if match_col not in df.columns:
        return _select_with_tolerance(df, balacc_tolerance, balacc_col, return_col)

    best_match = df[match_col].max()
    band1 = df[df[match_col] >= best_match - anchor_tolerance]
    if len(band1) == 0:
        band1 = df[df[match_col] == best_match]

    best_balacc = band1[balacc_col].max()
    band2 = band1[band1[balacc_col] >= best_balacc - balacc_tolerance]
    if len(band2) == 0:
        band2 = band1[band1[balacc_col] == best_balacc]

    band2_sorted = band2.sort_values(
        [return_col, match_col, balacc_col], ascending=False)
    return band2_sorted.index[0]


def grid_search_ensemble(feat, close, buy_pool, sell_pool, *,
                          k_buy_range, k_sell_range,
                          vote_ratio_buy, vote_ratio_sell,
                          cost, min_trades=1,
                          horizon, dd_limit, ru_limit,
                          stop_loss_pct,
                          anchor_mode=False,
                          anchor_safe_buy=None,
                          anchor_safe_sell=None,
                          oos_start_idx=None):
    close_arr = close.values.astype(np.float64)
    n_days    = len(close_arr)
    sl_pct    = float(stop_loss_pct) if (stop_loss_pct is not None and stop_loss_pct > 0) else 0.0

    use_oos = (oos_start_idx is not None and 0 < oos_start_idx < n_days)
    if use_oos:
        is_len = oos_start_idx
        years  = is_len / 252.0
    else:
        is_len = n_days
        years  = n_days / 252.0

    buy_sigs_all  = np.zeros((n_days, len(buy_pool)),  dtype=np.uint8)
    sell_sigs_all = np.zeros((n_days, len(sell_pool)), dtype=np.uint8)
    for k, (_, row) in enumerate(buy_pool.iterrows()):
        buy_sigs_all[:, k] = _to_signal_array(feat, row)
    for k, (_, row) in enumerate(sell_pool.iterrows()):
        sell_sigs_all[:, k] = _to_signal_array(feat, row)

    # ★ 가중 투표 가중치 (성공률 비례, 평균1 정규화). 일반투표면 전부 1.0
    if globals().get('USE_WEIGHTED_VOTE', False):
        buy_w_full  = compute_vote_weights(buy_pool['score'].values,
                                           globals().get('WEIGHT_MAX_RATIO', 1.6))
        sell_w_full = compute_vote_weights(sell_pool['score'].values,
                                           globals().get('WEIGHT_MAX_RATIO', 1.6))
    else:
        buy_w_full  = np.ones(len(buy_pool),  dtype=np.float64)
        sell_w_full = np.ones(len(sell_pool), dtype=np.float64)

    safe_buy, safe_sell, evaluable = _compute_safe_arrays(
        close_arr, horizon, dd_limit, ru_limit)
    if anchor_mode and anchor_safe_buy is not None and anchor_safe_sell is not None:
        safe_buy, safe_sell, evaluable = _apply_anchor_correction(
            safe_buy, safe_sell, evaluable, anchor_safe_buy, anchor_safe_sell)

    has_anchor = (anchor_mode and anchor_safe_buy is not None and anchor_safe_sell is not None)
    if has_anchor:
        anchor_buy_pass  = np.ascontiguousarray(anchor_safe_buy.astype(np.uint8))
        anchor_sell_pass = np.ascontiguousarray(anchor_safe_sell.astype(np.uint8))
        n_anchor_buy_total  = int(anchor_buy_pass.sum())
        n_anchor_sell_total = int(anchor_sell_pass.sum())
    else:
        anchor_buy_pass  = np.zeros(0, dtype=np.uint8)
        anchor_sell_pass = np.zeros(0, dtype=np.uint8)
        n_anchor_buy_total = n_anchor_sell_total = 0

    if use_oos:
        close_is  = np.ascontiguousarray(close_arr[:is_len])
        safe_buy_is  = np.ascontiguousarray(safe_buy[:is_len])
        safe_sell_is = np.ascontiguousarray(safe_sell[:is_len])
        eval_is      = np.ascontiguousarray(evaluable[:is_len])
        if has_anchor:
            anc_buy_is  = np.ascontiguousarray(anchor_buy_pass[:is_len])
            anc_sell_is = np.ascontiguousarray(anchor_sell_pass[:is_len])
        else:
            anc_buy_is = anc_sell_is = np.zeros(0, dtype=np.uint8)
        close_oos = np.ascontiguousarray(close_arr[is_len:])
        safe_buy_oos  = np.ascontiguousarray(safe_buy[is_len:])
        safe_sell_oos = np.ascontiguousarray(safe_sell[is_len:])
        eval_oos      = np.ascontiguousarray(evaluable[is_len:])
        empty_anc     = np.zeros(0, dtype=np.uint8)
        oos_years     = max(len(close_oos) / 252.0, 1e-9)
    else:
        close_is = close_arr; safe_buy_is = safe_buy
        safe_sell_is = safe_sell; eval_is = evaluable
        anc_buy_is = anchor_buy_pass; anc_sell_is = anchor_sell_pass

    if HAS_NUMBA:
        z = np.zeros((is_len, 1), dtype=np.uint8)
        zw = np.ones(1, dtype=np.float64)
        _simulate_ensemble(close_is, z, z, 1, 1, cost,
                            safe_buy_is, safe_sell_is, eval_is, sl_pct,
                            anc_buy_is, anc_sell_is, zw, zw)

    valid_k_buy  = [k for k in k_buy_range  if k <= len(buy_pool)]
    valid_k_sell = [k for k in k_sell_range if k <= len(sell_pool)]

    rows = []
    for K_b in valid_k_buy:
        buy_mat_full = np.ascontiguousarray(buy_sigs_all[:, :K_b])
        for K_s in valid_k_sell:
            sell_mat_full = np.ascontiguousarray(sell_sigs_all[:, :K_s])
            if use_oos:
                buy_mat  = np.ascontiguousarray(buy_mat_full[:is_len])
                sell_mat = np.ascontiguousarray(sell_mat_full[:is_len])
                buy_mat_oos  = np.ascontiguousarray(buy_mat_full[is_len:])
                sell_mat_oos = np.ascontiguousarray(sell_mat_full[is_len:])
            else:
                buy_mat = buy_mat_full; sell_mat = sell_mat_full
            buy_w_k  = np.ascontiguousarray(buy_w_full[:K_b])
            sell_w_k = np.ascontiguousarray(sell_w_full[:K_s])
            v_b_set = sorted(set(max(1, int(round(K_b * r))) for r in vote_ratio_buy))
            v_s_set = sorted(set(max(1, int(round(K_s * r))) for r in vote_ratio_sell))
            for v_b in v_b_set:
                for v_s in v_s_set:
                    (tr, nt, nw, srr, sqrr, mdd,
                     n_eval, n_bc, n_sc, n_stop,
                     b_tp, b_fp, b_fn,
                     s_tp, s_fp, s_fn,
                     n_ab_tot, n_ab_mat,
                     n_as_tot, n_as_mat) = _simulate_ensemble(
                        close_is, buy_mat, sell_mat, v_b, v_s, cost,
                        safe_buy_is, safe_sell_is, eval_is, sl_pct,
                        anc_buy_is, anc_sell_is, buy_w_k, sell_w_k)
                    if nt < min_trades: continue
                    if use_oos:
                        (oos_tr, oos_nt, oos_nw, _osr, _osqr, oos_mdd,
                         _one, _obc, _osc, _ostop,
                         _b1,_b2,_b3,_s1,_s2,_s3,
                         _x1,_x2,_x3,_x4) = _simulate_ensemble(
                            close_oos, buy_mat_oos, sell_mat_oos, v_b, v_s, cost,
                            safe_buy_oos, safe_sell_oos, eval_oos, sl_pct,
                            empty_anc, empty_anc, buy_w_k, sell_w_k)
                        oos_return    = oos_tr
                        oos_n_trades  = oos_nt
                        oos_win_rate  = (oos_nw / oos_nt) if oos_nt > 0 else np.nan
                        oos_max_dd    = oos_mdd
                    else:
                        oos_return    = np.nan
                        oos_n_trades  = 0
                        oos_win_rate  = np.nan
                        oos_max_dd    = np.nan
                    wr = nw / nt
                    avg = srr / nt
                    if nt > 1:
                        var = max(sqrr / nt - avg * avg, 0.0)
                        std = np.sqrt(var)
                        tpy = nt / years
                        sharpe = (avg * tpy) / (std * np.sqrt(tpy)) if std > 0 else 0.0
                    else:
                        sharpe = 0.0
                    cagr = (1 + tr) ** (1 / years) - 1 if (years > 0 and (1 + tr) > 0) else tr
                    avg_buy_score  = float(buy_pool.iloc[:K_b]['score'].mean())
                    avg_sell_score = float(sell_pool.iloc[:K_s]['score'].mean())
                    b_tn = n_eval - b_tp - b_fp - b_fn
                    s_tn = n_eval - s_tp - s_fp - s_fn
                    b_tpr = b_tp / (b_tp + b_fn) if (b_tp + b_fn) > 0 else 0.0
                    b_tnr = b_tn / (b_tn + b_fp) if (b_tn + b_fp) > 0 else 0.0
                    s_tpr = s_tp / (s_tp + s_fn) if (s_tp + s_fn) > 0 else 0.0
                    s_tnr = s_tn / (s_tn + s_fp) if (s_tn + s_fp) > 0 else 0.0
                    buy_sr  = (b_tpr + b_tnr) / 2.0
                    sell_sr = (s_tpr + s_tnr) / 2.0
                    avg_sr  = (buy_sr + sell_sr) / 2.0
                    buy_acc  = n_bc / n_eval if n_eval > 0 else 0.0
                    sell_acc = n_sc / n_eval if n_eval > 0 else 0.0
                    if has_anchor and n_ab_tot > 0:
                        buy_match  = float(n_ab_mat) / float(n_ab_tot)
                    else:
                        buy_match = 1.0
                    if has_anchor and n_as_tot > 0:
                        sell_match = float(n_as_mat) / float(n_as_tot)
                    else:
                        sell_match = 1.0
                    avg_match = (buy_match + sell_match) / 2.0
                    if use_oos and not np.isnan(oos_return):
                        combined_return = tr + oos_return
                    else:
                        combined_return = tr
                    rows.append((K_b, K_s, v_b, v_s,
                                  tr, cagr, nt, wr, avg, sharpe, mdd,
                                  avg_buy_score, avg_sell_score,
                                  v_b / K_b, v_s / K_s,
                                  buy_sr, sell_sr, avg_sr, n_eval, n_stop,
                                  buy_acc, sell_acc,
                                  buy_match, sell_match, avg_match,
                                  oos_return, oos_n_trades,
                                  oos_win_rate, oos_max_dd, combined_return))

    return pd.DataFrame(rows, columns=[
        'K_buy', 'K_sell', 'vote_buy', 'vote_sell',
        'total_return', 'cagr', 'n_trades', 'win_rate',
        'avg_return_per_trade', 'sharpe_like', 'max_drawdown',
        'avg_buy_score', 'avg_sell_score',
        'vote_ratio_buy', 'vote_ratio_sell',
        'buy_success_rate', 'sell_success_rate', 'avg_success_rate',
        'n_eval_days', 'n_stop_triggered',
        'buy_accuracy_plain', 'sell_accuracy_plain',
        'anchor_buy_match_rate', 'anchor_sell_match_rate', 'anchor_avg_match_rate',
        'oos_return', 'oos_n_trades',
        'oos_win_rate', 'oos_max_drawdown', 'combined_return',
    ])


def apply_mdd_and_trade_filters(grid_df, mdd_limit_pct, min_trades_daily, allow_fallback=True):
    """MDD 한도·최소거래수 필터. 만족 조합이 없으면(요청) 에러 대신
       ① 거래수 조건 만족분 중, ② 최대낙폭(0에 가까운) 가장 낮은 것들로 폴백한다."""
    df = grid_df.copy()
    if len(df) == 0:
        return df
    df_t = df[df['n_trades'] >= min_trades_daily]
    base = df_t if len(df_t) > 0 else (df if allow_fallback else df_t)   # 거래수 만족 없으면 전체로 폴백
    if mdd_limit_pct is not None:
        ok = base[base['max_drawdown'] >= -abs(mdd_limit_pct)]
        if len(ok) > 0:
            return ok
        # ★ MDD 한도 만족 조합 없음 → 최대낙폭 가장 낮은(0에 가까운) 조합으로 폴백
        if allow_fallback and len(base) > 0:
            _fb = base.sort_values('max_drawdown', ascending=False).head(max(5, min(20, len(base))))
            globals()['_MDD_FALLBACK_USED'] = True
            return _fb
        return ok
    return base


def meta_grid_search(feat, close, *,
                      meta_grid, indicators,
                      n_thresholds, horizon, dd_limit, ru_limit,
                      k_buy_range, k_sell_range,
                      vote_ratio_buy, vote_ratio_sell,
                      cost, min_trades_daily,
                      mdd_limit_pct, select_by,
                      stop_loss_pct,
                      selection_tolerance=0.0,
                      anchor_match_priority=False,
                      anchor_match_tolerance=0.05,
                      anchor_mode=False,
                      anchor_safe_buy=None,
                      anchor_safe_sell=None,
                      oos_start_idx=None,
                      bh_ret=None,
                      exclude_below_bh=True):
    score_cache = {}
    use_anchor = bool(anchor_mode and anchor_safe_buy is not None and anchor_safe_sell is not None)

    use_oos = (oos_start_idx is not None and 0 < oos_start_idx < len(close))
    if use_oos:
        feat_score  = feat.iloc[:oos_start_idx]
        close_score = close.iloc[:oos_start_idx]
        ab  = anchor_safe_buy[:oos_start_idx]  if use_anchor else None
        asu = anchor_safe_sell[:oos_start_idx] if use_anchor else None
    else:
        feat_score  = feat
        close_score = close
        ab  = anchor_safe_buy if use_anchor else None
        asu = anchor_safe_sell if use_anchor else None

    def get_scores(wilson_z, pct_range, min_signals):
        key = (wilson_z, pct_range, min_signals)
        if key in score_cache:
            return score_cache[key]
        pct_low, pct_high = pct_range
        buy_df, sell_df = evaluate_buy_sell_scores(
            feat_score, close_score, indicators=indicators,
            n_thresholds=n_thresholds,
            pct_low=pct_low, pct_high=pct_high,
            horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
            min_signals=min_signals, wilson_z=wilson_z,
            anchor_buy_arr=ab, anchor_sell_arr=asu,
        )
        score_cache[key] = (buy_df, sell_df)
        return buy_df, sell_df

    # ★ 성공률 우선 풀 (요청) — pct무관하게 1회만 계산해 모든 메타조합에서 공용.
    succ_buy_full = succ_sell_full = succ_buy_dedup = succ_sell_dedup = None
    if POOL_SELECT_BY_SUCCESS:
        print(f"  ★ 성공률 우선 풀 선출 ON — 컷오프 {POOL_SUCCESS_MIN_RATE*100:.0f}%, "
              f"최소신호 {POOL_SUCCESS_MIN_SIG}, 분위 {POOL_SUCCESS_WIDE_PCT}")
        succ_buy_full, succ_sell_full, succ_buy_dedup, succ_sell_dedup = select_pool_by_success(
            feat_score, close_score, indicators=indicators, n_thresholds=n_thresholds,
            horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit, wilson_z=1.0)
        _nb = 0 if succ_buy_dedup is None else len(succ_buy_dedup)
        _ns = 0 if succ_sell_dedup is None else len(succ_sell_dedup)
        print(f"     ▷ 성공률 {POOL_SUCCESS_MIN_RATE*100:.0f}%+ 지표: 매수 {_nb}개, 매도 {_ns}개")
        # 시트용으로 전역 저장 (write_excel에서 읽음) — 기존 _LAST_*_MAP 패턴과 동일.
        try:
            globals()['_LAST_SUCCESS_POOL'] = (
                succ_buy_dedup.copy()  if succ_buy_dedup  is not None else None,
                succ_sell_dedup.copy() if succ_sell_dedup is not None else None)
        except Exception:
            globals()['_LAST_SUCCESS_POOL'] = (succ_buy_dedup, succ_sell_dedup)

        # ★ K 하한 자동 인하 (요청) — 정예(희소) 풀은 K가 10부터면 동의 임계(K×vote)를 못 채워 거래 0.
        #   K 하한을 POOL_SUCCESS_K_FLOOR까지 낮춰 소수 동의로도 신호가 나게 함(상한은 유지).
        _kf = int(POOL_SUCCESS_K_FLOOR)
        try:
            _kbmax = max(k_buy_range)  if len(k_buy_range)  else 30
            _ksmax = max(k_sell_range) if len(k_sell_range) else 30
            if min(k_buy_range) > _kf:
                k_buy_range  = list(range(_kf, _kbmax + 1))
            if min(k_sell_range) > _kf:
                k_sell_range = list(range(_kf, _ksmax + 1))
            print(f"     ▷ 성공률 우선 풀 → K 하한 {_kf}로 인하 "
                  f"(K_buy {min(k_buy_range)}~{max(k_buy_range)}, K_sell {min(k_sell_range)}~{max(k_sell_range)})")
        except Exception as _ek:
            print(f"     ⚠ K 하한 인하 실패(무시): {_ek}")
    else:
        globals()['_LAST_SUCCESS_POOL'] = (None, None)

    combos = list(itertools.product(
        meta_grid['wilson_z'],
        meta_grid['pct_range'],
        meta_grid['min_signals'],
        meta_grid['corr_limit'],
        meta_grid['top_n_pool'],
    ))
    total = len(combos)
    print(f"  메타 그리드 총 {total}개 조합")
    _prio_p = globals().get('SELECTION_PRIORITY', 'balacc_return')
    if _prio_p == 'winrate_return':
        _wt = globals().get('WINRATE_TOLERANCE', 0.10)
        print(f"  ★ 선정: 일별거래 승률 -{_wt*100:.0f}%p 밴드 → 누적수익 최고")
    elif _prio_p == 'sell_mdd_return':
        _st = globals().get('SELL_SUCCESS_TOLERANCE', 0.02); _mt = globals().get('MDD_TOLERANCE', 0.01)
        print(f"  ★ 선정: 매도성공률 -{_st*100:.0f}%p 밴드 → MDD 최저 -{_mt*100:.0f}%p 밴드 → 누적수익 최고")
    elif _prio_p == 'avgband_mdd_return':
        print(f"  ★ 선정: 평균성공률 top band({selection_tolerance*100:.1f}%p) 내 → MDD 최저(동률시 수익 최대)")
    elif _prio_p == 'independent':
        print(f"  ★ 선정: 매수성공률 최고 설정 + 매도성공률 최고 설정 따로 찾아 합침 (독립 최적화)")
    elif _prio_p == 'stability':
        print(f"  ★ 선정: 안정성 종합점수(매도·매수성공·수익·MDD방어 가중 기하평균) 최대")
    elif _prio_p == 'sell_buy_return':
        print(f"  ★ 선정 우선순위: 매도성공률 → 매수성공률 → 누적수익 (밴드 {selection_tolerance*100:.2f}%p)")
    elif selection_tolerance > 0:
        print(f"  ★ Tolerance Band: 평균 BalAcc top - {selection_tolerance*100:.2f}%p 이내 중 수익률 최대")
    else:
        print(f"  ★ Strict: 1차 평균 BalAcc → 2차 수익률")
    use_match_priority = bool(anchor_match_priority and anchor_mode and
                               anchor_safe_buy is not None and anchor_safe_sell is not None)
    if use_match_priority:
        print(f"  ⚓ ANCHOR_MATCH_PRIORITY ON — 3단계 우선순위:")
        print(f"     1차: 매칭률 top - {anchor_match_tolerance*100:.1f}%p 이내")
        print(f"     2차: BalAcc top - {selection_tolerance*100:.2f}%p 이내")
        print(f"     3차: 수익률 최대")

    meta_rows = []
    all_passed_list = []      # ★ 모든 메타조합의 통과 그리드를 메타변수 붙여 누적 (통합 테이블)
    n_below_bh = 0            # ★ B&H 미달로 제외된 조합 수
    best_state = None
    best_overall_match = -np.inf
    best_overall_balacc = -np.inf
    best_overall_return = -np.inf
    n_no_pass = 0
    t0 = time.time()

    for ci, (wz, pr, ms, cl, tnp) in enumerate(combos):
        meta = dict(
            wilson_z=wz, pct_low=pr[0], pct_high=pr[1],
            min_signals=ms, corr_limit=cl,
            top_n_pool_buy=tnp, top_n_pool_sell=tnp,
        )
        buy_df, sell_df = get_scores(wz, pr, ms)
        # ★ buy_df/sell_df는 (성공률 우선 ON이면) evaluate_buy_sell_scores에서
        #   이미 '지표당 1행·성공률 내림차순'으로 정리됨 → diversify가 성공률 우선 풀을 만듦.
        buy_pool  = diversify_candidates(feat_score, buy_df,  top_n=tnp, corr_limit=cl)
        sell_pool = diversify_candidates(feat_score, sell_df, top_n=tnp, corr_limit=cl)

        # ★ 보정용 전체 후보 풀 저장 (요청) — top_n_pool은 그대로 100, 보정만 전체 탐색.
        #   diversify로 100개 추리기 전의 '점수 매긴 전체 후보(buy_df/sell_df)'를 메타키별 보관.
        #   보정 함수가 이 큰 풀에서 미매칭일 보정에 쓸 미사용 지표를 찾는다 (그리드 속도 영향 없음).
        try:
            _fc_key = (round(float(wz),4), int(pr[0]), int(pr[1]), round(float(cl),4))
            globals().setdefault('_LAST_FULL_CAND_MAP', {})[_fc_key] = (
                buy_df.reset_index(drop=True), sell_df.reset_index(drop=True))
        except Exception:
            pass

        inner_df = grid_search_ensemble(
            feat, close, buy_pool, sell_pool,
            k_buy_range=k_buy_range, k_sell_range=k_sell_range,
            vote_ratio_buy=vote_ratio_buy, vote_ratio_sell=vote_ratio_sell,
            cost=cost, min_trades=1,
            horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
            stop_loss_pct=stop_loss_pct,
            anchor_mode=anchor_mode,
            anchor_safe_buy=anchor_safe_buy,
            anchor_safe_sell=anchor_safe_sell,
            oos_start_idx=oos_start_idx,
        )
        passed = apply_mdd_and_trade_filters(inner_df, mdd_limit_pct, min_trades_daily)

        # ★ Buy&Hold 미달 조합 제외 (전략 수익이 B&H 이하면 버림)
        if exclude_below_bh and bh_ret is not None and len(passed) > 0:
            before_n = len(passed)
            passed = passed[passed['total_return'] > bh_ret].copy()
            n_below_bh += (before_n - len(passed))

        # ★ 이 메타조합의 통과 그리드에 메타변수 컬럼 붙여서 통합 테이블에 누적
        if len(passed) > 0:
            tagged = passed.copy()
            tagged['meta_wilson_z']   = wz
            tagged['meta_pct_low']    = pr[0]
            tagged['meta_pct_high']   = pr[1]
            tagged['meta_min_signals']= ms
            tagged['meta_corr_limit'] = cl
            tagged['meta_top_n_pool'] = tnp
            all_passed_list.append(tagged)

        if len(passed) == 0:
            n_no_pass += 1
            meta_rows.append({
                **meta,
                'n_buy_pool': len(buy_pool), 'n_sell_pool': len(sell_pool),
                'n_inner_total': len(inner_df), 'n_inner_passed': 0,
                'best_return': np.nan, 'best_mdd': np.nan, 'best_n_trades': 0,
                'best_K_buy': np.nan, 'best_K_sell': np.nan,
                'best_vote_buy': np.nan, 'best_vote_sell': np.nan,
                'best_win_rate': np.nan, 'best_sharpe': np.nan,
                'best_buy_sr': np.nan, 'best_sell_sr': np.nan, 'best_avg_sr': np.nan,
                'best_buy_match': np.nan, 'best_sell_match': np.nan, 'best_avg_match': np.nan,
            })
        else:
            if use_match_priority:
                best_idx = _select_with_anchor_priority(
                    passed,
                    anchor_tolerance=anchor_match_tolerance,
                    balacc_tolerance=selection_tolerance,
                    return_col='combined_return')
            else:
                best_idx = _select_with_tolerance(
                    passed, selection_tolerance,
                    primary='avg_success_rate', secondary='combined_return')
            best_in = passed.loc[best_idx].to_dict()
            meta_rows.append({
                **meta,
                'n_buy_pool': len(buy_pool), 'n_sell_pool': len(sell_pool),
                'n_inner_total': len(inner_df), 'n_inner_passed': len(passed),
                'best_return': best_in['total_return'],
                'best_oos_return': best_in.get('oos_return', np.nan),
                'best_combined_return': best_in.get('combined_return', best_in['total_return']),
                'best_mdd': best_in['max_drawdown'],
                'best_n_trades': int(best_in['n_trades']),
                'best_K_buy': int(best_in['K_buy']),
                'best_K_sell': int(best_in['K_sell']),
                'best_vote_buy': int(best_in['vote_buy']),
                'best_vote_sell': int(best_in['vote_sell']),
                'best_win_rate': best_in['win_rate'],
                'best_sharpe': best_in['sharpe_like'],
                'best_buy_sr':  best_in['buy_success_rate'],
                'best_sell_sr': best_in['sell_success_rate'],
                'best_avg_sr':  best_in['avg_success_rate'],
                'best_buy_match':  best_in.get('anchor_buy_match_rate',  np.nan),
                'best_sell_match': best_in.get('anchor_sell_match_rate', np.nan),
                'best_avg_match':  best_in.get('anchor_avg_match_rate',  np.nan),
            })
            this_balacc = best_in['avg_success_rate']
            this_return = best_in.get('combined_return', best_in['total_return'])
            this_match  = best_in.get('anchor_avg_match_rate', 1.0)

            update = False
            if best_state is None:
                update = True
            elif use_match_priority:
                if this_match > best_overall_match + anchor_match_tolerance:
                    update = True
                elif this_match >= best_overall_match - anchor_match_tolerance:
                    if this_balacc > best_overall_balacc + selection_tolerance:
                        update = True
                    elif this_balacc >= best_overall_balacc - selection_tolerance:
                        if this_return > best_overall_return:
                            update = True
            else:
                _prio = globals().get('SELECTION_PRIORITY', 'balacc_return')
                # ★ 일별거래 승률 밴드(-10%p) → 수익 (요청). meta 레벨도 동일.
                if _prio == 'winrate_return':
                    _wtol = globals().get('WINRATE_TOLERANCE', 0.10)
                    this_win = best_in.get('win_rate', 0.0)
                    bs = best_state[3] if best_state else None
                    bs_win = bs.get('win_rate', -np.inf) if bs else -np.inf
                    if this_win > bs_win + _wtol:
                        update = True
                    elif this_win >= bs_win - _wtol:
                        if this_return > best_overall_return:
                            update = True
                # ★ 매도성공 밴드 → MDD최저 밴드 → 수익 (요청). meta 레벨도 동일 우선순위.
                elif _prio == 'sell_mdd_return':
                    _stol = globals().get('SELL_SUCCESS_TOLERANCE', 0.02)
                    _mtol = globals().get('MDD_TOLERANCE', 0.01)
                    this_sell = best_in['sell_success_rate']
                    this_mdd  = best_in['max_drawdown']
                    bs = best_state[3] if best_state else None
                    bs_sell = bs['sell_success_rate'] if bs else -np.inf
                    bs_mdd  = bs['max_drawdown']      if bs else -np.inf
                    # 1차 매도성공률 밴드(-2%p)
                    if this_sell > bs_sell + _stol:
                        update = True
                    elif this_sell >= bs_sell - _stol:
                        # 2차 MDD 밴드(-1%p) — 덜 빠진 쪽
                        if this_mdd > bs_mdd + _mtol:
                            update = True
                        elif this_mdd >= bs_mdd - _mtol:
                            # 3차 수익
                            if this_return > best_overall_return:
                                update = True
                # ★ 평균성공 밴드 → MDD최저 → 수익 (요청). meta 레벨도 동일 우선순위.
                elif _prio == 'avgband_mdd_return':
                    this_avg = best_in['avg_success_rate']
                    this_mdd = best_in['max_drawdown']
                    bs = best_state[3] if best_state else None
                    bs_avg = bs['avg_success_rate'] if bs else -np.inf
                    bs_mdd = bs['max_drawdown']     if bs else -np.inf
                    # 1차 평균성공률 밴드
                    if this_avg > bs_avg + selection_tolerance:
                        update = True
                    elif this_avg >= bs_avg - selection_tolerance:
                        # 2차 MDD (0.1%p 단위로 비교) — 덜 빠진 쪽
                        if round(this_mdd, 3) > round(bs_mdd, 3):
                            update = True
                        elif round(this_mdd, 3) == round(bs_mdd, 3):
                            # 3차 수익
                            if this_return > best_overall_return:
                                update = True
                # ★ 독립 최적화 — meta 레벨은 best_inner의 (매수성공+매도성공) 합이
                #   가장 큰 메타 조합을 채택. 동률이면 수익으로.
                elif _prio == 'independent':
                    this_pair = best_in['buy_success_rate'] + best_in['sell_success_rate']
                    bs = best_state[3] if best_state else None
                    bs_pair = (bs['buy_success_rate'] + bs['sell_success_rate']) if bs else -np.inf
                    if this_pair > bs_pair + 1e-9:
                        update = True
                    elif this_pair >= bs_pair - 1e-9:
                        if this_return > best_overall_return:
                            update = True
                # ★ 안정성 모드 — meta 레벨은 정규화 기준이 없어, best_inner의
                #   매도성공 → 평균성공 → MDD방어 → 수익 순 밴드 비교로 stability에 준해 갱신
                elif _prio == 'stability':
                    this_sell = best_in['sell_success_rate']
                    this_avg  = best_in['avg_success_rate']
                    this_mdd  = best_in['max_drawdown']     # 음수, 클수록(0근처) 방어 좋음
                    bs = best_state[3] if best_state else None
                    bs_sell = bs['sell_success_rate'] if bs else -np.inf
                    bs_avg  = bs['avg_success_rate']  if bs else -np.inf
                    bs_mdd  = bs['max_drawdown']      if bs else -np.inf
                    if this_sell > bs_sell + selection_tolerance:
                        update = True
                    elif this_sell >= bs_sell - selection_tolerance:
                        if this_avg > bs_avg + selection_tolerance:
                            update = True
                        elif this_avg >= bs_avg - selection_tolerance:
                            # MDD 방어가 뚜렷이 좋으면 갱신, 비슷하면 수익으로
                            if this_mdd > bs_mdd + 0.01:        # 1%p 이상 덜 빠짐
                                update = True
                            elif this_mdd >= bs_mdd - 0.01:
                                if this_return > best_overall_return:
                                    update = True
                elif _prio == 'sell_buy_return':
                    this_sell = best_in['sell_success_rate']
                    this_buy  = best_in['buy_success_rate']
                    bs_sell = best_state[3]['sell_success_rate'] if best_state else -np.inf
                    bs_buy  = best_state[3]['buy_success_rate']  if best_state else -np.inf
                    if this_sell > bs_sell + selection_tolerance:
                        update = True
                    elif this_sell >= bs_sell - selection_tolerance:
                        if this_buy > bs_buy + selection_tolerance:
                            update = True
                        elif this_buy >= bs_buy - selection_tolerance:
                            if this_return > best_overall_return:
                                update = True
                else:
                    if this_balacc > best_overall_balacc + selection_tolerance:
                        update = True
                    elif this_balacc >= best_overall_balacc - selection_tolerance:
                        if this_return > best_overall_return:
                            update = True
                        elif this_return == best_overall_return and this_balacc > best_overall_balacc:
                            update = True

            if update:
                best_state = (meta.copy(), inner_df.copy(), passed.copy(),
                              best_in, buy_pool.copy(), sell_pool.copy())
                if this_match > best_overall_match:
                    best_overall_match = this_match
                if this_balacc > best_overall_balacc:
                    best_overall_balacc = this_balacc
                best_overall_return = this_return

        if (ci + 1) % max(total // 10, 1) == 0 or ci == total - 1:
            el = time.time() - t0
            est = el * total / (ci + 1)
            if best_state is not None:
                if use_match_priority:
                    cur = f"Match {best_overall_match*100:.1f}% / BalAcc {best_overall_balacc*100:.1f}% / Ret {best_overall_return*100:+.2f}%"
                else:
                    cur = f"BalAcc {best_overall_balacc*100:.1f}% / Ret {best_overall_return*100:+.2f}%"
            else:
                cur = "—"
            print(f"    ▸ {ci+1}/{total} ({(ci+1)/total*100:>3.0f}%)  "
                  f"경과 {el:>4.0f}s / 예상 {est:>4.0f}s  "
                  f"현재 best: {cur}  (통과없음 {n_no_pass}건)")

    if globals().get('_MDD_FALLBACK_USED'):
        print("  ⚠ MDD 한도 만족 조합 없음 → 최대낙폭 가장 낮은 조합으로 폴백 진행 (에러 대신).")
        globals().pop('_MDD_FALLBACK_USED', None)

    if best_state is None:
        # 여기까지 오면 통과 조합이 전무(=inner 그리드 자체가 빈 경우). MDD 미달은 위 필터에서 폴백됨.
        if all_passed_list:
            _cat = pd.concat(all_passed_list, ignore_index=True)
            _pick = _cat.sort_values('max_drawdown', ascending=False).head(1)
            print("  ⚠ 통과 조합 전무 → 최대낙폭 최저 1개로 폴백 진행.")
            best_state = ({}, _cat, _pick, _pick.iloc[0].to_dict(),
                          float(_pick.iloc[0].get('avg_success_rate', 0.0) or 0.0),
                          float(_pick.iloc[0].get('total_return', 0.0) or 0.0), 1.0)
        else:
            raise RuntimeError(
                "그리드 결과가 비었습니다 (지표/신호 없음). 풀·MIN_TRADES 설정을 확인하세요."
            )

    all_meta_df = pd.DataFrame(meta_rows)
    valid = all_meta_df.dropna(subset=['best_avg_sr']).copy()
    no_pass = all_meta_df[all_meta_df['best_avg_sr'].isna()].copy()
    ret_key = 'best_combined_return' if (use_oos and 'best_combined_return' in valid.columns) else 'best_return'
    if len(valid) > 0 and selection_tolerance > 0:
        top_balacc = valid['best_avg_sr'].max()
        in_band  = valid[valid['best_avg_sr'] >= top_balacc - selection_tolerance]
        out_band = valid[valid['best_avg_sr'] <  top_balacc - selection_tolerance]
        in_sorted  = in_band.sort_values([ret_key, 'best_avg_sr'], ascending=False)
        out_sorted = out_band.sort_values(['best_avg_sr', ret_key], ascending=False)
        meta_results_df = pd.concat([in_sorted, out_sorted, no_pass]).reset_index(drop=True)
    else:
        meta_results_df = pd.concat([
            valid.sort_values(['best_avg_sr', ret_key], ascending=False),
            no_pass
        ]).reset_index(drop=True)

    best_meta, inner_all, inner_passed, best_inner, buy_pool, sell_pool = best_state

    # ★ 통합 테이블 — 모든 메타조합의 통과 그리드(메타변수 포함)
    if all_passed_list:
        combined_all = pd.concat(all_passed_list, ignore_index=True)
        # ★ 최종 선정을 '통합 테이블 전체'에서 수행 (모든 메타×그리드를 한 풀에서 비교)
        try:
            comb_idx = _select_with_tolerance(
                combined_all, selection_tolerance,
                primary='avg_success_rate', secondary='combined_return')
            comb_best = combined_all.loc[comb_idx]
            # 통합 best의 메타변수로 best_meta 갱신 (재현 일치 위해)
            best_meta = dict(
                wilson_z=float(comb_best['meta_wilson_z']),
                pct_low=int(comb_best['meta_pct_low']),
                pct_high=int(comb_best['meta_pct_high']),
                min_signals=int(comb_best['meta_min_signals']),
                corr_limit=float(comb_best['meta_corr_limit']),
                top_n_pool_buy=int(comb_best['meta_top_n_pool']),
                top_n_pool_sell=int(comb_best['meta_top_n_pool']),
            )
            best_inner = comb_best.to_dict()
            print(f"\n  ★ 통합 테이블({len(combined_all)}개)에서 최종 선정 — "
                  f"wilson_z={best_meta['wilson_z']}, pct=({best_meta['pct_low']},{best_meta['pct_high']}), "
                  f"corr={best_meta['corr_limit']}")
            if n_below_bh > 0:
                print(f"     (Buy&Hold 미달 {n_below_bh}개 조합 제외됨)")
        except Exception as _e:
            print(f"  ⚠ 통합 선정 중 경고: {_e} — 메타별 best 사용")
    else:
        combined_all = inner_passed.copy() if inner_passed is not None else pd.DataFrame()

    # 통합 테이블을 정렬해서 반환 (선정 기준 순)
    combined_sorted = _sort_combined_table(combined_all, selection_tolerance)
    return (meta_results_df, inner_all, combined_sorted, best_meta, best_inner,
            buy_pool, sell_pool)


def _sort_combined_table(df, tol):
    """통합 그리드 테이블을 현재 SELECTION_PRIORITY 기준으로 정렬."""
    if df is None or len(df) == 0:
        return df
    prio = globals().get('SELECTION_PRIORITY', 'balacc_return')
    d = df.copy()
    ret_col = 'total_return' if 'total_return' in d.columns else 'combined_return'
    try:
        if prio == 'winrate_return':
            d = d.sort_values(['win_rate', ret_col], ascending=[False, False])
        elif prio == 'sell_mdd_return':
            d = d.sort_values(['sell_success_rate', 'max_drawdown', ret_col],
                              ascending=[False, False, False])
        elif prio == 'avgband_mdd_return':
            d = d.sort_values(['avg_success_rate', 'max_drawdown', ret_col],
                              ascending=[False, False, False])
        elif prio == 'independent':
            d['_pair'] = d['buy_success_rate'] + d['sell_success_rate']
            d = d.sort_values(['_pair', ret_col], ascending=[False, False]).drop(columns=['_pair'])
        else:
            d = d.sort_values(['avg_success_rate', ret_col], ascending=[False, False])
    except Exception:
        pass
    return d.reset_index(drop=True)


def daily_ensemble_backtest(feat, close, buy_pool, sell_pool,
                              K_buy, K_sell, vote_buy, vote_sell, *,
                              cost=None,
                              horizon=None,
                              dd_limit=None,
                              ru_limit=None,
                              stop_loss_pct=None,
                              anchor_mode=False,
                              anchor_safe_buy=None,
                              anchor_safe_sell=None,
                              buy_w_override=None,
                              sell_w_override=None):
    import numpy as np
    import pandas as pd
    # 기본값 — 원본 전역 상수 사용
    if cost is None:      cost      = COST_PER_TRADE
    if horizon is None:   horizon   = HORIZON_DAYS
    if dd_limit is None:  dd_limit  = DRAWDOWN_LIMIT_BUY
    if ru_limit is None:  ru_limit  = RUNUP_LIMIT_SELL

    # ★ 거래일 필터는 run_ensemble_search 진입 시 이미 수행됨(공휴일 제외).
    #   여기서는 들어온 close를 그대로 사용. 단, 단독 호출 대비 잔여 NaN만 안전 처리.
    if isinstance(close, pd.Series) and close.isna().any():
        valid_mask = close.notna().values
        close = close[valid_mask]
        feat = feat.iloc[valid_mask]
        if anchor_safe_buy is not None and len(anchor_safe_buy) == len(valid_mask):
            anchor_safe_buy = np.asarray(anchor_safe_buy)[valid_mask]
        if anchor_safe_sell is not None and len(anchor_safe_sell) == len(valid_mask):
            anchor_safe_sell = np.asarray(anchor_safe_sell)[valid_mask]
    cl = close
    n  = len(feat); dates = feat.index
    buy_used  = buy_pool.iloc[:K_buy].reset_index(drop=True)
    sell_used = sell_pool.iloc[:K_sell].reset_index(drop=True)
    buy_mat  = np.zeros((n, K_buy),  dtype=np.uint8)
    sell_mat = np.zeros((n, K_sell), dtype=np.uint8)
    for k, (_, row) in enumerate(buy_used.iterrows()):
        buy_mat[:, k] = _to_signal_array(feat, row)
    for k, (_, row) in enumerate(sell_used.iterrows()):
        sell_mat[:, k] = _to_signal_array(feat, row)

    # ★ 변경2: 가중 투표 가중치 (성공률 비례). 일반투표면 전부 1.0
    if globals().get('USE_WEIGHTED_VOTE', False):
        buy_w  = compute_vote_weights(buy_used['score'].values,
                                      globals().get('WEIGHT_MAX_RATIO', 1.6))
        sell_w = compute_vote_weights(sell_used['score'].values,
                                      globals().get('WEIGHT_MAX_RATIO', 1.6))
        weighted = True
    else:
        buy_w  = np.ones(K_buy,  dtype=np.float64)
        sell_w = np.ones(K_sell, dtype=np.float64)
        weighted = False
    # ★ 앵커 미매칭 보정 (요청) — 외부에서 지표 가중치를 주입하면 그걸 사용.
    #   특정 지표 점수(가중치)를 조정해 충돌 패배·간발의 차 미매칭을 보정하기 위함.
    if buy_w_override is not None and len(buy_w_override) == K_buy:
        buy_w = np.asarray(buy_w_override, dtype=np.float64)
        weighted = True
    if sell_w_override is not None and len(sell_w_override) == K_sell:
        sell_w = np.asarray(sell_w_override, dtype=np.float64)
        weighted = True

    # ★ 재현용 (요청) — 실제 사용된 투표 가중치를 풀에 부착해 두면 엑셀에 저장 가능.
    #   재현 시 이 가중치를 그대로 주입하면 보정(추가지표+가중치)이 정확히 재현된다.
    try:
        buy_used  = buy_used.copy();  buy_used['vote_weight']  = np.asarray(buy_w, dtype=np.float64)
        sell_used = sell_used.copy(); sell_used['vote_weight'] = np.asarray(sell_w, dtype=np.float64)
    except Exception:
        pass

    sl_active = stop_loss_pct is not None and stop_loss_pct > 0
    sl = float(stop_loss_pct) if sl_active else 0.0

    # ★ 변경4: 향후 안전 배열 (성공/실패 판정용) — 미리 계산
    close_vals0 = cl.values.astype(np.float64)
    safe_buy0, safe_sell0, eval0 = _compute_safe_arrays(
        close_vals0, horizon, dd_limit, ru_limit)
    if anchor_mode and anchor_safe_buy is not None and anchor_safe_sell is not None:
        safe_buy0, safe_sell0, eval0 = _apply_anchor_correction(
            safe_buy0, safe_sell0, eval0, anchor_safe_buy, anchor_safe_sell)

    pos = 0; entry_price = np.nan; entry_date = pd.NaT; entry_idx = -1
    stop_price_cur = np.nan
    sum_daily = 0.0
    up_sum_daily = 0.0
    equity = 1.0; n_trades = 0; n_wins = 0
    rows = []; trades = []
    peak = 1.0; mdd = 0.0
    worst_trade_so_far = 0.0   # ★ 그 시점까지 발생한 개별 거래 최대 손실(소수, 음수)
    n_conflicts = 0
    n_conflict_buy_won = 0
    n_conflict_sell_won = 0
    n_stop_triggered = 0
    pos_post = np.zeros(n, dtype=np.int8)
    prev_close = np.nan

    # ★ 전체 매수/매도 성공·실패 집계 카운터
    #   [정답인 날만 평가] 매수: 올랐어야 하는 날 중 신호 ON으로 맞춘 비율(놓침은 실패).
    #   '안 올랐는데 신호 없음'은 평가에서 제외 — 상승장 무임승차 방지(요청).
    n_buy_on_total = 0; n_buy_success = 0; n_buy_fail = 0          # 신호 ON만 (적중률용)
    n_sell_on_total = 0; n_sell_success = 0; n_sell_fail = 0
    n_buy_eval_all = 0; n_buy_correct_all = 0                      # ON+OFF 전체 (정확도용)
    n_sell_eval_all = 0; n_sell_correct_all = 0

    for i in range(n):
        d = dates[i]; price = float(cl.iloc[i])
        if pos == 1 and entry_idx >= 0 and i > entry_idx and pd.notna(prev_close) and prev_close > 0:
            _dret = price / prev_close - 1.0
            sum_daily += _dret
            if _dret > 0.0: up_sum_daily += _dret
            equity = 1.0 + sum_daily

        # ★ 변경2: 가중 카운트
        b_count = float(np.dot(buy_mat[i].astype(np.float64), buy_w))
        s_count = float(np.dot(sell_mat[i].astype(np.float64), sell_w))
        b_raw = int(buy_mat[i].sum())     # 원시 신호 개수(공식 표시용)
        s_raw = int(sell_mat[i].sum())
        b_on = b_count >= vote_buy
        s_on = s_count >= vote_sell

        b_strength = b_count / K_buy if K_buy > 0 else 0.0
        s_strength = s_count / K_sell if K_sell > 0 else 0.0

        # ★ 변경3: 점수 계산식 문자열 (중립 등 표시용)
        if weighted:
            b_formula = f"가중Σ {b_count:.2f} (신호 {b_raw}/{K_buy} × 성공률가중) {'≥' if b_on else '<'} {vote_buy}"
            s_formula = f"가중Σ {s_count:.2f} (신호 {s_raw}/{K_sell} × 성공률가중) {'≥' if s_on else '<'} {vote_sell}"
        else:
            b_formula = f"단순합 {int(b_count)} (신호 {b_raw}/{K_buy} × 1표) {'≥' if b_on else '<'} {vote_buy}"
            s_formula = f"단순합 {int(s_count)} (신호 {s_raw}/{K_sell} × 1표) {'≥' if s_on else '<'} {vote_sell}"

        # ★ 이 날 매수/매도 신호의 성공·실패 판정 (first-touch 정답 기준)
        #   [정답인 날만 평가 — 요청 반영]
        #   매수 정답일(올랐어야): 신호 ON → ✓성공 / 신호 OFF → ✗놓침
        #   매수 정답 아닌 날(안 올랐음): 신호 ON → ✗실패 / 신호 OFF → 평가 제외(공란)
        #   매도도 대칭. '안 내렸는데 매도 안 함'을 성공으로 치지 않는다(부풀리기 방지).
        buy_result = ''; sell_result = ''
        if eval0[i] == 1:
            sb = safe_buy0[i] == 1     # 매수 정답일(올랐어야)
            ss = safe_sell0[i] == 1    # 매도 정답일(내렸어야)
            # ── 매수 ──
            if b_on:
                n_buy_on_total += 1
                if sb:
                    buy_result = '✓성공'; n_buy_success += 1
                    n_buy_correct_all += 1; n_buy_eval_all += 1
                else:
                    buy_result = '✗실패'; n_buy_fail += 1
                    n_buy_eval_all += 1
            else:
                if sb:
                    buy_result = '✗놓침'                    # 올랐어야 했는데 신호 없음
                    n_buy_eval_all += 1
                # else: 안 올랐고 신호도 없음 → 평가 제외(카운트 안 함)
            # ── 매도 ──
            if s_on:
                n_sell_on_total += 1
                if ss:
                    sell_result = '✓성공'; n_sell_success += 1
                    n_sell_correct_all += 1; n_sell_eval_all += 1
                else:
                    sell_result = '✗실패'; n_sell_fail += 1
                    n_sell_eval_all += 1
            else:
                if ss:
                    sell_result = '✗놓침'                   # 내렸어야 했는데 신호 없음
                    n_sell_eval_all += 1
                # else: 안 내렸고 신호도 없음 → 평가 제외(카운트 안 함)

        stopped_today = False
        stop_ret_pct = np.nan
        if pos == 1 and sl_active and pd.notna(entry_price):
            if (price / entry_price - 1.0) <= -sl:
                ret_raw = price / entry_price - 1.0
                ret_net = (1 + ret_raw) * (1 - cost) - 1
                sum_daily -= cost
                equity = 1.0 + sum_daily
                n_trades += 1
                n_stop_triggered += 1
                if ret_net < worst_trade_so_far: worst_trade_so_far = ret_net  # ★ 최대 거래손실 갱신
                if ret_net > 0: n_wins += 1
                trades.append({
                    'trade_no': n_trades,
                    'entry_date': entry_date, 'entry_price': entry_price,
                    'exit_date': d, 'exit_price': price,
                    'days_held': i - entry_idx,
                    'gross_return_%': ret_raw * 100,
                    'net_return_%': ret_net * 100,
                    'cum_equity': equity, 'cum_return_%': sum_daily * 100,
                    'win': ret_net > 0,
                    'exit_reason': '손절매',
                    'stop_price': float(entry_price * (1 - sl)),
                    'exit_idx': i, 'entry_idx': entry_idx,
                })
                stop_ret_pct = ret_net * 100
                stopped_today = True
                pos = 0
                entry_price = np.nan; entry_date = pd.NaT; entry_idx = -1
                stop_price_cur = np.nan

        unr_eq = 1.0 + sum_daily
        if pos == 1 and pd.notna(entry_price):
            unr_pct = (price / entry_price - 1) * 100
            days_h = i - entry_idx
        else:
            unr_pct = np.nan
            days_h = 0

        if unr_eq > peak: peak = unr_eq
        cur_dd = unr_eq / peak - 1
        if cur_dd < mdd: mdd = cur_dd

        action = ''
        do_buy = False
        do_sell = False
        conflict = b_on and s_on
        if conflict and not stopped_today: n_conflicts += 1

        # ★ 변경3: 중립(둘 다 OFF) 공식 표시
        neutral_note = ''
        if (not b_on) and (not s_on) and not stopped_today:
            neutral_note = f"중립 [매수 {b_formula} / 매도 {s_formula}]"

        if stopped_today:
            action = f'⛔ 손절매 ({stop_ret_pct:+.2f}%, 종가청산)'
        elif (i + 1) < n:
            # ★ 충돌 처리 정책 (요청): '매도 우선'
            #   - 현금 중 매수신호인데 매도신호도 같이 뜨면(충돌) → 매수 보류 (잘못된 매수 상쇄)
            #   - 보유 중 매도신호면, 매수신호가 같이 떠도(충돌) → 청산 (올바른 매도는 안 상쇄)
            _sell_priority = globals().get('CONFLICT_SELL_PRIORITY', True)
            if pos == 0 and b_on:
                if conflict and (_sell_priority or s_strength > b_strength):
                    action = f'⚔ 충돌 → 매수 보류 (매도 우선, S={s_strength:.0%}/B={b_strength:.0%})'
                    n_conflict_sell_won += 1
                else:
                    do_buy = True
                    if conflict:
                        action = f'⚔ 충돌 B={b_strength:.0%}≥S={s_strength:.0%} → 매수'
                        n_conflict_buy_won += 1
                    else:
                        action = '매수신호 → 익일 매수'
            elif pos == 1 and s_on:
                if conflict and (not _sell_priority) and b_strength > s_strength:
                    action = f'⚔ 충돌 B={b_strength:.0%}>S={s_strength:.0%} → 청산 보류'
                    n_conflict_buy_won += 1
                else:
                    do_sell = True
                    if conflict:
                        action = f'⚔ 충돌 → 청산 (매도 우선, S={s_strength:.0%}/B={b_strength:.0%})'
                        n_conflict_sell_won += 1
                    else:
                        action = '매도신호 → 익일 청산'
            elif neutral_note:
                action = neutral_note   # ★ 중립일 때 공식 표시
        else:
            # ★ 마지막 날(실행일, 요청) — 다음날이 없어 '실행'은 못 하지만, 그날의 신호 판정
            #   (충돌/매수/매도/중립)을 일별 백테스트와 '똑같은 문구'로 표시한다.
            #   do_buy/do_sell·카운터·실제매매는 건드리지 않음(통계·결과 불변).
            _sp_last = globals().get('CONFLICT_SELL_PRIORITY', True)
            if pos == 0 and b_on:
                if conflict and (_sp_last or s_strength > b_strength):
                    action = f'⚔ 충돌 → 매수 보류 (매도 우선, S={s_strength:.0%}/B={b_strength:.0%})'
                elif conflict:
                    action = f'⚔ 충돌 B={b_strength:.0%}≥S={s_strength:.0%} → 매수'
                else:
                    action = '매수신호 → 익일 매수'
            elif pos == 1 and s_on:
                if conflict and (not _sp_last) and b_strength > s_strength:
                    action = f'⚔ 충돌 B={b_strength:.0%}>S={s_strength:.0%} → 청산 보류'
                elif conflict:
                    action = f'⚔ 충돌 → 청산 (매도 우선, S={s_strength:.0%}/B={b_strength:.0%})'
                else:
                    action = '매도신호 → 익일 청산'
            elif neutral_note:
                action = neutral_note

        rows.append({
            'date': d, 'close': price,
            'buy_count': b_count, 'buy_raw': b_raw, 'buy_on': b_on,
            'sell_count': s_count, 'sell_raw': s_raw, 'sell_on': s_on,
            'buy_formula': b_formula, 'sell_formula': s_formula,
            'buy_result': buy_result, 'sell_result': sell_result,   # ★ 변경4
            'position_pre': 1 if pos == 1 else 0,
            'entry_price': entry_price if pos == 1 else np.nan,
            'stop_price':  stop_price_cur if pos == 1 else np.nan,
            'days_held': days_h,
            'unrealized_pct': unr_pct,
            'realized_pct': stop_ret_pct if stopped_today else np.nan,
            'equity': unr_eq,
            'cum_return_pct': sum_daily * 100,
            'running_mdd_pct': worst_trade_so_far * 100,   # ★ 그 시점까지 최대 거래손실
            'action': action,
        })

        if do_buy:
            entry_price = float(cl.iloc[i + 1])
            entry_date  = dates[i + 1]; entry_idx = i + 1
            sum_daily -= cost; equity = 1.0 + sum_daily; pos = 1
            stop_price_cur = entry_price * (1 - sl) if sl_active else np.nan
        elif do_sell:
            ex_p = float(cl.iloc[i + 1]); ex_d = dates[i + 1]
            ret_raw = ex_p / entry_price - 1
            ret_net = (1 + ret_raw) * (1 - cost) - 1
            if price > 0:
                _dret = ex_p / price - 1.0
                sum_daily += _dret
                if _dret > 0.0: up_sum_daily += _dret
            sum_daily -= cost
            equity = 1.0 + sum_daily
            n_trades += 1
            if ret_net > 0: n_wins += 1
            if ret_net < worst_trade_so_far: worst_trade_so_far = ret_net  # ★ 최대 거래손실 갱신
            trades.append({
                'trade_no': n_trades,
                'entry_date': entry_date, 'entry_price': entry_price,
                'exit_date': ex_d, 'exit_price': ex_p,
                'days_held': (i + 1) - entry_idx,
                'gross_return_%': ret_raw * 100,
                'net_return_%': ret_net * 100,
                'cum_equity': equity, 'cum_return_%': sum_daily * 100,
                'win': ret_net > 0,
                'exit_reason': '매도신호',
                'stop_price': float(entry_price * (1 - sl)) if sl_active else np.nan,
                'exit_idx': i + 1, 'entry_idx': entry_idx,
            })
            pos = 0; entry_price = np.nan; entry_date = pd.NaT; entry_idx = -1
            stop_price_cur = np.nan

        pos_post[i] = pos
        prev_close = price

    daily = pd.DataFrame(rows)
    trades_df = pd.DataFrame(trades)

    for tr in trades:
        if tr['exit_reason'] == '손절매': continue
        ix = tr['exit_idx']
        if 0 <= ix < len(daily):
            daily.at[ix, 'realized_pct']   = tr['net_return_%']
            daily.at[ix, 'equity']         = tr['cum_equity']
            daily.at[ix, 'cum_return_pct'] = tr['cum_return_%']
            daily.at[ix, 'unrealized_pct'] = np.nan
            daily.at[ix, 'days_held']      = tr['days_held']
            ca = daily.at[ix, 'action']
            tag = f"청산 체결 ({tr['net_return_%']:+.2f}%)"
            daily.at[ix, 'action'] = tag if not ca else f"{ca} / {tag}"
    for tr in trades:
        ix = tr['entry_idx']
        if 0 <= ix < len(daily):
            ca = daily.at[ix, 'action']
            daily.at[ix, 'action'] = '매수 체결' if not ca else f"{ca} / 매수 체결"

    # ── 기존 balanced accuracy 계산 (그대로) ──
    close_vals = cl.values.astype(np.float64)
    safe_buy_arr, safe_sell_arr, eval_arr = _compute_safe_arrays(
        close_vals, horizon, dd_limit, ru_limit)
    if anchor_mode and anchor_safe_buy is not None and anchor_safe_sell is not None:
        safe_buy_arr, safe_sell_arr, eval_arr = _apply_anchor_correction(
            safe_buy_arr, safe_sell_arr, eval_arr,
            anchor_safe_buy, anchor_safe_sell)

    n_days_total = len(close_vals)
    n_eval = 0
    n_buy_correct = 0; n_sell_correct = 0
    n_buy_signal_on = 0; n_sell_signal_on = 0
    b_tp = b_fp = b_fn = b_tn = 0
    s_tp = s_fp = s_fn = s_tn = 0
    n_anchor_buy = 0; n_anchor_buy_caught = 0
    n_anchor_sell = 0; n_anchor_sell_caught = 0
    anchor_buy_diagnosis = []
    anchor_sell_diagnosis = []
    for i in range(n_days_total):
        if eval_arr[i] == 0: continue
        b_count_i = float(np.dot(buy_mat[i].astype(np.float64), buy_w))     # ★ 가중
        s_count_i = float(np.dot(sell_mat[i].astype(np.float64), sell_w))
        b_on_i = b_count_i >= vote_buy
        s_on_i = s_count_i >= vote_sell
        sb = safe_buy_arr[i] == 1
        ss = safe_sell_arr[i] == 1
        n_eval += 1
        if b_on_i: n_buy_signal_on += 1
        if s_on_i: n_sell_signal_on += 1
        if b_on_i and sb: b_tp += 1
        elif b_on_i and (not sb): b_fp += 1
        elif (not b_on_i) and sb: b_fn += 1
        else: b_tn += 1
        if s_on_i and ss: s_tp += 1
        elif s_on_i and (not ss): s_fp += 1
        elif (not s_on_i) and ss: s_fn += 1
        else: s_tn += 1
        if (b_on_i and sb) or ((not b_on_i) and (not sb)):
            n_buy_correct += 1
        if (s_on_i and ss) or ((not s_on_i) and (not ss)):
            n_sell_correct += 1
        if anchor_mode and anchor_safe_buy is not None:
            if anchor_safe_buy[i] == 1:
                n_anchor_buy += 1
                # ★ 타이밍 반영 (요청): 신호는 정답일보다 먼저 떠야 정답일에 포지션이 맞음.
                #   신호→익일매수 구조 + 앵커 정답일이 며칠에 걸친 바닥 구간인 점을 반영해,
                #   '정답일 당일 ~ +W일' 중 하루라도 실제 보유면 매칭.
                #   (충돌·실제 매매가 모두 반영된 pos_post 기준 — 신호만 보는 게 아님)
                _w = int(globals().get('ANCHOR_MATCH_WINDOW', 1))
                caught = any(pos_post[i+dd] == 1
                             for dd in range(0, _w+1) if i+dd < n_days_total)
                if caught: n_anchor_buy_caught += 1
                target_date = dates[i+1] if i+1 < n_days_total else dates[i]
                if caught:
                    status = '✓ 잡힘 (long 포지션 보유)'
                elif b_on_i and s_on_i:
                    status = f'✗ 매수신호 떴으나 매도가 더 강해 unmatched (B={b_count_i:.2f}/{vote_buy}, S={s_count_i:.2f}/{vote_sell})'
                elif b_on_i:
                    status = f'✗ 매수신호 떴으나 거래 안 됨 — 현재 cash'
                else:
                    status = f'✗ 신호 안 뜸 (buy {b_count_i:.2f} < {vote_buy})'
                anchor_buy_diagnosis.append({
                    'target_date': target_date, 'signal_day': dates[i],
                    'buy_count': b_count_i, 'vote_buy_required': vote_buy,
                    'signal_on': b_on_i, 'executed_match': caught, 'status': status,
                })
            if anchor_safe_sell[i] == 1:
                n_anchor_sell += 1
                # ★ 매도 매칭 = '고점 근처에서 팔았나' (요청, 엄격하되 현실적):
                #   - held_before: 고점 전 최근 LB일 중 보유한 적 있나 (고점까지 들고 왔나)
                #   - cash_after : 고점 당일~+W일 중 현금이 됐나 (고점에서 청산했나)
                #   고점 직전 1일만 보면 추세 중 잠깐 빠짐에도 실패 처리돼 너무 낮음 → LB로 완화.
                _w = int(globals().get('ANCHOR_MATCH_WINDOW', 1))
                _lb = int(globals().get('ANCHOR_SELL_HOLD_LOOKBACK', 5))
                held_before = any(pos_post[i-dd] == 1
                                  for dd in range(1, _lb+1) if i-dd >= 0)
                cash_after  = any(pos_post[i+dd] == 0
                                  for dd in range(0, _w+1) if i+dd < n_days_total)
                caught = held_before and cash_after
                if caught: n_anchor_sell_caught += 1
                target_date = dates[i+1] if i+1 < n_days_total else dates[i]
                if caught:
                    status = '✓ 잡힘 (고점 근처에서 청산)'
                elif not held_before:
                    status = '✗ 고점 전에 이미 청산 (스윙 일찍 빠짐) — 큰 상승 놓침'
                elif s_on_i and b_on_i:
                    status = f'✗ 매도신호 떴으나 매수가 더 강해 unmatched (S={s_count_i:.2f}/{vote_sell}, B={b_count_i:.2f}/{vote_buy})'
                elif s_on_i:
                    status = f'✗ 매도신호 떴으나 청산 안 됨'
                else:
                    status = f'✗ 신호 안 뜸 (sell {s_count_i:.2f} < {vote_sell})'
                anchor_sell_diagnosis.append({
                    'target_date': target_date, 'signal_day': dates[i],
                    'sell_count': s_count_i, 'vote_sell_required': vote_sell,
                    'signal_on': s_on_i, 'executed_match': caught, 'status': status,
                })

    b_tpr = b_tp / (b_tp + b_fn) if (b_tp + b_fn) > 0 else 0.0
    b_tnr = b_tn / (b_tn + b_fp) if (b_tn + b_fp) > 0 else 0.0
    s_tpr = s_tp / (s_tp + s_fn) if (s_tp + s_fn) > 0 else 0.0
    s_tnr = s_tn / (s_tn + s_fp) if (s_tn + s_fp) > 0 else 0.0
    buy_succ_rate  = (b_tpr + b_tnr) / 2.0
    sell_succ_rate = (s_tpr + s_tnr) / 2.0
    avg_succ_rate  = (buy_succ_rate + sell_succ_rate) / 2.0
    buy_acc_plain  = n_buy_correct  / n_eval if n_eval > 0 else 0.0
    sell_acc_plain = n_sell_correct / n_eval if n_eval > 0 else 0.0

    last = rows[-1]
    if pos == 1:
        cur = dict(
            position='보유 중 (LONG)', position_emoji='📈',
            entry_date=entry_date, entry_price=entry_price,
            stop_price=stop_price_cur,
            current_price=last['close'], days_held=last['days_held'],
            unrealized_pct=last['unrealized_pct'],
            equity=last['equity'], cum_return_pct=last['cum_return_pct'],
            buy_count_now=last['buy_count'], sell_count_now=last['sell_count'],
            buy_on_now=last['buy_on'],       sell_on_now=last['sell_on'],
        )
    else:
        cur = dict(
            position='현금 (CASH)', position_emoji='💵',
            entry_date=None, entry_price=None, stop_price=None,
            current_price=last['close'], days_held=0, unrealized_pct=None,
            equity=last['equity'], cum_return_pct=last['cum_return_pct'],
            buy_count_now=last['buy_count'], sell_count_now=last['sell_count'],
            buy_on_now=last['buy_on'],       sell_on_now=last['sell_on'],
        )
    cur['n_trades'] = n_trades; cur['n_wins'] = n_wins
    cur['win_rate'] = n_wins / n_trades if n_trades > 0 else 0.0
    cur['n_buy_evaluated']  = n_eval
    cur['n_buy_success']    = n_buy_correct
    cur['buy_success_rate'] = buy_succ_rate
    cur['n_sell_evaluated'] = n_eval
    cur['n_sell_success']   = n_sell_correct
    cur['sell_success_rate']= sell_succ_rate
    cur['avg_success_rate'] = avg_succ_rate
    cur['n_buy_signal_on']  = n_buy_signal_on
    cur['n_sell_signal_on'] = n_sell_signal_on
    cur['n_conflicts']      = n_conflicts
    cur['n_conflict_buy_won']  = n_conflict_buy_won
    cur['n_conflict_sell_won'] = n_conflict_sell_won
    cur['n_stop_triggered']    = n_stop_triggered
    cur['stop_loss_pct']       = float(stop_loss_pct) if sl_active else None
    cur['anchor_mode']         = bool(anchor_mode)
    cur['n_anchor_buy']        = n_anchor_buy
    cur['n_anchor_buy_caught'] = n_anchor_buy_caught
    cur['n_anchor_sell']       = n_anchor_sell
    cur['n_anchor_sell_caught']= n_anchor_sell_caught
    # ★ 실제 거래 기준 앵커 매칭률 (요청) — 비율로 저장
    cur['anchor_buy_match_rate']  = (n_anchor_buy_caught / n_anchor_buy) if n_anchor_buy > 0 else 0.0
    cur['anchor_sell_match_rate'] = (n_anchor_sell_caught / n_anchor_sell) if n_anchor_sell > 0 else 0.0
    cur['anchor_avg_match_rate']  = (cur['anchor_buy_match_rate'] + cur['anchor_sell_match_rate']) / 2.0
    # ★ 앵커 전략 수익 (요청) — '단기 최저점에서 매수, 단기 최고점에서 매도'로 얻을 수 있는
    #   사후적 최대 수익. 과거 차트가 다 있으므로, 거래비용을 넘어서는 모든 상승 구간을
    #   포착하는 최적 스윙 매매로 계산한다 (B&H보다 훨씬 높아야 정상).
    #   방법: 거래비용(cost) 이상 오르는 연속 구간만 매매 → 그 구간들의 상승률 합(복리).
    _cost = float(globals().get('COST_PER_TRADE', 0.0)) if 'cost' not in dir() else float(cost)
    _ap = cl.values.astype(np.float64)
    _equity = 1.0
    _i2 = 0
    _nn = len(_ap)
    while _i2 < _nn - 1:
        # 다음 거래일부터 '오르는 한' 계속 보유 (단기 저점→고점)
        if _ap[_i2+1] > _ap[_i2]:
            j = _i2
            # j가 국소 저점: 여기서 매수, 가격이 더 안 오를 때까지(국소 고점) 보유
            k = j
            while k < _nn - 1 and _ap[k+1] >= _ap[k]:
                k += 1
            # j에서 사서 k에서 팜 (단기 저점→단기 고점)
            gain = _ap[k] / _ap[j] - 1.0
            net = (1.0 - _cost) * (1.0 + gain) * (1.0 - _cost) - 1.0  # 매수·매도 비용 차감
            if net > 0:
                _equity *= (1.0 + net)
            _i2 = k + 1
        else:
            _i2 += 1
    cur['anchor_strategy_return'] = _equity - 1.0   # 사후 최적 스윙 누적수익(복리, 소수)
    cur['anchor_buy_diagnosis']  = anchor_buy_diagnosis
    cur['anchor_sell_diagnosis'] = anchor_sell_diagnosis
    cur['buy_accuracy_plain']  = buy_acc_plain
    cur['sell_accuracy_plain'] = sell_acc_plain
    cur['horizon']  = horizon
    cur['dd_limit'] = dd_limit
    cur['ru_limit'] = ru_limit
    cur['last_date'] = last['date']
    cur['K_buy'] = K_buy; cur['K_sell'] = K_sell
    cur['vote_buy'] = vote_buy; cur['vote_sell'] = vote_sell
    # ★ 최대 낙폭 정의 변경 (요청) — 자산곡선 고점대비 하락이 아니라,
    #   '개별 일별 거래에서 발생한 최대 손실'을 최대 낙폭으로 사용.
    #   (처음 투자 시작하는 입장에선 쌓아둔 이익이 없어, 한 거래의 손실이 곧 계좌 위험)
    #   거래가 1건도 없으면 0.0.
    if len(trades) > 0:
        _trade_rets = [t['net_return_%'] / 100.0 for t in trades]   # 각 거래 실현손익(소수)
        worst_trade = min(_trade_rets)                # 가장 큰 손실 (음수)
        max_trade_loss = worst_trade if worst_trade < 0 else 0.0
    else:
        max_trade_loss = 0.0
    cur['max_drawdown']      = max_trade_loss   # ★ 이제 '최대 거래 손실' 의미
    cur['equity_mdd']        = mdd              # (참고) 자산곡선 고점대비 하락 — 보존
    cur['max_trade_loss']    = max_trade_loss   # 명시적 키
    cur['up_cum_return_pct'] = up_sum_daily * 100.0
    cur['weighted_vote'] = weighted     # ★ 변경2
    # ★ 변경4: 전체 매수/매도 성공·실패 집계
    cur['n_buy_on_total']  = n_buy_on_total
    cur['n_buy_success_cnt'] = n_buy_success
    cur['n_buy_fail_cnt']    = n_buy_fail
    cur['n_sell_on_total'] = n_sell_on_total
    cur['n_sell_success_cnt'] = n_sell_success
    cur['n_sell_fail_cnt']    = n_sell_fail
    cur['buy_signal_hit_rate']  = n_buy_success  / n_buy_on_total  if n_buy_on_total  > 0 else 0.0
    cur['sell_signal_hit_rate'] = n_sell_success / n_sell_on_total if n_sell_on_total > 0 else 0.0
    # ★ 정답일 적중률 (정답인 날만 분모. ON적중=성공, OFF=놓침(실패). 안정답인 날 제외)
    cur['n_buy_eval_all']     = n_buy_eval_all
    cur['n_buy_correct_all']  = n_buy_correct_all
    cur['n_sell_eval_all']    = n_sell_eval_all
    cur['n_sell_correct_all'] = n_sell_correct_all
    cur['buy_acc_all']  = n_buy_correct_all  / n_buy_eval_all  if n_buy_eval_all  > 0 else 0.0
    cur['sell_acc_all'] = n_sell_correct_all / n_sell_eval_all if n_sell_eval_all > 0 else 0.0
    # ★ feature3 (요청) — 일별 시트에 '사용된 모든 지표'의 그날 수치 + 신호(0/1) 부착.
    #   날짜 기준 정렬(길이 달라도 안전). cur에 (종류,지표명,수치컬럼,신호컬럼) 목록 저장.
    try:
        _used_ind_cols = []
        _fidx = feat.index
        for _k, (_, _br) in enumerate(buy_used.iterrows()):
            _nm = str(_br['indicator'])
            _vc = f'_INDV_B{_k}'; _sc = f'_INDS_B{_k}'
            daily[_vc] = daily['date'].map(pd.Series(feat[_nm].values, index=_fidx))
            daily[_sc] = daily['date'].map(pd.Series(buy_mat[:, _k], index=_fidx))
            _used_ind_cols.append(('BUY', _nm, _vc, _sc))
        for _k, (_, _sr) in enumerate(sell_used.iterrows()):
            _nm = str(_sr['indicator'])
            _vc = f'_INDV_S{_k}'; _sc = f'_INDS_S{_k}'
            daily[_vc] = daily['date'].map(pd.Series(feat[_nm].values, index=_fidx))
            daily[_sc] = daily['date'].map(pd.Series(sell_mat[:, _k], index=_fidx))
            _used_ind_cols.append(('SELL', _nm, _vc, _sc))
        cur['used_ind_cols'] = _used_ind_cols
    except Exception as _eic:
        cur['used_ind_cols'] = []
        print(f"  ⚠ 일별 지표컬럼 부착 실패(무시): {_eic}")
    return daily, trades_df, cur, buy_used, sell_used


# ════════════════════════════════════════════════════════════════
#                       Excel 출력
# ════════════════════════════════════════════════════════════════
_HDR  = PatternFill("solid", fgColor="1F3864")
_GOOD = PatternFill("solid", fgColor="C6EFCE")
_MID  = PatternFill("solid", fgColor="FFEB9C")
_BAD  = PatternFill("solid", fgColor="FFC7CE")
_ALT  = PatternFill("solid", fgColor="F2F2F2")
_HL   = PatternFill("solid", fgColor="FFF2CC")
_SIGY = PatternFill("solid", fgColor="FFFF00")   # ★ 지표 신호 충족 = 노란색 (요청)
_BUY  = PatternFill("solid", fgColor="C6EFCE")
_SELL = PatternFill("solid", fgColor="FFC7CE")
_HOLD = PatternFill("solid", fgColor="DDEBF7")
_CASH = PatternFill("solid", fgColor="F5F5F5")
_CONF = PatternFill("solid", fgColor="FFE699")
_BAND = PatternFill("solid", fgColor="FFF7E0")
_WB_  = Font(bold=True, color='FFFFFF')
_TH   = Border(left=Side('thin', color='BDBDBD'), right=Side('thin', color='BDBDBD'),
               top=Side('thin', color='BDBDBD'), bottom=Side('thin', color='BDBDBD'))


def _hdr(ws, row, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row, ci); c.value = h; c.fill = _HDR; c.font = _WB_
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = _TH


def _ret_fill(ret, bh_ret):
    if ret > bh_ret + 0.10: return _GOOD
    if ret > bh_ret:         return _MID
    if ret > 0:              return _MID
    return _BAD


def _mdd_fill(mdd, mdd_limit_pct):
    if mdd_limit_pct is None:
        if mdd > -0.05: return _GOOD
        if mdd > -0.15: return _MID
        return _BAD
    limit = -abs(mdd_limit_pct)
    if mdd >= limit / 2: return _GOOD
    if mdd >= limit:     return _MID
    return _BAD


def _success_fill(rate):
    if rate >= 0.7: return _GOOD
    if rate >= 0.5: return _MID
    return _BAD


def _norm_date_set(dlist):
    s = set()
    if dlist is None: return s
    for d in dlist:
        try: s.add(pd.Timestamp(d).normalize())
        except Exception: pass
    return s


def _net_signal_k_search(feat, close_ser, buy_pool, sell_pool, *,
                         ticker='', max_k_candidates=600, oos_start=None,
                         n_buy=None, n_sell=None, search_counts=False, weight_exp=1.0, fixed_k=None,
                         select_by='full'):
    """★ 순신호 K 최적화 + OOS 검증 (요청):
       b(순신호) = (상위 n_buy 매수지표 가중합) − (상위 n_sell 매도지표 가중합).
       net > K → 롱(매수/보유), net ≤ K → 현금(매도). 신호는 다음날 반영.
       ★ (n_buy, n_sell, K) 모두 'OOS 수익 최고'로 선택(요청). OOS 없으면 학습수익 폴백.
       - search_counts=True: 지표개수(n_buy,n_sell)까지 탐색 → 최적 개수 자동 선정.
       - n_buy/n_sell 지정: 그 개수로 고정(그리드 행별 K_buy/K_sell 적용용).
       (수익=상승·하락률 단순 합산, 손절 미적용)
    """
    if feat is None or close_ser is None or buy_pool is None or sell_pool is None:
        return None
    try:
        dates = list(feat.index)
        n = len(dates)
        if n < 10 or len(buy_pool) == 0 or len(sell_pool) == 0:
            return None
        close = pd.Series(close_ser).reindex(dates).values.astype(float)

        _wtd = bool(globals().get('NET_SIGNAL_WEIGHTED', False))
        _wcol = str(globals().get('NET_SIGNAL_WEIGHT_COL', 'success_rate'))
        _net_is_float = _wtd
        _wexp = float(weight_exp) if weight_exp else 1.0
        def _wt_of(row):
            if not _wtd: return 1.0
            try:
                v = row.get(_wcol)
                if v is None or (isinstance(v, float) and np.isnan(v)): return 1.0
                v = float(v)
                return (v ** _wexp) if _wexp != 1.0 else v   # ★ 가중 스킴: 성공률^g
            except Exception:
                return 1.0
        # 지표별 (가중) 신호 배열 — 풀 순서(성공률 우선) 유지
        # ★ 요청 ②: 한 지표가 임계별 여러 성공률 보유 시, '그날 켜진 것 중 최고 성공률'을 그 지표 점수로.
        #   (단일임계 풀이면 지표당 1행 → 기존과 동일하게 동작. 안전)
        _multi = bool(globals().get('NET_MULTI_THRESHOLD_WEIGHT', False))
        def _build_sigs(pool):
            sigs = []
            if _multi and ('indicator' in pool.columns) and pool['indicator'].duplicated().any():
                for _ind, grp in pool.groupby('indicator', sort=False):
                    arr = np.zeros(n)
                    for _, row in grp.iterrows():
                        try:
                            s = np.nan_to_num(_to_signal_array(feat, row).astype(float))
                            arr = np.maximum(arr, _wt_of(row) * s)   # 켜진 임계 중 최고 가중
                        except Exception: pass
                    sigs.append(arr)
            else:
                for _, row in pool.iterrows():
                    try: sigs.append(_wt_of(row) * np.nan_to_num(_to_signal_array(feat, row).astype(float)))
                    except Exception: pass
            return sigs
        buy_sigs  = _build_sigs(buy_pool)
        sell_sigs = _build_sigs(sell_pool)
        nB = len(buy_sigs); nS = len(sell_sigs)
        if nB == 0 or nS == 0: return None
        buy_cum  = np.cumsum(np.array(buy_sigs),  axis=0)   # buy_cum[k-1] = 상위 k개 합
        sell_cum = np.cumsum(np.array(sell_sigs), axis=0)
        def _net_for(nb, ns):
            nb = max(1, min(int(nb), nB)); ns = max(1, min(int(ns), nS))
            return buy_cum[nb-1] - sell_cum[ns-1], nb, ns

        r = np.zeros(n)
        for t in range(1, n):
            p0 = close[t-1]; p1 = close[t]
            if p0 and p0 > 0 and not np.isnan(p1) and not np.isnan(p0):
                r[t] = p1 / p0 - 1.0

        # ── OOS 분할 ──
        oos_idx = n
        if oos_start is not None:
            try:
                _os = pd.Timestamp(oos_start).normalize()
                for t in range(n):
                    if pd.Timestamp(dates[t]).normalize() >= _os:
                        oos_idx = t; break
            except Exception:
                oos_idx = n
        has_oos = (10 <= oos_idx < n - 1)
        if not has_oos: oos_idx = n
        train_hi = oos_idx if has_oos else n

        # ── threshold(K) 탐색 (벡터화) ── 기존 방식: net[s-1]>K면 그날(s) 포지션 → ret[s] 포착.
        def _search_threshold(net):
            net_prev = np.empty(n); net_prev[0] = net[0]; net_prev[1:] = net[:-1]   # 1일 지연(기존)
            kmin = float(np.nanmin(net)); kmax = float(np.nanmax(net))
            if kmax <= kmin: kmax = kmin + 1.0
            if _net_is_float:
                _grid = np.linspace(kmin, kmax, min(max_k_candidates, 250))
                _qs = np.unique(np.nanpercentile(net, np.linspace(1, 99, 50)))
                ks = sorted(set(np.round(np.concatenate([_grid, _qs]), 4).tolist()))
            else:
                _kl = int(np.floor(kmin)); _kh = int(np.ceil(kmax))
                ks = (sorted(set(np.linspace(_kl, _kh, max_k_candidates).astype(int).tolist()))
                      if (_kh - _kl) > max_k_candidates else list(range(_kl, _kh + 1)))
            best = None; best_sel = None; table = []
            for K in ks:
                pos = (net_prev > K).astype(float); pos[0] = 0.0
                _hr = pos * r
                tr = float(np.sum(pos[:train_hi] * r[:train_hi]))
                oo = float(np.sum(pos[oos_idx:] * r[oos_idx:])) if has_oos else None
                fu = float(np.sum(_hr))
                hd = float(np.sum(_hr[_hr < 0]))
                hd_oos = (float(np.sum(_hr[oos_idx:][_hr[oos_idx:] < 0])) if has_oos else None)
                dl = int(pos[:train_hi].sum())
                _Kv = (round(float(K), 4) if _net_is_float else int(K))
                table.append((_Kv, tr, oo, fu, dl, hd, hd_oos))
                _sel = (oo if (select_by == 'oos' and has_oos and oo is not None) else fu)   # 전체 or OOS 기준
                if best_sel is None or _sel > best_sel:
                    best_sel = _sel; best = (_Kv, tr, oo, fu, dl)
            return best, table, (best_sel if best_sel is not None else -1e18), (kmin, kmax)

        # ── 지표개수(n_buy, n_sell) 후보 ──
        if search_counts:
            def _cand(mx):
                return list(range(1, int(mx) + 1))   # 1~풀크기(최대 100) 전부 탐색
            nb_list = _cand(nB); ns_list = _cand(nS)
        elif (n_buy is not None) or (n_sell is not None):
            nb_list = [int(n_buy) if n_buy else nB]; ns_list = [int(n_sell) if n_sell else nS]
        else:
            nb_list = [nB]; ns_list = [nS]

        overall = None
        def _eval(nb, ns, ov):
            net_c, nbu, nsu = _net_for(nb, ns)
            best_c, table_c, sel_c, krng = _search_threshold(net_c)
            if ov is None or sel_c > ov['sel']:
                ov = dict(sel=sel_c, best=best_c, table=table_c, net=net_c,
                          nb=nbu, ns=nsu, krng=krng)
            return ov, sel_c
        if search_counts and len(nb_list) * len(ns_list) > 400:
            # ★ 독립(좌표하강) 탐색 — 100×100=10000 → (100+100+100)≈300. 1~100 범위 그대로.
            _bns = ns_list[len(ns_list) // 2]        # ns 초기값(중앙)
            _best_nb = nb_list[0]; _bsc = -1e18
            for nb in nb_list:                       # ① nb 훑기 (ns 고정)
                overall, _sc = _eval(nb, _bns, overall)
                if _sc > _bsc: _bsc = _sc; _best_nb = nb
            _best_ns = _bns; _bsc = -1e18
            for ns in ns_list:                       # ② ns 훑기 (best nb 고정)
                overall, _sc = _eval(_best_nb, ns, overall)
                if _sc > _bsc: _bsc = _sc; _best_ns = ns
            for nb in nb_list:                       # ③ nb 재훑기 (best ns 고정) — 마무리
                overall, _sc = _eval(nb, _best_ns, overall)
        else:
            for nb in nb_list:
                for ns in ns_list:
                    overall, _ = _eval(nb, ns, overall)

        net = overall['net']; best = overall['best']; table = overall['table']
        n_buy_opt = overall['nb']; n_sell_opt = overall['ns']
        kmin, kmax = overall['krng']
        buy_count = buy_cum[n_buy_opt-1]; sell_count = sell_cum[n_sell_opt-1]

        best_k = best[0]
        if fixed_k is not None:      # ★ 재현: 원본 K 그대로 (탐색 안 함)
            best_k = float(fixed_k)
        # ★ 매수·매도 둘 다 '신호 본 날 종가'에 체결 (요청 A):
        #   net[s]>K 본 날(s) 종가에 매수 → 소유 pos_own[s]=net[s]>K.  net[s]≤K 본 날 종가에 매도.
        #   수익은 '전일 소유분'이 당일 등락 포착: daily_ret[s]=pos_own[s-1]*r[s].
        #   → 진입가=매수신호일 종가, 청산가=매도신호일 종가, 실현%=(청산가/진입가-1) 로 정확히 일치.
        pos_own = (net > best_k).astype(float); pos_own[0] = 0.0   # 표시·체결일·체결가
        pos_ret = np.zeros(n); pos_ret[1:] = pos_own[:-1]          # 수익 포착(전일 소유)
        pos = pos_own
        bh_full  = float(np.sum(r)); bh_train = float(np.sum(r[:train_hi]))
        bh_oos   = float(np.sum(r[oos_idx:])) if has_oos else None
        n_trades = int(np.sum(np.abs(np.diff(pos_own)) > 0)) if n > 1 else 0
        n_trades_oos = int(np.sum(np.abs(np.diff(pos_own[oos_idx:])) > 0)) if (has_oos and n - oos_idx > 1) else 0
        _held_r = pos_ret * r
        held_down_full  = float(np.sum(_held_r[_held_r < 0]))
        held_down_train = float(np.sum(_held_r[:train_hi][_held_r[:train_hi] < 0]))
        held_down_oos   = (float(np.sum(_held_r[oos_idx:][_held_r[oos_idx:] < 0])) if has_oos else None)
        daily_ret = pos_ret * r
        cum_run = np.cumsum(daily_ret)
        held_down_run = np.cumsum(np.where(daily_ret < 0, daily_ret, 0.0))
        action = []
        for t in range(n):
            if t == 0: action.append('—'); continue
            if pos_own[t] == 1 and pos_own[t-1] == 0:   action.append('매수')   # 신호 본 날 종가 체결
            elif pos_own[t] == 0 and pos_own[t-1] == 1: action.append('매도')
            elif pos_own[t] == 1:                        action.append('보유')
            else:                                        action.append('현금')

        oos_flag = np.zeros(n, dtype=int)
        if has_oos: oos_flag[oos_idx:] = 1
        daily = pd.DataFrame({
            'date': dates, 'price': close,
            'buy_count': (np.round(buy_count, 3) if _net_is_float else buy_count.astype(int)),
            'sell_count': (np.round(sell_count, 3) if _net_is_float else sell_count.astype(int)),
            'net': (np.round(net, 4) if _net_is_float else net.astype(int)),
            'position': pos.astype(int), 'is_oos': oos_flag,
            'day_ret': daily_ret, 'cum_ret': cum_run,
            'held_down_run': held_down_run, 'action': action,
        })
        return {
            'best_k': best_k, 'n_buy_opt': n_buy_opt, 'n_sell_opt': n_sell_opt,
            'train_cum': best[1], 'oos_cum': best[2], 'full_cum': best[3],
            'best_days_long': best[4], 'n_trades': n_trades, 'n_trades_oos': n_trades_oos,
            'held_down_full': held_down_full, 'held_down_train': held_down_train,
            'held_down_oos': held_down_oos,
            'bh_train': bh_train, 'bh_oos': bh_oos, 'bh_full': bh_full,
            'has_oos': has_oos, 'oos_idx': oos_idx,
            'oos_start_date': (pd.Timestamp(dates[oos_idx]).strftime('%Y-%m-%d') if has_oos else None),
            'k_table': table, 'daily': daily,
            'buy_pool_n': nB, 'sell_pool_n': nS,
            'net_min': (round(float(kmin),2) if _net_is_float else int(kmin)),
            'net_max': (round(float(kmax),2) if _net_is_float else int(kmax)),
            'weighted': _net_is_float, 'searched_counts': bool(search_counts),
        }
    except Exception as _e:
        print(f"  ⚠ 순신호 K 최적화 실패(무시): {_e}")
        return None


def _net_kl_search(net, r, *, mdd_limit=None, k_grid=None, fixed_kl=None, fixed_kl_mdd=None):
    """★ K/L 2임계 히스테리시스. net_prev≥K → 매수(롱), net_prev≤L → 매도(현금), 사이는 직전 포지션 유지 (K≥L).
       (K,L) 격자를 훑어 ① 전체수익 최고 ② MDD 한도 지키는 수익 최고 를 둘 다 찾는다.
       fixed_kl=(K,L) 주면 탐색 없이 그 값 그대로(재현). fixed_kl_mdd=(K,L)은 ②행 재현용.
       반환: dict{best_ret:(K,L,ret,mdd,dl,daily_pos), best_mdd:(...) or None, grid_n}."""
    net = np.asarray(net, float); r = np.asarray(r, float)
    n = len(net)
    net_prev = np.empty(n); net_prev[0] = net[0]; net_prev[1:] = net[:-1]

    def _run(K, L):
        pos = np.zeros(n); cur = 0.0
        for s in range(n):
            v = net_prev[s]
            if v >= K: cur = 1.0
            elif v <= L: cur = 0.0
            pos[s] = cur
        pos[0] = 0.0
        hr = pos * r
        ret = float(np.sum(hr))
        cum = np.cumsum(hr); run_max = np.maximum.accumulate(cum)
        mdd = float(np.min(cum - run_max)) if n else 0.0
        return ret, mdd, int(pos.sum()), pos

    if fixed_kl is not None:      # ★ 재현: 원본 (K,L) 그대로
        _K, _L = float(fixed_kl[0]), float(fixed_kl[1])
        ret, mdd, dl, pos = _run(_K, _L)
        _br = (_K, _L, ret, mdd, dl, pos)
        _bm = None
        if fixed_kl_mdd is not None:
            _Km, _Lm = float(fixed_kl_mdd[0]), float(fixed_kl_mdd[1])
            r2, m2, d2, p2 = _run(_Km, _Lm)
            _bm = (_Km, _Lm, r2, m2, d2, p2)
        return {'best_ret': _br, 'best_mdd': _bm, 'grid_n': 1,
                'top': [{'K': _br[0], 'L': _br[1], 'ret': _br[2], 'mdd': _br[3], 'dl': _br[4], 'pos': _br[5]}]}

    if k_grid is None:
        lo = float(np.nanmin(net)); hi = float(np.nanmax(net))
        if hi <= lo: hi = lo + 1.0
        _is_int = np.allclose(net, np.round(net))
        if _is_int:
            k_grid = list(range(int(np.floor(lo)), int(np.ceil(hi)) + 1))
        else:
            k_grid = sorted(set(np.round(np.concatenate([
                np.linspace(lo, hi, 60), np.nanpercentile(net, np.linspace(1, 99, 40))]), 4).tolist()))
    kg = sorted(k_grid)

    best_ret = None; best_mdd = None; cnt = 0; _all = []
    for K in kg:
        for L in kg:
            if L > K:
                continue
            cnt += 1
            ret, mdd, dl, pos = _run(K, L)
            _all.append((ret, mdd, dl, K, L, pos))
            if best_ret is None or ret > best_ret[2]:
                best_ret = (K, L, ret, mdd, dl, pos)
            if mdd_limit is not None and mdd >= -abs(mdd_limit):
                if best_mdd is None or ret > best_mdd[2]:
                    best_mdd = (K, L, ret, mdd, dl, pos)
    # 상위 (K,L) 조합 순위표 (수익 내림차순, 최대 40개)
    _all.sort(key=lambda x: x[0], reverse=True)
    top = [{'K': a[3], 'L': a[4], 'ret': a[0], 'mdd': a[1], 'dl': a[2], 'pos': a[5]} for a in _all[:40]]
    return {'best_ret': best_ret, 'best_mdd': best_mdd, 'grid_n': cnt, 'top': top}


def _write_indicator_matrix_sheet(ws, pool, feat, close_ser,
                                   anchor_buy_set, anchor_sell_set,
                                   ticker, kind_label):
    """★ feature2 (요청) — 지표 신호 매트릭스 시트.
       행=날짜, 열=[날짜, 종가] + 풀의 각 지표.
       지표 셀: 그날 수치 + 신호 충족 시 노란색. 행: 앵커 매수일=초록 / 매도일=빨강."""
    if pool is None or len(pool) == 0 or feat is None:
        ws.cell(1, 1).value = f'{ticker} — {kind_label} 지표 신호 (풀 없음)'
        return
    dates = feat.index
    inds  = [str(r['indicator']) for _, r in pool.iterrows()]
    # 각 지표의 신호배열(0/1) + 수치
    sig_cols  = []
    feat_vals = {}
    for _, prow in pool.iterrows():
        nm = str(prow['indicator'])
        try:    sig_cols.append(np.asarray(_to_signal_array(feat, prow)).astype(np.uint8))
        except Exception: sig_cols.append(np.zeros(len(dates), dtype=np.uint8))
        feat_vals[nm] = feat[nm].values if nm in feat.columns else np.full(len(dates), np.nan)
    ws.cell(1, 1).value = (f'{ticker} — {kind_label} 지표 신호 매트릭스 '
                           f'({len(inds)}개 지표 × {len(dates)}일)  '
                           f'｜ 노랑=신호(마지막날 제외·평가불가), 행 초록=다음날↑(매수자리), 빨강=다음날↓(매도자리)')
    ws.cell(1, 1).font = Font(bold=True, size=12, color='1F3864')
    _hdr(ws, 3, ['날짜', f'{ticker}종가'] + inds)
    close_vals = (close_ser.reindex(dates).values if close_ser is not None
                  else np.full(len(dates), np.nan))
    for di in range(len(dates)):
        r = di + 4
        dt = dates[di]
        # ★ 행 색(요청): 앵커 무시 — '신호 후 horizon일 이내' 종가 변동으로 색칠.
        #   horizon일 이내 최고가가 +DRAWDOWN_LIMIT_BUY 이상 → 이 날이 매수자리 → 초록
        #   horizon일 이내 최저가가 -RUNUP_LIMIT_SELL 이상 하락 → 매도자리 → 빨강
        #   (= 그 날 신호의 정답. horizon=1이면 다음날, 5면 5일 이내)
        _ddb = float(globals().get('DRAWDOWN_LIMIT_BUY', 0.01))
        _rus = float(globals().get('RUNUP_LIMIT_SELL', 0.01))
        _hz = int(globals().get('HORIZON_DAYS', 1)); _hz = _hz if _hz >= 1 else 1
        row_fill = None
        if di < len(dates) - 1 and pd.notna(close_vals[di]) and close_vals[di] > 0:
            _base = close_vals[di]
            _end = di + _hz
            if _end > len(dates) - 1: _end = len(dates) - 1
            _mx = -1.0e18; _mn = 1.0e18
            for _j in range(di + 1, _end + 1):
                if pd.notna(close_vals[_j]):
                    _r = close_vals[_j] / _base - 1.0
                    if _r > _mx: _mx = _r
                    if _r < _mn: _mn = _r
            if _mx >= _ddb:
                row_fill = _BUY      # horizon 내 +상승 도달 → 매수자리 → 초록
            elif _mn <= -_rus:
                row_fill = _SELL     # horizon 내 -하락 도달 → 매도자리 → 빨강
        c = ws.cell(r, 1)
        c.value = dt.date() if hasattr(dt, 'date') else dt
        c.number_format = 'YYYY-MM-DD'; c.font = Font(size=8); c.border = _TH
        c.alignment = Alignment(horizontal='center')
        if row_fill is not None: c.fill = row_fill
        c = ws.cell(r, 2)
        if pd.notna(close_vals[di]):
            c.value = round(float(close_vals[di]), 2); c.number_format = '#,##0.00'
        c.font = Font(size=8); c.border = _TH; c.alignment = Alignment(horizontal='right')
        if row_fill is not None: c.fill = row_fill
        for k, nm in enumerate(inds):
            cc = ws.cell(r, 3 + k)
            v = feat_vals[nm][di]
            if pd.notna(v):
                cc.value = round(float(v), 4); cc.number_format = '0.0000'
            cc.font = Font(size=8); cc.border = _TH; cc.alignment = Alignment(horizontal='right')
            if sig_cols[k][di] == 1 and di < len(dates) - 1:
                cc.fill = _SIGY            # 신호 충족 → 노란색 (마지막 날은 다음날 없어 평가불가 → 제외, 앙상블과 일치)
            elif row_fill is not None:
                cc.fill = row_fill
    widths = [12, 11] + [13] * len(inds)
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'C4'; ws.row_dimensions[3].height = 70


def _write_daily_rows(ws, daily, cur, mdd_limit_pct):
    """일별 행 작성 (일별 시트 / OOS 시트 공통).
       컬럼: 날짜,종가,매수카운트,매수ON,매수성공,매도카운트,매도ON,매도성공,
            포지션,액션,진입가,손절가,보유일,미실현%,실현%,누적자산,누적수익%,진행최대손실%,
            매수공식,매도공식 (총 20열) — 액션을 포지션 다음으로 이동(요청)"""
    for ri, row in daily.iterrows():
        r = ri + 5
        c = ws.cell(r, 1); c.value = row['date'].date()
        c.number_format = 'YYYY-MM-DD'
        c.font = Font(bold=True, size=9, color='1F3864')
        c.alignment = Alignment(horizontal='center'); c.border = _TH
        c = ws.cell(r, 2); c.value = round(row['close'], 2)
        c.number_format = '#,##0.00'
        c.alignment = Alignment(horizontal='right'); c.border = _TH
        c = ws.cell(r, 3); c.value = round(float(row['buy_count']), 2)
        c.alignment = Alignment(horizontal='center'); c.border = _TH
        c.font = Font(size=10, bold=bool(row['buy_on']))
        if row['buy_on']: c.fill = _BUY
        c = ws.cell(r, 4)
        if row['buy_on']:
            c.value = 'ON'; c.fill = _BUY; c.font = Font(bold=True, size=10, color='006100')
        else:
            c.value = '-'; c.font = Font(size=9, color='888888')
        c.alignment = Alignment(horizontal='center'); c.border = _TH
        # 매수성공 (변경4)
        c = ws.cell(r, 5)
        br = str(row.get('buy_result', ''))
        c.value = br if br else '-'
        if '✓' in br:        # ✓성공(ON적중)
            c.fill = _GOOD; c.font = Font(bold=True, size=10, color='006100')
        elif '✗' in br:      # ✗실패(ON오답) 또는 ✗놓침(정답인데 신호없음)
            c.fill = _BAD;  c.font = Font(bold=True, size=10, color='C00000')
        else:
            c.font = Font(size=9, color='888888')
        c.alignment = Alignment(horizontal='center'); c.border = _TH
        c = ws.cell(r, 6); c.value = round(float(row['sell_count']), 2)
        c.alignment = Alignment(horizontal='center'); c.border = _TH
        c.font = Font(size=10, bold=bool(row['sell_on']))
        if row['sell_on']: c.fill = _SELL
        c = ws.cell(r, 7)
        if row['sell_on']:
            c.value = 'ON'; c.fill = _SELL; c.font = Font(bold=True, size=10, color='C00000')
        else:
            c.value = '-'; c.font = Font(size=9, color='888888')
        c.alignment = Alignment(horizontal='center'); c.border = _TH
        # 매도성공 (변경4)
        c = ws.cell(r, 8)
        sr = str(row.get('sell_result', ''))
        c.value = sr if sr else '-'
        if '✓' in sr:
            c.fill = _GOOD; c.font = Font(bold=True, size=10, color='006100')
        elif '✗' in sr:
            c.fill = _BAD;  c.font = Font(bold=True, size=10, color='C00000')
        else:
            c.font = Font(size=9, color='888888')
        c.alignment = Alignment(horizontal='center'); c.border = _TH
        c = ws.cell(r, 9)
        if row['position_pre'] == 1:
            c.value = '보유'; c.fill = _HOLD; c.font = Font(bold=True, size=9, color='C00000')
        else:
            c.value = '현금'; c.fill = _CASH; c.font = Font(size=9, color='606060')
        c.alignment = Alignment(horizontal='center'); c.border = _TH
        # ★ 액션 (포지션 다음으로 이동)
        c = ws.cell(r, 10)
        if row['action']:
            c.value = row['action']
            if '손절매' in row['action']:
                c.fill = _BAD; c.font = Font(bold=True, size=10, color='C00000')
            elif '충돌' in row['action']:
                c.fill = _CONF; c.font = Font(bold=True, size=9, color='806000')
            elif '중립' in row['action']:
                c.font = Font(size=8, color='888888')
            elif '매수' in row['action']:
                c.fill = _BUY; c.font = Font(bold=True, size=9, color='006100')
            elif '매도' in row['action'] or '청산' in row['action']:
                c.fill = _SELL; c.font = Font(bold=True, size=9, color='C00000')
        c.alignment = Alignment(horizontal='left'); c.border = _TH
        c = ws.cell(r, 11)
        if pd.notna(row['entry_price']):
            c.value = round(row['entry_price'], 2); c.number_format = '#,##0.00'
        c.alignment = Alignment(horizontal='right'); c.border = _TH; c.font = Font(size=9)
        c = ws.cell(r, 12)
        if pd.notna(row.get('stop_price', np.nan)):
            c.value = round(float(row['stop_price']), 2); c.number_format = '#,##0.00'
            c.font = Font(size=9, color='C00000', bold=True)
        c.alignment = Alignment(horizontal='right'); c.border = _TH
        c = ws.cell(r, 13)
        c.value = int(row['days_held']) if row['days_held'] > 0 else ''
        c.alignment = Alignment(horizontal='center'); c.border = _TH; c.font = Font(size=9)
        c = ws.cell(r, 14)
        if pd.notna(row['unrealized_pct']):
            v = float(row['unrealized_pct'])
            c.value = round(v, 2); c.number_format = '0.00"%"'
            c.font = Font(size=9, color='006100' if v > 0 else 'C00000')
        c.alignment = Alignment(horizontal='right'); c.border = _TH
        c = ws.cell(r, 15)
        if pd.notna(row['realized_pct']):
            v = float(row['realized_pct'])
            c.value = round(v, 2); c.number_format = '0.00"%"'
            c.font = Font(bold=True, size=10, color='006100' if v > 0 else 'C00000')
        c.alignment = Alignment(horizontal='right'); c.border = _TH
        c = ws.cell(r, 16); c.value = round(row['equity'], 4)
        c.number_format = '0.0000'
        c.alignment = Alignment(horizontal='right'); c.border = _TH; c.font = Font(size=9)
        c = ws.cell(r, 17)
        v = float(row['cum_return_pct'])
        c.value = round(v, 2); c.number_format = '0.00"%"'
        c.alignment = Alignment(horizontal='right'); c.border = _TH
        c.font = Font(bold=True, size=9, color='006100' if v > 0 else 'C00000')
        c = ws.cell(r, 18)
        v = float(row['running_mdd_pct'])
        c.value = round(v, 2); c.number_format = '0.00"%"'
        c.alignment = Alignment(horizontal='right'); c.border = _TH
        c.font = Font(size=9, color='C00000' if v < 0 else '888888')
        if mdd_limit_pct is not None and v < -abs(mdd_limit_pct):
            c.fill = _BAD; c.font = Font(bold=True, size=9, color='C00000')
        # ★ feature3 — 매수공식/매도공식 제거, 사용지표별 '수치 + 신호면 노란색' 컬럼으로 교체
        _uic = cur.get('used_ind_cols', []) if isinstance(cur, dict) else []
        for _j, (_kind, _nm, _vc, _sc) in enumerate(_uic):
            cc = ws.cell(r, 19 + _j)
            _v = row.get(_vc, np.nan)
            if pd.notna(_v):
                try: cc.value = round(float(_v), 4)
                except Exception: cc.value = _v
                cc.number_format = '0.0000'
            cc.alignment = Alignment(horizontal='right'); cc.border = _TH; cc.font = Font(size=8)
            if row.get(_sc, 0) == 1:        # 신호 충족 → 노란색
                cc.fill = _SIGY
    _uic = cur.get('used_ind_cols', []) if isinstance(cur, dict) else []
    widths = [12, 10, 13, 8, 9, 13, 8, 9, 8, 34, 10, 10, 8, 10, 10, 10, 12, 10]
    widths += [14] * len(_uic)
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 32
    ws.freeze_panes = 'B5'


def _write_logic_verification_sheet(wb, feat, close_ser, ticker, *,
                                    horizon, dd_limit, ru_limit):
    """★ '검증_예측로직' 시트 (요청) — 적용된 개선사항이 의도대로 기능하는지 실행 시점에
       독립 재계산으로 교차검증해 기록.
       ① 기능 점검: 성공률 독립 재계산 일치 / KL·net>K 수익 재계산 일치 / 지연 인과성(미래참조 無)
          / 리드탐색·스킬필터·홀드아웃가드 적용 통계
       ② 기저확률(base rate) 표: 한도×선행일별 — '성공률 착시'(h 길수록 저절로 상승) 근거 제시
       ③④ 풀 지표별 예측력: 성공률·기저·스킬·배율·최적선행·리드프로파일·훈련/홀드아웃·순열백분위
       ⑤ 종합 판정 + 통계적 주의(선택편향·소표본)"""
    g = globals()
    ws = wb.create_sheet('검증_예측로직'); ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = (f'{ticker} — 예측 로직 검증 (모든 수치는 시트 생성 시점에 독립 재계산한 교차검증 값) '
                           f'｜ 스킬 = 성공률 − 기저확률(같은 조건에서 아무 날이나 골라도 맞는 확률). '
                           f'스킬 ≤ 0 이면 그 지표는 시장 기본 움직임 이상의 예측력이 없음')
    ws.cell(1, 1).font = Font(bold=True, size=12, color='1F3864')

    _OK  = PatternFill('solid', fgColor='C6EFCE')
    _WRN = PatternFill('solid', fgColor='FFC7CE')
    _MID = PatternFill('solid', fgColor='FFF2CC')

    if feat is None or close_ser is None:
        ws.cell(3, 1).value = '⚠ feat/close 미전달 — 검증 생략'
        return

    dates = feat.index
    close_arr = np.asarray(pd.Series(close_ser).reindex(dates).values, dtype=np.float64)
    n = len(close_arr)
    hz = max(1, int(horizon or g.get('HORIZON_DAYS', 1)))
    horizons = sorted(set(int(h) for h in (g.get('LEAD_HORIZONS') or [1, 2, 3, 5])) | {hz})
    use_bar = bool(g.get('LEAD_TRIPLE_BARRIER', False))
    f_hold = min(max(float(g.get('POOL_HOLDOUT_FRACTION', 0.30)), 0.0), 0.9)
    split = min(max(int(round(n * (1.0 - f_hold))), 1), n)
    perm_n = max(0, int(g.get('VERIFY_PERM_N', 60)))
    perm_rows_cap = max(0, int(g.get('VERIFY_PERM_MAX_ROWS', 40)))
    rng = np.random.default_rng(42)

    _mp = g.get('_KNET_MULTI_POOL')
    buy_pool = sell_pool = None
    if _mp and isinstance(_mp, tuple) and len(_mp) >= 3 and _mp[0] == ticker:
        buy_pool, sell_pool = _mp[1], _mp[2]

    # ── 행별 예측력 재계산 (지연 적용된 '실거래 신호' 기준) ──
    def _row_metrics(row, is_buy, do_perm):
        rd = row if isinstance(row, dict) else row.to_dict()
        default_limit = float(DRAWDOWN_LIMIT_BUY if is_buy else RUNUP_LIMIT_SELL)
        try:
            _sl = rd.get('sel_limit', default_limit)
            limit = float(_sl) if _sl is not None and np.isfinite(float(_sl)) else default_limit
        except Exception:
            limit = default_limit
        try:
            sig = _to_signal_array(feat, rd)              # 지연 포함 실거래 신호
            sig_raw = _to_signal_array_raw(feat, rd)      # 지연 인과성 점검용
        except Exception:
            return None
        d = 0
        try:
            d = int(rd.get('lead_shift', 0) or 0)
        except Exception:
            d = 0
        hit0, ev0 = _hit_flags_cached(close_arr, hz, limit, 1 if is_buy else 0, 0)
        nn, ok, sr = _success_on(hit0, ev0, sig)
        _, _, base = _success_on(hit0, ev0, None)
        skill = (sr - base) if (np.isfinite(sr) and np.isfinite(base)) else np.nan
        lift = (sr / base) if (np.isfinite(sr) and base and base > 0) else np.nan
        n_tr, _, sr_tr = _success_on(hit0, ev0, sig, 0, split)
        _, _, b_tr = _success_on(hit0, ev0, None, 0, split)
        n_ho, _, sr_ho = _success_on(hit0, ev0, sig, split, n)
        _, _, b_ho = _success_on(hit0, ev0, None, split, n)
        sk_ho = (sr_ho - b_ho) if (n_ho >= 3 and np.isfinite(sr_ho) and np.isfinite(b_ho)) else np.nan
        # 리드 프로파일 (원신호 기준 — '몇 일 전이 가장 정확한가')
        best_h = None; prof_txt = '—'
        if bool(g.get('LEAD_TIME_SEARCH', False)):
            prof = _lead_profile(close_arr, sig_raw, limit, is_buy, horizons, use_bar,
                                 min_n=max(3, int(g.get('POOL_SUCCESS_MIN_SIG', 10)) // 2))
            bs = -1e18
            for p in prof:
                if np.isfinite(p['skill']) and p['skill'] > bs + 1e-12:
                    bs = p['skill']; best_h = p['h']
            prof_txt = ' | '.join(
                (f"{p['h']}일:{p['sr']*100:.0f}%({p['skill']*100:+.0f}p)"
                 if np.isfinite(p['skill']) else f"{p['h']}일:—") for p in prof)
        # 순열(원형시프트) 검정 — 신호 '타이밍'을 무작위로 돌려도 이만한 스킬이 나오는지
        perm_pct = np.nan
        if do_perm and perm_n > 0 and np.isfinite(skill) and nn >= 5 and n > 40:
            offs = rng.integers(15, n - 15, size=perm_n)
            worse = 0; cnt = 0
            for off in offs:
                sp = np.roll(sig, int(off))
                pn, _, psr = _success_on(hit0, ev0, sp)
                if pn >= 3 and np.isfinite(psr):
                    cnt += 1
                    if skill > (psr - base) + 1e-12:
                        worse += 1
            perm_pct = (worse / cnt) if cnt else np.nan
        # 지연 인과성: 지연 적용 신호가 정확히 sig_raw를 d일 민 것인지 (미래참조 無 구조 확인)
        causal_ok = True
        if d > 0:
            causal_ok = bool(np.all(sig[:d] == 0) and np.all(sig[d:] == sig_raw[:len(sig_raw) - d]))
        return dict(ind=str(rd.get('indicator', '')), direction=str(rd.get('direction', '')),
                    threshold=rd.get('threshold'), limit=limit, d=d,
                    n=nn, sr=sr, base=base, skill=skill, lift=lift,
                    best_lead=best_h, prof=prof_txt,
                    n_tr=n_tr, sr_tr=sr_tr, n_ho=n_ho, sr_ho=sr_ho, skill_ho=sk_ho,
                    perm=perm_pct, causal_ok=causal_ok,
                    sr_stored=rd.get('success_rate'))

    buy_m = []; sell_m = []
    for pool, is_buy, dest in [(buy_pool, True, buy_m), (sell_pool, False, sell_m)]:
        if pool is None or len(pool) == 0:
            continue
        for i, (_, row) in enumerate(pool.iterrows()):
            m = _row_metrics(row, is_buy, do_perm=(i < perm_rows_cap))
            if m is not None:
                dest.append(m)

    # ════ ① 기능 점검 ════
    r = 3
    ws.cell(r, 1).value = '① 적용 개선사항 & 기능 점검 (실행 시점 자체 교차검증)'
    ws.cell(r, 1).font = Font(bold=True, size=12, color='1F3864'); r += 1
    _hdr(ws, r, ['점검 항목', '결과', '판정']); r += 1

    def _put_check(label, result, ok, warn_txt='⚠ 확인 필요'):
        nonlocal r
        ws.cell(r, 1).value = label
        ws.cell(r, 2).value = result
        c = ws.cell(r, 3)
        if ok is None:
            c.value = '— (해당 없음)'
        else:
            c.value = '✓ 정상' if ok else warn_txt
            c.fill = _OK if ok else _WRN
            c.font = Font(bold=True, color='006100' if ok else '9C0006')
        r += 1

    # (a) 성공률 독립 재계산 일치 — 저장 성공률 vs 재계산 (다른 코드경로 교차검증).
    #     지연 d>0 행도 저장 성공률이 '지연 적용 신호' 기준이므로 그대로 비교 가능.
    diffs = []
    for m in (buy_m + sell_m):
        if m['sr_stored'] is not None and np.isfinite(m['sr']):
            try:
                sv = float(m['sr_stored'])
                if sv > 1.5: sv /= 100.0
                if np.isfinite(sv):
                    diffs.append(abs(sv - m['sr']))
            except Exception:
                pass
    if diffs:
        _md = max(diffs)
        _put_check('성공률 독립 재계산 일치 (지연 반영, 선출 vs 검증 별도 코드)',
                   f'{len(diffs)}행 비교, 최대 오차 {_md*100:.4f}%p', _md < 0.005)
    else:
        _put_check('성공률 독립 재계산 일치', '비교 가능한 행 없음', None)

    # (b) KL 대표조합 수익 재계산 일치
    _kl = g.get('_KNET_KL'); _nf = g.get('_KNET_FULL')
    if _kl and _kl.get('best_ret') and _nf and _nf.get('daily') is not None:
        try:
            _d = _nf['daily']; _price = np.asarray(_d['price'].values, float)
            _rr = np.zeros(len(_price)); _rr[1:] = _price[1:] / _price[:-1] - 1.0
            _rr[~np.isfinite(_rr)] = 0.0
            _pos = np.asarray(_kl['best_ret'][5], float)
            _re = float(np.sum(_pos * _rr)); _st = float(_kl['best_ret'][2])
            _put_check('KL 순신호 대표조합 수익 재계산 일치',
                       f'시트 {_st*100:+.4f}% vs 재계산 {_re*100:+.4f}%', abs(_re - _st) < 1e-8)
        except Exception as _e:
            _put_check('KL 순신호 수익 재계산', f'실패: {_e}', False)
    else:
        _put_check('KL 순신호 수익 재계산', 'KL 결과 없음', None)

    # (c) net>K 일별수익 정합 — position(전일 소유)×당일 등락 재계산 vs day_ret
    if _nf and _nf.get('daily') is not None:
        try:
            _d = _nf['daily']; _price = np.asarray(_d['price'].values, float)
            _rr = np.zeros(len(_price)); _rr[1:] = _price[1:] / _price[:-1] - 1.0
            _rr[~np.isfinite(_rr)] = 0.0
            _pos = np.asarray(_d['position'].values, float)
            _dr2 = np.zeros(len(_price)); _dr2[1:] = _pos[:-1] * _rr[1:]
            _md = float(np.max(np.abs(_dr2 - np.asarray(_d['day_ret'].values, float))))
            _put_check("net>K 일별수익 정합 (진입=신호일 종가, 수익=전일 소유×당일 등락)",
                       f'최대 오차 {_md:.2e}', _md < 1e-9)
        except Exception as _e:
            _put_check('net>K 일별수익 정합', f'실패: {_e}', False)

    # (d) 신호 지연 인과성 — 지연 적용 신호가 원신호를 정확히 d일 민 것인지 (미래참조 無)
    _shifted = [m for m in (buy_m + sell_m) if m['d'] > 0]
    if _shifted:
        _all_ok = all(m['causal_ok'] for m in _shifted)
        _put_check('신호 지연 인과성 (sig[t-d]→t, 앞구간 0 채움 = 미래참조 없음)',
                   f'지연 적용 {len(_shifted)}행 전부 구조 확인', _all_ok)
    else:
        _put_check('신호 지연 인과성', '지연 적용 행 없음 (개선폭·홀드아웃 조건 미충족)', None)

    # (e) 선출 필터 적용 통계 (분석 실행 시 수집 — 재현 모드는 없음)
    _es = g.get('_ENRICH_STATS') or {}
    for _side, _lb in [('buy', '매수'), ('sell', '매도')]:
        st = _es.get(_side)
        if st:
            _put_check(f'{_lb} 풀 선출 필터 (스킬>{float(g.get("POOL_MIN_SKILL",0.0))*100:.0f}%p '
                       f'& 홀드아웃 가드 {float(g.get("POOL_HOLDOUT_MIN_SKILL",-0.1))*100:.0f}%p)',
                       f"{st['n_in']}행 → {st['n_out']}행 (스킬미달 -{st['drop_skill']} / "
                       f"홀드아웃 -{st['drop_holdout']} / 지연채택 {st['n_shift']}행"
                       f"{' / ⚠폴백' if st.get('fell_back') else ''})",
                       not st.get('fell_back'), warn_txt='⚠ 필터 폴백(전량 탈락→해제)')
        else:
            _put_check(f'{_lb} 풀 선출 필터', '통계 없음 (재현 모드 또는 필터 OFF)', None)

    # (f) 설정 스냅샷
    ws.cell(r, 1).value = '적용 설정'
    ws.cell(r, 2).value = (f"리드탐색={g.get('LEAD_TIME_SEARCH')} (h={g.get('LEAD_HORIZONS')}) · "
                           f"삼중배리어={g.get('LEAD_TRIPLE_BARRIER')} · "
                           f"지연정렬={g.get('LEAD_SHIFT_ENABLED')}(최대{g.get('LEAD_SHIFT_MAX')}일, "
                           f"개선≥{float(g.get('LEAD_SHIFT_MIN_GAIN',0.05))*100:.0f}%p) · "
                           f"스킬필터={g.get('POOL_REQUIRE_SKILL')} · "
                           f"홀드아웃={f_hold*100:.0f}% (분할 {pd.Timestamp(dates[split-1]).date() if split-1 < n else '—'} 이후)")
    r += 2

    # ════ ② 기저확률 표 ════
    ws.cell(r, 1).value = ('② 기저확률(base rate) — 아무 날이나 골라도 h일 이내 한도에 도달할 확률. '
                           'h(선행일)가 길수록 성공률이 저절로 오르므로, 지표 성공률은 반드시 같은 h의 기저확률과 비교(=스킬)')
    ws.cell(r, 1).font = Font(bold=True, size=12, color='1F3864'); r += 1
    _limits = sorted(set([float(x) for x in (g.get('STAGE_SUCCESS_LIMIT') or [dd_limit])] + [float(dd_limit)]))
    _hdr(ws, r, ['한도'] + [f'{h}일 매수기저' for h in horizons] + [f'{h}일 매도기저' for h in horizons]); r += 1
    for L in _limits:
        ws.cell(r, 1).value = f'±{L*100:.0f}%'
        ci = 2
        for is_buy in (1, 0):
            for h in horizons:
                hit, ev = _hit_flags_cached(close_arr, h, L, is_buy, 1 if use_bar else 0)
                _, _, b = _success_on(hit, ev, None)
                ws.cell(r, ci).value = f'{b*100:.1f}%' if np.isfinite(b) else '—'
                ci += 1
        r += 1
    ws.cell(r, 1).value = ('※ 판정 기준: ' + ('삼중배리어(유리 한도가 불리 한도보다 먼저 도달)' if use_bar
                                              else '기존(한도 도달 여부만)') +
                           ' — 체결 지평(1일) 풀 선출 기준은 기존 정의 그대로 유지')
    r += 2

    # ════ ③④ 풀 지표별 예측력 ════
    def _write_pool_table(title, metrics):
        nonlocal r
        ws.cell(r, 1).value = title
        ws.cell(r, 1).font = Font(bold=True, size=12, color='1F3864'); r += 1
        _hdr(ws, r, ['#', '지표', '방향', '임계치', '한도', '지연d', '신호수', '성공률', '기저확률',
                     '스킬(성공-기저)', '배율(성공/기저)', '최적선행일', '리드 프로파일 (h일: 성공률(스킬p))',
                     '훈련 성공률', '홀드아웃 성공률', '홀드아웃 스킬', '순열백분위', '판정'])
        r += 1
        _cap = 80
        for i, m in enumerate(metrics[:_cap], 1):
            verdict = '—'
            fill = None
            if np.isfinite(m['skill']):
                if m['skill'] > 0 and (not np.isfinite(m['perm']) or m['perm'] >= 0.90) \
                        and (not np.isfinite(m['skill_ho']) or m['skill_ho'] > -1e-12):
                    verdict = '✓ 유효(우연 대비 우위)'; fill = _OK
                elif m['skill'] > 0:
                    verdict = '△ 스킬 있으나 확증 약함'; fill = _MID
                else:
                    verdict = '⚠ 기저확률 이하'; fill = _WRN
            vals = [i, m['ind'], m['direction'],
                    (round(float(m['threshold']), 6) if m['threshold'] is not None and pd.notna(m['threshold']) else '—'),
                    f"±{m['limit']*100:.0f}%", m['d'], m['n'],
                    (f"{m['sr']*100:.1f}%" if np.isfinite(m['sr']) else '—'),
                    (f"{m['base']*100:.1f}%" if np.isfinite(m['base']) else '—'),
                    (f"{m['skill']*100:+.1f}%p" if np.isfinite(m['skill']) else '—'),
                    (f"{m['lift']:.2f}배" if np.isfinite(m['lift']) else '—'),
                    (f"{m['best_lead']}일" if m['best_lead'] else '—'),
                    m['prof'],
                    (f"{m['sr_tr']*100:.1f}% ({m['n_tr']})" if np.isfinite(m['sr_tr']) else '—'),
                    (f"{m['sr_ho']*100:.1f}% ({m['n_ho']})" if np.isfinite(m['sr_ho']) else f"— ({m['n_ho']})"),
                    (f"{m['skill_ho']*100:+.1f}%p" if np.isfinite(m['skill_ho']) else '— (표본<3)'),
                    (f"{m['perm']*100:.0f}%" if np.isfinite(m['perm']) else '—'),
                    verdict]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(r, ci); c.value = v; c.font = Font(size=9)
            if fill is not None:
                ws.cell(r, 18).fill = fill
            if np.isfinite(m['skill']):
                ws.cell(r, 10).font = Font(size=9, bold=True,
                                           color=('006100' if m['skill'] > 0 else '9C0006'))
            r += 1
        if len(metrics) > _cap:
            ws.cell(r, 1).value = f'… 외 {len(metrics)-_cap}행 (성공률 하위 생략)'; r += 1
        r += 1

    _write_pool_table(f'③ 매수 풀 지표별 예측력 검증 (체결 지평 {hz}일 · 실거래 신호=지연 적용 기준 독립 재계산)', buy_m)
    _write_pool_table(f'④ 매도 풀 지표별 예측력 검증 (체결 지평 {hz}일)', sell_m)

    # ════ ⑤ 종합 판정 ════
    ws.cell(r, 1).value = '⑤ 종합 판정'
    ws.cell(r, 1).font = Font(bold=True, size=12, color='1F3864'); r += 1
    def _med(vals):
        v = [x for x in vals if np.isfinite(x)]
        return float(np.median(v)) if v else np.nan
    for lb, ms in [('매수', buy_m), ('매도', sell_m)]:
        if not ms:
            ws.cell(r, 1).value = f'{lb} 풀'; ws.cell(r, 2).value = '풀 없음'; r += 1
            continue
        med_sk  = _med([m['skill'] for m in ms])
        med_tr  = _med([m['sr_tr'] for m in ms])
        med_ho  = _med([m['sr_ho'] for m in ms])
        med_sho = _med([m['skill_ho'] for m in ms])
        perms = [m['perm'] for m in ms if np.isfinite(m['perm'])]
        pos_sk = sum(1 for m in ms if np.isfinite(m['skill']) and m['skill'] > 0)
        strong = sum(1 for p in perms if p >= 0.90)
        ws.cell(r, 1).value = f'{lb} 풀 ({len(ms)}행)'
        ws.cell(r, 2).value = (
            f"스킬 중앙값 {med_sk*100:+.1f}%p (양수 {pos_sk}/{len(ms)}행) · "
            f"훈련 {med_tr*100:.1f}% → 홀드아웃 {med_ho*100:.1f}% "
            f"(감쇠 {(med_tr-med_ho)*100:+.1f}%p, 홀드아웃 스킬 중앙값 "
            f"{med_sho*100:+.1f}%p) · " if np.isfinite(med_ho) else
            f"스킬 중앙값 {med_sk*100:+.1f}%p (양수 {pos_sk}/{len(ms)}행) · 홀드아웃 표본 부족 · ")
        ws.cell(r, 2).value += (f"순열검정 백분위≥90%: {strong}/{len(perms)}행"
                                if perms else "순열검정 대상 없음")
        _ok = np.isfinite(med_sk) and med_sk > 0 and (not np.isfinite(med_sho) or med_sho > -0.05)
        c = ws.cell(r, 3)
        c.value = '✓ 예측력 있음' if _ok else '⚠ 예측력 약함/미확인'
        c.fill = _OK if _ok else _WRN
        c.font = Font(bold=True, color='006100' if _ok else '9C0006')
        r += 1
    r += 1
    for note in [
        '※ 주의 1 (선택 편향): 수천 지표×임계에서 성공률 최고를 고르면 우연히 높은 것이 섞임 — 순열백분위·홀드아웃이 낮은 행은 신뢰 금물.',
        '※ 주의 2 (소표본): 신호수·홀드아웃 신호가 적은 행의 성공률은 불안정 — 신호수 열을 함께 볼 것.',
        '※ 주의 3 (기간): 홀드아웃도 같은 데이터의 뒤쪽 일부일 뿐, 진짜 미래가 아님 — 실전은 새 데이터 재분석으로 재확인 권장.',
    ]:
        ws.cell(r, 1).value = note; ws.cell(r, 1).font = Font(size=9, italic=True, color='808080'); r += 1

    for ci, w in enumerate([34, 42, 9, 12, 8, 7, 8, 9, 10, 13, 12, 10, 60, 13, 15, 13, 10, 22], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'
    _nb = len(buy_m); _ns = len(sell_m)
    print(f"  ✅ '검증_예측로직' 시트 작성 — 매수 {_nb}행 / 매도 {_ns}행 재계산 검증")


def _kl_pos_held_dd(pos, price):
    """KL 백테스트 결과의 '보유중 최대 하락률' 계산 (진입가 대비 저점)."""
    held_max_dd = 0.0
    ei = None
    n = len(pos)
    for i in range(n):
        if pos[i] == 1 and (i == 0 or pos[i - 1] == 0):
            ei = i
        if pos[i] == 0 and i > 0 and pos[i - 1] == 1 and ei is not None:
            seg = price[ei:i + 1]
            if len(seg) >= 2:
                dd = float(seg.min() / seg[0] - 1.0)
                if dd < held_max_dd:
                    held_max_dd = dd
            ei = None
    if ei is not None:
        seg = price[ei:]
        if len(seg) >= 2:
            dd = float(seg.min() / seg[0] - 1.0)
            if dd < held_max_dd:
                held_max_dd = dd
    return held_max_dd


def run_selection_ab_verification(feat, close_full, ticker, *, top_n=None, quick=False):
    """★ 지표 선출 개선 A/B 검증 (요청).

    각 개선 flag의 ON/OFF 성적을 KL 순신호 백테스트로 비교해 개별 기여도를 측정.
    - Baseline: 모든 신규 flag OFF (원본 로직: raw success_rate 정렬, 감점 없음)
    - 변형: flag 하나씩 ON (개별 marginal gain)
    - Combined: 현재 defaults (모두 반영된 상태)

    각 변형별 KL 백테스트 결과 기록: 전체수익%, MDD%, 보유중하락%, 거래횟수, 승률%.
    반환: list[dict] — 엑셀 시트로 저장 가능.
    """
    g = globals()
    src = g.get('_KNET_PRERANK_POOL')
    if not src or not isinstance(src, tuple):
        return None
    buy_src, sell_src = src
    if buy_src is None or sell_src is None or len(buy_src) == 0 or len(sell_src) == 0:
        return None
    top_n = int(top_n or g.get('TOP_N_POOL', g.get('MAX_POOL', 100)) or 100)
    dd_limit = float((g.get('STAGE_SUCCESS_LIMIT') or [0.01])[0])
    hz = int(g.get('HORIZON_DAYS', 1))
    mdd_limit = g.get('MAX_DRAWDOWN_LIMIT_PCT')

    # 원래 flag 스냅샷 (복원용)
    _snap = {k: g.get(k) for k in ('POOL_RANK_BY', 'USE_ADVERSE_PENALTY',
                                    'USE_HOLDOUT_DECAY_PENALTY',
                                    'MAX_THRESHOLDS_PER_INDICATOR')}

    # 검증 변형 정의: (라벨, flag 딕셔너리)
    variants = [
        ('Baseline (신규 flag 전부 OFF)', dict(POOL_RANK_BY='success', USE_ADVERSE_PENALTY=False,
                                             USE_HOLDOUT_DECAY_PENALTY=False,
                                             MAX_THRESHOLDS_PER_INDICATOR=0)),
        ('SEL-1 정렬=Expected (SR×움직임)', dict(POOL_RANK_BY='expected', USE_ADVERSE_PENALTY=False,
                                                 USE_HOLDOUT_DECAY_PENALTY=False,
                                                 MAX_THRESHOLDS_PER_INDICATOR=0)),
        ('SEL-1 정렬=Wilson 하한', dict(POOL_RANK_BY='wilson', USE_ADVERSE_PENALTY=False,
                                        USE_HOLDOUT_DECAY_PENALTY=False,
                                        MAX_THRESHOLDS_PER_INDICATOR=0)),
        ('SEL-1 정렬=Skill (순 알파)', dict(POOL_RANK_BY='skill', USE_ADVERSE_PENALTY=False,
                                            USE_HOLDOUT_DECAY_PENALTY=False,
                                            MAX_THRESHOLDS_PER_INDICATOR=0)),
        ('SEL-2 불리방향 감점', dict(POOL_RANK_BY='success', USE_ADVERSE_PENALTY=True,
                                     USE_HOLDOUT_DECAY_PENALTY=False,
                                     MAX_THRESHOLDS_PER_INDICATOR=0)),
        ('SEL-3 지표당 임계 3개 제한', dict(POOL_RANK_BY='success', USE_ADVERSE_PENALTY=False,
                                            USE_HOLDOUT_DECAY_PENALTY=False,
                                            MAX_THRESHOLDS_PER_INDICATOR=3)),
        ('SEL-4 홀드아웃 감쇠 감점', dict(POOL_RANK_BY='success', USE_ADVERSE_PENALTY=False,
                                          USE_HOLDOUT_DECAY_PENALTY=True,
                                          MAX_THRESHOLDS_PER_INDICATOR=0)),
    ]
    if not quick:
        variants.append(('★ 전부 적용 (현재 defaults)', dict(_snap)))

    results = []
    _t0 = time.time()
    for label, flags in variants:
        # flag 적용
        for k, v in flags.items():
            g[k] = v
        try:
            # 풀 재구성 (스냅샷 재사용)
            _b = _limit_thresholds_per_indicator(buy_src.copy(),  verbose=False)
            _s = _limit_thresholds_per_indicator(sell_src.copy(), verbose=False)
            _b = _rank_pool_by_selection(_b, True,  dd_limit=dd_limit, verbose=False)
            _s = _rank_pool_by_selection(_s, False, dd_limit=dd_limit, verbose=False)
            _b = _b.head(top_n).reset_index(drop=True)
            _s = _s.head(top_n).reset_index(drop=True)
            # 순신호(net) 계산 — SEARCH_WEIGHT_SCHEME는 baseline과 공평 비교 위해 고정 g=1.0 사용
            _nsd = _net_signal_k_search(feat, close_full, _b, _s, ticker=ticker,
                                         oos_start=None, n_buy=None, n_sell=None,
                                         search_counts=True, weight_exp=1.0, select_by='full')
            if _nsd is None or _nsd.get('daily') is None:
                results.append(dict(variant=label, ret=None, mdd=None, held_dd=None,
                                     n_trades=None, win_rate=None, buy_rows=len(_b),
                                     sell_rows=len(_s), status='net 계산 실패'))
                continue
            _d = _nsd['daily']
            _net = _d['net'].values.astype(float)
            _price = _d['price'].values.astype(float)
            _rr = np.zeros(len(_price)); _rr[1:] = _price[1:] / _price[:-1] - 1.0
            _rr[~np.isfinite(_rr)] = 0.0
            _kl = _net_kl_search(_net, _rr, mdd_limit=mdd_limit)
            best = _kl.get('best_ret')
            if best is None:
                results.append(dict(variant=label, ret=None, mdd=None, held_dd=None,
                                     n_trades=None, win_rate=None, buy_rows=len(_b),
                                     sell_rows=len(_s), status='KL 결과 없음'))
                continue
            K, L, ret, mdd, dl, pos = best[:6]
            # 거래 통계
            tr = []; ei = None
            for i in range(len(pos)):
                if pos[i] == 1 and (i == 0 or pos[i - 1] == 0):
                    ei = i
                if pos[i] == 0 and i > 0 and pos[i - 1] == 1 and ei is not None:
                    tr.append(float(np.sum(_rr[ei + 1:i + 1]))); ei = None
            if ei is not None:
                tr.append(float(np.sum(_rr[ei + 1:])))
            nt = len(tr); wins = sum(1 for t in tr if t > 0)
            wr = (wins / nt * 100) if nt else 0.0
            held_dd = _kl_pos_held_dd(pos, _price)
            results.append(dict(variant=label, ret=float(ret), mdd=float(mdd),
                                 held_dd=float(held_dd), n_trades=int(nt), win_rate=float(wr),
                                 buy_rows=len(_b), sell_rows=len(_s),
                                 K=float(K), L=float(L), status='OK'))
        except Exception as e:
            results.append(dict(variant=label, ret=None, mdd=None, held_dd=None,
                                 n_trades=None, win_rate=None, buy_rows=0, sell_rows=0,
                                 status=f'실패: {str(e)[:60]}'))

    # 원래 flag 복원
    for k, v in _snap.items():
        g[k] = v

    _elapsed = time.time() - _t0
    # 콘솔 요약
    print("\n  " + "─" * 70)
    print(f"  🧪 지표 선출 A/B 검증 결과 ({_elapsed:.1f}초)")
    print("  " + "─" * 70)
    _base = next((r for r in results if 'Baseline' in r['variant']), None)
    _base_ret = _base['ret'] if _base and _base['ret'] is not None else 0.0
    _base_dd = _base['held_dd'] if _base and _base['held_dd'] is not None else 0.0
    for r in results:
        if r['ret'] is None:
            print(f"    · {r['variant']:36}  ✗ {r['status']}"); continue
        _dret = (r['ret'] - _base_ret) * 100
        _ddd = (r['held_dd'] - _base_dd) * 100
        _marker = '★' if 'Baseline' not in r['variant'] and _dret > 0 and _ddd >= -0.5 else ' '
        print(f"  {_marker} · {r['variant']:36}  수익 {r['ret']*100:+6.1f}% (Δ{_dret:+5.1f}%p)  "
              f"보유중하락 {r['held_dd']*100:5.1f}% (Δ{_ddd:+4.1f}%p)  거래 {r['n_trades']}회")
    print("  " + "─" * 70 + "\n")
    return results


def _write_selection_ab_sheet(wb, ab_results, *, mdd_limit=None):
    """A/B 검증 결과를 엑셀 시트로 저장."""
    if not ab_results:
        return
    ws = wb.create_sheet('지표선출_AB검증'); ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = ('🧪 지표 선출·점수 개선 아이디어의 KL 백테스트 성적 비교 '
                            '— 각 flag가 실제로 수익률·보유중하락 개선하는지 검증')
    ws.cell(1, 1).font = Font(bold=True, size=13, color='1F3864')
    ws.merge_cells('A1:K1')
    ws.cell(2, 1).value = ('★ Baseline = 신규 flag 전부 OFF (원본 로직: raw success_rate 정렬). '
                            'Δ 컬럼은 Baseline 대비 개선폭. 수익↑ & 보유중하락↑(덜 음수) 이면 개선.')
    ws.cell(2, 1).font = Font(italic=True, size=10, color='606060')
    ws.merge_cells('A2:K2')

    HDR = PatternFill('solid', fgColor='1F3864')
    OK = PatternFill('solid', fgColor='C6EFCE')
    ALT = PatternFill('solid', fgColor='F2F2F2')

    heads = ['변형(Variant)', '풀 매수/매도', 'K', 'L', '전체수익%',
             'Δ수익%p', 'MDD%', '보유중하락%', 'Δ보유중하락%p', '거래', '승률%']
    for ci, h in enumerate(heads, 1):
        c = ws.cell(4, ci); c.value = h
        c.font = Font(bold=True, color='FFFFFF'); c.fill = HDR
        c.alignment = Alignment(horizontal='center')

    _base = next((r for r in ab_results if 'Baseline' in r['variant']), None)
    _bret = _base['ret'] if _base and _base['ret'] is not None else 0.0
    _bdd = _base['held_dd'] if _base and _base['held_dd'] is not None else 0.0

    for ri, r in enumerate(ab_results):
        rr = 5 + ri
        is_base = ('Baseline' in r['variant'])
        vals = [r['variant']]
        vals.append(f"{r.get('buy_rows', 0)}/{r.get('sell_rows', 0)}")
        if r['ret'] is None:
            vals += ['—', '—', '—', '—', '—', '—', '—', '—', '—']
            ws.cell(rr, 1).value = f"{r['variant']}  ({r.get('status','?')})"
            ws.cell(rr, 1).font = Font(color='C00000')
            for ci in range(2, 12):
                ws.cell(rr, ci).value = '—'
            continue
        _dret_p = ('' if is_base else f" ({(r['ret']-_bret)*100:+.2f})")
        _ddd_p = ('' if is_base else f" ({(r['held_dd']-_bdd)*100:+.2f})")
        vals += [
            round(r.get('K', 0), 3), round(r.get('L', 0), 3),
            f"{r['ret']*100:+.2f}",
            '—' if is_base else f"{(r['ret']-_bret)*100:+.2f}",
            f"{r['mdd']*100:.2f}",
            f"{r['held_dd']*100:.2f}",
            '—' if is_base else f"{(r['held_dd']-_bdd)*100:+.2f}",
            r['n_trades'], round(r['win_rate'], 1),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(rr, ci); c.value = v
            c.alignment = Alignment(horizontal='center' if ci > 1 else 'left')
            c.font = Font(size=10, bold=is_base)
            if ri % 2 == 1: c.fill = ALT
        # 개선된 행 강조
        if not is_base and r['ret'] > _bret and r['held_dd'] >= _bdd - 0.005:
            for ci in range(1, 12): ws.cell(rr, ci).fill = OK

    # 하단 판정 요약
    _final = 5 + len(ab_results) + 1
    ws.cell(_final, 1).value = '📋 판정 요약'
    ws.cell(_final, 1).font = Font(bold=True, size=11, color='1F3864')
    _final += 1
    if _base:
        improved = [r for r in ab_results
                    if not ('Baseline' in r['variant']) and r['ret'] is not None
                    and r['ret'] > _bret and r['held_dd'] >= _bdd - 0.005]
        best_var = max([r for r in ab_results if r['ret'] is not None],
                        key=lambda r: r['ret']) if ab_results else None
        ws.cell(_final, 1).value = (f"개선 변형 {len(improved)}개 (Baseline 대비 수익↑ & 보유중하락 유지). "
                                     f"최고 수익: {best_var['variant']} ({best_var['ret']*100:+.2f}%)"
                                     if best_var else "")
        ws.cell(_final, 1).font = Font(size=10)
        _final += 1
        # 개선된 flag 조합 안내
        if improved:
            _tips = ' · '.join([r['variant'].split('(')[0].strip() for r in improved[:3]])
            ws.cell(_final, 1).value = f"→ 실전 활용 후보: {_tips}"
            ws.cell(_final, 1).font = Font(size=10, color='006100')

    for ci, w in enumerate([34, 12, 8, 8, 11, 11, 9, 12, 14, 8, 8], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A5'



def write_excel(meta_results_df, inner_all, inner_passed,
                best_meta, best_inner, buy_pool, sell_pool,
                daily, trades, cur, buy_used, sell_used, bh_ret, bh_cagr, *,
                ticker, output_file, horizon, dd_limit, ru_limit,
                mdd_limit_pct, min_trades_daily, stop_loss_pct,
                selection_tolerance=0.0,
                anchor_match_priority_arg=False,
                anchor_match_tolerance_arg=0.0,
                anchor_mode=False,
                anchor_buy_dates=None, anchor_sell_dates=None,
                auto_anchor=False,
                oos_enabled=False,
                oos_daily=None, oos_trades=None, oos_cur=None,
                bh_up_ret=None, anchor_prio=None,
                feat=None, close_full=None):
    wb = Workbook(); wb.remove(wb.active)

    # 1. 현재 포지션
    ws = wb.create_sheet('★ 현재 포지션', 0); ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = f'{ticker} — 앙상블 현재 포지션  ({cur["last_date"].date()})'
    ws.cell(1, 1).font = Font(bold=True, size=18, color='1F3864')
    ws.merge_cells('A1:D1')
    mdd_t = f"제한 -{abs(mdd_limit_pct)*100:.1f}%" if mdd_limit_pct is not None else "제한 없음"
    ws.cell(2, 1).value = (f'★ 매수 Top-{cur["K_buy"]}/{cur["vote_buy"]}투표  '
                            f'/ 매도 Top-{cur["K_sell"]}/{cur["vote_sell"]}투표  '
                            f'/ MDD {mdd_t}  / 최소거래 {min_trades_daily}회')
    ws.cell(2, 1).font = Font(italic=True, size=10, color='606060')
    ws.merge_cells('A2:D2')

    pr = 4
    ws.cell(pr, 1).value = '현재 포지션'
    ws.cell(pr, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.cell(pr, 2).value = f'{cur["position_emoji"]}  {cur["position"]}'
    if '보유' in cur['position']:
        ws.cell(pr, 2).fill = _HOLD
        ws.cell(pr, 2).font = Font(bold=True, size=18, color='C00000')
    else:
        ws.cell(pr, 2).fill = _CASH
        ws.cell(pr, 2).font = Font(bold=True, size=18, color='006100')
    ws.merge_cells(start_row=pr, start_column=2, end_row=pr, end_column=4)
    ws.row_dimensions[pr].height = 30

    if '보유' in cur['position']:
        details = [('진입 날짜', cur['entry_date'].date() if pd.notna(cur['entry_date']) else '-'),
                   ('진입 가격', f"${cur['entry_price']:.2f}"),
                   ('현재 가격', f"${cur['current_price']:.2f}"),
                   ('보유 일수', f"{cur['days_held']}일"),
                   ('미실현 손익', f"{cur['unrealized_pct']:+.2f}%")]
        if stop_loss_pct is not None and stop_loss_pct > 0 and cur.get('stop_price'):
            sp = cur['stop_price']
            dist_pct = (cur['current_price'] / sp - 1) * 100
            details.append(
                (f'⛔ 손절가 (-{stop_loss_pct*100:.1f}%)',
                  f"${sp:.2f}  (현재가 대비 +{dist_pct:.2f}% 여유)")
            )
    else:
        details = [('현재 가격', f"${cur['current_price']:.2f}"),
                   ('상태', '매수 신호 대기 중')]
    details += [
        ('', ''),
        ('전략 누적 자산', f"{cur['equity']:.4f}"),
        ('전략 누적 수익률', f"{cur['cum_return_pct']:+.2f}%"),
        ('Buy & Hold', f"{bh_ret*100:+.2f}%"),
        ('📈 상승일만 합산 (전략)', f"{cur.get('up_cum_return_pct', 0):+.2f}%"),
        ('📈 상승일만 합산 (B&H)', f"{bh_up_ret*100:+.2f}%" if bh_up_ret is not None else '—'),
        ('총 거래 수', f"{cur['n_trades']}회"),
        ('승률 (수익 거래)', f"{cur['win_rate']*100:.1f}% ({cur['n_wins']}/{cur['n_trades']})"),
        ('최대 거래손실 (단일거래)', f"{cur['max_drawdown']*100:.2f}%"),
    ]
    label_buy  = f'★ 매수신호 Balanced Acc (-{dd_limit*100:.1f}% 회피'
    label_sell = f'★ 매도신호 Balanced Acc (+{ru_limit*100:.1f}% 회피'
    if anchor_mode:
        label_buy  += ' + ⚓ANCHOR 보정)'
        label_sell += ' + ⚓ANCHOR 보정)'
    else:
        label_buy  += ')'
        label_sell += ')'
    sel_label = (f'★ 평균 Balanced Acc (Top band {selection_tolerance*100:.1f}%p 내 수익최대)'
                  if selection_tolerance > 0 else '★ 평균 Balanced Acc (선정 1차)')
    details += [
        ('', ''),
        (label_buy,
          f"{cur['buy_success_rate']*100:.1f}% (plain {cur['buy_accuracy_plain']*100:.1f}%, ON {cur['n_buy_signal_on']}일)"),
        (label_sell,
          f"{cur['sell_success_rate']*100:.1f}% (plain {cur['sell_accuracy_plain']*100:.1f}%, ON {cur['n_sell_signal_on']}일)"),
        (sel_label,
          f"{cur['avg_success_rate']*100:.1f}%"),
    ]
    # ★ 변경4: 전체 매수/매도 성공·실패 집계 + 투표 방식
    details += [
        ('', ''),
        ('★ 매수신호 성공/실패 (ON만)',
          f"{cur.get('n_buy_success_cnt',0)} 성공 / {cur.get('n_buy_fail_cnt',0)} 실패 "
          f"(신호 {cur.get('n_buy_on_total',0)}회, 적중 {cur.get('buy_signal_hit_rate',0)*100:.1f}%)"),
        ('★ 매수 정답일 적중률',
          f"{cur.get('buy_acc_all',0)*100:.1f}% "
          f"({cur.get('n_buy_correct_all',0)}/{cur.get('n_buy_eval_all',0)}일 — 올랐어야 한 날 중 적중, 놓침=실패)"),
        ('★ 매도신호 성공/실패 (ON만)',
          f"{cur.get('n_sell_success_cnt',0)} 성공 / {cur.get('n_sell_fail_cnt',0)} 실패 "
          f"(신호 {cur.get('n_sell_on_total',0)}회, 적중 {cur.get('sell_signal_hit_rate',0)*100:.1f}%)"),
        ('★ 매도 정답일 적중률',
          f"{cur.get('sell_acc_all',0)*100:.1f}% "
          f"({cur.get('n_sell_correct_all',0)}/{cur.get('n_sell_eval_all',0)}일 — 내렸어야 한 날 중 적중, 놓침=실패)"),
        ('★ 성공 판정 방식',
          'First-Touch (익일진입, 기간내 +목표/-손절 먼저 닿는 쪽)'),
        ('★ 투표 방식',
          '가중 투표 (성공률 비례)' if cur.get('weighted_vote') else '일반 투표 (모두 1표)'),
    ]
    if anchor_mode:
        details += [
            (f'⚓ 매수 정답일 매칭',
              f"{cur['n_anchor_buy_caught']}/{cur['n_anchor_buy']} "
              f"({(cur['n_anchor_buy_caught']/cur['n_anchor_buy']*100) if cur['n_anchor_buy']>0 else 0:.1f}%)"),
            (f'⚓ 매도 정답일 매칭',
              f"{cur['n_anchor_sell_caught']}/{cur['n_anchor_sell']} "
              f"({(cur['n_anchor_sell_caught']/cur['n_anchor_sell']*100) if cur['n_anchor_sell']>0 else 0:.1f}%)"),
        ]
    details += [
        (f'★ 신호 충돌 발생',
          f"{cur['n_conflicts']}일 (매수우세 {cur['n_conflict_buy_won']} / 매도우세 {cur['n_conflict_sell_won']})"),
        (f'⛔ 손절매 발동',
          f"{cur['n_stop_triggered']}회"
          + (f" (한도 -{stop_loss_pct*100:.1f}%)" if stop_loss_pct else " (손절매 비활성)")),
        ('', ''),
        ('★ MDD 한도', f"-{abs(mdd_limit_pct)*100:.2f}%" if mdd_limit_pct is not None else "없음"),
        ('★ 최소 거래수', f"{min_trades_daily}회"),
        ('★ 손절매 한도', f"-{stop_loss_pct*100:.2f}%" if (stop_loss_pct is not None and stop_loss_pct > 0) else "없음"),
        ('★ Tolerance Band', f"{selection_tolerance*100:.2f}%p" if selection_tolerance > 0 else "OFF (strict)"),
        ('★ ANCHOR 매칭 우선', f"ON ({anchor_match_tolerance_arg*100:.1f}%p)" if anchor_match_priority_arg and anchor_mode else "OFF"),
        ('★ ANCHOR 보정',
          ('ON ⚓ AUTO (자동 계산)' if auto_anchor else 'ON ⚓ (수동 입력)') if anchor_mode else 'OFF'),
    ]
    for i, (k, v) in enumerate(details):
        r = 6 + i
        ws.cell(r, 1).value = k
        ws.cell(r, 1).font = Font(bold=True, size=11, color='1F3864')
        ws.cell(r, 2).value = v
        ws.cell(r, 2).font = Font(size=11)
        if k.startswith('★') or k.startswith('⛔') or k.startswith('⚓'):
            ws.cell(r, 1).fill = _HL; ws.cell(r, 2).fill = _HL
            ws.cell(r, 2).font = Font(bold=True, size=12, color='C00000')
            if '매수신호' in k and 'Balanced' in k:
                ws.cell(r, 2).fill = _success_fill(cur['buy_success_rate'])
            elif '매도신호' in k and 'Balanced' in k:
                ws.cell(r, 2).fill = _success_fill(cur['sell_success_rate'])
            elif '평균' in k and 'Balanced' in k:
                ws.cell(r, 2).fill = _success_fill(cur['avg_success_rate'])
            elif k.startswith('⛔'):
                ws.cell(r, 2).font = Font(bold=True, size=13, color='C00000')

    sr = 6 + len(details) + 2
    ws.cell(sr, 1).value = '⚡ 오늘의 앙상블 투표'
    ws.cell(sr, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=4)
    _hdr(ws, sr + 2, ['신호', 'ON 개수', '필요(투표)', '결과', ''])

    r = sr + 3
    ws.cell(r, 1).value = '🟢 매수'; ws.cell(r, 1).font = Font(size=11, bold=True)
    ws.cell(r, 2).value = f"{cur['buy_count_now']:.2f}/{cur['K_buy']}"
    ws.cell(r, 3).value = f"{cur['vote_buy']}개 이상"
    ws.cell(r, 4).value = 'ON ✓' if cur['buy_on_now'] else 'OFF'
    if cur['buy_on_now']:
        ws.cell(r, 4).fill = _BUY
        ws.cell(r, 4).font = Font(size=12, bold=True, color='006100')
    for ci in range(1, 5):
        ws.cell(r, ci).border = _TH
        ws.cell(r, ci).alignment = Alignment(horizontal='center')

    r += 1
    ws.cell(r, 1).value = '🔴 매도'; ws.cell(r, 1).font = Font(size=11, bold=True)
    ws.cell(r, 2).value = f"{cur['sell_count_now']:.2f}/{cur['K_sell']}"
    ws.cell(r, 3).value = f"{cur['vote_sell']}개 이상"
    ws.cell(r, 4).value = 'ON ✓' if cur['sell_on_now'] else 'OFF'
    if cur['sell_on_now']:
        ws.cell(r, 4).fill = _SELL
        ws.cell(r, 4).font = Font(size=12, bold=True, color='C00000')
    for ci in range(1, 5):
        ws.cell(r, ci).border = _TH
        ws.cell(r, ci).alignment = Alignment(horizontal='center')

    b_str_now = cur['buy_count_now'] / cur['K_buy'] if cur['K_buy'] > 0 else 0
    s_str_now = cur['sell_count_now'] / cur['K_sell'] if cur['K_sell'] > 0 else 0
    r += 1
    ws.cell(r, 1).value = '신호 강도'
    ws.cell(r, 1).font = Font(size=10, italic=True, color='606060')
    ws.cell(r, 2).value = f"B={b_str_now:.0%} / S={s_str_now:.0%}"
    ws.cell(r, 3).value = '둘다 ON 시 강한 쪽 우선'
    ws.cell(r, 3).font = Font(size=9, italic=True, color='606060')
    if cur['buy_on_now'] and cur['sell_on_now']:
        ws.cell(r, 4).value = '⚔ 충돌'
        ws.cell(r, 4).fill = _CONF
        ws.cell(r, 4).font = Font(size=11, bold=True, color='806000')
    for ci in range(1, 5):
        ws.cell(r, ci).border = _TH
        ws.cell(r, ci).alignment = Alignment(horizontal='center')

    ar = r + 2
    ws.cell(ar, 1).value = '➡ 다음 액션'
    ws.cell(ar, 1).font = Font(bold=True, size=13, color='1F3864')
    if '보유' in cur['position']:
        if cur['sell_on_now']:
            if cur['buy_on_now'] and b_str_now > s_str_now:
                msg, color = ('⚔ 충돌→매수 우세 → 청산 보류 (보유 유지)', '806000')
            else:
                msg, color = ('🔴 매도 ON → 다음 거래일 청산', 'C00000')
        else:
            msg, color = ('📈 보유 유지', '0070C0')
    else:
        if cur['buy_on_now']:
            if cur['sell_on_now'] and s_str_now > b_str_now:
                msg, color = ('⚔ 충돌→매도 우세 → 매수 보류 (현금 유지)', '806000')
            else:
                msg, color = ('🟢 매수 ON → 다음 거래일 진입', '006100')
        else:
            msg, color = ('💵 현금 유지', '606060')
    ws.cell(ar, 2).value = msg
    ws.cell(ar, 2).font = Font(bold=True, size=12, color=color)
    ws.cell(ar, 2).fill = _HL
    ws.merge_cells(start_row=ar, start_column=2, end_row=ar, end_column=4)
    ws.row_dimensions[ar].height = 25

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 18

    # 2. 사용된 설정
    ws = wb.create_sheet('사용된 설정'); ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = f'{ticker} — 메타 + 앙상블 자동 선정'
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells('A1:B1')
    if globals().get('SELECTION_PRIORITY', 'balacc_return') == 'sell_buy_return':
        sel_method = f'매도성공률 → 매수성공률 → 누적수익 (밴드 {selection_tolerance*100:.1f}%p)'
    else:
        sel_method = (f'1차 평균 BalAcc top band {selection_tolerance*100:.1f}%p 내 → 2차 수익률 최대'
                       if selection_tolerance > 0
                       else '1차 평균 BalAcc → 2차 수익률 (strict)')
    anchor_method = 'OFF'
    if anchor_mode:
        anchor_method = (f"ON ⚓ AUTO 자동계산 (매수 {len(anchor_buy_dates or [])}일, 매도 {len(anchor_sell_dates or [])}일)"
                         if auto_anchor else
                         f"ON ⚓ 수동입력 (매수 {len(anchor_buy_dates or [])}일, 매도 {len(anchor_sell_dates or [])}일)")
    def _bi_get(k, default=None):
        """best_inner에서 키를 안전하게 추출 (재현 모드는 일부 키가 없을 수 있음)."""
        try:
            if hasattr(best_inner, 'get'):
                v = best_inner.get(k, default)
            else:
                v = best_inner[k] if k in best_inner else default
        except Exception:
            return default
        try:
            if v is None or (pd.notna(v) is False):
                return default
        except Exception:
            pass
        return v

    info = [
        ('티커', ticker),
        ('백테스트 기간', f"{daily['date'].iloc[0].date()} ~ {daily['date'].iloc[-1].date()}"),
        ('총 거래일 수', f"{len(daily)}일"),
        ('거래 비용 (왕복)', f"{COST_PER_TRADE*100:.2f}%"),
        ('', ''),
        ('★ 평가 기준 (고정)', ''),
        ('  Horizon', f"{horizon}일"),
        ('  매수 손실 한도', f"-{dd_limit*100:.1f}% 회피 확률"),
        ('  매도 상승 한도', f"+{ru_limit*100:.1f}% 회피 확률"),
        ('', ''),
        ('★ 예측로직 개선 (검증은 검증_예측로직 시트)', ''),
        ('  리드타임 탐색', (f"ON — 선행일 후보 {globals().get('LEAD_HORIZONS')}"
                              if globals().get('LEAD_TIME_SEARCH', False) else 'OFF')),
        ('  신호 지연 정렬', (f"ON — 최대 {globals().get('LEAD_SHIFT_MAX')}일 "
                               f"(훈련 스킬 +{float(globals().get('LEAD_SHIFT_MIN_GAIN',0.05))*100:.0f}%p↑ 시 채택)"
                               if globals().get('LEAD_SHIFT_ENABLED', False) else 'OFF')),
        ('  선출단계 리드 반영', ('ON — 풀 선출 점수화에서 지연 동시 탐색 (선행 지표 풀 진입 가능)'
                                   if (globals().get('LEAD_SELECT_IN_SCORING', True)
                                       and globals().get('LEAD_SHIFT_ENABLED', False)) else 'OFF')),
        ('  스킬 필터 (기저확률 초과)', (f"ON — 스킬 ≥ {float(globals().get('POOL_MIN_SKILL',0.0))*100:.0f}%p"
                                          if globals().get('POOL_REQUIRE_SKILL', False) else 'OFF')),
        ('  홀드아웃 가드', (f"ON — 뒤 {float(globals().get('POOL_HOLDOUT_FRACTION',0.3))*100:.0f}% 스킬 "
                              f"{float(globals().get('POOL_HOLDOUT_MIN_SKILL',-0.1))*100:.0f}%p 미만 제외"
                              if globals().get('POOL_HOLDOUT_GUARD', False) else 'OFF')),
        ('  삼중배리어 (리드 판정)', 'ON' if globals().get('LEAD_TRIPLE_BARRIER', False) else 'OFF'),
        ('', ''),
        ('★ 자동 선정 — 메타 변수', ''),
        ('  WILSON_Z', f"{best_meta['wilson_z']}"),
        ('  PCT 분위 범위', f"({best_meta['pct_low']:.0f}, {best_meta['pct_high']:.0f})"),
        ('  MIN_SIGNALS', f"{best_meta['min_signals']}"),
        ('  DIVERSITY_CORR_LIMIT', f"{best_meta['corr_limit']}"),
        ('  TOP_N_POOL_BUY', f"{best_meta['top_n_pool_buy']}"),
        ('  TOP_N_POOL_SELL', f"{best_meta['top_n_pool_sell']}"),
        ('', ''),
        ('★ 자동 선정 — 앙상블 구성', ''),
        ('  K_buy', f"{int(best_inner['K_buy'])}개"),
        ('  vote_buy', f"{int(best_inner['vote_buy'])}개 이상 ON"),
        ('  K_sell', f"{int(best_inner['K_sell'])}개"),
        ('  vote_sell', f"{int(best_inner['vote_sell'])}개 이상 ON"),
        ('  투표 방식', '가중 투표 (성공률 비례)' if cur.get('weighted_vote') else '일반 투표 (모두 1표)'),
        ('', ''),
        ('★ 충돌 해결 규칙', '둘 다 ON 시 count/K 비율 큰 쪽 우선'),
        ('', ''),
        ('★ 필터 제약', ''),
        ('  MDD 한도', f"-{abs(mdd_limit_pct)*100:.2f}%" if mdd_limit_pct is not None else "없음"),
        ('  최소 거래수', f"{min_trades_daily}회"),
        ('  손절매 한도', f"-{stop_loss_pct*100:.2f}%" if (stop_loss_pct is not None and stop_loss_pct > 0) else "없음"),
        ('', ''),
        ('── 결과 ──', ''),
        ('  B&H 누적', f"{bh_ret*100:+.2f}%"),
        ('  B&H CAGR', f"{bh_cagr*100:+.2f}%"),
        ('  전략 누적', f"{cur['cum_return_pct']:+.2f}%"),
        ('  vs B&H', f"{cur['cum_return_pct'] - bh_ret*100:+.2f}%p"),
        ('  📈 상승일만 합산 (전략)', f"{cur.get('up_cum_return_pct', 0):+.2f}%"),
        ('  📈 상승일만 합산 (B&H)', f"{bh_up_ret*100:+.2f}%" if bh_up_ret is not None else '—'),
        ('  거래 / 승률', f"{cur['n_trades']}회 / {cur['win_rate']*100:.1f}%"),
        ('  손절매 발동', f"{cur['n_stop_triggered']}회"),
        ('  Sharpe-like', (f"{_bi_get('sharpe_like'):.2f}"
                            if _bi_get('sharpe_like') is not None else '—')),
        ('  최대 거래손실', f"{cur['max_drawdown']*100:.2f}%"),
        ('', ''),
        ('★ 선정 우선순위', sel_method),
        ('★ ANCHOR 보정', anchor_method),
        ('', ''),
        (f'  ★ 매수신호 성공/실패 (전체)',
          f"{cur.get('n_buy_success_cnt',0)} / {cur.get('n_buy_fail_cnt',0)} (신호 {cur.get('n_buy_on_total',0)}회)"),
        (f'  ★ 매도신호 성공/실패 (전체)',
          f"{cur.get('n_sell_success_cnt',0)} / {cur.get('n_sell_fail_cnt',0)} (신호 {cur.get('n_sell_on_total',0)}회)"),
        (f'  ★ 매수신호 Balanced Acc',
          f"{cur['buy_success_rate']*100:.1f}% (plain {cur['buy_accuracy_plain']*100:.1f}%, ON {cur['n_buy_signal_on']}일)"),
        (f'  ★ 매도신호 Balanced Acc',
          f"{cur['sell_success_rate']*100:.1f}% (plain {cur['sell_accuracy_plain']*100:.1f}%, ON {cur['n_sell_signal_on']}일)"),
        (f'  ★ 평균 Balanced Acc',
          f"{cur['avg_success_rate']*100:.1f}%"),
    ]
    if anchor_mode:
        info += [
            (f'  ⚓ 매수 정답일 매칭',
              f"{cur['n_anchor_buy_caught']}/{cur['n_anchor_buy']}"),
            (f'  ⚓ 매도 정답일 매칭',
              f"{cur['n_anchor_sell_caught']}/{cur['n_anchor_sell']}"),
        ]
    info += [
        (f'  ★ 충돌 발생일',
          f"{cur['n_conflicts']}일 (B승 {cur['n_conflict_buy_won']} / S승 {cur['n_conflict_sell_won']})"),
    ]
    for ri, (k, v) in enumerate(info, 3):
        ws.cell(ri, 1).value = k
        ws.cell(ri, 1).font = Font(bold=True, size=11, color='1F3864')
        ws.cell(ri, 2).value = v
        ws.cell(ri, 2).font = Font(size=11)
        if k.startswith('★') or '★' in k or k.startswith('⚓') or '⚓' in k:
            ws.cell(ri, 1).fill = _HL; ws.cell(ri, 2).fill = _HL
            ws.cell(ri, 2).font = Font(bold=True, size=12, color='C00000')
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 65

    # 3. 메타 그리드 결과
    ws = wb.create_sheet('메타_그리드_결과'); ws.sheet_view.showGridLines = False
    sort_label = sel_method
    ws.cell(1, 1).value = f'메타 변수 조합별 결과 ({len(meta_results_df)}개) — {sort_label}'
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells('A1:Q1')
    _hdr(ws, 3, ['#', 'wilson_z', 'pct', 'min_sig', 'corr', 'pool',
                 '풀 매수', '풀 매도', '통과/전체',
                 '평균성공', '매수성공', '매도성공',
                 '★ 수익', 'MDD', '거래', 'K_b/v_b', 'K_s/v_s'])
    for ri, row in meta_results_df.head(80).iterrows():
        r = ri + 4
        is_best = (pd.notna(row['best_K_buy']) and
                   row['wilson_z'] == best_meta['wilson_z'] and
                   row['pct_low'] == best_meta['pct_low'] and
                   row['min_signals'] == best_meta['min_signals'] and
                   row['corr_limit'] == best_meta['corr_limit'] and
                   row['top_n_pool_buy'] == best_meta['top_n_pool_buy'] and
                   int(row['best_K_buy']) == int(best_inner['K_buy']))
        marker = '★' if is_best else ''
        vals = [
            f"{ri+1}{marker}",
            row['wilson_z'],
            f"({int(row['pct_low'])},{int(row['pct_high'])})",
            int(row['min_signals']),
            row['corr_limit'],
            int(row['top_n_pool_buy']),
            int(row['n_buy_pool']),
            int(row['n_sell_pool']),
            f"{int(row['n_inner_passed'])}/{int(row['n_inner_total'])}",
            f"{row['best_avg_sr']*100:.1f}%"  if pd.notna(row['best_avg_sr'])  else '—',
            f"{row['best_buy_sr']*100:.1f}%"  if pd.notna(row['best_buy_sr'])  else '—',
            f"{row['best_sell_sr']*100:.1f}%" if pd.notna(row['best_sell_sr']) else '—',
            f"{row['best_return']*100:+.2f}%" if pd.notna(row['best_return']) else '—',
            f"{row['best_mdd']*100:.2f}%" if pd.notna(row['best_mdd']) else '—',
            int(row['best_n_trades']) if pd.notna(row['best_mdd']) else '—',
            f"{int(row['best_K_buy'])}/{int(row['best_vote_buy'])}" if pd.notna(row['best_K_buy']) else '—',
            f"{int(row['best_K_sell'])}/{int(row['best_vote_sell'])}" if pd.notna(row['best_K_sell']) else '—',
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = _TH
            c.alignment = Alignment(horizontal='center'); c.font = Font(size=10)
            if is_best:
                c.fill = _HL
                c.font = Font(bold=True, color='C00000', size=11)
            elif ri % 2 == 1:
                c.fill = _ALT
        if pd.notna(row['best_avg_sr']):
            ws.cell(r, 10).fill = _success_fill(row['best_avg_sr'])
            ws.cell(r, 11).fill = _success_fill(row['best_buy_sr'])
            ws.cell(r, 12).fill = _success_fill(row['best_sell_sr'])
            if is_best:
                for cc in (10, 11, 12):
                    ws.cell(r, cc).font = Font(bold=True, color='C00000', size=11)
        if pd.notna(row['best_return']):
            ws.cell(r, 13).fill = _ret_fill(row['best_return'], bh_ret)
            if is_best:
                ws.cell(r, 13).font = Font(bold=True, color='C00000', size=12)
    for ci, w in enumerate([6, 10, 10, 9, 8, 8, 9, 9, 11, 11, 10, 10, 11, 10, 8, 11, 11], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'

    # 4. 매수 앙상블 지표
    ws = wb.create_sheet('매수_앙상블_지표'); ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = f'매수 앙상블 {len(buy_used)}개 지표'
    ws.cell(1, 1).font = Font(bold=True, size=13, color='1F3864')
    ws.merge_cells('A1:K1')
    _hdr(ws, 3, ['#', '지표', '방향', '임계치', '분위', '신호수', '성공수', '성공률', '점수', '🔁가중치', '지연(일)'])
    for ri, row in buy_used.reset_index(drop=True).iterrows():
        r = ri + 4
        vals = [ri + 1, row['indicator'], row['direction'],
                round(row['threshold'], 6), f"{row['pct_label']:.0f}%",
                int(row['n_signals']), int(row['n_success']),
                f"{row['success_rate']*100:.2f}%", round(row['score'], 4),
                round(float(row['vote_weight']), 6) if 'vote_weight' in row.index and pd.notna(row.get('vote_weight')) else '—',
                int(row['lead_shift']) if 'lead_shift' in row.index and pd.notna(row.get('lead_shift')) else 0]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = _TH
            c.alignment = Alignment(horizontal='center'); c.font = Font(size=10)
            if ri % 2 == 1: c.fill = _ALT
    for ci, w in enumerate([5, 32, 8, 14, 8, 10, 10, 10, 10, 11, 9], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'

    # 5. 매도 앙상블 지표
    ws = wb.create_sheet('매도_앙상블_지표'); ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = f'매도 앙상블 {len(sell_used)}개 지표'
    ws.cell(1, 1).font = Font(bold=True, size=13, color='1F3864')
    ws.merge_cells('A1:K1')
    _hdr(ws, 3, ['#', '지표', '방향', '임계치', '분위', '신호수', '성공수', '성공률', '점수', '🔁가중치', '지연(일)'])
    for ri, row in sell_used.reset_index(drop=True).iterrows():
        r = ri + 4
        vals = [ri + 1, row['indicator'], row['direction'],
                round(row['threshold'], 6), f"{row['pct_label']:.0f}%",
                int(row['n_signals']), int(row['n_success']),
                f"{row['success_rate']*100:.2f}%", round(row['score'], 4),
                round(float(row['vote_weight']), 6) if 'vote_weight' in row.index and pd.notna(row.get('vote_weight')) else '—',
                int(row['lead_shift']) if 'lead_shift' in row.index and pd.notna(row.get('lead_shift')) else 0]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = _TH
            c.alignment = Alignment(horizontal='center'); c.font = Font(size=10)
            if ri % 2 == 1: c.fill = _ALT
    for ci, w in enumerate([5, 32, 8, 14, 8, 10, 10, 10, 10, 11, 9], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'

    # 6. 내부 그리드 통과
    ws = wb.create_sheet('내부_그리드_통과'); ws.sheet_view.showGridLines = False
    mdd_t2 = f"MDD ≥ -{abs(mdd_limit_pct)*100:.1f}%" if mdd_limit_pct is not None else "MDD 제한 없음"
    use_match_p = bool(anchor_match_priority_arg and anchor_mode and 'anchor_avg_match_rate' in inner_passed.columns)
    inner_sort_label = (f'매칭률 ±{anchor_match_tolerance_arg*100:.1f}%p → BalAcc → 수익률'
                         if use_match_p else sort_label)
    ws.cell(1, 1).value = (f'전체 그리드 통합 (모든 메타변수 조합) — {mdd_t2}, 거래수 ≥ {min_trades_daily}회, B&H 초과만  '
                           f'({len(inner_passed)}개) — {inner_sort_label}')
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells('A1:AA1')
    _hdr(ws, 3, ['#', 'wilson_z', 'pct_low', 'pct_high', 'corr_limit', 'min_sig', 'pool',
                 'K_buy', 'vote_buy', 'K_sell', 'vote_sell',
                 '✅실제승률', '✅실제최대손실', '✅실제누적수익',
                 '📍실행일포지션', '🎬실행일액션', '📍매수카운트', '📍매도카운트',
                 '📅최근매수일', '💲진입가', '📅최근매도일', '💲매도가',
                 '🎯K최적값', '🎯K기준수익(합산)', '🎯K거래수',
                 '🎯보유중하락(전체)', '🎯지표수(매수/매도)'])
    # 통합 테이블은 meta_grid_search에서 이미 선정기준대로 정렬돼 옴 → 그대로 표시
    disp = inner_passed.head(TOP_N_GRID_OUT).reset_index(drop=True)
    n_in_band = 0
    def _g(row, k, default=np.nan):
        try: return row[k]
        except Exception: return default

    # ★★ 요청: 각 그리드 행에 '순신호 K 기준'(행의 K_buy/K_sell 지표개수 적용) 수익·거래·OOS·보유중하락 컬럼 추가.
    #   같은 wilson이라도 K_buy/K_sell이 다르면 풀 슬라이스가 달라 결과가 달라짐 → 행마다 구분됨.
    #   캐시키 = (메타조합, K_buy, K_sell) 로 동일 행만 1회 계산(속도).
    globals().pop('_KNET_BEST_POOL', None); globals().pop('_KNET_BEST_K', None)
    globals().pop('_KNET_BEST_NB', None); globals().pop('_KNET_BEST_NS', None)
    _knet_cache = {}
    _pool_map_k = globals().get('_LAST_POOL_MAP', None)
    def _meta_key_of(row):
        try:
            return (round(float(_g(row, 'meta_wilson_z')), 4),
                    int(_g(row, 'meta_pct_low')), int(_g(row, 'meta_pct_high')),
                    round(float(_g(row, 'meta_corr_limit')), 4))
        except Exception:
            return None
    def _row_kbuy_ksell(row):
        try: return int(_g(row, 'K_buy')), int(_g(row, 'K_sell'))
        except Exception: return None, None
    if feat is not None and close_full is not None and _pool_map_k:
        _oos_k = globals().get('OOS_START')
        _kret = []; _ktr = []; _koos = []; _kk = []; _khd_tr = []; _khd_oos = []; _knb = []
        # ★ K값 실행 상태 (col31-38 채움용): 마지막날 포지션/액션/카운트 + 최근 매수/매도
        _rpos = []; _ract = []; _rbs = []; _rss = []; _rbd = []; _rbp = []; _rsd = []; _rsp = []
        def _knet_exec_state(_r):
            try:
                d = _r['daily']; m = len(d)
                last = d.iloc[m-1]
                rpos = '롱' if int(last['position']) == 1 else '현금'
                ract = str(last['action'])
                rbs = float(last['buy_count']); rss = float(last['sell_count'])
                # 최근 매수(0→1 진입=익일 체결) / 매도(1→0)
                pos = d['position'].values.astype(int)
                dts = list(d['date']); px = d['price'].values.astype(float)
                lbd = lbp = lsd = lsp = None
                for t in range(1, m):
                    if pos[t] == 1 and pos[t-1] == 0: lbd = dts[t]; lbp = px[t]
                    elif pos[t] == 0 and pos[t-1] == 1: lsd = dts[t]; lsp = px[t]
                return rpos, ract, rbs, rss, lbd, lbp, lsd, lsp
            except Exception:
                return None, None, None, None, None, None, None, None
        def _push_empty():
            _kret.append(np.nan); _ktr.append(np.nan); _koos.append(np.nan)
            _kk.append(np.nan); _khd_tr.append(np.nan); _khd_oos.append(np.nan); _knb.append(np.nan)
            for _L in (_rpos, _ract, _rbs, _rss, _rbd, _rbp, _rsd, _rsp): _L.append(None)
        # ★ 메모리/시간 보호: 그리드 행별 net>K는 상위 N행만 계산 (나머지 '—').
        #   net>K 시트는 합친 풀(_KNET_MULTI_POOL)로 별도 계산하므로 여기는 그리드 시트 표시·폴백용일 뿐.
        _KNET_ROW_CAP = int(globals().get('KNET_GRID_ROW_CAP', 100))
        for _ri, (_, _row) in enumerate(disp.iterrows()):
            if _ri >= _KNET_ROW_CAP:      # 캡 초과 → 계산 생략(빈값)
                _push_empty(); continue
            _mk = _meta_key_of(_row); _kb, _ks = _row_kbuy_ksell(_row)
            _pools = _pool_map_k.get(_mk) if _mk else None
            if not _pools or _kb is None:
                _push_empty(); continue
            _ck = (_mk, _kb, _ks)
            if _ck not in _knet_cache:
                _bp, _sp = _pools
                try:
                    _knet_cache[_ck] = _net_signal_k_search(
                        feat, close_full, _bp, _sp, ticker=ticker, oos_start=_oos_k,
                        n_buy=_kb, n_sell=_ks)
                except Exception:
                    _knet_cache[_ck] = None
            _r = _knet_cache[_ck]
            if _r is None:
                _push_empty()
            else:
                _kret.append(_r['full_cum']); _ktr.append(_r['n_trades'])
                _koos.append(_r['oos_cum'] if _r['oos_cum'] is not None else np.nan)
                _kk.append(_r['best_k'])
                _khd_tr.append(_r.get('held_down_train', np.nan))
                _khd_oos.append(_r['held_down_oos'] if _r.get('held_down_oos') is not None else np.nan)
                _knb.append(f"{_r['n_buy_opt']}/{_r['n_sell_opt']}")
                _es = _knet_exec_state(_r)
                _rpos.append(_es[0]); _ract.append(_es[1]); _rbs.append(_es[2]); _rss.append(_es[3])
                _rbd.append(_es[4]); _rbp.append(_es[5]); _rsd.append(_es[6]); _rsp.append(_es[7])
        disp['knet_ret'] = _kret; disp['knet_trades'] = _ktr
        disp['knet_oos'] = _koos; disp['knet_k'] = _kk
        disp['knet_hd_train'] = _khd_tr; disp['knet_hd_oos'] = _khd_oos
        disp['knet_nbns'] = _knb
        disp['knet_rpos'] = _rpos; disp['knet_ract'] = _ract
        disp['knet_rbs'] = _rbs; disp['knet_rss'] = _rss
        disp['knet_lbd'] = _rbd; disp['knet_lbp'] = _rbp
        disp['knet_lsd'] = _rsd; disp['knet_lsp'] = _rsp
        # ★ 최적 그리드 = 'K기준 OOS 수익' 최고
        _sort_col = 'knet_oos' if disp['knet_oos'].notna().any() else 'knet_ret'
        if disp[_sort_col].notna().any():
            disp = disp.sort_values(_sort_col, ascending=False, na_position='last').reset_index(drop=True)
            _sort_kor = 'K기준 OOS수익' if _sort_col == 'knet_oos' else 'K기준 전체수익(합산)'
            ws.cell(1, 1).value = (f'전체 그리드 통합 — ★ {_sort_kor} 높은 순 정렬·선정  '
                                   f'({len(inner_passed)}개)')
            # 최적 그리드(1행)의 풀+K_buy/K_sell을 순신호/일별 시트가 쓰도록 저장 (일관성)
            try:
                _bk = _meta_key_of(disp.iloc[0])
                _bkb, _bks = _row_kbuy_ksell(disp.iloc[0])
                if _bk and _pool_map_k and _pool_map_k.get(_bk):
                    globals()['_KNET_BEST_POOL'] = _pool_map_k[_bk]
                    globals()['_KNET_BEST_NB'] = _bkb
                    globals()['_KNET_BEST_NS'] = _bks
                    _bestrow_k = disp.iloc[0].get('knet_k') if 'knet_k' in disp.columns else None
                    globals()['_KNET_BEST_K'] = _bestrow_k
            except Exception:
                pass
    _knet_cache.clear()   # ★ daily 표까지 담긴 캐시 즉시 비움 (메모리 회수)
    _has_knet = ('knet_ret' in disp.columns and disp['knet_ret'].notna().any())

    for ri, row in disp.iterrows():
        r = ri + 4
        # ★ K/vote + 메타변수(wilson_z·corr·pct)까지 모두 일치해야 best (같은 K/vote라도
        #   메타변수 다르면 풀이 달라 수익이 다름 — 메타변수 무시하면 엉뚱한 행에 ★가 찍힘)
        def _bi(k):
            try:
                v = best_inner[k]
                return v
            except Exception:
                try: return best_inner.get(k)
                except Exception: return None
        _bi_wz = _bi('meta_wilson_z'); _bi_cl = _bi('meta_corr_limit')
        _bi_pl = _bi('meta_pct_low');  _bi_ph = _bi('meta_pct_high')
        _row_wz = _g(row,'meta_wilson_z'); _row_cl = _g(row,'meta_corr_limit')
        _row_pl = _g(row,'meta_pct_low'); _row_ph = _g(row,'meta_pct_high')
        _meta_match = True
        if _bi_wz is not None and pd.notna(_row_wz):
            _meta_match = (abs(float(_row_wz)-float(_bi_wz))<1e-9 and
                           (_bi_cl is None or abs(float(_row_cl)-float(_bi_cl))<1e-9) and
                           (_bi_pl is None or int(_row_pl)==int(_bi_pl)) and
                           (_bi_ph is None or int(_row_ph)==int(_bi_ph)))
        is_best = (int(row['K_buy']) == int(best_inner['K_buy']) and
                   int(row['K_sell']) == int(best_inner['K_sell']) and
                   int(row['vote_buy']) == int(best_inner['vote_buy']) and
                   int(row['vote_sell']) == int(best_inner['vote_sell']) and
                   _meta_match)
        # ★ 요청: K기준 수익 정렬이 적용됐으면 ★ = 1행(K기준 수익 최고) = 최적 그리드
        if _has_knet:
            is_best = (ri == 0)
        marker = '★' if is_best else ''
        wz_v   = _g(row, 'meta_wilson_z')
        pl_v   = _g(row, 'meta_pct_low')
        ph_v   = _g(row, 'meta_pct_high')
        cl_v   = _g(row, 'meta_corr_limit')
        ms_v   = _g(row, 'meta_min_signals')
        pool_v = _g(row, 'meta_top_n_pool')
        bm_v   = _g(row, 'anchor_buy_match_rate')
        sm_v   = _g(row, 'anchor_sell_match_rate')
        vals = [
            f"{ri+1}{marker}",
            f"{wz_v:.2f}"  if pd.notna(wz_v)  else '—',
            int(pl_v)      if pd.notna(pl_v)  else '—',
            int(ph_v)      if pd.notna(ph_v)  else '—',
            f"{cl_v:.2f}"  if pd.notna(cl_v)  else '—',
            int(ms_v)      if pd.notna(ms_v)  else '—',
            int(pool_v)    if pd.notna(pool_v)else '—',
            int(row['K_buy']), int(row['vote_buy']),
            int(row['K_sell']), int(row['vote_sell']),
            (f"{_g(row,'real_win_rate')*100:.1f}%"     if pd.notna(_g(row,'real_win_rate'))     else '—'),
            (f"{_g(row,'real_max_drawdown')*100:.2f}%" if pd.notna(_g(row,'real_max_drawdown')) else '—'),
            (f"{_g(row,'real_total_return')*100:+.2f}%" if pd.notna(_g(row,'real_total_return')) else '—'),
            # ★ 실행일 K값 상태 + 최근 K값 매수/매도 (요청: net>K 기준)
            (str(_g(row,'knet_rpos')) if (_g(row,'knet_rpos') is not None and str(_g(row,'knet_rpos'))!='nan') else '—'),
            (str(_g(row,'knet_ract')) if (_g(row,'knet_ract') is not None and str(_g(row,'knet_ract'))!='nan') else '—'),
            (f"{float(_g(row,'knet_rbs')):.2f}"  if pd.notna(_g(row,'knet_rbs'))  else '—'),
            (f"{float(_g(row,'knet_rss')):.2f}"  if pd.notna(_g(row,'knet_rss'))  else '—'),
            (pd.Timestamp(_g(row,'knet_lbd')).date()  if pd.notna(_g(row,'knet_lbd'))  else '—'),
            (f"{float(_g(row,'knet_lbp')):.2f}"  if pd.notna(_g(row,'knet_lbp'))  else '—'),
            (pd.Timestamp(_g(row,'knet_lsd')).date()  if pd.notna(_g(row,'knet_lsd'))  else '—'),
            (f"{float(_g(row,'knet_lsp')):.2f}"  if pd.notna(_g(row,'knet_lsp'))  else '—'),
            # 🎯 K기준 (순신호 net>K, 상승·하락률 합산) — OOS 미사용
            (round(float(_g(row,'knet_k')),3) if pd.notna(_g(row,'knet_k')) else '—'),
            (f"{_g(row,'knet_ret')*100:+.2f}%"        if pd.notna(_g(row,'knet_ret')) else '—'),
            (int(_g(row,'knet_trades'))               if pd.notna(_g(row,'knet_trades')) else '—'),
            (f"{_g(row,'knet_hd_train')*100:+.2f}%"   if pd.notna(_g(row,'knet_hd_train')) else '—'),
            (str(_g(row,'knet_nbns')) if _g(row,'knet_nbns') is not None and str(_g(row,'knet_nbns'))!='nan' else '—'),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = _TH
            c.alignment = Alignment(horizontal='center'); c.font = Font(size=10)
            if is_best:
                c.fill = _HL
                c.font = Font(bold=True, color='C00000', size=11)
            elif ri % 2 == 1:
                c.fill = _ALT
        # 실제 수치 색칠 (12승률, 13최대손실, 14누적수익)
        _rw = _g(row, 'real_win_rate'); _rm = _g(row, 'real_max_drawdown'); _rt = _g(row, 'real_total_return')
        if pd.notna(_rw): ws.cell(r, 12).fill = _success_fill(_rw)
        if pd.notna(_rm): ws.cell(r, 13).fill = _mdd_fill(_rm, mdd_limit_pct)
        if pd.notna(_rt): ws.cell(r, 14).fill = _ret_fill(_rt, bh_ret)
        # 실행일 K값 포지션(15) 색칠 — 롱/현금
        _rp = _g(row, 'knet_rpos')
        if isinstance(_rp, str) and _rp == '롱':   ws.cell(r, 15).fill = _HOLD
        elif isinstance(_rp, str) and _rp == '현금': ws.cell(r, 15).fill = _CASH
        # 🎯 K 색칠 (24 전체수익)
        if pd.notna(_g(row,'knet_ret')): ws.cell(r, 24).fill = _ret_fill(_g(row,'knet_ret'), bh_ret)
        if is_best:
            ws.cell(r, 14).font = Font(bold=True, color='C00000', size=12)
            ws.cell(r, 23).font = Font(bold=True, color='C00000', size=12)
    for ci, w in enumerate([6, 9, 8, 8, 9, 8, 7, 8, 9, 8, 9,
                            11, 13, 13,
                            12, 14, 11, 11, 13, 10, 13, 10,
                            10, 15, 10, 14, 14], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'

    # ★ 순신호 K 결과 — 전체수익 최고 K / OOS수익 최고 K 2개 계산 (풀은 공유).
    _nsd_main = None; _nsd_full = None; _nsd_oos = None
    try:
        _replay_fixed_done = False
        _mp = globals().get('_KNET_MULTI_POOL')
        _use_multi = (_mp and isinstance(_mp, tuple) and _mp[0] == ticker
                      and _mp[1] is not None and _mp[2] is not None
                      and len(_mp[1]) > 0 and len(_mp[2]) > 0)
        if _use_multi:
            _mbp, _msp = _mp[1], _mp[2]; _nb = _ns = None; _sc_cnt = True
            print(f"  ★ net>K = 합친 다중임계 풀 (매수 {len(_mbp)}행 / 매도 {len(_msp)}행)")
        else:
            _kp = globals().get('_KNET_BEST_POOL')
            _mbp, _msp = (_kp if (_kp and _kp[0] is not None and _kp[1] is not None) else (buy_pool, sell_pool))
            _nb = globals().get('_KNET_BEST_NB'); _ns = globals().get('_KNET_BEST_NS'); _sc_cnt = False
        _has_oos_now = (globals().get('OOS_START') is not None)
        _fixed = globals().get('_KNET_REPLAY_FIXED')

        def _compute(sel_by, fk=None, fnb=None, fns=None, fg=None):
            """sel_by='full'|'oos'. fk 있으면 고정(재현), 없으면 g 탐색."""
            if fk is not None:
                c = _net_signal_k_search(feat, close_full, _mbp, _msp, ticker=ticker,
                                         oos_start=globals().get('OOS_START'),
                                         n_buy=fnb, n_sell=fns, search_counts=False,
                                         weight_exp=(fg or 1.0), fixed_k=fk, select_by=sel_by)
                return c, (fg or 1.0)
            _gs = (list(globals().get('NET_WEIGHT_SCHEMES', [1.0]))
                   if (globals().get('SEARCH_WEIGHT_SCHEME', False) and globals().get('NET_SIGNAL_WEIGHTED', False))
                   else [1.0])
            _bg = 1.0; _bsc = -1e18; _bc = None
            for _g_ in _gs:
                c = _net_signal_k_search(feat, close_full, _mbp, _msp, ticker=ticker,
                                         oos_start=globals().get('OOS_START'),
                                         n_buy=_nb, n_sell=_ns, search_counts=_sc_cnt,
                                         weight_exp=_g_, select_by=sel_by)
                _m = ((c.get('oos_cum') if sel_by == 'oos' else c.get('full_cum')) if c else None)
                _m = -1e18 if _m is None else _m
                if _m > _bsc: _bsc = _m; _bg = _g_; _bc = c
            if _bc is None:
                _bc = _net_signal_k_search(feat, close_full, _mbp, _msp, ticker=ticker,
                                           oos_start=globals().get('OOS_START'),
                                           n_buy=_nb, n_sell=_ns, search_counts=_sc_cnt, select_by=sel_by)
            return _bc, _bg

        if _fixed and isinstance(_fixed, dict) and _fixed.get('k_full') is not None:
            # ★ 재현: 저장된 전체/OOS K 그대로 (탐색 0)
            _nsd_full, _gf = _compute('full', _fixed.get('k_full'), _fixed.get('nb_full'),
                                      _fixed.get('ns_full'), _fixed.get('g_full'))
            if _has_oos_now and _fixed.get('k_oos') is not None:
                _nsd_oos, _go = _compute('oos', _fixed.get('k_oos'), _fixed.get('nb_oos'),
                                         _fixed.get('ns_oos'), _fixed.get('g_oos'))
            print(f"  ♻ 재현: 저장된 K(전체={_fixed.get('k_full')}, OOS={_fixed.get('k_oos')}) 그대로 적용")
            _replay_fixed_done = True
            globals()['_KNET_BEST_WEXP'] = (_fixed.get('g_full') or 1.0)
        else:
            _nsd_full, _gf = _compute('full')
            globals()['_KNET_BEST_WEXP'] = _gf
            print(f"  ★ 전체수익 최고 K = {_nsd_full['best_k'] if _nsd_full else '—'} "
                  f"(전체 {(_nsd_full['full_cum']*100 if _nsd_full else 0):+.1f}%)")
            if _has_oos_now:
                _nsd_oos, _go = _compute('oos')
                print(f"  ★ OOS수익 최고 K = {_nsd_oos['best_k'] if _nsd_oos else '—'} "
                      f"(OOS {(_nsd_oos['oos_cum']*100 if (_nsd_oos and _nsd_oos.get('oos_cum') is not None) else 0):+.1f}%)")
        _nsd_main = _nsd_full   # 기본(하위호환)
        globals()['_KNET_FULL'] = _nsd_full; globals()['_KNET_OOS'] = _nsd_oos
    except Exception as _em:
        print(f"  ⚠ 순신호 K 계산 실패(무시): {_em}")

    # 7. 거래 내역 — ★ K값 거래만 (요청: net>K 매수 → net≤K 매도 라운드트립). 레거시 거래 제외.
    ws = wb.create_sheet('거래 내역'); ws.sheet_view.showGridLines = False
    _ntr = []
    if _nsd_main is not None:
        d = _nsd_main['daily']
        pos = d['position'].values.astype(int)
        prices = d['price'].values.astype(float)
        dts = list(d['date']); rets = d['day_ret'].values.astype(float)
        ent_i = None
        for t in range(1, len(pos)):
            if pos[t] == 1 and pos[t-1] == 0:
                ent_i = t
            elif pos[t] == 0 and pos[t-1] == 1 and ent_i is not None:
                _ex = '매도신호'
                _ioos = int(d.iloc[ent_i]['is_oos']) == 1
                _ntr.append((ent_i, t, float(np.sum(rets[ent_i:t+1])), _ex, _ioos)); ent_i = None
        if ent_i is not None:
            _ntr.append((ent_i, len(pos)-1, float(np.sum(rets[ent_i:len(pos)])), '보유중', int(d.iloc[ent_i]['is_oos'])==1))
    _bk_t = (round(_nsd_main['best_k'],3) if (_nsd_main and _nsd_main.get('weighted')) else (_nsd_main['best_k'] if _nsd_main else '-'))
    ws.cell(1, 1).value = f'K값 거래 내역 — 총 {len(_ntr)}건 (net>K({_bk_t}) 매수 → net≤K 매도, 수익=구간 일별수익 합산)'
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864'); ws.merge_cells('A1:I1')
    _hdr(ws, 3, ['#', '진입일', '진입가', '청산일', '청산가', '보유일', '수익%(합산)', '누적수익%', '구간'])
    _cum = 0.0
    for _ix, (ei, xi, segret, exr, ioos) in enumerate(_ntr):
        r = _ix + 4; _cum += segret
        vals = [_ix+1, pd.Timestamp(dts[ei]).strftime('%Y-%m-%d'), f"${prices[ei]:.2f}",
                pd.Timestamp(dts[xi]).strftime('%Y-%m-%d'), f"${prices[xi]:.2f}",
                int(xi-ei), f"{segret*100:+.2f}%", f"{_cum*100:+.2f}%", 'OOS' if ioos else '학습']
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = _TH
            c.alignment = Alignment(horizontal='center'); c.font = Font(size=10)
            if _ix % 2 == 1: c.fill = _ALT
        ws.cell(r, 7).fill = _GOOD if segret > 0 else _BAD
        ws.cell(r, 7).font = Font(bold=True, size=10, color='006100' if segret > 0 else 'C00000')
    for ci, w in enumerate([5, 12, 11, 12, 11, 8, 13, 13, 7], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'
    print(f"  ✓ 거래 내역: K값 거래 {len(_ntr)}건")
    try:
        _pf = float(globals().get('CORRECT_PROFIT_FLOOR', 0.01))
        _fidx = daily['date'] if (daily is not None and hasattr(daily, 'columns') and 'date' in daily.columns) else []
        diag_df = diagnose_trades(trades, close=None, feat_index=_fidx,
                                  anchor_prio=anchor_prio, profit_floor=_pf)
    except Exception as _de:
        diag_df = None
    if diag_df is not None and len(diag_df) > 0:
        ws = wb.create_sheet('🔧 거래 진단'); ws.sheet_view.showGridLines = False
        n_wrong = int(diag_df['보정대상'].sum())
        ws.cell(1, 1).value = (f'거래 진단 — 총 {len(diag_df)}건 중 보정 대상(손실 또는 +{_pf*100:.0f}% 이하 수익) {n_wrong}건  '
                               f'※ 앵커(우선순위1 최대수익스윙 / 우선순위2 {_pf*100:.0f}%이상)와 매칭 진단')
        ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
        ws.merge_cells('A1:K1')
        _hdr(ws, 3, ['#', '진입일', '청산일', '비용후수익%', '보유일', '청산사유',
                     '매수 진단', '매도 진단', '매수앵커일치', '매도앵커일치', '판정'])
        for ri, drow in diag_df.iterrows():
            r = ri + 4
            _ed = drow['entry_date']; _xd = drow['exit_date']
            vals = [int(drow['trade_no']),
                    (_ed.date() if pd.notna(_ed) else '—'),
                    (_xd.date() if pd.notna(_xd) else '—'),
                    f"{drow['net_return_%']:+.2f}%",
                    int(drow['days_held']), drow['exit_reason'],
                    drow['매수_진단'], drow['매도_진단'],
                    ('✓' if drow['매수_앵커일치'] else '✗'),
                    ('✓' if drow['매도_앵커일치'] else '✗'),
                    drow['판정']]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(r, ci); c.value = v; c.border = _TH
                c.alignment = Alignment(horizontal='center'); c.font = Font(size=10)
                if ri % 2 == 1: c.fill = _ALT
            if bool(drow['보정대상']):
                for ci in range(1, 12):
                    ws.cell(r, ci).fill = PatternFill('solid', fgColor='FCE4D6')
                ws.cell(r, 4).font = Font(bold=True, color='C00000', size=10)
        for ci, w in enumerate([6, 13, 13, 12, 8, 12, 22, 22, 12, 12, 28], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = 'A4'
    if anchor_mode and (cur.get('anchor_buy_diagnosis') or cur.get('anchor_sell_diagnosis')):
        ws = wb.create_sheet('⚓ ANCHOR 진단'); ws.sheet_view.showGridLines = False
        buy_diag = cur.get('anchor_buy_diagnosis', [])
        sell_diag = cur.get('anchor_sell_diagnosis', [])
        n_buy_miss  = sum(1 for d in buy_diag  if not d.get('executed_match', d.get('signal_on', False)))
        n_sell_miss = sum(1 for d in sell_diag if not d.get('executed_match', d.get('signal_on', False)))

        title_extra = ' (AUTO 자동계산)' if auto_anchor else ' (수동 입력)'
        ws.cell(1, 1).value = (f'⚓ ANCHOR 정답일 진단{title_extra} — '
                                f'매수 {len(buy_diag)}개 중 잡힘 '
                                f'{len(buy_diag)-n_buy_miss}/{n_buy_miss} 못잡힘,  '
                                f'매도 {len(sell_diag)}개 중 잡힘 '
                                f'{len(sell_diag)-n_sell_miss}/{n_sell_miss} 못잡힘')
        ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
        ws.merge_cells('A1:F1')

        ws.cell(3, 1).value = '※ 매칭률은 실제 일별 거래 결과 기준 — 신호만 떠도 충돌 해결로 반대 방향 갔으면 unmatched'
        ws.cell(3, 1).font = Font(size=10, color='606060', italic=True)
        ws.merge_cells('A3:F3')
        ws.cell(4, 1).value = '    개선 방법: MAX_INDICATORS↑, N_THRESHOLDS↑, K_BUY/SELL_RANGE 확장, VOTE_RATIO 낮은 비율 추가, top_n_pool↑'
        ws.cell(4, 1).font = Font(size=10, color='606060', italic=True)
        ws.merge_cells('A4:F4')

        ws.cell(6, 1).value = f'■ 매수 정답일 진단 ({len(buy_diag)}개)'
        ws.cell(6, 1).font = Font(bold=True, size=12, color='006100')
        _hdr(ws, 7, ['#', '정답 매수일', '신호 떠야 했던 날',
                      f'buy_count (/{cur["K_buy"]})', f'필요 vote ({cur["vote_buy"]})', '결과'])
        for ri, d in enumerate(buy_diag):
            r = 8 + ri
            caught = d.get('executed_match', d.get('signal_on', False))
            vals = [ri+1,
                    d['target_date'].date() if hasattr(d['target_date'], 'date') else d['target_date'],
                    d['signal_day'].date() if hasattr(d['signal_day'], 'date') else d['signal_day'],
                    round(float(d['buy_count']), 2),
                    int(cur['vote_buy']),
                    d['status']]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(r, ci); c.value = v; c.border = _TH
                c.alignment = Alignment(horizontal='center'); c.font = Font(size=10)
                if ri % 2 == 1: c.fill = _ALT
            if caught:
                ws.cell(r, 6).fill = _GOOD
                ws.cell(r, 6).font = Font(bold=True, size=10, color='006100')
            else:
                ws.cell(r, 6).fill = _BAD
                ws.cell(r, 6).font = Font(bold=True, size=10, color='C00000')

        sell_start = 8 + len(buy_diag) + 3
        ws.cell(sell_start - 1, 1).value = f'■ 매도 정답일 진단 ({len(sell_diag)}개)'
        ws.cell(sell_start - 1, 1).font = Font(bold=True, size=12, color='C00000')
        _hdr(ws, sell_start, ['#', '정답 매도일', '신호 떠야 했던 날',
                               f'sell_count (/{cur["K_sell"]})', f'필요 vote ({cur["vote_sell"]})', '결과'])
        for ri, d in enumerate(sell_diag):
            r = sell_start + 1 + ri
            caught = d.get('executed_match', d.get('signal_on', False))
            vals = [ri+1,
                    d['target_date'].date() if hasattr(d['target_date'], 'date') else d['target_date'],
                    d['signal_day'].date() if hasattr(d['signal_day'], 'date') else d['signal_day'],
                    round(float(d['sell_count']), 2),
                    int(cur['vote_sell']),
                    d['status']]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(r, ci); c.value = v; c.border = _TH
                c.alignment = Alignment(horizontal='center'); c.font = Font(size=10)
                if ri % 2 == 1: c.fill = _ALT
            if caught:
                ws.cell(r, 6).fill = _GOOD
                ws.cell(r, 6).font = Font(bold=True, size=10, color='006100')
            else:
                ws.cell(r, 6).fill = _BAD
                ws.cell(r, 6).font = Font(bold=True, size=10, color='C00000')

        for ci, w in enumerate([5, 14, 14, 14, 14, 38], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = 'A8'

    # 8. 일별 백테스트 — ★ 레거시 형식(컬럼/서식) 유지 + 값은 K값 순신호(net>K)로 채움 (요청).
    #    매수카운트/매수ON/매도카운트/매도ON/포지션/액션/진입가/보유일/미실현/실현/누적수익 = K값 기준.
    for _si, (_shname, _nsd) in enumerate([('일별 백테스트', globals().get('_KNET_FULL') or _nsd_main)]):
        ws = wb.create_sheet(_shname, _si); ws.sheet_view.showGridLines = False
        if _nsd is None:
            ws.cell(1, 1).value = f'{_shname} — 순신호 K 계산 실패'
            ws.cell(1, 1).font = Font(bold=True, size=12, color='C00000'); continue
        _wtd = _nsd.get('weighted')
        _bk = (round(_nsd['best_k'], 3) if _wtd else _nsd['best_k'])
        _wtag = '점수가중' if _wtd else '단순개수'
        ws.cell(1, 1).value = (f'{_shname} — net({_wtag})>K({_bk})면 매수·보유, net≤K면 매도 (신호일 종가). '
                               f'K는 전체수익 최고. 수익=상승·하락률 합산.')
        ws.cell(1, 1).font = Font(bold=True, size=13, color='1F3864'); ws.merge_cells('A1:P1')
        def _p2(x): return (f"{x*100:+.2f}%" if x is not None else '—')
        ws.cell(2, 1).value = (f"★최적K={_bk} | 지표수 {_nsd['n_buy_opt']}/{_nsd['n_sell_opt']} | "
                               f"전체 {_p2(_nsd['full_cum'])} / OOS {_p2(_nsd.get('oos_cum'))} (B&H {_p2(_nsd['bh_full'])}) | "
                               f"거래 {_nsd['n_trades']}회 | 보유중하락 {_p2(_nsd['held_down_full'])}")
        ws.cell(2, 1).font = Font(bold=True, color='C00000'); ws.merge_cells('A2:P2')
        _hdr(ws, 4, ['날짜', f'{ticker}종가', '매수카운트', '매수ON', '매도카운트', '매도ON',
                     '포지션', '액션', '진입가', '보유일', '미실현%', '실현%',
                     '누적수익%(합산)', '순신호 net', '보유중하락 누적%', '구간'])
        d = _nsd['daily']
        pos = d['position'].values.astype(int)
        prices = d['price'].values.astype(float)
        dts = list(d['date']); rets = d['day_ret'].values.astype(float)
        _entry_px = None; _entry_i = None
        for i in range(len(d)):
            r = 5 + i; row = d.iloc[i]
            _p = pos[i]; _act = str(row['action'])
            if _act == '매수': _entry_px = prices[i]; _entry_i = i
            _held = (i - _entry_i) if (_p == 1 and _entry_i is not None) else None
            _unreal = ((prices[i] / _entry_px - 1.0) if (_p == 1 and _entry_px) else None)
            _real = None
            if _act == '매도' and _entry_px:
                _real = float(np.sum(rets[_entry_i:i+1])) if _entry_i is not None else None
                _entry_px = None; _entry_i = None
            ws.cell(r, 1).value = pd.Timestamp(row['date']).strftime('%Y-%m-%d')
            ws.cell(r, 2).value = (round(float(prices[i]), 2) if pd.notna(prices[i]) else None)
            ws.cell(r, 3).value = (round(float(row['buy_count']), 2) if _wtd else int(row['buy_count']))
            ws.cell(r, 4).value = ('✓' if float(row['net']) > _bk else '')
            ws.cell(r, 5).value = (round(float(row['sell_count']), 2) if _wtd else int(row['sell_count']))
            ws.cell(r, 6).value = ('✓' if not (float(row['net']) > _bk) else '')
            ws.cell(r, 7).value = ('롱' if _p == 1 else '현금')
            ws.cell(r, 8).value = _act
            ws.cell(r, 9).value = (f"${_entry_px:.2f}" if (_p == 1 and _entry_px) else
                                   (f"${prices[i]:.2f}" if _act == '매수' else ''))
            ws.cell(r, 10).value = (_held if _held is not None else '')
            ws.cell(r, 11).value = (round(_unreal*100, 2) if _unreal is not None else '')
            ws.cell(r, 12).value = (round(_real*100, 2) if _real is not None else '')
            ws.cell(r, 13).value = round(float(row['cum_ret']) * 100, 2)
            ws.cell(r, 14).value = (round(float(row['net']), 3) if _wtd else int(row['net']))
            ws.cell(r, 15).value = round(float(row['held_down_run']) * 100, 2)
            ws.cell(r, 16).value = ('OOS' if int(row['is_oos']) == 1 else '학습')
            if _p == 1: ws.cell(r, 7).fill = PatternFill('solid', fgColor='C6EFCE')
            if _act == '매수':   ws.cell(r, 8).fill = PatternFill('solid', fgColor='C6EFCE')
            elif _act == '매도': ws.cell(r, 8).fill = PatternFill('solid', fgColor='FFC7CE')
            if _real is not None: ws.cell(r, 12).fill = (_GOOD if _real > 0 else _BAD)
            if int(row['is_oos']) == 1: ws.cell(r, 1).fill = PatternFill('solid', fgColor='DDEBF7')
        for ci, w in enumerate([12, 10, 10, 7, 10, 7, 7, 9, 10, 7, 10, 10, 14, 11, 14, 7], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = 'A5'
        print(f"  ✓ {_shname}: K={_bk}, 거래 {_nsd['n_trades']}회")

    # 9. ★ OOS 일별 거래
    if oos_enabled and oos_daily is not None and len(oos_daily) > 0:
        ws = wb.create_sheet('OOS 일별 거래'); ws.sheet_view.showGridLines = False
        oc = oos_cur if oos_cur is not None else cur
        oos_bh = 0.0
        oos_bh_up = 0.0
        try:
            oos_bh = _bh_sum_return(oos_daily['close'].values) * 100
            oos_bh_up = _bh_up_sum_return(oos_daily['close'].values) * 100
        except Exception:
            oos_bh_up = 0.0
        ws.cell(1, 1).value = (f'🔬 OOS 검증 일별 거래 — {len(oos_daily)}일  '
                               f'[{oos_daily["date"].iloc[0].date()} ~ {oos_daily["date"].iloc[-1].date()}]')
        ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
        ws.merge_cells('A1:T1')
        ws.cell(2, 1).value = (f'OOS 누적 {oc["cum_return_pct"]:+.2f}%  /  B&H {oos_bh:+.2f}%  /  '
                               f'📈 상승일만 전략 {oc.get("up_cum_return_pct", 0):+.2f}% / B&H {oos_bh_up:+.2f}%  /  '
                               f'거래 {oc["n_trades"]}회  /  승률 {oc["win_rate"]*100:.1f}%  '
                               f'(IS 학습구간에서 찾은 조합을 검증구간에 그대로 적용 — 손 안 댐)')
        ws.cell(2, 1).font = Font(italic=True, size=10, color='C00000')
        ws.merge_cells('A2:T2')
        _oind_hdr = [(('매수:' if _k == 'BUY' else '매도:') + _nm)
                     for (_k, _nm, _vc, _sc) in oc.get('used_ind_cols', [])]
        _hdr(ws, 4, ['날짜', f'{ticker}종가',
                     f'매수카운트(/{oc["K_buy"]})', '매수ON', '매수성공',
                     f'매도카운트(/{oc["K_sell"]})', '매도ON', '매도성공',
                     '포지션', '액션', '진입가', '⛔ 손절가', '보유일',
                     '미실현%', '실현%', '누적자산', '누적수익%', '진행최대손실%']
                    + _oind_hdr)
        _write_daily_rows(ws, oos_daily, oc, mdd_limit_pct)

    # ─── ★ feature2 (요청) — 매수/매도 지표 신호 매트릭스 (각각 별도 시트) ───
    try:
        _abset = _norm_date_set(anchor_buy_dates)
        _asset = _norm_date_set(anchor_sell_dates)
        ws = wb.create_sheet('매수 지표 신호'); ws.sheet_view.showGridLines = False
        _write_indicator_matrix_sheet(ws, buy_pool, feat, close_full,
                                      _abset, _asset, ticker, '매수')
        ws = wb.create_sheet('매도 지표 신호'); ws.sheet_view.showGridLines = False
        _write_indicator_matrix_sheet(ws, sell_pool, feat, close_full,
                                      _abset, _asset, ticker, '매도')
    except Exception as _eim:
        print(f"  ⚠ 지표 신호 매트릭스 시트 작성 실패(무시): {_eim}")

    # ─── 7b. 성공률 우선 선출 시트 (요청) — 성공률로 먼저 뽑은 지표 정리 ───
    try:
        _sp = globals().get('_LAST_SUCCESS_POOL', (None, None))
        _spb, _sps = (_sp if isinstance(_sp, (tuple, list)) and len(_sp) == 2 else (None, None))

        def _write_success_sheet(ws, df, kind_label):
            ws.cell(1, 1).value = (f'{ticker} — {kind_label} 성공률 우선 선출 '
                                   f'(성공률 {POOL_SUCCESS_MIN_RATE*100:.0f}%+ · 최소신호 {POOL_SUCCESS_MIN_SIG} · '
                                   f'분위 {POOL_SUCCESS_WIDE_PCT})  ※성공률 내림차순, 동률은 점수(Wilson)')
            ws.cell(1, 1).font = Font(bold=True, size=11)
            heads = ['순위', '지표', '방향', '임계치', '분위', '신호수', '성공수', '성공률', '점수', '지연(일)']
            _hdr(ws, 3, heads)
            if df is None or len(df) == 0:
                ws.cell(4, 1).value = '조건(성공률/최소신호)을 만족하는 지표 없음'
                return
            for i, (_, r) in enumerate(df.iterrows(), start=1):
                rr = 3 + i
                try:
                    _ld = int(r.get('lead_shift', 0)) if pd.notna(r.get('lead_shift', 0)) else 0
                except Exception:
                    _ld = 0
                vals = [i, r['indicator'], r['direction'],
                        round(float(r['threshold']), 6),
                        (f"{float(r['pct_label']):.2f}" if r['direction'] in ('>=', '<=')
                         else f"z{r['direction'][1:]}{r['pct_label']:g}"),
                        int(r['n_signals']), int(r['n_success']),
                        f"{float(r['success_rate'])*100:.2f}%",
                        round(float(r['score']), 4), _ld]
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(rr, ci); c.value = v
                    if ci == 8:  # 성공률 강조
                        c.font = Font(bold=True)
            widths = [6, 30, 8, 12, 12, 9, 9, 10, 10, 9]
            for ci, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.freeze_panes = 'A4'

        ws = wb.create_sheet('성공률 우선 매수'); ws.sheet_view.showGridLines = False
        _write_success_sheet(ws, _spb, '매수')
        ws = wb.create_sheet('성공률 우선 매도'); ws.sheet_view.showGridLines = False
        _write_success_sheet(ws, _sps, '매도')
    except Exception as _esp:
        print(f"  ⚠ 성공률 우선 선출 시트 작성 실패(무시): {_esp}")

    # ─── 7c. (요청) 단일-K 순신호 K최적화 시트는 KL 순신호로 대체 → 생성 안 함 ───
    try:
        _sheets = []   # 비움 → 단일-K 시트 미생성 (KL 순신호로 대체)
        def _pct(x): return (f"{x*100:+.2f}%" if x is not None else '—')
        for _shname, _ns, _basis in _sheets:
            if _ns is None:
                continue
            ws = wb.create_sheet(_shname); ws.sheet_view.showGridLines = False
            _hoos = _ns['has_oos']
            ws.cell(1, 1).value = (
                f'{ticker} — 순신호 K 최적화  ｜ net=매수카운트−매도카운트, net>K면 롱·아니면 현금(신호일 종가 기준). '
                f'K는 [{_basis} 수익] 최고로 선정 (수익=상승·하락률 합산)')
            ws.cell(1, 1).font = Font(bold=True, size=11)
            ws.cell(3, 1).value = f'★ 최적 K ({_basis} 수익 기준)'; ws.cell(3, 2).value = _ns['best_k']
            ws.cell(3, 1).font = Font(bold=True); ws.cell(3, 2).font = Font(bold=True, color='C00000')
            # 구간별 요약 (학습/OOS/전체) — OOS 복원
            _hh = ws.cell(5, 1); _hh.value = '구간'; _hh.font = Font(bold=True)
            for _cc, _lab in [(2, '전략수익(합산)'), (3, 'B&H'), (4, '보유중하락 누적')]:
                ws.cell(5, _cc).value = _lab; ws.cell(5, _cc).font = Font(bold=True)
            ws.cell(6, 1).value = '학습'
            ws.cell(6, 2).value = _pct(_ns.get('train_cum')); ws.cell(6, 3).value = _pct(_ns.get('bh_train'))
            ws.cell(6, 4).value = _pct(_ns.get('held_down_train'))
            ws.cell(7, 1).value = '★ OOS'; ws.cell(7, 1).font = Font(bold=True, color='1F6F1F')
            ws.cell(7, 2).value = _pct(_ns.get('oos_cum')); ws.cell(7, 3).value = _pct(_ns.get('bh_oos'))
            ws.cell(7, 4).value = _pct(_ns.get('held_down_oos'))
            for _cc in (2, 3, 4): ws.cell(7, _cc).font = Font(bold=True, color='1F6F1F')
            ws.cell(8, 1).value = '전체'; ws.cell(8, 1).font = Font(bold=True)
            ws.cell(8, 2).value = _pct(_ns.get('full_cum')); ws.cell(8, 3).value = _pct(_ns.get('bh_full'))
            ws.cell(8, 4).value = _pct(_ns.get('held_down_full'))
            ws.cell(8, 4).font = Font(bold=True, color='C00000')
            ws.cell(9, 1).value = '거래 횟수(전체/OOS)'; ws.cell(9, 2).value = f"{_ns['n_trades']}회 / {_ns.get('n_trades_oos','—')}회"
            ws.cell(10, 1).value = '풀 크기(매수/매도)'; ws.cell(10, 2).value = f"{_ns['buy_pool_n']} / {_ns['sell_pool_n']}"
            ws.cell(11, 1).value = 'net 범위'; ws.cell(11, 2).value = f"{_ns['net_min']} ~ {_ns['net_max']}"

            # K vs 수익 표 (왼쪽) — 학습/OOS/전체 + 비율 (OOS 컬럼 복원)
            _hdr(ws, 13, ['K', '학습수익%', 'OOS수익%', '전체수익%', '보유중하락%(전체)',
                          'OOS보유중하락%', '전체 수익/하락비', 'OOS 수익/하락비', '롱일수'])
            for i, _trow in enumerate(_ns['k_table']):
                K, tr, oo, fu, dl = _trow[0], _trow[1], _trow[2], _trow[3], _trow[4]
                hd = _trow[5] if len(_trow) > 5 else None
                hd_oos = _trow[6] if len(_trow) > 6 else None
                _rt_full = (fu / abs(hd)) if (hd is not None and abs(hd) > 1e-9) else None
                _rt_oos = (oo / abs(hd_oos)) if (oo is not None and hd_oos is not None and abs(hd_oos) > 1e-9) else None
                rr = 14 + i
                ws.cell(rr, 1).value = K
                ws.cell(rr, 2).value = (round(tr * 100, 2) if tr is not None else None)
                ws.cell(rr, 3).value = (round(oo * 100, 2) if oo is not None else None)
                ws.cell(rr, 4).value = round(fu * 100, 2)
                ws.cell(rr, 5).value = (round(hd * 100, 2) if hd is not None else None)
                ws.cell(rr, 6).value = (round(hd_oos * 100, 2) if hd_oos is not None else None)
                ws.cell(rr, 7).value = (round(_rt_full, 2) if _rt_full is not None else '—')
                ws.cell(rr, 8).value = (round(_rt_oos, 2) if _rt_oos is not None else '—')
                ws.cell(rr, 9).value = dl
                if hd is not None: ws.cell(rr, 5).font = Font(color='C00000')
                if hd_oos is not None: ws.cell(rr, 6).font = Font(color='C00000')
                if K == _ns['best_k']:
                    for cc in range(1, 10):
                        ws.cell(rr, cc).fill = PatternFill('solid', fgColor='FFF2CC')
                        ws.cell(rr, cc).font = Font(bold=True)

            # 일별 백테스트 표 (오른쪽, col11~)
            d = _ns['daily']
            _hdr2 = ['날짜', f'{ticker}종가', '매수카운트', '매도카운트', 'net',
                     f'포지션(net>{(round(_ns["best_k"],3) if _ns.get("weighted") else _ns["best_k"])})',
                     '액션', '일별수익%', '누적수익%(합산)', '보유중하락%', '구간']
            _c0 = 11
            for ci, h in enumerate(_hdr2, start=_c0):
                c = ws.cell(13, ci); c.value = h; c.fill = _HDR; c.font = _WB_
            for i in range(len(d)):
                rr = 14 + i; row = d.iloc[i]
                ws.cell(rr, _c0).value   = pd.Timestamp(row['date']).strftime('%Y-%m-%d')
                ws.cell(rr, _c0+1).value = (round(float(row['price']), 2) if pd.notna(row['price']) else None)
                ws.cell(rr, _c0+2).value = int(row['buy_count'])
                ws.cell(rr, _c0+3).value = int(row['sell_count'])
                ws.cell(rr, _c0+4).value = (round(float(row['net']), 3) if _ns.get('weighted') else int(row['net']))
                ws.cell(rr, _c0+5).value = int(row['position'])
                ws.cell(rr, _c0+6).value = str(row['action'])
                ws.cell(rr, _c0+7).value = round(float(row['day_ret']) * 100, 3)
                ws.cell(rr, _c0+8).value = round(float(row['cum_ret']) * 100, 2)
                ws.cell(rr, _c0+9).value = round(float(row['held_down_run']) * 100, 2)
                ws.cell(rr, _c0+10).value = ('OOS' if int(row['is_oos']) == 1 else '학습')
                if int(row['position']) == 1: ws.cell(rr, _c0+5).fill = PatternFill('solid', fgColor='C6EFCE')
                _act = str(row['action'])
                if _act == '매수':   ws.cell(rr, _c0+6).fill = PatternFill('solid', fgColor='C6EFCE')
                elif _act == '매도': ws.cell(rr, _c0+6).fill = PatternFill('solid', fgColor='FFC7CE')
                if pd.notna(row['day_ret']) and float(row['day_ret']) < 0 and int(row['position']) == 1:
                    ws.cell(rr, _c0+7).fill = PatternFill('solid', fgColor='FFC7CE')
                if int(row['is_oos']) == 1: ws.cell(rr, _c0).fill = PatternFill('solid', fgColor='DDEBF7')
            for ci, w in enumerate([6, 11, 11, 11, 13, 11, 13, 12, 11, 12, 13, 11, 11, 12, 13, 9], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.freeze_panes = 'A14'
            print(f"  ✓ {_shname}: 최적 K={_ns['best_k']} | 전체 {_pct(_ns.get('full_cum'))} / OOS {_pct(_ns.get('oos_cum'))} "
                  f"| 거래 {_ns['n_trades']}회")
    except Exception as _ek:
        import traceback; print(f"  ⚠ 순신호 K최적화 시트 작성 실패(무시): {_ek}")

    # ─── 7d. 순신호 K/L 2임계 최적화 (net≥K 매수 / net≤L 매도 / 사이 유지) ───
    try:
        _kln = globals().get('_KNET_FULL') or _nsd_main
        if _kln is not None and _kln.get('daily') is not None:
            _d = _kln['daily']
            _net = _d['net'].values.astype(float)
            _price = _d['price'].values.astype(float)
            _dates = list(_d['date'])
            _wtd_kl = _kln.get('weighted')
            _rr = np.zeros(len(_price)); _rr[1:] = _price[1:] / _price[:-1] - 1.0
            _rr[~np.isfinite(_rr)] = 0.0
            _mddlim = globals().get('MAX_DRAWDOWN_LIMIT_PCT')
            _klfix = globals().get('_KNET_KL_FIXED')
            if _klfix and _klfix.get('kl_k') is not None:
                _kl = _net_kl_search(_net, _rr,
                                     fixed_kl=(_klfix['kl_k'], _klfix['kl_l']),
                                     fixed_kl_mdd=((_klfix['kl_k_mdd'], _klfix['kl_l_mdd'])
                                                   if _klfix.get('kl_k_mdd') is not None else None))
                print(f"  ♻ 재현: 원본 K/L(K={_klfix['kl_k']}, L={_klfix['kl_l']}) 그대로")
            else:
                _kl = _net_kl_search(_net, _rr, mdd_limit=_mddlim)
            globals()['_KNET_KL'] = _kl

            # (요청) KL 순신호 시트 — (K,L) 조합별 컬럼 결과. 일별 백테스트 옆(index 1)에 배치.
            ws = wb.create_sheet('KL 순신호', 1); ws.sheet_view.showGridLines = False
            _mtxt = (f'MDD 한도 {_mddlim*100:.0f}% 이내' if _mddlim is not None else 'MDD 한도 없음')
            ws.cell(1, 1).value = (f'{ticker} — KL 순신호 (net≥K 매수·롱 / net≤L 매도·현금 / 사이 유지, K≥L). '
                                   f'수익=상승·하락률 합산. {_mtxt}. ★=대표조합')
            ws.cell(1, 1).font = Font(bold=True, size=12, color='1F3864')

            def _kl_stats(pos, K=None, L=None):
                tr = []; ei = None; last_buy = None; last_sell = None
                # ★ 보유중 하락: 각 롱 구간 내 진입가 대비 저점 하락률 — 구간 최악값.
                held_max_dd = 0.0
                for i in range(len(pos)):
                    if pos[i] == 1 and (i == 0 or pos[i-1] == 0):
                        ei = i; last_buy = i
                    if pos[i] == 0 and i > 0 and pos[i-1] == 1:
                        last_sell = i
                        if ei is not None:
                            tr.append(float(np.sum(_rr[ei+1:i+1])))
                            _seg = _price[ei:i+1] if ei < len(_price) else _price[ei:]
                            if len(_seg) >= 2:
                                _dd = float(_seg.min() / _seg[0] - 1.0)
                                if _dd < held_max_dd: held_max_dd = _dd
                            ei = None
                if ei is not None:
                    tr.append(float(np.sum(_rr[ei+1:])))
                    _seg = _price[ei:]
                    if len(_seg) >= 2:
                        _dd = float(_seg.min() / _seg[0] - 1.0)
                        if _dd < held_max_dd: held_max_dd = _dd
                nt = len(tr); wins = sum(1 for t in tr if t > 0)
                wr = (wins / nt * 100) if nt else 0.0
                lb = (pd.Timestamp(_dates[last_buy]).strftime('%Y-%m-%d') if last_buy is not None else '-')
                ls = (pd.Timestamp(_dates[last_sell]).strftime('%Y-%m-%d') if last_sell is not None else '-')
                lbp = (round(float(_price[last_buy]), 2) if last_buy is not None else '-')
                lsp = (round(float(_price[last_sell]), 2) if last_sell is not None else '-')

                # ★ 매수·매도 정확도 (요청) — 신호일 t 판정: net[t] vs K/L → 다음날 t+1의 return.
                #   매수 정확 = net[t]≥K일 때 r[t+1]>0 (매수했는데 다음날 상승 = 맞음)
                #   매도 정확 = net[t]≤L일 때 r[t+1]<0 (매도했는데 다음날 하락 = 맞음)
                #   무포지션·중립(L<net<K) 상태는 미판정 (분모 제외).
                buy_acc = sell_acc = None
                n_buy_sig = n_sell_sig = 0
                buy_hit = sell_hit = 0
                if K is not None and L is not None and len(_net) >= 2 and len(_rr) >= 2:
                    # net[t]와 r[t+1] 정렬. 마지막 net는 미래 r 없어 스킵.
                    _n = min(len(_net), len(_rr) - 1)
                    for t in range(_n):
                        if _net[t] >= K:                              # 매수/보유 신호
                            n_buy_sig += 1
                            if _rr[t + 1] > 0: buy_hit += 1
                        elif _net[t] <= L:                             # 매도/현금 신호
                            n_sell_sig += 1
                            if _rr[t + 1] < 0: sell_hit += 1
                    if n_buy_sig  > 0: buy_acc  = buy_hit  / n_buy_sig  * 100
                    if n_sell_sig > 0: sell_acc = sell_hit / n_sell_sig * 100
                return nt, wr, lb, lbp, ls, lsp, held_max_dd, buy_acc, sell_acc, n_buy_sig, n_sell_sig

            _hdr(ws, 3, ['순위', 'K(매수임계)', 'L(매도임계)', '전체수익%', '최대낙폭%', '보유중하락%',
                         '거래횟수', '승률%',
                         '매수정확도%', '매수신호일수', '매도정확도%', '매도신호일수',
                         '최근매수일', '매수가', '최근매도일', '매도가', '비고'])
            _br = _kl.get('best_ret'); _bm = _kl.get('best_mdd')
            # 대표조합(최대수익·MDD내최고)을 항상 맨 위에 + 나머지 상위조합 (중복 제거)
            _disp = []
            def _as_t(tup):
                return {'K':tup[0],'L':tup[1],'ret':tup[2],'mdd':tup[3],'dl':tup[4],'pos':tup[5]}
            if _br is not None: _disp.append(_as_t(_br))
            if _bm is not None and (_br is None or abs(_bm[0]-_br[0])>1e-9 or abs(_bm[1]-_br[1])>1e-9):
                _disp.append(_as_t(_bm))
            _seen={(round(x['K'],4),round(x['L'],4)) for x in _disp}
            for t in _kl.get('top', []):
                if (round(t['K'],4),round(t['L'],4)) not in _seen:
                    _disp.append(t); _seen.add((round(t['K'],4),round(t['L'],4)))
            for ti, t in enumerate(_disp):
                r = 4 + ti
                nt, wr, lb, lbp, ls, lsp, held_dd, buy_acc, sell_acc, n_bs, n_ss = _kl_stats(
                    t['pos'], K=float(t['K']), L=float(t['L']))
                _tag = ''
                if _br and abs(t['K']-_br[0])<1e-9 and abs(t['L']-_br[1])<1e-9: _tag = '★최대수익'
                if _bm and abs(t['K']-_bm[0])<1e-9 and abs(t['L']-_bm[1])<1e-9: _tag = (_tag+' ★MDD내최고').strip()
                ws.cell(r,1).value = ti+1
                ws.cell(r,2).value = round(float(t['K']),3); ws.cell(r,3).value = round(float(t['L']),3)
                ws.cell(r,4).value = round(t['ret']*100,2); ws.cell(r,4).font = Font(bold=True, color='C00000')
                ws.cell(r,5).value = round(t['mdd']*100,2)
                ws.cell(r,6).value = round(held_dd*100, 2)                    # ★ 보유중하락%
                if held_dd < -0.10: ws.cell(r,6).font = Font(color='C00000')  # 10% 초과 강조
                ws.cell(r,7).value = nt; ws.cell(r,8).value = round(wr,1)
                # ★ 매수/매도 정확도 (요청) — net vs K/L 판정 다음날 방향 일치 비율
                ws.cell(r,9).value  = (round(buy_acc, 1)  if buy_acc  is not None else '-')
                ws.cell(r,10).value = n_bs
                ws.cell(r,11).value = (round(sell_acc, 1) if sell_acc is not None else '-')
                ws.cell(r,12).value = n_ss
                # 정확도가 50% 미만이면 빨간색 (동전 던지기보다 못함)
                if buy_acc  is not None and buy_acc  < 50: ws.cell(r,9).font  = Font(color='C00000', bold=True)
                if sell_acc is not None and sell_acc < 50: ws.cell(r,11).font = Font(color='C00000', bold=True)
                if buy_acc  is not None and buy_acc  > 60: ws.cell(r,9).font  = Font(color='006100', bold=True)
                if sell_acc is not None and sell_acc > 60: ws.cell(r,11).font = Font(color='006100', bold=True)
                # 최근 매매 정보
                ws.cell(r,13).value = lb; ws.cell(r,14).value = lbp
                ws.cell(r,15).value = ls; ws.cell(r,16).value = lsp
                ws.cell(r,17).value = (_tag + (' K=L' if abs(t['K']-t['L'])<1e-9 else '')).strip()
                if _tag:
                    for c in range(1,18): ws.cell(r,c).fill = PatternFill('solid', fgColor='FFF2CC')
                    ws.cell(r,17).font = Font(bold=True, color='1F6F1F')
            for ci, w in enumerate([5,12,12,11,11,12,9,8, 12,11,12,11, 13,10,13,10,22], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.freeze_panes = 'A4'
            ws.merge_cells('A1:Q1')      # 17개 컬럼

            _bt = _kl['best_ret']; _bm = _kl['best_mdd']
            print(f"  ✓ KL 순신호: 최대수익 K={_bt[0]:.3f}/L={_bt[1]:.3f} ({_bt[2]*100:+.1f}%, MDD {_bt[3]*100:.1f}%)"
                  + (f" | MDD한도내 K={_bm[0]:.3f}/L={_bm[1]:.3f}" if _bm else " | (MDD 폴백)"))
    except Exception as _ekl:
        import traceback; traceback.print_exc()
        print(f"  ⚠ 순신호 K/L 최적화 시트 작성 실패(무시): {_ekl}")

    # ─── 7e. ★ 지표 선출 A/B 검증 (요청) ───
    #   각 개선 flag ON/OFF의 KL 백테스트 성적(전체수익·MDD·보유중하락·거래·승률)을 비교.
    #   SELECTION_AB_VERIFY=False면 건너뜀.
    try:
        if globals().get('SELECTION_AB_VERIFY', True) and feat is not None and close_full is not None:
            _quick = bool(globals().get('SELECTION_AB_QUICK', False))
            _ab = run_selection_ab_verification(feat, close_full, ticker,
                                                 top_n=globals().get('TOP_N_POOL'),
                                                 quick=_quick)
            if _ab:
                _write_selection_ab_sheet(wb, _ab, mdd_limit=globals().get('MAX_DRAWDOWN_LIMIT_PCT'))
                globals()['_LAST_AB_RESULT'] = _ab
    except Exception as _abe:
        import traceback; traceback.print_exc()
        print(f"  ⚠ 지표 선출 A/B 검증 실패(무시): {_abe}")

    # ─── 8. 메타조합별 지표 풀 (그리드 번호 재현 정확도용) ───
    #   각 메타조합(wilson_z/pct/corr)이 선별한 매수·매도 지표 전체를 저장.
    #   그리드 번호 재현 시, 그 번호의 메타조합 풀을 여기서 읽어 '그대로' 사용 → 정확히 재현.
    _pool_map = globals().get('_LAST_POOL_MAP', None)
    if _pool_map:
        try:
            ws = wb.create_sheet('메타조합별_지표'); ws.sheet_view.showGridLines = False
            ws.cell(1, 1).value = ('메타조합별 지표 풀 — 그리드 번호 재현 시 그 조합의 메타변수에 맞는 '
                                   '지표를 여기서 읽어 정확히 재현합니다 (지표 재선별 안 함)')
            ws.cell(1, 1).font = Font(bold=True, size=12, color='1F3864')
            r = 3
            for (wz, pl, ph, corr), (bp, sp) in _pool_map.items():
                ws.cell(r, 1).value = f'★ wilson_z={wz} / pct=({pl},{ph}) / corr={corr}'
                ws.cell(r, 1).font = Font(bold=True, size=11, color='C00000')
                r += 1
                for side, pool in [('매수', bp), ('매도', sp)]:
                    if pool is None or len(pool) == 0: continue
                    ws.cell(r, 1).value = f'  [{side}] {len(pool)}개'
                    ws.cell(r, 1).font = Font(bold=True, size=10)
                    r += 1
                    _hdr(ws, r, ['side', '지표', '방향', '임계치', '분위', '신호수', '성공수', '성공률', '점수', '지연(일)'])
                    r += 1
                    for _, prow in pool.iterrows():
                        ws.cell(r, 1).value = side
                        ws.cell(r, 2).value = str(prow.get('indicator', ''))
                        ws.cell(r, 3).value = str(prow.get('direction', '>='))
                        ws.cell(r, 4).value = float(prow.get('threshold', 0.0))
                        ws.cell(r, 5).value = float(prow.get('pct_label', 50.0)) if 'pct_label' in prow else ''
                        ws.cell(r, 6).value = float(prow.get('n_signals', 0))
                        ws.cell(r, 7).value = float(prow.get('n_success', 0))
                        ws.cell(r, 8).value = float(prow.get('success_rate', 0.0))
                        ws.cell(r, 9).value = float(prow.get('score', 0.0))
                        try:
                            ws.cell(r, 10).value = int(prow.get('lead_shift', 0) or 0)
                        except Exception:
                            ws.cell(r, 10).value = 0
                        r += 1
                    r += 1
            for ci, w in enumerate([8, 28, 6, 12, 8, 8, 8, 10, 10, 9], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
        except Exception as _e:
            print(f"  ⚠ 메타조합별_지표 시트 작성 실패: {_e}")

    # ★ k순신호 재현풀 시트 — 엑셀만으로 정확 재현 (별도 pkl 불필요).
    #   net>K가 쓴 '합친 다중임계 풀'(지표·방향·임계·성공률) + K/지표수/g 를 그대로 기록.
    try:
        _mp = globals().get('_KNET_MULTI_POOL')
        if _mp and _mp[0] == ticker and _mp[1] is not None and _nsd_main:
            wsr = wb.create_sheet('k순신호 재현풀'); wsr.sheet_view.showGridLines = False
            wsr.cell(1, 1).value = (f'{ticker} — k순신호 재현용 풀/설정 (이 시트만 있으면 재현모드가 그대로 복원). '
                                    f'수정 금지.')
            wsr.cell(1, 1).font = Font(bold=True, size=11, color='1F3864')
            # 고정 파라미터 (2~3행) — 전체수익K / OOS수익K 2개
            _nf = globals().get('_KNET_FULL'); _no = globals().get('_KNET_OOS')
            wsr.cell(2, 1).value = 'PARAM'
            for c, h in enumerate(['k_full', 'nb_full', 'ns_full', 'g_full',
                                   'k_oos', 'nb_oos', 'ns_oos', 'g_oos'], 2):
                wsr.cell(2, c).value = h
            wsr.cell(3, 1).value = 'VALUE'
            def _pv(d, key):
                return (d.get(key) if (d and d.get(key) is not None) else None)
            _gf = float(globals().get('_KNET_BEST_WEXP', 1.0))
            wsr.cell(3, 2).value = _pv(_nf, 'best_k'); wsr.cell(3, 3).value = _pv(_nf, 'n_buy_opt')
            wsr.cell(3, 4).value = _pv(_nf, 'n_sell_opt'); wsr.cell(3, 5).value = _gf
            wsr.cell(3, 6).value = _pv(_no, 'best_k'); wsr.cell(3, 7).value = _pv(_no, 'n_buy_opt')
            wsr.cell(3, 8).value = _pv(_no, 'n_sell_opt'); wsr.cell(3, 9).value = _gf
            # K/L 2임계 재현값 (10~13열)
            for c, h in enumerate(['kl_k', 'kl_l', 'kl_k_mdd', 'kl_l_mdd'], 10):
                wsr.cell(2, c).value = h
            _klr = globals().get('_KNET_KL')
            if _klr and _klr.get('best_ret'):
                wsr.cell(3, 10).value = round(float(_klr['best_ret'][0]), 4)
                wsr.cell(3, 11).value = round(float(_klr['best_ret'][1]), 4)
            if _klr and _klr.get('best_mdd'):
                wsr.cell(3, 12).value = round(float(_klr['best_mdd'][0]), 4)
                wsr.cell(3, 13).value = round(float(_klr['best_mdd'][1]), 4)
            for c in range(1, 14):
                wsr.cell(2, c).font = Font(bold=True, color='FFFFFF'); wsr.cell(2, c).fill = PatternFill('solid', fgColor='1F3864')
            # 풀 테이블 (5행 헤더~)
            hr = 5
            for c, h in enumerate(['구분', '지표(indicator)', '방향(direction)', '임계치(threshold)', '성공률(success_rate)',
                                   '지연(lead_shift)', '선출한도(sel_limit)', '최적선행(best_lead)'], 1):
                cell = wsr.cell(hr, c); cell.value = h
                cell.font = Font(bold=True, color='FFFFFF'); cell.fill = PatternFill('solid', fgColor='548235')
            _rr = hr + 1
            def _pf(v, cast=float):
                try:
                    if v is None or (isinstance(v, float) and np.isnan(v)): return None
                    return cast(v)
                except Exception:
                    return None
            for _side, _pool in [('매수', _mp[1]), ('매도', _mp[2])]:
                if _pool is None: continue
                for _, prow in _pool.iterrows():
                    wsr.cell(_rr, 1).value = _side
                    wsr.cell(_rr, 2).value = str(prow.get('indicator'))
                    wsr.cell(_rr, 3).value = str(prow.get('direction'))
                    wsr.cell(_rr, 4).value = float(prow.get('threshold')) if pd.notna(prow.get('threshold')) else None
                    wsr.cell(_rr, 5).value = float(prow.get('success_rate')) if pd.notna(prow.get('success_rate')) else None
                    wsr.cell(_rr, 6).value = _pf(prow.get('lead_shift'), int) or 0    # ★ 신호 지연 재현
                    wsr.cell(_rr, 7).value = _pf(prow.get('sel_limit'))               # ★ 선출 한도 (검증용)
                    wsr.cell(_rr, 8).value = _pf(prow.get('best_lead'), int)          # ★ 최적 선행일 (참고)
                    _rr += 1
            for ci, w in enumerate([6, 30, 12, 12, 14, 12, 12, 12], 1):
                wsr.column_dimensions[get_column_letter(ci)].width = w
            print(f"  💾 'k순신호 재현풀' 시트 저장 — 엑셀만으로 재현 가능 (매수 {len(_mp[1])}행 / 매도 {len(_mp[2])}행)")
    except Exception as _pe:
        print(f"  ⚠ k순신호 재현풀 시트 저장 실패(무시): {_pe}")

    # ─── ★ 검증_예측로직 시트 (요청) — 개선사항 기능 점검 + 지표별 예측력 독립 재계산 ───
    try:
        if globals().get('VERIFY_SHEET_ENABLED', True):
            _write_logic_verification_sheet(wb, feat, close_full, ticker,
                                            horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit)
    except Exception as _ve:
        import traceback; traceback.print_exc()
        print(f"  ⚠ 검증_예측로직 시트 작성 실패(무시): {_ve}")

    wb.save(output_file)


# ════════════════════════════════════════════════════════════════
#                  메모리 변수 자동 로드
# ════════════════════════════════════════════════════════════════
def _resolve_data():
    g = globals()
    feat   = g.get('_pair_feat',  g.get('feat'))
    close  = g.get('_pair_close', g.get('close'))
    ticker = g.get('_pair_ticker', g.get('TICKER'))
    if feat is not None and close is not None and ticker is not None:
        print(f"  ✓ 메모리 로드: feat={feat.shape}, close={len(close)}일, ticker={ticker}")
        return feat, close, ticker
    print("  ⚠ 변수 없음 → 기존 코드 함수 시도")
    needed = ['download_data', 'download_fred_data', 'compute_features',
              'TICKER', 'DOWNLOAD_START', 'EVAL_START']
    miss = [n for n in needed if n not in g]
    if miss:
        raise RuntimeError(f"필요: {miss}")
    closes, ohlcv = g['download_data'](start=g['DOWNLOAD_START'])
    fred_df = g['download_fred_data'](start=g['DOWNLOAD_START'])
    if len(fred_df) == 0: fred_df = None
    feat = g['compute_features'](ohlcv, closes, fred_df=fred_df)
    feat = feat[feat.index >= pd.Timestamp(g['EVAL_START'])]
    close = ohlcv[g['TICKER']]['Close'].reindex(feat.index)
    return feat, close, g['TICKER']


# ════════════════════════════════════════════════════════════════
#                       메인 진입
# ════════════════════════════════════════════════════════════════
def _final_pick_by_real(pool, win_tol):
    """실거래 검증 결과 pool에서 최종 1개 선정 (요청 개정):
       SELECT_BY_CORRECTED_RETURN=True이고 corr_total_return 컬럼이 있으면
         → '보정 후 수익률'이 가장 높은 조합을 선정 (동률 시 보정후 승률).
       아니면 OOS_SELECT_BY_OOS_RETURN=True → OOS 수익, 그것도 없으면 전체 실제 수익.
       반환: (best_inner_dict, sel_row)"""
    use_corr = globals().get('SELECT_BY_CORRECTED_RETURN', True)
    use_oos  = globals().get('OOS_SELECT_BY_OOS_RETURN', True)
    # ★ 보정후 승률 하한 + 허용오차 비교 선정 (요청):
    #   1) 보정후 실제 승률(sel_win_rate) ≥ SELECT_WIN_FLOOR(90%) 중 보정후 수익(sel_total_return) 최고 = 1차 선정.
    #   2) 그 1차 선정의 승률에서 WINRATE_TOLERANCE 이내(승률이 최대 그만큼 낮아도 됨) 후보와 비교해,
    #      보정후 수익이 더 높은 게 있으면 그것으로 재선정. (90% 못 넘는 후보 없으면 최고승률 기준 폴백)
    if 'sel_win_rate' in pool.columns and 'sel_total_return' in pool.columns and len(pool) > 0:
        floor = float(globals().get('SELECT_WIN_FLOOR', 0.90))
        above = pool[pool['sel_win_rate'] >= floor]
        if len(above) > 0:
            b0 = above.sort_values(['sel_total_return', 'sel_win_rate'],
                                   ascending=[False, False]).reset_index(drop=True)
            w0 = float(b0.iloc[0]['sel_win_rate'])          # 1차 선정의 보정후 승률
            band = pool[pool['sel_win_rate'] >= w0 - win_tol]  # 승률 허용오차 이내(아래로) 후보
            b = band.sort_values(['sel_total_return', 'sel_win_rate'],
                                 ascending=[False, False]).reset_index(drop=True)
            _by = '\ubcf4\uc815\ud6c4\uc2b9\ub960\ud558\ud55c+\ud5c8\uc6a9\uc624\ucc28'
        else:
            # 90% 넘는 후보 없음 → 최고승률 -win_tol 밴드 내 수익 최고
            wmax = float(pool['sel_win_rate'].max())
            band = pool[pool['sel_win_rate'] >= wmax - win_tol]
            b = band.sort_values(['sel_total_return', 'sel_win_rate'],
                                 ascending=[False, False]).reset_index(drop=True)
            _by = '\ubcf4\uc815\ud6c4\ucd5c\uace0\uc2b9\ub960\ubc34\ub4dc(90%\ubbf8\ub2ec)'
    elif use_corr and 'corr_total_return' in pool.columns and pool['corr_total_return'].notna().any():
        _wcol = 'corr_win_rate' if 'corr_win_rate' in pool.columns else 'real_win_rate'
        b = pool[pool['corr_total_return'].notna()].sort_values(
            ['corr_total_return', _wcol], ascending=[False, False]).reset_index(drop=True)
        _by = '\ubcf4\uc815\ud6c4'
    elif use_oos and 'oos_total_return' in pool.columns and pool['oos_total_return'].notna().any():
        b = pool[pool['oos_total_return'].notna()].sort_values(
            ['oos_total_return', 'oos_win_rate' if 'oos_win_rate' in pool.columns else 'real_win_rate'],
            ascending=[False, False]).reset_index(drop=True)
        _by = 'OOS'
    else:
        b = pool.sort_values(['real_total_return', 'real_win_rate'],
                             ascending=[False, False]).reset_index(drop=True)
        _by = '\uc804\uccb4'
    sel = b.iloc[0]
    best_inner = {'K_buy': int(sel['K_buy']), 'vote_buy': int(sel['vote_buy']),
                  'K_sell': int(sel['K_sell']), 'vote_sell': int(sel['vote_sell'])}
    for _k in ('sharpe_like', 'total_return', 'avg_success_rate',
               'buy_success_rate', 'sell_success_rate', 'win_rate',
               'max_drawdown', 'n_trades'):
        if _k in sel.index:
            best_inner[_k] = sel[_k]
    _oosr = sel.get('oos_total_return', float('nan'))
    print(f"  \u2713 \uc2e4\uac70\ub798 \uac80\uc99d \uc644\ub8cc \u2014 {_by} \uc218\uc775\ub960 \ucd5c\uace0 \uc870\ud569 \uc120\uc815")
    if _by.startswith('\ubcf4\uc815\ud6c4'):
        _cr = sel.get('sel_total_return', sel.get('corr_total_return', float('nan')))
        _applied = bool(sel.get('corr_applied', False))
        _sw = sel.get('sel_win_rate', sel.get('real_win_rate', float('nan')))
        print(f"     \ucd5c\uc885: K_buy={best_inner['K_buy']}/v{best_inner['vote_buy']}, "
              f"K_sell={best_inner['K_sell']}/v{best_inner['vote_sell']}  "
              f"(\uc120\uc815\uc218\uc775 {_cr*100:+.2f}%{'(\ubcf4\uc815\ucc44\ud0dd)' if _applied else '(\ubcf4\uc815\uc548\ub428)'}, "
              f"\uc6d0\ub798\uc218\uc775 {sel['real_total_return']*100:+.2f}%, "
              f"\uc120\uc815\uc2b9\ub960 {_sw*100:.1f}%, "
              f"\uc120\uc815\ucd5c\ub300\uc190\uc2e4 {sel.get('sel_max_drawdown', sel['real_max_drawdown'])*100:.2f}%)")
    else:
        print(f"     \ucd5c\uc885: K_buy={best_inner['K_buy']}/v{best_inner['vote_buy']}, "
              f"K_sell={best_inner['K_sell']}/v{best_inner['vote_sell']}  "
              f"(OOS\uc218\uc775 {_oosr*100:+.2f}%, OOS\ub9e4\ub3c4\uc815\ud655\ub3c4 {sel.get('oos_sell_success', float('nan'))*100:.1f}%, "
              f"OOS\ub9e4\uc218\uc815\ud655\ub3c4 {sel.get('oos_buy_success', float('nan'))*100:.1f}%, "
              f"\uc804\uccb4\uc218\uc775 {sel['real_total_return']*100:+.2f}%)")
    return best_inner, sel


def _pick_verify_candidates(tbl, bh_ret=None):
    """실거래 검증 대상 후보를 고른다 (요청):
       - 그리드 최고 수익률과 B&H 수익률 '사이'에 있는 후보 (B&H ≤ 그리드수익 ≤ 최고)
       - 그중 그리드 수익 높은 순으로 정렬 → 상위 '절반'만 검증
       - VERIFY_TOP_N 으로 상한도 둔다 (절반이 너무 많으면 컷).
    """
    top_n = int(globals().get('VERIFY_TOP_N', 10000))
    ret_col = 'total_return' if 'total_return' in tbl.columns else (
              'combined_return' if 'combined_return' in tbl.columns else None)
    cand = tbl.copy()
    if ret_col is None:
        return cand.head(top_n).reset_index(drop=True), "수익컬럼 없음 → 상위"

    # 그리드 최고 수익과 B&H 사이 구간만
    grid_max = cand[ret_col].max()
    note = ""
    if bh_ret is not None and pd.notna(bh_ret):
        lo, hi = min(bh_ret, grid_max), max(bh_ret, grid_max)
        band = cand[(cand[ret_col] >= lo) & (cand[ret_col] <= hi)]
        if len(band) > 0:
            cand = band
            note = f"B&H({bh_ret*100:+.1f}%)~최고({grid_max*100:+.1f}%) 구간"
        else:
            note = "구간 후보 없음 → 전체"
    # 수익 높은 순 → 절반만
    cand = cand.sort_values(ret_col, ascending=False).reset_index(drop=True)
    half = max(1, len(cand) // 2)
    cand = cand.head(min(half, top_n)).reset_index(drop=True)
    note = (note + f", 수익 상위 절반 {len(cand)}개").strip(', ')
    return cand, note


def _extract_runday_info(_d, _t):
    """실행일(가장 최근 거래일) 기준 실제 포지션·액션·점수와 최근 매수/매도 정보 추출 (요청).
       반환 키: run_pos(1보유/0현금), run_action(일별 백테스트 액션), run_buy_score, run_sell_score,
                recent_buy_date, recent_buy_price, recent_sell_date, recent_sell_price."""
    info = {'run_pos': np.nan, 'run_action': None,
            'run_buy_score': np.nan, 'run_sell_score': np.nan,
            'recent_buy_date': None, 'recent_buy_price': np.nan,
            'recent_sell_date': None, 'recent_sell_price': np.nan}
    try:
        if _d is not None and len(_d) > 0:
            last = _d.iloc[-1]
            info['run_pos']        = int(last.get('position_pre', 0))
            _act = last.get('action', '')
            info['run_action']     = str(_act).strip() if (_act is not None and str(_act).strip()) else '관망'
            info['run_buy_score']  = float(last.get('buy_count', np.nan))
            info['run_sell_score'] = float(last.get('sell_count', np.nan))
            # 현재 보유 중이면 '열린 포지션' 진입가/매수일
            if info['run_pos'] == 1 and pd.notna(last.get('entry_price', np.nan)):
                info['recent_buy_price'] = float(last['entry_price'])
                if 'action' in _d.columns:
                    _bm = _d['action'].astype(str).str.contains('매수', na=False)
                    if _bm.any():
                        info['recent_buy_date'] = _d.loc[_bm, 'date'].iloc[-1]
        if _t is not None and len(_t) > 0:
            lt = _t.iloc[-1]
            # 가장 최근 청산(매도)
            info['recent_sell_date']  = lt.get('exit_date', None)
            info['recent_sell_price'] = float(lt.get('exit_price', np.nan))
            # 현금 상태면 최근 매수 = 마지막으로 청산된 거래의 진입
            if info['run_pos'] != 1:
                info['recent_buy_date']  = lt.get('entry_date', None)
                info['recent_buy_price'] = float(lt.get('entry_price', np.nan))
    except Exception:
        pass
    return info


def _verify_staged_candidates(merged_table, feat, close, pool_map, *,
                              horizon, dd_limit, ru_limit, stop_loss_pct, anchor_mode,
                              bh_ret=None, anchor_safe_buy=None, anchor_safe_sell=None):
    """staged merged_table의 후보를 각 메타조합 풀로 실제 일별 백테스트.
       실제 MDD 기준 통과분 중 실제승률 -10%p 범위에서 수익 최고 선정.
       반환: (best_inner, merged_table_with_real_cols)"""
    win_tol = globals().get('WINRATE_TOLERANCE', 0.10)
    top_n   = int(globals().get('VERIFY_TOP_N', 10000))
    mdd_lim = globals().get('VERIFY_MAX_DRAWDOWN_LIMIT', None)

    cand, _note = _pick_verify_candidates(merged_table, bh_ret=bh_ret)
    print(f"\n  🔬 실거래 검증 — {_note} → {len(cand)}개를 실제 일별 백테스트로 재계산 "
          f"(모든 변수 반복 끝난 뒤 1회)")
    try: _ZCACHE.clear()   # ★ 검증 전 z-캐시 정리 (요청)
    except Exception: pass

    real_rows = []
    _t0 = time.time()
    _miss_pool = 0
    for _i, _r in cand.iterrows():
        _key = (round(float(_r.get('meta_wilson_z', 0)),4),
                int(_r.get('meta_pct_low', 0)), int(_r.get('meta_pct_high', 0)),
                round(float(_r.get('meta_corr_limit', 0)),4))
        pools = pool_map.get(_key)
        if pools is None:
            _miss_pool += 1
            continue
        bp, sp = pools
        try:
            _d, _t, _cur, _bu, _su = daily_ensemble_backtest(
                feat, close, bp, sp,
                K_buy=int(_r['K_buy']), K_sell=int(_r['K_sell']),
                vote_buy=int(_r['vote_buy']), vote_sell=int(_r['vote_sell']),
                cost=COST_PER_TRADE, horizon=horizon,
                dd_limit=dd_limit, ru_limit=ru_limit,
                stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
                anchor_safe_buy=anchor_safe_buy, anchor_safe_sell=anchor_safe_sell)
        except Exception:
            continue
        row = _r.to_dict()
        row['real_win_rate']      = float(_cur.get('win_rate', 0.0))
        row['real_max_drawdown']  = float(_cur.get('max_drawdown', 0.0))
        row['real_total_return']  = float(_cur.get('cum_return_pct', 0.0)) / 100.0
        row['real_n_trades']      = int(_cur.get('n_trades', 0))
        row['real_buy_success']   = float(_cur.get('buy_success_rate', 0.0))
        row['real_sell_success']  = float(_cur.get('sell_success_rate', 0.0))
        row['real_avg_success']   = float(_cur.get('avg_success_rate', 0.0))
        row['real_buy_match']     = float(_cur.get('anchor_buy_match_rate', 0.0))
        row['real_sell_match']    = float(_cur.get('anchor_sell_match_rate', 0.0))
        row['real_avg_match']     = float(_cur.get('anchor_avg_match_rate', 0.0))
        row['real_anchor_return'] = float(_cur.get('anchor_strategy_return', 0.0))
        # ★ 실행일(가장 최근일) 실제 포지션·점수 + 최근 매수/매도 정보 (요청)
        _ri = _extract_runday_info(_d, _t)
        row.update(_ri)
        # ★ OOS 기능 OFF (요청) — OOS 백테스트 생략, 컬럼은 NaN
        if globals().get('OOS_SELECT_BY_OOS_RETURN', False):
            try:
                _oos_m = int(globals().get('OOS_MONTHS', 1))
                _last_dt = feat.index.max()
                _oos_start = _last_dt - pd.DateOffset(months=_oos_m)
                _omask = feat.index >= _oos_start
                _omask = np.asarray(_omask)
                _feat_oos = feat.loc[_omask]
                _close_oos = close.reindex(_feat_oos.index)
                _asb_o = (anchor_safe_buy[_omask] if (anchor_safe_buy is not None and len(anchor_safe_buy)==len(_omask)) else None)
                _ass_o = (anchor_safe_sell[_omask] if (anchor_safe_sell is not None and len(anchor_safe_sell)==len(_omask)) else None)
                if len(_feat_oos) >= 5:
                    _do, _to, _curo, _buo, _suo = daily_ensemble_backtest(
                        _feat_oos, _close_oos, bp, sp,
                        K_buy=int(_r['K_buy']), K_sell=int(_r['K_sell']),
                        vote_buy=int(_r['vote_buy']), vote_sell=int(_r['vote_sell']),
                        cost=COST_PER_TRADE, horizon=horizon,
                        dd_limit=dd_limit, ru_limit=ru_limit,
                        stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
                        anchor_safe_buy=_asb_o, anchor_safe_sell=_ass_o)
                    row['oos_total_return'] = float(_curo.get('cum_return_pct', 0.0)) / 100.0
                    row['oos_buy_success']  = float(_curo.get('buy_success_rate', 0.0))
                    row['oos_sell_success'] = float(_curo.get('sell_success_rate', 0.0))
                    row['oos_win_rate']     = float(_curo.get('win_rate', 0.0))
                    row['oos_n_trades']     = int(_curo.get('n_trades', 0))
                else:
                    row['oos_total_return'] = float('nan')
                    row['oos_buy_success'] = row['oos_sell_success'] = float('nan')
                    row['oos_win_rate'] = float('nan'); row['oos_n_trades'] = 0
            except Exception:
                row['oos_total_return'] = float('nan')
                row['oos_buy_success'] = row['oos_sell_success'] = float('nan')
                row['oos_win_rate'] = float('nan'); row['oos_n_trades'] = 0
        real_rows.append(row)
        if (_i + 1) % 1000 == 0:
            print(f"     ... {_i+1}/{len(cand)} 검증  (경과 {time.time()-_t0:.0f}초)")

    if _miss_pool > 0:
        print(f"     ℹ 풀 정보를 못 찾아 건너뛴 후보 {_miss_pool}개 (메타조합 캐시 불일치)")
    if not real_rows:
        print(f"  ⚠ 실거래 검증 후보가 모두 실패 → 그리드 근사 기준 사용")
        return None, merged_table

    verified = pd.DataFrame(real_rows)

    # ★ 보정 통합 (요청) — 수익 상위 N개 후보에 지표 가중치 보정 적용 후 재백테스트.
    #   보정 후 수익/손실/승률을 corr_* 컬럼에 기록. 보정 후 수익으로 최종 선정.
    if globals().get('USE_ANCHOR_MATCH_CORRECTION', True):
        verified = _apply_correction_to_candidates(
            verified, feat, close, pool_map,
            horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
            stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
            anchor_safe_buy=anchor_safe_buy, anchor_safe_sell=anchor_safe_sell)

    # ★ 선정 필터 수치를 '보정 후'로 (요청) — 보정 적용된 조합은 corr_*, 아니면 real_* 사용.
    #   sel_mdd / sel_win / sel_ret = 선정에 쓸 '유효 수치'.
    _use_corr_sel = globals().get('SELECT_FILTER_BY_CORRECTED', True)
    def _eff_col(colc, colr):
        base = verified[colr].astype(float).values
        if _use_corr_sel and colc in verified.columns:
            _ca = verified.get('corr_applied', pd.Series(False, index=verified.index))
            _ca = _ca.fillna(False).astype(bool).values
            cval = verified[colc].astype(float).values
            return np.where(_ca & ~np.isnan(cval), cval, base)
        return base
    verified['sel_max_drawdown'] = _eff_col('corr_max_drawdown', 'real_max_drawdown')
    verified['sel_win_rate']     = _eff_col('corr_win_rate',     'real_win_rate')
    verified['sel_total_return'] = _eff_col('corr_total_return', 'real_total_return')

    pool = verified
    if mdd_lim is not None:
        safe = verified[verified['sel_max_drawdown'] >= mdd_lim]    # ★ 보정후 최대손실로 필터
        if len(safe) > 0:
            pool = safe
            print(f"  🛡 (보정후) 최대 거래손실 {mdd_lim*100:.1f}% 이하(안전) 후보 {len(safe)}개로 선정 "
                  f"(전체 검증 {len(verified)}개 중)")
        else:
            print(f"  ⚠ (보정후) 최대 거래손실 {mdd_lim*100:.1f}% 이하 후보 없음 → 가장 손실 얕은 조합 선정")
            pool = verified.sort_values('sel_max_drawdown', ascending=False)

    best_inner, sel = _final_pick_by_real(pool, win_tol)
    out = verified.sort_values('sel_total_return', ascending=False).reset_index(drop=True)
    print(f"     실제 승률 {sel['real_win_rate']*100:.1f}%, "
          f"실제 평균성공 {sel['real_avg_success']*100:.1f}% "
          f"(매수 {sel['real_buy_success']*100:.1f}% / 매도 {sel['real_sell_success']*100:.1f}%), "
          f"실제 최대손실 {sel['real_max_drawdown']*100:.2f}%, "
          f"실제 수익 {sel['real_total_return']*100:+.2f}%")
    return best_inner, out


def _verify_candidates_by_daily(inner_passed, feat, close, buy_pool, sell_pool, *,
                                horizon, dd_limit, ru_limit, stop_loss_pct,
                                anchor_mode, anchor_safe_buy, anchor_safe_sell,
                                best_inner_fallback=None, bh_ret=None):
    """그리드 승률 상위 후보들을 '실제 일별 백테스트'로 한 번에 재계산하고,
       실제 최대거래손실(MDD) 기준을 통과한 것 중 실제 승률 -10%p 범위에서 수익 최고를 선정.
       반환: (best_inner, inner_passed_with_real_cols)
    """
    win_tol = globals().get('WINRATE_TOLERANCE', 0.10)
    top_n   = int(globals().get('VERIFY_TOP_N', 10000))
    mdd_lim = globals().get('VERIFY_MAX_DRAWDOWN_LIMIT', None)

    cand, _note = _pick_verify_candidates(inner_passed, bh_ret=bh_ret)
    print(f"\n  🔬 실거래 검증 — {_note} → {len(cand)}개를 실제 일별 백테스트로 재계산 "
          f"(모든 변수 반복 끝난 뒤 1회)")
    try: _ZCACHE.clear()   # ★ 검증 전 z-캐시 정리 (요청)
    except Exception: pass
    real_rows = []
    _t_start = time.time()
    for _i, _r in cand.iterrows():
        _kb, _vb = int(_r['K_buy']), int(_r['vote_buy'])
        _ks, _vs = int(_r['K_sell']), int(_r['vote_sell'])
        try:
            _d, _t, _cur, _bu, _su = daily_ensemble_backtest(
                feat, close, buy_pool, sell_pool,
                K_buy=_kb, K_sell=_ks, vote_buy=_vb, vote_sell=_vs,
                cost=COST_PER_TRADE, horizon=horizon,
                dd_limit=dd_limit, ru_limit=ru_limit,
                stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
                anchor_safe_buy=anchor_safe_buy, anchor_safe_sell=anchor_safe_sell)
        except Exception:
            continue
        row = _r.to_dict()
        row['real_win_rate']      = float(_cur.get('win_rate', 0.0))
        row['real_max_drawdown']  = float(_cur.get('max_drawdown', 0.0))
        row['real_total_return']  = float(_cur.get('cum_return_pct', 0.0)) / 100.0
        row['real_n_trades']      = int(_cur.get('n_trades', 0))
        row['real_buy_success']   = float(_cur.get('buy_success_rate', 0.0))
        row['real_sell_success']  = float(_cur.get('sell_success_rate', 0.0))
        row['real_avg_success']   = float(_cur.get('avg_success_rate', 0.0))
        row['real_buy_match']     = float(_cur.get('anchor_buy_match_rate', 0.0))
        row['real_sell_match']    = float(_cur.get('anchor_sell_match_rate', 0.0))
        row['real_avg_match']     = float(_cur.get('anchor_avg_match_rate', 0.0))
        row['real_anchor_return'] = float(_cur.get('anchor_strategy_return', 0.0))
        # ★ 실행일 실제 포지션·점수 + 최근 매수/매도 정보 (요청)
        row.update(_extract_runday_info(_d, _t))
        # ★ OOS 기능 OFF (요청) — OOS 백테스트 생략, 컬럼 NaN
        if globals().get('OOS_SELECT_BY_OOS_RETURN', False):
            try:
                _oos_m = int(globals().get('OOS_MONTHS', 1))
                _last_dt = feat.index.max()
                _oos_start = _last_dt - pd.DateOffset(months=_oos_m)
                _omask = feat.index >= _oos_start
                _omask = np.asarray(_omask)
                _feat_oos = feat.loc[_omask]
                _close_oos = close.reindex(_feat_oos.index)
                _asb_o = (anchor_safe_buy[_omask] if (anchor_safe_buy is not None and len(anchor_safe_buy)==len(_omask)) else None)
                _ass_o = (anchor_safe_sell[_omask] if (anchor_safe_sell is not None and len(anchor_safe_sell)==len(_omask)) else None)
                if len(_feat_oos) >= 5:
                    _do, _to, _curo, _buo, _suo = daily_ensemble_backtest(
                        _feat_oos, _close_oos, buy_pool, sell_pool,
                        K_buy=int(_r['K_buy']), K_sell=int(_r['K_sell']),
                        vote_buy=int(_r['vote_buy']), vote_sell=int(_r['vote_sell']),
                        cost=COST_PER_TRADE, horizon=horizon,
                        dd_limit=dd_limit, ru_limit=ru_limit,
                        stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
                        anchor_safe_buy=_asb_o, anchor_safe_sell=_ass_o)
                    row['oos_total_return'] = float(_curo.get('cum_return_pct', 0.0)) / 100.0
                    row['oos_buy_success']  = float(_curo.get('buy_success_rate', 0.0))
                    row['oos_sell_success'] = float(_curo.get('sell_success_rate', 0.0))
                    row['oos_win_rate']     = float(_curo.get('win_rate', 0.0))
                    row['oos_n_trades']     = int(_curo.get('n_trades', 0))
                else:
                    row['oos_total_return'] = float('nan')
                    row['oos_buy_success'] = row['oos_sell_success'] = float('nan')
                    row['oos_win_rate'] = float('nan'); row['oos_n_trades'] = 0
            except Exception:
                row['oos_total_return'] = float('nan')
                row['oos_buy_success'] = row['oos_sell_success'] = float('nan')
                row['oos_win_rate'] = float('nan'); row['oos_n_trades'] = 0
        real_rows.append(row)
        if (_i + 1) % 1000 == 0:
            print(f"     ... {_i+1}/{len(cand)} 검증  (경과 {time.time()-_t_start:.0f}초)")

    if not real_rows:
        print(f"  ⚠ 실거래 검증 후보가 모두 실패 → 그리드 근사 기준 사용")
        return best_inner_fallback, inner_passed

    verified = pd.DataFrame(real_rows)

    # ★ 보정 통합 (요청) — 일반 경로(STAGED OFF)도 보정 적용. 단일 풀을 키 dict로 감싼다.
    if globals().get('USE_ANCHOR_MATCH_CORRECTION', True):
        try:
            _single_map = {}
            for _idx, _vr in verified.iterrows():
                _k = (round(float(_vr.get('meta_wilson_z', 0)),4),
                      int(_vr.get('meta_pct_low', 0)), int(_vr.get('meta_pct_high', 0)),
                      round(float(_vr.get('meta_corr_limit', 0)),4))
                _single_map[_k] = (buy_pool, sell_pool)
            if not _single_map:  # 메타 컬럼이 없으면 더미 키
                _single_map = {(0.0,0,0,0.0): (buy_pool, sell_pool)}
                verified['meta_wilson_z'] = 0.0; verified['meta_pct_low'] = 0
                verified['meta_pct_high'] = 0; verified['meta_corr_limit'] = 0.0
            verified = _apply_correction_to_candidates(
                verified, feat, close, _single_map,
                horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
                stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
                anchor_safe_buy=anchor_safe_buy, anchor_safe_sell=anchor_safe_sell)
        except Exception as _ce:
            print(f"  ⚠ 보정 적용 실패(일반 경로): {_ce}")

    # ★ 선정 필터 수치를 '보정 후'로 (요청) — 보정 적용된 조합은 corr_*, 아니면 real_*.
    _use_corr_sel = globals().get('SELECT_FILTER_BY_CORRECTED', True)
    def _eff_col(colc, colr):
        base = verified[colr].astype(float).values
        if _use_corr_sel and colc in verified.columns:
            _ca = verified.get('corr_applied', pd.Series(False, index=verified.index))
            _ca = _ca.fillna(False).astype(bool).values
            cval = verified[colc].astype(float).values
            return np.where(_ca & ~np.isnan(cval), cval, base)
        return base
    verified['sel_max_drawdown'] = _eff_col('corr_max_drawdown', 'real_max_drawdown')
    verified['sel_win_rate']     = _eff_col('corr_win_rate',     'real_win_rate')
    verified['sel_total_return'] = _eff_col('corr_total_return', 'real_total_return')

    # ★ 1) (보정 후) 최대 거래손실(MDD) 기준 필터 (요청)
    pool = verified
    if mdd_lim is not None:
        safe = verified[verified['sel_max_drawdown'] >= mdd_lim]
        if len(safe) > 0:
            pool = safe
            print(f"  🛡 (보정후) 최대 거래손실 {mdd_lim*100:.1f}% 이하(안전) 후보 {len(safe)}개로 선정 "
                  f"(전체 검증 {len(verified)}개 중)")
        else:
            print(f"  ⚠ (보정후) 최대 거래손실 {mdd_lim*100:.1f}% 이하 후보가 없음 "
                  f"→ 가장 손실 얕은 조합으로 선정 (기준 완화)")
            pool = verified.sort_values('sel_max_drawdown', ascending=False)

    best_inner, sel = _final_pick_by_real(pool, win_tol)
    inner_passed_out = verified.sort_values('sel_total_return', ascending=False).reset_index(drop=True)

    print(f"     실제 승률 {sel['real_win_rate']*100:.1f}%, "
          f"실제 평균성공 {sel['real_avg_success']*100:.1f}% "
          f"(매수 {sel['real_buy_success']*100:.1f}% / 매도 {sel['real_sell_success']*100:.1f}%), "
          f"실제 최대거래손실 {sel['real_max_drawdown']*100:.2f}%, "
          f"실제 누적수익 {sel['real_total_return']*100:+.2f}%")
    return best_inner, inner_passed_out


def run_ensemble_search(*, eval_start=EVAL_START,
                         horizon=HORIZON_DAYS,
                         dd_limit=DRAWDOWN_LIMIT_BUY,
                         ru_limit=RUNUP_LIMIT_SELL,
                         n_thresholds=N_THRESHOLDS,
                         max_indicators=MAX_INDICATORS,
                         k_buy_range=K_BUY_RANGE,
                         k_sell_range=K_SELL_RANGE,
                         vote_ratio_buy=VOTE_RATIO_BUY,
                         vote_ratio_sell=VOTE_RATIO_SELL,
                         min_trades_daily=MIN_TRADES_DAILY,
                         max_drawdown_limit_pct='__USE_GLOBAL__',
                         stop_loss_pct=STOP_LOSS_PCT,
                         selection_tolerance=SELECTION_TOLERANCE,
                         anchor_match_priority=ANCHOR_MATCH_PRIORITY,
                         anchor_match_tolerance=ANCHOR_MATCH_TOLERANCE,
                         anchor_mode=ANCHOR_MODE,
                         auto_anchor=AUTO_ANCHOR,
                         auto_anchor_window=AUTO_ANCHOR_WINDOW,
                         auto_anchor_lookforward=AUTO_ANCHOR_LOOKFORWARD,
                         auto_anchor_min_rise=AUTO_ANCHOR_MIN_RISE,
                         auto_anchor_min_drop=AUTO_ANCHOR_MIN_DROP,
                         auto_anchor_price_tolerance=AUTO_ANCHOR_PRICE_TOLERANCE,
                         auto_anchor_max_dates=AUTO_ANCHOR_MAX_DATES,
                         anchor_buy_dates=None,
                         anchor_sell_dates=None,
                         meta_grid=META_GRID,
                         select_by=SELECT_BY,
                         oos_enabled=OOS_ENABLED,
                         oos_start=OOS_START,
                         write_output=True,
                         output_file=None,
                         inject_combined_table=None,
                         force_best_combo=None,
                         force_corr=None,
                         inject_pools=None):
    print('=' * 72)
    print('  매수/매도 앙상블 — 메타 그리드 자동 튜닝')
    print('=' * 72)
    # ★ 요청: 적용된 탐색 설정 로그 (확인용)
    print('  [탐색 설정]')
    print(f'    · 성공 판정: 신호 다음날(HORIZON={HORIZON_DAYS}일 이내) 종가 ±한도 도달')
    if globals().get('SEARCH_SUCCESS_LIMIT', False):
        _sl = globals().get('STAGE_SUCCESS_LIMIT', [DRAWDOWN_LIMIT_BUY])
        print(f'    · 상승/하락률 한도 탐색: {[f"{x*100:.0f}%" for x in _sl]} (각 한도로 성공률 계산 → 최적 선정)')
    else:
        print(f'    · 상승/하락률 한도: 고정 {DRAWDOWN_LIMIT_BUY*100:.0f}%')
    print(f'    · 분위(pct) 탐색: {STAGE_PCT_RANGE}  (0,100=전체)')
    print(f'    · 윌슨 z 탐색: {STAGE_WILSON_Z}')
    print(f'    · 풀 조건: 신호수 > {POOL_SUCCESS_MIN_SIG}개  &  성공률 ≥ 사용지표 최소성공률(컷 {POOL_SUCCESS_MIN_RATE*100:.0f}%)')
    if globals().get('SEARCH_WEIGHT_SCHEME', False):
        print(f'    · 가중 스킴 탐색(윌슨 후): weight=p**g, g∈{globals().get("NET_WEIGHT_SCHEMES", [1.0])}')
    print(f'    · 임계치별 다중 성공률 가중: {globals().get("NET_MULTI_THRESHOLD_WEIGHT", False)} '
          f'(그날 켜진 임계 중 최고 성공률 적용)')
    # ★ 미래 예측 로직 개선 (요청) — 리드타임/스킬/홀드아웃 설정 로그
    if globals().get('LEAD_TIME_SEARCH', False):
        print(f'    · 리드타임 탐색: ON — 선행일 후보 {globals().get("LEAD_HORIZONS")} '
              f'(스킬=성공률-기저확률 최대 h, 삼중배리어={globals().get("LEAD_TRIPLE_BARRIER", False)})')
    if globals().get('LEAD_SHIFT_ENABLED', False):
        print(f'    · 신호 지연 정렬: ON — 최대 {globals().get("LEAD_SHIFT_MAX")}일, '
              f'훈련 스킬 +{float(globals().get("LEAD_SHIFT_MIN_GAIN",0.05))*100:.0f}%p↑ & 홀드아웃 유지 시만 채택'
              + (' (★ 풀 선출 단계에서 지연 동시 탐색 — 선행 지표가 처음부터 풀에 진입)'
                 if globals().get('LEAD_SELECT_IN_SCORING', True) else ''))
    if globals().get('POOL_REQUIRE_SKILL', False):
        print(f'    · 스킬 필터: ON — 성공률이 기저확률을 {float(globals().get("POOL_MIN_SKILL",0.0))*100:.0f}%p 초과해야 풀 채택')
    if globals().get('POOL_HOLDOUT_GUARD', False):
        print(f'    · 홀드아웃 가드: ON — 뒤 {float(globals().get("POOL_HOLDOUT_FRACTION",0.3))*100:.0f}% 구간 스킬 '
              f'{float(globals().get("POOL_HOLDOUT_MIN_SKILL",-0.1))*100:.0f}%p 미만 지표 제외')
    if globals().get('VERIFY_SHEET_ENABLED', True):
        print(f"    · 검증 시트: ON — '검증_예측로직' 시트에 기능 점검·지표별 예측력 독립 재계산 기록")
    print('=' * 72)

    # ★ MDD 한도 — sentinel이면 호출 시점의 전역 MAX_DRAWDOWN_LIMIT_PCT를 다시 읽음
    #   (Colab에서 변수만 바꾸고 함수 재정의 안 해도 최신값 반영되도록)
    if max_drawdown_limit_pct == '__USE_GLOBAL__':
        max_drawdown_limit_pct = globals().get('MAX_DRAWDOWN_LIMIT_PCT', None)

    if anchor_buy_dates is None:  anchor_buy_dates  = list(ANCHOR_BUY_DATES)
    if anchor_sell_dates is None: anchor_sell_dates = list(ANCHOR_SELL_DATES)

    feat, close, ticker = _resolve_data()
    mask = feat.index >= pd.Timestamp(eval_start)
    feat = feat.loc[mask]
    close = close.reindex(feat.index)
    # ★ 거래하지 않는 날(공휴일 등 종가 NaN) 제거 — ffill로 가짜 변동 만들지 않음(요청).
    #   여기서 한 번 제거하면 anchor·grid·daily 모든 단계가 같은 길이의 거래일만 사용
    #   → 신규상장 종목(CRWV 등) 길이 불일치/broadcast 충돌 방지.
    _valid = close.notna()
    if _valid.sum() < len(close):
        _removed = len(close) - int(_valid.sum())
        feat = feat[_valid.values]
        close = close[_valid.values]
        print(f"  ℹ 거래 없는 날(종가 결측) {_removed}일 제외 → {len(close)}일")
    print(f"  기간: {feat.index[0].date()} ~ {feat.index[-1].date()}  ({len(feat)}일)")

    oos_start_idx = None
    if oos_enabled:
        oos_ts = pd.Timestamp(oos_start)
        pos_arr = np.where(feat.index >= oos_ts)[0]
        if len(pos_arr) == 0:
            print(f"  ⚠ OOS_START({oos_start})가 데이터 범위 밖 (전부 IS) → OOS 비활성화")
            oos_enabled = False
        elif pos_arr[0] == 0:
            print(f"  ⚠ OOS_START({oos_start})가 데이터 시작보다 이르거나 같음 (IS 없음) → OOS 비활성화")
            oos_enabled = False
        else:
            oos_start_idx = int(pos_arr[0])
            is_n = oos_start_idx
            oos_n = len(feat) - oos_start_idx
            print(f"  ★ OOS 검증 ON — IS(학습) {is_n}일 [{feat.index[0].date()}~{feat.index[oos_start_idx-1].date()}]  "
                  f"/ OOS(검증) {oos_n}일 [{feat.index[oos_start_idx].date()}~{feat.index[-1].date()}]")
            if is_n < 60 or oos_n < 20:
                print(f"     ⚠ IS 또는 OOS 구간이 짧음 (IS {is_n}, OOS {oos_n}) — 신뢰도 주의")
    else:
        print(f"  ★ OOS 검증: OFF (전체 기간 단일 백테스트)")

    if max_drawdown_limit_pct is not None:
        print(f"  ★ MDD 한도: -{abs(max_drawdown_limit_pct):.2f}%")
    else:
        print(f"  ★ MDD 한도: 없음")
    print(f"  ★ 최소 거래수: {min_trades_daily}회")
    if stop_loss_pct is not None and stop_loss_pct > 0:
        print(f"  ⛔ 손절매 한도: -{stop_loss_pct*100:.2f}%")
    else:
        print(f"  ⛔ 손절매: 없음")
    print(f"  ★ 충돌 해결: count/K 비율 큰 쪽 우선")
    if globals().get('USE_WEIGHTED_VOTE', False):
        print(f"  ★ 투표 방식: 가중 투표 (성공률 비례, 상한 {globals().get('WEIGHT_MAX_RATIO',1.6)}배)")
    else:
        print(f"  ★ 투표 방식: 일반 투표 (모두 1표)")
    _prio_re = globals().get('SELECTION_PRIORITY', 'balacc_return')
    if _prio_re == 'stability':
        print(f"  ★ 선정: 안정성 종합점수(매도·매수성공·수익·MDD방어 가중 기하평균) 최대")
    elif _prio_re == 'sell_buy_return':
        print(f"  ★ 선정 우선순위: 매도성공률 → 매수성공률 → 누적수익 (밴드 {selection_tolerance*100:.2f}%p)")
    elif selection_tolerance > 0:
        print(f"  ★ 선정 기준: 평균 BalAcc top - {selection_tolerance*100:.2f}%p 이내 중 수익률 최대 (Tolerance Band)")
    else:
        print(f"  ★ 선정 기준: 1차 평균 BalAcc → 2차 수익률 (strict)")
    if anchor_match_priority and anchor_mode:
        print(f"  ⚓ 매칭률 우선순위 ON — 매칭률 ±{anchor_match_tolerance*100:.1f}%p → BalAcc → 수익률")

    _anchor_prio = None
    if anchor_mode and auto_anchor:
        print(f"  ⚓ AUTO ANCHOR 자동 계산:")
        print(f"     window={auto_anchor_window}일, lookforward={auto_anchor_lookforward}일")
        print(f"     매수 정답: 좌우 {auto_anchor_window}일 local min + 그 후 +{auto_anchor_min_rise*100:.1f}% 이상 상승")
        print(f"     매도 정답: 좌우 {auto_anchor_window}일 local max + 그 후 -{auto_anchor_min_drop*100:.1f}% 이상 하락")
        print(f"     가격 근접 확장: base ±{auto_anchor_price_tolerance*100:.2f}% 이내 모든 날 포함")
        anchor_buy_dates, anchor_sell_dates, _anchor_prio = auto_compute_anchor_dates(
            feat.index, close,
            window=auto_anchor_window,
            lookforward=auto_anchor_lookforward,
            min_rise_after_buy=auto_anchor_min_rise,
            min_drop_after_sell=auto_anchor_min_drop,
            price_tolerance=auto_anchor_price_tolerance,
            max_dates=auto_anchor_max_dates,
            return_priority=True,
        )
        print(f"     ▷ 자동 계산 결과: 매수 정답 {len(anchor_buy_dates)}일, 매도 정답 {len(anchor_sell_dates)}일")
        if len(anchor_buy_dates) > 0:
            preview_b = anchor_buy_dates[:5] + (['...'] if len(anchor_buy_dates) > 5 else [])
            print(f"       매수 예시: {preview_b}")
        if len(anchor_sell_dates) > 0:
            preview_s = anchor_sell_dates[:5] + (['...'] if len(anchor_sell_dates) > 5 else [])
            print(f"       매도 예시: {preview_s}")
    elif anchor_mode:
        print(f"  ⚓ ANCHOR 수동 입력 — 매수 정답 {len(anchor_buy_dates)}일, 매도 정답 {len(anchor_sell_dates)}일")
    else:
        print(f"  ⚓ ANCHOR 보정: OFF")

    print(f"\n  메타 그리드:")
    for k, v in meta_grid.items():
        print(f"    {k:14}: {v}")

    if PREFILTER_ENABLED and inject_pools is None:
        feat_filtered = prefilter_indicators(
            feat, close, horizon=horizon,
            min_corr=PREFILTER_MIN_CORR,
            min_variance_rel=PREFILTER_MIN_VARIANCE_REL,
            max_nan_ratio=PREFILTER_MAX_NAN_RATIO,
        )
        if len(feat_filtered.columns) < 30:
            print(f"     ⚠ 필터 후 지표 {len(feat_filtered.columns)}개 < 30 → 원본 {len(feat.columns)}개 그대로 사용")
        else:
            feat = feat_filtered
    elif inject_pools is not None:
        print(f"  ♻ 재현 모드 — 사전 필터 건너뜀 (엑셀 지표 풀의 지표를 그대로 사용)")
    else:
        print(f"  ℹ 사전 필터 비활성화 — 전체 {len(feat.columns)}개 지표 사용")

    indicators = _select_indicators(feat, max_indicators)
    print(f"\n  후보 지표: {len(indicators)}개")

    anchor_safe_buy = anchor_safe_sell = None
    if anchor_mode:
        anchor_safe_buy, anchor_safe_sell = _compute_anchor_arrays(
            feat.index, anchor_buy_dates, anchor_sell_dates)
        n_b = int(anchor_safe_buy.sum())
        n_s = int(anchor_safe_sell.sum())
        if n_b == 0 and n_s == 0:
            print("  ⚠ ANCHOR 정답일이 데이터 범위 내 0개 — ANCHOR 보정 비활성화")
            anchor_mode = False
            anchor_safe_buy = anchor_safe_sell = None
        else:
            print(f"  ⚓ ANCHOR 매칭 — 매수 {n_b}일, 매도 {n_s}일 (데이터 범위 내)")

    # ★ B&H 수익률 먼저 계산 (meta_grid_search에 넘겨 B&H 미달 조합 제외용)
    close_arr = close.values.astype(np.float64)
    if oos_enabled and oos_start_idx is not None and 0 < oos_start_idx < len(close_arr):
        bh_ret_for_filter = _bh_sum_return(close_arr[:oos_start_idx])
    else:
        bh_ret_for_filter = _bh_sum_return(close_arr)

    print(f"\n[메타 그리드 탐색]   (B&H 미달 조합 제외, 기준 B&H={bh_ret_for_filter*100:+.2f}%)")
    if inject_pools is not None:
        # ★ 재현 모드 — 엑셀에서 읽은 지표 풀을 그대로 사용 (지표 재선별 안 함).
        #   meta_grid_search(지표 선별·그리드 탐색) 전체를 건너뛰고, 주입된 풀 +
        #   force_best_combo의 K/vote로 곧장 daily 백테스트 → 원본과 동일 지표·조합 재현.
        buy_pool, sell_pool = inject_pools
        print(f"  ♻ 재현 모드 — 엑셀의 지표 풀 그대로 사용 "
              f"(매수 {len(buy_pool)}개 / 매도 {len(sell_pool)}개 지표, 재선별 안 함)")
        # ★ 재현: 원본 스냅샷에 k순신호 풀이 있으면 그대로 사용(탐색 0). 없을 때만 재구성.
        _mp0 = globals().get('_KNET_MULTI_POOL')
        if _mp0 and _mp0[0] == ticker and _mp0[1] is not None:
            print(f"  ♻ 재현: 원본 k순신호 풀 그대로 사용 (탐색 생략) "
                  f"— 매수 {_mp0[1]['indicator'].nunique()} / 매도 {_mp0[2]['indicator'].nunique()}지표")
        else:
            try:
                _build_and_pick_knet_pool(feat, close, indicators=indicators,
                                          n_thresholds=n_thresholds, horizon=horizon, ticker=ticker)
            except Exception as _re:
                print(f"  ⚠ 재현용 합친 풀 생성 실패(폴백): {_re}")
        if force_best_combo is not None:
            best_inner = {
                'K_buy': int(force_best_combo['K_buy']),
                'vote_buy': int(force_best_combo['vote_buy']),
                'K_sell': int(force_best_combo['K_sell']),
                'vote_sell': int(force_best_combo['vote_sell']),
            }
        else:
            best_inner = {'K_buy': k_buy_range[0], 'vote_buy': int(k_buy_range[0]*vote_ratio_buy[0]),
                          'K_sell': k_sell_range[0], 'vote_sell': int(k_sell_range[0]*vote_ratio_sell[0])}
        # best_meta는 주입된 메타 그리드값으로 채움(엑셀 표시용)
        _mg0 = meta_grid
        best_meta = {
            'wilson_z': _mg0['wilson_z'][0],
            'pct_low': _mg0['pct_range'][0][0], 'pct_high': _mg0['pct_range'][0][1],
            'min_signals': _mg0['min_signals'][0], 'corr_limit': _mg0['corr_limit'][0],
            'top_n_pool_buy': len(buy_pool), 'top_n_pool_sell': len(sell_pool),
        }
        meta_results_df = pd.DataFrame()
        inner_all = pd.DataFrame()
        inner_passed = inject_combined_table if inject_combined_table is not None else pd.DataFrame()
    else:
        def _run_meta(_dd, _ru):
            return meta_grid_search(
                feat, close,
                meta_grid=meta_grid, indicators=indicators,
                n_thresholds=n_thresholds, horizon=horizon,
                dd_limit=_dd, ru_limit=_ru,
                k_buy_range=k_buy_range, k_sell_range=k_sell_range,
                vote_ratio_buy=vote_ratio_buy, vote_ratio_sell=vote_ratio_sell,
                cost=COST_PER_TRADE,
                min_trades_daily=min_trades_daily,
                mdd_limit_pct=max_drawdown_limit_pct,
                select_by=select_by,
                stop_loss_pct=stop_loss_pct,
                selection_tolerance=selection_tolerance,
                anchor_match_priority=anchor_match_priority,
                anchor_match_tolerance=anchor_match_tolerance,
                anchor_mode=anchor_mode,
                anchor_safe_buy=anchor_safe_buy,
                anchor_safe_sell=anchor_safe_sell,
                oos_start_idx=oos_start_idx if oos_enabled else None,
                bh_ret=bh_ret_for_filter,
                exclude_below_bh=globals().get('EXCLUDE_BELOW_BH', True),
            )

        # ★ 상승/하락률 한도 탐색 — '성공률 풀 선출'만 1~5% 반복(그리드 미실행).
        #   결과를 티커별 캐시 → 단계적 튜닝(pct→wilson→corr)이 재호출해도 '처음 1회'만.
        _mpc = globals().get('_KNET_MULTI_POOL')
        if (_mpc is not None and isinstance(_mpc, tuple) and _mpc[0] == ticker and _mpc[1] is not None):
            print(f"\n  ── 1~5% 통합 다중임계 풀: 캐시 재사용 (재탐색 생략) ──")
        else:
            _build_and_pick_knet_pool(feat, close, indicators=indicators,
                                      n_thresholds=n_thresholds, horizon=horizon, ticker=ticker)
        # 그리드 내부 성공평가용 대표 한도 = 리스트 중앙값 (net>K는 합친 풀 사용 → 그리드-투표 시트에만 영향)
        _reps = sorted(globals().get('STAGE_SUCCESS_LIMIT', [dd_limit]) or [dd_limit])
        dd_limit = ru_limit = _reps[len(_reps) // 2]
        # ★ 그리드 딱 1회 (윌슨/그리드는 여기서만) — 그리드-투표 시트용. net>K는 합친 풀 사용.
        if globals().get('SKIP_GRID_VOTE', False):
            print("  ⚡ SKIP_GRID_VOTE — 그리드-투표 K탐색 최소화(1×1). net>K는 합친 풀로 정확 계산됨.")
            _kb_save, _ks_save = k_buy_range, k_sell_range
            k_buy_range = [k_buy_range[0] if k_buy_range else 2]
            k_sell_range = [k_sell_range[0] if k_sell_range else 2]
            try:
                meta_results_df, inner_all, inner_passed, best_meta, best_inner, buy_pool, sell_pool = _run_meta(dd_limit, dd_limit)
            finally:
                k_buy_range, k_sell_range = _kb_save, _ks_save
        else:
            meta_results_df, inner_all, inner_passed, best_meta, best_inner, buy_pool, sell_pool = _run_meta(dd_limit, dd_limit)

    # ★ staged가 통합 테이블에서 고른 정확한 조합을 강제 (현재 포지션=★1등 일치 보장).
    #   meta_grid_search가 자체 선정한 best_inner와 staged의 최종 선택이 어긋나는 것을 방지.
    #   (inject_pools 재현 모드에서는 위에서 이미 best_inner를 force값으로 설정했으므로 건너뜀)
    if force_best_combo is not None and inject_pools is None:
        fb_kb = int(force_best_combo['K_buy']); fb_vb = int(force_best_combo['vote_buy'])
        fb_ks = int(force_best_combo['K_sell']); fb_vs = int(force_best_combo['vote_sell'])
        src_tbl = inject_combined_table if inject_combined_table is not None else inner_passed
        matched = None
        if src_tbl is not None and len(src_tbl) > 0:
            m = src_tbl[(src_tbl['K_buy']==fb_kb) & (src_tbl['vote_buy']==fb_vb) &
                        (src_tbl['K_sell']==fb_ks) & (src_tbl['vote_sell']==fb_vs)]
            if len(m) > 0:
                matched = m.iloc[0].to_dict()
        if matched is not None:
            best_inner = matched
            print(f"  ★ 최종 조합 강제 적용: K_buy={fb_kb}/v{fb_vb}, K_sell={fb_ks}/v{fb_vs} "
                  f"(현재 포지션이 ★1등과 동일하게 계산됨)")
        else:
            print(f"  ⚠ force_best_combo({fb_kb}/{fb_vb}, {fb_ks}/{fb_vs})를 테이블에서 못 찾음 — meta 자체 best 사용")

    bh_ret  = _bh_sum_return(close_arr)
    bh_up_ret = _bh_up_sum_return(close_arr)
    bh_cagr = (1 + bh_ret) ** (252 / len(close_arr)) - 1 if (1 + bh_ret) > 0 else bh_ret

    # ★ 승률 후보 실거래 검증 (요청) — 그리드는 빠른 근사라 실제 일별거래와 차이날 수 있음.
    #   1) 그리드 '승률' 상위 N개(기본 10000)를 골라 실제 일별 백테스트로 재계산
    #   2) 실제 승률 최고에서 -10%p 범위로 후보를 잡고
    #   3) 그 안에서 실제 수익률 높은 순으로 정렬 → 1등을 최종 선정
    #   (force_best_combo / inject_pools 모드는 이미 조합이 정해졌으므로 건너뜀)
    # ★ 실거래 검증은 '변수 반복마다' 하지 않는다 (요청).
    #   - staged 단계별 호출: write_output=False & inject_combined_table=None → 검증 안 함
    #   - 단독 분석(staged 아님): write_output=True & inject_combined_table=None → 검증 함
    #   - staged 최종 호출: staged가 이미 merged_table에서 실거래 검증을 끝내고
    #     force_best_combo로 그 결과를 넘기므로 여기선 안 함.
    _is_standalone = (write_output and inject_combined_table is None)
    _verify = (globals().get('VERIFY_BY_DAILY_BACKTEST', False)
               and force_best_combo is None and inject_pools is None
               and _is_standalone
               and inner_passed is not None and len(inner_passed) > 0)
    if _verify:
        best_inner, inner_passed = _verify_candidates_by_daily(
            inner_passed, feat, close, buy_pool, sell_pool,
            horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
            stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
            anchor_safe_buy=anchor_safe_buy, anchor_safe_sell=anchor_safe_sell,
            best_inner_fallback=best_inner, bh_ret=bh_ret)

    if inject_pools is None:
        print(f"\n  ─ 메타 그리드 Top 10 ─")
        print(f"  {'#':>2} {'w_z':>5}  {'pct':>8}  {'min_s':>5} {'corr':>5} {'pool':>4}  "
              f"{'평균성공':>8} {'매수':>6} {'매도':>6}  {'수익%':>8}  {'MDD%':>7}  {'거래':>4}")
        for i, r in meta_results_df.head(10).iterrows():
            ret_s = f"{r['best_return']*100:+.2f}" if pd.notna(r['best_return']) else "  —  "
            mdd_s = f"{r['best_mdd']*100:.2f}" if pd.notna(r['best_mdd']) else "  —  "
            nt_s  = f"{int(r['best_n_trades'])}" if pd.notna(r['best_mdd']) else "—"
            avg_s = f"{r['best_avg_sr']*100:.1f}%"  if pd.notna(r['best_avg_sr'])  else "  —  "
            b_s   = f"{r['best_buy_sr']*100:.1f}%"  if pd.notna(r['best_buy_sr'])  else "  —  "
            s_s   = f"{r['best_sell_sr']*100:.1f}%" if pd.notna(r['best_sell_sr']) else "  —  "
            print(f"  {i+1:>2}  {r['wilson_z']:>5}  "
                  f"({int(r['pct_low']):>2},{int(r['pct_high']):>2})  "
                  f"{int(r['min_signals']):>5}  {r['corr_limit']:>5}  "
                  f"{int(r['top_n_pool_buy']):>4}  "
                  f"{avg_s:>8} {b_s:>6} {s_s:>6}  "
                  f"{ret_s:>8}  {mdd_s:>7}  {nt_s:>4}")

        print(f"\n  ★ 1등 메타: WILSON_Z={best_meta['wilson_z']}, "
              f"PCT=({best_meta['pct_low']:.0f},{best_meta['pct_high']:.0f}), "
              f"MIN_SIG={best_meta['min_signals']}, CORR={best_meta['corr_limit']}, "
              f"POOL={best_meta['top_n_pool_buy']}")
        print(f"  ★ 1등 앙상블: K_buy={int(best_inner['K_buy'])}/v={int(best_inner['vote_buy'])}, "
              f"K_sell={int(best_inner['K_sell'])}/v={int(best_inner['vote_sell'])}")
        print(f"     평균성공 {best_inner['avg_success_rate']*100:.1f}% "
              f"(매수 {best_inner['buy_success_rate']*100:.1f}% / 매도 {best_inner['sell_success_rate']*100:.1f}%)")
        has_match = ('anchor_avg_match_rate' in best_inner)
        if has_match:
            match_v = best_inner.get('anchor_avg_match_rate', np.nan) if hasattr(best_inner, 'get') else best_inner['anchor_avg_match_rate']
            if pd.notna(match_v):
                print(f"     ⚓ 매칭률 {match_v*100:.1f}%")
        print(f"     누적 {best_inner['total_return']*100:+.2f}%  vs B&H {bh_ret*100:+.2f}% "
              f"({(best_inner['total_return']-bh_ret)*100:+.2f}%p)")
        print(f"     MDD {best_inner['max_drawdown']*100:.2f}%, "
              f"거래 {int(best_inner['n_trades'])}회, 승률 {best_inner['win_rate']*100:.1f}%")
    else:
        print(f"  ★ 재현 조합: K_buy={int(best_inner['K_buy'])}/v={int(best_inner['vote_buy'])}, "
              f"K_sell={int(best_inner['K_sell'])}/v={int(best_inner['vote_sell'])} "
              f"(엑셀 지표 풀 그대로, 일별 백테스트로 재현)")

    if len(inner_passed) > 0:
        cand = inner_passed.copy()
        if anchor_match_priority and anchor_mode and 'anchor_avg_match_rate' in cand.columns:
            top_m = cand['anchor_avg_match_rate'].max()
            cand = cand[cand['anchor_avg_match_rate'] >= top_m - anchor_match_tolerance]
        if selection_tolerance > 0:
            top_b = cand['avg_success_rate'].max()
            cand_band = cand[cand['avg_success_rate'] >= top_b - selection_tolerance]
        else:
            cand_band = cand
        if len(cand_band) > 1:
            top_ret_in_band = cand_band.sort_values('total_return', ascending=False).head(5)
            print(f"\n  ─ 밴드 안 수익률 상위 5개 (참고용, 모두 동등하게 좋은 후보) ─")
            print(f"  {'BalAcc':>7} {'매칭률':>7} {'수익률':>10} {'MDD':>8} {'K_b/v_b':>9} {'K_s/v_s':>9} {'거래':>4}")
            for _, r in top_ret_in_band.iterrows():
                m_str = f"{r['anchor_avg_match_rate']*100:.1f}%" if 'anchor_avg_match_rate' in r.index else '—'
                is_selected_str = ' ★' if (int(r['K_buy'])==int(best_inner['K_buy']) and
                                            int(r['K_sell'])==int(best_inner['K_sell']) and
                                            int(r['vote_buy'])==int(best_inner['vote_buy']) and
                                            int(r['vote_sell'])==int(best_inner['vote_sell'])) else ''
                print(f"  {r['avg_success_rate']*100:>6.1f}% {m_str:>7} {r['total_return']*100:>+8.2f}% "
                      f"{r['max_drawdown']*100:>+6.2f}% "
                      f"{int(r['K_buy']):>4}/{int(r['vote_buy']):<3} {int(r['K_sell']):>4}/{int(r['vote_sell']):<3} "
                      f"{int(r['n_trades']):>4}{is_selected_str}")
            print(f"  ※ 수익률 비중을 더 두고 싶다면 SELECTION_TOLERANCE를 키우세요 (현재 {selection_tolerance*100:.1f}%p)")

    print(f"\n[일별 백테스트]")
    # ★ 방어 (요청) — 검증 루프에서 쌓인 z-스코어 캐시를 비우고 최종 백테스트를 돌린다.
    #   (키는 이미 내용 기반이라 stale 위험은 없지만, 검증=최종 일치를 한 번 더 보장)
    try: _ZCACHE.clear()
    except Exception: pass
    # ★ 보정 채택된 조합이면 보정된 풀+가중치로 백테스트 → 일별/거래내역/현재포지션 모두 보정 반영 (요청)
    _bt_buy_pool, _bt_sell_pool = buy_pool, sell_pool
    _bt_Kb = int(best_inner['K_buy']); _bt_Ks = int(best_inner['K_sell'])
    _bt_bw = _bt_sw = None
    if force_corr is not None and force_corr.get('buy_pool') is not None:
        _bt_buy_pool = force_corr['buy_pool']; _bt_sell_pool = force_corr['sell_pool']
        _bt_Kb = len(_bt_buy_pool); _bt_Ks = len(_bt_sell_pool)
        _bt_bw = force_corr.get('buy_w'); _bt_sw = force_corr.get('sell_w')
        print(f"  🔧 보정 적용된 풀로 백테스트 — 매수 {_bt_Kb}개(추가 {force_corr.get('n_added_buy',0)}), "
              f"매도 {_bt_Ks}개(추가 {force_corr.get('n_added_sell',0)}) + 가중치 보정")
    daily, trades, cur, buy_used, sell_used = daily_ensemble_backtest(
        feat, close, _bt_buy_pool, _bt_sell_pool,
        K_buy=_bt_Kb,
        K_sell=_bt_Ks,
        vote_buy=int(best_inner['vote_buy']),
        vote_sell=int(best_inner['vote_sell']),
        horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
        stop_loss_pct=stop_loss_pct,
        anchor_mode=anchor_mode,
        anchor_safe_buy=anchor_safe_buy,
        anchor_safe_sell=anchor_safe_sell,
        buy_w_override=_bt_bw, sell_w_override=_bt_sw,
    )

    oos_daily = oos_trades = oos_cur = None
    if oos_enabled and oos_start_idx is not None:
        feat_oos  = feat.iloc[oos_start_idx:]
        close_oos = close.iloc[oos_start_idx:]
        try:
            oos_daily, oos_trades, oos_cur, _bu, _su = daily_ensemble_backtest(
                feat_oos, close_oos, buy_pool, sell_pool,
                K_buy=int(best_inner['K_buy']),
                K_sell=int(best_inner['K_sell']),
                vote_buy=int(best_inner['vote_buy']),
                vote_sell=int(best_inner['vote_sell']),
                horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
                stop_loss_pct=stop_loss_pct,
                anchor_mode=False,
                anchor_safe_buy=None, anchor_safe_sell=None,
            )
            oos_bh = _bh_sum_return(close_oos.values) * 100
            oos_bh_up = _bh_up_sum_return(close_oos.values) * 100
            print()
            print('  ' + '─' * 68)
            print(f'  🔬 OOS 검증 결과 ({close_oos.index[0].date()} ~ {close_oos.index[-1].date()}, {len(close_oos)}일)')
            print(f'     OOS 누적 {oos_cur["cum_return_pct"]:+.2f}%  vs B&H {oos_bh:+.2f}%  '
                  f'거래 {oos_cur["n_trades"]}회  승률 {oos_cur["win_rate"]*100:.1f}%')
            print(f'     📈 상승일만: 전략 {oos_cur["up_cum_return_pct"]:+.2f}%  vs B&H {oos_bh_up:+.2f}%')
            print(f'     ※ IS 누적 {cur["cum_return_pct"]:+.2f}% 대비 OOS가 비슷하면 신뢰↑, 크게 무너지면 과적합 의심')
        except Exception as e:
            print(f'  ⚠ OOS 일별 백테스트 실패: {e}')
            oos_daily = oos_trades = oos_cur = None

    print()
    print('  ' + '─' * 68)
    print(f'  📅  {ticker}  현재 포지션  ({cur["last_date"].date()})')
    print('  ' + '─' * 68)
    print(f'    {cur["position_emoji"]}  {cur["position"]}')
    if '보유' in cur['position']:
        print(f'    진입일: {cur["entry_date"].date()}   진입가: ${cur["entry_price"]:.2f}   '
              f'현재가: ${cur["current_price"]:.2f}')
        print(f'    보유일: {cur["days_held"]}일   미실현: {cur["unrealized_pct"]:+.2f}%')
        if cur.get('stop_price'):
            print(f'    ⛔ 손절가: ${cur["stop_price"]:.2f} '
                  f'(현재가 대비 +{(cur["current_price"]/cur["stop_price"]-1)*100:.2f}% 여유)')
    else:
        print(f'    현재가: ${cur["current_price"]:.2f}')
    print('  ' + '─' * 68)
    print(f'    누적 {cur["cum_return_pct"]:+.2f}%   B&H {bh_ret*100:+.2f}%   '
          f'MDD {cur["max_drawdown"]*100:.2f}%   거래 {cur["n_trades"]}회 (손절매 {cur["n_stop_triggered"]}회)')
    print(f'    📈 상승일만 합산: 전략 {cur["up_cum_return_pct"]:+.2f}%   B&H {bh_up_ret*100:+.2f}%  '
          f'(보유 중 양(+)의 일별 변동률만 합산)')
    # ★ 성공/실패 집계 출력 (first-touch, 정답인 날만 평가)
    print(f'    ✅ [ON만 적중률] 매수 {cur.get("n_buy_success_cnt",0)}/{cur.get("n_buy_on_total",0)} '
          f'({cur.get("buy_signal_hit_rate",0)*100:.1f}%)   '
          f'매도 {cur.get("n_sell_success_cnt",0)}/{cur.get("n_sell_on_total",0)} '
          f'({cur.get("sell_signal_hit_rate",0)*100:.1f}%)')
    print(f'    ✅ [정답일 적중률] 매수 {cur.get("buy_acc_all",0)*100:.1f}% '
          f'({cur.get("n_buy_correct_all",0)}/{cur.get("n_buy_eval_all",0)}일)   '
          f'매도 {cur.get("sell_acc_all",0)*100:.1f}% '
          f'({cur.get("n_sell_correct_all",0)}/{cur.get("n_sell_eval_all",0)}일)  '
          f'※ 올라야/내려야 했던 날만 분모, 놓침=실패 (무임승차 제외)')
    succ_label = ' (⚓ANCHOR 보정)' if anchor_mode else ''
    print(f'    매수신호 BalAcc{succ_label} {cur["buy_success_rate"]*100:.1f}% (plain {cur["buy_accuracy_plain"]*100:.1f}%, ON {cur["n_buy_signal_on"]})  '
          f'매도신호 BalAcc {cur["sell_success_rate"]*100:.1f}% (plain {cur["sell_accuracy_plain"]*100:.1f}%, ON {cur["n_sell_signal_on"]})  '
          f'평균 BalAcc {cur["avg_success_rate"]*100:.1f}%  /  충돌 {cur["n_conflicts"]}일')
    if anchor_mode:
        b_rate = (cur['n_anchor_buy_caught']/cur['n_anchor_buy']*100) if cur['n_anchor_buy']>0 else 0
        s_rate = (cur['n_anchor_sell_caught']/cur['n_anchor_sell']*100) if cur['n_anchor_sell']>0 else 0
        print(f'    ⚓ 정답일 매칭: 매수 {cur["n_anchor_buy_caught"]}/{cur["n_anchor_buy"]} ({b_rate:.1f}%)   '
              f'매도 {cur["n_anchor_sell_caught"]}/{cur["n_anchor_sell"]} ({s_rate:.1f}%)')
    print('  ' + '─' * 68)
    b_str_now = cur['buy_count_now']  / cur['K_buy']  if cur['K_buy']  > 0 else 0
    s_str_now = cur['sell_count_now'] / cur['K_sell'] if cur['K_sell'] > 0 else 0
    print(f'    🟢 매수: {cur["buy_count_now"]:.2f}/{cur["K_buy"]} ON  (필요 {cur["vote_buy"]}, 강도 {b_str_now:.0%})  '
          f'→ {"ON ✓" if cur["buy_on_now"] else "OFF"}')
    print(f'    🔴 매도: {cur["sell_count_now"]:.2f}/{cur["K_sell"]} ON  (필요 {cur["vote_sell"]}, 강도 {s_str_now:.0%})  '
          f'→ {"ON ✓" if cur["sell_on_now"] else "OFF"}')
    if cur['buy_on_now'] and cur['sell_on_now']:
        winner = "매수" if b_str_now >= s_str_now else "매도"
        print(f'    ⚔ 충돌 발생 → 강도 더 큰 [{winner}] 우세')
    print('  ' + '─' * 68)

    if not write_output:
        print('  ⏩ (중간 탐색 — Excel 저장 생략)\n')
        print('=' * 72)
        return (meta_results_df, inner_all, inner_passed,
                best_meta, best_inner, buy_pool, sell_pool,
                daily, trades, cur)

    if output_file is None:
        today_str = datetime.now().strftime('%Y-%m-%d')
        output_file = os.path.join(SCRIPT_DIR, f'ensemble_search_{ticker}_{today_str}.xlsx')
    print(f"\n  Excel 저장: {output_file}")
    # ★ 데이터 스냅샷 저장 (요청) — 재현 정확도용. 재현 때 이 데이터를 그대로 쓰면
    #   FRED 수정·vintage·하루 더 받음 등으로 외부데이터가 달라지는 문제가 사라진다.
    #   (재현 실행(inject_pools 있음)에서는 저장 안 함 — 원본 분석에서만)
    if globals().get('SAVE_DATA_SNAPSHOT', True) and inject_pools is None:
        try:
            _snap = os.path.splitext(output_file)[0] + '_data.pkl'
            _kd = {'feat': feat, 'close': close, 'ticker': ticker}
            # ★ k순신호 재현용: 원본이 고른 합친 풀 + K + 지표수 + 가중g 를 함께 저장 → 재현 때 탐색 없이 그대로.
            try:
                _mp = globals().get('_KNET_MULTI_POOL')
                if _mp and _mp[0] == ticker and _mp[1] is not None:
                    _kd['knet_buy'] = _mp[1]; _kd['knet_sell'] = _mp[2]
                _nf = globals().get('_KNET_FULL'); _no = globals().get('_KNET_OOS')
                if _nf:
                    _kd['knet_k'] = _nf.get('best_k'); _kd['knet_nb'] = _nf.get('n_buy_opt'); _kd['knet_ns'] = _nf.get('n_sell_opt')
                if _no:
                    _kd['knet_k_oos'] = _no.get('best_k'); _kd['knet_nb_oos'] = _no.get('n_buy_opt'); _kd['knet_ns_oos'] = _no.get('n_sell_opt')
                _kd['knet_g'] = globals().get('_KNET_BEST_WEXP', 1.0)
                _kd['knet_g_oos'] = globals().get('_KNET_BEST_WEXP', 1.0)
            except Exception:
                pass
            pd.to_pickle(_kd, _snap)
            print(f"  💾 데이터 스냅샷 저장: {os.path.basename(_snap)} "
                  f"({len(close)}일, 마지막 {str(close.index[-1])[:10]}) — 재현 정확도용"
                  f"{' +k순신호 풀/K' if 'knet_buy' in _kd else ''}")
        except Exception as _se:
            print(f"  ⚠ 데이터 스냅샷 저장 실패(무시): {_se}")
    _grid_table_for_excel = inject_combined_table if inject_combined_table is not None else inner_passed
    # ★ 보증 (요청) — 최종 백테스트 cur로 ★(선정) 행의 실제*/sel_*(/보정후*)를 덮어써서,
    #   그리드 선정 행이 '일별 백테스트 시트와 정확히 같은 수치'가 되게 한다.
    #   (검증 백테스트와 최종 백테스트가 어떤 이유로든 어긋나도, 표시·기록은 최종 기준으로 통일)
    try:
        _gt = _grid_table_for_excel
        if _gt is not None and len(_gt) > 0 and best_inner is not None and \
           all(k in _gt.columns for k in ('K_buy','vote_buy','K_sell','vote_sell')):
            _m = ((_gt['K_buy'].astype(float).round().astype(int) == int(best_inner['K_buy'])) &
                  (_gt['vote_buy'].astype(float).round().astype(int) == int(best_inner['vote_buy'])) &
                  (_gt['K_sell'].astype(float).round().astype(int) == int(best_inner['K_sell'])) &
                  (_gt['vote_sell'].astype(float).round().astype(int) == int(best_inner['vote_sell'])))
            if _m.any():
                _fin = {
                    'real_win_rate':      float(cur.get('win_rate', 0.0)),
                    'real_max_drawdown':  float(cur.get('max_drawdown', 0.0)),
                    'real_total_return':  float(cur.get('cum_return_pct', 0.0)) / 100.0,
                    'real_n_trades':      int(cur.get('n_trades', 0)),
                    'real_buy_success':   float(cur.get('buy_success_rate', 0.0)),
                    'real_sell_success':  float(cur.get('sell_success_rate', 0.0)),
                    'real_avg_success':   float(cur.get('avg_success_rate', 0.0)),
                    'real_buy_match':     float(cur.get('anchor_buy_match_rate', 0.0)),
                    'real_sell_match':    float(cur.get('anchor_sell_match_rate', 0.0)),
                    'real_avg_match':     float(cur.get('anchor_avg_match_rate', 0.0)),
                    'real_anchor_return': float(cur.get('anchor_strategy_return', 0.0)),
                }
                for _c, _v in _fin.items():
                    if _c in _gt.columns:
                        _gt.loc[_m, _c] = _v
                for _sc, _k in [('sel_win_rate', 'real_win_rate'),
                                ('sel_max_drawdown', 'real_max_drawdown'),
                                ('sel_total_return', 'real_total_return')]:
                    if _sc in _gt.columns:
                        _gt.loc[_m, _sc] = _fin[_k]
                _corrected = force_corr is not None and force_corr.get('buy_pool') is not None
                if _corrected:
                    for _cc, _k in [('corr_total_return', 'real_total_return'),
                                    ('corr_max_drawdown', 'real_max_drawdown'),
                                    ('corr_win_rate', 'real_win_rate'),
                                    ('corr_buy_success', 'real_buy_success'),
                                    ('corr_sell_success', 'real_sell_success'),
                                    ('corr_buy_match', 'real_buy_match'),
                                    ('corr_sell_match', 'real_sell_match')]:
                        if _cc in _gt.columns:
                            _gt.loc[_m, _cc] = _fin[_k]
                # 실행일 포지션·액션·최근매매도 최종 일별/거래 기준으로 갱신
                try:
                    for _rk, _rv in _extract_runday_info(daily, trades).items():
                        if _rk in _gt.columns:
                            _gt.loc[_m, _rk] = _rv
                except Exception:
                    pass
                print(f"  🔗 선정(★) 행 수치를 최종 일별 백테스트와 동기화 "
                      f"(승률 {_fin['real_win_rate']*100:.1f}%, 최대손실 {_fin['real_max_drawdown']*100:.2f}%, "
                      f"수익 {_fin['real_total_return']*100:+.2f}%)")
    except Exception as _sync_e:
        print(f"  ⚠ 선정행 최종수치 동기화 실패(무시): {_sync_e}")
    write_excel(meta_results_df, inner_all, _grid_table_for_excel,
                best_meta, best_inner, buy_pool, sell_pool,
                daily, trades, cur, buy_used, sell_used, bh_ret, bh_cagr,
                ticker=ticker, output_file=output_file,
                horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
                mdd_limit_pct=max_drawdown_limit_pct,
                min_trades_daily=min_trades_daily,
                stop_loss_pct=stop_loss_pct,
                selection_tolerance=selection_tolerance,
                anchor_match_priority_arg=anchor_match_priority,
                anchor_match_tolerance_arg=anchor_match_tolerance,
                anchor_mode=anchor_mode,
                anchor_buy_dates=anchor_buy_dates if anchor_mode else None,
                anchor_sell_dates=anchor_sell_dates if anchor_mode else None,
                auto_anchor=auto_anchor,
                oos_enabled=oos_enabled,
                oos_daily=oos_daily, oos_trades=oos_trades, oos_cur=oos_cur,
                bh_up_ret=bh_up_ret, anchor_prio=_anchor_prio,
                feat=feat, close_full=close)
    print('  ✓ 완료\n')
    print('=' * 72)

    import inspect
    caller_frame = inspect.stack()[1]
    caller_name = caller_frame.function if caller_frame else ''
    if caller_name not in ('run_multi_ticker_analysis', 'staged_meta_tune', 'replay_grid_combo') and AUTO_DOWNLOAD_EXCEL:
        _auto_download_excels([output_file])

    return (meta_results_df, inner_all, inner_passed,
            best_meta, best_inner, buy_pool, sell_pool,
            daily, trades, cur)


def staged_meta_tune(*, base_meta_grid=None,
                      stage_pct_range=None,
                      stage_wilson_z=None,
                      stage_wilson_refine_step=None,
                      stage_corr_limit=None,
                      selection_tolerance=None,
                      **run_kwargs):
    """
    ★ 단계적 메타 변수 자동 튜닝 (중복 실행 제거 버전).
    각 단계(pct_range / wilson_z / corr_limit)에서 후보를 실행하되,
    이미 돌린 (wz,pct,corr) 조합은 캐시에서 꺼내 재실행하지 않는다.
    선정 기준: 안정성 종합 점수(매도·매수성공·수익·MDD방어 가중 기하평균) 최대.
    ★ 최종 결과는 '다시 안 돌리고' 단계에서 이미 돌린 best 조합의 결과를 그대로 쓰고,
       Excel만 그 결과로 1회 저장한다 (중복 실행 0).
    """
    if base_meta_grid is None:            base_meta_grid = META_GRID
    if stage_pct_range is None:           stage_pct_range = STAGE_PCT_RANGE
    if stage_wilson_z is None:            stage_wilson_z = STAGE_WILSON_Z
    if stage_wilson_refine_step is None:  stage_wilson_refine_step = STAGE_WILSON_REFINE_STEP
    if stage_corr_limit is None:          stage_corr_limit = STAGE_CORR_LIMIT
    if selection_tolerance is None:       selection_tolerance = SELECTION_TOLERANCE

    base_ms   = base_meta_grid.get('min_signals', [10])[0]
    base_pool = base_meta_grid.get('top_n_pool', [150])[0]
    base_wz   = base_meta_grid.get('wilson_z', [1.7])[0]
    base_pct  = base_meta_grid.get('pct_range', [(10, 90)])[0]
    base_corr = base_meta_grid.get('corr_limit', [0.8])[0]

    output_file = run_kwargs.pop('output_file', None)

    import numpy as _np
    _swt = globals().get('STABILITY_WEIGHTS', (0.30, 0.25, 0.20, 0.25))
    _swt = _np.asarray(_swt, dtype=float); _swt = _swt / _swt.sum()

    def _mk_grid(wz, pct, corr):
        return {'wilson_z': [wz], 'pct_range': [pct], 'min_signals': [base_ms],
                'corr_limit': [corr], 'top_n_pool': [base_pool]}

    # ★ 결과 캐시 — 같은 (wz,pct,corr)는 한 번만 실행
    _cache = {}
    def _run_cached(wz, pct, corr):
        """캐시에 있으면 재사용, 없으면 실행(Excel 저장 안 함).
           반환: dict(...) 또는 None(통과 조합 0개 → 스킵).
           ★ 통과 조합이 없어 meta_grid_search가 에러를 내면, 멈추지 말고
             이 조합만 스킵(None)하고 나머지 단계는 계속 진행한다."""
        key = (round(float(wz), 6), tuple(pct), round(float(corr), 6))
        if key in _cache:
            return _cache[key]
        try:
            res = run_ensemble_search(
                meta_grid=_mk_grid(wz, pct, corr),
                write_output=False, output_file=None,
                **run_kwargs)
        except RuntimeError as _e:
            # MDD 한도/거래수/B&H 미달 등으로 통과 조합이 하나도 없는 경우
            print(f'    ⚠ (wz={wz}, pct={pct}, corr={corr}) 통과 조합 없음 → 이 조합 스킵')
            _cache[key] = None
            return None
        cur = res[-1]
        # res[2] = 이 메타조합의 통합 그리드 테이블(메타변수·B&H필터 포함)
        combined_tbl = res[2] if len(res) > 2 else None
        rec = {
            'sell': float(cur.get('sell_success_rate', 0.0)),
            'buy':  float(cur.get('buy_success_rate', 0.0)),
            'avg':  float(cur.get('avg_success_rate', 0.0)),
            'ret':  float(cur.get('cum_return_pct', float('-inf'))),
            'mdd':  float(cur.get('max_drawdown', -1.0)),
            'win':  float(cur.get('win_rate', 0.0)),
            'res':  res, 'wz': wz, 'pct': pct, 'corr': corr,
            'combined': combined_tbl,
        }
        _cache[key] = rec
        return rec

    _prio_st = globals().get('SELECTION_PRIORITY', 'balacc_return')

    def _pick_best(recs):
        """recs: list of dict (None은 스킵된 조합 → 제외).
           independent: (매수성공+매도성공) 합 최대(동률시 수익).
           그 외(stability): 안정성 종합점수 최대."""
        recs = [r for r in recs if r is not None]   # ★ 스킵된 조합 제외
        if not recs: return None
        if len(recs) == 1: return recs[0]
        tol = globals().get('SELECTION_TOLERANCE', 0.04)
        # ★ 일별거래 승률 밴드(-10%p) → 누적수익 최고 — 요청
        if _prio_st == 'winrate_return':
            wtol = globals().get('WINRATE_TOLERANCE', 0.10)
            best_win = max(r['win'] for r in recs)
            b1 = [r for r in recs if r['win'] >= best_win - wtol]
            b1.sort(key=lambda r: (r['ret'], r['win']), reverse=True)
            return b1[0]
        # ★ 매도성공 밴드(-2%p) → MDD최저 밴드(-1%p) → 누적수익 최고 — 요청, 기본
        if _prio_st == 'sell_mdd_return':
            stol = globals().get('SELL_SUCCESS_TOLERANCE', 0.02)
            mtol = globals().get('MDD_TOLERANCE', 0.01)
            best_sell = max(r['sell'] for r in recs)
            b1 = [r for r in recs if r['sell'] >= best_sell - stol]
            best_mdd = max(r['mdd'] for r in b1)
            b2 = [r for r in b1 if r['mdd'] >= best_mdd - mtol]
            b2.sort(key=lambda r: (r['ret'], r['sell']), reverse=True)
            return b2[0]
        # ★ 평균성공 밴드 → MDD최저(0.1%p 동률시 수익) — 요청
        if _prio_st == 'avgband_mdd_return':
            best_avg = max(r['avg'] for r in recs)
            band = [r for r in recs if r['avg'] >= best_avg - tol]
            best_mdd_r = max(round(r['mdd'], 3) for r in band)   # 덜 빠진 쪽
            grp = [r for r in band if round(r['mdd'], 3) == best_mdd_r]
            grp.sort(key=lambda r: (r['ret'], r['avg']), reverse=True)
            return grp[0]
        # ★ 독립 최적화 — 매수성공+매도성공 합이 최대인 변수조합
        if _prio_st == 'independent':
            best = None; best_key = None
            for r in recs:
                key = (r['buy'] + r['sell'], r['ret'])
                if best_key is None or key > best_key:
                    best_key = key; best = r
            return best
        rets = [r['ret'] for r in recs]; mdds = [r['mdd'] for r in recs]
        ret_lo, ret_hi = min(rets), max(rets)
        mdd_worst, mdd_best = min(mdds), max(mdds)
        def _nrm(x, lo, hi):
            if hi - lo < 1e-12: return 0.5
            return min(1.0, max(0.0, (x - lo) / (hi - lo)))
        eps = 1e-6
        best = None; best_sc = -1.0
        for r in recs:
            s = min(1.0, max(0.0, r['sell'])); b = min(1.0, max(0.0, r['buy']))
            rr = _nrm(r['ret'], ret_lo, ret_hi); m = _nrm(r['mdd'], mdd_worst, mdd_best)
            comps = (max(s,eps), max(b,eps), max(rr,eps), max(m,eps))
            sc = float(_np.prod([c ** wi for c, wi in zip(comps, _swt)]))
            r['_score'] = sc
            if sc > best_sc:
                best_sc = sc; best = r
        return best

    def _fmt(r):
        if r is None:
            return "(통과 조합 없음 — 스킵)"
        return (f"매도 {r['sell']*100:.1f}% / 매수 {r['buy']*100:.1f}% / 평균 {r['avg']*100:.1f}% "
                f"/ 수익 {r['ret']:+.2f}% / MDD {r['mdd']*100:.2f}%")

    print('\n' + '█' * 72)
    print('  ★★★  단계적 메타 변수 자동 튜닝 시작 (중복 실행 없음)  ★★★')
    _ps = globals().get('SELECTION_PRIORITY','')
    if _ps == 'winrate_return':
        _wt=globals().get('WINRATE_TOLERANCE',0.10)
        print(f'  ★ 선정: 일별거래 승률 -{_wt*100:.0f}%p 밴드 → 누적수익 최고')
    elif _ps == 'sell_mdd_return':
        _st=globals().get('SELL_SUCCESS_TOLERANCE',0.02); _mt=globals().get('MDD_TOLERANCE',0.01)
        print(f'  ★ 선정: 매도성공률 -{_st*100:.0f}%p 밴드 → MDD 최저 -{_mt*100:.0f}%p 밴드 → 누적수익 최고')
    elif _ps == 'avgband_mdd_return':
        print(f'  ★ 선정: 평균성공률 top band 내 → MDD 최저(동률시 수익 최대)')
    elif _ps == 'independent':
        print(f'  ★ 선정: 매수성공률 최고 + 매도성공률 최고 설정 따로 찾아 합침 (독립 최적화)')
    else:
        print(f'  ★ 선정: 안정성 종합점수(매도·매수성공·수익·MDD방어) 최대')
    print(f'    1단계 pct_range {len(stage_pct_range)}개 → '
          f'2단계 wilson_z {len(stage_wilson_z)}개(+재확인) → '
          f'3단계 corr_limit {len(stage_corr_limit)}개')
    print('█' * 72)

    # ─── 1단계: pct_range ───
    print(f'\n┌─ [1단계/3] pct_range — 후보 {len(stage_pct_range)}개 '
          f'(wilson_z={base_wz}, corr_limit={base_corr}) ─┐')
    recs1 = []
    for i, pct in enumerate(stage_pct_range, 1):
        print(f'│  ▸ [1단계] {i}/{len(stage_pct_range)} pct={pct} ...')
        r = _run_cached(base_wz, pct, base_corr)
        recs1.append(r)
        print(f'│    완료 — {_fmt(r)}')
    best1 = _pick_best(recs1)
    if best1 is None:
        best_pct = base_pct
        print(f'└─ [1단계] 통과 조합 없음 → 시작 pct_range={base_pct} 유지하고 계속 ─┘')
    else:
        best_pct = best1['pct']
        print(f'└─ [1단계 완료] best pct_range = {best_pct}  ({_fmt(best1)}) ─┘')

    # ─── 2단계: wilson_z (best_pct 고정) ───
    print(f'\n┌─ [2단계/3] wilson_z — 후보 {len(stage_wilson_z)}개 (pct={best_pct}) ─┐')
    recs2 = []
    for i, wz in enumerate(stage_wilson_z, 1):
        dup = ' (캐시 재사용)' if (round(float(wz),6),tuple(best_pct),round(float(base_corr),6)) in _cache else ''
        print(f'│  ▸ [2단계] {i}/{len(stage_wilson_z)} wz={wz}{dup} ...')
        r = _run_cached(wz, best_pct, base_corr)
        recs2.append(r)
        print(f'│    완료 — {_fmt(r)}')
    best2 = _pick_best(recs2)
    if best2 is None:
        best_wz = base_wz
        print(f'└─ [2단계] 통과 조합 없음 → 시작 wilson_z={base_wz} 유지하고 계속 ─┘')
    else:
        best_wz = best2['wz']
        # 재확인 — best_wz - step (새 값일 때만 실행)
        refine_wz = round(best_wz - stage_wilson_refine_step, 4)
        rkey = (round(float(refine_wz),6), tuple(best_pct), round(float(base_corr),6))
        if rkey not in _cache:
            print(f'│  ▸ [2단계 재확인] {best_wz} - {stage_wilson_refine_step} = {refine_wz} ...')
            rr = _run_cached(refine_wz, best_pct, base_corr)
            best2b = _pick_best([best2, rr])
            if best2b is not None and best2b['wz'] == refine_wz:
                print(f'│    재확인값 {refine_wz} 채택 — {_fmt(rr)}')
                best_wz = refine_wz; best2 = best2b
            else:
                print(f'│    기존 {best_wz} 유지')
        print(f'└─ [2단계 완료] best wilson_z = {best_wz}  ({_fmt(best2)}) ─┘')

    # ─── 3단계: corr_limit (best_pct, best_wz 고정) ───
    print(f'\n┌─ [3단계/3] corr_limit — 후보 {len(stage_corr_limit)}개 '
          f'(pct={best_pct}, wz={best_wz}) ─┐')
    recs3 = []
    for i, corr in enumerate(stage_corr_limit, 1):
        dup = ' (캐시 재사용)' if (round(float(best_wz),6),tuple(best_pct),round(float(corr),6)) in _cache else ''
        print(f'│  ▸ [3단계] {i}/{len(stage_corr_limit)} corr={corr}{dup} ...')
        r = _run_cached(best_wz, best_pct, corr)
        recs3.append(r)
        print(f'│    완료 — {_fmt(r)}')
    best3 = _pick_best(recs3)
    if best3 is None:
        best_corr = base_corr
        print(f'└─ [3단계] 통과 조합 없음 → 시작 corr_limit={base_corr} 유지 ─┘')
    else:
        best_corr = best3['corr']
        print(f'└─ [3단계 완료] best corr_limit = {best_corr}  ({_fmt(best3)}) ─┘')

    # ─── 단계에서 돌린 모든 결과를 하나의 통합 테이블로 합침 (None=스킵 제외) ───
    all_recs = [r for r in _cache.values() if r is not None]
    combined_parts = [r['combined'] for r in all_recs
                      if r.get('combined') is not None and len(r['combined']) > 0]
    if combined_parts:
        merged_table = pd.concat(combined_parts, ignore_index=True)
        dedup_keys = [c for c in ['meta_wilson_z','meta_pct_low','meta_pct_high',
                                   'meta_corr_limit','meta_min_signals','meta_top_n_pool',
                                   'K_buy','vote_buy','K_sell','vote_sell']
                      if c in merged_table.columns]
        if dedup_keys:
            merged_table = merged_table.drop_duplicates(subset=dedup_keys).reset_index(drop=True)
        merged_table = _sort_combined_table(merged_table, selection_tolerance)
    else:
        merged_table = None

    # ─── 통과 조합이 하나도 없으면: 에러로 멈추지 말고 안내 후 종료 ───
    if merged_table is None or len(merged_table) == 0:
        print('\n' + '█' * 72)
        print('  ⚠ 모든 메타조합에서 통과 조합이 없습니다 (MDD 한도/거래수/B&H 미달).')
        print('     → 엑셀 생성을 건너뜁니다. 다음을 완화해 다시 시도하세요:')
        print(f'        - MAX_DRAWDOWN_LIMIT_PCT (현재 {globals().get("MAX_DRAWDOWN_LIMIT_PCT")})')
        print(f'        - MIN_TRADES_DAILY (현재 {globals().get("MIN_TRADES_DAILY")})')
        print(f'        - EXCLUDE_BELOW_BH (현재 {globals().get("EXCLUDE_BELOW_BH")}) — B&H 미달 제외 끄기')
        print('█' * 72 + '\n')
        # RuntimeError 대신 None 반환 → 상위(run_multi)에서 다음 티커로 계속
        return None

    # ─── 합친 테이블에서 최종 선정 ───
    # ★ 실거래 검증 (요청) — 변수 반복 중엔 안 하고, 여기서 모인 merged_table로 1회만.
    #   각 후보 행의 메타조합(wz,pct,corr)에 해당하는 지표 풀을 _cache에서 찾아
    #   실제 일별 백테스트 → 실제 MDD 기준 통과분 중 승률밴드→수익 최고 선정.
    _do_verify = (globals().get('VERIFY_BY_DAILY_BACKTEST', False)
                  and len(merged_table) > 0)
    sel = None
    if _do_verify:
        try:
            _feat_v, _close_v, _tk_v = _resolve_data()
            _mask_v = _feat_v.index >= pd.Timestamp(run_kwargs.get('eval_start', EVAL_START))
            _feat_v = _feat_v.loc[_mask_v]; _close_v = _close_v.reindex(_feat_v.index)
            _valid_v = _close_v.notna()
            _feat_v = _feat_v[_valid_v.values]; _close_v = _close_v[_valid_v.values]
            # 메타조합별 풀 맵 (wz,pct_low,pct_high,corr) → (buy_pool, sell_pool)
            pool_map = {}
            for _rec in all_recs:
                _res = _rec.get('res')
                if _res is None or len(_res) < 7: continue
                _key = (round(float(_rec['wz']),4), int(_rec['pct'][0]), int(_rec['pct'][1]),
                        round(float(_rec['corr']),4))
                pool_map[_key] = (_res[5], _res[6])   # buy_pool, sell_pool
            globals()['_LAST_POOL_MAP'] = pool_map   # ★ 엑셀 '메타조합별_지표' 시트 작성용
            _bh_v = _bh_sum_return(_close_v.values.astype(np.float64))
            # ★ 앵커 배열 생성 — 검증 daily 백테스트에서 매칭률·앵커수익을 계산하려면 필수.
            #   (이게 없으면 daily 내부에서 매칭 집계가 통째로 스킵돼 매칭률이 0%로 나옴)
            _asb_v = _ass_v = None
            _amode_v = run_kwargs.get('anchor_mode', ANCHOR_MODE)
            if _amode_v and globals().get('AUTO_ANCHOR', True):
                try:
                    _abd, _asd = auto_compute_anchor_dates(
                        _feat_v.index, _close_v,
                        window=globals().get('AUTO_ANCHOR_WINDOW', 1),
                        lookforward=globals().get('AUTO_ANCHOR_LOOKFORWARD', 1),
                        min_rise_after_buy=globals().get('AUTO_ANCHOR_MIN_RISE', 0.01),
                        min_drop_after_sell=globals().get('AUTO_ANCHOR_MIN_DROP', 0.01),
                        price_tolerance=globals().get('AUTO_ANCHOR_PRICE_TOLERANCE', 0.01),
                        max_dates=globals().get('AUTO_ANCHOR_MAX_DATES', None))
                    _asb_v, _ass_v = _compute_anchor_arrays(_feat_v.index, _abd, _asd)
                    if int(_asb_v.sum()) == 0 and int(_ass_v.sum()) == 0:
                        _asb_v = _ass_v = None
                    else:
                        print(f"  ⚓ 검증용 ANCHOR 매칭 — 매수 {int(_asb_v.sum())}일, 매도 {int(_ass_v.sum())}일")
                except Exception as _ae:
                    print(f"  ⚠ 검증용 ANCHOR 배열 생성 실패 ({_ae}) → 매칭률은 0으로 표시될 수 있음")
                    _asb_v = _ass_v = None
            best_inner_v, verified_tbl = _verify_staged_candidates(
                merged_table, _feat_v, _close_v, pool_map,
                horizon=run_kwargs.get('horizon', HORIZON_DAYS),
                dd_limit=run_kwargs.get('dd_limit', DRAWDOWN_LIMIT_BUY),
                ru_limit=run_kwargs.get('ru_limit', RUNUP_LIMIT_SELL),
                stop_loss_pct=run_kwargs.get('stop_loss_pct', STOP_LOSS_PCT),
                anchor_mode=_amode_v,
                bh_ret=_bh_v, anchor_safe_buy=_asb_v, anchor_safe_sell=_ass_v)
            if best_inner_v is not None:
                merged_table = verified_tbl   # 실제 컬럼 포함 + 실제수익 정렬
                # best_inner_v와 일치하는 merged_table 행을 sel로
                sel = merged_table.iloc[0]
                for _ci in range(len(merged_table)):
                    _rr = merged_table.iloc[_ci]
                    if (int(_rr['K_buy'])==int(best_inner_v['K_buy']) and
                        int(_rr['vote_buy'])==int(best_inner_v['vote_buy']) and
                        int(_rr['K_sell'])==int(best_inner_v['K_sell']) and
                        int(_rr['vote_sell'])==int(best_inner_v['vote_sell'])):
                        sel = _rr; break
        except Exception as _ve:
            print(f"  ⚠ staged 실거래 검증 실패 ({_ve}) → 그리드 근사 기준으로 선정")
            sel = None

    if sel is None:
        sel_idx = _select_with_tolerance(merged_table, selection_tolerance,
                                          primary='avg_success_rate', secondary='combined_return')
        sel = merged_table.loc[sel_idx]
    fb = {'wz': float(sel['meta_wilson_z']),
          'pct': (int(sel['meta_pct_low']), int(sel['meta_pct_high'])),
          'corr': float(sel['meta_corr_limit']),
          'sell': float(sel['sell_success_rate']), 'buy': float(sel['buy_success_rate']),
          'avg': float(sel['avg_success_rate']), 'ret': float(sel['total_return']*100),
          'mdd': float(sel['max_drawdown'])}

    n_skipped = sum(1 for v in _cache.values() if v is None)
    print('\n' + '█' * 72)
    print(f'  ★ 최종 선정 — 단계에서 돌린 {len(combined_parts)}개 메타조합의 '
          f'통합 테이블({len(merged_table)}개 그리드)에서 1등')
    if n_skipped > 0:
        print(f'     (통과 조합 없어 스킵된 메타조합 {n_skipped}개 제외)')
    print(f'    pct_range  = {fb["pct"]}')
    print(f'    wilson_z   = {fb["wz"]}')
    print(f'    corr_limit = {fb["corr"]}')
    print(f'    {_fmt(fb)}')
    print('█' * 72)

    # ★ Excel 저장 — 최종 best 조합으로 1회 실행, 그리드 시트엔 merged_table 주입
    print(f'\n  📊 최종 조합으로 Excel 생성 (그리드 시트=단계 전 결과 통합)...')

    # ★ 선정 조합이 보정 채택됐으면, 보정 가중치를 재계산해 최종 백테스트에 주입 (요청).
    #   → 일별 백테스트·거래내역·현재포지션 시트가 모두 '보정된 결과'로 그려진다.
    _force_corr = None
    if bool(sel.get('corr_applied', False)) and globals().get('USE_ANCHOR_MATCH_CORRECTION', True):
        try:
            _bselC = {'K_buy': int(sel['K_buy']), 'vote_buy': int(sel['vote_buy']),
                      'K_sell': int(sel['K_sell']), 'vote_sell': int(sel['vote_sell'])}
            _keyC = (round(float(fb['wz']),4), int(fb['pct'][0]), int(fb['pct'][1]), round(float(fb['corr']),4))
            _poolsC = globals().get('_LAST_POOL_MAP', {}).get(_keyC)
            _fcC = globals().get('_LAST_FULL_CAND_MAP', {}).get(_keyC, (None, None))
            if _poolsC is not None:
                _bpC, _spC = _poolsC
                _ffC, _ccC, _tkC = _resolve_data()
                _mC = _ffC.index >= pd.Timestamp(run_kwargs.get('eval_start', EVAL_START))
                _ffC = _ffC.loc[_mC]; _ccC = _ccC.reindex(_ffC.index)
                _vmC = _ccC.notna(); _ffC = _ffC[_vmC.values]; _ccC = _ccC[_vmC.values]
                _abdC, _asdC = auto_compute_anchor_dates(
                    _ffC.index, _ccC, window=globals().get('AUTO_ANCHOR_WINDOW',1),
                    lookforward=globals().get('AUTO_ANCHOR_LOOKFORWARD',1),
                    min_rise_after_buy=globals().get('AUTO_ANCHOR_MIN_RISE',0.01),
                    min_drop_after_sell=globals().get('AUTO_ANCHOR_MIN_DROP',0.01),
                    price_tolerance=globals().get('AUTO_ANCHOR_PRICE_TOLERANCE',0.01),
                    max_dates=globals().get('AUTO_ANCHOR_MAX_DATES',None))
                _asbC, _assC = _compute_anchor_arrays(_ffC.index, _abdC, _asdC)
                _augBC, _augSC, _bwC, _swC, _mcC = anchor_match_correct_weights(
                    _ffC, _ccC, _bpC, _spC, _bselC,
                    anchor_mode=run_kwargs.get('anchor_mode', ANCHOR_MODE),
                    anchor_safe_buy=_asbC, anchor_safe_sell=_assC,
                    full_cand_buy=_fcC[0], full_cand_sell=_fcC[1], verbose=False)
                if _mcC.get('adopted', False) and _bwC is not None:
                    _force_corr = {'buy_pool': _augBC, 'sell_pool': _augSC,
                                   'buy_w': _bwC, 'sell_w': _swC,
                                   'n_added_buy': _mcC.get('n_added_buy', 0),
                                   'n_added_sell': _mcC.get('n_added_sell', 0)}
                    print(f"  🔧 선정 조합 보정 재계산 완료 → 최종 시트에 보정 결과 반영 "
                          f"(수익 {_mcC['ret_before']:+.1f}%→{_mcC['ret_after']:+.1f}%, "
                          f"MDD {_mcC['mdd_before']*100:.2f}%→{_mcC['mdd_after']*100:.2f}%)")
        except Exception as _ce:
            print(f"  ⚠ 선정 조합 보정 재계산 실패(기존 결과로 출력): {_ce}")

    try:
        final_res = run_ensemble_search(
            meta_grid=_mk_grid(fb['wz'], fb['pct'], fb['corr']),
            write_output=True, output_file=output_file,
            inject_combined_table=merged_table,
            force_best_combo={
                'K_buy': int(sel['K_buy']), 'vote_buy': int(sel['vote_buy']),
                'K_sell': int(sel['K_sell']), 'vote_sell': int(sel['vote_sell']),
            },
            force_corr=_force_corr,
            **run_kwargs)
    except RuntimeError as _e:
        # 만약 최종 best 조합조차 단독으로는 통과 못 하는 경우(드묾):
        # merged_table은 이미 있으니, 통과 조합 1개를 강제로 쓰도록 안내
        print(f'  ⚠ 최종 조합 단독 실행 실패: {_e}')
        print(f'     merged_table({len(merged_table)}개)은 유효하나 best 조합 재실행이 막힘.')
        print(f'     MAX_DRAWDOWN_LIMIT_PCT을 약간 완화해 재시도하세요.')
        return None

    print('\n' + '█' * 72)
    print(f'  ✅ 단계 튜닝 최종: {_fmt(fb)}')
    print(f'     (pct_range={fb["pct"]}, wilson_z={fb["wz"]}, corr_limit={fb["corr"]})')
    print(f'     ※ 단계 탐색 {len(_cache)}회 (스킵 {n_skipped}) + 최종 Excel 1회')
    print(f'     ※ 엑셀의 내부_그리드_통과 시트에 단계 전 결과가 모두 담김')
    print('█' * 72 + '\n')

    return final_res


# ════════════════════════════════════════════════════════════════════════
#   ★ 그리드 번호로 특정 조합 재현 → 일별 거래 Excel 생성
#     경로는 OUTPUT_DIR 고정, 파일명 + 그리드 번호만 입력.
#     '사용된 설정' 시트의 설정값(Horizon/손실한도/비용/메타변수 등)을 그대로 사용.
# ════════════════════════════════════════════════════════════════════════
def _parse_used_settings(wb):
    """'사용된 설정' 시트에서 백테스트/메타 설정값을 읽어 dict로 반환."""
    out = {}
    if '사용된 설정' not in wb.sheetnames:
        return out
    ws = wb['사용된 설정']
    kv = {}
    for r in range(1, ws.max_row + 1):
        k = ws.cell(r, 1).value; v = ws.cell(r, 2).value
        if k is None or v is None: continue
        kv[str(k).strip()] = str(v).strip()

    def _num(s):
        """문자열에서 첫 번째 숫자(부호·소수점 포함)만 추출."""
        if s is None: return None
        import re as _re
        m = _re.search(r'[-+]?\d+(?:\.\d+)?', str(s))
        if not m: return None
        try: return float(m.group(0))
        except Exception: return None

    for k, v in kv.items():
        kl = k.replace(' ', '')
        if 'Horizon' in k:
            n = _num(v);  out['horizon'] = int(n) if n is not None else None
        elif '매수손실한도' in kl:
            n = _num(v);  out['dd_limit'] = abs(n)/100.0 if n is not None else None
        elif '매도상승한도' in kl:
            n = _num(v);  out['ru_limit'] = abs(n)/100.0 if n is not None else None
        elif '거래비용' in kl:
            n = _num(v);  out['cost'] = n/100.0 if n is not None else None
        elif 'WILSON_Z' in k:
            out['wilson_z'] = _num(v)
        elif 'PCT' in k and '범위' in k:
            s = v.replace('(', '').replace(')', '').replace('[', '').replace(']', '')
            try:
                ps = [int(float(x.strip())) for x in s.split(',') if x.strip()]
                if len(ps) == 2: out['pct_range'] = tuple(ps)
            except Exception: pass
        elif 'MIN_SIGNALS' in k:
            n = _num(v);  out['min_signals'] = int(n) if n is not None else None
        elif 'DIVERSITY_CORR_LIMIT' in k or 'CORR_LIMIT' in k:
            out['corr_limit'] = _num(v)
        elif 'TOP_N_POOL_BUY' in k:
            n = _num(v);  out['top_n_pool'] = int(n) if n is not None else None
        elif 'MDD한도' in kl:
            n = _num(v);  out['mdd_limit_pct'] = abs(n) if n is not None else None
        elif '최소거래수' in kl:
            n = _num(v);  out['min_trades_daily'] = int(n) if n is not None else None
        elif '손절매한도' in kl:
            n = _num(v);  out['stop_loss_pct'] = abs(n)/100.0 if n is not None else None
        elif kl == 'K_buy':
            n = _num(v);  out['K_buy'] = int(n) if n is not None else None
        elif kl == 'vote_buy':
            n = _num(v);  out['vote_buy'] = int(n) if n is not None else None
        elif kl == 'K_sell':
            n = _num(v);  out['K_sell'] = int(n) if n is not None else None
        elif kl == 'vote_sell':
            n = _num(v);  out['vote_sell'] = int(n) if n is not None else None
        elif k.strip() == '티커' or kl == '티커':
            out['ticker'] = str(v).strip()
    return out


def replay_grid_combo(filename, grid_number=None, *,
                       feat=None, close=None,
                       output_dir=None,
                       **override_kwargs):
    """
    OUTPUT_DIR 안의 결과 Excel(filename)에서 grid_number 조합을 재현해
    일별 거래 Excel을 생성한다. 경로는 OUTPUT_DIR 고정 — 파일명과 번호만 입력.

    '사용된 설정' 시트의 설정값(Horizon, 손실/상승 한도, 거래비용, 메타변수,
    MDD 한도, 최소거래수, 손절매)을 그대로 읽어 그 조건으로 재현한다.

    Parameters
    ----------
    filename : str       OUTPUT_DIR 안의 결과 Excel 파일명 (예: 'ensemble_search_VRT_2026-06-01.xlsx')
                         전체 경로를 줘도 됨.
    grid_number : int|str  재현할 그리드 번호 ('14', '17●' 등 — 숫자만 인식)
    feat, close : 선택    원본 지표/종가. None이면 글로벌(_pair_feat/_pair_close)에서 탐색.
    output_dir : 선택     기본 OUTPUT_DIR. 다른 폴더 쓰려면 지정.
    override_kwargs       읽은 설정을 덮어쓸 인자.
    """
    from openpyxl import load_workbook

    # ★ 재현 대상이 ★최적 행이면 '보정 적용 풀(앙상블 시트)'을 쓰고, 아니면 base 메타풀 사용.
    #   자동재현(grid_number=None)은 항상 ★최적이므로 True.
    _is_best_row = (grid_number is None)

    base_dir = output_dir if output_dir is not None else OUTPUT_DIR
    # filename이 전체 경로면 그대로, 아니면 OUTPUT_DIR과 결합
    if os.path.isabs(filename) or os.path.dirname(filename):
        excel_path = filename
    else:
        excel_path = os.path.join(base_dir, filename)
    if not os.path.exists(excel_path):
        raise RuntimeError(f"파일을 찾을 수 없습니다: {excel_path}\n"
                           f"  OUTPUT_DIR={base_dir} 안에 파일명이 맞는지 확인하세요.")

    wb = load_workbook(excel_path, data_only=True)

    # ★ 사용된 설정 (백테스트 조건 + ★최적 K/vote)
    used = _parse_used_settings(wb)

    # ★ 원본 엑셀의 마지막 백테스트 날짜 — 재현 시 '같은 범위'로 잘라 정확 재현 (요청)
    _orig_last_date = None
    if globals().get('REPLAY_MATCH_ORIGINAL_RANGE', True) and '일별 백테스트' in wb.sheetnames:
        try:
            _wsd = wb['일별 백테스트']
            for _r in range(_wsd.max_row, 4, -1):
                _dv = _wsd.cell(_r, 1).value
                if _dv is not None:
                    _orig_last_date = pd.Timestamp(_dv); break
            if _orig_last_date is not None:
                print(f"  📅 원본 마지막 날짜 = {str(_orig_last_date)[:10]} → 재현도 이 날짜까지(같은 범위)")
        except Exception:
            _orig_last_date = None

    # ─── grid_number 가 None 이면: '내부_그리드_통과' 시트의 ★표시 행을 읽어 재현 ───
    #   (요청) ★를 어느 행에 넣든 그 행으로 재현. ★가 없고 데이터가 한 줄뿐이면 그 행으로.
    #   이 경우 '사용된 설정' 시트로는 절대 재현하지 않는다.
    if grid_number is None:
        if '내부_그리드_통과' not in wb.sheetnames:
            raise RuntimeError("'내부_그리드_통과' 시트가 없습니다 — 자동 재현 불가.")
        ws = wb['내부_그리드_통과']
        hdr = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(3, c).value
            if v is not None: hdr[str(v).strip()] = c
        for k in ['#', 'K_buy', 'vote_buy', 'K_sell', 'vote_sell']:
            if k not in hdr:
                raise RuntimeError(f"'내부_그리드_통과'에 '{k}' 컬럼이 없습니다. (헤더: {list(hdr)})")
        data_rows = []; star_rows = []
        for r in range(4, ws.max_row + 1):
            cell = ws.cell(r, hdr['#']).value
            if cell is None: continue
            if ws.cell(r, hdr['K_buy']).value in (None, ''): continue   # 데이터 행만
            data_rows.append(r)
            if '★' in str(cell): star_rows.append(r)
        # 자동 ★(최적행)은 # 셀이 '굵은 빨강(C00000)'으로 강조됨 → 자동/수동 구분해 '수동 ★' 우선
        def _is_auto_star(r):
            try:
                fnt = ws.cell(r, hdr['#']).font
                col = (fnt.color.rgb if (fnt and fnt.color) else '') or ''
                return bool(fnt and fnt.bold) and str(col).endswith('C00000')
            except Exception:
                return False
        manual_stars = [r for r in star_rows if not _is_auto_star(r)]
        if len(manual_stars) == 1:
            target = manual_stars[0]
            print(f"  ♻ 자동 재현 — 사용자가 찍은 ★행 사용 "
                  f"(#{str(ws.cell(target, hdr['#']).value).strip()})")
        elif len(manual_stars) > 1:
            _nums = [str(ws.cell(r, hdr['#']).value).strip() for r in manual_stars]
            raise RuntimeError(
                f"사용자가 찍은 ★가 {len(manual_stars)}개입니다: {', '.join(_nums)}\n"
                f"  재현할 행 하나에만 ★를 남기고 나머지는 지운 뒤 다시 실행하세요.")
        elif len(star_rows) == 1:
            target = star_rows[0]
            print(f"  ♻ 자동 재현 — ★표시(최적)행 사용 "
                  f"(#{str(ws.cell(target, hdr['#']).value).strip()})")
        elif len(star_rows) > 1:
            _nums = [str(ws.cell(r, hdr['#']).value).strip() for r in star_rows]
            raise RuntimeError(
                f"'내부_그리드_통과'에 ★가 {len(star_rows)}개인데 자동/수동 구분이 안 됩니다: {', '.join(_nums)}\n"
                f"  재현할 행 하나만 ★로 남기고 나머지는 지운 뒤 다시 실행하세요.")
        elif len(data_rows) == 1:
            target = data_rows[0]
            print(f"  ♻ 자동 재현 — ★ 없음, 데이터가 한 줄뿐 → 그 행 사용 "
                  f"(#{str(ws.cell(target, hdr['#']).value).strip()})")
        else:
            raise RuntimeError(
                f"'내부_그리드_통과'에 ★표시 행이 없고 데이터가 {len(data_rows)}줄입니다.\n"
                f"  ★를 원하는 행의 # 옆에 넣거나, 필터링으로 한 줄만 남겨 주세요.\n"
                f"  (요청대로 '사용된 설정' 시트로는 재현하지 않습니다.)")
        # ★ 보정이 '채택'된 행이면 보정 앙상블 풀(매수/매도_앙상블_지표)을, 아니면 base 메타풀
        #   ('메타조합별_지표'의 그 행 메타조합 풀)을 K로 잘라 그리드 평가를 그대로 재현.
        #   (#5211처럼 보정채택='—'이면 그리드에서 쓴 그 메타풀+K+vote 그대로여야 일치)
        _corr_col = hdr.get('🔧보정채택')
        _corr_val = ws.cell(target, _corr_col).value if _corr_col else None
        _corr_str = str(_corr_val).strip() if _corr_val is not None else ''
        _is_best_row = _corr_str not in ('', '—', '–', '-', 'None', 'nan')
        K_buy    = int(float(ws.cell(target, hdr['K_buy']).value))
        vote_buy = int(float(ws.cell(target, hdr['vote_buy']).value))
        K_sell   = int(float(ws.cell(target, hdr['K_sell']).value))
        vote_sell = int(float(ws.cell(target, hdr['vote_sell']).value))
        # ★ 보정 채택 행: 그리드 K_buy/vote = '보정 전' base 조합. 보정후(예: 346%)를 재현하려면
        #   보정 단계(지표 추가+재가중)를 다시 적용해야 한다. (가중치 저장 시 _has_w 경로가 처리)
        print(f"     (K_buy={K_buy}/v{vote_buy}, K_sell={K_sell}/v{vote_sell}, "
              f"보정채택='{_corr_str or '—'}'"
              f"{' → 보정후 재현은 저장된 가중치 필요' if _is_best_row else ' → 보정 없음(그리드 그대로)'})")
    else:
        gn_clean = ''.join(ch for ch in str(grid_number) if ch.isdigit())
        if not gn_clean:
            raise ValueError(f"그리드 번호를 해석할 수 없습니다: {grid_number!r}")
        gn = int(gn_clean)

        if '내부_그리드_통과' not in wb.sheetnames:
            raise RuntimeError("'내부_그리드_통과' 시트가 없습니다.")
        ws = wb['내부_그리드_통과']

        hdr = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(3, c).value
            if v is not None: hdr[str(v).strip()] = c
        for k in ['#', 'K_buy', 'vote_buy', 'K_sell', 'vote_sell']:
            if k not in hdr:
                raise RuntimeError(f"'내부_그리드_통과'에 '{k}' 컬럼이 없습니다. (헤더: {list(hdr)})")

        target = None
        for r in range(4, ws.max_row + 1):
            cell = ws.cell(r, hdr['#']).value
            if cell is None: continue
            digits = ''.join(ch for ch in str(cell) if ch.isdigit())
            if digits and int(digits) == gn:
                target = r; _is_best_row = ('★' in str(cell)); break
        if target is None:
            raise RuntimeError(f"#{gn} 번호를 찾지 못했습니다.")

        K_buy   = int(float(ws.cell(target, hdr['K_buy']).value))
        vote_buy = int(float(ws.cell(target, hdr['vote_buy']).value))
        K_sell  = int(float(ws.cell(target, hdr['K_sell']).value))
        vote_sell = int(float(ws.cell(target, hdr['vote_sell']).value))

    # ★ 메타변수 — 그리드 행에 직접 기록된 값을 우선 사용 (행마다 다를 수 있음).
    #   자동 재현 모드(target=None)면 _cellf는 None을 돌려주고 used 값으로 폴백.
    def _cellf(colname):
        if target is not None and colname in hdr:
            v = ws.cell(target, hdr[colname]).value
            try: return float(str(v).replace('%','').strip())
            except Exception: return None
        return None
    row_wz   = _cellf('wilson_z')
    row_pl   = _cellf('pct_low')
    row_ph   = _cellf('pct_high')
    row_cl   = _cellf('corr_limit')
    row_ms   = _cellf('min_sig')
    row_pool = _cellf('pool')

    mg = dict(META_GRID)
    # 행값 우선, 없으면 사용된 설정으로 폴백
    wz_use   = row_wz   if row_wz   is not None else used.get('wilson_z')
    cl_use   = row_cl   if row_cl   is not None else used.get('corr_limit')
    ms_use   = row_ms   if row_ms   is not None else used.get('min_signals')
    pool_use = row_pool if row_pool is not None else used.get('top_n_pool')
    if row_pl is not None and row_ph is not None:
        pct_use = (int(row_pl), int(row_ph))
    else:
        pct_use = used.get('pct_range')
    if wz_use   is not None: mg['wilson_z']   = [wz_use]
    if pct_use  is not None: mg['pct_range']  = [pct_use]
    if cl_use   is not None: mg['corr_limit'] = [cl_use]
    if ms_use   is not None: mg['min_signals']= [int(ms_use)]
    if pool_use is not None: mg['top_n_pool'] = [int(pool_use)]
    _gn_label = f"#{gn} " if grid_number is not None else "★최적 "
    print(f"    ({_gn_label}메타변수: wilson_z={wz_use}, pct={pct_use}, corr={cl_use}, min_sig={ms_use}, pool={pool_use})")

    g = globals()
    if feat is None:
        feat = g.get('_pair_feat')
        if feat is None: feat = g.get('feat')
    if close is None:
        close = g.get('_pair_close')
        if close is None: close = g.get('close')
    # ★ 메모리에 없으면 티커로 데이터 자동 확보 (다운로드 함수/yfinance 등 사용)
    # ★ 데이터 스냅샷 우선 (요청) — 원본 분석이 남긴 '..._data.pkl'이 같은 폴더에 있으면
    #   그걸 써서 '완전히 동일한 데이터'로 정확 재현 (재다운로드 안 함 → FRED수정/vintage 차단).
    #   스냅샷이 없거나 REPLAY_USE_SNAPSHOT=False면 기존처럼 재다운로드(연장 가능, 미세차이 가능).
    if (feat is None or close is None) and globals().get('REPLAY_USE_SNAPSHOT', True):
        _snap = os.path.splitext(excel_path)[0] + '_data.pkl'
        if os.path.exists(_snap):
            try:
                _dsnap = pd.read_pickle(_snap)
                feat = _dsnap.get('feat'); close = _dsnap.get('close')
                if feat is not None and close is not None and len(close) > 0:
                    print(f"  💾 데이터 스냅샷 사용: {os.path.basename(_snap)} "
                          f"({len(close)}일, 마지막 {str(close.index[-1])[:10]}) — 재다운로드 없이 정확 재현")
                    # ★ 원본 k순신호 풀·K·지표수·g 로드 → 재현 때 (wilson×corr) 탐색 안 하고 그대로 사용
                    if _dsnap.get('knet_buy') is not None and _dsnap.get('knet_sell') is not None:
                        _tk = _dsnap.get('ticker')
                        globals()['_KNET_MULTI_POOL'] = (_tk, _dsnap['knet_buy'], _dsnap['knet_sell'])
                        globals()['_KNET_REPLAY_FIXED'] = {
                            'k_full': _dsnap.get('knet_k'), 'nb_full': _dsnap.get('knet_nb'),
                            'ns_full': _dsnap.get('knet_ns'), 'g_full': _dsnap.get('knet_g', 1.0),
                            'k_oos': _dsnap.get('knet_k_oos'), 'nb_oos': _dsnap.get('knet_nb_oos'),
                            'ns_oos': _dsnap.get('knet_ns_oos'), 'g_oos': _dsnap.get('knet_g_oos', 1.0)}
                        print(f"  ♻ 원본 k순신호 풀·K 로드 — 재현 시 탐색 없이 그대로 사용 "
                              f"(K={_dsnap.get('knet_k')}, 매수{_dsnap['knet_buy']['indicator'].nunique()}"
                              f"/매도{_dsnap['knet_sell']['indicator'].nunique()}지표)")
                else:
                    feat = close = None
            except Exception as _le:
                print(f"  ⚠ 스냅샷 로드 실패({_le}) — 재다운로드로 진행")
                feat = close = None
        else:
            print(f"  ℹ 데이터 스냅샷 없음({os.path.basename(_snap)}) — 재다운로드로 진행 "
                  f"(외부데이터 차이로 미세하게 다를 수 있음)")

    if feat is None or close is None:
        ticker = used.get('ticker')
        if ticker is None:
            # 파일명에서 티커 추출 시도 (ensemble_search_VRT_... → VRT)
            import re as _re
            mtk = _re.search(r'ensemble_search_([A-Za-z0-9.\-]+)_', os.path.basename(excel_path))
            if mtk: ticker = mtk.group(1)
        if ticker is not None and '_resolve_data_for_ticker' in g and callable(g['_resolve_data_for_ticker']):
            print(f"  ℹ 메모리에 데이터 없음 → 티커 '{ticker}'로 자동 로드 시도...")
            try:
                feat, close = g['_resolve_data_for_ticker'](ticker, end_date=_orig_last_date)
                print(f"  ✓ '{ticker}' 데이터 로드 성공 ({len(close)}일)")
            except Exception as _e:
                raise RuntimeError(
                    f"티커 '{ticker}' 데이터 자동 로드 실패: {_e}\n"
                    f"  download_data/compute_features 함수가 정의돼 있거나 yfinance가 설치돼 있어야 합니다.\n"
                    f"  또는 replay_grid_combo(..., feat=내_feat, close=내_close)로 직접 전달하세요.")
        else:
            raise RuntimeError(
                "원본 데이터(feat, close)를 찾을 수 없고 자동 로드도 불가합니다.\n"
                f"  티커={ticker}, 다운로드 함수 존재={'_resolve_data_for_ticker' in g}\n"
                "  feat=, close= 로 직접 전달하거나 download_data/compute_features를 준비하세요.")

    base = os.path.splitext(os.path.basename(excel_path))[0]
    if grid_number is None:
        output_file = os.path.join(base_dir, f"{base}__replay_best.xlsx")
    else:
        output_file = os.path.join(base_dir, f"{base}__replay_grid{gn}.xlsx")

    print('═' * 72)
    if grid_number is None:
        print(f'  ★ 최적 조합 자동 재현 → 일별 거래 Excel')
    else:
        print(f'  ★ 그리드 #{gn} 재현 → 일별 거래 Excel')
    print(f'    파일: {excel_path}')
    print(f'    티커: {used.get("ticker", "(파일명에서 추출)")}')
    print(f'    조합: K_buy={K_buy}/vote={vote_buy}, K_sell={K_sell}/vote={vote_sell}')
    print(f'    설정(사용된 설정 시트 그대로):')
    print(f'      Horizon={used.get("horizon")}, dd_limit={used.get("dd_limit")}, '
          f'ru_limit={used.get("ru_limit")}, cost={used.get("cost")}')
    print(f'      WILSON_Z={used.get("wilson_z")}, PCT={used.get("pct_range")}, '
          f'CORR={used.get("corr_limit")}, MIN_SIG={used.get("min_signals")}')
    print(f'      MDD한도={used.get("mdd_limit_pct")}, 최소거래={used.get("min_trades_daily")}, '
          f'손절매={used.get("stop_loss_pct")}')
    print(f'    출력: {output_file}')
    print('═' * 72)

    kwargs = dict(
        meta_grid=mg,
        k_buy_range=[K_buy], k_sell_range=[K_sell],
        vote_ratio_buy=[vote_buy / K_buy] if K_buy > 0 else [1.0],
        vote_ratio_sell=[vote_sell / K_sell] if K_sell > 0 else [1.0],
        write_output=True, output_file=output_file,
        # ★ 이 그리드의 정확한 조합을 강제 → 현재 포지션 = 재현 조합 일치
        force_best_combo={'K_buy': K_buy, 'vote_buy': vote_buy,
                          'K_sell': K_sell, 'vote_sell': vote_sell},
        # inject_pools는 아래에서 엑셀 지표 풀을 구성한 뒤 추가함
        # ★ 재현은 '이미 선정된 그 조합 하나'를 그대로 다시 계산하는 것.
        #   선정용 필터(MDD 한도·최소거래수)를 끄지 않으면, 재현 결과가 원본과 미세하게
        #   달라(데이터 갱신 등) 그 조합이 필터에 걸려 '통과 0개'로 재현 실패함.
        max_drawdown_limit_pct=None,   # MDD 한도 무력화 (그 조합 그대로 봐야 함)
        min_trades_daily=1,            # 최소거래수 무력화
    )
    # 읽은 설정 반영 (run_ensemble_search 인자명에 맞춰).
    #   ※ cost는 전역 COST_PER_TRADE로 따로 처리.
    #   ※ mdd_limit / min_trades는 위에서 의도적으로 껐으므로 여기서 덮어쓰지 않음(목록 제외).
    for src_k, dst_k in [('horizon','horizon'), ('dd_limit','dd_limit'),
                          ('ru_limit','ru_limit'),
                          ('stop_loss_pct','stop_loss_pct')]:
        if used.get(src_k) is not None:
            kwargs[dst_k] = used[src_k]
    kwargs.update(override_kwargs)

    g['_pair_feat'] = feat
    g['_pair_close'] = close

    # ★ 엑셀의 매수/매도 앙상블 지표 시트를 읽어 '그 지표 그대로' 풀 구성 → 재현 정확도 보장.
    #   (메타변수로 지표를 다시 선별하지 않고, 원본이 쓴 지표를 그대로 사용)
    def _read_pool(sheet_name):
        if sheet_name not in wb.sheetnames:
            return None
        wsp = wb[sheet_name]
        # 헤더(3행): #, 지표, 방향, 임계치, 분위, 신호수, 성공수, 성공률, 점수
        h = {}
        for c in range(1, wsp.max_column + 1):
            v = wsp.cell(3, c).value
            if v is not None: h[str(v).strip()] = c
        need = ['지표', '방향', '임계치']
        if not all(k in h for k in need):
            return None
        rows = []
        for r in range(4, wsp.max_row + 1):
            ind = wsp.cell(r, h['지표']).value
            if ind is None or str(ind).strip() == '':
                continue
            def _gv(col, default=None):
                if col in h:
                    return wsp.cell(r, h[col]).value
                return default
            def _gf(col, default=0.0):
                v = _gv(col)
                try: return float(str(v).replace('%','').strip())
                except Exception: return default
            n_sig = _gf('신호수', 0.0)
            n_suc = _gf('성공수', 0.0)
            sr = _gf('성공률', 0.0)
            if sr > 1.5: sr = sr / 100.0   # '96.00%' → 0.96
            _vw = _gv('🔁가중치')
            try: _vwf = float(str(_vw).replace('%','').strip())
            except Exception: _vwf = np.nan
            rows.append({
                'indicator': str(ind).strip(),
                'direction': str(_gv('방향', '>=')).strip(),
                'threshold': _gf('임계치', 0.0),
                'pct_label': _gf('분위', 50.0),
                'n_signals': n_sig,
                'n_success': n_suc,
                'success_rate': sr,
                'avg_extreme': 0.01,
                'score': _gf('점수', sr),
                'vote_weight': _vwf,   # ★ 재현용 저장 가중치 (없으면 NaN)
                'lead_shift': int(_gf('지연(일)', 0.0)),   # ★ 지연 재현 (구버전 파일=0)
            })
        if not rows:
            return None
        return pd.DataFrame(rows)

    def _read_meta_pool(wz, pl, ph, corr):
        """메타조합별_지표 시트에서 (wz,pct,corr)에 맞는 매수/매도 풀을 읽는다.
           그리드 번호 재현 시, 그 번호의 메타조합 지표를 '그대로' 쓰기 위함."""
        sn = '메타조합별_지표'
        if sn not in wb.sheetnames:
            return None, None
        wsp = wb[sn]
        # 블록 헤더 '★ wilson_z=... / pct=(..) / corr=..' 를 찾아 매칭
        target_hdr = None
        rows_by_block = {}
        cur_block = None; cur_side = None; cols = None
        bp_rows = []; sp_rows = []
        def _close(s):
            try: return float(str(s).replace('%','').strip())
            except Exception: return 0.0
        import re as _re
        for r in range(1, wsp.max_row + 1):
            a = wsp.cell(r, 1).value
            if a is None: continue
            s = str(a)
            if s.startswith('★ wilson_z='):
                m = _re.search(r'wilson_z=([\d.]+).*pct=\((\d+),(\d+)\).*corr=([\d.]+)', s)
                if m:
                    bwz, bpl, bph, bcorr = float(m.group(1)), int(m.group(2)), int(m.group(3)), float(m.group(4))
                    cur_block = (round(bwz,4)==round(float(wz),4) and bpl==int(pl)
                                 and bph==int(ph) and round(bcorr,4)==round(float(corr),4))
                cur_side = None
                continue
            if not cur_block:
                continue
            if s.strip().startswith('[매수]'): cur_side = 'buy'; continue
            if s.strip().startswith('[매도]'): cur_side = 'sell'; continue
            if s == 'side':   # 헤더행
                continue
            # 데이터행: side, 지표, 방향, 임계치, 분위, 신호수, 성공수, 성공률, 점수
            side = s.strip()
            ind = wsp.cell(r, 2).value
            if ind is None or str(ind).strip() == '': continue
            rec = {'indicator': str(ind).strip(),
                   'direction': str(wsp.cell(r,3).value or '>=').strip(),
                   'threshold': _close(wsp.cell(r,4).value),
                   'pct_label': _close(wsp.cell(r,5).value),
                   'n_signals': _close(wsp.cell(r,6).value),
                   'n_success': _close(wsp.cell(r,7).value),
                   'success_rate': _close(wsp.cell(r,8).value),
                   'avg_extreme': 0.01,
                   'score': _close(wsp.cell(r,9).value),
                   'lead_shift': int(_close(wsp.cell(r,10).value))}   # ★ 지연 재현 (구버전=0)
            if side == '매수': bp_rows.append(rec)
            elif side == '매도': sp_rows.append(rec)
        bp = pd.DataFrame(bp_rows) if bp_rows else None
        sp = pd.DataFrame(sp_rows) if sp_rows else None
        return bp, sp

    # ★ 그리드 번호 재현이면, 그 번호의 메타조합 풀을 '메타조합별_지표' 시트에서 우선 읽음.
    #   (없으면 ★최적의 매수/매도_앙상블_지표 시트로 폴백)
    #   단, ★최적 행 재현이면 보정 적용된 '앙상블 시트(가중치 포함)'를 우선 사용해 보정까지 정확 재현.
    buy_pool_xl = sell_pool_xl = None
    if not _is_best_row:
        _mb, _ms = _read_meta_pool(wz_use, pct_use[0] if pct_use else 5,
                                   pct_use[1] if pct_use else 95, cl_use if cl_use is not None else 0.2)
        if _mb is not None and _ms is not None and len(_mb) > 0 and len(_ms) > 0:
            buy_pool_xl, sell_pool_xl = _mb, _ms
            _gnlbl = f"#{grid_number}" if grid_number is not None else "★행"
            print(f"  ♻ {_gnlbl}(보정 미채택)의 메타조합 지표를 '메타조합별_지표' 시트에서 읽음 "
                  f"(매수 {len(_mb)} / 매도 {len(_ms)}) — base 풀, K로 잘라 그리드 그대로 재현")
    if buy_pool_xl is None or sell_pool_xl is None:
        # ★최적 행 또는 자동재현 → 앙상블 시트(보정 적용 풀 + 🔁가중치)로 정확 재현
        buy_pool_xl  = _read_pool('매수_앙상블_지표')
        sell_pool_xl = _read_pool('매도_앙상블_지표')
        if _is_best_row and buy_pool_xl is not None:
            print(f"  ♻ ★최적 조합 재현 — 앙상블 시트(보정 적용 풀 + 가중치) 사용")

    inject_pools = None
    if buy_pool_xl is not None and sell_pool_xl is not None \
       and len(buy_pool_xl) > 0 and len(sell_pool_xl) > 0:
        # 엑셀 지표 중 현재 feat에 존재하는 것만 사용 (데이터 갱신으로 사라진 지표 방어)
        bcols = set(feat.columns)
        bp_use = buy_pool_xl[buy_pool_xl['indicator'].isin(bcols)].reset_index(drop=True)
        sp_use = sell_pool_xl[sell_pool_xl['indicator'].isin(bcols)].reset_index(drop=True)
        miss_b = len(buy_pool_xl) - len(bp_use)
        miss_s = len(sell_pool_xl) - len(sp_use)
        if len(bp_use) >= K_buy and len(sp_use) >= K_sell:
            inject_pools = (bp_use, sp_use)
            print(f"  ♻ 엑셀 지표 그대로 사용 — 매수풀 {len(bp_use)}개, 매도풀 {len(sp_use)}개")
            if miss_b or miss_s:
                print(f"     (현재 데이터에 없는 지표 제외: 매수 {miss_b}, 매도 {miss_s})")
        else:
            print(f"  ⚠ 엑셀 지표가 K값보다 적음(매수 {len(bp_use)}<{K_buy} 또는 매도 {len(sp_use)}<{K_sell})")
            print(f"     → 지표를 메타변수로 다시 선별해 재현(정확도 다소 하락 가능)")
    else:
        print(f"  ⚠ 앙상블 지표 시트를 읽지 못함 → 메타변수로 지표 재선별")

    kwargs['inject_pools'] = inject_pools   # ★ 엑셀 지표 풀 (None이면 메타변수로 재선별)

    # ★ 'k순신호 재현풀' 시트 읽기 → net>K를 원본 K/풀/지표수/g 그대로 재현 (탐색 0, pkl 불필요).
    try:
        import openpyxl as _oxl
        _wbk = _oxl.load_workbook(excel_path, read_only=True, data_only=True)
        if 'k순신호 재현풀' in _wbk.sheetnames:
            _wsk = _wbk['k순신호 재현풀']
            _rows = list(_wsk.iter_rows(values_only=True))
            # PARAM 헤더(2행)+VALUE(3행) → 헤더명으로 매핑 (전체K/OOS K)
            _hdrp = None; _valp = None
            for _r in _rows[:5]:
                if not _r: continue
                if str(_r[0]).strip() == 'PARAM': _hdrp = _r
                elif str(_r[0]).strip() == 'VALUE': _valp = _r
            _pm = {}
            if _hdrp and _valp:
                for _c in range(1, min(len(_hdrp), len(_valp))):
                    _k = _hdrp[_c]
                    if _k is not None: _pm[str(_k).strip()] = _valp[_c]
            _hi = None
            for _i, _r in enumerate(_rows):
                sv = [str(v).strip() if v is not None else '' for v in _r]
                if any('indicator' in s for s in sv) and '구분' in sv:
                    _hi = _i; break
            _bb = []; _ss = []
            if _hi is not None:
                for _r in _rows[_hi + 1:]:
                    if not _r or _r[0] is None or _r[1] is None: continue
                    side = str(_r[0]).strip()
                    rec = {'indicator': _r[1], 'direction': _r[2],
                           'threshold': _r[3], 'success_rate': _r[4]}
                    # ★ 신호 지연/선출한도/최적선행 재현 (구버전 파일은 열 없음 → 0/None)
                    try:
                        rec['lead_shift'] = int(_r[5]) if (len(_r) > 5 and _r[5] is not None) else 0
                    except Exception:
                        rec['lead_shift'] = 0
                    try:
                        rec['sel_limit'] = float(_r[6]) if (len(_r) > 6 and _r[6] is not None) else None
                    except Exception:
                        rec['sel_limit'] = None
                    try:
                        rec['best_lead'] = int(_r[7]) if (len(_r) > 7 and _r[7] is not None) else None
                    except Exception:
                        rec['best_lead'] = None
                    if '매수' in side: _bb.append(rec)
                    elif '매도' in side: _ss.append(rec)
            if _bb and _ss:
                _tkr = used.get('ticker') or (ticker if 'ticker' in dir() else None)
                _bdf = pd.DataFrame(_bb); _sdf = pd.DataFrame(_ss)
                if feat is not None:
                    _cols = set(feat.columns)
                    _bdf = _bdf[_bdf['indicator'].isin(_cols)].reset_index(drop=True)
                    _sdf = _sdf[_sdf['indicator'].isin(_cols)].reset_index(drop=True)
                globals()['_KNET_MULTI_POOL'] = (_tkr, _bdf, _sdf)
                # 구버전(best_k 단일) 호환: k_full 없으면 best_k 사용
                _kf = _pm.get('k_full', _pm.get('best_k'))
                globals()['_KNET_REPLAY_FIXED'] = {
                    'k_full': _kf, 'nb_full': _pm.get('nb_full', _pm.get('n_buy')),
                    'ns_full': _pm.get('ns_full', _pm.get('n_sell')), 'g_full': _pm.get('g_full', _pm.get('weight_g', 1.0)),
                    'k_oos': _pm.get('k_oos'), 'nb_oos': _pm.get('nb_oos'),
                    'ns_oos': _pm.get('ns_oos'), 'g_oos': _pm.get('g_oos', 1.0)}
                # ★ K/L 2임계 재현값
                if _pm.get('kl_k') is not None:
                    globals()['_KNET_KL_FIXED'] = {
                        'kl_k': _pm.get('kl_k'), 'kl_l': _pm.get('kl_l'),
                        'kl_k_mdd': _pm.get('kl_k_mdd'), 'kl_l_mdd': _pm.get('kl_l_mdd')}
                else:
                    globals().pop('_KNET_KL_FIXED', None)
                print(f"  ♻ 'k순신호 재현풀' 시트 로드 — 탐색 없이 그대로 재현 "
                      f"(K전체={_kf}, K_OOS={_pm.get('k_oos')}, K/L={_pm.get('kl_k')}/{_pm.get('kl_l')}, "
                      f"매수 {len(_bdf)}행 / 매도 {len(_sdf)}행)")
        _wbk.close()
    except Exception as _ke:
        print(f"  ⚠ k순신호 재현풀 시트 읽기 실패(재구성으로 진행): {_ke}")

    # ★ 보정 정확 재현 (요청) — 엑셀 앙상블 시트에 저장된 '🔁가중치'가 있으면, 그 풀과 가중치를
    #   force_corr로 넘겨 최종 백테스트가 '저장된 그대로'(추가지표+보정가중치) 재현되게 한다.
    #   (재현 중에는 보정을 다시 돌리지 않음 — 결정성/데이터 갱신 무관하게 원본 일치)
    _replay_force_corr = None
    if inject_pools is not None:
        _bp_inj, _sp_inj = inject_pools
        _has_w = ('vote_weight' in _bp_inj.columns and _bp_inj['vote_weight'].notna().any()
                  and 'vote_weight' in _sp_inj.columns and _sp_inj['vote_weight'].notna().any())
        if _has_w:
            _bw_arr = _bp_inj['vote_weight'].astype(float).values
            _sw_arr = _sp_inj['vote_weight'].astype(float).values
            # 풀 크기 = 저장된 전체(=보정 후 augmented). K도 그 길이로 강제.
            _replay_force_corr = {'buy_pool': _bp_inj, 'sell_pool': _sp_inj,
                                  'buy_w': _bw_arr, 'sell_w': _sw_arr,
                                  'n_added_buy': 0, 'n_added_sell': 0}
            kwargs['force_corr'] = _replay_force_corr
            kwargs['force_best_combo'] = {'K_buy': len(_bp_inj), 'vote_buy': vote_buy,
                                          'K_sell': len(_sp_inj), 'vote_sell': vote_sell}
            print(f"  ♻ 저장된 보정 가중치로 정확 재현 — 매수풀 {len(_bp_inj)}개 / 매도풀 {len(_sp_inj)}개 "
                  f"(가중치 그대로 주입, 보정 재실행 안 함)")
        else:
            print(f"  ℹ 엑셀에 '🔁가중치'가 없어(구버전 파일) 점수로 가중치 재계산 — "
                  f"보정 가중치는 정확 재현 안 될 수 있음")

    # ★ B&H 미달 제외도 재현 중에는 끈다 (그 조합이 B&H 미달이어도 그대로 봐야 함)
    _exbh_saved = globals().get('EXCLUDE_BELOW_BH')
    # 거래비용은 전역 COST_PER_TRADE 사용 → 재현 시 사용된 설정값으로 임시 교체 후 복원
    _cost_saved = globals().get('COST_PER_TRADE')
    _cost_used = used.get('cost')
    try:
        globals()['EXCLUDE_BELOW_BH'] = False
        if _cost_used is not None:
            globals()['COST_PER_TRADE'] = _cost_used
        result = run_ensemble_search(**kwargs)
    finally:
        globals()['EXCLUDE_BELOW_BH'] = _exbh_saved
        if _cost_used is not None:
            globals()['COST_PER_TRADE'] = _cost_saved

    _done_label = f'그리드 #{gn}' if grid_number is not None else '★최적 조합'
    print(f'\n  ✅ {_done_label} 일별 Excel 생성 완료 → {output_file}')
    if AUTO_DOWNLOAD_EXCEL:
        _auto_download_excels([output_file])
    return result

def find_latest_excel(ticker, *, output_dir=None):
    """해당 티커의 '가장 최근 날짜' 일반 분석 엑셀을 찾는다.
       - 파일명 형식: ensemble_search_{TICKER}_{YYYY-MM-DD}.xlsx
       - '__replay'가 들어간 재현본은 제외 (요청: 2번 그리드/재현 파일 아님)
       - 날짜가 가장 늦은 파일을 반환. 없으면 None.
    """
    import glob, re as _re
    base_dir = output_dir if output_dir is not None else OUTPUT_DIR
    pat = os.path.join(base_dir, f"ensemble_search_{ticker}_*.xlsx")
    cands = []
    for p in glob.glob(pat):
        name = os.path.basename(p)
        if '__replay' in name:        # 재현본 제외
            continue
        m = _re.search(rf"ensemble_search_{_re.escape(ticker)}_(\d{{4}}-\d{{2}}-\d{{2}})\.xlsx$", name)
        if not m:
            continue
        cands.append((m.group(1), p))   # (날짜문자열, 경로)
    if not cands:
        return None
    cands.sort(key=lambda x: x[0])      # 날짜 오름차순
    return cands[-1][1]                 # 가장 최근


def replay_latest_best(ticker, *, output_dir=None, feat=None, close=None, **override_kwargs):
    """티커의 가장 최근 일반 분석 엑셀을 자동으로 찾아, 그 엑셀의 ★최적 조합·지표를
       그대로 재현한다. (엑셀 기간까지 동일, 그 이후 새 거래일은 같은 로직으로 이어 계산)"""
    latest = find_latest_excel(ticker, output_dir=output_dir)
    if latest is None:
        raise RuntimeError(
            f"'{ticker}'의 일반 분석 엑셀을 찾지 못했습니다.\n"
            f"  {output_dir or OUTPUT_DIR} 안에 ensemble_search_{ticker}_YYYY-MM-DD.xlsx 가 있어야 합니다.\n"
            f"  (먼저 새 분석을 1회 실행해 엑셀을 만들어야 재현할 수 있습니다.)")
    print(f"  📂 '{ticker}' 가장 최근 분석 엑셀: {os.path.basename(latest)}")
    # grid_number=None → ★최적 조합 자동 재현
    return replay_grid_combo(latest, None, feat=feat, close=close,
                             output_dir=output_dir, **override_kwargs)


def catboost_correct_actions(feat, close, buy_pool, sell_pool, best_inner, *,
                             horizon=None, dd_limit=None, ru_limit=None,
                             stop_loss_pct=None, anchor_mode=True,
                             anchor_safe_buy=None, anchor_safe_sell=None):
    """선정 조합의 신호를 CatBoost로 보정한다 (요청) — '미래 판단'에 쓰는 게 목적.
       핵심: 타겟을 사후 앵커가 아니라 '그 시점에 검증 가능한 미래 결과'로 둔다.
         - 타겟 = 다음 H일 보유 시 비용 넘는 수익이 났나 (1=보유가 이득 / 0=현금이 이득)
           → 실전에서도 똑같이 '오늘 사면 H일 뒤 이득인가'로 검증되는 라벨이라 일반화 가능.
         - 피처 = 매수/매도 신호강도, 최근 수익/변동성 (그 시점까지 정보만)
         - walk-forward 학습 + 마지막 20%는 out-of-sample(OOS)로 떼어 '진짜 미래 성능' 측정
         - 마지막 날 예측 = 오늘 보유/현금 권고 (실전 판단)
       반환: dict (OOS 성능 포함). 미래 판단에 쓰려면 OOS 개선 여부를 꼭 보라.
    """
    if not globals().get('USE_CATBOOST_CORRECTION', False):
        return None
    try:
        from catboost import CatBoostClassifier
    except Exception as _e:
        print(f"  \u26a0 CatBoost \ubbf8\uc124\uce58 \u2014 \ubcf4\uc815 \uac74\ub108\ub700 (pip install catboost). ({_e})")
        return None

    import numpy as _np
    horizon = horizon if horizon is not None else globals().get('HORIZON_DAYS', 1)
    dd_limit = dd_limit if dd_limit is not None else globals().get('DRAWDOWN_LIMIT_BUY', 0.01)
    ru_limit = ru_limit if ru_limit is not None else globals().get('RUNUP_LIMIT_SELL', 0.01)
    stop_loss_pct = stop_loss_pct if stop_loss_pct is not None else globals().get('STOP_LOSS_PCT', 0.05)
    cost = globals().get('COST_PER_TRADE', 0.004)
    H = int(globals().get('CATBOOST_TARGET_HORIZON', 5))   # 미래 수익 판단 기간

    # 1) 선정 조합으로 daily 백테스트 → 일별 신호
    daily, trades, cur, _bu, _su = daily_ensemble_backtest(
        feat, close, buy_pool, sell_pool,
        K_buy=int(best_inner['K_buy']), K_sell=int(best_inner['K_sell']),
        vote_buy=int(best_inner['vote_buy']), vote_sell=int(best_inner['vote_sell']),
        cost=cost, horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
        stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
        anchor_safe_buy=anchor_safe_buy, anchor_safe_sell=anchor_safe_sell)
    if daily is None or len(daily) < 80:
        print("  \u26a0 CatBoost \ubcf4\uc815 \u2014 \ub370\uc774\ud130 \ubd80\uc871")
        return None

    cl = close.reindex(feat.index).astype(float)
    px = cl.values.astype(_np.float64)
    n = len(px)

    # 2) 타겟 y[i] = i일에 보유 시작해 H일 뒤까지 비용 넘는 수익이 나면 1 (미래 결과 — 검증 가능)
    y = _np.full(n, -1, dtype=_np.int64)
    for i in range(n - H):
        fwd = px[i+H] / px[i] - 1.0
        y[i] = 1 if fwd > (2*cost) else 0
    # 마지막 H일은 미래수익 미확정 → 학습/평가에서 제외 (y=-1)

    # 3) 피처 — 그 시점까지 정보만 (look-ahead 없음)
    bc = daily['buy_count'].values.astype(_np.float64)  if 'buy_count' in daily else _np.zeros(n)
    sc = daily['sell_count'].values.astype(_np.float64) if 'sell_count' in daily else _np.zeros(n)
    bc = bc[:n]; sc = sc[:n]
    if len(bc) < n: bc = _np.pad(bc, (0, n-len(bc)))
    if len(sc) < n: sc = _np.pad(sc, (0, n-len(sc)))
    ret1=_np.zeros(n); ret5=_np.zeros(n); vol10=_np.zeros(n); mom10=_np.zeros(n)
    for i in range(1, n):
        ret1[i]=px[i]/px[i-1]-1.0
        if i>=5:  ret5[i]=px[i]/px[i-5]-1.0
        if i>=10:
            seg=px[i-10:i+1]
            vol10[i]=_np.std(_np.diff(seg)/seg[:-1]) if len(seg)>1 else 0.0
            mom10[i]=px[i]/px[i-10]-1.0
    Kb=max(int(best_inner['K_buy']),1); Ks=max(int(best_inner['K_sell']),1)
    X=_np.column_stack([bc, sc, bc/Kb, sc/Ks, bc-sc, ret1, ret5, vol10, mom10])

    base_hold = (daily['position_pre'].values.astype(float)[:n]
                 if 'position_pre' in daily else _np.zeros(n))

    # 4) OOS 분할 — 마지막 20%는 학습에 안 쓰고 '진짜 미래'로 평가
    # OOS 분할 — OOS_MONTHS(최근 N개월)를 우선 사용, 없으면 OOS_FRACTION
    _oos_m = int(globals().get('OOS_MONTHS', 0))
    if _oos_m > 0:
        _oos_start = feat.index.max() - pd.DateOffset(months=_oos_m)
        split = int((feat.index < _oos_start).sum())
        oos_frac = (n - split) / n if n > 0 else 0.2
    else:
        oos_frac = float(globals().get('CATBOOST_OOS_FRACTION', 0.2))
        split = int(n * (1 - oos_frac))
    min_train = int(globals().get('CATBOOST_MIN_TRAIN_DAYS', 120))
    thr = float(globals().get('CATBOOST_PROB_THRESHOLD', 0.5))
    if split <= min_train or (n - H) <= split:
        print(f"  \u26a0 CatBoost \ubcf4\uc815 \u2014 \ud559\uc2b5/OOS \ubd84\ud560 \ubd88\uac00 (n={n})")
        return None

    pred_hold = _np.full(n, _np.nan)
    # walk-forward (split 이후 구간을 과거로만 학습해 예측)
    step = max(10, (n - min_train)//10)
    t = min_train
    while t < n - H:
        end = min(t+step, n - H)
        mask = y[:t] >= 0
        Xtr, ytr = X[:t][mask], y[:t][mask]
        if len(_np.unique(ytr)) < 2:
            pred_hold[t:end] = ytr[-1] if len(ytr) else 0; t=end; continue
        try:
            m=CatBoostClassifier(iterations=150, depth=4, learning_rate=0.1,
                                 loss_function='Logloss', verbose=False,
                                 random_seed=42, allow_writing_files=False)
            m.fit(Xtr, ytr)
            proba=m.predict_proba(X[t:end])[:,1]
            pred_hold[t:end]=(proba>=thr).astype(float)
        except Exception:
            pred_hold[t:end]=ytr[-1] if len(ytr) else 0
        t=end

    # 5) 수익 시뮬 (단순합산)
    def _sim(hold, lo, hi):
        pos=0; s=0.0; prev=_np.nan
        for i in range(lo, hi):
            if pos==1 and i>0 and not _np.isnan(prev) and prev>0:
                s += (px[i]/prev - 1.0)
            w=hold[i]
            if not _np.isnan(w):
                if w>=0.5 and pos==0: s-=cost; pos=1
                elif w<0.5 and pos==1: s-=cost; pos=0
            prev=px[i]
        return s
    corr_hold = base_hold.copy()
    for i in range(n):
        if not _np.isnan(pred_hold[i]): corr_hold[i]=pred_hold[i]

    # 전체 + OOS 구간 각각 평가 (★OOS가 진짜 미래 성능)
    ret_before_all = _sim(base_hold, 0, n)
    ret_after_all  = _sim(corr_hold, 0, n)
    ret_before_oos = _sim(base_hold, split, n)
    ret_after_oos  = _sim(corr_hold, split, n)
    # OOS 예측 적중률 (보정이 미래를 맞췄나)
    oos_mask = (y[split:n-H] >= 0)
    oos_pred = pred_hold[split:n-H][oos_mask]
    oos_true = y[split:n-H][oos_mask]
    acc = float(_np.mean((oos_pred>=0.5).astype(int) == oos_true)) if len(oos_true)>0 else float('nan')

    # 6) 마지막 날 실전 권고 (오늘 보유/현금?)
    today_reco = None
    try:
        mask=y>=0
        mf=CatBoostClassifier(iterations=150, depth=4, learning_rate=0.1,
                              loss_function='Logloss', verbose=False,
                              random_seed=42, allow_writing_files=False)
        mf.fit(X[mask], y[mask])
        p_today=float(mf.predict_proba(X[-1:])[:,1][0])
        today_reco = ('\ubcf4\uc720(\ub9e4\uc218)' if p_today>=thr else '\ud604\uae08(\ub9e4\ub3c4)', p_today)
    except Exception:
        pass

    print(f"  \U0001f916 CatBoost \ubcf4\uc815 (\ud0c0\uac9f=\ubbf8\ub798 {H}\uc77c \uc218\uc775) \u2014 \u2605OOS(\ucd5c\uadfc {oos_frac*100:.0f}%, \ud559\uc2b5\uc548\ud568) \uc131\ub2a5:")
    print(f"     OOS \uc218\uc775: \ubcf4\uc815\uc804 {ret_before_oos*100:+.1f}% \u2192 \ubcf4\uc815\ud6c4 {ret_after_oos*100:+.1f}%  "
          f"(OOS \uc608\uce21 \uc801\uc911\ub960 {acc*100:.0f}%)")
    if today_reco:
        print(f"     \u2192 \uc624\ub298 \uad8c\uace0: {today_reco[0]} (\ubcf4\uc720\ud655\ub960 {today_reco[1]*100:.0f}%)")
    if ret_after_oos > ret_before_oos:
        print(f"     \u2705 OOS\uc5d0\uc11c \ubcf4\uc815\uc774 \uac1c\uc120 \u2192 \ubbf8\ub798 \ud310\ub2e8\uc5d0 \ub3c4\uc6c0 \uac00\ub2a5\uc131")
    else:
        print(f"     \u26a0 OOS\uc5d0\uc11c \ubcf4\uc815\uc774 \uac1c\uc120 \uc548 \ub428 \u2192 \uc774 \uc885\ubaa9\uc740 \ubcf4\uc815 \ud6a8\uacfc \uc5c6\uc74c (\uc801\uc6a9 \uc8fc\uc758)")
    return {
        'ret_before_all': ret_before_all, 'ret_after_all': ret_after_all,
        'ret_before_oos': ret_before_oos, 'ret_after_oos': ret_after_oos,
        'oos_accuracy': acc, 'oos_improved': bool(ret_after_oos > ret_before_oos),
        'today_reco': today_reco, 'target_horizon': H,
        'dates': feat.index, 'pred_hold': pred_hold,
        'base_hold': base_hold, 'corr_hold': corr_hold,
    }


def catboost_correct_from_excel(filename, *, feat=None, close=None, output_dir=None):
    """엑셀의 ★최적 조합을 읽어 CatBoost 액션 보정을 적용한다 (요청, 노트북에서 호출).
       사용 예: mod.catboost_correct_from_excel('ensemble_search_CAT_2026-06-09.xlsx',
                                                 feat=내_feat, close=내_close)
    """
    base_dir = output_dir if output_dir is not None else OUTPUT_DIR
    path = filename if os.path.isabs(filename) else os.path.join(base_dir, filename)
    if not os.path.exists(path):
        raise RuntimeError(f"파일 없음: {path}")
    wb = load_workbook(path, data_only=True)
    used = _parse_used_settings(wb)
    for _k in ('K_buy', 'vote_buy', 'K_sell', 'vote_sell'):
        if used.get(_k) is None:
            raise RuntimeError(f"'사용된 설정'에서 {_k}를 못 읽음")
    best = {'K_buy': int(used['K_buy']), 'vote_buy': int(used['vote_buy']),
            'K_sell': int(used['K_sell']), 'vote_sell': int(used['vote_sell'])}
    # 지표 풀 — 엑셀의 매수/매도 앙상블 지표 시트
    def _rp(sheet):
        if sheet not in wb.sheetnames: return None
        ws = wb[sheet]; h = {}
        for c in range(1, ws.max_column+1):
            v = ws.cell(3, c).value
            if v is not None: h[str(v).strip()] = c
        rows = []
        for r in range(4, ws.max_row+1):
            ind = ws.cell(r, h.get('지표', 2)).value
            if ind is None or str(ind).strip() == '': continue
            def _cf(name, d=0.0):
                if name in h:
                    try: return float(str(ws.cell(r, h[name]).value).replace('%','').strip())
                    except Exception: return d
                return d
            rows.append({'indicator': str(ind).strip(),
                         'direction': str(ws.cell(r, h.get('방향', 3)).value or '>=').strip(),
                         'threshold': _cf('임계치'), 'pct_label': _cf('분위', 50.0),
                         'n_signals': _cf('신호수'), 'n_success': _cf('성공수'),
                         'success_rate': _cf('성공률'), 'avg_extreme': 0.01, 'score': _cf('점수'),
                         'lead_shift': int(_cf('지연(일)', 0.0))})
        return pd.DataFrame(rows) if rows else None
    bp = _rp('매수_앙상블_지표'); sp = _rp('매도_앙상블_지표')
    if bp is None or sp is None:
        raise RuntimeError("앙상블 지표 시트를 못 읽음")
    if feat is None or close is None:
        feat, close, _tk = _resolve_data()
    g = globals(); g['USE_CATBOOST_CORRECTION'] = True
    return catboost_correct_actions(feat, close, bp, sp, best, anchor_mode=g.get('ANCHOR_MODE', True))


def _apply_correction_to_candidates(verified, feat, close, pool_map, *,
                                    horizon, dd_limit, ru_limit, stop_loss_pct,
                                    anchor_mode, anchor_safe_buy, anchor_safe_sell):
    """수익 상위 N개 검증 후보에 지표 가중치 보정을 적용하고 보정 후 지표를 재계산 (요청).
       corr_* 컬럼 추가: corr_total_return / corr_max_drawdown / corr_win_rate /
       corr_buy_success / corr_sell_success / corr_buy_match / corr_sell_match /
       corr_applied(보정 채택 여부). 보정 안 한 후보는 real_* 값을 그대로 corr_*에 복사.
    """
    import numpy as _np
    topn = int(globals().get('CORRECT_VERIFY_TOP_N', 20))
    # 기본: corr_* = real_* (보정 안 됨)
    verified = verified.copy()
    verified['corr_total_return'] = verified['real_total_return']
    verified['corr_max_drawdown'] = verified['real_max_drawdown']
    verified['corr_win_rate']     = verified['real_win_rate']
    verified['corr_buy_success']  = verified['real_buy_success']
    verified['corr_sell_success'] = verified['real_sell_success']
    verified['corr_buy_match']    = verified['real_buy_match']
    verified['corr_sell_match']   = verified['real_sell_match']
    verified['corr_applied']      = False
    verified['corr_n_added_buy']  = 0
    verified['corr_n_added_sell'] = 0

    # 수익 상위 N개만 보정
    order = verified.sort_values('real_total_return', ascending=False).index[:topn]
    print(f"\n  🔧 보정 적용 — 수익 상위 {len(order)}개 후보에 지표 가중치 보정 시도 "
          f"(보정 후 수익/손실/승률 재계산)")
    _t0 = time.time()
    n_adopted = 0
    for cnt, idx in enumerate(order):
        _r = verified.loc[idx]
        _key = (round(float(_r.get('meta_wilson_z', 0)),4),
                int(_r.get('meta_pct_low', 0)), int(_r.get('meta_pct_high', 0)),
                round(float(_r.get('meta_corr_limit', 0)),4))
        pools = pool_map.get(_key)
        if pools is None:
            if cnt == 0:
                print(f"     ⚠ 풀 매칭 실패 — key={_key}, 사용가능 keys={list(pool_map.keys())[:3]}")
            continue
        bp, sp = pools
        # ★ 보정용 전체 후보 풀 (top_n_pool 100 무관, 전체 2800개 기반 점수 후보)
        _fcmap = globals().get('_LAST_FULL_CAND_MAP', {})
        _fc = _fcmap.get(_key, (None, None))
        _fcb, _fcs = _fc if _fc else (None, None)
        best_inner = {'K_buy': int(_r['K_buy']), 'vote_buy': int(_r['vote_buy']),
                      'K_sell': int(_r['K_sell']), 'vote_sell': int(_r['vote_sell'])}
        try:
            aug_bp, aug_sp, bw, sw, mc = anchor_match_correct_weights(
                feat, close, bp, sp, best_inner,
                horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
                stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
                anchor_safe_buy=anchor_safe_buy, anchor_safe_sell=anchor_safe_sell,
                full_cand_buy=_fcb, full_cand_sell=_fcs,
                verbose=(cnt == 0))
        except Exception as _e1:
            if cnt < 3:
                import traceback
                print(f"     ⚠ 보정 함수 에러(후보 {cnt}): {_e1}")
                traceback.print_exc()
            continue
        if not mc.get('adopted', False) or bw is None:
            continue
        # 보정된 가중치로 재백테스트해서 보정 후 전체 지표 산출
        try:
            _d, _t, _cur, _bu, _su = daily_ensemble_backtest(
                feat, close, aug_bp, aug_sp,
                K_buy=len(aug_bp), K_sell=len(aug_sp),
                vote_buy=int(_r['vote_buy']), vote_sell=int(_r['vote_sell']),
                cost=COST_PER_TRADE, horizon=horizon,
                dd_limit=dd_limit, ru_limit=ru_limit,
                stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
                anchor_safe_buy=anchor_safe_buy, anchor_safe_sell=anchor_safe_sell,
                buy_w_override=bw, sell_w_override=sw)
        except Exception as _e2:
            if cnt < 3:
                print(f"     ⚠ 보정후 재백테스트 에러(후보 {cnt}): {_e2}")
            continue
        _cr = float(_cur.get('cum_return_pct', 0.0)) / 100.0
        _cmdd = float(_cur.get('max_drawdown', 0.0))
        _real_mdd = float(_r.get('real_max_drawdown', 0.0))
        # 보정 후 수익이 원래보다 낮으면 채택 안 함 (수익 손해 방지)
        if _cr < float(_r['real_total_return']) - 1e-9:
            continue
        # ★ 보정 후 최대손실(MDD)이 원래보다 깊으면 채택 안 함 (요청: MDD 악화 금지)
        if _cmdd < _real_mdd - 1e-9:
            continue
        verified.at[idx, 'corr_total_return'] = _cr
        verified.at[idx, 'corr_max_drawdown'] = float(_cur.get('max_drawdown', 0.0))
        verified.at[idx, 'corr_win_rate']     = float(_cur.get('win_rate', 0.0))
        verified.at[idx, 'corr_buy_success']  = float(_cur.get('buy_success_rate', 0.0))
        verified.at[idx, 'corr_sell_success'] = float(_cur.get('sell_success_rate', 0.0))
        verified.at[idx, 'corr_buy_match']    = float(_cur.get('anchor_buy_match_rate', 0.0))
        verified.at[idx, 'corr_sell_match']   = float(_cur.get('anchor_sell_match_rate', 0.0))
        verified.at[idx, 'corr_applied']      = True
        verified.at[idx, 'corr_n_added_buy']  = int(mc.get('n_added_buy', 0))
        verified.at[idx, 'corr_n_added_sell'] = int(mc.get('n_added_sell', 0))
        n_adopted += 1
    print(f"     ✓ 보정 채택 {n_adopted}/{len(order)}개 "
          f"(수익 떨어지면 미채택, 경과 {time.time()-_t0:.0f}초)")
    return verified


def anchor_match_correct_weights(feat, close, full_buy_pool, full_sell_pool, best_inner, *,
                                 horizon=None, dd_limit=None, ru_limit=None,
                                 stop_loss_pct=None, anchor_mode=True,
                                 anchor_safe_buy=None, anchor_safe_sell=None,
                                 full_cand_buy=None, full_cand_sell=None,
                                 max_boost=0.8, step=0.1, max_add=15, verbose=True):
    """앵커 미매칭을 '미사용 지표 추가 + 큰 가중치 부여'로 보정 (요청 개편).
       철학(사용자 제안): 기존에 잘 맞은 날엔 영향을 주지 않으면서,
         '간발의 차로 미달(신호 안 뜸)'·'충돌 패배'한 미매칭일에만 그날 켜지는
         미사용 지표를 풀에 추가하고 충분한 가중치를 줘서 vote 임계를 넘긴다.
       ★ 보정 후보는 top_n_pool(100) 풀이 아니라 '점수 매긴 전체 후보(2800개 기반)'에서
         찾는다 (full_cand_buy/sell). top_n_pool은 그대로 두고 보정만 전체 탐색 (사용자 제안).
       - 추가 지표는 표적성(미매칭일 선택적 발화) 높은 것 우선 → 미래 일반화 + 다른날 영향 최소.
       - 추가 후 여러 가중치 크기를 시도해 '실제 수익이 가장 높아지는' 조합 채택.
       - 수익이 기존보다 낮으면 미채택(롤백).
       반환: (buy_pool_used, sell_pool_used, buy_w, sell_w, 결과dict)
    """
    import numpy as _np
    horizon = horizon if horizon is not None else globals().get('HORIZON_DAYS', 1)
    dd_limit = dd_limit if dd_limit is not None else globals().get('DRAWDOWN_LIMIT_BUY', 0.01)
    ru_limit = ru_limit if ru_limit is not None else globals().get('RUNUP_LIMIT_SELL', 0.01)
    stop_loss_pct = stop_loss_pct if stop_loss_pct is not None else globals().get('STOP_LOSS_PCT', 0.05)
    cost = globals().get('COST_PER_TRADE', 0.004)
    Kb = int(best_inner['K_buy']); Ks = int(best_inner['K_sell'])
    vb = int(best_inner['vote_buy']); vs = int(best_inner['vote_sell'])

    base_buy  = full_buy_pool.iloc[:Kb].reset_index(drop=True)
    base_sell = full_sell_pool.iloc[:Ks].reset_index(drop=True)

    # ★ extra(보정 후보) = 전체 후보 풀에서 'base에 없는 지표'. 전체가 없으면 작은 풀에서 폴백.
    def _make_extra(base_pool, small_full, big_cand):
        if big_cand is not None and len(big_cand) > 0:
            used = set(zip(base_pool['indicator'], base_pool['direction'],
                           _np.round(base_pool['threshold'].astype(float), 8)))
            keep = []
            for _, r in big_cand.iterrows():
                kkey = (r['indicator'], r['direction'], round(float(r['threshold']), 8))
                if kkey not in used:
                    keep.append(r)
            if keep:
                return pd.DataFrame(keep).reset_index(drop=True)
        return small_full.iloc[len(base_pool):].reset_index(drop=True)

    extra_buy  = _make_extra(base_buy,  full_buy_pool,  full_cand_buy)
    extra_sell = _make_extra(base_sell, full_sell_pool, full_cand_sell)

    def _run(bpool, spool, bwo=None, swo=None):
        return daily_ensemble_backtest(
            feat, close, bpool, spool, K_buy=len(bpool), K_sell=len(spool),
            vote_buy=vb, vote_sell=vs, cost=cost, horizon=horizon,
            dd_limit=dd_limit, ru_limit=ru_limit, stop_loss_pct=stop_loss_pct,
            anchor_mode=anchor_mode, anchor_safe_buy=anchor_safe_buy,
            anchor_safe_sell=anchor_safe_sell, buy_w_override=bwo, sell_w_override=swo)

    # 0) 기준(보정 전)
    d0, t0, cur0, bu0, su0 = _run(base_buy, base_sell)
    ret0 = float(cur0.get('cum_return_pct', 0.0))
    bm0 = float(cur0.get('anchor_buy_match_rate', 0.0))
    sm0 = float(cur0.get('anchor_sell_match_rate', 0.0))
    mdd0 = float(cur0.get('max_drawdown', 0.0))   # 기준 최대손실 (음수). 보정 후 이보다 깊으면 미채택.

    n = len(feat)
    W = int(globals().get('ANCHOR_MATCH_WINDOW', 2))
    sba = anchor_safe_buy; ssa = anchor_safe_sell
    pos_post = d0['position_pre'].values.astype(int)[:n] if 'position_pre' in d0 else _np.zeros(n, dtype=int)

    # ── 미매칭일 마스크 (앵커 정답인데 포지션 못 맞춘 날) ──
    miss_buy = _np.zeros(n, dtype=bool); miss_sell = _np.zeros(n, dtype=bool)
    matched_buy = _np.zeros(n, dtype=bool); matched_sell = _np.zeros(n, dtype=bool)
    if sba is not None and ssa is not None:
        for i in range(n):
            if i < len(sba) and sba[i] == 1:
                if any(pos_post[i+dd] == 1 for dd in range(0, W+1) if i+dd < n):
                    matched_buy[i] = True
                else:
                    miss_buy[i] = True
            if i < len(ssa) and ssa[i] == 1:
                if any(pos_post[i+dd] == 0 for dd in range(0, W+1) if i+dd < n):
                    matched_sell[i] = True
                else:
                    miss_sell[i] = True

    no_anchor = (sba is None or ssa is None or (miss_buy.sum()+miss_sell.sum()) == 0)
    if verbose:
        print(f"     [보정진단] 미매칭 매수 {int(miss_buy.sum())}일 / 매도 {int(miss_sell.sum())}일, "
              f"매칭 매수 {int(matched_buy.sum())} / 매도 {int(matched_sell.sum())}, "
              f"extra 매수 {len(extra_buy)} / 매도 {len(extra_sell)}, 기준수익 {ret0:+.1f}%")

    def _sig_arr(pool):
        return [_to_signal_array(feat, prow).astype(bool) for _, prow in pool.iterrows()]

    # ── 추가 지표 선택: 미매칭일에 자주 켜지고, 매칭/정상일엔 덜 켜지는 것 ──
    #   targeting score = (미매칭일 켜짐 비율) - 0.5*(매칭일 켜짐 비율)
    def _pick_extra(extra_pool, miss_mask, matched_mask):
        """미매칭일을 '실제로 덮는' 최소 지표만 선택 (요청: 항상 15개 추가 방지).
           표적성(미매칭에 선택적 발화) 높은 순으로 보되, '새 미매칭일을 추가로 덮는'
           지표만 채택하고 모든 미매칭일이 덮이면 멈춘다 → 추가 수가 +1, +2 처럼 가변.
           표적성 = 미매칭일 적중 / (그 외 날 발화 + eps).
        """
        if len(extra_pool) == 0 or miss_mask.sum() == 0:
            return [], _np.array([])
        sigs = _sig_arr(extra_pool)
        nmiss = max(1, int(miss_mask.sum()))
        other_mask = ~miss_mask
        nother = max(1, int(other_mask.sum()))
        scores = []
        for arr in sigs:
            hit_miss  = int((arr & miss_mask).sum()) / nmiss
            hit_other = int((arr & other_mask).sum()) / nother
            if hit_miss <= 0:
                scores.append(0.0); continue
            scores.append(hit_miss / (hit_other + 0.05))
        scores = _np.array(scores, dtype=float)
        order = _np.argsort(-scores)
        # ★ 그리디 셋커버: 표적성 높은 순으로, '새 미매칭일을 덮는' 지표만 채택.
        #   모든 미매칭일이 덮이거나 max_add에 도달하면 중단 → 최소 개수만 추가.
        covered = _np.zeros_like(miss_mask, dtype=bool)
        total_miss = int(miss_mask.sum())
        pick = []
        for o in order:
            if scores[o] < 1.0:
                break                       # 표적성 미달이면 더 안 봄
            new_cover = sigs[o] & miss_mask & ~covered
            if int(new_cover.sum()) == 0:
                continue                    # 새로 덮는 미매칭일 없음 → 불필요, 건너뜀
            pick.append(int(o))
            covered |= (sigs[o] & miss_mask)
            if int(covered.sum()) >= total_miss:
                break                       # 미매칭일 전부 덮음 → 종료
            if len(pick) >= max_add:
                break                       # 안전 상한
        if not pick:  # 표적성 1.0 넘는 게 없으면 표적성 상위 소수라도 (최대 3개)
            pick = [int(o) for o in order if scores[o] > 0][:3]
        return pick, scores

    pick_b, bsc = _pick_extra(extra_buy, miss_buy, matched_buy)
    pick_s, ssc = _pick_extra(extra_sell, miss_sell, matched_sell)

    add_buy  = extra_buy.iloc[pick_b].reset_index(drop=True) if pick_b else extra_buy.iloc[[]].reset_index(drop=True)
    add_sell = extra_sell.iloc[pick_s].reset_index(drop=True) if pick_s else extra_sell.iloc[[]].reset_index(drop=True)
    aug_buy  = pd.concat([base_buy, add_buy], ignore_index=True)
    aug_sell = pd.concat([base_sell, add_sell], ignore_index=True)
    if verbose:
        _bt = f"{max(bsc[pick_b]):.1f}" if len(pick_b) else "-"
        _st = f"{max(ssc[pick_s]):.1f}" if len(pick_s) else "-"
        print(f"     [보정진단] 추가 후보 지표(표적성≥1.0) — 매수 {len(add_buy)}개(최고표적성 {_bt}), "
              f"매도 {len(add_sell)}개(최고표적성 {_st})")

    # 기본 가중치 (기존 K개) — score 기반 또는 1.0
    def _wbase(pool):
        if globals().get('USE_WEIGHTED_VOTE', False):
            return compute_vote_weights(pool['score'].values, globals().get('WEIGHT_MAX_RATIO', 1.6))
        return _np.ones(len(pool))
    bw_core = _wbase(base_buy); sw_core = _wbase(base_sell)

    # ── 추가 지표에 줄 가중치를 여러 크기로 시도 (vote 임계를 넘기게) ──
    #   add_w 후보: 추가 지표 1개당 가중치. 충분히 커야 간발 미달(1~2점차)을 넘긴다.
    best = {'adopted': False, 'ret_before': ret0, 'bm_before': bm0, 'sm_before': sm0,
            'ret_after': ret0, 'bm_after': bm0, 'sm_after': sm0,
            'mdd_before': mdd0, 'mdd_after': mdd0,
            'bw': None, 'sw': None, 'n_added_buy': int(len(add_buy)), 'n_added_sell': int(len(add_sell))}

    if no_anchor or (len(add_buy) == 0 and len(add_sell) == 0):
        if verbose:
            print(f"  \u2139 \uc218\uc775 \ubcf4\uc815 \u2014 \ucd94\uac00\ud560 \ubbf8\uc0ac\uc6a9 \uc9c0\ud45c \uc5c6\uc74c \u2192 \uae30\uc874 \uadf8\ub9ac\ub4dc \uc720\uc9c0")
        return base_buy, base_sell, None, None, best

    nb = len(add_buy); ns = len(add_sell)
    # 추가 가중치: 작게(0.2)부터 — 기존 매매 영향 최소화하며 미달일만 살짝 밀어줌.
    #   촘촘히 올려 '수익 안 떨어지고 매칭 느는' 지점을 찾는다.

    # 기존 풀 내 지표 중 '미매칭일에 켜지는' 것 = 충돌일에 밀어줄 후보 (가중치 상향용)
    def _core_boost_dir(base_pool, miss_mask, matched_mask):
        if miss_mask.sum() == 0:
            return _np.zeros(len(base_pool))
        sigs = _sig_arr(base_pool)
        nmiss = max(1, int(miss_mask.sum()))
        d = []
        for arr in sigs:
            d.append(int((arr & miss_mask).sum()) / nmiss)   # 미매칭일에 켜지는 비율
        d = _np.array(d)
        return d/d.max() if d.max() > 0 else d
    bcore_dir = _core_boost_dir(base_buy, miss_buy, matched_buy)
    score_dir = _core_boost_dir(base_sell, miss_sell, matched_sell)

    _tried = []
    for add_w in (0.2, 0.3, 0.5, 0.8, 1.2, 1.8, 2.5):
        for core_boost in (0.0, 0.3, 0.6):
            bcore_w = bw_core * (1.0 + core_boost * bcore_dir)
            score_w = sw_core * (1.0 + core_boost * score_dir)
            bw = _np.concatenate([bcore_w, _np.full(nb, add_w)]) if nb > 0 else bcore_w
            sw = _np.concatenate([score_w, _np.full(ns, add_w)]) if ns > 0 else score_w
            d1, t1, cur1, _b1, _s1 = _run(aug_buy, aug_sell, bw, sw)
            ret1 = float(cur1.get('cum_return_pct', 0.0))
            bm1 = float(cur1.get('anchor_buy_match_rate', 0.0))
            sm1 = float(cur1.get('anchor_sell_match_rate', 0.0))
            mdd1 = float(cur1.get('max_drawdown', 0.0))
            _tried.append((ret1, mdd1))
            # ★ 채택 조건 (요청):
            #   1) 수익이 기존(ret0)보다 오를 것 (수익 최대화)
            #   2) 최대손실(MDD)이 기존(mdd0)보다 깊지 않을 것 — 같거나 더 얕아야 함
            #   둘 다 만족할 때만 채택. 수익만 오르고 손실이 깊어지면 미채택.
            mdd_ok = (mdd1 >= mdd0 - 1e-9)   # mdd는 음수: -2%보다 -6%가 작음 → mdd1>=mdd0이면 안 깊어짐
            if ret1 > ret0 + 1e-9 and mdd_ok and ret1 > best['ret_after'] + 1e-9:
                best.update({'adopted': True, 'ret_after': ret1, 'bm_after': bm1, 'sm_after': sm1,
                             'mdd_after': mdd1, 'bw': bw.copy(), 'sw': sw.copy()})
    if verbose and _tried:
        # MDD 제약(악화 안 됨)을 만족하는 시도 중 최고수익
        _ok = [r for (r, m) in _tried if m >= mdd0 - 1e-9]
        _mx_all = max(r for (r, m) in _tried)
        _mx_ok = max(_ok) if _ok else None
        if _mx_ok is not None and _mx_ok > ret0:
            print(f"     [보정진단] {len(_tried)}회 시도 — MDD유지하며 최고수익 {_mx_ok:+.1f}% "
                  f"(기준 {ret0:+.1f}%, MDD {mdd0*100:.2f}%) 개선됨✓")
        else:
            print(f"     [보정진단] {len(_tried)}회 시도 — 최고수익 {_mx_all:+.1f}%지만 "
                  f"MDD 악화로 미채택 (기준수익 {ret0:+.1f}%, 기준MDD {mdd0*100:.2f}%)")

    if best['adopted']:
        if verbose:
            print(f"  \U0001f3af \uc218\uc775 \ubcf4\uc815 (\ubbf8\uc0ac\uc6a9 \uc9c0\ud45c \ucd94\uac00 + \uac00\uc911\uce58 \ubd80\uc5ec, \ub0a0\uc9dc \ud655\uc7a5 \uc548\ud568):")
            print(f"     \uc218\uc775 {best['ret_before']:+.1f}% \u2192 {best['ret_after']:+.1f}% "
                  f"(+{best['ret_after']-best['ret_before']:.1f}%p)")
            print(f"     \ucd5c\ub300\uc190\uc2e4 {best['mdd_before']*100:.2f}% \u2192 {best['mdd_after']*100:.2f}% "
                  f"(\uc545\ud654 \uc5c6\uc74c \u2014 \uc870\uac74 \ucda9\uc871)")
            print(f"     \ub9e4\uc218\ub9e4\uce6d {best['bm_before']*100:.1f}% \u2192 {best['bm_after']*100:.1f}%, "
                  f"\ub9e4\ub3c4\ub9e4\uce6d {best['sm_before']*100:.1f}% \u2192 {best['sm_after']*100:.1f}%")
            print(f"     \ucd94\uac00\ub41c \ubbf8\uc0ac\uc6a9 \uc9c0\ud45c: \ub9e4\uc218 {best['n_added_buy']}\uac1c, \ub9e4\ub3c4 {best['n_added_sell']}\uac1c")
        return aug_buy, aug_sell, best['bw'], best['sw'], best
    else:
        if verbose:
            print(f"  \u2139 \uc218\uc775 \ubcf4\uc815 \u2014 \uac1c\uc120\ub418\ub294 \ucd94\uac00 \uc5c6\uc74c \u2192 \uae30\uc874 \uadf8\ub9ac\ub4dc \uc720\uc9c0")
        return base_buy, base_sell, None, None, best


def diagnose_trades(trades_df, close, feat_index, anchor_prio=None, *, profit_floor=0.01):
    """실제 거래 내역을 진단한다 (요청) — 앵커 날짜 외의 거래도 포함.
       각 거래에 대해:
        - net_return이 손실(<0)이거나 profit_floor(1%) 이하면 '틀린 거래' 후보
        - 매수일/매도일이 앵커(우선순위1 최대수익스윙 / 우선순위2 1%이상)와 맞는지
        - 잘못된 매수(저점 아닌 곳에서 삼) / 잘못된 매도(고점 아닌 곳에서 팜) 진단
       반환: 진단 DataFrame (거래별 1행) — Excel '거래_진단' 시트 + 보정 대상 식별용
    """
    import numpy as _np
    if trades_df is None or len(trades_df) == 0:
        return pd.DataFrame()
    norm_idx = pd.DatetimeIndex(feat_index).normalize()
    # 앵커 우선순위 집합 (날짜 문자열)
    ap = anchor_prio or {}
    buy_p1  = set(pd.Timestamp(d).normalize() for d in ap.get('buy_p1', set()))
    sell_p1 = set(pd.Timestamp(d).normalize() for d in ap.get('sell_p1', set()))
    buy_p2  = set(pd.Timestamp(d).normalize() for d in ap.get('buy_p2', set()))
    sell_p2 = set(pd.Timestamp(d).normalize() for d in ap.get('sell_p2', set()))

    rows = []
    for _, tr in trades_df.iterrows():
        ed = pd.Timestamp(tr['entry_date']).normalize() if pd.notna(tr.get('entry_date')) else None
        xd = pd.Timestamp(tr['exit_date']).normalize()  if pd.notna(tr.get('exit_date'))  else None
        ret = float(tr.get('net_return_%', 0.0)) / 100.0
        is_loss = ret < 0
        is_thin = (ret >= 0) and (ret <= profit_floor)   # 1% 이하 수익
        is_wrong = is_loss or is_thin

        # 매수일 진단
        if ed in buy_p1:    buy_status = '✓ 최대수익스윙 저점(P1)'
        elif ed in buy_p2:  buy_status = '○ 1%이상 저점(P2)'
        else:               buy_status = '✗ 앵커 아닌 곳에서 매수'
        # 매도일 진단
        if xd in sell_p1:    sell_status = '✓ 최대수익스윙 고점(P1)'
        elif xd in sell_p2:  sell_status = '○ 1%이상 고점(P2)'
        else:                sell_status = '✗ 앵커 아닌 곳에서 매도'

        buy_ok  = (ed in buy_p1) or (ed in buy_p2)
        sell_ok = (xd in sell_p1) or (xd in sell_p2)

        if is_loss:        verdict = '손실 거래 → 보정 대상'
        elif is_thin:      verdict = f'수익 {ret*100:.1f}%(≤{profit_floor*100:.0f}%) → 보정 대상'
        else:              verdict = '정상(수익)'

        rows.append({
            'trade_no': int(tr.get('trade_no', 0)),
            'entry_date': ed, 'exit_date': xd,
            'net_return_%': ret * 100.0,
            'days_held': int(tr.get('days_held', 0)),
            'exit_reason': tr.get('exit_reason', ''),
            '매수_진단': buy_status, '매도_진단': sell_status,
            '매수_앵커일치': buy_ok, '매도_앵커일치': sell_ok,
            '판정': verdict,
            '보정대상': is_wrong,
        })
    return pd.DataFrame(rows)


# ════════════════════════════════════════════════════════════════════════
#                  ★ 다중 티커 일괄 분석 시스템
# ════════════════════════════════════════════════════════════════════════
TICKERS = ['MPC']

AUTO_TUNE = {
    'EVAL_START':                  False,
    'HORIZON_DAYS':                False,
    'DRAWDOWN_LIMIT_BUY':          False,
    'RUNUP_LIMIT_SELL':            False,
    'AUTO_ANCHOR_WINDOW':          False,
    'AUTO_ANCHOR_LOOKFORWARD':     False,
    'AUTO_ANCHOR_MIN_RISE':        False,
    'AUTO_ANCHOR_MIN_DROP':        False,
    'AUTO_ANCHOR_PRICE_TOLERANCE': False,
    'wilson_z':                    False,
    'pct_range':                   False,
    'corr_limit':                  False,
}

MULTI_SUMMARY_FILE = 'ensemble_summary_all_tickers.xlsx'


def analyze_ticker_volatility(close):
    p = close.values.astype(np.float64)
    n = len(p)
    if n < 20:
        return {'daily_vol': 0.02, 'daily_vol_pct': 2.0, 'weekly_vol': 0.04,
                'n_days': n, 'typical_5d_move': 5.0, 'typical_10d_move': 7.0,
                'max_local_drawdown_5d': 3.0, 'max_local_runup_5d': 3.0,
                'price_dispersion': 0.2}
    rets = np.diff(p) / p[:-1]
    daily_vol = float(np.std(rets))
    daily_vol_pct = float(np.mean(np.abs(rets))) * 100.0

    moves_5d = []
    moves_10d = []
    dd_5d = []
    ru_5d = []
    for i in range(5, n - 5):
        win5 = p[i:i+6]
        moves_5d.append(abs(win5.max() / win5.min() - 1.0) * 100.0)
        start_p = p[i]
        if start_p > 0:
            end = min(i + 5, n - 1)
            future = p[i+1:end+1] if end >= i+1 else p[i:i+1]
            if len(future) > 0:
                dd_5d.append((future.min() / start_p - 1.0) * 100.0)
                ru_5d.append((future.max() / start_p - 1.0) * 100.0)
    for i in range(10, n - 10):
        win10 = p[i:i+11]
        moves_10d.append(abs(win10.max() / win10.min() - 1.0) * 100.0)

    weekly_rets = []
    for i in range(0, n - 5, 5):
        if p[i] > 0:
            weekly_rets.append(p[min(i+5, n-1)] / p[i] - 1.0)
    weekly_vol = float(np.std(weekly_rets)) if weekly_rets else daily_vol * np.sqrt(5)

    return {
        'daily_vol': daily_vol,
        'daily_vol_pct': daily_vol_pct,
        'weekly_vol': weekly_vol,
        'n_days': n,
        'typical_5d_move': float(np.median(moves_5d)) if moves_5d else 5.0,
        'typical_10d_move': float(np.median(moves_10d)) if moves_10d else 7.0,
        'max_local_drawdown_5d': float(np.median(dd_5d)) if dd_5d else -3.0,
        'max_local_runup_5d':    float(np.median(ru_5d)) if ru_5d else 3.0,
        'price_dispersion': float(np.std(p) / max(np.mean(p), 1e-10)),
    }


def auto_tune_variables(close, feat=None, *, auto_tune_flags=None):
    flags = auto_tune_flags or AUTO_TUNE
    stats = analyze_ticker_volatility(close)
    v   = stats['daily_vol']
    vp  = stats['daily_vol_pct']
    nd  = stats['n_days']
    m5  = stats['typical_5d_move']
    m10 = stats['typical_10d_move']

    tuned = {}

    if flags.get('EVAL_START'):
        try:
            start_idx = max(0, min(20, len(close) - 1))
            tuned['EVAL_START'] = close.index[start_idx].strftime('%Y-%m-%d')
        except Exception:
            tuned['EVAL_START'] = None

    if flags.get('HORIZON_DAYS'):
        if vp >= 3.0:    tuned['HORIZON_DAYS'] = 3
        elif vp >= 2.0:  tuned['HORIZON_DAYS'] = 4
        elif vp >= 1.0:  tuned['HORIZON_DAYS'] = 5
        else:            tuned['HORIZON_DAYS'] = 7

    if flags.get('DRAWDOWN_LIMIT_BUY'):
        dd = max(0.005, min(0.05, v * 0.5))
        tuned['DRAWDOWN_LIMIT_BUY'] = round(dd, 4)

    if flags.get('RUNUP_LIMIT_SELL'):
        ru = max(0.005, min(0.05, v * 0.5))
        tuned['RUNUP_LIMIT_SELL'] = round(ru, 4)

    if flags.get('AUTO_ANCHOR_WINDOW'):
        if vp >= 2.5:    tuned['AUTO_ANCHOR_WINDOW'] = 1
        elif vp >= 1.5:  tuned['AUTO_ANCHOR_WINDOW'] = 2
        elif vp >= 1.0:  tuned['AUTO_ANCHOR_WINDOW'] = 3
        else:            tuned['AUTO_ANCHOR_WINDOW'] = 5

    if flags.get('AUTO_ANCHOR_LOOKFORWARD'):
        if vp >= 2.5:    tuned['AUTO_ANCHOR_LOOKFORWARD'] = 1
        elif vp >= 1.5:  tuned['AUTO_ANCHOR_LOOKFORWARD'] = 2
        elif vp >= 1.0:  tuned['AUTO_ANCHOR_LOOKFORWARD'] = 3
        else:            tuned['AUTO_ANCHOR_LOOKFORWARD'] = 5

    if flags.get('AUTO_ANCHOR_MIN_RISE'):
        rise = max(0.015, min(0.08, m5 / 100.0 * 0.6))
        tuned['AUTO_ANCHOR_MIN_RISE'] = round(rise, 4)

    if flags.get('AUTO_ANCHOR_MIN_DROP'):
        drop = max(0.015, min(0.08, m5 / 100.0 * 0.6))
        tuned['AUTO_ANCHOR_MIN_DROP'] = round(drop, 4)

    if flags.get('AUTO_ANCHOR_PRICE_TOLERANCE'):
        tol = max(0.002, min(0.015, v * 0.3))
        tuned['AUTO_ANCHOR_PRICE_TOLERANCE'] = round(tol, 4)

    if flags.get('wilson_z'):
        if   nd < 100:  tuned['wilson_z'] = 1.96
        elif nd < 300:  tuned['wilson_z'] = 1.96
        else:           tuned['wilson_z'] = 1.96

    if flags.get('pct_range'):
        if vp >= 2.0:    tuned['pct_range'] = (5, 95)
        elif vp >= 1.0:  tuned['pct_range'] = (10, 90)
        else:            tuned['pct_range'] = (15, 85)

    if flags.get('corr_limit'):
        if   nd < 100:  tuned['corr_limit'] = 0.85
        elif nd < 300:  tuned['corr_limit'] = 0.80
        else:           tuned['corr_limit'] = 0.75

    return tuned, stats


def _restore_tuned_vars(raw):
    """변경5: _load_summary_state가 읽은 문자열 변수값을 effective_vars 형태로 복원."""
    if not raw:
        return {}
    def _f(v):
        if v in (None, '—', ''): return None
        s = str(v).replace('%', '').replace('+', '').strip()
        try: return float(s)
        except Exception: return None
    def _i(v):
        f = _f(v)
        return int(f) if f is not None else None
    def _pct(v):
        if v in (None, '—', ''): return None
        s = str(v).replace('(', '').replace(')', '').replace('[', '').replace(']', '')
        try:
            parts = [int(float(x.strip())) for x in s.split(',') if x.strip()]
            return tuple(parts) if len(parts) == 2 else None
        except Exception:
            return None
    out = {}
    if raw.get('EVAL_START') not in (None, '—', ''):
        out['EVAL_START'] = str(raw['EVAL_START']).strip()
    if _i(raw.get('HORIZON_DAYS')) is not None: out['HORIZON_DAYS'] = _i(raw['HORIZON_DAYS'])
    for src, dst in [('DRAWDOWN_LIMIT_BUY','DRAWDOWN_LIMIT_BUY'),
                     ('RUNUP_LIMIT_SELL','RUNUP_LIMIT_SELL'),
                     ('AUTO_ANCHOR_MIN_RISE','AUTO_ANCHOR_MIN_RISE'),
                     ('AUTO_ANCHOR_MIN_DROP','AUTO_ANCHOR_MIN_DROP'),
                     ('AUTO_ANCHOR_PRICE_TOLERANCE','AUTO_ANCHOR_PRICE_TOLERANCE')]:
        v = raw.get(src)
        f = _f(v)
        if f is not None:
            out[dst] = f / 100.0 if '%' in str(v) else f
    for src, dst in [('AUTO_ANCHOR_WINDOW','AUTO_ANCHOR_WINDOW'),
                     ('AUTO_ANCHOR_LOOKFORWARD','AUTO_ANCHOR_LOOKFORWARD'),
                     ('min_signals','min_signals'), ('top_n_pool','top_n_pool')]:
        if _i(raw.get(src)) is not None: out[dst] = _i(raw[src])
    if _f(raw.get('wilson_z')) is not None: out['wilson_z'] = _f(raw['wilson_z'])
    if _f(raw.get('corr_limit')) is not None: out['corr_limit'] = _f(raw['corr_limit'])
    if _pct(raw.get('pct_range')) is not None: out['pct_range'] = _pct(raw['pct_range'])
    return out


def _load_summary_state(summary_file):
    """이미 분석 완료된 티커 + (변경5) 사용된 변수값(tuned_vars) 복원."""
    if not os.path.exists(summary_file):
        return {}, []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(summary_file, data_only=True)
        if '결과 요약' not in wb.sheetnames:
            return {}, []
        ws = wb['결과 요약']
        done = {}
        order = []
        header_row = None
        for ri, row in enumerate(ws.iter_rows(values_only=True), 1):
            if ri == 1: continue
            if ri == 2:
                header_row = list(row); continue
            if row[0] is None: continue
            ticker = str(row[0]).strip()
            if not ticker or ticker == '티커': continue
            done[ticker] = dict(zip(header_row, row)) if header_row else {}
            order.append(ticker)

        # ★ 변경5: '사용된 변수값' 시트에서 tuned_vars 복원
        if '사용된 변수값' in wb.sheetnames:
            ws2 = wb['사용된 변수값']
            hdr2 = None
            for ri, row in enumerate(ws2.iter_rows(values_only=True), 1):
                if ri == 2:
                    hdr2 = list(row); continue
                if ri <= 2 or not row or row[0] is None: continue
                tk = str(row[0]).strip()
                if tk not in done: continue
                d2 = dict(zip(hdr2, row)) if hdr2 else {}
                done[tk]['_tuned_vars_raw'] = {
                    'EVAL_START': d2.get('EVAL_START'),
                    'HORIZON_DAYS': d2.get('HORIZON'),
                    'DRAWDOWN_LIMIT_BUY': d2.get('DD한도'),
                    'RUNUP_LIMIT_SELL': d2.get('RU한도'),
                    'AUTO_ANCHOR_WINDOW': d2.get('ANCHOR_WIN'),
                    'AUTO_ANCHOR_LOOKFORWARD': d2.get('ANCHOR_LF'),
                    'AUTO_ANCHOR_MIN_RISE': d2.get('MIN_RISE'),
                    'AUTO_ANCHOR_MIN_DROP': d2.get('MIN_DROP'),
                    'AUTO_ANCHOR_PRICE_TOLERANCE': d2.get('PRICE_TOL'),
                    'wilson_z': d2.get('wilson_z'),
                    'pct_range': d2.get('pct_range'),
                    'corr_limit': d2.get('corr_limit'),
                    'min_signals': d2.get('MIN_SIG'),
                    'top_n_pool': d2.get('TOP_N'),
                    'tuning_source': d2.get('튜닝상태'),
                }
        return done, order
    except Exception as e:
        print(f"  ⚠ 요약 파일 로드 실패: {e}")
        return {}, []


def _save_summary_excel(summary_records, summary_file, all_tickers, *,
                         tuning_log=None):
    wb = Workbook(); wb.remove(wb.active)

    ws = wb.create_sheet('결과 요약', 0); ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = (f'전체 티커 분석 요약 — {len(summary_records)}/{len(all_tickers)} 완료  '
                            f'({datetime.now().strftime("%Y-%m-%d %H:%M:%S")})')
    ws.cell(1, 1).font = Font(bold=True, size=16, color='1F3864')
    ws.merge_cells('A1:T1')

    headers = ['티커', '상태', '신호', '포지션', '오늘가',
               '누적수익%', 'B&H%', 'vsB&H%p', '📈상승일전략%', '📈상승일B&H%',
               'MDD%', '거래수', '승률%',
               'BalAcc%', '매수BalAcc%', '매도BalAcc%',
               '⚓매칭률%', '⚓매수%', '⚓매도%',
               'K_b/v_b', 'K_s/v_s', '분석일시']
    _hdr(ws, 2, headers)

    sig_fills = {'🟢 매수': _BUY, '🔴 매도': _SELL, '📈 보유': _HOLD,
                  '💵 현금': _CASH, '⚔ 충돌': _CONF, '— 에러': _BAD}
    for ri, ticker in enumerate(all_tickers, 3):
        if ticker not in summary_records:
            ws.cell(ri, 1).value = ticker
            ws.cell(ri, 1).font = Font(bold=True, size=11)
            ws.cell(ri, 2).value = '대기 중'
            ws.cell(ri, 2).font = Font(size=10, italic=True, color='888888')
            for ci in range(1, len(headers) + 1):
                ws.cell(ri, ci).border = _TH
                ws.cell(ri, ci).alignment = Alignment(horizontal='center')
            continue
        rec = summary_records[ticker]
        status = rec.get('status', '완료')
        signal = rec.get('signal', '—')
        is_err = (status == '에러')
        vals = [
            ticker,
            status, signal, rec.get('position', '—'),
            f"${rec['close']:.2f}" if rec.get('close') is not None else '—',
            f"{rec['cum_return_pct']:+.2f}" if rec.get('cum_return_pct') is not None else '—',
            f"{rec['bh_pct']:+.2f}" if rec.get('bh_pct') is not None else '—',
            f"{rec['vs_bh_pp']:+.2f}" if rec.get('vs_bh_pp') is not None else '—',
            f"{rec['up_cum_return_pct']:+.2f}" if rec.get('up_cum_return_pct') is not None else '—',
            f"{rec['bh_up_pct']:+.2f}" if rec.get('bh_up_pct') is not None else '—',
            f"{rec['mdd_pct']:.2f}" if rec.get('mdd_pct') is not None else '—',
            rec.get('n_trades', '—'),
            f"{rec['win_rate_pct']:.1f}" if rec.get('win_rate_pct') is not None else '—',
            f"{rec['balacc_pct']:.1f}" if rec.get('balacc_pct') is not None else '—',
            f"{rec['buy_balacc_pct']:.1f}" if rec.get('buy_balacc_pct') is not None else '—',
            f"{rec['sell_balacc_pct']:.1f}" if rec.get('sell_balacc_pct') is not None else '—',
            f"{rec['match_avg_pct']:.1f}" if rec.get('match_avg_pct') is not None else '—',
            f"{rec['match_buy_pct']:.1f}" if rec.get('match_buy_pct') is not None else '—',
            f"{rec['match_sell_pct']:.1f}" if rec.get('match_sell_pct') is not None else '—',
            rec.get('K_b_v_b', '—'),
            rec.get('K_s_v_s', '—'),
            rec.get('analyzed_at', '—'),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci); c.value = v; c.border = _TH
            c.alignment = Alignment(horizontal='center')
            c.font = Font(size=10, bold=(ci == 1 or ci == 3))
            if ri % 2 == 0: c.fill = _ALT
        sig_str = str(signal)
        for k, fl in sig_fills.items():
            if k in sig_str:
                ws.cell(ri, 3).fill = fl
                if '매수' in k: ws.cell(ri, 3).font = Font(bold=True, size=11, color='006100')
                elif '매도' in k: ws.cell(ri, 3).font = Font(bold=True, size=11, color='C00000')
                elif '충돌' in k: ws.cell(ri, 3).font = Font(bold=True, size=11, color='806000')
                break
        if is_err:
            ws.cell(ri, 2).fill = _BAD
            ws.cell(ri, 2).font = Font(bold=True, size=10, color='C00000')
        else:
            ws.cell(ri, 2).fill = _GOOD
            ws.cell(ri, 2).font = Font(bold=True, size=10, color='006100')
        if rec.get('cum_return_pct') is not None and rec.get('bh_pct') is not None:
            ws.cell(ri, 6).fill = _ret_fill(rec['cum_return_pct'] / 100.0, rec['bh_pct'] / 100.0)
            ws.cell(ri, 6).font = Font(bold=True, size=10,
                                          color='006100' if rec['cum_return_pct'] > 0 else 'C00000')
        if rec.get('up_cum_return_pct') is not None:
            ws.cell(ri, 9).font = Font(bold=True, size=10, color='006100')
        if rec.get('balacc_pct') is not None:
            ws.cell(ri, 14).fill = _success_fill(rec['balacc_pct'] / 100.0)
        if rec.get('match_avg_pct') is not None:
            ws.cell(ri, 17).fill = _success_fill(rec['match_avg_pct'] / 100.0)

    widths = [10, 10, 12, 14, 10, 11, 10, 10, 12, 12, 9, 8, 8, 9, 12, 12, 10, 9, 9, 11, 11, 18]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'B3'

    ws = wb.create_sheet('사용된 변수값'); ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = '티커별 사용된 변수값 (자동 튜닝 결과)'
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells('A1:N1')
    var_headers = ['티커', 'EVAL_START', 'HORIZON', 'DD한도', 'RU한도',
                    'ANCHOR_WIN', 'ANCHOR_LF', 'MIN_RISE', 'MIN_DROP',
                    'PRICE_TOL', 'wilson_z', 'pct_range', 'corr_limit',
                    'MIN_SIG', 'TOP_N', 'K_b/v_b', 'K_s/v_s', '튜닝상태']
    _hdr(ws, 2, var_headers)
    for ri, ticker in enumerate(all_tickers, 3):
        if ticker not in summary_records:
            ws.cell(ri, 1).value = ticker
            ws.cell(ri, 1).font = Font(bold=True, size=11)
            continue
        rec = summary_records[ticker]
        tv = rec.get('tuned_vars', {})
        is_tuned = rec.get('tuning_applied', False)
        tuning_src = rec.get('tuning_source', 'AUTO' if is_tuned else '기본값')
        kb = tv.get('K_buy'); vb = tv.get('vote_buy')
        ks = tv.get('K_sell'); vs = tv.get('vote_sell')
        if kb is not None and vb is not None:
            kbvb = f"{int(kb)}/{int(vb)}"
        else:
            kbvb = str(rec.get('K_b_v_b', '—'))
        if ks is not None and vs is not None:
            ksvs = f"{int(ks)}/{int(vs)}"
        else:
            ksvs = str(rec.get('K_s_v_s', '—'))
        vals = [ticker,
                str(tv.get('EVAL_START', '—')),
                tv.get('HORIZON_DAYS', '—'),
                f"{tv.get('DRAWDOWN_LIMIT_BUY', 0)*100:.2f}%" if tv.get('DRAWDOWN_LIMIT_BUY') else '—',
                f"{tv.get('RUNUP_LIMIT_SELL', 0)*100:.2f}%" if tv.get('RUNUP_LIMIT_SELL') else '—',
                tv.get('AUTO_ANCHOR_WINDOW', '—'),
                tv.get('AUTO_ANCHOR_LOOKFORWARD', '—'),
                f"{tv.get('AUTO_ANCHOR_MIN_RISE', 0)*100:.2f}%" if tv.get('AUTO_ANCHOR_MIN_RISE') else '—',
                f"{tv.get('AUTO_ANCHOR_MIN_DROP', 0)*100:.2f}%" if tv.get('AUTO_ANCHOR_MIN_DROP') else '—',
                f"{tv.get('AUTO_ANCHOR_PRICE_TOLERANCE', 0)*100:.2f}%" if tv.get('AUTO_ANCHOR_PRICE_TOLERANCE') else '—',
                tv.get('wilson_z', '—'),
                str(tv.get('pct_range', '—')),
                tv.get('corr_limit', '—'),
                tv.get('min_signals', '—') if tv.get('min_signals') is not None else '—',
                tv.get('top_n_pool', '—') if tv.get('top_n_pool') is not None else '—',
                kbvb,
                ksvs,
                tuning_src,
                ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci); c.value = v; c.border = _TH
            c.alignment = Alignment(horizontal='center')
            c.font = Font(size=10, bold=(ci == 1))
            if ri % 2 == 0: c.fill = _ALT
        status_col = len(var_headers)
        if tuning_src == 'AUTO':
            ws.cell(ri, status_col).fill = _GOOD
            ws.cell(ri, status_col).font = Font(bold=True, size=10, color='006100')
        elif tuning_src == 'EXCEL':
            ws.cell(ri, status_col).fill = _HOLD
            ws.cell(ri, status_col).font = Font(bold=True, size=10, color='0070C0')
        else:
            ws.cell(ri, status_col).fill = _MID
    var_widths = [10, 13, 9, 8, 8, 11, 11, 10, 10, 10, 9, 10, 10, 8, 7, 9, 9, 11]
    for ci, w in enumerate(var_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'B3'

    ws = wb.create_sheet('통계 분석'); ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = '티커별 변동성/통계 분석'
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells('A1:H1')
    stat_headers = ['티커', '일변동%', '주변동%', '5일평균변동%', '10일평균변동%',
                     '데이터일수', '5일평균낙폭%', '5일평균상승%']
    _hdr(ws, 2, stat_headers)
    for ri, ticker in enumerate(all_tickers, 3):
        if ticker not in summary_records:
            ws.cell(ri, 1).value = ticker
            continue
        rec = summary_records[ticker]
        s = rec.get('volatility_stats', {})
        vals = [ticker,
                f"{s.get('daily_vol_pct', 0):.2f}" if s else '—',
                f"{s.get('weekly_vol', 0)*100:.2f}" if s else '—',
                f"{s.get('typical_5d_move', 0):.2f}" if s else '—',
                f"{s.get('typical_10d_move', 0):.2f}" if s else '—',
                s.get('n_days', '—') if s else '—',
                f"{s.get('max_local_drawdown_5d', 0):.2f}" if s else '—',
                f"{s.get('max_local_runup_5d', 0):.2f}" if s else '—']
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci); c.value = v; c.border = _TH
            c.alignment = Alignment(horizontal='center')
            c.font = Font(size=10, bold=(ci == 1))
            if ri % 2 == 0: c.fill = _ALT
    for ci, w in enumerate([10, 11, 11, 14, 15, 12, 14, 14], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'B3'

    wb.save(summary_file)


def run_multi_ticker_analysis(tickers=None, *,
                               auto_tune_flags=None,
                               summary_file=None,
                               resume=True,
                               data_resolver=None,
                               per_ticker_overrides=None,
                               **run_kwargs):
    if tickers is None:        tickers = TICKERS
    if auto_tune_flags is None: auto_tune_flags = AUTO_TUNE
    if summary_file is None:    summary_file = os.path.join(SCRIPT_DIR, MULTI_SUMMARY_FILE)

    print('═' * 72)
    print(f'  ★ 다중 티커 분석 — {len(tickers)}개 종목')
    print(f'    티커: {tickers}')
    print(f'    이어서 진행: {resume}')
    on_vars = [k for k, v in auto_tune_flags.items() if v]
    off_vars = [k for k, v in auto_tune_flags.items() if not v]
    print(f'    자동 튜닝 ON ({len(on_vars)}): {on_vars}')
    if off_vars: print(f'    자동 튜닝 OFF ({len(off_vars)}): {off_vars}')
    print('═' * 72)

    existing, existing_order = _load_summary_state(summary_file)
    if existing:
        print(f"\n  ℹ 기존 요약 파일 발견 — {len(existing)}개 티커 기록 있음")
        print(f"    기록된 티커: {list(existing.keys())}")

    summary_records = {}
    for tk in existing:
        r = existing[tk]
        summary_records[tk] = {
            'status':    r.get('상태', '완료'),
            'signal':    r.get('신호', '—'),
            'position':  r.get('포지션', '—'),
            'close':     _try_float(str(r.get('오늘가', '')).replace('$','')),
            'cum_return_pct': _try_float(r.get('누적수익%')),
            'bh_pct':         _try_float(r.get('B&H%')),
            'up_cum_return_pct': _try_float(r.get('📈상승일전략%')),
            'bh_up_pct':         _try_float(r.get('📈상승일B&H%')),
            'vs_bh_pp':       _try_float(r.get('vsB&H%p')),
            'mdd_pct':        _try_float(r.get('MDD%')),
            'n_trades':       r.get('거래수'),
            'win_rate_pct':   _try_float(r.get('승률%')),
            'balacc_pct':     _try_float(r.get('BalAcc%')),
            'buy_balacc_pct': _try_float(r.get('매수BalAcc%')),
            'sell_balacc_pct':_try_float(r.get('매도BalAcc%')),
            'match_avg_pct':  _try_float(r.get('⚓매칭률%')),
            'match_buy_pct':  _try_float(r.get('⚓매수%')),
            'match_sell_pct': _try_float(r.get('⚓매도%')),
            'K_b_v_b':        r.get('K_b/v_b'),
            'K_s_v_s':        r.get('K_s/v_s'),
            'analyzed_at':    r.get('분석일시'),
            # ★ 변경5: tuned_vars 복원 (빈 dict 대신 실제 복원)
            'tuned_vars':     _restore_tuned_vars(r.get('_tuned_vars_raw')),
            'tuning_applied': bool(r.get('_tuned_vars_raw')),
            'tuning_source':  (r.get('_tuned_vars_raw') or {}).get('tuning_source', '기존유지'),
            'volatility_stats': {},
        }

    all_tickers_acc = list(existing_order) + [t for t in tickers if t not in existing]

    if data_resolver is None:
        data_resolver = _resolve_data_for_ticker

    t_total = time.time()
    n_done = sum(1 for tk in tickers if tk in existing)
    n_total = len(tickers)
    today_str = datetime.now().strftime('%Y-%m-%d')

    for ti, ticker in enumerate(tickers, 1):
        print(f"\n{'━' * 72}")
        print(f"  [{ti}/{n_total}] {ticker} 분석 시작")
        print('━' * 72)
        if resume and ticker in existing:
            print(f"  ⏩ 이미 완료됨 — 스킵")
            continue
        _free_global_caches()   # ★ 메모리 누적 방지 (Colab 여러 번 실행 끊김 예방)
        t_start = time.time()
        try:
            feat_t, close_t = data_resolver(ticker)
            if feat_t is None or close_t is None or len(close_t) < 30:
                raise RuntimeError(f"데이터 부족 또는 미로드 (n={len(close_t) if close_t is not None else 0})")

            stats = analyze_ticker_volatility(close_t)

            use_overrides = bool(per_ticker_overrides and ticker in per_ticker_overrides
                                  and per_ticker_overrides[ticker])
            if use_overrides:
                tuned = {}
                tuning_applied = False
                tuning_source = 'EXCEL'
                ov = per_ticker_overrides[ticker]
                print(f"\n  📋 Excel 변수값 모드 ({ticker}) — 자동 튜닝 OFF")
                for k, v in ov.items():
                    print(f"     {k:32}: {v}")
            else:
                tuned, _ = auto_tune_variables(close_t, feat_t, auto_tune_flags=auto_tune_flags)
                tuning_applied = bool(tuned)
                tuning_source = 'AUTO' if tuning_applied else '기본값'
                if tuning_applied:
                    print(f"\n  ⚙ 자동 튜닝 결과 ({ticker}):")
                    for k, v in tuned.items():
                        print(f"     {k:32}: {v}")
            print(f"  📊 통계: 일변동 {stats['daily_vol_pct']:.2f}%, "
                  f"5일변동 {stats['typical_5d_move']:.2f}%, n={stats['n_days']}")

            kwargs = dict(run_kwargs)
            if 'EVAL_START' in tuned and 'eval_start' not in kwargs:
                kwargs['eval_start'] = tuned['EVAL_START']
            mapping = {
                'HORIZON_DAYS':                'horizon',
                'DRAWDOWN_LIMIT_BUY':          'dd_limit',
                'RUNUP_LIMIT_SELL':            'ru_limit',
                'AUTO_ANCHOR_WINDOW':          'auto_anchor_window',
                'AUTO_ANCHOR_LOOKFORWARD':     'auto_anchor_lookforward',
                'AUTO_ANCHOR_MIN_RISE':        'auto_anchor_min_rise',
                'AUTO_ANCHOR_MIN_DROP':        'auto_anchor_min_drop',
                'AUTO_ANCHOR_PRICE_TOLERANCE': 'auto_anchor_price_tolerance',
            }
            for tk_k, rn_k in mapping.items():
                if tk_k in tuned and rn_k not in kwargs:
                    kwargs[rn_k] = tuned[tk_k]
            if any(k in tuned for k in ('wilson_z', 'pct_range', 'corr_limit')):
                base_mg = dict(META_GRID)
                if 'wilson_z'   in tuned: base_mg['wilson_z']   = [tuned['wilson_z']]
                if 'pct_range'  in tuned: base_mg['pct_range']  = [tuned['pct_range']]
                if 'corr_limit' in tuned: base_mg['corr_limit'] = [tuned['corr_limit']]
                kwargs['meta_grid'] = base_mg

            if use_overrides:
                ov = per_ticker_overrides[ticker]
                for k, v in ov.items():
                    if k == 'meta_grid':
                        kwargs['meta_grid'] = v
                    else:
                        kwargs[k] = v

            globals()['_pair_feat']   = feat_t
            globals()['_pair_close']  = close_t
            globals()['_pair_ticker'] = ticker

            individual_file = os.path.join(SCRIPT_DIR, f'ensemble_search_{ticker}_{today_str}.xlsx')
            kwargs.setdefault('output_file', individual_file)

            if (not use_overrides) and STAGED_META_TUNE:
                base_mg = kwargs.pop('meta_grid', META_GRID)
                print(f"\n  🔬 STAGED_META_TUNE ON ({ticker}) — 단계적 메타 변수 자동 튜닝 수행")
                result = staged_meta_tune(base_meta_grid=base_mg, **kwargs)
            else:
                result = run_ensemble_search(**kwargs)

            # ★ 통과 조합이 없어 staged가 None을 반환하면: 에러 없이 이 티커 스킵
            if result is None:
                print(f"  ⏭ {ticker}: 통과 조합 없음 → 엑셀 생성 스킵, 다음 티커로 진행")
                prev = summary_records.get(ticker)
                if prev is not None and prev.get('status') == '완료':
                    print(f"     (기존 완료 기록 유지)")
                else:
                    summary_records[ticker] = {
                        'status': '통과없음', 'signal': '— 통과조합없음', 'position': '—',
                        'close': None, 'analyzed_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'tuned_vars': {}, 'tuning_applied': False, 'volatility_stats': {},
                    }
                if globals().get('WRITE_SUMMARY_FILE', False):
                    try:
                        _save_summary_excel(summary_records, summary_file, all_tickers_acc)
                    except Exception:
                        pass
                continue

            (meta_results_df, inner_all, inner_passed,
             best_meta, best_inner, buy_pool, sell_pool,
             daily, trades, cur) = result

            bh_ret = _bh_sum_return(close_t.values) * 100.0
            bh_up_ret_t = _bh_up_sum_return(close_t.values) * 100.0
            if '보유' in cur['position']:
                if cur['sell_on_now'] and (not cur['buy_on_now'] or
                                            cur['sell_count_now']/max(cur['K_sell'],1) >= cur['buy_count_now']/max(cur['K_buy'],1)):
                    signal = '🔴 매도'
                elif cur['buy_on_now'] and cur['sell_on_now']:
                    signal = '⚔ 충돌'
                else:
                    signal = '📈 보유'
            else:
                if cur['buy_on_now'] and (not cur['sell_on_now'] or
                                            cur['buy_count_now']/max(cur['K_buy'],1) >= cur['sell_count_now']/max(cur['K_sell'],1)):
                    signal = '🟢 매수'
                elif cur['buy_on_now'] and cur['sell_on_now']:
                    signal = '⚔ 충돌'
                else:
                    signal = '💵 현금'

            match_avg  = best_inner.get('anchor_avg_match_rate',  None) if hasattr(best_inner, 'get') else None
            match_buy  = best_inner.get('anchor_buy_match_rate',  None) if hasattr(best_inner, 'get') else None
            match_sell = best_inner.get('anchor_sell_match_rate', None) if hasattr(best_inner, 'get') else None

            mg_used = kwargs.get('meta_grid', META_GRID)
            effective_vars = {
                'EVAL_START':                  kwargs.get('eval_start', EVAL_START),
                'HORIZON_DAYS':                kwargs.get('horizon', HORIZON_DAYS),
                'DRAWDOWN_LIMIT_BUY':          kwargs.get('dd_limit', DRAWDOWN_LIMIT_BUY),
                'RUNUP_LIMIT_SELL':            kwargs.get('ru_limit', RUNUP_LIMIT_SELL),
                'AUTO_ANCHOR_WINDOW':          kwargs.get('auto_anchor_window', AUTO_ANCHOR_WINDOW),
                'AUTO_ANCHOR_LOOKFORWARD':     kwargs.get('auto_anchor_lookforward', AUTO_ANCHOR_LOOKFORWARD),
                'AUTO_ANCHOR_MIN_RISE':        kwargs.get('auto_anchor_min_rise', AUTO_ANCHOR_MIN_RISE),
                'AUTO_ANCHOR_MIN_DROP':        kwargs.get('auto_anchor_min_drop', AUTO_ANCHOR_MIN_DROP),
                'AUTO_ANCHOR_PRICE_TOLERANCE': kwargs.get('auto_anchor_price_tolerance', AUTO_ANCHOR_PRICE_TOLERANCE),
                'wilson_z':   best_meta['wilson_z'],
                'pct_range':  (int(best_meta['pct_low']), int(best_meta['pct_high'])),
                'corr_limit': best_meta['corr_limit'],
                'min_signals':   int(best_meta.get('min_signals', 0)) if best_meta.get('min_signals') is not None else None,
                'top_n_pool':    int(best_meta.get('top_n_pool_buy', 0)) if best_meta.get('top_n_pool_buy') is not None else None,
                'K_buy':     int(best_inner['K_buy']),
                'vote_buy':  int(best_inner['vote_buy']),
                'K_sell':    int(best_inner['K_sell']),
                'vote_sell': int(best_inner['vote_sell']),
            }

            new_record = {
                'status': '완료',
                'signal': signal,
                'position': cur['position'],
                'close': cur['current_price'],
                'cum_return_pct': cur['cum_return_pct'],
                'bh_pct': bh_ret,
                'vs_bh_pp': cur['cum_return_pct'] - bh_ret,
                'up_cum_return_pct': cur.get('up_cum_return_pct', 0.0),
                'bh_up_pct': bh_up_ret_t,
                'mdd_pct': cur['max_drawdown'] * 100.0,
                'n_trades': cur['n_trades'],
                'win_rate_pct': cur['win_rate'] * 100.0,
                'balacc_pct':      cur['avg_success_rate']  * 100.0,
                'buy_balacc_pct':  cur['buy_success_rate']  * 100.0,
                'sell_balacc_pct': cur['sell_success_rate'] * 100.0,
                'match_avg_pct':  match_avg  * 100.0 if match_avg  is not None and pd.notna(match_avg)  else None,
                'match_buy_pct':  match_buy  * 100.0 if match_buy  is not None and pd.notna(match_buy)  else None,
                'match_sell_pct': match_sell * 100.0 if match_sell is not None and pd.notna(match_sell) else None,
                'K_b_v_b': f"{int(best_inner['K_buy'])}/{int(best_inner['vote_buy'])}",
                'K_s_v_s': f"{int(best_inner['K_sell'])}/{int(best_inner['vote_sell'])}",
                'analyzed_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'tuned_vars': effective_vars,
                'tuning_applied': tuning_applied,
                'tuning_source': tuning_source,
                'volatility_stats': stats,
            }

            # ★ 동일 티커 재실행 비교 — 그리드 선정과 같은 우선순위:
            #   매도성공률 → 평균성공률 → 누적수익 (밴드 SELECTION_TOLERANCE)
            prev = summary_records.get(ticker)
            _tol = SELECTION_TOLERANCE
            def _is_new_better(prev_rec, new_rec, tol=_tol):
                if prev_rec is None:
                    return True
                ps = prev_rec.get('sell_balacc_pct'); pa = prev_rec.get('balacc_pct'); pr = prev_rec.get('cum_return_pct')
                ns = new_rec.get('sell_balacc_pct');  na = new_rec.get('balacc_pct');  nr = new_rec.get('cum_return_pct')
                if ps is None or pa is None or pr is None:
                    return True
                if ns is None or na is None or nr is None:
                    return False
                tolp = tol * 100.0
                if ns > ps + tolp: return True
                if ns < ps - tolp: return False
                if na > pa + tolp: return True
                if na < pa - tolp: return False
                return nr > pr
            prev_ret = prev.get('cum_return_pct') if prev else None
            new_ret = new_record['cum_return_pct']
            if not _is_new_better(prev, new_record):
                print(f"  ℹ {ticker}: 기존(매도 {prev.get('sell_balacc_pct',0):.1f}%/평균 {prev.get('balacc_pct',0):.1f}%/수익 {prev_ret:+.2f}%) 우선 → 기존 유지")
                summary_records[ticker]['analyzed_at'] = new_record['analyzed_at']
            else:
                if prev_ret is not None:
                    print(f"  ✓ {ticker}: 신규(매도 {new_record.get('sell_balacc_pct',0):.1f}%/평균 {new_record.get('balacc_pct',0):.1f}%/수익 {new_ret:+.2f}%) 우선순위상 더 나음 → 갱신")
                summary_records[ticker] = new_record
            el = time.time() - t_start
            print(f"\n  ✓ {ticker} 완료 — {el:.1f}초, 수익 {cur['cum_return_pct']:+.2f}%, 신호 {signal}")
        except Exception as e:
            import traceback
            el = time.time() - t_start
            print(f"\n  ✗ {ticker} 에러 ({el:.1f}초): {e}")
            traceback.print_exc()
            prev = summary_records.get(ticker)
            if prev is not None and prev.get('status') == '완료':
                print(f"  ℹ {ticker}: 이번 분석 에러지만 기존 완료 기록 유지")
            else:
                summary_records[ticker] = {
                    'status': '에러',
                    'signal': '— 에러',
                    'position': '—',
                    'close': None,
                    'analyzed_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'error_msg': str(e),
                    'tuned_vars': {},
                    'tuning_applied': False,
                    'volatility_stats': {},
                }

        if globals().get('WRITE_SUMMARY_FILE', False):
            try:
                _save_summary_excel(summary_records, summary_file, all_tickers_acc)
                n_done += 1
                print(f"  💾 요약 저장됨 ({summary_file}) — {n_done}/{n_total} 진행")
            except Exception as e:
                print(f"  ⚠ 요약 저장 실패: {e}")
        else:
            n_done += 1

        # ★ 메모리 정리 — 한 종목이 끝나면 큰 객체를 비워 RAM 누적/폭증 방지.
        #   (73개 × 2800피처를 연속 돌리면 정리 없이는 RAM이 계속 쌓여 끊김)
        try: del result
        except Exception: pass
        try: del meta_results_df, inner_all, inner_passed, buy_pool, sell_pool, daily, trades
        except Exception: pass
        try: del feat_t, close_t
        except Exception: pass
        for _gk in ('_pair_feat', '_pair_close'):
            if _gk in globals():
                globals()[_gk] = None
        import gc as _gc
        _gc.collect()

    t_el = time.time() - t_total
    n_done_total = len([r for r in summary_records.values() if r.get('status') == '완료'])
    print(f"\n{'═' * 72}")
    print(f"  ★ 이번 실행 완료 — {n_total}개 처리  ({t_el/60:.1f}분 소요)")
    print(f"    완료: 전체 {len(all_tickers_acc)}개 티커 중 {n_done_total}개")
    print('═' * 72)

    if AUTO_DOWNLOAD_EXCEL:
        today_str = datetime.now().strftime('%Y-%m-%d')
        files_to_download = []   # ★ summary 파일 안 만듦 → 종목별 엑셀만
        for ticker in tickers:
            candidate = os.path.join(SCRIPT_DIR, f'ensemble_search_{ticker}_{today_str}.xlsx')
            if os.path.exists(candidate):
                files_to_download.append(candidate)
            else:
                import glob
                matches = sorted(glob.glob(os.path.join(SCRIPT_DIR, f'ensemble_search_{ticker}_*.xlsx')),
                                 key=os.path.getmtime, reverse=True)
                if matches:
                    files_to_download.append(matches[0])
        _auto_download_excels(files_to_download)

    return summary_records


def _auto_download_excels(file_paths, *, verbose=True):
    # 생성된 파일 경로를 전역에 누적 (노트북 셀에서 직접 다운로드할 수 있도록)
    g = globals()
    if '_GENERATED_FILES' not in g:
        g['_GENERATED_FILES'] = []
    for fp in file_paths:
        if fp and os.path.exists(fp) and fp not in g['_GENERATED_FILES']:
            g['_GENERATED_FILES'].append(fp)

    if not AUTO_DOWNLOAD_EXCEL:
        if verbose:
            print("  ℹ AUTO_DOWNLOAD_EXCEL=False — 자동 다운로드 건너뜀")
        return

    existing = [fp for fp in file_paths if fp and os.path.exists(fp)]
    missing  = [fp for fp in file_paths if fp and not os.path.exists(fp)]

    if verbose:
        print(f"\n{'─' * 72}")
        print(f"  📥 Excel 자동 다운로드 — {len(existing)}개 파일")
        print(f"{'─' * 72}")
        for fp in existing:
            sz_kb = os.path.getsize(fp) / 1024
            print(f"     ✓ {fp}  ({sz_kb:.1f} KB)")
        if missing:
            for fp in missing:
                print(f"     ⚠ 누락: {fp}")

    try:
        from google.colab import files
        is_colab = True
    except ImportError:
        is_colab = False

    if is_colab:
        for fp in existing:
            try:
                files.download(fp)
                if verbose:
                    print(f"     📥 다운로드 시작: {fp}")
            except Exception as e:
                print(f"     ⚠ 다운로드 실패 ({fp}): {e}")
        if verbose:
            print(f"  ✓ Colab 다운로드 트리거 완료\n")
    else:
        if verbose:
            print(f"  ℹ Colab 환경 아님 — 위 경로에서 직접 가져가세요\n")


def _try_float(x):
    if x is None: return None
    try:
        s = str(x).replace('%','').replace('+','').strip()
        return float(s)
    except Exception:
        return None


def _parse_overrides_from_excel(summary_file, tickers, *, verbose=True):
    if not os.path.exists(summary_file):
        if verbose: print(f"  ℹ 요약 파일 없음 — Excel 오버라이드 비활성")
        return {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(summary_file, data_only=True)
        if '사용된 변수값' not in wb.sheetnames:
            if verbose: print(f"  ⚠ '사용된 변수값' 시트 없음")
            return {}
        ws = wb['사용된 변수값']
        header_row = None
        for ri, row in enumerate(ws.iter_rows(values_only=True), 1):
            if ri == 2:
                header_row = list(row)
                break
        if not header_row:
            return {}

        kv_map = {}
        if '결과 요약' in wb.sheetnames:
            ws2 = wb['결과 요약']
            hdr2 = None
            for ri, row in enumerate(ws2.iter_rows(values_only=True), 1):
                if ri == 2:
                    hdr2 = list(row); continue
                if ri <= 2 or not row or row[0] is None: continue
                tk = str(row[0]).strip()
                if not tk or tk == '티커': continue
                d2 = dict(zip(hdr2, row)) if hdr2 else {}
                def _split_kv(s):
                    try:
                        a, b = str(s).split('/')
                        return int(float(a.strip())), int(float(b.strip()))
                    except Exception:
                        return None, None
                kb, vb = _split_kv(d2.get('K_b/v_b'))
                ks, vs = _split_kv(d2.get('K_s/v_s'))
                if kb is not None and ks is not None:
                    kv_map[tk] = (kb, vb, ks, vs)

        def _pct_to_ratio(v):
            if v is None: return None
            s = str(v).strip().replace('+','')
            if s == '—' or s == '': return None
            has_pct = '%' in s
            try:
                f = float(s.replace('%','').strip())
                return f / 100.0 if has_pct else f
            except Exception:
                return None

        def _maybe_int(v):
            if v is None: return None
            try:
                s = str(v).strip()
                if s == '—' or s == '': return None
                return int(float(s))
            except Exception:
                return None

        def _maybe_float(v):
            if v is None: return None
            try:
                s = str(v).strip()
                if s == '—' or s == '': return None
                return float(s)
            except Exception:
                return None

        def _parse_pct_range(v):
            if v is None: return None
            s = str(v).strip()
            if s == '—' or s == '': return None
            try:
                s = s.replace('(','').replace(')','').replace('[','').replace(']','')
                parts = [int(float(x.strip())) for x in s.split(',') if x.strip()]
                if len(parts) == 2: return tuple(parts)
            except Exception:
                return None
            return None

        result = {}
        for ri, row in enumerate(ws.iter_rows(values_only=True), 1):
            if ri <= 2: continue
            if not row or row[0] is None: continue
            ticker = str(row[0]).strip()
            if ticker not in tickers: continue
            d = dict(zip(header_row, row))

            ov = {}
            v = d.get('EVAL_START')
            if v not in (None, '—', ''): ov['eval_start'] = str(v).strip()
            v = _maybe_int(d.get('HORIZON'))
            if v is not None: ov['horizon'] = v
            v = _pct_to_ratio(d.get('DD한도'))
            if v is not None: ov['dd_limit'] = v
            v = _pct_to_ratio(d.get('RU한도'))
            if v is not None: ov['ru_limit'] = v
            v = _maybe_int(d.get('ANCHOR_WIN'))
            if v is not None: ov['auto_anchor_window'] = v
            v = _maybe_int(d.get('ANCHOR_LF'))
            if v is not None: ov['auto_anchor_lookforward'] = v
            v = _pct_to_ratio(d.get('MIN_RISE'))
            if v is not None: ov['auto_anchor_min_rise'] = v
            v = _pct_to_ratio(d.get('MIN_DROP'))
            if v is not None: ov['auto_anchor_min_drop'] = v
            v = _pct_to_ratio(d.get('PRICE_TOL'))
            if v is not None: ov['auto_anchor_price_tolerance'] = v

            mg_part = {}
            v = _maybe_float(d.get('wilson_z'))
            if v is not None: mg_part['wilson_z'] = [v]
            v = _parse_pct_range(d.get('pct_range'))
            if v is not None: mg_part['pct_range'] = [v]
            v = _maybe_float(d.get('corr_limit'))
            if v is not None: mg_part['corr_limit'] = [v]
            v = _maybe_int(d.get('MIN_SIG'))
            if v is not None: mg_part['min_signals'] = [v]
            v = _maybe_int(d.get('TOP_N'))
            if v is not None: mg_part['top_n_pool'] = [v]
            if mg_part:
                mg = dict(META_GRID)
                mg.update(mg_part)
                ov['meta_grid'] = mg

            kb = vb = ks = vs = None
            def _split_kv_local(s):
                try:
                    a, b = str(s).split('/')
                    return int(float(a.strip())), int(float(b.strip()))
                except Exception:
                    return None, None
            kb, vb = _split_kv_local(d.get('K_b/v_b'))
            ks, vs = _split_kv_local(d.get('K_s/v_s'))
            if (kb is None or ks is None) and ticker in kv_map:
                kb, vb, ks, vs = kv_map[ticker]
            if kb is not None and ks is not None:
                ov['k_buy_range']  = [kb]
                ov['k_sell_range'] = [ks]
                ov['vote_ratio_buy']  = [vb / kb] if kb > 0 else [1.0]
                ov['vote_ratio_sell'] = [vs / ks] if ks > 0 else [1.0]

            if ov:
                result[ticker] = ov
        return result
    except Exception as e:
        if verbose: print(f"  ⚠ Excel 변수 파싱 실패: {e}")
        import traceback; traceback.print_exc()
        return {}


def _resolve_data_for_ticker(ticker, end_date=None):
    g = globals()
    pdm = g.get('_pair_data_map')
    if isinstance(pdm, dict) and ticker in pdm:
        entry = pdm[ticker]
        if isinstance(entry, tuple) and len(entry) == 2:
            return entry[0], entry[1]
        if isinstance(entry, dict):
            return entry.get('feat'), entry.get('close')

    if 'load_ticker_data' in g and callable(g['load_ticker_data']):
        try:
            return g['load_ticker_data'](ticker)
        except Exception as e:
            print(f"  ⚠ load_ticker_data({ticker}) 실패: {e}")

    cache = g.setdefault('_multi_ticker_cache', {})
    if ticker in cache:
        print(f"  ℹ 캐시 사용: {ticker}")
        return cache[ticker]

    needed = ['download_data', 'compute_features']
    if all(n in g for n in needed):
        try:
            start = g.get('DOWNLOAD_START', '2020-01-01')

            old_ticker = g.get('TICKER')
            g['TICKER'] = ticker
            print(f"  ⚙ TICKER 글로벌 변경: {old_ticker} → {ticker}")

            dl_func = g['download_data']
            dl_varnames = dl_func.__code__.co_varnames[:dl_func.__code__.co_argcount]
            try:
                if 'tickers' in dl_varnames:
                    closes, ohlcv = dl_func(tickers=[ticker], start=start)
                else:
                    closes, ohlcv = dl_func(start=start)
            except TypeError:
                closes, ohlcv = dl_func(start=start)

            # ★ 최신 거래일 보충 (요청) — download_data 데이터가 '완료된 세션을 빠뜨릴 만큼'
            #   오래됐을 때만(오늘보다 2영업일 이상 전) yfinance로 다시 받는다.
            #   오늘 17일에 데이터가 16일까지면 = 어제(최신 완료 세션)이므로 정상 → 재조회 안 함.
            #   (1영업일 전은 '오늘장 미마감'이라 정상이므로 건드리지 않음)
            _did_yf_refetch = False
            _orig_ohlcv_t = ohlcv.get(ticker)      # 보충 실패 시 되돌릴 원본
            _orig_close_t = closes.get(ticker) if hasattr(closes, 'get') else None
            _need_yf = ticker not in ohlcv
            if not _need_yf and end_date is None:
                try:
                    _last0 = ohlcv[ticker].index[-1]
                    _today0 = pd.Timestamp(datetime.now().date())
                    if len(pd.bdate_range(_last0.normalize(), _today0)) - 1 >= 2:
                        _need_yf = True
                        print(f"  ℹ {ticker} download_data 데이터가 {str(_last0)[:10]}까지 — "
                              f"최신 거래일 보충 위해 yfinance 재조회")
                except Exception:
                    pass
            if _need_yf:
                if ticker not in ohlcv:
                    print(f"  ⚠ {ticker}가 download_data 결과에 없음 — yfinance 직접 호출")
                try:
                    import yfinance as yf
                    # ★ end를 '내일'로 명시 — yfinance는 end 미지정 시 당일을 빠뜨릴 수 있음.
                    #   (특히 장 마감 전/시간대 차이로 전날까지만 받아지는 문제 완화)
                    _end = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')
                    df = yf.download(ticker, start=start, end=_end,
                                     progress=False, auto_adjust=False)
                    if df is None or len(df) == 0:
                        raise RuntimeError(f"yfinance에서 {ticker} 데이터 0건")
                    if isinstance(df.columns, pd.MultiIndex):
                        df.columns = df.columns.get_level_values(0)
                    ohlcv[ticker] = df
                    closes[ticker] = df['Close'] if 'Close' in df.columns else df.iloc[:, 0]
                    _did_yf_refetch = True
                    _last = df.index[-1]
                    _today = pd.Timestamp(datetime.now().date())
                    _bdays = len(pd.bdate_range(_last.normalize(), _today)) - 1
                    print(f"  ✓ yfinance로 {ticker} 다운로드 성공 ({len(df)}일, 마지막 {str(_last)[:10]})")
                    if _bdays >= 1:
                        print(f"  ⚠ 최신 거래일 누락 가능 — 마지막 데이터가 {str(_last)[:10]}로 "
                              f"오늘({_today.date()})보다 {_bdays}영업일 전입니다.")
                        print(f"     원인: 미국장 마감 전 실행 or yfinance 반영 지연. "
                              f"장 마감 후(한국시간 다음날 오전 이후) 다시 받으면 최신일이 포함됩니다.")
                except Exception as _yf_e:
                    if ticker not in ohlcv:
                        raise RuntimeError(
                            f"{ticker}가 download_data 결과에 없고 yfinance 보충도 실패했습니다: {_yf_e}\n"
                            f"  download_data 함수의 티커 리스트에 '{ticker}'를 추가하거나,\n"
                            f"  yfinance 설치: pip install yfinance")
                    else:
                        print(f"  ⚠ 최신 거래일 보충 실패({_yf_e}) — "
                              f"download_data 데이터({str(ohlcv[ticker].index[-1])[:10]}까지)로 진행")

            fred_df = None
            if 'download_fred_data' in g:
                try: fred_df = g['download_fred_data'](start=start)
                except Exception: fred_df = None
                if fred_df is not None and len(fred_df) == 0: fred_df = None

            cf_func = g['compute_features']
            def _call_cf():
                try:
                    return cf_func(ohlcv, closes, fred_df=fred_df, ticker=ticker)
                except TypeError:
                    return cf_func(ohlcv, closes, fred_df=fred_df)

            # ★ 연장용 전체 계산 (요청, 기본 OFF) — 자르기 '전'에 새 거래일 행을 확보.
            #   ⚠ 반드시 '복사본'으로 계산한다. 그래야 compute_features가 입력을 건드려도
            #   아래 '자른 정확 재현' 계산이 오염되지 않는다 (과거 신호가 어긋나던 문제 차단).
            feat_full = close_full = None
            if end_date is not None and globals().get('REPLAY_EXTEND_TO_TODAY', False):
                try:
                    _oh = {t: (df.copy() if hasattr(df, 'copy') else df) for t, df in ohlcv.items()}
                    _cl = closes.copy() if hasattr(closes, 'copy') else dict(closes)
                    _fr = fred_df.copy() if fred_df is not None else None
                    try:
                        feat_full = cf_func(_oh, _cl, fred_df=_fr, ticker=ticker)
                    except TypeError:
                        feat_full = cf_func(_oh, _cl, fred_df=_fr)
                    if ticker in _oh:
                        close_full = _oh[ticker]['Close'].reindex(feat_full.index)
                    else:
                        feat_full = None
                except Exception as _fe:
                    print(f"  ⚠ 연장용 전체 계산 실패(연장 생략): {_fe}")
                    feat_full = close_full = None

            # ★ 원본 범위로 자르기 (요청) — end_date까지로 잘라 '정확 재현'. 전체구간 순위/정규화
            #   지표가 원본과 '같은 범위'로 계산돼 신호가 정확히 일치 (하루 더 받아 순위 바뀌던 문제 해결).
            if end_date is not None:
                try:
                    _ed = pd.Timestamp(end_date)
                    for _t in list(ohlcv.keys()):
                        try: ohlcv[_t] = ohlcv[_t].loc[ohlcv[_t].index <= _ed]
                        except Exception: pass
                    if hasattr(closes, 'columns'):              # DataFrame
                        closes = closes.loc[closes.index <= _ed]
                    elif isinstance(closes, dict):
                        for _t in list(closes.keys()):
                            try: closes[_t] = closes[_t].loc[closes[_t].index <= _ed]
                            except Exception: pass
                    if fred_df is not None:
                        try: fred_df = fred_df.loc[fred_df.index <= _ed]
                        except Exception: pass
                    print(f"  ✂ 원본 범위로 자름: {str(_ed)[:10]}까지 — 같은 범위로 계산해 정확 재현")
                except Exception as _te:
                    print(f"  ⚠ 원본 범위 자르기 실패(무시): {_te}")

            try:
                feat = _call_cf()
            except Exception as _cf_e:
                # ★ yfinance 보충 데이터가 타 종목(download_data)과 인덱스가 안 맞아 깨진 경우
                #   ('Can only compare identically-labeled Series' 등) → 원본으로 되돌려 재계산.
                if _did_yf_refetch and _orig_ohlcv_t is not None:
                    print(f"  ⚠ yfinance 보충 데이터로 피처 계산 실패 — "
                          f"download_data 원본({str(_orig_ohlcv_t.index[-1])[:10]}까지)으로 되돌려 재계산")
                    print(f"     (사유: {_cf_e})")
                    ohlcv[ticker] = _orig_ohlcv_t
                    if _orig_close_t is not None:
                        closes[ticker] = _orig_close_t
                    feat = _call_cf()
                else:
                    raise

            if ticker not in ohlcv:
                raise RuntimeError(f"{ticker} OHLCV 없음 (yfinance fallback 실패)")
            close = ohlcv[ticker]['Close'].reindex(feat.index)

            # ★ 새 거래일 덧붙이기 (요청) — 원본 마지막날 '이후' 행을 전체계산본에서 가져와 이어붙인다.
            #   → 과거는 정확 재현(자른 것), 새 거래일은 같은 로직으로 계산한 결과.
            if feat_full is not None and close_full is not None and end_date is not None:
                try:
                    _ed = pd.Timestamp(end_date)
                    _newidx = feat_full.index[feat_full.index > _ed]
                    if len(_newidx) > 0:
                        feat = pd.concat([feat, feat_full.reindex(columns=feat.columns).loc[_newidx]])
                        close = pd.concat([close, close_full.loc[_newidx]])
                        print(f"  ➕ 원본 이후 {len(_newidx)}일({str(_newidx[0])[:10]}~{str(_newidx[-1])[:10]}) "
                              f"같은 로직으로 이어 계산 → 최종 {str(feat.index[-1])[:10]}까지")
                    else:
                        print(f"  ℹ 원본 이후 새 거래일이 데이터에 없음 — 원본 날짜까지만 (연장할 데이터 없음)")
                except Exception as _xe:
                    print(f"  ⚠ 새 거래일 덧붙이기 실패(무시): {_xe}")

            cache[ticker] = (feat, close)
            return feat, close
        except Exception as e:
            raise RuntimeError(f"데이터 로드 실패 ({ticker}): {e}")

    raise RuntimeError(
        f"{ticker} 데이터를 로드할 수 없습니다. 다음 중 하나를 준비하세요:\n"
        f"  1. _pair_data_map = {{'AAPL': (feat, close), 'MPC': (feat, close), ...}}\n"
        f"  2. def load_ticker_data(ticker): return feat, close  # 글로벌 함수\n"
        f"  3. download_data / compute_features 함수 (기존 파이프라인)")


# ════════════════════════════════════════════════════════════════
#   ★ 모드 4 — 드라이브 폴더의 티커별 ★최적 조합을 '현재까지' 일괄 재현 + 요약 (요청)
# ════════════════════════════════════════════════════════════════
def _build_selected_summary(source_paths, out_dir, date_label, *, sheet_name='내부_그리드_통과'):
    """각 소스 엑셀의 내부 그리드 ★(선정) 행을 읽어 티커별 요약 엑셀을 만든다.
       (스트리밍 읽기 + ★ 없으면 건너뜀)."""
    import re as _re
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    FRE = _re.compile(r'ensemble_search_(.+)_(\d{4}-\d{2}-\d{2})\.xlsx$')

    def _read_star(path):
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            print(f"    ⚠ 열기 실패: {os.path.basename(path)} — {e}"); return None, None
        if sheet_name not in wb.sheetnames:
            print(f"    ⚠ '{sheet_name}' 시트 없음: {os.path.basename(path)}"); wb.close(); return None, None
        ws = wb[sheet_name]; header = None; sel = None; saw = False
        for row in ws.iter_rows(values_only=True):
            if header is None:
                sv = [str(v).strip() if v is not None else '' for v in row]
                if '#' in sv and 'K_buy' in sv:
                    header = [s if s else f'col{i+1}' for i, s in enumerate(sv)]
                continue
            c0 = row[0] if row else None
            if c0 is None: continue
            saw = True
            if '★' in str(c0): sel = list(row); break
        wb.close()
        if sel is None:
            print(f"    – ★ 없음 — 건너뜀: {os.path.basename(path)}" if (header and saw)
                  else f"    ⚠ 헤더/데이터 없음 — 건너뜀: {os.path.basename(path)}")
            return None, None
        return header, sel

    rows = []; master = None
    for p in source_paths:
        m = FRE.search(os.path.basename(p)); tk = m.group(1) if m else os.path.basename(p)
        dt = m.group(2) if m else ''
        header, vals = _read_star(p)
        if header is None: continue
        if master is None: master = header
        rec = {'티커': tk, '실행일': dt}
        rec.update({h: v for h, v in zip(header, vals)})
        rows.append(rec)
        print(f"    ✓ {tk:<6} ★행 읽음 (그리드 #{str(vals[0]).replace('★','').strip()})")
    if not rows:
        print("    ❌ 요약할 ★행이 없습니다."); return None

    out_cols = ['티커', '실행일'] + master
    wb = Workbook(); ws = wb.active; ws.title = '티커별_선정요약'; ws.sheet_view.showGridLines = False
    THIN = Side(style='thin', color='D9D9D9'); BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HF = PatternFill('solid', fgColor='1F3864'); ALT = PatternFill('solid', fgColor='F2F2F2')
    POS = PatternFill('solid', fgColor='E8F5E9'); NEG = PatternFill('solid', fgColor='FFEBEE')
    ws.cell(1, 1).value = f'티커별 선정(★) 요약 — {date_label} — {len(rows)}개 티커'
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out_cols))
    for ci, h in enumerate(out_cols, 1):
        c = ws.cell(3, ci); c.value = h; c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = HF; c.border = BRD; c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    pos_i = out_cols.index('📍실행일포지션') + 1 if '📍실행일포지션' in out_cols else None
    ret_i = [out_cols.index(c) + 1 for c in ('🔧보정후수익', '✅실제누적수익') if c in out_cols]
    for ri, rec in enumerate(rows):
        r = ri + 4
        for ci, h in enumerate(out_cols, 1):
            c = ws.cell(r, ci); c.value = rec.get(h, '—'); c.border = BRD; c.font = Font(size=10)
            c.alignment = Alignment(horizontal='center')
            if ri % 2 == 1: c.fill = ALT
        ws.cell(r, 1).font = Font(bold=True, size=11, color='1F3864')
        if pos_i is not None:
            pv = str(rec.get('📍실행일포지션', ''))
            if '보유' in pv: ws.cell(r, pos_i).fill = POS
            elif '현금' in pv: ws.cell(r, pos_i).fill = NEG
        for rx in ret_i: ws.cell(r, rx).font = Font(bold=True, size=10, color='C00000')
    for ci, h in enumerate(out_cols, 1):
        w = 8 if h == '티커' else (12 if h == '실행일' else 12)
        if h in ('🎬실행일액션', '📅최근매수일', '📅최근매도일', '🔧추가지표수'): w = 14
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 34; ws.freeze_panes = 'C4'
    out_path = os.path.join(out_dir, f'summary_selected_{date_label}.xlsx')
    wb.save(out_path)
    print(f"  ✅ 요약 저장: {out_path}  ({len(rows)}개 티커)")
    globals().setdefault('_GENERATED_FILES', []).append(out_path)
    return out_path


def run_mode4_drive_reproduce_all(drive_dir=None, *, date_subfolder=None, **override_kwargs):
    """모드 4 — 드라이브 ensemble_analysis 폴더에서 '티커별 가장 최근' 분석 엑셀을 모두 찾아,
       각 엑셀의 ★최적(내부 그리드 별표) 조합을 '현재까지' 일별 백테스트로 그대로 재현해
       새 엑셀을 만들고, 티커별 선정행 요약 엑셀도 만든다. 모든 결과물은 '날짜 폴더'에 저장."""
    import glob, re as _re
    g = globals()
    src = drive_dir or g.get('RUN_MODE4_DRIVE_DIR', '/content/drive/MyDrive/ensemble_analysis')
    if not os.path.isdir(src):
        print(f"  ✗ 드라이브 폴더를 찾을 수 없습니다: {src}")
        print(f"    먼저 드라이브 마운트: from google.colab import drive; drive.mount('/content/drive')")
        return []
    # 티커별 가장 최근 파일 (재현본 __replay 제외)
    latest = {}
    for p in glob.glob(os.path.join(src, "ensemble_search_*_*.xlsx")):
        name = os.path.basename(p)
        if '__replay' in name: continue
        m = _re.search(r"ensemble_search_(.+)_(\d{4}-\d{2}-\d{2})\.xlsx$", name)
        if not m: continue
        tk, dt = m.group(1), m.group(2)
        if tk not in latest or dt > latest[tk][0]:
            latest[tk] = (dt, p)
    if not latest:
        print(f"  ✗ {src} 에서 ensemble_search_<티커>_<날짜>.xlsx 형식 파일을 못 찾았습니다.")
        return []
    print(f"  📂 드라이브 폴더: {src}  — 티커 {len(latest)}개")

    today = datetime.now().strftime('%Y-%m-%d')
    sub = date_subfolder or today
    out_dir = os.path.join(src, sub)            # ★ 드라이브 안에 '날짜 폴더' 생성
    os.makedirs(out_dir, exist_ok=True)
    print(f"  📁 출력 날짜 폴더: {out_dir}\n")

    # 각 티커 ★최적 조합을 '현재까지' 재현 → 날짜 폴더에 엑셀 생성
    for tk, (dt, path) in sorted(latest.items()):
        print(f"  ─ [{tk}] {os.path.basename(path)} → 현재까지 재현 ─")
        # ★ 핵심 — 티커마다 '현재 날짜까지' 데이터를 새로 받도록 메모리 캐시를 비운다.
        #   안 비우면 이전 실행/이전 티커의 옛 데이터를 재사용해 원본과 같은 날짜까지만 나옴.
        for _k in ('_pair_feat', '_pair_close', '_pair_ticker'):
            g[_k] = None
        _mc = g.get('_multi_ticker_cache')
        if isinstance(_mc, dict):
            _mc.pop(tk, None)
        _pdm = g.get('_pair_data_map')
        if isinstance(_pdm, dict):
            _pdm.pop(tk, None)
        _free_global_caches(keep_pool_map=False)   # z-캐시 등도 정리
        try:
            # feat=None,close=None 명시 → 반드시 티커로 '신선한' 데이터를 새로 확보
            replay_grid_combo(path, None, output_dir=out_dir,
                              feat=None, close=None, **override_kwargs)
        except Exception as e:
            print(f"    ⚠ {tk} 재현 실패(건너뜀): {e}")

    # 티커별 선정행 요약 엑셀 (소스 ★행 기준) — 날짜 폴더에 저장
    print(f"\n  📑 티커별 선정행 요약 생성...")
    _build_selected_summary([p for _, (_, p) in sorted(latest.items())], out_dir, today)
    return g.get('_GENERATED_FILES', [])


def main():
    """RUN_* 변수(코드 상단 또는 노트북 셀에서 mod.RUN_MODE=... 로 변경)에 따라 실행.
       노트북에서: 모듈 로드 → mod.RUN_MODE 등 설정 → mod.main() 호출."""
    g = globals()
    g['_GENERATED_FILES'] = []   # 이번 실행에서 만든 엑셀 경로 누적
    _free_global_caches()        # ★ 이전 실행의 캐시 정리 (Colab 반복 실행 시 메모리 누적·끊김 방지)
    _mode = str(g.get('RUN_MODE', 3)).strip()
    print(f"\n[실행 모드] RUN_MODE = {_mode}")
    print("  1=새 분석  /  2=그리드 번호 재현  /  3=최근 엑셀 ★최적 자동 재현")

    if _mode == '2':
        print(f"\n  파일 경로는 OUTPUT_DIR 고정: {OUTPUT_DIR}")
        _fn = str(g.get('RUN_REPLAY_FILE', '')).strip().strip('"').strip("'")
        _gn = str(g.get('RUN_REPLAY_GRID_NUMBER', '')).strip()
        print(f"  재현 파일: {_fn}  /  그리드 번호: {_gn}")
        print("  ℹ 데이터가 메모리에 없으면 티커로 자동 다운로드를 시도합니다")
        try:
            replay_grid_combo(_fn, _gn)
        except Exception as _e:
            print(f"\n  ✗ 재현 실패: {_e}")
            print("    feat/close가 메모리에 있는지 확인하세요. 예:")
            print("      replay_grid_combo('파일명.xlsx', 14, feat=내_feat, close=내_close)")

    elif _mode == '1':
        # 분석 대상 티커 — RUN_TICKERS 가 있으면 그것, 없으면 기본 TICKERS
        _run_tickers = g.get('RUN_TICKERS', None)
        if _run_tickers:
            tickers_to_run = [str(t).strip().upper() for t in _run_tickers if str(t).strip()]
        else:
            tickers_to_run = TICKERS
        print(f"\n→ 분석 대상: {tickers_to_run}")

        summary_path = os.path.join(SCRIPT_DIR, MULTI_SUMMARY_FILE)
        overrides = None
        if g.get('RUN_USE_EXCEL_OVERRIDES', False) and os.path.exists(summary_path):
            print(f"\n기존 요약 파일에서 변수값 사용 시도: {summary_path}")
            overrides = _parse_overrides_from_excel(summary_path, tickers_to_run)
            if overrides:
                found = list(overrides.keys())
                missing = [t for t in tickers_to_run if t not in overrides]
                print(f"  ✓ Excel 변수값 적용: {found}")
                if missing:
                    print(f"  ℹ Excel에 없는 티커 (기본값 사용): {missing}")
            else:
                print(f"  ⚠ Excel에서 사용 가능한 변수값 못 찾음 → 기본값 사용")
                overrides = None
        else:
            print(f"→ 코드 상단 변수값(또는 AUTO_TUNE) 사용")

        run_multi_ticker_analysis(
            tickers=tickers_to_run,
            per_ticker_overrides=overrides,
            resume=False,
        )

    elif _mode == '4':
        # ── 모드 4 — 드라이브 폴더의 티커별 ★최적 조합을 '현재까지' 일괄 재현 + 요약 ──
        print(f"\n  드라이브 폴더의 티커별 가장 최근 분석 엑셀을 모두 찾아,")
        print(f"  각 ★최적 조합을 '현재까지' 재현하고 티커별 요약 엑셀을 만듭니다 (날짜 폴더에 저장).")
        print(f"  ℹ 드라이브가 마운트돼 있어야 합니다: from google.colab import drive; drive.mount('/content/drive')")
        try:
            run_mode4_drive_reproduce_all(
                g.get('RUN_MODE4_DRIVE_DIR', None),
                date_subfolder=g.get('RUN_MODE4_DATE_FOLDER', None))
        except Exception as _e:
            print(f"\n  ✗ 모드 4 실패: {_e}")

    else:
        # ── 모드 3 (기본) — 티커의 가장 최근 분석 엑셀에서 ★최적 조합 자동 재현 ──
        print(f"\n  엑셀 폴더(OUTPUT_DIR): {OUTPUT_DIR}")
        _tk = str(g.get('RUN_REPLAY_TICKER', '')).strip().upper()
        print(f"  '{_tk}'의 '가장 최근 일반 분석 엑셀'을 자동으로 찾아")
        print(f"  ★최적으로 선정된 변수·지표를 그대로 재현합니다 (재현본 __replay 파일은 제외).")
        if not _tk:
            print("  ⚠ RUN_REPLAY_TICKER 가 비어 종료합니다.")
        else:
            print("\n  ℹ 데이터가 메모리에 없으면 티커로 자동 다운로드를 시도합니다.")
            try:
                replay_latest_best(_tk)
            except Exception as _e:
                print(f"\n  ✗ 재현 실패: {_e}")

    # ── 생성된 파일 안내 + 셀에서 직접 다운로드할 수 있도록 경로 반환 ──
    made = g.get('_GENERATED_FILES', [])
    if made:
        print(f"\n{'═'*72}")
        print(f"  📦 생성된 엑셀 {len(made)}개:")
        for fp in made:
            print(f"     • {fp}")
        print(f"  💡 자동 다운로드가 안 뜨면, 노트북에서 아래로 직접 받으세요:")
        print(f"       from google.colab import files")
        print(f"       for f in mod.get_generated_files(): files.download(f)")
        print(f"{'═'*72}")
    return made


def get_generated_files():
    """이번 실행에서 만든 엑셀 파일 경로 목록 (노트북 셀에서 직접 다운로드용)."""
    return list(globals().get('_GENERATED_FILES', []))


if __name__ == '__main__':
    main()
