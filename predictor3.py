# @title
"""
XLK 하락 예측 임계치 탐색기 (완전 독립 실행)
==================================================
데이터 다운로드 → 450개 지표 계산 → (n, m) 그리드 탐색 →
450개 지표별 최적 임계치 → Train/Test 검증 → Excel 출력

모든 로직이 이 한 파일에 포함되어 있어 다른 파일 없이 바로 실행 가능.

의존성:  pip install yfinance pandas openpyxl numpy tqdm
실행:    python xlk_drop_predictor.py
출력:    xlk_drop_predictor_result.xlsx
"""
import warnings; warnings.filterwarnings('ignore')
import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from tqdm import tqdm
except ImportError:
    class _DummyBar:
        def __init__(self, total=None, desc=None, **kw):
            self.total = total or 0; self.n = 0; self.desc = desc or ''
        def update(self, k=1):
            self.n += k
            if self.total:
                pct = 100 * self.n / self.total
                print(f"\r  {self.desc} {self.n}/{self.total} ({pct:.0f}%)",
                      end='', flush=True)
        def close(self): print()
        def __iter__(self): return iter([])
    def tqdm(it=None, total=None, desc=None, **kw):
        return it if it is not None else _DummyBar(total=total, desc=desc, **kw)


# ════════════════════════════════════════════════════════════════
#                         설정
# ════════════════════════════════════════════════════════════════

# 11개 SPDR 섹터 ETF 전체 + 시장지수 + 매크로 자산
PEERS          = [
    # 11개 섹터 ETF (S&P 500 GICS 분류)
    'XLK',  # Technology
    'XLV',  # Health Care
    'XLF',  # Financials
    'XLY',  # Consumer Discretionary
    'XLP',  # Consumer Staples
    'XLE',  # Energy
    'XLI',  # Industrials
    'XLB',  # Materials
    'XLU',  # Utilities
    'XLRE', # Real Estate
    'XLC',  # Communication Services
    # 시장 지수 / 스타일
    'SPY', 'QQQ', 'IWM', 'DIA',
    # 테크 하위
    'SMH', 'SOXX', 'VGT', 'IGV',
    # 매크로 자산
    '^VIX', '^TNX', '^IRX',
    'GLD',          # 금
    'TLT',          # 장기국채
    'HYG',          # 하이일드 회사채
    'LQD',          # 투자등급 회사채
    'UUP',          # 달러
    'USO',          # 원유
    'DBC',          # 종합 원자재
    'SPLV',         # 저변동성
    'SPHB',         # 고베타
        # ★ 신규 추가: 경기선행/방산/퀄리티 개별주
    'PCAR',   # Paccar (트럭 제조 — 화물·경기 선행)
    'LMT',    # Lockheed Martin (방산 — 위험회피 순환)
    # ★ 신규 추가: 퀄리티·배당 팩터 ETF
    'NOBL',   # S&P500 배당귀족 (퀄리티 로테이션)
    'VIG',    # Vanguard 배당성장
    'QUAL',   # iShares MSCI Quality Factor
    # ★ 신규 추가: 금리·채권 심화
    'IEF',    # 7-10년 국채
    'BIL',    # 1-3개월 단기채 (무위험수익률 프록시)
    'TIP',    # TIPS (물가연동채 — 인플레 기대)
    'SHY',    # 1-3년 국채
    '^FVX',   # 5년 국채 금리
    '^TYX',   # 30년 국채 금리
    # ★ 신규 추가: 신용 스트레스
    'KRE',    # 지역은행 ETF (신용 경색 선행)
    'JNK',    # SPDR HY채권 (HYG 보완)
    # ★ 신규 추가: 경기선행 섹터
    'XHB',    # 주택건설업 (금리 민감)
    'IYT',    # 운송 (경제활동 온도계)
    'XRT',    # 소매 (소비자 심리 반영)
    'XME',    # 금속·광업 (원자재 수요)
    # ★ 신규 추가: 인플레·원자재 심화
    'CPER',   # 구리 ETF (성장 선행지표)
    'PDBC',   # 다각화 원자재
    'RINF',   # 인플레이션 기대 ETF
    # ★ 신규 추가: 변동성 구조
    'VXX',    # VIX 선물 ETN
    'UVXY',   # Ultra VIX Short-Term
    'SVXY',   # 역 VIX (공포 역발상)
    'SQQQ',   # 3× 인버스 QQQ (기관 숏 포지션 프록시)
    'SH',     # 1× 인버스 S&P500
    'FXY',    # 엔화 ETF (캐리 언와인드 선행)
    'GDX',    # 골드 마이너 (금보다 민감한 위험회피)
    'EEM',    # 이머징마켓 (글로벌 위험선호 선행)
    'RSP',    # S&P500 이퀄웨이트 (집중도 vs 브레드스)
    'VIXY',   # VIX 단기 선물 ETF
    # ★ XLK 대형 컴포넌트 (내부 선행 신호)
    'NVDA',   # 엔비디아 (반도체/AI 대장주)
    'AAPL',   # 애플
    'MSFT',   # 마이크로소프트 (XLK 최대 비중 ~21%)
    'AVGO',   # 브로드컴
        # ★ 신규 추가: 팩터 ETF / 글로벌 시장
    'MTUM',   # 모멘텀 팩터 ETF (기관 청산 선행)
    'FXI',    # 중국 대형주 (글로벌 위험 전이)
    'EWG',    # 독일 ETF (유럽 경기 프록시)
    'BKLN',   # 레버리지 론 ETF (shadow banking)
    'TQQQ',   # 3× QQQ (레버리지 투자자 심리)
    'MCHI',   # MSCI China (FXI 보완)
    '^VIX9D', '^VIX3M', '^VIX6M', '^SKEW', '^MOVE',   # ^MOVE는 야후에서 누락 잦음(없어도 됨)
    'GOOGL', 'META', 'TSLA'                            # NVDA/MSFT/AAPL/AMZN은 대개 이미 있음
]

EVAL_START     = '2022-01-01'         # 평가 시작일
DOWNLOAD_START = '2020-01-01'         # 지표 계산 히스토리 확보용

# (n, m) 그리드  ※ 단기 예측만 의미있음 → n≤5일 강제
N_RANGE = [1, 2, 3, 4, 5]                                  # 미래 일수 (최대 5일)
M_RANGE = [1.0, 1.5, 2.0, 2.5, 3.0, 4.0, 5.0]              # 누적 하락률 % (1~5%만, 7%+는 단기엔 비현실적)

# 임계치 스윕
N_THRESHOLDS   = 100
PCTL_LO, PCTL_HI = 1, 99              # 이상치 배제 분위

# 예측 유효성 제약
MIN_PRED_RATE  = 0.02                 # 예측 양성률 ≥ 2%
MAX_PRED_RATE  = 0.50                 # 예측 양성률 ≤ 50%
MIN_ACTUAL_POS = 10                   # 실제 하락 샘플 ≥ 10

TOP_K_FOR_SCORE = 20                  # (n, m) 점수 = 상위 K개 F1 평균
TRAIN_RATIO     = 0.70                # 시간순 분할 비율

# 일별 예측 설정
DAILY_START          = '2022-01-01'   # 일별 예측 출력 시작일
ENSEMBLE_TOP_K       = 10             # (기본값; 실행 시 K 스윕으로 자동 최적화)
ENSEMBLE_VOTE_THRESH = 0.5            # 다수결 임계치 (유효지표 중 50% 이상)
K_RANGE = [i for i in range(1,51)]
# 다양성 필터 (Top 지표들이 같은 신호의 변형이면 제외 → 진짜 다양한 지표 선택)
DIVERSITY_FILTER       = True          # True면 Top-K 내 상관 너무 높은 지표 자동 제거
DIVERSITY_CORR_LIMIT   = 0.85          # 상관 |0.85| 이상이면 동질 지표로 간주, 후순위 제거
RISE_OFFSET_ENABLED  = True       # 상승 앙상블로 하락 신호 강도 상쇄
RISE_K_RANGE         = K_RANGE    # 상승 Top-K 후보 (드롭과 동일하게 스윕)
RISE_VOTE_THRESH     = 0.5        # 상승 K 스윕 시 사용 (상승 예측 임계)
RISE_WEIGHT_RANGE    = [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7,
                        0.8, 0.9, 1.0, 1.25, 1.5]
                                   # 후보 weight들 — 각각 적용해 일별 F1 계산 → 최적 자동 선정
                                   # 0.0 = 상쇄 없음(baseline)  /  1.5 = 강한 상쇄
RISE_OFFSET_WEIGHT   = 1.0         # fallback 기본값 (스윕 실패 또는 상쇄 비활성화 시)


# ════════════════════════════════════════════════════════════════
#                      데이터 다운로드
# ════════════════════════════════════════════════════════════════
def download_data(start='2020-01-01'):
    # PEERS에 TICKER가 포함되어 있으면 자동 제거 (자기비교 방지)
    peers_clean = [p for p in PEERS if p != TICKER]

    # 동의어 그룹 — 같은 그룹 내 자산들은 거의 동일 (corr > 0.95)
    # TICKER가 어떤 그룹에 속하면 그 그룹의 다른 자산들도 PEERS에서 제거
    EQUIVALENT_GROUPS = [
        {'SMH', 'SOXX'},                  # 반도체 ETF (corr ~0.99)
        {'VGT', 'XLK', 'QQQ'},            # 광역 테크 (corr ~0.95+)
        {'IGV'},                          # 소프트웨어
        {'SPY', 'IVV', 'VOO', 'DIA'},     # 광역 시장
    ]
    excluded_equivs = set()
    for group in EQUIVALENT_GROUPS:
        if TICKER in group:
            excluded_equivs |= (group - {TICKER})
    peers_clean = [p for p in peers_clean if p not in excluded_equivs]

    all_tickers = [TICKER] + peers_clean
    all_tickers = list(dict.fromkeys(all_tickers))   # 중복 제거 (순서 유지)

    if len(peers_clean) < len(PEERS):
        removed = set(PEERS) - set(peers_clean) - {TICKER}
        print(f"  ℹ TICKER({TICKER})와 동의어/자기비교 자산 제거: {sorted(removed) if removed else 'TICKER만'}")
    print(f"  다운로드: {all_tickers}")
    raw = yf.download(all_tickers, start=start, auto_adjust=True, progress=False)
    ohlcv = {}
    for tk in all_tickers:
        try:
            df = raw.xs(tk, axis=1, level=1)[['Open', 'High', 'Low', 'Close', 'Volume']]
            if df['Close'].notna().sum() > 100:
                ohlcv[tk] = df
        except Exception:
            pass
    closes = pd.DataFrame({tk: ohlcv[tk]['Close'] for tk in ohlcv})
    return closes, ohlcv

# ════════════════════════════════════════════════════════════════
#            FRED 경제지표 다운로드 (신규 추가)
# ════════════════════════════════════════════════════════════════
import pandas as pd
import time
import os

def download_fred_data(start='2020-01-01', api_key='5a586c94ed745a6193f625c0620f5da4', max_retries=2, retry_wait=1.0):
    """
    FRED 경제지표 다운로드 (fredapi 공식 API 사용 — pandas_datareader보다 안정적·빠름).
    실패한 시리즈는 ID·설명·이유를 함께 표시하고, 일시적 실패는 재시도한다.

    API 키 준비 (셋 중 하나):
      1) 함수 인자:        download_fred_data(api_key='발급받은키')
      2) 환경변수:         os.environ['FRED_API_KEY'] = '발급받은키'
      3) Colab 비밀변수:   from google.colab import userdata; userdata.get('FRED_API_KEY')
    키 발급(무료, 즉시): https://fredaccount.stlouisfed.org/apikeys

    fredapi 미설치 또는 키 없으면 빈 DataFrame 반환 (지표 계산 건너뜀).
    설치: pip install fredapi
    """
    try:
        from fredapi import Fred
    except ImportError:
        print("  ⚠ fredapi 미설치 → FRED 경제지표 건너뜀")
        print("    (설치 후 재실행: pip install fredapi)")
        return pd.DataFrame()

    # API 키 확보 — 인자 > 환경변수 > Colab userdata 순
    if api_key is None:
        api_key = os.environ.get('FRED_API_KEY')
    if api_key is None:
        try:
            from google.colab import userdata
            api_key = userdata.get('FRED_API_KEY')
        except Exception:
            pass
    if not api_key:
        print("  ⚠ FRED API 키 없음 → FRED 경제지표 건너뜀")
        print("    키 발급(무료): https://fredaccount.stlouisfed.org/apikeys")
        print("    설정 예: download_fred_data(api_key='발급받은키')")
        return pd.DataFrame()

    try:
        fred = Fred(api_key=api_key)
    except Exception as e:
        print(f"  ⚠ FRED 초기화 실패: {e} → FRED 경제지표 건너뜀")
        return pd.DataFrame()

    daily_series = {
        'T10Y2Y':           '10Y-2Y 수익률 스프레드',
        'T10Y3M':           '10Y-3M 수익률 스프레드',
        'T5YIE':            '5년 BEI 기대인플레이션',
        'T10YIE':           '10년 BEI 기대인플레이션',
        'BAMLH0A0HYM2':     'HY 신용 OAS 스프레드',
        'BAMLC0A0CM':       'IG 신용 OAS 스프레드',
        'BAMLHYH0A0HYM2TRIV': 'HY 채권 총수익 지수',
        'EFFR':             '실효연방기금금리',
        'SOFR':             'SOFR (단기금리)',
    }
    weekly_series = {
        'ICSA':     '주간 신규 실업수당 청구',
        'CCSA':     '계속 실업수당 청구',
        'WRMFSL':   '연준 지급준비금 잔액',
    }
    monthly_series = {
        'UNRATE':      '실업률',
        'U6RATE':      'U6 광의실업률',
        'CPIAUCSL':    'CPI 전체',
        'CPILFESL':    'Core CPI',
        'PCEPI':       'PCE 인플레이션',
        'PCEPILFE':    'Core PCE',
        'CFNAI':       '시카고 연준 국가활동지수 (구 NAPM 대체)',
        'NEWORDER':    '비국방 자본재 신규주문 (구 NAPMNOI 대체)',
        'MANEMP':      '제조업 고용자수 (구 NAPMEI 대체)',
        'PPIACO':      '생산자물가지수 전체 (구 NAPMPI 대체)',
        'UMCSENT':     '미시간 소비자신뢰',
        'MICH':        '미시간 1년 기대인플레이션',
        'HOUST':       '주택착공',
        'PERMIT':      '주택건설허가',
        'INDPRO':      '산업생산지수',
        'TCU':         '설비가동률',
        'PAYEMS':      '비농업 취업자',
        'JTSJOL':      'JOLTS 구인건수',
        'TOTALSL':     '소비자 신용 총액',
        'RSAFS':       '소매판매 (계절조정)',
        'DSPIC96':     '실질 가처분소득',
        'PSAVERT':     '개인저축률',
        'CSUSHPINSA':  'Case-Shiller 주택가격',
        'TWEXBGSMTH':  '달러지수 (Fed 무역가중)',
        'M2SL':        'M2 통화량',
        'WALCL':       '연준 자산총액',
    }
    frames = {}
    all_series = {**daily_series, **weekly_series, **monthly_series}
    ok = 0
    fail_list = []   # (series_id, desc, 이유)
    for series_id, desc in all_series.items():
        last_err = None
        for attempt in range(max_retries + 1):
            try:
                s = fred.get_series(series_id, observation_start=start)
                if s is None or s.dropna().empty:
                    last_err = '데이터 없음(빈 시리즈)'
                    break   # 빈 데이터는 재시도해도 동일 → 중단
                frames[series_id] = s
                last_err = None
                break
            except Exception as e:
                last_err = f'{type(e).__name__}: {str(e)[:70]}'
                if attempt < max_retries:
                    time.sleep(retry_wait)   # 일시적 실패 재시도
                    continue
        if last_err is None:
            ok += 1
        else:
            fail_list.append((series_id, desc, last_err))

    print(f"  FRED: 성공 {ok}개 / 실패 {len(fail_list)}개 (총 {len(all_series)}개 시도)")
    if fail_list:
        print(f"  ── FRED 실패 목록 ──")
        for series_id, desc, reason in fail_list:
            print(f"     ✗ {series_id} ({desc}): {reason}")

    if not frames:
        return pd.DataFrame()
    raw = pd.DataFrame(frames)
    biz_idx = pd.bdate_range(start=start, end=pd.Timestamp.today())
    fred_df = raw.reindex(biz_idx).ffill().bfill()
    fred_df.index.name = 'Date'
    return fred_df

# ════════════════════════════════════════════════════════════════
#                  기술적 지표 헬퍼 함수
# ════════════════════════════════════════════════════════════════
def calc_tr(hi, lo, cl):
    return pd.concat([hi - lo, (hi - cl.shift()).abs(), (lo - cl.shift()).abs()],
                     axis=1).max(axis=1)

def calc_atr(hi, lo, cl, period):
    tr = calc_tr(hi, lo, cl)
    return tr.ewm(com=period - 1, adjust=False).mean(), tr

def calc_rsi(cl, period):
    d = cl.diff()
    g = d.clip(lower=0).ewm(com=period - 1, adjust=False).mean()
    ls = (-d.clip(upper=0)).ewm(com=period - 1, adjust=False).mean()
    rs = g / ls.replace(0, np.nan)
    return 100 - 100 / (1 + rs)

def calc_cci(hi, lo, cl, period):
    tp = (hi + lo + cl) / 3
    ma = tp.rolling(period).mean()
    md = tp.rolling(period).apply(lambda x: np.mean(np.abs(x - x.mean())))
    return (tp - ma) / (0.015 * md.replace(0, np.nan))

def calc_stoch_k(hi, lo, cl, period):
    ll = lo.rolling(period).min()
    hh = hi.rolling(period).max()
    return (cl - ll) / (hh - ll).replace(0, np.nan) * 100

def calc_williams(hi, lo, cl, period):
    hh = hi.rolling(period).max()
    ll = lo.rolling(period).min()
    return (hh - cl) / (hh - ll).replace(0, np.nan) * -100

def calc_zscore(s, period):
    return (s - s.rolling(period).mean()) / s.rolling(period).std().replace(0, np.nan)

def calc_pctrank(s, period):
    return s.rolling(period).apply(lambda x: pd.Series(x).rank(pct=True).iloc[-1],
                                   raw=False)

def calc_linreg_slope(s, period):
    t = np.arange(period)
    def _s(x):
        if np.isnan(x).any(): return np.nan
        return np.polyfit(t, x, 1)[0] / (abs(x.mean()) or 1)
    return s.rolling(period).apply(_s, raw=True)

def calc_dema(s, period):
    e1 = s.ewm(span=period, adjust=False).mean()
    e2 = e1.ewm(span=period, adjust=False).mean()
    return 2 * e1 - e2

def calc_tema(s, period):
    e1 = s.ewm(span=period, adjust=False).mean()
    e2 = e1.ewm(span=period, adjust=False).mean()
    e3 = e2.ewm(span=period, adjust=False).mean()
    return 3 * e1 - 3 * e2 + e3

def calc_hma(s, period):
    half = s.rolling(period // 2).apply(
        lambda x: np.average(x, weights=range(1, len(x) + 1)), raw=True)
    full = s.rolling(period).apply(
        lambda x: np.average(x, weights=range(1, len(x) + 1)), raw=True)
    raw_v = 2 * half - full
    sq = int(np.sqrt(period))
    return raw_v.rolling(sq).apply(
        lambda x: np.average(x, weights=range(1, len(x) + 1)), raw=True)

def calc_adx(hi, lo, cl, period):
    tr = calc_tr(hi, lo, cl)
    ph = hi.shift(); pl = lo.shift()
    pdm = np.where((hi - ph) > (pl - lo), (hi - ph).clip(lower=0), 0.0)
    mdm = np.where((pl - lo) > (hi - ph), (pl - lo).clip(lower=0), 0.0)
    atr_s = tr.ewm(com=period - 1, adjust=False).mean()
    pdi = pd.Series(pdm, index=cl.index).ewm(com=period - 1, adjust=False).mean() / atr_s * 100
    mdi = pd.Series(mdm, index=cl.index).ewm(com=period - 1, adjust=False).mean() / atr_s * 100
    dx  = (pdi - mdi).abs() / (pdi + mdi).replace(0, np.nan) * 100
    adx = dx.ewm(com=period - 1, adjust=False).mean()
    return adx, pdi, mdi

def calc_psar(hi_arr, lo_arr, cl_arr):
    psar = np.full(len(cl_arr), np.nan)
    af = 0.02; max_af = 0.20; bull = True
    sar = lo_arr[0]; ep = hi_arr[0]
    for i in range(1, len(cl_arr)):
        if bull:
            sar = sar + af * (ep - sar)
            sar = min(sar, lo_arr[i - 1], lo_arr[max(0, i - 2)])
            if lo_arr[i] < sar:
                bull = False; sar = ep; ep = lo_arr[i]; af = 0.02
            elif hi_arr[i] > ep:
                ep = hi_arr[i]; af = min(af + 0.02, max_af)
        else:
            sar = sar + af * (ep - sar)
            sar = max(sar, hi_arr[i - 1], hi_arr[max(0, i - 2)])
            if hi_arr[i] > sar:
                bull = True; sar = ep; ep = hi_arr[i]; af = 0.02
            elif lo_arr[i] < ep:
                ep = lo_arr[i]; af = min(af + 0.02, max_af)
        psar[i] = sar
    return pd.Series(psar, index=pd.RangeIndex(len(cl_arr)))

def fetch_intraday(ticker, interval='5m', period='60d', tries=3):
    """야후에서 일중 OHLCV를 받아온다 (무료: 5분봉 최근 60일). 실패 시 빈 DataFrame."""
    import yfinance as yf
    import time
    for k in range(tries):
        try:
            df = yf.download(ticker, period=period, interval=interval,
                             progress=False, auto_adjust=True, prepost=False)
            if df is not None and len(df) > 0:
                if isinstance(df.columns, pd.MultiIndex):
                    df.columns = df.columns.get_level_values(0)
                df = df.rename(columns=str.title)
                return df[['Open', 'High', 'Low', 'Close', 'Volume']].dropna()
        except Exception as e:
            if k == tries - 1:
                print(f"[intraday] {ticker} 수집 실패: {e}")
            time.sleep(1.5)
    return pd.DataFrame()


def intraday_daily_features(intraday):
    """5분봉 → 거래일별 미시구조 피처 (확장판 42개). 일별 인덱스로 반환."""
    if intraday is None or len(intraday) == 0:
        return pd.DataFrame()
    df = intraday.copy()
    try:
        if df.index.tz is not None:
            df.index = df.index.tz_convert('America/New_York')
    except Exception:
        pass
    df['date'] = df.index.normalize()
    df['hm'] = df.index.hour * 60 + df.index.minute
    df['bar_ret'] = df['Close'].pct_change()
    _daily_close = df['Close'].groupby(df['date']).last()
    _prev_close_map = _daily_close.shift(1)
    out = {}
    for d, g in df.groupby('date'):
        g = g.sort_index()
        if len(g) < 12:
            continue
        o = g['Open'].iloc[0]; c = g['Close'].iloc[-1]
        hi = g['High'].max(); lo = g['Low'].min()
        rng = (hi - lo) if (hi - lo) != 0 else np.nan
        vsum = g['Volume'].sum()
        vv = g['Volume'].values.astype(float)
        cl = g['Close'].values.astype(float)
        br = g['bar_ret'].dropna().values
        n = len(g)
        tp = (g['High'] + g['Low'] + g['Close']) / 3
        vwap = (tp * g['Volume']).sum() / (vsum if vsum > 0 else np.nan)
        hm = g['hm'].values
        rec = {}

        # ── 기존 15개 (id_) ──────────────────────────────────────
        rec['id_close_loc'] = (c - lo) / rng
        rec['id_close_vs_vwap'] = (c / vwap - 1) if (vwap and vwap > 0) else np.nan
        path = np.abs(np.diff(cl)).sum()
        rec['id_trend_efficiency'] = (c - o) / path if path > 0 else 0.0
        first = cl[min(6, n - 1)] / o - 1
        last = c / cl[max(0, n - 7)] - 1
        rec['id_open_drive'] = first
        rec['id_close_drive'] = last
        rec['id_smart_dumb'] = last - first
        early_v = vv[:6].sum(); late_v = vv[-6:].sum()
        rec['id_late_vol_ratio'] = late_v / vsum if vsum > 0 else np.nan
        rec['id_vol_smile'] = (early_v + late_v) / vsum if vsum > 0 else np.nan
        rec['id_realized_vol'] = np.sqrt(np.nansum(br ** 2)) if len(br) else np.nan
        rec['id_intraday_skew'] = pd.Series(br).skew() if len(br) > 3 else np.nan
        cummax = np.maximum.accumulate(cl)
        rec['id_intraday_maxdd'] = float(np.min(cl / cummax - 1))
        rec['id_up_bar_ratio'] = float((br > 0).mean()) if len(br) else np.nan
        mid = (g['High'] + g['Low']) / 2
        tick = np.sign(g['Close'] - mid)
        rec['id_vw_order_flow'] = float((tick * g['Volume']).sum() / (vsum if vsum > 0 else np.nan))
        rec['id_closing_vol'] = (np.std(br[-3:]) / (np.std(br) + 1e-9)) if len(br) > 5 else np.nan
        rec['id_above_vwap_ratio'] = float((g['Close'] > vwap).mean()) if (vwap and vwap > 0) else np.nan

        # ── 신규 A. 시간대별 수익 분해 (세션 구간) ───────────────
        def seg_ret(lo_m, hi_m):
            m = (hm >= lo_m) & (hm < hi_m)
            if m.sum() < 2: return np.nan
            seg = cl[m]
            return seg[-1] / seg[0] - 1
        rec['idx_ret_open30'] = seg_ret(570, 600)      # 09:30~10:00
        rec['idx_ret_morning'] = seg_ret(600, 720)     # 10:00~12:00
        rec['idx_ret_lunch'] = seg_ret(720, 810)       # 12:00~13:30
        rec['idx_ret_afternoon'] = seg_ret(810, 930)   # 13:30~15:30
        rec['idx_ret_close30'] = seg_ret(930, 960)     # 15:30~16:00
        _am = seg_ret(570, 720); _pm = seg_ret(810, 960)
        rec['idx_pm_minus_am'] = (_pm - _am) if (pd.notna(_pm) and pd.notna(_am)) else np.nan

        # ── 신규 B. 일중 추세 반전/지속 ──────────────────────────
        rec['idx_open_close_align'] = float(np.sign(_am) == np.sign(c - o)) if pd.notna(_am) else np.nan
        rec['idx_reversal_down'] = float(pd.notna(_am) and pd.notna(_pm) and _am > 0 and _pm < 0)
        rec['idx_reversal_up'] = float(pd.notna(_am) and pd.notna(_pm) and _am < 0 and _pm > 0)
        hi_idx = int(np.argmax(g['High'].values)); lo_idx = int(np.argmin(g['Low'].values))
        rec['idx_high_time_frac'] = hi_idx / n          # 0=장초 1=장마감
        rec['idx_low_time_frac'] = lo_idx / n
        rec['idx_high_after_low'] = float(hi_idx > lo_idx)   # 저점먼저→고점(장중 상승)

        # ── 신규 C. 갭/시가 위치 동역학 ──────────────────────────
        rec['idx_open_loc_in_day'] = (o - lo) / rng
        rec['idx_close_above_open'] = float(c > o)
        rec['idx_intraday_return'] = c / o - 1

        # ── 신규 D. 누적 델타 곡선 ───────────────────────────────
        bar_mid = (g['High'].values + g['Low'].values) / 2
        bar_dir = np.sign(g['Close'].values - bar_mid)
        signed_v = bar_dir * vv
        cum_delta = np.cumsum(signed_v)
        rec['idx_cum_delta_end'] = cum_delta[-1] / (vsum if vsum > 0 else np.nan)
        rec['idx_delta_late_vs_early'] = (signed_v[-n//3:].sum() - signed_v[:n//3].sum()) / (vsum if vsum > 0 else np.nan)
        rec['idx_price_delta_diverge'] = float((c > o) and (cum_delta[-1] < 0))

        # ── 신규 E. 변동성 일중 분포 ─────────────────────────────
        if len(br) > 12:
            rec['idx_vol_open_vs_close'] = (np.std(br[:n//4]) / (np.std(br[-n//4:]) + 1e-9))
        else:
            rec['idx_vol_open_vs_close'] = np.nan
        rec['idx_max_bar_move'] = float(np.max(np.abs(br))) if len(br) else np.nan
        if len(br) > 5:
            sq = np.sort(br ** 2)[::-1]
            rec['idx_vol_concentration'] = sq[:3].sum() / (np.sum(br ** 2) + 1e-12)
        else:
            rec['idx_vol_concentration'] = np.nan

        # ── 신규 F. 거래량-가격 일중 관계 ────────────────────────
        simple_mean = cl.mean()
        rec['idx_vwap_vs_mean'] = (vwap / simple_mean - 1) if simple_mean > 0 else np.nan
        big = vv > (vv.mean() + vv.std())
        if big.sum() > 0:
            rec['idx_big_bar_dir'] = float(np.sign(g['Close'].values - g['Open'].values)[big].mean())
        else:
            rec['idx_big_bar_dir'] = np.nan
        rec['idx_volume_trend'] = float(np.corrcoef(np.arange(n), vv)[0, 1]) if n > 3 else np.nan

        # ── 신규 G. 종가 무렵 행동 (마감 경매 프록시) ────────────
        rec['idx_closing_flow'] = signed_v[-6:].sum() / (vv[-6:].sum() + 1e-9)
        rec['idx_close_at_high'] = float(c >= hi * 0.999)
        rec['idx_close_at_low'] = float(c <= lo * 1.001)

        # ── v3 신뢰성 신규: 일중모멘텀/점프분해/자기상관/오버나이트/VWAP정밀/유동성 ──
        _pc = _prev_close_map.get(d, np.nan)
        def _seg(a, b):
            m = (hm >= a) & (hm < b)
            if m.sum() < 2: return np.nan
            s = cl[m]; return s[-1] / s[0] - 1
        # H. 일중 모멘텀 (마지막 30분 = 익일 예측: 학술 검증된 신호)
        _last30 = _seg(930, 960)
        rec['idm_last30_ret'] = _last30
        rec['idm_first30_ret'] = _seg(570, 600)
        rec['idm_first_last_sum'] = (rec['idm_first30_ret'] + _last30
                                     if pd.notna(rec['idm_first30_ret']) and pd.notna(_last30) else np.nan)
        rec['idm_last30_vol_share'] = vv[hm >= 930].sum() / (vsum if vsum > 0 else np.nan)
        # I. 점프 vs 연속 변동성 분해 (바이파워 변동)
        if len(br) > 5:
            _bv = np.sum(br ** 2)
            _bp = (np.pi / 2) * np.sum(np.abs(br[1:]) * np.abs(br[:-1]))
            _jump = max(_bv - _bp, 0.0)
            rec['idm_jump_share'] = _jump / (_bv + 1e-12)
            rec['idm_continuous_vol'] = np.sqrt(max(_bp, 0))
            _neg = br[br < 0]
            rec['idm_neg_semivar'] = np.sum(_neg ** 2) / (_bv + 1e-12)
            rec['idm_signed_jump'] = (np.sum(br[br > 0] ** 2) - np.sum(_neg ** 2)) / (_bv + 1e-12)
        else:
            rec['idm_jump_share'] = rec['idm_continuous_vol'] = np.nan
            rec['idm_neg_semivar'] = rec['idm_signed_jump'] = np.nan
        # J. 일중 자기상관 (평균회귀 vs 추세)
        if len(br) > 10:
            _b1 = br[1:]; _b0 = br[:-1]; _sd = np.std(_b1) * np.std(_b0)
            rec['idm_autocorr1'] = float(np.mean((_b1 - _b1.mean()) * (_b0 - _b0.mean())) / (_sd + 1e-12)) if _sd > 0 else np.nan
        else:
            rec['idm_autocorr1'] = np.nan
        rec['idm_run_persistence'] = float(np.mean(np.sign(br[1:]) == np.sign(br[:-1]))) if len(br) > 3 else np.nan
        # K. 오버나이트 vs 인트라데이 분해
        if pd.notna(_pc) and _pc > 0:
            rec['idm_overnight_ret'] = o / _pc - 1
            rec['idm_intraday_ret'] = c / o - 1
            rec['idm_overnight_faded'] = float((o / _pc - 1) > 0 and (c / o - 1) < 0)
            rec['idm_on_id_same_dir'] = float(np.sign(o / _pc - 1) == np.sign(c / o - 1))
        else:
            rec['idm_overnight_ret'] = rec['idm_intraday_ret'] = np.nan
            rec['idm_overnight_faded'] = rec['idm_on_id_same_dir'] = np.nan
        # L. VWAP 정밀 (시간가중 체류 / 재탈환 / z밴드)
        if vwap and vwap > 0:
            _w = np.linspace(0.5, 1.5, n)
            rec['idm_vwap_time_weighted'] = float(np.sum((cl > vwap) * _w) / np.sum(_w))
            rec['idm_close_reclaim_vwap'] = float(cl[-1] > vwap and cl[max(0, n-6)] < vwap)
            rec['idm_close_lose_vwap'] = float(cl[-1] < vwap and cl[max(0, n-6)] > vwap)
            _vwstd = np.sqrt(((((tp.values - vwap) ** 2) * vv).sum()) / (vsum if vsum > 0 else np.nan))
            rec['idm_close_vwap_z'] = (c - vwap) / (_vwstd + 1e-9) if _vwstd > 0 else np.nan
        else:
            rec['idm_vwap_time_weighted'] = rec['idm_close_reclaim_vwap'] = np.nan
            rec['idm_close_lose_vwap'] = rec['idm_close_vwap_z'] = np.nan
        # M. 일중 유동성/충격 (Amihud 일중판)
        if len(br) > 5:
            _dbar = np.abs(cl) * vv
            _brf = g['bar_ret'].values
            _illiq = np.abs(_brf[1:]) / (_dbar[1:] + 1e-9)
            rec['idm_intraday_illiq'] = float(np.nanmean(_illiq) * 1e9)
            _ui = np.abs(br[br > 0]).mean() if (br > 0).any() else np.nan
            _di = np.abs(br[br < 0]).mean() if (br < 0).any() else np.nan
            rec['idm_impact_asym'] = (_di / _ui) if (pd.notna(_ui) and _ui > 0) else np.nan
        else:
            rec['idm_intraday_illiq'] = rec['idm_impact_asym'] = np.nan

        out[d] = rec

    if not out:
        return pd.DataFrame()
    res = pd.DataFrame(out).T
    res.index = pd.to_datetime(res.index).tz_localize(None).normalize()
    res = res.sort_index()

    # ── N. 다일 집계로 신뢰성 강화 (단일일 노이즈 완화 — 핵심) ──
    for _base in ['idm_last30_ret', 'idm_signed_jump', 'idm_autocorr1', 'idm_intraday_ret',
                  'idm_close_vwap_z', 'idm_neg_semivar']:
        if _base in res.columns:
            res[f'{_base}_5davg'] = res[_base].rolling(5).mean()
            res[f'{_base}_10davg'] = res[_base].rolling(10).mean()
    if 'idm_last30_ret' in res.columns:
        res['idm_last30_consistency_5d'] = np.sign(res['idm_last30_ret']).rolling(5).mean()
    if 'idm_intraday_ret' in res.columns:
        res['idm_intraday_cum_5d'] = res['idm_intraday_ret'].rolling(5).sum()
    if 'idm_jump_share' in res.columns:
        res['idm_jump_freq_10d'] = (res['idm_jump_share'] > 0.3).rolling(10).mean()
    if 'idm_close_vwap_z' in res.columns:
        res['idm_above_vwap_freq_5d'] = (res['idm_close_vwap_z'] > 0).rolling(5).mean()
    return res



# ════════════════════════════════════════════════════════════════
#                  지표 계산 (~450개)
# ════════════════════════════════════════════════════════════════
def compute_features(ohlcv, closes, fred_df=None):
    df = ohlcv[TICKER].copy()
    cl = df['Close']; hi = df['High']; lo = df['Low']
    op = df['Open'];  vo = df['Volume']
    feat = pd.DataFrame(index=df.index)

    atr14, tr14 = calc_atr(hi, lo, cl, 14)

    # ── 1. 수익률 / 모멘텀 ─────────────────────────────
    for p in [1, 2, 3, 4, 5, 7, 10, 14, 15, 20, 25, 30, 40, 60, 90, 120]:
        feat[f'ret_{p}d'] = cl.pct_change(p)

    r1 = cl.pct_change(1)
    for p in [10, 20, 60]:
        feat[f'ret_zscore_{p}d']  = calc_zscore(r1, p)
        feat[f'ret_pctrank_{p}d'] = calc_pctrank(r1, p)

    for p in [3, 5, 6, 7, 9, 10, 12, 14, 20, 21, 28]:
        feat[f'rsi_{p}'] = calc_rsi(cl, p)

    for p in [9, 14, 21]:
        feat[f'rsi_{p}_zscore20'] = calc_zscore(calc_rsi(cl, p), 20)

    for fast, slow, sig in [(5, 13, 4), (8, 17, 9), (12, 26, 9), (19, 39, 9)]:
        ef = cl.ewm(span=fast, adjust=False).mean()
        es = cl.ewm(span=slow, adjust=False).mean()
        mc = ef - es; sg = mc.ewm(span=sig, adjust=False).mean()
        lbl = f'macd_{fast}_{slow}'
        feat[f'{lbl}_norm']    = mc / cl
        feat[f'{lbl}_hist']    = (mc - sg) / cl
        feat[f'{lbl}_cross']   = np.sign(mc - sg)
        feat[f'{lbl}_hist_ch'] = ((mc - sg) - (mc - sg).shift(1)) / cl

    for p in [3, 5, 7, 10, 14, 20, 30, 60]:
        feat[f'roc_{p}'] = (cl - cl.shift(p)) / cl.shift(p)

    for p in [10, 14, 20, 30, 40]:
        feat[f'cci_{p}'] = calc_cci(hi, lo, cl, p)

    for p in [10, 14, 20, 28]:
        feat[f'williams_{p}'] = calc_williams(hi, lo, cl, p)

    for p in [5, 9, 14, 21]:
        k = calc_stoch_k(hi, lo, cl, p)
        d = k.rolling(3).mean()
        feat[f'stoch_k_{p}']  = k
        feat[f'stoch_d_{p}']  = d
        feat[f'stoch_kd_{p}'] = k - d

    for p in [9, 14, 18]:
        feat[f'trix_{p}'] = calc_tema(cl, p).pct_change(1)

    for p in [9, 14, 20]:
        dd = cl.diff()
        su = dd.clip(lower=0).rolling(p).sum()
        sd = (-dd.clip(upper=0)).rolling(p).sum()
        feat[f'cmo_{p}'] = (su - sd) / (su + sd).replace(0, np.nan) * 100

    for p in [10, 14, 20]:
        feat[f'dpo_{p}'] = cl - cl.rolling(p).mean().shift(p // 2 + 1)

    for fast, slow in [(9, 26), (12, 26)]:
        ef = cl.ewm(span=fast, adjust=False).mean()
        es = cl.ewm(span=slow, adjust=False).mean()
        feat[f'ppo_{fast}_{slow}'] = (ef - es) / es.replace(0, np.nan) * 100

    for p in [5, 10, 20, 30]:
        feat[f'linreg_slope_{p}'] = calc_linreg_slope(cl, p)

    for p in [3, 5, 10, 20]:
        feat[f'vol_wt_ret_{p}'] = (cl.pct_change() * vo).rolling(p).sum() / vo.rolling(p).sum()

    # ── 2. 변동성 ─────────────────────────────────────
    for p in [5, 7, 10, 14, 20, 30]:
        atr_p, _ = calc_atr(hi, lo, cl, p)
        feat[f'atr_{p}']        = atr_p / cl
        feat[f'atr_{p}_zscore'] = calc_zscore(atr_p / cl, 60)

    for p in [5, 10, 20, 30, 60, 90]:
        feat[f'hist_vol_{p}'] = cl.pct_change().rolling(p).std() * np.sqrt(252)

    for sh, lg in [(5, 20), (10, 60), (20, 60)]:
        sv = cl.pct_change().rolling(sh).std()
        lv = cl.pct_change().rolling(lg).std()
        feat[f'hvol_ratio_{sh}_{lg}'] = sv / lv.replace(0, np.nan)

    gk = 0.5 * (np.log(hi / lo) ** 2) - (2 * np.log(2) - 1) * (np.log(cl / op) ** 2)
    feat['garman_5']  = gk.rolling(5).mean()
    feat['garman_20'] = gk.rolling(20).mean()

    feat['parkinson_20'] = np.sqrt(
        1 / (4 * 20 * np.log(2)) * (np.log(hi / lo) ** 2).rolling(20).sum() * 252)

    for p, k in [(10, 2.0), (10, 1.5), (20, 1.0), (20, 1.5),
                 (20, 2.0), (20, 2.5), (30, 2.0)]:
        ma = cl.rolling(p).mean(); sd = cl.rolling(p).std()
        feat[f'bb_pct_{p}_{k}']   = (cl - (ma - k * sd)) / (2 * k * sd).replace(0, np.nan)
        feat[f'bb_width_{p}_{k}'] = 2 * k * sd / ma.replace(0, np.nan)

    for p, m in [(10, 1.5), (20, 2.0), (20, 1.5)]:
        atr_p, _ = calc_atr(hi, lo, cl, p)
        mid = cl.ewm(span=p, adjust=False).mean()
        feat[f'keltner_pct_{p}_{m}']   = (cl - (mid - m * atr_p)) / (2 * m * atr_p).replace(0, np.nan)
        feat[f'keltner_width_{p}_{m}'] = 2 * m * atr_p / mid.replace(0, np.nan)

    for p in [10, 20, 55]:
        dh = hi.rolling(p).max(); dl = lo.rolling(p).min()
        feat[f'donchian_pct_{p}']   = (cl - dl) / (dh - dl).replace(0, np.nan)
        feat[f'donchian_width_{p}'] = (dh - dl) / cl

    for p in [14, 20]:
        roll_max = cl.rolling(p).max()
        feat[f'ulcer_{p}'] = np.sqrt(((cl - roll_max) / roll_max * 100) ** 2).rolling(p).mean()

    feat['intraday_range_norm'] = (hi - lo) / cl
    feat['true_range_norm']     = tr14 / cl
    feat['intraday_range_5ma']  = ((hi - lo) / cl).rolling(5).mean()

    vix_c = closes.get('^VIX')
    if vix_c is not None:
        feat['vix_level']      = vix_c
        feat['vix_1d_chg']     = vix_c.pct_change(1)
        feat['vix_5d_chg']     = vix_c.pct_change(5)
        feat['vix_zscore_20']  = calc_zscore(vix_c, 20)
        feat['vix_zscore_60']  = calc_zscore(vix_c, 60)
        feat['vix_spike_20']   = vix_c / vix_c.rolling(20).mean() - 1
        feat['vix_pctrank_60'] = calc_pctrank(vix_c, 60)
        feat['vix_term_slope'] = (vix_c - vix_c.shift(5)) / 5

    # ── 3. 거래량 ─────────────────────────────────────
    for p in [3, 5, 10, 20, 30, 60]:
        feat[f'vol_ratio_{p}'] = vo / vo.rolling(p).mean().replace(0, np.nan)

    for sh, lg in [(3, 10), (5, 20), (5, 60), (10, 30), (10, 60)]:
        feat[f'vol_trend_{sh}_{lg}'] = vo.rolling(sh).mean() / vo.rolling(lg).mean() - 1

    for p in [10, 20, 60]:
        feat[f'vol_zscore_{p}'] = calc_zscore(vo, p)

    for p in [10, 20]:
        up_v = vo.where(cl.diff() > 0, 0.0).rolling(p).sum()
        dn_v = vo.where(cl.diff() < 0, 0.0).rolling(p).sum()
        feat[f'vol_up_dn_ratio_{p}'] = up_v / dn_v.replace(0, np.nan)

    obv = (np.sign(cl.diff()) * vo).fillna(0).cumsum()
    for p in [5, 10, 20]:
        obv_ma = obv.rolling(p).mean()
        feat[f'obv_{p}_dist'] = (obv - obv_ma) / obv_ma.abs().replace(0, np.nan)
    feat['obv_slope_5']  = obv.diff(5)  / vo.rolling(5).mean().replace(0, np.nan)
    feat['obv_slope_10'] = obv.diff(10) / vo.rolling(10).mean().replace(0, np.nan)

    for p in [5, 10, 20]:
        mfm = ((cl - lo) - (hi - cl)) / (hi - lo).replace(0, np.nan)
        feat[f'cmf_{p}'] = (mfm * vo).rolling(p).sum() / vo.rolling(p).sum().replace(0, np.nan)

    for p in [7, 10, 14, 20]:
        tp = (hi + lo + cl) / 3; mf = tp * vo
        pos = mf.where(tp > tp.shift(), 0.0)
        neg = mf.where(tp < tp.shift(), 0.0)
        mfr = pos.rolling(p).sum() / neg.rolling(p).sum().replace(0, np.nan)
        feat[f'mfi_{p}'] = 100 - 100 / (1 + mfr)

    for p in [5, 10, 20, 60]:
        vwap = (cl * vo).rolling(p).sum() / vo.rolling(p).sum().replace(0, np.nan)
        feat[f'vwap_{p}_dist'] = cl / vwap - 1

    feat['force_idx_2']  = cl.diff(1) * vo
    feat['force_idx_13'] = (cl.diff(1) * vo).ewm(span=13, adjust=False).mean()

    bp  = (hi + lo) / 2 - (hi.shift() + lo.shift()) / 2
    box = (vo / 1e6) / (hi - lo).replace(0, np.nan)
    feat['ease_mov_14'] = (bp / box.replace(0, np.nan)).rolling(14).mean() / cl

    ad_line = ((2 * cl - lo - hi) / (hi - lo).replace(0, np.nan) * vo).cumsum()
    feat['ad_line_20_dist'] = calc_zscore(ad_line, 20)

    # ── 4. 추세 ───────────────────────────────────────
    for p in [5, 8, 10, 13, 20, 21, 34, 50, 55, 89, 100, 150, 200]:
        feat[f'sma_{p}_dist'] = cl / cl.rolling(p).mean() - 1

    for p in [5, 8, 10, 13, 20, 21, 26, 34, 50, 55, 89]:
        feat[f'ema_{p}_dist'] = cl / cl.ewm(span=p, adjust=False).mean() - 1

    for p in [10, 20, 50]:
        feat[f'dema_{p}_dist'] = cl / calc_dema(cl, p) - 1
    for p in [10, 20]:
        feat[f'tema_{p}_dist'] = cl / calc_tema(cl, p) - 1
    for p in [10, 20, 50]:
        feat[f'hma_{p}_dist']  = cl / calc_hma(cl, p) - 1

    for p in [10, 20]:
        vwma = (cl * vo).rolling(p).sum() / vo.rolling(p).sum().replace(0, np.nan)
        feat[f'vwma_{p}_dist'] = cl / vwma - 1

    for sh, lg in [(5, 20), (10, 50), (20, 50), (50, 200), (20, 200),
                    (5, 50), (8, 21), (13, 34)]:
        feat[f'sma_{sh}_{lg}_gap'] = cl.rolling(sh).mean() / cl.rolling(lg).mean() - 1
    feat['ema_12_26_gap'] = (cl.ewm(span=12, adjust=False).mean() /
                              cl.ewm(span=26, adjust=False).mean() - 1)
    feat['ema_8_21_gap']  = (cl.ewm(span=8,  adjust=False).mean() /
                              cl.ewm(span=21, adjust=False).mean() - 1)

    for p in [10, 14, 20]:
        adx_v, pdi, mdi = calc_adx(hi, lo, cl, p)
        feat[f'adx_{p}']      = adx_v
        feat[f'di_gap_{p}']   = pdi - mdi
        feat[f'di_ratio_{p}'] = pdi / mdi.replace(0, np.nan)

    for p in [10, 14, 20, 25]:
        feat[f'aroon_up_{p}']  = hi.rolling(p + 1).apply(lambda x: x.argmax(), raw=True) / p * 100
        feat[f'aroon_dn_{p}']  = lo.rolling(p + 1).apply(lambda x: x.argmin(), raw=True) / p * 100
        feat[f'aroon_osc_{p}'] = feat[f'aroon_up_{p}'] - feat[f'aroon_dn_{p}']

    psar_s = calc_psar(hi.values, lo.values, cl.values)
    psar_s.index = cl.index
    feat['psar_dist']   = cl / psar_s.replace(0, np.nan) - 1
    feat['psar_signal'] = np.sign(cl - psar_s)

    for p in [14, 21]:
        vm_p = (hi - lo.shift()).abs().rolling(p).sum()
        vm_m = (lo - hi.shift()).abs().rolling(p).sum()
        atr_sum = tr14.rolling(p).sum()
        feat[f'vortex_diff_{p}'] = (vm_p - vm_m) / atr_sum.replace(0, np.nan)

    ema9 = (hi - lo).ewm(span=9, adjust=False).mean()
    feat['mass_idx_25'] = (ema9 / ema9.ewm(span=9, adjust=False).mean().replace(0, np.nan)).rolling(25).sum()

    for p in [9, 26, 52]:
        ichi_mid = (hi.rolling(p).max() + lo.rolling(p).min()) / 2
        feat[f'ichimoku_{p}_dist'] = cl / ichi_mid.replace(0, np.nan) - 1

    for mult in [2.0, 3.0]:
        mid = (hi + lo) / 2
        feat[f'supertrend_upper_{mult}'] = (cl - (mid + mult * atr14)) / cl
        feat[f'supertrend_lower_{mult}'] = (cl - (mid - mult * atr14)) / cl

    feat['dist_52w_high'] = cl / hi.rolling(252).max() - 1
    feat['dist_52w_low']  = cl / lo.rolling(252).min() - 1
    feat['pos_52w_range'] = (cl - lo.rolling(252).min()) / \
                             (hi.rolling(252).max() - lo.rolling(252).min()).replace(0, np.nan)

    for p in [10, 20, 60]:
        feat[f'channel_pos_{p}'] = (cl - lo.rolling(p).min()) / \
                                    (hi.rolling(p).max() - lo.rolling(p).min()).replace(0, np.nan)

    for p in [10, 20]:
        r = cl.pct_change()
        feat[f'sharpe_like_{p}'] = r.rolling(p).mean() / r.rolling(p).std().replace(0, np.nan)

    for p in [10, 20, 60]:
        feat[f'close_pctrank_{p}'] = calc_pctrank(cl, p)

    # 스퀴즈
    bb_up = cl.rolling(20).mean() + 2 * cl.rolling(20).std()
    bb_lo = cl.rolling(20).mean() - 2 * cl.rolling(20).std()
    atr20, _ = calc_atr(hi, lo, cl, 20)
    ema20 = cl.ewm(span=20, adjust=False).mean()
    kc_up = ema20 + 1.5 * atr20; kc_lo = ema20 - 1.5 * atr20
    feat['squeeze_flag'] = ((bb_up < kc_up) & (bb_lo > kc_lo)).astype(float)
    feat['squeeze_hist'] = calc_linreg_slope(
        cl - (hi.rolling(20).max() + lo.rolling(20).min()) / 2, 5)

    # ── 5. 상대강도 / Breadth ─────────────────────────
    peer_map = {
        'spy': 'SPY', 'qqq': 'QQQ', 'smh': 'SMH', 'soxx': 'SOXX',
        'vgt': 'VGT', 'igv': 'IGV', 'xlc': 'XLC', 'xlf': 'XLF',
    }
    # TICKER 자체는 제외 (SPY가 TICKER면 rel_spy_*는 항상 0이 됨)
    peer_map = {k: v for k, v in peer_map.items() if v != TICKER}
    for lbl, sym in peer_map.items():
        peer = closes.get(sym)
        if peer is None: continue
        for p in [5, 10, 20, 60]:
            feat[f'rel_{lbl}_{p}d'] = cl.pct_change(p) - peer.pct_change(p)
        feat[f'beta_{lbl}_20d'] = (cl.pct_change().rolling(20).cov(peer.pct_change()) /
                                    peer.pct_change().rolling(20).var().replace(0, np.nan))
        feat[f'beta_{lbl}_60d'] = (cl.pct_change().rolling(60).cov(peer.pct_change()) /
                                    peer.pct_change().rolling(60).var().replace(0, np.nan))
        feat[f'corr_{lbl}_10d'] = cl.pct_change().rolling(10).corr(peer.pct_change())
        feat[f'corr_{lbl}_20d'] = cl.pct_change().rolling(20).corr(peer.pct_change())
        feat[f'corr_{lbl}_60d'] = cl.pct_change().rolling(60).corr(peer.pct_change())
        feat[f'rs_{lbl}_52w']   = (cl / cl.shift(252)) / (peer / peer.shift(252).replace(0, np.nan))

    xlf = closes.get('XLF'); xlc = closes.get('XLC')
    if xlf is not None: feat['spread_etf_xlk_xlf'] = cl.pct_change(20) - xlf.pct_change(20)
    if xlc is not None: feat['spread_etf_xlk_xlc'] = cl.pct_change(20) - xlc.pct_change(20)

    # ── 6. 매크로 ─────────────────────────────────────
    tnx = closes.get('^TNX'); irx = closes.get('^IRX')
    gld = closes.get('GLD');  tlt = closes.get('TLT')

    if tnx is not None:
        feat['tnx_level']       = tnx
        feat['tnx_zscore_60']   = calc_zscore(tnx, 60)
        feat['tnx_pctrank_252'] = calc_pctrank(tnx, 252)
        feat['tnx_sma20_dist']  = tnx / tnx.rolling(20).mean() - 1
        for p in [1, 5, 10, 20, 60]:
            feat[f'tnx_{p}d_chg'] = tnx.diff(p)

    if irx is not None:
        feat['irx_level']  = irx
        feat['irx_5d_chg'] = irx.diff(5)

    if tnx is not None and irx is not None:
        feat['term_spread']          = tnx - irx
        feat['macro_yield_slope_5d'] = (tnx - irx).diff(5)

    if gld is not None:
        feat['gld_5d_ret']   = gld.pct_change(5)
        feat['gld_10d_ret']  = gld.pct_change(10)
        feat['gld_corr_20d'] = cl.pct_change().rolling(20).corr(gld.pct_change())

    if tlt is not None:
        feat['tlt_5d_ret']   = tlt.pct_change(5)
        feat['tlt_20d_ret']  = tlt.pct_change(20)
        feat['tlt_corr_20d'] = cl.pct_change().rolling(20).corr(tlt.pct_change())
        feat['macro_xlk_tlt_diff'] = cl.pct_change(20) - tlt.pct_change(20)

    # ── 7. 가격 패턴 / 캔들 ─────────────────────────────
    body     = cl - op
    body_pct = body / op.replace(0, np.nan)
    rng      = hi - lo
    up_shad  = hi - pd.concat([cl, op], axis=1).max(axis=1)
    dn_shad  = pd.concat([cl, op], axis=1).min(axis=1) - lo

    feat['body_size_pct']    = body_pct.abs()
    feat['body_direction']   = np.sign(body_pct)
    feat['shadow_upper_pct'] = up_shad / op.replace(0, np.nan)
    feat['shadow_lower_pct'] = dn_shad / op.replace(0, np.nan)
    feat['shadow_ratio']     = up_shad / dn_shad.replace(0, np.nan)
    feat['candle_size_pct']  = rng / op.replace(0, np.nan)
    feat['body_to_range']    = body.abs() / rng.replace(0, np.nan)

    feat['doji_flag']          = (body_pct.abs() < 0.002).astype(float)
    feat['hammer_flag']        = ((dn_shad > 2 * body.abs()) & (up_shad < body.abs())).astype(float)
    feat['shooting_star_flag'] = ((up_shad > 2 * body.abs()) & (dn_shad < body.abs())).astype(float)

    prev_body = body.shift()
    feat['engulf_bull'] = ((body > 0) & (prev_body < 0) & (op < cl.shift()) & (cl > op.shift())).astype(float)
    feat['engulf_bear'] = ((body < 0) & (prev_body > 0) & (op > cl.shift()) & (cl < op.shift())).astype(float)
    feat['harami_bull'] = ((body > 0) & (body.shift() < 0) & (cl < op.shift()) & (op > cl.shift())).astype(float)

    feat['gap_up']   = (op / cl.shift() - 1).clip(lower=0)
    feat['gap_down'] = (op / cl.shift() - 1).clip(upper=0).abs()
    feat['gap_net']  = op / cl.shift() - 1
    feat['gap_fill'] = ((cl - op) * np.sign(op - cl.shift())).clip(lower=0) / rng.replace(0, np.nan)

    pivot = (hi.shift() + lo.shift() + cl.shift()) / 3
    feat['pivot_dist']    = cl / pivot.replace(0, np.nan) - 1
    feat['pivot_r1_dist'] = cl / (2 * pivot - lo.shift()).replace(0, np.nan) - 1
    feat['pivot_s1_dist'] = cl / (2 * pivot - hi.shift()).replace(0, np.nan) - 1

    for p in [1, 2, 3, 5]:
        feat[f'pattern_vs_{p}ago'] = (cl - cl.shift(p)) / (hi.shift(p) - lo.shift(p)).replace(0, np.nan)

    sign_ret = np.sign(cl.diff())
    streak = sign_ret.groupby((sign_ret != sign_ret.shift()).cumsum()).cumcount() + 1
    feat['up_streak']   = streak.where(sign_ret > 0, 0)
    feat['down_streak'] = streak.where(sign_ret < 0, 0)

    # ── 8. 추가 지표 ──────────────────────────────────
    for p in [9, 13, 20, 26]:
        ema_p = cl.ewm(span=p, adjust=False).mean()
        feat[f'elder_bull_{p}'] = hi - ema_p
        feat[f'elder_bear_{p}'] = lo - ema_p

    for p in [10, 14]:
        hl_ema = (hi - lo).ewm(span=p, adjust=False).mean()
        feat[f'chaikin_vol_{p}'] = (hl_ema - hl_ema.shift(p)) / hl_ema.shift(p).replace(0, np.nan)

    for base, chg in [(20, 5), (60, 20)]:
        hv = cl.pct_change().rolling(base).std() * np.sqrt(252)
        feat[f'hist_vol_{base}_roc_{chg}'] = (hv - hv.shift(chg)) / hv.shift(chg).replace(0, np.nan)

    # NVI / PVI
    nvi = pd.Series(1000.0, index=cl.index)
    pvi = pd.Series(1000.0, index=cl.index)
    for i in range(1, len(cl)):
        r = (cl.iloc[i] - cl.iloc[i - 1]) / cl.iloc[i - 1] if cl.iloc[i - 1] != 0 else 0
        if vo.iloc[i] < vo.iloc[i - 1]:
            nvi.iloc[i] = nvi.iloc[i - 1] * (1 + r)
        else:
            nvi.iloc[i] = nvi.iloc[i - 1]
        if vo.iloc[i] > vo.iloc[i - 1]:
            pvi.iloc[i] = pvi.iloc[i - 1] * (1 + r)
        else:
            pvi.iloc[i] = pvi.iloc[i - 1]
    feat['nvi_255_dist'] = nvi / nvi.rolling(255).mean() - 1
    feat['pvi_255_dist'] = pvi / pvi.rolling(255).mean() - 1

    for lbl, sym in [('xlv', 'XLV'), ('xle', 'XLE')]:
        if sym == TICKER: continue   # 자기비교 방지
        peer = closes.get(sym)
        if peer is None: continue
        for p in [5, 10, 20, 60]:
            feat[f'rel_{lbl}_{p}d'] = cl.pct_change(p) - peer.pct_change(p)
        feat[f'corr_{lbl}_20d'] = cl.pct_change().rolling(20).corr(peer.pct_change())

    for p in [20, 60, 252]:
        feat[f'price_zscore_{p}'] = calc_zscore(cl, p)

    for p in [20, 60]:
        feat[f'ret_skew_{p}'] = cl.pct_change().rolling(p).skew()
        feat[f'ret_kurt_{p}'] = cl.pct_change().rolling(p).kurt()

    feat['down_streak_zscore'] = calc_zscore(feat['down_streak'], 60)
    feat['up_streak_zscore']   = calc_zscore(feat['up_streak'], 60)

    # ══════════════════════════════════════════════════
    #  9. 하락 예측 특화 지표 (DROP PREDICTION SUITE)
    # ══════════════════════════════════════════════════

    # 9a. 드로다운 스위트 — "고점 대비 얼마나 빠졌는가"
    for p in [10, 20, 60, 120]:
        rolling_max_p = cl.rolling(p).max()
        feat[f'dd_from_{p}d_high']    = cl / rolling_max_p - 1           # 음수(-) / 0
        feat[f'days_since_{p}d_high'] = cl.rolling(p).apply(
            lambda x: float(len(x) - 1 - np.argmax(x)), raw=True
        )
    # 30일 롤링 윈도우 내 최대 드로다운
    def _max_dd(x):
        cum = (1 + x).cumprod()
        return float(cum.div(cum.cummax()).sub(1).min())
    feat['max_dd_30d'] = cl.pct_change().rolling(30).apply(_max_dd, raw=False)
    feat['max_dd_60d'] = cl.pct_change().rolling(60).apply(_max_dd, raw=False)

    # 9b. VIX 레짐
    if '^VIX' in closes.columns:
        vix = closes['^VIX']
        feat['vix_above_sma50']  = (vix > vix.rolling(50).mean()).astype(float)
        feat['vix_above_sma20']  = (vix > vix.rolling(20).mean()).astype(float)
        feat['vix_gt_20']        = (vix > 20).astype(float)
        feat['vix_gt_25']        = (vix > 25).astype(float)
        feat['vix_gt_30']        = (vix > 30).astype(float)
        feat['vix_of_vix_20']    = vix.pct_change().rolling(20).std() * np.sqrt(252)
        feat['vix_mom_5_20']     = vix.pct_change(5) / vix.rolling(20).std().replace(0, np.nan)
        feat['vix_accel_5']      = vix.pct_change(5) - vix.pct_change(10)   # 가속도

    # 9c. 크로스자산 스트레스 (안전자산 상대강세 = 위험회피)
    if 'TLT' in closes.columns:
        for p in [5, 10, 20]:
            feat[f'tlt_lead_{p}d'] = closes['TLT'].pct_change(p) - cl.pct_change(p)
    if 'GLD' in closes.columns:
        for p in [5, 10, 20]:
            feat[f'gld_lead_{p}d'] = closes['GLD'].pct_change(p) - cl.pct_change(p)

    # 위험회피 복합 (TLT↑ + GLD↑ + VIX↑ 동시발생 카운트)
    if all(s in closes.columns for s in ['TLT', 'GLD', '^VIX']):
        tlt_up = (closes['TLT'].pct_change(5) > 0).astype(float)
        gld_up = (closes['GLD'].pct_change(5) > 0).astype(float)
        vix_up = (closes['^VIX'].pct_change(5) > 0).astype(float)
        feat['risk_off_composite_5d']  = tlt_up + gld_up + vix_up
        tlt_up20 = (closes['TLT'].pct_change(20) > 0).astype(float)
        gld_up20 = (closes['GLD'].pct_change(20) > 0).astype(float)
        vix_up20 = (closes['^VIX'].pct_change(20) > 0).astype(float)
        feat['risk_off_composite_20d'] = tlt_up20 + gld_up20 + vix_up20

    # 금리 상승 + 주식 약세 동시 (매우 위험한 조합)
    if '^TNX' in closes.columns:
        tnx = closes['^TNX']
        tnx_rising = (tnx.diff(5) > 0).astype(float)
        spy_weak   = (cl.pct_change(5) < 0).astype(float)
        feat['yield_up_equity_down'] = tnx_rising * spy_weak

    # 9d. 모멘텀 다이버전스 (가격은 버티는데 모멘텀 죽는 경우)
    rsi14_s = calc_rsi(cl, 14)
    for p in [10, 20]:
        price_at_high = (cl >= cl.rolling(p).max() * 0.995).astype(float)  # 고점 부근
        rsi_below     = (rsi14_s < rsi14_s.rolling(p).max().shift(1)).astype(float)
        feat[f'bear_div_rsi_{p}'] = price_at_high * rsi_below

    # 모멘텀 감속 (단기 < 장기)
    feat['mom_decel_5_20']  = cl.pct_change(5)  - cl.pct_change(20)
    feat['mom_decel_10_60'] = cl.pct_change(10) - cl.pct_change(60)
    # 가격 2차 도함수 (가속 → 감속)
    sma5 = cl.rolling(5).mean()
    feat['price_accel_5']  = sma5.diff(5) - sma5.diff(5).shift(5)

    # 9e. 마켓 마이크로구조 — 약한 종가
    close_in_range = (cl - lo) / (hi - lo).replace(0, np.nan)
    feat['close_range_pos']       = close_in_range
    feat['close_range_pos_5ma']   = close_in_range.rolling(5).mean()
    feat['weak_close_ratio_10']   = (close_in_range < 0.3).astype(float).rolling(10).mean()
    feat['weak_close_ratio_20']   = (close_in_range < 0.3).astype(float).rolling(20).mean()
    # 거래량 동반 하락 (distribution day)
    heavy_red = ((cl < op) & (vo > vo.rolling(20).mean() * 1.3)).astype(float)
    feat['distribution_days_20']  = heavy_red.rolling(20).sum()
    feat['distribution_days_5']   = heavy_red.rolling(5).sum()

    # 9f. 섹터 분산/브레드스 (시장 내부 건강도)
    _sector_syms = [s for s in ['XLK', 'XLF', 'XLV', 'XLE', 'XLC']
                    if s in closes.columns and s != TICKER]
    if len(_sector_syms) >= 3:
        _sec_df_5  = pd.DataFrame({s: closes[s].pct_change(5)  for s in _sector_syms})
        _sec_df_20 = pd.DataFrame({s: closes[s].pct_change(20) for s in _sector_syms})
        feat['sector_dispersion_5d']   = _sec_df_5.std(axis=1)
        feat['sector_dispersion_20d']  = _sec_df_20.std(axis=1)
        # 섹터별 52주 고점 대비 평균 낙폭 (건강도)
        _sec_dd = pd.DataFrame({
            s: closes[s] / closes[s].rolling(252).max() - 1
            for s in _sector_syms
        })
        feat['sector_avg_dd_52w']   = _sec_dd.mean(axis=1)
        feat['sector_worst_dd_52w'] = _sec_dd.min(axis=1)
        # 20일 고점 근처 섹터 수 (많을수록 건강)
        _sec_near20 = pd.DataFrame({
            s: (closes[s] / closes[s].rolling(20).max() - 1 > -0.02).astype(float)
            for s in _sector_syms
        })
        feat['sector_near_20d_high_cnt'] = _sec_near20.sum(axis=1)

    # ══════════════════════════════════════════════════
    #  10. 11개 섹터 ETF 브레드스 + 매크로 + 베어랠리 구분
    #      (~117개 신규 지표, 기존 지표와 중복 없음)
    # ══════════════════════════════════════════════════
    SEC11 = ['XLK','XLV','XLF','XLY','XLP','XLE','XLI','XLB','XLU','XLRE','XLC']
    sec_avail = [s for s in SEC11 if s in closes.columns]

    # ── 10A. 11개 섹터 브레드스 (33개) ───────────────────
    if len(sec_avail) >= 5:
        sec_close = pd.DataFrame({s: closes[s] for s in sec_avail})

        # 섹터들이 SMA 위에 있는 비율/카운트
        for ma_p in [20, 50, 200]:
            above_ma = pd.DataFrame({
                s: (sec_close[s] > sec_close[s].rolling(ma_p).mean()).astype(float)
                for s in sec_avail
            })
            feat[f'sectors_above_sma{ma_p}_count'] = above_ma.sum(axis=1)
            feat[f'sectors_above_sma{ma_p}_pct']   = above_ma.mean(axis=1)

        # n일 양의 수익률 섹터 수/비율
        for p in [5, 10, 20, 60]:
            pos_ret = pd.DataFrame({
                s: (sec_close[s].pct_change(p) > 0).astype(float) for s in sec_avail
            })
            feat[f'sectors_pos_{p}d_count'] = pos_ret.sum(axis=1)
            feat[f'sectors_pos_{p}d_pct']   = pos_ret.mean(axis=1)

        # McClellan 진동자 근사: (advances - declines) EMA
        sec_chg = sec_close.pct_change()
        adv_dec = (sec_chg > 0).astype(float).sum(axis=1) - \
                  (sec_chg < 0).astype(float).sum(axis=1)
        feat['sectors_breadth_McClellan'] = (
            adv_dec.ewm(span=19, adjust=False).mean()
            - adv_dec.ewm(span=39, adjust=False).mean()
        )
        # AD 라인 z-score
        ad_line_sec = adv_dec.cumsum()
        feat['sectors_advance_decline_line'] = calc_zscore(ad_line_sec, 60)

        # 상승/하락 섹터 비율
        for p in [5, 20]:
            up_n   = (sec_close.pct_change(p) > 0).astype(float).sum(axis=1)
            down_n = (sec_close.pct_change(p) < 0).astype(float).sum(axis=1)
            feat[f'sectors_ad_ratio_{p}d'] = up_n / (up_n + down_n).replace(0, np.nan)

        # 신고가/신저가 섹터 수
        for p in [20, 60]:
            new_h = pd.DataFrame({
                s: (sec_close[s] >= sec_close[s].rolling(p).max() - 1e-9).astype(float)
                for s in sec_avail
            })
            new_l = pd.DataFrame({
                s: (sec_close[s] <= sec_close[s].rolling(p).min() + 1e-9).astype(float)
                for s in sec_avail
            })
            feat[f'sectors_new_{p}d_high_count'] = new_h.sum(axis=1)
            feat[f'sectors_new_{p}d_low_count']  = new_l.sum(axis=1)
            feat[f'sectors_high_low_diff_{p}']   = new_h.sum(axis=1) - new_l.sum(axis=1)

        # 섹터별 최근 고점 대비 낙폭의 중간값/평균
        sec_dd = pd.DataFrame({
            s: sec_close[s] / sec_close[s].rolling(252).max() - 1
            for s in sec_avail
        })
        feat['sectors_median_dd_from_high'] = sec_dd.median(axis=1)

        for p in [20, 60]:
            sec_dd_p = pd.DataFrame({
                s: sec_close[s] / sec_close[s].rolling(p).max() - 1
                for s in sec_avail
            })
            feat[f'sectors_max_drawdown_{p}d'] = sec_dd_p.mean(axis=1)

        # 섹터간 평균 상관 (위기때 1로 수렴)
        sec_chg_d = sec_close.pct_change()
        for p in [20, 60]:
            corr_mat = sec_chg_d.rolling(p).corr().dropna()
            # 평균 상관: 매일 11x11 상관행렬의 비대각 평균
            def _avg_offdiag_corr(group):
                m = group.values
                n_ = m.shape[0]
                if n_ < 2: return np.nan
                mask = ~np.eye(n_, dtype=bool)
                return float(np.nanmean(m[mask]))
            # 빠른 근사: 각 섹터 vs 시장평균 상관의 평균
            sec_mean = sec_chg_d.mean(axis=1)
            corrs = pd.DataFrame({
                s: sec_chg_d[s].rolling(p).corr(sec_mean) for s in sec_avail
            })
            feat[f'sectors_corr_avg_{p}'] = corrs.mean(axis=1)

        # 섹터 분산 z-score (60일 기준)
        sec_disp_60 = sec_chg_d.rolling(60).std().mean(axis=1)
        feat['sectors_dispersion_zscore_60'] = calc_zscore(sec_disp_60, 60)

        # 섹터 수익률 분포 왜도
        feat['sectors_skew_returns_20d'] = sec_chg_d.rolling(20).mean().apply(
            lambda x: pd.Series(x).skew(), axis=1
        ) if False else sec_close.pct_change(20).apply(
            lambda x: pd.Series(x).skew() if x.notna().any() else np.nan, axis=1
        )

        # 최고-최저 섹터 갭
        for p in [5, 20]:
            sec_ret_p = sec_close.pct_change(p)
            feat[f'sectors_top_minus_bottom_{p}d'] = (
                sec_ret_p.max(axis=1) - sec_ret_p.min(axis=1)
            )

    # ── 10B. 방어주 vs 공격주 로테이션 (12개) ─────────────
    DEFENSIVE  = [s for s in ['XLU', 'XLP', 'XLV']        if s in closes.columns]
    AGGRESSIVE = [s for s in ['XLK', 'XLY', 'XLF', 'XLC'] if s in closes.columns]

    if len(DEFENSIVE) >= 2 and len(AGGRESSIVE) >= 2:
        def_ret = pd.concat([closes[s].pct_change() for s in DEFENSIVE], axis=1).mean(axis=1)
        agg_ret = pd.concat([closes[s].pct_change() for s in AGGRESSIVE], axis=1).mean(axis=1)

        for p in [5, 20, 60]:
            def_p = (1 + def_ret).rolling(p).apply(np.prod, raw=True) - 1
            agg_p = (1 + agg_ret).rolling(p).apply(np.prod, raw=True) - 1
            feat[f'defensive_aggressive_ratio_{p}d'] = def_p - agg_p

        # 방어주 선도 (5/20일)
        for p in [5, 20]:
            def_lead = ((def_ret.rolling(p).mean() > 0) &
                        (agg_ret.rolling(p).mean() < 0)).astype(float)
            feat[f'defensive_lead_{p}d'] = def_lead.rolling(p).mean()

        # 개별 비율
        if 'XLU' in closes.columns and 'XLK' in closes.columns:
            r = closes['XLU'] / closes['XLK']
            feat['xlu_xlk_ratio_dist20'] = r / r.rolling(20).mean() - 1
            feat['xlu_outperform_60d']   = (closes['XLU'].pct_change(60) -
                                             closes['XLK'].pct_change(60))
        if 'XLP' in closes.columns and 'XLY' in closes.columns:
            r = closes['XLP'] / closes['XLY']
            feat['xlp_xly_ratio_dist20'] = r / r.rolling(20).mean() - 1
            feat['xlp_outperform_60d']   = (closes['XLP'].pct_change(60) -
                                             closes['XLY'].pct_change(60))
        if 'XLV' in closes.columns and 'XLK' in closes.columns:
            r = closes['XLV'] / closes['XLK']
            feat['xlv_xlk_ratio_dist20'] = r / r.rolling(20).mean() - 1

        # 종합 로테이션 점수 (0~3, 높을수록 방어주 강세 = 위험회피)
        score = pd.Series(0.0, index=cl.index)
        if 'XLU' in closes.columns and 'XLK' in closes.columns:
            score += (closes['XLU'].pct_change(20) > closes['XLK'].pct_change(20)).astype(float)
        if 'XLP' in closes.columns and 'XLY' in closes.columns:
            score += (closes['XLP'].pct_change(20) > closes['XLY'].pct_change(20)).astype(float)
        if 'XLV' in closes.columns and 'XLK' in closes.columns:
            score += (closes['XLV'].pct_change(20) > closes['XLK'].pct_change(20)).astype(float)
        feat['rotation_to_defensive_score'] = score

        # SPLV vs SPHB (저변동성 vs 고베타)
        if 'SPLV' in closes.columns and 'SPHB' in closes.columns:
            feat['low_vol_high_beta_spread_20d'] = (closes['SPLV'].pct_change(20) -
                                                     closes['SPHB'].pct_change(20))

    # ── 10C. 베어마켓 랠리 vs 진짜 상승 구분 (16개) ────────
    sma200 = cl.rolling(200).mean()
    in_downtrend = (cl < sma200).astype(float)

    # 상승일 거래량 강도
    up_day = (cl.pct_change() > 0).astype(float)
    for p in [5, 10]:
        rally_vol = (vo * up_day).rolling(p).mean() / vo.rolling(p).mean().replace(0, np.nan)
        feat[f'rally_volume_strength_{p}'] = rally_vol

    # 상승일에 같이 오른 섹터 수 (브레드스 품질)
    if len(sec_avail) >= 5:
        sec_chg_for_rally = sec_close.pct_change()
        sectors_up_today = (sec_chg_for_rally > 0).astype(float).sum(axis=1)
        for p in [5, 10]:
            # 시장이 오른 날의 섹터 동조도 평균
            spy_up = (cl.pct_change() > 0).astype(float)
            quality = (sectors_up_today * spy_up).rolling(p).sum() / \
                      spy_up.rolling(p).sum().replace(0, np.nan)
            feat[f'rally_breadth_quality_{p}'] = quality

    # 200일선 아래에서의 상승 = 의심스러움
    feat['rally_in_downtrend_flag'] = ((cl.pct_change(5) > 0) & (cl < sma200)).astype(float)

    # 약한 거래량 상승 (가짜 상승 신호)
    vol_below_avg = (vo < vo.rolling(20).mean()).astype(float)
    weak_up = up_day * vol_below_avg
    for p in [5, 10]:
        feat[f'low_volume_rally_{p}'] = weak_up.rolling(p).sum()

    # 고VIX 환경에서 상승 (반등이 위험할수록)
    if '^VIX' in closes.columns:
        vix_high = (closes['^VIX'] > 20).astype(float)
        rally_high_vix = up_day * vix_high
        for p in [5, 10]:
            feat[f'rally_with_high_vix_{p}'] = rally_high_vix.rolling(p).sum()

    # 베어 랠리 종합 점수 (0~5)
    bear_score = pd.Series(0.0, index=cl.index)
    bear_score += in_downtrend
    bear_score += (cl.pct_change(5) > 0).astype(float) * in_downtrend  # 하락 추세 중 단기 상승
    if 'rally_volume_strength_5' in feat.columns:
        bear_score += (feat['rally_volume_strength_5'] < 0.9).astype(float)
    if '^VIX' in closes.columns:
        bear_score += (closes['^VIX'] > 20).astype(float)
    bear_score += (cl < cl.rolling(50).mean()).astype(float)
    feat['bear_market_rally_score'] = bear_score

    # 반등 크기 / 직전 낙폭 (작으면 = 약한 반등)
    rolling_min_20 = cl.rolling(20).min()
    rolling_max_20 = cl.rolling(20).max()
    bounce_size = cl - rolling_min_20
    prior_dd = rolling_max_20 - rolling_min_20
    feat['rally_vs_dd_ratio'] = bounce_size / prior_dd.replace(0, np.nan)

    # 더 낮은 고점 (lower high) 카운트
    rolling_max_5  = cl.rolling(5).max()
    lower_high = (rolling_max_5 < rolling_max_5.shift(5)).astype(float)
    feat['lower_high_count_60'] = lower_high.rolling(60).sum()

    # 20일 고점에서 거부 (REJECTION) - 과거 정보만 사용
    # 원래 버그: shift(-3)로 미래 3일치 가격을 봐서 누설 → 수정
    # 새 정의: 오늘이 20일 고점 근처(99% 이상)면서 종가가 시가보다 낮음 (intraday 거부)
    near_20d_high = (cl >= cl.rolling(20).max() * 0.99)
    rejected_today = near_20d_high & (cl < op)
    feat['rally_failed_at_resistance_20'] = rejected_today.astype(float).rolling(20).sum()

    # 좁은 반등 (브레드스 약함)
    if 'sectors_pos_5d_pct' in feat.columns:
        feat['narrow_rally_advance_decline_5d'] = (
            (cl.pct_change(5) > 0).astype(float) *
            (1 - feat['sectors_pos_5d_pct'])
        )

    # IBD followthrough day 근사: 강한 거래량 + 큰 상승
    big_up = (cl.pct_change() > 0.0125).astype(float)
    high_vol = (vo > vo.rolling(50).mean() * 1.05).astype(float)
    feat['rally_followthrough_day'] = big_up * high_vol

    # 진짜 반등 점수 (높을수록 진짜)
    real_score = pd.Series(0.0, index=cl.index)
    real_score += (cl > sma200).astype(float)
    real_score += (cl > cl.rolling(50).mean()).astype(float)
    if 'sectors_pos_5d_pct' in feat.columns:
        real_score += (feat['sectors_pos_5d_pct'] > 0.6).astype(float)
    if '^VIX' in closes.columns:
        real_score += (closes['^VIX'] < 20).astype(float)
    real_score += (vo > vo.rolling(20).mean()).astype(float) * up_day
    feat['real_rally_score'] = real_score

    # ── 10D. 매크로/금리 심화 (12개) ──────────────────────
    if '^TNX' in closes.columns and '^IRX' in closes.columns:
        feat['yield_curve_2s10s_proxy_chg5'] = (closes['^TNX'] - closes['^IRX']).diff(5)

    if '^TNX' in closes.columns and 'GLD' in closes.columns:
        feat['real_yield_proxy_dist20'] = (
            closes['^TNX'].pct_change(20) - closes['GLD'].pct_change(20)
        )

    if 'HYG' in closes.columns and 'LQD' in closes.columns:
        for p in [5, 20]:
            # HYG 약세 - LQD 약세 = 신용 스프레드 확대
            feat[f'credit_spread_proxy_{p}d'] = (
                closes['LQD'].pct_change(p) - closes['HYG'].pct_change(p)
            )

    if 'UUP' in closes.columns:
        for p in [5, 20]:
            feat[f'dxy_proxy_chg_{p}'] = closes['UUP'].pct_change(p)

    if 'USO' in closes.columns:
        for p in [5, 20]:
            feat[f'oil_chg_{p}d'] = closes['USO'].pct_change(p)

    if 'GLD' in closes.columns and 'DBC' in closes.columns:
        # Copper/Gold 대신 DBC/GLD 사용
        cu_au = closes['DBC'] / closes['GLD']
        feat['copper_gold_ratio_chg'] = cu_au.pct_change(20)

    if '^TNX' in closes.columns:
        feat['tnx_breakout_60d'] = (
            closes['^TNX'] >= closes['^TNX'].rolling(60).max() * 0.999
        ).astype(float)

    if '^IRX' in closes.columns:
        feat['fed_pivot_proxy'] = -closes['^IRX'].diff(20)  # IRX 급락 = 비둘기파

    if '^TNX' in closes.columns:
        feat['stagflation_score'] = (
            (closes['^TNX'].diff(20) > 0).astype(float) *
            (cl.pct_change(20) < 0).astype(float)
        )

    # ── 10E. 시장 미시구조 (18개) ─────────────────────────
    # 고점 돌파 후 실패
    breakout = (hi >= hi.rolling(20).max().shift(1))
    bk_fail  = breakout & (cl < op)
    feat['high_low_breakout_failure_20'] = bk_fail.astype(float).rolling(20).sum()
    intraday_rev = ((op - cl).abs() > (hi - lo) * 0.7).astype(float)
    for p in [5, 10]:
        feat[f'intraday_reversal_{p}d'] = intraday_rev.rolling(p).sum()

    # OHLC 약세 패턴 (시가 = 고가, 종가 = 저가 부근)
    open_high_close_low = ((op >= hi * 0.998) & (cl <= lo * 1.002) & (cl < op)).astype(float)
    for p in [5, 10]:
        feat[f'open_low_close_red_{p}'] = open_high_close_low.rolling(p).sum()

    # 꼬리 위험 점수 (왜도+첨도 결합)
    ret_d = cl.pct_change()
    feat['tail_risk_score_20'] = -ret_d.rolling(20).skew() + ret_d.rolling(20).kurt()

    # 변동성 군집 (ARCH 효과)
    feat['vol_clustering_score'] = (ret_d.abs().rolling(5).mean() /
                                     ret_d.abs().rolling(60).mean().replace(0, np.nan))

    # 추세 효율성 (Kaufman ER)
    for p in [10, 20]:
        direction = (cl - cl.shift(p)).abs()
        volatility = cl.diff().abs().rolling(p).sum()
        feat[f'kaufman_efficiency_{p}'] = direction / volatility.replace(0, np.nan)

    for p in [20, 60]:
        net_move = (cl - cl.shift(p)).abs()
        path_len = cl.diff().abs().rolling(p).sum()
        feat[f'trend_efficiency_{p}'] = net_move / path_len.replace(0, np.nan)

    # Hurst 지수 근사 (간단한 R/S 비)
    def _hurst_approx(x):
        if len(x) < 10 or np.isnan(x).any(): return np.nan
        mean_x = np.mean(x); dev = np.cumsum(x - mean_x)
        R = np.max(dev) - np.min(dev); S = np.std(x)
        return float(R / (S + 1e-12)) / np.log(len(x))
    feat['fractal_dim_estimate_20'] = ret_d.rolling(20).apply(_hurst_approx, raw=True)

    # 범위 확장
    today_range = hi - lo
    for p in [5, 10]:
        feat[f'range_expansion_{p}'] = today_range / today_range.rolling(p).mean().replace(0, np.nan)

    # NR4/NR7 (좁은 범위)
    for p in [4, 7]:
        feat[f'narrow_range_{p}'] = (today_range <= today_range.rolling(p).min() + 1e-9).astype(float)

    # 내부일 (전일 high/low 안에 있음)
    inside = ((hi <= hi.shift()) & (lo >= lo.shift())).astype(float)
    feat['inside_day_count_10'] = inside.rolling(10).sum()
    outside = ((hi > hi.shift()) & (lo < lo.shift())).astype(float)
    feat['outside_day_count_10'] = outside.rolling(10).sum()

    # ── 10F. 옵션/공포 심화 (7개) ────────────────────────
    if '^VIX' in closes.columns:
        vix = closes['^VIX']
        # VIX 만기 곡선 근사 (5일 이평/20일 이평)
        feat['vix_curve_proxy'] = vix.rolling(5).mean() / vix.rolling(20).mean().replace(0, np.nan)
        feat['vix_zscore_252'] = calc_zscore(vix, 252)
        feat['vix_pctile_504'] = calc_pctrank(vix, 504)
        # VIX와 가격 디커플링 (정상은 음의 상관, 깨지면 위험)
        feat['vix_disconnect_20d'] = (cl.pct_change().rolling(20).corr(vix.pct_change()) + 1)

    # 수익률 분포 왜도 (옵션 skew 프록시)
    feat['skew_proxy_via_returns'] = -ret_d.rolling(60).skew()

    # 공포-탐욕 종합 (0~100)
    fg = pd.Series(50.0, index=cl.index)
    if '^VIX' in closes.columns:
        fg -= (closes['^VIX'] - 15) * 1.5
    fg += cl.pct_change(20) * 100 * 2
    feat['fear_greed_composite'] = fg.clip(0, 100)

    # 거래량 패턴으로 풋콜 비율 추론
    feat['put_call_proxy_via_volume'] = (
        (vo * (cl < op)).rolling(20).sum() /
        (vo * (cl > op)).rolling(20).sum().replace(0, np.nan)
    )

    # ── 10G. 레짐 (8개) ─────────────────────────────────
    for p in [20, 60]:
        # 추세장 강도 = 절대 누적 수익률 / 변동성
        cum_ret = (cl / cl.shift(p) - 1).abs()
        vol_p   = ret_d.rolling(p).std() * np.sqrt(p)
        feat[f'regime_trending_score_{p}'] = cum_ret / vol_p.replace(0, np.nan)
        # 평균회귀 강도 = 1/추세 (높을수록 평균회귀)
        feat[f'regime_meanrev_score_{p}'] = vol_p / (cum_ret + 1e-6)

    # Hurst proxy 60일
    feat['hurst_proxy_60'] = ret_d.rolling(60).apply(_hurst_approx, raw=True)

    # 추세 품질 (ADX 변형)
    for p in [20, 60]:
        feat[f'trend_quality_{p}'] = (cl - cl.shift(p)).abs() / (
            (hi.rolling(p).max() - lo.rolling(p).min()).replace(0, np.nan)
        )

    # 레짐 변화 신호 (20일 표준편차의 표준편차)
    feat['regime_change_signal_20'] = ret_d.rolling(20).std().rolling(20).std()

    # ── 10H. 다이버전스 / 이상 신호 (6개) ────────────────
    # 가격↑ 거래량↓ (전형적 약세 다이버전스)
    feat['price_volume_div_20'] = (
        (cl.pct_change(20) > 0).astype(float) *
        (vo.rolling(20).mean() < vo.rolling(60).mean()).astype(float)
    )

    # 가격↑ 섹터브레드스↓
    if 'sectors_pos_20d_pct' in feat.columns:
        for p in [20, 60]:
            key = f'sectors_pos_{p if p in [5,10,20,60] else 20}d_pct'
            if key in feat.columns:
                feat[f'price_breadth_div_{p}'] = (
                    (cl.pct_change(p) > 0).astype(float) *
                    (feat[key] < 0.5).astype(float)
                )

    # VIX-SPX 음의 상관 깨짐
    if '^VIX' in closes.columns:
        feat['vix_spx_corr_breakdown_20'] = (
            cl.pct_change().rolling(20).corr(closes['^VIX'].pct_change()) > 0
        ).astype(float)

    # RSI 다이버전스 (가격 신고가, RSI 신고가 못만듦)
    rsi14 = calc_rsi(cl, 14)
    price_hh = (cl >= cl.rolling(20).max() - 1e-9).astype(float)
    rsi_no_hh = (rsi14 < rsi14.rolling(20).max().shift(5)).astype(float)
    feat['rsi_divergence_higher_high_20'] = (price_hh * rsi_no_hh).rolling(20).sum()

    # MACD 히스토그램 약화
    macd_line = cl.ewm(span=12, adjust=False).mean() - cl.ewm(span=26, adjust=False).mean()
    macd_sig  = macd_line.ewm(span=9, adjust=False).mean()
    macd_hist = macd_line - macd_sig
    feat['macd_histogram_decay_count'] = (
        (macd_hist.diff() < 0) & (macd_hist > 0)
    ).astype(float).rolling(10).sum()

    # ── 10I. 계절성 / 시점 (5개) ─────────────────────────
    feat['month_of_year']        = pd.Series(cl.index.month, index=cl.index).astype(float)
    feat['day_of_month']         = pd.Series(cl.index.day,   index=cl.index).astype(float)
    # 거래일 of month
    bd_of_month = (
        pd.Series(1, index=cl.index)
        .groupby([cl.index.year, cl.index.month])
        .cumsum()
    )
    feat['trading_day_of_month'] = bd_of_month.astype(float)
    # 요일 더미
    dow = pd.Series(cl.index.dayofweek, index=cl.index)
    feat['is_monday_or_friday']  = ((dow == 0) | (dow == 4)).astype(float)
    # FOMC 6주 사이클 근사 (대략적인 위치)
    days_since_2000 = (cl.index - pd.Timestamp('2000-01-01')).days
    feat['days_to_fomc_proxy']   = pd.Series(days_since_2000 % 42, index=cl.index).astype(float)

    # ══════════════════════════════════════════════════
    #  11. 이벤트 충격 선행 지표 (Event Precursors)
    #  ── 외부 조사로 검증된 FN 패턴 분석 결과 추가
    #  · 2025-04-02 Liberation Day, 2025-01-27 DeepSeek,
    #    2025-10-10 Mag7 -$770B 등 이벤트 직전 패턴
    # ══════════════════════════════════════════════════

    # 11a. 시장 복잡도 지수 (Complacency / Fragility Index)
    # 이론: 큰 충격 직전엔 변동성이 비정상적으로 낮고 "너무 평온"한 상태가 많음
    # → 낮은 변동성 + 높은 가격 + 높은 모멘텀이 동시 발생 = 취약함
    if '^VIX' in closes.columns:
        vix = closes['^VIX']
        # VIX 60일 분위수 + 가격 상승 모멘텀 + 거래량 평온
        vix_low_pctile = 1 - calc_pctrank(vix, 60)   # VIX 저분위 (높을수록 안심)
        price_high_pctile = calc_pctrank(cl, 60)     # 가격 고분위
        vol_low = (vo < vo.rolling(60).mean()).astype(float)
        # 세 조건이 동시에 충족 → "복잡도" (취약함의 척도)
        feat['complacency_score'] = vix_low_pctile.fillna(0) * price_high_pctile.fillna(0) * vol_low

    # 11b. 동시 가격 분기 (Coiling Pattern)
    # 이론: 변동성 압축이 길어질수록 큰 방향성 움직임 발생 가능성 ↑
    # 수십일간 좁은 박스권이면 "터질 준비" 상태
    # 변동성 / 변동성의 60일 평균 — 1보다 작을수록 압축
    vol_5d  = cl.pct_change().rolling(5).std()
    vol_60d = cl.pct_change().rolling(60).std()
    feat['volatility_compression'] = (1 - vol_5d / vol_60d.replace(0, np.nan)).clip(0, 1)
    # 30일 가격 범위 / ATR — 좁을수록 압축
    price_range_30 = cl.rolling(30).max() - cl.rolling(30).min()
    atr_30 = (hi - lo).rolling(30).mean()
    feat['range_compression_30'] = price_range_30 / (atr_30 * 30).replace(0, np.nan)

    # 11c. 비대칭 위험 점수 (Asymmetric Risk)
    # 이론: 상승은 천천히, 하락은 빠르게 → 큰 충격 직전엔 일일 음봉 빈도가 점진적으로 늘어남
    # 최근 20일 음봉 비율 vs 최근 60일 음봉 비율
    down_day = (cl.pct_change() < 0).astype(float)
    feat['negative_day_acceleration'] = (
        down_day.rolling(20).mean() - down_day.rolling(60).mean()
    )
    # 일평균 음봉 크기 vs 양봉 크기 — 음봉이 더 크면 위험
    ret_s = cl.pct_change()
    avg_down = ret_s.where(ret_s < 0).abs().rolling(20, min_periods=3).mean()
    avg_up   = ret_s.where(ret_s > 0).rolling(20, min_periods=3).mean()
    feat['down_up_size_ratio_20'] = avg_down / avg_up.replace(0, np.nan)

    # 11d. 매그니피센트7 의존도 + 집중도 위험
    # XLK는 NVDA, AAPL, MSFT 3개로 38%+ 비중. 이들 중 하나가 흔들리면 ETF 전체 위험
    # 프록시: SMH(반도체) vs XLK 일별 수익률 차이의 변동성
    if 'SMH' in closes.columns and TICKER != 'SMH':
        smh_diff = (cl.pct_change() - closes['SMH'].pct_change()).abs().rolling(20).mean()
        # 차이 변동성이 갑자기 커지면 SMH(=NVDA 영향)가 XLK 전체를 흔들고 있다는 신호
        feat['mag7_concentration_risk_20'] = (
            smh_diff / smh_diff.rolling(60).mean().replace(0, np.nan)
        )
    # 한 종목이 전체를 견인하는 정도 (XLK vs IGV, SOXX 합성)
    if 'IGV' in closes.columns and TICKER != 'IGV':
        # XLK와 IGV의 상관 변화 — 갑자기 분리되면 어느 한 쪽이 망가지는 중
        corr_xlk_igv = cl.pct_change().rolling(10).corr(closes['IGV'].pct_change())
        feat['tech_subsector_decorrelation'] = (
            corr_xlk_igv.rolling(60).mean() - corr_xlk_igv
        )

    # 11e. 갭 폭발 위험 (Gap Risk Score)
    # 이론: 외부 뉴스 충격은 야간/주말에 발생 → 갭 빈도/크기 누적이 위험 신호
    open_gap = (op - cl.shift()) / cl.shift()
    # 최근 20일 평균 갭 크기 (절댓값)
    feat['avg_gap_size_20'] = open_gap.abs().rolling(20).mean()
    # 음의 갭 누적 (최근 5일 음의 갭 합)
    neg_gap = open_gap.where(open_gap < 0).fillna(0)
    feat['cumulative_negative_gap_5d']  = neg_gap.rolling(5).sum()
    feat['cumulative_negative_gap_20d'] = neg_gap.rolling(20).sum()
    # 큰 음의 갭 빈도 — 한 달에 -1% 이상 하락 갭이 몇 번?
    big_neg_gap = (open_gap < -0.01).astype(float)
    feat['big_negative_gap_count_20'] = big_neg_gap.rolling(20).sum()

    # 11f. 주말/연휴 직전 위험 — 외부 충격은 주말·연휴에 빈발
    # 금요일 종가 기준 신호 (주말 갭 다운 위험)
    feat['is_friday'] = (dow == 4).astype(float)
    # 금요일 종가에서 RSI가 70+면 "지친 상승"이라 월요일 갭다운 위험
    rsi14 = calc_rsi(cl, 14)
    feat['friday_overbought'] = ((dow == 4) & (rsi14 > 70)).astype(float)
    # 화요일 = 주요 경제지표 발표일 (CPI, 고용 등)
    feat['is_tuesday_or_wednesday'] = ((dow == 1) | (dow == 2)).astype(float)

    # 11g. 빠른 모멘텀 손실 신호 (Velocity Loss)
    # 이론: 큰 하락 직전엔 단기 모멘텀이 갑자기 죽기 시작 (decelerate)
    mom_5d  = cl.pct_change(5)
    mom_20d = cl.pct_change(20)
    # 5일 모멘텀이 양수에서 음수로 돌아서면 (5일 전엔 상승 → 지금 하락)
    feat['momentum_inversion_5d'] = (
        (mom_5d.shift(5) > 0) & (mom_5d < 0)
    ).astype(float)
    # 20일 모멘텀 대비 5일 모멘텀이 급격히 약해진 정도
    feat['momentum_velocity_loss'] = mom_20d - mom_5d * 4   # 정규화

    # ══════════════════════════════════════════════════════════
    #  12. 신규 ETF/주식 상대강도 지표 (~55개)
    #      — PCAR, LMT, NOBL, KRE, IYT, XHB, TIP, IEF 등
    # ══════════════════════════════════════════════════════════

    # ── 12A. PCAR (경기민감 트럭 제조) ────────────────────────
    pcar = closes.get('PCAR')
    if pcar is not None and TICKER != 'PCAR':
        for p in [5, 10, 20, 60]:
            feat[f'pcar_rel_{p}d']    = cl.pct_change(p) - pcar.pct_change(p)
        feat['pcar_corr_20d']          = cl.pct_change().rolling(20).corr(pcar.pct_change())
        feat['pcar_corr_60d']          = cl.pct_change().rolling(60).corr(pcar.pct_change())
        feat['pcar_zscore_20']         = calc_zscore(pcar, 20)
        feat['pcar_below_sma50']       = (pcar < pcar.rolling(50).mean()).astype(float)
        # PCAR 하락 선행 (경기 둔화 신호)
        feat['pcar_lead_10d']          = pcar.pct_change(10).shift(5)   # 5일 선행

    # ── 12B. LMT (방산 — 위험회피 순환 신호) ───────────────────
    lmt = closes.get('LMT')
    if lmt is not None and TICKER != 'LMT':
        for p in [5, 20, 60]:
            feat[f'lmt_rel_{p}d']     = lmt.pct_change(p) - cl.pct_change(p)
        feat['lmt_corr_20d']           = cl.pct_change().rolling(20).corr(lmt.pct_change())
        feat['lmt_outperform_20d']     = (lmt.pct_change(20) > cl.pct_change(20)).astype(float)
        feat['lmt_outperform_60d']     = (lmt.pct_change(60) > cl.pct_change(60)).astype(float)
        # 방산 강세 = 지정학/위험회피 신호 → XLK 약세 선행
        feat['defense_rotation_score'] = (
            (lmt.pct_change(20) > 0).astype(float) *
            (cl.pct_change(20) < lmt.pct_change(20)).astype(float)
        )

    # ── 12C. NOBL (배당귀족 — 퀄리티 로테이션) ─────────────────
    nobl = closes.get('NOBL')
    spy  = closes.get('SPY')
    qqq  = closes.get('QQQ')
    if nobl is not None:
        for p in [5, 20, 60]:
            feat[f'nobl_vs_spy_{p}d'] = nobl.pct_change(p) - (spy.pct_change(p) if spy is not None else 0)
            feat[f'nobl_vs_xlk_{p}d'] = nobl.pct_change(p) - cl.pct_change(p)
        feat['nobl_outperform_flag_20'] = (nobl.pct_change(20) > cl.pct_change(20)).astype(float)
        feat['nobl_corr_20d']           = cl.pct_change().rolling(20).corr(nobl.pct_change())
        nobl_ratio = nobl / cl
        feat['nobl_xlk_ratio_zscore_60'] = calc_zscore(nobl_ratio, 60)

    # ── 12D. VIG / QUAL (퀄리티 팩터) ──────────────────────────
    vig  = closes.get('VIG')
    qual = closes.get('QUAL')
    if vig is not None and TICKER not in ('VIG',):
        feat['vig_rel_xlk_20d']  = vig.pct_change(20) - cl.pct_change(20)
        feat['vig_rel_xlk_60d']  = vig.pct_change(60) - cl.pct_change(60)
        feat['vig_outperform_20'] = (vig.pct_change(20) > cl.pct_change(20)).astype(float)
    if qual is not None:
        feat['qual_rel_xlk_20d'] = qual.pct_change(20) - cl.pct_change(20)
        feat['qual_momentum_20'] = calc_zscore(qual.pct_change(), 20)

    # ── 12E. KRE (지역은행 — 신용 경색 선행) ───────────────────
    kre = closes.get('KRE')
    if kre is not None:
        for p in [5, 10, 20, 60]:
            feat[f'kre_ret_{p}d']      = kre.pct_change(p)
        feat['kre_zscore_20']           = calc_zscore(kre.pct_change(), 20)
        feat['kre_below_sma50']         = (kre < kre.rolling(50).mean()).astype(float)
        feat['kre_drawdown_20d']        = kre / kre.rolling(20).max() - 1
        # 은행주 약세 = 신용 위험 선행 (XLK 하락과 동행)
        feat['kre_xlk_corr_20d']        = cl.pct_change().rolling(20).corr(kre.pct_change())
        if spy is not None:
            feat['kre_rel_spy_20d']     = kre.pct_change(20) - spy.pct_change(20)

    # ── 12F. IYT (운송 — 경제활동 온도계) ──────────────────────
    iyt = closes.get('IYT')
    if iyt is not None:
        for p in [5, 20, 60]:
            feat[f'iyt_ret_{p}d']       = iyt.pct_change(p)
        feat['iyt_zscore_20']            = calc_zscore(iyt.pct_change(), 20)
        feat['iyt_below_sma50']          = (iyt < iyt.rolling(50).mean()).astype(float)
        feat['iyt_rel_xlk_20d']          = iyt.pct_change(20) - cl.pct_change(20)
        # Dow Theory: 운송 + 산업주 동시 확인 (불일치 = 추세 약화)
        xli = closes.get('XLI')
        if xli is not None:
            feat['dow_theory_divergence_20'] = (
                (cl.pct_change(20) > 0).astype(float) *          # XLK 상승
                (iyt.pct_change(20) < 0).astype(float) *         # 운송 하락
                (xli.pct_change(20) < 0).astype(float)           # 산업 하락
            )

    # ── 12G. XHB (주택건설 — 금리 민감도 측정) ─────────────────
    xhb = closes.get('XHB')
    if xhb is not None:
        for p in [10, 20, 60]:
            feat[f'xhb_ret_{p}d']       = xhb.pct_change(p)
        feat['xhb_zscore_20']            = calc_zscore(xhb.pct_change(), 20)
        feat['xhb_drawdown_60d']         = xhb / xhb.rolling(60).max() - 1
        feat['xhb_rel_xlk_20d']          = xhb.pct_change(20) - cl.pct_change(20)

    # ── 12H. XRT (소매 — 소비자 심리) ─────────────────────────
    xrt = closes.get('XRT')
    if xrt is not None:
        for p in [5, 20, 60]:
            feat[f'xrt_ret_{p}d']       = xrt.pct_change(p)
        feat['xrt_zscore_20']            = calc_zscore(xrt.pct_change(), 20)
        feat['xrt_rel_xlk_20d']          = xrt.pct_change(20) - cl.pct_change(20)

    # ── 12I. TIP / IEF / BIL (채권 구조) ──────────────────────
    tip = closes.get('TIP'); ief = closes.get('IEF'); bil = closes.get('BIL')
    shy = closes.get('SHY')
    if tip is not None:
        for p in [5, 10, 20]:
            feat[f'tip_ret_{p}d']       = tip.pct_change(p)
        feat['tip_zscore_20']            = calc_zscore(tip.pct_change(), 20)
        feat['tip_xl_relative_20d']      = tip.pct_change(20) - cl.pct_change(20)
        # TIP 강세 + XLK 약세 = 인플레 우려 속 테크 회피
        feat['inflation_fear_regime']    = (
            (tip.pct_change(20) > 0).astype(float) *
            (cl.pct_change(20) < 0).astype(float)
        )
    if ief is not None and bil is not None:
        # IEF vs BIL 스프레드 (중기 금리 리스크)
        feat['ief_bil_spread_20d']       = ief.pct_change(20) - bil.pct_change(20)
        feat['duration_risk_signal']     = (ief.pct_change(5) < -0.005).astype(float)
    if ief is not None and tip is not None:
        # 실질금리 프록시 (IEF - TIP 수익률 차)
        feat['real_rate_proxy_20d']      = ief.pct_change(20) - tip.pct_change(20)
    if bil is not None:
        feat['bil_5d_ret']               = bil.pct_change(5)
        # BIL 급등 = 단기 자금 도피
        feat['bil_safe_haven_signal']    = (bil.pct_change(5) > bil.pct_change(5).rolling(60).mean() +
                                            bil.pct_change(5).rolling(60).std()).astype(float)

    # ── 12J. CPER / XME (구리·금속 경기선행) ────────────────────
    cper = closes.get('CPER'); xme = closes.get('XME')
    gld  = closes.get('GLD')
    if cper is not None:
        for p in [5, 20, 60]:
            feat[f'cper_ret_{p}d']       = cper.pct_change(p)
        feat['cper_zscore_20']            = calc_zscore(cper.pct_change(), 20)
        # 구리/금 비율 (성장 선행 — 기존 DBC/GLD와 다름: 구리 단독 사용)
        if gld is not None:
            cu_au = cper / gld
            feat['copper_gold_ratio_pure_20'] = cu_au.pct_change(20)
            feat['copper_gold_zscore_60']     = calc_zscore(cu_au, 60)
    if xme is not None and TICKER != 'XME':
        feat['xme_ret_20d']              = xme.pct_change(20)
        feat['xme_rel_xlk_20d']          = xme.pct_change(20) - cl.pct_change(20)

    # ── 12K. VXX / UVXY (변동성 선물 구조) ─────────────────────
    vxx  = closes.get('VXX'); uvxy = closes.get('UVXY')
    svxy = closes.get('SVXY')
    if vxx is not None:
        feat['vxx_5d_ret']               = vxx.pct_change(5)
        feat['vxx_20d_ret']              = vxx.pct_change(20)
        feat['vxx_zscore_20']            = calc_zscore(vxx.pct_change(), 20)
        feat['vxx_above_sma20']          = (vxx > vxx.rolling(20).mean()).astype(float)
        # VXX / VIX 비율 → 콘탱고/백워데이션 프록시 (일별 데이터 기반)
        vix_c = closes.get('^VIX')
        if vix_c is not None:
            ratio_v = vxx / vix_c.replace(0, np.nan)
            feat['vxx_vix_ratio_zscore_20'] = calc_zscore(ratio_v, 20)
    if uvxy is not None and svxy is not None:
        # UVXY vs SVXY 강도차 (공포 vs 안도의 실시간 스코어)
        feat['uvxy_svxy_spread_5d']      = uvxy.pct_change(5) - svxy.pct_change(5)
    # JNK vs HYG 비교 (HYG 기존 있으므로 JNK 추가)
    jnk = closes.get('JNK'); hyg = closes.get('HYG')
    if jnk is not None and hyg is not None:
        feat['jnk_hyg_spread_20d']       = jnk.pct_change(20) - hyg.pct_change(20)
        feat['jnk_zscore_20']            = calc_zscore(jnk.pct_change(), 20)
    # 5Y / 30Y 금리
    fvx = closes.get('^FVX'); tyx = closes.get('^TYX')
    tnx_c = closes.get('^TNX')
    if fvx is not None and tyx is not None:
        feat['term_5y30y_spread']        = tyx - fvx
        for p in [5, 20]:
            feat[f'term_5y30y_chg_{p}d'] = (tyx - fvx).diff(p)
    if fvx is not None and tnx_c is not None:
        feat['term_5y10y_spread']        = tnx_c - fvx
    # RINF (인플레 기대)
    rinf = closes.get('RINF')
    if rinf is not None:
        feat['rinf_5d_ret']              = rinf.pct_change(5)
        feat['rinf_20d_ret']             = rinf.pct_change(20)
        feat['rinf_zscore_20']           = calc_zscore(rinf.pct_change(), 20)

    # ── 12L. 종합 매크로 스트레스 지수 (새 ETF 조합) ────────────
    # 높을수록 매크로 위험 환경
    macro_stress = pd.Series(0.0, index=cl.index)
    if kre  is not None: macro_stress += (kre.pct_change(20) < -0.05).astype(float)
    if iyt  is not None: macro_stress += (iyt.pct_change(20) < -0.03).astype(float)
    if jnk  is not None: macro_stress += (jnk.pct_change(5)  < -0.01).astype(float)
    if vxx  is not None: macro_stress += (vxx.pct_change(5)  > 0.10).astype(float)
    if nobl is not None: macro_stress += (nobl.pct_change(20) > cl.pct_change(20) + 0.02).astype(float)
    feat['macro_etf_stress_composite']   = macro_stress

    # ══════════════════════════════════════════════════════════
    #  13. FRED 실물 경제지표 (~55개)
    #      fred_df: download_fred_data() 반환값 (일별 리샘플링 완료)
    # ══════════════════════════════════════════════════════════
    if fred_df is not None and len(fred_df) > 0:
        # 공통 인덱스 맞추기
        fred_al = fred_df.reindex(feat.index).ffill().bfill()

        def _fred(col):
            """fred_al에 컬럼 존재 여부 안전 조회."""
            return fred_al[col] if col in fred_al.columns else None

        # ── 13A. 수익률 곡선 (FRED 공식 데이터) ─────────────────
        t10y2y = _fred('T10Y2Y'); t10y3m = _fred('T10Y3M')
        if t10y2y is not None:
            feat['fred_t10y2y_level']      = t10y2y
            feat['fred_t10y2y_chg5d']      = t10y2y.diff(5)
            feat['fred_t10y2y_chg20d']     = t10y2y.diff(20)
            feat['fred_t10y2y_inverted']   = (t10y2y < 0).astype(float)
            feat['fred_t10y2y_zscore_60']  = calc_zscore(t10y2y, 60)
            # 역전에서 복원 (가장 위험한 구간 = 역전 해소 직후)
            feat['fred_t10y2y_uninvert']   = (
                (t10y2y.shift(5) < 0) & (t10y2y >= 0)
            ).astype(float)
        if t10y3m is not None:
            feat['fred_t10y3m_level']      = t10y3m
            feat['fred_t10y3m_inverted']   = (t10y3m < 0).astype(float)
            feat['fred_t10y3m_chg20d']     = t10y3m.diff(20)
            feat['fred_recession_signal']  = (
                (t10y3m < 0) & (t10y2y < 0 if t10y2y is not None else True)
            ).astype(float)
        if t10y2y is not None and t10y3m is not None:
            feat['fred_dual_invert_score'] = (
                (t10y2y < 0).astype(float) + (t10y3m < 0).astype(float)
            )

        # ── 13B. 신용 스프레드 (실제 OAS) ──────────────────────
        hy_oas = _fred('BAMLH0A0HYM2'); ig_oas = _fred('BAMLC0A0CM')
        if hy_oas is not None:
            feat['fred_hy_oas']            = hy_oas
            feat['fred_hy_oas_chg5d']      = hy_oas.diff(5)
            feat['fred_hy_oas_chg20d']     = hy_oas.diff(20)
            feat['fred_hy_oas_zscore_60']  = calc_zscore(hy_oas, 60)
            feat['fred_hy_oas_pctrank_252']= calc_pctrank(hy_oas, 252)
            feat['fred_hy_oas_spike']      = (hy_oas > hy_oas.rolling(60).mean() +
                                               hy_oas.rolling(60).std() * 1.5).astype(float)
        if ig_oas is not None:
            feat['fred_ig_oas']            = ig_oas
            feat['fred_ig_oas_chg5d']      = ig_oas.diff(5)
            feat['fred_ig_oas_zscore_60']  = calc_zscore(ig_oas, 60)
        if hy_oas is not None and ig_oas is not None:
            feat['fred_hy_ig_spread']      = hy_oas - ig_oas
            feat['fred_hy_ig_spread_chg']  = (hy_oas - ig_oas).diff(5)
            feat['fred_hy_ig_spread_zscore'] = calc_zscore(hy_oas - ig_oas, 60)

        # ── 13C. 기대인플레이션 ────────────────────────────────
        bei5 = _fred('T5YIE'); bei10 = _fred('T10YIE')
        if bei5 is not None:
            feat['fred_bei5_level']        = bei5
            feat['fred_bei5_chg20d']       = bei5.diff(20)
            feat['fred_bei5_zscore_60']    = calc_zscore(bei5, 60)
            feat['fred_bei5_above_3pct']   = (bei5 > 3.0).astype(float)
        if bei10 is not None:
            feat['fred_bei10_level']       = bei10
            feat['fred_bei10_chg20d']      = bei10.diff(20)
        if bei5 is not None and bei10 is not None:
            feat['fred_bei_slope_5_10']    = bei10 - bei5
            feat['fred_bei_slope_chg20']   = (bei10 - bei5).diff(20)

        # ── 13D. 실업률 ─────────────────────────────────────────
        unrate = _fred('UNRATE'); u6 = _fred('U6RATE')
        if unrate is not None:
            feat['fred_unrate']            = unrate
            feat['fred_unrate_3m_chg']     = unrate.diff(63)   # ~3개월
            feat['fred_unrate_6m_chg']     = unrate.diff(126)
            feat['fred_unrate_rising']     = (unrate.diff(63) > 0.3).astype(float)
            feat['fred_unrate_zscore_252'] = calc_zscore(unrate, 252)
            # Sahm Rule 프록시: 최근 3개월 평균이 12개월 최저보다 0.5% 이상↑
            feat['fred_sahm_proxy']        = (
                unrate.rolling(63).mean() - unrate.rolling(252).min()
            )
            feat['fred_sahm_trigger']      = (feat['fred_sahm_proxy'] >= 0.5).astype(float)
        if u6 is not None:
            feat['fred_u6_level']          = u6
            feat['fred_u6_3m_chg']         = u6.diff(63)
            feat['fred_u6_unrate_gap']     = (u6 - unrate) if unrate is not None else u6

        # ── 13E. CPI / 인플레이션 ──────────────────────────────
        cpi = _fred('CPIAUCSL'); core_cpi = _fred('CPILFESL')
        pce = _fred('PCEPI');    core_pce = _fred('PCEPILFE')
        if cpi is not None:
            feat['fred_cpi_yoy']           = cpi.pct_change(252) * 100
            feat['fred_cpi_mom']           = cpi.pct_change(21)  * 100
            feat['fred_cpi_acceleration']  = cpi.pct_change(252).diff(63)
            feat['fred_cpi_above_5pct']    = (cpi.pct_change(252)*100 > 5).astype(float)
            feat['fred_cpi_falling']       = (cpi.pct_change(252).diff(63) < 0).astype(float)
        if core_cpi is not None:
            feat['fred_core_cpi_yoy']      = core_cpi.pct_change(252) * 100
            feat['fred_core_cpi_accel']    = core_cpi.pct_change(252).diff(63)
        if cpi is not None and core_cpi is not None:
            feat['fred_energy_inflation']  = (cpi.pct_change(252) - core_cpi.pct_change(252)) * 100
        if core_pce is not None:
            feat['fred_core_pce_yoy']      = core_pce.pct_change(252) * 100
            feat['fred_core_pce_above_2']  = (core_pce.pct_change(252)*100 > 2).astype(float)
        # 스태그플레이션 지수 (인플레 높고 성장 둔화)
        if core_cpi is not None and unrate is not None:
            feat['fred_stagflation_index'] = (
                calc_zscore(core_cpi.pct_change(252), 252) +
                calc_zscore(unrate.diff(63), 252)
            )

        # ── 13F. PMI (ISM 제조업) ──────────────────────────────
        pmi = _fred('NAPM'); pmi_orders = _fred('NAPMNOI')
        pmi_emp = _fred('NAPMEI'); pmi_price = _fred('NAPMPI')
        if pmi is not None:
            feat['fred_pmi_level']         = pmi
            feat['fred_pmi_chg3m']         = pmi.diff(63)
            feat['fred_pmi_below_50']      = (pmi < 50).astype(float)
            feat['fred_pmi_below_45']      = (pmi < 45).astype(float)
            feat['fred_pmi_zscore_252']    = calc_zscore(pmi, 252)
            feat['fred_pmi_falling_3m']    = (pmi.diff(63) < -2).astype(float)
        if pmi_orders is not None:
            feat['fred_pmi_orders']        = pmi_orders
            feat['fred_pmi_orders_below50'] = (pmi_orders < 50).astype(float)
            feat['fred_pmi_orders_chg3m']  = pmi_orders.diff(63)
        if pmi is not None and pmi_orders is not None:
            # 신규주문 > PMI = 모멘텀 양호; 반대 = 둔화 신호
            feat['fred_pmi_orders_pmi_gap'] = pmi_orders - pmi
        if pmi_emp is not None:
            feat['fred_pmi_employment']    = pmi_emp
            feat['fred_pmi_emp_below50']   = (pmi_emp < 50).astype(float)
        if pmi_price is not None:
            feat['fred_pmi_price_above60'] = (pmi_price > 60).astype(float)
            feat['fred_pmi_price_zscore']  = calc_zscore(pmi_price, 252)

        # ── 13G. 소비자 신뢰 ───────────────────────────────────
        umcsi = _fred('UMCSENT'); umcsi_inf = _fred('MICH')
        if umcsi is not None:
            feat['fred_umcsi_level']       = umcsi
            feat['fred_umcsi_3m_chg']      = umcsi.diff(63)
            feat['fred_umcsi_zscore_252']  = calc_zscore(umcsi, 252)
            feat['fred_umcsi_below_60']    = (umcsi < 60).astype(float)
            feat['fred_umcsi_falling_3m']  = (umcsi.diff(63) < -5).astype(float)
            feat['fred_umcsi_pctrank_252'] = calc_pctrank(umcsi, 252)
        if umcsi_inf is not None:
            feat['fred_umcsi_infl_exp']    = umcsi_inf
            feat['fred_umcsi_infl_above3'] = (umcsi_inf > 3.0).astype(float)

        # ── 13H. 주간 실업수당 청구 ────────────────────────────
        icsa = _fred('ICSA'); ccsa = _fred('CCSA')
        if icsa is not None:
            feat['fred_icsa_level']        = icsa / 1e3   # 천 명 단위
            feat['fred_icsa_4w_ma']        = icsa.rolling(20).mean() / 1e3
            feat['fred_icsa_yoy']          = icsa.pct_change(252) * 100
            feat['fred_icsa_spike']        = (icsa > icsa.rolling(60).mean() * 1.2).astype(float)
            feat['fred_icsa_zscore_252']   = calc_zscore(icsa, 252)
            feat['fred_icsa_rising_4w']    = (icsa.diff(20) > 10000).astype(float)
        if ccsa is not None:
            feat['fred_ccsa_zscore_252']   = calc_zscore(ccsa, 252)
            feat['fred_ccsa_rising']       = (ccsa.diff(20) > 0).astype(float)

        # ── 13I. 주택·산업 실물 지표 ───────────────────────────
        houst  = _fred('HOUST');  indpro  = _fred('INDPRO')
        payems = _fred('PAYEMS'); tcu = _fred('TCU')
        retail = _fred('RSAFS')
        if houst is not None:
            feat['fred_houst_yoy']         = houst.pct_change(252) * 100
            feat['fred_houst_below_1m']    = (houst < 1000).astype(float)  # 1백만 호
        if indpro is not None:
            feat['fred_indpro_yoy']        = indpro.pct_change(252) * 100
            feat['fred_indpro_falling']    = (indpro.pct_change(63) < -0.01).astype(float)
            feat['fred_indpro_zscore_252'] = calc_zscore(indpro, 252)
        if tcu is not None:
            feat['fred_tcu_level']         = tcu
            feat['fred_tcu_below_78']      = (tcu < 78).astype(float)
            feat['fred_tcu_zscore_252']    = calc_zscore(tcu, 252)
        if payems is not None:
            feat['fred_payems_3m_avg']     = payems.rolling(63).mean()
            feat['fred_payems_falling']    = (payems.diff(21) < 0).astype(float)
        if retail is not None:
            feat['fred_retail_yoy']        = retail.pct_change(252) * 100
            feat['fred_retail_zscore']     = calc_zscore(retail, 252)

        # ── 13J. 통화·연준 지표 ────────────────────────────────
        m2   = _fred('M2SL'); walcl = _fred('WALCL')
        effr = _fred('EFFR'); psavr = _fred('PSAVERT')
        if m2 is not None:
            feat['fred_m2_yoy']            = m2.pct_change(252) * 100
            feat['fred_m2_growth_falling'] = (m2.pct_change(252).diff(63) < 0).astype(float)
            feat['fred_m2_zscore_252']     = calc_zscore(m2.pct_change(252), 252)
        if walcl is not None:
            feat['fred_fed_bs_yoy']        = walcl.pct_change(252) * 100
            feat['fred_fed_bs_shrinking']  = (walcl.diff(63) < 0).astype(float)
        if effr is not None and bei5 is not None:
            feat['fred_real_rate_effr_bei5'] = effr - bei5   # 실질 정책금리
            feat['fred_tight_real_rate']   = (effr - bei5 > 2.0).astype(float)
        if psavr is not None:
            feat['fred_saving_rate_low']   = (psavr < 4.0).astype(float)
            feat['fred_saving_rate_zscore'] = calc_zscore(psavr, 252)

        # ── 13K. 복합 경기선행지수 ─────────────────────────────
        # 여러 경제지표를 표준화 후 합산 → 경기 모멘텀 점수
        recession_score = pd.Series(0.0, index=feat.index)
        if t10y3m is not None:   recession_score += (t10y3m < 0).astype(float)
        if t10y2y is not None:   recession_score += (t10y2y < 0).astype(float)
        if hy_oas is not None:   recession_score += (hy_oas > hy_oas.rolling(252).mean() +
                                                      hy_oas.rolling(252).std()).astype(float)
        if unrate is not None:   recession_score += feat.get('fred_sahm_trigger', pd.Series(0, index=feat.index))
        if pmi is not None:      recession_score += (pmi < 50).astype(float)
        if umcsi is not None:    recession_score += (umcsi < umcsi.rolling(252).mean()).astype(float)
        if icsa is not None:     recession_score += feat.get('fred_icsa_spike', pd.Series(0, index=feat.index))
        feat['fred_composite_recession_score'] = recession_score

        # 인플레 스트레스 점수
        inflation_stress = pd.Series(0.0, index=feat.index)
        if cpi is not None:      inflation_stress += (cpi.pct_change(252)*100 > 4).astype(float)
        if core_cpi is not None: inflation_stress += (core_cpi.pct_change(252)*100 > 3).astype(float)
        if bei5 is not None:     inflation_stress += (bei5 > 2.8).astype(float)
        if pmi_price is not None: inflation_stress += (pmi_price > 65).astype(float)
        feat['fred_inflation_stress_score'] = inflation_stress

    # ══════════════════════════════════════════════════════════
    #  14. 수학적·통계적 심화 지표 (~75개)
    #      — 누적 상승률 가속도, 엔트로피, 자기상관,
    #        Amihud 비유동성, 다운사이드 리스크, OU 평균회귀 등
    # ══════════════════════════════════════════════════════════
    ret_d   = cl.pct_change()
    log_ret = np.log(cl / cl.shift(1))

    # ── 14A. 누적 수익률 기울기 & 기울기 변화량 ────────────────
    for n_cum in [10, 20, 60]:
        # 누적 수익률
        cum_ret_n = cl.pct_change(n_cum)
        # 기울기 (n_cum일 창에서 선형회귀 기울기 / 가격 수준)
        slope_n = calc_linreg_slope(cl, n_cum)
        feat[f'cum_ret_{n_cum}d']           = cum_ret_n          # (기존 ret_*와 다름: pct_change 직접)
        feat[f'slope_{n_cum}d']             = slope_n
        # 기울기 변화량 (가속도: 오늘 기울기 - n일 전 기울기)
        feat[f'slope_accel_{n_cum}d']       = slope_n - slope_n.shift(n_cum // 2)
        # 2차 가속도 (기울기 변화량의 변화량)
        feat[f'slope_jerk_{n_cum}d']        = (slope_n - slope_n.shift(n_cum // 2)) - \
                                               (slope_n.shift(n_cum // 2) - slope_n.shift(n_cum))
        # 누적 수익률 기울기 (log scale)
        log_cum = np.log(cl / cl.shift(n_cum))
        log_slope = calc_linreg_slope(log_cum, n_cum // 2)
        feat[f'log_slope_{n_cum}d']         = log_slope
        feat[f'log_slope_accel_{n_cum}d']   = log_slope - log_slope.shift(n_cum // 4)

    # 누적 상승률 체감 지수: 최근 5일 누적 / 최근 20일 누적
    for (sh, lg) in [(5, 20), (10, 60), (20, 60)]:
        feat[f'gain_decel_{sh}_{lg}d']      = cl.pct_change(sh) / (cl.pct_change(lg) + 1e-6) - 1

    # ── 14B. 수익률 자기상관 (ACF) ──────────────────────────────
    for p in [10, 20, 60]:
        # lag-1 자기상관 (양수 = 추세; 음수 = 평균회귀)
        feat[f'ret_acf1_{p}d']              = ret_d.rolling(p).apply(
            lambda x: pd.Series(x).autocorr(lag=1) if len(x) > 2 else np.nan, raw=False
        )
    # lag-2, lag-3 (단기 반전 체크)
    for lag in [2, 3]:
        feat[f'ret_acf{lag}_20d']           = ret_d.rolling(20).apply(
            lambda x: pd.Series(x).autocorr(lag=lag) if len(x) > lag else np.nan, raw=False
        )

    # ── 14C. 샤논 엔트로피 (수익률 분포 무작위성) ──────────────
    def _entropy(x, bins=10):
        x = x[~np.isnan(x)]
        if len(x) < 10: return np.nan
        counts, _ = np.histogram(x, bins=bins)
        p = counts / counts.sum()
        p = p[p > 0]
        return float(-np.sum(p * np.log(p)))

    for p in [20, 60]:
        feat[f'ret_entropy_{p}d']           = ret_d.rolling(p).apply(
            lambda x: _entropy(x), raw=True
        )
    # 엔트로피 변화율 (상승 = 불확실성 증가)
    feat['ret_entropy_accel_20_60']         = (
        feat['ret_entropy_20d'] - feat.get('ret_entropy_60d', feat['ret_entropy_20d'])
    ) if 'ret_entropy_20d' in feat.columns else pd.Series(np.nan, index=cl.index)

    # ── 14D. 다운사이드 리스크 (Sortino / Omega) ────────────────
    for p in [20, 60]:
        _neg = ret_d.where(ret_d < 0).fillna(0)
        _pos = ret_d.where(ret_d > 0).fillna(0)
        downside_dev = _neg.rolling(p).apply(lambda x: np.sqrt(np.mean(x**2)), raw=True)
        feat[f'sortino_like_{p}d']          = (
            ret_d.rolling(p).mean() / downside_dev.replace(0, np.nan)
        )
        # Omega 비율 (양수 수익 합 / 음수 수익 절대합)
        feat[f'omega_ratio_{p}d']           = (
            _pos.rolling(p).sum() / _neg.abs().rolling(p).sum().replace(0, np.nan)
        )
        # 최대 단일 음봉 크기
        feat[f'max_single_loss_{p}d']       = ret_d.where(ret_d < 0).rolling(p).min()

    # ── 14E. Amihud 비유동성 지수 ───────────────────────────────
    dollar_vol = cl * vo
    amihud = ret_d.abs() / dollar_vol.replace(0, np.nan) * 1e6
    for p in [5, 20, 60]:
        feat[f'amihud_illiquidity_{p}d']    = amihud.rolling(p).mean()
    feat['amihud_zscore_60d']               = calc_zscore(amihud, 60)
    feat['amihud_spike']                    = (
        amihud > amihud.rolling(60).mean() + amihud.rolling(60).std() * 2
    ).astype(float)

    # ── 14F. 점프(Jump) 탐지 ───────────────────────────────────
    # Bipower variation 기반 점프 분리
    bpv_5 = (ret_d.abs() * ret_d.abs().shift(1)).rolling(5).mean() * np.pi / 2
    rv_5  = (ret_d ** 2).rolling(5).mean()
    feat['jump_component_5d']               = (rv_5 - bpv_5).clip(lower=0) / rv_5.replace(0, np.nan)
    feat['jump_flag_3sigma']                = (
        ret_d.abs() > ret_d.rolling(60).std() * 3
    ).astype(float)
    feat['jump_count_20d']                  = feat['jump_flag_3sigma'].rolling(20).sum()

    # ── 14G. Ornstein-Uhlenbeck 평균회귀 속도 ─────────────────
    def _ou_speed(x):
        if len(x) < 10 or np.isnan(x).any(): return np.nan
        y = x[1:] - x[:-1]   # delta
        x_lag = x[:-1] - x.mean()
        if x_lag.std() < 1e-10: return np.nan
        beta = np.cov(y, x_lag)[0, 1] / np.var(x_lag)
        return float(-beta)   # 양수 = 평균회귀

    for p in [20, 60]:
        feat[f'ou_mean_reversion_{p}d']     = cl.pct_change().rolling(p).apply(
            _ou_speed, raw=True
        )

    # ── 14H. 볼린저 밴드·켈트너 폭 변화율 ─────────────────────
    for (p, k) in [(20, 2.0), (20, 1.5)]:
        ma_v = cl.rolling(p).mean(); sd_v = cl.rolling(p).std()
        bb_w = 2 * k * sd_v / ma_v.replace(0, np.nan)
        feat[f'bb_width_roc_{p}_{k}_10d']   = bb_w / bb_w.shift(10) - 1
        feat[f'bb_width_roc_{p}_{k}_20d']   = bb_w / bb_w.shift(20) - 1
        feat[f'bb_width_pctrank_{p}_{k}']   = calc_pctrank(bb_w, 60)

    for (p, m) in [(20, 2.0), (20, 1.5)]:
        atr_p, _ = calc_atr(hi, lo, cl, p)
        mid_k = cl.ewm(span=p, adjust=False).mean()
        kc_w  = 2 * m * atr_p / mid_k.replace(0, np.nan)
        feat[f'keltner_width_roc_{p}_{m}']  = kc_w / kc_w.shift(20) - 1

    # ── 14I. 이동 Sharpe 기울기 ────────────────────────────────
    for p in [20, 60]:
        sh_like = ret_d.rolling(p).mean() / ret_d.rolling(p).std().replace(0, np.nan)
        feat[f'sharpe_slope_{p}d']          = sh_like - sh_like.shift(p // 2)
        feat[f'sharpe_accel_{p}d']          = (sh_like - sh_like.shift(p // 2)) - \
                                               (sh_like.shift(p // 2) - sh_like.shift(p))

    # ── 14J. 방향 일관성 (Streak consistency) ──────────────────
    sign_r = np.sign(ret_d)
    for p in [10, 20, 60]:
        # 같은 방향으로 가는 비율 (1 = 완전 일방향)
        feat[f'dir_consistency_{p}d']       = sign_r.rolling(p).sum().abs() / p
        # 연속 방향 변화 빈도 (높으면 지그재그)
        feat[f'dir_change_freq_{p}d']       = (sign_r.diff() != 0).rolling(p).mean()
    # 연속 상승일 z-score (기존 up_streak과 별개: 지수이평 필터 적용)
    feat['ema_streak_signal_5']             = (
        cl.pct_change().ewm(span=5, adjust=False).mean()
    )
    feat['ema_streak_signal_20']            = (
        cl.pct_change().ewm(span=20, adjust=False).mean()
    )

    # ── 14K. VaR / CVaR (Expected Shortfall) ───────────────────
    for p in [20, 60]:
        for q in [5, 10]:
            feat[f'var_{q}pct_{p}d']        = ret_d.rolling(p).quantile(q / 100)
            # CVaR (조건부 VaR = 하위 q% 평균)
            feat[f'cvar_{q}pct_{p}d']       = ret_d.rolling(p).apply(
                lambda x: x[x <= np.percentile(x, q)].mean() if len(x) > q else np.nan,
                raw=True
            )

    # ── 14L. 켈리 기준 비율 ────────────────────────────────────
    for p in [20, 60]:
        mu   = ret_d.rolling(p).mean()
        var_ = ret_d.rolling(p).var().replace(0, np.nan)
        feat[f'kelly_criterion_{p}d']       = mu / var_
        feat[f'kelly_positive_{p}d']        = (feat[f'kelly_criterion_{p}d'] > 0).astype(float)

    # ── 14M. 볼린저 밴드 터치 빈도 ─────────────────────────────
    for (p, k) in [(20, 2.0),]:
        ma_v = cl.rolling(p).mean(); sd_v = cl.rolling(p).std()
        upper = ma_v + k * sd_v; lower = ma_v - k * sd_v
        touch_upper = (cl >= upper).astype(float)
        touch_lower = (cl <= lower).astype(float)
        for win in [10, 20]:
            feat[f'bb_upper_touch_{win}d']  = touch_upper.rolling(win).sum()
            feat[f'bb_lower_touch_{win}d']  = touch_lower.rolling(win).sum()
            feat[f'bb_upper_touch_rate_{win}d'] = touch_upper.rolling(win).mean()

    # ── 14N. 선형 추세 잔차 (Detrended price) ──────────────────
    for p in [20, 60]:
        t_idx = np.arange(p)
        def _ols_resid(x):
            if np.isnan(x).any(): return np.nan
            slope_v, intercept = np.polyfit(t_idx, x, 1)
            return float(x[-1] - (slope_v * (p - 1) + intercept))
        feat[f'linear_detrend_{p}d']        = cl.rolling(p).apply(_ols_resid, raw=True) / cl

    # ── 14O. 실현 상관구조 변화 ─────────────────────────────────
    # 가격의 스무스니스 (연속적일수록 추세; 들쭉날쭉할수록 노이즈)
    def _smoothness(x):
        if len(x) < 5 or np.isnan(x).any(): return np.nan
        d1 = np.diff(x)
        d2 = np.diff(d1)
        return float(np.std(d2) / (np.std(d1) + 1e-10))

    feat['price_smoothness_20d']            = cl.rolling(20).apply(_smoothness, raw=True)
    feat['price_smoothness_60d']            = cl.rolling(60).apply(_smoothness, raw=True)

    # ── 14P. 누적 이익/손실 비대칭 ────────────────────────────
    for p in [20, 60]:
        cum_gain  = ret_d.where(ret_d > 0, 0).rolling(p).sum()
        cum_loss  = ret_d.where(ret_d < 0, 0).abs().rolling(p).sum()
        feat[f'gain_loss_ratio_{p}d']       = cum_gain / cum_loss.replace(0, np.nan)
        feat[f'gain_loss_ratio_chg_{p}d']   = (
            feat[f'gain_loss_ratio_{p}d'] - feat[f'gain_loss_ratio_{p}d'].shift(p // 2)
        )
        # 상승 회수 vs 하락 회수 비율
        up_cnt   = (ret_d > 0).rolling(p).sum()
        dn_cnt   = (ret_d < 0).rolling(p).sum()
        feat[f'win_rate_{p}d']              = up_cnt / p
        feat[f'win_rate_chg_{p}d']          = (
            feat[f'win_rate_{p}d'] - feat[f'win_rate_{p}d'].shift(p // 2)
        )

    # ── 14Q. 평균 음봉 크기 가속도 ─────────────────────────────
    avg_neg_ret = ret_d.where(ret_d < 0).abs().rolling(20).mean()
    feat['avg_neg_ret_accel_20d']           = avg_neg_ret - avg_neg_ret.shift(10)
    feat['avg_neg_ret_zscore_60d']          = calc_zscore(avg_neg_ret, 60)
    # 음봉 중 -1% 초과 비율 (대형 음봉 빈도)
    feat['large_neg_ret_rate_20d']          = (ret_d < -0.01).rolling(20).mean()
    feat['large_neg_ret_rate_60d']          = (ret_d < -0.01).rolling(60).mean()

    # ── 14R. 고점 이후 시간·속도 복합 지표 ─────────────────────
    for p in [20, 60]:
        _peak    = cl.rolling(p).max()
        _dd_curr = cl / _peak - 1   # 현재 드로다운 깊이
        _days    = feat.get(f'days_since_{p}d_high',
                            cl.rolling(p).apply(lambda x: float(len(x)-1-np.argmax(x)), raw=True))
        # 드로다운 속도 (깊이 / 경과일)
        feat[f'dd_velocity_{p}d']           = _dd_curr / (_days + 1)
        # 드로다운 가속도 (속도 변화)
        feat[f'dd_acceleration_{p}d']       = feat[f'dd_velocity_{p}d'] - \
                                               feat[f'dd_velocity_{p}d'].shift(5)

    # ── 14S. 지수 평활 추세 이탈 (EMA residual) ─────────────────
    for p in [10, 20, 50]:
        ema_p = cl.ewm(span=p, adjust=False).mean()
        resid = (cl - ema_p) / ema_p
        feat[f'ema_residual_std_{p}d']      = resid / resid.rolling(60).std().replace(0, np.nan)
        feat[f'ema_residual_zscore_{p}d']   = calc_zscore(resid, 60)

    # ── 14T. 가격 모멘텀 복합 점수 (비선형) ────────────────────
    # 각 기간의 모멘텀 방향 합산 (부호 기반, 중복 없음)
    mom_score = pd.Series(0.0, index=cl.index)
    for p in [5, 10, 20, 60]:
        mom_score += np.sign(cl.pct_change(p))
    feat['multi_period_mom_score']          = mom_score   # -4 ~ +4
    feat['multi_period_mom_all_neg']        = (mom_score == -4).astype(float)
    feat['multi_period_mom_all_pos']        = (mom_score == 4).astype(float)
    # 모멘텀 점수 변화
    feat['multi_period_mom_score_chg5d']    = mom_score - mom_score.shift(5)
    feat['multi_period_mom_score_chg20d']   = mom_score - mom_score.shift(20)

    # ══════════════════════════════════════════════════════════
    #  15. 스마트머니 / 내부자 선행 신호 (~95개)
    #
    #  이론적 근거:
    #  · 관세/AI 충격 등 "서프라이즈" 발표 전에도
    #    기관·내부자는 이미 헤지·청산·로테이션을 한다.
    #  · Liberation Day(25.4.2) 직전: GLD 신고점, FXY 강세,
    #    KRE 약세, HYG 하락이 3~7일 선행
    #  · DeepSeek(25.1.27) 직전: NVDA 분배일 누적,
    #    SQQQ 거래량 급등, 콤플레이센시 극대화
    # ══════════════════════════════════════════════════════════

    # ── 15A. 인버스 ETF 기관 숏 포지션 추적 ────────────────────
    sqqq_c = closes.get('SQQQ')
    sh_c   = closes.get('SH')

    if sqqq_c is not None:
        sqqq_ohlcv = ohlcv.get('SQQQ')
        for p in [3, 5, 10]:
            feat[f'sqqq_ret_{p}d']         = sqqq_c.pct_change(p)
        feat['sqqq_above_sma10']            = (sqqq_c > sqqq_c.rolling(10).mean()).astype(float)
        feat['sqqq_above_sma20']            = (sqqq_c > sqqq_c.rolling(20).mean()).astype(float)
        feat['sqqq_zscore_20']              = calc_zscore(sqqq_c.pct_change(), 20)
        feat['sqqq_consecutive_up_3d']      = (
            (sqqq_c.pct_change() > 0) &
            (sqqq_c.pct_change().shift(1) > 0) &
            (sqqq_c.pct_change().shift(2) > 0)
        ).astype(float)

        if sqqq_ohlcv is not None:
            sqqq_vol = sqqq_ohlcv['Volume']
            feat['sqqq_vol_ratio_5d']       = sqqq_vol / sqqq_vol.rolling(20).mean().replace(0, np.nan)
            feat['sqqq_vol_zscore_20']      = calc_zscore(sqqq_vol, 20)
            feat['sqqq_vol_surge_flag']     = (sqqq_vol > sqqq_vol.rolling(20).mean() * 2.0).astype(float)
            # 핵심: SQQQ 거래량 급등 + 가격 상승 동시 = 기관 숏 누적
            feat['sqqq_institutional_short'] = (
                (sqqq_vol > sqqq_vol.rolling(20).mean() * 1.5) &
                (sqqq_c.pct_change() > 0)
            ).astype(float)
            feat['sqqq_inst_short_3d_sum']  = feat['sqqq_institutional_short'].rolling(3).sum()
            feat['sqqq_inst_short_5d_sum']  = feat['sqqq_institutional_short'].rolling(5).sum()
            # 거래량 가중 SQQQ 상승 강도
            feat['sqqq_vol_weighted_5d']    = (
                (sqqq_c.pct_change() * sqqq_vol).rolling(5).sum() /
                sqqq_vol.rolling(5).sum().replace(0, np.nan)
            )

        # SQQQ vs QQQ 이론치 괴리 (레버리지 이상 수요 탐지)
        qqq_c = closes.get('QQQ')
        if qqq_c is not None:
            theoretical_5d   = -3 * qqq_c.pct_change(5)
            feat['sqqq_vs_theoretical_5d'] = sqqq_c.pct_change(5) - theoretical_5d
            feat['sqqq_excess_demand']      = (feat['sqqq_vs_theoretical_5d'] > 0.03).astype(float)

    if sh_c is not None:
        sh_ohlcv = ohlcv.get('SH')
        for p in [3, 5]:
            feat[f'sh_ret_{p}d']            = sh_c.pct_change(p)
        feat['sh_above_sma10']              = (sh_c > sh_c.rolling(10).mean()).astype(float)
        if sh_ohlcv is not None:
            sh_vol = sh_ohlcv['Volume']
            feat['sh_vol_ratio_5d']         = sh_vol / sh_vol.rolling(20).mean().replace(0, np.nan)
            feat['sh_institutional_short']   = (
                (sh_vol > sh_vol.rolling(20).mean() * 1.5) & (sh_c.pct_change() > 0)
            ).astype(float)
            feat['sh_inst_short_3d']        = feat['sh_institutional_short'].rolling(3).sum()

    # ── 15B. 엔화(FXY) 캐리 언와인드 선행 신호 ─────────────────
    # 캐리 트레이드(엔화 빌려 고수익 자산 투자) 언와인드 =
    # 위험자산 청산의 가장 빠른 선행 신호
    fxy_c = closes.get('FXY')
    if fxy_c is not None:
        for p in [3, 5, 10, 20]:
            feat[f'fxy_ret_{p}d']           = fxy_c.pct_change(p)
        feat['fxy_above_sma20']             = (fxy_c > fxy_c.rolling(20).mean()).astype(float)
        feat['fxy_above_sma50']             = (fxy_c > fxy_c.rolling(50).mean()).astype(float)
        feat['fxy_zscore_20']               = calc_zscore(fxy_c.pct_change(), 20)
        feat['fxy_zscore_60']               = calc_zscore(fxy_c.pct_change(), 60)
        feat['fxy_pctrank_60']              = calc_pctrank(fxy_c, 60)

        # 엔화 3일 연속 강세 = 캐리 언와인드 초기 신호
        feat['fxy_3d_consecutive_up']       = (
            (fxy_c.pct_change() > 0) &
            (fxy_c.pct_change().shift(1) > 0) &
            (fxy_c.pct_change().shift(2) > 0)
        ).astype(float)

        # 엔화 급등 가속도 (이중 상승 = 기관 대규모 청산)
        feat['fxy_accel_3d']                = fxy_c.pct_change(3) - fxy_c.pct_change(6).shift(3)
        feat['fxy_acceleration_flag']       = (feat['fxy_accel_3d'] > 0.005).astype(float)

        # 엔화 강세 + XLK 상승 공존 = 가장 위험한 조합
        # (캐리 청산 시작됐는데 주가는 아직 모름)
        feat['yen_carry_unwind_silent_5d']  = (
            (fxy_c.pct_change(5) > 0.008) & (cl.pct_change(5) > 0)
        ).astype(float)

        # FXY + GLD 동시 강세 (안전자산 쌍두마차)
        gld_c2 = closes.get('GLD')
        if gld_c2 is not None:
            feat['fxy_gld_both_up_5d']      = (
                (fxy_c.pct_change(5) > 0.005) & (gld_c2.pct_change(5) > 0.01)
            ).astype(float)
            feat['fxy_gld_both_up_3d']      = (
                (fxy_c.pct_change(3) > 0.003) & (gld_c2.pct_change(3) > 0.005)
            ).astype(float)

        # FXY + TLT 동시 강세 (채권+엔화 동반 = 극도 위험회피)
        tlt_c2 = closes.get('TLT')
        if tlt_c2 is not None:
            feat['fxy_tlt_both_up_5d']      = (
                (fxy_c.pct_change(5) > 0.005) & (tlt_c2.pct_change(5) > 0.005)
            ).astype(float)

    # ── 15C. 골드 마이너(GDX) 선행 신호 ───────────────────────
    # GDX는 금(GLD)보다 베타 2~3배 → 기관 공포 포지션 더 일찍 반영
    gdx_c = closes.get('GDX')
    gld_c = closes.get('GLD')

    if gdx_c is not None:
        for p in [3, 5, 10, 20]:
            feat[f'gdx_ret_{p}d']           = gdx_c.pct_change(p)
        feat['gdx_zscore_20']               = calc_zscore(gdx_c.pct_change(), 20)
        feat['gdx_above_sma20']             = (gdx_c > gdx_c.rolling(20).mean()).astype(float)
        feat['gdx_above_sma50']             = (gdx_c > gdx_c.rolling(50).mean()).astype(float)
        feat['gdx_drawdown_20d']            = gdx_c / gdx_c.rolling(20).max() - 1
        feat['gdx_pctrank_60']              = calc_pctrank(gdx_c, 60)

        if gld_c is not None:
            gdx_gld_ratio = gdx_c / gld_c.replace(0, np.nan)
            feat['gdx_gld_ratio_zscore_20'] = calc_zscore(gdx_gld_ratio, 20)
            feat['gdx_gld_ratio_zscore_60'] = calc_zscore(gdx_gld_ratio, 60)
            # GDX가 GLD보다 먼저 강해짐 = 기관이 레버리지 금 매수 (강한 위험회피)
            feat['gdx_leads_gld_5d']        = gdx_c.pct_change(5) - gld_c.pct_change(5)
            feat['gdx_premium_flag']        = (
                feat['gdx_gld_ratio_zscore_60'] > 1.5
            ).astype(float)

        # GDX 강세 + XLK 약세 동반 (가장 강한 역방향 신호)
        feat['gdx_xlk_divergence_5d']       = (
            (gdx_c.pct_change(5) > 0.03) & (cl.pct_change(5) < 0)
        ).astype(float)
        feat['gdx_xlk_divergence_10d']      = (
            (gdx_c.pct_change(10) > 0.05) & (cl.pct_change(10) < 0.01)
        ).astype(float)

    # ── 15D. 이머징마켓(EEM) 글로벌 위험선호 선행 ──────────────
    eem_c = closes.get('EEM')
    spy_c = closes.get('SPY')

    if eem_c is not None:
        for p in [5, 10, 20, 60]:
            feat[f'eem_ret_{p}d']           = eem_c.pct_change(p)
        feat['eem_zscore_20']               = calc_zscore(eem_c.pct_change(), 20)
        feat['eem_below_sma50']             = (eem_c < eem_c.rolling(50).mean()).astype(float)
        feat['eem_drawdown_20d']            = eem_c / eem_c.rolling(20).max() - 1
        feat['eem_drawdown_60d']            = eem_c / eem_c.rolling(60).max() - 1
        feat['eem_pctrank_60']              = calc_pctrank(eem_c, 60)

        if spy_c is not None:
            for p in [5, 20]:
                feat[f'eem_vs_spy_{p}d']    = eem_c.pct_change(p) - spy_c.pct_change(p)
            # EM 심하게 뒤처짐 + 미국 버팀 = 글로벌 위기 미국 전이 예정
            feat['eem_lagging_spy_flag_20d'] = (
                eem_c.pct_change(20) < spy_c.pct_change(20) - 0.05
            ).astype(float)

        # EEM 약세 + XLK 강세 = 전이 위험 가장 높은 패턴
        feat['eem_xlk_divergence_20d']      = (
            (eem_c.pct_change(20) < -0.03) & (cl.pct_change(20) > 0)
        ).astype(float)
        feat['eem_xlk_divergence_5d']       = (
            (eem_c.pct_change(5) < -0.01) & (cl.pct_change(5) > 0.01)
        ).astype(float)

    # ── 15E. 이퀄웨이트(RSP) vs 시가총액(SPY) 집중도 위험 ───────
    # RSP < SPY = 소수 대형주가 지수 받치는 중 (취약한 상승)
    rsp_c = closes.get('RSP')
    if rsp_c is not None and spy_c is not None:
        for p in [5, 10, 20, 60]:
            feat[f'rsp_vs_spy_{p}d']        = rsp_c.pct_change(p) - spy_c.pct_change(p)
        # SPY 오르는데 RSP는 못 따라옴 = 시가총액 집중 위험
        feat['cap_concentration_flag_20d']  = (
            (spy_c.pct_change(20) > 0.02) &
            (rsp_c.pct_change(20) < spy_c.pct_change(20) - 0.03)
        ).astype(float)
        feat['rsp_spy_gap_zscore_60']       = calc_zscore(
            rsp_c.pct_change(20) - spy_c.pct_change(20), 60
        )
        # RSP/SPY 비율 기울기 (하락 = 점점 소수 종목 의존)
        rsp_spy_ratio = rsp_c / spy_c
        feat['rsp_spy_ratio_slope_10d']     = calc_linreg_slope(rsp_spy_ratio, 10)
        feat['rsp_spy_ratio_below_sma50']   = (
            rsp_spy_ratio < rsp_spy_ratio.rolling(50).mean()
        ).astype(float)
        feat['rsp_spy_ratio_pctrank_60']    = calc_pctrank(rsp_spy_ratio, 60)

    # ── 15F. NVDA / MSFT / AAPL 내부 약화 선행 ─────────────────
    nvda_c = closes.get('NVDA')
    msft_c = closes.get('MSFT')
    aapl_c = closes.get('AAPL')
    avgo_c = closes.get('AVGO')

    if nvda_c is not None:
        for p in [3, 5, 10, 20]:
            feat[f'nvda_ret_{p}d']          = nvda_c.pct_change(p)
        feat['nvda_rsi_14']                 = calc_rsi(nvda_c, 14)
        feat['nvda_rsi_9']                  = calc_rsi(nvda_c, 9)
        feat['nvda_zscore_20']              = calc_zscore(nvda_c.pct_change(), 20)
        feat['nvda_drawdown_20d']           = nvda_c / nvda_c.rolling(20).max() - 1
        feat['nvda_drawdown_60d']           = nvda_c / nvda_c.rolling(60).max() - 1
        feat['nvda_above_sma50']            = (nvda_c > nvda_c.rolling(50).mean()).astype(float)
        for p in [5, 10, 20]:
            feat[f'nvda_vs_xlk_{p}d']       = nvda_c.pct_change(p) - cl.pct_change(p)
        # NVDA 내부 약화: XLK는 오르는데 NVDA는 하락
        feat['nvda_xlk_weakness_5d']        = (
            (cl.pct_change(5) > 0) & (nvda_c.pct_change(5) < -0.01)
        ).astype(float)
        # NVDA RSI 과매수 (70+) 상태에서 가격 정체
        feat['nvda_overbought_stall']       = (
            (feat['nvda_rsi_14'] > 70) &
            (nvda_c.pct_change(3).abs() < 0.02)
        ).astype(float)

        nvda_ohlcv_d = ohlcv.get('NVDA')
        if nvda_ohlcv_d is not None:
            nvda_vol = nvda_ohlcv_d['Volume']
            feat['nvda_vol_ratio_20d']      = nvda_vol / nvda_vol.rolling(20).mean().replace(0, np.nan)
            # 분배일: 높은 거래량 + 음봉
            nvda_dist = (
                (nvda_c.pct_change() < -0.01) &
                (nvda_vol > nvda_vol.rolling(20).mean() * 1.3)
            ).astype(float)
            feat['nvda_distribution_day']   = nvda_dist
            feat['nvda_dist_days_10d']      = nvda_dist.rolling(10).sum()
            feat['nvda_dist_days_20d']      = nvda_dist.rolling(20).sum()
            # 약한 랠리: 소량 거래량 + 가격 상승 (기관 비참여)
            feat['nvda_low_vol_rally_5d']   = (
                (nvda_c.pct_change(5) > 0.02) &
                (nvda_vol.rolling(5).mean() < nvda_vol.rolling(20).mean() * 0.75)
            ).astype(float)
            # OBV 다이버전스: 가격 상승 + OBV 하락 (스마트머니 팔고 있음)
            nvda_obv = (np.sign(nvda_c.diff()) * nvda_vol).fillna(0).cumsum()
            feat['nvda_obv_diverge_10d']    = (
                (nvda_c.pct_change(10) > 0.03) &
                (nvda_obv.diff(10) < 0)
            ).astype(float)

    if msft_c is not None:
        for p in [5, 10, 20]:
            feat[f'msft_ret_{p}d']          = msft_c.pct_change(p)
        feat['msft_rsi_14']                 = calc_rsi(msft_c, 14)
        feat['msft_drawdown_20d']           = msft_c / msft_c.rolling(20).max() - 1
        for p in [5, 20]:
            feat[f'msft_vs_xlk_{p}d']       = msft_c.pct_change(p) - cl.pct_change(p)
        feat['msft_below_sma50']            = (msft_c < msft_c.rolling(50).mean()).astype(float)
        msft_ohlcv_d = ohlcv.get('MSFT')
        if msft_ohlcv_d is not None:
            msft_vol = msft_ohlcv_d['Volume']
            msft_dist = (
                (msft_c.pct_change() < -0.01) &
                (msft_vol > msft_vol.rolling(20).mean() * 1.3)
            ).astype(float)
            feat['msft_dist_days_10d']      = msft_dist.rolling(10).sum()

    if aapl_c is not None:
        for p in [5, 10, 20]:
            feat[f'aapl_ret_{p}d']          = aapl_c.pct_change(p)
        feat['aapl_rsi_14']                 = calc_rsi(aapl_c, 14)
        feat['aapl_drawdown_20d']           = aapl_c / aapl_c.rolling(20).max() - 1
        for p in [5, 20]:
            feat[f'aapl_vs_xlk_{p}d']       = aapl_c.pct_change(p) - cl.pct_change(p)

    # 빅3 동시 내부 약화 (NVDA + MSFT + AAPL 모두 XLK보다 부진)
    big3_available = [x for x in [nvda_c, msft_c, aapl_c] if x is not None]
    if len(big3_available) >= 2:
        big3_weak_score = pd.Series(0.0, index=cl.index)
        for stk in big3_available:
            big3_weak_score += (stk.pct_change(5) < cl.pct_change(5) - 0.01).astype(float)
        feat['big3_concurrent_weakness_5d'] = big3_weak_score
        feat['big3_all_weak_5d']            = (big3_weak_score == len(big3_available)).astype(float)

        # 빅3 RSI 동시 과매수 (조정 직전 전형 패턴)
        rsi_list = []
        if nvda_c is not None and 'nvda_rsi_14' in feat.columns:
            rsi_list.append(feat['nvda_rsi_14'] > 70)
        if msft_c is not None and 'msft_rsi_14' in feat.columns:
            rsi_list.append(feat['msft_rsi_14'] > 70)
        if aapl_c is not None and 'aapl_rsi_14' in feat.columns:
            rsi_list.append(feat['aapl_rsi_14'] > 70)
        if len(rsi_list) >= 2:
            feat['big3_all_overbought']     = pd.concat(rsi_list, axis=1).all(axis=1).astype(float)

    # ── 15G. VIX 선물 구조 (백워데이션 = 극도 공포) ────────────
    vixy_c = closes.get('VIXY')
    vxx_c  = closes.get('VXX')
    vix_cv = closes.get('^VIX')

    if vixy_c is not None and vix_cv is not None:
        vixy_vix = vixy_c / vix_cv.replace(0, np.nan)
        feat['vixy_vix_ratio']              = vixy_vix
        feat['vixy_vix_ratio_zscore_20']    = calc_zscore(vixy_vix, 20)
        # 백워데이션: 선물 < 현물 = 패닉 극대화 신호
        feat['vixy_backwardation_flag']     = (vixy_c < vix_cv).astype(float)
        feat['vixy_ret_5d']                 = vixy_c.pct_change(5)
        feat['vixy_above_sma20']            = (vixy_c > vixy_c.rolling(20).mean()).astype(float)

    if vix_cv is not None:
        # VVIX 프록시: VIX의 단기 변동성 (공포의 공포)
        feat['vvix_proxy_10d']              = vix_cv.pct_change().rolling(10).std() * np.sqrt(252)
        feat['vvix_proxy_zscore_60']        = calc_zscore(feat['vvix_proxy_10d'], 60)
        # 조용히 VIX 상승 (시장은 모르는데 VIX가 먼저 움직임)
        feat['vix_silent_rise_5d']          = (
            (vix_cv.pct_change(5) > 0.12) &    # VIX 5일새 12% 이상 상승
            (cl.pct_change(5) > -0.01)          # 하지만 주가는 아직 하락 안 함
        ).astype(float)
        # 변동성 위험 프리미엄 (실현변동성 대비 내재변동성)
        rv20 = cl.pct_change().rolling(20).std() * np.sqrt(252) * 100
        feat['vol_risk_premium']            = vix_cv - rv20
        feat['vrp_zscore_60']               = calc_zscore(feat['vol_risk_premium'], 60)
        # VRP 음수 = 시장이 실제보다 위험 과소평가
        feat['vrp_negative_flag']           = (feat['vol_risk_premium'] < -2).astype(float)
        feat['market_underpricing_risk']    = (
            feat['vol_risk_premium'] < feat['vol_risk_premium'].rolling(252).quantile(0.15)
        ).astype(float)

    # ── 15H. 다중 안전자산 동시 선행 플로우 (핵심 신호) ─────────
    # 이론: 기관이 포지션 정리하면 여러 안전자산이 동시에 강해짐
    #       주가는 아직 하락 안 했는데 TLT + GLD + FXY + VIX 동시 상승
    gld_c3 = closes.get('GLD'); tlt_c3 = closes.get('TLT')
    fxy_c3 = closes.get('FXY'); hyg_c2 = closes.get('HYG')
    kre_c2 = closes.get('KRE')

    safe_flow = pd.Series(0.0, index=cl.index)

    if tlt_c3 is not None:
        safe_flow += (tlt_c3.pct_change(3) > 0.004).astype(float)         # TLT 3일 상승
        safe_flow += (tlt_c3.rolling(5).mean() > tlt_c3.rolling(20).mean()).astype(float)
    if gld_c3 is not None:
        safe_flow += (gld_c3.pct_change(3) > 0.004).astype(float)         # GLD 3일 상승
        safe_flow += (gld_c3 >= gld_c3.rolling(20).max() * 0.995).astype(float)  # 고점 근처
    if fxy_c3 is not None:
        safe_flow += (fxy_c3.pct_change(3) > 0.003).astype(float)         # 엔화 강세
        safe_flow += (fxy_c3 > fxy_c3.rolling(20).mean()).astype(float)
    if vix_cv is not None:
        safe_flow += (vix_cv.pct_change(3) > 0.06).astype(float)          # VIX 조용히 상승
        safe_flow += (vix_cv > vix_cv.rolling(20).mean()).astype(float)
    if hyg_c2 is not None:
        safe_flow += (hyg_c2.pct_change(3) < -0.005).astype(float)        # HYG 약세
    if kre_c2 is not None:
        safe_flow += (kre_c2.pct_change(5) < -0.02).astype(float)         # KRE 약세

    feat['safe_haven_flow_score']           = safe_flow      # 0~12
    # 핵심: 주가 버티는데 안전자산 4개+ 동시 신호 = 가장 강력한 선행 경고
    feat['price_up_safehaven_4plus']        = (
        (cl.pct_change(5) > -0.01) & (safe_flow >= 4)
    ).astype(float)
    feat['price_up_safehaven_6plus']        = (
        (cl.pct_change(5) > -0.01) & (safe_flow >= 6)
    ).astype(float)
    feat['safe_haven_flow_rolling5d']       = safe_flow.rolling(5).mean()
    feat['safe_haven_flow_accel']           = safe_flow - safe_flow.shift(5)
    feat['safe_haven_flow_zscore_60']       = calc_zscore(safe_flow, 60)

    # ── 15I. 방어주 조용한 기관 선매수 패턴 ────────────────────
    xlp_c2 = closes.get('XLP'); xlu_c2 = closes.get('XLU')
    xlv_c2 = closes.get('XLV'); xly_c2 = closes.get('XLY')

    def_rot_score = pd.Series(0.0, index=cl.index)
    if xlp_c2 is not None and xly_c2 is not None:
        def_rot_score += (xlp_c2.pct_change(5) > xly_c2.pct_change(5) + 0.01).astype(float)
        xlp_xly_r = xlp_c2 / xly_c2
        def_rot_score += (calc_linreg_slope(xlp_xly_r, 10) > 0).astype(float)
        feat['xlp_xly_slope_accel']         = (
            calc_linreg_slope(xlp_xly_r, 5) - calc_linreg_slope(xlp_xly_r, 20)
        )
    if xlu_c2 is not None and TICKER != 'XLU':
        def_rot_score += (xlu_c2.pct_change(5) > 0.01).astype(float)
        def_rot_score += (xlu_c2 > xlu_c2.rolling(20).mean()).astype(float)
        feat['xlu_above_sma20']             = (xlu_c2 > xlu_c2.rolling(20).mean()).astype(float)
        feat['xlu_vs_xlk_5d']               = xlu_c2.pct_change(5) - cl.pct_change(5)
    if xlv_c2 is not None:
        def_rot_score += (xlv_c2.pct_change(5) > cl.pct_change(5) + 0.01).astype(float)

    feat['institutional_defensive_rotation'] = def_rot_score
    # 방어주 3개 동시 강세 + 테크 약세
    if xlp_c2 is not None and xlu_c2 is not None and xlv_c2 is not None:
        feat['triple_defensive_vs_tech']    = (
            (xlp_c2.pct_change(5) > 0) &
            (xlu_c2.pct_change(5) > 0) &
            (xlv_c2.pct_change(5) > 0) &
            (cl.pct_change(5) < 0)
        ).astype(float)

    # ── 15J. 기관 분배(Distribution) 심화 패턴 ─────────────────
    # 스마트머니가 조용히 팔면서 가격은 유지 → 결국 무너짐

    # 스탈스(Stalls): 거래량 많은데 가격 못 오름 (매도 흡수)
    stall = (
        (cl.pct_change().abs() < 0.003) &
        (vo > vo.rolling(20).mean() * 1.25)
    ).astype(float)
    feat['stall_day_count_5d']              = stall.rolling(5).sum()
    feat['stall_day_count_10d']             = stall.rolling(10).sum()
    feat['stall_day_count_20d']             = stall.rolling(20).sum()

    # 클라이맥스 매도: 2× 거래량 + 큰 음봉 (기관 청산)
    climax_sell = (
        (cl < op) &
        (vo > vo.rolling(20).mean() * 2.0) &
        (cl.pct_change() < -0.015)
    ).astype(float)
    feat['climax_sell_3d']                  = climax_sell.rolling(3).sum()
    feat['climax_sell_5d']                  = climax_sell.rolling(5).sum()
    feat['climax_sell_20d']                 = climax_sell.rolling(20).sum()

    # 가격 신고점 + 거래량 감소 (관심 식음)
    price_new_high_20 = (cl >= cl.rolling(20).max() - 1e-9)
    low_vol_flag      = (vo < vo.rolling(20).mean() * 0.8)
    feat['new_high_low_vol_20d']            = (price_new_high_20 & low_vol_flag).astype(float).rolling(20).sum()

    # 상승일 거래량 비율 악화 (기관이 상승에 참여 안 함)
    up_vol = (vo * (cl.pct_change() > 0).astype(float)).rolling(10).sum()
    total_vol = vo.rolling(10).sum().replace(0, np.nan)
    feat['up_vol_ratio_10d']                = up_vol / total_vol
    feat['up_vol_ratio_zscore_60']          = calc_zscore(feat['up_vol_ratio_10d'], 60)
    feat['up_vol_ratio_deteriorating']      = (
        feat['up_vol_ratio_10d'] < feat['up_vol_ratio_10d'].rolling(20).mean() -
        feat['up_vol_ratio_10d'].rolling(20).std()
    ).astype(float)

    # OBV 가격 다이버전스 (기존 OBV와 다름: 단기 5일 기준)
    obv_short = (np.sign(cl.diff()) * vo).fillna(0).cumsum()
    feat['obv_price_div_5d']                = (
        (cl.pct_change(5) > 0.02) & (obv_short.diff(5) < 0)
    ).astype(float)
    feat['obv_price_div_10d']               = (
        (cl.pct_change(10) > 0.03) & (obv_short.diff(10) < 0)
    ).astype(float)

    # ── 15K. 신용시장 스마트머니 선행 ──────────────────────────
    hyg_c3 = closes.get('HYG'); lqd_c = closes.get('LQD')

    if hyg_c3 is not None:
        # HYG 하락 가속도 (이미 있는 HYG 지표와 다른 각도)
        hyg_slope = calc_linreg_slope(hyg_c3, 10)
        feat['hyg_slope_accel_5d']          = hyg_slope - hyg_slope.shift(5)
        # HYG 하락 + XLK 상승 = 신용 선행 경고 중
        feat['credit_leads_equity_flag_5d'] = (
            (hyg_c3.pct_change(5) < -0.01) & (cl.pct_change(5) > 0)
        ).astype(float)
        feat['credit_leads_equity_flag_3d'] = (
            (hyg_c3.pct_change(3) < -0.005) & (cl.pct_change(3) > 0)
        ).astype(float)
        hyg_od = ohlcv.get('HYG')
        if hyg_od is not None:
            hyg_vol_d = hyg_od['Volume']
            hyg_obv   = (np.sign(hyg_c3.diff()) * hyg_vol_d).fillna(0).cumsum()
            feat['hyg_obv_diverge_10d']     = (
                (hyg_c3.pct_change(10) > 0) & (hyg_obv.diff(10) < 0)
            ).astype(float)

    if lqd_c is not None and hyg_c3 is not None:
        # IG(LQD)가 HY(HYG)보다 먼저 하락 = 정교한 기관 먼저 빠짐
        feat['ig_leads_hy_breakdown_5d']    = (
            (lqd_c.pct_change(5) < -0.005) & (hyg_c3.pct_change(5) > -0.005)
        ).astype(float)

    # ── 15L. 갭 패턴 스마트머니 신호 ──────────────────────────
    gap_net = op / cl.shift() - 1
    # 갭 업 후 당일 반전 (기관이 갭 업을 이용해 매도)
    gap_up_reversal = (
        (gap_net > 0.005) & (cl < op)
    ).astype(float)
    feat['gap_up_reversal_3d']              = gap_up_reversal.rolling(3).sum()
    feat['gap_up_reversal_5d']              = gap_up_reversal.rolling(5).sum()
    feat['gap_up_reversal_10d']             = gap_up_reversal.rolling(10).sum()

    # 저항선 근처에서 갭 업 반전 (고점 부근 공급 압박)
    near_60d_high = (cl >= cl.rolling(60).max() * 0.98)
    feat['resistance_gap_reversal_10d']     = (
    (gap_net > 0.005) & (cl < op) & near_60d_high
    ).astype(float).rolling(10).sum()

    # 야간 갭 음수 누적 (외부 충격 프록시: 관세/AI 뉴스는 주로 오버나이트)
    neg_gap = gap_net.where(gap_net < -0.005).fillna(0)
    feat['overnight_neg_gap_cum_5d']        = neg_gap.rolling(5).sum()
    feat['overnight_neg_gap_count_20d']     = (gap_net < -0.005).astype(float).rolling(20).sum()
    feat['overnight_neg_gap_zscore_60']     = calc_zscore(neg_gap.abs(), 60)

    # ── 15M. 스마트머니 종합 스코어 ───────────────────────────
    sm_total = pd.Series(0.0, index=cl.index)

    # 인버스 ETF
    if 'sqqq_institutional_short' in feat.columns:
        sm_total += feat.get('sqqq_inst_short_3d_sum', pd.Series(0.0, index=cl.index)).clip(0, 2)
    if 'sh_inst_short_3d' in feat.columns:
        sm_total += feat['sh_inst_short_3d'].clip(0, 1)

    # 안전자산 선행
    sm_total += (safe_flow >= 5).astype(float) * 2
    sm_total += (safe_flow >= 3).astype(float)

    # 엔화 캐리 언와인드
    if 'fxy_3d_consecutive_up' in feat.columns:
        sm_total += feat['fxy_3d_consecutive_up']
    if 'yen_carry_unwind_silent_5d' in feat.columns:
        sm_total += feat['yen_carry_unwind_silent_5d']

    # 대형주 내부 약화
    if 'big3_concurrent_weakness_5d' in feat.columns:
        sm_total += (feat['big3_concurrent_weakness_5d'] >= 2).astype(float)
    if 'nvda_dist_days_10d' in feat.columns:
        sm_total += (feat['nvda_dist_days_10d'] >= 2).astype(float)

    # 방어주 조용한 선매수
    sm_total += (feat['institutional_defensive_rotation'] >= 3).astype(float)

    # 분배일 누적
    sm_total += (feat.get('climax_sell_5d', pd.Series(0.0, index=cl.index)) >= 1).astype(float)
    sm_total += (feat.get('stall_day_count_10d', pd.Series(0.0, index=cl.index)) >= 2).astype(float)

    # 신용 선행
    if 'credit_leads_equity_flag_5d' in feat.columns:
        sm_total += feat['credit_leads_equity_flag_5d']

    # EEM 괴리
    if 'eem_xlk_divergence_5d' in feat.columns:
        sm_total += feat['eem_xlk_divergence_5d']

    # GDX 강세
    if 'gdx_xlk_divergence_5d' in feat.columns:
        sm_total += feat['gdx_xlk_divergence_5d']

    # 집중도 위험
    if 'cap_concentration_flag_20d' in feat.columns:
        sm_total += feat['cap_concentration_flag_20d']

    feat['smart_money_total_score']         = sm_total

    # N일 누적 (신호 지속성이 높을수록 신뢰도 ↑)
    for p in [3, 5, 10]:
        feat[f'smart_money_rolling_{p}d']   = sm_total.rolling(p).mean()

    # 임박 하락 경보: 스마트머니 4+ & 과매수
    feat['imminent_drop_alert_rsi60']       = (
        (sm_total >= 4) & (calc_rsi(cl, 14) > 60)
    ).astype(float)
    feat['imminent_drop_alert_rsi70']       = (
        (sm_total >= 3) & (calc_rsi(cl, 14) > 70)
    ).astype(float)
    # 5일 연속 스마트머니 누적 (가장 신뢰도 높은 신호)
    feat['smart_money_5d_persistent']       = (
        sm_total.rolling(5).min() >= 2
    ).astype(float)

    # ══════════════════════════════════════════════════════════════
    #  16. 투자 대가 밸류에이션 지표
    # ══════════════════════════════════════════════════════════════

    # ── 16A. 버핏 지표 프록시 ────────────────────────────────────
    # 이론: 시총/GDP > 1 → 고평가. SPY 200일 괴리로 근사
    spy_c2 = closes.get('SPY')
    if spy_c2 is not None:
        spy_200ma = spy_c2.rolling(200).mean()
        spy_overval = spy_c2 / spy_200ma.replace(0, np.nan) - 1
        feat['buffett_spy_overval_200ma']   = spy_overval
        feat['buffett_spy_overval_zscore']  = calc_zscore(spy_overval, 252)
        feat['buffett_spy_dd_from_252h']    = spy_c2 / spy_c2.rolling(252).max() - 1
        # 고평가 + 금리 상승 = 가장 위험한 버핏 신호
        if '^TNX' in closes.columns:
            tnx_rising_b = (closes['^TNX'].diff(20) > 0.2).astype(float)
            overval_flag  = (spy_overval > 0.15).astype(float)
            feat['buffett_overval_rising_rate'] = tnx_rising_b * overval_flag

    # ── 16B. 그레이엄 안전마진 ────────────────────────────────────
    # sqrt(22.5 × EPS × BPS) 근사: 현재가 vs 200일 -2σ 하단
    bb_lower_200 = cl.rolling(200).mean() - 2 * cl.rolling(200).std()
    feat['graham_safety_margin_200'] = (cl - bb_lower_200) / cl.replace(0, np.nan)
    feat['graham_overval_flag']      = (feat['graham_safety_margin_200'] > 0.30).astype(float)
    feat['graham_pv_ratio_52w']      = cl / lo.rolling(252).min().replace(0, np.nan)
    feat['graham_pv_excess_2x']      = (feat['graham_pv_ratio_52w'] > 2.0).astype(float)

    # ── 16C. Fed 모델 (이익수익률 vs 10년물 금리) ─────────────────
    # 이익수익률 < 국채 금리 → 주식 고평가 (하락 위험)
    if '^TNX' in closes.columns:
        price_to_52w_avg    = cl / cl.rolling(252).mean().replace(0, np.nan)
        earnings_yield_p    = 1 / price_to_52w_avg.replace(0, np.nan)
        bond_yield_p        = closes['^TNX'] / 100
        feat['fed_model_spread']         = earnings_yield_p - bond_yield_p
        feat['fed_model_negative']       = (feat['fed_model_spread'] < 0).astype(float)
        feat['fed_model_spread_zscore']  = calc_zscore(feat['fed_model_spread'], 252)
        feat['fed_model_worsening_20d']  = feat['fed_model_spread'].diff(20)
        feat['fed_model_neg_rising_rate']= (
            (feat['fed_model_spread'] < 0) & (closes['^TNX'].diff(5) > 0.1)
        ).astype(float)

    # ── 16D. 켈리 기준 (Kelly Criterion) ─────────────────────────
    # Kelly = (b×p − q) / b  →  < 0 이면 베팅하지 말 것
    for kp in [20, 60]:
        win_rate_k = (ret_d > 0).rolling(kp).mean()
        avg_win_k  = ret_d.where(ret_d > 0).rolling(kp).mean().abs()
        avg_loss_k = ret_d.where(ret_d < 0).rolling(kp).mean().abs()
        b_k        = avg_win_k / avg_loss_k.replace(0, np.nan)
        kelly_v    = (b_k * win_rate_k - (1 - win_rate_k)) / b_k.replace(0, np.nan)
        feat[f'kelly_signal_{kp}d']        = kelly_v
        feat[f'kelly_negative_{kp}d']      = (kelly_v < 0).astype(float)
        feat[f'kelly_deteriorating_{kp}d'] = (kelly_v - kelly_v.shift(kp // 2) < -0.10).astype(float)

    # ── 16E. 소로스 반사성 이론 (가격 가속도가 극에 달할 때 반전) ───
    p1d  = cl.pct_change(5)
    p2d  = p1d.diff(5)
    feat['soros_reflexivity_accel']  = p2d
    feat['soros_reflexivity_spike']  = (
        p2d.abs() > p2d.abs().rolling(60).mean() * 2
    ).astype(float)
    feat['soros_bubble_score']       = (p2d / (p1d.abs() + 1e-6)).clip(-10, 10)

    # ── 16F. 드러켄밀러 유동성 지표 (유동성 감소 → 위험자산 선행 하락) ─
    if 'BIL' in closes.columns and 'TLT' in closes.columns:
        liq_proxy = closes['BIL'].pct_change(20) - closes['TLT'].pct_change(20)
        feat['druckenmiller_liquidity_20d'] = liq_proxy
        feat['druckenmiller_tightening']    = (liq_proxy < 0).astype(float)
    if '^TNX' in closes.columns:
        tnx_accel_d = closes['^TNX'].diff(5) - closes['^TNX'].diff(10).shift(5)
        feat['druckenmiller_rate_shock']    = (tnx_accel_d > 0.15).astype(float)

    # ── 16G. 샤프(Sharpe) CAPE 근사 (고평가 국면 구분) ─────────────
    # Shiller CAPE = 현재 가격 / 10년 인플레 조정 평균 이익
    # ETF 프록시: 현재가 / 10년 롤링 최저가 (내재가치 하한선)
    rolling_min_10y = cl.rolling(min(252 * 5, len(cl))).min()
    feat['shiller_cape_proxy']      = cl / rolling_min_10y.replace(0, np.nan)
    feat['shiller_cape_extreme']    = (feat['shiller_cape_proxy'] > feat['shiller_cape_proxy'].rolling(252).quantile(0.90)).astype(float)
    feat['shiller_cape_zscore_252'] = calc_zscore(feat['shiller_cape_proxy'], 252)

    # ── 16H. 토빈 Q 근사 (시장가치 vs 대체비용) ──────────────────
    # Tobin Q > 1 → 고평가. 프록시: 현재가 / 52주 볼린저 중심
    tobin_q_proxy = cl / cl.rolling(252).mean().replace(0, np.nan)
    feat['tobin_q_proxy']           = tobin_q_proxy
    feat['tobin_q_above_1_2']       = (tobin_q_proxy > 1.2).astype(float)
    feat['tobin_q_zscore']          = calc_zscore(tobin_q_proxy, 252)
    feat['tobin_q_declining_20d']   = (tobin_q_proxy.diff(20) < -0.05).astype(float)

    # ══════════════════════════════════════════════════════════════
    #  17. 인플레이션 / 물가 충격 지표
    # ══════════════════════════════════════════════════════════════

    # ── 17A. 원자재 인플레이션 복합 지수 ────────────────────────
    uso_ci  = closes.get('USO')
    cper_ci = closes.get('CPER')
    dbc_ci  = closes.get('DBC')
    gld_ci  = closes.get('GLD')
    pdbc_ci = closes.get('PDBC')

    comm_score = pd.Series(0.0, index=cl.index)
    comm_cnt   = 0
    for sym_ci, s_ci in [('USO', uso_ci), ('CPER', cper_ci),
                          ('DBC', dbc_ci), ('GLD', gld_ci)]:
        if s_ci is None: continue
        z = calc_zscore(s_ci.pct_change(20).fillna(0), 60).fillna(0)
        comm_score += z; comm_cnt += 1
    if comm_cnt > 0:
        feat['commodity_inflation_composite'] = comm_score / comm_cnt
        feat['commodity_inflation_high']      = (feat['commodity_inflation_composite'] > 1.5).astype(float)

    # 오일 충격 플래그 (30일 +20% 이상)
    if uso_ci is not None:
        feat['oil_30d_chg']     = uso_ci.pct_change(30)
        feat['oil_shock_30d']   = (uso_ci.pct_change(30) > 0.20).astype(float)
        feat['oil_spike_5d']    = (uso_ci.pct_change(5)  > 0.08).astype(float)
        for p in [10, 20]:
            feat[f'oil_tech_inverse_{p}d'] = uso_ci.pct_change(p) - cl.pct_change(p)

    # ── 17B. 인플레 기대 충격 (BEI Proxy: IEF - TIP) ────────────
    tip_ci = closes.get('TIP'); ief_ci = closes.get('IEF')
    if tip_ci is not None and ief_ci is not None:
        bei_p = ief_ci.pct_change(20) - tip_ci.pct_change(20)
        feat['bei_proxy_20d']       = bei_p
        feat['bei_spike_up']        = (bei_p > bei_p.rolling(60).mean() + bei_p.rolling(60).std()).astype(float)
        feat['bei_spike_down']      = (bei_p < bei_p.rolling(60).mean() - bei_p.rolling(60).std()).astype(float)
        feat['bei_accel_5d']        = bei_p.diff(5)
        feat['bei_zscore_60d']      = calc_zscore(bei_p, 60)
        # BEI 급등 + 금리 상승 = 스태그플레이션 전조
        if '^TNX' in closes.columns:
            feat['stagflation_bei_tnx'] = (
                feat['bei_spike_up'] * (closes['^TNX'].diff(10) > 0.2).astype(float)
            )

    # ── 17C. 실질 금리 충격 (Real Rate Shock) ────────────────────
    # 실질금리 급등 = 성장주 DCF 할인율 폭등 → XLK 직격
    if '^TNX' in closes.columns:
        tnx_ci = closes['^TNX']
        feat['nominal_rate_shock_5d']  = (tnx_ci.diff(5)  > 0.20).astype(float)
        feat['nominal_rate_shock_20d'] = (tnx_ci.diff(20) > 0.50).astype(float)
        feat['rate_vol_20d_v2']        = tnx_ci.diff().rolling(20).std()
        feat['rate_vol_zscore_v2']     = calc_zscore(feat['rate_vol_20d_v2'].fillna(0), 60)
        if tip_ci is not None:
            real_shock = tnx_ci.diff(10) + tip_ci.pct_change(10) * 10
            feat['real_rate_shock_10d']  = real_shock
            feat['real_rate_shock_flag'] = (
                real_shock > real_shock.rolling(60).mean() + real_shock.rolling(60).std()
            ).astype(float)

    # ── 17D. 달러 Wrecking Ball (강달러 + 긴축 = 글로벌 유동성 위축) ─
    uup_ci = closes.get('UUP')
    if uup_ci is not None:
        for p in [10, 20, 60]:
            feat[f'dxy_momentum_{p}d'] = uup_ci.pct_change(p)
        feat['dxy_above_sma50']      = (uup_ci > uup_ci.rolling(50).mean()).astype(float)
        feat['dxy_zscore_60d']       = calc_zscore(uup_ci.pct_change(20).fillna(0), 60)
        dxy_accel = uup_ci.pct_change(10) - uup_ci.pct_change(20).shift(10)
        feat['dxy_acceleration_flag']= (dxy_accel > 0.02).astype(float)
        if dbc_ci is not None:
            feat['dollar_commodity_stress'] = (
                (uup_ci.pct_change(20) > 0.02) & (dbc_ci.pct_change(20) < -0.03)
            ).astype(float)

    # ── 17E. PPI 근사 vs CPI 프록시 갭 (생산자-소비자 압력 갭) ────
    # PPI 프록시: 원자재(DBC) 상승 / CPI 프록시: 소매(XRT) 상승
    xrt_ci = closes.get('XRT')
    if dbc_ci is not None and xrt_ci is not None:
        ppi_proxy = dbc_ci.pct_change(60)
        cpi_proxy = xrt_ci.pct_change(60)
        feat['ppi_cpi_gap_60d']       = ppi_proxy - cpi_proxy
        feat['ppi_squeeze_flag']      = (feat['ppi_cpi_gap_60d'] > 0.05).astype(float)
        feat['ppi_cpi_gap_zscore']    = calc_zscore(feat['ppi_cpi_gap_60d'].fillna(0), 60)

    # ── 17F. 스태그플레이션 종합 점수 (2025 Liberation Day 유형) ──
    stag_score = pd.Series(0.0, index=cl.index)
    if uso_ci  is not None: stag_score += (uso_ci.pct_change(20) > 0.08).astype(float)
    if '^TNX' in closes.columns: stag_score += (closes['^TNX'].diff(20) > 0.3).astype(float)
    if uup_ci  is not None: stag_score += feat.get('dxy_above_sma50',
                                          pd.Series(0.0, index=cl.index))
    stag_score += (cl.pct_change(20) < -0.03).astype(float)
    if 'KRE' in closes.columns: stag_score += (closes['KRE'].pct_change(10) < -0.03).astype(float)
    feat['stagflation_composite_v2'] = stag_score

# ══════════════════════════════════════════════════════════════
    #  18. 신규 핵심 하락 선행 지표 (~200개)
    #      — 팩터 로테이션·역사적 붕괴 패턴·금융조건지수·
    #        차트패턴 계량화·수학 심화(적분/고차미분/위상공간/
    #        정보이론)·글로벌 전이·미시구조·레짐 변화
    # ══════════════════════════════════════════════════════════════
    _ret1 = cl.pct_change()

    # ── 18A. 팩터 ETF 로테이션 (기관 스마트머니 선행) ────────────
    # 이론: MTUM(모멘텀 팩터) 청산은 헤지펀드 디레버리징의 첫 신호
    mtum_c = closes.get('MTUM')
    splv_c2 = closes.get('SPLV')
    sphb_c2 = closes.get('SPHB')

    if mtum_c is not None:
        for p in [5, 10, 20, 60]:
            feat[f'mtum_ret_{p}d'] = mtum_c.pct_change(p)
        feat['mtum_rsi_14']         = calc_rsi(mtum_c, 14)
        feat['mtum_dd_20d']         = mtum_c / mtum_c.rolling(20).max() - 1
        feat['mtum_dd_60d']         = mtum_c / mtum_c.rolling(60).max() - 1
        feat['mtum_vs_xlk_5d']      = mtum_c.pct_change(5) - cl.pct_change(5)
        feat['mtum_vs_xlk_20d']     = mtum_c.pct_change(20) - cl.pct_change(20)
        feat['mtum_below_sma50']    = (mtum_c < mtum_c.rolling(50).mean()).astype(float)
        feat['mtum_pctrank_60']     = calc_pctrank(mtum_c, 60)
        feat['mtum_factor_crash']   = (feat['mtum_pctrank_60'] < 0.20).astype(float)
        feat['mtum_zscore_20']      = calc_zscore(mtum_c.pct_change(), 20)
        # MTUM 하락 + XLK 상승 = 가장 위험한 팩터 다이버전스
        feat['mtum_xlk_diverge_5d'] = (
            (mtum_c.pct_change(5) < -0.02) & (cl.pct_change(5) > 0)
        ).astype(float)
        # 모멘텀 팩터 연속 하락 (3일)
        feat['mtum_3d_consecutive_down'] = (
            (mtum_c.pct_change() < 0) &
            (mtum_c.pct_change().shift(1) < 0) &
            (mtum_c.pct_change().shift(2) < 0)
        ).astype(float)

    if splv_c2 is not None and sphb_c2 is not None:
        lvhb_r = splv_c2 / sphb_c2.replace(0, np.nan)
        feat['splv_sphb_ratio_accel_5d']    = lvhb_r.pct_change(5) - lvhb_r.pct_change(10).shift(5)
        feat['splv_sphb_ratio_zscore_60']   = calc_zscore(lvhb_r, 60)
        feat['splv_sphb_pctrank_252']       = calc_pctrank(lvhb_r, 252)
        # 3일 연속 저변동성 우세 = 기관 방어 전환 가속
        feat['splv_dominates_3d'] = (
            (splv_c2.pct_change() > sphb_c2.pct_change()) &
            (splv_c2.pct_change().shift(1) > sphb_c2.pct_change().shift(1)) &
            (splv_c2.pct_change().shift(2) > sphb_c2.pct_change().shift(2))
        ).astype(float)

    # SQQQ 거래량 52주 분위수 (사상 최대치 접근 = 기관 풋 패닉 매수)
    sqqq_ov = ohlcv.get('SQQQ')
    if sqqq_ov is not None:
        sv52 = sqqq_ov['Volume']
        feat['sqqq_vol_pctrank_252']   = calc_pctrank(sv52, 252)
        feat['sqqq_vol_52w_extreme']   = (feat['sqqq_vol_pctrank_252'] > 0.90).astype(float)
        feat['sqqq_vol_new_high_20d']  = (sv52 >= sv52.rolling(20).max() - 1).astype(float)

    # ── 18B. 역사적 주요 붕괴 패턴 재현 지표 ────────────────────
    # ── [2000 닷컴] 극단 모멘텀 소진 패턴
    six_mo_ret  = cl.pct_change(126)
    feat['dotcom_topping_pattern'] = (
        (six_mo_ret > 0.25) & (cl.pct_change(5) < -0.02)
    ).astype(float)
    feat['dotcom_momentum_exhaust'] = (
        (six_mo_ret > 0.20) &
        (calc_rsi(cl, 14) < calc_rsi(cl, 14).shift(10)) &
        (cl.pct_change(20) < cl.pct_change(20).shift(20))
    ).astype(float)
    # 6개월 수익률 분위수 극단 (상위 5% = 과열)
    feat['six_month_ret_pctrank'] = calc_pctrank(six_mo_ret.fillna(0), 252)
    feat['dotcom_overshoot_flag'] = (feat['six_month_ret_pctrank'] > 0.95).astype(float)

    # ── [2008 GFC] 은행-신용 동반 붕괴 지문
    _kre = closes.get('KRE'); _hyg = closes.get('HYG')
    if _kre is not None and _hyg is not None:
        feat['gfc_bank_credit_score'] = (
            (_kre.pct_change(20) < -0.08).astype(float) +
            (_hyg.pct_change(10) < -0.03).astype(float) +
            (calc_rsi(_kre, 14) < 35).astype(float)
        )
        feat['gfc_pattern_trigger'] = (feat['gfc_bank_credit_score'] >= 2).astype(float)
        # 은행주 + 신용 동시 60일 신저가 (리먼 패턴)
        feat['dual_new_low_60d_gfc'] = (
            (_kre <= _kre.rolling(60).min() * 1.01) &
            (_hyg <= _hyg.rolling(60).min() * 1.01)
        ).astype(float)

    # ── [2020 코로나] 유동성 소멸 지문
    # 특징: VIX 2일 40%+ 급등 + 거래량 3배 폭발
    _vix_d2_spike = pd.Series(0.0, index=cl.index)
    if '^VIX' in closes.columns:
        _vix_d2_spike = (closes['^VIX'].pct_change(2) > 0.40).astype(float)
    feat['covid_vol_liquidity_crash'] = (
        _vix_d2_spike * (vo / vo.rolling(60).mean() > 3.0).astype(float)
    )
    # 유동성 진공: 일중 범위 확대 + 거래량 급감 (bid-ask 스프레드 급팽창 proxy)
    feat['liquidity_vacuum_signal'] = (
        ((hi - lo) / cl > ((hi - lo) / cl).rolling(20).mean() * 1.5) &
        (vo / vo.rolling(20).mean() < 0.6)
    ).astype(float)
    feat['liquidity_vacuum_5d_sum'] = feat['liquidity_vacuum_signal'].rolling(5).sum()

    # ── [2022 금리 충격] 실질금리 급등 → 성장주 DCF 붕괴
    _tnx = closes.get('^TNX')
    if _tnx is not None:
        feat['rate_shock_60bp_20d']       = (_tnx.diff(20) > 0.60).astype(float)
        feat['rate_shock_100bp_60d']      = (_tnx.diff(60) > 1.00).astype(float)
        feat['rate_rise_velocity_pctrank']= calc_pctrank(_tnx.diff(20).fillna(0), 252)
        # 고평가 + 금리 충격 = 닷컴 2.0 시나리오
        feat['valuation_reset_signal'] = (
            feat['rate_shock_60bp_20d'] *
            (cl / cl.rolling(252).mean() > 1.20).astype(float)
        )
        # 금리 변동성 (당일 등락의 20일 표준편차) — 기존 rate_vol_20d_v2와 이름 다름
        feat['rate_daily_vol_20d']  = _tnx.diff(1).rolling(20).std()
        feat['rate_vol_pctrank_252']= calc_pctrank(feat['rate_daily_vol_20d'].fillna(0), 252)

    # ── [2025 관세 연쇄 충격] 달러+원자재+신흥국 복합
    _uup = closes.get('UUP'); _eem = closes.get('EEM'); _uso = closes.get('USO')
    if _uup is not None and _eem is not None:
        _ts = pd.Series(0.0, index=cl.index)
        _ts += (_uup.pct_change(10) > 0.02).astype(float)
        _ts += (_eem.pct_change(10) < -0.03).astype(float)
        if _uso is not None:
            _ts += (_uso.pct_change(10) < -0.05).astype(float)
        feat['tariff_global_shock_score'] = _ts
        feat['tariff_shock_trigger']      = (_ts >= 2).astype(float)

    # ── 18C. 금융 조건 지수 (FCI) ────────────────────────────────
    # 이론: FCI 악화는 주식 하락 6~12주 선행
    fci = pd.Series(0.0, index=cl.index); fci_n = 0
    if _tnx is not None:
        fci += calc_zscore(_tnx.ffill(), 252).fillna(0); fci_n += 1
    _lqd = closes.get('LQD'); _hyg2 = closes.get('HYG')
    if _lqd is not None and _hyg2 is not None:
        # 신용 스프레드 확대 = FCI 악화
        fci -= calc_zscore((_lqd.pct_change(20) - _hyg2.pct_change(20)).fillna(0), 252).fillna(0)
        fci_n += 1
    if _uup is not None:
        fci += calc_zscore(_uup.pct_change(20).fillna(0), 252).fillna(0); fci_n += 1
    if '^VIX' in closes.columns:
        fci += calc_zscore(closes['^VIX'].ffill(), 252).fillna(0); fci_n += 1
    if fci_n > 0:
        feat['fci_composite']          = fci / fci_n
        feat['fci_tightening']         = (feat['fci_composite'] > 1.0).astype(float)
        feat['fci_extreme_tightening'] = (feat['fci_composite'] > 2.0).astype(float)
        feat['fci_accel_20d']          = feat['fci_composite'] - feat['fci_composite'].shift(20)
        feat['fci_accel_flag']         = (feat['fci_accel_20d'] > 0.5).astype(float)

    # 신용 임펄스 (Credit Impulse): 신용 변화율의 변화율
    # = 경기 둔화 6개월 선행 (BIS 검증 지표)
    if _lqd is not None:
        lqd_m60 = _lqd.pct_change(60)
        feat['credit_impulse_60d']      = lqd_m60 - lqd_m60.shift(60)
        feat['credit_impulse_negative'] = (feat['credit_impulse_60d'] < -0.05).astype(float)
        feat['credit_impulse_zscore']   = calc_zscore(feat['credit_impulse_60d'].fillna(0), 252)

    # Repo 시장 스트레스 (BIL 급등 + TLT 약세 = 단기 자금 수요 폭발)
    _bil = closes.get('BIL'); _tlt = closes.get('TLT')
    if _bil is not None and _tlt is not None:
        _repo = _bil.pct_change(5) - _tlt.pct_change(5)
        feat['repo_stress_proxy']  = _repo
        feat['repo_stress_flag']   = (
            _repo > _repo.rolling(60).mean() + _repo.rolling(60).std() * 1.5
        ).astype(float)
        # 조용한 repo 위기: BIL 상승 + 주가 버팀 (시장은 아직 모름)
        feat['stealthy_repo_stress'] = (
            (_bil.pct_change(3) > 0.001) & (cl.pct_change(3) > -0.01)
        ).astype(float)
        feat['repo_stress_zscore_60'] = calc_zscore(_repo.fillna(0), 60)

    # BKLN (레버리지 론) 약세 = 쉐도우 뱅킹 신용 경색
    _bkln = closes.get('BKLN')
    if _bkln is not None:
        feat['bkln_ret_10d']        = _bkln.pct_change(10)
        feat['bkln_ret_20d']        = _bkln.pct_change(20)
        feat['bkln_below_sma50']    = (_bkln < _bkln.rolling(50).mean()).astype(float)
        feat['bkln_dd_20d']         = _bkln / _bkln.rolling(20).max() - 1
        feat['bkln_zscore_20']      = calc_zscore(_bkln.pct_change(), 20)
        # BKLN 하락 + XLK 상승 = 신용 경색 선행 경고
        feat['bkln_xlk_diverge_5d'] = (
            (_bkln.pct_change(5) < -0.01) & (cl.pct_change(5) > 0)
        ).astype(float)

    # ── 18D. 차트 패턴 계량화 ────────────────────────────────────
    # ── [헤드앤숄더] 계량화
    head_h    = hi.rolling(20).max()
    l_shoulder= hi.shift(30).rolling(20).max()
    neckline  = lo.rolling(20).min()
    feat['head_shoulders_forming'] = (
        (head_h > l_shoulder * 1.03) &
        (cl < head_h * 0.98) &
        (cl > neckline)
    ).astype(float)
    feat['head_shoulders_break'] = (
        feat['head_shoulders_forming'].shift(5).rolling(10).max() == 1
    ).astype(float) * (cl < neckline).astype(float)

    # ── [상승 쐐기 (Rising Wedge)] 고점 기울기 < 저점 기울기
    def _slope20(s):
        return s.rolling(20).apply(
            lambda x: (np.polyfit(np.arange(20), x, 1)[0] /
                       (abs(float(np.mean(x))) + 1e-10))
            if not np.isnan(x).any() else np.nan, raw=True
        )
    hi_sl20 = _slope20(hi); lo_sl20 = _slope20(lo)
    feat['rising_wedge_flag'] = (
        (hi_sl20 > 0) & (lo_sl20 > 0) &
        (hi_sl20 < lo_sl20) &           # 저점이 더 가파름 = 쐐기
        (cl.pct_change(20) > 0.03)
    ).astype(float)
    feat['rising_wedge_strength'] = (lo_sl20 - hi_sl20).clip(lower=0)

    # ── [더블 탑] 고점 2회 + 하락
    prev_h20 = hi.shift(20).rolling(20).max()
    curr_h5  = hi.rolling(5).max()
    feat['double_top_flag'] = (
        ((curr_h5 / prev_h20.replace(0, np.nan) - 1).abs() < 0.02) &
        (cl.pct_change(5) < -0.01)
    ).astype(float)

    # ── [베어 플래그] 급락 후 약한 반등
    feat['bear_flag_pattern'] = (
        cl.pct_change(10).shift(5) < -0.05        # 10일 전 급락
    ).astype(float) * (
        (cl.pct_change(3) > 0) & (cl.pct_change(3) < 0.02)  # 지금 약한 반등
    ).astype(float)

    # ── [데스크로스 심화]
    sma50_v = cl.rolling(50).mean(); sma200_v = cl.rolling(200).mean()
    feat['death_cross_active']      = (sma50_v < sma200_v).astype(float)
    feat['death_cross_depth']       = (sma200_v - sma50_v) / sma200_v.replace(0, np.nan)
    feat['death_cross_momentum']    = (
        (sma50_v < sma200_v) & (cl.pct_change(20) < -0.05)
    ).astype(float)
    feat['death_cross_vol_confirm'] = (
        (sma50_v < sma200_v) & (vo > vo.rolling(20).mean() * 1.2)
    ).astype(float)

    # ── [불 트랩] 골든크로스 직후 재하락
    golden_cross = (
        (sma50_v > sma200_v) & (sma50_v.shift(10) < sma200_v.shift(10))
    )
    feat['golden_cross_bull_trap'] = (
        golden_cross.rolling(15).max().fillna(0) == 1
    ).astype(float) * (cl.pct_change(5) < -0.03).astype(float)

    # ── [채널 하단 이탈] 지지선 붕괴
    for _cp in [20, 50, 100]:
        _lower = lo.rolling(_cp).min()
        feat[f'support_break_{_cp}d'] = (
            (cl < _lower.shift(1)) & (cl.shift(1) >= _lower.shift(2))
        ).astype(float)

    # ── [변동성 채널 이탈] 기존 Keltner 하단 이탈 + 거래량 확인
    for _kp, _km in [(20, 2.0), (20, 1.5)]:
        _kmid = cl.ewm(span=_kp, adjust=False).mean()
        _katr, _ = calc_atr(hi, lo, cl, _kp)
        _klo  = _kmid - _km * _katr
        feat[f'keltner_break_{_kp}_{_km}'] = (
            (cl < _klo) & (vo > vo.rolling(20).mean())
        ).astype(float)

    # ── 18E. 수학 심화 (적분·고차미분·위상공간·정보이론) ────────
    # ── [적분] 드로다운 면적 (Area Under Drawdown Curve)
    # 깊이 × 기간 = 고통 총량 (단순 최대낙폭보다 더 정밀)
    for _dp in [20, 60]:
        _rmax = cl.rolling(_dp).max()
        _dd   = (cl / _rmax - 1).clip(upper=0)
        feat[f'drawdown_area_{_dp}d']       = _dd.rolling(_dp).sum()
        feat[f'drawdown_area_accel_{_dp}d'] = (
            feat[f'drawdown_area_{_dp}d'] -
            feat[f'drawdown_area_{_dp}d'].shift(_dp // 2)
        )
        # 드로다운 면적 분위수 (역사적으로 얼마나 심각한가)
        feat[f'drawdown_area_pctrank_{_dp}d'] = calc_pctrank(
            feat[f'drawdown_area_{_dp}d'].fillna(0), 252
        )

    # ── [고차미분] 가격 3차 도함수 (Jerk = 가속도 변화율)
    _pvel  = cl.diff(1)                    # 1차: 속도
    _pacc  = _pvel.diff(1)                 # 2차: 가속도
    _pjerk = _pacc.diff(1)                 # 3차: 저크
    feat['price_velocity_5ma']  = _pvel.rolling(5).mean() / cl
    feat['price_accel_5ma']     = _pacc.rolling(5).mean() / cl
    feat['price_jerk_5ma']      = _pjerk.rolling(5).mean() / cl
    feat['price_jerk_spike']    = (
        _pjerk.abs() > _pjerk.abs().rolling(60).mean() * 3
    ).astype(float)
    # 가속도 부호 반전 (상승 가속 → 하락 가속 전환)
    feat['accel_sign_flip_to_neg'] = (
        (_pacc.rolling(3).mean() < 0) &
        (_pacc.rolling(3).mean().shift(5) > 0)
    ).astype(float)

    # ── [적분 기반 모멘텀] 임펄스 = 속도의 지수가중 누적
    for _ip in [10, 20]:
        _impulse = _pvel.ewm(span=_ip, adjust=False).mean() * _ip
        feat[f'price_impulse_{_ip}d']          = _impulse / cl
        feat[f'price_impulse_sign_flip_{_ip}d'] = (
            (np.sign(_impulse) < 0) & (np.sign(_impulse.shift(_ip)) > 0)
        ).astype(float)

    # ── [Lyapunov 지수 근사] 시스템 카오스도 측정
    # 양수 = 예측 불가한 카오틱 상태 = 큰 움직임 임박
    def _lyapunov(x):
        x = x[~np.isnan(x)]
        if len(x) < 10: return np.nan
        d = np.abs(np.diff(x))
        return float(np.mean(np.log(d + 1e-10)))

    feat['lyapunov_proxy_20d']  = _ret1.rolling(20).apply(_lyapunov, raw=True)
    feat['lyapunov_proxy_60d']  = _ret1.rolling(60).apply(_lyapunov, raw=True)
    feat['lyapunov_rising_20d'] = (
        feat['lyapunov_proxy_20d'] > feat['lyapunov_proxy_60d']
    ).astype(float)
    feat['lyapunov_zscore_60']  = calc_zscore(feat['lyapunov_proxy_20d'].fillna(0), 60)

    # ── [재현 정량화 (Recurrence Rate)] 현재가가 과거와 얼마나 유사한가
    def _recurrence(x, eps=0.02):
        x = x[~np.isnan(x)]
        if len(x) < 10: return np.nan
        return float(np.sum(np.abs(x[:-1] - x[-1]) / (abs(x[-1]) + 1e-10) < eps) / max(len(x)-1, 1))

    feat['recurrence_rate_20d']  = cl.rolling(20).apply(_recurrence, raw=True)
    feat['recurrence_rate_60d']  = cl.rolling(60).apply(_recurrence, raw=True)
    # 재현율 급등 = 과거 저항/지지 수준 근접 (반전 가능성)
    feat['recurrence_rate_spike'] = (
        feat['recurrence_rate_20d'] > feat['recurrence_rate_20d'].rolling(60).mean() * 2
    ).astype(float)

    # ── [위상공간 재구성 (Takens Embedding)] lag-1,2,3 상태공간 거리
    _l1 = _ret1; _l2 = _ret1.shift(1); _l3 = _ret1.shift(2)
    _phase_dist = np.sqrt(_l1**2 + _l2.fillna(0)**2 + _l3.fillna(0)**2)
    feat['phase_space_dist_5ma']   = _phase_dist.rolling(5).mean()
    feat['phase_space_expansion']  = (
        feat['phase_space_dist_5ma'] / feat['phase_space_dist_5ma'].rolling(20).mean() - 1
    )
    feat['phase_space_zscore_60']  = calc_zscore(_phase_dist, 60)

    # ── [푸리에 지배 주기] 가격 사이클 붕괴 = 레짐 전환
    def _dominant_period(x):
        x = x[~np.isnan(x)]
        if len(x) < 20: return np.nan
        f = np.abs(np.fft.rfft(x - np.mean(x)))
        idx = np.argmax(f[1:]) + 1
        return float(len(x) / max(idx, 1))

    feat['fourier_dominant_period_60d'] = _ret1.rolling(60).apply(_dominant_period, raw=True)
    feat['fourier_period_shortening']   = (
        feat['fourier_dominant_period_60d'] <
        feat['fourier_dominant_period_60d'].shift(20) * 0.7
    ).astype(float)

    # ── [전이 엔트로피 proxy] VIX → XLK 인과 강도
    if '^VIX' in closes.columns:
        _vd = np.sign(closes['^VIX'].diff())
        _xd = np.sign(cl.diff())
        # VIX 상승 → 다음날 XLK 하락하는 비율 (20일 이동)
        feat['vix_to_xlk_transfer_20d'] = (
            (_vd.shift(1) == 1) & (_xd == -1)
        ).astype(float).rolling(20).mean()
        feat['vix_xlk_transfer_zscore'] = calc_zscore(
            feat['vix_to_xlk_transfer_20d'].fillna(0), 60
        )

    # ── [Kolmogorov 복잡도 근사] 런-길이 기반 패턴 무질서도
    def _kolmogorov(x):
        x = x[~np.isnan(x)]
        if len(x) < 10: return np.nan
        b = (x > np.median(x)).astype(int)
        runs = 1 + int(np.sum(np.diff(b) != 0))
        return float(runs / len(x))

    feat['kolmogorov_complexity_20d'] = _ret1.rolling(20).apply(_kolmogorov, raw=True)
    feat['kolmogorov_spike'] = (
        feat['kolmogorov_complexity_20d'] >
        feat['kolmogorov_complexity_20d'].rolling(60).mean() +
        feat['kolmogorov_complexity_20d'].rolling(60).std() * 1.5
    ).astype(float)

    # ── [Roll's Spread 추정] 연속 수익률 자기공분산 기반 bid-ask 추정
    # Roll(1984): S = 2√(-Cov(r_t, r_{t-1}))
    _cov_roll = _ret1.rolling(20).apply(
        lambda x: np.cov(x[1:], x[:-1])[0, 1] if len(x) > 5 else np.nan, raw=True
    )
    feat['roll_spread_est']       = 2 * np.sqrt((-_cov_roll).clip(lower=0))
    feat['roll_spread_zscore_60'] = calc_zscore(feat['roll_spread_est'].fillna(0), 60)
    feat['roll_spread_widening']  = (
        feat['roll_spread_est'] > feat['roll_spread_est'].rolling(20).mean() * 1.5
    ).astype(float)

    # ── [VPIN proxy] 거래량 기반 정보 거래 확률
    # 높을수록 내부자/기관 정보 거래 비중 증가 = 곧 큰 움직임
    _upv = (vo * (cl > cl.shift()).astype(float)).fillna(0)
    _dnv = (vo * (cl < cl.shift()).astype(float)).fillna(0)
    for _vp in [10, 20]:
        _tv = (_upv + _dnv).rolling(_vp).sum().replace(0, np.nan)
        _vpin = (_upv - _dnv).abs().rolling(_vp).sum() / _tv
        feat[f'vpin_proxy_{_vp}d']    = _vpin
        feat[f'vpin_high_{_vp}d']     = (_vpin > 0.5).astype(float)
    feat['vpin_rising_5d'] = (
        feat['vpin_proxy_10d'] > feat['vpin_proxy_10d'].shift(5)
    ).astype(float)

    # ── 18F. 글로벌 전이 위험 ────────────────────────────────────
    # ── 중국 (FXI / MCHI)
    _fxi = closes.get('FXI'); _mchi = closes.get('MCHI')
    for _sym, _lbl in [(_fxi, 'fxi'), (_mchi, 'mchi')]:
        if _sym is None: continue
        for p in [5, 20, 60]:
            feat[f'{_lbl}_ret_{p}d']   = _sym.pct_change(p)
        feat[f'{_lbl}_dd_20d']         = _sym / _sym.rolling(20).max() - 1
        feat[f'{_lbl}_below_sma200']   = (_sym < _sym.rolling(200).mean()).astype(float)
        feat[f'{_lbl}_zscore_20']      = calc_zscore(_sym.pct_change(), 20)
        # 중국 선행 하락 (미국보다 먼저 떨어짐)
        feat[f'{_lbl}_leads_xlk_5d']   = (
            (_sym.pct_change(5) < -0.03) & (cl.pct_change(5) > -0.01)
        ).astype(float)

    # ── 유럽 (EWG: Germany)
    _ewg = closes.get('EWG')
    if _ewg is not None:
        for p in [5, 20, 60]:
            feat[f'ewg_ret_{p}d']   = _ewg.pct_change(p)
        feat['ewg_below_sma50']     = (_ewg < _ewg.rolling(50).mean()).astype(float)
        feat['ewg_dd_20d']          = _ewg / _ewg.rolling(20).max() - 1
        feat['ewg_zscore_20']       = calc_zscore(_ewg.pct_change(), 20)
        _spy2 = closes.get('SPY')
        if _spy2 is not None:
            feat['ewg_vs_spy_20d']  = _ewg.pct_change(20) - _spy2.pct_change(20)
            feat['europe_us_diverge_20d'] = (
                (_ewg.pct_change(20) < -0.04) & (_spy2.pct_change(20) > 0.01)
            ).astype(float)

    # ── 글로벌 동시 붕괴 스코어
    _global_risk = pd.Series(0.0, index=cl.index)
    for _gs in [_ewg, _fxi, closes.get('EEM'), closes.get('DBC')]:
        if _gs is not None:
            _global_risk += (_gs.pct_change(5) < -0.02).astype(float)
    feat['global_simultaneous_drop_score'] = _global_risk
    feat['global_4asset_drop_flag']        = (_global_risk >= 3).astype(float)

    # 달러 강세 + 신흥국 약세 복합 (EM 자본 이탈)
    if _uup is not None and closes.get('EEM') is not None:
        _eem2 = closes.get('EEM')
        for p in [5, 10, 20]:
            feat[f'em_dollar_stress_{p}d'] = _uup.pct_change(p) - _eem2.pct_change(p)
        feat['em_dollar_stress_extreme'] = (
            feat['em_dollar_stress_20d'] >
            feat['em_dollar_stress_20d'].rolling(60).mean() +
            feat['em_dollar_stress_20d'].rolling(60).std() * 2
        ).astype(float)

    # ── 18G. 유동성 미시구조 심화 ────────────────────────────────
    # ── Kyle's Lambda (가격충격계수 = 대형 매도 취약성 측정)
    # λ = |수익률| / 달러거래량 — 높을수록 큰 주문에 취약
    _dv  = cl * vo
    _kyl = _ret1.abs() / _dv.replace(0, np.nan) * 1e8
    feat['kyle_lambda_5d']       = _kyl.rolling(5).mean()
    feat['kyle_lambda_20d']      = _kyl.rolling(20).mean()
    feat['kyle_lambda_zscore_60']= calc_zscore(_kyl, 60)
    feat['kyle_lambda_spike']    = (
        _kyl > _kyl.rolling(60).mean() + _kyl.rolling(60).std() * 2
    ).astype(float)
    feat['kyle_lambda_pctrank_252'] = calc_pctrank(_kyl.fillna(0), 252)

    # ── Hasbrouck 정보 비율 proxy (가격 영향 지속성)
    # 5일 가격 반응 / 당일 가격 충격 (1보다 크면 정보 영구 반영)
    _5d_impact = _ret1.rolling(5).sum()
    _1d_impact = _ret1.abs()
    feat['info_ratio_proxy_5d'] = _5d_impact.abs() / (_1d_impact.rolling(5).mean() + 1e-10)
    feat['info_ratio_high_flag'] = (feat['info_ratio_proxy_5d'] > 2.0).astype(float)

    # ── Amihud 비유동성 가속도 (기존 amihud와 이름 다름)
    _amihud_new = _ret1.abs() / _dv.replace(0, np.nan) * 1e6
    feat['amihud_accel_5d'] = (
        _amihud_new.rolling(5).mean() - _amihud_new.rolling(20).mean()
    ) / _amihud_new.rolling(20).std().replace(0, np.nan)
    feat['amihud_sudden_illiquid'] = (feat['amihud_accel_5d'] > 2.0).astype(float)

    # ── 18H. 통화정책 사이클 선행 지표 ──────────────────────────
    _irx = closes.get('^IRX'); _tnx2 = closes.get('^TNX')

    if _irx is not None:
        feat['fed_easing_expectation']  = -(_irx.diff(20))     # 양수 = 금리 인하 기대
        feat['fed_easing_strong']       = (_irx.diff(20) < -0.20).astype(float)
        feat['fed_panic_cut_signal']    = (_irx.diff(5) < -0.15).astype(float)  # 5일 15bp = 긴급인하
        feat['irx_pctrank_252']         = calc_pctrank(_irx, 252)

    if _irx is not None and _tnx2 is not None:
        _yc = _tnx2 - _irx
        feat['yield_curve_raw']         = _yc
        # 수익률 곡선의 2차 도함수 (가팔라지는 속도의 변화)
        feat['yield_curve_2nd_deriv']   = _yc.diff(5) - _yc.diff(10).shift(5)
        feat['yield_curve_steep_fast']  = (
            (_yc.diff(20) > 0.30) & (_yc < 0.50)
        ).astype(float)
        # 역전 해소 직후 = 역사적으로 경기침체 시작 시점
        feat['yield_uninvert_timing']   = (
            (_yc.shift(60) < 0) & (_yc > 0) & (_yc.diff(20) > 0.15)
        ).astype(float)
        # 통화정책 실수 지표 (실질금리가 자연이자율보다 과도하게 높음)
        feat['overtightening_proxy']    = _irx - 2.5   # 중립금리 2.5% 가정
        feat['overtightening_flag']     = (feat['overtightening_proxy'] > 1.5).astype(float)

    # ── 18I. 기업 이익 질 & 스마트머니 내부 신호 ────────────────
    # ── 어닝 퀄리티 프록시: XLK vs 개별 대형주 갭 추적
    # 어닝 발표 후 갭 하락 = "sell the news" 정착 (고점 징후)
    _nvda = closes.get('NVDA')
    if _nvda is not None:
        _nvda_gap = _nvda.pct_change() - _nvda.pct_change().shift(1)
        feat['nvda_gap_direction_20d']     = np.sign(_nvda_gap).rolling(20).sum()
        feat['nvda_sellnews_count_20d']    = (
            (_nvda.pct_change() < -0.02) &
            (_nvda.pct_change(5).shift(1) > 0.03)  # 5일 상승 후 하락
        ).astype(float).rolling(20).sum()
        feat['nvda_gap_trend_zscore']      = calc_zscore(_nvda_gap.fillna(0), 63)

    # NOBL(배당귀족) 신저가 = 퀄리티 주식마저 팔림 (최악 국면 신호)
    _nobl = closes.get('NOBL')
    if _nobl is not None:
        feat['nobl_new_low_60d']        = (_nobl <= _nobl.rolling(60).min() * 1.01).astype(float)
        feat['nobl_new_low_20d']        = (_nobl <= _nobl.rolling(20).min() * 1.01).astype(float)
        feat['nobl_dd_rate_5d']         = _nobl.pct_change(5)
        _spy3 = closes.get('SPY')
        if _spy3 is not None:
            feat['nobl_underperform_severe'] = (
                _nobl.pct_change(10) < _spy3.pct_change(10) - 0.03
            ).astype(float)

    # ── 18J. VIX 구조 심화 ───────────────────────────────────────
    if '^VIX' in closes.columns:
        _vix = closes['^VIX']
        # VIX 연속 상승일 수 (streak 카운터)
        _vup = (_vix.diff() > 0).astype(float)
        _vstreak = _vup.groupby((_vup != _vup.shift()).cumsum()).cumcount() + 1
        feat['vix_consecutive_up_days']  = _vstreak.where(_vup > 0, 0)
        feat['vix_streak_5_plus']        = (feat['vix_consecutive_up_days'] >= 5).astype(float)
        # VIX 상승 속도 분위수
        feat['vix_roc_5d_pctrank_252']   = calc_pctrank(_vix.pct_change(5).fillna(0), 252)
        feat['vix_roc_extreme_up']       = (feat['vix_roc_5d_pctrank_252'] > 0.90).astype(float)
        # VIX 조용한 상승 + 주가 버팀 = 가장 위험한 선행 패턴
        feat['vix_silent_creep_5d']      = (
            (_vix.pct_change(5) > 0.08) & (cl.pct_change(5) > -0.005)
        ).astype(float)
        # VIX 스파이크 후 허위 진정 (Dead Cat Bounce 위험)
        feat['vix_false_calm_signal']    = (
            (_vix.shift(10) > 28) & (_vix < 22) & (cl.pct_change(5) > 0.02)
        ).astype(float)
        # 실현변동성 대비 내재변동성 갭 (기존 vol_risk_premium과 계산 방식 다름)
        _rv10 = _ret1.rolling(10).std() * np.sqrt(252) * 100
        feat['iv_rv_gap_10d']           = _vix - _rv10
        feat['iv_rv_gap_negative']      = (feat['iv_rv_gap_10d'] < -3).astype(float)  # VIX가 실현변동성보다 낮음 = 과소평가

    # ── 18K. 채권-주식 상관 레짐 전환 ────────────────────────────
    # 이론: 정상(디플레) = 음의 상관 / 인플레 레짐 = 양의 상관
    # 양의 상관으로 전환 = 둘 다 하락하는 스태그플레이션 위험
    _tlt2 = closes.get('TLT')
    if _tlt2 is not None:
        _be_corr20 = _ret1.rolling(20).corr(_tlt2.pct_change())
        _be_corr60 = _ret1.rolling(60).corr(_tlt2.pct_change())
        feat['bond_equity_corr_20d']      = _be_corr20
        feat['bond_equity_corr_60d']      = _be_corr60
        feat['bond_equity_positive_corr'] = (_be_corr20 > 0.3).astype(float)
        # 상관 레짐 반전 (음 → 양 = 인플레 레짐 전환)
        feat['bond_equity_regime_flip']   = (
            (_be_corr20 > 0) & (_be_corr20.shift(20) < -0.3)
        ).astype(float)
        feat['bond_equity_corr_zscore']   = calc_zscore(_be_corr20.fillna(0), 252)

    # ── 18L. 섹터 집중도 심화 ────────────────────────────────────
    _sec11 = [s for s in ['XLK','XLF','XLV','XLY','XLP','XLE','XLI','XLB','XLU','XLRE','XLC']
              if s in closes.columns]
    if len(_sec11) >= 5:
        _sdf = pd.DataFrame({s: closes[s] for s in _sec11})
        _sr20 = pd.DataFrame({s: closes[s].pct_change(20) for s in _sec11})

        # Herfindahl 집중도 지수 proxy (수익률 제곱합 / 분산)
        _sq = _sr20.apply(lambda x: x**2)
        _hhi_denom = (_sr20.std(axis=1)**2 * len(_sec11) + 1e-10)
        feat['sector_herfindahl_20d']     = _sq.sum(axis=1) / _hhi_denom
        feat['sector_hhi_zscore_252']     = calc_zscore(feat['sector_herfindahl_20d'].fillna(0), 252)

        # 섹터 60일 신저가 동시 개수
        _snl60 = pd.DataFrame({
            s: (closes[s] <= closes[s].rolling(60).min() * 1.01).astype(float)
            for s in _sec11
        })
        feat['sector_sim_60d_low_count']  = _snl60.sum(axis=1)
        feat['sector_4plus_60d_low']      = (feat['sector_sim_60d_low_count'] >= 4).astype(float)

        # 섹터 수익률 극단 분화 (소수만 오르고 다수 하락 = 취약한 상승)
        _srange20 = _sr20.max(axis=1) - _sr20.min(axis=1)
        feat['sector_bifurcation_zscore'] = calc_zscore(_srange20, 252)
        feat['sector_extreme_bifurcation']= (feat['sector_bifurcation_zscore'] > 2.0).astype(float)

        # XLK만 오르고 나머지 하락 = 집중 위험 (과거 닷컴 패턴)
        _others = [s for s in _sec11 if s != 'XLK']
        if len(_others) >= 4:
            _other_neg = pd.DataFrame({
                s: (closes[s].pct_change(20) < -0.01).astype(float)
                for s in _others
            })
            feat['xlk_only_rising_flag'] = (
                (cl.pct_change(20) > 0.02) &
                (_other_neg.sum(axis=1) >= len(_others) * 0.7)
            ).astype(float)

    # ── 18M. 종합 붕괴 위험 지수 (Crash Risk Composite) ─────────
    # 이 지표 하나로 여러 위험 신호 동시 점등 여부 확인
    _cr = pd.Series(0.0, index=cl.index)
    # 밸류에이션 위험
    _cr += feat.get('dotcom_overshoot_flag',     pd.Series(0.0, index=cl.index))
    _cr += feat.get('valuation_reset_signal',    pd.Series(0.0, index=cl.index))
    _cr += feat.get('graham_overval_flag',        pd.Series(0.0, index=cl.index))
    # 유동성/신용 위험
    _cr += feat.get('fci_tightening',            pd.Series(0.0, index=cl.index))
    _cr += feat.get('repo_stress_flag',          pd.Series(0.0, index=cl.index))
    _cr += feat.get('credit_impulse_negative',   pd.Series(0.0, index=cl.index))
    # 차트 패턴
    _cr += feat.get('death_cross_active',        pd.Series(0.0, index=cl.index))
    _cr += feat.get('head_shoulders_break',      pd.Series(0.0, index=cl.index))
    _cr += feat.get('rising_wedge_flag',         pd.Series(0.0, index=cl.index))
    # 팩터/기관 이탈
    _cr += feat.get('mtum_factor_crash',         pd.Series(0.0, index=cl.index))
    _cr += feat.get('splv_dominates_3d',         pd.Series(0.0, index=cl.index))
    # 글로벌 전이
    _cr += feat.get('global_4asset_drop_flag',   pd.Series(0.0, index=cl.index))
    _cr += feat.get('tariff_shock_trigger',      pd.Series(0.0, index=cl.index))
    # 채권-주식 레짐
    _cr += feat.get('bond_equity_regime_flip',   pd.Series(0.0, index=cl.index))
    # 유동성 미시구조
    _cr += feat.get('kyle_lambda_spike',         pd.Series(0.0, index=cl.index))
    _cr += feat.get('vix_silent_creep_5d',       pd.Series(0.0, index=cl.index))

    feat['crash_risk_composite_v2'] = _cr
    feat['crash_risk_high_v2']      = (_cr >= 5).astype(float)
    feat['crash_risk_extreme_v2']   = (_cr >= 8).astype(float)
    # 5일 연속 누적 (지속적 위험 = 가장 신뢰도 높음)
    feat['crash_risk_persistent_5d']= (_cr.rolling(5).min() >= 3).astype(float)
    # 정리

        # ══════════════════════════════════════════════════════════════
    #  19. 월가·내부자 + 개미 심리 + 투자 대가 공식 (~200개)
    # ══════════════════════════════════════════════════════════════

    # ─────────────────────── 공통 재사용 변수 ─────────────────────
    _ret1_19    = cl.pct_change()
    _vol_avg20  = vo.rolling(20).mean().replace(0, np.nan)
    _vol_ratio  = vo / _vol_avg20                         # 거래량 비율
    _atr14_19, _= calc_atr(hi, lo, cl, 14)
    _rsi14_19   = calc_rsi(cl, 14)
    _from_52h   = cl / cl.rolling(252).max().replace(0, np.nan) - 1   # 52주 고점 대비
    _sma200_19  = cl.rolling(200).mean()
    _vix_19     = closes.get('^VIX')
    _tnx_19     = closes.get('^TNX')
    _hyg_19     = closes.get('HYG')
    _tlt_19     = closes.get('TLT')
    _gld_19     = closes.get('GLD')
    _bil_19     = closes.get('BIL')
    _dbc_19     = closes.get('DBC')

    # ── 19A. 옵션 시장 스마트머니 (Put/Call + IV Skew) ────────────
    if _vix_19 is not None:
        _vix = _vix_19
        _rv20 = _ret1_19.rolling(20).std() * np.sqrt(252) * 100
        _iv_skew = _vix - _rv20                               # IV - RV 갭

        # IV Skew 기울기 (급등 = 풋 수요 폭발)
        feat['iv_skew_slope_5d']         = _iv_skew.diff(5)
        feat['iv_skew_slope_10d']        = _iv_skew.diff(10)
        feat['iv_skew_zscore_60d_v2']    = calc_zscore(_iv_skew.fillna(0), 60)
        feat['iv_skew_pctrank_252_v2']   = calc_pctrank(_iv_skew.fillna(0), 252)

        # 풋 폭발 신호: Skew 급등 + VIX 상승 동시
        feat['put_explosion_signal']     = (
            (_iv_skew.diff(5) > _iv_skew.diff(5).rolling(60).mean().fillna(0) * 1.5) &
            (_vix.pct_change(5) > 0.10)
        ).astype(float)

        # OpEx Week 효과 (3번째 금요일 근처: 변동성 압축 → 폭발 직전)
        _dom = pd.Series(cl.index.day, index=cl.index)
        feat['opex_week_flag']           = ((_dom >= 15) & (_dom <= 21)).astype(float)
        feat['post_opex_vol_expansion']  = (
            feat['opex_week_flag'].shift(3).fillna(0) *
            (_ret1_19.rolling(3).std() /
             _ret1_19.rolling(20).std().replace(0, np.nan)).fillna(1)
        )

        # VIX 콘탱고/백워데이션 (선물 구조 프록시)
        _vix_contango                    = _vix.rolling(5).mean() / _vix.rolling(20).mean().replace(0, np.nan)
        feat['vix_contango_proxy']       = _vix_contango
        feat['vix_backwardation_v2']     = (_vix_contango < 0.95).astype(float)
        feat['vix_term_premium']         = _vix.rolling(20).mean() - _vix.rolling(5).mean()
        feat['vix_term_premium_zscore']  = calc_zscore(feat['vix_term_premium'].fillna(0), 60)

        # 감마 스퀴즈 위험: 좁은 범위 + VIX 낮음
        _vix_norm = (_vix / _vix.rolling(252).quantile(0.90).replace(0, np.nan)).fillna(1)
        feat['gamma_squeeze_risk']       = (
            (1 - _ret1_19.rolling(5).std() / _ret1_19.rolling(60).std().replace(0, np.nan)) *
            (1 - _vix_norm)
        ).clip(0, 1)

        # 내재변동성 급등 (시장 참여자의 하락 보험 구매 폭발)
        feat['iv_spike_3sigma']          = (
            _vix > _vix.rolling(60).mean() + 3 * _vix.rolling(60).std()
        ).astype(float)
        feat['iv_creep_20d']             = (           # 조용한 VIX 상승 (이미 vix_silent_rise_5d 있으나 20일 버전)
            (_vix.pct_change(20) > 0.20) & (cl.pct_change(20) > -0.02)
        ).astype(float)

    # ── 19B. 다크풀·블록 거래 프록시 ────────────────────────────
    # 비정상 거래량 스파이크 → 블록 매도
    feat['block_trade_vol_3x']           = (_vol_ratio > 3.0).astype(float)
    feat['block_trade_vol_5x']           = (_vol_ratio > 5.0).astype(float)
    feat['block_trade_count_20d']        = feat['block_trade_vol_3x'].rolling(20).sum()

    # 숨겨진 기관 매도: 거래량 폭발인데 가격은 안 움직임
    _hidden_sell = (
        (_vol_ratio > 2.0) & (_ret1_19.abs() < 0.003)
    ).astype(float)
    feat['hidden_inst_sell_5d']          = _hidden_sell.rolling(5).sum()
    feat['hidden_inst_sell_10d']         = _hidden_sell.rolling(10).sum()

    # 가격 충격/거래량 비율 역전 (정상: 큰 가격 = 큰 거래량)
    feat['low_impact_high_vol_5d']       = (
        (_vol_ratio > 1.5) & (_ret1_19.abs() < 0.003)
    ).astype(float).rolling(5).sum()
    feat['price_vol_impact_inversion']   = (
        _ret1_19.abs() / (_vol_ratio.replace(0, 1) + 1e-6)
    )
    feat['impact_inversion_zscore_60']   = calc_zscore(
        feat['price_vol_impact_inversion'].fillna(0), 60
    )

    # 장 마감 직전 대형 거래 누적 (다크풀 프록시: 지속적인 묻지마 매도)
    _absorb = (
        (_vol_ratio > 1.8) & (_ret1_19 >= -0.002) & (cl < op)  # 거래량 많은데 소폭 음봉
    ).astype(float)
    feat['institutional_absorption_5d']  = _absorb.rolling(5).sum()
    feat['institutional_absorption_10d'] = _absorb.rolling(10).sum()

    # ── 19C. 내부자 거래 패턴 프록시 ─────────────────────────────
    # 어닝 시즌: 1·4·7·10월 고거래량 + 가격 정체
    _month    = pd.Series(cl.index.month, index=cl.index)
    _earn_ssn = _month.isin([1, 4, 7, 10]).astype(float)
    feat['earnings_season_vol_spike']    = _earn_ssn * (_vol_ratio > 1.3).astype(float)
    feat['pre_earnings_insider_proxy']   = (
        _earn_ssn * (_vol_ratio > 1.3).astype(float) * (_ret1_19.abs() < 0.003).astype(float)
    )

    # 52주 고점 근처 대량 매도 (내부자 전형 청산 패턴)
    _near_52h = (_from_52h > -0.03).astype(float)
    _top_sell  = _near_52h * (_vol_ratio > 1.5).astype(float)
    feat['insider_sell_at_top_10d']      = _top_sell.rolling(10).sum()
    feat['insider_sell_at_top_20d']      = _top_sell.rolling(20).sum()

    # 연속 고거래량 음봉 누적 (분산매도)
    _inst_dist = (_ret1_19 < -0.005) & (_vol_ratio > 1.5)
    feat['insider_distribution_10d']     = _inst_dist.astype(float).rolling(10).sum()
    feat['insider_distribution_20d']     = _inst_dist.astype(float).rolling(20).sum()
    feat['insider_dist_acceleration']    = (
        _inst_dist.astype(float).rolling(10).sum() -
        _inst_dist.astype(float).rolling(20).sum() / 2
    )

    # OBV 고급 다이버전스: 가격 고점 + OBV 추세 하락 (기관 조용히 팔기 시작)
    _obv_19    = (np.sign(cl.diff()) * vo).fillna(0).cumsum()
    _obv_slope = calc_linreg_slope(_obv_19, 10)
    feat['obv_slope_neg_at_top']         = (
        (_near_52h.astype(bool)) & (_obv_slope < 0)
    ).astype(float)
    feat['obv_slope_diverge_20d']        = (
        (cl.pct_change(20) > 0.03) & (_obv_slope < 0)
    ).astype(float)

    # ── 19D. 개미 심리 지표 ────────────────────────────────────
    # FOMO 스코어: 52주 고점 근처 + 거래량 + 20일 상승 + RSI 과매수
    _fomo = pd.Series(0.0, index=cl.index)
    _fomo += _near_52h
    _fomo += (_vol_ratio > 1.5).astype(float)
    _fomo += (cl.pct_change(20) > 0.15).astype(float)
    _fomo += (_rsi14_19 > 65).astype(float)
    if _vix_19 is not None:
        _fomo += (_vix_19 < 15).astype(float)
    feat['retail_fomo_score']            = _fomo

    # FOMO 소진: FOMO 극점인데 더 오르지 않음 → 반전 직전
    feat['retail_fomo_exhaustion']       = (
        (_fomo >= 3) & (cl.pct_change(5).abs() < 0.01)
    ).astype(float)

    # 패닉 스코어
    _panic = pd.Series(0.0, index=cl.index)
    if _vix_19 is not None:
        _panic += (_vix_19 > 25).astype(float)
        _panic += (_vix_19.pct_change(3) > 0.25).astype(float)
    _panic += (_vol_ratio > 2.0).astype(float)
    _panic += (cl.pct_change(3) < -0.04).astype(float)
    feat['retail_panic_score']           = _panic
    feat['retail_panic_extreme']         = (_panic >= 3).astype(float)

    # TQQQ 레버리지 개미 과열 신호
    _tqqq = closes.get('TQQQ')
    _tqqq_ov = ohlcv.get('TQQQ')
    if _tqqq is not None:
        feat['tqqq_ret_5d_v2']           = _tqqq.pct_change(5)
        feat['tqqq_rsi_14']              = calc_rsi(_tqqq, 14)
        feat['tqqq_5d_consecutive_up']   = (
            _ret1_19.rolling(5).apply(           # TQQQ 연속 5일 양봉
                lambda x: 1.0 if (_tqqq.pct_change().reindex(cl.index).iloc[
                    max(0, len(x)-5):len(x)] > 0).all() else 0.0,
                raw=False)
        )
        # 더 단순한 대안
        _tqqq_up5 = (_tqqq.pct_change() > 0).astype(float).rolling(5).sum()
        feat['tqqq_5d_all_up']           = (_tqqq_up5 == 5).astype(float)

        if _tqqq_ov is not None:
            _tv = _tqqq_ov['Volume']
            _tv_ratio = _tv / _tv.rolling(20).mean().replace(0, np.nan)
            feat['tqqq_vol_surge_flag']  = (_tv_ratio > 2.0).astype(float)
            feat['tqqq_vol_surge_5d']    = feat['tqqq_vol_surge_flag'].rolling(5).sum()
            # 레버리지 개미 대거 진입: 거래량 폭발 + 가격 상승
            feat['leveraged_retail_crowding'] = (
                (_tv_ratio > 2.0) & (_tqqq.pct_change(5) > 0.05)
            ).astype(float)
            feat['leveraged_retail_crowd_3d'] = feat['leveraged_retail_crowding'].rolling(3).sum()

            # TQQQ:QQQ 거래량 비율 (레버리지 투기 온도계)
            _qqq_ov = ohlcv.get('QQQ')
            if _qqq_ov is not None:
                _qq_vol = _qqq_ov['Volume']
                feat['tqqq_qqq_vol_ratio']       = _tv / _qq_vol.replace(0, np.nan)
                feat['tqqq_qqq_vol_zscore_60']   = calc_zscore(
                    feat['tqqq_qqq_vol_ratio'].fillna(0), 60
                )

    # ── 19E. 버핏 심화 — 주식 vs 채권 수익률 비교 ───────────────
    if _tnx_19 is not None:
        _bond_yield   = _tnx_19 / 100
        _earn_yield   = cl.rolling(200).mean() / cl.replace(0, np.nan)   # 이익수익률 프록시
        feat['buffett_eq_bond_ratio']    = _earn_yield / _bond_yield.replace(0, np.nan)
        feat['buffett_bonds_attractive'] = (feat['buffett_eq_bond_ratio'] < 1.0).astype(float)
        feat['buffett_eq_bond_zscore']   = calc_zscore(feat['buffett_eq_bond_ratio'].fillna(0), 252)
        feat['buffett_eq_bond_deterior'] = (feat['buffett_eq_bond_ratio'].diff(20) < -0.05).astype(float)

    # 현금 선호 지수 (BIL 강세 = 단기채 수요 = 위험회피 = 버핏식 현금 모드)
    if _bil_19 is not None:
        feat['cash_preference_index']    = _bil_19.pct_change(20) * 1000
        feat['buffett_cash_mode']        = (_bil_19.pct_change(5) > 0.001).astype(float)
        feat['buffett_cash_mode_accel']  = (
            _bil_19.pct_change(5) > _bil_19.pct_change(10).shift(5)
        ).astype(float)

    # ── 19F. 피터 린치 CANSLIM 심화 ──────────────────────────────
    # [I] 기관 후원: 상승일 거래량 비율 지속 개선
    _upvol_20    = (vo * (_ret1_19 > 0).astype(float)).rolling(20).sum()
    _upvol_ratio = _upvol_20 / vo.rolling(20).sum().replace(0, np.nan)
    feat['canslim_inst_buy_20d']         = _upvol_ratio
    feat['canslim_inst_buy_improving']   = (_upvol_ratio > _upvol_ratio.shift(20)).astype(float)
    feat['canslim_inst_buy_deterior']    = (_upvol_ratio < _upvol_ratio.shift(20)).astype(float)

    # [M] 분배일 누적 (O'Neil: 25일 내 5회+ → 시장 압박)
    _dist_oneil = (
        (cl < op) & (_ret1_19 < -0.001) & (vo >= _vol_avg20)
    ).astype(float)
    feat['canslim_dist_days_25d']        = _dist_oneil.rolling(25).sum()
    feat['canslim_sell_signal_5dist']    = (feat['canslim_dist_days_25d'] >= 5).astype(float)
    feat['canslim_market_pressure_3d']   = (feat['canslim_dist_days_25d'] >= 3).astype(float)

    # [L] 선두주자 붕괴: NVDA·MSFT·AAPL 중 2개+ 50일선 하회
    _big3_below50 = pd.Series(0.0, index=cl.index)
    for _bsym in ['NVDA', 'MSFT', 'AAPL']:
        _bs = closes.get(_bsym)
        if _bs is not None:
            _big3_below50 += (_bs < _bs.rolling(50).mean()).astype(float)
    feat['canslim_leaders_below_50ma']   = _big3_below50
    feat['canslim_leaders_break_2plus']  = (_big3_below50 >= 2).astype(float)
    feat['canslim_leaders_all_break']    = (_big3_below50 >= 3).astype(float)

    # [C] 최근 실적 서프라이즈 프록시: 어닝 시즌 직후 큰 양봉 또는 갭업
    feat['canslim_earnings_reaction']    = (
        _earn_ssn * (cl.pct_change() > 0.02) * (_vol_ratio > 1.5)
    ).astype(float)

    # ── 19G. 제시 리버모어 피벗 포인트 계량화 ─────────────────
    _prev_hi20 = hi.rolling(20).max().shift(1)
    _prev_lo20 = lo.rolling(20).min().shift(1)

    # 저항선 반복 시험 (3회+ = 삼중 천장 위험)
    _near_resist = (cl >= _prev_hi20 * 0.99).astype(float)
    feat['livermore_resist_test_60d']    = _near_resist.rolling(60).sum()
    feat['livermore_triple_top']         = (feat['livermore_resist_test_60d'] >= 3).astype(float)

    # 지지선 붕괴 후 회복 실패 (리버모어의 가장 강한 하락 신호)
    _supp_break = (cl < _prev_lo20).astype(float)
    feat['livermore_support_break']      = _supp_break
    feat['livermore_failed_recovery']    = (
        _supp_break.shift(3).rolling(5).max().fillna(0) == 1
    ).astype(float) * _supp_break

    # 조정 깊이 + 약한 반등 (10% 조정 후 5% 미만 반등)
    _corr_depth  = (cl.rolling(20).max() - cl) / cl.rolling(20).max().replace(0, np.nan)
    feat['livermore_correction_10pct']   = (_corr_depth > 0.10).astype(float)
    feat['livermore_weak_recovery']      = (
        (_corr_depth.shift(10) > 0.10) & (cl.pct_change(10) < 0.05)
    ).astype(float)

    # 더 낮은 고점 + 더 낮은 저점 연속 (하락 추세 확정)
    _lower_hi20  = (hi.rolling(5).max() < hi.rolling(5).max().shift(20)).astype(float)
    _lower_lo20  = (lo.rolling(5).min() < lo.rolling(5).min().shift(20)).astype(float)
    feat['livermore_downtrend_confirm']  = (_lower_hi20 * _lower_lo20)
    feat['livermore_downtrend_streak']   = feat['livermore_downtrend_confirm'].rolling(20).sum()

    # ── 19H. 하워드 막스 — 사이클 과열 스코어 ─────────────────
    _marks_score = pd.Series(0.0, index=cl.index)
    _marks_score += (cl / _sma200_19.replace(0, np.nan) > 1.15).astype(float)    # 200일선 15%+
    _marks_score += (cl.pct_change(252) > 0.25).astype(float)                    # 1년 25%+ 상승
    if _vix_19 is not None:
        _marks_score += (_vix_19 < 15).astype(float)
    _marks_score += (_vol_ratio < 0.70).astype(float)                            # 거래량 감소 (관심 식음)
    _marks_score += (
        _ret1_19.rolling(20).std() < _ret1_19.rolling(60).std() * 0.70
    ).astype(float)                                                               # 변동성 압축
    _marks_score += (_rsi14_19 > 70).astype(float)
    feat['marks_cycle_overheat_score']   = _marks_score
    feat['marks_cycle_peak_flag']        = (_marks_score >= 4).astype(float)

    # 2층 사고: 가격은 고점인데 모멘텀은 식음 (군중과 반대 신호)
    feat['marks_second_level_diverge']   = (
        (cl >= cl.rolling(60).max() * 0.97) &
        (_rsi14_19 < _rsi14_19.rolling(20).mean().fillna(50))
    ).astype(float)
    feat['marks_complacency_peak']       = (
        (_marks_score >= 3) & feat['marks_second_level_diverge'].astype(bool)
    ).astype(float)

    # 시장 사이클 위치 (0~1, 1에 가까울수록 꼭대기)
    _cycle_pos = pd.Series(0.0, index=cl.index)
    for _cp_p in [60, 120, 252]:
        _cycle_pos += calc_pctrank(cl, _cp_p).fillna(0.5)
    feat['marks_cycle_position_score']   = _cycle_pos / 3  # 0~1

    # ── 19I. 레이 달리오 — 부채사이클 + 올웨더 스트레스 ─────────
    _dalio_debt = pd.Series(0.0, index=cl.index)
    if _hyg_19 is not None:
        _dalio_debt += (_hyg_19.pct_change(20) < -0.03).astype(float)
    if _tnx_19 is not None:
        _dalio_debt += (_tnx_19.diff(20) > 0.30).astype(float)
    _dalio_debt += (cl.pct_change(20) < -0.05).astype(float)
    feat['dalio_debt_cycle_stress']      = _dalio_debt
    feat['dalio_deleveraging_signal']    = (_dalio_debt >= 2).astype(float)

    # 올웨더 스트레스: 주식+채권+원자재 동반 하락 (극단 상황)
    _aw = pd.Series(0.0, index=cl.index)
    _aw += (cl.pct_change(10) < -0.05).astype(float)
    if _tlt_19 is not None:
        _aw += (_tlt_19.pct_change(10) < -0.03).astype(float)
    if _dbc_19 is not None:
        _aw += (_dbc_19.pct_change(10) < -0.03).astype(float)
    feat['dalio_allweather_stress']      = _aw
    feat['dalio_everything_falls']       = (_aw >= 2).astype(float)

    # 달리오의 아름다운 디레버리징 조건 역방향 (추악한 디레버리징 신호)
    # 조건: 성장 하락 + 부채 부담 증가 + 통화 완화 불충분
    _ugly_delev = pd.Series(0.0, index=cl.index)
    _ugly_delev += (cl.pct_change(60) < -0.10).astype(float)
    if _hyg_19 is not None:
        _ugly_delev += (_hyg_19.pct_change(20) < -0.05).astype(float)
    if _vix_19 is not None:
        _ugly_delev += (_vix_19 > 30).astype(float)
    feat['dalio_ugly_deleveraging']      = _ugly_delev
    feat['dalio_ugly_deleverage_flag']   = (_ugly_delev >= 2).astype(float)

    # ── 19J. 마이클 버리 — 신용 선행 갭 ─────────────────────────
    if _hyg_19 is not None:
        _credit_gap = _hyg_19.pct_change(20) - cl.pct_change(20)
        feat['burry_credit_equity_gap_20d']   = _credit_gap
        feat['burry_credit_leads_down_20d']   = (
            (_hyg_19.pct_change(20) < -0.02) & (cl.pct_change(20) > 0)
        ).astype(float)
        feat['burry_credit_gap_zscore']       = calc_zscore(_credit_gap.fillna(0), 252)

        # 신용 가속 하락 (5일이 10일보다 빠름 = 초기 붕괴 신호)
        feat['burry_credit_accel_fall']       = (
            (_hyg_19.pct_change(5) < -0.01) &
            (_hyg_19.pct_change(10) < _hyg_19.pct_change(5) * 2)
        ).astype(float)

        # HYG 5일 연속 약세 + XLK 버팀 (버리의 '모두가 틀릴 때' 신호)
        _hyg_down5 = (_hyg_19.pct_change() < 0).astype(float).rolling(5).sum()
        feat['burry_credit_5d_down_equity_up'] = (
            (_hyg_down5 >= 4) & (cl.pct_change(5) > 0)
        ).astype(float)

    # ── 19K. 폴 튜더 존스 — 200일선 + 5:1 리스크/리워드 ─────────
    _ptj_dist    = cl / _sma200_19.replace(0, np.nan) - 1
    feat['tudor_200ma_distance']         = _ptj_dist
    feat['tudor_above_200ma_flag']       = (_ptj_dist > 0).astype(float)
    feat['tudor_far_above_200ma']        = (_ptj_dist > 0.20).astype(float)  # 20%↑ = 위험 구간
    feat['tudor_200ma_trend_up']         = (
        _sma200_19 > _sma200_19.shift(20)
    ).astype(float)
    feat['tudor_200ma_trend_down']       = (
        _sma200_19 < _sma200_19.shift(20)
    ).astype(float)

    # 베어마켓 진입 (52주 고점 -20%)
    feat['tudor_bear_market_flag']       = (_from_52h < -0.20).astype(float)
    feat['tudor_correction_10pct']       = (_from_52h < -0.10).astype(float)
    feat['tudor_drawdown_accel_5d']      = _from_52h - _from_52h.shift(5)   # 낙폭 가속

    # 5:1 리스크/리워드: 하락 기대 크기 vs 상승 기대 크기
    _big_losses = _ret1_19.rolling(20).apply(
        lambda x: x[x < 0].mean() if (x < 0).any() else 0.0, raw=False
    )
    _big_gains  = _ret1_19.rolling(20).apply(
        lambda x: x[x > 0].mean() if (x > 0).any() else 1e-6, raw=False
    )
    feat['tudor_risk_reward_20d']        = _big_losses.abs() / _big_gains.abs().replace(0, 1e-6)
    feat['tudor_bad_risk_reward']        = (feat['tudor_risk_reward_20d'] > 1.0).astype(float)

    # ── 19L. 존 템플턴 — 낙관론 극점 역발상 ─────────────────────
    _temp_opt = pd.Series(0.0, index=cl.index)
    _temp_opt += (cl / cl.rolling(252).max().replace(0, np.nan) > 0.95).astype(float)
    _temp_opt += (cl.pct_change(60) > 0.20).astype(float)
    if _vix_19 is not None:
        _temp_opt += (_vix_19 < 15).astype(float)
    _temp_opt += (_rsi14_19 > 70).astype(float)
    _temp_opt += (_vol_ratio < 0.80).astype(float)    # 거래량 감소 = 관심 식음
    _temp_opt += (calc_rsi(cl, 9) > 75).astype(float)
    feat['templeton_optimism_peak']      = _temp_opt
    feat['templeton_sell_signal']        = (_temp_opt >= 4).astype(float)

    # 군중 쏠림 (모두 같은 방향 = 역발상 기회)
    feat['templeton_crowded_long']       = (
        (cl.pct_change(20) > 0.08) & (_rsi14_19 > 65) & (_vol_ratio > 1.0)
    ).astype(float)
    feat['templeton_max_pessimism_flag'] = (  # 역방향: 반등 기회
        (cl.pct_change(20) < -0.15) & (_rsi14_19 < 30) & (_vol_ratio > 2.0)
    ).astype(float)

    # ── 19M. 수학 심화 공식 10가지 ─────────────────────────────

    # ─ (1) 시그모이드 변환 RSI (비선형 강조: 극단 구간 신호 확대) ─
    _rsi_c       = (_rsi14_19.fillna(50) - 50) / 10    # 중심화
    feat['sigmoid_rsi_14']              = 1 / (1 + np.exp(-_rsi_c))    # 0~1
    feat['sigmoid_rsi_overbought']      = (feat['sigmoid_rsi_14'] > 0.85).astype(float)
    feat['sigmoid_rsi_oversold']        = (feat['sigmoid_rsi_14'] < 0.15).astype(float)
    feat['sigmoid_rsi_slope_5d']        = feat['sigmoid_rsi_14'].diff(5)
    feat['sigmoid_rsi_peak_5d']         = (
        (feat['sigmoid_rsi_slope_5d'] < 0) &
        (feat['sigmoid_rsi_14'] > 0.75)
    ).astype(float)  # 시그모이드 꺾임 = 조정 시작

    # ─ (2) 로그 거래량 Z-score (극단 스파이크 완화, 분포 정규화) ─
    _log_vol     = np.log1p(vo)
    feat['log_vol_zscore_60d_v2']       = calc_zscore(_log_vol, 60)
    feat['log_vol_mean_rev']            = calc_zscore(_log_vol, 20) - calc_zscore(_log_vol, 60)
    feat['log_vol_pctrank_252_v2']      = calc_pctrank(_log_vol, 252)

    # ─ (3) ATR 정규화 모멘텀 (변동성 조정 후 모멘텀 비교) ─────────
    for _p in [5, 10, 20]:
        _raw_ret     = cl.pct_change(_p)
        _atr_norm    = (_atr14_19 * np.sqrt(_p) / cl).replace(0, np.nan)
        feat[f'atr_adj_mom_{_p}d']      = _raw_ret / _atr_norm
        feat[f'atr_adj_mom_zscore_{_p}d'] = calc_zscore(
            feat[f'atr_adj_mom_{_p}d'].fillna(0), 60
        )
    feat['atr_adj_mom_negative_cross'] = (
        (feat['atr_adj_mom_5d'] < 0) & (feat['atr_adj_mom_20d'] > 0)  # 단기 음전
    ).astype(float)

    # ─ (4) BB %B의 지수이평 다이버전스 (스무스 과매수/과매도) ──────
    _ma20_v  = cl.rolling(20).mean()
    _sd20_v  = cl.rolling(20).std()
    _bbpct   = (cl - (_ma20_v - 2 * _sd20_v)) / (4 * _sd20_v).replace(0, np.nan)
    _bbpct_e = _bbpct.ewm(span=5, adjust=False).mean()
    feat['bb_pct_ema5']                 = _bbpct_e
    feat['bb_pct_ema5_overbought']      = (_bbpct_e > 0.80).astype(float)
    feat['bb_pct_ema5_bearish_div']     = (
        (_bbpct > 0.70) & (_bbpct_e < _bbpct_e.shift(5))  # 원래는 높지만 EMA 꺾임
    ).astype(float)

    # ─ (5) 거래량 가중 모멘텀 VWMO (기관 참여 품질 반영) ───────────
    for _vp in [5, 10, 20]:
        _vwmo = (_ret1_19 * vo).rolling(_vp).sum() / vo.rolling(_vp).sum().replace(0, np.nan)
        feat[f'vwmo_{_vp}d']            = _vwmo
        feat[f'vwmo_{_vp}d_zscore_60']  = calc_zscore(_vwmo.fillna(0), 60)
    feat['vwmo_5_20_crossunder']        = (
        (feat['vwmo_5d'] < feat['vwmo_20d']) &
        (feat['vwmo_5d'].shift(1) >= feat['vwmo_20d'].shift(1))
    ).astype(float)

    # ─ (6) 이동 정보비율 (Information Ratio: 초과수익/추적오차) ─────
    for _irp in [20, 60]:
        _excess = _ret1_19 - _ret1_19.rolling(252).mean()
        _te     = _excess.rolling(_irp).std().replace(0, np.nan)
        _ir     = _excess.rolling(_irp).mean() / _te
        feat[f'info_ratio_moving_{_irp}d']  = _ir
        feat[f'ir_deteriorating_{_irp}d']   = (
            _ir < _ir.shift(_irp // 2) - 0.10
        ).astype(float)
    feat['info_ratio_20_below_60']      = (
        feat['info_ratio_moving_20d'] < feat['info_ratio_moving_60d'] - 0.15
    ).astype(float)

    # ─ (7) 수정 샤프 지수 (Adjusted Sharpe: 왜도·첨도 패널티 포함) ─
    for _asp in [20, 60]:
        _mu   = _ret1_19.rolling(_asp).mean()
        _sig  = _ret1_19.rolling(_asp).std().replace(0, np.nan)
        _sk   = _ret1_19.rolling(_asp).skew().fillna(0)
        _ku   = _ret1_19.rolling(_asp).kurt().fillna(0)
        _sr   = _mu / _sig
        feat[f'modified_sharpe_{_asp}d'] = _sr * (
            1 + (_sk / 6) * _sr.fillna(0) -
            ((_ku - 3) / 24) * (_sr.fillna(0) ** 2)
        )
        feat[f'mod_sharpe_neg_{_asp}d']  = (feat[f'modified_sharpe_{_asp}d'] < 0).astype(float)
        feat[f'mod_sharpe_deterior_{_asp}d'] = (
            feat[f'modified_sharpe_{_asp}d'] < feat[f'modified_sharpe_{_asp}d'].shift(_asp // 2)
        ).astype(float)

    # ─ (8) 조화평균 다기간 모멘텀 (극단값에 강건한 다기간 종합) ────
    _m5  = cl.pct_change(5)
    _m10 = cl.pct_change(10)
    _m20 = cl.pct_change(20)
    _m60 = cl.pct_change(60)
    _denom_5_20  = (_m5.abs() + _m20.abs()).replace(0, np.nan)
    _denom_all   = (_m5.abs() + _m10.abs() + _m20.abs() + _m60.abs()).replace(0, np.nan)
    feat['harmonic_mom_5_20']           = (
        2 * _m5 * _m20 / _denom_5_20 * np.sign(_m5.fillna(0) + _m20.fillna(0))
    )
    feat['harmonic_mom_all_4']          = (
        4 * (_m5 * _m10 * _m20 * _m60).fillna(0) / (_denom_all * 4) *
        np.sign((_m5.fillna(0) + _m10.fillna(0) + _m20.fillna(0) + _m60.fillna(0)))
    )
    feat['harmonic_mom_negative']       = (feat['harmonic_mom_all_4'] < -0.005).astype(float)
    feat['harmonic_mom_all_bear']       = (
        (_m5 < 0) & (_m10 < 0) & (_m20 < 0) & (_m60 < 0)
    ).astype(float)

    # ─ (9) 리스크 패리티 강제 청산 신호 ────────────────────────────
    # 주식+채권 변동성 동시 급등 → 패리티 펀드 강제 청산 → 동반 하락
    _vol_eq = _ret1_19.rolling(20).std() * np.sqrt(252)
    if _tnx_19 is not None:
        _vol_bd = _tnx_19.diff(1).rolling(20).std() * np.sqrt(252)
        feat['risk_parity_stress_ratio'] = _vol_eq / _vol_bd.replace(0, np.nan)
        feat['rp_forced_unwind_signal']  = (
            (_vol_eq > _vol_eq.rolling(60).mean() * 1.5) &
            (_vol_bd > _vol_bd.rolling(60).mean() * 1.5)
        ).astype(float)
        feat['rp_unwind_zscore_60']      = calc_zscore(
            feat['risk_parity_stress_ratio'].fillna(0), 60
        )

    # ─ (10) 거래량 가중 RSI (VRSI) — 기존 RSI보다 기관 자금흐름 반영 ─
    def _calc_vrsi(close, volume, period=14):
        delta    = close.diff()
        _up_v    = volume.where(delta > 0, 0.0)
        _dn_v    = volume.where(delta < 0, 0.0)
        _avg_up  = _up_v.ewm(com=period - 1, adjust=False).mean()
        _avg_dn  = _dn_v.ewm(com=period - 1, adjust=False).mean()
        _rs      = _avg_up / _avg_dn.replace(0, np.nan)
        return 100 - 100 / (1 + _rs)

    feat['vrsi_14']                      = _calc_vrsi(cl, vo, 14)
    feat['vrsi_9']                       = _calc_vrsi(cl, vo, 9)
    feat['vrsi_overbought_70']           = (feat['vrsi_14'] > 70).astype(float)
    feat['vrsi_bearish_div']             = (
        (cl >= cl.rolling(20).max() * 0.99) &
        (feat['vrsi_14'] < feat['vrsi_14'].rolling(20).max().shift(5).fillna(50))
    ).astype(float)
    feat['vrsi_slope_5d']                = feat['vrsi_14'].diff(5)
    # VRSI vs 일반 RSI 괴리: 양수 = 거래량은 하락 지지, 음수 = 거래량이 상승 배신
    feat['vrsi_rsi_gap']                 = _rsi14_19.fillna(50) - feat['vrsi_14'].fillna(50)
    feat['vrsi_rsi_gap_negative']        = (feat['vrsi_rsi_gap'] < -10).astype(float)

    # ─ (11) 람다 손실/이익 비율 (큰 하락일 vs 큰 상승일 비율) ────────
    for _lp in [20, 60]:
        _q20 = _ret1_19.rolling(_lp).quantile(0.20)
        _q80 = _ret1_19.rolling(_lp).quantile(0.80)
        _l_mean = _ret1_19.rolling(_lp).apply(
            lambda x, q=None: x[x < np.percentile(x, 20)].mean()
            if (x < np.percentile(x, 20)).any() else 0.0, raw=True
        )
        _g_mean = _ret1_19.rolling(_lp).apply(
            lambda x: x[x > np.percentile(x, 80)].mean()
            if (x > np.percentile(x, 80)).any() else 1e-6, raw=True
        )
        feat[f'lambda_loss_gain_{_lp}d'] = _l_mean.abs() / _g_mean.abs().replace(0, 1e-6)
        feat[f'lambda_ratio_bad_{_lp}d'] = (feat[f'lambda_loss_gain_{_lp}d'] > 1.2).astype(float)

    # ── 19N. 종합 하락 경보 지수 v3 (모든 신규 신호 결합) ────────
    _alert3 = pd.Series(0.0, index=cl.index)

    # 옵션/내부자 신호
    _alert3 += feat.get('put_explosion_signal',      pd.Series(0.0, index=cl.index))
    _alert3 += feat.get('block_trade_count_20d',     pd.Series(0.0, index=cl.index)).clip(0, 1)
    _alert3 += feat.get('insider_distribution_20d',  pd.Series(0.0, index=cl.index)).clip(0, 1)
    _alert3 += feat.get('insider_sell_at_top_10d',   pd.Series(0.0, index=cl.index)).clip(0, 1)
    _alert3 += feat.get('obv_slope_neg_at_top',      pd.Series(0.0, index=cl.index))
    _alert3 += feat.get('hidden_inst_sell_5d',       pd.Series(0.0, index=cl.index)).clip(0, 1)

    # 개미 과열
    _alert3 += feat.get('retail_fomo_exhaustion',    pd.Series(0.0, index=cl.index))
    _alert3 += feat.get('leveraged_retail_crowd_3d', pd.Series(0.0, index=cl.index)).clip(0, 1)
    _alert3 += feat.get('tqqq_5d_all_up',            pd.Series(0.0, index=cl.index))

    # 투자 대가 신호
    _alert3 += feat.get('templeton_sell_signal',     pd.Series(0.0, index=cl.index))
    _alert3 += feat.get('marks_cycle_peak_flag',     pd.Series(0.0, index=cl.index))
    _alert3 += feat.get('canslim_sell_signal_5dist', pd.Series(0.0, index=cl.index))
    _alert3 += feat.get('tudor_far_above_200ma',     pd.Series(0.0, index=cl.index))
    _alert3 += feat.get('livermore_triple_top',      pd.Series(0.0, index=cl.index))
    _alert3 += feat.get('dalio_deleveraging_signal', pd.Series(0.0, index=cl.index))
    _alert3 += feat.get('burry_credit_leads_down_20d', pd.Series(0.0, index=cl.index))

    # 수학 심화 신호
    _alert3 += feat.get('sigmoid_rsi_overbought',    pd.Series(0.0, index=cl.index))
    _alert3 += feat.get('vrsi_bearish_div',          pd.Series(0.0, index=cl.index))
    _alert3 += feat.get('mod_sharpe_neg_20d',        pd.Series(0.0, index=cl.index))
    _alert3 += feat.get('harmonic_mom_all_bear',     pd.Series(0.0, index=cl.index))
    _alert3 += feat.get('vwmo_5_20_crossunder',      pd.Series(0.0, index=cl.index))
    _alert3 += feat.get('rp_forced_unwind_signal',   pd.Series(0.0, index=cl.index))
    _alert3 += feat.get('tudor_bad_risk_reward',     pd.Series(0.0, index=cl.index))
    _alert3 += feat.get('buffett_bonds_attractive',  pd.Series(0.0, index=cl.index))

    feat['drop_alert_comprehensive_v3']  = _alert3
    feat['drop_alert_high_v3']           = (_alert3 >= 6).astype(float)
    feat['drop_alert_extreme_v3']        = (_alert3 >= 10).astype(float)
    feat['drop_alert_rolling5d_v3']      = _alert3.rolling(5).mean()
    feat['drop_alert_persistent_v3']     = (_alert3.rolling(5).min() >= 3).astype(float)
    feat['drop_alert_accel_v3']          = _alert3 - _alert3.shift(5)       # 경보 가속도
    feat['drop_alert_zscore_60_v3']      = calc_zscore(_alert3, 60)

    _max_hist = min(252 * 5, len(cl))   # 최대 5년치 (데이터 부족 대응)

    for _ny, _nd in [(1, 252), (2, 504), (3, 756), (5, _max_hist)]:
        # ① 현재가 / N년 롤링 평균 (버핏 지표 스타일)
        _mean_n = cl.rolling(_nd, min_periods=_nd // 2).mean()
        feat[f'cape_to_mean_{_ny}y']   = cl / _mean_n.replace(0, np.nan)
        # 그 비율이 얼마나 극단적인가 (z-score, 분위수)
        feat[f'cape_to_mean_{_ny}y_zscore252'] = calc_zscore(
            (cl / _mean_n.replace(0, np.nan)).fillna(method='ffill'), 252)
        feat[f'cape_to_mean_{_ny}y_pctrank252'] = calc_pctrank(
            (cl / _mean_n.replace(0, np.nan)).fillna(0), 252)

        # ② 현재가 / N년 롤링 중앙값 (이상치에 강건)
        _med_n = cl.rolling(_nd, min_periods=_nd // 2).median()
        feat[f'cape_to_median_{_ny}y'] = cl / _med_n.replace(0, np.nan)

        # ③ 현재가 / N년 기하평균 (복리 수익률 관점 고평가도)
        #    기하평균 = exp(log(cl)의 롤링 평균)
        _log_mean = np.log(cl.replace(0, np.nan)).rolling(_nd, min_periods=_nd // 2).mean()
        _geo_mean = np.exp(_log_mean)
        feat[f'cape_to_geomean_{_ny}y'] = cl / _geo_mean.replace(0, np.nan)

    # ── 19B. 로그-선형 추세 이탈도 (Trend Deviation CAPE) ───────────
    # 장기 로그선형 추세선 위로 얼마나 올라갔는가
    # 닷컴·2022년처럼 추세선 위로 크게 올라갔을 때 붕괴

    for _nd in [252, 504, 756]:
        _log_cl = np.log(cl.replace(0, np.nan))
        _t_arr  = np.arange(_nd, dtype=float)

        def _log_trend_ratio(x, t=_t_arr):
            """현재가가 N일 로그선형 추세 대비 몇 배인지"""
            if len(x) < _nd // 2 or np.isnan(x).any():
                return np.nan
            slope, intercept = np.polyfit(t[:len(x)], x, 1)
            trend_last = slope * (len(x) - 1) + intercept
            return float(np.exp(x[-1] - trend_last))   # >1 = 추세 위 / <1 = 추세 아래

        feat[f'log_trend_ratio_{_nd}d'] = _log_cl.rolling(_nd, min_periods=_nd // 2).apply(
            _log_trend_ratio, raw=True)
        # 추세 대비 얼마나 극단적인가
        _ltr = feat[f'log_trend_ratio_{_nd}d']
        feat[f'log_trend_ratio_{_nd}d_zscore252'] = calc_zscore(_ltr.fillna(1.0), 252)
        feat[f'log_trend_ratio_{_nd}d_pctrank252'] = calc_pctrank(_ltr.fillna(1.0), 252)

    # ── 19C. 연환산 수익률 과열도 (Return-CAPE) ──────────────────────
    # 현재 N년 연환산 수익률이 역사적으로 얼마나 높은가
    # 수익률이 과거 평균보다 훨씬 높으면 → 평균회귀 하락 위험

    for _ny in [1, 2, 3, 5]:
        _nd = min(252 * _ny, len(cl) - 1)
        if _nd < 50:
            continue
        # N년 연환산 수익률
        _ann_ret = (cl / cl.shift(_nd)) ** (1.0 / _ny) - 1
        feat[f'ann_ret_cagr_{_ny}y']            = _ann_ret
        # 역사적 분위수 (역대 최고 수준이면 → 과열)
        feat[f'ann_ret_cagr_{_ny}y_pctrank252'] = calc_pctrank(_ann_ret.fillna(0), 252)
        feat[f'ann_ret_cagr_{_ny}y_pctrank504'] = calc_pctrank(_ann_ret.fillna(0), 504)
        # 수익률 가속: 단기 수익률이 장기 수익률보다 얼마나 빠른가
        if _ny >= 2:
            _nd_short = min(252, len(cl) - 1)
            _ann_ret_1y = (cl / cl.shift(_nd_short)) ** 1.0 - 1
            feat[f'return_accel_1y_vs_{_ny}y']  = _ann_ret_1y - _ann_ret

    # ── 19D. CAPE 가속도 (버블 팽창 속도) ────────────────────────────
    # CAPE 비율 자체가 얼마나 빠르게 올라가고 있는가
    # 닷컴 직전처럼 CAPE가 가속 상승 → 붕괴 직전 신호

    for _ny, _nd in [(2, 504), (3, 756), (5, _max_hist)]:
        _cape_m = feat.get(f'cape_to_mean_{_ny}y')
        if _cape_m is None:
            continue
        # CAPE 기울기 (20일, 60일)
        feat[f'cape_mean_{_ny}y_slope_20d']  = _cape_m - _cape_m.shift(20)
        feat[f'cape_mean_{_ny}y_slope_60d']  = _cape_m - _cape_m.shift(60)
        # CAPE 가속도 (기울기의 변화)
        slope20 = _cape_m - _cape_m.shift(20)
        feat[f'cape_mean_{_ny}y_accel']      = slope20 - slope20.shift(20)
        # CAPE가 역사적 고점 분위수 돌파
        feat[f'cape_mean_{_ny}y_above90pct'] = (
            _cape_m > _cape_m.rolling(252).quantile(0.90)
        ).astype(float)
        feat[f'cape_mean_{_ny}y_above95pct'] = (
            _cape_m > _cape_m.rolling(252).quantile(0.95)
        ).astype(float)
        # CAPE 정상화 (과열 후 수렴 시작 = 하락 신호)
        feat[f'cape_mean_{_ny}y_cooling']    = (
            (_cape_m.shift(20) > _cape_m.rolling(252).quantile(0.85)) &  # 직전 과열
            (_cape_m < _cape_m.shift(20))                                  # 지금 식고 있음
        ).astype(float)

    # ── 19E. 크로스에셋 상대 고평가도 (XLK vs SPY CAPE 갭) ───────────
    # XLK가 시장 전체보다 얼마나 더 비싼가
    # 기술주가 시장 대비 과도하게 비싸면 → 상대 하락 위험

    _spy2 = closes.get('SPY')
    _qqq2 = closes.get('QQQ')

    if _spy2 is not None:
        for _ny, _nd in [(2, 504), (3, 756), (5, _max_hist)]:
            _xlk_cape = feat.get(f'cape_to_mean_{_ny}y')
            _spy_mean = _spy2.rolling(_nd, min_periods=_nd // 2).mean()
            _spy_cape = _spy2 / _spy_mean.replace(0, np.nan)
            if _xlk_cape is not None:
                # XLK CAPE - SPY CAPE (양수 = XLK가 더 비쌈)
                feat[f'xlk_vs_spy_cape_gap_{_ny}y']   = _xlk_cape - _spy_cape
                # 비율 (1 이상 = XLK가 상대적으로 더 고평가)
                feat[f'xlk_vs_spy_cape_ratio_{_ny}y']  = _xlk_cape / _spy_cape.replace(0, np.nan)
                # 이 갭이 역사적으로 극단적인가
                _gap = _xlk_cape - _spy_cape
                feat[f'xlk_spy_cape_gap_{_ny}y_zscore252'] = calc_zscore(_gap.fillna(0), 252)
                feat[f'xlk_spy_cape_gap_{_ny}y_above90']   = (
                    _gap > _gap.rolling(252).quantile(0.90)
                ).astype(float)

    # ── 19F. 포물선 버블 탐지 (Log-Parabolic Extension) ──────────────
    # 정상적인 상승 = 로그 수익률이 선형(등속)
    # 버블    = 로그 수익률이 가속(포물선) → 붕괴 전조

    _log_cl2 = np.log(cl.replace(0, np.nan))

    # 로그 수익률 비율: 단기 / 중기 (1 이상이면 가속 중)
    for _sh, _lg in [(20, 60), (20, 120), (60, 252), (60, 504)]:
        _r_sh = _log_cl2.diff(_sh)
        _r_lg = _log_cl2.diff(_lg)
        # 단기 연환산 vs 장기 연환산 비율
        _ratio = (_r_sh * (_lg / _sh)) / (_r_lg.replace(0, np.nan))
        feat[f'log_accel_ratio_{_sh}_{_lg}d']        = _ratio
        # 가속 극단 플래그 (상위 90% 이상)
        feat[f'log_accel_extreme_{_sh}_{_lg}d']      = (
            _ratio > _ratio.rolling(252).quantile(0.90)
        ).astype(float)

    # 2차 로그 도함수 (로그 수익률이 빨라지고 있는가)
    _lr20 = _log_cl2.diff(20)    # 20일 로그 수익률
    feat['log_ret_2nd_deriv_20d']  = _lr20 - _lr20.shift(20)     # 가속도
    feat['log_ret_3rd_deriv_20d']  = ((_lr20 - _lr20.shift(20)) -
                                       (_lr20.shift(20) - _lr20.shift(40)))   # 저크
    # 가속도가 역사적으로 극단적인가
    feat['log_ret_accel_pctrank252'] = calc_pctrank(
        feat['log_ret_2nd_deriv_20d'].fillna(0), 252)
    feat['log_ret_accel_above90']    = (
        feat['log_ret_2nd_deriv_20d'] >
        feat['log_ret_2nd_deriv_20d'].rolling(252).quantile(0.90)
    ).astype(float)
    # 가속도가 역전 (포물선 꼭대기 통과)
    feat['log_accel_peak_reversal']  = (
        (feat['log_ret_2nd_deriv_20d'].shift(5) > 0) &
        (feat['log_ret_2nd_deriv_20d'] < 0) &
        (feat['log_ret_2nd_deriv_20d'].shift(5) >
         feat['log_ret_2nd_deriv_20d'].shift(5).rolling(60).quantile(0.70))
    ).astype(float)

    # ── 19G. 거래량 조정 고평가도 (Volume-Adjusted CAPE) ────────────
    # 고평가 + 거래량 감소 = 약한 상승 (기관 비참여) → 붕괴 취약
    # 고평가 + 거래량 급증 = 클라이맥스 매수 → 천장 신호

    for _ny, _nd in [(2, 504), (3, 756)]:
        _cape_m2 = feat.get(f'cape_to_mean_{_ny}y')
        if _cape_m2 is None:
            continue
        # CAPE가 높은데 거래량이 20일 평균 이하 (약한 상승)
        _high_cape2 = (_cape_m2 > _cape_m2.rolling(252).quantile(0.75)).astype(float)
        _low_vol2   = (vo < vo.rolling(20).mean()).astype(float)
        _high_vol2  = (vo > vo.rolling(20).mean() * 1.5).astype(float)
        feat[f'cape_high_vol_low_{_ny}y']   = _high_cape2 * _low_vol2   # 약한 고평가
        feat[f'cape_climax_buy_{_ny}y']     = _high_cape2 * _high_vol2  # 클라이맥스 매수

        # CAPE 변화 vs 거래량 변화 비율
        _cape_chg = _cape_m2.pct_change(20).fillna(0)
        _vol_chg  = vo.rolling(20).mean() / vo.rolling(60).mean().replace(0, np.nan)
        feat[f'cape_vol_diverge_{_ny}y']    = _cape_chg - (_vol_chg - 1)

    # 가격 대비 누적 거래대금 비율 (얼마나 많은 돈이 현재 가격을 만들었는가)
    # 가격은 높은데 최근 거래량은 적으면 → 희박한 수요
    _dollar_vol_20d = (cl * vo).rolling(20).sum()
    _dollar_vol_252d = (cl * vo).rolling(252).sum()
    feat['price_per_dollar_vol_ratio'] = cl / (_dollar_vol_20d / _dollar_vol_252d * cl.mean()).replace(0, np.nan)

    # ── 19H. 복합 고평가 종합 점수 (CAPE Composite Score) ────────────
    # 여러 CAPE 지표를 표준화 후 합산
    # 이 점수가 높을수록 하락 위험 ↑

    _cape_composite = pd.Series(0.0, index=cl.index)
    _cape_cnt = 0

    # 기존 shiller_cape_proxy (원본에 있음)
    if 'shiller_cape_proxy' in feat.columns:
        _sc = feat['shiller_cape_proxy']
        _cape_composite += calc_zscore(_sc.fillna(method='ffill'), 252).fillna(0)
        _cape_cnt += 1

    # 새로 만든 CAPE들
    for _key in [f'cape_to_mean_2y', f'cape_to_mean_3y', f'cape_to_mean_5y',
                  f'cape_to_geomean_3y', f'log_trend_ratio_504d',
                  f'ann_ret_cagr_3y']:
        if _key in feat.columns:
            _v = feat[_key].fillna(method='ffill')
            _v_z = calc_zscore(_v, 252).fillna(0)
            _cape_composite += _v_z
            _cape_cnt += 1

    if _cape_cnt > 0:
        feat['cape_composite_score']       = _cape_composite / _cape_cnt
        feat['cape_composite_pctrank252']  = calc_pctrank(
            feat['cape_composite_score'].fillna(0), 252)
        feat['cape_composite_above90']     = (
            feat['cape_composite_score'] >
            feat['cape_composite_score'].rolling(252).quantile(0.90)
        ).astype(float)
        feat['cape_composite_above95']     = (
            feat['cape_composite_score'] >
            feat['cape_composite_score'].rolling(252).quantile(0.95)
        ).astype(float)
        # 복합 CAPE 기울기 (식어가고 있는가)
        feat['cape_composite_slope_20d']   = (feat['cape_composite_score'] -
                                               feat['cape_composite_score'].shift(20))
        feat['cape_composite_cooling_flag'] = (
            (feat['cape_composite_score'].shift(10) >
             feat['cape_composite_score'].rolling(252).quantile(0.80)) &
            (feat['cape_composite_slope_20d'] < 0)
        ).astype(float)
        # "고평가인데 모멘텀이 죽고 있다" — 가장 강력한 하락 선행 신호
        if 'rsi_14' not in feat.columns:
            _rsi14_cape = calc_rsi(cl, 14)
        else:
            _rsi14_cape = feat['rsi_14']
        feat['cape_high_rsi_falling'] = (
            (feat['cape_composite_score'] >
             feat['cape_composite_score'].rolling(252).quantile(0.75)) &
            (_rsi14_cape < _rsi14_cape.shift(5)) &
            (_rsi14_cape.shift(5) > 60)
        ).astype(float)

# ══════════════════════════════════════════════════════════════
    #  20. 캔들 패턴 + OHLC 수학 공식 상승/하락 지표 (대량 추가, ~95개)
    #      — n일 종가/시가/고가/저가만으로 만드는 계량 지표.
    #        기존 섹션과 중복 없는 새 이름(cdl_/pos_in_range_ 등)만 사용.
    # ══════════════════════════════════════════════════════════════
    _o = op; _h = hi; _l = lo; _c = cl; _v = vo
    _rng     = (_h - _l).replace(0, np.nan)                  # 당일 전체 범위
    _body    = (_c - _o)                                     # 몸통(부호 포함)
    _absbody = _body.abs()
    _uwick   = _h - pd.concat([_c, _o], axis=1).max(axis=1)  # 위꼬리
    _lwick   = pd.concat([_c, _o], axis=1).min(axis=1) - _l  # 아래꼬리
    _pc      = _c.shift(1)                                   # 전일 종가
    _po      = _o.shift(1); _ph = _h.shift(1); _pl = _l.shift(1)
    _pbody   = _body.shift(1)
    _hl2     = (_h + _l) / 2
    _hlc3    = (_h + _l + _c) / 3                            # typical price
    _ohlc4   = (_o + _h + _l + _c) / 4

    # ── 20A. 캔들 구조 비율 (방향성 있는 연속값) ────────────────
    feat['cdl_body_to_range']        = _body / _rng                       # +상승몸통 / -하락몸통
    feat['cdl_upwick_ratio']         = _uwick / _rng                      # 위꼬리 비중(하락압력)
    feat['cdl_lowwick_ratio']        = _lwick / _rng                      # 아래꼬리 비중(상승압력)
    feat['cdl_wick_balance']         = (_lwick - _uwick) / _rng           # +면 매수우위
    feat['cdl_close_loc']            = (_c - _l) / _rng                   # 0(저가)~1(고가) 종가 위치
    feat['cdl_open_loc']             = (_o - _l) / _rng
    feat['cdl_body_center_loc']      = ((_c + _o) / 2 - _l) / _rng        # 몸통 중심 위치
    feat['cdl_close_vs_hlc3']        = _c / _hlc3 - 1
    feat['cdl_close_vs_ohlc4']       = _c / _ohlc4 - 1
    feat['cdl_co_gap_norm']          = (_o - _pc) / _pc.replace(0, np.nan)  # 시가갭
    feat['cdl_wick_asym']            = (_uwick - _lwick) / (_uwick + _lwick).replace(0, np.nan)

    # ── 20B. 단일 캔들 패턴 플래그 (방향 명시) ──────────────────
    _small_body = _absbody <= _rng * 0.1
    feat['cdl_doji']                 = _small_body.astype(float)
    feat['cdl_dragonfly_doji']       = (_small_body & (_uwick <= _rng * 0.1) & (_lwick >= _rng * 0.6)).astype(float)  # 상승
    feat['cdl_gravestone_doji']      = (_small_body & (_lwick <= _rng * 0.1) & (_uwick >= _rng * 0.6)).astype(float)  # 하락
    feat['cdl_hammer']               = ((_lwick >= _absbody * 2) & (_uwick <= _absbody * 0.5) & (_body > 0)).astype(float)  # 상승반전
    feat['cdl_hanging_man']          = ((_lwick >= _absbody * 2) & (_uwick <= _absbody * 0.5) & (_c < _pc)).astype(float)   # 하락반전
    feat['cdl_inverted_hammer']      = ((_uwick >= _absbody * 2) & (_lwick <= _absbody * 0.5) & (_body > 0)).astype(float)
    feat['cdl_shooting_star']        = ((_uwick >= _absbody * 2) & (_lwick <= _absbody * 0.5) & (_body < 0)).astype(float)  # 하락
    feat['cdl_marubozu_bull']        = ((_body > 0) & (_uwick <= _rng * 0.05) & (_lwick <= _rng * 0.05)).astype(float)      # 강상승
    feat['cdl_marubozu_bear']        = ((_body < 0) & (_uwick <= _rng * 0.05) & (_lwick <= _rng * 0.05)).astype(float)      # 강하락
    feat['cdl_spinning_top']         = ((_absbody <= _rng * 0.3) & (_uwick >= _rng * 0.3) & (_lwick >= _rng * 0.3)).astype(float)
    feat['cdl_high_wave']            = ((_absbody <= _rng * 0.2) & (_rng > _rng.rolling(20).mean())).astype(float)
    feat['cdl_belt_hold_bull']       = ((_body > 0) & (_o <= _l * 1.001) & (_body >= _rng * 0.6)).astype(float)
    feat['cdl_belt_hold_bear']       = ((_body < 0) & (_o >= _h * 0.999) & (_absbody >= _rng * 0.6)).astype(float)

    # ── 20C. 2봉 패턴 ───────────────────────────────────────────
    feat['cdl_engulf_bull']          = ((_body > 0) & (_pbody < 0) & (_c >= _po) & (_o <= _pc)).astype(float)
    feat['cdl_engulf_bear']          = ((_body < 0) & (_pbody > 0) & (_o >= _pc) & (_c <= _po)).astype(float)
    feat['cdl_harami_bull']          = ((_pbody < 0) & (_body > 0) & (_h <= _ph) & (_l >= _pl)).astype(float)
    feat['cdl_harami_bear']          = ((_pbody > 0) & (_body < 0) & (_h <= _ph) & (_l >= _pl)).astype(float)
    feat['cdl_piercing']             = ((_pbody < 0) & (_body > 0) & (_o < _pl) & (_c > (_po + _pc) / 2) & (_c < _po)).astype(float)
    feat['cdl_dark_cloud']           = ((_pbody > 0) & (_body < 0) & (_o > _ph) & (_c < (_po + _pc) / 2) & (_c > _po)).astype(float)
    feat['cdl_tweezer_bottom']       = ((_l - _pl).abs() <= _rng * 0.05).astype(float) * (_pbody < 0).astype(float) * (_body > 0).astype(float)
    feat['cdl_tweezer_top']          = ((_h - _ph).abs() <= _rng * 0.05).astype(float) * (_pbody > 0).astype(float) * (_body < 0).astype(float)
    feat['cdl_kicking_bull']         = ((_pbody < 0) & (_body > 0) & (_o > _ph)).astype(float)
    feat['cdl_kicking_bear']         = ((_pbody > 0) & (_body < 0) & (_o < _pl)).astype(float)
    feat['cdl_gap_up_follow']        = ((_o > _ph) & (_body > 0)).astype(float)
    feat['cdl_gap_down_follow']      = ((_o < _pl) & (_body < 0)).astype(float)

    # ── 20D. 3봉 패턴 ───────────────────────────────────────────
    _b1 = _body.shift(2); _b2 = _body.shift(1); _b3 = _body
    _c1 = _c.shift(2); _c2 = _c.shift(1)
    _o1 = _o.shift(2)                                 # 2일 전 시가
    _mid1 = (_c1 + _o1) / 2                            # 2일 전 캔들 몸통 중간
    feat['cdl_morning_star']         = ((_b1 < 0) & (_absbody.shift(1) <= _rng.shift(1) * 0.3) & (_b3 > 0) & (_c > _mid1)).astype(float)
    feat['cdl_evening_star']         = ((_b1 > 0) & (_absbody.shift(1) <= _rng.shift(1) * 0.3) & (_b3 < 0) & (_c < _mid1)).astype(float)
    feat['cdl_three_white_soldiers'] = ((_b1 > 0) & (_b2 > 0) & (_b3 > 0) & (_c > _c2) & (_c2 > _c1)).astype(float)
    feat['cdl_three_black_crows']    = ((_b1 < 0) & (_b2 < 0) & (_b3 < 0) & (_c < _c2) & (_c2 < _c1)).astype(float)
    feat['cdl_three_inside_up']      = ((_b1 < 0) & (_b2 > 0) & (_b3 > 0) & (_c > _c1)).astype(float)
    feat['cdl_three_inside_down']    = ((_b1 > 0) & (_b2 < 0) & (_b3 < 0) & (_c < _c1)).astype(float)
    feat['cdl_3bar_body_dir_sum']    = np.sign(_b1).fillna(0) + np.sign(_b2).fillna(0) + np.sign(_b3).fillna(0)

    # ── 20E. n일 OHLC 수학 통계 (방향성 모멘텀/압력) ───────────
    for p in [3, 5, 10, 20]:
        _hh = _h.rolling(p).max(); _ll = _l.rolling(p).min()
        feat[f'pos_in_range_{p}d']       = (_c - _ll) / (_hh - _ll).replace(0, np.nan)        # 0~1 종가 위치
        feat[f'bull_day_ratio_{p}d']     = (_c > _o).rolling(p).mean()                        # 매수우위 빈도
        feat[f'avg_body_ratio_{p}d']     = (_body / _rng).rolling(p).mean()                   # 추세 강도(부호)
        feat[f'wick_pressure_{p}d']      = (_lwick.rolling(p).sum() - _uwick.rolling(p).sum()) / _c  # 꼬리 압력 누적
        _new_hi = (_h > _h.shift(1)).rolling(p).sum()
        _new_lo = (_l < _l.shift(1)).rolling(p).sum()
        feat[f'hi_lo_break_diff_{p}d']   = (_new_hi - _new_lo) / p                            # 고/저 경신 차
        feat[f'range_norm_mom_{p}d']     = (_c - _c.shift(p)) / _rng.rolling(p).mean().replace(0, np.nan)  # 변동성 정규화 모멘텀
        feat[f'intraday_drive_{p}d']     = _body.rolling(p).sum() / _rng.rolling(p).sum().replace(0, np.nan)  # 장중 추진력

    # ── 20F. 갭 분석 (시가 vs 전일 종가) ────────────────────────
    _gap = (_o - _pc) / _pc.replace(0, np.nan)
    for p in [5, 10, 20]:
        feat[f'gap_up_ratio_{p}d']       = (_gap > 0.002).rolling(p).mean()                   # 상승갭 빈도
        feat[f'gap_down_ratio_{p}d']     = (_gap < -0.002).rolling(p).mean()                  # 하락갭 빈도
        feat[f'net_gap_sum_{p}d']        = _gap.rolling(p).sum()                              # 순 갭 누적
    feat['gap_up_filled_bear']       = ((_gap > 0.003) & (_c < _pc)).astype(float)            # 상승갭 메움(약세)
    feat['gap_down_filled_bull']     = ((_gap < -0.003) & (_c > _pc)).astype(float)           # 하락갭 메움(강세)

    # ── 20G. 진폭/변동성 구조 (OHLC 기반, 방향) ─────────────────
    for p in [5, 10, 20]:
        feat[f'range_expansion_ratio_{p}d'] = _rng / _rng.rolling(p).mean().replace(0, np.nan)  # 변동성 확장/수축
        _intra = (_h - _l) / _c
        _inter = (_c / _pc - 1).abs()
        feat[f'intra_inter_vol_{p}d']    = _intra.rolling(p).mean() / _inter.rolling(p).mean().replace(0, np.nan)  # 장중/종가간 변동
    for p in [10, 20]:
        _net = _c - _c.shift(p)
        _path = (_c - _c.shift(1)).abs().rolling(p).sum()
        feat[f'signed_efficiency_{p}d']  = _net / _path.replace(0, np.nan)                    # 부호 있는 추세효율

    # ── 20H. 캔들 종합 강도 점수 (상승+ / 하락-) ────────────────
    _bull_score = (
        feat['cdl_hammer'] + feat['cdl_marubozu_bull'] + feat['cdl_engulf_bull'] +
        feat['cdl_piercing'] + feat['cdl_morning_star'] + feat['cdl_three_white_soldiers'] +
        feat['cdl_dragonfly_doji'] + feat['cdl_three_inside_up'] + feat['cdl_belt_hold_bull']
    )
    _bear_score = (
        feat['cdl_shooting_star'] + feat['cdl_marubozu_bear'] + feat['cdl_engulf_bear'] +
        feat['cdl_dark_cloud'] + feat['cdl_evening_star'] + feat['cdl_three_black_crows'] +
        feat['cdl_gravestone_doji'] + feat['cdl_three_inside_down'] + feat['cdl_belt_hold_bear'] +
        feat['cdl_hanging_man']
    )
    feat['cdl_bull_pattern_score']   = _bull_score
    feat['cdl_bear_pattern_score']   = _bear_score
    feat['cdl_net_pattern_score']    = _bull_score - _bear_score
    for p in [5, 10]:
        feat[f'cdl_net_score_sum_{p}d'] = (_bull_score - _bear_score).rolling(p).sum()

# ══════════════════════════════════════════════════════════════
    #  21. 캔들/OHLC 수학 공식 지표 — 확장판 (~93개, 섹션20과 중복 없음)
    # ══════════════════════════════════════════════════════════════
    _o = op; _h = hi; _l = lo; _c = cl
    _rng     = (_h - _l).replace(0, np.nan)
    _body    = (_c - _o)
    _absbody = _body.abs()
    _uw      = _h - pd.concat([_c, _o], axis=1).max(axis=1)
    _lw      = pd.concat([_c, _o], axis=1).min(axis=1) - _l
    _pc      = _c.shift(1); _po = _o.shift(1); _ph = _h.shift(1); _pl = _l.shift(1)
    _tp      = (_h + _l + _c) / 3
    _wclose  = (_h + _l + 2 * _c) / 4               # weighted close
    _med     = (_h + _l) / 2

    # ── 21A. Heikin-Ashi (평활 캔들 — 추세 방향 명확) ───────────
    _ha_close = (_o + _h + _l + _c) / 4
    _ha_open  = ((_o.shift(1) + _c.shift(1)) / 2).fillna((_o + _c) / 2)
    _ha_body  = _ha_close - _ha_open
    feat['ha_body_norm']          = _ha_body / _c                       # +상승 / -하락
    feat['ha_body_to_range']      = _ha_body / _rng
    feat['ha_trend_dir']          = np.sign(_ha_body)
    for p in [5, 10, 20]:
        feat[f'ha_bull_ratio_{p}d']   = (_ha_body > 0).rolling(p).mean()    # 평활 상승일 비율
        feat[f'ha_body_mean_{p}d']    = (_ha_body / _c).rolling(p).mean()   # 평활 추세 강도

    # ── 21B. 캔들 모멘텀/연속성 (방향성) ────────────────────────
    _bsign = np.sign(_body)
    _grp   = (_bsign != _bsign.shift()).cumsum()
    _run   = _bsign.groupby(_grp).cumcount() + 1
    feat['body_bull_run']         = _run.where(_bsign > 0, 0)
    feat['body_bear_run']         = _run.where(_bsign < 0, 0)
    feat['body_run_signed']       = _run * _bsign                        # +연속상승 / -연속하락
    feat['body_size_accel_5d']    = (_absbody / _c).rolling(5).mean() - (_absbody / _c).rolling(20).mean()
    feat['higher_high_run_10d']   = (_h > _ph).rolling(10).sum() / 10
    feat['lower_low_run_10d']     = (_l < _pl).rolling(10).sum() / 10
    feat['hh_ll_balance_10d']     = ((_h > _ph).rolling(10).sum() - (_l < _pl).rolling(10).sum()) / 10
    feat['hh_hl_uptrend_10d']     = ((_h > _ph) & (_l > _pl)).rolling(10).mean()
    feat['lh_ll_downtrend_10d']   = ((_h < _ph) & (_l < _pl)).rolling(10).mean()

    # ── 21C. 가격 위치/축 기반 수학 지표 ────────────────────────
    for p in [5, 10, 20, 50]:
        _hh = _h.rolling(p).max(); _ll = _l.rolling(p).min()
        _mid = (_hh + _ll) / 2
        feat[f'close_vs_midrange_{p}d']  = (_c - _mid) / (_hh - _ll).replace(0, np.nan)  # +상단/-하단
        feat[f'tp_mom_{p}d']             = _tp / _tp.shift(p) - 1
    for p in [5, 10, 20]:
        feat[f'wclose_mom_{p}d']         = _wclose / _wclose.shift(p) - 1
    for p in [10, 20]:
        feat[f'medprice_slope_{p}d']     = (_med - _med.shift(p)) / _c

    # ── 21D. 일중 강도/압력 수학 (종가·시가·고저 조합) ──────────
    _clv = ((_c - _l) - (_h - _c)) / _rng                              # -1(저가마감)~+1(고가마감)
    feat['clv']                   = _clv
    for p in [5, 10, 20]:
        feat[f'clv_mean_{p}d']        = _clv.rolling(p).mean()
    feat['intraday_push']         = _body / _rng
    for p in [5, 10]:
        feat[f'intraday_push_mean_{p}d'] = (_body / _rng).rolling(p).mean()
    feat['buy_pressure']          = (_c - _l) / _rng                     # 0~1
    feat['sell_pressure']         = (_h - _c) / _rng                     # 0~1
    feat['net_pressure']          = ((_c - _l) - (_h - _c)) / _rng
    for p in [5, 10, 20]:
        feat[f'net_pressure_sum_{p}d'] = (((_c - _l) - (_h - _c)) / _rng).rolling(p).sum()
    feat['gap_recovery']          = ((_o - _pc) / _pc.replace(0, np.nan)) * -1 + (_body / _c)

    # ── 21E. OHLC 변동성/범위 수학 ──────────────────────────────
    _gk_daily = 0.5 * (np.log(_h / _l) ** 2) - (2 * np.log(2) - 1) * (np.log(_c / _o) ** 2)
    feat['gk_vol_daily']          = _gk_daily
    feat['gk_vol_daily_z60']      = (_gk_daily - _gk_daily.rolling(60).mean()) / _gk_daily.rolling(60).std().replace(0, np.nan)
    _rs_daily = (np.log(_h / _c) * np.log(_h / _o) + np.log(_l / _c) * np.log(_l / _o))
    feat['rs_vol_daily']          = _rs_daily
    for p in [10, 20]:
        feat[f'rs_vol_{p}d']          = _rs_daily.rolling(p).mean()
    feat['co_hl_ratio']           = _absbody / _rng
    for p in [10, 20]:
        feat[f'co_hl_ratio_mean_{p}d'] = (_absbody / _rng).rolling(p).mean()
    _overnight = (_o / _pc - 1)
    _daytime   = (_c / _o - 1)
    for p in [10, 20]:
        feat[f'overnight_vol_{p}d']   = _overnight.rolling(p).std()
        feat[f'daytime_vol_{p}d']     = _daytime.rolling(p).std()
        feat[f'overnight_day_ratio_{p}d'] = (_overnight.rolling(p).std() /
                                              _daytime.rolling(p).std().replace(0, np.nan))
    for p in [10, 20]:
        feat[f'overnight_ret_sum_{p}d'] = _overnight.rolling(p).sum()
        feat[f'daytime_ret_sum_{p}d']   = _daytime.rolling(p).sum()

    # ── 21F. 추가 캔들 패턴 (섹션20에 없는 것) ──────────────────
    _b1 = _body.shift(2); _b2 = _body.shift(1); _b3 = _body
    feat['cdl_rising_three']      = ((_b1 > 0) & (_b2 < 0) & (_b3 > 0) &
                                     (_c > _c.shift(2)) & (_absbody.shift(1) < _absbody.shift(2))).astype(float)
    feat['cdl_falling_three']     = ((_b1 < 0) & (_b2 > 0) & (_b3 < 0) &
                                     (_c < _c.shift(2)) & (_absbody.shift(1) < _absbody.shift(2))).astype(float)
    feat['cdl_stick_sandwich']    = (((_c.shift(2) - _c).abs() <= _rng * 0.05) &
                                     (_b2 > 0) & (_b1 < 0)).astype(float)
    feat['cdl_island_top']        = ((_o.shift(1) > _ph) & (_o < _l.shift(1))).astype(float)
    feat['cdl_island_bottom']     = ((_o.shift(1) < _pl) & (_o > _h.shift(1))).astype(float)
    feat['cdl_pin_bar_bear']      = ((_uw >= _rng * 0.6) & (_absbody <= _rng * 0.25)).astype(float)
    feat['cdl_pin_bar_bull']      = ((_lw >= _rng * 0.6) & (_absbody <= _rng * 0.25)).astype(float)
    feat['cdl_outside_bull']      = ((_h > _ph) & (_l < _pl) & (_body > 0)).astype(float)
    feat['cdl_outside_bear']      = ((_h > _ph) & (_l < _pl) & (_body < 0)).astype(float)
    feat['cdl_inside_bar']        = ((_h <= _ph) & (_l >= _pl)).astype(float)
    _is_doji = (_absbody <= _rng * 0.1)
    feat['cdl_doji_count_5d']     = _is_doji.rolling(5).sum()

    # ── 21G. 프랙탈/스윙 구조 (Bill Williams 프랙탈) ────────────
    _swing_hi = ((_h.shift(2) > _h.shift(4)) & (_h.shift(2) > _h.shift(3)) &
                 (_h.shift(2) > _h.shift(1)) & (_h.shift(2) > _h)).astype(float)
    _swing_lo = ((_l.shift(2) < _l.shift(4)) & (_l.shift(2) < _l.shift(3)) &
                 (_l.shift(2) < _l.shift(1)) & (_l.shift(2) < _l)).astype(float)
    feat['fractal_swing_high']    = _swing_hi
    feat['fractal_swing_low']     = _swing_lo
    feat['fractal_swing_balance_20d'] = (_swing_lo.rolling(20).sum() - _swing_hi.rolling(20).sum())
    feat['bars_since_swing_low_20']  = _swing_lo.rolling(20).apply(
        lambda x: float(len(x) - 1 - np.argmax(x[::-1])) if x.sum() > 0 else 20.0, raw=True)
    feat['bars_since_swing_high_20'] = _swing_hi.rolling(20).apply(
        lambda x: float(len(x) - 1 - np.argmax(x[::-1])) if x.sum() > 0 else 20.0, raw=True)

    # ── 21H. 피보나치/되돌림 수학 (스윙 기반) ───────────────────
    for p in [20, 50]:
        _hh = _h.rolling(p).max(); _ll = _l.rolling(p).min()
        _fib_range = (_hh - _ll).replace(0, np.nan)
        _retr = (_c - _ll) / _fib_range                                  # 0~1 되돌림 위치
        feat[f'fib_retrace_pos_{p}d']    = _retr
        feat[f'near_fib_618_{p}d']       = (np.abs(_retr - 0.618) < 0.03).astype(float)
        feat[f'near_fib_382_{p}d']       = (np.abs(_retr - 0.382) < 0.03).astype(float)
        feat[f'near_fib_500_{p}d']       = (np.abs(_retr - 0.5) < 0.03).astype(float)

    # ── 21I. 캔들 종합 강도 점수 v2 (섹션20 점수와 별개) ────────
    _bull2 = (feat['cdl_rising_three'] + feat['cdl_pin_bar_bull'] +
              feat['cdl_outside_bull'] + feat['cdl_island_bottom'])
    _bear2 = (feat['cdl_falling_three'] + feat['cdl_pin_bar_bear'] +
              feat['cdl_outside_bear'] + feat['cdl_island_top'])
    feat['cdl_v2_bull_score']     = _bull2
    feat['cdl_v2_bear_score']     = _bear2
    feat['cdl_v2_net_score']      = _bull2 - _bear2
    for p in [5, 10]:
        feat[f'cdl_v2_net_sum_{p}d']  = (_bull2 - _bear2).rolling(p).sum()
    feat['ohlc_direction_composite'] = (
        np.sign(_ha_body).fillna(0) +
        np.sign(_clv).fillna(0) +
        np.sign(_body).fillna(0) +
        np.sign((_c - _l) - (_h - _c)).fillna(0)
    )

# ══════════════════════════════════════════════════════════════
    #  22. 캔들/OHLC 수학 공식 — 3차 확장 (~69개, 섹션20·21과 중복 없음)
    # ══════════════════════════════════════════════════════════════
    _o = op; _h = hi; _l = lo; _c = cl; _v = vo
    _rng     = (_h - _l).replace(0, np.nan)
    _body    = (_c - _o)
    _absbody = _body.abs()
    _uw      = _h - pd.concat([_c, _o], axis=1).max(axis=1)
    _lw      = pd.concat([_c, _o], axis=1).min(axis=1) - _l
    _pc      = _c.shift(1); _po = _o.shift(1); _ph = _h.shift(1); _pl = _l.shift(1)
    _ret1    = _c.pct_change()

    # ── 22A. 진봉(true body) 위치·중첩 수학 ─────────────────────
    _today_hi_body = pd.concat([_c, _o], axis=1).max(axis=1)
    _today_lo_body = pd.concat([_c, _o], axis=1).min(axis=1)
    _prev_hi_body  = pd.concat([_pc, _po], axis=1).max(axis=1)
    _prev_lo_body  = pd.concat([_pc, _po], axis=1).min(axis=1)
    _overlap = (pd.concat([_today_hi_body, _prev_hi_body], axis=1).min(axis=1) -
                pd.concat([_today_lo_body, _prev_lo_body], axis=1).max(axis=1))
    _union   = (pd.concat([_today_hi_body, _prev_hi_body], axis=1).max(axis=1) -
                pd.concat([_today_lo_body, _prev_lo_body], axis=1).min(axis=1)).replace(0, np.nan)
    feat['body_overlap_ratio']    = (_overlap / _union).clip(-1, 1)        # 음수=갭/분리
    feat['body_shift_dir']        = np.sign((_c + _o) / 2 - (_pc + _po) / 2)   # 몸통 중심 이동방향
    feat['body_size_ratio_prev']  = _absbody / _absbody.shift(1).replace(0, np.nan)  # 몸통 확대=추세강화
    for p in [5, 10, 20]:
        feat[f'true_body_mom_{p}d']   = (_c / _o - 1).rolling(p).sum()     # 갭 제외 순수 일중추세

    # ── 22B. 그림자(꼬리) 동역학 ────────────────────────────────
    _wick_net = (_lw - _uw) / _rng
    for p in [5, 10, 20]:
        feat[f'wick_net_mean_{p}d']   = _wick_net.rolling(p).mean()         # +매수꼬리우위
        feat[f'wick_net_slope_{p}d']  = _wick_net.rolling(p).mean() - _wick_net.rolling(p).mean().shift(p)
    feat['long_upper_wick_count_10d'] = (_uw > _rng * 0.5).rolling(10).sum()  # 천장 거부
    feat['long_lower_wick_count_10d'] = (_lw > _rng * 0.5).rolling(10).sum()  # 바닥 지지
    feat['wick_dominance_10d']        = ((_lw > _rng * 0.5).rolling(10).sum() -
                                         (_uw > _rng * 0.5).rolling(10).sum())
    feat['both_wick_long_count_10d']  = ((_uw > _rng * 0.35) & (_lw > _rng * 0.35)).rolling(10).sum()

    # ── 22C. 고가-저가 채널 기하 (방향) ─────────────────────────
    for p in [10, 20, 50]:
        _hh = _h.rolling(p).max(); _ll = _l.rolling(p).min()
        _hi_slope = (_hh - _hh.shift(p)) / _c
        _lo_slope = (_ll - _ll.shift(p)) / _c
        feat[f'channel_converge_{p}d']   = _hi_slope - _lo_slope            # <0 수렴 / >0 확대
        feat[f'channel_skew_{p}d']       = (_hi_slope + _lo_slope) / 2      # 채널 전체 방향
        feat[f'channel_mid_dist_{p}d']   = (_c - (_hh + _ll) / 2) / (_hh - _ll).replace(0, np.nan)

    # ── 22D. OHLC 비선형 변환 (로그·차분·교차곱) ────────────────
    feat['log_hl_range']          = np.log(_h / _l)
    for p in [10, 20]:
        feat[f'log_hl_range_mean_{p}d']  = np.log(_h / _l).rolling(p).mean()
        feat[f'log_hl_range_z_{p}d']     = ((np.log(_h / _l) - np.log(_h / _l).rolling(p).mean()) /
                                            np.log(_h / _l).rolling(p).std().replace(0, np.nan))
    feat['log_co_vs_hl']          = np.log(_c / _o) / np.log(_h / _l).replace(0, np.nan)  # 일중 효율
    feat['close_2nd_diff_norm']   = (_c - 2 * _c.shift(1) + _c.shift(2)) / _c             # 종가 가속도
    feat['hl_cross_mom_5d']       = ((_h - _ph) + (_l - _pl)).rolling(5).sum() / _c
    feat['hl_cross_mom_10d']      = ((_h - _ph) + (_l - _pl)).rolling(10).sum() / _c

    # ── 22E. 캔들 일관성/변동 점수 ──────────────────────────────
    for p in [5, 10, 20]:
        feat[f'candle_dir_consistency_{p}d'] = np.sign(_body).rolling(p).sum() / p  # +1전부상승/-1전부하락
    feat['body_dir_flips_10d']    = (np.sign(_body).diff() != 0).rolling(10).sum()  # 지그재그도
    feat['body_dir_flips_20d']    = (np.sign(_body).diff() != 0).rolling(20).sum()
    _cl_loc = (_c - _l) / _rng
    feat['close_loc_trend_10d']   = _cl_loc.rolling(5).mean() - _cl_loc.rolling(20).mean()  # 마감강도 추세

    # ── 22F. 갭 정밀 분석 (시가갭 4분류) ────────────────────────
    _gap = (_o - _pc) / _pc.replace(0, np.nan)
    feat['gap_up_bull_count_20d']   = ((_gap > 0.001) & (_body > 0)).rolling(20).sum()
    feat['gap_up_bear_count_20d']   = ((_gap > 0.001) & (_body < 0)).rolling(20).sum()   # 상승갭 소진
    feat['gap_dn_bull_count_20d']   = ((_gap < -0.001) & (_body > 0)).rolling(20).sum()  # 하락갭 회복
    feat['gap_dn_bear_count_20d']   = ((_gap < -0.001) & (_body < 0)).rolling(20).sum()
    feat['gap_weighted_dir_10d']    = (_gap * np.sign(_body)).rolling(10).sum()          # 갭크기 가중방향
    feat['unfilled_gap_up']         = ((_gap > 0.002) & (_l > _pc)).astype(float)        # 강한 상승추세
    feat['unfilled_gap_down']       = ((_gap < -0.002) & (_h < _pc)).astype(float)       # 강한 하락추세

    # ── 22G. 종가 분포/마감강도 정규화 (OHLC 기반) ──────────────
    _ohlc_mean = (_o + _h + _l + _c) / 4
    feat['close_vs_ohlc_mean']    = (_c - _ohlc_mean) / _rng
    for p in [5, 10, 20]:
        feat[f'close_vs_ohlc_mean_{p}d'] = ((_c - _ohlc_mean) / _rng).rolling(p).mean()
    feat['open_vs_prev_range']    = (_o - _pl) / (_ph - _pl).replace(0, np.nan)          # 시가 갭방향 정밀
    feat['close_in_prev_range']   = (_c - _pl) / (_ph - _pl).replace(0, np.nan)          # 추세 연속성
    for p in [3, 5]:
        feat[f'close_above_open_streak_{p}d'] = (_c > _o).rolling(p).apply(
            lambda x: 1.0 if x.all() else 0.0, raw=True)

    # ── 22H. Pivot 확장 (R2/R3/S2/S3 + Fib + Camarilla) ─────────
    _pivot = (_ph + _pl + _pc) / 3
    _pr    = (_ph - _pl)                                       # 전일 범위
    feat['pivot_r2_dist']         = _c / (_pivot + _pr).replace(0, np.nan) - 1
    feat['pivot_s2_dist']         = _c / (_pivot - _pr).replace(0, np.nan) - 1
    feat['pivot_r3_dist']         = _c / (_ph + 2 * (_pivot - _pl)).replace(0, np.nan) - 1
    feat['pivot_s3_dist']         = _c / (_pl - 2 * (_ph - _pivot)).replace(0, np.nan) - 1
    feat['pivot_fib_r1_dist']     = _c / (_pivot + 0.382 * _pr).replace(0, np.nan) - 1
    feat['pivot_fib_s1_dist']     = _c / (_pivot - 0.382 * _pr).replace(0, np.nan) - 1
    feat['cam_h4_dist']           = _c / (_pc + _pr * 1.1 / 2).replace(0, np.nan) - 1
    feat['cam_l4_dist']           = _c / (_pc - _pr * 1.1 / 2).replace(0, np.nan) - 1
    _above_pivot = (_c > _pivot).astype(float)
    feat['close_above_pivot_streak'] = _above_pivot.groupby(
        (_above_pivot != _above_pivot.shift()).cumsum()).cumcount() + 1

    # ── 22I. 종합 방향 점수 v3 (새 신호 결합) ───────────────────
    _dir3 = (
        np.sign(feat['wick_net_mean_5d']).fillna(0) +
        np.sign(feat['channel_skew_20d']).fillna(0) +
        np.sign(feat['close_vs_ohlc_mean']).fillna(0) +
        np.sign(feat['true_body_mom_5d']).fillna(0) +
        np.sign(feat['gap_weighted_dir_10d']).fillna(0)
    )
    feat['ohlc_dir_score_v3']     = _dir3                              # -5 ~ +5
    feat['ohlc_dir_v3_all_bull']  = (_dir3 >= 4).astype(float)
    feat['ohlc_dir_v3_all_bear']  = (_dir3 <= -4).astype(float)
    for p in [5, 10]:
        feat[f'ohlc_dir_v3_sum_{p}d'] = _dir3.rolling(p).sum()

# ══════════════════════════════════════════════════════════════
    #  23. 캔들/수학 + 하락 정밀 예측 지표 (~58개, 섹션20·21·22와 중복 없음)
    #      접두사: dn_(약세누적) exh_(소진) brk_(붕괴) vd_(변동성하락)
    #              dpc_(약세다이버전스) drop_(하락트리거/확률)
    # ══════════════════════════════════════════════════════════════
    _o = op; _h = hi; _l = lo; _c = cl; _v = vo
    _rng     = (_h - _l).replace(0, np.nan)
    _body    = (_c - _o)
    _absbody = _body.abs()
    _uw      = _h - pd.concat([_c, _o], axis=1).max(axis=1)
    _lw      = pd.concat([_c, _o], axis=1).min(axis=1) - _l
    _pc      = _c.shift(1); _po = _o.shift(1); _ph = _h.shift(1); _pl = _l.shift(1)
    _ret1    = _c.pct_change()
    _rsi14   = calc_rsi(_c, 14)

    # ── 23A. 약세 캔들 누적 압력 (하락 빈도·강도 시계열) ────────
    _bear_candle = (_c < _o).astype(float)                              # 음봉
    _strong_bear = ((_c < _o) & (_absbody > _rng * 0.6)).astype(float)  # 강한 음봉
    _weak_close  = ((_c - _l) / _rng < 0.3).astype(float)              # 저가권 마감
    for p in [5, 10, 20]:
        feat[f'dn_bear_candle_ratio_{p}d'] = _bear_candle.rolling(p).mean()
        feat[f'dn_strong_bear_cnt_{p}d']   = _strong_bear.rolling(p).sum()
        feat[f'dn_weak_close_ratio_{p}d']  = _weak_close.rolling(p).mean()
    _up_body = _body.where(_body > 0)
    _dn_body = _body.where(_body < 0).abs()
    for p in [10, 20]:
        _up_mean = _up_body.rolling(p, min_periods=2).mean()
        _dn_mean = _dn_body.rolling(p, min_periods=2).mean()
        feat[f'dn_body_dominance_{p}d'] = _dn_mean / _up_mean.replace(0, np.nan)  # >1 하락우세
    feat['dn_bear_accel_5_20'] = _bear_candle.rolling(5).mean() - _bear_candle.rolling(20).mean()

    # ── 23B. 고점 거부/소진 패턴 (천장 하락 선행) ───────────────
    _new_hi_try = (_h >= _h.rolling(20).max() - 1e-9)
    _rejected   = _new_hi_try & (_uw > _absbody) & (_c < _o)
    feat['exh_high_rejection_10d']  = _rejected.astype(float).rolling(10).sum()
    feat['exh_high_rejection_20d']  = _rejected.astype(float).rolling(20).sum()
    _shrinking_up = ((_body > 0) & (_body.shift(1) > 0) & (_body.shift(2) > 0) &
                     (_absbody < _absbody.shift(1)) & (_absbody.shift(1) < _absbody.shift(2)))
    feat['exh_rising_momentum_fade'] = _shrinking_up.astype(float)        # 상승 소진
    feat['exh_rising_fade_10d']      = _shrinking_up.astype(float).rolling(10).sum()
    _up_no_vol = ((_body > 0) & (_v < _v.rolling(20).mean() * 0.8)).astype(float)
    feat['exh_up_without_volume_10d'] = _up_no_vol.rolling(10).sum()      # 거래량 없는 상승(가짜)
    feat['exh_close_far_from_high_10d'] = ((_h - _c) / _rng > 0.5).rolling(10).mean()  # 상단저항

    # ── 23C. 지지 붕괴/하락 가속 (브레이크다운) ─────────────────
    for p in [10, 20, 50]:
        _supp = _l.rolling(p).min().shift(1)
        feat[f'brk_support_break_{p}d'] = ((_c < _supp) & (_pc >= _supp)).astype(float)
        feat[f'brk_below_support_cnt_{p}d'] = (_c < _supp).rolling(10).sum()
    _new_low = (_l < _l.shift(1)).astype(float)
    feat['brk_new_low_streak'] = _new_low.groupby((_new_low != _new_low.shift()).cumsum()).cumcount() + 1
    feat['brk_new_low_streak'] = feat['brk_new_low_streak'].where(_new_low > 0, 0)
    feat['brk_new_low_cnt_10d'] = _new_low.rolling(10).sum()
    _gap = (_o - _pc) / _pc.replace(0, np.nan)
    feat['brk_gap_down_no_recover'] = ((_gap < -0.003) & (_c < _o)).astype(float)
    feat['brk_gap_down_persist_10d'] = ((_gap < -0.003) & (_c < _o)).rolling(10).sum()
    for p in [5, 10]:
        feat[f'brk_close_under_low_{p}d'] = (_c < _l.rolling(p).min().shift(1)).astype(float)

    # ── 23D. 변동성 확대 + 하락 동반 (위험 가속) ────────────────
    _range_exp = _rng / _rng.rolling(20).mean()
    feat['vd_range_exp_with_drop'] = ((_range_exp > 1.5) & (_c < _o)).astype(float)
    feat['vd_range_exp_drop_10d']  = ((_range_exp > 1.5) & (_c < _o)).rolling(10).sum()
    feat['vd_drop_efficiency']     = (_c - _h) / _rng                   # 고가 대비 종가(음수=약세)
    for p in [5, 10]:
        feat[f'vd_drop_efficiency_{p}d'] = ((_c - _h) / _rng).rolling(p).mean()
    feat['vd_range_spike_20d']     = (_rng >= _rng.rolling(20).max() - 1e-9).astype(float)
    _dn_rng = _rng.where(_c < _o)
    _up_rng = _rng.where(_c > _o)
    for p in [20]:
        _dn_rng_mean = _dn_rng.rolling(p, min_periods=2).mean()
        _up_rng_mean = _up_rng.rolling(p, min_periods=2).mean()
        feat[f'vd_down_vol_premium_{p}d'] = _dn_rng_mean / _up_rng_mean.replace(0, np.nan)  # >1 하락격렬

    # ── 23E. 약세 다이버전스 정밀 (가격 vs 내부강도) ────────────
    _cl_loc = (_c - _l) / _rng
    _price_hh = (_c >= _c.rolling(20).max() - 1e-9)
    feat['dpc_price_hh_weak_close'] = (_price_hh & (_cl_loc < 0.5)).astype(float)
    feat['dpc_hh_weak_close_20d']   = (_price_hh & (_cl_loc < 0.5)).rolling(20).sum()
    feat['dpc_high_up_close_flat']  = ((_h > _ph) & (_c <= _pc)).astype(float)  # 분배
    feat['dpc_high_up_close_flat_10d'] = ((_h > _ph) & (_c <= _pc)).rolling(10).sum()
    feat['dpc_rsi_bear_div_20d'] = (_price_hh & (_rsi14 < _rsi14.rolling(20).max().shift(3))).astype(float).rolling(20).sum()
    feat['dpc_up_momentum_slowing'] = ((_c > _pc) & ((_c / _pc - 1) < (_pc / _c.shift(2) - 1))).astype(float)

    # ── 23F. 하락 종합 압력 점수 (캔들+구조 결합) ───────────────
    _drop_pressure = (
        (feat['dn_bear_candle_ratio_5d'] > 0.6).astype(float) +
        (feat['exh_high_rejection_10d'] >= 2).astype(float) +
        (feat['brk_below_support_cnt_20d'] >= 3).astype(float) +
        (feat['vd_range_exp_drop_10d'] >= 2).astype(float) +
        (feat['dpc_hh_weak_close_20d'] >= 3).astype(float) +
        (feat['dn_bear_accel_5_20'] > 0.1).astype(float)
    )
    feat['drop_pressure_score']    = _drop_pressure                    # 0~6
    feat['drop_pressure_high']     = (_drop_pressure >= 4).astype(float)
    for p in [3, 5, 10]:
        feat[f'drop_pressure_sum_{p}d'] = _drop_pressure.rolling(p).sum()
    feat['drop_pressure_accel_5d'] = _drop_pressure - _drop_pressure.shift(5)
    feat['drop_pressure_zscore_60'] = calc_zscore(_drop_pressure, 60)

    # ── 23G. 캔들 기반 단기 반전(하락) 트리거 ───────────────────
    feat['drop_overbought_bear_candle'] = ((_rsi14 > 70) & (_c < _o) & (_uw > _absbody)).astype(float)
    _surge3 = (_c.pct_change(3) > 0.05)
    feat['drop_after_surge_first_red'] = (_surge3.shift(1) & (_c < _o)).astype(float)  # 급등후 첫음봉
    feat['drop_consec_upper_wick'] = ((_uw > _absbody) & (_uw.shift(1) > _absbody.shift(1))).astype(float)
    feat['drop_bull_trap_engulf'] = ((_body.shift(1) > _rng.shift(1) * 0.6) &
                                     (_body < 0) & (_c < _po.shift(0).fillna(_o))).astype(float)

    # ── 23H. 누적 하락 확률 프록시 (OHLC 통계 결합) ─────────────
    for p in [10, 20]:
        _neg_body_ratio = (_body < 0).rolling(p).mean()
        _low_close_ratio = (_cl_loc < 0.4).rolling(p).mean()
        feat[f'drop_prob_proxy_{p}d'] = (_neg_body_ratio + _low_close_ratio) / 2
    feat['drop_dual_decline_5d'] = ((_c < _pc) & (_l < _pl)).rolling(5).mean()   # 종가+저가 동시하락
    feat['drop_dual_decline_10d'] = ((_c < _pc) & (_l < _pl)).rolling(10).mean()

    # ══════════════════════════════════════════════════════════════
    #  24. 종목별 하락 정밀 예측 지표 (~88개, 섹션20~23과 중복 없음)
    #      접두사: cr_(크래시) ds_(분배) lq_(유동성) rv_(반전)
    #              tr_(추세붕괴) px_(가격구조) sq_(수급)
    # ══════════════════════════════════════════════════════════════
    _o = op; _h = hi; _l = lo; _c = cl; _v = vo
    _rng     = (_h - _l).replace(0, np.nan)
    _body    = (_c - _o)
    _absbody = _body.abs()
    _uw      = _h - pd.concat([_c, _o], axis=1).max(axis=1)
    _lw      = pd.concat([_c, _o], axis=1).min(axis=1) - _l
    _pc      = _c.shift(1); _po = _o.shift(1); _ph = _h.shift(1); _pl = _l.shift(1)
    _ret1    = _c.pct_change()
    _logret  = np.log(_c / _pc)
    _rsi14   = calc_rsi(_c, 14)
    _atr14, _tr = calc_atr(_h, _l, _c, 14)
    _vma20   = _v.rolling(20).mean().replace(0, np.nan)
    _vratio  = _v / _vma20

    # ── 24A. 크래시 선행 — 꼬리위험/하방 비대칭 (수학) ──────────
    _neg_ret = _ret1.where(_ret1 < 0, 0.0)
    for p in [10, 20, 60]:
        feat[f'cr_downside_dev_{p}d'] = np.sqrt((_neg_ret ** 2).rolling(p).mean())  # 하방 반편차
    _pos_ret = _ret1.where(_ret1 > 0, 0.0)
    for p in [20, 60]:
        _dvar = (_neg_ret ** 2).rolling(p).mean()
        _uvar = (_pos_ret ** 2).rolling(p).mean().replace(0, np.nan)
        feat[f'cr_down_up_var_ratio_{p}d'] = _dvar / _uvar              # >1 하락격렬(크래시취약)
    for p in [20, 60]:
        feat[f'cr_ret_skew_{p}d'] = _logret.rolling(p).skew()           # 음수=하락꼬리위험
    for p in [20, 60]:
        _var5 = _ret1.rolling(p).quantile(0.05)
        feat[f'cr_var5_breach_{p}d'] = (_ret1 <= _var5).astype(float)
        feat[f'cr_var5_breach_cnt_{p}d'] = (_ret1 <= _var5).rolling(p).sum()
    _crash_idx = (-_logret.rolling(20).skew()).fillna(0) + _logret.rolling(20).kurt().fillna(0) * 0.3
    feat['cr_crash_risk_index_20d'] = _crash_idx
    feat['cr_crash_risk_zscore_60'] = calc_zscore(_crash_idx, 60)
    feat['cr_crash_risk_rising_5d'] = (_crash_idx > _crash_idx.shift(5)).astype(float)

    # ── 24B. 분배(Distribution) — 기관 매도 흔적 (수급) ─────────
    _dist_day = ((_ret1 < -0.002) & (_vratio > 1.1)).astype(float)
    for p in [10, 20, 25]:
        feat[f'ds_dist_day_cnt_{p}d'] = _dist_day.rolling(p).sum()
    feat['ds_dist_cluster_5d'] = _dist_day.rolling(5).sum()
    feat['ds_dist_accel_10_25'] = _dist_day.rolling(10).sum() / 10 - _dist_day.rolling(25).sum() / 25
    _up_vol = _v.where(_ret1 > 0, 0.0)
    _dn_vol = _v.where(_ret1 < 0, 0.0)
    for p in [10, 20]:
        feat[f'ds_down_vol_ratio_{p}d'] = (_dn_vol.rolling(p).sum() /
                                           (_up_vol.rolling(p).sum() + _dn_vol.rolling(p).sum()).replace(0, np.nan))
    feat['ds_heavy_weak_close'] = ((_vratio > 1.5) & ((_c - _l) / _rng < 0.4)).astype(float)
    feat['ds_heavy_weak_close_10d'] = feat['ds_heavy_weak_close'].rolling(10).sum()
    _obv = (np.sign(_c.diff()) * _v).fillna(0).cumsum()
    feat['ds_obv_slope_10d'] = (_obv - _obv.shift(10)) / _vma20
    feat['ds_obv_slope_neg'] = (feat['ds_obv_slope_10d'] < 0).astype(float)
    feat['ds_price_up_obv_down_10d'] = ((_c.pct_change(10) > 0.01) & (_obv.diff(10) < 0)).astype(float)
    _mfm = ((_c - _l) - (_h - _c)) / _rng
    _adl = (_mfm * _v).fillna(0).cumsum()
    feat['ds_adl_slope_10d'] = (_adl - _adl.shift(10)) / (_v.rolling(10).sum().replace(0, np.nan))
    feat['ds_adl_falling'] = (feat['ds_adl_slope_10d'] < 0).astype(float)

    # ── 24C. 유동성/충격 (대량매도 취약성) ──────────────────────
    _dollar_vol = (_c * _v).replace(0, np.nan)
    _amihud = _ret1.abs() / _dollar_vol * 1e9
    for p in [5, 20]:
        feat[f'lq_amihud_{p}d'] = _amihud.rolling(p).mean()            # 높을수록 급락 취약
    feat['lq_amihud_zscore_60'] = calc_zscore(_amihud, 60)
    feat['lq_amihud_spike'] = (_amihud > _amihud.rolling(60).mean() + _amihud.rolling(60).std() * 2).astype(float)
    feat['lq_volume_drying_20d'] = (_v.rolling(5).mean() / _v.rolling(20).mean().replace(0, np.nan))
    feat['lq_volume_dry_flag'] = (feat['lq_volume_drying_20d'] < 0.7).astype(float)  # 매수세 고갈
    _impact = _ret1.abs() / _vratio.replace(0, np.nan)
    feat['lq_impact_down_10d'] = _impact.where(_ret1 < 0).rolling(10, min_periods=2).mean()
    feat['lq_impact_up_10d'] = _impact.where(_ret1 > 0).rolling(10, min_periods=2).mean()
    feat['lq_impact_asym_10d'] = (feat['lq_impact_down_10d'] /
                                  feat['lq_impact_up_10d'].replace(0, np.nan))
    _overnight = (_o / _pc - 1)
    feat['lq_overnight_risk_20d'] = _overnight.abs().rolling(20).mean()
    feat['lq_neg_overnight_cnt_20d'] = (_overnight < -0.005).rolling(20).sum()

    # ── 24D. 반전(천장→하락) 트리거 (수학+캔들) ─────────────────
    _ext = (_c - _c.rolling(20).mean()) / _atr14.replace(0, np.nan)
    feat['rv_atr_extension_20d'] = _ext
    feat['rv_overextended_flag'] = (_ext > 2.5).astype(float)            # 과확장=반전위험
    feat['rv_ext_reversal'] = ((_ext.shift(1) > 2.5) & (_ext < _ext.shift(1))).astype(float)
    _accel = _c.pct_change(5) - _c.pct_change(5).shift(5)
    feat['rv_parabolic_accel'] = _accel
    feat['rv_blowoff_top'] = ((_c.pct_change(10) > 0.10) & (_accel < 0) & (_accel.shift(1) > 0)).astype(float)
    feat['rv_rsi_extreme_turn'] = ((_rsi14.shift(1) > 75) & (_rsi14 < _rsi14.shift(1)) & (_c < _o)).astype(float)
    feat['rv_climax_reversal'] = ((_c.pct_change(5) > 0.05) & (_vratio > 2.0) & (_c < _o)).astype(float)
    _bb_up = _c.rolling(20).mean() + 2 * _c.rolling(20).std()
    feat['rv_bb_upper_reject'] = ((_h > _bb_up) & (_c < _bb_up) & (_c < _o)).astype(float)
    feat['rv_bb_upper_reject_10d'] = feat['rv_bb_upper_reject'].rolling(10).sum()

    # ── 24E. 추세 붕괴 (상승추세 종료 감지) ─────────────────────
    _sma20 = _c.rolling(20).mean(); _sma50 = _c.rolling(50).mean()
    feat['tr_break_sma20'] = ((_c < _sma20) & (_pc >= _sma20.shift(1))).astype(float)
    feat['tr_below_sma20_cnt_10d'] = (_c < _sma20).rolling(10).sum()
    feat['tr_sma20_below_sma50'] = (_sma20 < _sma50).astype(float)
    feat['tr_sma20_cross_down'] = ((_sma20 < _sma50) & (_sma20.shift(1) >= _sma50.shift(1))).astype(float)
    _sma20_slope = (_sma20 - _sma20.shift(5)) / _c
    feat['tr_sma20_slope'] = _sma20_slope
    feat['tr_sma20_slope_neg'] = (_sma20_slope < 0).astype(float)
    feat['tr_sma20_slope_turning'] = ((_sma20_slope < 0) & (_sma20_slope.shift(3) > 0)).astype(float)
    _roll_hi5 = _h.rolling(5).max()
    feat['tr_lower_high_form'] = (_roll_hi5 < _roll_hi5.shift(5)).astype(float)
    feat['tr_lower_high_cnt_20d'] = (_roll_hi5 < _roll_hi5.shift(5)).rolling(20).sum()
    _net_move = _c - _c.shift(10)
    _path = (_c.diff().abs()).rolling(10).sum().replace(0, np.nan)
    feat['tr_signed_efficiency_neg_10d'] = ((_net_move / _path) < -0.3).astype(float)

    # ── 24F. 가격 구조 위험 (고점 대비/낙폭 수학) ───────────────
    for p in [20, 60]:
        _peak = _c.rolling(p).max()
        _dd = _c / _peak - 1
        feat[f'px_drawdown_{p}d'] = _dd
        feat[f'px_dd_accel_{p}d'] = _dd - _dd.shift(5)                  # 음수=낙폭심화
    _peak20 = _c.rolling(20).max()
    _days_since_peak = _c.rolling(20).apply(lambda x: float(len(x) - 1 - np.argmax(x)), raw=True)
    feat['px_fast_drop_from_peak'] = ((_c / _peak20 - 1) < -0.03) & (_days_since_peak <= 3)
    feat['px_fast_drop_from_peak'] = feat['px_fast_drop_from_peak'].astype(float)
    _hi252 = _c.rolling(252, min_periods=60).max()
    feat['px_near_52h_bear'] = ((_c > _hi252 * 0.97) & (_c < _o) & (_uw > _absbody)).astype(float)
    feat['px_near_52h_bear_10d'] = feat['px_near_52h_bear'].rolling(10).sum()
    for p in [60, 120]:
        feat[f'px_price_pctrank_{p}d'] = _c.rolling(p).apply(
            lambda x: pd.Series(x).rank(pct=True).iloc[-1], raw=False)
    for p in [20, 60]:
        feat[f'px_close_zscore_{p}d'] = calc_zscore(_c, p)

    # ── 24G. 수급 약화 (매수세 고갈) ────────────────────────────
    _tp = (_h + _l + _c) / 3
    _mf = _tp * _v
    _pos_mf = _mf.where(_tp > _tp.shift(), 0.0).rolling(14).sum()
    _neg_mf = _mf.where(_tp < _tp.shift(), 0.0).rolling(14).sum().replace(0, np.nan)
    _mfi = 100 - 100 / (1 + _pos_mf / _neg_mf)
    feat['sq_mfi_14'] = _mfi
    feat['sq_mfi_overbought_turn'] = ((_mfi.shift(1) > 80) & (_mfi < _mfi.shift(1))).astype(float)
    feat['sq_mfi_falling_5d'] = (_mfi < _mfi.shift(5)).astype(float)
    _cmf = (_mfm * _v).rolling(20).sum() / _v.rolling(20).sum().replace(0, np.nan)
    feat['sq_cmf_20'] = _cmf
    feat['sq_cmf_negative'] = (_cmf < 0).astype(float)
    feat['sq_cmf_turning_neg'] = ((_cmf < 0) & (_cmf.shift(3) > 0)).astype(float)
    _cl_loc = (_c - _l) / _rng
    feat['sq_buy_pressure_trend'] = _cl_loc.rolling(5).mean() - _cl_loc.rolling(20).mean()
    feat['sq_buy_pressure_falling'] = (feat['sq_buy_pressure_trend'] < -0.1).astype(float)
    _force = _c.diff() * _v
    _force_ema = _force.ewm(span=13, adjust=False).mean()
    feat['sq_force_index_neg'] = (_force_ema < 0).astype(float)
    feat['sq_force_index_norm'] = _force_ema / (_c * _vma20)

    # ── 24H. 종합 하락 예측 점수 v4 (모든 신규 신호 결합) ───────
    _crash_score = (
        (feat['cr_down_up_var_ratio_20d'] > 1.3).astype(float) +
        (feat['cr_ret_skew_20d'] < -0.5).astype(float) +
        (feat['ds_dist_day_cnt_20d'] >= 4).astype(float) +
        (feat['ds_price_up_obv_down_10d']) +
        (feat['lq_amihud_spike']) +
        (feat['rv_overextended_flag']) +
        (feat['tr_below_sma20_cnt_10d'] >= 5).astype(float) +
        (feat['tr_sma20_slope_neg']) +
        (feat['px_dd_accel_20d'] < -0.02).astype(float) +
        (feat['sq_cmf_negative']) +
        (feat['sq_mfi_falling_5d'])
    )
    feat['cr_drop_score_v4'] = _crash_score                            # 0~11
    feat['cr_drop_score_v4_high'] = (_crash_score >= 6).astype(float)
    feat['cr_drop_score_v4_extreme'] = (_crash_score >= 8).astype(float)
    for p in [3, 5, 10]:
        feat[f'cr_drop_score_v4_sum_{p}d'] = _crash_score.rolling(p).sum()
    feat['cr_drop_score_v4_accel'] = _crash_score - _crash_score.shift(5)
    feat['cr_drop_score_v4_zscore60'] = calc_zscore(_crash_score, 60)
    feat['cr_drop_score_v4_persist_5d'] = (_crash_score.rolling(5).min() >= 4).astype(float)

# ══════════════════════════════════════════════════════════════
    #  25. 종목별 하락 예측 — 5차 확장 (~72개, 섹션20~24와 중복 없음)
    #      접두사: dq_(추세품질) ac_(가속) dv_(다이버전스) vol_dn_(변동성)
    #              gp_(갭) bd_(MA붕괴) wk_(오실레이터약화) ce_(종합)
    # ══════════════════════════════════════════════════════════════
    _o = op; _h = hi; _l = lo; _c = cl; _v = vo
    _rng     = (_h - _l).replace(0, np.nan)
    _body    = (_c - _o)
    _absbody = _body.abs()
    _uw      = _h - pd.concat([_c, _o], axis=1).max(axis=1)
    _lw      = pd.concat([_c, _o], axis=1).min(axis=1) - _l
    _pc      = _c.shift(1); _po = _o.shift(1); _ph = _h.shift(1); _pl = _l.shift(1)
    _ret1    = _c.pct_change()
    _rsi14   = calc_rsi(_c, 14)
    _atr14, _trv = calc_atr(_h, _l, _c, 14)
    _vma20   = _v.rolling(20).mean().replace(0, np.nan)

    # ── 25A. 추세 품질 악화 (상승의 질 저하 = 하락 선행) ─────────
    for p in [10, 20]:
        _uv = _v.where(_ret1 > 0).rolling(p, min_periods=2).mean()
        _dv = _v.where(_ret1 < 0).rolling(p, min_periods=2).mean()
        feat[f'dq_up_dn_vol_quality_{p}d'] = _uv / _dv.replace(0, np.nan)   # <1 상승에 힘없음
    for p in [10, 20]:
        _ug = _ret1.where(_ret1 > 0).rolling(p, min_periods=2).mean()
        _dg = _ret1.where(_ret1 < 0).abs().rolling(p, min_periods=2).mean()
        feat[f'dq_gain_loss_size_{p}d'] = _ug / _dg.replace(0, np.nan)      # <1 하락이 더 큼
    _up_consist = (_ret1 > 0).rolling(10).mean()
    feat['dq_up_consist_decline'] = _up_consist.shift(10) - _up_consist     # 양수=상승일 줄어듦
    for p in [10, 20]:
        feat[f'dq_body_mean_dir_{p}d'] = (_body / _c).rolling(p).mean()     # 음전=추세 하락전환
    for p in [20]:
        _net = (_c - _c.shift(p)).abs()
        _gross = _c.diff().abs().rolling(p).sum().replace(0, np.nan)
        feat[f'dq_trend_noise_{p}d'] = 1 - (_net / _gross)                 # 1에 가까울수록 노이즈
    _cl_loc = (_c - _l) / _rng
    feat['dq_strong_close_ratio_decline'] = (
        (_cl_loc > 0.7).rolling(10).mean() - (_cl_loc > 0.7).rolling(20).mean())

    # ── 25B. 모멘텀 가속도/꺾임 포착 ────────────────────────────
    for p in [5, 10, 20]:
        _mom = _c.pct_change(p)
        feat[f'ac_mom_{p}d'] = _mom
        feat[f'ac_mom_accel_{p}d'] = _mom - _mom.shift(p)                  # 모멘텀 가속도
    _m10 = _c.pct_change(10)
    _accel10 = _m10 - _m10.shift(10)
    feat['ac_accel_sign_flip_neg'] = ((_accel10 < 0) & (_accel10.shift(5) > 0)).astype(float)
    feat['ac_rsi_velocity_5d'] = _rsi14 - _rsi14.shift(5)
    feat['ac_rsi_decel'] = ((_rsi14 - _rsi14.shift(5)) < (_rsi14.shift(5) - _rsi14.shift(10))).astype(float)
    feat['ac_mom_cross_down'] = ((_c.pct_change(5) < _c.pct_change(20)) &
                                 (_c.pct_change(5).shift(1) >= _c.pct_change(20).shift(1))).astype(float)
    feat['ac_accel_zscore_60'] = calc_zscore(_accel10, 60)
    feat['ac_sharp_decel_flag'] = (feat['ac_accel_zscore_60'] < -1.5).astype(float)

    # ── 25C. 다이버전스 심화 (가격 vs 다중 지표) ────────────────
    _price_hh20 = (_c >= _c.rolling(20).max() - 1e-9)
    _atr_mom = (_c - _c.shift(10)) / _atr14.replace(0, np.nan)
    feat['dv_price_hh_mom_weak'] = (_price_hh20 & (_atr_mom < _atr_mom.rolling(20).max().shift(3))).astype(float)
    feat['dv_price_hh_mom_weak_20d'] = feat['dv_price_hh_mom_weak'].rolling(20).sum()
    feat['dv_hh_vol_decline'] = (_price_hh20 & (_v < _vma20 * 0.8)).astype(float)
    feat['dv_hh_vol_decline_20d'] = feat['dv_hh_vol_decline'].rolling(20).sum()
    feat['dv_failed_breakout_20d'] = ((_h >= _h.rolling(20).max().shift(1)) & (_c < _o)).rolling(20).sum()
    _macd = _c.ewm(span=12, adjust=False).mean() - _c.ewm(span=26, adjust=False).mean()
    _macd_sig = _macd.ewm(span=9, adjust=False).mean()
    _macd_hist = _macd - _macd_sig
    feat['dv_macd_hist_fade'] = ((_macd_hist > 0) & (_macd_hist < _macd_hist.shift(1)) &
                                 (_c.pct_change(5) > 0)).astype(float)
    feat['dv_macd_hist_fade_10d'] = feat['dv_macd_hist_fade'].rolling(10).sum()
    feat['dv_macd_cross_down'] = ((_macd < _macd_sig) & (_macd.shift(1) >= _macd_sig.shift(1))).astype(float)

    # ── 25D. 변동성 레짐 전환 (하락 직전 변동성 패턴) ───────────
    _vol5 = _ret1.rolling(5).std(); _vol20 = _ret1.rolling(20).std()
    feat['vol_dn_compression'] = (_vol5 / _vol20.replace(0, np.nan))
    feat['vol_dn_squeeze_release_bear'] = ((_vol5 / _vol20 > 1.3) &
                                            (_vol5.shift(3) / _vol20.shift(3) < 0.8) &
                                            (_c < _o)).astype(float)
    feat['vol_dn_atr_spike'] = (_atr14 > _atr14.rolling(60).mean() + _atr14.rolling(60).std() * 1.5).astype(float)
    feat['vol_dn_atr_rising_5d'] = (_atr14 > _atr14.shift(5)).astype(float)
    feat['vol_dn_vol_of_vol_20d'] = _vol20.rolling(20).std() / _vol20.rolling(20).mean().replace(0, np.nan)
    _big_down = (_ret1 < -0.02).astype(float)
    feat['vol_dn_shock_count_10d'] = _big_down.rolling(10).sum()
    feat['vol_dn_shock_count_20d'] = _big_down.rolling(20).sum()
    feat['vol_dn_range_trend'] = (_rng / _c).rolling(5).mean() - (_rng / _c).rolling(20).mean()

    # ── 25E. 갭 기반 하락 신호 (오버나이트 위험) ────────────────
    _gap = (_o - _pc) / _pc.replace(0, np.nan)
    feat['gp_down_gap_freq_20d'] = (_gap < -0.003).rolling(20).mean()
    feat['gp_down_gap_freq_rising'] = ((_gap < -0.003).rolling(10).mean() >
                                       (_gap < -0.003).rolling(20).mean()).astype(float)
    feat['gp_gap_up_fade'] = ((_gap > 0.003) & (_c < _o)).astype(float)
    feat['gp_gap_up_fade_10d'] = feat['gp_gap_up_fade'].rolling(10).sum()
    feat['gp_cum_gap_10d'] = _gap.rolling(10).sum()
    feat['gp_cum_gap_neg'] = (feat['gp_cum_gap_10d'] < 0).astype(float)
    feat['gp_big_down_gap'] = (_gap < -0.02).astype(float)
    feat['gp_big_down_gap_20d'] = (_gap < -0.02).rolling(20).sum()
    feat['gp_gap_vol_zscore'] = calc_zscore(_gap.abs(), 60)

    # ── 25F. 다단계 추세 붕괴 확인 (여러 MA 동시) ───────────────
    _sma10 = _c.rolling(10).mean(); _sma20 = _c.rolling(20).mean()
    _sma50 = _c.rolling(50).mean(); _sma100 = _c.rolling(100).mean()
    _below_count = ((_c < _sma10).astype(float) + (_c < _sma20).astype(float) +
                    (_c < _sma50).astype(float) + (_c < _sma100).astype(float))
    feat['bd_below_ma_count'] = _below_count                              # 0~4 (많을수록 약세)
    feat['bd_below_all_ma'] = (_below_count == 4).astype(float)
    feat['bd_below_ma_rising'] = (_below_count > _below_count.shift(5)).astype(float)
    _bullish_align = (_sma10 > _sma20) & (_sma20 > _sma50)
    feat['bd_bullish_align'] = _bullish_align.astype(float)
    feat['bd_align_breakdown'] = (_bullish_align.shift(1) & ~_bullish_align).astype(float)
    feat['bd_all_ma_falling'] = ((_sma10 < _sma10.shift(3)) & (_sma20 < _sma20.shift(3)) &
                                 (_sma50 < _sma50.shift(3))).astype(float)
    feat['bd_ma_break_velocity'] = _below_count - _below_count.shift(10)

    # ── 25G. 내부 강도 약화 (오실레이터 종합) ───────────────────
    _ll14 = _l.rolling(14).min(); _hh14 = _h.rolling(14).max()
    _stoch_k = (_c - _ll14) / (_hh14 - _ll14).replace(0, np.nan) * 100
    feat['wk_stoch_overbought_turn'] = ((_stoch_k.shift(1) > 80) & (_stoch_k < _stoch_k.shift(1))).astype(float)
    feat['wk_stoch_falling'] = (_stoch_k < _stoch_k.shift(3)).astype(float)
    _wr = (_hh14 - _c) / (_hh14 - _ll14).replace(0, np.nan) * -100
    feat['wk_williams_weak'] = ((_wr.shift(1) > -20) & (_wr < _wr.shift(1))).astype(float)
    _tp = (_h + _l + _c) / 3
    _cci = (_tp - _tp.rolling(20).mean()) / (0.015 * _tp.rolling(20).apply(
        lambda x: np.mean(np.abs(x - x.mean())), raw=True).replace(0, np.nan))
    feat['wk_cci_overbought_turn'] = ((_cci.shift(1) > 100) & (_cci < _cci.shift(1))).astype(float)
    feat['wk_cci_negative'] = (_cci < 0).astype(float)
    feat['wk_multi_osc_falling'] = ((_rsi14 < _rsi14.shift(3)) & (_stoch_k < _stoch_k.shift(3)) &
                                    (_cci < _cci.shift(3))).astype(float)
    feat['wk_multi_osc_falling_5d'] = feat['wk_multi_osc_falling'].rolling(5).sum()
    feat['wk_rsi_below_50'] = (_rsi14 < 50).astype(float)
    feat['wk_rsi_cross_50_down'] = ((_rsi14 < 50) & (_rsi14.shift(1) >= 50)).astype(float)

    # ── 25H. 종합 하락 예측 점수 v5 (전 신호 결합) ──────────────
    _ce = (
        (feat['dq_gain_loss_size_20d'] < 0.9).astype(float) +
        (feat['dq_trend_noise_20d'] > 0.7).astype(float) +
        (feat['ac_accel_sign_flip_neg']) +
        (feat['ac_mom_cross_down']) +
        (feat['dv_price_hh_mom_weak_20d'] >= 2).astype(float) +
        (feat['dv_macd_cross_down']) +
        (feat['vol_dn_squeeze_release_bear']) +
        (feat['vol_dn_shock_count_20d'] >= 3).astype(float) +
        (feat['gp_cum_gap_neg']) +
        (feat['bd_below_ma_count'] >= 3).astype(float) +
        (feat['bd_align_breakdown']) +
        (feat['wk_multi_osc_falling']) +
        (feat['wk_rsi_below_50'])
    )
    feat['ce_drop_score_v5'] = _ce                                     # 0~13
    feat['ce_drop_score_v5_high'] = (_ce >= 7).astype(float)
    feat['ce_drop_score_v5_extreme'] = (_ce >= 10).astype(float)
    for p in [3, 5, 10]:
        feat[f'ce_drop_score_v5_sum_{p}d'] = _ce.rolling(p).sum()
    feat['ce_drop_score_v5_accel'] = _ce - _ce.shift(5)
    feat['ce_drop_score_v5_zscore60'] = calc_zscore(_ce, 60)
    feat['ce_drop_score_v5_persist_5d'] = (_ce.rolling(5).min() >= 5).astype(float)
    feat['ce_drop_score_v5_rising'] = (_ce > _ce.shift(3)).astype(float)

# ══════════════════════════════════════════════════════════════
    #  26. 종목별 하락 예측 — 6차 확장 (~74개, 섹션20~25와 중복 없음)
    #      접두사: st_(통계/분포) it_(정보이론) ms_(미시구조)
    #              ts_(시간구조) rg_(레짐) tl_(꼬리위험) cp_(종합)
    # ══════════════════════════════════════════════════════════════
    _o = op; _h = hi; _l = lo; _c = cl; _v = vo
    _rng     = (_h - _l).replace(0, np.nan)
    _body    = (_c - _o)
    _pc      = _c.shift(1); _po = _o.shift(1); _ph = _h.shift(1); _pl = _l.shift(1)
    _ret1    = _c.pct_change()
    _logret  = np.log(_c / _pc)
    _rsi14   = calc_rsi(_c, 14)
    _atr14, _trv = calc_atr(_h, _l, _c, 14)
    _vma20   = _v.rolling(20).mean().replace(0, np.nan)

    # ── 26A. 통계/분포 기반 하락 위험 (수익률 분포 모양) ────────
    for p in [20, 60]:
        feat[f'st_kurtosis_{p}d'] = _logret.rolling(p).kurt()          # 첨도(꼬리 두께)
        feat[f'st_kurt_rising_{p}d'] = (_logret.rolling(p).kurt() >
                                        _logret.rolling(p).kurt().shift(5)).astype(float)
    for p in [20, 60]:
        _neg = _ret1.where(_ret1 < 0, 0.0)
        _dd_dev = np.sqrt((_neg ** 2).rolling(p).mean())
        _tot_dev = _ret1.rolling(p).std().replace(0, np.nan)
        feat[f'st_downside_share_{p}d'] = _dd_dev / _tot_dev           # 하락 변동성 집중도
    for p in [20, 60]:
        feat[f'st_neg_day_excess_{p}d'] = (_ret1 < 0).rolling(p).mean() - 0.5  # 왼쪽 치우침
    for p in [20, 60]:
        feat[f'st_cvar_10_{p}d'] = _ret1.rolling(p).apply(             # 하위10% 평균(ES)
            lambda x: x[x <= np.percentile(x, 10)].mean() if len(x) >= 10 else np.nan, raw=True)
    for p in [20, 60]:
        feat[f'st_mean_median_gap_{p}d'] = (_ret1.rolling(p).mean() - _ret1.rolling(p).median())  # 음수=왼쪽꼬리
    _vol20 = _ret1.rolling(20).std()
    feat['st_vol_trend_20d'] = _vol20 - _vol20.shift(10)
    feat['st_vol_rising_flag'] = (_vol20 > _vol20.shift(10)).astype(float)

    # ── 26B. 정보이론 (예측불가성/무질서도 증가 = 위험) ─────────
    def _entropy(x, bins=8):
        x = x[~np.isnan(x)]
        if len(x) < 10: return np.nan
        cnt, _e = np.histogram(x, bins=bins)
        pr = cnt / cnt.sum()
        pr = pr[pr > 0]
        return float(-np.sum(pr * np.log(pr)))
    for p in [20, 60]:
        feat[f'it_entropy_{p}d'] = _ret1.rolling(p).apply(lambda x: _entropy(x), raw=True)
    feat['it_entropy_rising_20d'] = (feat['it_entropy_20d'] > feat['it_entropy_20d'].shift(5)).astype(float)
    _sign = np.sign(_ret1)
    feat['it_sign_flip_rate_20d'] = (_sign.diff() != 0).rolling(20).mean()  # 부호전환 빈도
    feat['it_sign_flip_rising'] = (feat['it_sign_flip_rate_20d'] >
                                   feat['it_sign_flip_rate_20d'].shift(10)).astype(float)
    def _binary_entropy(x):
        x = x[~np.isnan(x)]
        if len(x) < 10: return np.nan
        b = (x > np.median(x)).astype(int)
        runs = 1 + int(np.sum(np.diff(b) != 0))
        return float(runs / len(x))
    feat['it_pattern_complexity_20d'] = _ret1.rolling(20).apply(_binary_entropy, raw=True)
    feat['it_complexity_spike'] = (feat['it_pattern_complexity_20d'] >
                                   feat['it_pattern_complexity_20d'].rolling(60).mean() +
                                   feat['it_pattern_complexity_20d'].rolling(60).std()).astype(float)
    for p in [20]:
        feat[f'it_autocorr1_{p}d'] = _ret1.rolling(p).apply(
            lambda x: pd.Series(x).autocorr(lag=1) if len(x) > 3 else np.nan, raw=False)
    feat['it_autocorr_turning_neg'] = ((feat['it_autocorr1_20d'] < 0) &
                                       (feat['it_autocorr1_20d'].shift(5) > 0)).astype(float)

    # ── 26C. 미시구조 (체결 압력/스프레드 프록시) ───────────────
    _beta = (np.log(_h / _l) ** 2) + (np.log(_ph / _pl) ** 2)
    _gamma = np.log(pd.concat([_h, _ph], axis=1).max(axis=1) /
                    pd.concat([_l, _pl], axis=1).min(axis=1)) ** 2
    _alpha = ((np.sqrt(2 * _beta) - np.sqrt(_beta)) / (3 - 2 * np.sqrt(2)) -
              np.sqrt(_gamma / (3 - 2 * np.sqrt(2))))
    _cs_spread = 2 * (np.exp(_alpha) - 1) / (1 + np.exp(_alpha))
    feat['ms_cs_spread'] = _cs_spread.clip(lower=0)                    # Corwin-Schultz 스프레드
    feat['ms_cs_spread_zscore_60'] = calc_zscore(_cs_spread, 60)
    feat['ms_cs_spread_widening'] = (_cs_spread > _cs_spread.rolling(20).mean() * 1.5).astype(float)
    feat['ms_intraday_reversal_down'] = ((_h - _o > (_h - _l) * 0.5) & (_c < _o)).astype(float)
    feat['ms_intraday_rev_down_10d'] = feat['ms_intraday_reversal_down'].rolling(10).sum()
    feat['ms_close_below_vwap_proxy'] = (_c < (_h + _l + _c) / 3).astype(float)
    feat['ms_weak_close_freq_10d'] = feat['ms_close_below_vwap_proxy'].rolling(10).mean()
    _cov = _ret1.rolling(20).apply(
        lambda x: np.cov(x[1:], x[:-1])[0, 1] if len(x) > 5 else np.nan, raw=True)
    feat['ms_roll_spread'] = 2 * np.sqrt((-_cov).clip(lower=0))        # Roll 스프레드
    feat['ms_roll_spread_rising'] = (feat['ms_roll_spread'] >
                                     feat['ms_roll_spread'].rolling(20).mean() * 1.3).astype(float)
    _signed_vol = _v * np.sign(_ret1)
    feat['ms_vol_imbalance_10d'] = _signed_vol.rolling(10).sum() / _v.rolling(10).sum().replace(0, np.nan)
    feat['ms_vol_imbalance_neg'] = (feat['ms_vol_imbalance_10d'] < -0.2).astype(float)  # 매도 집중

    # ── 26D. 시간구조 (며칠에 걸친 누적 약세 패턴) ──────────────
    _down_day = (_c < _pc).astype(float)
    feat['ts_down_streak'] = _down_day.groupby((_down_day != _down_day.shift()).cumsum()).cumcount() + 1
    feat['ts_down_streak'] = feat['ts_down_streak'].where(_down_day > 0, 0)
    feat['ts_down_streak_3plus'] = (feat['ts_down_streak'] >= 3).astype(float)
    for p in [5, 10, 20]:
        feat[f'ts_down_density_{p}d'] = _down_day.rolling(p).mean()    # 약세 밀도
    feat['ts_lower_lows_10d'] = (_l.rolling(5).min() < _l.rolling(5).min().shift(5)).astype(float)
    feat['ts_lower_lows_cnt_20d'] = (_l.rolling(5).min() < _l.rolling(5).min().shift(5)).rolling(20).sum()
    _local_high = _h.rolling(5).max()
    feat['ts_failed_rebound'] = ((_c.pct_change(3) > 0) & (_local_high < _local_high.shift(5))).astype(float)
    feat['ts_persistent_weakness'] = ((_c.pct_change(5) < 0) & (_c.pct_change(10) < 0)).astype(float)
    feat['ts_persistent_weak_20d'] = feat['ts_persistent_weakness'].rolling(20).mean()
    _days_since_high = _c.rolling(60).apply(lambda x: float(len(x) - 1 - np.argmax(x)), raw=True)
    feat['ts_days_since_60d_high'] = _days_since_high
    feat['ts_stale_high_flag'] = (_days_since_high > 20).astype(float)  # 고점 갱신 못함

    # ── 26E. 레짐 전환 (상승장→하락장 통계 감지) ────────────────
    for p in [20, 60]:
        _trend_str = (_c - _c.shift(p)).abs() / (_ret1.rolling(p).std() * np.sqrt(p)).replace(0, np.nan)
        feat[f'rg_trend_strength_{p}d'] = _trend_str
    feat['rg_trend_weakening'] = (feat['rg_trend_strength_20d'] < feat['rg_trend_strength_20d'].shift(10)).astype(float)
    _ema20 = _c.ewm(span=20, adjust=False).mean()
    _ema50 = _c.ewm(span=50, adjust=False).mean()
    feat['rg_regime_score'] = np.sign(_c - _ema20) + np.sign(_ema20 - _ema50) + np.sign(_c.pct_change(20))
    feat['rg_regime_bearish'] = (feat['rg_regime_score'] <= -2).astype(float)
    feat['rg_regime_turning_down'] = ((feat['rg_regime_score'] < 0) &
                                      (feat['rg_regime_score'].shift(5) >= 0)).astype(float)
    _vol_now = _ret1.rolling(10).std()
    _vol_base = _ret1.rolling(60).std()
    feat['rg_vol_regime_ratio'] = _vol_now / _vol_base.replace(0, np.nan)
    feat['rg_vol_regime_shift_up'] = (feat['rg_vol_regime_ratio'] > 1.5).astype(float)
    feat['rg_neg_mom_persistence'] = ((_c.pct_change(5) < 0) & (_c.pct_change(10) < 0) &
                                      (_c.pct_change(20) < 0)).astype(float)

    # ── 26F. 꼬리위험 정밀 (극단 손실 확률) ─────────────────────
    for p in [10, 20]:
        feat[f'tl_max_loss_{p}d'] = _ret1.rolling(p).min()            # 최대 단일 손실
    _big_loss = (_ret1 < -0.03).astype(float)
    feat['tl_big_loss_recent_10d'] = _big_loss.rolling(10).sum()
    feat['tl_big_loss_recent_20d'] = _big_loss.rolling(20).sum()
    _loss_only = _ret1.where(_ret1 < 0).abs()
    feat['tl_loss_size_growing'] = (_loss_only.rolling(5, min_periods=2).mean() >
                                    _loss_only.rolling(20, min_periods=2).mean()).astype(float)
    _ret_std60 = _ret1.rolling(60).std()
    feat['tl_3sigma_down'] = (_ret1 < -3 * _ret_std60).astype(float)
    feat['tl_3sigma_down_60d'] = (_ret1 < -3 * _ret_std60).rolling(60).sum()
    feat['tl_loss_clustering'] = (_ret1.abs() > 2 * _ret_std60).rolling(5).sum()
    feat['tl_weak_recovery'] = ((_ret1.shift(1) < -0.02) & (_ret1 < 0.005)).astype(float)

    # ── 26G. 종합 하락 예측 점수 v6 ─────────────────────────────
    _cp = (
        (feat['st_kurt_rising_20d']) +
        (feat['st_cvar_10_20d'] < -0.02).astype(float) +
        (feat['it_entropy_rising_20d']) +
        (feat['it_autocorr_turning_neg']) +
        (feat['ms_cs_spread_widening']) +
        (feat['ms_vol_imbalance_neg']) +
        (feat['ts_down_streak_3plus']) +
        (feat['ts_persistent_weakness']) +
        (feat['rg_regime_bearish']) +
        (feat['rg_vol_regime_shift_up']) +
        (feat['tl_big_loss_recent_20d'] >= 2).astype(float) +
        (feat['tl_loss_size_growing'])
    )
    feat['cp_drop_score_v6'] = _cp                                     # 0~12
    feat['cp_drop_score_v6_high'] = (_cp >= 6).astype(float)
    feat['cp_drop_score_v6_extreme'] = (_cp >= 9).astype(float)
    for p in [3, 5, 10]:
        feat[f'cp_drop_score_v6_sum_{p}d'] = _cp.rolling(p).sum()
    feat['cp_drop_score_v6_accel'] = _cp - _cp.shift(5)
    feat['cp_drop_score_v6_zscore60'] = calc_zscore(_cp, 60)
    feat['cp_drop_score_v6_persist_5d'] = (_cp.rolling(5).min() >= 4).astype(float)
    feat['cp_drop_score_v6_rising'] = (_cp > _cp.shift(3)).astype(float)


# ══════════════════════════════════════════════════════════════
    #  27. 종목별 하락 예측 — 7차 확장 (~67개, 섹션20~26과 중복 없음)
    #      접두사: fr_(프랙탈/카오스) en_(에너지/관성) sp_(스펙트럼)
    #              vp_(거래량-가격) dt_(분포꼬리) cy_(사이클) zz_(종합)
    # ══════════════════════════════════════════════════════════════
    _o = op; _h = hi; _l = lo; _c = cl; _v = vo
    _rng     = (_h - _l).replace(0, np.nan)
    _body    = (_c - _o)
    _pc      = _c.shift(1)
    _ret1    = _c.pct_change()
    _logret  = np.log(_c / _pc)
    _rsi14   = calc_rsi(_c, 14)
    _atr14, _trv = calc_atr(_h, _l, _c, 14)
    _vma20   = _v.rolling(20).mean().replace(0, np.nan)

    # ── 27A. 프랙탈/카오스 (불안정성 증가 = 큰 움직임 임박) ─────
    def _hurst_rs(x):
        x = x[~np.isnan(x)]
        if len(x) < 16: return np.nan
        mean_x = np.mean(x)
        dev = np.cumsum(x - mean_x)
        R = np.max(dev) - np.min(dev)
        S = np.std(x)
        if S < 1e-12: return np.nan
        return float(np.log(R / S + 1e-12) / np.log(len(x)))
    for p in [30, 60]:
        feat[f'fr_hurst_{p}d'] = _ret1.rolling(p).apply(_hurst_rs, raw=True)  # <0.5 평균회귀(추세끝)
    feat['fr_hurst_falling'] = (feat['fr_hurst_30d'] < feat['fr_hurst_30d'].shift(10)).astype(float)
    feat['fr_hurst_below_05'] = (feat['fr_hurst_30d'] < 0.5).astype(float)
    def _lyap(x):
        x = x[~np.isnan(x)]
        if len(x) < 10: return np.nan
        d = np.abs(np.diff(x))
        d = d[d > 0]
        if len(d) < 3: return np.nan
        return float(np.mean(np.log(d + 1e-10)))
    feat['fr_lyapunov_20d'] = _ret1.rolling(20).apply(_lyap, raw=True)     # 카오스/예측불가
    feat['fr_lyapunov_rising'] = (feat['fr_lyapunov_20d'] > feat['fr_lyapunov_20d'].shift(5)).astype(float)
    feat['fr_lyapunov_zscore_60'] = calc_zscore(feat['fr_lyapunov_20d'], 60)
    def _fractal_dim(x):
        x = x[~np.isnan(x)]
        if len(x) < 10: return np.nan
        n = len(x)
        L = np.sum(np.abs(np.diff(x)))
        d = np.max(np.abs(x - x[0]))
        if d < 1e-12 or L < 1e-12: return np.nan
        return float(np.log(n) / (np.log(n) + np.log(d / L)))
    feat['fr_fractal_dim_20d'] = _c.rolling(20).apply(_fractal_dim, raw=True)  # 경로 거칠기
    feat['fr_fractal_dim_rising'] = (feat['fr_fractal_dim_20d'] > feat['fr_fractal_dim_20d'].shift(5)).astype(float)
    feat['fr_path_roughness_20d'] = (_c.diff().abs().rolling(20).sum() /
                                     (_c - _c.shift(20)).abs().replace(0, np.nan))

    # ── 27B. 에너지/관성 물리 모델 (모멘텀 소진) ────────────────
    _ke = 0.5 * (_v / _vma20) * (_ret1 ** 2)
    feat['en_kinetic_energy'] = _ke                                   # 운동에너지
    feat['en_ke_zscore_60'] = calc_zscore(_ke, 60)
    feat['en_ke_dissipating'] = (_ke.rolling(5).mean() < _ke.rolling(20).mean()).astype(float)
    _momentum_phys = _ret1 * (_v / _vma20)
    for p in [5, 10]:
        feat[f'en_momentum_phys_{p}d'] = _momentum_phys.rolling(p).sum()
    feat['en_momentum_decaying'] = (_momentum_phys.rolling(5).sum() < 0).astype(float)
    _pe = (_c - _c.rolling(60).mean()) / _atr14.replace(0, np.nan)
    feat['en_potential_energy'] = _pe                                # 평균회귀 압력(고점=떨어질힘)
    feat['en_pe_high_flag'] = (_pe > 2.0).astype(float)
    feat['en_friction_20d'] = _ret1.rolling(20).std() / (_c.pct_change(20).abs() + 1e-6)
    feat['en_friction_rising'] = (feat['en_friction_20d'] > feat['en_friction_20d'].shift(10)).astype(float)
    _velocity = _c.pct_change(5)
    feat['en_velocity_decay'] = ((_velocity < _velocity.shift(5)) & (_velocity.shift(5) > 0)).astype(float)

    # ── 27C. 스펙트럼/주파수 (사이클 붕괴) ──────────────────────
    def _dom_period(x):
        x = x[~np.isnan(x)]
        if len(x) < 20: return np.nan
        f = np.abs(np.fft.rfft(x - np.mean(x)))
        if len(f) < 2: return np.nan
        idx = np.argmax(f[1:]) + 1
        return float(len(x) / max(idx, 1))
    feat['sp_dom_period_40d'] = _ret1.rolling(40).apply(_dom_period, raw=True)
    feat['sp_period_shortening'] = (feat['sp_dom_period_40d'] > feat['sp_dom_period_40d'].shift(10) * 0.8).astype(float)
    def _hf_energy(x):
        x = x[~np.isnan(x)]
        if len(x) < 20: return np.nan
        f = np.abs(np.fft.rfft(x - np.mean(x))) ** 2
        if f.sum() < 1e-12: return np.nan
        half = len(f) // 2
        return float(f[half:].sum() / f.sum())
    feat['sp_hf_energy_40d'] = _ret1.rolling(40).apply(_hf_energy, raw=True)  # 고주파 비중(불안정)
    feat['sp_hf_energy_rising'] = (feat['sp_hf_energy_40d'] > feat['sp_hf_energy_40d'].shift(10)).astype(float)
    def _spec_entropy(x):
        x = x[~np.isnan(x)]
        if len(x) < 20: return np.nan
        f = np.abs(np.fft.rfft(x - np.mean(x))) ** 2
        if f.sum() < 1e-12: return np.nan
        pr = f / f.sum()
        pr = pr[pr > 0]
        return float(-np.sum(pr * np.log(pr)))
    feat['sp_spectral_entropy_40d'] = _ret1.rolling(40).apply(_spec_entropy, raw=True)
    feat['sp_spec_entropy_rising'] = (feat['sp_spectral_entropy_40d'] >
                                      feat['sp_spectral_entropy_40d'].shift(10)).astype(float)

    # ── 27D. 거래량-가격 동조 붕괴 (수급 이상) ──────────────────
    for p in [10, 20]:
        feat[f'vp_pv_corr_{p}d'] = _c.rolling(p).corr(_v)             # 가격-거래량 상관
    feat['vp_pv_corr_neg'] = (feat['vp_pv_corr_20d'] < -0.2).astype(float)
    _v_trend = _v.rolling(5).mean() / _v.rolling(20).mean()
    feat['vp_price_up_vol_down'] = ((_c.pct_change(10) > 0) & (_v_trend < 0.9)).astype(float)
    feat['vp_price_up_vol_down_20d'] = feat['vp_price_up_vol_down'].rolling(20).sum()
    _vwap20 = (_c * _v).rolling(20).sum() / _v.rolling(20).sum().replace(0, np.nan)
    feat['vp_vwap_dist_20d'] = _c / _vwap20 - 1
    feat['vp_below_vwap'] = (_c < _vwap20).astype(float)
    feat['vp_vwap_cross_down'] = ((_c < _vwap20) & (_pc >= _vwap20.shift(1))).astype(float)
    for p in [10, 20]:
        feat[f'vp_vw_return_{p}d'] = (_ret1 * _v).rolling(p).sum() / _v.rolling(p).sum().replace(0, np.nan)
    feat['vp_vw_return_neg'] = (feat['vp_vw_return_10d'] < 0).astype(float)
    _high_vol_day = (_v > _vma20 * 1.5)
    feat['vp_highvol_down_ratio_20d'] = ((_high_vol_day & (_ret1 < 0)).rolling(20).sum() /
                                         _high_vol_day.rolling(20).sum().replace(0, np.nan))

    # ── 27E. 분포 꼬리 동역학 (극단 위험 변화) ──────────────────
    for p in [20, 60]:
        _skew = _logret.rolling(p).skew()
        feat[f'dt_skew_{p}d'] = _skew
        feat[f'dt_skew_turning_neg_{p}d'] = ((_skew < 0) & (_skew.shift(5) > 0)).astype(float)
    for p in [20, 60]:
        feat[f'dt_tail_asym_{p}d'] = _ret1.rolling(p).apply(
            lambda x: (abs(np.percentile(x, 5)) / (np.percentile(x, 95) + 1e-9))
            if len(x) >= 10 else np.nan, raw=True)
    feat['dt_left_tail_heavy'] = (feat['dt_tail_asym_20d'] > 1.3).astype(float)
    for p in [20]:
        _dn_std = _ret1.where(_ret1 < 0).rolling(p, min_periods=3).std()
        _up_std = _ret1.where(_ret1 > 0).rolling(p, min_periods=3).std()
        feat[f'dt_vol_asym_{p}d'] = _dn_std / _up_std.replace(0, np.nan)
    _ext_dn = (_ret1 < _ret1.rolling(60).quantile(0.05)).rolling(20).sum()
    _ext_up = (_ret1 > _ret1.rolling(60).quantile(0.95)).rolling(20).sum()
    feat['dt_extreme_dn_excess_20d'] = _ext_dn - _ext_up
    for p in [60]:
        _sk = _logret.rolling(p).skew()
        _ku = _logret.rolling(p).kurt()
        feat[f'dt_jarque_bera_{p}d'] = p / 6 * (_sk ** 2 + (_ku ** 2) / 4)
    feat['dt_non_normal_flag'] = (feat['dt_jarque_bera_60d'] > 10).astype(float)

    # ── 27F. 사이클/계절 위치 (고점 사이클 성숙도) ──────────────
    _pctile_sum = pd.Series(0.0, index=_c.index); _cnt = 0
    for p in [60, 120, 250]:
        _pctile_sum = _pctile_sum + _c.rolling(p, min_periods=p // 2).apply(
            lambda x: pd.Series(x).rank(pct=True).iloc[-1], raw=False).fillna(0.5)
        _cnt += 1
    feat['cy_cycle_position'] = _pctile_sum / _cnt                    # 고점영역 성숙도
    feat['cy_cycle_top_zone'] = (feat['cy_cycle_position'] > 0.85).astype(float)
    feat['cy_cycle_turning'] = ((feat['cy_cycle_position'] < feat['cy_cycle_position'].shift(5)) &
                                (feat['cy_cycle_position'].shift(5) > 0.8)).astype(float)
    _above_sma50 = (_c > _c.rolling(50).mean()).astype(float)
    feat['cy_uptrend_age'] = _above_sma50.groupby((_above_sma50 != _above_sma50.shift()).cumsum()).cumcount() + 1
    feat['cy_uptrend_age'] = feat['cy_uptrend_age'].where(_above_sma50 > 0, 0)
    feat['cy_uptrend_overaged'] = (feat['cy_uptrend_age'] > 60).astype(float)
    _surge = (_c.pct_change(20) > 0.15)
    feat['cy_post_surge_flag'] = _surge.rolling(20).max().fillna(0)
    _dom = pd.Series(_c.index.day, index=_c.index)
    feat['cy_month_end_zone'] = (_dom >= 25).astype(float)

    # ── 27G. 종합 하락 예측 점수 v7 ─────────────────────────────
    _zz = (
        (feat['fr_hurst_falling']) +
        (feat['fr_lyapunov_rising']) +
        (feat['en_ke_dissipating']) +
        (feat['en_momentum_decaying']) +
        (feat['en_pe_high_flag']) +
        (feat['sp_hf_energy_rising']) +
        (feat['vp_pv_corr_neg']) +
        (feat['vp_below_vwap']) +
        (feat['vp_vw_return_neg']) +
        (feat['dt_left_tail_heavy']) +
        (feat['dt_extreme_dn_excess_20d'] > 1).astype(float) +
        (feat['cy_cycle_top_zone']) +
        (feat['cy_uptrend_overaged'])
    )
    feat['zz_drop_score_v7'] = _zz                                     # 0~13
    feat['zz_drop_score_v7_high'] = (_zz >= 7).astype(float)
    feat['zz_drop_score_v7_extreme'] = (_zz >= 10).astype(float)
    for p in [3, 5, 10]:
        feat[f'zz_drop_score_v7_sum_{p}d'] = _zz.rolling(p).sum()
    feat['zz_drop_score_v7_accel'] = _zz - _zz.shift(5)
    feat['zz_drop_score_v7_zscore60'] = calc_zscore(_zz, 60)
    feat['zz_drop_score_v7_persist_5d'] = (_zz.rolling(5).min() >= 5).astype(float)
    feat['zz_drop_score_v7_rising'] = (_zz > _zz.shift(3)).astype(float)

    # ══════════════════════════════════════════════════════════════
    #  28. 저상관 하락 예측 지표 — 입력 소스 분산으로 상호 상관 최소화 (~69개)
    #      (평균 절대상관 ~0.21, 수익률과 평균 |corr| ~0.17)
    #      vo_(거래량단독) rk_(순위/robust) sg_(부호열) cal_(캘린더)
    #      qt_(분위) rb_(범위전용) iv_(저변동성반전) lc_(저상관종합)
    # ══════════════════════════════════════════════════════════════
    _o = op; _h = hi; _l = lo; _c = cl; _v = vo
    _rng     = (_h - _l).replace(0, np.nan)
    _pc      = _c.shift(1)
    _ret1    = _c.pct_change()

    # ── 28A. 거래량 단독 신호 (가격과 독립) ─────────────────────
    _vlog = np.log1p(_v)
    feat['vo_log_zscore_20'] = calc_zscore(_vlog, 20)
    feat['vo_log_zscore_60'] = calc_zscore(_vlog, 60)
    feat['vo_trend_5_20'] = _v.rolling(5).mean() / _v.rolling(20).mean().replace(0, np.nan) - 1
    feat['vo_trend_20_60'] = _v.rolling(20).mean() / _v.rolling(60).mean().replace(0, np.nan) - 1
    feat['vo_volume_vol_20'] = _v.pct_change().rolling(20).std()       # 거래량 변동성
    feat['vo_volume_vol_rising'] = (feat['vo_volume_vol_20'] > feat['vo_volume_vol_20'].shift(10)).astype(float)
    feat['vo_pctrank_60'] = _v.rolling(60).apply(lambda x: pd.Series(x).rank(pct=True).iloc[-1], raw=False)
    feat['vo_pctrank_120'] = _v.rolling(120).apply(lambda x: pd.Series(x).rank(pct=True).iloc[-1], raw=False)
    feat['vo_surge_freq_20'] = (_v > _v.rolling(60).mean() * 2).rolling(20).mean()
    feat['vo_accel_5'] = _v.rolling(5).mean().pct_change(5)
    feat['vo_accel_spike'] = (feat['vo_accel_5'] > 0.5).astype(float)
    feat['vo_concentration_10'] = _v.rolling(10).max() / _v.rolling(10).sum().replace(0, np.nan)

    # ── 28B. 순위/로버스트 통계 (이상치 강건, 가격레벨 독립) ────
    for p in [20, 60]:
        feat[f'rk_close_rank_{p}'] = _c.rolling(p).apply(
            lambda x: pd.Series(x).rank(pct=True).iloc[-1], raw=False)
    feat['rk_rank_drop_5'] = feat['rk_close_rank_20'].shift(5) - feat['rk_close_rank_20']
    feat['rk_rank_falling'] = (feat['rk_close_rank_20'] < feat['rk_close_rank_20'].shift(5)).astype(float)
    for p in [20]:
        _med = _c.rolling(p).median()
        _mad = (_c - _med).abs().rolling(p).median().replace(0, np.nan)
        feat[f'rk_robust_z_{p}'] = (_c - _med) / (1.4826 * _mad)       # MAD 기반 robust z
    feat['rk_robust_z_extreme'] = (feat['rk_robust_z_20'] > 2.5).astype(float)
    feat['rk_ret_rank_60'] = _ret1.rolling(60).apply(
        lambda x: pd.Series(x).rank(pct=True).iloc[-1], raw=False)
    feat['rk_ret_low_rank'] = (feat['rk_ret_rank_60'] < 0.1).astype(float)
    def _spearman_trend(x):
        x = x[~np.isnan(x)]
        n = len(x)
        if n < 10: return np.nan
        rank = pd.Series(x).rank().values
        t = np.arange(n)
        return float(np.corrcoef(rank, t)[0, 1])
    feat['rk_spearman_trend_20'] = _c.rolling(20).apply(_spearman_trend, raw=True)
    feat['rk_spearman_neg'] = (feat['rk_spearman_trend_20'] < -0.3).astype(float)

    # ── 28C. 부호열 패턴 (수익률 부호만 사용 — 크기 독립) ───────
    _sign = np.sign(_ret1).fillna(0)
    for p in [5, 10, 20]:
        feat[f'sg_sign_sum_{p}'] = _sign.rolling(p).sum()
    feat['sg_flip_count_10'] = (_sign.diff() != 0).rolling(10).sum()
    feat['sg_three_down'] = ((_sign == -1) & (_sign.shift(1) == -1) & (_sign.shift(2) == -1)).astype(float)
    _grp = (_sign != _sign.shift()).cumsum()
    _runlen = _sign.groupby(_grp).cumcount() + 1
    _dn_run = _runlen.where(_sign < 0, 0)
    _up_run = _runlen.where(_sign > 0, 0)
    feat['sg_dn_run_max_20'] = _dn_run.rolling(20).max()
    feat['sg_up_run_max_20'] = _up_run.rolling(20).max()
    feat['sg_run_asym_20'] = _dn_run.rolling(20).max() - _up_run.rolling(20).max()
    def _sign_entropy(x):
        x = x[~np.isnan(x)]
        if len(x) < 8: return np.nan
        p_up = np.mean(x > 0)
        p_dn = np.mean(x < 0)
        e = 0.0
        for pp in [p_up, p_dn]:
            if pp > 0: e -= pp * np.log(pp)
        return float(e)
    feat['sg_sign_entropy_20'] = _ret1.rolling(20).apply(_sign_entropy, raw=True)
    _w = np.arange(1, 11)
    feat['sg_weighted_dn_10'] = (_sign == -1).rolling(10).apply(
        lambda x: np.sum(x * _w) / _w.sum(), raw=True)

    # ── 28D. 캘린더/시간 효과 (가격과 완전 독립) ────────────────
    _dow = pd.Series(_c.index.dayofweek, index=_c.index)
    _dom = pd.Series(_c.index.day, index=_c.index)
    _month = pd.Series(_c.index.month, index=_c.index)
    feat['cal_is_monday'] = (_dow == 0).astype(float)
    feat['cal_is_friday'] = (_dow == 4).astype(float)
    feat['cal_month_end'] = (_dom >= 26).astype(float)
    feat['cal_month_start'] = (_dom <= 3).astype(float)
    feat['cal_quarter_end'] = ((_month.isin([3, 6, 9, 12])) & (_dom >= 25)).astype(float)
    feat['cal_weak_season'] = (_month.isin([9, 10])).astype(float)     # 역사적 약세 계절
    feat['cal_january'] = (_month == 1).astype(float)
    feat['cal_opex_week'] = ((_dom >= 15) & (_dom <= 21)).astype(float)
    _days_idx = np.arange(len(_c))
    feat['cal_cycle_21'] = pd.Series(_days_idx % 21, index=_c.index).astype(float)

    # ── 28E. 분위 기반 변동성/위치 (절대레벨 독립, robust) ──────
    feat['qt_range_pctrank_60'] = (_rng / _c).rolling(60).apply(
        lambda x: pd.Series(x).rank(pct=True).iloc[-1], raw=False)
    feat['qt_range_high_flag'] = (feat['qt_range_pctrank_60'] > 0.85).astype(float)
    for p in [20, 60]:
        _q25 = _c.rolling(p).quantile(0.25)
        _q75 = _c.rolling(p).quantile(0.75)
        feat[f'qt_iqr_pos_{p}'] = (_c - _q25) / (_q75 - _q25).replace(0, np.nan)
    feat['qt_above_iqr_60'] = (feat['qt_iqr_pos_60'] > 1.0).astype(float)
    for p in [20]:
        _rq = _ret1.rolling(p).quantile(0.75) - _ret1.rolling(p).quantile(0.25)
        feat[f'qt_ret_iqr_{p}'] = _rq
        feat[f'qt_ret_iqr_rising_{p}'] = (_rq > _rq.shift(10)).astype(float)
    feat['qt_extreme_combo'] = ((feat['qt_iqr_pos_60'] > 1.0) & (feat['qt_range_pctrank_60'] > 0.8)).astype(float)

    # ── 28F. 범위 전용 지표 (종가 무관, 고저만) ─────────────────
    feat['rb_high_falling_10'] = (_h.rolling(5).max() < _h.rolling(5).max().shift(5)).astype(float)
    feat['rb_high_lower_cnt_20'] = (_h < _h.shift(1)).rolling(20).sum()
    feat['rb_low_breaking_10'] = (_l < _l.rolling(10).min().shift(1)).astype(float)
    feat['rb_low_lower_cnt_20'] = (_l < _l.shift(1)).rolling(20).sum()
    _hl_mid = (_h + _l) / 2
    feat['rb_midpoint_trend_10'] = _hl_mid / _hl_mid.shift(10) - 1
    feat['rb_midpoint_falling'] = (_hl_mid < _hl_mid.shift(5)).astype(float)
    feat['rb_high_vs_open'] = (_h - _o) / _rng
    feat['rb_high_near_open_10'] = ((_h - _o) / _rng < 0.3).rolling(10).mean()
    feat['rb_range_compress_60'] = (_rng / _c) / (_rng / _c).rolling(60).mean().replace(0, np.nan)

    # ── 28G. 저변동성 함정/반전 (다른 위험 차원) ────────────────
    _vol10 = _ret1.rolling(10).std()
    _vol60 = _ret1.rolling(60).std()
    feat['iv_low_vol_flag'] = (_vol10 < _vol60 * 0.6).astype(float)    # 폭풍전야
    feat['iv_vol_compression_60'] = _vol10 / _vol60.replace(0, np.nan)
    feat['iv_calm_at_top'] = ((_vol10 < _vol60 * 0.7) &
                              (feat['rk_close_rank_60'] > 0.8)).astype(float)
    feat['iv_vol_trough_turn'] = ((_vol10 > _vol10.shift(3)) &
                                  (_vol10.shift(3) < _vol60 * 0.6)).astype(float)
    _narrow = (_rng <= _rng.rolling(10).min() * 1.1)
    feat['iv_narrow_then_drop'] = (_narrow.shift(1) & (_ret1 < -0.005)).astype(float)

    # ── 28H. 저상관 종합 하락 점수 (서로 다른 차원 결합) ────────
    _lc = (
        (feat['vo_log_zscore_60'] > 1.5).astype(float) +        # 거래량 차원
        (feat['rk_spearman_neg']) +                              # 순위 추세 차원
        (feat['sg_three_down']) +                                # 부호열 차원
        (feat['cal_weak_season']) +                              # 캘린더 차원
        (feat['qt_extreme_combo']) +                             # 분위 차원
        (feat['rb_midpoint_falling']) +                          # 범위 차원
        (feat['iv_calm_at_top'])                                 # 저변동 차원
    )
    feat['lc_drop_score_v8'] = _lc                              # 0~7
    feat['lc_drop_score_v8_high'] = (_lc >= 4).astype(float)
    for p in [5, 10]:
        feat[f'lc_drop_score_v8_sum_{p}d'] = _lc.rolling(p).sum()
    feat['lc_drop_score_v8_rising'] = (_lc > _lc.shift(3)).astype(float)
    feat['lc_drop_score_v8_zscore60'] = calc_zscore(_lc, 60)


    # ══════════════════════════════════════════════════════════════
    #  29. 섹터/산업 상대강도 기반 상승·하락 예측 (~64개, 저상관·중복없음)
    #      (평균 절대상관 ~0.23, 종목수익률과 |corr| ~0.13 — 상대 정보라 독립)
    #      접두사: rsx_(상대강도) sru_(섹터상대상승) srd_(섹터상대하락)
    #              brd_(브레드스) rot_(로테이션) ssp_(섹터스프레드)
    #              led_(선도/후행) scm_(섹터종합)
    #      ※ closes(시장/섹터 ETF) + TICKER 사용 — 함수 내 이미 가용
    # ══════════════════════════════════════════════════════════════
    _c29 = cl
    _ret29 = _c29.pct_change()
    SECTORS_29 = ['XLK','XLV','XLF','XLY','XLP','XLE','XLI','XLB','XLU','XLRE','XLC']
    _avail_sec29 = [s for s in SECTORS_29 if s in closes.columns and s != TICKER]

    # ── 29A. 시장 대비 상대강도(RS) — 비율선/기울기/가속 ────────
    _bench29 = None
    for b in ['SPY', 'QQQ', 'RSP']:
        if b in closes.columns and b != TICKER:
            _bench29 = closes[b]; break
    if _bench29 is not None:
        _rs_line = _c29 / _bench29.replace(0, np.nan)
        feat['rsx_rs_line_zscore_60'] = calc_zscore(_rs_line, 60)
        for p in [10, 20, 50]:
            feat[f'rsx_rs_slope_{p}d'] = _rs_line / _rs_line.shift(p) - 1
        feat['rsx_rs_below_sma20'] = (_rs_line < _rs_line.rolling(20).mean()).astype(float)
        feat['rsx_rs_below_sma50'] = (_rs_line < _rs_line.rolling(50).mean()).astype(float)
        feat['rsx_rs_new_high_20'] = (_rs_line >= _rs_line.rolling(20).max() - 1e-12).astype(float)
        feat['rsx_rs_new_low_20']  = (_rs_line <= _rs_line.rolling(20).min() + 1e-12).astype(float)
        _rs_mom = _rs_line.pct_change(10)
        feat['rsx_rs_accel'] = _rs_mom - _rs_mom.shift(10)
        feat['rsx_rs_accel_neg'] = (feat['rsx_rs_accel'] < 0).astype(float)
        _underperf = (_ret29 < _bench29.pct_change()).astype(float)
        for p in [10, 20]:
            feat[f'rsx_underperf_ratio_{p}d'] = _underperf.rolling(p).mean()
        feat['rsx_persistent_underperf'] = (_underperf.rolling(10).mean() > 0.65).astype(float)
        feat['rsx_stock_dn_mkt_up'] = ((_ret29 < 0) & (_bench29.pct_change() > 0)).astype(float)
        feat['rsx_stock_dn_mkt_up_10d'] = feat['rsx_stock_dn_mkt_up'].rolling(10).sum()
        feat['rsx_stock_up_mkt_dn'] = ((_ret29 > 0) & (_bench29.pct_change() < 0)).astype(float)
        feat['rsx_stock_up_mkt_dn_10d'] = feat['rsx_stock_up_mkt_dn'].rolling(10).sum()

    # ── 29B. 소속 섹터 추정 후 섹터 대비 상대 위치 ──────────────
    _best_sec29 = None; _best_corr29 = -2
    if _avail_sec29:
        for s in _avail_sec29:
            c = _ret29.rolling(60).corr(closes[s].pct_change()).mean()
            if pd.notna(c) and c > _best_corr29:
                _best_corr29 = c; _best_sec29 = s
    if _best_sec29 is not None:
        _sec = closes[_best_sec29]
        _sec_ret = _sec.pct_change()
        for p in [5, 10, 20]:
            feat[f'sru_vs_sector_mom_{p}d'] = _c29.pct_change(p) - _sec.pct_change(p)
        _sec_rs = _c29 / _sec.replace(0, np.nan)
        feat['sru_sector_rs_zscore_60'] = calc_zscore(_sec_rs, 60)
        feat['sru_sector_rs_slope_20'] = _sec_rs / _sec_rs.shift(20) - 1
        feat['srd_both_falling_5d'] = ((_c29.pct_change(5) < 0) & (_sec.pct_change(5) < 0) &
                                       (_c29.pct_change(5) < _sec.pct_change(5))).astype(float)
        feat['srd_laggard_in_sector'] = ((_sec.pct_change(10) > 0) & (_c29.pct_change(10) < 0)).astype(float)
        feat['srd_laggard_20d'] = feat['srd_laggard_in_sector'].rolling(20).sum()
        feat['sru_leader_in_sector'] = ((_sec.pct_change(10) < 0) & (_c29.pct_change(10) > 0)).astype(float)
        feat['srd_sector_beta_60'] = (_ret29.rolling(60).cov(_sec_ret) /
                                      _sec_ret.rolling(60).var().replace(0, np.nan))
        feat['srd_high_beta_flag'] = (feat['srd_sector_beta_60'] > 1.3).astype(float)
        feat['srd_sector_decorr_20'] = _ret29.rolling(20).corr(_sec_ret)
        feat['srd_decorr_drop'] = (feat['srd_sector_decorr_20'] < 0.3).astype(float)

    # ── 29C. 시장 브레드스 (섹터 전반 건강도 — 시스템 위험) ─────
    if len(_avail_sec29) >= 5:
        _sec_df = pd.DataFrame({s: closes[s] for s in _avail_sec29})
        _above50 = pd.DataFrame({s: (_sec_df[s] > _sec_df[s].rolling(50).mean()).astype(float)
                                 for s in _avail_sec29})
        feat['brd_sectors_above_sma50_pct'] = _above50.mean(axis=1)
        feat['brd_breadth_weak'] = (_above50.mean(axis=1) < 0.4).astype(float)
        feat['brd_breadth_falling'] = (_above50.mean(axis=1) < _above50.mean(axis=1).shift(10)).astype(float)
        _pos20 = pd.DataFrame({s: (_sec_df[s].pct_change(20) > 0).astype(float)
                               for s in _avail_sec29})
        feat['brd_sectors_pos20_pct'] = _pos20.mean(axis=1)
        feat['brd_most_sectors_down'] = (_pos20.mean(axis=1) < 0.3).astype(float)
        _newlow = pd.DataFrame({s: (_sec_df[s] <= _sec_df[s].rolling(60).min() * 1.01).astype(float)
                                for s in _avail_sec29})
        feat['brd_sectors_newlow_cnt'] = _newlow.sum(axis=1)
        feat['brd_systemic_stress'] = (_newlow.sum(axis=1) >= 3).astype(float)
        _sec_chg = _sec_df.pct_change()
        _sec_mean = _sec_chg.mean(axis=1)
        _avg_corr = pd.DataFrame({s: _sec_chg[s].rolling(20).corr(_sec_mean)
                                  for s in _avail_sec29}).mean(axis=1)
        feat['brd_sector_corr_avg_20'] = _avg_corr
        feat['brd_corr_spike'] = (_avg_corr > 0.8).astype(float)
        if _bench29 is not None:
            feat['brd_divergence'] = ((_bench29.pct_change(20) > 0) &
                                      (_above50.mean(axis=1) < _above50.mean(axis=1).shift(20))).astype(float)

    # ── 29D. 방어/공격 로테이션 (위험선호 변화) ─────────────────
    _defs29 = [s for s in ['XLU','XLP','XLV'] if s in closes.columns and s != TICKER]
    _aggs29 = [s for s in ['XLK','XLY','XLF','XLC'] if s in closes.columns and s != TICKER]
    if len(_defs29) >= 2 and len(_aggs29) >= 2:
        _def_ret = pd.concat([closes[s].pct_change() for s in _defs29], axis=1).mean(axis=1)
        _agg_ret = pd.concat([closes[s].pct_change() for s in _aggs29], axis=1).mean(axis=1)
        for p in [5, 20]:
            _dp = (1 + _def_ret).rolling(p).apply(np.prod, raw=True) - 1
            _ap = (1 + _agg_ret).rolling(p).apply(np.prod, raw=True) - 1
            feat[f'rot_def_minus_agg_{p}d'] = _dp - _ap
        feat['rot_to_defensive'] = (feat['rot_def_minus_agg_20d'] > 0).astype(float)
        feat['rot_defensive_turning'] = ((feat['rot_def_minus_agg_5d'] > 0) &
                                         (feat['rot_def_minus_agg_20d'] < 0)).astype(float)
        feat['rot_def_accel'] = feat['rot_def_minus_agg_5d'] - feat['rot_def_minus_agg_20d']
        feat['rot_risk_for_stock'] = (feat['rot_def_minus_agg_20d'] > 0.01).astype(float)

    # ── 29E. 섹터 스프레드/순환 (산업 모멘텀) ───────────────────
    if len(_avail_sec29) >= 5:
        _sec_df2 = pd.DataFrame({s: closes[s] for s in _avail_sec29})
        _sec_ret20 = _sec_df2.pct_change(20)
        feat['ssp_sector_dispersion_20'] = _sec_ret20.std(axis=1)
        feat['ssp_top_bottom_gap_20'] = _sec_ret20.max(axis=1) - _sec_ret20.min(axis=1)
        if _best_sec29 is not None:
            _sec_rank = _sec_ret20.rank(axis=1, pct=True)
            if _best_sec29 in _sec_rank.columns:
                feat['ssp_sector_rank_pct'] = _sec_rank[_best_sec29]
                feat['ssp_sector_bottom_quartile'] = (_sec_rank[_best_sec29] < 0.25).astype(float)
        feat['ssp_dispersion_falling'] = (_sec_ret20.std(axis=1) < _sec_ret20.std(axis=1).shift(10)).astype(float)

    # ── 29F. 선도-후행 (경기 사이클 ETF 선행 신호) ──────────────
    _lead_signals = pd.Series(0.0, index=_c29.index)
    _lead_cnt = 0
    for _ls in ['SMH', 'IYT', 'IWM', 'XLF', 'KRE']:
        if _ls in closes.columns and _ls != TICKER:
            _lead_signals = _lead_signals + (closes[_ls].pct_change(10) < -0.03).astype(float)
            _lead_cnt += 1
    if _lead_cnt > 0:
        feat['led_cyclical_weakness'] = _lead_signals / _lead_cnt
        feat['led_cyclical_weak_flag'] = (_lead_signals >= max(2, _lead_cnt // 2)).astype(float)
    if 'IWM' in closes.columns and 'SPY' in closes.columns and TICKER not in ('IWM', 'SPY'):
        feat['led_smallcap_rs_20'] = closes['IWM'].pct_change(20) - closes['SPY'].pct_change(20)
        feat['led_smallcap_weak'] = (feat['led_smallcap_rs_20'] < -0.03).astype(float)
    if 'SPHB' in closes.columns and 'SPLV' in closes.columns:
        feat['led_beta_appetite_20'] = closes['SPHB'].pct_change(20) - closes['SPLV'].pct_change(20)
        feat['led_risk_off_flag'] = (feat['led_beta_appetite_20'] < -0.02).astype(float)

    # ── 29G. 섹터 종합 상승/하락 점수 ───────────────────────────
    _scm_dn = pd.Series(0.0, index=_c29.index)
    for _k in ['rsx_persistent_underperf', 'srd_laggard_in_sector', 'brd_breadth_weak',
               'brd_corr_spike', 'rot_to_defensive', 'led_cyclical_weak_flag',
               'ssp_sector_bottom_quartile']:
        if _k in feat.columns:
            _scm_dn = _scm_dn + feat[_k]
    feat['scm_sector_drop_score'] = _scm_dn
    feat['scm_sector_drop_high'] = (_scm_dn >= 4).astype(float)
    feat['scm_sector_drop_rising'] = (_scm_dn > _scm_dn.shift(3)).astype(float)
    for p in [5, 10]:
        feat[f'scm_sector_drop_sum_{p}d'] = _scm_dn.rolling(p).sum()
    _scm_up = pd.Series(0.0, index=_c29.index)
    for _k in ['rsx_stock_up_mkt_dn', 'sru_leader_in_sector', 'rsx_rs_new_high_20']:
        if _k in feat.columns:
            _scm_up = _scm_up + feat[_k]
    feat['scm_sector_rise_score'] = _scm_up
    feat['scm_sector_net_score'] = _scm_up - _scm_dn


# ══════════════════════════════════════════════════════════════
    #  30. 섹터/산업 상대 — 6차 확장 (~56개, 섹션29와도 중복없음·저상관)
    #      (평균 절대상관 ~0.20, 종목수익률과 |corr| ~0.10 — 가장 독립적)
    #      접두사: dsp_(섹터분산내위치) bta_(조건부베타) crk_(상관레짐)
    #              chn_(섹터채널상태) qrt_(섹터분위전이) mcr_(매크로축)
    #              tlt_(채권/안전자산축) scn_(신규섹터종합)
    #      ※ closes + TICKER 사용 — 함수 내 이미 가용
    # ══════════════════════════════════════════════════════════════
    _c30 = cl
    _ret30 = _c30.pct_change()
    SECTORS_30 = ['XLK','XLV','XLF','XLY','XLP','XLE','XLI','XLB','XLU','XLRE','XLC']
    _sec_av30 = [s for s in SECTORS_30 if s in closes.columns and s != TICKER]

    # 소속 섹터 추정 (절대편차 최소 = 가장 유사한 섹터)
    _best30 = None; _bestv30 = 1e18
    if _sec_av30:
        for s in _sec_av30:
            diff = (_ret30 - closes[s].pct_change()).abs().rolling(60).mean().mean()
            if pd.notna(diff) and diff < _bestv30:
                _bestv30 = diff; _best30 = s

    # ── 30A. 섹터 분산 내 종목 위치 (z-score 거리) ──────────────
    if len(_sec_av30) >= 5:
        _peer_ret = pd.DataFrame({s: closes[s].pct_change() for s in _sec_av30})
        _peer_mean = _peer_ret.mean(axis=1)
        for p in [5, 20]:
            _stock_cum = _c30.pct_change(p)
            _peer_cum = (1 + _peer_ret).rolling(p).apply(np.prod, raw=True) - 1
            feat[f'dsp_zdist_from_peers_{p}d'] = (_stock_cum - _peer_cum.mean(axis=1)) / _peer_cum.std(axis=1).replace(0, np.nan)
        _below_peer = (_ret30 < _peer_mean).astype(float)
        feat['dsp_below_peer_streak'] = _below_peer.groupby(
            (_below_peer != _below_peer.shift()).cumsum()).cumcount() + 1
        feat['dsp_below_peer_streak'] = feat['dsp_below_peer_streak'].where(_below_peer > 0, 0)
        feat['dsp_below_peer_ratio_20'] = _below_peer.rolling(20).mean()
        _stock_vs_each = pd.DataFrame({s: (_c30.pct_change(20) > closes[s].pct_change(20)).astype(float)
                                       for s in _sec_av30})
        feat['dsp_beat_peer_ratio_20'] = _stock_vs_each.mean(axis=1)   # 높으면 상승, 낮으면 하락
        feat['dsp_worst_in_peers'] = (_stock_vs_each.mean(axis=1) < 0.2).astype(float)
        feat['dsp_best_in_peers'] = (_stock_vs_each.mean(axis=1) > 0.8).astype(float)

    # ── 30B. 조건부 베타 (하락장 베타 vs 상승장 베타) ───────────
    _bench30 = None
    for b in ['SPY', 'QQQ', 'RSP']:
        if b in closes.columns and b != TICKER:
            _bench30 = closes[b]; break
    if _bench30 is not None:
        _bret = _bench30.pct_change()
        for p in [60]:
            _dn_mask = _bret < 0
            _x_dn = _bret.where(_dn_mask); _y_dn = _ret30.where(_dn_mask)
            feat[f'bta_downside_beta_{p}d'] = (_y_dn.rolling(p, min_periods=10).cov(_x_dn) /
                                               _x_dn.rolling(p, min_periods=10).var().replace(0, np.nan))
            _up_mask = _bret > 0
            _x_up = _bret.where(_up_mask); _y_up = _ret30.where(_up_mask)
            feat[f'bta_upside_beta_{p}d'] = (_y_up.rolling(p, min_periods=10).cov(_x_up) /
                                             _x_up.rolling(p, min_periods=10).var().replace(0, np.nan))
            feat[f'bta_beta_asym_{p}d'] = feat[f'bta_downside_beta_{p}d'] - feat[f'bta_upside_beta_{p}d']
        feat['bta_bad_asym_flag'] = (feat['bta_beta_asym_60d'] > 0.3).astype(float)
        feat['bta_high_downside_beta'] = (feat['bta_downside_beta_60d'] > 1.3).astype(float)
        _mkt_dn_sum = _bret.where(_bret < 0).rolling(60, min_periods=10).sum()
        _stk_when_mkt_dn = _ret30.where(_bret < 0).rolling(60, min_periods=10).sum()
        feat['bta_downside_capture_60'] = _stk_when_mkt_dn / _mkt_dn_sum.replace(0, np.nan)
        feat['bta_high_dn_capture'] = (feat['bta_downside_capture_60'] > 1.1).astype(float)

    # ── 30C. 상관 레짐 변화 (섹터/시장과의 동조 변화) ───────────
    if _bench30 is not None:
        _bret = _bench30.pct_change()
        _corr_short = _ret30.rolling(20).corr(_bret)
        _corr_long = _ret30.rolling(60).corr(_bret)
        feat['crk_corr_regime_shift'] = _corr_short - _corr_long
        feat['crk_corr_rising'] = (_corr_short > _corr_long + 0.2).astype(float)
        feat['crk_corr_breaking'] = (_corr_short < _corr_long - 0.2).astype(float)
        feat['crk_corr_zscore_60'] = calc_zscore(_corr_short, 60)
    if _best30 is not None:
        _sec = closes[_best30]
        _scorr20 = _ret30.rolling(20).corr(_sec.pct_change())
        feat['crk_sector_corr_20'] = _scorr20
        feat['crk_sector_decoupling'] = (_scorr20 < 0.4).astype(float)
        feat['crk_decouple_and_drop'] = ((_scorr20 < 0.4) & (_ret30.rolling(5).mean() < 0)).astype(float)

    # ── 30D. 섹터 채널 내 위치 (소속 섹터 ETF의 기술적 상태) ────
    if _best30 is not None:
        _sec = closes[_best30]
        feat['chn_sector_pos_252'] = (_sec - _sec.rolling(252, min_periods=60).min()) / \
                                     (_sec.rolling(252, min_periods=60).max() -
                                      _sec.rolling(252, min_periods=60).min()).replace(0, np.nan)
        feat['chn_sector_near_low'] = (feat['chn_sector_pos_252'] < 0.2).astype(float)
        feat['chn_sector_below_sma50'] = (_sec < _sec.rolling(50).mean()).astype(float)
        feat['chn_sector_below_sma200'] = (_sec < _sec.rolling(200, min_periods=60).mean()).astype(float)
        feat['chn_sector_drawdown_60'] = _sec / _sec.rolling(60).max() - 1
        feat['chn_sector_deep_dd'] = (feat['chn_sector_drawdown_60'] < -0.1).astype(float)
        feat['chn_sector_mom_neg_20'] = (_sec.pct_change(20) < 0).astype(float)
        _sec_vol = _sec.pct_change().rolling(20).std()
        feat['chn_sector_vol_spike'] = (_sec_vol > _sec_vol.rolling(60).mean() * 1.5).astype(float)

    # ── 30E. 섹터 분위 전이 (섹터 순위 모멘텀) ──────────────────
    if len(_sec_av30) >= 5 and _best30 is not None:
        _sec_df = pd.DataFrame({s: closes[s] for s in _sec_av30})
        _rank20 = _sec_df.pct_change(20).rank(axis=1, pct=True)
        if _best30 in _rank20.columns:
            _my_rank = _rank20[_best30]
            feat['qrt_sector_rank_20'] = _my_rank
            feat['qrt_sector_rank_falling'] = (_my_rank < _my_rank.shift(10)).astype(float)
            feat['qrt_sector_rank_drop_5'] = _my_rank.shift(5) - _my_rank
            feat['qrt_sector_top_to_bottom'] = ((_my_rank < 0.4) & (_my_rank.shift(20) > 0.7)).astype(float)
            feat['qrt_sector_bottom_zone'] = (_my_rank < 0.3).astype(float)
            feat['qrt_sector_top_zone'] = (_my_rank > 0.7).astype(float)
        feat['qrt_rank_dispersion'] = _rank20.std(axis=1)

    # ── 30F. 매크로 축 상대 (금리/변동성/달러 대비) ─────────────
    _tnx30 = closes.get('^TNX'); _vix30 = closes.get('^VIX'); _uup30 = closes.get('UUP')
    if _tnx30 is not None:
        _tnx_chg = _tnx30.diff()
        feat['mcr_rate_sensitivity_60'] = _ret30.rolling(60).corr(_tnx_chg)
        feat['mcr_rate_vulnerable'] = (feat['mcr_rate_sensitivity_60'] < -0.2).astype(float)
        feat['mcr_rate_spike_risk'] = ((_tnx30.diff(5) > 0.15) & (feat['mcr_rate_sensitivity_60'] < -0.1)).astype(float)
    if _vix30 is not None:
        feat['mcr_vix_sensitivity_20'] = _ret30.rolling(20).corr(_vix30.pct_change())
        feat['mcr_vix_vulnerable'] = (feat['mcr_vix_sensitivity_20'] < -0.4).astype(float)
        feat['mcr_high_vix_regime'] = ((_vix30 > _vix30.rolling(50).mean()) & (_vix30 > 20)).astype(float)
    if _uup30 is not None:
        feat['mcr_dollar_sensitivity_60'] = _ret30.rolling(60).corr(_uup30.pct_change())
        feat['mcr_dollar_vulnerable'] = ((feat['mcr_dollar_sensitivity_60'] < -0.2) &
                                         (_uup30.pct_change(20) > 0.02)).astype(float)

    # ── 30G. 채권/안전자산 축 (위험회피 시 종목 위치) ───────────
    _tltc30 = closes.get('TLT'); _gldc30 = closes.get('GLD'); _hygc30 = closes.get('HYG')
    if _tltc30 is not None:
        feat['tlt_bond_up_stock_down'] = ((_tltc30.pct_change(5) > 0.01) & (_ret30.rolling(5).mean() < 0)).astype(float)
        feat['tlt_bond_stock_corr_20'] = _ret30.rolling(20).corr(_tltc30.pct_change())
    if _hygc30 is not None:
        feat['tlt_hy_weak_signal'] = (_hygc30.pct_change(10) < -0.02).astype(float)
        feat['tlt_hy_lead_warning'] = ((_hygc30.pct_change(10) < -0.01) & (_ret30.rolling(10).mean() > 0)).astype(float)
    if _gldc30 is not None:
        feat['tlt_gold_up_stock_down'] = ((_gldc30.pct_change(10) > 0.03) & (_c30.pct_change(10) < 0)).astype(float)

    # ── 30H. 신규 섹터 종합 점수 (상승/하락 양방향) ─────────────
    _dn30 = pd.Series(0.0, index=_c30.index)
    for _k in ['dsp_worst_in_peers', 'bta_bad_asym_flag', 'crk_corr_rising',
               'chn_sector_below_sma50', 'qrt_sector_bottom_zone', 'mcr_vix_vulnerable',
               'tlt_hy_weak_signal']:
        if _k in feat.columns:
            _dn30 = _dn30 + feat[_k]
    feat['scn_drop_score'] = _dn30
    feat['scn_drop_high'] = (_dn30 >= 4).astype(float)
    feat['scn_drop_rising'] = (_dn30 > _dn30.shift(3)).astype(float)
    for p in [5, 10]:
        feat[f'scn_drop_sum_{p}d'] = _dn30.rolling(p).sum()
    _up30 = pd.Series(0.0, index=_c30.index)
    for _k in ['dsp_best_in_peers', 'qrt_sector_top_zone']:
        if _k in feat.columns:
            _up30 = _up30 + feat[_k]
    if 'chn_sector_below_sma50' in feat.columns:
        _up30 = _up30 + (1 - feat['chn_sector_below_sma50'])
    feat['scn_rise_score'] = _up30
    feat['scn_net_score'] = _up30 - _dn30

# ══════════════════════════════════════════════════════════════
    #  31. 섹터/산업 상대 — 7차 (~63개, 섹션29·30과도 중복없음·최저상관)
    #      (평균 절대상관 ~0.17, 종목수익률과 |corr| ~0.05 — 거의 독립)
    #      접두사: lag_(섹터선후행) coh_(응집/이탈) brt_(브레드스추세)
    #              pair_(핵심페어) flo_(자금흐름로테이션) rsk_(위험프리미엄)
    #              sdv_(섹터다이버전스) snw_(신규종합)
    #      ※ closes + TICKER 사용
    # ══════════════════════════════════════════════════════════════
    _c31 = cl
    _ret31 = _c31.pct_change()
    SECTORS_31 = ['XLK','XLV','XLF','XLY','XLP','XLE','XLI','XLB','XLU','XLRE','XLC']
    _sec_av31 = [s for s in SECTORS_31 if s in closes.columns and s != TICKER]
    _best31 = None; _bc31 = -2
    if _sec_av31:
        for s in _sec_av31:
            c = _ret31.rolling(90).corr(closes[s].pct_change()).mean()
            if pd.notna(c) and c > _bc31: _bc31 = c; _best31 = s

    # ── 31A. 섹터 선·후행 (리드-래그 시차) ──────────────────────
    if _best31 is not None:
        _sec = closes[_best31]; _sret = _sec.pct_change()
        for lag in [1, 2, 3]:
            feat[f'lag_lead_sector_{lag}d'] = _ret31.rolling(40).corr(_sret.shift(lag))    # 종목 선행
            feat[f'lag_follow_sector_{lag}d'] = _ret31.rolling(40).corr(_sret.shift(-lag))  # 종목 후행
        feat['lag_sector_led_down'] = ((_sret.shift(1) < -0.01) & (_sret.shift(2) < -0.005) &
                                       (_ret31 > -0.005)).astype(float)
        feat['lag_sector_led_down_5d'] = feat['lag_sector_led_down'].rolling(5).sum()
        feat['lag_sector_led_up'] = ((_sret.shift(1) > 0.01) & (_sret.shift(2) > 0.005) &
                                     (_ret31 < 0.005)).astype(float)
        feat['lag_mom_gap_shifted'] = _c31.pct_change(5) - _sec.pct_change(5).shift(3)

    # ── 31B. 섹터 응집/이탈 (peer 동조 붕괴) ────────────────────
    if len(_sec_av31) >= 5:
        _peer = pd.DataFrame({s: closes[s].pct_change() for s in _sec_av31})
        _peer_mean = _peer.mean(axis=1)
        _track_err = (_ret31 - _peer_mean).rolling(20).std()
        feat['coh_tracking_error_20'] = _track_err
        feat['coh_track_err_zscore_60'] = calc_zscore(_track_err, 60)
        feat['coh_track_err_rising'] = (_track_err > _track_err.shift(10)).astype(float)
        feat['coh_cum_excess_20'] = (_ret31 - _peer_mean).rolling(20).sum()
        feat['coh_negative_drift'] = (feat['coh_cum_excess_20'] < -0.03).astype(float)
        _peer_disp = _peer.std(axis=1)
        feat['coh_peer_dispersion_zscore'] = calc_zscore(_peer_disp, 60)
        _stock_vol = _ret31.rolling(20).std()
        _peer_vol = _peer_mean.rolling(20).std().replace(0, np.nan)
        feat['coh_rel_volatility_20'] = _stock_vol / _peer_vol
        feat['coh_high_rel_vol'] = (feat['coh_rel_volatility_20'] > 1.5).astype(float)

    # ── 31C. 브레드스 추세 (시장 건강도 시간 변화) ──────────────
    if len(_sec_av31) >= 5:
        _sd = pd.DataFrame({s: closes[s] for s in _sec_av31})
        _nh = pd.DataFrame({s: (_sd[s] >= _sd[s].rolling(20).max() - 1e-12).astype(float) for s in _sec_av31}).sum(axis=1)
        _nl = pd.DataFrame({s: (_sd[s] <= _sd[s].rolling(20).min() + 1e-12).astype(float) for s in _sec_av31}).sum(axis=1)
        feat['brt_net_new_high_20'] = (_nh - _nl) / len(_sec_av31)
        feat['brt_net_nh_falling'] = (feat['brt_net_new_high_20'] < feat['brt_net_new_high_20'].shift(5)).astype(float)
        _adv = (_sd.pct_change() > 0).sum(axis=1) - (_sd.pct_change() < 0).sum(axis=1)
        _ad_line = _adv.cumsum()
        feat['brt_ad_line_slope_10'] = (_ad_line - _ad_line.shift(10)) / len(_sec_av31)
        feat['brt_ad_line_falling'] = (feat['brt_ad_line_slope_10'] < 0).astype(float)
        _pos_ratio = (_sd.pct_change() > 0).mean(axis=1)
        feat['brt_breadth_momentum_10'] = _pos_ratio.rolling(10).mean() - _pos_ratio.rolling(20).mean()
        feat['brt_breadth_deteriorating'] = (feat['brt_breadth_momentum_10'] < -0.1).astype(float)
        feat['brt_thrust_down'] = (_pos_ratio.rolling(5).mean() < 0.3).astype(float)
        feat['brt_thrust_up'] = (_pos_ratio.rolling(5).mean() > 0.7).astype(float)

    # ── 31D. 핵심 페어 (위험선호 직접 측정) ─────────────────────
    _pairs31 = [('XLY', 'XLP', 'disc_stap'), ('XLK', 'XLU', 'tech_util'),
                ('XLF', 'XLRE', 'fin_re'), ('XLI', 'XLB', 'ind_mat')]
    for a, b, lbl in _pairs31:
        if a in closes.columns and b in closes.columns and a != TICKER and b != TICKER:
            _ratio = closes[a] / closes[b].replace(0, np.nan)
            feat[f'pair_{lbl}_ratio_z60'] = calc_zscore(_ratio, 60)
            feat[f'pair_{lbl}_slope_20'] = _ratio / _ratio.shift(20) - 1
            feat[f'pair_{lbl}_risk_off'] = (_ratio < _ratio.rolling(20).mean()).astype(float)

    # ── 31E. 자금흐름 로테이션 가속 ─────────────────────────────
    if len(_sec_av31) >= 5:
        _sd2 = pd.DataFrame({s: closes[s] for s in _sec_av31})
        _sec_mom5 = _sd2.pct_change(5)
        _sec_mom20 = _sd2.pct_change(20)
        if _best31 is not None and _best31 in _sec_mom5.columns:
            _rank5 = _sec_mom5.rank(axis=1, pct=True)[_best31]
            _rank20 = _sec_mom20.rank(axis=1, pct=True)[_best31]
            feat['flo_rank_momentum_shift'] = _rank5 - _rank20
            feat['flo_outflow_accel'] = (feat['flo_rank_momentum_shift'] < -0.2).astype(float)
            feat['flo_inflow_accel'] = (feat['flo_rank_momentum_shift'] > 0.2).astype(float)
        feat['flo_rotation_intensity'] = _sec_mom20.std(axis=1)
        feat['flo_strong_rotation'] = (_sec_mom20.std(axis=1) >
                                       _sec_mom20.std(axis=1).rolling(60).mean() * 1.3).astype(float)
        _def_m = pd.concat([_sec_mom20[s] for s in ['XLU','XLP','XLV'] if s in _sec_mom20.columns], axis=1).mean(axis=1)
        _agg_m = pd.concat([_sec_mom20[s] for s in ['XLK','XLY','XLF'] if s in _sec_mom20.columns], axis=1).mean(axis=1)
        feat['flo_defensive_inflow'] = _def_m - _agg_m
        feat['flo_defensive_dominant'] = (feat['flo_defensive_inflow'] > 0.02).astype(float)

    # ── 31F. 위험 프리미엄 축 (섹터 위험조정 상대) ──────────────
    if _best31 is not None:
        _sec = closes[_best31]; _sret = _sec.pct_change()
        for p in [20, 60]:
            _stk_sharpe = _ret31.rolling(p).mean() / _ret31.rolling(p).std().replace(0, np.nan)
            _sec_sharpe = _sret.rolling(p).mean() / _sret.rolling(p).std().replace(0, np.nan)
            feat[f'rsk_sharpe_vs_sector_{p}'] = _stk_sharpe - _sec_sharpe
        feat['rsk_sharpe_underperform'] = (feat['rsk_sharpe_vs_sector_20'] < -0.1).astype(float)
        _stk_dn = _ret31.where(_ret31 < 0).rolling(40, min_periods=5).std()
        _sec_dn = _sret.where(_sret < 0).rolling(40, min_periods=5).std()
        feat['rsk_downvol_ratio_40'] = _stk_dn / _sec_dn.replace(0, np.nan)
        feat['rsk_high_downvol_vs_sector'] = (feat['rsk_downvol_ratio_40'] > 1.3).astype(float)

    # ── 31G. 섹터 다이버전스 (종목 vs 섹터 신호 불일치) ─────────
    if _best31 is not None:
        _sec = closes[_best31]
        _stk_hh = (_c31 >= _c31.rolling(20).max() - 1e-12)
        _sec_hh = (_sec >= _sec.rolling(20).max() - 1e-12)
        feat['sdv_stock_hh_sector_not'] = (_stk_hh & ~_sec_hh).astype(float)
        feat['sdv_stock_hh_alone_10d'] = (_stk_hh & ~_sec_hh).rolling(10).sum()
        _sec_ll = (_sec <= _sec.rolling(20).min() + 1e-12)
        feat['sdv_sector_ll_stock_not'] = (_sec_ll & (_c31 > _c31.rolling(20).min() * 1.02)).astype(float)
        feat['sdv_mom_divergence_20'] = _c31.pct_change(20) - _sec.pct_change(20)
        feat['sdv_positive_div'] = (feat['sdv_mom_divergence_20'] > 0.05).astype(float)
        feat['sdv_negative_div'] = (feat['sdv_mom_divergence_20'] < -0.05).astype(float)

    # ── 31H. 신규 섹터 종합 (상승/하락) ─────────────────────────
    _dn31 = pd.Series(0.0, index=_c31.index)
    for _k in ['lag_sector_led_down', 'coh_negative_drift', 'brt_breadth_deteriorating',
               'flo_outflow_accel', 'rsk_sharpe_underperform', 'sdv_negative_div',
               'coh_high_rel_vol']:
        if _k in feat.columns: _dn31 = _dn31 + feat[_k]
    feat['snw_drop_score'] = _dn31
    feat['snw_drop_high'] = (_dn31 >= 4).astype(float)
    feat['snw_drop_rising'] = (_dn31 > _dn31.shift(3)).astype(float)
    for p in [5, 10]:
        feat[f'snw_drop_sum_{p}d'] = _dn31.rolling(p).sum()
    _up31 = pd.Series(0.0, index=_c31.index)
    for _k in ['lag_sector_led_up', 'flo_inflow_accel', 'sdv_positive_div', 'brt_thrust_up']:
        if _k in feat.columns: _up31 = _up31 + feat[_k]
    feat['snw_rise_score'] = _up31
    feat['snw_net_score'] = _up31 - _dn31

# ══════════════════════════════════════════════════════════════
    #  32. 섹터/산업 상대 — 8차 (~47개, 섹션29·30·31과도 중복없음·저상관)
    #      (종목수익률과 |corr| ~0.11 — 거의 독립)
    #      접두사: idr_(산업내상대드로다운) cvx_(공분산/잔차구조)
    #              brk2_(섹터지지붕괴) vlr_(거래량상대) rer_(상대수익레짐)
    #              ddc_(드로다운동조) tnr_(추세동조) snx_(신규종합)
    #      ※ closes + TICKER 사용
    # ══════════════════════════════════════════════════════════════
    _c32 = cl
    _ret32 = _c32.pct_change()
    SECTORS_32 = ['XLK','XLV','XLF','XLY','XLP','XLE','XLI','XLB','XLU','XLRE','XLC']
    _sec_av32 = [s for s in SECTORS_32 if s in closes.columns and s != TICKER]
    _best32 = None; _bc32 = -2
    if _sec_av32:
        for s in _sec_av32:
            c = _ret32.rolling(90).corr(closes[s].pct_change()).mean()
            if pd.notna(c) and c > _bc32: _bc32 = c; _best32 = s

    # ── 32A. 산업 내 상대 드로다운 (peer 대비 낙폭) ─────────────
    if len(_sec_av32) >= 5:
        _stk_dd = _c32 / _c32.rolling(60).max() - 1
        _peer_dd = pd.DataFrame({s: closes[s] / closes[s].rolling(60).max() - 1 for s in _sec_av32}).mean(axis=1)
        feat['idr_relative_drawdown_60'] = _stk_dd - _peer_dd           # 음수=종목이 더 깊이 빠짐
        feat['idr_deeper_dd_flag'] = (feat['idr_relative_drawdown_60'] < -0.03).astype(float)
        _stk_newlow = (_c32 <= _c32.rolling(60).min() * 1.01)
        _peer_newlow = pd.DataFrame({s: (closes[s] <= closes[s].rolling(60).min() * 1.01).astype(float)
                                     for s in _sec_av32}).mean(axis=1)
        feat['idr_stock_newlow_alone'] = (_stk_newlow & (_peer_newlow < 0.3)).astype(float)
        feat['idr_stock_newlow_alone_10d'] = (_stk_newlow & (_peer_newlow < 0.3)).rolling(10).sum()
        _stk_recov = _c32 / _c32.rolling(20).min() - 1
        _peer_recov = pd.DataFrame({s: closes[s] / closes[s].rolling(20).min() - 1 for s in _sec_av32}).mean(axis=1)
        feat['idr_weak_recovery_20'] = _stk_recov - _peer_recov
        feat['idr_lagging_recovery'] = (feat['idr_weak_recovery_20'] < -0.02).astype(float)

    # ── 32B. 공분산 구조 (베타 변화/잔차 위험) ──────────────────
    _bench32 = None
    for b in ['SPY', 'QQQ', 'RSP']:
        if b in closes.columns and b != TICKER: _bench32 = closes[b]; break
    if _bench32 is not None:
        _bret = _bench32.pct_change()
        _beta20 = _ret32.rolling(20).cov(_bret) / _bret.rolling(20).var().replace(0, np.nan)
        _beta60 = _ret32.rolling(60).cov(_bret) / _bret.rolling(60).var().replace(0, np.nan)
        feat['cvx_beta_shift_20_60'] = _beta20 - _beta60                # 양수=베타 상승(위험↑)
        feat['cvx_beta_rising'] = (_beta20 > _beta60 + 0.3).astype(float)
        _pred = _beta60 * _bret
        _resid = _ret32 - _pred
        feat['cvx_residual_vol_20'] = _resid.rolling(20).std()          # 개별(비체계적) 위험
        feat['cvx_residual_vol_zscore'] = calc_zscore(_resid.rolling(20).std(), 60)
        feat['cvx_idio_risk_spike'] = (feat['cvx_residual_vol_zscore'] > 1.5).astype(float)
        feat['cvx_resid_cumsum_20'] = _resid.rolling(20).sum()          # 알파
        feat['cvx_negative_alpha'] = (feat['cvx_resid_cumsum_20'] < -0.03).astype(float)
        _var_stk = _ret32.rolling(60).var().replace(0, np.nan)
        feat['cvx_r_squared_60'] = (_beta60 ** 2 * _bret.rolling(60).var()) / _var_stk
        feat['cvx_low_r2_flag'] = (feat['cvx_r_squared_60'] < 0.2).astype(float)

    # ── 32C. 섹터 지지선 동반 붕괴 ──────────────────────────────
    if _best32 is not None:
        _sec = closes[_best32]
        _sec_brk = (_sec < _sec.rolling(50).mean()) & (_sec.shift(1) >= _sec.rolling(50).mean().shift(1))
        feat['brk2_sector_sma50_break'] = _sec_brk.astype(float)
        feat['brk2_both_break_sma50'] = (_sec_brk & (_c32 < _c32.rolling(50).mean())).astype(float)
        feat['brk2_sector_new_low_20'] = (_sec <= _sec.rolling(20).min() * 1.005).astype(float)
        feat['brk2_sector_new_low_60'] = (_sec <= _sec.rolling(60).min() * 1.005).astype(float)
        _sec_dn = (_sec.pct_change() < 0).astype(float)
        feat['brk2_sector_down_streak'] = _sec_dn.groupby(
            (_sec_dn != _sec_dn.shift()).cumsum()).cumcount() + 1
        feat['brk2_sector_down_streak'] = feat['brk2_sector_down_streak'].where(_sec_dn > 0, 0)
        feat['brk2_sector_down_3plus'] = (feat['brk2_sector_down_streak'] >= 3).astype(float)

    # ── 32D. 거래량 상대 (섹터 대비 거래 활동) ──────────────────
    if _best32 is not None:
        _sec = closes[_best32]
        _vma20 = vo.rolling(20).mean().replace(0, np.nan)
        feat['vlr_volsurge_sector_down'] = ((vo > _vma20 * 1.5) & (_sec.pct_change(5) < 0)).astype(float)
        feat['vlr_volsurge_sector_down_10d'] = feat['vlr_volsurge_sector_down'].rolling(10).sum()
        feat['vlr_vol_up_sec_weak'] = ((vo.rolling(5).mean() > _vma20) &
                                       (_sec.pct_change(10) < -0.02)).astype(float)
        _rel_ret = _ret32 - _sec.pct_change()
        feat['vlr_vw_rel_return_10'] = (_rel_ret * vo).rolling(10).sum() / vo.rolling(10).sum().replace(0, np.nan)
        feat['vlr_vw_rel_negative'] = (feat['vlr_vw_rel_return_10'] < 0).astype(float)

    # ── 32E. 상대 수익 레짐 (상대강세/약세 지속성) ──────────────
    if _best32 is not None:
        _sec = closes[_best32]
        _rel = _c32 / _sec.replace(0, np.nan)
        _rel_sma20 = _rel.rolling(20).mean()
        _rel_sma60 = _rel.rolling(60).mean()
        feat['rer_rs_regime'] = np.sign(_rel - _rel_sma20) + np.sign(_rel_sma20 - _rel_sma60)
        feat['rer_rs_bearish'] = (feat['rer_rs_regime'] <= -2).astype(float)
        feat['rer_rs_turning_down'] = ((feat['rer_rs_regime'] < 0) &
                                       (feat['rer_rs_regime'].shift(5) >= 0)).astype(float)
        _rel_dn = (_rel < _rel.shift(1)).astype(float)
        feat['rer_rs_down_ratio_20'] = _rel_dn.rolling(20).mean()
        feat['rer_persistent_rs_weak'] = (_rel_dn.rolling(10).mean() > 0.65).astype(float)
        feat['rer_rs_volatility_20'] = _rel.pct_change().rolling(20).std()

    # ── 32F. 드로다운 동조 (섹터 동시 낙폭 = 시스템) ────────────
    if len(_sec_av32) >= 5:
        _peer_in_dd = pd.DataFrame({s: ((closes[s] / closes[s].rolling(60).max() - 1) < -0.05).astype(float)
                                    for s in _sec_av32}).mean(axis=1)
        feat['ddc_peers_in_dd_ratio'] = _peer_in_dd
        feat['ddc_widespread_dd'] = (_peer_in_dd > 0.5).astype(float)
        feat['ddc_dd_spreading'] = (_peer_in_dd > _peer_in_dd.shift(10)).astype(float)
        _stk_dd = _c32 / _c32.rolling(60).max() - 1
        feat['ddc_systemic_dd'] = ((_stk_dd < -0.05) & (_peer_in_dd > 0.4)).astype(float)

    # ── 32G. 추세 동조 (섹터 추세 vs 종목 추세) ─────────────────
    if _best32 is not None:
        _sec = closes[_best32]
        _sec_slope = (_sec.rolling(50).mean() - _sec.rolling(50).mean().shift(10))
        _stk_slope = (_c32.rolling(50).mean() - _c32.rolling(50).mean().shift(10))
        feat['tnr_both_down_trend'] = ((_sec_slope < 0) & (_stk_slope < 0)).astype(float)
        feat['tnr_sector_down_stock_up'] = ((_sec_slope < 0) & (_stk_slope > 0)).astype(float)  # 상대강세
        feat['tnr_sector_up_stock_down'] = ((_sec_slope > 0) & (_stk_slope < 0)).astype(float)  # 상대약세
        feat['tnr_sector_trend_flip_down'] = ((_sec_slope < 0) & (_sec_slope.shift(5) > 0)).astype(float)

    # ── 32H. 신규 섹터 종합 (상승/하락 양방향) ──────────────────
    _dn32 = pd.Series(0.0, index=_c32.index)
    for _k in ['idr_deeper_dd_flag', 'cvx_negative_alpha', 'brk2_both_break_sma50',
               'vlr_vw_rel_negative', 'rer_rs_bearish', 'ddc_systemic_dd', 'tnr_sector_up_stock_down']:
        if _k in feat.columns: _dn32 = _dn32 + feat[_k]
    feat['snx_drop_score'] = _dn32
    feat['snx_drop_high'] = (_dn32 >= 4).astype(float)
    feat['snx_drop_rising'] = (_dn32 > _dn32.shift(3)).astype(float)
    for p in [5, 10]:
        feat[f'snx_drop_sum_{p}d'] = _dn32.rolling(p).sum()
    _up32 = pd.Series(0.0, index=_c32.index)
    if 'tnr_sector_down_stock_up' in feat.columns:
        _up32 = _up32 + feat['tnr_sector_down_stock_up']
    if 'idr_lagging_recovery' in feat.columns:
        _up32 = _up32 + (1 - feat['idr_lagging_recovery'])
    feat['snx_rise_score'] = _up32
    feat['snx_net_score'] = _up32 - _dn32

    ##
# ══════════════════════════════════════════════════════════════
    #  33. 섹터/산업 상대 — 9차 (~51개, 섹션29~32와도 중복없음·저상관)
    #      (종목수익률과 |corr| ~0.12 — 거의 독립)
    #      접두사: dpr_(횡단면백분위) tqa_(추세품질상대) flw_(자금흐름폭)
    #              cvg_(변동성동조) gpr_(갭상대) crd_(신용/위험축)
    #              esd_(승률스트릭상대) sny_(신규종합)
    #      ※ closes + TICKER 사용
    # ══════════════════════════════════════════════════════════════
    _c33 = cl
    _ret33 = _c33.pct_change()
    SECTORS_33 = ['XLK','XLV','XLF','XLY','XLP','XLE','XLI','XLB','XLU','XLRE','XLC']
    _sec_av33 = [s for s in SECTORS_33 if s in closes.columns and s != TICKER]
    _best33 = None; _bc33 = -2
    if _sec_av33:
        for s in _sec_av33:
            c = _ret33.rolling(90).corr(closes[s].pct_change()).mean()
            if pd.notna(c) and c > _bc33: _bc33 = c; _best33 = s

    # ── 33A. 섹터 분산 내 횡단면 백분위 (cross-sectional rank) ──
    if len(_sec_av33) >= 5 and _best33 is not None:
        _all = pd.DataFrame({s: closes[s].pct_change() for s in _sec_av33})
        _all[TICKER + '_self'] = _ret33
        for p in [5, 20]:
            _cum = (1 + _all).rolling(p).apply(np.prod, raw=True) - 1
            _rank = _cum.rank(axis=1, pct=True)
            feat[f'dpr_xsec_rank_{p}d'] = _rank[TICKER + '_self']
        feat['dpr_bottom_decile'] = (feat['dpr_xsec_rank_20d'] < 0.1).astype(float)
        feat['dpr_top_decile'] = (feat['dpr_xsec_rank_20d'] > 0.9).astype(float)
        feat['dpr_rank_slope_10'] = feat['dpr_xsec_rank_20d'] - feat['dpr_xsec_rank_20d'].shift(10)
        feat['dpr_rank_deteriorating'] = (feat['dpr_rank_slope_10'] < -0.2).astype(float)
        _sec_only = pd.DataFrame({s: closes[s].pct_change(20) for s in _sec_av33})
        feat['dpr_zscore_in_dist_20'] = (_c33.pct_change(20) - _sec_only.mean(axis=1)) / _sec_only.std(axis=1).replace(0, np.nan)
        feat['dpr_outlier_low'] = (feat['dpr_zscore_in_dist_20'] < -1.5).astype(float)

    # ── 33B. 추세 품질 상대 (섹터 대비 추세 효율) ───────────────
    if _best33 is not None:
        _sec = closes[_best33]
        for p in [20]:
            _stk_eff = (_c33 - _c33.shift(p)).abs() / _c33.diff().abs().rolling(p).sum().replace(0, np.nan)
            _sec_eff = (_sec - _sec.shift(p)).abs() / _sec.diff().abs().rolling(p).sum().replace(0, np.nan)
            feat[f'tqa_efficiency_gap_{p}'] = _stk_eff - _sec_eff
        _stk_dir = np.sign(_c33 - _c33.shift(20))
        feat['tqa_weaker_downtrend'] = ((feat['tqa_efficiency_gap_20'] < 0) & (_stk_dir < 0)).astype(float)
        _rel_ret = _ret33 - _sec.pct_change()
        feat['tqa_rel_sharpe_20'] = _rel_ret.rolling(20).mean() / _rel_ret.rolling(20).std().replace(0, np.nan)
        feat['tqa_rel_sharpe_neg'] = (feat['tqa_rel_sharpe_20'] < -0.1).astype(float)
        feat['tqa_rel_consistency_20'] = (_rel_ret < 0).rolling(20).mean()
        feat['tqa_persistent_lag'] = (feat['tqa_rel_consistency_20'] > 0.65).astype(float)

    # ── 33C. 자금흐름 폭 (섹터 내 광범위 약세) ──────────────────
    if len(_sec_av33) >= 5:
        _sd = pd.DataFrame({s: closes[s] for s in _sec_av33})
        def _rsi_simple(x, p=14):
            d = x.diff()
            g = d.clip(lower=0).ewm(com=p-1, adjust=False).mean()
            l = (-d.clip(upper=0)).ewm(com=p-1, adjust=False).mean()
            return 100 - 100/(1 + g/l.replace(0, np.nan))
        _rsi_below = pd.DataFrame({s: (_rsi_simple(_sd[s]) < 50).astype(float) for s in _sec_av33}).mean(axis=1)
        feat['flw_sectors_rsi_below50_pct'] = _rsi_below
        feat['flw_broad_weakness'] = (_rsi_below > 0.6).astype(float)
        feat['flw_weakness_rising'] = (_rsi_below > _rsi_below.shift(10)).astype(float)
        _avg_mom = _sd.pct_change(20).mean(axis=1)
        feat['flw_avg_sector_mom_20'] = _avg_mom
        feat['flw_negative_breadth_mom'] = (_avg_mom < 0).astype(float)
        feat['flw_mom_accel'] = _avg_mom - _avg_mom.shift(10)
        feat['flw_broad_deceleration'] = (feat['flw_mom_accel'] < -0.02).astype(float)

    # ── 33D. 변동성 동조 (섹터 변동성과 종목 변동성) ────────────
    if _best33 is not None:
        _sec = closes[_best33]
        _stk_vol = _ret33.rolling(20).std()
        _sec_vol = _sec.pct_change().rolling(20).std()
        feat['cvg_vol_ratio_20'] = _stk_vol / _sec_vol.replace(0, np.nan)
        feat['cvg_vol_ratio_zscore'] = calc_zscore(_stk_vol / _sec_vol.replace(0, np.nan), 60)
        feat['cvg_both_vol_spike'] = ((_stk_vol > _stk_vol.rolling(60).mean() * 1.3) &
                                      (_sec_vol > _sec_vol.rolling(60).mean() * 1.3)).astype(float)
        feat['cvg_sector_vol_lead'] = ((_sec_vol > _sec_vol.rolling(60).mean() * 1.3) &
                                       (_stk_vol < _stk_vol.rolling(60).mean() * 1.1)).astype(float)
        feat['cvg_vol_corr_40'] = _stk_vol.rolling(40).corr(_sec_vol)

    # ── 33E. 갭 상대 (섹터 대비 갭 행동) ────────────────────────
    if _best33 is not None:
        _sec = closes[_best33]
        _stk_gap = (op / _c33.shift(1) - 1)
        _sec_gap = (_sec / _sec.shift(1) - 1)
        feat['gpr_stock_gap_down_alone'] = ((_stk_gap < -0.005) & (_sec_gap > -0.002)).astype(float)
        feat['gpr_stock_gap_down_alone_10d'] = feat['gpr_stock_gap_down_alone'].rolling(10).sum()
        feat['gpr_relative_gap'] = _stk_gap - _sec_gap
        feat['gpr_rel_gap_cum_10'] = feat['gpr_relative_gap'].rolling(10).sum()
        feat['gpr_rel_gap_negative'] = (feat['gpr_rel_gap_cum_10'] < -0.02).astype(float)

    # ── 33F. 신용/위험 축 (HY·신용스프레드 프록시) ──────────────
    _hyg33 = closes.get('HYG'); _lqd33 = closes.get('LQD'); _tlt33 = closes.get('TLT')
    if _hyg33 is not None and _lqd33 is not None:
        _credit = _hyg33 / _lqd33.replace(0, np.nan)
        feat['crd_hy_ig_ratio_z60'] = calc_zscore(_credit, 60)
        feat['crd_credit_risk_off'] = (_credit < _credit.rolling(20).mean()).astype(float)
        feat['crd_credit_deteriorating'] = (_credit.pct_change(10) < -0.01).astype(float)
        feat['crd_credit_warn_for_stock'] = ((_credit.pct_change(10) < -0.005) &
                                             (_ret33.rolling(10).mean() > 0)).astype(float)
    if _hyg33 is not None and _lqd33 is None:
        feat['crd_hy_trend_z60'] = calc_zscore(_hyg33, 60)
        feat['crd_hy_falling'] = (_hyg33.pct_change(10) < -0.01).astype(float)
    if _tlt33 is not None:
        feat['crd_duration_bid'] = (_tlt33.pct_change(20) > 0.03).astype(float)
        feat['crd_flight_to_safety'] = ((_tlt33.pct_change(10) > 0.02) & (_ret33.rolling(10).mean() < 0)).astype(float)

    # ── 33G. 승률/스트릭 상대 (섹터 대비 일별 승패) ─────────────
    if _best33 is not None:
        _sec = closes[_best33]
        _sret = _sec.pct_change()
        _beat = (_ret33 > _sret).astype(float)
        for p in [10, 20]:
            feat[f'esd_beat_sector_ratio_{p}'] = _beat.rolling(p).mean()
        feat['esd_losing_to_sector'] = (_beat.rolling(20).mean() < 0.4).astype(float)
        _lose = (_ret33 < _sret).astype(float)
        feat['esd_lose_streak'] = _lose.groupby((_lose != _lose.shift()).cumsum()).cumcount() + 1
        feat['esd_lose_streak'] = feat['esd_lose_streak'].where(_lose > 0, 0)
        feat['esd_lose_streak_3plus'] = (feat['esd_lose_streak'] >= 3).astype(float)
        feat['esd_beat_ratio_falling'] = (_beat.rolling(10).mean() < _beat.rolling(20).mean()).astype(float)
        feat['esd_beat_turning_up'] = ((_beat.rolling(5).mean() > 0.6) &
                                       (_beat.rolling(20).mean() < 0.45)).astype(float)

    # ── 33H. 신규 섹터 종합 (상승/하락 양방향) ──────────────────
    _dn33 = pd.Series(0.0, index=_c33.index)
    for _k in ['dpr_bottom_decile', 'tqa_persistent_lag', 'flw_broad_weakness',
               'cvg_sector_vol_lead', 'gpr_rel_gap_negative', 'crd_credit_risk_off',
               'esd_losing_to_sector']:
        if _k in feat.columns: _dn33 = _dn33 + feat[_k]
    feat['sny_drop_score'] = _dn33
    feat['sny_drop_high'] = (_dn33 >= 4).astype(float)
    feat['sny_drop_rising'] = (_dn33 > _dn33.shift(3)).astype(float)
    for p in [5, 10]:
        feat[f'sny_drop_sum_{p}d'] = _dn33.rolling(p).sum()
    _up33 = pd.Series(0.0, index=_c33.index)
    for _k in ['dpr_top_decile', 'esd_beat_turning_up']:
        if _k in feat.columns: _up33 = _up33 + feat[_k]
    feat['sny_rise_score'] = _up33
    feat['sny_net_score'] = _up33 - _dn33

# ══════════════════════════════════════════════════════════════
    #  34. 변동성레짐 적응 + 기관/개미 심리 프록시 (~74개, 중복없음·저상관)
    #      (평균 절대상관 ~0.23, 종목수익률과 |corr| ~0.14)
    #      접두사: vreg_(변동성레짐분류) hva_(고변동적응) lva_(저변동적응)
    #              inst_(기관흔적) reta_(개미심리) sent_(파생/공포탐욕)
    #              smt_(스마트머니) pmix_(심리종합)
    #      ※ closes + TICKER 사용 (^VIX/^VVIX 있으면 파생심리 활성)
    # ══════════════════════════════════════════════════════════════
    _o34 = op; _h34 = hi; _l34 = lo; _c34 = cl; _v34 = vo
    _rng34  = (_h34 - _l34).replace(0, np.nan)
    _pc34   = _c34.shift(1)
    _ret34  = _c34.pct_change()
    _vma20_34 = _v34.rolling(20).mean().replace(0, np.nan)
    _atr_pct34 = ((_h34 - _l34) / _c34).rolling(14).mean()
    _close_loc34 = (_c34 - _l34) / _rng34

    # ── 34A. 변동성 레짐 분류 (종목 타입 자동 식별) ─────────────
    _vol60_34 = _ret34.rolling(60).std()
    _vol_pctrank34 = _vol60_34.rolling(250, min_periods=60).apply(
        lambda x: pd.Series(x).rank(pct=True).iloc[-1], raw=False)
    feat['vreg_vol_level_pctrank'] = _vol_pctrank34
    feat['vreg_is_high_vol'] = (_vol_pctrank34 > 0.7).astype(float)        # 고변동 종목 국면
    feat['vreg_is_low_vol'] = (_vol_pctrank34 < 0.3).astype(float)         # 저변동 종목 국면
    feat['vreg_vol_expanding'] = (_vol60_34 > _vol60_34.shift(20) * 1.3).astype(float)
    feat['vreg_vol_contracting'] = (_vol60_34 < _vol60_34.shift(20) * 0.7).astype(float)
    feat['vreg_vol_zscore_120'] = calc_zscore(_vol60_34, 120)
    _hv34 = (_vol_pctrank34 > 0.6).astype(float)
    feat['vreg_high_vol_persist'] = _hv34.rolling(20).sum()
    feat['vreg_atr_pct_level'] = _atr_pct34
    feat['vreg_atr_pct_zscore'] = calc_zscore(_atr_pct34, 120)

    # ── 34B. 고변동성 종목 적응 규칙 (ATR 정규화 신호) ──────────
    _atr14_34 = _rng34.ewm(com=13, adjust=False).mean()
    _ret_in_atr34 = (_c34 - _pc34) / _atr14_34.replace(0, np.nan)
    feat['hva_ret_in_atr'] = _ret_in_atr34
    feat['hva_big_drop_atr'] = (_ret_in_atr34 < -1.5).astype(float)        # ATR 1.5배 하락
    feat['hva_big_drop_atr_10d'] = (_ret_in_atr34 < -1.5).rolling(10).sum()
    feat['hva_trend_confirmed_dn'] = ((_c34 < _c34.rolling(30).mean()) &
                                      (_c34.rolling(5).mean() < _c34.rolling(30).mean())).astype(float)
    for p in [10, 20]:
        feat[f'hva_vol_adj_mom_{p}'] = _c34.pct_change(p) / (_vol60_34 * np.sqrt(p)).replace(0, np.nan)
    feat['hva_vol_blowup_down'] = ((_vol60_34 > _vol60_34.rolling(60).mean() * 1.5) &
                                   (_ret34.rolling(5).mean() < 0)).astype(float)
    feat['hva_vol_cluster'] = (_ret34.abs() > _ret34.abs().rolling(20).mean() * 2).rolling(5).sum()

    # ── 34C. 저변동성 종목 적응 규칙 (작은 변화도 신호) ─────────
    _small_z34 = (_c34 - _c34.rolling(20).mean()) / _c34.rolling(20).std().replace(0, np.nan)
    feat['lva_small_breakdown'] = (_small_z34 < -1.0).astype(float)        # 저변동엔 1σ도 신호
    feat['lva_quiet_then_drop'] = ((_vol60_34 < _vol60_34.rolling(120).median()) &
                                   (_ret34 < -_vol60_34)).astype(float)
    feat['lva_first_crack'] = ((_atr_pct34.shift(1) < _atr_pct34.rolling(60).median()) &
                               (_rng34 / _c34 > _atr_pct34.rolling(60).median() * 2)).astype(float)
    feat['lva_volume_awakening'] = ((_vol60_34 < _vol60_34.rolling(120).median()) &
                                    (_v34 > _vma20_34 * 2)).astype(float)
    feat['lva_volume_awakening_10d'] = feat['lva_volume_awakening'].rolling(10).sum()
    _bb_width34 = (_c34.rolling(20).std() * 4) / _c34.rolling(20).mean()
    feat['lva_squeeze_extreme'] = (_bb_width34 < _bb_width34.rolling(120).quantile(0.1)).astype(float)
    feat['lva_squeeze_then_dn'] = (feat['lva_squeeze_extreme'].shift(1).astype(bool) & (_ret34 < 0)).astype(float)

    # ── 34D. 기관 흔적 (대량거래/체결 패턴 추정) ────────────────
    feat['inst_absorption'] = ((_v34 > _vma20_34 * 1.5) & (_ret34.abs() < _atr_pct34 * 0.5)).astype(float)
    feat['inst_absorption_10d'] = feat['inst_absorption'].rolling(10).sum()
    feat['inst_buy_close_heavy'] = ((_v34 > _vma20_34 * 1.3) & (_close_loc34 > 0.7)).astype(float)
    feat['inst_sell_close_heavy'] = ((_v34 > _vma20_34 * 1.3) & (_close_loc34 < 0.3)).astype(float)
    feat['inst_net_close_pressure_10d'] = (feat['inst_buy_close_heavy'].rolling(10).sum() -
                                           feat['inst_sell_close_heavy'].rolling(10).sum())
    feat['inst_block_trade'] = (_v34 > _v34.rolling(60).mean() + _v34.rolling(60).std() * 2).astype(float)
    feat['inst_block_down_10d'] = ((feat['inst_block_trade'] > 0) & (_ret34 < 0)).rolling(10).sum()
    _heavy_signed34 = np.where(_v34 > _vma20_34 * 1.5, np.sign(_ret34), 0)
    feat['inst_smart_flow_20d'] = pd.Series(_heavy_signed34, index=_c34.index).rolling(20).sum()
    feat['inst_distribution_flag'] = (feat['inst_smart_flow_20d'] < -3).astype(float)
    _vwap34 = (_c34 * _v34).rolling(20).sum() / _v34.rolling(20).sum().replace(0, np.nan)
    feat['inst_above_vwap_streak'] = ((_c34 > _vwap34).astype(float).groupby(
        ((_c34 > _vwap34) != (_c34 > _vwap34).shift()).cumsum()).cumcount() + 1).where(_c34 > _vwap34, 0)

    # ── 34E. 개미(리테일) 심리 프록시 ───────────────────────────
    _gap34 = (_o34 / _pc34 - 1)
    feat['reta_fomo_chase'] = ((_gap34 > 0.02) & (_v34 > _vma20_34 * 1.5)).astype(float)
    feat['reta_fomo_chase_10d'] = feat['reta_fomo_chase'].rolling(10).sum()
    feat['reta_buy_exhaustion'] = ((_c34.shift(1) / _c34.shift(2) - 1 > 0.03) &
                                   (_v34.shift(1) > _vma20_34 * 1.5) & (_ret34 < 0)).astype(float)
    feat['reta_panic_sell'] = ((_ret34 < -0.03) & (_v34 > _vma20_34 * 2) & (_close_loc34 < 0.3)).astype(float)
    feat['reta_panic_sell_20d'] = feat['reta_panic_sell'].rolling(20).sum()
    feat['reta_capitulation_bounce'] = (feat['reta_panic_sell'].shift(1).astype(bool) & (_ret34 > 0)).astype(float)
    feat['reta_overheated_20d'] = ((_gap34 > 0.01) & (_v34 > _vma20_34 * 1.3)).rolling(20).sum()
    feat['reta_euphoria_flag'] = (feat['reta_overheated_20d'] > 5).astype(float)
    _small_up34 = ((_ret34 > 0) & (_ret34 < 0.01)).astype(float)
    feat['reta_retail_dca_20d'] = _small_up34.rolling(20).sum()

    # ── 34F. 파생/공포탐욕 (VIX 등 시장 심리) ───────────────────
    _vix34 = closes.get('^VIX')
    if _vix34 is not None:
        feat['sent_vix_level'] = _vix34
        feat['sent_vix_zscore_60'] = calc_zscore(_vix34, 60)
        feat['sent_vix_spike'] = (_vix34 > _vix34.rolling(20).mean() + _vix34.rolling(20).std() * 1.5).astype(float)
        feat['sent_fear_regime'] = (_vix34 > 25).astype(float)
        feat['sent_extreme_fear'] = (_vix34 > 35).astype(float)
        feat['sent_complacency'] = (_vix34 < 14).astype(float)
        feat['sent_complacency_at_high'] = ((_vix34 < 15) & (_c34 > _c34.rolling(60).max() * 0.97)).astype(float)
        feat['sent_vix_roc_5'] = _vix34.pct_change(5)
        feat['sent_vix_surging'] = (_vix34.pct_change(5) > 0.2).astype(float)
        feat['sent_vix_stock_both_up'] = ((_vix34.pct_change(5) > 0.05) & (_c34.pct_change(5) > 0)).astype(float)
    _vvix34 = closes.get('^VVIX') if closes.get('^VVIX') is not None else closes.get('VVIX')
    if _vvix34 is not None:
        feat['sent_vvix_zscore'] = calc_zscore(_vvix34, 60)
        feat['sent_vol_uncertainty'] = (_vvix34 > _vvix34.rolling(60).mean() * 1.2).astype(float)
    if _vix34 is not None:
        _realized34 = _ret34.rolling(20).std() * np.sqrt(252) * 100
        feat['sent_iv_rv_ratio'] = _vix34 / _realized34.replace(0, np.nan)
        feat['sent_hedging_demand'] = (feat['sent_iv_rv_ratio'] > 1.3).astype(float)
        feat['sent_low_hedge_complacent'] = (feat['sent_iv_rv_ratio'] < 0.9).astype(float)

    # ── 34G. 스마트머니 vs 개미 (장중 vs 오버나이트) ────────────
    _open_ret34 = (_o34 - _pc34) / _pc34.replace(0, np.nan)      # 오버나이트(갭, 개미 영향)
    _close_ret34 = (_c34 - _o34) / _o34.replace(0, np.nan)       # 장중(스마트머니 가설)
    feat['smt_intraday_strength_10'] = _close_ret34.rolling(10).mean()
    feat['smt_overnight_strength_10'] = _open_ret34.rolling(10).mean()
    feat['smt_smart_dumb_spread_20'] = (_close_ret34.rolling(20).sum() - _open_ret34.rolling(20).sum())
    feat['smt_smart_money_selling'] = (feat['smt_intraday_strength_10'] < 0).astype(float)
    feat['smt_dumb_buy_smart_sell'] = ((_open_ret34 > 0) & (_close_ret34 < 0)).astype(float)
    feat['smt_dumb_buy_smart_sell_10d'] = feat['smt_dumb_buy_smart_sell'].rolling(10).sum()
    feat['smt_smart_accumulation'] = ((_open_ret34 < 0) & (_close_ret34 > 0)).astype(float)
    feat['smt_smart_accum_10d'] = feat['smt_smart_accumulation'].rolling(10).sum()

    # ── 34H. 심리 종합 (상승/하락 양방향) ───────────────────────
    _dn34 = pd.Series(0.0, index=_c34.index)
    for _k in ['inst_distribution_flag', 'reta_euphoria_flag', 'reta_buy_exhaustion',
               'sent_complacency_at_high', 'smt_smart_money_selling', 'hva_vol_blowup_down',
               'inst_sell_close_heavy']:
        if _k in feat.columns: _dn34 = _dn34 + feat[_k]
    feat['pmix_drop_score'] = _dn34
    feat['pmix_drop_high'] = (_dn34 >= 4).astype(float)
    feat['pmix_drop_rising'] = (_dn34 > _dn34.shift(3)).astype(float)
    for p in [5, 10]:
        feat[f'pmix_drop_sum_{p}d'] = _dn34.rolling(p).sum()
    _up34 = pd.Series(0.0, index=_c34.index)
    for _k in ['smt_smart_accumulation', 'reta_capitulation_bounce', 'inst_buy_close_heavy']:
        if _k in feat.columns: _up34 = _up34 + feat[_k]
    feat['pmix_rise_score'] = _up34
    feat['pmix_net_score'] = _up34 - _dn34
    feat['pmix_capitulation_zone'] = ((_dn34 >= 3) & (feat.get('reta_panic_sell_20d',
                                       pd.Series(0.0, index=_c34.index)) >= 2)).astype(float)

# ══════════════════════════════════════════════════════════════
    #  35. 변동성레짐 적응 + 기관/개미 심리 — 2차 확장 (~67개, 섹션34와 중복없음)
    #      (평균 절대상관 ~0.28, 종목수익률과 |corr| ~0.10 — 거의 독립)
    #      접두사: vrt_(변동성타입별룰) bpc_(돌파/거짓돌파) acc_(매집/분산 Wyckoff)
    #              ord_(주문흐름프록시) crw_(군중쏠림) liq_(유동성심리)
    #              flt_(자금이탈) psc_(심리종합v2)
    #      ※ closes + TICKER 사용
    # ══════════════════════════════════════════════════════════════
    _o35 = op; _h35 = hi; _l35 = lo; _c35 = cl; _v35 = vo
    _rng35  = (_h35 - _l35).replace(0, np.nan)
    _pc35   = _c35.shift(1)
    _ret35  = _c35.pct_change()
    _vma20_35 = _v35.rolling(20).mean().replace(0, np.nan)
    _vma50_35 = _v35.rolling(50).mean().replace(0, np.nan)
    _close_loc35 = (_c35 - _l35) / _rng35
    _atr14_35 = _rng35.ewm(com=13, adjust=False).mean()
    _vol60_35 = _ret35.rolling(60).std()
    _vol_rank35 = _vol60_35.rolling(250, min_periods=60).apply(
        lambda x: pd.Series(x).rank(pct=True).iloc[-1], raw=False)

    # ── 35A. 변동성 타입별 차별화 규칙 (레짐 조건부 신호) ───────
    _zc35 = (_c35 - _c35.rolling(20).mean()) / _c35.rolling(20).std().replace(0, np.nan)
    feat['vrt_hv_oversold'] = ((_vol_rank35 > 0.6) & (_zc35 < -2.0)).astype(float)       # 고변동 과매도 반등
    feat['vrt_lv_trend_break'] = ((_vol_rank35 < 0.4) & (_c35 < _c35.rolling(50).mean()) &
                                  (_pc35 >= _c35.rolling(50).mean().shift(1))).astype(float)  # 저변동 추세이탈
    feat['vrt_regime_flip_down'] = ((_vol_rank35 > 0.5) & (_vol_rank35.shift(20) < 0.3) &
                                    (_ret35.rolling(5).mean() < 0)).astype(float)
    for p in [20, 60]:
        feat[f'vrt_return_per_risk_{p}'] = _c35.pct_change(p) / (_vol60_35 * np.sqrt(p)).replace(0, np.nan)
    _dd35 = _c35 / _c35.rolling(60).max() - 1
    feat['vrt_vol_norm_drawdown'] = _dd35 / (_vol60_35 * np.sqrt(60)).replace(0, np.nan)
    feat['vrt_vol_norm_dd_extreme'] = (feat['vrt_vol_norm_drawdown'] < -1.0).astype(float)
    feat['vrt_hv_vol_peak'] = ((_vol_rank35 > 0.7) & (_vol60_35 < _vol60_35.shift(3)) &
                               (_vol60_35.shift(3) > _vol60_35.rolling(60).mean() * 1.5)).astype(float)
    _dn_vol35 = _ret35.where(_ret35 < 0).rolling(40, min_periods=5).std()
    _up_vol35 = _ret35.where(_ret35 > 0).rolling(40, min_periods=5).std()
    feat['vrt_leverage_effect'] = _dn_vol35 / _up_vol35.replace(0, np.nan)               # 하락시 변동성↑
    feat['vrt_high_leverage_eff'] = (feat['vrt_leverage_effect'] > 1.4).astype(float)

    # ── 35B. 돌파/거짓돌파 (개미가 당하는 패턴) ─────────────────
    _hi20_35 = _h35.rolling(20).max()
    feat['bpc_false_breakout_up'] = ((_h35 > _hi20_35.shift(1)) & (_c35 < _hi20_35.shift(1)) &
                                     (_c35 < _o35)).astype(float)
    feat['bpc_false_breakout_10d'] = feat['bpc_false_breakout_up'].rolling(10).sum()
    _lo20_35 = _l35.rolling(20).min()
    feat['bpc_false_breakdown'] = ((_l35 < _lo20_35.shift(1)) & (_c35 > _lo20_35.shift(1)) &
                                   (_c35 > _o35)).astype(float)
    feat['bpc_breakout_no_volume'] = ((_c35 > _hi20_35.shift(1)) & (_v35 < _vma20_35)).astype(float)
    feat['bpc_resistance_reject_20d'] = ((_h35 >= _hi20_35.shift(1) * 0.99) & (_c35 < _o35)).rolling(20).sum()
    feat['bpc_newhigh_reversal'] = ((_c35.shift(1) >= _c35.rolling(60).max().shift(1) - 1e-9) &
                                    (_ret35 < -0.01)).astype(float)
    _gap35 = (_o35 / _pc35 - 1)
    feat['bpc_gap_down_continuation'] = ((_gap35 < -0.01) & (_c35 < _o35) & (_h35 < _pc35)).astype(float)

    # ── 35C. 매집/분산 정밀 (Wyckoff 단계 프록시) ───────────────
    feat['acc_accumulation'] = ((_ret35.rolling(10).mean().abs() < _vol60_35 * 0.3) &
                                (_v35.rolling(10).mean() < _vma50_35 * 0.9) &
                                (_c35 < _c35.rolling(60).mean())).astype(float)
    feat['acc_accumulation_20d'] = feat['acc_accumulation'].rolling(20).sum()
    feat['acc_distribution'] = ((_ret35.rolling(10).mean().abs() < _vol60_35 * 0.3) &
                                (_v35.rolling(10).mean() > _vma50_35 * 1.1) &
                                (_c35 > _c35.rolling(60).mean())).astype(float)
    feat['acc_distribution_20d'] = feat['acc_distribution'].rolling(20).sum()
    feat['acc_spring'] = ((_l35 < _l35.rolling(30).min().shift(1)) & (_c35 > _l35.rolling(30).min().shift(1)) &
                          (_close_loc35 > 0.6)).astype(float)
    feat['acc_upthrust'] = ((_h35 > _h35.rolling(30).max().shift(1)) & (_c35 < _h35.rolling(30).max().shift(1)) &
                            (_close_loc35 < 0.4)).astype(float)
    feat['acc_upthrust_10d'] = feat['acc_upthrust'].rolling(10).sum()
    _effort_result35 = (_ret35.abs()) / ((_v35 / _vma20_35).replace(0, np.nan))
    feat['acc_effort_no_result'] = (_effort_result35 < _effort_result35.rolling(60).quantile(0.2)).astype(float)
    feat['acc_selling_absorbed'] = ((_v35 > _vma20_35 * 1.5) & (_ret35 > -0.005) & (_close_loc35 > 0.5)).astype(float)

    # ── 35D. 주문흐름 프록시 (체결 방향 추정) ───────────────────
    _mid35 = (_h35 + _l35) / 2
    _tick_dir35 = np.sign(_c35 - _mid35)
    feat['ord_tick_pressure_10'] = pd.Series(_tick_dir35, index=_c35.index).rolling(10).mean()
    feat['ord_sell_pressure_dominant'] = (feat['ord_tick_pressure_10'] < -0.3).astype(float)
    _vw_dir35 = _tick_dir35 * (_v35 / _vma20_35)
    feat['ord_vw_flow_20'] = pd.Series(_vw_dir35, index=_c35.index).rolling(20).sum()
    feat['ord_vw_flow_negative'] = (feat['ord_vw_flow_20'] < 0).astype(float)
    _delta35 = (_c35 - _o35) / _rng35 * _v35
    feat['ord_cum_delta_20'] = _delta35.rolling(20).sum() / _v35.rolling(20).sum().replace(0, np.nan)
    feat['ord_delta_falling'] = (feat['ord_cum_delta_20'] < feat['ord_cum_delta_20'].shift(5)).astype(float)
    feat['ord_buy_into_resistance'] = ((feat['ord_tick_pressure_10'] > 0.2) &
                                       (_c35.pct_change(10) < 0.01)).astype(float)
    feat['ord_closing_strength_10'] = _close_loc35.rolling(10).mean()
    feat['ord_weak_closing_trend'] = (_close_loc35.rolling(5).mean() < _close_loc35.rolling(20).mean()).astype(float)

    # ── 35E. 군중 쏠림 (과도한 한쪽 = 역방향 위험) ──────────────
    _up_days35 = (_ret35 > 0).rolling(10).sum()
    feat['crw_one_sided_up'] = (_up_days35 >= 8).astype(float)
    feat['crw_one_sided_down'] = (_up_days35 <= 2).astype(float)
    _rsi35 = calc_rsi(_c35, 14)
    feat['crw_rsi_extreme_high_5d'] = (_rsi35 > 70).rolling(5).sum()
    feat['crw_rsi_extreme_low_5d'] = (_rsi35 < 30).rolling(5).sum()
    feat['crw_rsi_stuck_high'] = ((_rsi35 > 70).rolling(5).sum() >= 4).astype(float)
    feat['crw_euphoric_volume'] = ((_ret35 > 0.03) & (_v35 > _vma20_35 * 2)).astype(float)
    feat['crw_euphoric_20d'] = feat['crw_euphoric_volume'].rolling(20).sum()
    feat['crw_optimism_crack'] = ((_up_days35.shift(1) >= 8) & (_ret35 < 0)).astype(float)
    feat['crw_extreme_move'] = (_ret35.abs() > _vol60_35 * 3).astype(float)
    feat['crw_extreme_move_20d'] = feat['crw_extreme_move'].rolling(20).sum()

    # ── 35F. 유동성 심리 (거래 활발도와 심리) ───────────────────
    _dollar35 = _c35 * _v35
    feat['liq_dollar_vol_zscore'] = calc_zscore(_dollar35, 60)
    feat['liq_interest_surge'] = (_dollar35 > _dollar35.rolling(60).mean() * 2).astype(float)
    feat['liq_interest_fading'] = (_dollar35.rolling(10).mean() < _dollar35.rolling(60).mean() * 0.7).astype(float)
    feat['liq_rally_no_interest'] = ((_c35.pct_change(10) > 0.05) & (_v35.rolling(10).mean() < _vma50_35)).astype(float)
    feat['liq_volume_breakout'] = ((_v35 > _vma20_35 * 2.5) &
                                   (_v35.shift(1).rolling(10).mean() < _vma50_35 * 0.8)).astype(float)
    feat['liq_thin_drop'] = ((_ret35 < -0.02) & (_v35 < _vma20_35 * 0.7)).astype(float)
    feat['liq_thin_drop_10d'] = feat['liq_thin_drop'].rolling(10).sum()

    # ── 35G. 자금 이탈 (스마트머니 발 빼기 정밀) ────────────────
    _obv35 = (np.sign(_c35.diff()).fillna(0) * _v35).cumsum()
    _price_hh35 = (_c35 >= _c35.rolling(20).max() - 1e-9)
    _obv_hh35 = (_obv35 >= _obv35.rolling(20).max() - 1e-9)
    feat['flt_obv_divergence'] = (_price_hh35 & ~_obv_hh35).astype(float)
    feat['flt_obv_divergence_20d'] = feat['flt_obv_divergence'].rolling(20).sum()
    _vwap50_35 = (_c35 * _v35).rolling(50).sum() / _v35.rolling(50).sum().replace(0, np.nan)
    feat['flt_below_vwap50'] = (_c35 < _vwap50_35).astype(float)
    feat['flt_vwap50_breakdown'] = ((_c35 < _vwap50_35) & (_pc35 >= _vwap50_35.shift(1))).astype(float)
    _up_vol_avg35 = _v35.where(_ret35 > 0).rolling(20, min_periods=3).mean()
    _dn_vol_avg35 = _v35.where(_ret35 < 0).rolling(20, min_periods=3).mean()
    feat['flt_vol_flow_bearish'] = (_dn_vol_avg35 > _up_vol_avg35 * 1.2).astype(float)
    feat['flt_heavy_red_20d'] = ((_v35 > _vma20_35 * 1.3) & (_ret35 < -0.01)).rolling(20).sum()
    _smart_flow35 = pd.Series(np.where(_v35 > _vma20_35 * 1.3, np.sign(_ret35), 0),
                              index=_c35.index).rolling(20).sum()
    feat['flt_flow_turning_neg'] = ((_smart_flow35 < 0) & (_smart_flow35.shift(5) > 0)).astype(float)

    # ── 35H. 심리 종합 v2 (상승/하락 양방향) ────────────────────
    _dn35 = pd.Series(0.0, index=_c35.index)
    for _k in ['bpc_false_breakout_up', 'acc_distribution', 'acc_upthrust', 'ord_vw_flow_negative',
               'crw_one_sided_up', 'crw_optimism_crack', 'liq_rally_no_interest',
               'flt_obv_divergence', 'flt_vol_flow_bearish']:
        if _k in feat.columns: _dn35 = _dn35 + feat[_k]
    feat['psc_drop_score'] = _dn35
    feat['psc_drop_high'] = (_dn35 >= 5).astype(float)
    feat['psc_drop_rising'] = (_dn35 > _dn35.shift(3)).astype(float)
    for p in [5, 10]:
        feat[f'psc_drop_sum_{p}d'] = _dn35.rolling(p).sum()
    _up35 = pd.Series(0.0, index=_c35.index)
    for _k in ['bpc_false_breakdown', 'acc_accumulation', 'acc_spring', 'acc_selling_absorbed',
               'crw_one_sided_down']:
        if _k in feat.columns: _up35 = _up35 + feat[_k]
    feat['psc_rise_score'] = _up35
    feat['psc_net_score'] = _up35 - _dn35
    feat['psc_reversal_up_zone'] = ((feat['vrt_hv_oversold'] > 0) & (_up35 >= 2)).astype(float)

    # ══════════════════════════════════════════════════════════════
    #  36. 변동성레짐 적응 + 기관/개미 심리 — 3차 (~65개, 섹션34·35와 중복없음)
    #      (평균 절대상관 ~0.19, 종목수익률과 |corr| ~0.11 — 거의 독립)
    #      접두사: vbk_(변동성분해) skn_(왜도/콜백) trp_(개미트랩정밀)
    #              whl_(고래/대량흔적) pnc_(패닉/항복) cvd_(누적델타심화)
    #              ovn_(오버나이트심리) msc_(심리종합v3)
    #      ※ closes + TICKER 사용
    # ══════════════════════════════════════════════════════════════
    _o36 = op; _h36 = hi; _l36 = lo; _c36 = cl; _v36 = vo
    _rng36  = (_h36 - _l36).replace(0, np.nan)
    _pc36   = _c36.shift(1)
    _ret36  = _c36.pct_change()
    _logret36 = np.log(_c36 / _pc36)
    _vma20_36 = _v36.rolling(20).mean().replace(0, np.nan)
    _close_loc36 = (_c36 - _l36) / _rng36
    _vol60_36 = _ret36.rolling(60).std()
    _vol_rank36 = _vol60_36.rolling(250, min_periods=60).apply(
        lambda x: pd.Series(x).rank(pct=True).iloc[-1], raw=False)

    # ── 36A. 변동성 분해 (점프 vs 연속 / 일중 vs 야간) ──────────
    _jump36 = _logret36.where(_logret36.abs() > _logret36.rolling(60).std() * 2.5, 0.0)
    feat['vbk_jump_var_20'] = (_jump36 ** 2).rolling(20).sum()
    feat['vbk_jump_ratio_20'] = ((_jump36 ** 2).rolling(20).sum() /
                                 (_logret36 ** 2).rolling(20).sum().replace(0, np.nan))
    feat['vbk_jump_dominant'] = (feat['vbk_jump_ratio_20'] > 0.4).astype(float)
    _neg_jump36 = _logret36.where((_logret36 < 0) & (_logret36.abs() > _logret36.rolling(60).std() * 2.5), 0.0)
    feat['vbk_neg_jump_20'] = (_neg_jump36 ** 2).rolling(20).sum()
    feat['vbk_neg_jump_recent'] = (_neg_jump36.rolling(5).sum() < 0).astype(float)
    _intraday36 = np.log(_c36 / _o36)
    _overnight36 = np.log(_o36 / _pc36)
    feat['vbk_intraday_var_20'] = (_intraday36 ** 2).rolling(20).mean()
    feat['vbk_overnight_var_20'] = (_overnight36 ** 2).rolling(20).mean()
    feat['vbk_overnight_risk_ratio'] = ((_overnight36 ** 2).rolling(20).mean() /
                                        (_intraday36 ** 2).rolling(20).mean().replace(0, np.nan))
    feat['vbk_overnight_heavy'] = (feat['vbk_overnight_risk_ratio'] > 1.5).astype(float)
    _vol20_36 = _ret36.rolling(20).std()
    feat['vbk_vol_accel'] = (_vol20_36 - _vol20_36.shift(5)) - (_vol20_36.shift(5) - _vol20_36.shift(10))
    feat['vbk_vol_accelerating'] = (feat['vbk_vol_accel'] > 0).astype(float)
    _parkinson36 = np.sqrt((np.log(_h36 / _l36) ** 2).rolling(20).mean() / (4 * np.log(2)))
    feat['vbk_parkinson_vs_close'] = _parkinson36 / _vol20_36.replace(0, np.nan)
    feat['vbk_intraday_excess'] = (feat['vbk_parkinson_vs_close'] > 1.5).astype(float)

    # ── 36B. 왜도/콜백 (분포 비대칭 동역학) ─────────────────────
    _skew20_36 = _logret36.rolling(20).skew()
    feat['skn_skew_20'] = _skew20_36
    feat['skn_skew_turning_neg'] = ((_skew20_36 < 0) & (_skew20_36.shift(3) > 0.3)).astype(float)
    feat['skn_hv_neg_skew'] = ((_vol_rank36 > 0.6) & (_skew20_36 < -0.5)).astype(float)
    _pullback36 = (_c36.rolling(10).max() - _c36) / _c36.rolling(10).max()
    feat['skn_pullback_depth_10'] = _pullback36
    feat['skn_pullback_zscore'] = calc_zscore(_pullback36, 60)
    feat['skn_deep_pullback'] = (calc_zscore(_pullback36, 60) > 1.5).astype(float)
    _up_capture36 = _ret36.where(_ret36 > 0).rolling(40, min_periods=5).mean()
    _dn_capture36 = _ret36.where(_ret36 < 0).rolling(40, min_periods=5).mean().abs()
    feat['skn_capture_asym'] = _dn_capture36 / _up_capture36.replace(0, np.nan)
    feat['skn_bad_capture'] = (feat['skn_capture_asym'] > 1.3).astype(float)
    feat['skn_median_ret_neg_20'] = (_ret36.rolling(20).median() < 0).astype(float)

    # ── 36C. 개미 트랩 정밀 (행동재무 패턴) ─────────────────────
    _prior_high36 = _c36.rolling(60).max().shift(5)
    feat['trp_overhead_supply'] = ((_c36 > _prior_high36 * 0.97) & (_c36 < _prior_high36 * 1.03) &
                                   (_c36 < _o36)).astype(float)
    feat['trp_overhead_supply_10d'] = feat['trp_overhead_supply'].rolling(10).sum()
    feat['trp_stoploss_hunt'] = ((_l36 < _l36.rolling(20).min().shift(1)) &
                                 (_close_loc36 > 0.6) & (_ret36 > 0)).astype(float)
    feat['trp_momo_trap'] = ((_c36.shift(1) > _c36.rolling(20).max().shift(2)) &
                             (_ret36 < -0.015)).astype(float)
    feat['trp_falling_knife'] = ((_c36 < _c36.rolling(60).min().shift(1)) &
                                 (_ret36.rolling(3).sum() < -0.05)).astype(float)
    feat['trp_falling_knife_10d'] = feat['trp_falling_knife'].rolling(10).sum()
    feat['trp_dead_cat'] = ((_ret36.shift(2) < -0.04) & (_ret36.shift(1) > 0.01) & (_ret36 < 0)).astype(float)
    feat['trp_premature_exit'] = ((_ret36.rolling(5).apply(lambda x: (x > 0).sum(), raw=True) >= 4) &
                                  (_ret36.rolling(5).mean() < 0.005) & (_ret36 > 0.02)).astype(float)

    # ── 36D. 고래/대량 흔적 (대형 주문 추정) ────────────────────
    _vol_consistency36 = 1 - (_v36.rolling(5).std() / _v36.rolling(5).mean().replace(0, np.nan))
    feat['whl_iceberg_proxy'] = ((_vol_consistency36 > 0.7) & (_v36 > _vma20_36 * 1.2) &
                                 (_ret36.abs() < _vol60_36)).astype(float)
    _whale_day36 = (_v36 > _v36.rolling(60).mean() + _v36.rolling(60).std() * 2)
    feat['whl_whale_buy'] = (_whale_day36 & (_close_loc36 > 0.6) & (_ret36 > 0)).astype(float)
    feat['whl_whale_sell'] = (_whale_day36 & (_close_loc36 < 0.4) & (_ret36 < 0)).astype(float)
    feat['whl_net_whale_20d'] = (feat['whl_whale_buy'].rolling(20).sum() -
                                 feat['whl_whale_sell'].rolling(20).sum())
    feat['whl_follow_through'] = ((_v36.shift(1) > _vma20_36 * 2) &
                                  (np.sign(_ret36) == np.sign(_ret36.shift(1)))).astype(float)
    _dollar36 = _c36 * _v36
    feat['whl_stealth_accum'] = ((_dollar36.rolling(20).mean() > _dollar36.rolling(60).mean() * 1.2) &
                                 (_c36.pct_change(20).abs() < 0.03)).astype(float)
    feat['whl_support_break_volume'] = ((_v36 > _vma20_36 * 1.8) & (_ret36 < -0.02) &
                                        (_close_loc36 < 0.25)).astype(float)

    # ── 36E. 패닉/항복 단계 (바닥 형성 프록시) ──────────────────
    feat['pnc_capitulation_vol'] = ((_v36 > _v36.rolling(60).mean() * 3) & (_ret36 < -0.04)).astype(float)
    feat['pnc_capitulation_20d'] = feat['pnc_capitulation_vol'].rolling(20).sum()
    feat['pnc_selling_exhausted'] = ((_ret36.rolling(5).sum() < -0.08) & (_v36 < _vma20_36 * 0.7)).astype(float)
    _dn_streak36 = (_ret36 < 0).astype(float)
    _dn_run36 = _dn_streak36.groupby((_dn_streak36 != _dn_streak36.shift()).cumsum()).cumcount() + 1
    _dn_run36 = _dn_run36.where(_dn_streak36 > 0, 0)
    feat['pnc_long_down_streak'] = (_dn_run36 >= 5).astype(float)
    feat['pnc_streak_exhaustion'] = ((_dn_run36.shift(1) >= 4) & (_ret36 > 0)).astype(float)
    _rsi36 = calc_rsi(_c36, 14)
    feat['pnc_extreme_fear'] = ((_rsi36 < 25) & (_v36 > _vma20_36 * 1.5)).astype(float)
    feat['pnc_cascade_5d'] = (_ret36.rolling(5).sum() < -0.10).astype(float)
    feat['pnc_cascade_intensity'] = (-_ret36.rolling(5).sum()).clip(lower=0)

    # ── 36F. 누적 델타 심화 (매수/매도 압력 정밀) ───────────────
    _delta36 = ((_c36 - _l36) - (_h36 - _c36)) / _rng36 * _v36
    feat['cvd_cum_delta_zscore'] = calc_zscore(_delta36.rolling(10).sum(), 60)
    feat['cvd_bearish_divergence'] = ((_c36.pct_change(10) > 0.02) &
                                      (_delta36.rolling(10).sum() < 0)).astype(float)
    _delta_ma36 = _delta36.rolling(10).mean()
    feat['cvd_delta_flip_neg'] = ((_delta_ma36 < 0) & (_delta_ma36.shift(3) > 0)).astype(float)
    feat['cvd_failed_absorption'] = ((_delta36.rolling(5).sum() > 0) &
                                     (_c36.pct_change(5) < -0.01)).astype(float)
    feat['cvd_persistent_selling'] = (_delta36.rolling(20).apply(lambda x: (x < 0).sum(), raw=True) >= 13).astype(float)

    # ── 36G. 오버나이트 심리 (갭 행동 = 정보/심리) ──────────────
    _gap36 = (_o36 / _pc36 - 1)
    feat['ovn_gap_bias_20'] = np.sign(_gap36).rolling(20).mean()
    feat['ovn_persistent_gap_down'] = (np.sign(_gap36).rolling(10).mean() < -0.3).astype(float)
    _intraday_dir36 = np.sign(_c36 - _o36)
    feat['ovn_gap_fade_freq_20'] = ((np.sign(_gap36) > 0) & (_intraday_dir36 < 0)).rolling(20).mean()
    feat['ovn_gap_fade_dominant'] = (feat['ovn_gap_fade_freq_20'] > 0.5).astype(float)
    feat['ovn_cum_overnight_20'] = _gap36.rolling(20).sum()
    feat['ovn_overnight_bleeding'] = (feat['ovn_cum_overnight_20'] < -0.03).astype(float)
    feat['ovn_gap_vol_spike'] = (_gap36.abs().rolling(10).mean() >
                                 _gap36.abs().rolling(60).mean() * 1.5).astype(float)

    # ── 36H. 심리 종합 v3 (상승/하락 양방향) ────────────────────
    _dn36 = pd.Series(0.0, index=_c36.index)
    for _k in ['vbk_neg_jump_recent', 'skn_bad_capture', 'trp_overhead_supply', 'trp_dead_cat',
               'whl_whale_sell', 'whl_support_break_volume', 'cvd_bearish_divergence',
               'cvd_persistent_selling', 'ovn_gap_fade_dominant', 'ovn_overnight_bleeding']:
        if _k in feat.columns: _dn36 = _dn36 + feat[_k]
    feat['msc_drop_score'] = _dn36
    feat['msc_drop_high'] = (_dn36 >= 5).astype(float)
    feat['msc_drop_rising'] = (_dn36 > _dn36.shift(3)).astype(float)
    for p in [5, 10]:
        feat[f'msc_drop_sum_{p}d'] = _dn36.rolling(p).sum()
    _up36 = pd.Series(0.0, index=_c36.index)
    for _k in ['trp_stoploss_hunt', 'pnc_selling_exhausted', 'pnc_streak_exhaustion',
               'whl_stealth_accum', 'whl_whale_buy']:
        if _k in feat.columns: _up36 = _up36 + feat[_k]
    feat['msc_rise_score'] = _up36
    feat['msc_net_score'] = _up36 - _dn36
    feat['msc_bottom_zone'] = ((feat.get('pnc_capitulation_20d', pd.Series(0.0, index=_c36.index)) >= 1) &
                               (_up36 >= 2)).astype(float)

    # ══════════════════════════════════════════════════════════════
    #  38. 일중(5분봉) 미시구조 피처 결합 (확장 42개, 최근 60일만 값 / 그외 NaN)
    #      접두사: id_(기존15) idx_(신규27)  ※ 야후 무료 5분봉 60일 한계
    # ══════════════════════════════════════════════════════════════
    _ID_COLS = ['id_close_loc','id_close_vs_vwap','id_trend_efficiency','id_open_drive',
                'id_close_drive','id_smart_dumb','id_late_vol_ratio','id_vol_smile',
                'id_realized_vol','id_intraday_skew','id_intraday_maxdd','id_up_bar_ratio',
                'id_vw_order_flow','id_closing_vol','id_above_vwap_ratio',
                'idx_ret_open30','idx_ret_morning','idx_ret_lunch','idx_ret_afternoon',
                'idx_ret_close30','idx_pm_minus_am','idx_open_close_align','idx_reversal_down',
                'idx_reversal_up','idx_high_time_frac','idx_low_time_frac','idx_high_after_low',
                'idx_open_loc_in_day','idx_close_above_open','idx_intraday_return',
                'idx_cum_delta_end','idx_delta_late_vs_early','idx_price_delta_diverge',
                'idx_vol_open_vs_close','idx_max_bar_move','idx_vol_concentration',
                'idx_vwap_vs_mean','idx_big_bar_dir','idx_volume_trend',
                'idx_closing_flow','idx_close_at_high','idx_close_at_low']

    _ID_COLS += ['idm_last30_ret','idm_first30_ret','idm_first_last_sum','idm_last30_vol_share',
                 'idm_jump_share','idm_continuous_vol','idm_neg_semivar','idm_signed_jump',
                 'idm_autocorr1','idm_run_persistence','idm_overnight_ret','idm_intraday_ret',
                 'idm_overnight_faded','idm_on_id_same_dir','idm_vwap_time_weighted',
                 'idm_close_reclaim_vwap','idm_close_lose_vwap','idm_close_vwap_z',
                 'idm_intraday_illiq','idm_impact_asym',
                 'idm_last30_ret_5davg','idm_last30_ret_10davg','idm_signed_jump_5davg',
                 'idm_signed_jump_10davg','idm_autocorr1_5davg','idm_autocorr1_10davg',
                 'idm_intraday_ret_5davg','idm_intraday_ret_10davg','idm_close_vwap_z_5davg',
                 'idm_close_vwap_z_10davg','idm_neg_semivar_5davg','idm_neg_semivar_10davg',
                 'idm_last30_consistency_5d','idm_intraday_cum_5d','idm_jump_freq_10d',
                 'idm_above_vwap_freq_5d']
    try:
        _intra = fetch_intraday(TICKER, interval='5m', period='60d')
        _idf = intraday_daily_features(_intra)
        if _idf is not None and len(_idf) > 0:
            _idf = _idf.reindex(feat.index)
            for _col in _ID_COLS:
                feat[_col] = _idf[_col].values if _col in _idf.columns else np.nan
        else:
            for _col in _ID_COLS:
                feat[_col] = np.nan
    except Exception as _e:
        print(f"[intraday] 결합 건너뜀: {_e}")

# ══════════════════════════════════════════════════════════════
    #  39. 매수/바닥 진입 + 추세 전환 타이밍 (~56개, 기존과 중복 없음)
    #      기존 지표는 '하락 경고'에 집중되어 있어 '언제 사는가(상승 진입)'가 비어 있었음.
    #      이 섹션은 미래 '매수 시점' 예측을 직접 겨냥한다.
    #      접두사: bacc_(바닥매집완료) rev_(반전확인) tcyc_(추세사이클위치)
    #              vexp_(변동성팽창방향) mtf_(다중기간정렬) sqz_(스퀴즈해소방향)
    #      ※ 다른 섹션과 동일하게 op/hi/lo/cl/vo + calc_* 헬퍼만 사용. 새 외부데이터 불필요.
    _o = op; _h = hi; _l = lo; _c = cl; _v = vo
    _rng = (_h - _l).replace(0, np.nan)
    _body = (_c - _o)
    _pc = _c.shift(1)
    _ret1 = _c.pct_change()
    _rsi14 = calc_rsi(_c, 14)
    _atr14, _tr = calc_atr(_h, _l, _c, 14)
    _vma20 = _v.rolling(20).mean().replace(0, np.nan)
    _close_loc = (_c - _l) / _rng

    # ── 39A. 바닥 매집 완료 신호 (하락 끝 → 매수 준비) ────────────
    # 하락 후 변동성 수축 + 거래량 감소 + 가격 안정 = 매집 완료 임박
    _in_downtrend = (_c < _c.rolling(50).mean()).astype(float)
    _vol_dropping = (_v.rolling(10).mean() < _v.rolling(40).mean() * 0.8).astype(float)
    _range_settling = (_rng.rolling(10).mean() / _c < (_rng / _c).rolling(60).median() * 0.8).astype(float)
    feat['bacc_accumulation_setup'] = (_in_downtrend * _vol_dropping * _range_settling)
    feat['bacc_accumulation_20d'] = feat['bacc_accumulation_setup'].rolling(20).sum()
    # 저점 higher-low 형성 (바닥 다지기)
    _local_low = _l.rolling(10).min()
    feat['bacc_higher_low_form'] = ((_local_low > _local_low.shift(10)) & (_c < _c.rolling(50).mean())).astype(float)
    feat['bacc_higher_low_10d'] = feat['bacc_higher_low_form'].rolling(10).sum()
    # 매도 소진: 큰 음봉 뒤 거래량 동반 양봉 (반전 캔들)
    feat['bacc_selling_climax_reversal'] = (
        (_ret1.shift(1) < -0.025) & (_v.shift(1) > _vma20 * 1.5) &
        (_ret1 > 0.01) & (_close_loc > 0.6)
    ).astype(float)
    # RSI 강세 다이버전스: 가격 신저점 but RSI 더 높음 (하락 동력 소진)
    _price_ll = (_c <= _c.rolling(20).min() + 1e-9)
    _rsi_higher = (_rsi14 > _rsi14.rolling(20).min().shift(3))
    feat['bacc_bullish_divergence'] = (_price_ll & _rsi_higher).astype(float)
    feat['bacc_bullish_div_20d'] = feat['bacc_bullish_divergence'].rolling(20).sum()
    # 거래량 마름 후 첫 대량 양봉 (관심 복귀)
    feat['bacc_volume_revival'] = (
        (_v.shift(1).rolling(10).mean() < _vma20 * 0.7) &
        (_v > _vma20 * 1.5) & (_ret1 > 0)
    ).astype(float)
    # 매집 종합 점수
    feat['bacc_accum_score'] = (
        feat['bacc_higher_low_form'] + feat['bacc_selling_climax_reversal'] +
        feat['bacc_bullish_divergence'] + feat['bacc_volume_revival']
    )
    feat['bacc_accum_ready'] = (feat['bacc_accum_score'] >= 2).astype(float)

    # ── 39B. 반전 확인 (하락→상승 전환의 '확정' 신호) ───────────
    # 50일선 회복 (추세 전환 1차 확인)
    _sma50 = _c.rolling(50).mean()
    feat['rev_reclaim_sma50'] = ((_c > _sma50) & (_pc <= _sma50.shift(1))).astype(float)
    feat['rev_above_sma50_streak'] = (
        (_c > _sma50).astype(float).groupby(((_c > _sma50) != (_c > _sma50).shift()).cumsum()).cumcount() + 1
    ).where(_c > _sma50, 0)
    # 20일 고점 돌파 (단기 추세 전환)
    feat['rev_break_20d_high'] = ((_c > _c.rolling(20).max().shift(1)) & (_c.shift(1) < _c.rolling(20).max().shift(2))).astype(float)
    # 연속 상승일 + 거래량 증가 (follow-through)
    _up_days_5 = (_ret1 > 0).rolling(5).sum()
    feat['rev_followthrough'] = ((_up_days_5 >= 4) & (_v.rolling(5).mean() > _vma20)).astype(float)
    # MACD 골든크로스 + 0선 상향 (모멘텀 전환)
    _macd = _c.ewm(span=12, adjust=False).mean() - _c.ewm(span=26, adjust=False).mean()
    _macd_sig = _macd.ewm(span=9, adjust=False).mean()
    feat['rev_macd_bull_cross'] = ((_macd > _macd_sig) & (_macd.shift(1) <= _macd_sig.shift(1))).astype(float)
    feat['rev_macd_zero_cross_up'] = ((_macd > 0) & (_macd.shift(1) <= 0)).astype(float)
    # RSI 50 상향 돌파 (약세→강세 전환)
    feat['rev_rsi_cross_50_up'] = ((_rsi14 > 50) & (_rsi14.shift(1) <= 50)).astype(float)
    # 반전 종합 점수
    feat['rev_confirm_score'] = (
        feat['rev_reclaim_sma50'] + feat['rev_break_20d_high'] + feat['rev_followthrough'] +
        feat['rev_macd_bull_cross'] + feat['rev_rsi_cross_50_up']
    )
    feat['rev_confirmed'] = (feat['rev_confirm_score'] >= 2).astype(float)

    # ── 39C. 추세 사이클 위치 (상승 추세의 '나이'와 신선도) ──────
    # 상승 추세 시작 후 경과일 (신선한 추세 = 매수 유리)
    _uptrend = (_c > _sma50).astype(float)
    feat['tcyc_uptrend_age'] = _uptrend.groupby((_uptrend != _uptrend.shift()).cumsum()).cumcount().where(_uptrend > 0, 0)
    feat['tcyc_fresh_uptrend'] = ((feat['tcyc_uptrend_age'] > 0) & (feat['tcyc_uptrend_age'] <= 10)).astype(float)
    # 눌림목(pullback) in 상승추세 = 매수 기회
    _in_uptrend = (_c > _sma50) & (_sma50 > _sma50.shift(10))
    feat['tcyc_pullback_in_uptrend'] = (_in_uptrend & (_c < _c.rolling(10).max() * 0.97) & (_rsi14 < 50)).astype(float)
    feat['tcyc_pullback_10d'] = feat['tcyc_pullback_in_uptrend'].rolling(10).sum()
    # 추세 강도 (ADX 유사: 방향성 효율)
    _dir_move = (_c - _c.shift(20)).abs()
    _path = _c.diff().abs().rolling(20).sum().replace(0, np.nan)
    feat['tcyc_trend_strength_20'] = _dir_move / _path
    feat['tcyc_strong_trend'] = (feat['tcyc_trend_strength_20'] > 0.4).astype(float)
    # 추세 가속 (상승이 빨라짐)
    _mom20 = _c.pct_change(20)
    feat['tcyc_trend_accelerating'] = ((_mom20 > 0) & (_mom20 > _mom20.shift(10))).astype(float)
    # 황금 눌림목: 상승추세 + 50일선 근처 되돌림 + 반등 시작
    feat['tcyc_golden_pullback'] = (
        _in_uptrend & (_c <= _sma50 * 1.03) & (_c >= _sma50 * 0.98) & (_ret1 > 0)
    ).astype(float)

    # ── 39D. 변동성 팽창 방향 (압축→팽창 시 어느 쪽으로?) ────────
    _vol5 = _ret1.rolling(5).std()
    _vol20 = _ret1.rolling(20).std()
    _bb_width = (_c.rolling(20).std() * 4) / _c.rolling(20).mean()
    # 변동성 압축 후 팽창 시작 + 상승 방향
    _was_compressed = (_bb_width.shift(1) < _bb_width.rolling(60).quantile(0.25).shift(1))
    feat['vexp_bullish_expansion'] = (_was_compressed & (_ret1 > _vol20) & (_close_loc > 0.6)).astype(float)
    feat['vexp_bearish_expansion'] = (_was_compressed & (_ret1 < -_vol20) & (_close_loc < 0.4)).astype(float)
    # 변동성 팽창 순방향 점수 (양수=상승팽창)
    feat['vexp_expansion_dir'] = feat['vexp_bullish_expansion'] - feat['vexp_bearish_expansion']
    feat['vexp_expansion_dir_10d'] = feat['vexp_expansion_dir'].rolling(10).sum()
    # ATR 확장 + 상승 (추세 시작의 에너지)
    feat['vexp_atr_expand_up'] = ((_atr14 > _atr14.shift(5) * 1.2) & (_c.pct_change(5) > 0)).astype(float)
    # 범위 확장 양봉 (강한 매수)
    feat['vexp_wide_range_bull'] = ((_rng / _c > (_rng / _c).rolling(20).mean() * 1.5) & (_body > 0) & (_close_loc > 0.7)).astype(float)
    feat['vexp_wide_range_bull_10d'] = feat['vexp_wide_range_bull'].rolling(10).sum()

    # ── 39E. 다중 시간프레임 정렬 (단기·중기·장기 동시 상승) ─────
    _sma10 = _c.rolling(10).mean(); _sma20 = _c.rolling(20).mean()
    _sma100 = _c.rolling(100).mean()
    # 정배열 (10>20>50>100)
    feat['mtf_bullish_alignment'] = ((_sma10 > _sma20) & (_sma20 > _sma50) & (_sma50 > _sma100)).astype(float)
    # 정배열 막 시작 (전환점)
    feat['mtf_alignment_start'] = (
        feat['mtf_bullish_alignment'].astype(bool) & ~feat['mtf_bullish_alignment'].shift(1).astype(bool)
    ).astype(float)
    # 다중 기간 모멘텀 동시 양수
    _mom_sum = (np.sign(_c.pct_change(5)) + np.sign(_c.pct_change(10)) +
                np.sign(_c.pct_change(20)) + np.sign(_c.pct_change(60)))
    feat['mtf_all_timeframe_up'] = (_mom_sum == 4).astype(float)
    feat['mtf_mom_alignment_score'] = _mom_sum
    # 가격이 모든 주요 이평 위
    _above_count = ((_c > _sma10).astype(float) + (_c > _sma20).astype(float) +
                    (_c > _sma50).astype(float) + (_c > _sma100).astype(float))
    feat['mtf_above_all_ma'] = (_above_count == 4).astype(float)
    feat['mtf_above_ma_count'] = _above_count
    feat['mtf_above_ma_rising'] = (_above_count > _above_count.shift(5)).astype(float)

    # ── 39F. 스퀴즈 해소 방향 (TTM 스퀴즈 유사) ─────────────────
    _atr20, _ = calc_atr(_h, _l, _c, 20)
    _kc_upper = _c.ewm(span=20, adjust=False).mean() + 1.5 * _atr20
    _kc_lower = _c.ewm(span=20, adjust=False).mean() - 1.5 * _atr20
    _bb_upper = _c.rolling(20).mean() + 2 * _c.rolling(20).std()
    _bb_lower = _c.rolling(20).mean() - 2 * _c.rolling(20).std()
    _in_squeeze = (_bb_upper < _kc_upper) & (_bb_lower > _kc_lower)
    feat['sqz_in_squeeze'] = _in_squeeze.astype(float)
    feat['sqz_squeeze_count'] = _in_squeeze.astype(float).rolling(20).sum()
    # 스퀴즈 해소 (압축 끝) + 방향
    _squeeze_fired = _in_squeeze.shift(1) & ~_in_squeeze
    _mom_dir = _c - (_h.rolling(20).max() + _l.rolling(20).min()) / 2
    feat['sqz_fired_bullish'] = (_squeeze_fired & (_mom_dir > 0)).astype(float)
    feat['sqz_fired_bearish'] = (_squeeze_fired & (_mom_dir < 0)).astype(float)
    feat['sqz_fired_dir_10d'] = (feat['sqz_fired_bullish'] - feat['sqz_fired_bearish']).rolling(10).sum()
    # 스퀴즈 모멘텀 (압축 중 방향 기울기)
    feat['sqz_momentum'] = calc_linreg_slope(_mom_dir, 5)
    feat['sqz_momentum_positive'] = (feat['sqz_momentum'] > 0).astype(float)

    # ── 39G. 매수 타이밍 종합 점수 (상승 진입 신호) ──────────────
    _buy_score = (
        feat['bacc_accum_ready'] +
        feat['rev_confirmed'] +
        feat['tcyc_golden_pullback'] +
        feat['vexp_bullish_expansion'] +
        feat['mtf_alignment_start'] +
        feat['sqz_fired_bullish'] +
        feat['rev_macd_bull_cross'] +
        feat['tcyc_fresh_uptrend']
    )
    feat['mtf_buy_timing_score'] = _buy_score
    feat['mtf_buy_signal_high'] = (_buy_score >= 3).astype(float)
    feat['mtf_buy_signal_strong'] = (_buy_score >= 5).astype(float)
    feat['mtf_buy_score_rising'] = (_buy_score > _buy_score.shift(3)).astype(float)
    for p in [3, 5, 10]:
        feat[f'mtf_buy_score_sum_{p}d'] = _buy_score.rolling(p).sum()
    feat['mtf_buy_score_zscore60'] = calc_zscore(_buy_score, 60)     

# ══════════════════════════════════════════════════════════════
    #  40. 캔들 패턴 대량 확장 (~76개, 섹션 20·21·22와 중복 없음)
    #      고전 캔들스틱 패턴 중 아직 없는 것 + 강세/약세 방향 종합점수.
    #      접두사: cx_ (candle extended) — 기존 cdl_ 과 이름 완전 분리.
    #      ※ 다른 섹션과 동일하게 op/hi/lo/cl/vo 만 사용. 새 외부데이터 불필요.
    # ══════════════════════════════════════════════════════════════
    _o = op; _h = hi; _l = lo; _c = cl; _v = vo
    _rng  = (_h - _l).replace(0, np.nan)
    _body = (_c - _o)
    _abs  = _body.abs()
    _uw   = _h - pd.concat([_c, _o], axis=1).max(axis=1)   # 위꼬리
    _lw   = pd.concat([_c, _o], axis=1).min(axis=1) - _l   # 아래꼬리
    # 전일/2~4일 전
    _pc, _po, _ph, _pl = _c.shift(1), _o.shift(1), _h.shift(1), _l.shift(1)
    _pb = _body.shift(1); _pabs = _abs.shift(1)
    _c2, _o2, _h2, _l2 = _c.shift(2), _o.shift(2), _h.shift(2), _l.shift(2)
    _b2 = _body.shift(2)
    _c3, _o3 = _c.shift(3), _o.shift(3); _b3 = _body.shift(3)
    _c4 = _c.shift(4); _b4 = _body.shift(4)
    # 몸통 상/하단
    _tbody = pd.concat([_c, _o], axis=1).max(axis=1)
    _bbody = pd.concat([_c, _o], axis=1).min(axis=1)
    _doji  = (_abs <= _rng * 0.1)

    # ── 40A. 스타 계열 (갭 + 작은 몸통 반전) ───────────────────────
    feat['cx_doji_star_bull'] = ((_pb < 0) & (_pabs > _rng.shift(1) * 0.5) & _doji & (_tbody < _bbody.shift(1))).astype(float)
    feat['cx_doji_star_bear'] = ((_pb > 0) & (_pabs > _rng.shift(1) * 0.5) & _doji & (_bbody > _tbody.shift(1))).astype(float)
    feat['cx_morning_doji_star'] = ((_b2 < 0) & (_abs.shift(1) <= _rng.shift(1) * 0.1) & (_body > 0) & (_c > (_c2 + _o2) / 2)).astype(float)
    feat['cx_evening_doji_star'] = ((_b2 > 0) & (_abs.shift(1) <= _rng.shift(1) * 0.1) & (_body < 0) & (_c < (_c2 + _o2) / 2)).astype(float)
    feat['cx_tri_star_bull'] = (_doji & (_abs.shift(1) <= _rng.shift(1) * 0.1) & (_abs.shift(2) <= _rng.shift(2) * 0.1) & (_l < _l.shift(1))).astype(float)
    feat['cx_tri_star_bear'] = (_doji & (_abs.shift(1) <= _rng.shift(1) * 0.1) & (_abs.shift(2) <= _rng.shift(2) * 0.1) & (_h > _h.shift(1))).astype(float)
    feat['cx_abandoned_baby_bull'] = ((_b2 < 0) & (_abs.shift(1) <= _rng.shift(1) * 0.1) & (_h.shift(1) < _l2) & (_l > _h.shift(1)) & (_body > 0)).astype(float)
    feat['cx_abandoned_baby_bear'] = ((_b2 > 0) & (_abs.shift(1) <= _rng.shift(1) * 0.1) & (_l.shift(1) > _h2) & (_h < _l.shift(1)) & (_body < 0)).astype(float)

    # ── 40B. 도지 변형 ─────────────────────────────────────────────
    feat['cx_long_legged_doji'] = (_doji & (_uw > _rng * 0.35) & (_lw > _rng * 0.35)).astype(float)
    feat['cx_rickshaw_man'] = (_doji & (_uw > _rng * 0.4) & (_lw > _rng * 0.4) &
                               (((_tbody + _bbody) / 2 - _l) / _rng).between(0.4, 0.6)).astype(float)
    feat['cx_high_wave_doji'] = (_doji & (_rng > _rng.rolling(20).mean() * 1.3)).astype(float)

    # ── 40C. 마루보즈 변형 (꼬리 없는 강한 봉) ─────────────────────
    feat['cx_closing_marubozu_bull'] = ((_body > 0) & (_uw <= _rng * 0.03) & (_lw > _rng * 0.03)).astype(float)
    feat['cx_closing_marubozu_bear'] = ((_body < 0) & (_lw <= _rng * 0.03) & (_uw > _rng * 0.03)).astype(float)
    feat['cx_opening_marubozu_bull'] = ((_body > 0) & (_lw <= _rng * 0.03) & (_uw > _rng * 0.03)).astype(float)
    feat['cx_opening_marubozu_bear'] = ((_body < 0) & (_uw <= _rng * 0.03) & (_lw > _rng * 0.03)).astype(float)

    # ── 40D. 갭 연속 패턴 (Tasuki / Side-by-side) ──────────────────
    feat['cx_tasuki_gap_up'] = ((_pb > 0) & (_po > _c2) & (_body < 0) & (_o < _pc) & (_o > _po) & (_c > _po)).astype(float)
    feat['cx_tasuki_gap_down'] = ((_pb < 0) & (_po < _c2) & (_body > 0) & (_o > _pc) & (_o < _po) & (_c < _po)).astype(float)
    feat['cx_side_by_side_white'] = ((_pb > 0) & (_body > 0) & (_po > _c2) & (_o > _pc * 0.99) & (_o < _pc * 1.01)).astype(float)

    # ── 40E. 넥라인 계열 (On-neck/In-neck/Thrusting) ───────────────
    feat['cx_on_neck'] = ((_pb < 0) & (_body > 0) & (_o < _pl) & ((_c - _pl).abs() < _rng * 0.05)).astype(float)
    feat['cx_in_neck'] = ((_pb < 0) & (_body > 0) & (_o < _pl) & (_c > _pc) & (_c < _pc + _rng * 0.1)).astype(float)
    feat['cx_thrusting'] = ((_pb < 0) & (_body > 0) & (_o < _pl) & (_c > (_po + _pc) / 2 * 0.5) & (_c < (_po + _pc) / 2)).astype(float)

    # ── 40F. 카운터어택 / 세퍼레이팅 (종가·시가 일치) ──────────────
    feat['cx_counterattack_bull'] = ((_pb < 0) & (_body > 0) & ((_c - _pc).abs() < _rng * 0.05)).astype(float)
    feat['cx_counterattack_bear'] = ((_pb > 0) & (_body < 0) & ((_c - _pc).abs() < _rng * 0.05)).astype(float)
    feat['cx_separating_lines_bull'] = ((_pb < 0) & (_body > 0) & ((_o - _po).abs() < _rng * 0.05)).astype(float)
    feat['cx_separating_lines_bear'] = ((_pb > 0) & (_body < 0) & ((_o - _po).abs() < _rng * 0.05)).astype(float)

    # ── 40G. 매칭 / 호밍 (저점·고점 일치) ──────────────────────────
    feat['cx_matching_low'] = ((_pb < 0) & (_body < 0) & ((_c - _pc).abs() < _rng * 0.03)).astype(float)
    feat['cx_matching_high'] = ((_pb > 0) & (_body > 0) & ((_c - _pc).abs() < _rng * 0.03)).astype(float)
    feat['cx_homing_pigeon'] = ((_pb < 0) & (_body < 0) & (_h <= _ph) & (_l >= _pl) & (_pabs > _abs)).astype(float)

    # ── 40H. 3봉 반전 (Three Outside, Two Crows) ───────────────────
    feat['cx_three_outside_up'] = ((_pb < 0) & (_body > 0) & (_o <= _pc) & (_c >= _po) & (_c > _pc)).astype(float)
    feat['cx_three_outside_down'] = ((_pb > 0) & (_body < 0) & (_o >= _pc) & (_c <= _po)).astype(float)
    feat['cx_two_crows'] = ((_b2 > 0) & (_b2.abs() > _rng.shift(2) * 0.5) & (_pb < 0) & (_po > _c2) & (_body < 0) & (_o < _po) & (_o > _pc) & (_c < _c2) & (_c > _o2)).astype(float)
    feat['cx_upside_gap_two_crows'] = ((_b2 > 0) & (_pb < 0) & (_po > _c2) & (_body < 0) & (_c < _pc) & (_o > _po)).astype(float)

    # ── 40I. 3선 타격 / 매트홀드 (지속 패턴) ───────────────────────
    feat['cx_three_line_strike_bull'] = ((_b3 > 0) & (_b2 > 0) & (_pb > 0) & (_c3 < _c2) & (_c2 < _pc) & (_body < 0) & (_o > _pc) & (_c < _c3)).astype(float)
    feat['cx_three_line_strike_bear'] = ((_b3 < 0) & (_b2 < 0) & (_pb < 0) & (_c3 > _c2) & (_c2 > _pc) & (_body > 0) & (_o < _pc) & (_c > _c3)).astype(float)
    feat['cx_mat_hold'] = ((_b3 > 0) & (_b3.abs() > _rng.shift(3) * 0.5) & (_body > 0) & (_c > _h.shift(1)) & (_c > _h2) & (_c2 < _c3) & (_pc < _c3)).astype(float)

    # ── 40J. 사다리 / 딜리버레이션 / 어드밴스블록 (소진) ───────────
    feat['cx_ladder_bottom'] = ((_b4 < 0) & (_b3 < 0) & (_b2 < 0) & (_pb < 0) & (_body > 0) & (_o > _po)).astype(float)
    feat['cx_ladder_top'] = ((_b4 > 0) & (_b3 > 0) & (_b2 > 0) & (_pb > 0) & (_body < 0) & (_o < _po)).astype(float)
    feat['cx_deliberation'] = ((_b2 > 0) & (_pb > 0) & (_body > 0) & (_pabs > _rng.shift(1) * 0.5) & (_abs < _pabs * 0.5) & (_o > _pc)).astype(float)
    feat['cx_advance_block'] = ((_b2 > 0) & (_pb > 0) & (_body > 0) & (_abs < _pabs) & (_pabs < _b2.abs()) & (_uw > _abs * 0.5)).astype(float)
    feat['cx_concealing_baby'] = ((_b3 < 0) & (_b2 < 0) & (_pb < 0) & (_body < 0) & (_h > _ph)).astype(float)

    # ── 40K. 브레이크어웨이 (갭 추세 시작) ─────────────────────────
    feat['cx_breakaway_bull'] = ((_b4 < 0) & (_o3 < _c4) & (_b2 < 0) & (_pb < 0) & (_body > 0) & (_c > _o3)).astype(float)
    feat['cx_breakaway_bear'] = ((_b4 > 0) & (_o3 > _c4) & (_b2 > 0) & (_pb > 0) & (_body < 0) & (_c < _o3)).astype(float)

    # ── 40L. 타워 / 프라이팬 / 덤플링 (라운딩 반전) ────────────────
    feat['cx_tower_bottom'] = ((_c.shift(5) > _c.shift(4)) & (_c.rolling(3).mean() < _c.shift(4)) & (_body > 0) & (_c > _c.shift(5) * 0.98)).astype(float)
    feat['cx_tower_top'] = ((_c.shift(5) < _c.shift(4)) & (_c.rolling(3).mean() > _c.shift(4)) & (_body < 0) & (_c < _c.shift(5) * 1.02)).astype(float)
    _min10 = _l.rolling(10).min()
    feat['cx_fry_pan_bottom'] = ((_c > _c.rolling(10).mean()) & (_min10 == _min10.shift(3)) & (_c.pct_change(10) > 0) & (_l > _min10 * 1.01)).astype(float)
    _max10 = _h.rolling(10).max()
    feat['cx_dumpling_top'] = ((_c < _c.rolling(10).mean()) & (_max10 == _max10.shift(3)) & (_c.pct_change(10) < 0) & (_h < _max10 * 0.99)).astype(float)

    # ── 40M. 키 / 아웃사이드 / 2봉 리버설 ──────────────────────────
    feat['cx_key_reversal_up'] = ((_l < _pl) & (_o < _pc) & (_c > _pc) & (_c > _po)).astype(float)
    feat['cx_key_reversal_down'] = ((_h > _ph) & (_o > _pc) & (_c < _pc) & (_c < _po)).astype(float)
    feat['cx_outside_reversal_up'] = ((_h > _ph) & (_l < _pl) & (_c > _pc) & (_body > 0)).astype(float)
    feat['cx_outside_reversal_down'] = ((_h > _ph) & (_l < _pl) & (_c < _pc) & (_body < 0)).astype(float)
    feat['cx_two_bar_reversal_up'] = ((_pb < 0) & (_pabs > _rng.shift(1) * 0.6) & (_body > 0) & (_abs > _rng * 0.6) & (_c > _po)).astype(float)
    feat['cx_two_bar_reversal_down'] = ((_pb > 0) & (_pabs > _rng.shift(1) * 0.6) & (_body < 0) & (_abs > _rng * 0.6) & (_c < _po)).astype(float)

    # ── 40N. 3봉 플레이 / 히카케 (변형 패턴) ───────────────────────
    feat['cx_three_bar_play_bull'] = ((_b2 > 0) & (_b2.abs() > _rng.shift(2) * 0.6) & (_abs.shift(1) < _b2.abs() * 0.5) & (_body > 0) & (_c > _h2)).astype(float)
    feat['cx_three_bar_play_bear'] = ((_b2 < 0) & (_b2.abs() > _rng.shift(2) * 0.6) & (_abs.shift(1) < _b2.abs() * 0.5) & (_body < 0) & (_c < _l2)).astype(float)
    _inside1 = (_ph <= _h2) & (_pl >= _l2)
    feat['cx_hikkake_bull'] = (_inside1.shift(1) & (_l.shift(1) < _pl) & (_c > _ph)).astype(float)
    feat['cx_hikkake_bear'] = (_inside1.shift(1) & (_h.shift(1) > _ph) & (_c < _pl)).astype(float)

    # ── 40O. 유니크 3리버 / 라스트 엔걸핑 / 스틱 리버설 ────────────
    feat['cx_unique_three_river'] = ((_b2 < 0) & (_b2.abs() > _rng.shift(2) * 0.5) & (_pb < 0) & (_l.shift(1) < _l2) & (_pc > _c2) & (_body > 0) & (_abs < _rng * 0.3)).astype(float)
    feat['cx_last_engulf_bottom'] = ((_pb < 0) & (_body < 0) & (_o >= _pc) & (_c <= _po) & (_c < _c.rolling(20).min().shift(1) * 1.02)).astype(float)
    feat['cx_last_engulf_top'] = ((_pb > 0) & (_body > 0) & (_o <= _pc) & (_c >= _po) & (_c > _c.rolling(20).max().shift(1) * 0.98)).astype(float)
    feat['cx_stick_reversal_up'] = ((_c.shift(1) < _c.shift(2)) & (_c.shift(2) < _c.shift(3)) & (_body > 0) & (_c > _o.shift(2))).astype(float)
    feat['cx_stick_reversal_down'] = ((_c.shift(1) > _c.shift(2)) & (_c.shift(2) > _c.shift(3)) & (_body < 0) & (_c < _o.shift(2))).astype(float)

    # ── 40P. 레코드 세션 (연속 신고가/신저가 = 소진) ───────────────
    _new_high = (_h > _h.shift(1))
    _new_low = (_l < _l.shift(1))
    feat['cx_record_high_sessions'] = _new_high.rolling(8).sum()
    feat['cx_eight_new_highs'] = (_new_high.rolling(8).sum() >= 7).astype(float)
    feat['cx_record_low_sessions'] = _new_low.rolling(8).sum()
    feat['cx_eight_new_lows'] = (_new_low.rolling(8).sum() >= 7).astype(float)

    # ── 40Q. 패턴 종합 점수 (강세/약세 결합) ───────────────────────
    _cx_bull_cols = ['cx_doji_star_bull', 'cx_morning_doji_star', 'cx_tri_star_bull', 'cx_abandoned_baby_bull',
                     'cx_closing_marubozu_bull', 'cx_tasuki_gap_down', 'cx_counterattack_bull', 'cx_matching_low',
                     'cx_homing_pigeon', 'cx_three_outside_up', 'cx_three_line_strike_bull', 'cx_ladder_bottom',
                     'cx_breakaway_bull', 'cx_tower_bottom', 'cx_fry_pan_bottom', 'cx_key_reversal_up',
                     'cx_outside_reversal_up', 'cx_two_bar_reversal_up', 'cx_three_bar_play_bull', 'cx_hikkake_bull',
                     'cx_unique_three_river', 'cx_last_engulf_bottom', 'cx_stick_reversal_up', 'cx_separating_lines_bull']
    _cx_bear_cols = ['cx_doji_star_bear', 'cx_evening_doji_star', 'cx_tri_star_bear', 'cx_abandoned_baby_bear',
                     'cx_closing_marubozu_bear', 'cx_tasuki_gap_up', 'cx_counterattack_bear', 'cx_matching_high',
                     'cx_two_crows', 'cx_upside_gap_two_crows', 'cx_three_outside_down', 'cx_three_line_strike_bear',
                     'cx_ladder_top', 'cx_breakaway_bear', 'cx_tower_top', 'cx_dumpling_top', 'cx_key_reversal_down',
                     'cx_outside_reversal_down', 'cx_two_bar_reversal_down', 'cx_three_bar_play_bear', 'cx_hikkake_bear',
                     'cx_deliberation', 'cx_advance_block', 'cx_last_engulf_top', 'cx_stick_reversal_down', 'cx_separating_lines_bear']
    feat['cx_bull_pattern_score'] = feat[[c for c in _cx_bull_cols if c in feat.columns]].sum(axis=1)
    feat['cx_bear_pattern_score'] = feat[[c for c in _cx_bear_cols if c in feat.columns]].sum(axis=1)
    feat['cx_net_pattern_score'] = feat['cx_bull_pattern_score'] - feat['cx_bear_pattern_score']
    for p in [5, 10]:
        feat[f'cx_bull_score_sum_{p}d'] = feat['cx_bull_pattern_score'].rolling(p).sum()
        feat[f'cx_bear_score_sum_{p}d'] = feat['cx_bear_pattern_score'].rolling(p).sum()
        feat[f'cx_net_score_sum_{p}d'] = feat['cx_net_pattern_score'].rolling(p).sum()
    feat['cx_strong_bull_signal'] = (feat['cx_bull_pattern_score'] >= 2).astype(float)
    feat['cx_strong_bear_signal'] = (feat['cx_bear_pattern_score'] >= 2).astype(float)

# ══════════════════════════════════════════════════════════════
    #  41. 1일 선행 하락 경보 — 기존 미포함 정밀 지표 (54개, 중복 없음)
    #  ※ 신규 티커 필요(없으면 .get()이 None→해당 지표만 건너뜀, 에러 없음):
    #     ^VIX9D, ^VIX3M(구 ^VXV), ^VIX6M, ^SKEW, ^MOVE(선택),
    #     GOOGL, META, TSLA, AMZN  (NVDA/MSFT/AAPL은 이미 사용)
    #  ※ 역사 한계: ^VIX3M~2007, ^VIX9D~2011 — 그 이전 연도는 NaN(정상).
    #     ^SKEW는 1990+라 닷컴까지 커버. 메가캡 상관은 가용 종목이
    #     4개 이상일 때만 계산(초기엔 NVDA/MSFT/AAPL/AMZN으로 동작).
    #  접두사: vts_(VIX기간구조) skw_(SKEW) mcc_(메가캡상관)
    #          hbg_(브레드스churn) bvol_(국채변동성) pre1_(종합)
    # ══════════════════════════════════════════════════════════════
    _c41 = cl
    _ret41 = _c41.pct_change()

    # ── 41A. VIX 실제 기간구조 (백워데이션 = 가장 강한 1일 선행) ──
    _vix9d = closes.get('^VIX9D')
    _vix1m = closes.get('^VIX')
    _vix3m = closes.get('^VIX3M')
    if _vix3m is None:
        _vix3m = closes.get('^VXV')
    _vix6m = closes.get('^VIX6M')
    if _vix9d is not None and _vix1m is not None:
        _r91 = _vix9d / _vix1m.replace(0, np.nan)
        feat['vts_vix9d_1m_ratio'] = _r91
        feat['vts_backwardation_9d'] = (_r91 > 1.0).astype(float)
        feat['vts_backwardation_9d_strong'] = (_r91 > 1.05).astype(float)
        feat['vts_9d_1m_zscore_60'] = calc_zscore(_r91, 60)
        feat['vts_9d_1m_spike'] = (_r91 > _r91.rolling(60).mean() + _r91.rolling(60).std() * 1.5).astype(float)
    if _vix1m is not None and _vix3m is not None:
        _r13 = _vix1m / _vix3m.replace(0, np.nan)
        feat['vts_vix1m_3m_ratio'] = _r13
        feat['vts_backwardation_1m_3m'] = (_r13 > 1.0).astype(float)
        feat['vts_1m_3m_zscore_60'] = calc_zscore(_r13, 60)
        feat['vts_term_slope_chg5'] = _r13 - _r13.shift(5)
        feat['vts_flip_to_backward'] = ((_r13 > 1.0) & (_r13.shift(1) <= 1.0)).astype(float)
    if _vix3m is not None and _vix6m is not None:
        feat['vts_vix3m_6m_ratio'] = _vix3m / _vix6m.replace(0, np.nan)
    if _vix9d is not None and _vix1m is not None and _vix3m is not None:
        feat['vts_full_backwardation'] = ((_vix9d > _vix1m) & (_vix1m > _vix3m)).astype(float)
        feat['vts_full_backward_3d'] = feat['vts_full_backwardation'].rolling(3).sum()

    # ── 41B. CBOE SKEW (꼬리위험 가격 = 기관 대형 풋 수요) ──────────
    _skew = closes.get('^SKEW')
    if _skew is not None:
        feat['skw_level'] = _skew
        feat['skw_zscore_60'] = calc_zscore(_skew, 60)
        feat['skw_pctrank_252'] = calc_pctrank(_skew, 252)
        feat['skw_above_145'] = (_skew > 145).astype(float)
        feat['skw_above_150'] = (_skew > 150).astype(float)
        feat['skw_spike_5d'] = (_skew.diff(5) > _skew.diff(5).rolling(60).std() * 2).astype(float)
        if _vix1m is not None:
            feat['skw_complacent_tail'] = ((_skew > 140) & (_vix1m < 16)).astype(float)
            feat['skw_vix_ratio'] = _skew / _vix1m.replace(0, np.nan)
            feat['skw_vix_ratio_z60'] = calc_zscore(_skew / _vix1m.replace(0, np.nan), 60)

    # ── 41C. 메가캡 내부 상관 급등 (동반청산 = 시스템 위험) ─────────
    _megas = {}
    for _m in ['NVDA', 'MSFT', 'AAPL', 'AMZN', 'GOOGL', 'META', 'TSLA']:
        _s = closes.get(_m)
        if _s is not None and _m != TICKER:
            _megas[_m] = _s.pct_change()
    if len(_megas) >= 4:
        _mdf = pd.DataFrame(_megas)
        _mmean = _mdf.mean(axis=1)
        _corrs = pd.DataFrame({k: _mdf[k].rolling(20).corr(_mmean) for k in _mdf.columns})
        _avgc = _corrs.mean(axis=1)
        feat['mcc_avg_corr_20'] = _avgc
        feat['mcc_corr_high'] = (_avgc > 0.8).astype(float)
        feat['mcc_corr_spike'] = (_avgc > _avgc.rolling(60).mean() + _avgc.rolling(60).std() * 1.5).astype(float)
        feat['mcc_corr_zscore_60'] = calc_zscore(_avgc, 60)
        feat['mcc_corr_up_megas_dn'] = ((_avgc > 0.75) & (_mmean.rolling(3).mean() < 0)).astype(float)
        feat['mcc_breadth_neg_5d'] = (_mdf < 0).sum(axis=1).rolling(5).mean()
        feat['mcc_all_red'] = ((_mdf < 0).sum(axis=1) >= len(_megas)).astype(float)
        feat['mcc_all_red_3d'] = feat['mcc_all_red'].rolling(3).sum()

    # ── 41D. 브레드스 churn (Hindenburg / Titanic = 천장 혼조) ─────
    SEC41 = ['XLK', 'XLV', 'XLF', 'XLY', 'XLP', 'XLE', 'XLI', 'XLB', 'XLU', 'XLRE', 'XLC']
    _sa41 = [s for s in SEC41 if s in closes.columns and s != TICKER]
    if len(_sa41) >= 7:
        _sd41 = pd.DataFrame({s: closes[s] for s in _sa41})
        _nh = pd.DataFrame({s: (_sd41[s] >= _sd41[s].rolling(252, min_periods=60).max() * 0.999).astype(float)
                            for s in _sa41}).sum(axis=1)
        _nl = pd.DataFrame({s: (_sd41[s] <= _sd41[s].rolling(252, min_periods=60).min() * 1.001).astype(float)
                            for s in _sa41}).sum(axis=1)
        _n41 = len(_sa41)
        _chg41 = _sd41.pct_change()
        _adv41 = (_chg41 > 0).sum(axis=1) - (_chg41 < 0).sum(axis=1)
        _mcl41 = _adv41.ewm(span=19, adjust=False).mean() - _adv41.ewm(span=39, adjust=False).mean()
        _mkt41 = _sd41.mean(axis=1)
        _mkt_up41 = (_mkt41 > _mkt41.rolling(50).mean()).astype(float)
        feat['hbg_new_high_cnt'] = _nh
        feat['hbg_new_low_cnt'] = _nl
        feat['hbg_both_elevated'] = ((_nh >= max(2, _n41 * 0.2)) & (_nl >= max(2, _n41 * 0.2))).astype(float)
        feat['hbg_hindenburg_flag'] = (feat['hbg_both_elevated'].astype(bool) & (_mkt_up41 > 0) & (_mcl41 < 0)).astype(float)
        feat['hbg_hindenburg_10d'] = feat['hbg_hindenburg_flag'].rolling(10).sum()
        feat['hbg_titanic_flag'] = ((_nl > _nh) & (_mkt_up41 > 0) &
                                    (_mkt41 >= _mkt41.rolling(60).max() * 0.97)).astype(float)
        feat['hbg_nl_minus_nh'] = (_nl - _nh) / _n41
        feat['hbg_churn_score'] = feat['hbg_both_elevated'] + feat['hbg_titanic_flag'] + (_mcl41 < 0).astype(float)

    # ── 41E. 국채 변동성 (MOVE = 채권發 선행; 금리'수준'·'수익률변동성'과 다른 각도) ──
    _move = closes.get('^MOVE')
    if _move is not None:
        feat['bvol_move_level'] = _move
        feat['bvol_move_zscore_60'] = calc_zscore(_move, 60)
        feat['bvol_move_spike'] = (_move > _move.rolling(60).mean() + _move.rolling(60).std() * 1.5).astype(float)
        feat['bvol_move_above_120'] = (_move > 120).astype(float)
        feat['bvol_move_5d_chg'] = _move.pct_change(5)
    _tlt41 = closes.get('TLT')
    if _tlt41 is not None:
        _tltvol = _tlt41.pct_change().rolling(20).std() * np.sqrt(252)
        feat['bvol_tlt_realvol_20'] = _tltvol
        feat['bvol_tlt_vol_zscore'] = calc_zscore(_tltvol, 120)
        feat['bvol_tlt_vol_spike'] = (_tltvol > _tltvol.rolling(120).mean() + _tltvol.rolling(120).std() * 1.5).astype(float)
        feat['bvol_bondvol_stock_calm'] = ((_tltvol > _tltvol.rolling(60).mean() * 1.3) &
                                           (_ret41.rolling(5).mean() > -0.005)).astype(float)

    # ── 41F. 1일 선행 하락 종합 경보 (위 신호 결합) ────────────────
    _pre1 = pd.Series(0.0, index=_c41.index)
    for _k in ['vts_backwardation_1m_3m', 'vts_full_backwardation', 'skw_complacent_tail',
               'mcc_corr_up_megas_dn', 'hbg_hindenburg_flag', 'hbg_titanic_flag',
               'bvol_tlt_vol_spike', 'bvol_move_spike']:
        if _k in feat.columns:
            _pre1 = _pre1 + feat[_k].fillna(0)
    feat['pre1_predrop_score'] = _pre1
    feat['pre1_predrop_high'] = (_pre1 >= 3).astype(float)
    feat['pre1_predrop_extreme'] = (_pre1 >= 5).astype(float)
    feat['pre1_predrop_rising'] = (_pre1 > _pre1.shift(3)).astype(float)
    for _p in [3, 5]:
        feat[f'pre1_predrop_sum_{_p}d'] = _pre1.rolling(_p).sum()
    feat['pre1_predrop_zscore_60'] = calc_zscore(_pre1, 60)

# ══════════════════════════════════════════════════════════════
    #  42. 하락 '유형'별 1일 선행 경보 — 기존 미포함 카테고리 (70개, 중복 없음)
    #  ※ 추가 데이터 필요(없으면 .get()/columns 체크로 해당 블록만 건너뜀):
    #     FRED: 'NFCI','ANFCI','STLFSI4','TEDRATE','DCPF3M','DTB3'
    #     티커: '^CPC','^CPCE'(풋콜), 'BTC-USD'
    #  접두사: fnd_(자금경색) pcr_(풋콜) xast_(크로스에셋) btc_(크립토)
    #          mga_(메가캡에어포켓) pre2_(유형별종합)
    # ══════════════════════════════════════════════════════════════
    _c42 = cl
    _ret42 = _c42.pct_change()

    # ── 42A. [금융위기형] 단기자금 경색 + 금융스트레스지수 (FRED) ──
    if fred_df is not None and len(fred_df) > 0:
        _fa = fred_df.reindex(feat.index).ffill().bfill()
        def _f42(c):
            return _fa[c] if c in _fa.columns else None
        _nfci = _f42('NFCI'); _anfci = _f42('ANFCI'); _stl = _f42('STLFSI4')
        if _nfci is not None:
            feat['fnd_nfci_level'] = _nfci
            feat['fnd_nfci_tight'] = (_nfci > 0).astype(float)
            feat['fnd_nfci_zscore_252'] = calc_zscore(_nfci, 252)
            feat['fnd_nfci_rising_4w'] = (_nfci.diff(20) > 0).astype(float)
            feat['fnd_nfci_jump'] = (_nfci.diff(5) > _nfci.diff(5).rolling(120).std() * 2).astype(float)
        if _anfci is not None:
            feat['fnd_anfci_level'] = _anfci
            feat['fnd_anfci_positive'] = (_anfci > 0).astype(float)
            feat['fnd_anfci_rising'] = (_anfci.diff(20) > 0).astype(float)
        if _stl is not None:
            feat['fnd_stlfsi_level'] = _stl
            feat['fnd_stlfsi_positive'] = (_stl > 0).astype(float)
            feat['fnd_stlfsi_zscore_252'] = calc_zscore(_stl, 252)
            feat['fnd_stlfsi_spike'] = (_stl > _stl.rolling(252).mean() + _stl.rolling(252).std() * 1.5).astype(float)
        _ted = _f42('TEDRATE')
        if _ted is not None:
            feat['fnd_ted_level'] = _ted
            feat['fnd_ted_above_50bp'] = (_ted > 0.5).astype(float)
            feat['fnd_ted_zscore_252'] = calc_zscore(_ted, 252)
            feat['fnd_ted_widening_5d'] = (_ted.diff(5) > 0.1).astype(float)
            feat['fnd_ted_spike'] = (_ted > _ted.rolling(120).mean() + _ted.rolling(120).std() * 2).astype(float)
        _cpf = _f42('DCPF3M'); _tb3 = _f42('DTB3')
        if _cpf is not None and _tb3 is not None:
            _cpspr = _cpf - _tb3
            feat['fnd_cp_bill_spread'] = _cpspr
            feat['fnd_cp_bill_zscore_252'] = calc_zscore(_cpspr, 252)
            feat['fnd_cp_bill_widening'] = (_cpspr.diff(5) > 0.1).astype(float)
            feat['fnd_cp_stress_flag'] = (_cpspr > _cpspr.rolling(252).mean() + _cpspr.rolling(252).std() * 2).astype(float)
        _fndsc = pd.Series(0.0, index=feat.index)
        for _k in ['fnd_nfci_tight', 'fnd_anfci_positive', 'fnd_stlfsi_positive',
                   'fnd_ted_above_50bp', 'fnd_cp_stress_flag']:
            if _k in feat.columns:
                _fndsc = _fndsc + feat[_k].fillna(0)
        feat['fnd_funding_stress_score'] = _fndsc
        feat['fnd_funding_stress_high'] = (_fndsc >= 2).astype(float)
        feat['fnd_stress_rising_stock_calm'] = ((_fndsc > _fndsc.shift(10)) &
                                                (_ret42.rolling(5).mean() > -0.005)).astype(float)

    # ── 42B. [심리극단형] 실제 CBOE 풋/콜 비율 ──
    _pc_tot = closes.get('^CPC'); _pc_eq = closes.get('^CPCE')
    for _pcser, _lbl in [(_pc_tot, 'cpc'), (_pc_eq, 'cpce')]:
        if _pcser is not None:
            feat[f'pcr_{_lbl}_level'] = _pcser
            feat[f'pcr_{_lbl}_zscore_60'] = calc_zscore(_pcser, 60)
            feat[f'pcr_{_lbl}_pctrank_252'] = calc_pctrank(_pcser, 252)
            feat[f'pcr_{_lbl}_ma5'] = _pcser.rolling(5).mean()
            feat[f'pcr_{_lbl}_complacent_low'] = (calc_pctrank(_pcser, 252) < 0.15).astype(float)
            feat[f'pcr_{_lbl}_fear_high'] = (calc_pctrank(_pcser, 252) > 0.90).astype(float)
            feat[f'pcr_{_lbl}_rising_5d'] = (_pcser.rolling(5).mean() > _pcser.rolling(20).mean()).astype(float)
    if _pc_eq is not None:
        feat['pcr_complacency_at_high'] = ((calc_pctrank(_pc_eq, 252) < 0.15) &
                                           (_c42 >= _c42.rolling(60).max() * 0.97)).astype(float)

    # ── 42C. [시스템/분산실패형] 크로스에셋 동반붕괴 ──
    _ASSETS42 = {'SPY': closes.get('SPY'), 'TLT': closes.get('TLT'), 'GLD': closes.get('GLD'),
                 'DBC': closes.get('DBC'), 'EEM': closes.get('EEM'), 'HYG': closes.get('HYG')}
    _av42 = {k: v.pct_change() for k, v in _ASSETS42.items() if v is not None and k != TICKER}
    if len(_av42) >= 4:
        _adf42 = pd.DataFrame(_av42)
        _amean42 = _adf42.mean(axis=1)
        _acorr42 = pd.DataFrame({k: _adf42[k].rolling(20).corr(_amean42) for k in _adf42.columns}).mean(axis=1)
        feat['xast_avg_corr_20'] = _acorr42
        feat['xast_corr_high'] = (_acorr42 > 0.6).astype(float)
        feat['xast_corr_zscore_60'] = calc_zscore(_acorr42, 60)
        _ndown42 = (_adf42.rolling(5).sum() < 0).sum(axis=1)
        feat['xast_assets_down_5d'] = _ndown42
        feat['xast_no_hiding_place'] = (_ndown42 >= max(4, len(_av42) - 1)).astype(float)
        _safe_down = pd.Series(0.0, index=feat.index)
        if 'TLT' in _adf42.columns:
            _safe_down = _safe_down + (_adf42['TLT'].rolling(5).sum() < 0).astype(float)
        if 'GLD' in _adf42.columns:
            _safe_down = _safe_down + (_adf42['GLD'].rolling(5).sum() < 0).astype(float)
        feat['xast_safe_assets_down'] = _safe_down
        feat['xast_correlation_breakdown'] = ((_acorr42 > 0.55) & (_amean42.rolling(5).mean() < 0)).astype(float)

    # ── 42D. [위험선호 카나리아형] 크립토 (BTC) ──
    _btc = closes.get('BTC-USD')
    if _btc is not None:
        _btcr = _btc.pct_change()
        feat['btc_ret_5d'] = _btc.pct_change(5)
        feat['btc_drawdown_20d'] = _btc / _btc.rolling(20).max() - 1
        feat['btc_below_sma50'] = (_btc < _btc.rolling(50).mean()).astype(float)
        feat['btc_realvol_20'] = _btcr.rolling(20).std() * np.sqrt(365)
        feat['btc_vol_spike'] = (_btcr.rolling(10).std() > _btcr.rolling(60).std() * 1.5).astype(float)
        feat['btc_crash_5d'] = (_btc.pct_change(5) < -0.12).astype(float)
        feat['btc_leads_equity_down'] = ((_btc.pct_change(5) < -0.08) & (_c42.pct_change(5) > -0.01)).astype(float)
        feat['btc_equity_both_down'] = ((_btc.pct_change(3) < -0.05) & (_c42.pct_change(3) < 0)).astype(float)

    # ── 42E. [집중도/메가캡 에어포켓형] ──
    _megas42 = {}
    for _m in ['NVDA', 'MSFT', 'AAPL', 'AMZN', 'GOOGL', 'META', 'TSLA']:
        _sm = closes.get(_m)
        if _sm is not None and _m != TICKER:
            _megas42[_m] = _sm
    if len(_megas42) >= 4:
        _mc42 = pd.DataFrame(_megas42)
        _mr42 = _mc42.pct_change()
        _ngap = pd.DataFrame({k: ((_mc42[k] / _mc42[k].shift(1) - 1) < -0.015).astype(float) for k in _mc42.columns}).sum(axis=1)
        feat['mga_gapdown_count'] = _ngap
        feat['mga_gapdown_cluster'] = (_ngap >= max(3, len(_megas42) // 2)).astype(float)
        feat['mga_single_airpocket'] = (_mr42.min(axis=1) < -0.04).astype(float)
        feat['mga_single_airpocket_5d'] = feat['mga_single_airpocket'].rolling(5).sum()
        feat['mga_worst_mega_ret'] = _mr42.min(axis=1)
        _below20 = pd.DataFrame({k: (_mc42[k] < _mc42[k].rolling(20).mean()).astype(float) for k in _mc42.columns}).sum(axis=1)
        feat['mga_below_sma20_count'] = _below20
        feat['mga_leadership_break'] = (_below20 >= max(4, int(len(_megas42) * 0.6))).astype(float)
        feat['mga_mega_weak_index_firm'] = ((_mr42.mean(axis=1).rolling(5).mean() < 0) &
                                            (_c42.pct_change(5) > 0)).astype(float)
        _all_up_prev = (_mr42.shift(1) > 0).sum(axis=1) >= len(_megas42) - 1
        feat['mga_blowoff_then_red'] = (_all_up_prev & ((_mr42 < 0).sum(axis=1) >= len(_megas42) - 1)).astype(float)

    # ── 42F. 유형별 1일 선행 종합 경보 v2 ──
    _pre2 = pd.Series(0.0, index=_c42.index)
    for _k in ['fnd_funding_stress_high', 'fnd_cp_stress_flag', 'pcr_complacency_at_high',
               'xast_no_hiding_place', 'xast_correlation_breakdown', 'btc_leads_equity_down',
               'mga_gapdown_cluster', 'mga_single_airpocket', 'mga_leadership_break']:
        if _k in feat.columns:
            _pre2 = _pre2 + feat[_k].fillna(0)
    feat['pre2_bytype_score'] = _pre2
    feat['pre2_bytype_high'] = (_pre2 >= 3).astype(float)
    feat['pre2_bytype_extreme'] = (_pre2 >= 5).astype(float)
    feat['pre2_bytype_rising'] = (_pre2 > _pre2.shift(3)).astype(float)
    for _p in [3, 5]:
        feat[f'pre2_bytype_sum_{_p}d'] = _pre2.rolling(_p).sum()
    feat['pre2_bytype_zscore_60'] = calc_zscore(_pre2, 60)

    feat.replace([np.inf, -np.inf], np.nan, inplace=True)
    print(f"  계산된 피처 수: {len(feat.columns)}개")
    return feat


# ════════════════════════════════════════════════════════════════
#                     타겟 & 스윕
# ════════════════════════════════════════════════════════════════
def make_target(close: pd.Series, n: int, m_pct: float) -> pd.Series:
    """t 시점 기준 미래 n일 누적수익률이 -m_pct% 이하면 1."""
    future_ret = close.shift(-n) / close - 1.0
    y = (future_ret <= -m_pct / 100.0).astype(float)
    y[future_ret.isna()] = np.nan
    return y

def make_rise_target(close: pd.Series, n: int, m_pct: float) -> pd.Series:
    """
    t 시점 기준 미래 n일 누적수익률이 +m_pct% 이상이면 1 (상승 타겟).
    하락 타겟(make_target)의 정반대 방향 — 같은 (n, m)으로 상승 예측 학습용.
    """
    future_ret = close.shift(-n) / close - 1.0
    y = (future_ret >= m_pct / 100.0).astype(float)
    y[future_ret.isna()] = np.nan
    return y

def sweep_feature_vec(x, y):
    """단일 지표 벡터화 임계치 스윕 → 최고 F1."""
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    mask = ~np.isnan(x) & ~np.isnan(y)
    x_, y_ = x[mask], y[mask].astype(int)
    N = len(x_)
    if N < 50: return None
    n_pos = int(y_.sum())
    if n_pos < MIN_ACTUAL_POS or n_pos == N: return None

    lo_v, hi_v = np.percentile(x_, [PCTL_LO, PCTL_HI])
    if hi_v <= lo_v: return None
    thresholds = np.linspace(lo_v, hi_v, N_THRESHOLDS)
    base_rate = n_pos / N

    # 스윕 범위 메타정보 (어디서 어디까지 얼마 단위로 탐색했는지)
    step = (hi_v - lo_v) / (N_THRESHOLDS - 1) if N_THRESHOLDS > 1 else 0.0
    raw_min = float(np.nanmin(x_))
    raw_max = float(np.nanmax(x_))

    best, best_f1 = None, -1.0
    for direction in ('>=', '<='):
        if direction == '>=':
            pred = x_[None, :] >= thresholds[:, None]
        else:
            pred = x_[None, :] <= thresholds[:, None]
        n_pred = pred.sum(axis=1).astype(float)
        tp = (pred & (y_[None, :] == 1)).sum(axis=1).astype(float)
        with np.errstate(divide='ignore', invalid='ignore'):
            precision = np.where(n_pred > 0, tp / n_pred, 0.0)
            recall    = tp / n_pos
            f1 = np.where(precision + recall > 0,
                          2 * precision * recall / (precision + recall), 0.0)
        pred_rate = n_pred / N
        valid = (pred_rate >= MIN_PRED_RATE) & (pred_rate <= MAX_PRED_RATE) & (tp > 0)
        f1_valid = np.where(valid, f1, -1.0)
        idx = int(f1_valid.argmax())
        if f1_valid[idx] > best_f1:
            best_f1 = f1_valid[idx]
            best = dict(
                threshold=float(thresholds[idx]),
                direction=direction,
                precision=float(precision[idx]),
                recall=float(recall[idx]),
                f1=float(f1[idx]),
                lift=float(precision[idx] / base_rate) if base_rate > 0 else np.nan,
                tp=int(tp[idx]),
                fp=int(n_pred[idx] - tp[idx]),
                fn=int(n_pos - tp[idx]),
                tn=int(N - n_pred[idx] - (n_pos - tp[idx])),
                n_samples=N,
                base_rate=float(base_rate),
                pred_rate=float(pred_rate[idx]),
                # 탐색 범위 메타데이터
                sweep_min=float(lo_v),
                sweep_max=float(hi_v),
                sweep_step=float(step),
                sweep_n=int(N_THRESHOLDS),
                raw_min=raw_min,
                raw_max=raw_max,
                threshold_idx=int(idx),     # 0~39 중 어디에서 최적인지
            )
    return best if best_f1 > 0 else None


def evaluate_all_features(feat_df, y):
    rows = []
    y_vals = y.values if hasattr(y, 'values') else y
    for col in feat_df.columns:
        r = sweep_feature_vec(feat_df[col].values, y_vals)
        if r is not None:
            r['feature'] = col
            rows.append(r)
    if not rows: return pd.DataFrame()
    df = pd.DataFrame(rows)
    cols = ['feature', 'direction', 'threshold', 'f1', 'precision', 'recall',
            'lift', 'base_rate', 'pred_rate', 'tp', 'fp', 'fn', 'tn', 'n_samples',
            'sweep_min', 'sweep_max', 'sweep_step', 'sweep_n',
            'raw_min', 'raw_max', 'threshold_idx']
    return df[cols].sort_values('f1', ascending=False).reset_index(drop=True)


def grid_search(feat_df, close, n_range, m_range, verbose=True):
    """(n, m) 그리드 탐색 — 조합마다 450개 지표 전부 임계치 스윕."""
    import time
    grid_rows, all_results = [], {}
    total_combos = len(n_range) * len(m_range)
    total_sweeps = 0            # 실제 수행된 (지표 × (n,m)) 조합 수
    n_feat = len(feat_df.columns)
    t_global = time.time()

    print(f"  총 {total_combos}개 (n, m) 조합 × 지표 {n_feat}개 = "
          f"최대 {total_combos * n_feat:,}회 지표-스윕")
    print(f"  임계치 단계 {N_THRESHOLDS}개 × 방향 2 → "
          f"최대 {total_combos * n_feat * N_THRESHOLDS * 2:,}회 임계치 평가")
    print()

    done = 0
    for n in n_range:
        for m in m_range:
            done += 1
            t0 = time.time()
            y = make_target(close, n, m)
            common = feat_df.index.intersection(y.dropna().index)
            if len(common) < 100:
                if verbose:
                    print(f"    [{done:2d}/{total_combos}] n={n:>2}일 m={m:>4.1f}%  "
                          f"→ skip (샘플 {len(common)}일 부족)")
                continue
            y_sub = y.loc[common]
            n_actual_pos = int(y_sub.sum())
            if n_actual_pos < MIN_ACTUAL_POS:
                if verbose:
                    print(f"    [{done:2d}/{total_combos}] n={n:>2}일 m={m:>4.1f}%  "
                          f"→ skip (양성 {n_actual_pos}건 < {MIN_ACTUAL_POS})")
                continue
            feat_sub = feat_df.loc[common]
            res = evaluate_all_features(feat_sub, y_sub)
            total_sweeps += n_feat   # ← 이 조합에서 450개 지표 전부 스윕했음을 카운트
            if len(res) == 0:
                if verbose:
                    print(f"    [{done:2d}/{total_combos}] n={n:>2}일 m={m:>4.1f}%  "
                          f"→ 유효 지표 없음")
                continue
            top = res.nlargest(TOP_K_FOR_SCORE, 'f1')
            top_f1   = float(top['f1'].mean())
            top_lift = float(top['lift'].mean())
            pos_rate = float(y_sub.mean())
            # ─────────────────────────────────────────────────────
            # 종합 점수 — 핵심 원칙:
            #   1) F1이 주축. F1 < 0.30이면 사실상 동전던지기로 간주 (강한 페널티)
            #   2) Lift는 부수적 보너스 (이전엔 너무 큰 영향)
            #   3) 양성률 [3%, 30%] 밖이면 강한 페널티
            #      - 너무 희귀(< 3%): F1이 우연히 부풀려질 수 있음
            #      - 너무 흔함(> 30%): 매일 "하락"이라 의미 없음
            #
            # 공식: F1^1.5 × lift_bonus × stability
            #   F1^1.5: F1 0.5 → 0.354, F1 0.3 → 0.164, F1 0.2 → 0.089 (작은 F1 강하게 페널티)
            #   lift_bonus: 1 + 0.3 * log1p(max(Lift-1.2, 0))  (Lift 1.2 미만이면 보너스 없음)
            #   stability: 양성률 페널티
            # ─────────────────────────────────────────────────────
            f1_core = top_f1 ** 1.5
            lift_bonus = 1 + 0.3 * np.log1p(max(top_lift - 1.2, 0))
            if pos_rate < 0.03:
                stability = (pos_rate / 0.03) ** 2          # 1% → 0.11배
            elif pos_rate <= 0.25:
                stability = 1.0                              # 3~25% → 만점
            elif pos_rate <= 0.40:
                stability = max(0.4, 1.0 - (pos_rate - 0.25) * 2)  # 40% → 0.7배
            else:
                stability = 0.3                              # 40%+ → 0.3배 (의미 거의 없음)
            composite_score = f1_core * lift_bonus * stability
            grid_rows.append(dict(
                n=n, m_pct=m,
                positive_rate=pos_rate,
                n_positives=n_actual_pos,
                n_samples=int(len(y_sub)),
                n_valid_features=int(len(res)),
                top20_mean_f1=top_f1,
                top20_mean_precision=float(top['precision'].mean()),
                top20_mean_lift=top_lift,
                composite_score=composite_score,
                max_f1=float(res['f1'].max()),
                best_feature=res.iloc[0]['feature'],
            ))
            all_results[(n, m)] = res
            elapsed = time.time() - t0
            if verbose:
                print(f"    [{done:2d}/{total_combos}] n={n:>2}일 m={m:>4.1f}%  "
                      f"양성 {n_actual_pos:>3}/{len(y_sub)} ({y_sub.mean():6.1%})  "
                      f"유효지표 {len(res):>3}/{n_feat}  "
                      f"최대F1 {res['f1'].max():.4f}  Top20평균F1 {top['f1'].mean():.4f}  "
                      f"[{elapsed:>5.2f}s]")

    total_elapsed = time.time() - t_global
    print()
    print(f"  ✓ 그리드 완료: 실제 지표-스윕 {total_sweeps:,}회  "
          f"(= 임계치 평가 {total_sweeps * N_THRESHOLDS * 2:,}회)  "
          f"총 {total_elapsed:.1f}초")

    return pd.DataFrame(grid_rows), all_results


def train_test_validate(feat_df, y, train_ratio=TRAIN_RATIO, embargo_days=None):
    """
    시간순 Train/Test 분할 + Embargo (완충 구간).

    embargo_days: Train 끝과 Test 시작 사이의 제외 일수.
        타겟이 미래 n일 누적수익률이므로, Train 마지막 n일의 y는 Test 첫 n일의
        가격을 사용해 만들어짐 → embargo_days는 최소 n+1일 권장.
        None이면 호출측에서 best_n 기반으로 설정 (main에서 처리).
    """
    n = len(y)
    if embargo_days is None:
        embargo_days = 0
    split = int(n * train_ratio)
    tr_end = split
    te_start = min(split + embargo_days, n)
    tr_idx = y.index[:tr_end]
    te_idx = y.index[te_start:]

    train_res = evaluate_all_features(feat_df.loc[tr_idx], y.loc[tr_idx])
    if len(train_res) == 0: return None

    rows = []
    y_te_all = y.loc[te_idx].values
    for _, r in train_res.iterrows():
        col, t, d = r['feature'], r['threshold'], r['direction']
        x_te = feat_df.loc[te_idx, col].values
        mask = ~np.isnan(x_te) & ~np.isnan(y_te_all)
        x_te_, y_te_ = x_te[mask], y_te_all[mask].astype(int)
        if len(y_te_) < 20 or y_te_.sum() < 2:
            rows.append(dict(feature=col, test_f1=np.nan, test_prec=np.nan,
                             test_recall=np.nan, test_tp=0, test_fp=0, test_fn=0))
            continue
        pred = (x_te_ >= t) if d == '>=' else (x_te_ <= t)
        tp = int(((pred == 1) & (y_te_ == 1)).sum())
        fp = int(((pred == 1) & (y_te_ == 0)).sum())
        fn = int(((pred == 0) & (y_te_ == 1)).sum())
        prec = tp / (tp + fp) if (tp + fp) > 0 else 0.0
        rec  = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        f1   = 2 * prec * rec / (prec + rec) if (prec + rec) > 0 else 0.0
        rows.append(dict(feature=col, test_f1=f1, test_prec=prec, test_recall=rec,
                         test_tp=tp, test_fp=fp, test_fn=fn))
    test_df = pd.DataFrame(rows)
    merged = train_res.merge(test_df, on='feature', how='left')
    return merged


def walk_forward_cv(feat_df, y, n_splits=5, embargo_days=0, top_k=20):
    """
    Walk-forward 시간순 교차검증.
    예: 데이터 1000일을 5-fold로 나눠
        Fold1: train[0:200],  embargo, test[200+e : 400]
        Fold2: train[0:400],  embargo, test[400+e : 600]
        ...
        Fold5: train[0:800],  embargo, test[800+e : 1000]
    각 fold마다 Top-K 지표의 평균 F1 계산 → fold들의 평균/표준편차 보고
    """
    n = len(y)
    fold_size = n // (n_splits + 1)
    if fold_size < 30:
        return None
    fold_results = []
    for k in range(1, n_splits + 1):
        tr_end = fold_size * k
        te_start = min(tr_end + embargo_days, n)
        te_end = min(te_start + fold_size, n)
        if te_end - te_start < 20: break
        tr_idx = y.index[:tr_end]
        te_idx = y.index[te_start:te_end]
        # Train에서 임계치 학습
        train_res = evaluate_all_features(feat_df.loc[tr_idx], y.loc[tr_idx])
        if len(train_res) == 0: continue
        top = train_res.head(top_k)
        # Test에서 동일 규칙 적용
        y_te = y.loc[te_idx].values
        f1s = []
        for _, r in top.iterrows():
            col = r['feature']; t = r['threshold']; d = r['direction']
            if col not in feat_df.columns: continue
            x_te = feat_df.loc[te_idx, col].values
            mask = ~np.isnan(x_te) & ~np.isnan(y_te)
            x_te_, y_te_ = x_te[mask], y_te[mask].astype(int)
            if len(y_te_) < 10 or y_te_.sum() < 2: continue
            pred = (x_te_ >= t) if d == '>=' else (x_te_ <= t)
            tp = int(((pred == 1) & (y_te_ == 1)).sum())
            fp = int(((pred == 1) & (y_te_ == 0)).sum())
            fn = int(((pred == 0) & (y_te_ == 1)).sum())
            prec = tp / (tp + fp) if (tp + fp) > 0 else 0.0
            rec  = tp / (tp + fn) if (tp + fn) > 0 else 0.0
            f1_v = 2 * prec * rec / (prec + rec) if (prec + rec) > 0 else 0.0
            f1s.append(f1_v)
        if f1s:
            fold_results.append(dict(
                fold=k,
                tr_start=str(tr_idx[0].date()),
                tr_end=str(tr_idx[-1].date()),
                te_start=str(te_idx[0].date()),
                te_end=str(te_idx[-1].date()),
                tr_size=len(tr_idx),
                te_size=len(te_idx),
                top_k_mean_f1=float(np.mean(f1s)),
                top_k_median_f1=float(np.median(f1s)),
            ))
    return pd.DataFrame(fold_results)


def permutation_test(x, y, n_perm=200, top_threshold_count=80):
    """
    Permutation test: y를 무작위 셔플한 후 동일한 임계치 스윕을 수행.

    무작위 셔플된 y에 대해서도 우연히 높은 F1이 나올 수 있는데,
    그 분포를 보고 "관찰된 F1이 우연일 확률" (p-value)을 추정.

    Returns: (observed_f1, null_f1_dist, p_value)
        p_value < 0.05 → 신호가 우연이 아닐 가능성 95%
        p_value >= 0.05 → 우연일 가능성 충분, 신뢰 어려움
    """
    x = np.asarray(x, dtype=float); y = np.asarray(y, dtype=float)
    mask = ~np.isnan(x) & ~np.isnan(y)
    x_, y_ = x[mask], y[mask].astype(int)
    if len(x_) < 100 or y_.sum() < 10: return None
    # 관찰값
    obs = sweep_feature_vec(x_, y_)
    if obs is None: return None
    # null 분포: y를 셔플하고 동일 스윕 반복
    rng = np.random.default_rng(42)
    null_f1 = []
    for _ in range(n_perm):
        y_shuffled = rng.permutation(y_)
        r = sweep_feature_vec(x_, y_shuffled)
        if r is not None:
            null_f1.append(r['f1'])
    null_f1 = np.array(null_f1) if null_f1 else np.array([0.0])
    p_val = float(np.mean(null_f1 >= obs['f1']))
    return dict(observed_f1=obs['f1'], null_mean=float(null_f1.mean()),
                null_max=float(null_f1.max()), p_value=p_val,
                n_perm=len(null_f1))


def _extract_topk_signal_strength(feat, best_res, top_k, dates):
    """
    Top-K 지표 (다양성 필터 적용) → 신호 매트릭스 + 신호 강도.
    드롭/상승 양쪽 앙상블에서 공통으로 사용.

    Returns: (signal_df, sig_count, sig_valid, sig_strength, top_k_res)
    """
    # ── 다양성 필터: 상관 너무 높은 지표는 후순위 제거 ──
    if DIVERSITY_FILTER and len(best_res) > top_k:
        candidates = best_res.head(min(top_k * 5, len(best_res))).reset_index(drop=True)
        cand_signals = pd.DataFrame(index=feat.index)
        for _, r in candidates.iterrows():
            col = r['feature']
            if col not in feat.columns: continue
            x = feat[col]
            if r['direction'] == '>=':
                sig = (x >= r['threshold']).astype(float)
            else:
                sig = (x <= r['threshold']).astype(float)
            sig[x.isna()] = np.nan
            cand_signals[col] = sig
        selected_cols = []
        for _, r in candidates.iterrows():
            col = r['feature']
            if col not in cand_signals.columns: continue
            if not selected_cols:
                selected_cols.append(col); continue
            sig_new = cand_signals[col]
            redundant = False
            for prev_col in selected_cols:
                c = sig_new.corr(cand_signals[prev_col])
                if pd.notna(c) and abs(c) >= DIVERSITY_CORR_LIMIT:
                    redundant = True; break
            if not redundant:
                selected_cols.append(col)
            if len(selected_cols) >= top_k: break
        if len(selected_cols) < top_k:
            for _, r in best_res.head(top_k * 3).iterrows():
                if r['feature'] not in selected_cols:
                    selected_cols.append(r['feature'])
                if len(selected_cols) >= top_k: break
        top_k_res = best_res[best_res['feature'].isin(selected_cols)] \
                       .sort_values('f1', ascending=False).head(top_k).reset_index(drop=True)
    else:
        top_k_res = best_res.head(top_k).reset_index(drop=True)

    # ── 신호 매트릭스 ──
    signal_df = pd.DataFrame(index=dates)
    for _, r in top_k_res.iterrows():
        col = r['feature']
        if col not in feat.columns: continue
        t = r['threshold']; d = r['direction']
        x = feat.loc[dates, col]
        if d == '>=':
            sig = (x >= t).astype(float)
        else:
            sig = (x <= t).astype(float)
        sig[x.isna()] = np.nan
        signal_df[col] = sig

    sig_valid = signal_df.notna().sum(axis=1)
    sig_count = signal_df.sum(axis=1, skipna=True)
    sig_strength = sig_count / sig_valid.replace(0, np.nan)
    return signal_df, sig_count, sig_valid, sig_strength, top_k_res


# ════════════════════════════════════════════════════════════════
#                   일별 예측 (2025년~현재)
# ════════════════════════════════════════════════════════════════
def build_daily_predictions(feat, close, best_res, best_nm,
                             top_k=None,
                             start=None,
                             vote_thresh=None):
    """
    최적 (n, m)과 Top-K 지표별 최적 임계치를 사용해
    start 이후 매일 개별 신호 + 앙상블 하락 예측 생성.

    None인 인자는 호출 시점의 모듈 전역값(DAILY_START 등)을 읽음
    (default 인자는 모듈 로드 시 박히므로 노트북 수정 반영 안됨).
    """
    if top_k       is None: top_k       = ENSEMBLE_TOP_K
    if start       is None: start       = DAILY_START
    if vote_thresh is None: vote_thresh = ENSEMBLE_VOTE_THRESH
    n, m = best_nm
    future_ret = close.shift(-n) / close - 1
    y_future = (future_ret <= -m / 100.0).astype(float)
    y_future[future_ret.isna()] = np.nan

    mask = feat.index >= pd.Timestamp(start)
    dates = feat.index[mask]
    if len(dates) == 0:
        return None, None

    # 다양성 필터: 상관 너무 높은 지표는 후순위 제거 → 진짜 다양한 Top-K 확보
    if DIVERSITY_FILTER and len(best_res) > top_k:
        candidates = best_res.head(min(top_k * 5, len(best_res))).reset_index(drop=True)
        cand_signals = pd.DataFrame(index=feat.index)
        for _, r in candidates.iterrows():
            col = r['feature']
            if col not in feat.columns: continue
            x = feat[col]
            if r['direction'] == '>=':
                sig = (x >= r['threshold']).astype(float)
            else:
                sig = (x <= r['threshold']).astype(float)
            sig[x.isna()] = np.nan
            cand_signals[col] = sig
        # F1 내림차순 그리디 선택: 이미 선택된 것과 상관 < limit인 것만
        selected_cols = []
        for _, r in candidates.iterrows():
            col = r['feature']
            if col not in cand_signals.columns: continue
            if not selected_cols:
                selected_cols.append(col); continue
            sig_new = cand_signals[col]
            redundant = False
            for prev_col in selected_cols:
                c = sig_new.corr(cand_signals[prev_col])
                if pd.notna(c) and abs(c) >= DIVERSITY_CORR_LIMIT:
                    redundant = True; break
            if not redundant:
                selected_cols.append(col)
            if len(selected_cols) >= top_k: break
        # 그래도 부족하면 원래 Top-K로 보충
        if len(selected_cols) < top_k:
            for _, r in best_res.head(top_k * 3).iterrows():
                if r['feature'] not in selected_cols:
                    selected_cols.append(r['feature'])
                if len(selected_cols) >= top_k: break
        top_k_res = best_res[best_res['feature'].isin(selected_cols)] \
                       .sort_values('f1', ascending=False).head(top_k).reset_index(drop=True)
    else:
        top_k_res = best_res.head(top_k).reset_index(drop=True)

    # 각 Top-K 지표에 대해 "그 날짜에 신호가 켜졌는가" (1/0/NaN) 벡터화 계산
    signal_df = pd.DataFrame(index=dates)
    for _, r in top_k_res.iterrows():
        col = r['feature']
        if col not in feat.columns:
            continue
        t = r['threshold']; d = r['direction']
        x = feat.loc[dates, col]
        if d == '>=':
            sig = (x >= t).astype(float)
        else:
            sig = (x <= t).astype(float)
        sig[x.isna()] = np.nan
        signal_df[col] = sig

    sig_valid    = signal_df.notna().sum(axis=1)
    sig_count    = signal_df.sum(axis=1, skipna=True)
    sig_strength = sig_count / sig_valid.replace(0, np.nan)
    ensemble_pred = (sig_strength >= vote_thresh).astype(float)
    # 유효 신호가 너무 적으면 예측 불가 (K에 비례, 최소 2개)
    min_valid = max(2, int(top_k * 0.3))
    ensemble_pred[sig_valid < min_valid] = np.nan

    daily = pd.DataFrame({
        'date':            dates,
        'close':           close.reindex(dates).values,
        'future_ret_pct':  future_ret.reindex(dates).values * 100,
        'actual_drop':     y_future.reindex(dates).values,
        'signal_count':    sig_count.values.astype(int),
        'signal_valid':    sig_valid.values.astype(int),
        'signal_strength': sig_strength.values,
        'ensemble_pred':   ensemble_pred.values,
    })

    def _result(row):
        if pd.isna(row['actual_drop']):   return '미관측'
        if pd.isna(row['ensemble_pred']): return '?'
        if row['ensemble_pred'] == 1 and row['actual_drop'] == 1: return 'TP ✓'
        if row['ensemble_pred'] == 1 and row['actual_drop'] == 0: return 'FP ✗'
        if row['ensemble_pred'] == 0 and row['actual_drop'] == 1: return 'FN ✗'
        return 'TN ✓'
    daily['result'] = daily.apply(_result, axis=1)

    # 개별 지표 신호 컬럼 부착
    for col in signal_df.columns:
        daily[f'_sig_{col}'] = signal_df[col].values

    return daily, top_k_res

def build_daily_predictions_with_rise(
    feat, close, best_drop_res, best_rise_res, best_nm,
    top_k_drop, top_k_rise,
    start=None, vote_thresh=None, rise_offset_weight=None,
):
    """
    드롭 신호 - 상승 신호 상쇄 → 순(net) 신호로 하락 예측.

    net_strength = drop_strength - rise_offset_weight × rise_strength
    ensemble_pred = (net_strength >= vote_thresh)

    예: 드롭 70% + 상승 60%, weight=1.0 → net=10% → 50% 미달 → 정상 예측 (FP 회피)
        드롭 70% + 상승 10%, weight=1.0 → net=60% → 50% 초과 → 하락 예측 (TP 유지)
    """
    if start              is None: start              = DAILY_START
    if vote_thresh        is None: vote_thresh        = ENSEMBLE_VOTE_THRESH
    if rise_offset_weight is None: rise_offset_weight = RISE_OFFSET_WEIGHT

    n, m = best_nm
    future_ret = close.shift(-n) / close - 1
    y_drop = (future_ret <= -m / 100.0).astype(float)
    y_drop[future_ret.isna()] = np.nan

    dates = feat.index[feat.index >= pd.Timestamp(start)]
    if len(dates) == 0:
        return None, None, None

    # 드롭 신호 추출 (Top-K 드롭 지표)
    drop_sig_df, drop_count, drop_valid, drop_strength, drop_topk_res = \
        _extract_topk_signal_strength(feat, best_drop_res, top_k_drop, dates)
    # 상승 신호 추출 (Top-K 상승 지표)
    rise_sig_df, rise_count, rise_valid, rise_strength, rise_topk_res = \
        _extract_topk_signal_strength(feat, best_rise_res, top_k_rise, dates)

    # 상승 신호 NaN을 0으로 처리 (상승 신호가 없는 날 = 상쇄 안 함)
    rise_strength_filled = rise_strength.fillna(0)
    net_strength = drop_strength - rise_offset_weight * rise_strength_filled

    # 최종 예측
    ensemble_pred = (net_strength >= vote_thresh).astype(float)
    min_valid = max(2, int(top_k_drop * 0.3))
    ensemble_pred[drop_valid < min_valid] = np.nan

    # 결과 DataFrame — 기존 sheet writer와 호환되도록 signal_count/signal_strength 동일 컬럼명 사용
    daily = pd.DataFrame({
        'date':              dates,
        'close':             close.reindex(dates).values,
        'future_ret_pct':    future_ret.reindex(dates).values * 100,
        'actual_drop':       y_drop.reindex(dates).values,
        'signal_count':      drop_count.values.astype(int),     # 호환: 드롭 카운트
        'signal_valid':      drop_valid.values.astype(int),
        'signal_strength':   net_strength.values,                # 호환: 이제는 net 강도
        # 추가 — 드롭/상승 분리 추적
        'drop_strength':     drop_strength.values,
        'rise_signal_count': rise_count.values.astype(int),
        'rise_signal_valid': rise_valid.values.astype(int),
        'rise_strength':     rise_strength.values,
        'net_strength':      net_strength.values,
        'ensemble_pred':     ensemble_pred.values,
    })

    def _result(row):
        if pd.isna(row['actual_drop']):   return '미관측'
        if pd.isna(row['ensemble_pred']): return '?'
        if row['ensemble_pred'] == 1 and row['actual_drop'] == 1: return 'TP ✓'
        if row['ensemble_pred'] == 1 and row['actual_drop'] == 0: return 'FP ✗'
        if row['ensemble_pred'] == 0 and row['actual_drop'] == 1: return 'FN ✗'
        return 'TN ✓'
    daily['result'] = daily.apply(_result, axis=1)

    # 개별 지표 신호 부착 (드롭은 _sig_ — 기존 sheet writer 호환)
    for col in drop_sig_df.columns:
        daily[f'_sig_{col}'] = drop_sig_df[col].values
    # 상승 신호는 별도 prefix로 부착 (디버그/조회용)
    for col in rise_sig_df.columns:
        daily[f'_rise_sig_{col}'] = rise_sig_df[col].values

    return daily, drop_topk_res, rise_topk_res


def summarize_daily(daily):
    """일별 예측의 관측 가능한 부분에 대한 혼동행렬 요약."""
    obs = daily[daily['actual_drop'].notna() & daily['ensemble_pred'].notna()]
    if len(obs) == 0:
        return None
    tp = int(((obs['ensemble_pred'] == 1) & (obs['actual_drop'] == 1)).sum())
    fp = int(((obs['ensemble_pred'] == 1) & (obs['actual_drop'] == 0)).sum())
    fn = int(((obs['ensemble_pred'] == 0) & (obs['actual_drop'] == 1)).sum())
    tn = int(((obs['ensemble_pred'] == 0) & (obs['actual_drop'] == 0)).sum())
    prec = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    rec  = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    f1   = 2 * prec * rec / (prec + rec) if (prec + rec) > 0 else 0.0
    base = (tp + fn) / (tp + fp + fn + tn) if (tp + fp + fn + tn) > 0 else 0.0
    lift = prec / base if base > 0 else np.nan
    return dict(n_obs=len(obs), tp=tp, fp=fp, fn=fn, tn=tn,
                precision=prec, recall=rec, f1=f1, base_rate=base, lift=lift)


def sweep_top_k(feat, close, best_res, best_nm,
                k_range=None,
                vote_thresh=None,
                eval_start=None):
    """
    Top-K 앙상블 사이즈를 스윕해 가장 예측력 좋은 K 찾기.
    평가 기간 기본: 전체 샘플(2022~).
    """
    if k_range     is None: k_range     = K_RANGE
    if vote_thresh is None: vote_thresh = ENSEMBLE_VOTE_THRESH
    if eval_start is None:
        eval_start = feat.index[0].strftime('%Y-%m-%d')
    n_avail = len(best_res)
    k_candidates = [k for k in k_range if k <= n_avail]

    rows = []
    for k in k_candidates:
        daily, _ = build_daily_predictions(
            feat, close, best_res, best_nm,
            top_k=k, start=eval_start, vote_thresh=vote_thresh)
        s = summarize_daily(daily)
        if s is None:
            continue
        rows.append(dict(
            top_k=k,
            f1=s['f1'], precision=s['precision'], recall=s['recall'],
            lift=s['lift'],
            tp=s['tp'], fp=s['fp'], fn=s['fn'], tn=s['tn'],
            n_obs=s['n_obs'], base_rate=s['base_rate'],
        ))
    df = pd.DataFrame(rows)
    if len(df) == 0:
        return df, ENSEMBLE_TOP_K, 0.0
    idx = df['f1'].idxmax()
    best_k = int(df.loc[idx, 'top_k'])
    best_f1 = float(df.loc[idx, 'f1'])
    return df, best_k, best_f1

def sweep_top_k_rise(feat, close, best_rise_res, best_nm,
                     k_range=None, vote_thresh=None, eval_start=None):
    """
    상승 Top-K 사이즈 스윕 — 상승 앙상블이 상승 자체를 얼마나 잘 예측하는가로 K 결정.
    드롭의 sweep_top_k와 같은 논리, 단지 타겟이 상승이고 매트릭이 상승 F1.
    """
    if k_range     is None: k_range     = RISE_K_RANGE
    if vote_thresh is None: vote_thresh = RISE_VOTE_THRESH
    if eval_start  is None: eval_start  = feat.index[0].strftime('%Y-%m-%d')

    n, m = best_nm
    future_ret = close.shift(-n) / close - 1
    y_rise = (future_ret >= m / 100.0).astype(float)
    y_rise[future_ret.isna()] = np.nan

    dates = feat.index[feat.index >= pd.Timestamp(eval_start)]
    n_avail = len(best_rise_res)
    k_candidates = [k for k in k_range if k <= n_avail]

    rows = []
    for k in k_candidates:
        _, _, _, rise_strength, _ = _extract_topk_signal_strength(
            feat, best_rise_res, k, dates)
        pred = (rise_strength >= vote_thresh).astype(float)
        pred[rise_strength.isna()] = np.nan

        obs = pd.DataFrame({
            'pred': pred.values,
            'y':    y_rise.reindex(dates).values
        }).dropna()
        if len(obs) == 0: continue
        tp = int(((obs['pred'] == 1) & (obs['y'] == 1)).sum())
        fp = int(((obs['pred'] == 1) & (obs['y'] == 0)).sum())
        fn = int(((obs['pred'] == 0) & (obs['y'] == 1)).sum())
        tn = int(((obs['pred'] == 0) & (obs['y'] == 0)).sum())
        prec = tp / (tp + fp) if (tp + fp) > 0 else 0.0
        rec  = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        f1   = 2 * prec * rec / (prec + rec) if (prec + rec) > 0 else 0.0
        base = (tp + fn) / max(tp + fp + fn + tn, 1)
        lift = prec / base if base > 0 else 0.0
        rows.append(dict(
            top_k=k, f1=f1, precision=prec, recall=rec, lift=lift,
            tp=tp, fp=fp, fn=fn, tn=tn, n_obs=len(obs), base_rate=base,
        ))

    df = pd.DataFrame(rows)
    if len(df) == 0:
        return df, ENSEMBLE_TOP_K, 0.0
    idx = df['f1'].idxmax()
    return df, int(df.loc[idx, 'top_k']), float(df.loc[idx, 'f1'])

def sweep_rise_offset_weight(
    feat, close, best_drop_res, best_rise_res, best_nm,
    top_k_drop, top_k_rise,
    weight_range=None, start=None, vote_thresh=None,
):
    """
    상승 신호 상쇄 가중치(weight) 스윕 — 일별 하락 예측 F1을 최대화하는 weight 탐색.
    weight=0.0 (= baseline, 상쇄 없음)부터 weight=1.5+ (강한 상쇄)까지 시도.

    각 weight마다 build_daily_predictions_with_rise를 호출하고 summarize_daily로
    F1 계산 → 최고 F1을 내는 weight 자동 선정.
    """
    if weight_range is None: weight_range = RISE_WEIGHT_RANGE
    if start        is None: start        = DAILY_START
    if vote_thresh  is None: vote_thresh  = ENSEMBLE_VOTE_THRESH

    rows = []
    for w in weight_range:
        daily, _, _ = build_daily_predictions_with_rise(
            feat, close, best_drop_res, best_rise_res, best_nm,
            top_k_drop=top_k_drop, top_k_rise=top_k_rise,
            start=start, vote_thresh=vote_thresh, rise_offset_weight=w)
        s = summarize_daily(daily) if daily is not None else None
        if s is None: continue
        rows.append(dict(
            weight=float(w),
            f1=s['f1'], precision=s['precision'], recall=s['recall'],
            lift=s['lift'],
            tp=s['tp'], fp=s['fp'], fn=s['fn'], tn=s['tn'],
            n_obs=s['n_obs'], base_rate=s['base_rate'],
        ))
    df = pd.DataFrame(rows)
    if len(df) == 0:
        return df, RISE_OFFSET_WEIGHT, 0.0
    idx = df['f1'].idxmax()
    return df, float(df.loc[idx, 'weight']), float(df.loc[idx, 'f1'])

# ════════════════════════════════════════════════════════════════
#         대안 예측 모델 5종 (기존 단일 임계치와 비교용)
# ════════════════════════════════════════════════════════════════
def _make_signal_matrix(feat, best_res, top_k, dates):
    """공통 헬퍼: Top-K 지표의 0/1 신호 매트릭스 (date × feature)."""
    top_k_res = best_res.head(top_k).reset_index(drop=True)
    sig_df = pd.DataFrame(index=dates)
    for _, r in top_k_res.iterrows():
        col = r['feature']
        if col not in feat.columns: continue
        x = feat.loc[dates, col]
        if r['direction'] == '>=':
            sig = (x >= r['threshold']).astype(float)
        else:
            sig = (x <= r['threshold']).astype(float)
        sig[x.isna()] = np.nan
        sig_df[col] = sig
    return sig_df


def _eval_predictions(pred, y, name='Model'):
    """공통 헬퍼: 예측 벡터로부터 혼동행렬과 메트릭 계산."""
    obs = pd.DataFrame({'pred': pred, 'y': y}).dropna()
    if len(obs) == 0:
        return None
    tp = int(((obs['pred'] == 1) & (obs['y'] == 1)).sum())
    fp = int(((obs['pred'] == 1) & (obs['y'] == 0)).sum())
    fn = int(((obs['pred'] == 0) & (obs['y'] == 1)).sum())
    tn = int(((obs['pred'] == 0) & (obs['y'] == 0)).sum())
    prec = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    rec  = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    f1   = 2 * prec * rec / (prec + rec) if (prec + rec) > 0 else 0.0
    base = (tp + fn) / max(tp + fp + fn + tn, 1)
    lift = prec / base if base > 0 else np.nan
    return dict(name=name, n_obs=len(obs), tp=tp, fp=fp, fn=fn, tn=tn,
                precision=prec, recall=rec, f1=f1, base_rate=base, lift=lift,
                pred_rate=(tp + fp) / max(len(obs), 1))


# ── 모델 1: 가중 다수결 (F1 가중치) ─────────────────────────────
def model_weighted_vote(feat, close, best_res, best_nm, top_k=10,
                         start=None, threshold=0.5):
    """
    각 지표의 신호를 F1 점수로 가중합 → 임계치 초과시 하락 예측.

    기존 다수결 (단순 평균)과 비교:
    - 기존: 10개 중 5개 ON → 0.5
    - 가중: F1 0.5 지표 5개 + F1 0.3 지표 5개 = (0.5×5 + 0.3×0)/(0.5+0.3)*10 = 가중 점수

    이론적 이점: 신뢰도 높은 지표가 더 큰 영향력 → 정밀도 향상 예상.
    """
    if start is None: start = DAILY_START
    n, m = best_nm
    future_ret = close.shift(-n) / close - 1
    y_future = (future_ret <= -m / 100.0).astype(float)
    y_future[future_ret.isna()] = np.nan

    dates = feat.index[feat.index >= pd.Timestamp(start)]
    sig_df = _make_signal_matrix(feat, best_res, top_k, dates)
    weights = best_res.head(top_k).set_index('feature')['f1'].reindex(sig_df.columns).values
    weights = weights / max(weights.sum(), 1e-9)

    # 가중 신호 점수
    sig_arr = sig_df.values
    valid_mask = ~np.isnan(sig_arr)
    weight_mat = weights[None, :] * valid_mask
    weight_sum = weight_mat.sum(axis=1)
    score = np.where(weight_sum > 0,
                     np.nansum(sig_arr * weights[None, :], axis=1) / weight_sum,
                     np.nan)
    pred = (score >= threshold).astype(float)
    pred[np.isnan(score)] = np.nan

    return _eval_predictions(pred, y_future.reindex(dates).values, '가중다수결')


# ── 모델 2: 동시 신호 (AND - 최소 K개 ON 필수) ──────────────────
def model_unanimous(feat, close, best_res, best_nm, top_k=5,
                     start=None, min_required=None):
    """
    상위 K개 지표 중 최소 N개가 동시에 ON일 때만 하락 예측.

    이론적 이점: 정밀도 극대화 (false alarm 최소).
    예: 5개 중 4개 동시 ON 요구 → 보수적이지만 정확
    """
    if start is None: start = DAILY_START
    if min_required is None: min_required = max(2, int(top_k * 0.8))   # 80% 이상
    n, m = best_nm
    future_ret = close.shift(-n) / close - 1
    y_future = (future_ret <= -m / 100.0).astype(float)
    y_future[future_ret.isna()] = np.nan

    dates = feat.index[feat.index >= pd.Timestamp(start)]
    sig_df = _make_signal_matrix(feat, best_res, top_k, dates)
    sig_count = sig_df.sum(axis=1, skipna=True)
    sig_valid = sig_df.notna().sum(axis=1)
    pred = (sig_count >= min_required).astype(float)
    pred[sig_valid < max(2, top_k // 2)] = np.nan

    return _eval_predictions(pred.values, y_future.reindex(dates).values,
                              f'동시신호({min_required}/{top_k})')


# ── 모델 3: 로지스틱 회귀 (Train구간 학습, Test 예측) ────────────
def model_logistic(feat, close, best_res, best_nm, top_k=20,
                    start=None, train_ratio=0.70, embargo_days=0):
    """
    Top-K 지표의 0/1 신호를 입력으로 로지스틱 회귀 학습.
    Train 구간에서 회귀 계수 추정 → 전 구간 예측.

    이론적 이점: 지표간 비선형 상호작용 학습. 단, 충분한 양성 샘플 필요.
    """
    if start is None: start = DAILY_START
    n, m = best_nm
    future_ret = close.shift(-n) / close - 1
    y = (future_ret <= -m / 100.0).astype(float)
    y[future_ret.isna()] = np.nan
    y_clean = y.dropna()
    if len(y_clean) < 200 or y_clean.sum() < 30:
        return None

    # 모든 데이터 기간으로 신호 매트릭스 생성
    sig_all = _make_signal_matrix(feat, best_res, top_k, y_clean.index).fillna(0)
    if sig_all.shape[1] == 0: return None

    # Train/Test 분할 (embargo 적용)
    n_data = len(y_clean)
    tr_end = int(n_data * train_ratio)
    te_start = min(tr_end + embargo_days, n_data)
    tr_idx = y_clean.index[:tr_end]
    te_idx = y_clean.index[te_start:]
    if len(te_idx) < 20: return None

    # 간단한 로지스틱 회귀 (수치적 안정성을 위해 sklearn 없이 numpy로)
    X_tr = sig_all.loc[tr_idx].values
    y_tr = y_clean.loc[tr_idx].values
    if y_tr.sum() < 5: return None
    # 절편 + L2 정규화 ridge logistic
    X_tr = np.column_stack([np.ones(len(X_tr)), X_tr])
    n_feat = X_tr.shape[1]
    w = np.zeros(n_feat)
    lr = 0.1; reg = 0.1
    for _ in range(200):
        z = X_tr @ w
        p = 1 / (1 + np.exp(-np.clip(z, -30, 30)))
        grad = X_tr.T @ (p - y_tr) / len(y_tr) + reg * w
        grad[0] = X_tr[:, 0] @ (p - y_tr) / len(y_tr)   # 절편 정규화 안 함
        w -= lr * grad
        if np.linalg.norm(grad) < 1e-5: break

    # Test 구간 예측
    dates_pred = feat.index[(feat.index >= pd.Timestamp(start)) &
                             (feat.index.isin(y_clean.index))]
    sig_pred = _make_signal_matrix(feat, best_res, top_k, dates_pred).fillna(0)
    X_pred = np.column_stack([np.ones(len(sig_pred)), sig_pred.values])
    z_pred = X_pred @ w
    p_pred = 1 / (1 + np.exp(-np.clip(z_pred, -30, 30)))
    pred = (p_pred >= 0.5).astype(float)

    y_pred_actual = y_clean.reindex(dates_pred).values
    return _eval_predictions(pred, y_pred_actual, '로지스틱회귀')


# ── 모델 4: 신호 누적 (최근 N일 누적 강도) ──────────────────────
def model_cumulative(feat, close, best_res, best_nm, top_k=10, lookback=3,
                      start=None, vote_thresh=0.5):
    """
    최근 lookback일 동안의 신호 강도 평균이 임계치 초과 → 하락 예측.

    이론적 이점: 일시적 노이즈 신호 감소, 지속적 신호만 잡음.
    예: 오늘만 5/10 ON이면 노이즈일 수 있지만, 3일 연속 평균 5/10 → 진짜 신호.
    """
    if start is None: start = DAILY_START
    n, m = best_nm
    future_ret = close.shift(-n) / close - 1
    y_future = (future_ret <= -m / 100.0).astype(float)
    y_future[future_ret.isna()] = np.nan

    dates = feat.index[feat.index >= pd.Timestamp(start)]
    sig_df = _make_signal_matrix(feat, best_res, top_k, dates)
    sig_strength = sig_df.sum(axis=1, skipna=True) / sig_df.notna().sum(axis=1).replace(0, np.nan)
    cum_strength = sig_strength.rolling(lookback).mean()
    pred = (cum_strength >= vote_thresh).astype(float)
    pred[cum_strength.isna()] = np.nan

    return _eval_predictions(pred.values, y_future.reindex(dates).values,
                              f'누적신호({lookback}일)')


# ── 모델 5: 다단계 예측 (강도별 신뢰도 분리) ────────────────────
def model_tiered(feat, close, best_res, best_nm, top_k=10,
                  start=None, low=0.4, high=0.7):
    """
    신호 강도를 3단계로 분류:
    - 강도 < low      → 정상 (예측: 0)
    - low <= 강도 < high → 약한 경고 (관찰만, 예측: 0)
    - 강도 >= high    → 강한 경고 (예측: 1)

    이론적 이점: 정밀도 극대화. high 구간에서만 행동하면 false alarm 줄어듦.
    """
    if start is None: start = DAILY_START
    n, m = best_nm
    future_ret = close.shift(-n) / close - 1
    y_future = (future_ret <= -m / 100.0).astype(float)
    y_future[future_ret.isna()] = np.nan

    dates = feat.index[feat.index >= pd.Timestamp(start)]
    sig_df = _make_signal_matrix(feat, best_res, top_k, dates)
    sig_strength = sig_df.sum(axis=1, skipna=True) / sig_df.notna().sum(axis=1).replace(0, np.nan)
    pred = (sig_strength >= high).astype(float)
    pred[sig_strength.isna()] = np.nan

    return _eval_predictions(pred.values, y_future.reindex(dates).values,
                              f'고강도신호(>={high:.0%})')


# ── 모델 0: 기존 방식 (단순 다수결) — 비교 기준점 ──────────────
def model_baseline(feat, close, best_res, best_nm, top_k=10,
                    start=None, vote_thresh=0.5):
    """기존 방식: Top-K 다수결 (>= vote_thresh)."""
    if start is None: start = DAILY_START
    n, m = best_nm
    future_ret = close.shift(-n) / close - 1
    y_future = (future_ret <= -m / 100.0).astype(float)
    y_future[future_ret.isna()] = np.nan

    dates = feat.index[feat.index >= pd.Timestamp(start)]
    sig_df = _make_signal_matrix(feat, best_res, top_k, dates)
    sig_strength = sig_df.sum(axis=1, skipna=True) / sig_df.notna().sum(axis=1).replace(0, np.nan)
    pred = (sig_strength >= vote_thresh).astype(float)
    pred[sig_strength.isna()] = np.nan

    return _eval_predictions(pred.values, y_future.reindex(dates).values,
                              f'기존(다수결≥{vote_thresh:.0%})')


def compare_all_models(feat, close, best_res, best_nm, top_k, start=None, embargo_days=0):
    """6개 모델 모두 실행 → 결과 비교 DataFrame 반환."""
    results = []
    # 0. 기존 다수결 (baseline)
    r = model_baseline(feat, close, best_res, best_nm, top_k=top_k, start=start)
    if r: results.append(r)
    # 1. 가중 다수결
    r = model_weighted_vote(feat, close, best_res, best_nm, top_k=top_k, start=start)
    if r: results.append(r)
    # 2. 동시신호 (80% 합의)
    r = model_unanimous(feat, close, best_res, best_nm, top_k=top_k, start=start)
    if r: results.append(r)
    # 3. 로지스틱 회귀 (top_k 더 많이 사용)
    r = model_logistic(feat, close, best_res, best_nm, top_k=min(top_k * 2, 20),
                        start=start, embargo_days=embargo_days)
    if r: results.append(r)
    # 4. 누적신호 (3일)
    r = model_cumulative(feat, close, best_res, best_nm, top_k=top_k,
                          lookback=3, start=start)
    if r: results.append(r)
    # 5. 고강도 신호 (>=70%)
    r = model_tiered(feat, close, best_res, best_nm, top_k=top_k, start=start,
                      low=0.4, high=0.7)
    if r: results.append(r)
    return pd.DataFrame(results)


# ════════════════════════════════════════════════════════════════
#                    Excel 출력 스타일
# ════════════════════════════════════════════════════════════════
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
SUBHDR_FILL = PatternFill("solid", fgColor="2E75B6")
GOOD_FILL   = PatternFill("solid", fgColor="C6EFCE")
MID_FILL    = PatternFill("solid", fgColor="FFEB9C")
BAD_FILL    = PatternFill("solid", fgColor="FFC7CE")
HIGHLIGHT   = PatternFill("solid", fgColor="FFF2CC")
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")
WHITE_BOLD  = Font(bold=True, color='FFFFFF')
_SIDE       = Side(style='thin', color='BDBDBD')
THIN        = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)
_THICK      = Side(style='thick', color='C00000')
THICK_RED   = Border(left=_THICK, right=_THICK, top=_THICK, bottom=_THICK)


def _write_header_row(ws, row, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row, ci)
        c.value = h; c.fill = HEADER_FILL; c.font = WHITE_BOLD
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = THIN


def _f1_fill(f1):
    if f1 >= 0.30: return GOOD_FILL
    if f1 >= 0.15: return MID_FILL
    return BAD_FILL


def write_summary_sheet(wb, feat, best_nm, best_res, grid_df):
    ws = wb.create_sheet('요약', 0)
    ws.sheet_view.showGridLines = False
    n, m = best_nm

    ws.cell(1, 1).value = f'{TICKER} 하락 예측 임계치 탐색 — 요약'
    ws.cell(1, 1).font = Font(bold=True, size=16, color='1F3864')
    ws.merge_cells('A1:D1')

    row_best = grid_df[(grid_df['n'] == n) & (grid_df['m_pct'] == m)].iloc[0]
    rows = [
        ('평가 기간',                 f"{feat.index[0].date()} ~ {feat.index[-1].date()}"),
        ('샘플 일수',                 len(feat)),
        ('사용 지표 수',              len(feat.columns)),
        ('',                          ''),
        ('★ 최적 n (미래 일수)',      f'{n} 일'),
        ('★ 최적 m (누적 하락률)',    f'{m} %'),
        ('→ 실제 하락 빈도 (양성률)', f"{row_best['positive_rate']:.2%} ({int(row_best['n_positives'])}건)"),
        ('→ Top 20 평균 F1',          f"{row_best['top20_mean_f1']:.4f}"),
        ('→ Top 20 평균 정밀도',      f"{row_best['top20_mean_precision']:.2%}"),
        ('→ Top 20 평균 Lift',        f"{row_best['top20_mean_lift']:.2f}×"),
        ('→ 종합 점수',               f"{row_best['composite_score']:.4f}"),
        ('   (산출 식)',              'F1^1.5 × (1+0.3·log(1+max(Lift-1.2,0))) × 양성률페널티'),
        ('→ 유효 지표 수',            f"{int(row_best['n_valid_features'])} / {len(feat.columns)}"),
        ('',                          ''),
        ('🥇 최고 단일 지표',         best_res.iloc[0]['feature']),
        ('   방향',                   best_res.iloc[0]['direction']),
        ('   임계치',                 f"{best_res.iloc[0]['threshold']:.6f}"),
        ('   F1',                     f"{best_res.iloc[0]['f1']:.4f}"),
        ('   정밀도 / 재현율',        f"{best_res.iloc[0]['precision']:.2%} / {best_res.iloc[0]['recall']:.2%}"),
        ('   Lift (기저 대비)',       f"{best_res.iloc[0]['lift']:.2f}×"),
    ]
    for ri, (k, v) in enumerate(rows, 3):
        c1, c2 = ws.cell(ri, 1), ws.cell(ri, 2)
        c1.value = k; c2.value = v
        c1.font = Font(bold=True, size=11, color='1F3864')
        c2.font = Font(size=11)
        if k.startswith('★'):
            c1.fill = HIGHLIGHT; c2.fill = HIGHLIGHT
            c2.font = Font(bold=True, size=12, color='C00000')

    r = len(rows) + 5
    ws.cell(r, 1).value = '🏆 Top 10 예측 지표'
    ws.cell(r, 1).font = Font(bold=True, size=13, color='1F3864')
    r += 2
    _write_header_row(ws, r, ['순위', '지표명', '방향', '임계치', 'F1', '정밀도', '재현율', 'Lift'])
    for ri, row in best_res.head(10).iterrows():
        rr = r + 1 + ri
        vals = [ri + 1, row['feature'], row['direction'],
                round(row['threshold'], 6), round(row['f1'], 4),
                f"{row['precision']:.2%}", f"{row['recall']:.2%}", f"{row['lift']:.2f}×"]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(rr, ci); c.value = v; c.border = THIN
            c.alignment = Alignment(horizontal='center' if ci in (1, 3) else 'right')
            c.font = Font(size=10, bold=(ci == 5))
            if ri % 2 == 1: c.fill = ALT_FILL
        ws.cell(rr, 5).fill = _f1_fill(row['f1'])

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 36
    for ci in range(3, 9):
        ws.column_dimensions[get_column_letter(ci)].width = 12


def write_grid_sheet(wb, grid_df, best_nm):
    ws = wb.create_sheet('(n,m)_그리드')
    ws.sheet_view.showGridLines = False

    ws.cell(1, 1).value = '(n, m) 그리드 탐색 — Top20 평균 F1 히트맵'
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.cell(2, 1).value = f"★ 최적: n = {best_nm[0]}일,  m = {best_nm[1]}%  (빨간 테두리)"
    ws.cell(2, 1).font = Font(bold=True, size=11, color='C00000')

    pivot = grid_df.pivot(index='n', columns='m_pct', values='top20_mean_f1')
    sr = 4
    hdr = ws.cell(sr, 1)
    hdr.value = 'n ＼ m(%)'; hdr.fill = HEADER_FILL; hdr.font = WHITE_BOLD
    hdr.alignment = Alignment(horizontal='center'); hdr.border = THIN
    for ci, mv in enumerate(pivot.columns, 2):
        c = ws.cell(sr, ci); c.value = f'{mv}%'
        c.fill = HEADER_FILL; c.font = WHITE_BOLD
        c.alignment = Alignment(horizontal='center'); c.border = THIN
    for ri, nv in enumerate(pivot.index, sr + 1):
        c = ws.cell(ri, 1); c.value = f'{nv}일'
        c.fill = SUBHDR_FILL; c.font = WHITE_BOLD
        c.alignment = Alignment(horizontal='center'); c.border = THIN
        for ci, mv in enumerate(pivot.columns, 2):
            v = pivot.loc[nv, mv]
            cell = ws.cell(ri, ci)
            if pd.notna(v):
                cell.value = round(float(v), 4); cell.number_format = '0.0000'
                cell.fill = _f1_fill(v)
                if (nv, mv) == best_nm:
                    cell.font = Font(bold=True, size=12, color='C00000')
                    cell.border = THICK_RED
                else:
                    cell.border = THIN
            cell.alignment = Alignment(horizontal='center')

    dr = sr + len(pivot) + 3
    ws.cell(dr, 1).value = '조합별 상세 (종합점수 = F1 × log(1+Lift-1) 내림차순)'
    ws.cell(dr, 1).font = Font(bold=True, size=12, color='1F3864')
    dr += 1
    hdrs = ['n', 'm(%)', '양성률', '양성수', '샘플수', '유효지표',
            'Top20평균F1', 'Top20평균Prec', 'Top20평균Lift', '종합점수', '최대F1', '최고지표']
    _write_header_row(ws, dr, hdrs)
    gs = grid_df.sort_values('composite_score', ascending=False).reset_index(drop=True)
    for ri, row in gs.iterrows():
        r = dr + 1 + ri
        vals = [int(row['n']), row['m_pct'],
                round(row['positive_rate'], 4), int(row['n_positives']),
                int(row['n_samples']), int(row['n_valid_features']),
                round(row['top20_mean_f1'], 4), round(row['top20_mean_precision'], 4),
                round(row['top20_mean_lift'], 3),
                round(row['composite_score'], 4),
                round(row['max_f1'], 4),
                row['best_feature']]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = THIN
            c.alignment = Alignment(horizontal='center')
            if ri % 2 == 1: c.fill = ALT_FILL
        if (int(row['n']), row['m_pct']) == best_nm:
            for ci in range(1, 13):
                ws.cell(r, ci).fill = HIGHLIGHT
                ws.cell(r, ci).font = Font(bold=True)

    widths = [10, 10, 10, 10, 10, 10, 12, 12, 12, 12, 10, 30]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def write_threshold_sheet(wb, result_df, best_nm):
    ws = wb.create_sheet('최적조합_전체지표')
    ws.sheet_view.showGridLines = False
    n, m = best_nm

    ws.cell(1, 1).value = f'최적 (n = {n}일, m = {m}%) — 전체 지표별 최적 임계치 (F1 내림차순)'
    ws.cell(1, 1).font = Font(bold=True, size=13, color='1F3864')
    ws.merge_cells('A1:L1')

    headers = ['순위', '지표명', '방향', '임계치', 'F1', '정밀도', '재현율',
               'Lift', '기저율', '예측률', 'TP', 'FP']
    _write_header_row(ws, 3, headers)
    for ri, row in result_df.iterrows():
        r = ri + 4
        vals = [ri + 1, row['feature'], row['direction'],
                round(row['threshold'], 6), round(row['f1'], 4),
                round(row['precision'], 4), round(row['recall'], 4),
                round(row['lift'], 3), round(row['base_rate'], 4),
                round(row['pred_rate'], 4), int(row['tp']), int(row['fp'])]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = THIN
            c.alignment = Alignment(horizontal='center' if ci in (1, 3) else 'right')
            c.font = Font(size=9)
            if ri % 2 == 1: c.fill = ALT_FILL
        ws.cell(r, 4).number_format = '0.000000'
        ws.cell(r, 5).fill = _f1_fill(row['f1'])
        ws.cell(r, 5).font = Font(size=10, bold=True)

    widths = [6, 32, 8, 14, 10, 10, 10, 10, 10, 10, 8, 8]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'


def write_sweep_range_sheet(wb, result_df, best_nm):
    """
    각 지표별로 어디서 어디까지 어떤 단위(step)로 임계치를 탐색했는지 기록.
    사용자 요청: '지표별로 어디서부터 어디까지 얼마만큼 조정하면서 임계치 찾았는지'
    """
    ws = wb.create_sheet('임계치_탐색범위')
    ws.sheet_view.showGridLines = False
    n, m = best_nm

    ws.cell(1, 1).value = (f'지표별 임계치 탐색 범위 — n={n}일, m={m}% 기준  '
                            f'(임계치 = 분위 1%~99% 범위를 등간격 N단계로 스캔)')
    ws.cell(1, 1).font = Font(bold=True, size=13, color='1F3864')
    ws.merge_cells('A1:N1')
    ws.cell(2, 1).value = ('스윕범위 = 1~99 분위 (이상치 배제) | 단계 = (max−min)/(N−1) | '
                            '선택위치 = 0~N−1 중 최적 인덱스')
    ws.cell(2, 1).font = Font(italic=True, size=10, color='606060')

    headers = ['순위', '지표명', '방향', '★최적임계치', '선택위치',
               '스윕_min', '스윕_max', '스윕_단계크기', '단계수',
               '원본_min', '원본_max', 'F1', '정밀도', 'Lift']
    _write_header_row(ws, 4, headers)

    for ri, row in result_df.iterrows():
        r = ri + 5
        # 선택 위치 시각화: 단계별로 ★ 표시
        thr_idx = int(row.get('threshold_idx', 0))
        sweep_n = int(row.get('sweep_n', N_THRESHOLDS))
        # 텍스트 게이지: 0~9 (선택 위치를 10개 칸으로 압축 표시)
        bar_len = 10
        bar_pos = int(round(thr_idx / max(sweep_n - 1, 1) * (bar_len - 1)))
        gauge = '─' * bar_len
        gauge = gauge[:bar_pos] + '★' + gauge[bar_pos+1:]
        sel_str = f"{thr_idx + 1}/{sweep_n}  [{gauge}]"

        vals = [
            ri + 1, row['feature'], row['direction'],
            round(row['threshold'], 6), sel_str,
            round(row['sweep_min'], 6), round(row['sweep_max'], 6),
            round(row['sweep_step'], 6), sweep_n,
            round(row['raw_min'], 6), round(row['raw_max'], 6),
            round(row['f1'], 4), round(row['precision'], 4), round(row['lift'], 3),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = THIN
            c.font = Font(size=9, name='Consolas' if ci == 5 else 'Arial')
            if ci in (1, 3):
                c.alignment = Alignment(horizontal='center')
            elif ci == 5:
                c.alignment = Alignment(horizontal='left')
            else:
                c.alignment = Alignment(horizontal='right')
            if ri % 2 == 1: c.fill = ALT_FILL
        # 숫자 포맷
        for ci in (4, 6, 7, 8, 10, 11):
            ws.cell(r, ci).number_format = '0.000000'
        ws.cell(r, 4).font = Font(size=9, bold=True, color='C00000')
        ws.cell(r, 12).fill = _f1_fill(row['f1'])

        # 선택위치가 끝단에 가까우면 경고: 임계치를 더 넓게 잡아야 할 수도
        if thr_idx <= 1 or thr_idx >= sweep_n - 2:
            ws.cell(r, 5).fill = PatternFill("solid", fgColor="FFE699")  # 노란색 경고
            ws.cell(r, 5).font = Font(size=9, name='Consolas',
                                       color='9C5700', bold=True)

    widths = [6, 32, 6, 14, 22, 12, 12, 14, 8, 12, 12, 8, 10, 8]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A5'


def write_top20_sheet(wb, result_df, best_nm):
    ws = wb.create_sheet('상위20_상세')
    ws.sheet_view.showGridLines = False
    n, m = best_nm

    ws.cell(1, 1).value = f'상위 20개 예측 지표 상세 — n={n}일, m={m}%'
    ws.cell(1, 1).font = Font(bold=True, size=13, color='1F3864')
    ws.merge_cells('A1:N1')

    headers = ['순위', '지표명', '방향', '임계치', 'F1', '정밀도', '재현율',
               'Lift', '기저하락률', '예측양성률', 'TP', 'FP', 'FN', 'TN']
    _write_header_row(ws, 3, headers)
    for ri, row in result_df.head(20).reset_index(drop=True).iterrows():
        r = ri + 4
        vals = [ri + 1, row['feature'], row['direction'],
                round(row['threshold'], 6), round(row['f1'], 4),
                round(row['precision'], 4), round(row['recall'], 4),
                round(row['lift'], 3), round(row['base_rate'], 4),
                round(row['pred_rate'], 4), int(row['tp']), int(row['fp']),
                int(row['fn']), int(row['tn'])]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = THIN
            c.alignment = Alignment(horizontal='center' if ci in (1, 3) else 'right')
            if ri % 2 == 1: c.fill = ALT_FILL
        ws.cell(r, 4).number_format = '0.000000'
        ws.cell(r, 5).fill = _f1_fill(row['f1'])

    widths = [6, 34, 8, 14, 10, 10, 10, 10, 12, 12, 8, 8, 8, 8]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'


def write_validation_sheet(wb, merged, best_nm):
    if merged is None or len(merged) == 0: return
    ws = wb.create_sheet('Train_Test_검증')
    ws.sheet_view.showGridLines = False
    n, m = best_nm

    ws.cell(1, 1).value = f'시간순 Train/Test 분할 검증 — n={n}일, m={m}% (Train 70% → Test 30%)'
    ws.cell(1, 1).font = Font(bold=True, size=13, color='1F3864')
    ws.merge_cells('A1:J1')
    ws.cell(2, 1).value = 'Train에서 학습한 임계치를 미관측 Test 구간에 그대로 적용한 결과'
    ws.cell(2, 1).font = Font(italic=True, size=10, color='606060')

    _write_header_row(ws, 4, ['순위', '지표명', '방향', '임계치',
                              'Train_F1', 'Train_Prec', 'Test_F1', 'Test_Prec',
                              'Test_Recall', '강건성'])

    disp = merged.sort_values('f1', ascending=False).head(50).reset_index(drop=True)
    for ri, row in disp.iterrows():
        r = ri + 5
        tf1 = row.get('test_f1'); tpr = row.get('test_prec'); trc = row.get('test_recall')
        tf1 = 0.0 if pd.isna(tf1) else float(tf1)
        tpr = 0.0 if pd.isna(tpr) else float(tpr)
        trc = 0.0 if pd.isna(trc) else float(trc)

        if row['f1'] > 0.001:
            ratio = tf1 / row['f1']
            if   ratio >= 0.70: robust, rfill = '강건 ✓',  GOOD_FILL
            elif ratio >= 0.40: robust, rfill = '중간',    MID_FILL
            else:               robust, rfill = '과적합?', BAD_FILL
        else:
            robust, rfill = '-', None

        vals = [ri + 1, row['feature'], row['direction'],
                round(row['threshold'], 6),
                round(row['f1'], 4), round(row['precision'], 4),
                round(tf1, 4), round(tpr, 4), round(trc, 4), robust]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = THIN
            c.alignment = Alignment(horizontal='center' if ci in (1, 3, 10) else 'right')
            if ri % 2 == 1 and ci != 10: c.fill = ALT_FILL
        if rfill: ws.cell(r, 10).fill = rfill
        ws.cell(r, 4).number_format = '0.000000'

    widths = [6, 34, 8, 14, 10, 10, 10, 10, 10, 12]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A5'


def write_models_comparison_sheet(wb, comp_df, best_nm, top_k):
    """6개 예측 모델 성능 비교 시트."""
    if comp_df is None or len(comp_df) == 0: return
    ws = wb.create_sheet('모델_비교')
    ws.sheet_view.showGridLines = False
    n, m = best_nm

    ws.cell(1, 1).value = (f'예측 방식 비교 — n={n}일, m={m}%, Top-{top_k} 지표 사용  '
                           f'(F1 내림차순 정렬, ★는 baseline 대비 개선됨)')
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells('A1:K1')
    ws.cell(2, 1).value = '6가지 예측 방식을 동일 데이터에 적용해 정확도 비교'
    ws.cell(2, 1).font = Font(italic=True, size=10, color='606060')

    headers = ['순위', '예측 방식', 'F1', '정밀도', '재현율', 'Lift',
               'TP', 'FP', 'FN', 'TN', '예측률']
    _write_header_row(ws, 4, headers)

    # baseline F1을 찾아서 비교
    baseline_f1 = 0
    for _, r in comp_df.iterrows():
        if '기존' in str(r['name']):
            baseline_f1 = r['f1']
            break

    sorted_df = comp_df.sort_values('f1', ascending=False).reset_index(drop=True)
    for ri, row in sorted_df.iterrows():
        r = ri + 5
        is_baseline = '기존' in str(row['name'])
        is_better = row['f1'] > baseline_f1 + 0.005   # 0.5% 이상 개선
        marker = '★' if is_better else ('●' if is_baseline else '')

        vals = [
            f"{ri + 1}{marker}",
            row['name'],
            round(row['f1'], 4),
            round(row['precision'], 4),
            round(row['recall'], 4),
            round(row['lift'], 3),
            int(row['tp']), int(row['fp']),
            int(row['fn']), int(row['tn']),
            round(row['pred_rate'], 4),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = THIN
            c.alignment = Alignment(horizontal='center')
            if is_baseline:
                c.fill = PatternFill("solid", fgColor="DDEBF7")  # 옅은 파란색
                c.font = Font(bold=True, size=10)
            elif is_better:
                c.fill = HIGHLIGHT
                c.font = Font(bold=True, color='C00000', size=10)
            else:
                if ri % 2 == 1: c.fill = ALT_FILL
            if ci == 3:  # F1 색상강조
                c.fill = _f1_fill(row['f1'])
                if is_better:
                    c.font = Font(bold=True, color='C00000', size=11)

    widths = [8, 26, 10, 10, 10, 10, 8, 8, 8, 8, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A5'

    # 설명 추가
    desc_row = len(sorted_df) + 7
    descriptions = [
        ("● 기존(다수결): Top-K 중 50%+ 신호 시 하락 예측", "DDEBF7"),
        ("가중다수결: F1 점수로 가중합 — 신뢰도 높은 지표 영향력 ↑", "F2F2F2"),
        ("동시신호(N/K): 80% 이상 합의 필수 — 보수적이지만 정밀", "F2F2F2"),
        ("로지스틱회귀: Train구간으로 회귀 학습 후 적용", "F2F2F2"),
        ("누적신호(3일): 최근 3일 평균 신호강도 — 노이즈 감소", "F2F2F2"),
        ("고강도신호(>=70%): 강한 합의일 때만 행동 — 정밀도 중시", "F2F2F2"),
    ]
    for i, (txt, color) in enumerate(descriptions):
        ws.cell(desc_row + i, 1).value = txt
        ws.cell(desc_row + i, 1).font = Font(size=9, italic=True, color='606060')
        ws.cell(desc_row + i, 1).fill = PatternFill("solid", fgColor=color)
        ws.merge_cells(start_row=desc_row+i, start_column=1,
                        end_row=desc_row+i, end_column=11)


def write_walkforward_sheet(wb, wf_df, best_nm):
    """Walk-forward CV 결과 시트."""
    if wf_df is None or len(wf_df) == 0: return
    ws = wb.create_sheet('WalkForward_CV')
    ws.sheet_view.showGridLines = False
    n, m = best_nm

    ws.cell(1, 1).value = (f'Walk-forward 시간순 교차검증 — n={n}일, m={m}%  '
                           f'(매 fold마다 Train구간 끝까지로 학습 → 다음 구간으로 검증)')
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells('A1:I1')

    desc = (f"단일 70:30 분할은 운에 따라 결과가 흔들림. {len(wf_df)}-fold로 시간 흐름에 따른 일관성을 검증.\n"
            f"평균과 표준편차가 낮을수록 강건한 신호.")
    ws.cell(2, 1).value = desc
    ws.cell(2, 1).font = Font(italic=True, size=10, color='606060')
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('A2:I2')
    ws.row_dimensions[2].height = 32

    headers = ['Fold', 'Train 시작', 'Train 끝', 'Test 시작', 'Test 끝',
               'Train 일수', 'Test 일수', 'Top-K 평균 F1', 'Top-K 중간값 F1']
    _write_header_row(ws, 4, headers)
    for ri, row in wf_df.iterrows():
        r = ri + 5
        vals = [int(row['fold']), row['tr_start'], row['tr_end'],
                row['te_start'], row['te_end'],
                int(row['tr_size']), int(row['te_size']),
                round(row['top_k_mean_f1'], 4), round(row['top_k_median_f1'], 4)]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = THIN
            c.alignment = Alignment(horizontal='center')
            if ri % 2 == 1: c.fill = ALT_FILL
        ws.cell(r, 8).fill = _f1_fill(row['top_k_mean_f1'])

    # 통계 요약
    stat_row = len(wf_df) + 6
    f1_mean = wf_df['top_k_mean_f1'].mean()
    f1_std  = wf_df['top_k_mean_f1'].std()
    cv_ratio = f1_std / max(f1_mean, 0.01)
    ws.cell(stat_row, 1).value = '통계 요약'
    ws.cell(stat_row, 1).font = Font(bold=True, size=11, color='1F3864')
    ws.cell(stat_row + 1, 1).value = '평균 F1'
    ws.cell(stat_row + 1, 2).value = round(f1_mean, 4)
    ws.cell(stat_row + 2, 1).value = '표준편차'
    ws.cell(stat_row + 2, 2).value = round(f1_std, 4)
    ws.cell(stat_row + 3, 1).value = '변동계수 (낮을수록 안정)'
    ws.cell(stat_row + 3, 2).value = round(cv_ratio, 3)
    if cv_ratio < 0.2:
        verdict = '✓ 매우 안정적인 신호'; color = '006100'
    elif cv_ratio < 0.4:
        verdict = '○ 보통 수준의 안정성'; color = '7F6000'
    else:
        verdict = '✗ 시간에 따라 크게 변동 (불안정)'; color = 'C00000'
    ws.cell(stat_row + 4, 1).value = '평가'
    ws.cell(stat_row + 4, 2).value = verdict
    ws.cell(stat_row + 4, 2).font = Font(bold=True, color=color)

    widths = [8, 12, 12, 12, 12, 12, 12, 16, 16]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def write_topk_sheet(wb, k_sweep_df, best_k, best_nm):
    """Top-K 앙상블 사이즈 스윕 결과 시트."""
    if k_sweep_df is None or len(k_sweep_df) == 0:
        return
    ws = wb.create_sheet('TopK_스윕')
    ws.sheet_view.showGridLines = False
    n, m = best_nm

    ws.cell(1, 1).value = f'Top-K 앙상블 사이즈 최적화 — n={n}일, m={m}% (전체 평가기간 기준)'
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells('A1:J1')
    ws.cell(2, 1).value = f"★ 최적 Top-K = {best_k}  (F1 기준 자동 선정)"
    ws.cell(2, 1).font = Font(bold=True, size=11, color='C00000')

    _write_header_row(ws, 4, ['Top-K', 'F1', '정밀도', '재현율', 'Lift',
                               'TP', 'FP', 'FN', 'TN', '관측일'])
    for ri, row in k_sweep_df.iterrows():
        r = ri + 5
        is_best = int(row['top_k']) == best_k
        vals = [int(row['top_k']),
                round(row['f1'], 4),
                round(row['precision'], 4),
                round(row['recall'], 4),
                round(row['lift'], 3),
                int(row['tp']), int(row['fp']),
                int(row['fn']), int(row['tn']),
                int(row['n_obs'])]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = THIN
            c.alignment = Alignment(horizontal='center')
            if is_best:
                c.fill = HIGHLIGHT
                c.font = Font(bold=True, color='C00000', size=11)
                if ci == 1:
                    c.value = f"★ {int(row['top_k'])}"
            elif ri % 2 == 1:
                c.fill = ALT_FILL
        if not is_best:
            ws.cell(r, 2).fill = _f1_fill(row['f1'])

    widths = [10, 10, 10, 10, 10, 8, 8, 8, 8, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A5'


def write_fn_analysis_sheet(wb, daily, best_nm, top_k):
    """
    FN(놓친 하락) 분석 시트 — 각 큰 하락을 놓친 이유 진단.

    표시: 큰 하락(미래수익률 -2% 이하) 중 우리가 못 잡은 날들과
          그날의 신호 분포, 다른 시장 지표 상태
    """
    if daily is None or len(daily) == 0:
        return
    n, m = best_nm

    # FN 케이스 필터링
    fn_cases = daily[
        (daily['actual_drop'] == 1) &
        (daily['ensemble_pred'] == 0)
    ].copy()
    if len(fn_cases) == 0: return

    # 미래 수익률 기준 정렬 (최악 순)
    fn_cases = fn_cases.sort_values('future_ret_pct').reset_index(drop=True)

    ws = wb.create_sheet('FN_놓친하락_분석')
    ws.sheet_view.showGridLines = False

    ws.cell(1, 1).value = (f'놓친 하락(FN) 분석 — n={n}일, m={m}%, Top-{top_k} 앙상블')
    ws.cell(1, 1).font = Font(bold=True, size=14, color='C00000')
    ws.merge_cells('A1:H1')

    desc = (f"실제로 {m}%+ 하락했지만 시스템이 예측 못한 날들 (총 {len(fn_cases)}건). "
            f"미래수익률 나쁜 순.\n"
            f"신호수가 0~2/{top_k}로 너무 낮은 날 = '외부 이벤트 충격으로 사전 신호 없음' 가능성 ↑")
    ws.cell(2, 1).value = desc
    ws.cell(2, 1).font = Font(italic=True, size=10, color='606060')
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('A2:H2')
    ws.row_dimensions[2].height = 32

    headers = ['순위', '날짜', '종가', f'미래{n}일수익률%',
               f'신호수/{top_k}', '신호강도', '진단', '주요 사건 (참고)']
    _write_header_row(ws, 4, headers)

    # 알려진 주요 이벤트 (외부 검증된 날짜)
    known_events = {
        '2025-01-23': '🔴 DeepSeek 충격 직전 (1/27 NVDA -17%)',
        '2025-01-24': '🔴 DeepSeek 발표 직전',
        '2025-02-19': '⚠ 관세 위협 시작',
        '2025-02-20': '⚠ 관세 위협 본격화',
        '2025-02-21': '⚠ 관세 + Walmart 가이던스',
        '2025-02-25': '🔴 NVDA 어닝 발표 (2/26)',
        '2025-02-26': '🔴 NVDA 어닝 -8%',
        '2025-02-27': '⚠ NVDA 후폭풍',
        '2025-02-28': '⚠ NVDA 후폭풍',
        '2025-03-04': '⚠ 관세 단계적 발효',
        '2025-03-05': '⚠ 관세 단계적 발효',
        '2025-03-06': '⚠ 관세 단계적 발효',
        '2025-03-07': '🔴 관세 시행 + AI 회의론',
        '2025-03-25': '⚠ 관세 위협 재가열',
        '2025-04-01': '🔴 Liberation Day 직전',
        '2025-04-02': '🔴🔴 Liberation Day! 트럼프 사상최대 관세',
        '2025-04-03': '🔴🔴 S&P -4.8%, NDX -6%',
        '2025-04-04': '🔴 China 보복 관세 발표',
        '2025-04-16': '⚠ 관세 추가 위협',
        '2025-08-19': '⚠ Jackson Hole 직전',
        '2025-10-08': '🔴 트럼프 100% China 관세 위협',
        '2025-10-09': '🔴 Mag7 -$770B',
        '2025-11-04': '⚠ NVDA 어닝 시즌 우려',
        '2025-11-12': '⚠ AI 거품 우려',
    }

    for ri, row in fn_cases.iterrows():
        r = ri + 5
        date_key = str(row['date'].date()) if hasattr(row['date'], 'date') else str(row['date'])[:10]
        sig_count = int(row['signal_count']) if pd.notna(row['signal_count']) else 0
        sig_strength = float(row['signal_strength']) if pd.notna(row['signal_strength']) else 0
        fret = float(row['future_ret_pct']) if pd.notna(row['future_ret_pct']) else 0

        # 진단
        if sig_count == 0:
            diag = '신호 0개 — 완전 깜깜'
            diag_color = 'C00000'
        elif sig_count <= 2:
            diag = '신호 미약 — 이벤트 충격 가능성'
            diag_color = 'C00000'
        elif sig_strength < 0.5:
            diag = '강도 부족 — 임계 50% 미달'
            diag_color = '9C5700'
        else:
            diag = '강도 충분했으나 분류 실패'
            diag_color = '0070C0'

        # 알려진 이벤트
        event = known_events.get(date_key, '')

        vals = [
            ri + 1,
            date_key,
            round(float(row['close']), 2) if pd.notna(row['close']) else '',
            round(fret, 2),
            f"{sig_count}/{top_k}",
            f"{sig_strength:.1%}" if pd.notna(row['signal_strength']) else '-',
            diag,
            event,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = THIN
            c.alignment = Alignment(horizontal='center' if ci in (1, 5, 6) else 'left')
            if ri % 2 == 1: c.fill = ALT_FILL

        # 미래 수익률 색상 (큰 하락일수록 진한 빨강)
        if fret <= -5:
            ws.cell(r, 4).fill = PatternFill("solid", fgColor="C00000")
            ws.cell(r, 4).font = Font(bold=True, color='FFFFFF', size=10)
        elif fret <= -3:
            ws.cell(r, 4).fill = BAD_FILL
            ws.cell(r, 4).font = Font(bold=True, color='C00000')
        else:
            ws.cell(r, 4).fill = PatternFill("solid", fgColor="FFE0E0")

        # 진단 색상
        ws.cell(r, 7).font = Font(bold=True, color=diag_color, size=10)
        if event.startswith('🔴'):
            ws.cell(r, 8).fill = PatternFill("solid", fgColor="FFE0E0")
            ws.cell(r, 8).font = Font(bold=True, size=9)

    # 통계 요약
    stat_row = len(fn_cases) + 6
    ws.cell(stat_row, 1).value = '─── 진단 통계 ───'
    ws.cell(stat_row, 1).font = Font(bold=True, size=12, color='1F3864')

    sig_zero = (fn_cases['signal_count'] == 0).sum() if 'signal_count' in fn_cases.columns else 0
    sig_low  = (fn_cases['signal_count'] <= 2).sum() if 'signal_count' in fn_cases.columns else 0
    big_drops = (fn_cases['future_ret_pct'] <= -3).sum()

    diag_stats = [
        (f"총 FN(놓친 하락) 건수", f"{len(fn_cases)}건"),
        (f"  └ -3%↓ 큰 하락 놓침", f"{big_drops}건"),
        (f"  └ 신호 0개로 완전 깜깜", f"{sig_zero}건 ({sig_zero/max(len(fn_cases),1)*100:.0f}%)"),
        (f"  └ 신호 ≤ 2개로 미약", f"{sig_low}건 ({sig_low/max(len(fn_cases),1)*100:.0f}%)"),
        ('', ''),
        ('💡 결론', f"{sig_low/max(len(fn_cases),1)*100:.0f}% 의 놓친 하락이 '신호 미약' 상태"
         f" → 외부 이벤트 충격(관세, AI 뉴스 등)으로 사전 시계열 패턴 없이 발생"),
        ('', ''),
        ('📌 권장사항', '뉴스/이벤트 모니터링이 시계열 분석으로는 본질적 한계'),
        ('   ', '특히 트럼프 트윗, NVDA 어닝, FOMC 등 알려진 위험일 회피 권장'),
    ]
    for i, (k, v) in enumerate(diag_stats, 1):
        rr = stat_row + i
        c1 = ws.cell(rr, 1); c2 = ws.cell(rr, 2)
        c1.value = k; c2.value = v
        if k.startswith('💡') or k.startswith('📌'):
            c1.font = Font(bold=True, size=11, color='C00000')
            c2.font = Font(bold=True, size=10, color='C00000')
        else:
            c1.font = Font(size=10, color='606060')
            c2.font = Font(size=10)
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=8)

    widths = [6, 12, 10, 14, 12, 12, 28, 40]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A5'


def write_daily_prediction_sheet(wb, daily, top_k_res, best_nm, summary,
                                  start_date=None,
                                  vote_thresh=None):
    """일별 예측 시트 — 날짜 × (요약 8열 + Top-K 개별 지표 신호)."""
    if start_date  is None: start_date  = DAILY_START
    if vote_thresh is None: vote_thresh = ENSEMBLE_VOTE_THRESH
    if daily is None or len(daily) == 0:
        return
    ws = wb.create_sheet('일별예측_2025~현재')
    ws.sheet_view.showGridLines = False
    n, m = best_nm
    top_k = len(top_k_res)
    sig_cols = [c for c in daily.columns if c.startswith('_sig_')]

    # 타이틀
    ws.cell(1, 1).value = (f'일별 하락 예측 ({start_date}~현재) — '
                           f'n={n}일, m={m}%, Top-{top_k} 앙상블 '
                           f'(유효지표 중 ≥{vote_thresh*100:.0f}% 신호 → 하락 예측)')
    ws.cell(1, 1).font = Font(bold=True, size=13, color='1F3864')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9 + top_k)

    # 요약 라인
    if summary is not None:
        # 신호 발동 빈도 진단 (대부분이 0/N이면 "지표가 너무 까다롭다"는 신호)
        strengths = daily['signal_strength'].dropna()
        pct_strong = (strengths >= vote_thresh).sum() / max(len(strengths), 1)
        pct_any    = (strengths > 0).sum()           / max(len(strengths), 1)
        pct_zero   = (strengths == 0).sum()          / max(len(strengths), 1)
        ws.cell(2, 1).value = (
            f"관측 {summary['n_obs']}일  │  "
            f"TP={summary['tp']}  FP={summary['fp']}  FN={summary['fn']}  TN={summary['tn']}  │  "
            f"정밀도 {summary['precision']:.2%}  "
            f"재현율 {summary['recall']:.2%}  "
            f"F1 {summary['f1']:.4f}  "
            f"Lift {summary['lift']:.2f}×  "
            f"(기저 {summary['base_rate']:.2%})"
        )
        ws.cell(2, 1).font = Font(bold=True, size=10, color='C00000')
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9 + top_k)
        # 신호 빈도 라인 (너무 자주/너무 안 켜지면 경고)
        if pct_zero > 0.85:
            freq_warning = '⚠ 지표가 너무 까다로움 (신호 ≥1 발동: <15%) — Top-K 줄이거나 임계 완화 권장'
            freq_color = 'C00000'
        elif pct_strong > 0.40:
            freq_warning = '⚠ 신호 너무 자주 발동 (≥50% 강도: >40%) — Top-K 늘리거나 임계 강화 권장'
            freq_color = 'C00000'
        else:
            freq_warning = '○ 정상 빈도'
            freq_color = '006100'
        ws.cell(3, 1).value = (
            f"신호 발동 빈도: 0/N={pct_zero:.1%} | 1+ 발동={pct_any:.1%} | "
            f"≥{vote_thresh*100:.0f}% 강도={pct_strong:.1%}  →  {freq_warning}"
        )
        ws.cell(3, 1).font = Font(italic=True, size=9, color=freq_color)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=9 + top_k)

    # 헤더
    # 변경 후
    fixed = ['날짜', f'{TICKER}종가', f'미래{n}일수익률%', '실제하락',
         '신호수', '유효수', '신호강도(drop)', '상승신호수', '상승유효수',
         '상승신호강도(rise)', '순신호강도(net)', '앙상블예측', '결과']
    # 지표명 짧게
    sig_hdrs = [c[5:] for c in sig_cols]
    _write_header_row(ws, 4, fixed + sig_hdrs)

    # 데이터 행
    TN_FILL = PatternFill("solid", fgColor="F5F5F5")
    FP_FILL = PatternFill("solid", fgColor="FFD0D0")
    FN_FILL = PatternFill("solid", fgColor="FFE8A0")
    for ri, row in daily.iterrows():
        r = ri + 5
        # 1. 날짜
        c = ws.cell(r, 1); c.value = row['date'].date()
        c.number_format = 'YYYY-MM-DD'
        c.font = Font(bold=True, size=9, color='1F3864')
        c.alignment = Alignment(horizontal='center'); c.border = THIN
        # 2. 종가
        c = ws.cell(r, 2)
        if pd.notna(row['close']):
            c.value = round(float(row['close']), 2); c.number_format = '#,##0.00'
        c.alignment = Alignment(horizontal='right'); c.border = THIN
        # 3. 미래 수익률
        c = ws.cell(r, 3)
        if pd.notna(row['future_ret_pct']):
            v = float(row['future_ret_pct'])
            c.value = round(v, 2); c.number_format = '0.00"%"'
            if v <= -m: c.fill = BAD_FILL; c.font = Font(bold=True, color='C00000')
            elif v < 0: c.fill = PatternFill("solid", fgColor="FFF0F0")
        c.alignment = Alignment(horizontal='right'); c.border = THIN
        # 4. 실제 하락
        c = ws.cell(r, 4)
        if pd.notna(row['actual_drop']):
            if row['actual_drop'] == 1:
                c.value = '하락'; c.fill = BAD_FILL; c.font = Font(bold=True)
            else:
                c.value = '-'
        else:
            c.value = '미래'; c.font = Font(italic=True, color='888888')
        c.alignment = Alignment(horizontal='center'); c.border = THIN
        # 5. 신호수
        c = ws.cell(r, 5); c.value = int(row['signal_count'])
        c.alignment = Alignment(horizontal='center'); c.border = THIN
        # 6. 유효수
        c = ws.cell(r, 6); c.value = int(row['signal_valid'])
        c.font = Font(size=9, color='888888')
        c.alignment = Alignment(horizontal='center'); c.border = THIN
        # 7. 신호강도
        c = ws.cell(r, 7)
        if pd.notna(row['signal_strength']):
            s = float(row['signal_strength'])
            c.value = round(s, 3); c.number_format = '0%'
            if s >= vote_thresh + 0.2:   c.fill = BAD_FILL;  c.font = Font(bold=True, color='C00000')
            elif s >= vote_thresh:       c.fill = FP_FILL
            elif s >= 0.3:                c.fill = MID_FILL
        c.alignment = Alignment(horizontal='center'); c.border = THIN
        # 변경 후: 7번 블록 끝나고 8번 시작 전에 아래 블록 삽입

        # 7-1. 상승 신호수
        c = ws.cell(r, 8)
        if 'rise_signal_count' in row.index and pd.notna(row.get('rise_signal_count')):
            c.value = int(row['rise_signal_count'])
        else:
            c.value = '-'
        c.alignment = Alignment(horizontal='center'); c.border = THIN

        # 7-2. 상승 유효수
        c = ws.cell(r, 9)
        if 'rise_signal_valid' in row.index and pd.notna(row.get('rise_signal_valid')):
            c.value = int(row['rise_signal_valid'])
            c.font = Font(size=9, color='888888')
        else:
            c.value = '-'
        c.alignment = Alignment(horizontal='center'); c.border = THIN

        # 7-3. 상승 신호강도
        c = ws.cell(r, 10)
        if 'rise_strength' in row.index and pd.notna(row.get('rise_strength')):
            rs = float(row['rise_strength'])
            c.value = round(rs, 3); c.number_format = '0%'
            if rs >= 0.5:
                c.fill = GOOD_FILL; c.font = Font(bold=True, color='006100')
        else:
            c.value = '-'
        c.alignment = Alignment(horizontal='center'); c.border = THIN

        # 7-4. 순 신호강도(net)
        c = ws.cell(r, 11)
        if 'net_strength' in row.index and pd.notna(row.get('net_strength')):
            ns = float(row['net_strength'])
            c.value = round(ns, 3); c.number_format = '0%'
            if ns >= vote_thresh:
                c.fill = BAD_FILL; c.font = Font(bold=True, color='C00000')
            elif ns < 0:
                c.fill = GOOD_FILL
        else:
            c.value = '-'
        c.alignment = Alignment(horizontal='center'); c.border = THIN
        # 8. 앙상블 예측
        c = ws.cell(r, 12)
        if pd.notna(row['ensemble_pred']):
            if row['ensemble_pred'] == 1:
                c.value = '⚠ 하락예측'; c.fill = BAD_FILL
                c.font = Font(bold=True, color='C00000')
            else:
                c.value = '정상'; c.font = Font(size=9, color='888888')
        else:
            c.value = '?'
        c.alignment = Alignment(horizontal='center'); c.border = THIN
        # 9. 결과
        c = ws.cell(r, 13); c.value = row['result']
        rs = row['result']
        if   'TP' in rs: c.fill = GOOD_FILL; c.font = Font(bold=True, color='006100')
        elif 'FP' in rs: c.fill = FP_FILL;   c.font = Font(bold=True, color='C00000')
        elif 'FN' in rs: c.fill = FN_FILL;   c.font = Font(bold=True, color='9C5700')
        elif 'TN' in rs: c.fill = TN_FILL;   c.font = Font(size=9, color='888888')
        else:            c.font = Font(size=9, italic=True, color='888888')
        c.alignment = Alignment(horizontal='center'); c.border = THIN
        # 10+. 개별 지표 신호
        for si, sc in enumerate(sig_cols):
            c = ws.cell(r, 14 + si)
            v = row[sc]
            if pd.notna(v):
                vi = int(v)
                c.value = vi
                if vi == 1:
                    c.fill = BAD_FILL
                    c.font = Font(bold=True, size=9, color='C00000')
                else:
                    c.font = Font(size=9, color='BBBBBB')
            else:
                c.value = '-'; c.font = Font(size=9, color='BBBBBB')
            c.alignment = Alignment(horizontal='center'); c.border = THIN

    # 컬럼 폭
    widths_fixed = [12, 10, 12, 10, 8, 8, 10, 8, 8, 12, 12, 14, 10]
    for ci, w in enumerate(widths_fixed, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for si in range(len(sig_cols)):
        ws.column_dimensions[get_column_letter(10 + si)].width = 14

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[4].height = 38
    ws.freeze_panes = 'B5'

# ══════════════════════════════════════════════════════════════════
#  지표 검증 도구 (워크포워드 + 다중검정 보정 + 상관 중복제거)
#  ※ 원본 코드는 한 줄도 수정 안 함. 이 블록을 파일 맨 아래에 붙이기만.
#    원본이 만든 feat / ohlcv / closes / TICKER 를 자동으로 찾아 실행.
# ══════════════════════════════════════════════════════════════════
import numpy as np
import pandas as pd

def _iv_make_target(close, horizon=5):
    return close.shift(-horizon) / close - 1

def _iv_ic(x, y):
    d = pd.concat([x, y], axis=1).dropna()
    if len(d) < 30 or d.iloc[:, 0].nunique() < 5:
        return np.nan, np.nan, 0
    xr = d.iloc[:, 0].rank(); yr = d.iloc[:, 1].rank()
    ic = np.corrcoef(xr, yr)[0, 1]
    n = len(d)
    t = ic * np.sqrt((n - 2) / max(1 - ic**2, 1e-9)) if abs(ic) < 1 else np.nan
    return ic, t, n

def _iv_fdr_bh(pvals, alpha=0.10):
    p = pvals.dropna().sort_values(); m = len(p)
    if m == 0: return pd.Series(dtype=bool)
    thresh = np.arange(1, m + 1) / m * alpha
    passed = p.values <= thresh
    keep = p.index[:np.max(np.where(passed)[0]) + 1] if passed.any() else pd.Index([])
    out = pd.Series(False, index=pvals.index); out.loc[keep] = True
    return out

def _iv_walk_forward(feat, target, n_splits=4):
    df = feat.loc[target.dropna().index.intersection(feat.index)]
    tgt = target.reindex(df.index); n = len(df)
    if n < (n_splits + 1) * 40:
        n_splits = max(2, n // 40 - 1)
    fold = n // (n_splits + 1); rows = {}
    for col in df.columns:
        is_ics, oos_ics = [], []
        for k in range(1, n_splits + 1):
            x_is, y_is = df[col].iloc[:fold*k], tgt.iloc[:fold*k]
            x_oos, y_oos = df[col].iloc[fold*k:fold*(k+1)], tgt.iloc[fold*k:fold*(k+1)]
            ic_is, _, _ = _iv_ic(x_is, y_is)
            ic_oos, _, n_oos = _iv_ic(x_oos, y_oos)
            if pd.notna(ic_is) and pd.notna(ic_oos) and n_oos >= 20:
                is_ics.append(ic_is); oos_ics.append(ic_oos)
        if len(oos_ics) >= 2:
            rows[col] = {'is_ic': np.mean(is_ics), 'oos_ic': np.mean(oos_ics),
                         'sign_consistency': np.mean(np.sign(is_ics) == np.sign(oos_ics)),
                         'n_folds': len(oos_ics)}
    return pd.DataFrame(rows).T

def _iv_dedup(feat, ranking, thresh=0.7):
    cols = [c for c in ranking.sort_values(ascending=False).index if c in feat.columns]
    sub = feat[cols].dropna()
    if len(sub) < 20 or len(cols) < 2: return cols
    cm = sub.corr().abs(); selected, removed = [], set()
    for c in cols:
        if c in removed: continue
        selected.append(c)
        removed.update(cm.index[(cm[c] > thresh) & (cm.index != c)])
    return selected

def validate_indicators(feat, close, horizon=5, alpha_fdr=0.10,
                        min_oos_ic=0.02, min_sign_consistency=0.75, corr_thresh=0.7):
    from math import erfc
    target = _iv_make_target(close, horizon)
    valid = [c for c in feat.columns
             if feat[c].notna().sum() >= 60 and feat[c].dropna().nunique() >= 5]
    feat = feat[valid]
    rows = {}
    for col in feat.columns:
        ic, t, n = _iv_ic(feat[col], target.reindex(feat.index))
        if pd.notna(ic) and n > 0:
            p = erfc(abs(t) / np.sqrt(2)) if pd.notna(t) else np.nan
            rows[col] = {'ic': ic, 'p': p, 'n': n}
    ic_tbl = pd.DataFrame(rows).T
    if len(ic_tbl) == 0:
        return {'n_input': len(feat.columns), 'n_fdr_pass': 0, 'n_robust': 0,
                'n_final': 0, 'final_indicators': [], 'survivors_table': pd.DataFrame()}
    ic_tbl['fdr_pass'] = _iv_fdr_bh(ic_tbl['p'], alpha_fdr)
    wf = _iv_walk_forward(feat, target)
    merged = ic_tbl.join(wf, how='inner')
    merged['robust'] = (merged['fdr_pass'] & (merged['oos_ic'].abs() >= min_oos_ic) &
                        (merged['sign_consistency'] >= min_sign_consistency) &
                        (np.sign(merged['is_ic']) == np.sign(merged['oos_ic'])))
    surv = merged[merged['robust']].copy()
    if len(surv) > 0:
        kept = [c for c in _iv_dedup(feat, surv['oos_ic'].abs(), corr_thresh) if c in surv.index]
        surv['kept'] = surv.index.isin(kept)
    else:
        kept = []
    surv = surv.sort_values('oos_ic', key=lambda s: s.abs(), ascending=False)
    return {'n_input': len(feat.columns), 'n_fdr_pass': int(ic_tbl['fdr_pass'].sum()),
            'n_robust': int(merged['robust'].sum()), 'n_final': len(kept),
            'final_indicators': kept, 'survivors_table': surv}

def _iv_autorun():
    g = globals()
    _feat = g.get('feat')
    if _feat is None or not isinstance(_feat, pd.DataFrame):
        print("[검증] feat(지표 DataFrame)를 못 찾음 — 원본에서 지표를 담은 변수명을 확인하세요.")
        print("       예: feat = compute_features(...) 로 받았다면 그 변수명이 'feat'이어야 함.")
        return
    _close = None
    if 'ohlcv' in g and isinstance(g['ohlcv'], pd.DataFrame) and 'Close' in g['ohlcv']:
        _close = g['ohlcv']['Close']
    elif 'closes' in g and 'TICKER' in g and g['TICKER'] in g['closes']:
        _close = g['closes'][g['TICKER']]
    if _close is None:
        print("[검증] 종가 시계열을 못 찾음 — ohlcv['Close'] 또는 closes[TICKER] 를 확인하세요.")
        return
    print("=" * 60)
    print("지표 검증 (워크포워드 + 다중검정 보정 + 상관 중복제거)")
    print("=" * 60)
    for h in [1, 5, 10]:
        r = validate_indicators(_feat, _close, horizon=h)
        print(f"\n[{h}일 예측] 입력 {r['n_input']} → FDR통과 {r['n_fdr_pass']} "
              f"→ 강건 {r['n_robust']} → 최종 {r['n_final']}")
        if r['n_final'] > 0:
            print(f"  최종 채택: {r['final_indicators'][:15]}")
            st = r['survivors_table']
            show = [c for c in ['ic', 'oos_ic', 'sign_consistency', 'kept'] if c in st.columns]
            print(st[show].head(12).round(4).to_string())
            try:
                st.to_csv(f"survivors_h{h}.csv")
            except Exception:
                pass
    print("\n" + "=" * 60)
    print("주의: 통과한 지표도 미래 수익을 보장하지 않습니다. 소액·모의로 반드시 추가 검증하세요.")

# @title
"""
매수/매도 앙상블 — 메타 그리드 자동 튜닝 + MDD 제한 + 손절매 + ANCHOR 보정
====================================================================
★ 기능
   1. META_GRID — 점수화/풀 구성 변수를 자동 그리드 탐색
   2. MIN_TRADES_DAILY — 일별 거래수 하한
   3. MAX_DRAWDOWN_LIMIT_PCT — 일별 MDD 한도
   4. 점수화 결과 캐싱
   5. 신호 충돌 해결 — 강도(count/K) 큰 쪽 우선
   6. Balanced Accuracy 평가 + plain accuracy 참고
   7. ★ Tolerance Band 정렬 — 정확도 top-band 이내 중 수익률 최대 (SELECTION_TOLERANCE)
   8. 손절매 (STOP_LOSS_PCT)
   9. ★ ANCHOR 자동 계산 (AUTO_ANCHOR) — 백테스트 데이터로 최적 매수/매도 정답일 자동 산출
"""

import warnings; warnings.filterwarnings('ignore')
import os, math, time, itertools
from datetime import datetime, timedelta
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = '/content/ensemble_analysis'   # ★ 로컬(Colab 세션) 저장 — Drive 저장 안 함

def _resolve_output_dir(target):
    """OUTPUT_DIR 해석 + 폴더 생성. (Google Drive 저장 안 함 — 로컬에만 저장 후 다운로드)"""
    if target is None:
        return '/content' if os.path.isdir('/content') else os.getcwd()
    # ★ Drive 경로가 지정돼 있어도 Drive에 저장하지 않고 로컬로 대체
    if target.startswith('/content/drive/'):
        local_alt = '/content/ensemble_analysis' if os.path.isdir('/content') else os.path.join(os.getcwd(), 'ensemble_analysis')
        print(f"  ℹ Drive 저장 비활성화 — 로컬 폴더에 저장 후 다운로드: {local_alt}")
        target = local_alt
    # 폴더 없으면 생성
    try:
        os.makedirs(target, exist_ok=True)
    except Exception as e:
        print(f"  ⚠ {target} 생성 실패: {e} → /content 로 fallback")
        return '/content' if os.path.isdir('/content') else os.getcwd()
    return target

SCRIPT_DIR = _resolve_output_dir(OUTPUT_DIR)
OUTPUT_DIR = SCRIPT_DIR   # ★ 이후 모든 참조(재현 모드 등)가 로컬 폴더를 가리키도록 동기화
print(f"  📂 출력 폴더(로컬): {SCRIPT_DIR}  — 실행 종료 후 자동 다운로드됩니다")

try:
    from numba import njit
    HAS_NUMBA = True
    print("  ℹ numba 감지 — JIT 가속")
except ImportError:
    HAS_NUMBA = False
    def njit(*args, **kw):
        if args and callable(args[0]): return args[0]
        return lambda f: f
    print("  ⚠ numba 미설치.  pip install numba 권장")


# ════════════════════════════════════════════════════════════════
#                            설정
# ════════════════════════════════════════════════════════════════
EVAL_START          = '2025-01-01'

OOS_ENABLED         = False          # OOS 검증 on/off
OOS_START           = '2025-10-01'   # 이 날부터 현재까지 = OOS 검증구간

HORIZON_DAYS        = 5
DRAWDOWN_LIMIT_BUY  = 0.25
RUNUP_LIMIT_SELL    = 0.25

N_THRESHOLDS        = 800
MAX_INDICATORS      = 1000

# ★ 성공률 우선 풀 선출 (요청) — 점수가 아니라 '성공률'로 먼저 지표를 선발한 뒤 그리드.
#   목적: pct(분위)가 달라 따로 나오던 고성공 지표를 누락 없이 한 풀에 모으고,
#         선발 기준을 점수(Wilson)→성공률 우선으로 바꿈. 가짜 100% 방지 위해 표본 가드 둠.
POOL_SELECT_BY_SUCCESS = True      # True면 풀을 성공률 우선으로 선출(아래 기준), False면 기존 점수순.
POOL_SUCCESS_MIN_RATE  = 0.60      # 성공률 컷오프 (요청: 0.60)
POOL_SUCCESS_MIN_SIG   = 10        # 최소 신호수 — 소표본 가짜 100% 방지 (요청: 10)
POOL_SUCCESS_WIDE_PCT  = (1, 99)   # 풀 평가용 분위 (요청: 1,99)
POOL_SUCCESS_K_FLOOR   = 2         # ★ 성공률 우선 시 K 하한 — 정예(희소) 풀은 소수 동의로도 신호나야 거래 발생.
                                   #   (기존 K_BUY_RANGE는 10부터라 정예풀에선 거래 0 → 자동으로 이 값까지 낮춤)

# ★ 미래 예측 정확도 강화 (요청) — 지표 신호 생성·평가 방식 보강.
#   목적: 과거 적합이 아니라 '미래에도 유지되는 신호'를 우대해 매수/매도 시점 예측력↑.
# (1) z-스코어 신호: 절대 임계 대신 롤링 z=(x-평균)/표준편차에도 임계를 걸어 신호 생성.
#     시장 레짐이 바뀌어도 '평균 대비 몇 시그마'는 의미가 유지됨 → 미래 일반화에 강함.
USE_ZSCORE_SIGNAL   = True    # True면 각 지표를 절대임계 + z스코어임계 두 방식으로 평가
ZSCORE_WINDOW       = 60      # z스코어 롤링 창(거래일). 약 3개월.
ZSCORE_THRESHOLDS   = [-2.5, -2.0, -1.5, -1.0, -0.5, 0.5, 1.0, 1.5, 2.0, 2.5]  # z 임계 후보 (요청 확장)
# (2) OOS 안정성 가중: 지표 점수를 '전체기간 Wilson'에만 의존하지 말고,
#     기간을 앞/뒤로 나눠 둘 다 좋은 지표(시간적으로 안정)에 가산점 → 과최적화 억제.
USE_OOS_STABILITY   = False   # ★ 요청: OOS 관련 기능 전부 OFF
OOS_STABILITY_WEIGHT = 0.3    # 안정성 가중 강도 (0=기존과 동일, 1=안정성 절반 반영)

# ★ 큰 상승/하락 적중 가산점 (요청) — 지표 신호 중 '큰 움직임'을 맞춘 비율이 높을수록
#   점수에 가산. 매수지표: 신호 뒤 큰 상승을 맞춘 비율, 매도지표: 큰 하락을 맞춘 비율.
#   목적: 자잘한 적중만 많은 지표보다 '굵직한 변동을 잡는' 지표를 우대 → 실전 수익 직결.
USE_BIG_MOVE_BONUS   = True   # True면 큰움직임 적중비율을 점수에 가산
BIG_MOVE_THRESHOLD   = 0.03   # 신호 뒤 horizon 내 유리방향 최대변동이 이 값(3%) 이상이면 '큰 움직임'
BIG_MOVE_BONUS_WEIGHT = 0.5   # 가산 강도: score *= (1 + W * 큰움직임적중비율)

K_BUY_RANGE         = [i for i in range(10, 100)]
K_SELL_RANGE        = [i for i in range(10, 100)]
VOTE_RATIO_BUY      = [0.2, 0.25, 0.3, 0.35, 0.4, 0.45, 0.5,
                       0.55, 0.6, 0.65, 0.7, 0.75, 0.8, 0.85]
VOTE_RATIO_SELL     = [0.2, 0.25, 0.3, 0.35, 0.4, 0.45, 0.5,
                       0.55, 0.6, 0.65, 0.7, 0.75, 0.8, 0.85]

COST_PER_TRADE      = 0.004

MIN_TRADES_DAILY    = 10
MAX_DRAWDOWN_LIMIT_PCT = None

STOP_LOSS_PCT       = 0.05

# ★ 선정 우선순위
#   'winrate_return'     : 일별거래 승률 최고 -10%p 밴드 → 그중 누적수익 최고 (요청, 기본)
#   'sell_mdd_return'    : 매도성공률 최고 -2%p 밴드 → MDD 최저 -1%p 밴드 → 누적수익 최고
#   'avgband_mdd_return' : 평균성공률 top band 안에서 → MDD 가장 낮은(0.1%p 단위 동률이면 수익)
#   'independent'     : 매수성공률 최고 설정 + 매도성공률 최고 설정을 따로 찾아 합침.
#                       성공률은 매수/매도가 독립이므로 각각 최댓값을 동시 달성. 수익·MDD는 따라옴.
#                       (MDD 한도가 켜져 있으면 '한도 통과분 중' 각각 최고를 찾음)
#   'stability'       : 매도성공·매수성공·누적수익·MDD방어를 종합한 안정성 점수 최대
#   'sell_buy_return' : 매도성공률 → 매수성공률 → 누적수익 순 (밴드 줄세우기)
#   'balacc_return'   : 기존 (평균 BalAcc → 수익률)
SELECTION_PRIORITY = 'winrate_return'

# ★ 안정성 종합 점수 가중치 (SELECTION_PRIORITY='stability'일 때 사용)
#   (매도성공률, 매수성공률, 누적수익, MDD방어) — 합이 1이 되도록 자동 정규화됨.
#   가중 기하평균이라 한 요소라도 후보군 내 최저면 점수가 크게 깎임 → 골고루 좋은 조합 선호.
#   하락 회피 강화하려면 매도성공률(첫째)·MDD방어(넷째) 비중을 올리세요.
STABILITY_WEIGHTS = (0.30, 0.25, 0.20, 0.25)

# ★ 가중 투표 (변경2) — 지표 성공률(Wilson)에 비례해 표 가중
#   USE_WEIGHTED_VOTE=False면 기존 일반 투표(모두 1표)
#   가중 강도는 점수 분포 편차에 따라 자동 조절(데이터가 결정), 상한 WEIGHT_MAX_RATIO
USE_WEIGHTED_VOTE = True
WEIGHT_MAX_RATIO  = 1.6   # (구) 선형 가중 상한 — 아래 희석방지 장치가 켜지면 보조 역할

# ★ 점수 희석 방지 장치 (요청) — 자잘한 다수 지표가 합쳐져 강한 소수 지표를 덮지 않도록,
#   가중치를 '절대 점수 차이'에 지수적으로 반응시킨다(소프트맥스형). 강한 지표일수록 표가
#   급증해 합산에서 살아남는다. 단 한 지표가 전부를 지배하지 않게 상·하한으로 '적당히' 제한.
#     - VOTE_WEIGHT_TEMP: 온도(작을수록 강한지표에 표 더 몰림). 0.15≈적당.
#     - VOTE_STRONG_CAP : 최고지표가 가질 수 있는 표 상한(평균=1 기준 배수). 과도지배 방지.
#   끄려면 USE_ANTIDILUTION=False (그러면 기존 선형 WEIGHT_MAX_RATIO 가중).
USE_ANTIDILUTION  = True
VOTE_WEIGHT_TEMP  = 0.15
VOTE_STRONG_CAP   = 4.0

SELECTION_TOLERANCE = 0.04

# ★ 'sell_mdd_return' 모드 밴드 폭
SELL_SUCCESS_TOLERANCE = 0.02   # 매도성공률 최고에서 이 차이(2%p)까지 후보
MDD_TOLERANCE          = 0.01   # MDD 최저에서 이 차이(1%p)까지 후보 (그중 수익 최고 선택)

# ★ 'winrate_return' 모드 밴드 폭 (요청) — 일별거래 승률 최고에서 이 차이(10%p)까지 후보
WINRATE_TOLERANCE      = 0.04

# ★ 승률 후보 실거래 검증 (요청) — 그리드는 빠른 근사라 실제 일별거래와 MDD·승률·수익이
#   다를 수 있음. 그래서 승률 상위 후보만 골라 '실제 일별 백테스트'를 돌려 진짜 수치를
#   구하고, 그중 실제 누적수익이 가장 높은 조합을 최종 선정한다.
VERIFY_BY_DAILY_BACKTEST = True   # True: 후보들을 실제 일별백테스트로 재검증 후 선정
VERIFY_TOP_N             = 10000  # 그리드 승률 상위 몇 개를 실제로 돌릴지 (많을수록 정확·느림)

# ★ 실거래 검증 후, '실제 최대 거래손실'이 이 값 이하(더 안전)인 후보만 선정 대상으로 (요청).
#   예: -0.03 이면 실제 단일거래 최대손실이 -3%보다 깊지 않은 조합만 후보.
#   None 이면 이 필터를 끈다. (단, 필터로 후보가 0개면 자동으로 필터를 완화해 최선을 고름)
VERIFY_MAX_DRAWDOWN_LIMIT = -0.03

# ★ 최종 선정 2차 밴드 (요청) — 실제 승률 밴드 후보 중, 실제 '평균 성공률' 최고에서
#   이 값(3%p) 이내를 다시 후보로 두고, 그중 실제 수익률 최고를 선정.
VERIFY_AVG_SUCCESS_TOLERANCE = 0.4

# ★ 수익 동률 판정 (요청) — 실제 수익률 차이가 이 값(3%p) 이내면 '동률'로 보고,
#   그중 실제 평균 성공률이 더 높은 조합을 선정.
VERIFY_RETURN_TIE = 0.03

# ★ 최종 선정 2차 밴드 (요청) — 실제 승률 밴드 후보 중, 실제 '평균 매칭률'(매수·매도 앵커 매칭)
#   최고에서 이 값(2%p) 이내를 다시 후보로 두고, 그중 실제 수익률 최고를 선정.
VERIFY_MATCH_TOLERANCE = 0.4

# ★ Buy&Hold 미달 조합 제외 (전략 누적수익이 B&H 이하면 후보에서 버림)
EXCLUDE_BELOW_BH = False

ANCHOR_MATCH_PRIORITY = False   # ★ 선정은 평균성공+MDD+수익 기준으로. 매칭 우선이 그걸 덮지 않도록 OFF
ANCHOR_MATCH_TOLERANCE = 0.10

ANCHOR_MODE = True

AUTO_ANCHOR = True
AUTO_ANCHOR_WINDOW     = 1
AUTO_ANCHOR_LOOKFORWARD = 1
AUTO_ANCHOR_MIN_RISE   = 0.01
AUTO_ANCHOR_MIN_DROP   = 0.01
AUTO_ANCHOR_PRICE_TOLERANCE = 0.01
AUTO_ANCHOR_MAX_DATES  = None

# ★ 앵커 매칭 타이밍 윈도우 (요청) — 신호는 정답일보다 먼저 떠야 정답일에 포지션이 맞음.
#   매칭 판정 시 '정답일 당일 ~ +N일' 사이에 실제 포지션이 맞으면 매칭으로 인정.
#   신호→익일체결 지연 + 앵커 정답일이 며칠에 걸친 바닥/천장 구간인 점을 반영.
#   0이면 당일만, 1이면 당일+익일, 2면 +2일까지. (충돌·실제매매는 그대로 반영)
ANCHOR_MATCH_WINDOW = 2

# ★ 매도 매칭의 '고점 전 보유' 조회 범위 (요청) — 고점 직전 W일만 보면 너무 빡빡해
#   매도매칭이 50%대로 낮게 나옴. 고점까지 오는 과정 최근 N일 중 보유한 적 있으면
#   '고점까지 들고 왔다'로 인정 (그 뒤 고점 근처 청산이면 매칭).
ANCHOR_SELL_HOLD_LOOKBACK = 5

# ★ 충돌 처리 정책 (요청) — '매도 우선':
#   True: 매수·매도 신호 충돌 시, 현금이면 매수 보류(잘못된 매수 상쇄),
#         보유면 청산(올바른 매도는 안 상쇄). 위험회피적, 손실 방지 우선.
#   False: 기존 강도 비교(b_strength vs s_strength).
CONFLICT_SELL_PRIORITY = False

# ★ 최종 선정 우선순위 (요청) — 매도 정확도 우선 → 매수 정확도 → 수익:
#   1) 실제 매도 성공률(정확도) 최고에서 -SELL_ACC_TOL 범위
#   2) 그중 실제 매수 성공률 최고에서 -BUY_ACC_TOL 범위
#   3) 그중 실제 수익 최고
VERIFY_SELL_ACC_TOLERANCE = 0.01   # 매도 정확도 1차 밴드 (3%p)
VERIFY_BUY_ACC_TOLERANCE  = 0.01   # 매수 정확도 2차 밴드 (3%p)

# ★ CatBoost 액션 보정 (요청) — 선정된 조합의 신호로 백테스트한 뒤, 앵커 정답과 안 맞는
#   매수/매도 액션을 머신러닝으로 보정해본다.
#   - 피처: 매수/매도 신호 강도, 최근 수익/변동성 등 (그 시점까지 정보만)
#   - 타겟: 앵커 정답(올라야 할 날=보유, 내려야 할 날=현금)
#   - walk-forward(시계열 분할)로 학습→이후구간 적용. 미래참조를 피하려 과거로만 학습.
#   ⚠ 앵커는 사후적 정답이라, 보정 결과는 '참고용'. 실전 일반화는 별도 검증 필요.
USE_CATBOOST_CORRECTION = False  # ★ 요청: CatBoost 보정 제거 — 앵커 기반 가중치 보정만 사용
CATBOOST_MIN_TRAIN_DAYS = 120    # 최소 학습일수 (이만큼 쌓인 뒤부터 보정 적용)
CATBOOST_PROB_THRESHOLD = 0.5    # 보유 확률이 이 값 이상이면 '보유'로 보정
CATBOOST_TARGET_HORIZON = 5      # ★ 타겟: 미래 N일 수익 (앵커 아님 — 실전 검증 가능한 라벨)
CATBOOST_OOS_FRACTION   = 0.2    # ★ 마지막 N%는 학습에서 빼고 '진짜 미래'로 성능 측정

# ★ OOS(out-of-sample) 검증 기간 (요청) — 최근 이 개월은 학습/탐색에서 빼고,
#   후보 그리드를 이 기간에서만 백테스트해 'OOS 수익/정확도'를 측정.
#   OOS 수익이 가장 높은 조합을 최종 선정 → 미래 일반화에 가까운 선택.
OOS_MONTHS = 1                   # 최근 N개월을 OOS로 (이전달~현재)
OOS_SELECT_BY_OOS_RETURN = False # ★ 요청: OOS 관련 기능 전부 OFF

# ★ 앵커 미매칭 보정 (요청) — 최적 그리드 선정 후, CatBoost 보정 '전'에 실행.
#   충돌 패배·간발의 차로 앵커와 안 맞은 매수/매도를 '지표 가중치 조정'으로 보정.
#   단, 조정 후 수익이 떨어지면 롤백(기존 그리드 유지) → 미매칭만 보정, 수익 손해 없음.
USE_ANCHOR_MATCH_CORRECTION = True
# ★ 검증 후보 보정 통합 (요청) — 실거래 검증 후보 중 수익 상위 N개에 지표 가중치 보정을
#   적용하고, 보정 후 수익/손실/승률을 재계산. 보정 후 수익이 가장 높은 조합을 선정.
#   (모든 후보에 보정하면 너무 느려서 수익 상위 N개로 제한)
CORRECT_VERIFY_TOP_N = 20        # 수익 상위 몇 개 후보에 보정을 적용할지
CORRECT_PROFIT_FLOOR = 0.01      # 거래 진단: 수익 이 값(1%) 이하 + 손실 거래를 '틀린 거래'로
SELECT_BY_CORRECTED_RETURN = True  # True면 보정 후 수익 최고로 선정
# ★ 후보 선정 필터를 '보정 후' 수치로 (요청) — 최대거래손실 한도·승률 밴드를 보정 적용된
#   조합은 보정후 수치로 판정. 보정 안 된 조합은 실제 수치 그대로. SELECT_BY_CORRECTED_RETURN과 짝.
SELECT_FILTER_BY_CORRECTED = True
# ★ 보정후 승률 하한 (요청) — 보정후 실제 거래 승률이 이 값 이상인 조합 중 보정후 수익 최고를
#   1차 선정. 이후 그 승률에서 WINRATE_TOLERANCE 이내 후보와 수익 비교해 더 높으면 재선정.
SELECT_WIN_FLOOR = 0.90

ANCHOR_BUY_DATES = [
]
ANCHOR_SELL_DATES = [
]

SELECT_BY           = 'total_return'
TOP_N_GRID_OUT      = 10000

META_GRID = {
    # ★ staged 방식의 '시작값'. 단계 탐색은 STAGE_PCT_RANGE / STAGE_WILSON_Z /
    #   STAGE_CORR_LIMIT 후보들을 순서대로 돌리며 좁힌다 (모든 조합 X).
    'wilson_z':    [1.65],
    'pct_range':   [(5, 95)],
    'min_signals': [10],
    'corr_limit':  [0.2],
    'top_n_pool':  [100],
}

STAGED_META_TUNE = True   # ★ True: pct_range → wilson_z → corr_limit 순으로 단계적 결정 (요청).
                          #   단계에서 돌린 결과들을 한 엑셀에 모두 모아 최종 판단.
STAGE_PCT_RANGE   = [(5, 95)]
STAGE_WILSON_Z    = [1.65, 1.75, 1.85, 1.95]
STAGE_WILSON_REFINE_STEP = 0.05
STAGE_CORR_LIMIT  = [0.2]

PREFILTER_ENABLED          = True
PREFILTER_MIN_CORR         = 0.005
PREFILTER_MIN_VARIANCE_REL = 1e-6
PREFILTER_MAX_NAN_RATIO    = 0.5

AUTO_DOWNLOAD_EXCEL = True

# ★ 재현 방식 (요청) — 모드 2/3가 예전에 정확히 맞던 방식과 동일하게.
#   재현 때 '원본 엑셀의 마지막 날짜까지'로 데이터를 잘라 compute_features에 넘긴다.
#   → 전체구간 순위(rank)·정규화 지표가 원본과 '같은 범위'로 계산돼 신호가 정확히 일치.
#   (재현이 하루 더 받아 순위가 다시 매겨지던 문제 해결. 연장은 안 함 = 원본 날짜까지 정확 재현)
REPLAY_MATCH_ORIGINAL_RANGE = True

# ★ 재현 후 새 거래일 연장 (요청) — 기본 끔(OFF). 정확 재현이 우선.
#   켜면(True) 과거(원본 마지막날)까지는 정확 재현 + 그 이후 새 거래일을 이어 계산해 덧붙인다.
#   ⚠ 단, 전체구간 순위 지표가 있으면 '정확 재현'과 '연장'을 100% 동시 보장하기 어렵다.
#   새 데이터까지 보려면 '새 분석'을 한 번 더 돌리는 것이 가장 정확하다.
REPLAY_EXTEND_TO_TODAY = False

# 데이터 스냅샷(.pkl 저장) 방식 — 기본 끔. (원본 범위 자르기로 충분)
#   True로 켜면 분석 때 feat/close를 ..._data.pkl로 저장하고 재현 때 그걸 그대로 사용.
SAVE_DATA_SNAPSHOT  = False
REPLAY_USE_SNAPSHOT = False

# ★ summary(전체 티커 요약) 파일 생성 여부 — 요청으로 기본 OFF.
#   끄면 종목별 분석 엑셀만 만들고, ensemble_summary_all_tickers.xlsx는 안 만든다.
WRITE_SUMMARY_FILE = False

# ════════════════════════════════════════════════════════════════
#   ★ 실행 모드 — input() 대신 여기서 직접 지정 (요청)
#     RUN_MODE 로 동작을 고른다. Colab에서 이 값만 바꿔 실행하면 됨.
# ════════════════════════════════════════════════════════════════
#   RUN_MODE = 1 : 새 분석 실행
#            = 2 : 기존 결과의 '그리드 번호'로 일별 Excel 재현
#            = 3 : 티커의 '가장 최근 분석 엑셀'에서 ★최적 조합 자동 재현 (지표·변수 그대로)
#            = 4 : 드라이브 폴더의 '티커별 가장 최근' 엑셀들을 모두 ★최적 재현(현재까지) + 요약
RUN_MODE = 3

# ── 모드 1 (새 분석) 설정 ──
#   분석할 티커 목록. None 이면 코드의 기본 TICKERS 사용.
RUN_TICKERS = None                 # 예: ['STX', 'GOOG', 'ORCL']  또는  None
#   기존 요약 엑셀의 변수값을 사용할지 (True/False)
RUN_USE_EXCEL_OVERRIDES = False

# ── 모드 2 (그리드 번호 재현) 설정 ──
RUN_REPLAY_FILE = 'ensemble_search_VRT_2026-06-01.xlsx'   # OUTPUT_DIR 안의 파일명
RUN_REPLAY_GRID_NUMBER = 14                                # 재현할 그리드 번호

# ── 모드 3 (최근 엑셀 ★최적 자동 재현) 설정 ──
RUN_REPLAY_TICKER = 'STX'          # 재현할 티커 (가장 최근 일반 분석 엑셀을 자동으로 찾음)

# ── 모드 4 (드라이브 폴더 일괄 ★재현 + 요약) 설정 ──
#   드라이브의 ensemble_analysis 폴더에서 '티커별 가장 최근' 분석 엑셀을 모두 찾아,
#   각 ★최적 조합을 '현재까지' 일별 백테스트로 재현하고 티커별 요약 엑셀까지 만든다.
#   결과물은 '날짜 폴더'(드라이브폴더/오늘날짜)에 저장. (먼저 drive.mount 필요)
RUN_MODE4_DRIVE_DIR   = '/content/drive/MyDrive/ensemble_analysis'  # 드라이브 폴더
RUN_MODE4_DATE_FOLDER = None       # 출력 날짜 폴더명. None=오늘 날짜 자동


def wilson_lower(k, n, z):
    if n == 0: return 0.0
    p = k / n
    denom  = 1.0 + z * z / n
    center = p + z * z / (2 * n)
    half   = z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n))
    return max(0.0, (center - half) / denom)


def compute_vote_weights(scores, max_ratio=1.6):
    """성공률(Wilson 점수)에 비례한 투표 가중치. 데이터가 강도를 결정.
       - 점수 분포 편차 크면 차등↑, 작으면 균등에 가깝게
       - 최고/최저 비율 max_ratio 이하 (온건)
       - 평균 1로 정규화 → 기존 vote 임계값 스케일 그대로 호환"""
    s = np.clip(np.asarray(scores, dtype=np.float64), 0.0, 1.0)
    n = len(s)
    if n == 0:
        return np.ones(0, dtype=np.float64)
    spread = float(s.max() - s.min())
    if spread < 1e-9:
        return np.ones(n, dtype=np.float64)

    # ★ 희석 방지 장치 (요청) — 절대 점수차에 지수 반응(소프트맥스형). 강한 지표 표가 급증해
    #   자잘한 다수에 묻히지 않음. 상·하한으로 단일 지표의 과도지배만 막는다(적당히).
    if globals().get('USE_ANTIDILUTION', False):
        T   = max(float(globals().get('VOTE_WEIGHT_TEMP', 0.15)), 1e-6)
        cap = float(globals().get('VOTE_STRONG_CAP', 4.0))
        w = np.exp((s - s.max()) / T)      # 0<w<=1, 최고점수=1 (수치 안정)
        w = w / w.mean()                   # 평균 1 정규화 (vote 비율 스케일 호환)
        w = np.clip(w, 1.0 / cap, cap)     # 과도지배/과도소멸 방지
        w = w / w.mean()
        return w.astype(np.float64)

    # (기존) 선형 가중 — 최고/최저 비율 max_ratio 이하
    z = (s - s.min()) / spread
    beta = min(1.0, float(np.std(s)) / 0.10)
    w = 1.0 + z * (max_ratio - 1.0) * beta
    w = w / w.mean()
    return w.astype(np.float64)


def compute_stability_scores(df, weights=None):
    """후보 DataFrame에 '안정성 종합 점수'(stability_score) 컬럼을 붙여 반환.
       구성: 매도성공률, 매수성공률, 누적수익(후보군 min-max), MDD방어(후보군 min-max).
       가중 기하평균 — 한 요소라도 후보군 내 최저 수준이면 점수가 크게 깎여
       '종합적으로 안정적인'(골고루 좋은) 조합이 상위로 온다.
       거래수는 별도 MIN_TRADES_DAILY 필터로 이미 보장되므로 점수에는 안 넣음."""
    import numpy as np
    if weights is None:
        weights = globals().get('STABILITY_WEIGHTS', (0.30, 0.25, 0.20, 0.25))
    w = np.asarray(weights, dtype=np.float64)
    w = w / w.sum()
    out = df.copy()
    if len(out) == 0:
        out['stability_score'] = []
        return out
    ret_col = 'total_return' if 'total_return' in out.columns else 'combined_return'
    rets = out[ret_col].astype(float).values
    mdds = out['max_drawdown'].astype(float).values   # 음수, 0에 가까울수록 좋음
    ret_lo, ret_hi = float(np.min(rets)), float(np.max(rets))
    mdd_worst, mdd_best = float(np.min(mdds)), float(np.max(mdds))

    def _nrm(x, lo, hi):
        if hi - lo < 1e-12: return 0.5
        return min(1.0, max(0.0, (x - lo) / (hi - lo)))

    eps = 1e-6
    scores = np.zeros(len(out), dtype=np.float64)
    sell = out['sell_success_rate'].astype(float).values
    buy  = out['buy_success_rate'].astype(float).values
    for i in range(len(out)):
        s = min(1.0, max(0.0, sell[i]))
        b = min(1.0, max(0.0, buy[i]))
        r = _nrm(rets[i], ret_lo, ret_hi)
        m = _nrm(mdds[i], mdd_worst, mdd_best)   # MDD 방어력 (0근처=1, 최악=0)
        comps = (max(s, eps), max(b, eps), max(r, eps), max(m, eps))
        scores[i] = float(np.prod([c ** wi for c, wi in zip(comps, w)]))
    out['stability_score'] = scores
    return out


def _bh_sum_return(close_arr):
    a = np.asarray(close_arr, dtype=np.float64)
    if len(a) < 2:
        return 0.0
    rets = a[1:] / a[:-1] - 1.0
    rets = rets[np.isfinite(rets)]
    return float(np.sum(rets))


def _bh_up_sum_return(close_arr):
    a = np.asarray(close_arr, dtype=np.float64)
    if len(a) < 2:
        return 0.0
    rets = a[1:] / a[:-1] - 1.0
    rets = rets[np.isfinite(rets)]
    return float(np.sum(rets[rets > 0.0]))


def _optimal_swing_points(prices, cost, price_tolerance=0.01):
    """진짜 최대 수익 스윙의 저점(매수)/고점(매도)을 찾는다 (요청).
       anchor_strategy_return과 동일한 로직: 단기 저점에서 사서 오르는 한 보유,
       단기 고점에서 청산. 단 거래비용(2*cost)을 넘는 상승 구간만 채택.
       그 저점/고점 직전 ±price_tolerance 이내 가격이면 같은 앵커로 포함.
       반환: (buy_idx_set, sell_idx_set) — 우선순위 1 (최우선 매칭 대상)
    """
    n = len(prices)
    buy_idx = set(); sell_idx = set()
    i = 0
    while i < n - 1:
        if prices[i+1] > prices[i]:
            j = i                      # j = 국소 저점
            k = j
            while k < n - 1 and prices[k+1] >= prices[k]:
                k += 1                 # k = 국소 고점
            gain = prices[k] / prices[j] - 1.0
            net = (1.0 - cost) * (1.0 + gain) * (1.0 - cost) - 1.0
            if net > 0:                # 비용 넘는 상승 구간만 채택
                base_b = prices[j]
                base_s = prices[k]
                # 저점 j + 직전 ±tol 이내 (j 주변, 같은 가격대)
                for m in range(max(0, j - 5), j + 1):
                    if base_b > 0 and abs(prices[m] / base_b - 1.0) <= price_tolerance:
                        buy_idx.add(m)
                # 고점 k + 직전 ±tol 이내
                for m in range(max(0, k - 5), k + 1):
                    if base_s > 0 and abs(prices[m] / base_s - 1.0) <= price_tolerance:
                        sell_idx.add(m)
            i = k + 1
        else:
            i += 1
    return buy_idx, sell_idx


def auto_compute_anchor_dates(dates, close, *,
                                window=5,
                                lookforward=5,
                                min_rise_after_buy=0.03,
                                min_drop_after_sell=0.03,
                                price_tolerance=0.005,
                                max_dates=None,
                                return_priority=False):
    """앵커 정답일 생성 (요청 개정):
       (우선순위1) 진짜 최대 수익 스윙의 저점(매수)/고점(매도) + 직전 ±tol 이내.
       (우선순위2) 기존 방식: 좌우 window 최저/최고 & lookforward 내 min_rise/drop 이상.
       두 종류를 합치되, 우선순위1을 최우선 매칭 대상으로 표시한다.
       return_priority=True면 (buy_dates, sell_dates, prio_dict) 반환.
       prio_dict: {'buy_p1':set, 'sell_p1':set, 'buy_p2':set, 'sell_p2':set} (날짜 문자열)
    """
    prices = close.values.astype(np.float64)
    n = len(prices)
    cost = float(globals().get('COST_PER_TRADE', 0.004))

    # ── 우선순위1: 최대 수익 스윙 저점/고점 (+ 직전 ±tol) ──
    sw_buy_idx, sw_sell_idx = _optimal_swing_points(prices, cost, price_tolerance)

    # ── 우선순위2: 기존 1% 이상 앵커 ──
    base_buys = []
    base_sells = []
    for j in range(window, n - lookforward):
        win_lo = max(0, j - window)
        win_hi = min(n - 1, j + window)
        win = prices[win_lo:win_hi + 1]
        cur = prices[j]
        future = prices[j:j + lookforward + 1]
        if cur == win.min() and len(future) > 1:
            max_rise = future.max() / cur - 1.0
            if max_rise >= min_rise_after_buy:
                base_buys.append((j, max_rise))
        if cur == win.max() and len(future) > 1:
            max_drop = future.min() / cur - 1.0
            if max_drop <= -min_drop_after_sell:
                base_sells.append((j, -max_drop))
    base_buys.sort(key=lambda p: -p[1])
    base_sells.sort(key=lambda p: -p[1])
    if max_dates is not None:
        base_buys  = base_buys[:max_dates]
        base_sells = base_sells[:max_dates]

    p2_buy_idx = set(); p2_sell_idx = set()
    n_buy_base = 0; n_sell_base = 0
    for j, _ in base_buys:
        base_p = prices[j]
        if base_p <= 0: continue
        n_buy_base += 1
        for k in range(max(0, j - window), min(n, j + window + 1)):
            if prices[k] > 0 and abs(prices[k] / base_p - 1.0) <= price_tolerance:
                p2_buy_idx.add(k)
    for j, _ in base_sells:
        base_p = prices[j]
        if base_p <= 0: continue
        n_sell_base += 1
        for k in range(max(0, j - window), min(n, j + window + 1)):
            if prices[k] > 0 and abs(prices[k] / base_p - 1.0) <= price_tolerance:
                p2_sell_idx.add(k)

    # 우선순위1이 우선 — p2에서 p1과 겹치는 건 p1로 귀속
    p2_buy_only  = p2_buy_idx  - sw_buy_idx
    p2_sell_only = p2_sell_idx - sw_sell_idx

    all_buy_idx  = sw_buy_idx  | p2_buy_only
    all_sell_idx = sw_sell_idx | p2_sell_only

    def _to_dates(idx_set):
        return sorted([dates[k].strftime('%Y-%m-%d') for k in idx_set])

    buy_dates  = _to_dates(all_buy_idx)
    sell_dates = _to_dates(all_sell_idx)

    print(f"     앵커: [우선순위1 최대수익스윙] 매수 {len(sw_buy_idx)}일 / 매도 {len(sw_sell_idx)}일,  "
          f"[우선순위2 {min_rise_after_buy*100:.0f}%이상] 매수 +{len(p2_buy_only)}일 / 매도 +{len(p2_sell_only)}일  "
          f"(직전 ±{price_tolerance*100:.1f}% 포함)")

    if return_priority:
        prio = {
            'buy_p1':  set(_to_dates(sw_buy_idx)),
            'sell_p1': set(_to_dates(sw_sell_idx)),
            'buy_p2':  set(_to_dates(p2_buy_only)),
            'sell_p2': set(_to_dates(p2_sell_only)),
        }
        return buy_dates, sell_dates, prio
    return buy_dates, sell_dates


@njit
def _eval_buy_signals(close_arr, signal_arr, horizon, dd_limit, anchor_buy_arr):
    # ★ 기준(요청): 매수신호 적중 = '신호 다음날 종가'가 '신호일 종가' 대비 +dd_limit(=1%) 이상 상승.
    #   ★ 앵커 오버라이드 제거(요청 A) — 앵커일이라도 다음날 +1% 못 오르면 '실패'.
    #     → 성공률 100%면 '모든 신호가 실제로 다음날 +1% 상승'을 의미.
    #   anchor_buy_arr / horizon 인자는 호출 호환 위해 남겨두나 더는 사용하지 않음.
    n = close_arr.shape[0]
    ns = 0; ok = 0; sum_ret = 0.0
    for i in range(n - 1):
        if signal_arr[i] != 1: continue
        base_p = close_arr[i]
        if base_p <= 0.0: continue
        nxt_p = close_arr[i + 1]
        ret = nxt_p / base_p - 1.0
        ns += 1; sum_ret += ret
        if ret >= dd_limit:
            ok += 1
    return ns, ok, sum_ret


@njit(cache=True)
def _eval_sell_signals(close_arr, signal_arr, horizon, ru_limit, anchor_sell_arr):
    # ★ 기준(요청): 매도신호 적중 = '신호 다음날 종가'가 '신호일 종가' 대비 -ru_limit(=1%) 이상 하락.
    #   ★ 앵커 오버라이드 제거(요청 A) — 앵커일이라도 다음날 -1% 못 내리면 '실패'.
    #   anchor_sell_arr / horizon 인자는 호출 호환 위해 남겨두나 미사용.
    n = close_arr.shape[0]
    ns = 0; ok = 0; sum_ret = 0.0
    for i in range(n - 1):
        if signal_arr[i] != 1: continue
        base_p = close_arr[i]
        if base_p <= 0.0: continue
        nxt_p = close_arr[i + 1]
        ret = nxt_p / base_p - 1.0
        ns += 1; sum_ret += ret
        if ret <= -ru_limit:
            ok += 1
    return ns, ok, sum_ret


@njit(cache=True)
def _eval_big_move_hits(close_arr, signal_arr, horizon, big_thr, is_buy):
    """신호 중 '큰 움직임'을 맞춘 비율 계산용 (요청).
       매수신호(is_buy=1): 익일 진입가 대비 이후 horizon 내 최대 상승률이 big_thr 이상이면 적중.
       매도신호(is_buy=0): 익일 기준가 대비 이후 horizon 내 최대 하락률이 -big_thr 이하면 적중
         (= 매도해서 큰 하락을 피했다).
       반환: (n_signals, n_big_hits)."""
    n = close_arr.shape[0]
    ns = 0; hit = 0
    for i in range(n - 1):
        if signal_arr[i] != 1: continue
        if i + 1 >= n: break
        base_p = close_arr[i + 1]
        if base_p <= 0.0: continue
        end = i + 1 + horizon
        if end >= n: end = n - 1
        if end <= i + 1: continue
        if is_buy == 1:
            mx = base_p
            for j in range(i + 2, end + 1):
                if close_arr[j] > mx: mx = close_arr[j]
            ns += 1
            if mx / base_p - 1.0 >= big_thr: hit += 1
        else:
            mn = base_p
            for j in range(i + 2, end + 1):
                if close_arr[j] < mn: mn = close_arr[j]
            ns += 1
            if mn / base_p - 1.0 <= -big_thr: hit += 1
    return ns, hit


@njit(cache=True)
def _compute_safe_arrays(close_arr, horizon, dd_limit, ru_limit):
    """
    ★ 기준 변경(요청) — 정답일을 '신호 다음날 종가' 기준으로.
      safe_buy[i]=1  : 다음날 종가가 신호일 종가 대비 +dd_limit(=1%) 이상 상승 → '올랐어야'(매수 정답)
      safe_sell[i]=1 : 다음날 종가가 -ru_limit(=1%) 이상 하락 → '내렸어야'(매도 정답)
      둘 다 아니면(미미한 변동) evaluable=0 → 성공/실패 평가에서 제외(부풀리기 방지).
    ※ horizon 인자는 기존 호출 호환 위해 남겨두나 더는 사용하지 않음(다음날만 본다).
      dd_limit=매수 상승목표(+1%), ru_limit=매도 하락목표(-1%).
    """
    n = close_arr.shape[0]
    safe_buy  = np.zeros(n, dtype=np.uint8)
    safe_sell = np.zeros(n, dtype=np.uint8)
    evaluable = np.zeros(n, dtype=np.uint8)
    for i in range(n - 1):
        base = close_arr[i]
        if base <= 0.0:
            continue
        ret = close_arr[i + 1] / base - 1.0
        if ret >= dd_limit:
            safe_buy[i] = 1; evaluable[i] = 1
        elif ret <= -ru_limit:
            safe_sell[i] = 1; evaluable[i] = 1
        # else: 미미한 변동 → evaluable[i] 그대로 0 (평가 제외)
    return safe_buy, safe_sell, evaluable


def _compute_anchor_arrays(dates, anchor_buy_dates, anchor_sell_dates):
    n = len(dates)
    anchor_buy  = np.zeros(n, dtype=np.uint8)
    anchor_sell = np.zeros(n, dtype=np.uint8)

    def _to_norm_set(date_list):
        s = set()
        for d in date_list:
            try:
                ts = pd.Timestamp(d).normalize()
                s.add(ts)
            except Exception:
                pass
        return s

    buy_set  = _to_norm_set(anchor_buy_dates)
    sell_set = _to_norm_set(anchor_sell_dates)

    norm_dates = pd.DatetimeIndex(dates).normalize()
    for i in range(n - 1):
        nd = norm_dates[i + 1]
        if nd in buy_set:  anchor_buy[i]  = 1
        if nd in sell_set: anchor_sell[i] = 1

    n_buy_total  = len(buy_set)
    n_sell_total = len(sell_set)
    n_buy_matched  = int(anchor_buy.sum())
    n_sell_matched = int(anchor_sell.sum())
    if n_buy_total > 0 and n_buy_matched < n_buy_total:
        print(f"  ⚠ ANCHOR: 매수 정답일 {n_buy_total}개 중 {n_buy_matched}개만 데이터 범위 내")
    if n_sell_total > 0 and n_sell_matched < n_sell_total:
        print(f"  ⚠ ANCHOR: 매도 정답일 {n_sell_total}개 중 {n_sell_matched}개만 데이터 범위 내")
    return anchor_buy, anchor_sell


def _apply_anchor_correction(safe_buy, safe_sell, evaluable,
                              anchor_buy, anchor_sell):
    safe_buy_c  = np.maximum(safe_buy,  anchor_buy)
    safe_sell_c = np.maximum(safe_sell, anchor_sell)
    eval_c = np.maximum(evaluable, np.maximum(anchor_buy, anchor_sell))
    if len(eval_c) > 0:
        eval_c[-1] = 0
    return safe_buy_c, safe_sell_c, eval_c


@njit(cache=True)
def _simulate_ensemble(close_arr, buy_sig_mat, sell_sig_mat,
                        vote_buy, vote_sell, cost,
                        safe_buy, safe_sell, evaluable,
                        stop_loss_pct,
                        anchor_buy_arr, anchor_sell_arr,
                        buy_w, sell_w):
    n   = close_arr.shape[0]
    K_b = buy_sig_mat.shape[1]
    K_s = sell_sig_mat.shape[1]
    pos = 0; ent = 0.0; entidx = -1
    sum_daily = 0.0
    nt = 0; nw = 0; sr = 0.0; sqr = 0.0
    me = 1.0; md = 0.0
    n_eval = 0
    n_buy_correct = 0
    n_sell_correct = 0
    n_buy_tp = 0; n_buy_fp = 0; n_buy_fn = 0
    n_sell_tp = 0; n_sell_fp = 0; n_sell_fn = 0
    n_stop_triggered = 0
    n_anchor_buy_total = 0
    n_anchor_buy_matched = 0
    n_anchor_sell_total = 0
    n_anchor_sell_matched = 0
    use_anchor_buy  = anchor_buy_arr.shape[0]  == n
    use_anchor_sell = anchor_sell_arr.shape[0] == n
    for i in range(n):
        if pos == 1 and entidx >= 0 and i > entidx and i >= 1:
            if close_arr[i - 1] > 0.0:
                sum_daily += close_arr[i] / close_arr[i - 1] - 1.0

        stopped_today = False
        if pos == 1 and stop_loss_pct > 0.0 and ent > 0.0:
            cur_p = close_arr[i]
            if cur_p > 0.0 and (cur_p / ent - 1.0) <= -stop_loss_pct:
                ret = cur_p / ent - 1.0
                sum_daily -= cost
                nt += 1
                if ret > 0.0: nw += 1
                sr += ret; sqr += ret * ret
                pos = 0
                ent = 0.0
                entidx = -1
                n_stop_triggered += 1
                stopped_today = True

        cur_eq = 1.0 + sum_daily
        if cur_eq > me: me = cur_eq
        dd = cur_eq / me - 1.0
        if dd < md: md = dd

        if i >= n - 1:
            if use_anchor_buy and anchor_buy_arr[i] == 1:
                n_anchor_buy_total += 1
                if pos == 1: n_anchor_buy_matched += 1
            if use_anchor_sell and anchor_sell_arr[i] == 1:
                n_anchor_sell_total += 1
                if pos == 0: n_anchor_sell_matched += 1
            continue

        b_count = 0.0
        for k in range(K_b):
            if buy_sig_mat[i, k] == 1: b_count += buy_w[k]
        s_count = 0.0
        for k in range(K_s):
            if sell_sig_mat[i, k] == 1: s_count += sell_w[k]

        b_on = b_count >= vote_buy
        s_on = s_count >= vote_sell

        if evaluable[i] == 1:
            n_eval += 1
            sb = safe_buy[i] == 1
            ss = safe_sell[i] == 1
            if (b_on and sb) or ((not b_on) and (not sb)):
                n_buy_correct += 1
            if (s_on and ss) or ((not s_on) and (not ss)):
                n_sell_correct += 1
            if b_on and sb: n_buy_tp += 1
            elif b_on and (not sb): n_buy_fp += 1
            elif (not b_on) and sb: n_buy_fn += 1
            if s_on and ss: n_sell_tp += 1
            elif s_on and (not ss): n_sell_fp += 1
            elif (not s_on) and ss: n_sell_fn += 1

        if stopped_today: continue

        if K_b > 0:
            b_strength = b_count / K_b
        else:
            b_strength = 0.0
        if K_s > 0:
            s_strength = s_count / K_s
        else:
            s_strength = 0.0

        if pos == 0:
            if b_on:
                if s_on and s_strength > b_strength:
                    pass
                else:
                    p = close_arr[i + 1]
                    if p > 0.0:
                        ent = p; pos = 1; entidx = i + 1
                        sum_daily -= cost
        else:
            if s_on:
                if b_on and b_strength > s_strength:
                    pass
                else:
                    p = close_arr[i + 1]
                    if p > 0.0 and ent > 0.0:
                        ret = p / ent - 1.0
                        if close_arr[i] > 0.0:
                            sum_daily += close_arr[i + 1] / close_arr[i] - 1.0
                        sum_daily -= cost
                        nt += 1
                        if ret > 0.0: nw += 1
                        sr += ret; sqr += ret * ret
                    pos = 0
                    ent = 0.0
                    entidx = -1

        if use_anchor_buy and anchor_buy_arr[i] == 1:
            n_anchor_buy_total += 1
            if pos == 1: n_anchor_buy_matched += 1
        if use_anchor_sell and anchor_sell_arr[i] == 1:
            n_anchor_sell_total += 1
            if pos == 0: n_anchor_sell_matched += 1
    if pos == 1 and ent > 0.0:
        ret = close_arr[n - 1] / ent - 1.0
        sum_daily -= cost
        nt += 1
        if ret > 0.0: nw += 1
        sr += ret; sqr += ret * ret
        cur_eq = 1.0 + sum_daily
        if cur_eq > me: me = cur_eq
        dd = cur_eq / me - 1.0
        if dd < md: md = dd
    return (sum_daily, nt, nw, sr, sqr, md,
            n_eval, n_buy_correct, n_sell_correct, n_stop_triggered,
            n_buy_tp, n_buy_fp, n_buy_fn,
            n_sell_tp, n_sell_fp, n_sell_fn,
            n_anchor_buy_total, n_anchor_buy_matched,
            n_anchor_sell_total, n_anchor_sell_matched)


def prefilter_indicators(feat, close, *,
                          horizon=None,
                          min_corr=None,
                          min_variance_rel=None,
                          max_nan_ratio=None,
                          verbose=True):
    if horizon          is None: horizon          = HORIZON_DAYS
    if min_corr         is None: min_corr         = PREFILTER_MIN_CORR
    if min_variance_rel is None: min_variance_rel = PREFILTER_MIN_VARIANCE_REL
    if max_nan_ratio    is None: max_nan_ratio    = PREFILTER_MAX_NAN_RATIO

    future_ret = close.shift(-horizon) / close - 1
    future_ret = future_ret.dropna()

    n_orig = len(feat.columns)
    keep = []
    drop_nan = drop_var = drop_corr = 0
    n_days = len(feat)

    for col in feat.columns:
        x = feat[col]
        nan_ratio = x.isna().sum() / max(n_days, 1)
        if nan_ratio > max_nan_ratio:
            drop_nan += 1
            continue

        x_valid = x.dropna()
        if len(x_valid) < 50:
            drop_nan += 1
            continue

        std_v  = float(x_valid.std())
        mean_v = float(x_valid.mean())
        if std_v < 1e-12:
            drop_var += 1
            continue
        rel_var = std_v / max(abs(mean_v), 1e-10)
        if abs(mean_v) > 1e-10 and rel_var < min_variance_rel:
            drop_var += 1
            continue

        common = x_valid.index.intersection(future_ret.index)
        if len(common) < 50:
            keep.append(col)
            continue
        try:
            corr = float(x.loc[common].corr(future_ret.loc[common]))
        except Exception:
            corr = np.nan
        if pd.isna(corr) or abs(corr) < min_corr:
            drop_corr += 1
            continue

        keep.append(col)

    if verbose:
        n_kept = len(keep)
        pct_kept = n_kept / max(n_orig, 1) * 100
        print(f"  🔍 지표 사전 필터: {n_orig} → {n_kept}개 ({pct_kept:.1f}% 유지)")
        print(f"     제거 — NaN과다: {drop_nan}, 저분산: {drop_var}, 저상관(|r|<{min_corr}): {drop_corr}")
        if n_kept < 30:
            print(f"     ⚠ 유지 지표 너무 적음. PREFILTER_MIN_CORR를 낮추거나 PREFILTER_ENABLED=False 권장")

    return feat[keep]


def _select_indicators(feat, max_n=None):
    valid = []
    for col in feat.columns:
        x = feat[col].values.astype(np.float64)
        m = ~np.isnan(x)
        if m.sum() < 100: continue
        v = x[m]
        sd = float(np.std(v))
        if sd < 1e-12: continue
        cv = sd / max(abs(float(np.mean(v))), 1e-10)
        valid.append((col, cv))
    if max_n and len(valid) > max_n:
        valid.sort(key=lambda p: -p[1])
        valid = valid[:max_n]
    return [c for c, _ in valid]


def _rolling_zscore(x, window):
    """롤링 z-스코어 = (x - 이동평균) / 이동표준편차. 미래참조 없음(과거 window만)."""
    s = pd.Series(x)
    mu = s.rolling(window, min_periods=max(10, window//3)).mean()
    sd = s.rolling(window, min_periods=max(10, window//3)).std(ddof=0)
    z = ((s - mu) / sd.replace(0.0, np.nan)).values
    return z


def _stability_adjusted_score(close_arr, sig_arr, horizon, limit, anchor_arr,
                              wilson_z, is_buy, min_signals):
    """지표 신호의 점수 = Wilson 하한 × (1 + w·시간안정성).
       시간안정성: 신호일을 앞/뒤 절반으로 나눠 각각 성공률을 구하고,
         둘 다 좋고 서로 비슷할수록(미래에도 유지될 가능성↑) 가산점.
       반환: (n, ok, sum, score). 신호 부족 시 score=음수로 사실상 제외.
    """
    evalf = _eval_buy_signals if is_buy else _eval_sell_signals
    n_all, ok_all, sum_all = evalf(close_arr, sig_arr, horizon, limit, anchor_arr)
    if n_all < min_signals:
        return n_all, ok_all, sum_all, -1.0
    base = wilson_lower(ok_all, n_all, wilson_z)

    # ★ 큰 움직임 적중 가산 (요청) — 신호 중 '큰 상승/하락'을 맞춘 비율이 높을수록 점수↑.
    #   자잘한 적중만 많은 지표보다 굵직한 변동을 잡는 지표를 우대 → 실전 수익 직결.
    if globals().get('USE_BIG_MOVE_BONUS', False):
        big_thr = float(globals().get('BIG_MOVE_THRESHOLD', 0.03))
        bw      = float(globals().get('BIG_MOVE_BONUS_WEIGHT', 0.5))
        bn, bhit = _eval_big_move_hits(close_arr, sig_arr, horizon, big_thr,
                                       1 if is_buy else 0)
        if bn > 0:
            big_ratio = bhit / bn          # 0~1
            base = base * (1.0 + bw * big_ratio)

    if not globals().get('USE_OOS_STABILITY', False):
        return n_all, ok_all, sum_all, base
    # 앞/뒤 절반으로 신호 분할 (시간 순)
    nd = len(sig_arr); half = nd // 2
    sig_a = sig_arr.copy(); sig_a[half:] = 0     # 앞 절반만
    sig_b = sig_arr.copy(); sig_b[:half] = 0     # 뒤 절반만
    na, oka, _ = evalf(close_arr, sig_a, horizon, limit, anchor_arr)
    nb, okb, _ = evalf(close_arr, sig_b, horizon, limit, anchor_arr)
    if na < 3 or nb < 3:
        return n_all, ok_all, sum_all, base   # 한쪽 표본 부족 → 안정성 평가 불가, 기본점수
    ra = oka / na; rb = okb / nb
    # 안정성 = 두 기간 모두 0.5 넘고 + 서로 가까움. [0,1] 정규화.
    both_good = min(ra, rb)                       # 둘 중 나쁜 쪽 (낮으면 불안정)
    consistency = 1.0 - min(1.0, abs(ra - rb) / 0.5)  # 차이 작을수록 1
    stability = max(0.0, (both_good - 0.5) * 2.0) * consistency  # 0~1
    w = float(globals().get('OOS_STABILITY_WEIGHT', 0.3))
    score = base * (1.0 + w * stability)
    return n_all, ok_all, sum_all, score


def evaluate_buy_sell_scores(feat, close, *, indicators,
                              n_thresholds, pct_low, pct_high,
                              horizon, dd_limit, ru_limit,
                              min_signals, wilson_z,
                              anchor_buy_arr=None, anchor_sell_arr=None):
    close_arr = close.values.astype(np.float64)
    n_days    = len(close_arr)
    pcts      = np.linspace(pct_low, pct_high, n_thresholds)
    use_z     = globals().get('USE_ZSCORE_SIGNAL', False)
    z_window  = int(globals().get('ZSCORE_WINDOW', 60))
    z_thrs    = globals().get('ZSCORE_THRESHOLDS', [-2.0, 2.0])

    if anchor_buy_arr is None:
        anchor_buy_arr = np.zeros(0, dtype=np.uint8)
    if anchor_sell_arr is None:
        anchor_sell_arr = np.zeros(0, dtype=np.uint8)

    if HAS_NUMBA:
        zero_sig = np.zeros(n_days, dtype=np.uint8)
        _eval_buy_signals(close_arr, zero_sig, horizon, dd_limit, anchor_buy_arr)
        _eval_sell_signals(close_arr, zero_sig, horizon, ru_limit, anchor_sell_arr)

    buy_rows = []; sell_rows = []
    for col in indicators:
        x = feat[col].values.astype(np.float64)
        valid = ~np.isnan(x)
        if valid.sum() < 100: continue

        # ── 신호 후보 목록: (방향라벨, 임계라벨, 임계값, 신호배열) ──
        sig_specs = []
        # (a) 기존 절대 임계 (백분위 기반)
        for p in pcts:
            thr = float(np.nanpercentile(x, p))
            sig_specs.append(('>=', float(p), thr, ((x >= thr) & valid).astype(np.uint8)))
            sig_specs.append(('<=', float(p), thr, ((x <= thr) & valid).astype(np.uint8)))
        # (b) z-스코어 임계 (롤링 정규화) — 미래 일반화에 강함 (요청)
        if use_z:
            z = _rolling_zscore(x, z_window)
            zvalid = ~np.isnan(z)
            for zt in z_thrs:
                if zt >= 0:
                    s = ((z >= zt) & zvalid).astype(np.uint8); d = 'z>='
                else:
                    s = ((z <= zt) & zvalid).astype(np.uint8); d = 'z<='
                sig_specs.append((d, float(zt), float(zt), s))

        for direction, plabel, thr, sig_arr in sig_specs:
            if int(sig_arr.sum()) < min_signals: continue
            bn, bok, bsum, bscore = _stability_adjusted_score(
                close_arr, sig_arr, horizon, dd_limit, anchor_buy_arr,
                wilson_z, True, min_signals)
            if bn >= min_signals and bscore >= 0:
                buy_rows.append((col, direction, plabel, thr,
                                  bn, bok, bok / bn,
                                  float(bsum / bn) if bn else 0.0, bscore))
            sn, sok, ssum, sscore = _stability_adjusted_score(
                close_arr, sig_arr, horizon, ru_limit, anchor_sell_arr,
                wilson_z, False, min_signals)
            if sn >= min_signals and sscore >= 0:
                sell_rows.append((col, direction, plabel, thr,
                                   sn, sok, sok / sn,
                                   float(ssum / sn) if sn else 0.0, sscore))

    cols = ['indicator', 'direction', 'pct_label', 'threshold',
            'n_signals', 'n_success', 'success_rate', 'avg_extreme', 'score']
    buy_df  = pd.DataFrame(buy_rows,  columns=cols)
    sell_df = pd.DataFrame(sell_rows, columns=cols)
    if globals().get('POOL_SELECT_BY_SUCCESS', False):
        # ★ 성공률 우선(요청) — 모든 풀 경로 공통 처리:
        #   ① 지표당 '최고 성공률' 1행만 남김(중복 임계 제거) ② 성공률 내림차순 정렬(동률은 점수).
        #   하드 컷오프(70%)는 여기 두지 않음 — 풀이 비어 거래 0이 되지 않게 best-available로 채움.
        #   (70% 컷오프는 '성공률 우선' 시트 표시용으로만 적용)
        def _succ_rank(df):
            if df is None or len(df) == 0:
                return (df.sort_values('score', ascending=False).reset_index(drop=True)
                        if df is not None else df)
            d = df.sort_values(['success_rate', 'score'], ascending=[False, False])
            d = d.drop_duplicates('indicator', keep='first').reset_index(drop=True)
            return d
        buy_df  = _succ_rank(buy_df)
        sell_df = _succ_rank(sell_df)
    else:
        buy_df  = buy_df.sort_values('score', ascending=False).reset_index(drop=True)
        sell_df = sell_df.sort_values('score', ascending=False).reset_index(drop=True)
    return buy_df, sell_df


def select_pool_by_success(feat, close, *, indicators, n_thresholds,
                           horizon, dd_limit, ru_limit, wilson_z=1.0):
    """★ 성공률 우선 풀 선출 (요청).
       - 넓은 분위(POOL_SUCCESS_WIDE_PCT) + z-스코어 후보 전체로 모든 지표를 평가 →
         pct(분위)가 달라 따로 나오던 고성공 지표를 누락 없이 한 번에 평가(z스코어는 pct무관).
       - 표본 가드: 신호수 >= POOL_SUCCESS_MIN_SIG (가짜 100% 방지).
       - 성공률 >= POOL_SUCCESS_MIN_RATE 만 후보로.
       - 성공률 내림차순 정렬(동률은 점수=Wilson 하한으로 타이브레이크).
       - 지표당 1행(최고 성공률)만 남긴 dedup 버전도 함께 반환(시트/표시용).
       반환: (buy_full, sell_full, buy_dedup, sell_dedup)
         *_full   : 성공률순 정렬된 전체 후보행 (diversify_candidates에 그대로 투입 → 풀)
         *_dedup  : 지표당 최고성공 1행 (성공률 우선 선출 시트용)
    """
    lo, hi = POOL_SUCCESS_WIDE_PCT
    bdf, sdf = evaluate_buy_sell_scores(
        feat, close, indicators=indicators, n_thresholds=n_thresholds,
        pct_low=lo, pct_high=hi, horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
        min_signals=POOL_SUCCESS_MIN_SIG, wilson_z=wilson_z,
        anchor_buy_arr=None, anchor_sell_arr=None,   # 오버라이드 제거됨 → 앵커 무관(순수 다음날 ±1%)
    )

    def _filt(df):
        if df is None or len(df) == 0:
            return df.copy() if df is not None else df
        d = df[(df['n_signals'] >= POOL_SUCCESS_MIN_SIG) &
               (df['success_rate'] >= POOL_SUCCESS_MIN_RATE)].copy()
        # 성공률 우선, 동률이면 점수(Wilson 하한)로 — 가짜 100% 강등
        d = d.sort_values(['success_rate', 'score'], ascending=[False, False]).reset_index(drop=True)
        return d

    buy_full  = _filt(bdf)
    sell_full = _filt(sdf)
    # 지표당 1행(성공률 최고) — 이미 성공률순 정렬이라 first가 최고
    buy_dedup  = (buy_full.drop_duplicates('indicator', keep='first').reset_index(drop=True)
                  if buy_full is not None and len(buy_full) else buy_full)
    sell_dedup = (sell_full.drop_duplicates('indicator', keep='first').reset_index(drop=True)
                  if sell_full is not None and len(sell_full) else sell_full)
    return buy_full, sell_full, buy_dedup, sell_dedup


_ZCACHE = {}

def _free_global_caches(*, keep_pool_map=False):
    """전역 캐시·맵을 비워 메모리 누적을 막는다 (Colab 여러 번 실행 시 끊김 방지).
       - _ZCACHE: z-스코어 캐시 (지표×기간 배열 — 가장 큼)
       - _LAST_FULL_CAND_MAP: 메타조합별 전체 후보 풀 (수천 행 × 메타키 — 큼)
       - _LAST_POOL_MAP: 메타조합별 선정 풀 (엑셀 재현용 — keep_pool_map이면 유지)
       각 티커/분석 시작 시 호출. gc까지 강제해 즉시 회수."""
    import gc
    g = globals()
    if '_ZCACHE' in g:
        g['_ZCACHE'].clear()
    if '_LAST_FULL_CAND_MAP' in g:
        g['_LAST_FULL_CAND_MAP'].clear()
    if not keep_pool_map and '_LAST_POOL_MAP' in g:
        g['_LAST_POOL_MAP'].clear()
    gc.collect()


def _get_zscore_cached(feat, col, window):
    """feat[col]의 z-스코어를 캐싱해 반환.
       ★ 키를 '내용 기반'으로 (요청) — 예전엔 id(feat)였는데, 파이썬 객체 id는 GC 후
       재사용돼 '다른 feat'인데 같은 키가 되어 stale z-스코어를 반환할 수 있었다(검증≠최종 원인).
       이제 (길이, 시작/끝 인덱스, 해당 컬럼 첫/끝값, 지표명, window)로 키를 만들어,
       데이터가 같으면 재사용·다르면 반드시 새로 계산한다."""
    try:
        idx = feat.index
        col_vals = feat[col].values
        n = len(col_vals)
        # 내용 식별자 — 동일 데이터면 동일, 한 행만 달라도 달라짐
        ckey = (n,
                (idx[0] if n else None), (idx[-1] if n else None),
                (float(col_vals[0]) if n and col_vals[0] == col_vals[0] else None),
                (float(col_vals[-1]) if n and col_vals[-1] == col_vals[-1] else None))
    except Exception:
        ckey = (id(feat),)   # 안전 폴백
    key = (ckey, col, window)
    c = _ZCACHE.get(key)
    if c is not None and len(c) == len(feat):
        return c
    z = _rolling_zscore(feat[col].values.astype(np.float64), window)
    # 캐시 폭주 방지: 항목이 너무 많으면 비움
    if len(_ZCACHE) > 6000:
        _ZCACHE.clear()
    _ZCACHE[key] = z
    return z


def _to_signal_array(feat, row):
    d = row['direction']
    # z-스코어 신호 (롤링 정규화) — 절대임계와 달리 분포 적응적, 미래 일반화에 강함
    if d == 'z>=' or d == 'z<=':
        z = _get_zscore_cached(feat, row['indicator'], int(globals().get('ZSCORE_WINDOW', 60)))
        zvalid = ~np.isnan(z)
        if d == 'z>=':
            return ((z >= row['threshold']) & zvalid).astype(np.uint8)
        return ((z <= row['threshold']) & zvalid).astype(np.uint8)
    # 기존 절대 임계
    x = feat[row['indicator']].values.astype(np.float64)
    valid = ~np.isnan(x)
    if d == '>=':
        return ((x >= row['threshold']) & valid).astype(np.uint8)
    return ((x <= row['threshold']) & valid).astype(np.uint8)


def diversify_candidates(feat, score_df, *, top_n, corr_limit):
    if len(score_df) == 0: return score_df.copy()
    pool = []
    pool_signals = []
    pool_indicators = set()
    for _, row in score_df.iterrows():
        if len(pool) >= top_n: break
        if row['indicator'] in pool_indicators: continue
        sig = _to_signal_array(feat, row).astype(np.float64)
        redundant = False
        for prev_sig in pool_signals:
            mask = ~np.isnan(sig) & ~np.isnan(prev_sig)
            if mask.sum() < 10: continue
            a = sig[mask]; b = prev_sig[mask]
            if a.std() < 1e-9 or b.std() < 1e-9: continue
            c = float(np.corrcoef(a, b)[0, 1])
            if abs(c) >= corr_limit:
                redundant = True; break
        if redundant: continue
        pool.append(row.to_dict())
        pool_signals.append(sig)
        pool_indicators.add(row['indicator'])
    return pd.DataFrame(pool)


def _select_with_tolerance(df, tolerance,
                            primary='avg_success_rate',
                            secondary='combined_return'):
    """
    SELECTION_PRIORITY='sell_buy_return'이면:
      1차 매도성공률(top-tol 밴드) → 2차 매수성공률(top-tol 밴드) → 3차 누적수익 최대
    아니면 기존: primary(top-tol 밴드) 중 secondary 최대
    """
    if len(df) == 0: return None
    if secondary not in df.columns:
        secondary = 'total_return'

    prio = globals().get('SELECTION_PRIORITY', 'balacc_return')

    # ★ 일별거래 승률 밴드(-10%p) → 그중 누적수익 최고 (요청)
    #   1) 거래 승률(수익 낸 거래 비율) 최고에서 10%p 안의 후보
    #   2) 그중 누적수익이 가장 높은 것
    if prio == 'winrate_return' and 'win_rate' in df.columns:
        win_tol = globals().get('WINRATE_TOLERANCE', 0.10)
        best_win = df['win_rate'].max()
        band = df[df['win_rate'] >= best_win - win_tol].copy()
        if len(band) == 0:
            band = df[df['win_rate'] == best_win].copy()
        ret_col = 'total_return' if 'total_return' in band.columns else 'combined_return'
        band = band.sort_values([ret_col, 'win_rate'], ascending=False)
        return band.index[0]

    # ★ 매도성공 밴드(-2%p) → MDD최저 밴드(-1%p) → 누적수익 최고 (요청, 기본)
    #   1) 매도성공률 최고에서 2%p 안의 후보
    #   2) 그중 MDD 가장 낮은(0에 가까운)에서 1%p 안의 후보 (비슷한 낙폭은 묶음)
    #   3) 그중 누적수익이 가장 높은 것
    if prio == 'sell_mdd_return' and 'sell_success_rate' in df.columns \
       and 'max_drawdown' in df.columns:
        sell_tol = globals().get('SELL_SUCCESS_TOLERANCE', 0.02)
        mdd_tol  = globals().get('MDD_TOLERANCE', 0.01)
        best_sell = df['sell_success_rate'].max()
        b1 = df[df['sell_success_rate'] >= best_sell - sell_tol].copy()
        if len(b1) == 0:
            b1 = df[df['sell_success_rate'] == best_sell].copy()
        best_mdd = b1['max_drawdown'].max()           # 가장 0에 가까운(덜 빠진)
        b2 = b1[b1['max_drawdown'] >= best_mdd - mdd_tol].copy()
        if len(b2) == 0:
            b2 = b1
        ret_col = 'total_return' if 'total_return' in b2.columns else 'combined_return'
        b2 = b2.sort_values([ret_col, 'sell_success_rate'], ascending=False)
        return b2.index[0]

    # ★ 평균성공 밴드 → MDD최저 → 수익최대 모드 (요청, 기본).
    #   1) 평균성공률 최고에서 tolerance(밴드) 안의 후보를 모음
    #   2) 그중 MDD가 가장 낮은(0에 가까운) 것 — 0.1%p 단위로 같으면 '비슷'으로 보고
    #   3) 그 동률 그룹에서 누적수익이 가장 높은 것
    if prio == 'avgband_mdd_return' and 'avg_success_rate' in df.columns \
       and 'max_drawdown' in df.columns:
        best_avg = df['avg_success_rate'].max()
        band = df[df['avg_success_rate'] >= best_avg - tolerance].copy()
        if len(band) == 0:
            band = df[df['avg_success_rate'] == best_avg].copy()
        band['_mdd_r'] = band['max_drawdown'].round(3)   # 0.1%p 단위
        best_mdd_r = band['_mdd_r'].max()                # 가장 0에 가까운(덜 빠진)
        mdd_grp = band[band['_mdd_r'] == best_mdd_r]
        ret_col = 'total_return' if 'total_return' in mdd_grp.columns else 'combined_return'
        mdd_grp = mdd_grp.sort_values([ret_col, 'avg_success_rate'], ascending=False)
        return mdd_grp.index[0]

    # ★ 독립 최적화 모드 — 매수성공률 최고 설정 + 매도성공률 최고 설정을 따로 찾아 합침.
    #   (매수성공률은 K_buy/vote_buy만으로, 매도성공률은 K_sell/vote_sell만으로 결정되므로
    #    둘은 독립. 각각 최고를 골라 한 조합으로 합치면 매수·매도 성공률 모두 최댓값 달성.)
    #   합친 (K_buy*, vote_buy*, K_sell*, vote_sell*) 행은 grid가 모든 조합을 돌렸으므로
    #   이미 df 안에 존재 → 그 행을 찾아 반환. 수익·MDD는 그 조합의 실제값을 그대로 씀.
    if prio == 'independent' and 'buy_success_rate' in df.columns \
       and 'sell_success_rate' in df.columns \
       and all(c in df.columns for c in ('K_buy','vote_buy','K_sell','vote_sell')):
        # 1) 매수성공률 최고 → 동률이면 매수 신호가 더 자주 뜨는(vote_buy/K_buy 낮은) 쪽
        best_buy_sr = df['buy_success_rate'].max()
        buy_cand = df[df['buy_success_rate'] >= best_buy_sr - 1e-9].copy()
        buy_cand['_vote_ratio_b'] = buy_cand['vote_buy'] / buy_cand['K_buy'].clip(lower=1)
        buy_cand = buy_cand.sort_values(['_vote_ratio_b'], ascending=True)
        Kb_star = int(buy_cand.iloc[0]['K_buy']); vb_star = int(buy_cand.iloc[0]['vote_buy'])
        # 2) 매도성공률 최고 → 동률이면 매도 신호가 더 자주 뜨는 쪽
        best_sell_sr = df['sell_success_rate'].max()
        sell_cand = df[df['sell_success_rate'] >= best_sell_sr - 1e-9].copy()
        sell_cand['_vote_ratio_s'] = sell_cand['vote_sell'] / sell_cand['K_sell'].clip(lower=1)
        sell_cand = sell_cand.sort_values(['_vote_ratio_s'], ascending=True)
        Ks_star = int(sell_cand.iloc[0]['K_sell']); vs_star = int(sell_cand.iloc[0]['vote_sell'])
        # 3) 합친 조합 행 찾기
        merged = df[(df['K_buy']==Kb_star) & (df['vote_buy']==vb_star) &
                    (df['K_sell']==Ks_star) & (df['vote_sell']==vs_star)]
        if len(merged) > 0:
            return merged.index[0]
        # 합친 조합이 df에 없으면(필터로 빠졌거나 그리드 누락): 매수최고 행으로 폴백
        return buy_cand.index[0]

    # ★ 안정성 종합 점수 모드 — 매도·매수성공·수익·MDD방어 가중 기하평균 최대
    if prio == 'stability' and 'sell_success_rate' in df.columns \
       and 'buy_success_rate' in df.columns and 'max_drawdown' in df.columns:
        scored = compute_stability_scores(df)
        scored = scored.sort_values(
            ['stability_score', 'sell_success_rate', secondary], ascending=False)
        return scored.index[0]

    if prio == 'sell_buy_return' \
       and 'sell_success_rate' in df.columns and 'buy_success_rate' in df.columns:
        best_sell = df['sell_success_rate'].max()
        b1 = df[df['sell_success_rate'] >= best_sell - tolerance]
        if len(b1) == 0: b1 = df[df['sell_success_rate'] == best_sell]
        best_buy = b1['buy_success_rate'].max()
        b2 = b1[b1['buy_success_rate'] >= best_buy - tolerance]
        if len(b2) == 0: b2 = b1[b1['buy_success_rate'] == best_buy]
        b2_sorted = b2.sort_values(
            [secondary, 'sell_success_rate', 'buy_success_rate'], ascending=False)
        return b2_sorted.index[0]

    best_primary = df[primary].max()
    band_min = best_primary - tolerance
    band = df[df[primary] >= band_min]
    if len(band) == 0:
        band = df[df[primary] == best_primary]
    band_sorted = band.sort_values([secondary, primary], ascending=False)
    return band_sorted.index[0]


def _select_with_anchor_priority(df, *,
                                  anchor_tolerance, balacc_tolerance,
                                  match_col='anchor_avg_match_rate',
                                  balacc_col='avg_success_rate',
                                  return_col='combined_return'):
    if len(df) == 0: return None
    if return_col not in df.columns:
        return_col = 'total_return'
    if match_col not in df.columns:
        return _select_with_tolerance(df, balacc_tolerance, balacc_col, return_col)

    best_match = df[match_col].max()
    band1 = df[df[match_col] >= best_match - anchor_tolerance]
    if len(band1) == 0:
        band1 = df[df[match_col] == best_match]

    best_balacc = band1[balacc_col].max()
    band2 = band1[band1[balacc_col] >= best_balacc - balacc_tolerance]
    if len(band2) == 0:
        band2 = band1[band1[balacc_col] == best_balacc]

    band2_sorted = band2.sort_values(
        [return_col, match_col, balacc_col], ascending=False)
    return band2_sorted.index[0]


def grid_search_ensemble(feat, close, buy_pool, sell_pool, *,
                          k_buy_range, k_sell_range,
                          vote_ratio_buy, vote_ratio_sell,
                          cost, min_trades=1,
                          horizon, dd_limit, ru_limit,
                          stop_loss_pct,
                          anchor_mode=False,
                          anchor_safe_buy=None,
                          anchor_safe_sell=None,
                          oos_start_idx=None):
    close_arr = close.values.astype(np.float64)
    n_days    = len(close_arr)
    sl_pct    = float(stop_loss_pct) if (stop_loss_pct is not None and stop_loss_pct > 0) else 0.0

    use_oos = (oos_start_idx is not None and 0 < oos_start_idx < n_days)
    if use_oos:
        is_len = oos_start_idx
        years  = is_len / 252.0
    else:
        is_len = n_days
        years  = n_days / 252.0

    buy_sigs_all  = np.zeros((n_days, len(buy_pool)),  dtype=np.uint8)
    sell_sigs_all = np.zeros((n_days, len(sell_pool)), dtype=np.uint8)
    for k, (_, row) in enumerate(buy_pool.iterrows()):
        buy_sigs_all[:, k] = _to_signal_array(feat, row)
    for k, (_, row) in enumerate(sell_pool.iterrows()):
        sell_sigs_all[:, k] = _to_signal_array(feat, row)

    # ★ 가중 투표 가중치 (성공률 비례, 평균1 정규화). 일반투표면 전부 1.0
    if globals().get('USE_WEIGHTED_VOTE', False):
        buy_w_full  = compute_vote_weights(buy_pool['score'].values,
                                           globals().get('WEIGHT_MAX_RATIO', 1.6))
        sell_w_full = compute_vote_weights(sell_pool['score'].values,
                                           globals().get('WEIGHT_MAX_RATIO', 1.6))
    else:
        buy_w_full  = np.ones(len(buy_pool),  dtype=np.float64)
        sell_w_full = np.ones(len(sell_pool), dtype=np.float64)

    safe_buy, safe_sell, evaluable = _compute_safe_arrays(
        close_arr, horizon, dd_limit, ru_limit)
    if anchor_mode and anchor_safe_buy is not None and anchor_safe_sell is not None:
        safe_buy, safe_sell, evaluable = _apply_anchor_correction(
            safe_buy, safe_sell, evaluable, anchor_safe_buy, anchor_safe_sell)

    has_anchor = (anchor_mode and anchor_safe_buy is not None and anchor_safe_sell is not None)
    if has_anchor:
        anchor_buy_pass  = np.ascontiguousarray(anchor_safe_buy.astype(np.uint8))
        anchor_sell_pass = np.ascontiguousarray(anchor_safe_sell.astype(np.uint8))
        n_anchor_buy_total  = int(anchor_buy_pass.sum())
        n_anchor_sell_total = int(anchor_sell_pass.sum())
    else:
        anchor_buy_pass  = np.zeros(0, dtype=np.uint8)
        anchor_sell_pass = np.zeros(0, dtype=np.uint8)
        n_anchor_buy_total = n_anchor_sell_total = 0

    if use_oos:
        close_is  = np.ascontiguousarray(close_arr[:is_len])
        safe_buy_is  = np.ascontiguousarray(safe_buy[:is_len])
        safe_sell_is = np.ascontiguousarray(safe_sell[:is_len])
        eval_is      = np.ascontiguousarray(evaluable[:is_len])
        if has_anchor:
            anc_buy_is  = np.ascontiguousarray(anchor_buy_pass[:is_len])
            anc_sell_is = np.ascontiguousarray(anchor_sell_pass[:is_len])
        else:
            anc_buy_is = anc_sell_is = np.zeros(0, dtype=np.uint8)
        close_oos = np.ascontiguousarray(close_arr[is_len:])
        safe_buy_oos  = np.ascontiguousarray(safe_buy[is_len:])
        safe_sell_oos = np.ascontiguousarray(safe_sell[is_len:])
        eval_oos      = np.ascontiguousarray(evaluable[is_len:])
        empty_anc     = np.zeros(0, dtype=np.uint8)
        oos_years     = max(len(close_oos) / 252.0, 1e-9)
    else:
        close_is = close_arr; safe_buy_is = safe_buy
        safe_sell_is = safe_sell; eval_is = evaluable
        anc_buy_is = anchor_buy_pass; anc_sell_is = anchor_sell_pass

    if HAS_NUMBA:
        z = np.zeros((is_len, 1), dtype=np.uint8)
        zw = np.ones(1, dtype=np.float64)
        _simulate_ensemble(close_is, z, z, 1, 1, cost,
                            safe_buy_is, safe_sell_is, eval_is, sl_pct,
                            anc_buy_is, anc_sell_is, zw, zw)

    valid_k_buy  = [k for k in k_buy_range  if k <= len(buy_pool)]
    valid_k_sell = [k for k in k_sell_range if k <= len(sell_pool)]

    rows = []
    for K_b in valid_k_buy:
        buy_mat_full = np.ascontiguousarray(buy_sigs_all[:, :K_b])
        for K_s in valid_k_sell:
            sell_mat_full = np.ascontiguousarray(sell_sigs_all[:, :K_s])
            if use_oos:
                buy_mat  = np.ascontiguousarray(buy_mat_full[:is_len])
                sell_mat = np.ascontiguousarray(sell_mat_full[:is_len])
                buy_mat_oos  = np.ascontiguousarray(buy_mat_full[is_len:])
                sell_mat_oos = np.ascontiguousarray(sell_mat_full[is_len:])
            else:
                buy_mat = buy_mat_full; sell_mat = sell_mat_full
            buy_w_k  = np.ascontiguousarray(buy_w_full[:K_b])
            sell_w_k = np.ascontiguousarray(sell_w_full[:K_s])
            v_b_set = sorted(set(max(1, int(round(K_b * r))) for r in vote_ratio_buy))
            v_s_set = sorted(set(max(1, int(round(K_s * r))) for r in vote_ratio_sell))
            for v_b in v_b_set:
                for v_s in v_s_set:
                    (tr, nt, nw, srr, sqrr, mdd,
                     n_eval, n_bc, n_sc, n_stop,
                     b_tp, b_fp, b_fn,
                     s_tp, s_fp, s_fn,
                     n_ab_tot, n_ab_mat,
                     n_as_tot, n_as_mat) = _simulate_ensemble(
                        close_is, buy_mat, sell_mat, v_b, v_s, cost,
                        safe_buy_is, safe_sell_is, eval_is, sl_pct,
                        anc_buy_is, anc_sell_is, buy_w_k, sell_w_k)
                    if nt < min_trades: continue
                    if use_oos:
                        (oos_tr, oos_nt, oos_nw, _osr, _osqr, oos_mdd,
                         _one, _obc, _osc, _ostop,
                         _b1,_b2,_b3,_s1,_s2,_s3,
                         _x1,_x2,_x3,_x4) = _simulate_ensemble(
                            close_oos, buy_mat_oos, sell_mat_oos, v_b, v_s, cost,
                            safe_buy_oos, safe_sell_oos, eval_oos, sl_pct,
                            empty_anc, empty_anc, buy_w_k, sell_w_k)
                        oos_return    = oos_tr
                        oos_n_trades  = oos_nt
                        oos_win_rate  = (oos_nw / oos_nt) if oos_nt > 0 else np.nan
                        oos_max_dd    = oos_mdd
                    else:
                        oos_return    = np.nan
                        oos_n_trades  = 0
                        oos_win_rate  = np.nan
                        oos_max_dd    = np.nan
                    wr = nw / nt
                    avg = srr / nt
                    if nt > 1:
                        var = max(sqrr / nt - avg * avg, 0.0)
                        std = np.sqrt(var)
                        tpy = nt / years
                        sharpe = (avg * tpy) / (std * np.sqrt(tpy)) if std > 0 else 0.0
                    else:
                        sharpe = 0.0
                    cagr = (1 + tr) ** (1 / years) - 1 if (years > 0 and (1 + tr) > 0) else tr
                    avg_buy_score  = float(buy_pool.iloc[:K_b]['score'].mean())
                    avg_sell_score = float(sell_pool.iloc[:K_s]['score'].mean())
                    b_tn = n_eval - b_tp - b_fp - b_fn
                    s_tn = n_eval - s_tp - s_fp - s_fn
                    b_tpr = b_tp / (b_tp + b_fn) if (b_tp + b_fn) > 0 else 0.0
                    b_tnr = b_tn / (b_tn + b_fp) if (b_tn + b_fp) > 0 else 0.0
                    s_tpr = s_tp / (s_tp + s_fn) if (s_tp + s_fn) > 0 else 0.0
                    s_tnr = s_tn / (s_tn + s_fp) if (s_tn + s_fp) > 0 else 0.0
                    buy_sr  = (b_tpr + b_tnr) / 2.0
                    sell_sr = (s_tpr + s_tnr) / 2.0
                    avg_sr  = (buy_sr + sell_sr) / 2.0
                    buy_acc  = n_bc / n_eval if n_eval > 0 else 0.0
                    sell_acc = n_sc / n_eval if n_eval > 0 else 0.0
                    if has_anchor and n_ab_tot > 0:
                        buy_match  = float(n_ab_mat) / float(n_ab_tot)
                    else:
                        buy_match = 1.0
                    if has_anchor and n_as_tot > 0:
                        sell_match = float(n_as_mat) / float(n_as_tot)
                    else:
                        sell_match = 1.0
                    avg_match = (buy_match + sell_match) / 2.0
                    if use_oos and not np.isnan(oos_return):
                        combined_return = tr + oos_return
                    else:
                        combined_return = tr
                    rows.append((K_b, K_s, v_b, v_s,
                                  tr, cagr, nt, wr, avg, sharpe, mdd,
                                  avg_buy_score, avg_sell_score,
                                  v_b / K_b, v_s / K_s,
                                  buy_sr, sell_sr, avg_sr, n_eval, n_stop,
                                  buy_acc, sell_acc,
                                  buy_match, sell_match, avg_match,
                                  oos_return, oos_n_trades,
                                  oos_win_rate, oos_max_dd, combined_return))

    return pd.DataFrame(rows, columns=[
        'K_buy', 'K_sell', 'vote_buy', 'vote_sell',
        'total_return', 'cagr', 'n_trades', 'win_rate',
        'avg_return_per_trade', 'sharpe_like', 'max_drawdown',
        'avg_buy_score', 'avg_sell_score',
        'vote_ratio_buy', 'vote_ratio_sell',
        'buy_success_rate', 'sell_success_rate', 'avg_success_rate',
        'n_eval_days', 'n_stop_triggered',
        'buy_accuracy_plain', 'sell_accuracy_plain',
        'anchor_buy_match_rate', 'anchor_sell_match_rate', 'anchor_avg_match_rate',
        'oos_return', 'oos_n_trades',
        'oos_win_rate', 'oos_max_drawdown', 'combined_return',
    ])


def apply_mdd_and_trade_filters(grid_df, mdd_limit_pct, min_trades_daily):
    df = grid_df.copy()
    df = df[df['n_trades'] >= min_trades_daily]
    if mdd_limit_pct is not None:
        df = df[df['max_drawdown'] >= -abs(mdd_limit_pct) / 100.0]
    return df


def meta_grid_search(feat, close, *,
                      meta_grid, indicators,
                      n_thresholds, horizon, dd_limit, ru_limit,
                      k_buy_range, k_sell_range,
                      vote_ratio_buy, vote_ratio_sell,
                      cost, min_trades_daily,
                      mdd_limit_pct, select_by,
                      stop_loss_pct,
                      selection_tolerance=0.0,
                      anchor_match_priority=False,
                      anchor_match_tolerance=0.05,
                      anchor_mode=False,
                      anchor_safe_buy=None,
                      anchor_safe_sell=None,
                      oos_start_idx=None,
                      bh_ret=None,
                      exclude_below_bh=True):
    score_cache = {}
    use_anchor = bool(anchor_mode and anchor_safe_buy is not None and anchor_safe_sell is not None)

    use_oos = (oos_start_idx is not None and 0 < oos_start_idx < len(close))
    if use_oos:
        feat_score  = feat.iloc[:oos_start_idx]
        close_score = close.iloc[:oos_start_idx]
        ab  = anchor_safe_buy[:oos_start_idx]  if use_anchor else None
        asu = anchor_safe_sell[:oos_start_idx] if use_anchor else None
    else:
        feat_score  = feat
        close_score = close
        ab  = anchor_safe_buy if use_anchor else None
        asu = anchor_safe_sell if use_anchor else None

    def get_scores(wilson_z, pct_range, min_signals):
        key = (wilson_z, pct_range, min_signals)
        if key in score_cache:
            return score_cache[key]
        pct_low, pct_high = pct_range
        buy_df, sell_df = evaluate_buy_sell_scores(
            feat_score, close_score, indicators=indicators,
            n_thresholds=n_thresholds,
            pct_low=pct_low, pct_high=pct_high,
            horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
            min_signals=min_signals, wilson_z=wilson_z,
            anchor_buy_arr=ab, anchor_sell_arr=asu,
        )
        score_cache[key] = (buy_df, sell_df)
        return buy_df, sell_df

    # ★ 성공률 우선 풀 (요청) — pct무관하게 1회만 계산해 모든 메타조합에서 공용.
    succ_buy_full = succ_sell_full = succ_buy_dedup = succ_sell_dedup = None
    if POOL_SELECT_BY_SUCCESS:
        print(f"  ★ 성공률 우선 풀 선출 ON — 컷오프 {POOL_SUCCESS_MIN_RATE*100:.0f}%, "
              f"최소신호 {POOL_SUCCESS_MIN_SIG}, 분위 {POOL_SUCCESS_WIDE_PCT}")
        succ_buy_full, succ_sell_full, succ_buy_dedup, succ_sell_dedup = select_pool_by_success(
            feat_score, close_score, indicators=indicators, n_thresholds=n_thresholds,
            horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit, wilson_z=1.0)
        _nb = 0 if succ_buy_dedup is None else len(succ_buy_dedup)
        _ns = 0 if succ_sell_dedup is None else len(succ_sell_dedup)
        print(f"     ▷ 성공률 {POOL_SUCCESS_MIN_RATE*100:.0f}%+ 지표: 매수 {_nb}개, 매도 {_ns}개")
        # 시트용으로 전역 저장 (write_excel에서 읽음) — 기존 _LAST_*_MAP 패턴과 동일.
        try:
            globals()['_LAST_SUCCESS_POOL'] = (
                succ_buy_dedup.copy()  if succ_buy_dedup  is not None else None,
                succ_sell_dedup.copy() if succ_sell_dedup is not None else None)
        except Exception:
            globals()['_LAST_SUCCESS_POOL'] = (succ_buy_dedup, succ_sell_dedup)

        # ★ K 하한 자동 인하 (요청) — 정예(희소) 풀은 K가 10부터면 동의 임계(K×vote)를 못 채워 거래 0.
        #   K 하한을 POOL_SUCCESS_K_FLOOR까지 낮춰 소수 동의로도 신호가 나게 함(상한은 유지).
        _kf = int(POOL_SUCCESS_K_FLOOR)
        try:
            _kbmax = max(k_buy_range)  if len(k_buy_range)  else 30
            _ksmax = max(k_sell_range) if len(k_sell_range) else 30
            if min(k_buy_range) > _kf:
                k_buy_range  = list(range(_kf, _kbmax + 1))
            if min(k_sell_range) > _kf:
                k_sell_range = list(range(_kf, _ksmax + 1))
            print(f"     ▷ 성공률 우선 풀 → K 하한 {_kf}로 인하 "
                  f"(K_buy {min(k_buy_range)}~{max(k_buy_range)}, K_sell {min(k_sell_range)}~{max(k_sell_range)})")
        except Exception as _ek:
            print(f"     ⚠ K 하한 인하 실패(무시): {_ek}")
    else:
        globals()['_LAST_SUCCESS_POOL'] = (None, None)

    combos = list(itertools.product(
        meta_grid['wilson_z'],
        meta_grid['pct_range'],
        meta_grid['min_signals'],
        meta_grid['corr_limit'],
        meta_grid['top_n_pool'],
    ))
    total = len(combos)
    print(f"  메타 그리드 총 {total}개 조합")
    _prio_p = globals().get('SELECTION_PRIORITY', 'balacc_return')
    if _prio_p == 'winrate_return':
        _wt = globals().get('WINRATE_TOLERANCE', 0.10)
        print(f"  ★ 선정: 일별거래 승률 -{_wt*100:.0f}%p 밴드 → 누적수익 최고")
    elif _prio_p == 'sell_mdd_return':
        _st = globals().get('SELL_SUCCESS_TOLERANCE', 0.02); _mt = globals().get('MDD_TOLERANCE', 0.01)
        print(f"  ★ 선정: 매도성공률 -{_st*100:.0f}%p 밴드 → MDD 최저 -{_mt*100:.0f}%p 밴드 → 누적수익 최고")
    elif _prio_p == 'avgband_mdd_return':
        print(f"  ★ 선정: 평균성공률 top band({selection_tolerance*100:.1f}%p) 내 → MDD 최저(동률시 수익 최대)")
    elif _prio_p == 'independent':
        print(f"  ★ 선정: 매수성공률 최고 설정 + 매도성공률 최고 설정 따로 찾아 합침 (독립 최적화)")
    elif _prio_p == 'stability':
        print(f"  ★ 선정: 안정성 종합점수(매도·매수성공·수익·MDD방어 가중 기하평균) 최대")
    elif _prio_p == 'sell_buy_return':
        print(f"  ★ 선정 우선순위: 매도성공률 → 매수성공률 → 누적수익 (밴드 {selection_tolerance*100:.2f}%p)")
    elif selection_tolerance > 0:
        print(f"  ★ Tolerance Band: 평균 BalAcc top - {selection_tolerance*100:.2f}%p 이내 중 수익률 최대")
    else:
        print(f"  ★ Strict: 1차 평균 BalAcc → 2차 수익률")
    use_match_priority = bool(anchor_match_priority and anchor_mode and
                               anchor_safe_buy is not None and anchor_safe_sell is not None)
    if use_match_priority:
        print(f"  ⚓ ANCHOR_MATCH_PRIORITY ON — 3단계 우선순위:")
        print(f"     1차: 매칭률 top - {anchor_match_tolerance*100:.1f}%p 이내")
        print(f"     2차: BalAcc top - {selection_tolerance*100:.2f}%p 이내")
        print(f"     3차: 수익률 최대")

    meta_rows = []
    all_passed_list = []      # ★ 모든 메타조합의 통과 그리드를 메타변수 붙여 누적 (통합 테이블)
    n_below_bh = 0            # ★ B&H 미달로 제외된 조합 수
    best_state = None
    best_overall_match = -np.inf
    best_overall_balacc = -np.inf
    best_overall_return = -np.inf
    n_no_pass = 0
    t0 = time.time()

    for ci, (wz, pr, ms, cl, tnp) in enumerate(combos):
        meta = dict(
            wilson_z=wz, pct_low=pr[0], pct_high=pr[1],
            min_signals=ms, corr_limit=cl,
            top_n_pool_buy=tnp, top_n_pool_sell=tnp,
        )
        buy_df, sell_df = get_scores(wz, pr, ms)
        # ★ buy_df/sell_df는 (성공률 우선 ON이면) evaluate_buy_sell_scores에서
        #   이미 '지표당 1행·성공률 내림차순'으로 정리됨 → diversify가 성공률 우선 풀을 만듦.
        buy_pool  = diversify_candidates(feat_score, buy_df,  top_n=tnp, corr_limit=cl)
        sell_pool = diversify_candidates(feat_score, sell_df, top_n=tnp, corr_limit=cl)

        # ★ 보정용 전체 후보 풀 저장 (요청) — top_n_pool은 그대로 100, 보정만 전체 탐색.
        #   diversify로 100개 추리기 전의 '점수 매긴 전체 후보(buy_df/sell_df)'를 메타키별 보관.
        #   보정 함수가 이 큰 풀에서 미매칭일 보정에 쓸 미사용 지표를 찾는다 (그리드 속도 영향 없음).
        try:
            _fc_key = (round(float(wz),4), int(pr[0]), int(pr[1]), round(float(cl),4))
            globals().setdefault('_LAST_FULL_CAND_MAP', {})[_fc_key] = (
                buy_df.reset_index(drop=True), sell_df.reset_index(drop=True))
        except Exception:
            pass

        inner_df = grid_search_ensemble(
            feat, close, buy_pool, sell_pool,
            k_buy_range=k_buy_range, k_sell_range=k_sell_range,
            vote_ratio_buy=vote_ratio_buy, vote_ratio_sell=vote_ratio_sell,
            cost=cost, min_trades=1,
            horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
            stop_loss_pct=stop_loss_pct,
            anchor_mode=anchor_mode,
            anchor_safe_buy=anchor_safe_buy,
            anchor_safe_sell=anchor_safe_sell,
            oos_start_idx=oos_start_idx,
        )
        passed = apply_mdd_and_trade_filters(inner_df, mdd_limit_pct, min_trades_daily)

        # ★ Buy&Hold 미달 조합 제외 (전략 수익이 B&H 이하면 버림)
        if exclude_below_bh and bh_ret is not None and len(passed) > 0:
            before_n = len(passed)
            passed = passed[passed['total_return'] > bh_ret].copy()
            n_below_bh += (before_n - len(passed))

        # ★ 이 메타조합의 통과 그리드에 메타변수 컬럼 붙여서 통합 테이블에 누적
        if len(passed) > 0:
            tagged = passed.copy()
            tagged['meta_wilson_z']   = wz
            tagged['meta_pct_low']    = pr[0]
            tagged['meta_pct_high']   = pr[1]
            tagged['meta_min_signals']= ms
            tagged['meta_corr_limit'] = cl
            tagged['meta_top_n_pool'] = tnp
            all_passed_list.append(tagged)

        if len(passed) == 0:
            n_no_pass += 1
            meta_rows.append({
                **meta,
                'n_buy_pool': len(buy_pool), 'n_sell_pool': len(sell_pool),
                'n_inner_total': len(inner_df), 'n_inner_passed': 0,
                'best_return': np.nan, 'best_mdd': np.nan, 'best_n_trades': 0,
                'best_K_buy': np.nan, 'best_K_sell': np.nan,
                'best_vote_buy': np.nan, 'best_vote_sell': np.nan,
                'best_win_rate': np.nan, 'best_sharpe': np.nan,
                'best_buy_sr': np.nan, 'best_sell_sr': np.nan, 'best_avg_sr': np.nan,
                'best_buy_match': np.nan, 'best_sell_match': np.nan, 'best_avg_match': np.nan,
            })
        else:
            if use_match_priority:
                best_idx = _select_with_anchor_priority(
                    passed,
                    anchor_tolerance=anchor_match_tolerance,
                    balacc_tolerance=selection_tolerance,
                    return_col='combined_return')
            else:
                best_idx = _select_with_tolerance(
                    passed, selection_tolerance,
                    primary='avg_success_rate', secondary='combined_return')
            best_in = passed.loc[best_idx].to_dict()
            meta_rows.append({
                **meta,
                'n_buy_pool': len(buy_pool), 'n_sell_pool': len(sell_pool),
                'n_inner_total': len(inner_df), 'n_inner_passed': len(passed),
                'best_return': best_in['total_return'],
                'best_oos_return': best_in.get('oos_return', np.nan),
                'best_combined_return': best_in.get('combined_return', best_in['total_return']),
                'best_mdd': best_in['max_drawdown'],
                'best_n_trades': int(best_in['n_trades']),
                'best_K_buy': int(best_in['K_buy']),
                'best_K_sell': int(best_in['K_sell']),
                'best_vote_buy': int(best_in['vote_buy']),
                'best_vote_sell': int(best_in['vote_sell']),
                'best_win_rate': best_in['win_rate'],
                'best_sharpe': best_in['sharpe_like'],
                'best_buy_sr':  best_in['buy_success_rate'],
                'best_sell_sr': best_in['sell_success_rate'],
                'best_avg_sr':  best_in['avg_success_rate'],
                'best_buy_match':  best_in.get('anchor_buy_match_rate',  np.nan),
                'best_sell_match': best_in.get('anchor_sell_match_rate', np.nan),
                'best_avg_match':  best_in.get('anchor_avg_match_rate',  np.nan),
            })
            this_balacc = best_in['avg_success_rate']
            this_return = best_in.get('combined_return', best_in['total_return'])
            this_match  = best_in.get('anchor_avg_match_rate', 1.0)

            update = False
            if best_state is None:
                update = True
            elif use_match_priority:
                if this_match > best_overall_match + anchor_match_tolerance:
                    update = True
                elif this_match >= best_overall_match - anchor_match_tolerance:
                    if this_balacc > best_overall_balacc + selection_tolerance:
                        update = True
                    elif this_balacc >= best_overall_balacc - selection_tolerance:
                        if this_return > best_overall_return:
                            update = True
            else:
                _prio = globals().get('SELECTION_PRIORITY', 'balacc_return')
                # ★ 일별거래 승률 밴드(-10%p) → 수익 (요청). meta 레벨도 동일.
                if _prio == 'winrate_return':
                    _wtol = globals().get('WINRATE_TOLERANCE', 0.10)
                    this_win = best_in.get('win_rate', 0.0)
                    bs = best_state[3] if best_state else None
                    bs_win = bs.get('win_rate', -np.inf) if bs else -np.inf
                    if this_win > bs_win + _wtol:
                        update = True
                    elif this_win >= bs_win - _wtol:
                        if this_return > best_overall_return:
                            update = True
                # ★ 매도성공 밴드 → MDD최저 밴드 → 수익 (요청). meta 레벨도 동일 우선순위.
                elif _prio == 'sell_mdd_return':
                    _stol = globals().get('SELL_SUCCESS_TOLERANCE', 0.02)
                    _mtol = globals().get('MDD_TOLERANCE', 0.01)
                    this_sell = best_in['sell_success_rate']
                    this_mdd  = best_in['max_drawdown']
                    bs = best_state[3] if best_state else None
                    bs_sell = bs['sell_success_rate'] if bs else -np.inf
                    bs_mdd  = bs['max_drawdown']      if bs else -np.inf
                    # 1차 매도성공률 밴드(-2%p)
                    if this_sell > bs_sell + _stol:
                        update = True
                    elif this_sell >= bs_sell - _stol:
                        # 2차 MDD 밴드(-1%p) — 덜 빠진 쪽
                        if this_mdd > bs_mdd + _mtol:
                            update = True
                        elif this_mdd >= bs_mdd - _mtol:
                            # 3차 수익
                            if this_return > best_overall_return:
                                update = True
                # ★ 평균성공 밴드 → MDD최저 → 수익 (요청). meta 레벨도 동일 우선순위.
                elif _prio == 'avgband_mdd_return':
                    this_avg = best_in['avg_success_rate']
                    this_mdd = best_in['max_drawdown']
                    bs = best_state[3] if best_state else None
                    bs_avg = bs['avg_success_rate'] if bs else -np.inf
                    bs_mdd = bs['max_drawdown']     if bs else -np.inf
                    # 1차 평균성공률 밴드
                    if this_avg > bs_avg + selection_tolerance:
                        update = True
                    elif this_avg >= bs_avg - selection_tolerance:
                        # 2차 MDD (0.1%p 단위로 비교) — 덜 빠진 쪽
                        if round(this_mdd, 3) > round(bs_mdd, 3):
                            update = True
                        elif round(this_mdd, 3) == round(bs_mdd, 3):
                            # 3차 수익
                            if this_return > best_overall_return:
                                update = True
                # ★ 독립 최적화 — meta 레벨은 best_inner의 (매수성공+매도성공) 합이
                #   가장 큰 메타 조합을 채택. 동률이면 수익으로.
                elif _prio == 'independent':
                    this_pair = best_in['buy_success_rate'] + best_in['sell_success_rate']
                    bs = best_state[3] if best_state else None
                    bs_pair = (bs['buy_success_rate'] + bs['sell_success_rate']) if bs else -np.inf
                    if this_pair > bs_pair + 1e-9:
                        update = True
                    elif this_pair >= bs_pair - 1e-9:
                        if this_return > best_overall_return:
                            update = True
                # ★ 안정성 모드 — meta 레벨은 정규화 기준이 없어, best_inner의
                #   매도성공 → 평균성공 → MDD방어 → 수익 순 밴드 비교로 stability에 준해 갱신
                elif _prio == 'stability':
                    this_sell = best_in['sell_success_rate']
                    this_avg  = best_in['avg_success_rate']
                    this_mdd  = best_in['max_drawdown']     # 음수, 클수록(0근처) 방어 좋음
                    bs = best_state[3] if best_state else None
                    bs_sell = bs['sell_success_rate'] if bs else -np.inf
                    bs_avg  = bs['avg_success_rate']  if bs else -np.inf
                    bs_mdd  = bs['max_drawdown']      if bs else -np.inf
                    if this_sell > bs_sell + selection_tolerance:
                        update = True
                    elif this_sell >= bs_sell - selection_tolerance:
                        if this_avg > bs_avg + selection_tolerance:
                            update = True
                        elif this_avg >= bs_avg - selection_tolerance:
                            # MDD 방어가 뚜렷이 좋으면 갱신, 비슷하면 수익으로
                            if this_mdd > bs_mdd + 0.01:        # 1%p 이상 덜 빠짐
                                update = True
                            elif this_mdd >= bs_mdd - 0.01:
                                if this_return > best_overall_return:
                                    update = True
                elif _prio == 'sell_buy_return':
                    this_sell = best_in['sell_success_rate']
                    this_buy  = best_in['buy_success_rate']
                    bs_sell = best_state[3]['sell_success_rate'] if best_state else -np.inf
                    bs_buy  = best_state[3]['buy_success_rate']  if best_state else -np.inf
                    if this_sell > bs_sell + selection_tolerance:
                        update = True
                    elif this_sell >= bs_sell - selection_tolerance:
                        if this_buy > bs_buy + selection_tolerance:
                            update = True
                        elif this_buy >= bs_buy - selection_tolerance:
                            if this_return > best_overall_return:
                                update = True
                else:
                    if this_balacc > best_overall_balacc + selection_tolerance:
                        update = True
                    elif this_balacc >= best_overall_balacc - selection_tolerance:
                        if this_return > best_overall_return:
                            update = True
                        elif this_return == best_overall_return and this_balacc > best_overall_balacc:
                            update = True

            if update:
                best_state = (meta.copy(), inner_df.copy(), passed.copy(),
                              best_in, buy_pool.copy(), sell_pool.copy())
                if this_match > best_overall_match:
                    best_overall_match = this_match
                if this_balacc > best_overall_balacc:
                    best_overall_balacc = this_balacc
                best_overall_return = this_return

        if (ci + 1) % max(total // 10, 1) == 0 or ci == total - 1:
            el = time.time() - t0
            est = el * total / (ci + 1)
            if best_state is not None:
                if use_match_priority:
                    cur = f"Match {best_overall_match*100:.1f}% / BalAcc {best_overall_balacc*100:.1f}% / Ret {best_overall_return*100:+.2f}%"
                else:
                    cur = f"BalAcc {best_overall_balacc*100:.1f}% / Ret {best_overall_return*100:+.2f}%"
            else:
                cur = "—"
            print(f"    ▸ {ci+1}/{total} ({(ci+1)/total*100:>3.0f}%)  "
                  f"경과 {el:>4.0f}s / 예상 {est:>4.0f}s  "
                  f"현재 best: {cur}  (통과없음 {n_no_pass}건)")

    if best_state is None:
        raise RuntimeError(
            f"MDD 한도 또는 거래수 ≥ {min_trades_daily} 인 조합을 찾지 못함.\n"
            f"MAX_DRAWDOWN_LIMIT_PCT 완화 또는 MIN_TRADES_DAILY 낮추세요."
        )

    all_meta_df = pd.DataFrame(meta_rows)
    valid = all_meta_df.dropna(subset=['best_avg_sr']).copy()
    no_pass = all_meta_df[all_meta_df['best_avg_sr'].isna()].copy()
    ret_key = 'best_combined_return' if (use_oos and 'best_combined_return' in valid.columns) else 'best_return'
    if len(valid) > 0 and selection_tolerance > 0:
        top_balacc = valid['best_avg_sr'].max()
        in_band  = valid[valid['best_avg_sr'] >= top_balacc - selection_tolerance]
        out_band = valid[valid['best_avg_sr'] <  top_balacc - selection_tolerance]
        in_sorted  = in_band.sort_values([ret_key, 'best_avg_sr'], ascending=False)
        out_sorted = out_band.sort_values(['best_avg_sr', ret_key], ascending=False)
        meta_results_df = pd.concat([in_sorted, out_sorted, no_pass]).reset_index(drop=True)
    else:
        meta_results_df = pd.concat([
            valid.sort_values(['best_avg_sr', ret_key], ascending=False),
            no_pass
        ]).reset_index(drop=True)

    best_meta, inner_all, inner_passed, best_inner, buy_pool, sell_pool = best_state

    # ★ 통합 테이블 — 모든 메타조합의 통과 그리드(메타변수 포함)
    if all_passed_list:
        combined_all = pd.concat(all_passed_list, ignore_index=True)
        # ★ 최종 선정을 '통합 테이블 전체'에서 수행 (모든 메타×그리드를 한 풀에서 비교)
        try:
            comb_idx = _select_with_tolerance(
                combined_all, selection_tolerance,
                primary='avg_success_rate', secondary='combined_return')
            comb_best = combined_all.loc[comb_idx]
            # 통합 best의 메타변수로 best_meta 갱신 (재현 일치 위해)
            best_meta = dict(
                wilson_z=float(comb_best['meta_wilson_z']),
                pct_low=int(comb_best['meta_pct_low']),
                pct_high=int(comb_best['meta_pct_high']),
                min_signals=int(comb_best['meta_min_signals']),
                corr_limit=float(comb_best['meta_corr_limit']),
                top_n_pool_buy=int(comb_best['meta_top_n_pool']),
                top_n_pool_sell=int(comb_best['meta_top_n_pool']),
            )
            best_inner = comb_best.to_dict()
            print(f"\n  ★ 통합 테이블({len(combined_all)}개)에서 최종 선정 — "
                  f"wilson_z={best_meta['wilson_z']}, pct=({best_meta['pct_low']},{best_meta['pct_high']}), "
                  f"corr={best_meta['corr_limit']}")
            if n_below_bh > 0:
                print(f"     (Buy&Hold 미달 {n_below_bh}개 조합 제외됨)")
        except Exception as _e:
            print(f"  ⚠ 통합 선정 중 경고: {_e} — 메타별 best 사용")
    else:
        combined_all = inner_passed.copy() if inner_passed is not None else pd.DataFrame()

    # 통합 테이블을 정렬해서 반환 (선정 기준 순)
    combined_sorted = _sort_combined_table(combined_all, selection_tolerance)
    return (meta_results_df, inner_all, combined_sorted, best_meta, best_inner,
            buy_pool, sell_pool)


def _sort_combined_table(df, tol):
    """통합 그리드 테이블을 현재 SELECTION_PRIORITY 기준으로 정렬."""
    if df is None or len(df) == 0:
        return df
    prio = globals().get('SELECTION_PRIORITY', 'balacc_return')
    d = df.copy()
    ret_col = 'total_return' if 'total_return' in d.columns else 'combined_return'
    try:
        if prio == 'winrate_return':
            d = d.sort_values(['win_rate', ret_col], ascending=[False, False])
        elif prio == 'sell_mdd_return':
            d = d.sort_values(['sell_success_rate', 'max_drawdown', ret_col],
                              ascending=[False, False, False])
        elif prio == 'avgband_mdd_return':
            d = d.sort_values(['avg_success_rate', 'max_drawdown', ret_col],
                              ascending=[False, False, False])
        elif prio == 'independent':
            d['_pair'] = d['buy_success_rate'] + d['sell_success_rate']
            d = d.sort_values(['_pair', ret_col], ascending=[False, False]).drop(columns=['_pair'])
        else:
            d = d.sort_values(['avg_success_rate', ret_col], ascending=[False, False])
    except Exception:
        pass
    return d.reset_index(drop=True)


def daily_ensemble_backtest(feat, close, buy_pool, sell_pool,
                              K_buy, K_sell, vote_buy, vote_sell, *,
                              cost=None,
                              horizon=None,
                              dd_limit=None,
                              ru_limit=None,
                              stop_loss_pct=None,
                              anchor_mode=False,
                              anchor_safe_buy=None,
                              anchor_safe_sell=None,
                              buy_w_override=None,
                              sell_w_override=None):
    import numpy as np
    import pandas as pd
    # 기본값 — 원본 전역 상수 사용
    if cost is None:      cost      = COST_PER_TRADE
    if horizon is None:   horizon   = HORIZON_DAYS
    if dd_limit is None:  dd_limit  = DRAWDOWN_LIMIT_BUY
    if ru_limit is None:  ru_limit  = RUNUP_LIMIT_SELL

    # ★ 거래일 필터는 run_ensemble_search 진입 시 이미 수행됨(공휴일 제외).
    #   여기서는 들어온 close를 그대로 사용. 단, 단독 호출 대비 잔여 NaN만 안전 처리.
    if isinstance(close, pd.Series) and close.isna().any():
        valid_mask = close.notna().values
        close = close[valid_mask]
        feat = feat.iloc[valid_mask]
        if anchor_safe_buy is not None and len(anchor_safe_buy) == len(valid_mask):
            anchor_safe_buy = np.asarray(anchor_safe_buy)[valid_mask]
        if anchor_safe_sell is not None and len(anchor_safe_sell) == len(valid_mask):
            anchor_safe_sell = np.asarray(anchor_safe_sell)[valid_mask]
    cl = close
    n  = len(feat); dates = feat.index
    buy_used  = buy_pool.iloc[:K_buy].reset_index(drop=True)
    sell_used = sell_pool.iloc[:K_sell].reset_index(drop=True)
    buy_mat  = np.zeros((n, K_buy),  dtype=np.uint8)
    sell_mat = np.zeros((n, K_sell), dtype=np.uint8)
    for k, (_, row) in enumerate(buy_used.iterrows()):
        buy_mat[:, k] = _to_signal_array(feat, row)
    for k, (_, row) in enumerate(sell_used.iterrows()):
        sell_mat[:, k] = _to_signal_array(feat, row)

    # ★ 변경2: 가중 투표 가중치 (성공률 비례). 일반투표면 전부 1.0
    if globals().get('USE_WEIGHTED_VOTE', False):
        buy_w  = compute_vote_weights(buy_used['score'].values,
                                      globals().get('WEIGHT_MAX_RATIO', 1.6))
        sell_w = compute_vote_weights(sell_used['score'].values,
                                      globals().get('WEIGHT_MAX_RATIO', 1.6))
        weighted = True
    else:
        buy_w  = np.ones(K_buy,  dtype=np.float64)
        sell_w = np.ones(K_sell, dtype=np.float64)
        weighted = False
    # ★ 앵커 미매칭 보정 (요청) — 외부에서 지표 가중치를 주입하면 그걸 사용.
    #   특정 지표 점수(가중치)를 조정해 충돌 패배·간발의 차 미매칭을 보정하기 위함.
    if buy_w_override is not None and len(buy_w_override) == K_buy:
        buy_w = np.asarray(buy_w_override, dtype=np.float64)
        weighted = True
    if sell_w_override is not None and len(sell_w_override) == K_sell:
        sell_w = np.asarray(sell_w_override, dtype=np.float64)
        weighted = True

    # ★ 재현용 (요청) — 실제 사용된 투표 가중치를 풀에 부착해 두면 엑셀에 저장 가능.
    #   재현 시 이 가중치를 그대로 주입하면 보정(추가지표+가중치)이 정확히 재현된다.
    try:
        buy_used  = buy_used.copy();  buy_used['vote_weight']  = np.asarray(buy_w, dtype=np.float64)
        sell_used = sell_used.copy(); sell_used['vote_weight'] = np.asarray(sell_w, dtype=np.float64)
    except Exception:
        pass

    sl_active = stop_loss_pct is not None and stop_loss_pct > 0
    sl = float(stop_loss_pct) if sl_active else 0.0

    # ★ 변경4: 향후 안전 배열 (성공/실패 판정용) — 미리 계산
    close_vals0 = cl.values.astype(np.float64)
    safe_buy0, safe_sell0, eval0 = _compute_safe_arrays(
        close_vals0, horizon, dd_limit, ru_limit)
    if anchor_mode and anchor_safe_buy is not None and anchor_safe_sell is not None:
        safe_buy0, safe_sell0, eval0 = _apply_anchor_correction(
            safe_buy0, safe_sell0, eval0, anchor_safe_buy, anchor_safe_sell)

    pos = 0; entry_price = np.nan; entry_date = pd.NaT; entry_idx = -1
    stop_price_cur = np.nan
    sum_daily = 0.0
    up_sum_daily = 0.0
    equity = 1.0; n_trades = 0; n_wins = 0
    rows = []; trades = []
    peak = 1.0; mdd = 0.0
    worst_trade_so_far = 0.0   # ★ 그 시점까지 발생한 개별 거래 최대 손실(소수, 음수)
    n_conflicts = 0
    n_conflict_buy_won = 0
    n_conflict_sell_won = 0
    n_stop_triggered = 0
    pos_post = np.zeros(n, dtype=np.int8)
    prev_close = np.nan

    # ★ 전체 매수/매도 성공·실패 집계 카운터
    #   [정답인 날만 평가] 매수: 올랐어야 하는 날 중 신호 ON으로 맞춘 비율(놓침은 실패).
    #   '안 올랐는데 신호 없음'은 평가에서 제외 — 상승장 무임승차 방지(요청).
    n_buy_on_total = 0; n_buy_success = 0; n_buy_fail = 0          # 신호 ON만 (적중률용)
    n_sell_on_total = 0; n_sell_success = 0; n_sell_fail = 0
    n_buy_eval_all = 0; n_buy_correct_all = 0                      # ON+OFF 전체 (정확도용)
    n_sell_eval_all = 0; n_sell_correct_all = 0

    for i in range(n):
        d = dates[i]; price = float(cl.iloc[i])
        if pos == 1 and entry_idx >= 0 and i > entry_idx and pd.notna(prev_close) and prev_close > 0:
            _dret = price / prev_close - 1.0
            sum_daily += _dret
            if _dret > 0.0: up_sum_daily += _dret
            equity = 1.0 + sum_daily

        # ★ 변경2: 가중 카운트
        b_count = float(np.dot(buy_mat[i].astype(np.float64), buy_w))
        s_count = float(np.dot(sell_mat[i].astype(np.float64), sell_w))
        b_raw = int(buy_mat[i].sum())     # 원시 신호 개수(공식 표시용)
        s_raw = int(sell_mat[i].sum())
        b_on = b_count >= vote_buy
        s_on = s_count >= vote_sell

        b_strength = b_count / K_buy if K_buy > 0 else 0.0
        s_strength = s_count / K_sell if K_sell > 0 else 0.0

        # ★ 변경3: 점수 계산식 문자열 (중립 등 표시용)
        if weighted:
            b_formula = f"가중Σ {b_count:.2f} (신호 {b_raw}/{K_buy} × 성공률가중) {'≥' if b_on else '<'} {vote_buy}"
            s_formula = f"가중Σ {s_count:.2f} (신호 {s_raw}/{K_sell} × 성공률가중) {'≥' if s_on else '<'} {vote_sell}"
        else:
            b_formula = f"단순합 {int(b_count)} (신호 {b_raw}/{K_buy} × 1표) {'≥' if b_on else '<'} {vote_buy}"
            s_formula = f"단순합 {int(s_count)} (신호 {s_raw}/{K_sell} × 1표) {'≥' if s_on else '<'} {vote_sell}"

        # ★ 이 날 매수/매도 신호의 성공·실패 판정 (first-touch 정답 기준)
        #   [정답인 날만 평가 — 요청 반영]
        #   매수 정답일(올랐어야): 신호 ON → ✓성공 / 신호 OFF → ✗놓침
        #   매수 정답 아닌 날(안 올랐음): 신호 ON → ✗실패 / 신호 OFF → 평가 제외(공란)
        #   매도도 대칭. '안 내렸는데 매도 안 함'을 성공으로 치지 않는다(부풀리기 방지).
        buy_result = ''; sell_result = ''
        if eval0[i] == 1:
            sb = safe_buy0[i] == 1     # 매수 정답일(올랐어야)
            ss = safe_sell0[i] == 1    # 매도 정답일(내렸어야)
            # ── 매수 ──
            if b_on:
                n_buy_on_total += 1
                if sb:
                    buy_result = '✓성공'; n_buy_success += 1
                    n_buy_correct_all += 1; n_buy_eval_all += 1
                else:
                    buy_result = '✗실패'; n_buy_fail += 1
                    n_buy_eval_all += 1
            else:
                if sb:
                    buy_result = '✗놓침'                    # 올랐어야 했는데 신호 없음
                    n_buy_eval_all += 1
                # else: 안 올랐고 신호도 없음 → 평가 제외(카운트 안 함)
            # ── 매도 ──
            if s_on:
                n_sell_on_total += 1
                if ss:
                    sell_result = '✓성공'; n_sell_success += 1
                    n_sell_correct_all += 1; n_sell_eval_all += 1
                else:
                    sell_result = '✗실패'; n_sell_fail += 1
                    n_sell_eval_all += 1
            else:
                if ss:
                    sell_result = '✗놓침'                   # 내렸어야 했는데 신호 없음
                    n_sell_eval_all += 1
                # else: 안 내렸고 신호도 없음 → 평가 제외(카운트 안 함)

        stopped_today = False
        stop_ret_pct = np.nan
        if pos == 1 and sl_active and pd.notna(entry_price):
            if (price / entry_price - 1.0) <= -sl:
                ret_raw = price / entry_price - 1.0
                ret_net = (1 + ret_raw) * (1 - cost) - 1
                sum_daily -= cost
                equity = 1.0 + sum_daily
                n_trades += 1
                n_stop_triggered += 1
                if ret_net < worst_trade_so_far: worst_trade_so_far = ret_net  # ★ 최대 거래손실 갱신
                if ret_net > 0: n_wins += 1
                trades.append({
                    'trade_no': n_trades,
                    'entry_date': entry_date, 'entry_price': entry_price,
                    'exit_date': d, 'exit_price': price,
                    'days_held': i - entry_idx,
                    'gross_return_%': ret_raw * 100,
                    'net_return_%': ret_net * 100,
                    'cum_equity': equity, 'cum_return_%': sum_daily * 100,
                    'win': ret_net > 0,
                    'exit_reason': '손절매',
                    'stop_price': float(entry_price * (1 - sl)),
                    'exit_idx': i, 'entry_idx': entry_idx,
                })
                stop_ret_pct = ret_net * 100
                stopped_today = True
                pos = 0
                entry_price = np.nan; entry_date = pd.NaT; entry_idx = -1
                stop_price_cur = np.nan

        unr_eq = 1.0 + sum_daily
        if pos == 1 and pd.notna(entry_price):
            unr_pct = (price / entry_price - 1) * 100
            days_h = i - entry_idx
        else:
            unr_pct = np.nan
            days_h = 0

        if unr_eq > peak: peak = unr_eq
        cur_dd = unr_eq / peak - 1
        if cur_dd < mdd: mdd = cur_dd

        action = ''
        do_buy = False
        do_sell = False
        conflict = b_on and s_on
        if conflict and not stopped_today: n_conflicts += 1

        # ★ 변경3: 중립(둘 다 OFF) 공식 표시
        neutral_note = ''
        if (not b_on) and (not s_on) and not stopped_today:
            neutral_note = f"중립 [매수 {b_formula} / 매도 {s_formula}]"

        if stopped_today:
            action = f'⛔ 손절매 ({stop_ret_pct:+.2f}%, 종가청산)'
        elif (i + 1) < n:
            # ★ 충돌 처리 정책 (요청): '매도 우선'
            #   - 현금 중 매수신호인데 매도신호도 같이 뜨면(충돌) → 매수 보류 (잘못된 매수 상쇄)
            #   - 보유 중 매도신호면, 매수신호가 같이 떠도(충돌) → 청산 (올바른 매도는 안 상쇄)
            _sell_priority = globals().get('CONFLICT_SELL_PRIORITY', True)
            if pos == 0 and b_on:
                if conflict and (_sell_priority or s_strength > b_strength):
                    action = f'⚔ 충돌 → 매수 보류 (매도 우선, S={s_strength:.0%}/B={b_strength:.0%})'
                    n_conflict_sell_won += 1
                else:
                    do_buy = True
                    if conflict:
                        action = f'⚔ 충돌 B={b_strength:.0%}≥S={s_strength:.0%} → 매수'
                        n_conflict_buy_won += 1
                    else:
                        action = '매수신호 → 익일 매수'
            elif pos == 1 and s_on:
                if conflict and (not _sell_priority) and b_strength > s_strength:
                    action = f'⚔ 충돌 B={b_strength:.0%}>S={s_strength:.0%} → 청산 보류'
                    n_conflict_buy_won += 1
                else:
                    do_sell = True
                    if conflict:
                        action = f'⚔ 충돌 → 청산 (매도 우선, S={s_strength:.0%}/B={b_strength:.0%})'
                        n_conflict_sell_won += 1
                    else:
                        action = '매도신호 → 익일 청산'
            elif neutral_note:
                action = neutral_note   # ★ 중립일 때 공식 표시
        else:
            # ★ 마지막 날(실행일, 요청) — 다음날이 없어 '실행'은 못 하지만, 그날의 신호 판정
            #   (충돌/매수/매도/중립)을 일별 백테스트와 '똑같은 문구'로 표시한다.
            #   do_buy/do_sell·카운터·실제매매는 건드리지 않음(통계·결과 불변).
            _sp_last = globals().get('CONFLICT_SELL_PRIORITY', True)
            if pos == 0 and b_on:
                if conflict and (_sp_last or s_strength > b_strength):
                    action = f'⚔ 충돌 → 매수 보류 (매도 우선, S={s_strength:.0%}/B={b_strength:.0%})'
                elif conflict:
                    action = f'⚔ 충돌 B={b_strength:.0%}≥S={s_strength:.0%} → 매수'
                else:
                    action = '매수신호 → 익일 매수'
            elif pos == 1 and s_on:
                if conflict and (not _sp_last) and b_strength > s_strength:
                    action = f'⚔ 충돌 B={b_strength:.0%}>S={s_strength:.0%} → 청산 보류'
                elif conflict:
                    action = f'⚔ 충돌 → 청산 (매도 우선, S={s_strength:.0%}/B={b_strength:.0%})'
                else:
                    action = '매도신호 → 익일 청산'
            elif neutral_note:
                action = neutral_note

        rows.append({
            'date': d, 'close': price,
            'buy_count': b_count, 'buy_raw': b_raw, 'buy_on': b_on,
            'sell_count': s_count, 'sell_raw': s_raw, 'sell_on': s_on,
            'buy_formula': b_formula, 'sell_formula': s_formula,
            'buy_result': buy_result, 'sell_result': sell_result,   # ★ 변경4
            'position_pre': 1 if pos == 1 else 0,
            'entry_price': entry_price if pos == 1 else np.nan,
            'stop_price':  stop_price_cur if pos == 1 else np.nan,
            'days_held': days_h,
            'unrealized_pct': unr_pct,
            'realized_pct': stop_ret_pct if stopped_today else np.nan,
            'equity': unr_eq,
            'cum_return_pct': sum_daily * 100,
            'running_mdd_pct': worst_trade_so_far * 100,   # ★ 그 시점까지 최대 거래손실
            'action': action,
        })

        if do_buy:
            entry_price = float(cl.iloc[i + 1])
            entry_date  = dates[i + 1]; entry_idx = i + 1
            sum_daily -= cost; equity = 1.0 + sum_daily; pos = 1
            stop_price_cur = entry_price * (1 - sl) if sl_active else np.nan
        elif do_sell:
            ex_p = float(cl.iloc[i + 1]); ex_d = dates[i + 1]
            ret_raw = ex_p / entry_price - 1
            ret_net = (1 + ret_raw) * (1 - cost) - 1
            if price > 0:
                _dret = ex_p / price - 1.0
                sum_daily += _dret
                if _dret > 0.0: up_sum_daily += _dret
            sum_daily -= cost
            equity = 1.0 + sum_daily
            n_trades += 1
            if ret_net > 0: n_wins += 1
            if ret_net < worst_trade_so_far: worst_trade_so_far = ret_net  # ★ 최대 거래손실 갱신
            trades.append({
                'trade_no': n_trades,
                'entry_date': entry_date, 'entry_price': entry_price,
                'exit_date': ex_d, 'exit_price': ex_p,
                'days_held': (i + 1) - entry_idx,
                'gross_return_%': ret_raw * 100,
                'net_return_%': ret_net * 100,
                'cum_equity': equity, 'cum_return_%': sum_daily * 100,
                'win': ret_net > 0,
                'exit_reason': '매도신호',
                'stop_price': float(entry_price * (1 - sl)) if sl_active else np.nan,
                'exit_idx': i + 1, 'entry_idx': entry_idx,
            })
            pos = 0; entry_price = np.nan; entry_date = pd.NaT; entry_idx = -1
            stop_price_cur = np.nan

        pos_post[i] = pos
        prev_close = price

    daily = pd.DataFrame(rows)
    trades_df = pd.DataFrame(trades)

    for tr in trades:
        if tr['exit_reason'] == '손절매': continue
        ix = tr['exit_idx']
        if 0 <= ix < len(daily):
            daily.at[ix, 'realized_pct']   = tr['net_return_%']
            daily.at[ix, 'equity']         = tr['cum_equity']
            daily.at[ix, 'cum_return_pct'] = tr['cum_return_%']
            daily.at[ix, 'unrealized_pct'] = np.nan
            daily.at[ix, 'days_held']      = tr['days_held']
            ca = daily.at[ix, 'action']
            tag = f"청산 체결 ({tr['net_return_%']:+.2f}%)"
            daily.at[ix, 'action'] = tag if not ca else f"{ca} / {tag}"
    for tr in trades:
        ix = tr['entry_idx']
        if 0 <= ix < len(daily):
            ca = daily.at[ix, 'action']
            daily.at[ix, 'action'] = '매수 체결' if not ca else f"{ca} / 매수 체결"

    # ── 기존 balanced accuracy 계산 (그대로) ──
    close_vals = cl.values.astype(np.float64)
    safe_buy_arr, safe_sell_arr, eval_arr = _compute_safe_arrays(
        close_vals, horizon, dd_limit, ru_limit)
    if anchor_mode and anchor_safe_buy is not None and anchor_safe_sell is not None:
        safe_buy_arr, safe_sell_arr, eval_arr = _apply_anchor_correction(
            safe_buy_arr, safe_sell_arr, eval_arr,
            anchor_safe_buy, anchor_safe_sell)

    n_days_total = len(close_vals)
    n_eval = 0
    n_buy_correct = 0; n_sell_correct = 0
    n_buy_signal_on = 0; n_sell_signal_on = 0
    b_tp = b_fp = b_fn = b_tn = 0
    s_tp = s_fp = s_fn = s_tn = 0
    n_anchor_buy = 0; n_anchor_buy_caught = 0
    n_anchor_sell = 0; n_anchor_sell_caught = 0
    anchor_buy_diagnosis = []
    anchor_sell_diagnosis = []
    for i in range(n_days_total):
        if eval_arr[i] == 0: continue
        b_count_i = float(np.dot(buy_mat[i].astype(np.float64), buy_w))     # ★ 가중
        s_count_i = float(np.dot(sell_mat[i].astype(np.float64), sell_w))
        b_on_i = b_count_i >= vote_buy
        s_on_i = s_count_i >= vote_sell
        sb = safe_buy_arr[i] == 1
        ss = safe_sell_arr[i] == 1
        n_eval += 1
        if b_on_i: n_buy_signal_on += 1
        if s_on_i: n_sell_signal_on += 1
        if b_on_i and sb: b_tp += 1
        elif b_on_i and (not sb): b_fp += 1
        elif (not b_on_i) and sb: b_fn += 1
        else: b_tn += 1
        if s_on_i and ss: s_tp += 1
        elif s_on_i and (not ss): s_fp += 1
        elif (not s_on_i) and ss: s_fn += 1
        else: s_tn += 1
        if (b_on_i and sb) or ((not b_on_i) and (not sb)):
            n_buy_correct += 1
        if (s_on_i and ss) or ((not s_on_i) and (not ss)):
            n_sell_correct += 1
        if anchor_mode and anchor_safe_buy is not None:
            if anchor_safe_buy[i] == 1:
                n_anchor_buy += 1
                # ★ 타이밍 반영 (요청): 신호는 정답일보다 먼저 떠야 정답일에 포지션이 맞음.
                #   신호→익일매수 구조 + 앵커 정답일이 며칠에 걸친 바닥 구간인 점을 반영해,
                #   '정답일 당일 ~ +W일' 중 하루라도 실제 보유면 매칭.
                #   (충돌·실제 매매가 모두 반영된 pos_post 기준 — 신호만 보는 게 아님)
                _w = int(globals().get('ANCHOR_MATCH_WINDOW', 1))
                caught = any(pos_post[i+dd] == 1
                             for dd in range(0, _w+1) if i+dd < n_days_total)
                if caught: n_anchor_buy_caught += 1
                target_date = dates[i+1] if i+1 < n_days_total else dates[i]
                if caught:
                    status = '✓ 잡힘 (long 포지션 보유)'
                elif b_on_i and s_on_i:
                    status = f'✗ 매수신호 떴으나 매도가 더 강해 unmatched (B={b_count_i:.2f}/{vote_buy}, S={s_count_i:.2f}/{vote_sell})'
                elif b_on_i:
                    status = f'✗ 매수신호 떴으나 거래 안 됨 — 현재 cash'
                else:
                    status = f'✗ 신호 안 뜸 (buy {b_count_i:.2f} < {vote_buy})'
                anchor_buy_diagnosis.append({
                    'target_date': target_date, 'signal_day': dates[i],
                    'buy_count': b_count_i, 'vote_buy_required': vote_buy,
                    'signal_on': b_on_i, 'executed_match': caught, 'status': status,
                })
            if anchor_safe_sell[i] == 1:
                n_anchor_sell += 1
                # ★ 매도 매칭 = '고점 근처에서 팔았나' (요청, 엄격하되 현실적):
                #   - held_before: 고점 전 최근 LB일 중 보유한 적 있나 (고점까지 들고 왔나)
                #   - cash_after : 고점 당일~+W일 중 현금이 됐나 (고점에서 청산했나)
                #   고점 직전 1일만 보면 추세 중 잠깐 빠짐에도 실패 처리돼 너무 낮음 → LB로 완화.
                _w = int(globals().get('ANCHOR_MATCH_WINDOW', 1))
                _lb = int(globals().get('ANCHOR_SELL_HOLD_LOOKBACK', 5))
                held_before = any(pos_post[i-dd] == 1
                                  for dd in range(1, _lb+1) if i-dd >= 0)
                cash_after  = any(pos_post[i+dd] == 0
                                  for dd in range(0, _w+1) if i+dd < n_days_total)
                caught = held_before and cash_after
                if caught: n_anchor_sell_caught += 1
                target_date = dates[i+1] if i+1 < n_days_total else dates[i]
                if caught:
                    status = '✓ 잡힘 (고점 근처에서 청산)'
                elif not held_before:
                    status = '✗ 고점 전에 이미 청산 (스윙 일찍 빠짐) — 큰 상승 놓침'
                elif s_on_i and b_on_i:
                    status = f'✗ 매도신호 떴으나 매수가 더 강해 unmatched (S={s_count_i:.2f}/{vote_sell}, B={b_count_i:.2f}/{vote_buy})'
                elif s_on_i:
                    status = f'✗ 매도신호 떴으나 청산 안 됨'
                else:
                    status = f'✗ 신호 안 뜸 (sell {s_count_i:.2f} < {vote_sell})'
                anchor_sell_diagnosis.append({
                    'target_date': target_date, 'signal_day': dates[i],
                    'sell_count': s_count_i, 'vote_sell_required': vote_sell,
                    'signal_on': s_on_i, 'executed_match': caught, 'status': status,
                })

    b_tpr = b_tp / (b_tp + b_fn) if (b_tp + b_fn) > 0 else 0.0
    b_tnr = b_tn / (b_tn + b_fp) if (b_tn + b_fp) > 0 else 0.0
    s_tpr = s_tp / (s_tp + s_fn) if (s_tp + s_fn) > 0 else 0.0
    s_tnr = s_tn / (s_tn + s_fp) if (s_tn + s_fp) > 0 else 0.0
    buy_succ_rate  = (b_tpr + b_tnr) / 2.0
    sell_succ_rate = (s_tpr + s_tnr) / 2.0
    avg_succ_rate  = (buy_succ_rate + sell_succ_rate) / 2.0
    buy_acc_plain  = n_buy_correct  / n_eval if n_eval > 0 else 0.0
    sell_acc_plain = n_sell_correct / n_eval if n_eval > 0 else 0.0

    last = rows[-1]
    if pos == 1:
        cur = dict(
            position='보유 중 (LONG)', position_emoji='📈',
            entry_date=entry_date, entry_price=entry_price,
            stop_price=stop_price_cur,
            current_price=last['close'], days_held=last['days_held'],
            unrealized_pct=last['unrealized_pct'],
            equity=last['equity'], cum_return_pct=last['cum_return_pct'],
            buy_count_now=last['buy_count'], sell_count_now=last['sell_count'],
            buy_on_now=last['buy_on'],       sell_on_now=last['sell_on'],
        )
    else:
        cur = dict(
            position='현금 (CASH)', position_emoji='💵',
            entry_date=None, entry_price=None, stop_price=None,
            current_price=last['close'], days_held=0, unrealized_pct=None,
            equity=last['equity'], cum_return_pct=last['cum_return_pct'],
            buy_count_now=last['buy_count'], sell_count_now=last['sell_count'],
            buy_on_now=last['buy_on'],       sell_on_now=last['sell_on'],
        )
    cur['n_trades'] = n_trades; cur['n_wins'] = n_wins
    cur['win_rate'] = n_wins / n_trades if n_trades > 0 else 0.0
    cur['n_buy_evaluated']  = n_eval
    cur['n_buy_success']    = n_buy_correct
    cur['buy_success_rate'] = buy_succ_rate
    cur['n_sell_evaluated'] = n_eval
    cur['n_sell_success']   = n_sell_correct
    cur['sell_success_rate']= sell_succ_rate
    cur['avg_success_rate'] = avg_succ_rate
    cur['n_buy_signal_on']  = n_buy_signal_on
    cur['n_sell_signal_on'] = n_sell_signal_on
    cur['n_conflicts']      = n_conflicts
    cur['n_conflict_buy_won']  = n_conflict_buy_won
    cur['n_conflict_sell_won'] = n_conflict_sell_won
    cur['n_stop_triggered']    = n_stop_triggered
    cur['stop_loss_pct']       = float(stop_loss_pct) if sl_active else None
    cur['anchor_mode']         = bool(anchor_mode)
    cur['n_anchor_buy']        = n_anchor_buy
    cur['n_anchor_buy_caught'] = n_anchor_buy_caught
    cur['n_anchor_sell']       = n_anchor_sell
    cur['n_anchor_sell_caught']= n_anchor_sell_caught
    # ★ 실제 거래 기준 앵커 매칭률 (요청) — 비율로 저장
    cur['anchor_buy_match_rate']  = (n_anchor_buy_caught / n_anchor_buy) if n_anchor_buy > 0 else 0.0
    cur['anchor_sell_match_rate'] = (n_anchor_sell_caught / n_anchor_sell) if n_anchor_sell > 0 else 0.0
    cur['anchor_avg_match_rate']  = (cur['anchor_buy_match_rate'] + cur['anchor_sell_match_rate']) / 2.0
    # ★ 앵커 전략 수익 (요청) — '단기 최저점에서 매수, 단기 최고점에서 매도'로 얻을 수 있는
    #   사후적 최대 수익. 과거 차트가 다 있으므로, 거래비용을 넘어서는 모든 상승 구간을
    #   포착하는 최적 스윙 매매로 계산한다 (B&H보다 훨씬 높아야 정상).
    #   방법: 거래비용(cost) 이상 오르는 연속 구간만 매매 → 그 구간들의 상승률 합(복리).
    _cost = float(globals().get('COST_PER_TRADE', 0.0)) if 'cost' not in dir() else float(cost)
    _ap = cl.values.astype(np.float64)
    _equity = 1.0
    _i2 = 0
    _nn = len(_ap)
    while _i2 < _nn - 1:
        # 다음 거래일부터 '오르는 한' 계속 보유 (단기 저점→고점)
        if _ap[_i2+1] > _ap[_i2]:
            j = _i2
            # j가 국소 저점: 여기서 매수, 가격이 더 안 오를 때까지(국소 고점) 보유
            k = j
            while k < _nn - 1 and _ap[k+1] >= _ap[k]:
                k += 1
            # j에서 사서 k에서 팜 (단기 저점→단기 고점)
            gain = _ap[k] / _ap[j] - 1.0
            net = (1.0 - _cost) * (1.0 + gain) * (1.0 - _cost) - 1.0  # 매수·매도 비용 차감
            if net > 0:
                _equity *= (1.0 + net)
            _i2 = k + 1
        else:
            _i2 += 1
    cur['anchor_strategy_return'] = _equity - 1.0   # 사후 최적 스윙 누적수익(복리, 소수)
    cur['anchor_buy_diagnosis']  = anchor_buy_diagnosis
    cur['anchor_sell_diagnosis'] = anchor_sell_diagnosis
    cur['buy_accuracy_plain']  = buy_acc_plain
    cur['sell_accuracy_plain'] = sell_acc_plain
    cur['horizon']  = horizon
    cur['dd_limit'] = dd_limit
    cur['ru_limit'] = ru_limit
    cur['last_date'] = last['date']
    cur['K_buy'] = K_buy; cur['K_sell'] = K_sell
    cur['vote_buy'] = vote_buy; cur['vote_sell'] = vote_sell
    # ★ 최대 낙폭 정의 변경 (요청) — 자산곡선 고점대비 하락이 아니라,
    #   '개별 일별 거래에서 발생한 최대 손실'을 최대 낙폭으로 사용.
    #   (처음 투자 시작하는 입장에선 쌓아둔 이익이 없어, 한 거래의 손실이 곧 계좌 위험)
    #   거래가 1건도 없으면 0.0.
    if len(trades) > 0:
        _trade_rets = [t['net_return_%'] / 100.0 for t in trades]   # 각 거래 실현손익(소수)
        worst_trade = min(_trade_rets)                # 가장 큰 손실 (음수)
        max_trade_loss = worst_trade if worst_trade < 0 else 0.0
    else:
        max_trade_loss = 0.0
    cur['max_drawdown']      = max_trade_loss   # ★ 이제 '최대 거래 손실' 의미
    cur['equity_mdd']        = mdd              # (참고) 자산곡선 고점대비 하락 — 보존
    cur['max_trade_loss']    = max_trade_loss   # 명시적 키
    cur['up_cum_return_pct'] = up_sum_daily * 100.0
    cur['weighted_vote'] = weighted     # ★ 변경2
    # ★ 변경4: 전체 매수/매도 성공·실패 집계
    cur['n_buy_on_total']  = n_buy_on_total
    cur['n_buy_success_cnt'] = n_buy_success
    cur['n_buy_fail_cnt']    = n_buy_fail
    cur['n_sell_on_total'] = n_sell_on_total
    cur['n_sell_success_cnt'] = n_sell_success
    cur['n_sell_fail_cnt']    = n_sell_fail
    cur['buy_signal_hit_rate']  = n_buy_success  / n_buy_on_total  if n_buy_on_total  > 0 else 0.0
    cur['sell_signal_hit_rate'] = n_sell_success / n_sell_on_total if n_sell_on_total > 0 else 0.0
    # ★ 정답일 적중률 (정답인 날만 분모. ON적중=성공, OFF=놓침(실패). 안정답인 날 제외)
    cur['n_buy_eval_all']     = n_buy_eval_all
    cur['n_buy_correct_all']  = n_buy_correct_all
    cur['n_sell_eval_all']    = n_sell_eval_all
    cur['n_sell_correct_all'] = n_sell_correct_all
    cur['buy_acc_all']  = n_buy_correct_all  / n_buy_eval_all  if n_buy_eval_all  > 0 else 0.0
    cur['sell_acc_all'] = n_sell_correct_all / n_sell_eval_all if n_sell_eval_all > 0 else 0.0
    # ★ feature3 (요청) — 일별 시트에 '사용된 모든 지표'의 그날 수치 + 신호(0/1) 부착.
    #   날짜 기준 정렬(길이 달라도 안전). cur에 (종류,지표명,수치컬럼,신호컬럼) 목록 저장.
    try:
        _used_ind_cols = []
        _fidx = feat.index
        for _k, (_, _br) in enumerate(buy_used.iterrows()):
            _nm = str(_br['indicator'])
            _vc = f'_INDV_B{_k}'; _sc = f'_INDS_B{_k}'
            daily[_vc] = daily['date'].map(pd.Series(feat[_nm].values, index=_fidx))
            daily[_sc] = daily['date'].map(pd.Series(buy_mat[:, _k], index=_fidx))
            _used_ind_cols.append(('BUY', _nm, _vc, _sc))
        for _k, (_, _sr) in enumerate(sell_used.iterrows()):
            _nm = str(_sr['indicator'])
            _vc = f'_INDV_S{_k}'; _sc = f'_INDS_S{_k}'
            daily[_vc] = daily['date'].map(pd.Series(feat[_nm].values, index=_fidx))
            daily[_sc] = daily['date'].map(pd.Series(sell_mat[:, _k], index=_fidx))
            _used_ind_cols.append(('SELL', _nm, _vc, _sc))
        cur['used_ind_cols'] = _used_ind_cols
    except Exception as _eic:
        cur['used_ind_cols'] = []
        print(f"  ⚠ 일별 지표컬럼 부착 실패(무시): {_eic}")
    return daily, trades_df, cur, buy_used, sell_used


# ════════════════════════════════════════════════════════════════
#                       Excel 출력
# ════════════════════════════════════════════════════════════════
_HDR  = PatternFill("solid", fgColor="1F3864")
_GOOD = PatternFill("solid", fgColor="C6EFCE")
_MID  = PatternFill("solid", fgColor="FFEB9C")
_BAD  = PatternFill("solid", fgColor="FFC7CE")
_ALT  = PatternFill("solid", fgColor="F2F2F2")
_HL   = PatternFill("solid", fgColor="FFF2CC")
_SIGY = PatternFill("solid", fgColor="FFFF00")   # ★ 지표 신호 충족 = 노란색 (요청)
_BUY  = PatternFill("solid", fgColor="C6EFCE")
_SELL = PatternFill("solid", fgColor="FFC7CE")
_HOLD = PatternFill("solid", fgColor="DDEBF7")
_CASH = PatternFill("solid", fgColor="F5F5F5")
_CONF = PatternFill("solid", fgColor="FFE699")
_BAND = PatternFill("solid", fgColor="FFF7E0")
_WB_  = Font(bold=True, color='FFFFFF')
_TH   = Border(left=Side('thin', color='BDBDBD'), right=Side('thin', color='BDBDBD'),
               top=Side('thin', color='BDBDBD'), bottom=Side('thin', color='BDBDBD'))


def _hdr(ws, row, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row, ci); c.value = h; c.fill = _HDR; c.font = _WB_
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = _TH


def _ret_fill(ret, bh_ret):
    if ret > bh_ret + 0.10: return _GOOD
    if ret > bh_ret:         return _MID
    if ret > 0:              return _MID
    return _BAD


def _mdd_fill(mdd, mdd_limit_pct):
    if mdd_limit_pct is None:
        if mdd > -0.05: return _GOOD
        if mdd > -0.15: return _MID
        return _BAD
    limit = -abs(mdd_limit_pct) / 100.0
    if mdd >= limit / 2: return _GOOD
    if mdd >= limit:     return _MID
    return _BAD


def _success_fill(rate):
    if rate >= 0.7: return _GOOD
    if rate >= 0.5: return _MID
    return _BAD


def _norm_date_set(dlist):
    s = set()
    if dlist is None: return s
    for d in dlist:
        try: s.add(pd.Timestamp(d).normalize())
        except Exception: pass
    return s


def _write_indicator_matrix_sheet(ws, pool, feat, close_ser,
                                   anchor_buy_set, anchor_sell_set,
                                   ticker, kind_label):
    """★ feature2 (요청) — 지표 신호 매트릭스 시트.
       행=날짜, 열=[날짜, 종가] + 풀의 각 지표.
       지표 셀: 그날 수치 + 신호 충족 시 노란색. 행: 앵커 매수일=초록 / 매도일=빨강."""
    if pool is None or len(pool) == 0 or feat is None:
        ws.cell(1, 1).value = f'{ticker} — {kind_label} 지표 신호 (풀 없음)'
        return
    dates = feat.index
    inds  = [str(r['indicator']) for _, r in pool.iterrows()]
    # 각 지표의 신호배열(0/1) + 수치
    sig_cols  = []
    feat_vals = {}
    for _, prow in pool.iterrows():
        nm = str(prow['indicator'])
        try:    sig_cols.append(np.asarray(_to_signal_array(feat, prow)).astype(np.uint8))
        except Exception: sig_cols.append(np.zeros(len(dates), dtype=np.uint8))
        feat_vals[nm] = feat[nm].values if nm in feat.columns else np.full(len(dates), np.nan)
    ws.cell(1, 1).value = (f'{ticker} — {kind_label} 지표 신호 매트릭스 '
                           f'({len(inds)}개 지표 × {len(dates)}일)  '
                           f'｜ 노랑=신호(마지막날 제외·평가불가), 행 초록=다음날↑(매수자리), 빨강=다음날↓(매도자리)')
    ws.cell(1, 1).font = Font(bold=True, size=12, color='1F3864')
    _hdr(ws, 3, ['날짜', f'{ticker}종가'] + inds)
    close_vals = (close_ser.reindex(dates).values if close_ser is not None
                  else np.full(len(dates), np.nan))
    for di in range(len(dates)):
        r = di + 4
        dt = dates[di]
        # ★ 행 색(요청): 앵커 무시 — '다음날 기준' 으로 색칠.
        #   다음날 종가가 이 날 대비 +DRAWDOWN_LIMIT_BUY 이상 상승 → 이 날이 '매수자리' → 초록
        #   다음날 종가가 -RUNUP_LIMIT_SELL 이상 하락 → 이 날이 '매도자리' → 빨강
        #   (= 그 날 신호의 정답. 매수신호가 초록행이면 다음날 올라 성공, 매도신호가 빨강행이면 성공)
        #   마지막 날은 다음날이 없어 무색.
        _ddb = float(globals().get('DRAWDOWN_LIMIT_BUY', 0.01))
        _rus = float(globals().get('RUNUP_LIMIT_SELL', 0.01))
        row_fill = None
        if di < len(dates) - 1 and pd.notna(close_vals[di]) and pd.notna(close_vals[di + 1]) and close_vals[di] > 0:
            _ret = close_vals[di + 1] / close_vals[di] - 1.0
            if _ret >= _ddb:
                row_fill = _BUY      # 다음날 +상승 → 이 날이 매수자리 → 초록
            elif _ret <= -_rus:
                row_fill = _SELL     # 다음날 -하락 → 이 날이 매도자리 → 빨강
        c = ws.cell(r, 1)
        c.value = dt.date() if hasattr(dt, 'date') else dt
        c.number_format = 'YYYY-MM-DD'; c.font = Font(size=8); c.border = _TH
        c.alignment = Alignment(horizontal='center')
        if row_fill is not None: c.fill = row_fill
        c = ws.cell(r, 2)
        if pd.notna(close_vals[di]):
            c.value = round(float(close_vals[di]), 2); c.number_format = '#,##0.00'
        c.font = Font(size=8); c.border = _TH; c.alignment = Alignment(horizontal='right')
        if row_fill is not None: c.fill = row_fill
        for k, nm in enumerate(inds):
            cc = ws.cell(r, 3 + k)
            v = feat_vals[nm][di]
            if pd.notna(v):
                cc.value = round(float(v), 4); cc.number_format = '0.0000'
            cc.font = Font(size=8); cc.border = _TH; cc.alignment = Alignment(horizontal='right')
            if sig_cols[k][di] == 1 and di < len(dates) - 1:
                cc.fill = _SIGY            # 신호 충족 → 노란색 (마지막 날은 다음날 없어 평가불가 → 제외, 앙상블과 일치)
            elif row_fill is not None:
                cc.fill = row_fill
    widths = [12, 11] + [13] * len(inds)
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'C4'; ws.row_dimensions[3].height = 70


def _write_daily_rows(ws, daily, cur, mdd_limit_pct):
    """일별 행 작성 (일별 시트 / OOS 시트 공통).
       컬럼: 날짜,종가,매수카운트,매수ON,매수성공,매도카운트,매도ON,매도성공,
            포지션,액션,진입가,손절가,보유일,미실현%,실현%,누적자산,누적수익%,진행최대손실%,
            매수공식,매도공식 (총 20열) — 액션을 포지션 다음으로 이동(요청)"""
    for ri, row in daily.iterrows():
        r = ri + 5
        c = ws.cell(r, 1); c.value = row['date'].date()
        c.number_format = 'YYYY-MM-DD'
        c.font = Font(bold=True, size=9, color='1F3864')
        c.alignment = Alignment(horizontal='center'); c.border = _TH
        c = ws.cell(r, 2); c.value = round(row['close'], 2)
        c.number_format = '#,##0.00'
        c.alignment = Alignment(horizontal='right'); c.border = _TH
        c = ws.cell(r, 3); c.value = round(float(row['buy_count']), 2)
        c.alignment = Alignment(horizontal='center'); c.border = _TH
        c.font = Font(size=10, bold=bool(row['buy_on']))
        if row['buy_on']: c.fill = _BUY
        c = ws.cell(r, 4)
        if row['buy_on']:
            c.value = 'ON'; c.fill = _BUY; c.font = Font(bold=True, size=10, color='006100')
        else:
            c.value = '-'; c.font = Font(size=9, color='888888')
        c.alignment = Alignment(horizontal='center'); c.border = _TH
        # 매수성공 (변경4)
        c = ws.cell(r, 5)
        br = str(row.get('buy_result', ''))
        c.value = br if br else '-'
        if '✓' in br:        # ✓성공(ON적중)
            c.fill = _GOOD; c.font = Font(bold=True, size=10, color='006100')
        elif '✗' in br:      # ✗실패(ON오답) 또는 ✗놓침(정답인데 신호없음)
            c.fill = _BAD;  c.font = Font(bold=True, size=10, color='C00000')
        else:
            c.font = Font(size=9, color='888888')
        c.alignment = Alignment(horizontal='center'); c.border = _TH
        c = ws.cell(r, 6); c.value = round(float(row['sell_count']), 2)
        c.alignment = Alignment(horizontal='center'); c.border = _TH
        c.font = Font(size=10, bold=bool(row['sell_on']))
        if row['sell_on']: c.fill = _SELL
        c = ws.cell(r, 7)
        if row['sell_on']:
            c.value = 'ON'; c.fill = _SELL; c.font = Font(bold=True, size=10, color='C00000')
        else:
            c.value = '-'; c.font = Font(size=9, color='888888')
        c.alignment = Alignment(horizontal='center'); c.border = _TH
        # 매도성공 (변경4)
        c = ws.cell(r, 8)
        sr = str(row.get('sell_result', ''))
        c.value = sr if sr else '-'
        if '✓' in sr:
            c.fill = _GOOD; c.font = Font(bold=True, size=10, color='006100')
        elif '✗' in sr:
            c.fill = _BAD;  c.font = Font(bold=True, size=10, color='C00000')
        else:
            c.font = Font(size=9, color='888888')
        c.alignment = Alignment(horizontal='center'); c.border = _TH
        c = ws.cell(r, 9)
        if row['position_pre'] == 1:
            c.value = '보유'; c.fill = _HOLD; c.font = Font(bold=True, size=9, color='C00000')
        else:
            c.value = '현금'; c.fill = _CASH; c.font = Font(size=9, color='606060')
        c.alignment = Alignment(horizontal='center'); c.border = _TH
        # ★ 액션 (포지션 다음으로 이동)
        c = ws.cell(r, 10)
        if row['action']:
            c.value = row['action']
            if '손절매' in row['action']:
                c.fill = _BAD; c.font = Font(bold=True, size=10, color='C00000')
            elif '충돌' in row['action']:
                c.fill = _CONF; c.font = Font(bold=True, size=9, color='806000')
            elif '중립' in row['action']:
                c.font = Font(size=8, color='888888')
            elif '매수' in row['action']:
                c.fill = _BUY; c.font = Font(bold=True, size=9, color='006100')
            elif '매도' in row['action'] or '청산' in row['action']:
                c.fill = _SELL; c.font = Font(bold=True, size=9, color='C00000')
        c.alignment = Alignment(horizontal='left'); c.border = _TH
        c = ws.cell(r, 11)
        if pd.notna(row['entry_price']):
            c.value = round(row['entry_price'], 2); c.number_format = '#,##0.00'
        c.alignment = Alignment(horizontal='right'); c.border = _TH; c.font = Font(size=9)
        c = ws.cell(r, 12)
        if pd.notna(row.get('stop_price', np.nan)):
            c.value = round(float(row['stop_price']), 2); c.number_format = '#,##0.00'
            c.font = Font(size=9, color='C00000', bold=True)
        c.alignment = Alignment(horizontal='right'); c.border = _TH
        c = ws.cell(r, 13)
        c.value = int(row['days_held']) if row['days_held'] > 0 else ''
        c.alignment = Alignment(horizontal='center'); c.border = _TH; c.font = Font(size=9)
        c = ws.cell(r, 14)
        if pd.notna(row['unrealized_pct']):
            v = float(row['unrealized_pct'])
            c.value = round(v, 2); c.number_format = '0.00"%"'
            c.font = Font(size=9, color='006100' if v > 0 else 'C00000')
        c.alignment = Alignment(horizontal='right'); c.border = _TH
        c = ws.cell(r, 15)
        if pd.notna(row['realized_pct']):
            v = float(row['realized_pct'])
            c.value = round(v, 2); c.number_format = '0.00"%"'
            c.font = Font(bold=True, size=10, color='006100' if v > 0 else 'C00000')
        c.alignment = Alignment(horizontal='right'); c.border = _TH
        c = ws.cell(r, 16); c.value = round(row['equity'], 4)
        c.number_format = '0.0000'
        c.alignment = Alignment(horizontal='right'); c.border = _TH; c.font = Font(size=9)
        c = ws.cell(r, 17)
        v = float(row['cum_return_pct'])
        c.value = round(v, 2); c.number_format = '0.00"%"'
        c.alignment = Alignment(horizontal='right'); c.border = _TH
        c.font = Font(bold=True, size=9, color='006100' if v > 0 else 'C00000')
        c = ws.cell(r, 18)
        v = float(row['running_mdd_pct'])
        c.value = round(v, 2); c.number_format = '0.00"%"'
        c.alignment = Alignment(horizontal='right'); c.border = _TH
        c.font = Font(size=9, color='C00000' if v < 0 else '888888')
        if mdd_limit_pct is not None and v < -abs(mdd_limit_pct):
            c.fill = _BAD; c.font = Font(bold=True, size=9, color='C00000')
        # ★ feature3 — 매수공식/매도공식 제거, 사용지표별 '수치 + 신호면 노란색' 컬럼으로 교체
        _uic = cur.get('used_ind_cols', []) if isinstance(cur, dict) else []
        for _j, (_kind, _nm, _vc, _sc) in enumerate(_uic):
            cc = ws.cell(r, 19 + _j)
            _v = row.get(_vc, np.nan)
            if pd.notna(_v):
                try: cc.value = round(float(_v), 4)
                except Exception: cc.value = _v
                cc.number_format = '0.0000'
            cc.alignment = Alignment(horizontal='right'); cc.border = _TH; cc.font = Font(size=8)
            if row.get(_sc, 0) == 1:        # 신호 충족 → 노란색
                cc.fill = _SIGY
    _uic = cur.get('used_ind_cols', []) if isinstance(cur, dict) else []
    widths = [12, 10, 13, 8, 9, 13, 8, 9, 8, 34, 10, 10, 8, 10, 10, 10, 12, 10]
    widths += [14] * len(_uic)
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 32
    ws.freeze_panes = 'B5'


def write_excel(meta_results_df, inner_all, inner_passed,
                best_meta, best_inner, buy_pool, sell_pool,
                daily, trades, cur, buy_used, sell_used, bh_ret, bh_cagr, *,
                ticker, output_file, horizon, dd_limit, ru_limit,
                mdd_limit_pct, min_trades_daily, stop_loss_pct,
                selection_tolerance=0.0,
                anchor_match_priority_arg=False,
                anchor_match_tolerance_arg=0.0,
                anchor_mode=False,
                anchor_buy_dates=None, anchor_sell_dates=None,
                auto_anchor=False,
                oos_enabled=False,
                oos_daily=None, oos_trades=None, oos_cur=None,
                bh_up_ret=None, anchor_prio=None,
                feat=None, close_full=None):
    wb = Workbook(); wb.remove(wb.active)

    # 1. 현재 포지션
    ws = wb.create_sheet('★ 현재 포지션', 0); ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = f'{ticker} — 앙상블 현재 포지션  ({cur["last_date"].date()})'
    ws.cell(1, 1).font = Font(bold=True, size=18, color='1F3864')
    ws.merge_cells('A1:D1')
    mdd_t = f"제한 -{abs(mdd_limit_pct):.1f}%" if mdd_limit_pct is not None else "제한 없음"
    ws.cell(2, 1).value = (f'★ 매수 Top-{cur["K_buy"]}/{cur["vote_buy"]}투표  '
                            f'/ 매도 Top-{cur["K_sell"]}/{cur["vote_sell"]}투표  '
                            f'/ MDD {mdd_t}  / 최소거래 {min_trades_daily}회')
    ws.cell(2, 1).font = Font(italic=True, size=10, color='606060')
    ws.merge_cells('A2:D2')

    pr = 4
    ws.cell(pr, 1).value = '현재 포지션'
    ws.cell(pr, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.cell(pr, 2).value = f'{cur["position_emoji"]}  {cur["position"]}'
    if '보유' in cur['position']:
        ws.cell(pr, 2).fill = _HOLD
        ws.cell(pr, 2).font = Font(bold=True, size=18, color='C00000')
    else:
        ws.cell(pr, 2).fill = _CASH
        ws.cell(pr, 2).font = Font(bold=True, size=18, color='006100')
    ws.merge_cells(start_row=pr, start_column=2, end_row=pr, end_column=4)
    ws.row_dimensions[pr].height = 30

    if '보유' in cur['position']:
        details = [('진입 날짜', cur['entry_date'].date() if pd.notna(cur['entry_date']) else '-'),
                   ('진입 가격', f"${cur['entry_price']:.2f}"),
                   ('현재 가격', f"${cur['current_price']:.2f}"),
                   ('보유 일수', f"{cur['days_held']}일"),
                   ('미실현 손익', f"{cur['unrealized_pct']:+.2f}%")]
        if stop_loss_pct is not None and stop_loss_pct > 0 and cur.get('stop_price'):
            sp = cur['stop_price']
            dist_pct = (cur['current_price'] / sp - 1) * 100
            details.append(
                (f'⛔ 손절가 (-{stop_loss_pct*100:.1f}%)',
                  f"${sp:.2f}  (현재가 대비 +{dist_pct:.2f}% 여유)")
            )
    else:
        details = [('현재 가격', f"${cur['current_price']:.2f}"),
                   ('상태', '매수 신호 대기 중')]
    details += [
        ('', ''),
        ('전략 누적 자산', f"{cur['equity']:.4f}"),
        ('전략 누적 수익률', f"{cur['cum_return_pct']:+.2f}%"),
        ('Buy & Hold', f"{bh_ret*100:+.2f}%"),
        ('📈 상승일만 합산 (전략)', f"{cur.get('up_cum_return_pct', 0):+.2f}%"),
        ('📈 상승일만 합산 (B&H)', f"{bh_up_ret*100:+.2f}%" if bh_up_ret is not None else '—'),
        ('총 거래 수', f"{cur['n_trades']}회"),
        ('승률 (수익 거래)', f"{cur['win_rate']*100:.1f}% ({cur['n_wins']}/{cur['n_trades']})"),
        ('최대 거래손실 (단일거래)', f"{cur['max_drawdown']*100:.2f}%"),
    ]
    label_buy  = f'★ 매수신호 Balanced Acc (-{dd_limit*100:.1f}% 회피'
    label_sell = f'★ 매도신호 Balanced Acc (+{ru_limit*100:.1f}% 회피'
    if anchor_mode:
        label_buy  += ' + ⚓ANCHOR 보정)'
        label_sell += ' + ⚓ANCHOR 보정)'
    else:
        label_buy  += ')'
        label_sell += ')'
    sel_label = (f'★ 평균 Balanced Acc (Top band {selection_tolerance*100:.1f}%p 내 수익최대)'
                  if selection_tolerance > 0 else '★ 평균 Balanced Acc (선정 1차)')
    details += [
        ('', ''),
        (label_buy,
          f"{cur['buy_success_rate']*100:.1f}% (plain {cur['buy_accuracy_plain']*100:.1f}%, ON {cur['n_buy_signal_on']}일)"),
        (label_sell,
          f"{cur['sell_success_rate']*100:.1f}% (plain {cur['sell_accuracy_plain']*100:.1f}%, ON {cur['n_sell_signal_on']}일)"),
        (sel_label,
          f"{cur['avg_success_rate']*100:.1f}%"),
    ]
    # ★ 변경4: 전체 매수/매도 성공·실패 집계 + 투표 방식
    details += [
        ('', ''),
        ('★ 매수신호 성공/실패 (ON만)',
          f"{cur.get('n_buy_success_cnt',0)} 성공 / {cur.get('n_buy_fail_cnt',0)} 실패 "
          f"(신호 {cur.get('n_buy_on_total',0)}회, 적중 {cur.get('buy_signal_hit_rate',0)*100:.1f}%)"),
        ('★ 매수 정답일 적중률',
          f"{cur.get('buy_acc_all',0)*100:.1f}% "
          f"({cur.get('n_buy_correct_all',0)}/{cur.get('n_buy_eval_all',0)}일 — 올랐어야 한 날 중 적중, 놓침=실패)"),
        ('★ 매도신호 성공/실패 (ON만)',
          f"{cur.get('n_sell_success_cnt',0)} 성공 / {cur.get('n_sell_fail_cnt',0)} 실패 "
          f"(신호 {cur.get('n_sell_on_total',0)}회, 적중 {cur.get('sell_signal_hit_rate',0)*100:.1f}%)"),
        ('★ 매도 정답일 적중률',
          f"{cur.get('sell_acc_all',0)*100:.1f}% "
          f"({cur.get('n_sell_correct_all',0)}/{cur.get('n_sell_eval_all',0)}일 — 내렸어야 한 날 중 적중, 놓침=실패)"),
        ('★ 성공 판정 방식',
          'First-Touch (익일진입, 기간내 +목표/-손절 먼저 닿는 쪽)'),
        ('★ 투표 방식',
          '가중 투표 (성공률 비례)' if cur.get('weighted_vote') else '일반 투표 (모두 1표)'),
    ]
    if anchor_mode:
        details += [
            (f'⚓ 매수 정답일 매칭',
              f"{cur['n_anchor_buy_caught']}/{cur['n_anchor_buy']} "
              f"({(cur['n_anchor_buy_caught']/cur['n_anchor_buy']*100) if cur['n_anchor_buy']>0 else 0:.1f}%)"),
            (f'⚓ 매도 정답일 매칭',
              f"{cur['n_anchor_sell_caught']}/{cur['n_anchor_sell']} "
              f"({(cur['n_anchor_sell_caught']/cur['n_anchor_sell']*100) if cur['n_anchor_sell']>0 else 0:.1f}%)"),
        ]
    details += [
        (f'★ 신호 충돌 발생',
          f"{cur['n_conflicts']}일 (매수우세 {cur['n_conflict_buy_won']} / 매도우세 {cur['n_conflict_sell_won']})"),
        (f'⛔ 손절매 발동',
          f"{cur['n_stop_triggered']}회"
          + (f" (한도 -{stop_loss_pct*100:.1f}%)" if stop_loss_pct else " (손절매 비활성)")),
        ('', ''),
        ('★ MDD 한도', f"-{abs(mdd_limit_pct):.2f}%" if mdd_limit_pct is not None else "없음"),
        ('★ 최소 거래수', f"{min_trades_daily}회"),
        ('★ 손절매 한도', f"-{stop_loss_pct*100:.2f}%" if (stop_loss_pct is not None and stop_loss_pct > 0) else "없음"),
        ('★ Tolerance Band', f"{selection_tolerance*100:.2f}%p" if selection_tolerance > 0 else "OFF (strict)"),
        ('★ ANCHOR 매칭 우선', f"ON ({anchor_match_tolerance_arg*100:.1f}%p)" if anchor_match_priority_arg and anchor_mode else "OFF"),
        ('★ ANCHOR 보정',
          ('ON ⚓ AUTO (자동 계산)' if auto_anchor else 'ON ⚓ (수동 입력)') if anchor_mode else 'OFF'),
    ]
    for i, (k, v) in enumerate(details):
        r = 6 + i
        ws.cell(r, 1).value = k
        ws.cell(r, 1).font = Font(bold=True, size=11, color='1F3864')
        ws.cell(r, 2).value = v
        ws.cell(r, 2).font = Font(size=11)
        if k.startswith('★') or k.startswith('⛔') or k.startswith('⚓'):
            ws.cell(r, 1).fill = _HL; ws.cell(r, 2).fill = _HL
            ws.cell(r, 2).font = Font(bold=True, size=12, color='C00000')
            if '매수신호' in k and 'Balanced' in k:
                ws.cell(r, 2).fill = _success_fill(cur['buy_success_rate'])
            elif '매도신호' in k and 'Balanced' in k:
                ws.cell(r, 2).fill = _success_fill(cur['sell_success_rate'])
            elif '평균' in k and 'Balanced' in k:
                ws.cell(r, 2).fill = _success_fill(cur['avg_success_rate'])
            elif k.startswith('⛔'):
                ws.cell(r, 2).font = Font(bold=True, size=13, color='C00000')

    sr = 6 + len(details) + 2
    ws.cell(sr, 1).value = '⚡ 오늘의 앙상블 투표'
    ws.cell(sr, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=4)
    _hdr(ws, sr + 2, ['신호', 'ON 개수', '필요(투표)', '결과', ''])

    r = sr + 3
    ws.cell(r, 1).value = '🟢 매수'; ws.cell(r, 1).font = Font(size=11, bold=True)
    ws.cell(r, 2).value = f"{cur['buy_count_now']:.2f}/{cur['K_buy']}"
    ws.cell(r, 3).value = f"{cur['vote_buy']}개 이상"
    ws.cell(r, 4).value = 'ON ✓' if cur['buy_on_now'] else 'OFF'
    if cur['buy_on_now']:
        ws.cell(r, 4).fill = _BUY
        ws.cell(r, 4).font = Font(size=12, bold=True, color='006100')
    for ci in range(1, 5):
        ws.cell(r, ci).border = _TH
        ws.cell(r, ci).alignment = Alignment(horizontal='center')

    r += 1
    ws.cell(r, 1).value = '🔴 매도'; ws.cell(r, 1).font = Font(size=11, bold=True)
    ws.cell(r, 2).value = f"{cur['sell_count_now']:.2f}/{cur['K_sell']}"
    ws.cell(r, 3).value = f"{cur['vote_sell']}개 이상"
    ws.cell(r, 4).value = 'ON ✓' if cur['sell_on_now'] else 'OFF'
    if cur['sell_on_now']:
        ws.cell(r, 4).fill = _SELL
        ws.cell(r, 4).font = Font(size=12, bold=True, color='C00000')
    for ci in range(1, 5):
        ws.cell(r, ci).border = _TH
        ws.cell(r, ci).alignment = Alignment(horizontal='center')

    b_str_now = cur['buy_count_now'] / cur['K_buy'] if cur['K_buy'] > 0 else 0
    s_str_now = cur['sell_count_now'] / cur['K_sell'] if cur['K_sell'] > 0 else 0
    r += 1
    ws.cell(r, 1).value = '신호 강도'
    ws.cell(r, 1).font = Font(size=10, italic=True, color='606060')
    ws.cell(r, 2).value = f"B={b_str_now:.0%} / S={s_str_now:.0%}"
    ws.cell(r, 3).value = '둘다 ON 시 강한 쪽 우선'
    ws.cell(r, 3).font = Font(size=9, italic=True, color='606060')
    if cur['buy_on_now'] and cur['sell_on_now']:
        ws.cell(r, 4).value = '⚔ 충돌'
        ws.cell(r, 4).fill = _CONF
        ws.cell(r, 4).font = Font(size=11, bold=True, color='806000')
    for ci in range(1, 5):
        ws.cell(r, ci).border = _TH
        ws.cell(r, ci).alignment = Alignment(horizontal='center')

    ar = r + 2
    ws.cell(ar, 1).value = '➡ 다음 액션'
    ws.cell(ar, 1).font = Font(bold=True, size=13, color='1F3864')
    if '보유' in cur['position']:
        if cur['sell_on_now']:
            if cur['buy_on_now'] and b_str_now > s_str_now:
                msg, color = ('⚔ 충돌→매수 우세 → 청산 보류 (보유 유지)', '806000')
            else:
                msg, color = ('🔴 매도 ON → 다음 거래일 청산', 'C00000')
        else:
            msg, color = ('📈 보유 유지', '0070C0')
    else:
        if cur['buy_on_now']:
            if cur['sell_on_now'] and s_str_now > b_str_now:
                msg, color = ('⚔ 충돌→매도 우세 → 매수 보류 (현금 유지)', '806000')
            else:
                msg, color = ('🟢 매수 ON → 다음 거래일 진입', '006100')
        else:
            msg, color = ('💵 현금 유지', '606060')
    ws.cell(ar, 2).value = msg
    ws.cell(ar, 2).font = Font(bold=True, size=12, color=color)
    ws.cell(ar, 2).fill = _HL
    ws.merge_cells(start_row=ar, start_column=2, end_row=ar, end_column=4)
    ws.row_dimensions[ar].height = 25

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 18

    # 2. 사용된 설정
    ws = wb.create_sheet('사용된 설정'); ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = f'{ticker} — 메타 + 앙상블 자동 선정'
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells('A1:B1')
    if globals().get('SELECTION_PRIORITY', 'balacc_return') == 'sell_buy_return':
        sel_method = f'매도성공률 → 매수성공률 → 누적수익 (밴드 {selection_tolerance*100:.1f}%p)'
    else:
        sel_method = (f'1차 평균 BalAcc top band {selection_tolerance*100:.1f}%p 내 → 2차 수익률 최대'
                       if selection_tolerance > 0
                       else '1차 평균 BalAcc → 2차 수익률 (strict)')
    anchor_method = 'OFF'
    if anchor_mode:
        anchor_method = (f"ON ⚓ AUTO 자동계산 (매수 {len(anchor_buy_dates or [])}일, 매도 {len(anchor_sell_dates or [])}일)"
                         if auto_anchor else
                         f"ON ⚓ 수동입력 (매수 {len(anchor_buy_dates or [])}일, 매도 {len(anchor_sell_dates or [])}일)")
    def _bi_get(k, default=None):
        """best_inner에서 키를 안전하게 추출 (재현 모드는 일부 키가 없을 수 있음)."""
        try:
            if hasattr(best_inner, 'get'):
                v = best_inner.get(k, default)
            else:
                v = best_inner[k] if k in best_inner else default
        except Exception:
            return default
        try:
            if v is None or (pd.notna(v) is False):
                return default
        except Exception:
            pass
        return v

    info = [
        ('티커', ticker),
        ('백테스트 기간', f"{daily['date'].iloc[0].date()} ~ {daily['date'].iloc[-1].date()}"),
        ('총 거래일 수', f"{len(daily)}일"),
        ('거래 비용 (왕복)', f"{COST_PER_TRADE*100:.2f}%"),
        ('', ''),
        ('★ 평가 기준 (고정)', ''),
        ('  Horizon', f"{horizon}일"),
        ('  매수 손실 한도', f"-{dd_limit*100:.1f}% 회피 확률"),
        ('  매도 상승 한도', f"+{ru_limit*100:.1f}% 회피 확률"),
        ('', ''),
        ('★ 자동 선정 — 메타 변수', ''),
        ('  WILSON_Z', f"{best_meta['wilson_z']}"),
        ('  PCT 분위 범위', f"({best_meta['pct_low']:.0f}, {best_meta['pct_high']:.0f})"),
        ('  MIN_SIGNALS', f"{best_meta['min_signals']}"),
        ('  DIVERSITY_CORR_LIMIT', f"{best_meta['corr_limit']}"),
        ('  TOP_N_POOL_BUY', f"{best_meta['top_n_pool_buy']}"),
        ('  TOP_N_POOL_SELL', f"{best_meta['top_n_pool_sell']}"),
        ('', ''),
        ('★ 자동 선정 — 앙상블 구성', ''),
        ('  K_buy', f"{int(best_inner['K_buy'])}개"),
        ('  vote_buy', f"{int(best_inner['vote_buy'])}개 이상 ON"),
        ('  K_sell', f"{int(best_inner['K_sell'])}개"),
        ('  vote_sell', f"{int(best_inner['vote_sell'])}개 이상 ON"),
        ('  투표 방식', '가중 투표 (성공률 비례)' if cur.get('weighted_vote') else '일반 투표 (모두 1표)'),
        ('', ''),
        ('★ 충돌 해결 규칙', '둘 다 ON 시 count/K 비율 큰 쪽 우선'),
        ('', ''),
        ('★ 필터 제약', ''),
        ('  MDD 한도', f"-{abs(mdd_limit_pct):.2f}%" if mdd_limit_pct is not None else "없음"),
        ('  최소 거래수', f"{min_trades_daily}회"),
        ('  손절매 한도', f"-{stop_loss_pct*100:.2f}%" if (stop_loss_pct is not None and stop_loss_pct > 0) else "없음"),
        ('', ''),
        ('── 결과 ──', ''),
        ('  B&H 누적', f"{bh_ret*100:+.2f}%"),
        ('  B&H CAGR', f"{bh_cagr*100:+.2f}%"),
        ('  전략 누적', f"{cur['cum_return_pct']:+.2f}%"),
        ('  vs B&H', f"{cur['cum_return_pct'] - bh_ret*100:+.2f}%p"),
        ('  📈 상승일만 합산 (전략)', f"{cur.get('up_cum_return_pct', 0):+.2f}%"),
        ('  📈 상승일만 합산 (B&H)', f"{bh_up_ret*100:+.2f}%" if bh_up_ret is not None else '—'),
        ('  거래 / 승률', f"{cur['n_trades']}회 / {cur['win_rate']*100:.1f}%"),
        ('  손절매 발동', f"{cur['n_stop_triggered']}회"),
        ('  Sharpe-like', (f"{_bi_get('sharpe_like'):.2f}"
                            if _bi_get('sharpe_like') is not None else '—')),
        ('  최대 거래손실', f"{cur['max_drawdown']*100:.2f}%"),
        ('', ''),
        ('★ 선정 우선순위', sel_method),
        ('★ ANCHOR 보정', anchor_method),
        ('', ''),
        (f'  ★ 매수신호 성공/실패 (전체)',
          f"{cur.get('n_buy_success_cnt',0)} / {cur.get('n_buy_fail_cnt',0)} (신호 {cur.get('n_buy_on_total',0)}회)"),
        (f'  ★ 매도신호 성공/실패 (전체)',
          f"{cur.get('n_sell_success_cnt',0)} / {cur.get('n_sell_fail_cnt',0)} (신호 {cur.get('n_sell_on_total',0)}회)"),
        (f'  ★ 매수신호 Balanced Acc',
          f"{cur['buy_success_rate']*100:.1f}% (plain {cur['buy_accuracy_plain']*100:.1f}%, ON {cur['n_buy_signal_on']}일)"),
        (f'  ★ 매도신호 Balanced Acc',
          f"{cur['sell_success_rate']*100:.1f}% (plain {cur['sell_accuracy_plain']*100:.1f}%, ON {cur['n_sell_signal_on']}일)"),
        (f'  ★ 평균 Balanced Acc',
          f"{cur['avg_success_rate']*100:.1f}%"),
    ]
    if anchor_mode:
        info += [
            (f'  ⚓ 매수 정답일 매칭',
              f"{cur['n_anchor_buy_caught']}/{cur['n_anchor_buy']}"),
            (f'  ⚓ 매도 정답일 매칭',
              f"{cur['n_anchor_sell_caught']}/{cur['n_anchor_sell']}"),
        ]
    info += [
        (f'  ★ 충돌 발생일',
          f"{cur['n_conflicts']}일 (B승 {cur['n_conflict_buy_won']} / S승 {cur['n_conflict_sell_won']})"),
    ]
    for ri, (k, v) in enumerate(info, 3):
        ws.cell(ri, 1).value = k
        ws.cell(ri, 1).font = Font(bold=True, size=11, color='1F3864')
        ws.cell(ri, 2).value = v
        ws.cell(ri, 2).font = Font(size=11)
        if k.startswith('★') or '★' in k or k.startswith('⚓') or '⚓' in k:
            ws.cell(ri, 1).fill = _HL; ws.cell(ri, 2).fill = _HL
            ws.cell(ri, 2).font = Font(bold=True, size=12, color='C00000')
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 65

    # 3. 메타 그리드 결과
    ws = wb.create_sheet('메타_그리드_결과'); ws.sheet_view.showGridLines = False
    sort_label = sel_method
    ws.cell(1, 1).value = f'메타 변수 조합별 결과 ({len(meta_results_df)}개) — {sort_label}'
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells('A1:Q1')
    _hdr(ws, 3, ['#', 'wilson_z', 'pct', 'min_sig', 'corr', 'pool',
                 '풀 매수', '풀 매도', '통과/전체',
                 '평균성공', '매수성공', '매도성공',
                 '★ 수익', 'MDD', '거래', 'K_b/v_b', 'K_s/v_s'])
    for ri, row in meta_results_df.head(80).iterrows():
        r = ri + 4
        is_best = (pd.notna(row['best_K_buy']) and
                   row['wilson_z'] == best_meta['wilson_z'] and
                   row['pct_low'] == best_meta['pct_low'] and
                   row['min_signals'] == best_meta['min_signals'] and
                   row['corr_limit'] == best_meta['corr_limit'] and
                   row['top_n_pool_buy'] == best_meta['top_n_pool_buy'] and
                   int(row['best_K_buy']) == int(best_inner['K_buy']))
        marker = '★' if is_best else ''
        vals = [
            f"{ri+1}{marker}",
            row['wilson_z'],
            f"({int(row['pct_low'])},{int(row['pct_high'])})",
            int(row['min_signals']),
            row['corr_limit'],
            int(row['top_n_pool_buy']),
            int(row['n_buy_pool']),
            int(row['n_sell_pool']),
            f"{int(row['n_inner_passed'])}/{int(row['n_inner_total'])}",
            f"{row['best_avg_sr']*100:.1f}%"  if pd.notna(row['best_avg_sr'])  else '—',
            f"{row['best_buy_sr']*100:.1f}%"  if pd.notna(row['best_buy_sr'])  else '—',
            f"{row['best_sell_sr']*100:.1f}%" if pd.notna(row['best_sell_sr']) else '—',
            f"{row['best_return']*100:+.2f}%" if pd.notna(row['best_return']) else '—',
            f"{row['best_mdd']*100:.2f}%" if pd.notna(row['best_mdd']) else '—',
            int(row['best_n_trades']) if pd.notna(row['best_mdd']) else '—',
            f"{int(row['best_K_buy'])}/{int(row['best_vote_buy'])}" if pd.notna(row['best_K_buy']) else '—',
            f"{int(row['best_K_sell'])}/{int(row['best_vote_sell'])}" if pd.notna(row['best_K_sell']) else '—',
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = _TH
            c.alignment = Alignment(horizontal='center'); c.font = Font(size=10)
            if is_best:
                c.fill = _HL
                c.font = Font(bold=True, color='C00000', size=11)
            elif ri % 2 == 1:
                c.fill = _ALT
        if pd.notna(row['best_avg_sr']):
            ws.cell(r, 10).fill = _success_fill(row['best_avg_sr'])
            ws.cell(r, 11).fill = _success_fill(row['best_buy_sr'])
            ws.cell(r, 12).fill = _success_fill(row['best_sell_sr'])
            if is_best:
                for cc in (10, 11, 12):
                    ws.cell(r, cc).font = Font(bold=True, color='C00000', size=11)
        if pd.notna(row['best_return']):
            ws.cell(r, 13).fill = _ret_fill(row['best_return'], bh_ret)
            if is_best:
                ws.cell(r, 13).font = Font(bold=True, color='C00000', size=12)
    for ci, w in enumerate([6, 10, 10, 9, 8, 8, 9, 9, 11, 11, 10, 10, 11, 10, 8, 11, 11], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'

    # 4. 매수 앙상블 지표
    ws = wb.create_sheet('매수_앙상블_지표'); ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = f'매수 앙상블 {len(buy_used)}개 지표'
    ws.cell(1, 1).font = Font(bold=True, size=13, color='1F3864')
    ws.merge_cells('A1:J1')
    _hdr(ws, 3, ['#', '지표', '방향', '임계치', '분위', '신호수', '성공수', '성공률', '점수', '🔁가중치'])
    for ri, row in buy_used.reset_index(drop=True).iterrows():
        r = ri + 4
        vals = [ri + 1, row['indicator'], row['direction'],
                round(row['threshold'], 6), f"{row['pct_label']:.0f}%",
                int(row['n_signals']), int(row['n_success']),
                f"{row['success_rate']*100:.2f}%", round(row['score'], 4),
                round(float(row['vote_weight']), 6) if 'vote_weight' in row.index and pd.notna(row.get('vote_weight')) else '—']
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = _TH
            c.alignment = Alignment(horizontal='center'); c.font = Font(size=10)
            if ri % 2 == 1: c.fill = _ALT
    for ci, w in enumerate([5, 32, 8, 14, 8, 10, 10, 10, 10, 11], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'

    # 5. 매도 앙상블 지표
    ws = wb.create_sheet('매도_앙상블_지표'); ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = f'매도 앙상블 {len(sell_used)}개 지표'
    ws.cell(1, 1).font = Font(bold=True, size=13, color='1F3864')
    ws.merge_cells('A1:J1')
    _hdr(ws, 3, ['#', '지표', '방향', '임계치', '분위', '신호수', '성공수', '성공률', '점수', '🔁가중치'])
    for ri, row in sell_used.reset_index(drop=True).iterrows():
        r = ri + 4
        vals = [ri + 1, row['indicator'], row['direction'],
                round(row['threshold'], 6), f"{row['pct_label']:.0f}%",
                int(row['n_signals']), int(row['n_success']),
                f"{row['success_rate']*100:.2f}%", round(row['score'], 4),
                round(float(row['vote_weight']), 6) if 'vote_weight' in row.index and pd.notna(row.get('vote_weight')) else '—']
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = _TH
            c.alignment = Alignment(horizontal='center'); c.font = Font(size=10)
            if ri % 2 == 1: c.fill = _ALT
    for ci, w in enumerate([5, 32, 8, 14, 8, 10, 10, 10, 10, 11], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'

    # 6. 내부 그리드 통과
    ws = wb.create_sheet('내부_그리드_통과'); ws.sheet_view.showGridLines = False
    mdd_t2 = f"MDD ≥ -{abs(mdd_limit_pct):.1f}%" if mdd_limit_pct is not None else "MDD 제한 없음"
    use_match_p = bool(anchor_match_priority_arg and anchor_mode and 'anchor_avg_match_rate' in inner_passed.columns)
    inner_sort_label = (f'매칭률 ±{anchor_match_tolerance_arg*100:.1f}%p → BalAcc → 수익률'
                         if use_match_p else sort_label)
    ws.cell(1, 1).value = (f'전체 그리드 통합 (모든 메타변수 조합) — {mdd_t2}, 거래수 ≥ {min_trades_daily}회, B&H 초과만  '
                           f'({len(inner_passed)}개) — {inner_sort_label}')
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells('A1:AV1')
    _hdr(ws, 3, ['#', 'wilson_z', 'pct_low', 'pct_high', 'corr_limit', 'min_sig', 'pool',
                 'K_buy', 'vote_buy', 'K_sell', 'vote_sell',
                 '✅실제승률', '✅실제최대손실', '✅실제누적수익',
                 '✅실제매수성공', '✅실제매도성공', '✅실제평균성공',
                 '✅실제매수매칭', '✅실제매도매칭', '✅실제평균매칭', '✅앵커수익',
                 '🔧보정후수익', '🔧보정후최대손실', '🔧보정후승률',
                 '🔧보정후매수성공', '🔧보정후매도성공',
                 '🔧보정후매수매칭', '🔧보정후매도매칭',
                 '🔧보정채택', '🔧추가지표수',
                 '📍실행일포지션', '🎬실행일액션', '📍매수점수', '📍매도점수',
                 '📅최근매수일', '💲진입가', '📅최근매도일', '💲매도가',
                 '🔬OOS승률', '🔬OOS최대손실', '🔬OOS수익률'])
    # 통합 테이블은 meta_grid_search에서 이미 선정기준대로 정렬돼 옴 → 그대로 표시
    disp = inner_passed.head(TOP_N_GRID_OUT).reset_index(drop=True)
    n_in_band = 0
    def _g(row, k, default=np.nan):
        try: return row[k]
        except Exception: return default
    for ri, row in disp.iterrows():
        r = ri + 4
        # ★ K/vote + 메타변수(wilson_z·corr·pct)까지 모두 일치해야 best (같은 K/vote라도
        #   메타변수 다르면 풀이 달라 수익이 다름 — 메타변수 무시하면 엉뚱한 행에 ★가 찍힘)
        def _bi(k):
            try:
                v = best_inner[k]
                return v
            except Exception:
                try: return best_inner.get(k)
                except Exception: return None
        _bi_wz = _bi('meta_wilson_z'); _bi_cl = _bi('meta_corr_limit')
        _bi_pl = _bi('meta_pct_low');  _bi_ph = _bi('meta_pct_high')
        _row_wz = _g(row,'meta_wilson_z'); _row_cl = _g(row,'meta_corr_limit')
        _row_pl = _g(row,'meta_pct_low'); _row_ph = _g(row,'meta_pct_high')
        _meta_match = True
        if _bi_wz is not None and pd.notna(_row_wz):
            _meta_match = (abs(float(_row_wz)-float(_bi_wz))<1e-9 and
                           (_bi_cl is None or abs(float(_row_cl)-float(_bi_cl))<1e-9) and
                           (_bi_pl is None or int(_row_pl)==int(_bi_pl)) and
                           (_bi_ph is None or int(_row_ph)==int(_bi_ph)))
        is_best = (int(row['K_buy']) == int(best_inner['K_buy']) and
                   int(row['K_sell']) == int(best_inner['K_sell']) and
                   int(row['vote_buy']) == int(best_inner['vote_buy']) and
                   int(row['vote_sell']) == int(best_inner['vote_sell']) and
                   _meta_match)
        marker = '★' if is_best else ''
        wz_v   = _g(row, 'meta_wilson_z')
        pl_v   = _g(row, 'meta_pct_low')
        ph_v   = _g(row, 'meta_pct_high')
        cl_v   = _g(row, 'meta_corr_limit')
        ms_v   = _g(row, 'meta_min_signals')
        pool_v = _g(row, 'meta_top_n_pool')
        bm_v   = _g(row, 'anchor_buy_match_rate')
        sm_v   = _g(row, 'anchor_sell_match_rate')
        vals = [
            f"{ri+1}{marker}",
            f"{wz_v:.2f}"  if pd.notna(wz_v)  else '—',
            int(pl_v)      if pd.notna(pl_v)  else '—',
            int(ph_v)      if pd.notna(ph_v)  else '—',
            f"{cl_v:.2f}"  if pd.notna(cl_v)  else '—',
            int(ms_v)      if pd.notna(ms_v)  else '—',
            int(pool_v)    if pd.notna(pool_v)else '—',
            int(row['K_buy']), int(row['vote_buy']),
            int(row['K_sell']), int(row['vote_sell']),
            (f"{_g(row,'real_win_rate')*100:.1f}%"     if pd.notna(_g(row,'real_win_rate'))     else '—'),
            (f"{_g(row,'real_max_drawdown')*100:.2f}%" if pd.notna(_g(row,'real_max_drawdown')) else '—'),
            (f"{_g(row,'real_total_return')*100:+.2f}%" if pd.notna(_g(row,'real_total_return')) else '—'),
            (f"{_g(row,'real_buy_success')*100:.1f}%"  if pd.notna(_g(row,'real_buy_success'))  else '—'),
            (f"{_g(row,'real_sell_success')*100:.1f}%" if pd.notna(_g(row,'real_sell_success')) else '—'),
            (f"{_g(row,'real_avg_success')*100:.1f}%"  if pd.notna(_g(row,'real_avg_success'))  else '—'),
            (f"{_g(row,'real_buy_match')*100:.1f}%"   if pd.notna(_g(row,'real_buy_match'))   else '—'),
            (f"{_g(row,'real_sell_match')*100:.1f}%"  if pd.notna(_g(row,'real_sell_match'))  else '—'),
            (f"{_g(row,'real_avg_match')*100:.1f}%"   if pd.notna(_g(row,'real_avg_match'))   else '—'),
            (f"{_g(row,'real_anchor_return')*100:+.2f}%" if pd.notna(_g(row,'real_anchor_return')) else '—'),
            (f"{_g(row,'corr_total_return')*100:+.2f}%" if pd.notna(_g(row,'corr_total_return')) else '—'),
            (f"{_g(row,'corr_max_drawdown')*100:.2f}%" if pd.notna(_g(row,'corr_max_drawdown')) else '—'),
            (f"{_g(row,'corr_win_rate')*100:.1f}%" if pd.notna(_g(row,'corr_win_rate')) else '—'),
            (f"{_g(row,'corr_buy_success')*100:.1f}%" if pd.notna(_g(row,'corr_buy_success')) else '—'),
            (f"{_g(row,'corr_sell_success')*100:.1f}%" if pd.notna(_g(row,'corr_sell_success')) else '—'),
            (f"{_g(row,'corr_buy_match')*100:.1f}%" if pd.notna(_g(row,'corr_buy_match')) else '—'),
            (f"{_g(row,'corr_sell_match')*100:.1f}%" if pd.notna(_g(row,'corr_sell_match')) else '—'),
            ('채택' if bool(_g(row,'corr_applied', False)) else '—'),
            (f"매수+{int(_g(row,'corr_n_added_buy',0))}/매도+{int(_g(row,'corr_n_added_sell',0))}"
                 if (pd.notna(_g(row,'corr_n_added_buy')) and (_g(row,'corr_n_added_buy',0) or _g(row,'corr_n_added_sell',0))) else '—'),
            # ★ 실행일 실제 포지션·점수 + 최근 매수/매도 (요청)
            ('보유' if _g(row,'run_pos')==1 else ('현금' if pd.notna(_g(row,'run_pos')) else '—')),
            (str(_g(row,'run_action')) if (_g(row,'run_action') is not None and pd.notna(_g(row,'run_action'))) else '—'),
            (f"{_g(row,'run_buy_score'):.2f}"  if pd.notna(_g(row,'run_buy_score'))  else '—'),
            (f"{_g(row,'run_sell_score'):.2f}" if pd.notna(_g(row,'run_sell_score')) else '—'),
            (pd.Timestamp(_g(row,'recent_buy_date')).date()  if pd.notna(_g(row,'recent_buy_date'))  else '—'),
            (f"{_g(row,'recent_buy_price'):.2f}"  if pd.notna(_g(row,'recent_buy_price'))  else '—'),
            (pd.Timestamp(_g(row,'recent_sell_date')).date() if pd.notna(_g(row,'recent_sell_date')) else '—'),
            (f"{_g(row,'recent_sell_price'):.2f}" if pd.notna(_g(row,'recent_sell_price')) else '—'),
            (f"{_g(row,'oos_win_rate')*100:.1f}%"     if pd.notna(_g(row,'oos_win_rate'))     else '—'),
            (f"{_g(row,'oos_max_drawdown')*100:.2f}%" if pd.notna(_g(row,'oos_max_drawdown')) else '—'),
            (f"{_g(row,'oos_return')*100:+.2f}%"      if pd.notna(_g(row,'oos_return'))      else '—'),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = _TH
            c.alignment = Alignment(horizontal='center'); c.font = Font(size=10)
            if is_best:
                c.fill = _HL
                c.font = Font(bold=True, color='C00000', size=11)
            elif ri % 2 == 1:
                c.fill = _ALT
        # 실제 수치 색칠 (12승률,13MDD,14수익,15매수성공,16매도성공,17평균성공,18매수매칭,19매도매칭,20평균매칭,21앵커수익)
        _rw = _g(row, 'real_win_rate'); _rm = _g(row, 'real_max_drawdown'); _rt = _g(row, 'real_total_return')
        if pd.notna(_rw): ws.cell(r, 12).fill = _success_fill(_rw)
        if pd.notna(_rm): ws.cell(r, 13).fill = _mdd_fill(_rm, mdd_limit_pct)
        if pd.notna(_rt): ws.cell(r, 14).fill = _ret_fill(_rt, bh_ret)
        _rb = _g(row, 'real_buy_success'); _rs2 = _g(row, 'real_sell_success'); _ra = _g(row, 'real_avg_success')
        if pd.notna(_rb):  ws.cell(r, 15).fill = _success_fill(_rb)
        if pd.notna(_rs2): ws.cell(r, 16).fill = _success_fill(_rs2)
        if pd.notna(_ra):  ws.cell(r, 17).fill = _success_fill(_ra)
        _rbm = _g(row, 'real_buy_match'); _rsm = _g(row, 'real_sell_match')
        _ravm = _g(row, 'real_avg_match'); _rar = _g(row, 'real_anchor_return')
        if pd.notna(_rbm):  ws.cell(r, 18).fill = _success_fill(_rbm)
        if pd.notna(_rsm):  ws.cell(r, 19).fill = _success_fill(_rsm)
        if pd.notna(_ravm): ws.cell(r, 20).fill = _success_fill(_ravm)
        if pd.notna(_rar):  ws.cell(r, 21).fill = _ret_fill(_rar, bh_ret)
        # 보정후 색칠 (22수익,23MDD,24승률,25매수성공,26매도성공,27매수매칭,28매도매칭,29채택)
        _cr = _g(row, 'corr_total_return'); _cm = _g(row, 'corr_max_drawdown'); _cw = _g(row, 'corr_win_rate')
        _cbs = _g(row, 'corr_buy_success'); _css = _g(row, 'corr_sell_success')
        _cbm = _g(row, 'corr_buy_match'); _csm = _g(row, 'corr_sell_match')
        _capplied = bool(_g(row, 'corr_applied', False))
        if pd.notna(_cr):  ws.cell(r, 22).fill = _ret_fill(_cr, bh_ret)
        if pd.notna(_cm):  ws.cell(r, 23).fill = _mdd_fill(_cm, mdd_limit_pct)
        if pd.notna(_cw):  ws.cell(r, 24).fill = _success_fill(_cw)
        if pd.notna(_cbs): ws.cell(r, 25).fill = _success_fill(_cbs)
        if pd.notna(_css): ws.cell(r, 26).fill = _success_fill(_css)
        if pd.notna(_cbm): ws.cell(r, 27).fill = _success_fill(_cbm)
        if pd.notna(_csm): ws.cell(r, 28).fill = _success_fill(_csm)
        if _capplied:      ws.cell(r, 29).fill = _HL
        # 🔬 OOS 색칠 (39승률, 40최대손실, 41수익) — 맨 끝 3개 컬럼
        _ow = _g(row, 'oos_win_rate'); _omd = _g(row, 'oos_max_drawdown'); _ortn = _g(row, 'oos_return')
        if pd.notna(_ow):   ws.cell(r, 39).fill = _success_fill(_ow)
        if pd.notna(_omd):  ws.cell(r, 40).fill = _mdd_fill(_omd, mdd_limit_pct)
        if pd.notna(_ortn): ws.cell(r, 41).fill = _ret_fill(_ortn, bh_ret)
        # 실행일 포지션(31) 색칠 — 보유/현금
        if _g(row,'run_pos') == 1: ws.cell(r, 31).fill = _HOLD
        elif pd.notna(_g(row,'run_pos')): ws.cell(r, 31).fill = _CASH
        if is_best:
            ws.cell(r, 14).font = Font(bold=True, color='C00000', size=12)
            if pd.notna(_cr):
                ws.cell(r, 22).font = Font(bold=True, color='C00000', size=12)
    for ci, w in enumerate([6, 9, 8, 8, 9, 8, 7, 8, 9, 8, 9,
                            11, 11, 11, 13, 12, 12, 12, 12, 12, 12,
                            12, 13, 12, 13, 13, 13, 13, 10, 14,
                            12, 18, 10, 10, 13, 10, 13, 10,
                            11, 13, 12], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'

    # 7. 거래 내역
    ws = wb.create_sheet('거래 내역'); ws.sheet_view.showGridLines = False
    n_stop = int(cur.get('n_stop_triggered', 0))
    ws.cell(1, 1).value = f'완결된 거래 — 총 {len(trades)}건 (손절매 {n_stop}건 포함)'
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells('A1:L1')
    _hdr(ws, 3, ['#', '진입일', '진입가', '⛔ 손절가', '청산사유',
                 '청산일', '청산가', '보유일',
                 '총수익%', '비용후수익%', '누적자산', '누적수익%'])
    for ri, row in trades.iterrows():
        r = ri + 4
        exit_reason = row.get('exit_reason', '매도신호')
        stop_p = row.get('stop_price', np.nan)
        sp_str = f"${stop_p:.2f}" if pd.notna(stop_p) else '-'
        vals = [int(row['trade_no']),
                row['entry_date'].date(), f"${row['entry_price']:.2f}",
                sp_str, exit_reason,
                row['exit_date'].date(),  f"${row['exit_price']:.2f}",
                int(row['days_held']),
                f"{row['gross_return_%']:+.2f}%",
                f"{row['net_return_%']:+.2f}%",
                f"{row['cum_equity']:.4f}",
                f"{row['cum_return_%']:+.2f}%"]
        is_stop = (exit_reason == '손절매')
        for ci, v in enumerate(vals, 1):
            c = ws.cell(r, ci); c.value = v; c.border = _TH
            c.alignment = Alignment(horizontal='center'); c.font = Font(size=10)
            if ri % 2 == 1: c.fill = _ALT
        if is_stop:
            ws.cell(r, 5).fill = _BAD
            ws.cell(r, 5).font = Font(bold=True, size=10, color='C00000')
        if row['net_return_%'] > 0:
            ws.cell(r, 10).fill = _GOOD
            ws.cell(r, 10).font = Font(bold=True, size=10, color='006100')
        else:
            ws.cell(r, 10).fill = _BAD
            ws.cell(r, 10).font = Font(bold=True, size=10, color='C00000')
    for ci, w in enumerate([5, 12, 10, 11, 11, 12, 10, 8, 10, 12, 10, 12], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'

    # 7-1. 🔧 거래 진단 (요청) — 앵커 외 거래 포함, 손실·1%이하 수익 거래를 '보정 대상'으로 식별
    try:
        _pf = float(globals().get('CORRECT_PROFIT_FLOOR', 0.01))
        _fidx = daily['date'] if (daily is not None and hasattr(daily, 'columns') and 'date' in daily.columns) else []
        diag_df = diagnose_trades(trades, close=None, feat_index=_fidx,
                                  anchor_prio=anchor_prio, profit_floor=_pf)
    except Exception as _de:
        diag_df = None
    if diag_df is not None and len(diag_df) > 0:
        ws = wb.create_sheet('🔧 거래 진단'); ws.sheet_view.showGridLines = False
        n_wrong = int(diag_df['보정대상'].sum())
        ws.cell(1, 1).value = (f'거래 진단 — 총 {len(diag_df)}건 중 보정 대상(손실 또는 +{_pf*100:.0f}% 이하 수익) {n_wrong}건  '
                               f'※ 앵커(우선순위1 최대수익스윙 / 우선순위2 {_pf*100:.0f}%이상)와 매칭 진단')
        ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
        ws.merge_cells('A1:K1')
        _hdr(ws, 3, ['#', '진입일', '청산일', '비용후수익%', '보유일', '청산사유',
                     '매수 진단', '매도 진단', '매수앵커일치', '매도앵커일치', '판정'])
        for ri, drow in diag_df.iterrows():
            r = ri + 4
            _ed = drow['entry_date']; _xd = drow['exit_date']
            vals = [int(drow['trade_no']),
                    (_ed.date() if pd.notna(_ed) else '—'),
                    (_xd.date() if pd.notna(_xd) else '—'),
                    f"{drow['net_return_%']:+.2f}%",
                    int(drow['days_held']), drow['exit_reason'],
                    drow['매수_진단'], drow['매도_진단'],
                    ('✓' if drow['매수_앵커일치'] else '✗'),
                    ('✓' if drow['매도_앵커일치'] else '✗'),
                    drow['판정']]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(r, ci); c.value = v; c.border = _TH
                c.alignment = Alignment(horizontal='center'); c.font = Font(size=10)
                if ri % 2 == 1: c.fill = _ALT
            if bool(drow['보정대상']):
                for ci in range(1, 12):
                    ws.cell(r, ci).fill = PatternFill('solid', fgColor='FCE4D6')
                ws.cell(r, 4).font = Font(bold=True, color='C00000', size=10)
        for ci, w in enumerate([6, 13, 13, 12, 8, 12, 22, 22, 12, 12, 28], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = 'A4'
    if anchor_mode and (cur.get('anchor_buy_diagnosis') or cur.get('anchor_sell_diagnosis')):
        ws = wb.create_sheet('⚓ ANCHOR 진단'); ws.sheet_view.showGridLines = False
        buy_diag = cur.get('anchor_buy_diagnosis', [])
        sell_diag = cur.get('anchor_sell_diagnosis', [])
        n_buy_miss  = sum(1 for d in buy_diag  if not d.get('executed_match', d.get('signal_on', False)))
        n_sell_miss = sum(1 for d in sell_diag if not d.get('executed_match', d.get('signal_on', False)))

        title_extra = ' (AUTO 자동계산)' if auto_anchor else ' (수동 입력)'
        ws.cell(1, 1).value = (f'⚓ ANCHOR 정답일 진단{title_extra} — '
                                f'매수 {len(buy_diag)}개 중 잡힘 '
                                f'{len(buy_diag)-n_buy_miss}/{n_buy_miss} 못잡힘,  '
                                f'매도 {len(sell_diag)}개 중 잡힘 '
                                f'{len(sell_diag)-n_sell_miss}/{n_sell_miss} 못잡힘')
        ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
        ws.merge_cells('A1:F1')

        ws.cell(3, 1).value = '※ 매칭률은 실제 일별 거래 결과 기준 — 신호만 떠도 충돌 해결로 반대 방향 갔으면 unmatched'
        ws.cell(3, 1).font = Font(size=10, color='606060', italic=True)
        ws.merge_cells('A3:F3')
        ws.cell(4, 1).value = '    개선 방법: MAX_INDICATORS↑, N_THRESHOLDS↑, K_BUY/SELL_RANGE 확장, VOTE_RATIO 낮은 비율 추가, top_n_pool↑'
        ws.cell(4, 1).font = Font(size=10, color='606060', italic=True)
        ws.merge_cells('A4:F4')

        ws.cell(6, 1).value = f'■ 매수 정답일 진단 ({len(buy_diag)}개)'
        ws.cell(6, 1).font = Font(bold=True, size=12, color='006100')
        _hdr(ws, 7, ['#', '정답 매수일', '신호 떠야 했던 날',
                      f'buy_count (/{cur["K_buy"]})', f'필요 vote ({cur["vote_buy"]})', '결과'])
        for ri, d in enumerate(buy_diag):
            r = 8 + ri
            caught = d.get('executed_match', d.get('signal_on', False))
            vals = [ri+1,
                    d['target_date'].date() if hasattr(d['target_date'], 'date') else d['target_date'],
                    d['signal_day'].date() if hasattr(d['signal_day'], 'date') else d['signal_day'],
                    round(float(d['buy_count']), 2),
                    int(cur['vote_buy']),
                    d['status']]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(r, ci); c.value = v; c.border = _TH
                c.alignment = Alignment(horizontal='center'); c.font = Font(size=10)
                if ri % 2 == 1: c.fill = _ALT
            if caught:
                ws.cell(r, 6).fill = _GOOD
                ws.cell(r, 6).font = Font(bold=True, size=10, color='006100')
            else:
                ws.cell(r, 6).fill = _BAD
                ws.cell(r, 6).font = Font(bold=True, size=10, color='C00000')

        sell_start = 8 + len(buy_diag) + 3
        ws.cell(sell_start - 1, 1).value = f'■ 매도 정답일 진단 ({len(sell_diag)}개)'
        ws.cell(sell_start - 1, 1).font = Font(bold=True, size=12, color='C00000')
        _hdr(ws, sell_start, ['#', '정답 매도일', '신호 떠야 했던 날',
                               f'sell_count (/{cur["K_sell"]})', f'필요 vote ({cur["vote_sell"]})', '결과'])
        for ri, d in enumerate(sell_diag):
            r = sell_start + 1 + ri
            caught = d.get('executed_match', d.get('signal_on', False))
            vals = [ri+1,
                    d['target_date'].date() if hasattr(d['target_date'], 'date') else d['target_date'],
                    d['signal_day'].date() if hasattr(d['signal_day'], 'date') else d['signal_day'],
                    round(float(d['sell_count']), 2),
                    int(cur['vote_sell']),
                    d['status']]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(r, ci); c.value = v; c.border = _TH
                c.alignment = Alignment(horizontal='center'); c.font = Font(size=10)
                if ri % 2 == 1: c.fill = _ALT
            if caught:
                ws.cell(r, 6).fill = _GOOD
                ws.cell(r, 6).font = Font(bold=True, size=10, color='006100')
            else:
                ws.cell(r, 6).fill = _BAD
                ws.cell(r, 6).font = Font(bold=True, size=10, color='C00000')

        for ci, w in enumerate([5, 14, 14, 14, 14, 38], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = 'A8'

    # 8. 일별 백테스트 (맨 앞 시트로 배치 — 요청)
    ws = wb.create_sheet('일별 백테스트', 0); ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = f'일별 백테스트 — {len(daily)}일'
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells('A1:T1')
    _ind_hdr = [(('매수:' if _k == 'BUY' else '매도:') + _nm)
                for (_k, _nm, _vc, _sc) in cur.get('used_ind_cols', [])]
    _hdr(ws, 4, ['날짜', f'{ticker}종가',
                 f'매수카운트(/{cur["K_buy"]})', '매수ON', '매수성공',
                 f'매도카운트(/{cur["K_sell"]})', '매도ON', '매도성공',
                 '포지션', '액션', '진입가', '⛔ 손절가', '보유일',
                 '미실현%', '실현%', '누적자산', '누적수익%', '진행최대손실%']
                + _ind_hdr)
    _write_daily_rows(ws, daily, cur, mdd_limit_pct)

    # 9. ★ OOS 일별 거래
    if oos_enabled and oos_daily is not None and len(oos_daily) > 0:
        ws = wb.create_sheet('OOS 일별 거래'); ws.sheet_view.showGridLines = False
        oc = oos_cur if oos_cur is not None else cur
        oos_bh = 0.0
        oos_bh_up = 0.0
        try:
            oos_bh = _bh_sum_return(oos_daily['close'].values) * 100
            oos_bh_up = _bh_up_sum_return(oos_daily['close'].values) * 100
        except Exception:
            oos_bh_up = 0.0
        ws.cell(1, 1).value = (f'🔬 OOS 검증 일별 거래 — {len(oos_daily)}일  '
                               f'[{oos_daily["date"].iloc[0].date()} ~ {oos_daily["date"].iloc[-1].date()}]')
        ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
        ws.merge_cells('A1:T1')
        ws.cell(2, 1).value = (f'OOS 누적 {oc["cum_return_pct"]:+.2f}%  /  B&H {oos_bh:+.2f}%  /  '
                               f'📈 상승일만 전략 {oc.get("up_cum_return_pct", 0):+.2f}% / B&H {oos_bh_up:+.2f}%  /  '
                               f'거래 {oc["n_trades"]}회  /  승률 {oc["win_rate"]*100:.1f}%  '
                               f'(IS 학습구간에서 찾은 조합을 검증구간에 그대로 적용 — 손 안 댐)')
        ws.cell(2, 1).font = Font(italic=True, size=10, color='C00000')
        ws.merge_cells('A2:T2')
        _oind_hdr = [(('매수:' if _k == 'BUY' else '매도:') + _nm)
                     for (_k, _nm, _vc, _sc) in oc.get('used_ind_cols', [])]
        _hdr(ws, 4, ['날짜', f'{ticker}종가',
                     f'매수카운트(/{oc["K_buy"]})', '매수ON', '매수성공',
                     f'매도카운트(/{oc["K_sell"]})', '매도ON', '매도성공',
                     '포지션', '액션', '진입가', '⛔ 손절가', '보유일',
                     '미실현%', '실현%', '누적자산', '누적수익%', '진행최대손실%']
                    + _oind_hdr)
        _write_daily_rows(ws, oos_daily, oc, mdd_limit_pct)

    # ─── ★ feature2 (요청) — 매수/매도 지표 신호 매트릭스 (각각 별도 시트) ───
    try:
        _abset = _norm_date_set(anchor_buy_dates)
        _asset = _norm_date_set(anchor_sell_dates)
        ws = wb.create_sheet('매수 지표 신호'); ws.sheet_view.showGridLines = False
        _write_indicator_matrix_sheet(ws, buy_pool, feat, close_full,
                                      _abset, _asset, ticker, '매수')
        ws = wb.create_sheet('매도 지표 신호'); ws.sheet_view.showGridLines = False
        _write_indicator_matrix_sheet(ws, sell_pool, feat, close_full,
                                      _abset, _asset, ticker, '매도')
    except Exception as _eim:
        print(f"  ⚠ 지표 신호 매트릭스 시트 작성 실패(무시): {_eim}")

    # ─── 7b. 성공률 우선 선출 시트 (요청) — 성공률로 먼저 뽑은 지표 정리 ───
    try:
        _sp = globals().get('_LAST_SUCCESS_POOL', (None, None))
        _spb, _sps = (_sp if isinstance(_sp, (tuple, list)) and len(_sp) == 2 else (None, None))

        def _write_success_sheet(ws, df, kind_label):
            ws.cell(1, 1).value = (f'{ticker} — {kind_label} 성공률 우선 선출 '
                                   f'(성공률 {POOL_SUCCESS_MIN_RATE*100:.0f}%+ · 최소신호 {POOL_SUCCESS_MIN_SIG} · '
                                   f'분위 {POOL_SUCCESS_WIDE_PCT})  ※성공률 내림차순, 동률은 점수(Wilson)')
            ws.cell(1, 1).font = Font(bold=True, size=11)
            heads = ['순위', '지표', '방향', '임계치', '분위', '신호수', '성공수', '성공률', '점수']
            _hdr(ws, 3, heads)
            if df is None or len(df) == 0:
                ws.cell(4, 1).value = '조건(성공률/최소신호)을 만족하는 지표 없음'
                return
            for i, (_, r) in enumerate(df.iterrows(), start=1):
                rr = 3 + i
                vals = [i, r['indicator'], r['direction'],
                        round(float(r['threshold']), 6),
                        (f"{float(r['pct_label']):.2f}" if r['direction'] in ('>=', '<=')
                         else f"z{r['direction'][1:]}{r['pct_label']:g}"),
                        int(r['n_signals']), int(r['n_success']),
                        f"{float(r['success_rate'])*100:.2f}%",
                        round(float(r['score']), 4)]
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(rr, ci); c.value = v
                    if ci == 8:  # 성공률 강조
                        c.font = Font(bold=True)
            widths = [6, 30, 8, 12, 12, 9, 9, 10, 10]
            for ci, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.freeze_panes = 'A4'

        ws = wb.create_sheet('성공률 우선 매수'); ws.sheet_view.showGridLines = False
        _write_success_sheet(ws, _spb, '매수')
        ws = wb.create_sheet('성공률 우선 매도'); ws.sheet_view.showGridLines = False
        _write_success_sheet(ws, _sps, '매도')
    except Exception as _esp:
        print(f"  ⚠ 성공률 우선 선출 시트 작성 실패(무시): {_esp}")

    # ─── 8. 메타조합별 지표 풀 (그리드 번호 재현 정확도용) ───
    #   각 메타조합(wilson_z/pct/corr)이 선별한 매수·매도 지표 전체를 저장.
    #   그리드 번호 재현 시, 그 번호의 메타조합 풀을 여기서 읽어 '그대로' 사용 → 정확히 재현.
    _pool_map = globals().get('_LAST_POOL_MAP', None)
    if _pool_map:
        try:
            ws = wb.create_sheet('메타조합별_지표'); ws.sheet_view.showGridLines = False
            ws.cell(1, 1).value = ('메타조합별 지표 풀 — 그리드 번호 재현 시 그 조합의 메타변수에 맞는 '
                                   '지표를 여기서 읽어 정확히 재현합니다 (지표 재선별 안 함)')
            ws.cell(1, 1).font = Font(bold=True, size=12, color='1F3864')
            r = 3
            for (wz, pl, ph, corr), (bp, sp) in _pool_map.items():
                ws.cell(r, 1).value = f'★ wilson_z={wz} / pct=({pl},{ph}) / corr={corr}'
                ws.cell(r, 1).font = Font(bold=True, size=11, color='C00000')
                r += 1
                for side, pool in [('매수', bp), ('매도', sp)]:
                    if pool is None or len(pool) == 0: continue
                    ws.cell(r, 1).value = f'  [{side}] {len(pool)}개'
                    ws.cell(r, 1).font = Font(bold=True, size=10)
                    r += 1
                    _hdr(ws, r, ['side', '지표', '방향', '임계치', '분위', '신호수', '성공수', '성공률', '점수'])
                    r += 1
                    for _, prow in pool.iterrows():
                        ws.cell(r, 1).value = side
                        ws.cell(r, 2).value = str(prow.get('indicator', ''))
                        ws.cell(r, 3).value = str(prow.get('direction', '>='))
                        ws.cell(r, 4).value = float(prow.get('threshold', 0.0))
                        ws.cell(r, 5).value = float(prow.get('pct_label', 50.0)) if 'pct_label' in prow else ''
                        ws.cell(r, 6).value = float(prow.get('n_signals', 0))
                        ws.cell(r, 7).value = float(prow.get('n_success', 0))
                        ws.cell(r, 8).value = float(prow.get('success_rate', 0.0))
                        ws.cell(r, 9).value = float(prow.get('score', 0.0))
                        r += 1
                    r += 1
            for ci, w in enumerate([8, 28, 6, 12, 8, 8, 8, 10, 10], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
        except Exception as _e:
            print(f"  ⚠ 메타조합별_지표 시트 작성 실패: {_e}")

    wb.save(output_file)


# ════════════════════════════════════════════════════════════════
#                  메모리 변수 자동 로드
# ════════════════════════════════════════════════════════════════
def _resolve_data():
    g = globals()
    feat   = g.get('_pair_feat',  g.get('feat'))
    close  = g.get('_pair_close', g.get('close'))
    ticker = g.get('_pair_ticker', g.get('TICKER'))
    if feat is not None and close is not None and ticker is not None:
        print(f"  ✓ 메모리 로드: feat={feat.shape}, close={len(close)}일, ticker={ticker}")
        return feat, close, ticker
    print("  ⚠ 변수 없음 → 기존 코드 함수 시도")
    needed = ['download_data', 'download_fred_data', 'compute_features',
              'TICKER', 'DOWNLOAD_START', 'EVAL_START']
    miss = [n for n in needed if n not in g]
    if miss:
        raise RuntimeError(f"필요: {miss}")
    closes, ohlcv = g['download_data'](start=g['DOWNLOAD_START'])
    fred_df = g['download_fred_data'](start=g['DOWNLOAD_START'])
    if len(fred_df) == 0: fred_df = None
    feat = g['compute_features'](ohlcv, closes, fred_df=fred_df)
    feat = feat[feat.index >= pd.Timestamp(g['EVAL_START'])]
    close = ohlcv[g['TICKER']]['Close'].reindex(feat.index)
    return feat, close, g['TICKER']


# ════════════════════════════════════════════════════════════════
#                       메인 진입
# ════════════════════════════════════════════════════════════════
def _final_pick_by_real(pool, win_tol):
    """실거래 검증 결과 pool에서 최종 1개 선정 (요청 개정):
       SELECT_BY_CORRECTED_RETURN=True이고 corr_total_return 컬럼이 있으면
         → '보정 후 수익률'이 가장 높은 조합을 선정 (동률 시 보정후 승률).
       아니면 OOS_SELECT_BY_OOS_RETURN=True → OOS 수익, 그것도 없으면 전체 실제 수익.
       반환: (best_inner_dict, sel_row)"""
    use_corr = globals().get('SELECT_BY_CORRECTED_RETURN', True)
    use_oos  = globals().get('OOS_SELECT_BY_OOS_RETURN', True)
    # ★ 보정후 승률 하한 + 허용오차 비교 선정 (요청):
    #   1) 보정후 실제 승률(sel_win_rate) ≥ SELECT_WIN_FLOOR(90%) 중 보정후 수익(sel_total_return) 최고 = 1차 선정.
    #   2) 그 1차 선정의 승률에서 WINRATE_TOLERANCE 이내(승률이 최대 그만큼 낮아도 됨) 후보와 비교해,
    #      보정후 수익이 더 높은 게 있으면 그것으로 재선정. (90% 못 넘는 후보 없으면 최고승률 기준 폴백)
    if 'sel_win_rate' in pool.columns and 'sel_total_return' in pool.columns and len(pool) > 0:
        floor = float(globals().get('SELECT_WIN_FLOOR', 0.90))
        above = pool[pool['sel_win_rate'] >= floor]
        if len(above) > 0:
            b0 = above.sort_values(['sel_total_return', 'sel_win_rate'],
                                   ascending=[False, False]).reset_index(drop=True)
            w0 = float(b0.iloc[0]['sel_win_rate'])          # 1차 선정의 보정후 승률
            band = pool[pool['sel_win_rate'] >= w0 - win_tol]  # 승률 허용오차 이내(아래로) 후보
            b = band.sort_values(['sel_total_return', 'sel_win_rate'],
                                 ascending=[False, False]).reset_index(drop=True)
            _by = '\ubcf4\uc815\ud6c4\uc2b9\ub960\ud558\ud55c+\ud5c8\uc6a9\uc624\ucc28'
        else:
            # 90% 넘는 후보 없음 → 최고승률 -win_tol 밴드 내 수익 최고
            wmax = float(pool['sel_win_rate'].max())
            band = pool[pool['sel_win_rate'] >= wmax - win_tol]
            b = band.sort_values(['sel_total_return', 'sel_win_rate'],
                                 ascending=[False, False]).reset_index(drop=True)
            _by = '\ubcf4\uc815\ud6c4\ucd5c\uace0\uc2b9\ub960\ubc34\ub4dc(90%\ubbf8\ub2ec)'
    elif use_corr and 'corr_total_return' in pool.columns and pool['corr_total_return'].notna().any():
        _wcol = 'corr_win_rate' if 'corr_win_rate' in pool.columns else 'real_win_rate'
        b = pool[pool['corr_total_return'].notna()].sort_values(
            ['corr_total_return', _wcol], ascending=[False, False]).reset_index(drop=True)
        _by = '\ubcf4\uc815\ud6c4'
    elif use_oos and 'oos_total_return' in pool.columns and pool['oos_total_return'].notna().any():
        b = pool[pool['oos_total_return'].notna()].sort_values(
            ['oos_total_return', 'oos_win_rate' if 'oos_win_rate' in pool.columns else 'real_win_rate'],
            ascending=[False, False]).reset_index(drop=True)
        _by = 'OOS'
    else:
        b = pool.sort_values(['real_total_return', 'real_win_rate'],
                             ascending=[False, False]).reset_index(drop=True)
        _by = '\uc804\uccb4'
    sel = b.iloc[0]
    best_inner = {'K_buy': int(sel['K_buy']), 'vote_buy': int(sel['vote_buy']),
                  'K_sell': int(sel['K_sell']), 'vote_sell': int(sel['vote_sell'])}
    for _k in ('sharpe_like', 'total_return', 'avg_success_rate',
               'buy_success_rate', 'sell_success_rate', 'win_rate',
               'max_drawdown', 'n_trades'):
        if _k in sel.index:
            best_inner[_k] = sel[_k]
    _oosr = sel.get('oos_total_return', float('nan'))
    print(f"  \u2713 \uc2e4\uac70\ub798 \uac80\uc99d \uc644\ub8cc \u2014 {_by} \uc218\uc775\ub960 \ucd5c\uace0 \uc870\ud569 \uc120\uc815")
    if _by.startswith('\ubcf4\uc815\ud6c4'):
        _cr = sel.get('sel_total_return', sel.get('corr_total_return', float('nan')))
        _applied = bool(sel.get('corr_applied', False))
        _sw = sel.get('sel_win_rate', sel.get('real_win_rate', float('nan')))
        print(f"     \ucd5c\uc885: K_buy={best_inner['K_buy']}/v{best_inner['vote_buy']}, "
              f"K_sell={best_inner['K_sell']}/v{best_inner['vote_sell']}  "
              f"(\uc120\uc815\uc218\uc775 {_cr*100:+.2f}%{'(\ubcf4\uc815\ucc44\ud0dd)' if _applied else '(\ubcf4\uc815\uc548\ub428)'}, "
              f"\uc6d0\ub798\uc218\uc775 {sel['real_total_return']*100:+.2f}%, "
              f"\uc120\uc815\uc2b9\ub960 {_sw*100:.1f}%, "
              f"\uc120\uc815\ucd5c\ub300\uc190\uc2e4 {sel.get('sel_max_drawdown', sel['real_max_drawdown'])*100:.2f}%)")
    else:
        print(f"     \ucd5c\uc885: K_buy={best_inner['K_buy']}/v{best_inner['vote_buy']}, "
              f"K_sell={best_inner['K_sell']}/v{best_inner['vote_sell']}  "
              f"(OOS\uc218\uc775 {_oosr*100:+.2f}%, OOS\ub9e4\ub3c4\uc815\ud655\ub3c4 {sel.get('oos_sell_success', float('nan'))*100:.1f}%, "
              f"OOS\ub9e4\uc218\uc815\ud655\ub3c4 {sel.get('oos_buy_success', float('nan'))*100:.1f}%, "
              f"\uc804\uccb4\uc218\uc775 {sel['real_total_return']*100:+.2f}%)")
    return best_inner, sel


def _pick_verify_candidates(tbl, bh_ret=None):
    """실거래 검증 대상 후보를 고른다 (요청):
       - 그리드 최고 수익률과 B&H 수익률 '사이'에 있는 후보 (B&H ≤ 그리드수익 ≤ 최고)
       - 그중 그리드 수익 높은 순으로 정렬 → 상위 '절반'만 검증
       - VERIFY_TOP_N 으로 상한도 둔다 (절반이 너무 많으면 컷).
    """
    top_n = int(globals().get('VERIFY_TOP_N', 10000))
    ret_col = 'total_return' if 'total_return' in tbl.columns else (
              'combined_return' if 'combined_return' in tbl.columns else None)
    cand = tbl.copy()
    if ret_col is None:
        return cand.head(top_n).reset_index(drop=True), "수익컬럼 없음 → 상위"

    # 그리드 최고 수익과 B&H 사이 구간만
    grid_max = cand[ret_col].max()
    note = ""
    if bh_ret is not None and pd.notna(bh_ret):
        lo, hi = min(bh_ret, grid_max), max(bh_ret, grid_max)
        band = cand[(cand[ret_col] >= lo) & (cand[ret_col] <= hi)]
        if len(band) > 0:
            cand = band
            note = f"B&H({bh_ret*100:+.1f}%)~최고({grid_max*100:+.1f}%) 구간"
        else:
            note = "구간 후보 없음 → 전체"
    # 수익 높은 순 → 절반만
    cand = cand.sort_values(ret_col, ascending=False).reset_index(drop=True)
    half = max(1, len(cand) // 2)
    cand = cand.head(min(half, top_n)).reset_index(drop=True)
    note = (note + f", 수익 상위 절반 {len(cand)}개").strip(', ')
    return cand, note


def _extract_runday_info(_d, _t):
    """실행일(가장 최근 거래일) 기준 실제 포지션·액션·점수와 최근 매수/매도 정보 추출 (요청).
       반환 키: run_pos(1보유/0현금), run_action(일별 백테스트 액션), run_buy_score, run_sell_score,
                recent_buy_date, recent_buy_price, recent_sell_date, recent_sell_price."""
    info = {'run_pos': np.nan, 'run_action': None,
            'run_buy_score': np.nan, 'run_sell_score': np.nan,
            'recent_buy_date': None, 'recent_buy_price': np.nan,
            'recent_sell_date': None, 'recent_sell_price': np.nan}
    try:
        if _d is not None and len(_d) > 0:
            last = _d.iloc[-1]
            info['run_pos']        = int(last.get('position_pre', 0))
            _act = last.get('action', '')
            info['run_action']     = str(_act).strip() if (_act is not None and str(_act).strip()) else '관망'
            info['run_buy_score']  = float(last.get('buy_count', np.nan))
            info['run_sell_score'] = float(last.get('sell_count', np.nan))
            # 현재 보유 중이면 '열린 포지션' 진입가/매수일
            if info['run_pos'] == 1 and pd.notna(last.get('entry_price', np.nan)):
                info['recent_buy_price'] = float(last['entry_price'])
                if 'action' in _d.columns:
                    _bm = _d['action'].astype(str).str.contains('매수', na=False)
                    if _bm.any():
                        info['recent_buy_date'] = _d.loc[_bm, 'date'].iloc[-1]
        if _t is not None and len(_t) > 0:
            lt = _t.iloc[-1]
            # 가장 최근 청산(매도)
            info['recent_sell_date']  = lt.get('exit_date', None)
            info['recent_sell_price'] = float(lt.get('exit_price', np.nan))
            # 현금 상태면 최근 매수 = 마지막으로 청산된 거래의 진입
            if info['run_pos'] != 1:
                info['recent_buy_date']  = lt.get('entry_date', None)
                info['recent_buy_price'] = float(lt.get('entry_price', np.nan))
    except Exception:
        pass
    return info


def _verify_staged_candidates(merged_table, feat, close, pool_map, *,
                              horizon, dd_limit, ru_limit, stop_loss_pct, anchor_mode,
                              bh_ret=None, anchor_safe_buy=None, anchor_safe_sell=None):
    """staged merged_table의 후보를 각 메타조합 풀로 실제 일별 백테스트.
       실제 MDD 기준 통과분 중 실제승률 -10%p 범위에서 수익 최고 선정.
       반환: (best_inner, merged_table_with_real_cols)"""
    win_tol = globals().get('WINRATE_TOLERANCE', 0.10)
    top_n   = int(globals().get('VERIFY_TOP_N', 10000))
    mdd_lim = globals().get('VERIFY_MAX_DRAWDOWN_LIMIT', None)

    cand, _note = _pick_verify_candidates(merged_table, bh_ret=bh_ret)
    print(f"\n  🔬 실거래 검증 — {_note} → {len(cand)}개를 실제 일별 백테스트로 재계산 "
          f"(모든 변수 반복 끝난 뒤 1회)")
    try: _ZCACHE.clear()   # ★ 검증 전 z-캐시 정리 (요청)
    except Exception: pass

    real_rows = []
    _t0 = time.time()
    _miss_pool = 0
    for _i, _r in cand.iterrows():
        _key = (round(float(_r.get('meta_wilson_z', 0)),4),
                int(_r.get('meta_pct_low', 0)), int(_r.get('meta_pct_high', 0)),
                round(float(_r.get('meta_corr_limit', 0)),4))
        pools = pool_map.get(_key)
        if pools is None:
            _miss_pool += 1
            continue
        bp, sp = pools
        try:
            _d, _t, _cur, _bu, _su = daily_ensemble_backtest(
                feat, close, bp, sp,
                K_buy=int(_r['K_buy']), K_sell=int(_r['K_sell']),
                vote_buy=int(_r['vote_buy']), vote_sell=int(_r['vote_sell']),
                cost=COST_PER_TRADE, horizon=horizon,
                dd_limit=dd_limit, ru_limit=ru_limit,
                stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
                anchor_safe_buy=anchor_safe_buy, anchor_safe_sell=anchor_safe_sell)
        except Exception:
            continue
        row = _r.to_dict()
        row['real_win_rate']      = float(_cur.get('win_rate', 0.0))
        row['real_max_drawdown']  = float(_cur.get('max_drawdown', 0.0))
        row['real_total_return']  = float(_cur.get('cum_return_pct', 0.0)) / 100.0
        row['real_n_trades']      = int(_cur.get('n_trades', 0))
        row['real_buy_success']   = float(_cur.get('buy_success_rate', 0.0))
        row['real_sell_success']  = float(_cur.get('sell_success_rate', 0.0))
        row['real_avg_success']   = float(_cur.get('avg_success_rate', 0.0))
        row['real_buy_match']     = float(_cur.get('anchor_buy_match_rate', 0.0))
        row['real_sell_match']    = float(_cur.get('anchor_sell_match_rate', 0.0))
        row['real_avg_match']     = float(_cur.get('anchor_avg_match_rate', 0.0))
        row['real_anchor_return'] = float(_cur.get('anchor_strategy_return', 0.0))
        # ★ 실행일(가장 최근일) 실제 포지션·점수 + 최근 매수/매도 정보 (요청)
        _ri = _extract_runday_info(_d, _t)
        row.update(_ri)
        # ★ OOS 기능 OFF (요청) — OOS 백테스트 생략, 컬럼은 NaN
        if globals().get('OOS_SELECT_BY_OOS_RETURN', False):
            try:
                _oos_m = int(globals().get('OOS_MONTHS', 1))
                _last_dt = feat.index.max()
                _oos_start = _last_dt - pd.DateOffset(months=_oos_m)
                _omask = feat.index >= _oos_start
                _omask = np.asarray(_omask)
                _feat_oos = feat.loc[_omask]
                _close_oos = close.reindex(_feat_oos.index)
                _asb_o = (anchor_safe_buy[_omask] if (anchor_safe_buy is not None and len(anchor_safe_buy)==len(_omask)) else None)
                _ass_o = (anchor_safe_sell[_omask] if (anchor_safe_sell is not None and len(anchor_safe_sell)==len(_omask)) else None)
                if len(_feat_oos) >= 5:
                    _do, _to, _curo, _buo, _suo = daily_ensemble_backtest(
                        _feat_oos, _close_oos, bp, sp,
                        K_buy=int(_r['K_buy']), K_sell=int(_r['K_sell']),
                        vote_buy=int(_r['vote_buy']), vote_sell=int(_r['vote_sell']),
                        cost=COST_PER_TRADE, horizon=horizon,
                        dd_limit=dd_limit, ru_limit=ru_limit,
                        stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
                        anchor_safe_buy=_asb_o, anchor_safe_sell=_ass_o)
                    row['oos_total_return'] = float(_curo.get('cum_return_pct', 0.0)) / 100.0
                    row['oos_buy_success']  = float(_curo.get('buy_success_rate', 0.0))
                    row['oos_sell_success'] = float(_curo.get('sell_success_rate', 0.0))
                    row['oos_win_rate']     = float(_curo.get('win_rate', 0.0))
                    row['oos_n_trades']     = int(_curo.get('n_trades', 0))
                else:
                    row['oos_total_return'] = float('nan')
                    row['oos_buy_success'] = row['oos_sell_success'] = float('nan')
                    row['oos_win_rate'] = float('nan'); row['oos_n_trades'] = 0
            except Exception:
                row['oos_total_return'] = float('nan')
                row['oos_buy_success'] = row['oos_sell_success'] = float('nan')
                row['oos_win_rate'] = float('nan'); row['oos_n_trades'] = 0
        real_rows.append(row)
        if (_i + 1) % 1000 == 0:
            print(f"     ... {_i+1}/{len(cand)} 검증  (경과 {time.time()-_t0:.0f}초)")

    if _miss_pool > 0:
        print(f"     ℹ 풀 정보를 못 찾아 건너뛴 후보 {_miss_pool}개 (메타조합 캐시 불일치)")
    if not real_rows:
        print(f"  ⚠ 실거래 검증 후보가 모두 실패 → 그리드 근사 기준 사용")
        return None, merged_table

    verified = pd.DataFrame(real_rows)

    # ★ 보정 통합 (요청) — 수익 상위 N개 후보에 지표 가중치 보정 적용 후 재백테스트.
    #   보정 후 수익/손실/승률을 corr_* 컬럼에 기록. 보정 후 수익으로 최종 선정.
    if globals().get('USE_ANCHOR_MATCH_CORRECTION', True):
        verified = _apply_correction_to_candidates(
            verified, feat, close, pool_map,
            horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
            stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
            anchor_safe_buy=anchor_safe_buy, anchor_safe_sell=anchor_safe_sell)

    # ★ 선정 필터 수치를 '보정 후'로 (요청) — 보정 적용된 조합은 corr_*, 아니면 real_* 사용.
    #   sel_mdd / sel_win / sel_ret = 선정에 쓸 '유효 수치'.
    _use_corr_sel = globals().get('SELECT_FILTER_BY_CORRECTED', True)
    def _eff_col(colc, colr):
        base = verified[colr].astype(float).values
        if _use_corr_sel and colc in verified.columns:
            _ca = verified.get('corr_applied', pd.Series(False, index=verified.index))
            _ca = _ca.fillna(False).astype(bool).values
            cval = verified[colc].astype(float).values
            return np.where(_ca & ~np.isnan(cval), cval, base)
        return base
    verified['sel_max_drawdown'] = _eff_col('corr_max_drawdown', 'real_max_drawdown')
    verified['sel_win_rate']     = _eff_col('corr_win_rate',     'real_win_rate')
    verified['sel_total_return'] = _eff_col('corr_total_return', 'real_total_return')

    pool = verified
    if mdd_lim is not None:
        safe = verified[verified['sel_max_drawdown'] >= mdd_lim]    # ★ 보정후 최대손실로 필터
        if len(safe) > 0:
            pool = safe
            print(f"  🛡 (보정후) 최대 거래손실 {mdd_lim*100:.1f}% 이하(안전) 후보 {len(safe)}개로 선정 "
                  f"(전체 검증 {len(verified)}개 중)")
        else:
            print(f"  ⚠ (보정후) 최대 거래손실 {mdd_lim*100:.1f}% 이하 후보 없음 → 가장 손실 얕은 조합 선정")
            pool = verified.sort_values('sel_max_drawdown', ascending=False)

    best_inner, sel = _final_pick_by_real(pool, win_tol)
    out = verified.sort_values('sel_total_return', ascending=False).reset_index(drop=True)
    print(f"     실제 승률 {sel['real_win_rate']*100:.1f}%, "
          f"실제 평균성공 {sel['real_avg_success']*100:.1f}% "
          f"(매수 {sel['real_buy_success']*100:.1f}% / 매도 {sel['real_sell_success']*100:.1f}%), "
          f"실제 최대손실 {sel['real_max_drawdown']*100:.2f}%, "
          f"실제 수익 {sel['real_total_return']*100:+.2f}%")
    return best_inner, out


def _verify_candidates_by_daily(inner_passed, feat, close, buy_pool, sell_pool, *,
                                horizon, dd_limit, ru_limit, stop_loss_pct,
                                anchor_mode, anchor_safe_buy, anchor_safe_sell,
                                best_inner_fallback=None, bh_ret=None):
    """그리드 승률 상위 후보들을 '실제 일별 백테스트'로 한 번에 재계산하고,
       실제 최대거래손실(MDD) 기준을 통과한 것 중 실제 승률 -10%p 범위에서 수익 최고를 선정.
       반환: (best_inner, inner_passed_with_real_cols)
    """
    win_tol = globals().get('WINRATE_TOLERANCE', 0.10)
    top_n   = int(globals().get('VERIFY_TOP_N', 10000))
    mdd_lim = globals().get('VERIFY_MAX_DRAWDOWN_LIMIT', None)

    cand, _note = _pick_verify_candidates(inner_passed, bh_ret=bh_ret)
    print(f"\n  🔬 실거래 검증 — {_note} → {len(cand)}개를 실제 일별 백테스트로 재계산 "
          f"(모든 변수 반복 끝난 뒤 1회)")
    try: _ZCACHE.clear()   # ★ 검증 전 z-캐시 정리 (요청)
    except Exception: pass
    real_rows = []
    _t_start = time.time()
    for _i, _r in cand.iterrows():
        _kb, _vb = int(_r['K_buy']), int(_r['vote_buy'])
        _ks, _vs = int(_r['K_sell']), int(_r['vote_sell'])
        try:
            _d, _t, _cur, _bu, _su = daily_ensemble_backtest(
                feat, close, buy_pool, sell_pool,
                K_buy=_kb, K_sell=_ks, vote_buy=_vb, vote_sell=_vs,
                cost=COST_PER_TRADE, horizon=horizon,
                dd_limit=dd_limit, ru_limit=ru_limit,
                stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
                anchor_safe_buy=anchor_safe_buy, anchor_safe_sell=anchor_safe_sell)
        except Exception:
            continue
        row = _r.to_dict()
        row['real_win_rate']      = float(_cur.get('win_rate', 0.0))
        row['real_max_drawdown']  = float(_cur.get('max_drawdown', 0.0))
        row['real_total_return']  = float(_cur.get('cum_return_pct', 0.0)) / 100.0
        row['real_n_trades']      = int(_cur.get('n_trades', 0))
        row['real_buy_success']   = float(_cur.get('buy_success_rate', 0.0))
        row['real_sell_success']  = float(_cur.get('sell_success_rate', 0.0))
        row['real_avg_success']   = float(_cur.get('avg_success_rate', 0.0))
        row['real_buy_match']     = float(_cur.get('anchor_buy_match_rate', 0.0))
        row['real_sell_match']    = float(_cur.get('anchor_sell_match_rate', 0.0))
        row['real_avg_match']     = float(_cur.get('anchor_avg_match_rate', 0.0))
        row['real_anchor_return'] = float(_cur.get('anchor_strategy_return', 0.0))
        # ★ 실행일 실제 포지션·점수 + 최근 매수/매도 정보 (요청)
        row.update(_extract_runday_info(_d, _t))
        # ★ OOS 기능 OFF (요청) — OOS 백테스트 생략, 컬럼 NaN
        if globals().get('OOS_SELECT_BY_OOS_RETURN', False):
            try:
                _oos_m = int(globals().get('OOS_MONTHS', 1))
                _last_dt = feat.index.max()
                _oos_start = _last_dt - pd.DateOffset(months=_oos_m)
                _omask = feat.index >= _oos_start
                _omask = np.asarray(_omask)
                _feat_oos = feat.loc[_omask]
                _close_oos = close.reindex(_feat_oos.index)
                _asb_o = (anchor_safe_buy[_omask] if (anchor_safe_buy is not None and len(anchor_safe_buy)==len(_omask)) else None)
                _ass_o = (anchor_safe_sell[_omask] if (anchor_safe_sell is not None and len(anchor_safe_sell)==len(_omask)) else None)
                if len(_feat_oos) >= 5:
                    _do, _to, _curo, _buo, _suo = daily_ensemble_backtest(
                        _feat_oos, _close_oos, buy_pool, sell_pool,
                        K_buy=int(_r['K_buy']), K_sell=int(_r['K_sell']),
                        vote_buy=int(_r['vote_buy']), vote_sell=int(_r['vote_sell']),
                        cost=COST_PER_TRADE, horizon=horizon,
                        dd_limit=dd_limit, ru_limit=ru_limit,
                        stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
                        anchor_safe_buy=_asb_o, anchor_safe_sell=_ass_o)
                    row['oos_total_return'] = float(_curo.get('cum_return_pct', 0.0)) / 100.0
                    row['oos_buy_success']  = float(_curo.get('buy_success_rate', 0.0))
                    row['oos_sell_success'] = float(_curo.get('sell_success_rate', 0.0))
                    row['oos_win_rate']     = float(_curo.get('win_rate', 0.0))
                    row['oos_n_trades']     = int(_curo.get('n_trades', 0))
                else:
                    row['oos_total_return'] = float('nan')
                    row['oos_buy_success'] = row['oos_sell_success'] = float('nan')
                    row['oos_win_rate'] = float('nan'); row['oos_n_trades'] = 0
            except Exception:
                row['oos_total_return'] = float('nan')
                row['oos_buy_success'] = row['oos_sell_success'] = float('nan')
                row['oos_win_rate'] = float('nan'); row['oos_n_trades'] = 0
        real_rows.append(row)
        if (_i + 1) % 1000 == 0:
            print(f"     ... {_i+1}/{len(cand)} 검증  (경과 {time.time()-_t_start:.0f}초)")

    if not real_rows:
        print(f"  ⚠ 실거래 검증 후보가 모두 실패 → 그리드 근사 기준 사용")
        return best_inner_fallback, inner_passed

    verified = pd.DataFrame(real_rows)

    # ★ 보정 통합 (요청) — 일반 경로(STAGED OFF)도 보정 적용. 단일 풀을 키 dict로 감싼다.
    if globals().get('USE_ANCHOR_MATCH_CORRECTION', True):
        try:
            _single_map = {}
            for _idx, _vr in verified.iterrows():
                _k = (round(float(_vr.get('meta_wilson_z', 0)),4),
                      int(_vr.get('meta_pct_low', 0)), int(_vr.get('meta_pct_high', 0)),
                      round(float(_vr.get('meta_corr_limit', 0)),4))
                _single_map[_k] = (buy_pool, sell_pool)
            if not _single_map:  # 메타 컬럼이 없으면 더미 키
                _single_map = {(0.0,0,0,0.0): (buy_pool, sell_pool)}
                verified['meta_wilson_z'] = 0.0; verified['meta_pct_low'] = 0
                verified['meta_pct_high'] = 0; verified['meta_corr_limit'] = 0.0
            verified = _apply_correction_to_candidates(
                verified, feat, close, _single_map,
                horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
                stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
                anchor_safe_buy=anchor_safe_buy, anchor_safe_sell=anchor_safe_sell)
        except Exception as _ce:
            print(f"  ⚠ 보정 적용 실패(일반 경로): {_ce}")

    # ★ 선정 필터 수치를 '보정 후'로 (요청) — 보정 적용된 조합은 corr_*, 아니면 real_*.
    _use_corr_sel = globals().get('SELECT_FILTER_BY_CORRECTED', True)
    def _eff_col(colc, colr):
        base = verified[colr].astype(float).values
        if _use_corr_sel and colc in verified.columns:
            _ca = verified.get('corr_applied', pd.Series(False, index=verified.index))
            _ca = _ca.fillna(False).astype(bool).values
            cval = verified[colc].astype(float).values
            return np.where(_ca & ~np.isnan(cval), cval, base)
        return base
    verified['sel_max_drawdown'] = _eff_col('corr_max_drawdown', 'real_max_drawdown')
    verified['sel_win_rate']     = _eff_col('corr_win_rate',     'real_win_rate')
    verified['sel_total_return'] = _eff_col('corr_total_return', 'real_total_return')

    # ★ 1) (보정 후) 최대 거래손실(MDD) 기준 필터 (요청)
    pool = verified
    if mdd_lim is not None:
        safe = verified[verified['sel_max_drawdown'] >= mdd_lim]
        if len(safe) > 0:
            pool = safe
            print(f"  🛡 (보정후) 최대 거래손실 {mdd_lim*100:.1f}% 이하(안전) 후보 {len(safe)}개로 선정 "
                  f"(전체 검증 {len(verified)}개 중)")
        else:
            print(f"  ⚠ (보정후) 최대 거래손실 {mdd_lim*100:.1f}% 이하 후보가 없음 "
                  f"→ 가장 손실 얕은 조합으로 선정 (기준 완화)")
            pool = verified.sort_values('sel_max_drawdown', ascending=False)

    best_inner, sel = _final_pick_by_real(pool, win_tol)
    inner_passed_out = verified.sort_values('sel_total_return', ascending=False).reset_index(drop=True)

    print(f"     실제 승률 {sel['real_win_rate']*100:.1f}%, "
          f"실제 평균성공 {sel['real_avg_success']*100:.1f}% "
          f"(매수 {sel['real_buy_success']*100:.1f}% / 매도 {sel['real_sell_success']*100:.1f}%), "
          f"실제 최대거래손실 {sel['real_max_drawdown']*100:.2f}%, "
          f"실제 누적수익 {sel['real_total_return']*100:+.2f}%")
    return best_inner, inner_passed_out


def run_ensemble_search(*, eval_start=EVAL_START,
                         horizon=HORIZON_DAYS,
                         dd_limit=DRAWDOWN_LIMIT_BUY,
                         ru_limit=RUNUP_LIMIT_SELL,
                         n_thresholds=N_THRESHOLDS,
                         max_indicators=MAX_INDICATORS,
                         k_buy_range=K_BUY_RANGE,
                         k_sell_range=K_SELL_RANGE,
                         vote_ratio_buy=VOTE_RATIO_BUY,
                         vote_ratio_sell=VOTE_RATIO_SELL,
                         min_trades_daily=MIN_TRADES_DAILY,
                         max_drawdown_limit_pct='__USE_GLOBAL__',
                         stop_loss_pct=STOP_LOSS_PCT,
                         selection_tolerance=SELECTION_TOLERANCE,
                         anchor_match_priority=ANCHOR_MATCH_PRIORITY,
                         anchor_match_tolerance=ANCHOR_MATCH_TOLERANCE,
                         anchor_mode=ANCHOR_MODE,
                         auto_anchor=AUTO_ANCHOR,
                         auto_anchor_window=AUTO_ANCHOR_WINDOW,
                         auto_anchor_lookforward=AUTO_ANCHOR_LOOKFORWARD,
                         auto_anchor_min_rise=AUTO_ANCHOR_MIN_RISE,
                         auto_anchor_min_drop=AUTO_ANCHOR_MIN_DROP,
                         auto_anchor_price_tolerance=AUTO_ANCHOR_PRICE_TOLERANCE,
                         auto_anchor_max_dates=AUTO_ANCHOR_MAX_DATES,
                         anchor_buy_dates=None,
                         anchor_sell_dates=None,
                         meta_grid=META_GRID,
                         select_by=SELECT_BY,
                         oos_enabled=OOS_ENABLED,
                         oos_start=OOS_START,
                         write_output=True,
                         output_file=None,
                         inject_combined_table=None,
                         force_best_combo=None,
                         force_corr=None,
                         inject_pools=None):
    print('=' * 72)
    print('  매수/매도 앙상블 — 메타 그리드 자동 튜닝')
    print('=' * 72)

    # ★ MDD 한도 — sentinel이면 호출 시점의 전역 MAX_DRAWDOWN_LIMIT_PCT를 다시 읽음
    #   (Colab에서 변수만 바꾸고 함수 재정의 안 해도 최신값 반영되도록)
    if max_drawdown_limit_pct == '__USE_GLOBAL__':
        max_drawdown_limit_pct = globals().get('MAX_DRAWDOWN_LIMIT_PCT', None)

    if anchor_buy_dates is None:  anchor_buy_dates  = list(ANCHOR_BUY_DATES)
    if anchor_sell_dates is None: anchor_sell_dates = list(ANCHOR_SELL_DATES)

    feat, close, ticker = _resolve_data()
    mask = feat.index >= pd.Timestamp(eval_start)
    feat = feat.loc[mask]
    close = close.reindex(feat.index)
    # ★ 거래하지 않는 날(공휴일 등 종가 NaN) 제거 — ffill로 가짜 변동 만들지 않음(요청).
    #   여기서 한 번 제거하면 anchor·grid·daily 모든 단계가 같은 길이의 거래일만 사용
    #   → 신규상장 종목(CRWV 등) 길이 불일치/broadcast 충돌 방지.
    _valid = close.notna()
    if _valid.sum() < len(close):
        _removed = len(close) - int(_valid.sum())
        feat = feat[_valid.values]
        close = close[_valid.values]
        print(f"  ℹ 거래 없는 날(종가 결측) {_removed}일 제외 → {len(close)}일")
    print(f"  기간: {feat.index[0].date()} ~ {feat.index[-1].date()}  ({len(feat)}일)")

    oos_start_idx = None
    if oos_enabled:
        oos_ts = pd.Timestamp(oos_start)
        pos_arr = np.where(feat.index >= oos_ts)[0]
        if len(pos_arr) == 0:
            print(f"  ⚠ OOS_START({oos_start})가 데이터 범위 밖 (전부 IS) → OOS 비활성화")
            oos_enabled = False
        elif pos_arr[0] == 0:
            print(f"  ⚠ OOS_START({oos_start})가 데이터 시작보다 이르거나 같음 (IS 없음) → OOS 비활성화")
            oos_enabled = False
        else:
            oos_start_idx = int(pos_arr[0])
            is_n = oos_start_idx
            oos_n = len(feat) - oos_start_idx
            print(f"  ★ OOS 검증 ON — IS(학습) {is_n}일 [{feat.index[0].date()}~{feat.index[oos_start_idx-1].date()}]  "
                  f"/ OOS(검증) {oos_n}일 [{feat.index[oos_start_idx].date()}~{feat.index[-1].date()}]")
            if is_n < 60 or oos_n < 20:
                print(f"     ⚠ IS 또는 OOS 구간이 짧음 (IS {is_n}, OOS {oos_n}) — 신뢰도 주의")
    else:
        print(f"  ★ OOS 검증: OFF (전체 기간 단일 백테스트)")

    if max_drawdown_limit_pct is not None:
        print(f"  ★ MDD 한도: -{abs(max_drawdown_limit_pct):.2f}%")
    else:
        print(f"  ★ MDD 한도: 없음")
    print(f"  ★ 최소 거래수: {min_trades_daily}회")
    if stop_loss_pct is not None and stop_loss_pct > 0:
        print(f"  ⛔ 손절매 한도: -{stop_loss_pct*100:.2f}%")
    else:
        print(f"  ⛔ 손절매: 없음")
    print(f"  ★ 충돌 해결: count/K 비율 큰 쪽 우선")
    if globals().get('USE_WEIGHTED_VOTE', False):
        print(f"  ★ 투표 방식: 가중 투표 (성공률 비례, 상한 {globals().get('WEIGHT_MAX_RATIO',1.6)}배)")
    else:
        print(f"  ★ 투표 방식: 일반 투표 (모두 1표)")
    _prio_re = globals().get('SELECTION_PRIORITY', 'balacc_return')
    if _prio_re == 'stability':
        print(f"  ★ 선정: 안정성 종합점수(매도·매수성공·수익·MDD방어 가중 기하평균) 최대")
    elif _prio_re == 'sell_buy_return':
        print(f"  ★ 선정 우선순위: 매도성공률 → 매수성공률 → 누적수익 (밴드 {selection_tolerance*100:.2f}%p)")
    elif selection_tolerance > 0:
        print(f"  ★ 선정 기준: 평균 BalAcc top - {selection_tolerance*100:.2f}%p 이내 중 수익률 최대 (Tolerance Band)")
    else:
        print(f"  ★ 선정 기준: 1차 평균 BalAcc → 2차 수익률 (strict)")
    if anchor_match_priority and anchor_mode:
        print(f"  ⚓ 매칭률 우선순위 ON — 매칭률 ±{anchor_match_tolerance*100:.1f}%p → BalAcc → 수익률")

    _anchor_prio = None
    if anchor_mode and auto_anchor:
        print(f"  ⚓ AUTO ANCHOR 자동 계산:")
        print(f"     window={auto_anchor_window}일, lookforward={auto_anchor_lookforward}일")
        print(f"     매수 정답: 좌우 {auto_anchor_window}일 local min + 그 후 +{auto_anchor_min_rise*100:.1f}% 이상 상승")
        print(f"     매도 정답: 좌우 {auto_anchor_window}일 local max + 그 후 -{auto_anchor_min_drop*100:.1f}% 이상 하락")
        print(f"     가격 근접 확장: base ±{auto_anchor_price_tolerance*100:.2f}% 이내 모든 날 포함")
        anchor_buy_dates, anchor_sell_dates, _anchor_prio = auto_compute_anchor_dates(
            feat.index, close,
            window=auto_anchor_window,
            lookforward=auto_anchor_lookforward,
            min_rise_after_buy=auto_anchor_min_rise,
            min_drop_after_sell=auto_anchor_min_drop,
            price_tolerance=auto_anchor_price_tolerance,
            max_dates=auto_anchor_max_dates,
            return_priority=True,
        )
        print(f"     ▷ 자동 계산 결과: 매수 정답 {len(anchor_buy_dates)}일, 매도 정답 {len(anchor_sell_dates)}일")
        if len(anchor_buy_dates) > 0:
            preview_b = anchor_buy_dates[:5] + (['...'] if len(anchor_buy_dates) > 5 else [])
            print(f"       매수 예시: {preview_b}")
        if len(anchor_sell_dates) > 0:
            preview_s = anchor_sell_dates[:5] + (['...'] if len(anchor_sell_dates) > 5 else [])
            print(f"       매도 예시: {preview_s}")
    elif anchor_mode:
        print(f"  ⚓ ANCHOR 수동 입력 — 매수 정답 {len(anchor_buy_dates)}일, 매도 정답 {len(anchor_sell_dates)}일")
    else:
        print(f"  ⚓ ANCHOR 보정: OFF")

    print(f"\n  메타 그리드:")
    for k, v in meta_grid.items():
        print(f"    {k:14}: {v}")

    if PREFILTER_ENABLED and inject_pools is None:
        feat_filtered = prefilter_indicators(
            feat, close, horizon=horizon,
            min_corr=PREFILTER_MIN_CORR,
            min_variance_rel=PREFILTER_MIN_VARIANCE_REL,
            max_nan_ratio=PREFILTER_MAX_NAN_RATIO,
        )
        if len(feat_filtered.columns) < 30:
            print(f"     ⚠ 필터 후 지표 {len(feat_filtered.columns)}개 < 30 → 원본 {len(feat.columns)}개 그대로 사용")
        else:
            feat = feat_filtered
    elif inject_pools is not None:
        print(f"  ♻ 재현 모드 — 사전 필터 건너뜀 (엑셀 지표 풀의 지표를 그대로 사용)")
    else:
        print(f"  ℹ 사전 필터 비활성화 — 전체 {len(feat.columns)}개 지표 사용")

    indicators = _select_indicators(feat, max_indicators)
    print(f"\n  후보 지표: {len(indicators)}개")

    anchor_safe_buy = anchor_safe_sell = None
    if anchor_mode:
        anchor_safe_buy, anchor_safe_sell = _compute_anchor_arrays(
            feat.index, anchor_buy_dates, anchor_sell_dates)
        n_b = int(anchor_safe_buy.sum())
        n_s = int(anchor_safe_sell.sum())
        if n_b == 0 and n_s == 0:
            print("  ⚠ ANCHOR 정답일이 데이터 범위 내 0개 — ANCHOR 보정 비활성화")
            anchor_mode = False
            anchor_safe_buy = anchor_safe_sell = None
        else:
            print(f"  ⚓ ANCHOR 매칭 — 매수 {n_b}일, 매도 {n_s}일 (데이터 범위 내)")

    # ★ B&H 수익률 먼저 계산 (meta_grid_search에 넘겨 B&H 미달 조합 제외용)
    close_arr = close.values.astype(np.float64)
    if oos_enabled and oos_start_idx is not None and 0 < oos_start_idx < len(close_arr):
        bh_ret_for_filter = _bh_sum_return(close_arr[:oos_start_idx])
    else:
        bh_ret_for_filter = _bh_sum_return(close_arr)

    print(f"\n[메타 그리드 탐색]   (B&H 미달 조합 제외, 기준 B&H={bh_ret_for_filter*100:+.2f}%)")
    if inject_pools is not None:
        # ★ 재현 모드 — 엑셀에서 읽은 지표 풀을 그대로 사용 (지표 재선별 안 함).
        #   meta_grid_search(지표 선별·그리드 탐색) 전체를 건너뛰고, 주입된 풀 +
        #   force_best_combo의 K/vote로 곧장 daily 백테스트 → 원본과 동일 지표·조합 재현.
        buy_pool, sell_pool = inject_pools
        print(f"  ♻ 재현 모드 — 엑셀의 지표 풀 그대로 사용 "
              f"(매수 {len(buy_pool)}개 / 매도 {len(sell_pool)}개 지표, 재선별 안 함)")
        if force_best_combo is not None:
            best_inner = {
                'K_buy': int(force_best_combo['K_buy']),
                'vote_buy': int(force_best_combo['vote_buy']),
                'K_sell': int(force_best_combo['K_sell']),
                'vote_sell': int(force_best_combo['vote_sell']),
            }
        else:
            best_inner = {'K_buy': k_buy_range[0], 'vote_buy': int(k_buy_range[0]*vote_ratio_buy[0]),
                          'K_sell': k_sell_range[0], 'vote_sell': int(k_sell_range[0]*vote_ratio_sell[0])}
        # best_meta는 주입된 메타 그리드값으로 채움(엑셀 표시용)
        _mg0 = meta_grid
        best_meta = {
            'wilson_z': _mg0['wilson_z'][0],
            'pct_low': _mg0['pct_range'][0][0], 'pct_high': _mg0['pct_range'][0][1],
            'min_signals': _mg0['min_signals'][0], 'corr_limit': _mg0['corr_limit'][0],
            'top_n_pool_buy': len(buy_pool), 'top_n_pool_sell': len(sell_pool),
        }
        meta_results_df = pd.DataFrame()
        inner_all = pd.DataFrame()
        inner_passed = inject_combined_table if inject_combined_table is not None else pd.DataFrame()
    else:
        meta_results_df, inner_all, inner_passed, best_meta, best_inner, buy_pool, sell_pool = meta_grid_search(
            feat, close,
            meta_grid=meta_grid, indicators=indicators,
            n_thresholds=n_thresholds, horizon=horizon,
            dd_limit=dd_limit, ru_limit=ru_limit,
            k_buy_range=k_buy_range, k_sell_range=k_sell_range,
            vote_ratio_buy=vote_ratio_buy, vote_ratio_sell=vote_ratio_sell,
            cost=COST_PER_TRADE,
            min_trades_daily=min_trades_daily,
            mdd_limit_pct=max_drawdown_limit_pct,
            select_by=select_by,
            stop_loss_pct=stop_loss_pct,
            selection_tolerance=selection_tolerance,
            anchor_match_priority=anchor_match_priority,
            anchor_match_tolerance=anchor_match_tolerance,
            anchor_mode=anchor_mode,
            anchor_safe_buy=anchor_safe_buy,
            anchor_safe_sell=anchor_safe_sell,
            oos_start_idx=oos_start_idx if oos_enabled else None,
            bh_ret=bh_ret_for_filter,
            exclude_below_bh=globals().get('EXCLUDE_BELOW_BH', True),
        )

    # ★ staged가 통합 테이블에서 고른 정확한 조합을 강제 (현재 포지션=★1등 일치 보장).
    #   meta_grid_search가 자체 선정한 best_inner와 staged의 최종 선택이 어긋나는 것을 방지.
    #   (inject_pools 재현 모드에서는 위에서 이미 best_inner를 force값으로 설정했으므로 건너뜀)
    if force_best_combo is not None and inject_pools is None:
        fb_kb = int(force_best_combo['K_buy']); fb_vb = int(force_best_combo['vote_buy'])
        fb_ks = int(force_best_combo['K_sell']); fb_vs = int(force_best_combo['vote_sell'])
        src_tbl = inject_combined_table if inject_combined_table is not None else inner_passed
        matched = None
        if src_tbl is not None and len(src_tbl) > 0:
            m = src_tbl[(src_tbl['K_buy']==fb_kb) & (src_tbl['vote_buy']==fb_vb) &
                        (src_tbl['K_sell']==fb_ks) & (src_tbl['vote_sell']==fb_vs)]
            if len(m) > 0:
                matched = m.iloc[0].to_dict()
        if matched is not None:
            best_inner = matched
            print(f"  ★ 최종 조합 강제 적용: K_buy={fb_kb}/v{fb_vb}, K_sell={fb_ks}/v{fb_vs} "
                  f"(현재 포지션이 ★1등과 동일하게 계산됨)")
        else:
            print(f"  ⚠ force_best_combo({fb_kb}/{fb_vb}, {fb_ks}/{fb_vs})를 테이블에서 못 찾음 — meta 자체 best 사용")

    bh_ret  = _bh_sum_return(close_arr)
    bh_up_ret = _bh_up_sum_return(close_arr)
    bh_cagr = (1 + bh_ret) ** (252 / len(close_arr)) - 1 if (1 + bh_ret) > 0 else bh_ret

    # ★ 승률 후보 실거래 검증 (요청) — 그리드는 빠른 근사라 실제 일별거래와 차이날 수 있음.
    #   1) 그리드 '승률' 상위 N개(기본 10000)를 골라 실제 일별 백테스트로 재계산
    #   2) 실제 승률 최고에서 -10%p 범위로 후보를 잡고
    #   3) 그 안에서 실제 수익률 높은 순으로 정렬 → 1등을 최종 선정
    #   (force_best_combo / inject_pools 모드는 이미 조합이 정해졌으므로 건너뜀)
    # ★ 실거래 검증은 '변수 반복마다' 하지 않는다 (요청).
    #   - staged 단계별 호출: write_output=False & inject_combined_table=None → 검증 안 함
    #   - 단독 분석(staged 아님): write_output=True & inject_combined_table=None → 검증 함
    #   - staged 최종 호출: staged가 이미 merged_table에서 실거래 검증을 끝내고
    #     force_best_combo로 그 결과를 넘기므로 여기선 안 함.
    _is_standalone = (write_output and inject_combined_table is None)
    _verify = (globals().get('VERIFY_BY_DAILY_BACKTEST', False)
               and force_best_combo is None and inject_pools is None
               and _is_standalone
               and inner_passed is not None and len(inner_passed) > 0)
    if _verify:
        best_inner, inner_passed = _verify_candidates_by_daily(
            inner_passed, feat, close, buy_pool, sell_pool,
            horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
            stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
            anchor_safe_buy=anchor_safe_buy, anchor_safe_sell=anchor_safe_sell,
            best_inner_fallback=best_inner, bh_ret=bh_ret)

    if inject_pools is None:
        print(f"\n  ─ 메타 그리드 Top 10 ─")
        print(f"  {'#':>2} {'w_z':>5}  {'pct':>8}  {'min_s':>5} {'corr':>5} {'pool':>4}  "
              f"{'평균성공':>8} {'매수':>6} {'매도':>6}  {'수익%':>8}  {'MDD%':>7}  {'거래':>4}")
        for i, r in meta_results_df.head(10).iterrows():
            ret_s = f"{r['best_return']*100:+.2f}" if pd.notna(r['best_return']) else "  —  "
            mdd_s = f"{r['best_mdd']*100:.2f}" if pd.notna(r['best_mdd']) else "  —  "
            nt_s  = f"{int(r['best_n_trades'])}" if pd.notna(r['best_mdd']) else "—"
            avg_s = f"{r['best_avg_sr']*100:.1f}%"  if pd.notna(r['best_avg_sr'])  else "  —  "
            b_s   = f"{r['best_buy_sr']*100:.1f}%"  if pd.notna(r['best_buy_sr'])  else "  —  "
            s_s   = f"{r['best_sell_sr']*100:.1f}%" if pd.notna(r['best_sell_sr']) else "  —  "
            print(f"  {i+1:>2}  {r['wilson_z']:>5}  "
                  f"({int(r['pct_low']):>2},{int(r['pct_high']):>2})  "
                  f"{int(r['min_signals']):>5}  {r['corr_limit']:>5}  "
                  f"{int(r['top_n_pool_buy']):>4}  "
                  f"{avg_s:>8} {b_s:>6} {s_s:>6}  "
                  f"{ret_s:>8}  {mdd_s:>7}  {nt_s:>4}")

        print(f"\n  ★ 1등 메타: WILSON_Z={best_meta['wilson_z']}, "
              f"PCT=({best_meta['pct_low']:.0f},{best_meta['pct_high']:.0f}), "
              f"MIN_SIG={best_meta['min_signals']}, CORR={best_meta['corr_limit']}, "
              f"POOL={best_meta['top_n_pool_buy']}")
        print(f"  ★ 1등 앙상블: K_buy={int(best_inner['K_buy'])}/v={int(best_inner['vote_buy'])}, "
              f"K_sell={int(best_inner['K_sell'])}/v={int(best_inner['vote_sell'])}")
        print(f"     평균성공 {best_inner['avg_success_rate']*100:.1f}% "
              f"(매수 {best_inner['buy_success_rate']*100:.1f}% / 매도 {best_inner['sell_success_rate']*100:.1f}%)")
        has_match = ('anchor_avg_match_rate' in best_inner)
        if has_match:
            match_v = best_inner.get('anchor_avg_match_rate', np.nan) if hasattr(best_inner, 'get') else best_inner['anchor_avg_match_rate']
            if pd.notna(match_v):
                print(f"     ⚓ 매칭률 {match_v*100:.1f}%")
        print(f"     누적 {best_inner['total_return']*100:+.2f}%  vs B&H {bh_ret*100:+.2f}% "
              f"({(best_inner['total_return']-bh_ret)*100:+.2f}%p)")
        print(f"     MDD {best_inner['max_drawdown']*100:.2f}%, "
              f"거래 {int(best_inner['n_trades'])}회, 승률 {best_inner['win_rate']*100:.1f}%")
    else:
        print(f"  ★ 재현 조합: K_buy={int(best_inner['K_buy'])}/v={int(best_inner['vote_buy'])}, "
              f"K_sell={int(best_inner['K_sell'])}/v={int(best_inner['vote_sell'])} "
              f"(엑셀 지표 풀 그대로, 일별 백테스트로 재현)")

    if len(inner_passed) > 0:
        cand = inner_passed.copy()
        if anchor_match_priority and anchor_mode and 'anchor_avg_match_rate' in cand.columns:
            top_m = cand['anchor_avg_match_rate'].max()
            cand = cand[cand['anchor_avg_match_rate'] >= top_m - anchor_match_tolerance]
        if selection_tolerance > 0:
            top_b = cand['avg_success_rate'].max()
            cand_band = cand[cand['avg_success_rate'] >= top_b - selection_tolerance]
        else:
            cand_band = cand
        if len(cand_band) > 1:
            top_ret_in_band = cand_band.sort_values('total_return', ascending=False).head(5)
            print(f"\n  ─ 밴드 안 수익률 상위 5개 (참고용, 모두 동등하게 좋은 후보) ─")
            print(f"  {'BalAcc':>7} {'매칭률':>7} {'수익률':>10} {'MDD':>8} {'K_b/v_b':>9} {'K_s/v_s':>9} {'거래':>4}")
            for _, r in top_ret_in_band.iterrows():
                m_str = f"{r['anchor_avg_match_rate']*100:.1f}%" if 'anchor_avg_match_rate' in r.index else '—'
                is_selected_str = ' ★' if (int(r['K_buy'])==int(best_inner['K_buy']) and
                                            int(r['K_sell'])==int(best_inner['K_sell']) and
                                            int(r['vote_buy'])==int(best_inner['vote_buy']) and
                                            int(r['vote_sell'])==int(best_inner['vote_sell'])) else ''
                print(f"  {r['avg_success_rate']*100:>6.1f}% {m_str:>7} {r['total_return']*100:>+8.2f}% "
                      f"{r['max_drawdown']*100:>+6.2f}% "
                      f"{int(r['K_buy']):>4}/{int(r['vote_buy']):<3} {int(r['K_sell']):>4}/{int(r['vote_sell']):<3} "
                      f"{int(r['n_trades']):>4}{is_selected_str}")
            print(f"  ※ 수익률 비중을 더 두고 싶다면 SELECTION_TOLERANCE를 키우세요 (현재 {selection_tolerance*100:.1f}%p)")

    print(f"\n[일별 백테스트]")
    # ★ 방어 (요청) — 검증 루프에서 쌓인 z-스코어 캐시를 비우고 최종 백테스트를 돌린다.
    #   (키는 이미 내용 기반이라 stale 위험은 없지만, 검증=최종 일치를 한 번 더 보장)
    try: _ZCACHE.clear()
    except Exception: pass
    # ★ 보정 채택된 조합이면 보정된 풀+가중치로 백테스트 → 일별/거래내역/현재포지션 모두 보정 반영 (요청)
    _bt_buy_pool, _bt_sell_pool = buy_pool, sell_pool
    _bt_Kb = int(best_inner['K_buy']); _bt_Ks = int(best_inner['K_sell'])
    _bt_bw = _bt_sw = None
    if force_corr is not None and force_corr.get('buy_pool') is not None:
        _bt_buy_pool = force_corr['buy_pool']; _bt_sell_pool = force_corr['sell_pool']
        _bt_Kb = len(_bt_buy_pool); _bt_Ks = len(_bt_sell_pool)
        _bt_bw = force_corr.get('buy_w'); _bt_sw = force_corr.get('sell_w')
        print(f"  🔧 보정 적용된 풀로 백테스트 — 매수 {_bt_Kb}개(추가 {force_corr.get('n_added_buy',0)}), "
              f"매도 {_bt_Ks}개(추가 {force_corr.get('n_added_sell',0)}) + 가중치 보정")
    daily, trades, cur, buy_used, sell_used = daily_ensemble_backtest(
        feat, close, _bt_buy_pool, _bt_sell_pool,
        K_buy=_bt_Kb,
        K_sell=_bt_Ks,
        vote_buy=int(best_inner['vote_buy']),
        vote_sell=int(best_inner['vote_sell']),
        horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
        stop_loss_pct=stop_loss_pct,
        anchor_mode=anchor_mode,
        anchor_safe_buy=anchor_safe_buy,
        anchor_safe_sell=anchor_safe_sell,
        buy_w_override=_bt_bw, sell_w_override=_bt_sw,
    )

    oos_daily = oos_trades = oos_cur = None
    if oos_enabled and oos_start_idx is not None:
        feat_oos  = feat.iloc[oos_start_idx:]
        close_oos = close.iloc[oos_start_idx:]
        try:
            oos_daily, oos_trades, oos_cur, _bu, _su = daily_ensemble_backtest(
                feat_oos, close_oos, buy_pool, sell_pool,
                K_buy=int(best_inner['K_buy']),
                K_sell=int(best_inner['K_sell']),
                vote_buy=int(best_inner['vote_buy']),
                vote_sell=int(best_inner['vote_sell']),
                horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
                stop_loss_pct=stop_loss_pct,
                anchor_mode=False,
                anchor_safe_buy=None, anchor_safe_sell=None,
            )
            oos_bh = _bh_sum_return(close_oos.values) * 100
            oos_bh_up = _bh_up_sum_return(close_oos.values) * 100
            print()
            print('  ' + '─' * 68)
            print(f'  🔬 OOS 검증 결과 ({close_oos.index[0].date()} ~ {close_oos.index[-1].date()}, {len(close_oos)}일)')
            print(f'     OOS 누적 {oos_cur["cum_return_pct"]:+.2f}%  vs B&H {oos_bh:+.2f}%  '
                  f'거래 {oos_cur["n_trades"]}회  승률 {oos_cur["win_rate"]*100:.1f}%')
            print(f'     📈 상승일만: 전략 {oos_cur["up_cum_return_pct"]:+.2f}%  vs B&H {oos_bh_up:+.2f}%')
            print(f'     ※ IS 누적 {cur["cum_return_pct"]:+.2f}% 대비 OOS가 비슷하면 신뢰↑, 크게 무너지면 과적합 의심')
        except Exception as e:
            print(f'  ⚠ OOS 일별 백테스트 실패: {e}')
            oos_daily = oos_trades = oos_cur = None

    print()
    print('  ' + '─' * 68)
    print(f'  📅  {ticker}  현재 포지션  ({cur["last_date"].date()})')
    print('  ' + '─' * 68)
    print(f'    {cur["position_emoji"]}  {cur["position"]}')
    if '보유' in cur['position']:
        print(f'    진입일: {cur["entry_date"].date()}   진입가: ${cur["entry_price"]:.2f}   '
              f'현재가: ${cur["current_price"]:.2f}')
        print(f'    보유일: {cur["days_held"]}일   미실현: {cur["unrealized_pct"]:+.2f}%')
        if cur.get('stop_price'):
            print(f'    ⛔ 손절가: ${cur["stop_price"]:.2f} '
                  f'(현재가 대비 +{(cur["current_price"]/cur["stop_price"]-1)*100:.2f}% 여유)')
    else:
        print(f'    현재가: ${cur["current_price"]:.2f}')
    print('  ' + '─' * 68)
    print(f'    누적 {cur["cum_return_pct"]:+.2f}%   B&H {bh_ret*100:+.2f}%   '
          f'MDD {cur["max_drawdown"]*100:.2f}%   거래 {cur["n_trades"]}회 (손절매 {cur["n_stop_triggered"]}회)')
    print(f'    📈 상승일만 합산: 전략 {cur["up_cum_return_pct"]:+.2f}%   B&H {bh_up_ret*100:+.2f}%  '
          f'(보유 중 양(+)의 일별 변동률만 합산)')
    # ★ 성공/실패 집계 출력 (first-touch, 정답인 날만 평가)
    print(f'    ✅ [ON만 적중률] 매수 {cur.get("n_buy_success_cnt",0)}/{cur.get("n_buy_on_total",0)} '
          f'({cur.get("buy_signal_hit_rate",0)*100:.1f}%)   '
          f'매도 {cur.get("n_sell_success_cnt",0)}/{cur.get("n_sell_on_total",0)} '
          f'({cur.get("sell_signal_hit_rate",0)*100:.1f}%)')
    print(f'    ✅ [정답일 적중률] 매수 {cur.get("buy_acc_all",0)*100:.1f}% '
          f'({cur.get("n_buy_correct_all",0)}/{cur.get("n_buy_eval_all",0)}일)   '
          f'매도 {cur.get("sell_acc_all",0)*100:.1f}% '
          f'({cur.get("n_sell_correct_all",0)}/{cur.get("n_sell_eval_all",0)}일)  '
          f'※ 올라야/내려야 했던 날만 분모, 놓침=실패 (무임승차 제외)')
    succ_label = ' (⚓ANCHOR 보정)' if anchor_mode else ''
    print(f'    매수신호 BalAcc{succ_label} {cur["buy_success_rate"]*100:.1f}% (plain {cur["buy_accuracy_plain"]*100:.1f}%, ON {cur["n_buy_signal_on"]})  '
          f'매도신호 BalAcc {cur["sell_success_rate"]*100:.1f}% (plain {cur["sell_accuracy_plain"]*100:.1f}%, ON {cur["n_sell_signal_on"]})  '
          f'평균 BalAcc {cur["avg_success_rate"]*100:.1f}%  /  충돌 {cur["n_conflicts"]}일')
    if anchor_mode:
        b_rate = (cur['n_anchor_buy_caught']/cur['n_anchor_buy']*100) if cur['n_anchor_buy']>0 else 0
        s_rate = (cur['n_anchor_sell_caught']/cur['n_anchor_sell']*100) if cur['n_anchor_sell']>0 else 0
        print(f'    ⚓ 정답일 매칭: 매수 {cur["n_anchor_buy_caught"]}/{cur["n_anchor_buy"]} ({b_rate:.1f}%)   '
              f'매도 {cur["n_anchor_sell_caught"]}/{cur["n_anchor_sell"]} ({s_rate:.1f}%)')
    print('  ' + '─' * 68)
    b_str_now = cur['buy_count_now']  / cur['K_buy']  if cur['K_buy']  > 0 else 0
    s_str_now = cur['sell_count_now'] / cur['K_sell'] if cur['K_sell'] > 0 else 0
    print(f'    🟢 매수: {cur["buy_count_now"]:.2f}/{cur["K_buy"]} ON  (필요 {cur["vote_buy"]}, 강도 {b_str_now:.0%})  '
          f'→ {"ON ✓" if cur["buy_on_now"] else "OFF"}')
    print(f'    🔴 매도: {cur["sell_count_now"]:.2f}/{cur["K_sell"]} ON  (필요 {cur["vote_sell"]}, 강도 {s_str_now:.0%})  '
          f'→ {"ON ✓" if cur["sell_on_now"] else "OFF"}')
    if cur['buy_on_now'] and cur['sell_on_now']:
        winner = "매수" if b_str_now >= s_str_now else "매도"
        print(f'    ⚔ 충돌 발생 → 강도 더 큰 [{winner}] 우세')
    print('  ' + '─' * 68)

    if not write_output:
        print('  ⏩ (중간 탐색 — Excel 저장 생략)\n')
        print('=' * 72)
        return (meta_results_df, inner_all, inner_passed,
                best_meta, best_inner, buy_pool, sell_pool,
                daily, trades, cur)

    if output_file is None:
        today_str = datetime.now().strftime('%Y-%m-%d')
        output_file = os.path.join(SCRIPT_DIR, f'ensemble_search_{ticker}_{today_str}.xlsx')
    print(f"\n  Excel 저장: {output_file}")
    # ★ 데이터 스냅샷 저장 (요청) — 재현 정확도용. 재현 때 이 데이터를 그대로 쓰면
    #   FRED 수정·vintage·하루 더 받음 등으로 외부데이터가 달라지는 문제가 사라진다.
    #   (재현 실행(inject_pools 있음)에서는 저장 안 함 — 원본 분석에서만)
    if globals().get('SAVE_DATA_SNAPSHOT', True) and inject_pools is None:
        try:
            _snap = os.path.splitext(output_file)[0] + '_data.pkl'
            pd.to_pickle({'feat': feat, 'close': close, 'ticker': ticker}, _snap)
            print(f"  💾 데이터 스냅샷 저장: {os.path.basename(_snap)} "
                  f"({len(close)}일, 마지막 {str(close.index[-1])[:10]}) — 재현 정확도용")
        except Exception as _se:
            print(f"  ⚠ 데이터 스냅샷 저장 실패(무시): {_se}")
    _grid_table_for_excel = inject_combined_table if inject_combined_table is not None else inner_passed
    # ★ 보증 (요청) — 최종 백테스트 cur로 ★(선정) 행의 실제*/sel_*(/보정후*)를 덮어써서,
    #   그리드 선정 행이 '일별 백테스트 시트와 정확히 같은 수치'가 되게 한다.
    #   (검증 백테스트와 최종 백테스트가 어떤 이유로든 어긋나도, 표시·기록은 최종 기준으로 통일)
    try:
        _gt = _grid_table_for_excel
        if _gt is not None and len(_gt) > 0 and best_inner is not None and \
           all(k in _gt.columns for k in ('K_buy','vote_buy','K_sell','vote_sell')):
            _m = ((_gt['K_buy'].astype(float).round().astype(int) == int(best_inner['K_buy'])) &
                  (_gt['vote_buy'].astype(float).round().astype(int) == int(best_inner['vote_buy'])) &
                  (_gt['K_sell'].astype(float).round().astype(int) == int(best_inner['K_sell'])) &
                  (_gt['vote_sell'].astype(float).round().astype(int) == int(best_inner['vote_sell'])))
            if _m.any():
                _fin = {
                    'real_win_rate':      float(cur.get('win_rate', 0.0)),
                    'real_max_drawdown':  float(cur.get('max_drawdown', 0.0)),
                    'real_total_return':  float(cur.get('cum_return_pct', 0.0)) / 100.0,
                    'real_n_trades':      int(cur.get('n_trades', 0)),
                    'real_buy_success':   float(cur.get('buy_success_rate', 0.0)),
                    'real_sell_success':  float(cur.get('sell_success_rate', 0.0)),
                    'real_avg_success':   float(cur.get('avg_success_rate', 0.0)),
                    'real_buy_match':     float(cur.get('anchor_buy_match_rate', 0.0)),
                    'real_sell_match':    float(cur.get('anchor_sell_match_rate', 0.0)),
                    'real_avg_match':     float(cur.get('anchor_avg_match_rate', 0.0)),
                    'real_anchor_return': float(cur.get('anchor_strategy_return', 0.0)),
                }
                for _c, _v in _fin.items():
                    if _c in _gt.columns:
                        _gt.loc[_m, _c] = _v
                for _sc, _k in [('sel_win_rate', 'real_win_rate'),
                                ('sel_max_drawdown', 'real_max_drawdown'),
                                ('sel_total_return', 'real_total_return')]:
                    if _sc in _gt.columns:
                        _gt.loc[_m, _sc] = _fin[_k]
                _corrected = force_corr is not None and force_corr.get('buy_pool') is not None
                if _corrected:
                    for _cc, _k in [('corr_total_return', 'real_total_return'),
                                    ('corr_max_drawdown', 'real_max_drawdown'),
                                    ('corr_win_rate', 'real_win_rate'),
                                    ('corr_buy_success', 'real_buy_success'),
                                    ('corr_sell_success', 'real_sell_success'),
                                    ('corr_buy_match', 'real_buy_match'),
                                    ('corr_sell_match', 'real_sell_match')]:
                        if _cc in _gt.columns:
                            _gt.loc[_m, _cc] = _fin[_k]
                # 실행일 포지션·액션·최근매매도 최종 일별/거래 기준으로 갱신
                try:
                    for _rk, _rv in _extract_runday_info(daily, trades).items():
                        if _rk in _gt.columns:
                            _gt.loc[_m, _rk] = _rv
                except Exception:
                    pass
                print(f"  🔗 선정(★) 행 수치를 최종 일별 백테스트와 동기화 "
                      f"(승률 {_fin['real_win_rate']*100:.1f}%, 최대손실 {_fin['real_max_drawdown']*100:.2f}%, "
                      f"수익 {_fin['real_total_return']*100:+.2f}%)")
    except Exception as _sync_e:
        print(f"  ⚠ 선정행 최종수치 동기화 실패(무시): {_sync_e}")
    write_excel(meta_results_df, inner_all, _grid_table_for_excel,
                best_meta, best_inner, buy_pool, sell_pool,
                daily, trades, cur, buy_used, sell_used, bh_ret, bh_cagr,
                ticker=ticker, output_file=output_file,
                horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
                mdd_limit_pct=max_drawdown_limit_pct,
                min_trades_daily=min_trades_daily,
                stop_loss_pct=stop_loss_pct,
                selection_tolerance=selection_tolerance,
                anchor_match_priority_arg=anchor_match_priority,
                anchor_match_tolerance_arg=anchor_match_tolerance,
                anchor_mode=anchor_mode,
                anchor_buy_dates=anchor_buy_dates if anchor_mode else None,
                anchor_sell_dates=anchor_sell_dates if anchor_mode else None,
                auto_anchor=auto_anchor,
                oos_enabled=oos_enabled,
                oos_daily=oos_daily, oos_trades=oos_trades, oos_cur=oos_cur,
                bh_up_ret=bh_up_ret, anchor_prio=_anchor_prio,
                feat=feat, close_full=close)
    print('  ✓ 완료\n')
    print('=' * 72)

    import inspect
    caller_frame = inspect.stack()[1]
    caller_name = caller_frame.function if caller_frame else ''
    if caller_name not in ('run_multi_ticker_analysis', 'staged_meta_tune', 'replay_grid_combo') and AUTO_DOWNLOAD_EXCEL:
        _auto_download_excels([output_file])

    return (meta_results_df, inner_all, inner_passed,
            best_meta, best_inner, buy_pool, sell_pool,
            daily, trades, cur)


def staged_meta_tune(*, base_meta_grid=None,
                      stage_pct_range=None,
                      stage_wilson_z=None,
                      stage_wilson_refine_step=None,
                      stage_corr_limit=None,
                      selection_tolerance=None,
                      **run_kwargs):
    """
    ★ 단계적 메타 변수 자동 튜닝 (중복 실행 제거 버전).
    각 단계(pct_range / wilson_z / corr_limit)에서 후보를 실행하되,
    이미 돌린 (wz,pct,corr) 조합은 캐시에서 꺼내 재실행하지 않는다.
    선정 기준: 안정성 종합 점수(매도·매수성공·수익·MDD방어 가중 기하평균) 최대.
    ★ 최종 결과는 '다시 안 돌리고' 단계에서 이미 돌린 best 조합의 결과를 그대로 쓰고,
       Excel만 그 결과로 1회 저장한다 (중복 실행 0).
    """
    if base_meta_grid is None:            base_meta_grid = META_GRID
    if stage_pct_range is None:           stage_pct_range = STAGE_PCT_RANGE
    if stage_wilson_z is None:            stage_wilson_z = STAGE_WILSON_Z
    if stage_wilson_refine_step is None:  stage_wilson_refine_step = STAGE_WILSON_REFINE_STEP
    if stage_corr_limit is None:          stage_corr_limit = STAGE_CORR_LIMIT
    if selection_tolerance is None:       selection_tolerance = SELECTION_TOLERANCE

    base_ms   = base_meta_grid.get('min_signals', [10])[0]
    base_pool = base_meta_grid.get('top_n_pool', [150])[0]
    base_wz   = base_meta_grid.get('wilson_z', [1.7])[0]
    base_pct  = base_meta_grid.get('pct_range', [(10, 90)])[0]
    base_corr = base_meta_grid.get('corr_limit', [0.8])[0]

    output_file = run_kwargs.pop('output_file', None)

    import numpy as _np
    _swt = globals().get('STABILITY_WEIGHTS', (0.30, 0.25, 0.20, 0.25))
    _swt = _np.asarray(_swt, dtype=float); _swt = _swt / _swt.sum()

    def _mk_grid(wz, pct, corr):
        return {'wilson_z': [wz], 'pct_range': [pct], 'min_signals': [base_ms],
                'corr_limit': [corr], 'top_n_pool': [base_pool]}

    # ★ 결과 캐시 — 같은 (wz,pct,corr)는 한 번만 실행
    _cache = {}
    def _run_cached(wz, pct, corr):
        """캐시에 있으면 재사용, 없으면 실행(Excel 저장 안 함).
           반환: dict(...) 또는 None(통과 조합 0개 → 스킵).
           ★ 통과 조합이 없어 meta_grid_search가 에러를 내면, 멈추지 말고
             이 조합만 스킵(None)하고 나머지 단계는 계속 진행한다."""
        key = (round(float(wz), 6), tuple(pct), round(float(corr), 6))
        if key in _cache:
            return _cache[key]
        try:
            res = run_ensemble_search(
                meta_grid=_mk_grid(wz, pct, corr),
                write_output=False, output_file=None,
                **run_kwargs)
        except RuntimeError as _e:
            # MDD 한도/거래수/B&H 미달 등으로 통과 조합이 하나도 없는 경우
            print(f'    ⚠ (wz={wz}, pct={pct}, corr={corr}) 통과 조합 없음 → 이 조합 스킵')
            _cache[key] = None
            return None
        cur = res[-1]
        # res[2] = 이 메타조합의 통합 그리드 테이블(메타변수·B&H필터 포함)
        combined_tbl = res[2] if len(res) > 2 else None
        rec = {
            'sell': float(cur.get('sell_success_rate', 0.0)),
            'buy':  float(cur.get('buy_success_rate', 0.0)),
            'avg':  float(cur.get('avg_success_rate', 0.0)),
            'ret':  float(cur.get('cum_return_pct', float('-inf'))),
            'mdd':  float(cur.get('max_drawdown', -1.0)),
            'win':  float(cur.get('win_rate', 0.0)),
            'res':  res, 'wz': wz, 'pct': pct, 'corr': corr,
            'combined': combined_tbl,
        }
        _cache[key] = rec
        return rec

    _prio_st = globals().get('SELECTION_PRIORITY', 'balacc_return')

    def _pick_best(recs):
        """recs: list of dict (None은 스킵된 조합 → 제외).
           independent: (매수성공+매도성공) 합 최대(동률시 수익).
           그 외(stability): 안정성 종합점수 최대."""
        recs = [r for r in recs if r is not None]   # ★ 스킵된 조합 제외
        if not recs: return None
        if len(recs) == 1: return recs[0]
        tol = globals().get('SELECTION_TOLERANCE', 0.04)
        # ★ 일별거래 승률 밴드(-10%p) → 누적수익 최고 — 요청
        if _prio_st == 'winrate_return':
            wtol = globals().get('WINRATE_TOLERANCE', 0.10)
            best_win = max(r['win'] for r in recs)
            b1 = [r for r in recs if r['win'] >= best_win - wtol]
            b1.sort(key=lambda r: (r['ret'], r['win']), reverse=True)
            return b1[0]
        # ★ 매도성공 밴드(-2%p) → MDD최저 밴드(-1%p) → 누적수익 최고 — 요청, 기본
        if _prio_st == 'sell_mdd_return':
            stol = globals().get('SELL_SUCCESS_TOLERANCE', 0.02)
            mtol = globals().get('MDD_TOLERANCE', 0.01)
            best_sell = max(r['sell'] for r in recs)
            b1 = [r for r in recs if r['sell'] >= best_sell - stol]
            best_mdd = max(r['mdd'] for r in b1)
            b2 = [r for r in b1 if r['mdd'] >= best_mdd - mtol]
            b2.sort(key=lambda r: (r['ret'], r['sell']), reverse=True)
            return b2[0]
        # ★ 평균성공 밴드 → MDD최저(0.1%p 동률시 수익) — 요청
        if _prio_st == 'avgband_mdd_return':
            best_avg = max(r['avg'] for r in recs)
            band = [r for r in recs if r['avg'] >= best_avg - tol]
            best_mdd_r = max(round(r['mdd'], 3) for r in band)   # 덜 빠진 쪽
            grp = [r for r in band if round(r['mdd'], 3) == best_mdd_r]
            grp.sort(key=lambda r: (r['ret'], r['avg']), reverse=True)
            return grp[0]
        # ★ 독립 최적화 — 매수성공+매도성공 합이 최대인 변수조합
        if _prio_st == 'independent':
            best = None; best_key = None
            for r in recs:
                key = (r['buy'] + r['sell'], r['ret'])
                if best_key is None or key > best_key:
                    best_key = key; best = r
            return best
        rets = [r['ret'] for r in recs]; mdds = [r['mdd'] for r in recs]
        ret_lo, ret_hi = min(rets), max(rets)
        mdd_worst, mdd_best = min(mdds), max(mdds)
        def _nrm(x, lo, hi):
            if hi - lo < 1e-12: return 0.5
            return min(1.0, max(0.0, (x - lo) / (hi - lo)))
        eps = 1e-6
        best = None; best_sc = -1.0
        for r in recs:
            s = min(1.0, max(0.0, r['sell'])); b = min(1.0, max(0.0, r['buy']))
            rr = _nrm(r['ret'], ret_lo, ret_hi); m = _nrm(r['mdd'], mdd_worst, mdd_best)
            comps = (max(s,eps), max(b,eps), max(rr,eps), max(m,eps))
            sc = float(_np.prod([c ** wi for c, wi in zip(comps, _swt)]))
            r['_score'] = sc
            if sc > best_sc:
                best_sc = sc; best = r
        return best

    def _fmt(r):
        if r is None:
            return "(통과 조합 없음 — 스킵)"
        return (f"매도 {r['sell']*100:.1f}% / 매수 {r['buy']*100:.1f}% / 평균 {r['avg']*100:.1f}% "
                f"/ 수익 {r['ret']:+.2f}% / MDD {r['mdd']*100:.2f}%")

    print('\n' + '█' * 72)
    print('  ★★★  단계적 메타 변수 자동 튜닝 시작 (중복 실행 없음)  ★★★')
    _ps = globals().get('SELECTION_PRIORITY','')
    if _ps == 'winrate_return':
        _wt=globals().get('WINRATE_TOLERANCE',0.10)
        print(f'  ★ 선정: 일별거래 승률 -{_wt*100:.0f}%p 밴드 → 누적수익 최고')
    elif _ps == 'sell_mdd_return':
        _st=globals().get('SELL_SUCCESS_TOLERANCE',0.02); _mt=globals().get('MDD_TOLERANCE',0.01)
        print(f'  ★ 선정: 매도성공률 -{_st*100:.0f}%p 밴드 → MDD 최저 -{_mt*100:.0f}%p 밴드 → 누적수익 최고')
    elif _ps == 'avgband_mdd_return':
        print(f'  ★ 선정: 평균성공률 top band 내 → MDD 최저(동률시 수익 최대)')
    elif _ps == 'independent':
        print(f'  ★ 선정: 매수성공률 최고 + 매도성공률 최고 설정 따로 찾아 합침 (독립 최적화)')
    else:
        print(f'  ★ 선정: 안정성 종합점수(매도·매수성공·수익·MDD방어) 최대')
    print(f'    1단계 pct_range {len(stage_pct_range)}개 → '
          f'2단계 wilson_z {len(stage_wilson_z)}개(+재확인) → '
          f'3단계 corr_limit {len(stage_corr_limit)}개')
    print('█' * 72)

    # ─── 1단계: pct_range ───
    print(f'\n┌─ [1단계/3] pct_range — 후보 {len(stage_pct_range)}개 '
          f'(wilson_z={base_wz}, corr_limit={base_corr}) ─┐')
    recs1 = []
    for i, pct in enumerate(stage_pct_range, 1):
        print(f'│  ▸ [1단계] {i}/{len(stage_pct_range)} pct={pct} ...')
        r = _run_cached(base_wz, pct, base_corr)
        recs1.append(r)
        print(f'│    완료 — {_fmt(r)}')
    best1 = _pick_best(recs1)
    if best1 is None:
        best_pct = base_pct
        print(f'└─ [1단계] 통과 조합 없음 → 시작 pct_range={base_pct} 유지하고 계속 ─┘')
    else:
        best_pct = best1['pct']
        print(f'└─ [1단계 완료] best pct_range = {best_pct}  ({_fmt(best1)}) ─┘')

    # ─── 2단계: wilson_z (best_pct 고정) ───
    print(f'\n┌─ [2단계/3] wilson_z — 후보 {len(stage_wilson_z)}개 (pct={best_pct}) ─┐')
    recs2 = []
    for i, wz in enumerate(stage_wilson_z, 1):
        dup = ' (캐시 재사용)' if (round(float(wz),6),tuple(best_pct),round(float(base_corr),6)) in _cache else ''
        print(f'│  ▸ [2단계] {i}/{len(stage_wilson_z)} wz={wz}{dup} ...')
        r = _run_cached(wz, best_pct, base_corr)
        recs2.append(r)
        print(f'│    완료 — {_fmt(r)}')
    best2 = _pick_best(recs2)
    if best2 is None:
        best_wz = base_wz
        print(f'└─ [2단계] 통과 조합 없음 → 시작 wilson_z={base_wz} 유지하고 계속 ─┘')
    else:
        best_wz = best2['wz']
        # 재확인 — best_wz - step (새 값일 때만 실행)
        refine_wz = round(best_wz - stage_wilson_refine_step, 4)
        rkey = (round(float(refine_wz),6), tuple(best_pct), round(float(base_corr),6))
        if rkey not in _cache:
            print(f'│  ▸ [2단계 재확인] {best_wz} - {stage_wilson_refine_step} = {refine_wz} ...')
            rr = _run_cached(refine_wz, best_pct, base_corr)
            best2b = _pick_best([best2, rr])
            if best2b is not None and best2b['wz'] == refine_wz:
                print(f'│    재확인값 {refine_wz} 채택 — {_fmt(rr)}')
                best_wz = refine_wz; best2 = best2b
            else:
                print(f'│    기존 {best_wz} 유지')
        print(f'└─ [2단계 완료] best wilson_z = {best_wz}  ({_fmt(best2)}) ─┘')

    # ─── 3단계: corr_limit (best_pct, best_wz 고정) ───
    print(f'\n┌─ [3단계/3] corr_limit — 후보 {len(stage_corr_limit)}개 '
          f'(pct={best_pct}, wz={best_wz}) ─┐')
    recs3 = []
    for i, corr in enumerate(stage_corr_limit, 1):
        dup = ' (캐시 재사용)' if (round(float(best_wz),6),tuple(best_pct),round(float(corr),6)) in _cache else ''
        print(f'│  ▸ [3단계] {i}/{len(stage_corr_limit)} corr={corr}{dup} ...')
        r = _run_cached(best_wz, best_pct, corr)
        recs3.append(r)
        print(f'│    완료 — {_fmt(r)}')
    best3 = _pick_best(recs3)
    if best3 is None:
        best_corr = base_corr
        print(f'└─ [3단계] 통과 조합 없음 → 시작 corr_limit={base_corr} 유지 ─┘')
    else:
        best_corr = best3['corr']
        print(f'└─ [3단계 완료] best corr_limit = {best_corr}  ({_fmt(best3)}) ─┘')

    # ─── 단계에서 돌린 모든 결과를 하나의 통합 테이블로 합침 (None=스킵 제외) ───
    all_recs = [r for r in _cache.values() if r is not None]
    combined_parts = [r['combined'] for r in all_recs
                      if r.get('combined') is not None and len(r['combined']) > 0]
    if combined_parts:
        merged_table = pd.concat(combined_parts, ignore_index=True)
        dedup_keys = [c for c in ['meta_wilson_z','meta_pct_low','meta_pct_high',
                                   'meta_corr_limit','meta_min_signals','meta_top_n_pool',
                                   'K_buy','vote_buy','K_sell','vote_sell']
                      if c in merged_table.columns]
        if dedup_keys:
            merged_table = merged_table.drop_duplicates(subset=dedup_keys).reset_index(drop=True)
        merged_table = _sort_combined_table(merged_table, selection_tolerance)
    else:
        merged_table = None

    # ─── 통과 조합이 하나도 없으면: 에러로 멈추지 말고 안내 후 종료 ───
    if merged_table is None or len(merged_table) == 0:
        print('\n' + '█' * 72)
        print('  ⚠ 모든 메타조합에서 통과 조합이 없습니다 (MDD 한도/거래수/B&H 미달).')
        print('     → 엑셀 생성을 건너뜁니다. 다음을 완화해 다시 시도하세요:')
        print(f'        - MAX_DRAWDOWN_LIMIT_PCT (현재 {globals().get("MAX_DRAWDOWN_LIMIT_PCT")})')
        print(f'        - MIN_TRADES_DAILY (현재 {globals().get("MIN_TRADES_DAILY")})')
        print(f'        - EXCLUDE_BELOW_BH (현재 {globals().get("EXCLUDE_BELOW_BH")}) — B&H 미달 제외 끄기')
        print('█' * 72 + '\n')
        # RuntimeError 대신 None 반환 → 상위(run_multi)에서 다음 티커로 계속
        return None

    # ─── 합친 테이블에서 최종 선정 ───
    # ★ 실거래 검증 (요청) — 변수 반복 중엔 안 하고, 여기서 모인 merged_table로 1회만.
    #   각 후보 행의 메타조합(wz,pct,corr)에 해당하는 지표 풀을 _cache에서 찾아
    #   실제 일별 백테스트 → 실제 MDD 기준 통과분 중 승률밴드→수익 최고 선정.
    _do_verify = (globals().get('VERIFY_BY_DAILY_BACKTEST', False)
                  and len(merged_table) > 0)
    sel = None
    if _do_verify:
        try:
            _feat_v, _close_v, _tk_v = _resolve_data()
            _mask_v = _feat_v.index >= pd.Timestamp(run_kwargs.get('eval_start', EVAL_START))
            _feat_v = _feat_v.loc[_mask_v]; _close_v = _close_v.reindex(_feat_v.index)
            _valid_v = _close_v.notna()
            _feat_v = _feat_v[_valid_v.values]; _close_v = _close_v[_valid_v.values]
            # 메타조합별 풀 맵 (wz,pct_low,pct_high,corr) → (buy_pool, sell_pool)
            pool_map = {}
            for _rec in all_recs:
                _res = _rec.get('res')
                if _res is None or len(_res) < 7: continue
                _key = (round(float(_rec['wz']),4), int(_rec['pct'][0]), int(_rec['pct'][1]),
                        round(float(_rec['corr']),4))
                pool_map[_key] = (_res[5], _res[6])   # buy_pool, sell_pool
            globals()['_LAST_POOL_MAP'] = pool_map   # ★ 엑셀 '메타조합별_지표' 시트 작성용
            _bh_v = _bh_sum_return(_close_v.values.astype(np.float64))
            # ★ 앵커 배열 생성 — 검증 daily 백테스트에서 매칭률·앵커수익을 계산하려면 필수.
            #   (이게 없으면 daily 내부에서 매칭 집계가 통째로 스킵돼 매칭률이 0%로 나옴)
            _asb_v = _ass_v = None
            _amode_v = run_kwargs.get('anchor_mode', ANCHOR_MODE)
            if _amode_v and globals().get('AUTO_ANCHOR', True):
                try:
                    _abd, _asd = auto_compute_anchor_dates(
                        _feat_v.index, _close_v,
                        window=globals().get('AUTO_ANCHOR_WINDOW', 1),
                        lookforward=globals().get('AUTO_ANCHOR_LOOKFORWARD', 1),
                        min_rise_after_buy=globals().get('AUTO_ANCHOR_MIN_RISE', 0.01),
                        min_drop_after_sell=globals().get('AUTO_ANCHOR_MIN_DROP', 0.01),
                        price_tolerance=globals().get('AUTO_ANCHOR_PRICE_TOLERANCE', 0.01),
                        max_dates=globals().get('AUTO_ANCHOR_MAX_DATES', None))
                    _asb_v, _ass_v = _compute_anchor_arrays(_feat_v.index, _abd, _asd)
                    if int(_asb_v.sum()) == 0 and int(_ass_v.sum()) == 0:
                        _asb_v = _ass_v = None
                    else:
                        print(f"  ⚓ 검증용 ANCHOR 매칭 — 매수 {int(_asb_v.sum())}일, 매도 {int(_ass_v.sum())}일")
                except Exception as _ae:
                    print(f"  ⚠ 검증용 ANCHOR 배열 생성 실패 ({_ae}) → 매칭률은 0으로 표시될 수 있음")
                    _asb_v = _ass_v = None
            best_inner_v, verified_tbl = _verify_staged_candidates(
                merged_table, _feat_v, _close_v, pool_map,
                horizon=run_kwargs.get('horizon', HORIZON_DAYS),
                dd_limit=run_kwargs.get('dd_limit', DRAWDOWN_LIMIT_BUY),
                ru_limit=run_kwargs.get('ru_limit', RUNUP_LIMIT_SELL),
                stop_loss_pct=run_kwargs.get('stop_loss_pct', STOP_LOSS_PCT),
                anchor_mode=_amode_v,
                bh_ret=_bh_v, anchor_safe_buy=_asb_v, anchor_safe_sell=_ass_v)
            if best_inner_v is not None:
                merged_table = verified_tbl   # 실제 컬럼 포함 + 실제수익 정렬
                # best_inner_v와 일치하는 merged_table 행을 sel로
                sel = merged_table.iloc[0]
                for _ci in range(len(merged_table)):
                    _rr = merged_table.iloc[_ci]
                    if (int(_rr['K_buy'])==int(best_inner_v['K_buy']) and
                        int(_rr['vote_buy'])==int(best_inner_v['vote_buy']) and
                        int(_rr['K_sell'])==int(best_inner_v['K_sell']) and
                        int(_rr['vote_sell'])==int(best_inner_v['vote_sell'])):
                        sel = _rr; break
        except Exception as _ve:
            print(f"  ⚠ staged 실거래 검증 실패 ({_ve}) → 그리드 근사 기준으로 선정")
            sel = None

    if sel is None:
        sel_idx = _select_with_tolerance(merged_table, selection_tolerance,
                                          primary='avg_success_rate', secondary='combined_return')
        sel = merged_table.loc[sel_idx]
    fb = {'wz': float(sel['meta_wilson_z']),
          'pct': (int(sel['meta_pct_low']), int(sel['meta_pct_high'])),
          'corr': float(sel['meta_corr_limit']),
          'sell': float(sel['sell_success_rate']), 'buy': float(sel['buy_success_rate']),
          'avg': float(sel['avg_success_rate']), 'ret': float(sel['total_return']*100),
          'mdd': float(sel['max_drawdown'])}

    n_skipped = sum(1 for v in _cache.values() if v is None)
    print('\n' + '█' * 72)
    print(f'  ★ 최종 선정 — 단계에서 돌린 {len(combined_parts)}개 메타조합의 '
          f'통합 테이블({len(merged_table)}개 그리드)에서 1등')
    if n_skipped > 0:
        print(f'     (통과 조합 없어 스킵된 메타조합 {n_skipped}개 제외)')
    print(f'    pct_range  = {fb["pct"]}')
    print(f'    wilson_z   = {fb["wz"]}')
    print(f'    corr_limit = {fb["corr"]}')
    print(f'    {_fmt(fb)}')
    print('█' * 72)

    # ★ Excel 저장 — 최종 best 조합으로 1회 실행, 그리드 시트엔 merged_table 주입
    print(f'\n  📊 최종 조합으로 Excel 생성 (그리드 시트=단계 전 결과 통합)...')

    # ★ 선정 조합이 보정 채택됐으면, 보정 가중치를 재계산해 최종 백테스트에 주입 (요청).
    #   → 일별 백테스트·거래내역·현재포지션 시트가 모두 '보정된 결과'로 그려진다.
    _force_corr = None
    if bool(sel.get('corr_applied', False)) and globals().get('USE_ANCHOR_MATCH_CORRECTION', True):
        try:
            _bselC = {'K_buy': int(sel['K_buy']), 'vote_buy': int(sel['vote_buy']),
                      'K_sell': int(sel['K_sell']), 'vote_sell': int(sel['vote_sell'])}
            _keyC = (round(float(fb['wz']),4), int(fb['pct'][0]), int(fb['pct'][1]), round(float(fb['corr']),4))
            _poolsC = globals().get('_LAST_POOL_MAP', {}).get(_keyC)
            _fcC = globals().get('_LAST_FULL_CAND_MAP', {}).get(_keyC, (None, None))
            if _poolsC is not None:
                _bpC, _spC = _poolsC
                _ffC, _ccC, _tkC = _resolve_data()
                _mC = _ffC.index >= pd.Timestamp(run_kwargs.get('eval_start', EVAL_START))
                _ffC = _ffC.loc[_mC]; _ccC = _ccC.reindex(_ffC.index)
                _vmC = _ccC.notna(); _ffC = _ffC[_vmC.values]; _ccC = _ccC[_vmC.values]
                _abdC, _asdC = auto_compute_anchor_dates(
                    _ffC.index, _ccC, window=globals().get('AUTO_ANCHOR_WINDOW',1),
                    lookforward=globals().get('AUTO_ANCHOR_LOOKFORWARD',1),
                    min_rise_after_buy=globals().get('AUTO_ANCHOR_MIN_RISE',0.01),
                    min_drop_after_sell=globals().get('AUTO_ANCHOR_MIN_DROP',0.01),
                    price_tolerance=globals().get('AUTO_ANCHOR_PRICE_TOLERANCE',0.01),
                    max_dates=globals().get('AUTO_ANCHOR_MAX_DATES',None))
                _asbC, _assC = _compute_anchor_arrays(_ffC.index, _abdC, _asdC)
                _augBC, _augSC, _bwC, _swC, _mcC = anchor_match_correct_weights(
                    _ffC, _ccC, _bpC, _spC, _bselC,
                    anchor_mode=run_kwargs.get('anchor_mode', ANCHOR_MODE),
                    anchor_safe_buy=_asbC, anchor_safe_sell=_assC,
                    full_cand_buy=_fcC[0], full_cand_sell=_fcC[1], verbose=False)
                if _mcC.get('adopted', False) and _bwC is not None:
                    _force_corr = {'buy_pool': _augBC, 'sell_pool': _augSC,
                                   'buy_w': _bwC, 'sell_w': _swC,
                                   'n_added_buy': _mcC.get('n_added_buy', 0),
                                   'n_added_sell': _mcC.get('n_added_sell', 0)}
                    print(f"  🔧 선정 조합 보정 재계산 완료 → 최종 시트에 보정 결과 반영 "
                          f"(수익 {_mcC['ret_before']:+.1f}%→{_mcC['ret_after']:+.1f}%, "
                          f"MDD {_mcC['mdd_before']*100:.2f}%→{_mcC['mdd_after']*100:.2f}%)")
        except Exception as _ce:
            print(f"  ⚠ 선정 조합 보정 재계산 실패(기존 결과로 출력): {_ce}")

    try:
        final_res = run_ensemble_search(
            meta_grid=_mk_grid(fb['wz'], fb['pct'], fb['corr']),
            write_output=True, output_file=output_file,
            inject_combined_table=merged_table,
            force_best_combo={
                'K_buy': int(sel['K_buy']), 'vote_buy': int(sel['vote_buy']),
                'K_sell': int(sel['K_sell']), 'vote_sell': int(sel['vote_sell']),
            },
            force_corr=_force_corr,
            **run_kwargs)
    except RuntimeError as _e:
        # 만약 최종 best 조합조차 단독으로는 통과 못 하는 경우(드묾):
        # merged_table은 이미 있으니, 통과 조합 1개를 강제로 쓰도록 안내
        print(f'  ⚠ 최종 조합 단독 실행 실패: {_e}')
        print(f'     merged_table({len(merged_table)}개)은 유효하나 best 조합 재실행이 막힘.')
        print(f'     MAX_DRAWDOWN_LIMIT_PCT을 약간 완화해 재시도하세요.')
        return None

    print('\n' + '█' * 72)
    print(f'  ✅ 단계 튜닝 최종: {_fmt(fb)}')
    print(f'     (pct_range={fb["pct"]}, wilson_z={fb["wz"]}, corr_limit={fb["corr"]})')
    print(f'     ※ 단계 탐색 {len(_cache)}회 (스킵 {n_skipped}) + 최종 Excel 1회')
    print(f'     ※ 엑셀의 내부_그리드_통과 시트에 단계 전 결과가 모두 담김')
    print('█' * 72 + '\n')

    return final_res


# ════════════════════════════════════════════════════════════════════════
#   ★ 그리드 번호로 특정 조합 재현 → 일별 거래 Excel 생성
#     경로는 OUTPUT_DIR 고정, 파일명 + 그리드 번호만 입력.
#     '사용된 설정' 시트의 설정값(Horizon/손실한도/비용/메타변수 등)을 그대로 사용.
# ════════════════════════════════════════════════════════════════════════
def _parse_used_settings(wb):
    """'사용된 설정' 시트에서 백테스트/메타 설정값을 읽어 dict로 반환."""
    out = {}
    if '사용된 설정' not in wb.sheetnames:
        return out
    ws = wb['사용된 설정']
    kv = {}
    for r in range(1, ws.max_row + 1):
        k = ws.cell(r, 1).value; v = ws.cell(r, 2).value
        if k is None or v is None: continue
        kv[str(k).strip()] = str(v).strip()

    def _num(s):
        """문자열에서 첫 번째 숫자(부호·소수점 포함)만 추출."""
        if s is None: return None
        import re as _re
        m = _re.search(r'[-+]?\d+(?:\.\d+)?', str(s))
        if not m: return None
        try: return float(m.group(0))
        except Exception: return None

    for k, v in kv.items():
        kl = k.replace(' ', '')
        if 'Horizon' in k:
            n = _num(v);  out['horizon'] = int(n) if n is not None else None
        elif '매수손실한도' in kl:
            n = _num(v);  out['dd_limit'] = abs(n)/100.0 if n is not None else None
        elif '매도상승한도' in kl:
            n = _num(v);  out['ru_limit'] = abs(n)/100.0 if n is not None else None
        elif '거래비용' in kl:
            n = _num(v);  out['cost'] = n/100.0 if n is not None else None
        elif 'WILSON_Z' in k:
            out['wilson_z'] = _num(v)
        elif 'PCT' in k and '범위' in k:
            s = v.replace('(', '').replace(')', '').replace('[', '').replace(']', '')
            try:
                ps = [int(float(x.strip())) for x in s.split(',') if x.strip()]
                if len(ps) == 2: out['pct_range'] = tuple(ps)
            except Exception: pass
        elif 'MIN_SIGNALS' in k:
            n = _num(v);  out['min_signals'] = int(n) if n is not None else None
        elif 'DIVERSITY_CORR_LIMIT' in k or 'CORR_LIMIT' in k:
            out['corr_limit'] = _num(v)
        elif 'TOP_N_POOL_BUY' in k:
            n = _num(v);  out['top_n_pool'] = int(n) if n is not None else None
        elif 'MDD한도' in kl:
            n = _num(v);  out['mdd_limit_pct'] = abs(n) if n is not None else None
        elif '최소거래수' in kl:
            n = _num(v);  out['min_trades_daily'] = int(n) if n is not None else None
        elif '손절매한도' in kl:
            n = _num(v);  out['stop_loss_pct'] = abs(n)/100.0 if n is not None else None
        elif kl == 'K_buy':
            n = _num(v);  out['K_buy'] = int(n) if n is not None else None
        elif kl == 'vote_buy':
            n = _num(v);  out['vote_buy'] = int(n) if n is not None else None
        elif kl == 'K_sell':
            n = _num(v);  out['K_sell'] = int(n) if n is not None else None
        elif kl == 'vote_sell':
            n = _num(v);  out['vote_sell'] = int(n) if n is not None else None
        elif k.strip() == '티커' or kl == '티커':
            out['ticker'] = str(v).strip()
    return out


def replay_grid_combo(filename, grid_number=None, *,
                       feat=None, close=None,
                       output_dir=None,
                       **override_kwargs):
    """
    OUTPUT_DIR 안의 결과 Excel(filename)에서 grid_number 조합을 재현해
    일별 거래 Excel을 생성한다. 경로는 OUTPUT_DIR 고정 — 파일명과 번호만 입력.

    '사용된 설정' 시트의 설정값(Horizon, 손실/상승 한도, 거래비용, 메타변수,
    MDD 한도, 최소거래수, 손절매)을 그대로 읽어 그 조건으로 재현한다.

    Parameters
    ----------
    filename : str       OUTPUT_DIR 안의 결과 Excel 파일명 (예: 'ensemble_search_VRT_2026-06-01.xlsx')
                         전체 경로를 줘도 됨.
    grid_number : int|str  재현할 그리드 번호 ('14', '17●' 등 — 숫자만 인식)
    feat, close : 선택    원본 지표/종가. None이면 글로벌(_pair_feat/_pair_close)에서 탐색.
    output_dir : 선택     기본 OUTPUT_DIR. 다른 폴더 쓰려면 지정.
    override_kwargs       읽은 설정을 덮어쓸 인자.
    """
    from openpyxl import load_workbook

    # ★ 재현 대상이 ★최적 행이면 '보정 적용 풀(앙상블 시트)'을 쓰고, 아니면 base 메타풀 사용.
    #   자동재현(grid_number=None)은 항상 ★최적이므로 True.
    _is_best_row = (grid_number is None)

    base_dir = output_dir if output_dir is not None else OUTPUT_DIR
    # filename이 전체 경로면 그대로, 아니면 OUTPUT_DIR과 결합
    if os.path.isabs(filename) or os.path.dirname(filename):
        excel_path = filename
    else:
        excel_path = os.path.join(base_dir, filename)
    if not os.path.exists(excel_path):
        raise RuntimeError(f"파일을 찾을 수 없습니다: {excel_path}\n"
                           f"  OUTPUT_DIR={base_dir} 안에 파일명이 맞는지 확인하세요.")

    wb = load_workbook(excel_path, data_only=True)

    # ★ 사용된 설정 (백테스트 조건 + ★최적 K/vote)
    used = _parse_used_settings(wb)

    # ★ 원본 엑셀의 마지막 백테스트 날짜 — 재현 시 '같은 범위'로 잘라 정확 재현 (요청)
    _orig_last_date = None
    if globals().get('REPLAY_MATCH_ORIGINAL_RANGE', True) and '일별 백테스트' in wb.sheetnames:
        try:
            _wsd = wb['일별 백테스트']
            for _r in range(_wsd.max_row, 4, -1):
                _dv = _wsd.cell(_r, 1).value
                if _dv is not None:
                    _orig_last_date = pd.Timestamp(_dv); break
            if _orig_last_date is not None:
                print(f"  📅 원본 마지막 날짜 = {str(_orig_last_date)[:10]} → 재현도 이 날짜까지(같은 범위)")
        except Exception:
            _orig_last_date = None

    # ─── grid_number 가 None 이면: '내부_그리드_통과' 시트의 ★표시 행을 읽어 재현 ───
    #   (요청) ★를 어느 행에 넣든 그 행으로 재현. ★가 없고 데이터가 한 줄뿐이면 그 행으로.
    #   이 경우 '사용된 설정' 시트로는 절대 재현하지 않는다.
    if grid_number is None:
        if '내부_그리드_통과' not in wb.sheetnames:
            raise RuntimeError("'내부_그리드_통과' 시트가 없습니다 — 자동 재현 불가.")
        ws = wb['내부_그리드_통과']
        hdr = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(3, c).value
            if v is not None: hdr[str(v).strip()] = c
        for k in ['#', 'K_buy', 'vote_buy', 'K_sell', 'vote_sell']:
            if k not in hdr:
                raise RuntimeError(f"'내부_그리드_통과'에 '{k}' 컬럼이 없습니다. (헤더: {list(hdr)})")
        data_rows = []; star_rows = []
        for r in range(4, ws.max_row + 1):
            cell = ws.cell(r, hdr['#']).value
            if cell is None: continue
            if ws.cell(r, hdr['K_buy']).value in (None, ''): continue   # 데이터 행만
            data_rows.append(r)
            if '★' in str(cell): star_rows.append(r)
        # 자동 ★(최적행)은 # 셀이 '굵은 빨강(C00000)'으로 강조됨 → 자동/수동 구분해 '수동 ★' 우선
        def _is_auto_star(r):
            try:
                fnt = ws.cell(r, hdr['#']).font
                col = (fnt.color.rgb if (fnt and fnt.color) else '') or ''
                return bool(fnt and fnt.bold) and str(col).endswith('C00000')
            except Exception:
                return False
        manual_stars = [r for r in star_rows if not _is_auto_star(r)]
        if len(manual_stars) == 1:
            target = manual_stars[0]
            print(f"  ♻ 자동 재현 — 사용자가 찍은 ★행 사용 "
                  f"(#{str(ws.cell(target, hdr['#']).value).strip()})")
        elif len(manual_stars) > 1:
            _nums = [str(ws.cell(r, hdr['#']).value).strip() for r in manual_stars]
            raise RuntimeError(
                f"사용자가 찍은 ★가 {len(manual_stars)}개입니다: {', '.join(_nums)}\n"
                f"  재현할 행 하나에만 ★를 남기고 나머지는 지운 뒤 다시 실행하세요.")
        elif len(star_rows) == 1:
            target = star_rows[0]
            print(f"  ♻ 자동 재현 — ★표시(최적)행 사용 "
                  f"(#{str(ws.cell(target, hdr['#']).value).strip()})")
        elif len(star_rows) > 1:
            _nums = [str(ws.cell(r, hdr['#']).value).strip() for r in star_rows]
            raise RuntimeError(
                f"'내부_그리드_통과'에 ★가 {len(star_rows)}개인데 자동/수동 구분이 안 됩니다: {', '.join(_nums)}\n"
                f"  재현할 행 하나만 ★로 남기고 나머지는 지운 뒤 다시 실행하세요.")
        elif len(data_rows) == 1:
            target = data_rows[0]
            print(f"  ♻ 자동 재현 — ★ 없음, 데이터가 한 줄뿐 → 그 행 사용 "
                  f"(#{str(ws.cell(target, hdr['#']).value).strip()})")
        else:
            raise RuntimeError(
                f"'내부_그리드_통과'에 ★표시 행이 없고 데이터가 {len(data_rows)}줄입니다.\n"
                f"  ★를 원하는 행의 # 옆에 넣거나, 필터링으로 한 줄만 남겨 주세요.\n"
                f"  (요청대로 '사용된 설정' 시트로는 재현하지 않습니다.)")
        # ★ 보정이 '채택'된 행이면 보정 앙상블 풀(매수/매도_앙상블_지표)을, 아니면 base 메타풀
        #   ('메타조합별_지표'의 그 행 메타조합 풀)을 K로 잘라 그리드 평가를 그대로 재현.
        #   (#5211처럼 보정채택='—'이면 그리드에서 쓴 그 메타풀+K+vote 그대로여야 일치)
        _corr_col = hdr.get('🔧보정채택')
        _corr_val = ws.cell(target, _corr_col).value if _corr_col else None
        _corr_str = str(_corr_val).strip() if _corr_val is not None else ''
        _is_best_row = _corr_str not in ('', '—', '–', '-', 'None', 'nan')
        K_buy    = int(float(ws.cell(target, hdr['K_buy']).value))
        vote_buy = int(float(ws.cell(target, hdr['vote_buy']).value))
        K_sell   = int(float(ws.cell(target, hdr['K_sell']).value))
        vote_sell = int(float(ws.cell(target, hdr['vote_sell']).value))
        print(f"     (K_buy={K_buy}/v{vote_buy}, K_sell={K_sell}/v{vote_sell}, "
              f"보정채택='{_corr_str or '—'}' → {'보정 앙상블 풀' if _is_best_row else 'base 메타풀(그리드 그대로)'})")
    else:
        gn_clean = ''.join(ch for ch in str(grid_number) if ch.isdigit())
        if not gn_clean:
            raise ValueError(f"그리드 번호를 해석할 수 없습니다: {grid_number!r}")
        gn = int(gn_clean)

        if '내부_그리드_통과' not in wb.sheetnames:
            raise RuntimeError("'내부_그리드_통과' 시트가 없습니다.")
        ws = wb['내부_그리드_통과']

        hdr = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(3, c).value
            if v is not None: hdr[str(v).strip()] = c
        for k in ['#', 'K_buy', 'vote_buy', 'K_sell', 'vote_sell']:
            if k not in hdr:
                raise RuntimeError(f"'내부_그리드_통과'에 '{k}' 컬럼이 없습니다. (헤더: {list(hdr)})")

        target = None
        for r in range(4, ws.max_row + 1):
            cell = ws.cell(r, hdr['#']).value
            if cell is None: continue
            digits = ''.join(ch for ch in str(cell) if ch.isdigit())
            if digits and int(digits) == gn:
                target = r; _is_best_row = ('★' in str(cell)); break
        if target is None:
            raise RuntimeError(f"#{gn} 번호를 찾지 못했습니다.")

        K_buy   = int(float(ws.cell(target, hdr['K_buy']).value))
        vote_buy = int(float(ws.cell(target, hdr['vote_buy']).value))
        K_sell  = int(float(ws.cell(target, hdr['K_sell']).value))
        vote_sell = int(float(ws.cell(target, hdr['vote_sell']).value))

    # ★ 메타변수 — 그리드 행에 직접 기록된 값을 우선 사용 (행마다 다를 수 있음).
    #   자동 재현 모드(target=None)면 _cellf는 None을 돌려주고 used 값으로 폴백.
    def _cellf(colname):
        if target is not None and colname in hdr:
            v = ws.cell(target, hdr[colname]).value
            try: return float(str(v).replace('%','').strip())
            except Exception: return None
        return None
    row_wz   = _cellf('wilson_z')
    row_pl   = _cellf('pct_low')
    row_ph   = _cellf('pct_high')
    row_cl   = _cellf('corr_limit')
    row_ms   = _cellf('min_sig')
    row_pool = _cellf('pool')

    mg = dict(META_GRID)
    # 행값 우선, 없으면 사용된 설정으로 폴백
    wz_use   = row_wz   if row_wz   is not None else used.get('wilson_z')
    cl_use   = row_cl   if row_cl   is not None else used.get('corr_limit')
    ms_use   = row_ms   if row_ms   is not None else used.get('min_signals')
    pool_use = row_pool if row_pool is not None else used.get('top_n_pool')
    if row_pl is not None and row_ph is not None:
        pct_use = (int(row_pl), int(row_ph))
    else:
        pct_use = used.get('pct_range')
    if wz_use   is not None: mg['wilson_z']   = [wz_use]
    if pct_use  is not None: mg['pct_range']  = [pct_use]
    if cl_use   is not None: mg['corr_limit'] = [cl_use]
    if ms_use   is not None: mg['min_signals']= [int(ms_use)]
    if pool_use is not None: mg['top_n_pool'] = [int(pool_use)]
    _gn_label = f"#{gn} " if grid_number is not None else "★최적 "
    print(f"    ({_gn_label}메타변수: wilson_z={wz_use}, pct={pct_use}, corr={cl_use}, min_sig={ms_use}, pool={pool_use})")

    g = globals()
    if feat is None:  feat  = g.get('_pair_feat')  or g.get('feat')
    if close is None: close = g.get('_pair_close') or g.get('close')
    # ★ 메모리에 없으면 티커로 데이터 자동 확보 (다운로드 함수/yfinance 등 사용)
    # ★ 데이터 스냅샷 우선 (요청) — 원본 분석이 남긴 '..._data.pkl'이 같은 폴더에 있으면
    #   그걸 써서 '완전히 동일한 데이터'로 정확 재현 (재다운로드 안 함 → FRED수정/vintage 차단).
    #   스냅샷이 없거나 REPLAY_USE_SNAPSHOT=False면 기존처럼 재다운로드(연장 가능, 미세차이 가능).
    if (feat is None or close is None) and globals().get('REPLAY_USE_SNAPSHOT', True):
        _snap = os.path.splitext(excel_path)[0] + '_data.pkl'
        if os.path.exists(_snap):
            try:
                _dsnap = pd.read_pickle(_snap)
                feat = _dsnap.get('feat'); close = _dsnap.get('close')
                if feat is not None and close is not None and len(close) > 0:
                    print(f"  💾 데이터 스냅샷 사용: {os.path.basename(_snap)} "
                          f"({len(close)}일, 마지막 {str(close.index[-1])[:10]}) — 재다운로드 없이 정확 재현")
                else:
                    feat = close = None
            except Exception as _le:
                print(f"  ⚠ 스냅샷 로드 실패({_le}) — 재다운로드로 진행")
                feat = close = None
        else:
            print(f"  ℹ 데이터 스냅샷 없음({os.path.basename(_snap)}) — 재다운로드로 진행 "
                  f"(외부데이터 차이로 미세하게 다를 수 있음)")

    if feat is None or close is None:
        ticker = used.get('ticker')
        if ticker is None:
            # 파일명에서 티커 추출 시도 (ensemble_search_VRT_... → VRT)
            import re as _re
            mtk = _re.search(r'ensemble_search_([A-Za-z0-9.\-]+)_', os.path.basename(excel_path))
            if mtk: ticker = mtk.group(1)
        if ticker is not None and '_resolve_data_for_ticker' in g and callable(g['_resolve_data_for_ticker']):
            print(f"  ℹ 메모리에 데이터 없음 → 티커 '{ticker}'로 자동 로드 시도...")
            try:
                feat, close = g['_resolve_data_for_ticker'](ticker, end_date=_orig_last_date)
                print(f"  ✓ '{ticker}' 데이터 로드 성공 ({len(close)}일)")
            except Exception as _e:
                raise RuntimeError(
                    f"티커 '{ticker}' 데이터 자동 로드 실패: {_e}\n"
                    f"  download_data/compute_features 함수가 정의돼 있거나 yfinance가 설치돼 있어야 합니다.\n"
                    f"  또는 replay_grid_combo(..., feat=내_feat, close=내_close)로 직접 전달하세요.")
        else:
            raise RuntimeError(
                "원본 데이터(feat, close)를 찾을 수 없고 자동 로드도 불가합니다.\n"
                f"  티커={ticker}, 다운로드 함수 존재={'_resolve_data_for_ticker' in g}\n"
                "  feat=, close= 로 직접 전달하거나 download_data/compute_features를 준비하세요.")

    base = os.path.splitext(os.path.basename(excel_path))[0]
    if grid_number is None:
        output_file = os.path.join(base_dir, f"{base}__replay_best.xlsx")
    else:
        output_file = os.path.join(base_dir, f"{base}__replay_grid{gn}.xlsx")

    print('═' * 72)
    if grid_number is None:
        print(f'  ★ 최적 조합 자동 재현 → 일별 거래 Excel')
    else:
        print(f'  ★ 그리드 #{gn} 재현 → 일별 거래 Excel')
    print(f'    파일: {excel_path}')
    print(f'    티커: {used.get("ticker", "(파일명에서 추출)")}')
    print(f'    조합: K_buy={K_buy}/vote={vote_buy}, K_sell={K_sell}/vote={vote_sell}')
    print(f'    설정(사용된 설정 시트 그대로):')
    print(f'      Horizon={used.get("horizon")}, dd_limit={used.get("dd_limit")}, '
          f'ru_limit={used.get("ru_limit")}, cost={used.get("cost")}')
    print(f'      WILSON_Z={used.get("wilson_z")}, PCT={used.get("pct_range")}, '
          f'CORR={used.get("corr_limit")}, MIN_SIG={used.get("min_signals")}')
    print(f'      MDD한도={used.get("mdd_limit_pct")}, 최소거래={used.get("min_trades_daily")}, '
          f'손절매={used.get("stop_loss_pct")}')
    print(f'    출력: {output_file}')
    print('═' * 72)

    kwargs = dict(
        meta_grid=mg,
        k_buy_range=[K_buy], k_sell_range=[K_sell],
        vote_ratio_buy=[vote_buy / K_buy] if K_buy > 0 else [1.0],
        vote_ratio_sell=[vote_sell / K_sell] if K_sell > 0 else [1.0],
        write_output=True, output_file=output_file,
        # ★ 이 그리드의 정확한 조합을 강제 → 현재 포지션 = 재현 조합 일치
        force_best_combo={'K_buy': K_buy, 'vote_buy': vote_buy,
                          'K_sell': K_sell, 'vote_sell': vote_sell},
        # inject_pools는 아래에서 엑셀 지표 풀을 구성한 뒤 추가함
        # ★ 재현은 '이미 선정된 그 조합 하나'를 그대로 다시 계산하는 것.
        #   선정용 필터(MDD 한도·최소거래수)를 끄지 않으면, 재현 결과가 원본과 미세하게
        #   달라(데이터 갱신 등) 그 조합이 필터에 걸려 '통과 0개'로 재현 실패함.
        max_drawdown_limit_pct=None,   # MDD 한도 무력화 (그 조합 그대로 봐야 함)
        min_trades_daily=1,            # 최소거래수 무력화
    )
    # 읽은 설정 반영 (run_ensemble_search 인자명에 맞춰).
    #   ※ cost는 전역 COST_PER_TRADE로 따로 처리.
    #   ※ mdd_limit / min_trades는 위에서 의도적으로 껐으므로 여기서 덮어쓰지 않음(목록 제외).
    for src_k, dst_k in [('horizon','horizon'), ('dd_limit','dd_limit'),
                          ('ru_limit','ru_limit'),
                          ('stop_loss_pct','stop_loss_pct')]:
        if used.get(src_k) is not None:
            kwargs[dst_k] = used[src_k]
    kwargs.update(override_kwargs)

    g['_pair_feat'] = feat
    g['_pair_close'] = close

    # ★ 엑셀의 매수/매도 앙상블 지표 시트를 읽어 '그 지표 그대로' 풀 구성 → 재현 정확도 보장.
    #   (메타변수로 지표를 다시 선별하지 않고, 원본이 쓴 지표를 그대로 사용)
    def _read_pool(sheet_name):
        if sheet_name not in wb.sheetnames:
            return None
        wsp = wb[sheet_name]
        # 헤더(3행): #, 지표, 방향, 임계치, 분위, 신호수, 성공수, 성공률, 점수
        h = {}
        for c in range(1, wsp.max_column + 1):
            v = wsp.cell(3, c).value
            if v is not None: h[str(v).strip()] = c
        need = ['지표', '방향', '임계치']
        if not all(k in h for k in need):
            return None
        rows = []
        for r in range(4, wsp.max_row + 1):
            ind = wsp.cell(r, h['지표']).value
            if ind is None or str(ind).strip() == '':
                continue
            def _gv(col, default=None):
                if col in h:
                    return wsp.cell(r, h[col]).value
                return default
            def _gf(col, default=0.0):
                v = _gv(col)
                try: return float(str(v).replace('%','').strip())
                except Exception: return default
            n_sig = _gf('신호수', 0.0)
            n_suc = _gf('성공수', 0.0)
            sr = _gf('성공률', 0.0)
            if sr > 1.5: sr = sr / 100.0   # '96.00%' → 0.96
            _vw = _gv('🔁가중치')
            try: _vwf = float(str(_vw).replace('%','').strip())
            except Exception: _vwf = np.nan
            rows.append({
                'indicator': str(ind).strip(),
                'direction': str(_gv('방향', '>=')).strip(),
                'threshold': _gf('임계치', 0.0),
                'pct_label': _gf('분위', 50.0),
                'n_signals': n_sig,
                'n_success': n_suc,
                'success_rate': sr,
                'avg_extreme': 0.01,
                'score': _gf('점수', sr),
                'vote_weight': _vwf,   # ★ 재현용 저장 가중치 (없으면 NaN)
            })
        if not rows:
            return None
        return pd.DataFrame(rows)

    def _read_meta_pool(wz, pl, ph, corr):
        """메타조합별_지표 시트에서 (wz,pct,corr)에 맞는 매수/매도 풀을 읽는다.
           그리드 번호 재현 시, 그 번호의 메타조합 지표를 '그대로' 쓰기 위함."""
        sn = '메타조합별_지표'
        if sn not in wb.sheetnames:
            return None, None
        wsp = wb[sn]
        # 블록 헤더 '★ wilson_z=... / pct=(..) / corr=..' 를 찾아 매칭
        target_hdr = None
        rows_by_block = {}
        cur_block = None; cur_side = None; cols = None
        bp_rows = []; sp_rows = []
        def _close(s):
            try: return float(str(s).replace('%','').strip())
            except Exception: return 0.0
        import re as _re
        for r in range(1, wsp.max_row + 1):
            a = wsp.cell(r, 1).value
            if a is None: continue
            s = str(a)
            if s.startswith('★ wilson_z='):
                m = _re.search(r'wilson_z=([\d.]+).*pct=\((\d+),(\d+)\).*corr=([\d.]+)', s)
                if m:
                    bwz, bpl, bph, bcorr = float(m.group(1)), int(m.group(2)), int(m.group(3)), float(m.group(4))
                    cur_block = (round(bwz,4)==round(float(wz),4) and bpl==int(pl)
                                 and bph==int(ph) and round(bcorr,4)==round(float(corr),4))
                cur_side = None
                continue
            if not cur_block:
                continue
            if s.strip().startswith('[매수]'): cur_side = 'buy'; continue
            if s.strip().startswith('[매도]'): cur_side = 'sell'; continue
            if s == 'side':   # 헤더행
                continue
            # 데이터행: side, 지표, 방향, 임계치, 분위, 신호수, 성공수, 성공률, 점수
            side = s.strip()
            ind = wsp.cell(r, 2).value
            if ind is None or str(ind).strip() == '': continue
            rec = {'indicator': str(ind).strip(),
                   'direction': str(wsp.cell(r,3).value or '>=').strip(),
                   'threshold': _close(wsp.cell(r,4).value),
                   'pct_label': _close(wsp.cell(r,5).value),
                   'n_signals': _close(wsp.cell(r,6).value),
                   'n_success': _close(wsp.cell(r,7).value),
                   'success_rate': _close(wsp.cell(r,8).value),
                   'avg_extreme': 0.01,
                   'score': _close(wsp.cell(r,9).value)}
            if side == '매수': bp_rows.append(rec)
            elif side == '매도': sp_rows.append(rec)
        bp = pd.DataFrame(bp_rows) if bp_rows else None
        sp = pd.DataFrame(sp_rows) if sp_rows else None
        return bp, sp

    # ★ 그리드 번호 재현이면, 그 번호의 메타조합 풀을 '메타조합별_지표' 시트에서 우선 읽음.
    #   (없으면 ★최적의 매수/매도_앙상블_지표 시트로 폴백)
    #   단, ★최적 행 재현이면 보정 적용된 '앙상블 시트(가중치 포함)'를 우선 사용해 보정까지 정확 재현.
    buy_pool_xl = sell_pool_xl = None
    if not _is_best_row:
        _mb, _ms = _read_meta_pool(wz_use, pct_use[0] if pct_use else 5,
                                   pct_use[1] if pct_use else 95, cl_use if cl_use is not None else 0.2)
        if _mb is not None and _ms is not None and len(_mb) > 0 and len(_ms) > 0:
            buy_pool_xl, sell_pool_xl = _mb, _ms
            _gnlbl = f"#{grid_number}" if grid_number is not None else "★행"
            print(f"  ♻ {_gnlbl}(보정 미채택)의 메타조합 지표를 '메타조합별_지표' 시트에서 읽음 "
                  f"(매수 {len(_mb)} / 매도 {len(_ms)}) — base 풀, K로 잘라 그리드 그대로 재현")
    if buy_pool_xl is None or sell_pool_xl is None:
        # ★최적 행 또는 자동재현 → 앙상블 시트(보정 적용 풀 + 🔁가중치)로 정확 재현
        buy_pool_xl  = _read_pool('매수_앙상블_지표')
        sell_pool_xl = _read_pool('매도_앙상블_지표')
        if _is_best_row and buy_pool_xl is not None:
            print(f"  ♻ ★최적 조합 재현 — 앙상블 시트(보정 적용 풀 + 가중치) 사용")

    inject_pools = None
    if buy_pool_xl is not None and sell_pool_xl is not None \
       and len(buy_pool_xl) > 0 and len(sell_pool_xl) > 0:
        # 엑셀 지표 중 현재 feat에 존재하는 것만 사용 (데이터 갱신으로 사라진 지표 방어)
        bcols = set(feat.columns)
        bp_use = buy_pool_xl[buy_pool_xl['indicator'].isin(bcols)].reset_index(drop=True)
        sp_use = sell_pool_xl[sell_pool_xl['indicator'].isin(bcols)].reset_index(drop=True)
        miss_b = len(buy_pool_xl) - len(bp_use)
        miss_s = len(sell_pool_xl) - len(sp_use)
        if len(bp_use) >= K_buy and len(sp_use) >= K_sell:
            inject_pools = (bp_use, sp_use)
            print(f"  ♻ 엑셀 지표 그대로 사용 — 매수풀 {len(bp_use)}개, 매도풀 {len(sp_use)}개")
            if miss_b or miss_s:
                print(f"     (현재 데이터에 없는 지표 제외: 매수 {miss_b}, 매도 {miss_s})")
        else:
            print(f"  ⚠ 엑셀 지표가 K값보다 적음(매수 {len(bp_use)}<{K_buy} 또는 매도 {len(sp_use)}<{K_sell})")
            print(f"     → 지표를 메타변수로 다시 선별해 재현(정확도 다소 하락 가능)")
    else:
        print(f"  ⚠ 앙상블 지표 시트를 읽지 못함 → 메타변수로 지표 재선별")

    kwargs['inject_pools'] = inject_pools   # ★ 엑셀 지표 풀 (None이면 메타변수로 재선별)

    # ★ 보정 정확 재현 (요청) — 엑셀 앙상블 시트에 저장된 '🔁가중치'가 있으면, 그 풀과 가중치를
    #   force_corr로 넘겨 최종 백테스트가 '저장된 그대로'(추가지표+보정가중치) 재현되게 한다.
    #   (재현 중에는 보정을 다시 돌리지 않음 — 결정성/데이터 갱신 무관하게 원본 일치)
    _replay_force_corr = None
    if inject_pools is not None:
        _bp_inj, _sp_inj = inject_pools
        _has_w = ('vote_weight' in _bp_inj.columns and _bp_inj['vote_weight'].notna().any()
                  and 'vote_weight' in _sp_inj.columns and _sp_inj['vote_weight'].notna().any())
        if _has_w:
            _bw_arr = _bp_inj['vote_weight'].astype(float).values
            _sw_arr = _sp_inj['vote_weight'].astype(float).values
            # 풀 크기 = 저장된 전체(=보정 후 augmented). K도 그 길이로 강제.
            _replay_force_corr = {'buy_pool': _bp_inj, 'sell_pool': _sp_inj,
                                  'buy_w': _bw_arr, 'sell_w': _sw_arr,
                                  'n_added_buy': 0, 'n_added_sell': 0}
            kwargs['force_corr'] = _replay_force_corr
            kwargs['force_best_combo'] = {'K_buy': len(_bp_inj), 'vote_buy': vote_buy,
                                          'K_sell': len(_sp_inj), 'vote_sell': vote_sell}
            print(f"  ♻ 저장된 보정 가중치로 정확 재현 — 매수풀 {len(_bp_inj)}개 / 매도풀 {len(_sp_inj)}개 "
                  f"(가중치 그대로 주입, 보정 재실행 안 함)")
        else:
            print(f"  ℹ 엑셀에 '🔁가중치'가 없어(구버전 파일) 점수로 가중치 재계산 — "
                  f"보정 가중치는 정확 재현 안 될 수 있음")

    # ★ B&H 미달 제외도 재현 중에는 끈다 (그 조합이 B&H 미달이어도 그대로 봐야 함)
    _exbh_saved = globals().get('EXCLUDE_BELOW_BH')
    # 거래비용은 전역 COST_PER_TRADE 사용 → 재현 시 사용된 설정값으로 임시 교체 후 복원
    _cost_saved = globals().get('COST_PER_TRADE')
    _cost_used = used.get('cost')
    try:
        globals()['EXCLUDE_BELOW_BH'] = False
        if _cost_used is not None:
            globals()['COST_PER_TRADE'] = _cost_used
        result = run_ensemble_search(**kwargs)
    finally:
        globals()['EXCLUDE_BELOW_BH'] = _exbh_saved
        if _cost_used is not None:
            globals()['COST_PER_TRADE'] = _cost_saved

    _done_label = f'그리드 #{gn}' if grid_number is not None else '★최적 조합'
    print(f'\n  ✅ {_done_label} 일별 Excel 생성 완료 → {output_file}')
    if AUTO_DOWNLOAD_EXCEL:
        _auto_download_excels([output_file])
    return result

def find_latest_excel(ticker, *, output_dir=None):
    """해당 티커의 '가장 최근 날짜' 일반 분석 엑셀을 찾는다.
       - 파일명 형식: ensemble_search_{TICKER}_{YYYY-MM-DD}.xlsx
       - '__replay'가 들어간 재현본은 제외 (요청: 2번 그리드/재현 파일 아님)
       - 날짜가 가장 늦은 파일을 반환. 없으면 None.
    """
    import glob, re as _re
    base_dir = output_dir if output_dir is not None else OUTPUT_DIR
    pat = os.path.join(base_dir, f"ensemble_search_{ticker}_*.xlsx")
    cands = []
    for p in glob.glob(pat):
        name = os.path.basename(p)
        if '__replay' in name:        # 재현본 제외
            continue
        m = _re.search(rf"ensemble_search_{_re.escape(ticker)}_(\d{{4}}-\d{{2}}-\d{{2}})\.xlsx$", name)
        if not m:
            continue
        cands.append((m.group(1), p))   # (날짜문자열, 경로)
    if not cands:
        return None
    cands.sort(key=lambda x: x[0])      # 날짜 오름차순
    return cands[-1][1]                 # 가장 최근


def replay_latest_best(ticker, *, output_dir=None, feat=None, close=None, **override_kwargs):
    """티커의 가장 최근 일반 분석 엑셀을 자동으로 찾아, 그 엑셀의 ★최적 조합·지표를
       그대로 재현한다. (엑셀 기간까지 동일, 그 이후 새 거래일은 같은 로직으로 이어 계산)"""
    latest = find_latest_excel(ticker, output_dir=output_dir)
    if latest is None:
        raise RuntimeError(
            f"'{ticker}'의 일반 분석 엑셀을 찾지 못했습니다.\n"
            f"  {output_dir or OUTPUT_DIR} 안에 ensemble_search_{ticker}_YYYY-MM-DD.xlsx 가 있어야 합니다.\n"
            f"  (먼저 새 분석을 1회 실행해 엑셀을 만들어야 재현할 수 있습니다.)")
    print(f"  📂 '{ticker}' 가장 최근 분석 엑셀: {os.path.basename(latest)}")
    # grid_number=None → ★최적 조합 자동 재현
    return replay_grid_combo(latest, None, feat=feat, close=close,
                             output_dir=output_dir, **override_kwargs)


def catboost_correct_actions(feat, close, buy_pool, sell_pool, best_inner, *,
                             horizon=None, dd_limit=None, ru_limit=None,
                             stop_loss_pct=None, anchor_mode=True,
                             anchor_safe_buy=None, anchor_safe_sell=None):
    """선정 조합의 신호를 CatBoost로 보정한다 (요청) — '미래 판단'에 쓰는 게 목적.
       핵심: 타겟을 사후 앵커가 아니라 '그 시점에 검증 가능한 미래 결과'로 둔다.
         - 타겟 = 다음 H일 보유 시 비용 넘는 수익이 났나 (1=보유가 이득 / 0=현금이 이득)
           → 실전에서도 똑같이 '오늘 사면 H일 뒤 이득인가'로 검증되는 라벨이라 일반화 가능.
         - 피처 = 매수/매도 신호강도, 최근 수익/변동성 (그 시점까지 정보만)
         - walk-forward 학습 + 마지막 20%는 out-of-sample(OOS)로 떼어 '진짜 미래 성능' 측정
         - 마지막 날 예측 = 오늘 보유/현금 권고 (실전 판단)
       반환: dict (OOS 성능 포함). 미래 판단에 쓰려면 OOS 개선 여부를 꼭 보라.
    """
    if not globals().get('USE_CATBOOST_CORRECTION', False):
        return None
    try:
        from catboost import CatBoostClassifier
    except Exception as _e:
        print(f"  \u26a0 CatBoost \ubbf8\uc124\uce58 \u2014 \ubcf4\uc815 \uac74\ub108\ub700 (pip install catboost). ({_e})")
        return None

    import numpy as _np
    horizon = horizon if horizon is not None else globals().get('HORIZON_DAYS', 1)
    dd_limit = dd_limit if dd_limit is not None else globals().get('DRAWDOWN_LIMIT_BUY', 0.01)
    ru_limit = ru_limit if ru_limit is not None else globals().get('RUNUP_LIMIT_SELL', 0.01)
    stop_loss_pct = stop_loss_pct if stop_loss_pct is not None else globals().get('STOP_LOSS_PCT', 0.05)
    cost = globals().get('COST_PER_TRADE', 0.004)
    H = int(globals().get('CATBOOST_TARGET_HORIZON', 5))   # 미래 수익 판단 기간

    # 1) 선정 조합으로 daily 백테스트 → 일별 신호
    daily, trades, cur, _bu, _su = daily_ensemble_backtest(
        feat, close, buy_pool, sell_pool,
        K_buy=int(best_inner['K_buy']), K_sell=int(best_inner['K_sell']),
        vote_buy=int(best_inner['vote_buy']), vote_sell=int(best_inner['vote_sell']),
        cost=cost, horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
        stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
        anchor_safe_buy=anchor_safe_buy, anchor_safe_sell=anchor_safe_sell)
    if daily is None or len(daily) < 80:
        print("  \u26a0 CatBoost \ubcf4\uc815 \u2014 \ub370\uc774\ud130 \ubd80\uc871")
        return None

    cl = close.reindex(feat.index).astype(float)
    px = cl.values.astype(_np.float64)
    n = len(px)

    # 2) 타겟 y[i] = i일에 보유 시작해 H일 뒤까지 비용 넘는 수익이 나면 1 (미래 결과 — 검증 가능)
    y = _np.full(n, -1, dtype=_np.int64)
    for i in range(n - H):
        fwd = px[i+H] / px[i] - 1.0
        y[i] = 1 if fwd > (2*cost) else 0
    # 마지막 H일은 미래수익 미확정 → 학습/평가에서 제외 (y=-1)

    # 3) 피처 — 그 시점까지 정보만 (look-ahead 없음)
    bc = daily['buy_count'].values.astype(_np.float64)  if 'buy_count' in daily else _np.zeros(n)
    sc = daily['sell_count'].values.astype(_np.float64) if 'sell_count' in daily else _np.zeros(n)
    bc = bc[:n]; sc = sc[:n]
    if len(bc) < n: bc = _np.pad(bc, (0, n-len(bc)))
    if len(sc) < n: sc = _np.pad(sc, (0, n-len(sc)))
    ret1=_np.zeros(n); ret5=_np.zeros(n); vol10=_np.zeros(n); mom10=_np.zeros(n)
    for i in range(1, n):
        ret1[i]=px[i]/px[i-1]-1.0
        if i>=5:  ret5[i]=px[i]/px[i-5]-1.0
        if i>=10:
            seg=px[i-10:i+1]
            vol10[i]=_np.std(_np.diff(seg)/seg[:-1]) if len(seg)>1 else 0.0
            mom10[i]=px[i]/px[i-10]-1.0
    Kb=max(int(best_inner['K_buy']),1); Ks=max(int(best_inner['K_sell']),1)
    X=_np.column_stack([bc, sc, bc/Kb, sc/Ks, bc-sc, ret1, ret5, vol10, mom10])

    base_hold = (daily['position_pre'].values.astype(float)[:n]
                 if 'position_pre' in daily else _np.zeros(n))

    # 4) OOS 분할 — 마지막 20%는 학습에 안 쓰고 '진짜 미래'로 평가
    # OOS 분할 — OOS_MONTHS(최근 N개월)를 우선 사용, 없으면 OOS_FRACTION
    _oos_m = int(globals().get('OOS_MONTHS', 0))
    if _oos_m > 0:
        _oos_start = feat.index.max() - pd.DateOffset(months=_oos_m)
        split = int((feat.index < _oos_start).sum())
        oos_frac = (n - split) / n if n > 0 else 0.2
    else:
        oos_frac = float(globals().get('CATBOOST_OOS_FRACTION', 0.2))
        split = int(n * (1 - oos_frac))
    min_train = int(globals().get('CATBOOST_MIN_TRAIN_DAYS', 120))
    thr = float(globals().get('CATBOOST_PROB_THRESHOLD', 0.5))
    if split <= min_train or (n - H) <= split:
        print(f"  \u26a0 CatBoost \ubcf4\uc815 \u2014 \ud559\uc2b5/OOS \ubd84\ud560 \ubd88\uac00 (n={n})")
        return None

    pred_hold = _np.full(n, _np.nan)
    # walk-forward (split 이후 구간을 과거로만 학습해 예측)
    step = max(10, (n - min_train)//10)
    t = min_train
    while t < n - H:
        end = min(t+step, n - H)
        mask = y[:t] >= 0
        Xtr, ytr = X[:t][mask], y[:t][mask]
        if len(_np.unique(ytr)) < 2:
            pred_hold[t:end] = ytr[-1] if len(ytr) else 0; t=end; continue
        try:
            m=CatBoostClassifier(iterations=150, depth=4, learning_rate=0.1,
                                 loss_function='Logloss', verbose=False,
                                 random_seed=42, allow_writing_files=False)
            m.fit(Xtr, ytr)
            proba=m.predict_proba(X[t:end])[:,1]
            pred_hold[t:end]=(proba>=thr).astype(float)
        except Exception:
            pred_hold[t:end]=ytr[-1] if len(ytr) else 0
        t=end

    # 5) 수익 시뮬 (단순합산)
    def _sim(hold, lo, hi):
        pos=0; s=0.0; prev=_np.nan
        for i in range(lo, hi):
            if pos==1 and i>0 and not _np.isnan(prev) and prev>0:
                s += (px[i]/prev - 1.0)
            w=hold[i]
            if not _np.isnan(w):
                if w>=0.5 and pos==0: s-=cost; pos=1
                elif w<0.5 and pos==1: s-=cost; pos=0
            prev=px[i]
        return s
    corr_hold = base_hold.copy()
    for i in range(n):
        if not _np.isnan(pred_hold[i]): corr_hold[i]=pred_hold[i]

    # 전체 + OOS 구간 각각 평가 (★OOS가 진짜 미래 성능)
    ret_before_all = _sim(base_hold, 0, n)
    ret_after_all  = _sim(corr_hold, 0, n)
    ret_before_oos = _sim(base_hold, split, n)
    ret_after_oos  = _sim(corr_hold, split, n)
    # OOS 예측 적중률 (보정이 미래를 맞췄나)
    oos_mask = (y[split:n-H] >= 0)
    oos_pred = pred_hold[split:n-H][oos_mask]
    oos_true = y[split:n-H][oos_mask]
    acc = float(_np.mean((oos_pred>=0.5).astype(int) == oos_true)) if len(oos_true)>0 else float('nan')

    # 6) 마지막 날 실전 권고 (오늘 보유/현금?)
    today_reco = None
    try:
        mask=y>=0
        mf=CatBoostClassifier(iterations=150, depth=4, learning_rate=0.1,
                              loss_function='Logloss', verbose=False,
                              random_seed=42, allow_writing_files=False)
        mf.fit(X[mask], y[mask])
        p_today=float(mf.predict_proba(X[-1:])[:,1][0])
        today_reco = ('\ubcf4\uc720(\ub9e4\uc218)' if p_today>=thr else '\ud604\uae08(\ub9e4\ub3c4)', p_today)
    except Exception:
        pass

    print(f"  \U0001f916 CatBoost \ubcf4\uc815 (\ud0c0\uac9f=\ubbf8\ub798 {H}\uc77c \uc218\uc775) \u2014 \u2605OOS(\ucd5c\uadfc {oos_frac*100:.0f}%, \ud559\uc2b5\uc548\ud568) \uc131\ub2a5:")
    print(f"     OOS \uc218\uc775: \ubcf4\uc815\uc804 {ret_before_oos*100:+.1f}% \u2192 \ubcf4\uc815\ud6c4 {ret_after_oos*100:+.1f}%  "
          f"(OOS \uc608\uce21 \uc801\uc911\ub960 {acc*100:.0f}%)")
    if today_reco:
        print(f"     \u2192 \uc624\ub298 \uad8c\uace0: {today_reco[0]} (\ubcf4\uc720\ud655\ub960 {today_reco[1]*100:.0f}%)")
    if ret_after_oos > ret_before_oos:
        print(f"     \u2705 OOS\uc5d0\uc11c \ubcf4\uc815\uc774 \uac1c\uc120 \u2192 \ubbf8\ub798 \ud310\ub2e8\uc5d0 \ub3c4\uc6c0 \uac00\ub2a5\uc131")
    else:
        print(f"     \u26a0 OOS\uc5d0\uc11c \ubcf4\uc815\uc774 \uac1c\uc120 \uc548 \ub428 \u2192 \uc774 \uc885\ubaa9\uc740 \ubcf4\uc815 \ud6a8\uacfc \uc5c6\uc74c (\uc801\uc6a9 \uc8fc\uc758)")
    return {
        'ret_before_all': ret_before_all, 'ret_after_all': ret_after_all,
        'ret_before_oos': ret_before_oos, 'ret_after_oos': ret_after_oos,
        'oos_accuracy': acc, 'oos_improved': bool(ret_after_oos > ret_before_oos),
        'today_reco': today_reco, 'target_horizon': H,
        'dates': feat.index, 'pred_hold': pred_hold,
        'base_hold': base_hold, 'corr_hold': corr_hold,
    }


def catboost_correct_from_excel(filename, *, feat=None, close=None, output_dir=None):
    """엑셀의 ★최적 조합을 읽어 CatBoost 액션 보정을 적용한다 (요청, 노트북에서 호출).
       사용 예: mod.catboost_correct_from_excel('ensemble_search_CAT_2026-06-09.xlsx',
                                                 feat=내_feat, close=내_close)
    """
    base_dir = output_dir if output_dir is not None else OUTPUT_DIR
    path = filename if os.path.isabs(filename) else os.path.join(base_dir, filename)
    if not os.path.exists(path):
        raise RuntimeError(f"파일 없음: {path}")
    wb = load_workbook(path, data_only=True)
    used = _parse_used_settings(wb)
    for _k in ('K_buy', 'vote_buy', 'K_sell', 'vote_sell'):
        if used.get(_k) is None:
            raise RuntimeError(f"'사용된 설정'에서 {_k}를 못 읽음")
    best = {'K_buy': int(used['K_buy']), 'vote_buy': int(used['vote_buy']),
            'K_sell': int(used['K_sell']), 'vote_sell': int(used['vote_sell'])}
    # 지표 풀 — 엑셀의 매수/매도 앙상블 지표 시트
    def _rp(sheet):
        if sheet not in wb.sheetnames: return None
        ws = wb[sheet]; h = {}
        for c in range(1, ws.max_column+1):
            v = ws.cell(3, c).value
            if v is not None: h[str(v).strip()] = c
        rows = []
        for r in range(4, ws.max_row+1):
            ind = ws.cell(r, h.get('지표', 2)).value
            if ind is None or str(ind).strip() == '': continue
            def _cf(name, d=0.0):
                if name in h:
                    try: return float(str(ws.cell(r, h[name]).value).replace('%','').strip())
                    except Exception: return d
                return d
            rows.append({'indicator': str(ind).strip(),
                         'direction': str(ws.cell(r, h.get('방향', 3)).value or '>=').strip(),
                         'threshold': _cf('임계치'), 'pct_label': _cf('분위', 50.0),
                         'n_signals': _cf('신호수'), 'n_success': _cf('성공수'),
                         'success_rate': _cf('성공률'), 'avg_extreme': 0.01, 'score': _cf('점수')})
        return pd.DataFrame(rows) if rows else None
    bp = _rp('매수_앙상블_지표'); sp = _rp('매도_앙상블_지표')
    if bp is None or sp is None:
        raise RuntimeError("앙상블 지표 시트를 못 읽음")
    if feat is None or close is None:
        feat, close, _tk = _resolve_data()
    g = globals(); g['USE_CATBOOST_CORRECTION'] = True
    return catboost_correct_actions(feat, close, bp, sp, best, anchor_mode=g.get('ANCHOR_MODE', True))


def _apply_correction_to_candidates(verified, feat, close, pool_map, *,
                                    horizon, dd_limit, ru_limit, stop_loss_pct,
                                    anchor_mode, anchor_safe_buy, anchor_safe_sell):
    """수익 상위 N개 검증 후보에 지표 가중치 보정을 적용하고 보정 후 지표를 재계산 (요청).
       corr_* 컬럼 추가: corr_total_return / corr_max_drawdown / corr_win_rate /
       corr_buy_success / corr_sell_success / corr_buy_match / corr_sell_match /
       corr_applied(보정 채택 여부). 보정 안 한 후보는 real_* 값을 그대로 corr_*에 복사.
    """
    import numpy as _np
    topn = int(globals().get('CORRECT_VERIFY_TOP_N', 20))
    # 기본: corr_* = real_* (보정 안 됨)
    verified = verified.copy()
    verified['corr_total_return'] = verified['real_total_return']
    verified['corr_max_drawdown'] = verified['real_max_drawdown']
    verified['corr_win_rate']     = verified['real_win_rate']
    verified['corr_buy_success']  = verified['real_buy_success']
    verified['corr_sell_success'] = verified['real_sell_success']
    verified['corr_buy_match']    = verified['real_buy_match']
    verified['corr_sell_match']   = verified['real_sell_match']
    verified['corr_applied']      = False
    verified['corr_n_added_buy']  = 0
    verified['corr_n_added_sell'] = 0

    # 수익 상위 N개만 보정
    order = verified.sort_values('real_total_return', ascending=False).index[:topn]
    print(f"\n  🔧 보정 적용 — 수익 상위 {len(order)}개 후보에 지표 가중치 보정 시도 "
          f"(보정 후 수익/손실/승률 재계산)")
    _t0 = time.time()
    n_adopted = 0
    for cnt, idx in enumerate(order):
        _r = verified.loc[idx]
        _key = (round(float(_r.get('meta_wilson_z', 0)),4),
                int(_r.get('meta_pct_low', 0)), int(_r.get('meta_pct_high', 0)),
                round(float(_r.get('meta_corr_limit', 0)),4))
        pools = pool_map.get(_key)
        if pools is None:
            if cnt == 0:
                print(f"     ⚠ 풀 매칭 실패 — key={_key}, 사용가능 keys={list(pool_map.keys())[:3]}")
            continue
        bp, sp = pools
        # ★ 보정용 전체 후보 풀 (top_n_pool 100 무관, 전체 2800개 기반 점수 후보)
        _fcmap = globals().get('_LAST_FULL_CAND_MAP', {})
        _fc = _fcmap.get(_key, (None, None))
        _fcb, _fcs = _fc if _fc else (None, None)
        best_inner = {'K_buy': int(_r['K_buy']), 'vote_buy': int(_r['vote_buy']),
                      'K_sell': int(_r['K_sell']), 'vote_sell': int(_r['vote_sell'])}
        try:
            aug_bp, aug_sp, bw, sw, mc = anchor_match_correct_weights(
                feat, close, bp, sp, best_inner,
                horizon=horizon, dd_limit=dd_limit, ru_limit=ru_limit,
                stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
                anchor_safe_buy=anchor_safe_buy, anchor_safe_sell=anchor_safe_sell,
                full_cand_buy=_fcb, full_cand_sell=_fcs,
                verbose=(cnt == 0))
        except Exception as _e1:
            if cnt < 3:
                import traceback
                print(f"     ⚠ 보정 함수 에러(후보 {cnt}): {_e1}")
                traceback.print_exc()
            continue
        if not mc.get('adopted', False) or bw is None:
            continue
        # 보정된 가중치로 재백테스트해서 보정 후 전체 지표 산출
        try:
            _d, _t, _cur, _bu, _su = daily_ensemble_backtest(
                feat, close, aug_bp, aug_sp,
                K_buy=len(aug_bp), K_sell=len(aug_sp),
                vote_buy=int(_r['vote_buy']), vote_sell=int(_r['vote_sell']),
                cost=COST_PER_TRADE, horizon=horizon,
                dd_limit=dd_limit, ru_limit=ru_limit,
                stop_loss_pct=stop_loss_pct, anchor_mode=anchor_mode,
                anchor_safe_buy=anchor_safe_buy, anchor_safe_sell=anchor_safe_sell,
                buy_w_override=bw, sell_w_override=sw)
        except Exception as _e2:
            if cnt < 3:
                print(f"     ⚠ 보정후 재백테스트 에러(후보 {cnt}): {_e2}")
            continue
        _cr = float(_cur.get('cum_return_pct', 0.0)) / 100.0
        _cmdd = float(_cur.get('max_drawdown', 0.0))
        _real_mdd = float(_r.get('real_max_drawdown', 0.0))
        # 보정 후 수익이 원래보다 낮으면 채택 안 함 (수익 손해 방지)
        if _cr < float(_r['real_total_return']) - 1e-9:
            continue
        # ★ 보정 후 최대손실(MDD)이 원래보다 깊으면 채택 안 함 (요청: MDD 악화 금지)
        if _cmdd < _real_mdd - 1e-9:
            continue
        verified.at[idx, 'corr_total_return'] = _cr
        verified.at[idx, 'corr_max_drawdown'] = float(_cur.get('max_drawdown', 0.0))
        verified.at[idx, 'corr_win_rate']     = float(_cur.get('win_rate', 0.0))
        verified.at[idx, 'corr_buy_success']  = float(_cur.get('buy_success_rate', 0.0))
        verified.at[idx, 'corr_sell_success'] = float(_cur.get('sell_success_rate', 0.0))
        verified.at[idx, 'corr_buy_match']    = float(_cur.get('anchor_buy_match_rate', 0.0))
        verified.at[idx, 'corr_sell_match']   = float(_cur.get('anchor_sell_match_rate', 0.0))
        verified.at[idx, 'corr_applied']      = True
        verified.at[idx, 'corr_n_added_buy']  = int(mc.get('n_added_buy', 0))
        verified.at[idx, 'corr_n_added_sell'] = int(mc.get('n_added_sell', 0))
        n_adopted += 1
    print(f"     ✓ 보정 채택 {n_adopted}/{len(order)}개 "
          f"(수익 떨어지면 미채택, 경과 {time.time()-_t0:.0f}초)")
    return verified


def anchor_match_correct_weights(feat, close, full_buy_pool, full_sell_pool, best_inner, *,
                                 horizon=None, dd_limit=None, ru_limit=None,
                                 stop_loss_pct=None, anchor_mode=True,
                                 anchor_safe_buy=None, anchor_safe_sell=None,
                                 full_cand_buy=None, full_cand_sell=None,
                                 max_boost=0.8, step=0.1, max_add=15, verbose=True):
    """앵커 미매칭을 '미사용 지표 추가 + 큰 가중치 부여'로 보정 (요청 개편).
       철학(사용자 제안): 기존에 잘 맞은 날엔 영향을 주지 않으면서,
         '간발의 차로 미달(신호 안 뜸)'·'충돌 패배'한 미매칭일에만 그날 켜지는
         미사용 지표를 풀에 추가하고 충분한 가중치를 줘서 vote 임계를 넘긴다.
       ★ 보정 후보는 top_n_pool(100) 풀이 아니라 '점수 매긴 전체 후보(2800개 기반)'에서
         찾는다 (full_cand_buy/sell). top_n_pool은 그대로 두고 보정만 전체 탐색 (사용자 제안).
       - 추가 지표는 표적성(미매칭일 선택적 발화) 높은 것 우선 → 미래 일반화 + 다른날 영향 최소.
       - 추가 후 여러 가중치 크기를 시도해 '실제 수익이 가장 높아지는' 조합 채택.
       - 수익이 기존보다 낮으면 미채택(롤백).
       반환: (buy_pool_used, sell_pool_used, buy_w, sell_w, 결과dict)
    """
    import numpy as _np
    horizon = horizon if horizon is not None else globals().get('HORIZON_DAYS', 1)
    dd_limit = dd_limit if dd_limit is not None else globals().get('DRAWDOWN_LIMIT_BUY', 0.01)
    ru_limit = ru_limit if ru_limit is not None else globals().get('RUNUP_LIMIT_SELL', 0.01)
    stop_loss_pct = stop_loss_pct if stop_loss_pct is not None else globals().get('STOP_LOSS_PCT', 0.05)
    cost = globals().get('COST_PER_TRADE', 0.004)
    Kb = int(best_inner['K_buy']); Ks = int(best_inner['K_sell'])
    vb = int(best_inner['vote_buy']); vs = int(best_inner['vote_sell'])

    base_buy  = full_buy_pool.iloc[:Kb].reset_index(drop=True)
    base_sell = full_sell_pool.iloc[:Ks].reset_index(drop=True)

    # ★ extra(보정 후보) = 전체 후보 풀에서 'base에 없는 지표'. 전체가 없으면 작은 풀에서 폴백.
    def _make_extra(base_pool, small_full, big_cand):
        if big_cand is not None and len(big_cand) > 0:
            used = set(zip(base_pool['indicator'], base_pool['direction'],
                           _np.round(base_pool['threshold'].astype(float), 8)))
            keep = []
            for _, r in big_cand.iterrows():
                kkey = (r['indicator'], r['direction'], round(float(r['threshold']), 8))
                if kkey not in used:
                    keep.append(r)
            if keep:
                return pd.DataFrame(keep).reset_index(drop=True)
        return small_full.iloc[len(base_pool):].reset_index(drop=True)

    extra_buy  = _make_extra(base_buy,  full_buy_pool,  full_cand_buy)
    extra_sell = _make_extra(base_sell, full_sell_pool, full_cand_sell)

    def _run(bpool, spool, bwo=None, swo=None):
        return daily_ensemble_backtest(
            feat, close, bpool, spool, K_buy=len(bpool), K_sell=len(spool),
            vote_buy=vb, vote_sell=vs, cost=cost, horizon=horizon,
            dd_limit=dd_limit, ru_limit=ru_limit, stop_loss_pct=stop_loss_pct,
            anchor_mode=anchor_mode, anchor_safe_buy=anchor_safe_buy,
            anchor_safe_sell=anchor_safe_sell, buy_w_override=bwo, sell_w_override=swo)

    # 0) 기준(보정 전)
    d0, t0, cur0, bu0, su0 = _run(base_buy, base_sell)
    ret0 = float(cur0.get('cum_return_pct', 0.0))
    bm0 = float(cur0.get('anchor_buy_match_rate', 0.0))
    sm0 = float(cur0.get('anchor_sell_match_rate', 0.0))
    mdd0 = float(cur0.get('max_drawdown', 0.0))   # 기준 최대손실 (음수). 보정 후 이보다 깊으면 미채택.

    n = len(feat)
    W = int(globals().get('ANCHOR_MATCH_WINDOW', 2))
    sba = anchor_safe_buy; ssa = anchor_safe_sell
    pos_post = d0['position_pre'].values.astype(int)[:n] if 'position_pre' in d0 else _np.zeros(n, dtype=int)

    # ── 미매칭일 마스크 (앵커 정답인데 포지션 못 맞춘 날) ──
    miss_buy = _np.zeros(n, dtype=bool); miss_sell = _np.zeros(n, dtype=bool)
    matched_buy = _np.zeros(n, dtype=bool); matched_sell = _np.zeros(n, dtype=bool)
    if sba is not None and ssa is not None:
        for i in range(n):
            if i < len(sba) and sba[i] == 1:
                if any(pos_post[i+dd] == 1 for dd in range(0, W+1) if i+dd < n):
                    matched_buy[i] = True
                else:
                    miss_buy[i] = True
            if i < len(ssa) and ssa[i] == 1:
                if any(pos_post[i+dd] == 0 for dd in range(0, W+1) if i+dd < n):
                    matched_sell[i] = True
                else:
                    miss_sell[i] = True

    no_anchor = (sba is None or ssa is None or (miss_buy.sum()+miss_sell.sum()) == 0)
    if verbose:
        print(f"     [보정진단] 미매칭 매수 {int(miss_buy.sum())}일 / 매도 {int(miss_sell.sum())}일, "
              f"매칭 매수 {int(matched_buy.sum())} / 매도 {int(matched_sell.sum())}, "
              f"extra 매수 {len(extra_buy)} / 매도 {len(extra_sell)}, 기준수익 {ret0:+.1f}%")

    def _sig_arr(pool):
        return [_to_signal_array(feat, prow).astype(bool) for _, prow in pool.iterrows()]

    # ── 추가 지표 선택: 미매칭일에 자주 켜지고, 매칭/정상일엔 덜 켜지는 것 ──
    #   targeting score = (미매칭일 켜짐 비율) - 0.5*(매칭일 켜짐 비율)
    def _pick_extra(extra_pool, miss_mask, matched_mask):
        """미매칭일을 '실제로 덮는' 최소 지표만 선택 (요청: 항상 15개 추가 방지).
           표적성(미매칭에 선택적 발화) 높은 순으로 보되, '새 미매칭일을 추가로 덮는'
           지표만 채택하고 모든 미매칭일이 덮이면 멈춘다 → 추가 수가 +1, +2 처럼 가변.
           표적성 = 미매칭일 적중 / (그 외 날 발화 + eps).
        """
        if len(extra_pool) == 0 or miss_mask.sum() == 0:
            return [], _np.array([])
        sigs = _sig_arr(extra_pool)
        nmiss = max(1, int(miss_mask.sum()))
        other_mask = ~miss_mask
        nother = max(1, int(other_mask.sum()))
        scores = []
        for arr in sigs:
            hit_miss  = int((arr & miss_mask).sum()) / nmiss
            hit_other = int((arr & other_mask).sum()) / nother
            if hit_miss <= 0:
                scores.append(0.0); continue
            scores.append(hit_miss / (hit_other + 0.05))
        scores = _np.array(scores, dtype=float)
        order = _np.argsort(-scores)
        # ★ 그리디 셋커버: 표적성 높은 순으로, '새 미매칭일을 덮는' 지표만 채택.
        #   모든 미매칭일이 덮이거나 max_add에 도달하면 중단 → 최소 개수만 추가.
        covered = _np.zeros_like(miss_mask, dtype=bool)
        total_miss = int(miss_mask.sum())
        pick = []
        for o in order:
            if scores[o] < 1.0:
                break                       # 표적성 미달이면 더 안 봄
            new_cover = sigs[o] & miss_mask & ~covered
            if int(new_cover.sum()) == 0:
                continue                    # 새로 덮는 미매칭일 없음 → 불필요, 건너뜀
            pick.append(int(o))
            covered |= (sigs[o] & miss_mask)
            if int(covered.sum()) >= total_miss:
                break                       # 미매칭일 전부 덮음 → 종료
            if len(pick) >= max_add:
                break                       # 안전 상한
        if not pick:  # 표적성 1.0 넘는 게 없으면 표적성 상위 소수라도 (최대 3개)
            pick = [int(o) for o in order if scores[o] > 0][:3]
        return pick, scores

    pick_b, bsc = _pick_extra(extra_buy, miss_buy, matched_buy)
    pick_s, ssc = _pick_extra(extra_sell, miss_sell, matched_sell)

    add_buy  = extra_buy.iloc[pick_b].reset_index(drop=True) if pick_b else extra_buy.iloc[[]].reset_index(drop=True)
    add_sell = extra_sell.iloc[pick_s].reset_index(drop=True) if pick_s else extra_sell.iloc[[]].reset_index(drop=True)
    aug_buy  = pd.concat([base_buy, add_buy], ignore_index=True)
    aug_sell = pd.concat([base_sell, add_sell], ignore_index=True)
    if verbose:
        _bt = f"{max(bsc[pick_b]):.1f}" if len(pick_b) else "-"
        _st = f"{max(ssc[pick_s]):.1f}" if len(pick_s) else "-"
        print(f"     [보정진단] 추가 후보 지표(표적성≥1.0) — 매수 {len(add_buy)}개(최고표적성 {_bt}), "
              f"매도 {len(add_sell)}개(최고표적성 {_st})")

    # 기본 가중치 (기존 K개) — score 기반 또는 1.0
    def _wbase(pool):
        if globals().get('USE_WEIGHTED_VOTE', False):
            return compute_vote_weights(pool['score'].values, globals().get('WEIGHT_MAX_RATIO', 1.6))
        return _np.ones(len(pool))
    bw_core = _wbase(base_buy); sw_core = _wbase(base_sell)

    # ── 추가 지표에 줄 가중치를 여러 크기로 시도 (vote 임계를 넘기게) ──
    #   add_w 후보: 추가 지표 1개당 가중치. 충분히 커야 간발 미달(1~2점차)을 넘긴다.
    best = {'adopted': False, 'ret_before': ret0, 'bm_before': bm0, 'sm_before': sm0,
            'ret_after': ret0, 'bm_after': bm0, 'sm_after': sm0,
            'mdd_before': mdd0, 'mdd_after': mdd0,
            'bw': None, 'sw': None, 'n_added_buy': int(len(add_buy)), 'n_added_sell': int(len(add_sell))}

    if no_anchor or (len(add_buy) == 0 and len(add_sell) == 0):
        if verbose:
            print(f"  \u2139 \uc218\uc775 \ubcf4\uc815 \u2014 \ucd94\uac00\ud560 \ubbf8\uc0ac\uc6a9 \uc9c0\ud45c \uc5c6\uc74c \u2192 \uae30\uc874 \uadf8\ub9ac\ub4dc \uc720\uc9c0")
        return base_buy, base_sell, None, None, best

    nb = len(add_buy); ns = len(add_sell)
    # 추가 가중치: 작게(0.2)부터 — 기존 매매 영향 최소화하며 미달일만 살짝 밀어줌.
    #   촘촘히 올려 '수익 안 떨어지고 매칭 느는' 지점을 찾는다.

    # 기존 풀 내 지표 중 '미매칭일에 켜지는' 것 = 충돌일에 밀어줄 후보 (가중치 상향용)
    def _core_boost_dir(base_pool, miss_mask, matched_mask):
        if miss_mask.sum() == 0:
            return _np.zeros(len(base_pool))
        sigs = _sig_arr(base_pool)
        nmiss = max(1, int(miss_mask.sum()))
        d = []
        for arr in sigs:
            d.append(int((arr & miss_mask).sum()) / nmiss)   # 미매칭일에 켜지는 비율
        d = _np.array(d)
        return d/d.max() if d.max() > 0 else d
    bcore_dir = _core_boost_dir(base_buy, miss_buy, matched_buy)
    score_dir = _core_boost_dir(base_sell, miss_sell, matched_sell)

    _tried = []
    for add_w in (0.2, 0.3, 0.5, 0.8, 1.2, 1.8, 2.5):
        for core_boost in (0.0, 0.3, 0.6):
            bcore_w = bw_core * (1.0 + core_boost * bcore_dir)
            score_w = sw_core * (1.0 + core_boost * score_dir)
            bw = _np.concatenate([bcore_w, _np.full(nb, add_w)]) if nb > 0 else bcore_w
            sw = _np.concatenate([score_w, _np.full(ns, add_w)]) if ns > 0 else score_w
            d1, t1, cur1, _b1, _s1 = _run(aug_buy, aug_sell, bw, sw)
            ret1 = float(cur1.get('cum_return_pct', 0.0))
            bm1 = float(cur1.get('anchor_buy_match_rate', 0.0))
            sm1 = float(cur1.get('anchor_sell_match_rate', 0.0))
            mdd1 = float(cur1.get('max_drawdown', 0.0))
            _tried.append((ret1, mdd1))
            # ★ 채택 조건 (요청):
            #   1) 수익이 기존(ret0)보다 오를 것 (수익 최대화)
            #   2) 최대손실(MDD)이 기존(mdd0)보다 깊지 않을 것 — 같거나 더 얕아야 함
            #   둘 다 만족할 때만 채택. 수익만 오르고 손실이 깊어지면 미채택.
            mdd_ok = (mdd1 >= mdd0 - 1e-9)   # mdd는 음수: -2%보다 -6%가 작음 → mdd1>=mdd0이면 안 깊어짐
            if ret1 > ret0 + 1e-9 and mdd_ok and ret1 > best['ret_after'] + 1e-9:
                best.update({'adopted': True, 'ret_after': ret1, 'bm_after': bm1, 'sm_after': sm1,
                             'mdd_after': mdd1, 'bw': bw.copy(), 'sw': sw.copy()})
    if verbose and _tried:
        # MDD 제약(악화 안 됨)을 만족하는 시도 중 최고수익
        _ok = [r for (r, m) in _tried if m >= mdd0 - 1e-9]
        _mx_all = max(r for (r, m) in _tried)
        _mx_ok = max(_ok) if _ok else None
        if _mx_ok is not None and _mx_ok > ret0:
            print(f"     [보정진단] {len(_tried)}회 시도 — MDD유지하며 최고수익 {_mx_ok:+.1f}% "
                  f"(기준 {ret0:+.1f}%, MDD {mdd0*100:.2f}%) 개선됨✓")
        else:
            print(f"     [보정진단] {len(_tried)}회 시도 — 최고수익 {_mx_all:+.1f}%지만 "
                  f"MDD 악화로 미채택 (기준수익 {ret0:+.1f}%, 기준MDD {mdd0*100:.2f}%)")

    if best['adopted']:
        if verbose:
            print(f"  \U0001f3af \uc218\uc775 \ubcf4\uc815 (\ubbf8\uc0ac\uc6a9 \uc9c0\ud45c \ucd94\uac00 + \uac00\uc911\uce58 \ubd80\uc5ec, \ub0a0\uc9dc \ud655\uc7a5 \uc548\ud568):")
            print(f"     \uc218\uc775 {best['ret_before']:+.1f}% \u2192 {best['ret_after']:+.1f}% "
                  f"(+{best['ret_after']-best['ret_before']:.1f}%p)")
            print(f"     \ucd5c\ub300\uc190\uc2e4 {best['mdd_before']*100:.2f}% \u2192 {best['mdd_after']*100:.2f}% "
                  f"(\uc545\ud654 \uc5c6\uc74c \u2014 \uc870\uac74 \ucda9\uc871)")
            print(f"     \ub9e4\uc218\ub9e4\uce6d {best['bm_before']*100:.1f}% \u2192 {best['bm_after']*100:.1f}%, "
                  f"\ub9e4\ub3c4\ub9e4\uce6d {best['sm_before']*100:.1f}% \u2192 {best['sm_after']*100:.1f}%")
            print(f"     \ucd94\uac00\ub41c \ubbf8\uc0ac\uc6a9 \uc9c0\ud45c: \ub9e4\uc218 {best['n_added_buy']}\uac1c, \ub9e4\ub3c4 {best['n_added_sell']}\uac1c")
        return aug_buy, aug_sell, best['bw'], best['sw'], best
    else:
        if verbose:
            print(f"  \u2139 \uc218\uc775 \ubcf4\uc815 \u2014 \uac1c\uc120\ub418\ub294 \ucd94\uac00 \uc5c6\uc74c \u2192 \uae30\uc874 \uadf8\ub9ac\ub4dc \uc720\uc9c0")
        return base_buy, base_sell, None, None, best


def diagnose_trades(trades_df, close, feat_index, anchor_prio=None, *, profit_floor=0.01):
    """실제 거래 내역을 진단한다 (요청) — 앵커 날짜 외의 거래도 포함.
       각 거래에 대해:
        - net_return이 손실(<0)이거나 profit_floor(1%) 이하면 '틀린 거래' 후보
        - 매수일/매도일이 앵커(우선순위1 최대수익스윙 / 우선순위2 1%이상)와 맞는지
        - 잘못된 매수(저점 아닌 곳에서 삼) / 잘못된 매도(고점 아닌 곳에서 팜) 진단
       반환: 진단 DataFrame (거래별 1행) — Excel '거래_진단' 시트 + 보정 대상 식별용
    """
    import numpy as _np
    if trades_df is None or len(trades_df) == 0:
        return pd.DataFrame()
    norm_idx = pd.DatetimeIndex(feat_index).normalize()
    # 앵커 우선순위 집합 (날짜 문자열)
    ap = anchor_prio or {}
    buy_p1  = set(pd.Timestamp(d).normalize() for d in ap.get('buy_p1', set()))
    sell_p1 = set(pd.Timestamp(d).normalize() for d in ap.get('sell_p1', set()))
    buy_p2  = set(pd.Timestamp(d).normalize() for d in ap.get('buy_p2', set()))
    sell_p2 = set(pd.Timestamp(d).normalize() for d in ap.get('sell_p2', set()))

    rows = []
    for _, tr in trades_df.iterrows():
        ed = pd.Timestamp(tr['entry_date']).normalize() if pd.notna(tr.get('entry_date')) else None
        xd = pd.Timestamp(tr['exit_date']).normalize()  if pd.notna(tr.get('exit_date'))  else None
        ret = float(tr.get('net_return_%', 0.0)) / 100.0
        is_loss = ret < 0
        is_thin = (ret >= 0) and (ret <= profit_floor)   # 1% 이하 수익
        is_wrong = is_loss or is_thin

        # 매수일 진단
        if ed in buy_p1:    buy_status = '✓ 최대수익스윙 저점(P1)'
        elif ed in buy_p2:  buy_status = '○ 1%이상 저점(P2)'
        else:               buy_status = '✗ 앵커 아닌 곳에서 매수'
        # 매도일 진단
        if xd in sell_p1:    sell_status = '✓ 최대수익스윙 고점(P1)'
        elif xd in sell_p2:  sell_status = '○ 1%이상 고점(P2)'
        else:                sell_status = '✗ 앵커 아닌 곳에서 매도'

        buy_ok  = (ed in buy_p1) or (ed in buy_p2)
        sell_ok = (xd in sell_p1) or (xd in sell_p2)

        if is_loss:        verdict = '손실 거래 → 보정 대상'
        elif is_thin:      verdict = f'수익 {ret*100:.1f}%(≤{profit_floor*100:.0f}%) → 보정 대상'
        else:              verdict = '정상(수익)'

        rows.append({
            'trade_no': int(tr.get('trade_no', 0)),
            'entry_date': ed, 'exit_date': xd,
            'net_return_%': ret * 100.0,
            'days_held': int(tr.get('days_held', 0)),
            'exit_reason': tr.get('exit_reason', ''),
            '매수_진단': buy_status, '매도_진단': sell_status,
            '매수_앵커일치': buy_ok, '매도_앵커일치': sell_ok,
            '판정': verdict,
            '보정대상': is_wrong,
        })
    return pd.DataFrame(rows)


# ════════════════════════════════════════════════════════════════════════
#                  ★ 다중 티커 일괄 분석 시스템
# ════════════════════════════════════════════════════════════════════════
TICKERS = ['MPC']

AUTO_TUNE = {
    'EVAL_START':                  False,
    'HORIZON_DAYS':                False,
    'DRAWDOWN_LIMIT_BUY':          False,
    'RUNUP_LIMIT_SELL':            False,
    'AUTO_ANCHOR_WINDOW':          False,
    'AUTO_ANCHOR_LOOKFORWARD':     False,
    'AUTO_ANCHOR_MIN_RISE':        False,
    'AUTO_ANCHOR_MIN_DROP':        False,
    'AUTO_ANCHOR_PRICE_TOLERANCE': False,
    'wilson_z':                    False,
    'pct_range':                   False,
    'corr_limit':                  False,
}

MULTI_SUMMARY_FILE = 'ensemble_summary_all_tickers.xlsx'


def analyze_ticker_volatility(close):
    p = close.values.astype(np.float64)
    n = len(p)
    if n < 20:
        return {'daily_vol': 0.02, 'daily_vol_pct': 2.0, 'weekly_vol': 0.04,
                'n_days': n, 'typical_5d_move': 5.0, 'typical_10d_move': 7.0,
                'max_local_drawdown_5d': 3.0, 'max_local_runup_5d': 3.0,
                'price_dispersion': 0.2}
    rets = np.diff(p) / p[:-1]
    daily_vol = float(np.std(rets))
    daily_vol_pct = float(np.mean(np.abs(rets))) * 100.0

    moves_5d = []
    moves_10d = []
    dd_5d = []
    ru_5d = []
    for i in range(5, n - 5):
        win5 = p[i:i+6]
        moves_5d.append(abs(win5.max() / win5.min() - 1.0) * 100.0)
        start_p = p[i]
        if start_p > 0:
            end = min(i + 5, n - 1)
            future = p[i+1:end+1] if end >= i+1 else p[i:i+1]
            if len(future) > 0:
                dd_5d.append((future.min() / start_p - 1.0) * 100.0)
                ru_5d.append((future.max() / start_p - 1.0) * 100.0)
    for i in range(10, n - 10):
        win10 = p[i:i+11]
        moves_10d.append(abs(win10.max() / win10.min() - 1.0) * 100.0)

    weekly_rets = []
    for i in range(0, n - 5, 5):
        if p[i] > 0:
            weekly_rets.append(p[min(i+5, n-1)] / p[i] - 1.0)
    weekly_vol = float(np.std(weekly_rets)) if weekly_rets else daily_vol * np.sqrt(5)

    return {
        'daily_vol': daily_vol,
        'daily_vol_pct': daily_vol_pct,
        'weekly_vol': weekly_vol,
        'n_days': n,
        'typical_5d_move': float(np.median(moves_5d)) if moves_5d else 5.0,
        'typical_10d_move': float(np.median(moves_10d)) if moves_10d else 7.0,
        'max_local_drawdown_5d': float(np.median(dd_5d)) if dd_5d else -3.0,
        'max_local_runup_5d':    float(np.median(ru_5d)) if ru_5d else 3.0,
        'price_dispersion': float(np.std(p) / max(np.mean(p), 1e-10)),
    }


def auto_tune_variables(close, feat=None, *, auto_tune_flags=None):
    flags = auto_tune_flags or AUTO_TUNE
    stats = analyze_ticker_volatility(close)
    v   = stats['daily_vol']
    vp  = stats['daily_vol_pct']
    nd  = stats['n_days']
    m5  = stats['typical_5d_move']
    m10 = stats['typical_10d_move']

    tuned = {}

    if flags.get('EVAL_START'):
        try:
            start_idx = max(0, min(20, len(close) - 1))
            tuned['EVAL_START'] = close.index[start_idx].strftime('%Y-%m-%d')
        except Exception:
            tuned['EVAL_START'] = None

    if flags.get('HORIZON_DAYS'):
        if vp >= 3.0:    tuned['HORIZON_DAYS'] = 3
        elif vp >= 2.0:  tuned['HORIZON_DAYS'] = 4
        elif vp >= 1.0:  tuned['HORIZON_DAYS'] = 5
        else:            tuned['HORIZON_DAYS'] = 7

    if flags.get('DRAWDOWN_LIMIT_BUY'):
        dd = max(0.005, min(0.05, v * 0.5))
        tuned['DRAWDOWN_LIMIT_BUY'] = round(dd, 4)

    if flags.get('RUNUP_LIMIT_SELL'):
        ru = max(0.005, min(0.05, v * 0.5))
        tuned['RUNUP_LIMIT_SELL'] = round(ru, 4)

    if flags.get('AUTO_ANCHOR_WINDOW'):
        if vp >= 2.5:    tuned['AUTO_ANCHOR_WINDOW'] = 1
        elif vp >= 1.5:  tuned['AUTO_ANCHOR_WINDOW'] = 2
        elif vp >= 1.0:  tuned['AUTO_ANCHOR_WINDOW'] = 3
        else:            tuned['AUTO_ANCHOR_WINDOW'] = 5

    if flags.get('AUTO_ANCHOR_LOOKFORWARD'):
        if vp >= 2.5:    tuned['AUTO_ANCHOR_LOOKFORWARD'] = 1
        elif vp >= 1.5:  tuned['AUTO_ANCHOR_LOOKFORWARD'] = 2
        elif vp >= 1.0:  tuned['AUTO_ANCHOR_LOOKFORWARD'] = 3
        else:            tuned['AUTO_ANCHOR_LOOKFORWARD'] = 5

    if flags.get('AUTO_ANCHOR_MIN_RISE'):
        rise = max(0.015, min(0.08, m5 / 100.0 * 0.6))
        tuned['AUTO_ANCHOR_MIN_RISE'] = round(rise, 4)

    if flags.get('AUTO_ANCHOR_MIN_DROP'):
        drop = max(0.015, min(0.08, m5 / 100.0 * 0.6))
        tuned['AUTO_ANCHOR_MIN_DROP'] = round(drop, 4)

    if flags.get('AUTO_ANCHOR_PRICE_TOLERANCE'):
        tol = max(0.002, min(0.015, v * 0.3))
        tuned['AUTO_ANCHOR_PRICE_TOLERANCE'] = round(tol, 4)

    if flags.get('wilson_z'):
        if   nd < 100:  tuned['wilson_z'] = 1.96
        elif nd < 300:  tuned['wilson_z'] = 1.96
        else:           tuned['wilson_z'] = 1.96

    if flags.get('pct_range'):
        if vp >= 2.0:    tuned['pct_range'] = (5, 95)
        elif vp >= 1.0:  tuned['pct_range'] = (10, 90)
        else:            tuned['pct_range'] = (15, 85)

    if flags.get('corr_limit'):
        if   nd < 100:  tuned['corr_limit'] = 0.85
        elif nd < 300:  tuned['corr_limit'] = 0.80
        else:           tuned['corr_limit'] = 0.75

    return tuned, stats


def _restore_tuned_vars(raw):
    """변경5: _load_summary_state가 읽은 문자열 변수값을 effective_vars 형태로 복원."""
    if not raw:
        return {}
    def _f(v):
        if v in (None, '—', ''): return None
        s = str(v).replace('%', '').replace('+', '').strip()
        try: return float(s)
        except Exception: return None
    def _i(v):
        f = _f(v)
        return int(f) if f is not None else None
    def _pct(v):
        if v in (None, '—', ''): return None
        s = str(v).replace('(', '').replace(')', '').replace('[', '').replace(']', '')
        try:
            parts = [int(float(x.strip())) for x in s.split(',') if x.strip()]
            return tuple(parts) if len(parts) == 2 else None
        except Exception:
            return None
    out = {}
    if raw.get('EVAL_START') not in (None, '—', ''):
        out['EVAL_START'] = str(raw['EVAL_START']).strip()
    if _i(raw.get('HORIZON_DAYS')) is not None: out['HORIZON_DAYS'] = _i(raw['HORIZON_DAYS'])
    for src, dst in [('DRAWDOWN_LIMIT_BUY','DRAWDOWN_LIMIT_BUY'),
                     ('RUNUP_LIMIT_SELL','RUNUP_LIMIT_SELL'),
                     ('AUTO_ANCHOR_MIN_RISE','AUTO_ANCHOR_MIN_RISE'),
                     ('AUTO_ANCHOR_MIN_DROP','AUTO_ANCHOR_MIN_DROP'),
                     ('AUTO_ANCHOR_PRICE_TOLERANCE','AUTO_ANCHOR_PRICE_TOLERANCE')]:
        v = raw.get(src)
        f = _f(v)
        if f is not None:
            out[dst] = f / 100.0 if '%' in str(v) else f
    for src, dst in [('AUTO_ANCHOR_WINDOW','AUTO_ANCHOR_WINDOW'),
                     ('AUTO_ANCHOR_LOOKFORWARD','AUTO_ANCHOR_LOOKFORWARD'),
                     ('min_signals','min_signals'), ('top_n_pool','top_n_pool')]:
        if _i(raw.get(src)) is not None: out[dst] = _i(raw[src])
    if _f(raw.get('wilson_z')) is not None: out['wilson_z'] = _f(raw['wilson_z'])
    if _f(raw.get('corr_limit')) is not None: out['corr_limit'] = _f(raw['corr_limit'])
    if _pct(raw.get('pct_range')) is not None: out['pct_range'] = _pct(raw['pct_range'])
    return out


def _load_summary_state(summary_file):
    """이미 분석 완료된 티커 + (변경5) 사용된 변수값(tuned_vars) 복원."""
    if not os.path.exists(summary_file):
        return {}, []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(summary_file, data_only=True)
        if '결과 요약' not in wb.sheetnames:
            return {}, []
        ws = wb['결과 요약']
        done = {}
        order = []
        header_row = None
        for ri, row in enumerate(ws.iter_rows(values_only=True), 1):
            if ri == 1: continue
            if ri == 2:
                header_row = list(row); continue
            if row[0] is None: continue
            ticker = str(row[0]).strip()
            if not ticker or ticker == '티커': continue
            done[ticker] = dict(zip(header_row, row)) if header_row else {}
            order.append(ticker)

        # ★ 변경5: '사용된 변수값' 시트에서 tuned_vars 복원
        if '사용된 변수값' in wb.sheetnames:
            ws2 = wb['사용된 변수값']
            hdr2 = None
            for ri, row in enumerate(ws2.iter_rows(values_only=True), 1):
                if ri == 2:
                    hdr2 = list(row); continue
                if ri <= 2 or not row or row[0] is None: continue
                tk = str(row[0]).strip()
                if tk not in done: continue
                d2 = dict(zip(hdr2, row)) if hdr2 else {}
                done[tk]['_tuned_vars_raw'] = {
                    'EVAL_START': d2.get('EVAL_START'),
                    'HORIZON_DAYS': d2.get('HORIZON'),
                    'DRAWDOWN_LIMIT_BUY': d2.get('DD한도'),
                    'RUNUP_LIMIT_SELL': d2.get('RU한도'),
                    'AUTO_ANCHOR_WINDOW': d2.get('ANCHOR_WIN'),
                    'AUTO_ANCHOR_LOOKFORWARD': d2.get('ANCHOR_LF'),
                    'AUTO_ANCHOR_MIN_RISE': d2.get('MIN_RISE'),
                    'AUTO_ANCHOR_MIN_DROP': d2.get('MIN_DROP'),
                    'AUTO_ANCHOR_PRICE_TOLERANCE': d2.get('PRICE_TOL'),
                    'wilson_z': d2.get('wilson_z'),
                    'pct_range': d2.get('pct_range'),
                    'corr_limit': d2.get('corr_limit'),
                    'min_signals': d2.get('MIN_SIG'),
                    'top_n_pool': d2.get('TOP_N'),
                    'tuning_source': d2.get('튜닝상태'),
                }
        return done, order
    except Exception as e:
        print(f"  ⚠ 요약 파일 로드 실패: {e}")
        return {}, []


def _save_summary_excel(summary_records, summary_file, all_tickers, *,
                         tuning_log=None):
    wb = Workbook(); wb.remove(wb.active)

    ws = wb.create_sheet('결과 요약', 0); ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = (f'전체 티커 분석 요약 — {len(summary_records)}/{len(all_tickers)} 완료  '
                            f'({datetime.now().strftime("%Y-%m-%d %H:%M:%S")})')
    ws.cell(1, 1).font = Font(bold=True, size=16, color='1F3864')
    ws.merge_cells('A1:T1')

    headers = ['티커', '상태', '신호', '포지션', '오늘가',
               '누적수익%', 'B&H%', 'vsB&H%p', '📈상승일전략%', '📈상승일B&H%',
               'MDD%', '거래수', '승률%',
               'BalAcc%', '매수BalAcc%', '매도BalAcc%',
               '⚓매칭률%', '⚓매수%', '⚓매도%',
               'K_b/v_b', 'K_s/v_s', '분석일시']
    _hdr(ws, 2, headers)

    sig_fills = {'🟢 매수': _BUY, '🔴 매도': _SELL, '📈 보유': _HOLD,
                  '💵 현금': _CASH, '⚔ 충돌': _CONF, '— 에러': _BAD}
    for ri, ticker in enumerate(all_tickers, 3):
        if ticker not in summary_records:
            ws.cell(ri, 1).value = ticker
            ws.cell(ri, 1).font = Font(bold=True, size=11)
            ws.cell(ri, 2).value = '대기 중'
            ws.cell(ri, 2).font = Font(size=10, italic=True, color='888888')
            for ci in range(1, len(headers) + 1):
                ws.cell(ri, ci).border = _TH
                ws.cell(ri, ci).alignment = Alignment(horizontal='center')
            continue
        rec = summary_records[ticker]
        status = rec.get('status', '완료')
        signal = rec.get('signal', '—')
        is_err = (status == '에러')
        vals = [
            ticker,
            status, signal, rec.get('position', '—'),
            f"${rec['close']:.2f}" if rec.get('close') is not None else '—',
            f"{rec['cum_return_pct']:+.2f}" if rec.get('cum_return_pct') is not None else '—',
            f"{rec['bh_pct']:+.2f}" if rec.get('bh_pct') is not None else '—',
            f"{rec['vs_bh_pp']:+.2f}" if rec.get('vs_bh_pp') is not None else '—',
            f"{rec['up_cum_return_pct']:+.2f}" if rec.get('up_cum_return_pct') is not None else '—',
            f"{rec['bh_up_pct']:+.2f}" if rec.get('bh_up_pct') is not None else '—',
            f"{rec['mdd_pct']:.2f}" if rec.get('mdd_pct') is not None else '—',
            rec.get('n_trades', '—'),
            f"{rec['win_rate_pct']:.1f}" if rec.get('win_rate_pct') is not None else '—',
            f"{rec['balacc_pct']:.1f}" if rec.get('balacc_pct') is not None else '—',
            f"{rec['buy_balacc_pct']:.1f}" if rec.get('buy_balacc_pct') is not None else '—',
            f"{rec['sell_balacc_pct']:.1f}" if rec.get('sell_balacc_pct') is not None else '—',
            f"{rec['match_avg_pct']:.1f}" if rec.get('match_avg_pct') is not None else '—',
            f"{rec['match_buy_pct']:.1f}" if rec.get('match_buy_pct') is not None else '—',
            f"{rec['match_sell_pct']:.1f}" if rec.get('match_sell_pct') is not None else '—',
            rec.get('K_b_v_b', '—'),
            rec.get('K_s_v_s', '—'),
            rec.get('analyzed_at', '—'),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci); c.value = v; c.border = _TH
            c.alignment = Alignment(horizontal='center')
            c.font = Font(size=10, bold=(ci == 1 or ci == 3))
            if ri % 2 == 0: c.fill = _ALT
        sig_str = str(signal)
        for k, fl in sig_fills.items():
            if k in sig_str:
                ws.cell(ri, 3).fill = fl
                if '매수' in k: ws.cell(ri, 3).font = Font(bold=True, size=11, color='006100')
                elif '매도' in k: ws.cell(ri, 3).font = Font(bold=True, size=11, color='C00000')
                elif '충돌' in k: ws.cell(ri, 3).font = Font(bold=True, size=11, color='806000')
                break
        if is_err:
            ws.cell(ri, 2).fill = _BAD
            ws.cell(ri, 2).font = Font(bold=True, size=10, color='C00000')
        else:
            ws.cell(ri, 2).fill = _GOOD
            ws.cell(ri, 2).font = Font(bold=True, size=10, color='006100')
        if rec.get('cum_return_pct') is not None and rec.get('bh_pct') is not None:
            ws.cell(ri, 6).fill = _ret_fill(rec['cum_return_pct'] / 100.0, rec['bh_pct'] / 100.0)
            ws.cell(ri, 6).font = Font(bold=True, size=10,
                                          color='006100' if rec['cum_return_pct'] > 0 else 'C00000')
        if rec.get('up_cum_return_pct') is not None:
            ws.cell(ri, 9).font = Font(bold=True, size=10, color='006100')
        if rec.get('balacc_pct') is not None:
            ws.cell(ri, 14).fill = _success_fill(rec['balacc_pct'] / 100.0)
        if rec.get('match_avg_pct') is not None:
            ws.cell(ri, 17).fill = _success_fill(rec['match_avg_pct'] / 100.0)

    widths = [10, 10, 12, 14, 10, 11, 10, 10, 12, 12, 9, 8, 8, 9, 12, 12, 10, 9, 9, 11, 11, 18]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'B3'

    ws = wb.create_sheet('사용된 변수값'); ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = '티커별 사용된 변수값 (자동 튜닝 결과)'
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells('A1:N1')
    var_headers = ['티커', 'EVAL_START', 'HORIZON', 'DD한도', 'RU한도',
                    'ANCHOR_WIN', 'ANCHOR_LF', 'MIN_RISE', 'MIN_DROP',
                    'PRICE_TOL', 'wilson_z', 'pct_range', 'corr_limit',
                    'MIN_SIG', 'TOP_N', 'K_b/v_b', 'K_s/v_s', '튜닝상태']
    _hdr(ws, 2, var_headers)
    for ri, ticker in enumerate(all_tickers, 3):
        if ticker not in summary_records:
            ws.cell(ri, 1).value = ticker
            ws.cell(ri, 1).font = Font(bold=True, size=11)
            continue
        rec = summary_records[ticker]
        tv = rec.get('tuned_vars', {})
        is_tuned = rec.get('tuning_applied', False)
        tuning_src = rec.get('tuning_source', 'AUTO' if is_tuned else '기본값')
        kb = tv.get('K_buy'); vb = tv.get('vote_buy')
        ks = tv.get('K_sell'); vs = tv.get('vote_sell')
        if kb is not None and vb is not None:
            kbvb = f"{int(kb)}/{int(vb)}"
        else:
            kbvb = str(rec.get('K_b_v_b', '—'))
        if ks is not None and vs is not None:
            ksvs = f"{int(ks)}/{int(vs)}"
        else:
            ksvs = str(rec.get('K_s_v_s', '—'))
        vals = [ticker,
                str(tv.get('EVAL_START', '—')),
                tv.get('HORIZON_DAYS', '—'),
                f"{tv.get('DRAWDOWN_LIMIT_BUY', 0)*100:.2f}%" if tv.get('DRAWDOWN_LIMIT_BUY') else '—',
                f"{tv.get('RUNUP_LIMIT_SELL', 0)*100:.2f}%" if tv.get('RUNUP_LIMIT_SELL') else '—',
                tv.get('AUTO_ANCHOR_WINDOW', '—'),
                tv.get('AUTO_ANCHOR_LOOKFORWARD', '—'),
                f"{tv.get('AUTO_ANCHOR_MIN_RISE', 0)*100:.2f}%" if tv.get('AUTO_ANCHOR_MIN_RISE') else '—',
                f"{tv.get('AUTO_ANCHOR_MIN_DROP', 0)*100:.2f}%" if tv.get('AUTO_ANCHOR_MIN_DROP') else '—',
                f"{tv.get('AUTO_ANCHOR_PRICE_TOLERANCE', 0)*100:.2f}%" if tv.get('AUTO_ANCHOR_PRICE_TOLERANCE') else '—',
                tv.get('wilson_z', '—'),
                str(tv.get('pct_range', '—')),
                tv.get('corr_limit', '—'),
                tv.get('min_signals', '—') if tv.get('min_signals') is not None else '—',
                tv.get('top_n_pool', '—') if tv.get('top_n_pool') is not None else '—',
                kbvb,
                ksvs,
                tuning_src,
                ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci); c.value = v; c.border = _TH
            c.alignment = Alignment(horizontal='center')
            c.font = Font(size=10, bold=(ci == 1))
            if ri % 2 == 0: c.fill = _ALT
        status_col = len(var_headers)
        if tuning_src == 'AUTO':
            ws.cell(ri, status_col).fill = _GOOD
            ws.cell(ri, status_col).font = Font(bold=True, size=10, color='006100')
        elif tuning_src == 'EXCEL':
            ws.cell(ri, status_col).fill = _HOLD
            ws.cell(ri, status_col).font = Font(bold=True, size=10, color='0070C0')
        else:
            ws.cell(ri, status_col).fill = _MID
    var_widths = [10, 13, 9, 8, 8, 11, 11, 10, 10, 10, 9, 10, 10, 8, 7, 9, 9, 11]
    for ci, w in enumerate(var_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'B3'

    ws = wb.create_sheet('통계 분석'); ws.sheet_view.showGridLines = False
    ws.cell(1, 1).value = '티커별 변동성/통계 분석'
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells('A1:H1')
    stat_headers = ['티커', '일변동%', '주변동%', '5일평균변동%', '10일평균변동%',
                     '데이터일수', '5일평균낙폭%', '5일평균상승%']
    _hdr(ws, 2, stat_headers)
    for ri, ticker in enumerate(all_tickers, 3):
        if ticker not in summary_records:
            ws.cell(ri, 1).value = ticker
            continue
        rec = summary_records[ticker]
        s = rec.get('volatility_stats', {})
        vals = [ticker,
                f"{s.get('daily_vol_pct', 0):.2f}" if s else '—',
                f"{s.get('weekly_vol', 0)*100:.2f}" if s else '—',
                f"{s.get('typical_5d_move', 0):.2f}" if s else '—',
                f"{s.get('typical_10d_move', 0):.2f}" if s else '—',
                s.get('n_days', '—') if s else '—',
                f"{s.get('max_local_drawdown_5d', 0):.2f}" if s else '—',
                f"{s.get('max_local_runup_5d', 0):.2f}" if s else '—']
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci); c.value = v; c.border = _TH
            c.alignment = Alignment(horizontal='center')
            c.font = Font(size=10, bold=(ci == 1))
            if ri % 2 == 0: c.fill = _ALT
    for ci, w in enumerate([10, 11, 11, 14, 15, 12, 14, 14], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'B3'

    wb.save(summary_file)


def run_multi_ticker_analysis(tickers=None, *,
                               auto_tune_flags=None,
                               summary_file=None,
                               resume=True,
                               data_resolver=None,
                               per_ticker_overrides=None,
                               **run_kwargs):
    if tickers is None:        tickers = TICKERS
    if auto_tune_flags is None: auto_tune_flags = AUTO_TUNE
    if summary_file is None:    summary_file = os.path.join(SCRIPT_DIR, MULTI_SUMMARY_FILE)

    print('═' * 72)
    print(f'  ★ 다중 티커 분석 — {len(tickers)}개 종목')
    print(f'    티커: {tickers}')
    print(f'    이어서 진행: {resume}')
    on_vars = [k for k, v in auto_tune_flags.items() if v]
    off_vars = [k for k, v in auto_tune_flags.items() if not v]
    print(f'    자동 튜닝 ON ({len(on_vars)}): {on_vars}')
    if off_vars: print(f'    자동 튜닝 OFF ({len(off_vars)}): {off_vars}')
    print('═' * 72)

    existing, existing_order = _load_summary_state(summary_file)
    if existing:
        print(f"\n  ℹ 기존 요약 파일 발견 — {len(existing)}개 티커 기록 있음")
        print(f"    기록된 티커: {list(existing.keys())}")

    summary_records = {}
    for tk in existing:
        r = existing[tk]
        summary_records[tk] = {
            'status':    r.get('상태', '완료'),
            'signal':    r.get('신호', '—'),
            'position':  r.get('포지션', '—'),
            'close':     _try_float(str(r.get('오늘가', '')).replace('$','')),
            'cum_return_pct': _try_float(r.get('누적수익%')),
            'bh_pct':         _try_float(r.get('B&H%')),
            'up_cum_return_pct': _try_float(r.get('📈상승일전략%')),
            'bh_up_pct':         _try_float(r.get('📈상승일B&H%')),
            'vs_bh_pp':       _try_float(r.get('vsB&H%p')),
            'mdd_pct':        _try_float(r.get('MDD%')),
            'n_trades':       r.get('거래수'),
            'win_rate_pct':   _try_float(r.get('승률%')),
            'balacc_pct':     _try_float(r.get('BalAcc%')),
            'buy_balacc_pct': _try_float(r.get('매수BalAcc%')),
            'sell_balacc_pct':_try_float(r.get('매도BalAcc%')),
            'match_avg_pct':  _try_float(r.get('⚓매칭률%')),
            'match_buy_pct':  _try_float(r.get('⚓매수%')),
            'match_sell_pct': _try_float(r.get('⚓매도%')),
            'K_b_v_b':        r.get('K_b/v_b'),
            'K_s_v_s':        r.get('K_s/v_s'),
            'analyzed_at':    r.get('분석일시'),
            # ★ 변경5: tuned_vars 복원 (빈 dict 대신 실제 복원)
            'tuned_vars':     _restore_tuned_vars(r.get('_tuned_vars_raw')),
            'tuning_applied': bool(r.get('_tuned_vars_raw')),
            'tuning_source':  (r.get('_tuned_vars_raw') or {}).get('tuning_source', '기존유지'),
            'volatility_stats': {},
        }

    all_tickers_acc = list(existing_order) + [t for t in tickers if t not in existing]

    if data_resolver is None:
        data_resolver = _resolve_data_for_ticker

    t_total = time.time()
    n_done = sum(1 for tk in tickers if tk in existing)
    n_total = len(tickers)
    today_str = datetime.now().strftime('%Y-%m-%d')

    for ti, ticker in enumerate(tickers, 1):
        print(f"\n{'━' * 72}")
        print(f"  [{ti}/{n_total}] {ticker} 분석 시작")
        print('━' * 72)
        if resume and ticker in existing:
            print(f"  ⏩ 이미 완료됨 — 스킵")
            continue
        _free_global_caches()   # ★ 메모리 누적 방지 (Colab 여러 번 실행 끊김 예방)
        t_start = time.time()
        try:
            feat_t, close_t = data_resolver(ticker)
            if feat_t is None or close_t is None or len(close_t) < 30:
                raise RuntimeError(f"데이터 부족 또는 미로드 (n={len(close_t) if close_t is not None else 0})")

            stats = analyze_ticker_volatility(close_t)

            use_overrides = bool(per_ticker_overrides and ticker in per_ticker_overrides
                                  and per_ticker_overrides[ticker])
            if use_overrides:
                tuned = {}
                tuning_applied = False
                tuning_source = 'EXCEL'
                ov = per_ticker_overrides[ticker]
                print(f"\n  📋 Excel 변수값 모드 ({ticker}) — 자동 튜닝 OFF")
                for k, v in ov.items():
                    print(f"     {k:32}: {v}")
            else:
                tuned, _ = auto_tune_variables(close_t, feat_t, auto_tune_flags=auto_tune_flags)
                tuning_applied = bool(tuned)
                tuning_source = 'AUTO' if tuning_applied else '기본값'
                if tuning_applied:
                    print(f"\n  ⚙ 자동 튜닝 결과 ({ticker}):")
                    for k, v in tuned.items():
                        print(f"     {k:32}: {v}")
            print(f"  📊 통계: 일변동 {stats['daily_vol_pct']:.2f}%, "
                  f"5일변동 {stats['typical_5d_move']:.2f}%, n={stats['n_days']}")

            kwargs = dict(run_kwargs)
            if 'EVAL_START' in tuned and 'eval_start' not in kwargs:
                kwargs['eval_start'] = tuned['EVAL_START']
            mapping = {
                'HORIZON_DAYS':                'horizon',
                'DRAWDOWN_LIMIT_BUY':          'dd_limit',
                'RUNUP_LIMIT_SELL':            'ru_limit',
                'AUTO_ANCHOR_WINDOW':          'auto_anchor_window',
                'AUTO_ANCHOR_LOOKFORWARD':     'auto_anchor_lookforward',
                'AUTO_ANCHOR_MIN_RISE':        'auto_anchor_min_rise',
                'AUTO_ANCHOR_MIN_DROP':        'auto_anchor_min_drop',
                'AUTO_ANCHOR_PRICE_TOLERANCE': 'auto_anchor_price_tolerance',
            }
            for tk_k, rn_k in mapping.items():
                if tk_k in tuned and rn_k not in kwargs:
                    kwargs[rn_k] = tuned[tk_k]
            if any(k in tuned for k in ('wilson_z', 'pct_range', 'corr_limit')):
                base_mg = dict(META_GRID)
                if 'wilson_z'   in tuned: base_mg['wilson_z']   = [tuned['wilson_z']]
                if 'pct_range'  in tuned: base_mg['pct_range']  = [tuned['pct_range']]
                if 'corr_limit' in tuned: base_mg['corr_limit'] = [tuned['corr_limit']]
                kwargs['meta_grid'] = base_mg

            if use_overrides:
                ov = per_ticker_overrides[ticker]
                for k, v in ov.items():
                    if k == 'meta_grid':
                        kwargs['meta_grid'] = v
                    else:
                        kwargs[k] = v

            globals()['_pair_feat']   = feat_t
            globals()['_pair_close']  = close_t
            globals()['_pair_ticker'] = ticker

            individual_file = os.path.join(SCRIPT_DIR, f'ensemble_search_{ticker}_{today_str}.xlsx')
            kwargs.setdefault('output_file', individual_file)

            if (not use_overrides) and STAGED_META_TUNE:
                base_mg = kwargs.pop('meta_grid', META_GRID)
                print(f"\n  🔬 STAGED_META_TUNE ON ({ticker}) — 단계적 메타 변수 자동 튜닝 수행")
                result = staged_meta_tune(base_meta_grid=base_mg, **kwargs)
            else:
                result = run_ensemble_search(**kwargs)

            # ★ 통과 조합이 없어 staged가 None을 반환하면: 에러 없이 이 티커 스킵
            if result is None:
                print(f"  ⏭ {ticker}: 통과 조합 없음 → 엑셀 생성 스킵, 다음 티커로 진행")
                prev = summary_records.get(ticker)
                if prev is not None and prev.get('status') == '완료':
                    print(f"     (기존 완료 기록 유지)")
                else:
                    summary_records[ticker] = {
                        'status': '통과없음', 'signal': '— 통과조합없음', 'position': '—',
                        'close': None, 'analyzed_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'tuned_vars': {}, 'tuning_applied': False, 'volatility_stats': {},
                    }
                if globals().get('WRITE_SUMMARY_FILE', False):
                    try:
                        _save_summary_excel(summary_records, summary_file, all_tickers_acc)
                    except Exception:
                        pass
                continue

            (meta_results_df, inner_all, inner_passed,
             best_meta, best_inner, buy_pool, sell_pool,
             daily, trades, cur) = result

            bh_ret = _bh_sum_return(close_t.values) * 100.0
            bh_up_ret_t = _bh_up_sum_return(close_t.values) * 100.0
            if '보유' in cur['position']:
                if cur['sell_on_now'] and (not cur['buy_on_now'] or
                                            cur['sell_count_now']/max(cur['K_sell'],1) >= cur['buy_count_now']/max(cur['K_buy'],1)):
                    signal = '🔴 매도'
                elif cur['buy_on_now'] and cur['sell_on_now']:
                    signal = '⚔ 충돌'
                else:
                    signal = '📈 보유'
            else:
                if cur['buy_on_now'] and (not cur['sell_on_now'] or
                                            cur['buy_count_now']/max(cur['K_buy'],1) >= cur['sell_count_now']/max(cur['K_sell'],1)):
                    signal = '🟢 매수'
                elif cur['buy_on_now'] and cur['sell_on_now']:
                    signal = '⚔ 충돌'
                else:
                    signal = '💵 현금'

            match_avg  = best_inner.get('anchor_avg_match_rate',  None) if hasattr(best_inner, 'get') else None
            match_buy  = best_inner.get('anchor_buy_match_rate',  None) if hasattr(best_inner, 'get') else None
            match_sell = best_inner.get('anchor_sell_match_rate', None) if hasattr(best_inner, 'get') else None

            mg_used = kwargs.get('meta_grid', META_GRID)
            effective_vars = {
                'EVAL_START':                  kwargs.get('eval_start', EVAL_START),
                'HORIZON_DAYS':                kwargs.get('horizon', HORIZON_DAYS),
                'DRAWDOWN_LIMIT_BUY':          kwargs.get('dd_limit', DRAWDOWN_LIMIT_BUY),
                'RUNUP_LIMIT_SELL':            kwargs.get('ru_limit', RUNUP_LIMIT_SELL),
                'AUTO_ANCHOR_WINDOW':          kwargs.get('auto_anchor_window', AUTO_ANCHOR_WINDOW),
                'AUTO_ANCHOR_LOOKFORWARD':     kwargs.get('auto_anchor_lookforward', AUTO_ANCHOR_LOOKFORWARD),
                'AUTO_ANCHOR_MIN_RISE':        kwargs.get('auto_anchor_min_rise', AUTO_ANCHOR_MIN_RISE),
                'AUTO_ANCHOR_MIN_DROP':        kwargs.get('auto_anchor_min_drop', AUTO_ANCHOR_MIN_DROP),
                'AUTO_ANCHOR_PRICE_TOLERANCE': kwargs.get('auto_anchor_price_tolerance', AUTO_ANCHOR_PRICE_TOLERANCE),
                'wilson_z':   best_meta['wilson_z'],
                'pct_range':  (int(best_meta['pct_low']), int(best_meta['pct_high'])),
                'corr_limit': best_meta['corr_limit'],
                'min_signals':   int(best_meta.get('min_signals', 0)) if best_meta.get('min_signals') is not None else None,
                'top_n_pool':    int(best_meta.get('top_n_pool_buy', 0)) if best_meta.get('top_n_pool_buy') is not None else None,
                'K_buy':     int(best_inner['K_buy']),
                'vote_buy':  int(best_inner['vote_buy']),
                'K_sell':    int(best_inner['K_sell']),
                'vote_sell': int(best_inner['vote_sell']),
            }

            new_record = {
                'status': '완료',
                'signal': signal,
                'position': cur['position'],
                'close': cur['current_price'],
                'cum_return_pct': cur['cum_return_pct'],
                'bh_pct': bh_ret,
                'vs_bh_pp': cur['cum_return_pct'] - bh_ret,
                'up_cum_return_pct': cur.get('up_cum_return_pct', 0.0),
                'bh_up_pct': bh_up_ret_t,
                'mdd_pct': cur['max_drawdown'] * 100.0,
                'n_trades': cur['n_trades'],
                'win_rate_pct': cur['win_rate'] * 100.0,
                'balacc_pct':      cur['avg_success_rate']  * 100.0,
                'buy_balacc_pct':  cur['buy_success_rate']  * 100.0,
                'sell_balacc_pct': cur['sell_success_rate'] * 100.0,
                'match_avg_pct':  match_avg  * 100.0 if match_avg  is not None and pd.notna(match_avg)  else None,
                'match_buy_pct':  match_buy  * 100.0 if match_buy  is not None and pd.notna(match_buy)  else None,
                'match_sell_pct': match_sell * 100.0 if match_sell is not None and pd.notna(match_sell) else None,
                'K_b_v_b': f"{int(best_inner['K_buy'])}/{int(best_inner['vote_buy'])}",
                'K_s_v_s': f"{int(best_inner['K_sell'])}/{int(best_inner['vote_sell'])}",
                'analyzed_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'tuned_vars': effective_vars,
                'tuning_applied': tuning_applied,
                'tuning_source': tuning_source,
                'volatility_stats': stats,
            }

            # ★ 동일 티커 재실행 비교 — 그리드 선정과 같은 우선순위:
            #   매도성공률 → 평균성공률 → 누적수익 (밴드 SELECTION_TOLERANCE)
            prev = summary_records.get(ticker)
            _tol = SELECTION_TOLERANCE
            def _is_new_better(prev_rec, new_rec, tol=_tol):
                if prev_rec is None:
                    return True
                ps = prev_rec.get('sell_balacc_pct'); pa = prev_rec.get('balacc_pct'); pr = prev_rec.get('cum_return_pct')
                ns = new_rec.get('sell_balacc_pct');  na = new_rec.get('balacc_pct');  nr = new_rec.get('cum_return_pct')
                if ps is None or pa is None or pr is None:
                    return True
                if ns is None or na is None or nr is None:
                    return False
                tolp = tol * 100.0
                if ns > ps + tolp: return True
                if ns < ps - tolp: return False
                if na > pa + tolp: return True
                if na < pa - tolp: return False
                return nr > pr
            prev_ret = prev.get('cum_return_pct') if prev else None
            new_ret = new_record['cum_return_pct']
            if not _is_new_better(prev, new_record):
                print(f"  ℹ {ticker}: 기존(매도 {prev.get('sell_balacc_pct',0):.1f}%/평균 {prev.get('balacc_pct',0):.1f}%/수익 {prev_ret:+.2f}%) 우선 → 기존 유지")
                summary_records[ticker]['analyzed_at'] = new_record['analyzed_at']
            else:
                if prev_ret is not None:
                    print(f"  ✓ {ticker}: 신규(매도 {new_record.get('sell_balacc_pct',0):.1f}%/평균 {new_record.get('balacc_pct',0):.1f}%/수익 {new_ret:+.2f}%) 우선순위상 더 나음 → 갱신")
                summary_records[ticker] = new_record
            el = time.time() - t_start
            print(f"\n  ✓ {ticker} 완료 — {el:.1f}초, 수익 {cur['cum_return_pct']:+.2f}%, 신호 {signal}")
        except Exception as e:
            import traceback
            el = time.time() - t_start
            print(f"\n  ✗ {ticker} 에러 ({el:.1f}초): {e}")
            traceback.print_exc()
            prev = summary_records.get(ticker)
            if prev is not None and prev.get('status') == '완료':
                print(f"  ℹ {ticker}: 이번 분석 에러지만 기존 완료 기록 유지")
            else:
                summary_records[ticker] = {
                    'status': '에러',
                    'signal': '— 에러',
                    'position': '—',
                    'close': None,
                    'analyzed_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'error_msg': str(e),
                    'tuned_vars': {},
                    'tuning_applied': False,
                    'volatility_stats': {},
                }

        if globals().get('WRITE_SUMMARY_FILE', False):
            try:
                _save_summary_excel(summary_records, summary_file, all_tickers_acc)
                n_done += 1
                print(f"  💾 요약 저장됨 ({summary_file}) — {n_done}/{n_total} 진행")
            except Exception as e:
                print(f"  ⚠ 요약 저장 실패: {e}")
        else:
            n_done += 1

        # ★ 메모리 정리 — 한 종목이 끝나면 큰 객체를 비워 RAM 누적/폭증 방지.
        #   (73개 × 2800피처를 연속 돌리면 정리 없이는 RAM이 계속 쌓여 끊김)
        try: del result
        except Exception: pass
        try: del meta_results_df, inner_all, inner_passed, buy_pool, sell_pool, daily, trades
        except Exception: pass
        try: del feat_t, close_t
        except Exception: pass
        for _gk in ('_pair_feat', '_pair_close'):
            if _gk in globals():
                globals()[_gk] = None
        import gc as _gc
        _gc.collect()

    t_el = time.time() - t_total
    n_done_total = len([r for r in summary_records.values() if r.get('status') == '완료'])
    print(f"\n{'═' * 72}")
    print(f"  ★ 이번 실행 완료 — {n_total}개 처리  ({t_el/60:.1f}분 소요)")
    print(f"    완료: 전체 {len(all_tickers_acc)}개 티커 중 {n_done_total}개")
    print('═' * 72)

    if AUTO_DOWNLOAD_EXCEL:
        today_str = datetime.now().strftime('%Y-%m-%d')
        files_to_download = []   # ★ summary 파일 안 만듦 → 종목별 엑셀만
        for ticker in tickers:
            candidate = os.path.join(SCRIPT_DIR, f'ensemble_search_{ticker}_{today_str}.xlsx')
            if os.path.exists(candidate):
                files_to_download.append(candidate)
            else:
                import glob
                matches = sorted(glob.glob(os.path.join(SCRIPT_DIR, f'ensemble_search_{ticker}_*.xlsx')),
                                 key=os.path.getmtime, reverse=True)
                if matches:
                    files_to_download.append(matches[0])
        _auto_download_excels(files_to_download)

    return summary_records


def _auto_download_excels(file_paths, *, verbose=True):
    # 생성된 파일 경로를 전역에 누적 (노트북 셀에서 직접 다운로드할 수 있도록)
    g = globals()
    if '_GENERATED_FILES' not in g:
        g['_GENERATED_FILES'] = []
    for fp in file_paths:
        if fp and os.path.exists(fp) and fp not in g['_GENERATED_FILES']:
            g['_GENERATED_FILES'].append(fp)

    if not AUTO_DOWNLOAD_EXCEL:
        if verbose:
            print("  ℹ AUTO_DOWNLOAD_EXCEL=False — 자동 다운로드 건너뜀")
        return

    existing = [fp for fp in file_paths if fp and os.path.exists(fp)]
    missing  = [fp for fp in file_paths if fp and not os.path.exists(fp)]

    if verbose:
        print(f"\n{'─' * 72}")
        print(f"  📥 Excel 자동 다운로드 — {len(existing)}개 파일")
        print(f"{'─' * 72}")
        for fp in existing:
            sz_kb = os.path.getsize(fp) / 1024
            print(f"     ✓ {fp}  ({sz_kb:.1f} KB)")
        if missing:
            for fp in missing:
                print(f"     ⚠ 누락: {fp}")

    try:
        from google.colab import files
        is_colab = True
    except ImportError:
        is_colab = False

    if is_colab:
        for fp in existing:
            try:
                files.download(fp)
                if verbose:
                    print(f"     📥 다운로드 시작: {fp}")
            except Exception as e:
                print(f"     ⚠ 다운로드 실패 ({fp}): {e}")
        if verbose:
            print(f"  ✓ Colab 다운로드 트리거 완료\n")
    else:
        if verbose:
            print(f"  ℹ Colab 환경 아님 — 위 경로에서 직접 가져가세요\n")


def _try_float(x):
    if x is None: return None
    try:
        s = str(x).replace('%','').replace('+','').strip()
        return float(s)
    except Exception:
        return None


def _parse_overrides_from_excel(summary_file, tickers, *, verbose=True):
    if not os.path.exists(summary_file):
        if verbose: print(f"  ℹ 요약 파일 없음 — Excel 오버라이드 비활성")
        return {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(summary_file, data_only=True)
        if '사용된 변수값' not in wb.sheetnames:
            if verbose: print(f"  ⚠ '사용된 변수값' 시트 없음")
            return {}
        ws = wb['사용된 변수값']
        header_row = None
        for ri, row in enumerate(ws.iter_rows(values_only=True), 1):
            if ri == 2:
                header_row = list(row)
                break
        if not header_row:
            return {}

        kv_map = {}
        if '결과 요약' in wb.sheetnames:
            ws2 = wb['결과 요약']
            hdr2 = None
            for ri, row in enumerate(ws2.iter_rows(values_only=True), 1):
                if ri == 2:
                    hdr2 = list(row); continue
                if ri <= 2 or not row or row[0] is None: continue
                tk = str(row[0]).strip()
                if not tk or tk == '티커': continue
                d2 = dict(zip(hdr2, row)) if hdr2 else {}
                def _split_kv(s):
                    try:
                        a, b = str(s).split('/')
                        return int(float(a.strip())), int(float(b.strip()))
                    except Exception:
                        return None, None
                kb, vb = _split_kv(d2.get('K_b/v_b'))
                ks, vs = _split_kv(d2.get('K_s/v_s'))
                if kb is not None and ks is not None:
                    kv_map[tk] = (kb, vb, ks, vs)

        def _pct_to_ratio(v):
            if v is None: return None
            s = str(v).strip().replace('+','')
            if s == '—' or s == '': return None
            has_pct = '%' in s
            try:
                f = float(s.replace('%','').strip())
                return f / 100.0 if has_pct else f
            except Exception:
                return None

        def _maybe_int(v):
            if v is None: return None
            try:
                s = str(v).strip()
                if s == '—' or s == '': return None
                return int(float(s))
            except Exception:
                return None

        def _maybe_float(v):
            if v is None: return None
            try:
                s = str(v).strip()
                if s == '—' or s == '': return None
                return float(s)
            except Exception:
                return None

        def _parse_pct_range(v):
            if v is None: return None
            s = str(v).strip()
            if s == '—' or s == '': return None
            try:
                s = s.replace('(','').replace(')','').replace('[','').replace(']','')
                parts = [int(float(x.strip())) for x in s.split(',') if x.strip()]
                if len(parts) == 2: return tuple(parts)
            except Exception:
                return None
            return None

        result = {}
        for ri, row in enumerate(ws.iter_rows(values_only=True), 1):
            if ri <= 2: continue
            if not row or row[0] is None: continue
            ticker = str(row[0]).strip()
            if ticker not in tickers: continue
            d = dict(zip(header_row, row))

            ov = {}
            v = d.get('EVAL_START')
            if v not in (None, '—', ''): ov['eval_start'] = str(v).strip()
            v = _maybe_int(d.get('HORIZON'))
            if v is not None: ov['horizon'] = v
            v = _pct_to_ratio(d.get('DD한도'))
            if v is not None: ov['dd_limit'] = v
            v = _pct_to_ratio(d.get('RU한도'))
            if v is not None: ov['ru_limit'] = v
            v = _maybe_int(d.get('ANCHOR_WIN'))
            if v is not None: ov['auto_anchor_window'] = v
            v = _maybe_int(d.get('ANCHOR_LF'))
            if v is not None: ov['auto_anchor_lookforward'] = v
            v = _pct_to_ratio(d.get('MIN_RISE'))
            if v is not None: ov['auto_anchor_min_rise'] = v
            v = _pct_to_ratio(d.get('MIN_DROP'))
            if v is not None: ov['auto_anchor_min_drop'] = v
            v = _pct_to_ratio(d.get('PRICE_TOL'))
            if v is not None: ov['auto_anchor_price_tolerance'] = v

            mg_part = {}
            v = _maybe_float(d.get('wilson_z'))
            if v is not None: mg_part['wilson_z'] = [v]
            v = _parse_pct_range(d.get('pct_range'))
            if v is not None: mg_part['pct_range'] = [v]
            v = _maybe_float(d.get('corr_limit'))
            if v is not None: mg_part['corr_limit'] = [v]
            v = _maybe_int(d.get('MIN_SIG'))
            if v is not None: mg_part['min_signals'] = [v]
            v = _maybe_int(d.get('TOP_N'))
            if v is not None: mg_part['top_n_pool'] = [v]
            if mg_part:
                mg = dict(META_GRID)
                mg.update(mg_part)
                ov['meta_grid'] = mg

            kb = vb = ks = vs = None
            def _split_kv_local(s):
                try:
                    a, b = str(s).split('/')
                    return int(float(a.strip())), int(float(b.strip()))
                except Exception:
                    return None, None
            kb, vb = _split_kv_local(d.get('K_b/v_b'))
            ks, vs = _split_kv_local(d.get('K_s/v_s'))
            if (kb is None or ks is None) and ticker in kv_map:
                kb, vb, ks, vs = kv_map[ticker]
            if kb is not None and ks is not None:
                ov['k_buy_range']  = [kb]
                ov['k_sell_range'] = [ks]
                ov['vote_ratio_buy']  = [vb / kb] if kb > 0 else [1.0]
                ov['vote_ratio_sell'] = [vs / ks] if ks > 0 else [1.0]

            if ov:
                result[ticker] = ov
        return result
    except Exception as e:
        if verbose: print(f"  ⚠ Excel 변수 파싱 실패: {e}")
        import traceback; traceback.print_exc()
        return {}


def _resolve_data_for_ticker(ticker, end_date=None):
    g = globals()
    pdm = g.get('_pair_data_map')
    if isinstance(pdm, dict) and ticker in pdm:
        entry = pdm[ticker]
        if isinstance(entry, tuple) and len(entry) == 2:
            return entry[0], entry[1]
        if isinstance(entry, dict):
            return entry.get('feat'), entry.get('close')

    if 'load_ticker_data' in g and callable(g['load_ticker_data']):
        try:
            return g['load_ticker_data'](ticker)
        except Exception as e:
            print(f"  ⚠ load_ticker_data({ticker}) 실패: {e}")

    cache = g.setdefault('_multi_ticker_cache', {})
    if ticker in cache:
        print(f"  ℹ 캐시 사용: {ticker}")
        return cache[ticker]

    needed = ['download_data', 'compute_features']
    if all(n in g for n in needed):
        try:
            start = g.get('DOWNLOAD_START', '2020-01-01')

            old_ticker = g.get('TICKER')
            g['TICKER'] = ticker
            print(f"  ⚙ TICKER 글로벌 변경: {old_ticker} → {ticker}")

            dl_func = g['download_data']
            dl_varnames = dl_func.__code__.co_varnames[:dl_func.__code__.co_argcount]
            try:
                if 'tickers' in dl_varnames:
                    closes, ohlcv = dl_func(tickers=[ticker], start=start)
                else:
                    closes, ohlcv = dl_func(start=start)
            except TypeError:
                closes, ohlcv = dl_func(start=start)

            # ★ 최신 거래일 보충 (요청) — download_data 데이터가 '완료된 세션을 빠뜨릴 만큼'
            #   오래됐을 때만(오늘보다 2영업일 이상 전) yfinance로 다시 받는다.
            #   오늘 17일에 데이터가 16일까지면 = 어제(최신 완료 세션)이므로 정상 → 재조회 안 함.
            #   (1영업일 전은 '오늘장 미마감'이라 정상이므로 건드리지 않음)
            _did_yf_refetch = False
            _orig_ohlcv_t = ohlcv.get(ticker)      # 보충 실패 시 되돌릴 원본
            _orig_close_t = closes.get(ticker) if hasattr(closes, 'get') else None
            _need_yf = ticker not in ohlcv
            if not _need_yf and end_date is None:
                try:
                    _last0 = ohlcv[ticker].index[-1]
                    _today0 = pd.Timestamp(datetime.now().date())
                    if len(pd.bdate_range(_last0.normalize(), _today0)) - 1 >= 2:
                        _need_yf = True
                        print(f"  ℹ {ticker} download_data 데이터가 {str(_last0)[:10]}까지 — "
                              f"최신 거래일 보충 위해 yfinance 재조회")
                except Exception:
                    pass
            if _need_yf:
                if ticker not in ohlcv:
                    print(f"  ⚠ {ticker}가 download_data 결과에 없음 — yfinance 직접 호출")
                try:
                    import yfinance as yf
                    # ★ end를 '내일'로 명시 — yfinance는 end 미지정 시 당일을 빠뜨릴 수 있음.
                    #   (특히 장 마감 전/시간대 차이로 전날까지만 받아지는 문제 완화)
                    _end = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')
                    df = yf.download(ticker, start=start, end=_end,
                                     progress=False, auto_adjust=False)
                    if df is None or len(df) == 0:
                        raise RuntimeError(f"yfinance에서 {ticker} 데이터 0건")
                    if isinstance(df.columns, pd.MultiIndex):
                        df.columns = df.columns.get_level_values(0)
                    ohlcv[ticker] = df
                    closes[ticker] = df['Close'] if 'Close' in df.columns else df.iloc[:, 0]
                    _did_yf_refetch = True
                    _last = df.index[-1]
                    _today = pd.Timestamp(datetime.now().date())
                    _bdays = len(pd.bdate_range(_last.normalize(), _today)) - 1
                    print(f"  ✓ yfinance로 {ticker} 다운로드 성공 ({len(df)}일, 마지막 {str(_last)[:10]})")
                    if _bdays >= 1:
                        print(f"  ⚠ 최신 거래일 누락 가능 — 마지막 데이터가 {str(_last)[:10]}로 "
                              f"오늘({_today.date()})보다 {_bdays}영업일 전입니다.")
                        print(f"     원인: 미국장 마감 전 실행 or yfinance 반영 지연. "
                              f"장 마감 후(한국시간 다음날 오전 이후) 다시 받으면 최신일이 포함됩니다.")
                except Exception as _yf_e:
                    if ticker not in ohlcv:
                        raise RuntimeError(
                            f"{ticker}가 download_data 결과에 없고 yfinance 보충도 실패했습니다: {_yf_e}\n"
                            f"  download_data 함수의 티커 리스트에 '{ticker}'를 추가하거나,\n"
                            f"  yfinance 설치: pip install yfinance")
                    else:
                        print(f"  ⚠ 최신 거래일 보충 실패({_yf_e}) — "
                              f"download_data 데이터({str(ohlcv[ticker].index[-1])[:10]}까지)로 진행")

            fred_df = None
            if 'download_fred_data' in g:
                try: fred_df = g['download_fred_data'](start=start)
                except Exception: fred_df = None
                if fred_df is not None and len(fred_df) == 0: fred_df = None

            cf_func = g['compute_features']
            def _call_cf():
                try:
                    return cf_func(ohlcv, closes, fred_df=fred_df, ticker=ticker)
                except TypeError:
                    return cf_func(ohlcv, closes, fred_df=fred_df)

            # ★ 연장용 전체 계산 (요청, 기본 OFF) — 자르기 '전'에 새 거래일 행을 확보.
            #   ⚠ 반드시 '복사본'으로 계산한다. 그래야 compute_features가 입력을 건드려도
            #   아래 '자른 정확 재현' 계산이 오염되지 않는다 (과거 신호가 어긋나던 문제 차단).
            feat_full = close_full = None
            if end_date is not None and globals().get('REPLAY_EXTEND_TO_TODAY', False):
                try:
                    _oh = {t: (df.copy() if hasattr(df, 'copy') else df) for t, df in ohlcv.items()}
                    _cl = closes.copy() if hasattr(closes, 'copy') else dict(closes)
                    _fr = fred_df.copy() if fred_df is not None else None
                    try:
                        feat_full = cf_func(_oh, _cl, fred_df=_fr, ticker=ticker)
                    except TypeError:
                        feat_full = cf_func(_oh, _cl, fred_df=_fr)
                    if ticker in _oh:
                        close_full = _oh[ticker]['Close'].reindex(feat_full.index)
                    else:
                        feat_full = None
                except Exception as _fe:
                    print(f"  ⚠ 연장용 전체 계산 실패(연장 생략): {_fe}")
                    feat_full = close_full = None

            # ★ 원본 범위로 자르기 (요청) — end_date까지로 잘라 '정확 재현'. 전체구간 순위/정규화
            #   지표가 원본과 '같은 범위'로 계산돼 신호가 정확히 일치 (하루 더 받아 순위 바뀌던 문제 해결).
            if end_date is not None:
                try:
                    _ed = pd.Timestamp(end_date)
                    for _t in list(ohlcv.keys()):
                        try: ohlcv[_t] = ohlcv[_t].loc[ohlcv[_t].index <= _ed]
                        except Exception: pass
                    if hasattr(closes, 'columns'):              # DataFrame
                        closes = closes.loc[closes.index <= _ed]
                    elif isinstance(closes, dict):
                        for _t in list(closes.keys()):
                            try: closes[_t] = closes[_t].loc[closes[_t].index <= _ed]
                            except Exception: pass
                    if fred_df is not None:
                        try: fred_df = fred_df.loc[fred_df.index <= _ed]
                        except Exception: pass
                    print(f"  ✂ 원본 범위로 자름: {str(_ed)[:10]}까지 — 같은 범위로 계산해 정확 재현")
                except Exception as _te:
                    print(f"  ⚠ 원본 범위 자르기 실패(무시): {_te}")

            try:
                feat = _call_cf()
            except Exception as _cf_e:
                # ★ yfinance 보충 데이터가 타 종목(download_data)과 인덱스가 안 맞아 깨진 경우
                #   ('Can only compare identically-labeled Series' 등) → 원본으로 되돌려 재계산.
                if _did_yf_refetch and _orig_ohlcv_t is not None:
                    print(f"  ⚠ yfinance 보충 데이터로 피처 계산 실패 — "
                          f"download_data 원본({str(_orig_ohlcv_t.index[-1])[:10]}까지)으로 되돌려 재계산")
                    print(f"     (사유: {_cf_e})")
                    ohlcv[ticker] = _orig_ohlcv_t
                    if _orig_close_t is not None:
                        closes[ticker] = _orig_close_t
                    feat = _call_cf()
                else:
                    raise

            if ticker not in ohlcv:
                raise RuntimeError(f"{ticker} OHLCV 없음 (yfinance fallback 실패)")
            close = ohlcv[ticker]['Close'].reindex(feat.index)

            # ★ 새 거래일 덧붙이기 (요청) — 원본 마지막날 '이후' 행을 전체계산본에서 가져와 이어붙인다.
            #   → 과거는 정확 재현(자른 것), 새 거래일은 같은 로직으로 계산한 결과.
            if feat_full is not None and close_full is not None and end_date is not None:
                try:
                    _ed = pd.Timestamp(end_date)
                    _newidx = feat_full.index[feat_full.index > _ed]
                    if len(_newidx) > 0:
                        feat = pd.concat([feat, feat_full.reindex(columns=feat.columns).loc[_newidx]])
                        close = pd.concat([close, close_full.loc[_newidx]])
                        print(f"  ➕ 원본 이후 {len(_newidx)}일({str(_newidx[0])[:10]}~{str(_newidx[-1])[:10]}) "
                              f"같은 로직으로 이어 계산 → 최종 {str(feat.index[-1])[:10]}까지")
                    else:
                        print(f"  ℹ 원본 이후 새 거래일이 데이터에 없음 — 원본 날짜까지만 (연장할 데이터 없음)")
                except Exception as _xe:
                    print(f"  ⚠ 새 거래일 덧붙이기 실패(무시): {_xe}")

            cache[ticker] = (feat, close)
            return feat, close
        except Exception as e:
            raise RuntimeError(f"데이터 로드 실패 ({ticker}): {e}")

    raise RuntimeError(
        f"{ticker} 데이터를 로드할 수 없습니다. 다음 중 하나를 준비하세요:\n"
        f"  1. _pair_data_map = {{'AAPL': (feat, close), 'MPC': (feat, close), ...}}\n"
        f"  2. def load_ticker_data(ticker): return feat, close  # 글로벌 함수\n"
        f"  3. download_data / compute_features 함수 (기존 파이프라인)")


# ════════════════════════════════════════════════════════════════
#   ★ 모드 4 — 드라이브 폴더의 티커별 ★최적 조합을 '현재까지' 일괄 재현 + 요약 (요청)
# ════════════════════════════════════════════════════════════════
def _build_selected_summary(source_paths, out_dir, date_label, *, sheet_name='내부_그리드_통과'):
    """각 소스 엑셀의 내부 그리드 ★(선정) 행을 읽어 티커별 요약 엑셀을 만든다.
       (스트리밍 읽기 + ★ 없으면 건너뜀)."""
    import re as _re
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    FRE = _re.compile(r'ensemble_search_(.+)_(\d{4}-\d{2}-\d{2})\.xlsx$')

    def _read_star(path):
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            print(f"    ⚠ 열기 실패: {os.path.basename(path)} — {e}"); return None, None
        if sheet_name not in wb.sheetnames:
            print(f"    ⚠ '{sheet_name}' 시트 없음: {os.path.basename(path)}"); wb.close(); return None, None
        ws = wb[sheet_name]; header = None; sel = None; saw = False
        for row in ws.iter_rows(values_only=True):
            if header is None:
                sv = [str(v).strip() if v is not None else '' for v in row]
                if '#' in sv and 'K_buy' in sv:
                    header = [s if s else f'col{i+1}' for i, s in enumerate(sv)]
                continue
            c0 = row[0] if row else None
            if c0 is None: continue
            saw = True
            if '★' in str(c0): sel = list(row); break
        wb.close()
        if sel is None:
            print(f"    – ★ 없음 — 건너뜀: {os.path.basename(path)}" if (header and saw)
                  else f"    ⚠ 헤더/데이터 없음 — 건너뜀: {os.path.basename(path)}")
            return None, None
        return header, sel

    rows = []; master = None
    for p in source_paths:
        m = FRE.search(os.path.basename(p)); tk = m.group(1) if m else os.path.basename(p)
        dt = m.group(2) if m else ''
        header, vals = _read_star(p)
        if header is None: continue
        if master is None: master = header
        rec = {'티커': tk, '실행일': dt}
        rec.update({h: v for h, v in zip(header, vals)})
        rows.append(rec)
        print(f"    ✓ {tk:<6} ★행 읽음 (그리드 #{str(vals[0]).replace('★','').strip()})")
    if not rows:
        print("    ❌ 요약할 ★행이 없습니다."); return None

    out_cols = ['티커', '실행일'] + master
    wb = Workbook(); ws = wb.active; ws.title = '티커별_선정요약'; ws.sheet_view.showGridLines = False
    THIN = Side(style='thin', color='D9D9D9'); BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HF = PatternFill('solid', fgColor='1F3864'); ALT = PatternFill('solid', fgColor='F2F2F2')
    POS = PatternFill('solid', fgColor='E8F5E9'); NEG = PatternFill('solid', fgColor='FFEBEE')
    ws.cell(1, 1).value = f'티커별 선정(★) 요약 — {date_label} — {len(rows)}개 티커'
    ws.cell(1, 1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out_cols))
    for ci, h in enumerate(out_cols, 1):
        c = ws.cell(3, ci); c.value = h; c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = HF; c.border = BRD; c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    pos_i = out_cols.index('📍실행일포지션') + 1 if '📍실행일포지션' in out_cols else None
    ret_i = [out_cols.index(c) + 1 for c in ('🔧보정후수익', '✅실제누적수익') if c in out_cols]
    for ri, rec in enumerate(rows):
        r = ri + 4
        for ci, h in enumerate(out_cols, 1):
            c = ws.cell(r, ci); c.value = rec.get(h, '—'); c.border = BRD; c.font = Font(size=10)
            c.alignment = Alignment(horizontal='center')
            if ri % 2 == 1: c.fill = ALT
        ws.cell(r, 1).font = Font(bold=True, size=11, color='1F3864')
        if pos_i is not None:
            pv = str(rec.get('📍실행일포지션', ''))
            if '보유' in pv: ws.cell(r, pos_i).fill = POS
            elif '현금' in pv: ws.cell(r, pos_i).fill = NEG
        for rx in ret_i: ws.cell(r, rx).font = Font(bold=True, size=10, color='C00000')
    for ci, h in enumerate(out_cols, 1):
        w = 8 if h == '티커' else (12 if h == '실행일' else 12)
        if h in ('🎬실행일액션', '📅최근매수일', '📅최근매도일', '🔧추가지표수'): w = 14
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 34; ws.freeze_panes = 'C4'
    out_path = os.path.join(out_dir, f'summary_selected_{date_label}.xlsx')
    wb.save(out_path)
    print(f"  ✅ 요약 저장: {out_path}  ({len(rows)}개 티커)")
    globals().setdefault('_GENERATED_FILES', []).append(out_path)
    return out_path


def run_mode4_drive_reproduce_all(drive_dir=None, *, date_subfolder=None, **override_kwargs):
    """모드 4 — 드라이브 ensemble_analysis 폴더에서 '티커별 가장 최근' 분석 엑셀을 모두 찾아,
       각 엑셀의 ★최적(내부 그리드 별표) 조합을 '현재까지' 일별 백테스트로 그대로 재현해
       새 엑셀을 만들고, 티커별 선정행 요약 엑셀도 만든다. 모든 결과물은 '날짜 폴더'에 저장."""
    import glob, re as _re
    g = globals()
    src = drive_dir or g.get('RUN_MODE4_DRIVE_DIR', '/content/drive/MyDrive/ensemble_analysis')
    if not os.path.isdir(src):
        print(f"  ✗ 드라이브 폴더를 찾을 수 없습니다: {src}")
        print(f"    먼저 드라이브 마운트: from google.colab import drive; drive.mount('/content/drive')")
        return []
    # 티커별 가장 최근 파일 (재현본 __replay 제외)
    latest = {}
    for p in glob.glob(os.path.join(src, "ensemble_search_*_*.xlsx")):
        name = os.path.basename(p)
        if '__replay' in name: continue
        m = _re.search(r"ensemble_search_(.+)_(\d{4}-\d{2}-\d{2})\.xlsx$", name)
        if not m: continue
        tk, dt = m.group(1), m.group(2)
        if tk not in latest or dt > latest[tk][0]:
            latest[tk] = (dt, p)
    if not latest:
        print(f"  ✗ {src} 에서 ensemble_search_<티커>_<날짜>.xlsx 형식 파일을 못 찾았습니다.")
        return []
    print(f"  📂 드라이브 폴더: {src}  — 티커 {len(latest)}개")

    today = datetime.now().strftime('%Y-%m-%d')
    sub = date_subfolder or today
    out_dir = os.path.join(src, sub)            # ★ 드라이브 안에 '날짜 폴더' 생성
    os.makedirs(out_dir, exist_ok=True)
    print(f"  📁 출력 날짜 폴더: {out_dir}\n")

    # 각 티커 ★최적 조합을 '현재까지' 재현 → 날짜 폴더에 엑셀 생성
    for tk, (dt, path) in sorted(latest.items()):
        print(f"  ─ [{tk}] {os.path.basename(path)} → 현재까지 재현 ─")
        # ★ 핵심 — 티커마다 '현재 날짜까지' 데이터를 새로 받도록 메모리 캐시를 비운다.
        #   안 비우면 이전 실행/이전 티커의 옛 데이터를 재사용해 원본과 같은 날짜까지만 나옴.
        for _k in ('_pair_feat', '_pair_close', '_pair_ticker'):
            g[_k] = None
        _mc = g.get('_multi_ticker_cache')
        if isinstance(_mc, dict):
            _mc.pop(tk, None)
        _pdm = g.get('_pair_data_map')
        if isinstance(_pdm, dict):
            _pdm.pop(tk, None)
        _free_global_caches(keep_pool_map=False)   # z-캐시 등도 정리
        try:
            # feat=None,close=None 명시 → 반드시 티커로 '신선한' 데이터를 새로 확보
            replay_grid_combo(path, None, output_dir=out_dir,
                              feat=None, close=None, **override_kwargs)
        except Exception as e:
            print(f"    ⚠ {tk} 재현 실패(건너뜀): {e}")

    # 티커별 선정행 요약 엑셀 (소스 ★행 기준) — 날짜 폴더에 저장
    print(f"\n  📑 티커별 선정행 요약 생성...")
    _build_selected_summary([p for _, (_, p) in sorted(latest.items())], out_dir, today)
    return g.get('_GENERATED_FILES', [])


def main():
    """RUN_* 변수(코드 상단 또는 노트북 셀에서 mod.RUN_MODE=... 로 변경)에 따라 실행.
       노트북에서: 모듈 로드 → mod.RUN_MODE 등 설정 → mod.main() 호출."""
    g = globals()
    g['_GENERATED_FILES'] = []   # 이번 실행에서 만든 엑셀 경로 누적
    _free_global_caches()        # ★ 이전 실행의 캐시 정리 (Colab 반복 실행 시 메모리 누적·끊김 방지)
    _mode = str(g.get('RUN_MODE', 3)).strip()
    print(f"\n[실행 모드] RUN_MODE = {_mode}")
    print("  1=새 분석  /  2=그리드 번호 재현  /  3=최근 엑셀 ★최적 자동 재현")

    if _mode == '2':
        print(f"\n  파일 경로는 OUTPUT_DIR 고정: {OUTPUT_DIR}")
        _fn = str(g.get('RUN_REPLAY_FILE', '')).strip().strip('"').strip("'")
        _gn = str(g.get('RUN_REPLAY_GRID_NUMBER', '')).strip()
        print(f"  재현 파일: {_fn}  /  그리드 번호: {_gn}")
        print("  ℹ 데이터가 메모리에 없으면 티커로 자동 다운로드를 시도합니다")
        try:
            replay_grid_combo(_fn, _gn)
        except Exception as _e:
            print(f"\n  ✗ 재현 실패: {_e}")
            print("    feat/close가 메모리에 있는지 확인하세요. 예:")
            print("      replay_grid_combo('파일명.xlsx', 14, feat=내_feat, close=내_close)")

    elif _mode == '1':
        # 분석 대상 티커 — RUN_TICKERS 가 있으면 그것, 없으면 기본 TICKERS
        _run_tickers = g.get('RUN_TICKERS', None)
        if _run_tickers:
            tickers_to_run = [str(t).strip().upper() for t in _run_tickers if str(t).strip()]
        else:
            tickers_to_run = TICKERS
        print(f"\n→ 분석 대상: {tickers_to_run}")

        summary_path = os.path.join(SCRIPT_DIR, MULTI_SUMMARY_FILE)
        overrides = None
        if g.get('RUN_USE_EXCEL_OVERRIDES', False) and os.path.exists(summary_path):
            print(f"\n기존 요약 파일에서 변수값 사용 시도: {summary_path}")
            overrides = _parse_overrides_from_excel(summary_path, tickers_to_run)
            if overrides:
                found = list(overrides.keys())
                missing = [t for t in tickers_to_run if t not in overrides]
                print(f"  ✓ Excel 변수값 적용: {found}")
                if missing:
                    print(f"  ℹ Excel에 없는 티커 (기본값 사용): {missing}")
            else:
                print(f"  ⚠ Excel에서 사용 가능한 변수값 못 찾음 → 기본값 사용")
                overrides = None
        else:
            print(f"→ 코드 상단 변수값(또는 AUTO_TUNE) 사용")

        run_multi_ticker_analysis(
            tickers=tickers_to_run,
            per_ticker_overrides=overrides,
            resume=False,
        )

    elif _mode == '4':
        # ── 모드 4 — 드라이브 폴더의 티커별 ★최적 조합을 '현재까지' 일괄 재현 + 요약 ──
        print(f"\n  드라이브 폴더의 티커별 가장 최근 분석 엑셀을 모두 찾아,")
        print(f"  각 ★최적 조합을 '현재까지' 재현하고 티커별 요약 엑셀을 만듭니다 (날짜 폴더에 저장).")
        print(f"  ℹ 드라이브가 마운트돼 있어야 합니다: from google.colab import drive; drive.mount('/content/drive')")
        try:
            run_mode4_drive_reproduce_all(
                g.get('RUN_MODE4_DRIVE_DIR', None),
                date_subfolder=g.get('RUN_MODE4_DATE_FOLDER', None))
        except Exception as _e:
            print(f"\n  ✗ 모드 4 실패: {_e}")

    else:
        # ── 모드 3 (기본) — 티커의 가장 최근 분석 엑셀에서 ★최적 조합 자동 재현 ──
        print(f"\n  엑셀 폴더(OUTPUT_DIR): {OUTPUT_DIR}")
        _tk = str(g.get('RUN_REPLAY_TICKER', '')).strip().upper()
        print(f"  '{_tk}'의 '가장 최근 일반 분석 엑셀'을 자동으로 찾아")
        print(f"  ★최적으로 선정된 변수·지표를 그대로 재현합니다 (재현본 __replay 파일은 제외).")
        if not _tk:
            print("  ⚠ RUN_REPLAY_TICKER 가 비어 종료합니다.")
        else:
            print("\n  ℹ 데이터가 메모리에 없으면 티커로 자동 다운로드를 시도합니다.")
            try:
                replay_latest_best(_tk)
            except Exception as _e:
                print(f"\n  ✗ 재현 실패: {_e}")

    # ── 생성된 파일 안내 + 셀에서 직접 다운로드할 수 있도록 경로 반환 ──
    made = g.get('_GENERATED_FILES', [])
    if made:
        print(f"\n{'═'*72}")
        print(f"  📦 생성된 엑셀 {len(made)}개:")
        for fp in made:
            print(f"     • {fp}")
        print(f"  💡 자동 다운로드가 안 뜨면, 노트북에서 아래로 직접 받으세요:")
        print(f"       from google.colab import files")
        print(f"       for f in mod.get_generated_files(): files.download(f)")
        print(f"{'═'*72}")
    return made


def get_generated_files():
    """이번 실행에서 만든 엑셀 파일 경로 목록 (노트북 셀에서 직접 다운로드용)."""
    return list(globals().get('_GENERATED_FILES', []))


if __name__ == '__main__':
    main()

